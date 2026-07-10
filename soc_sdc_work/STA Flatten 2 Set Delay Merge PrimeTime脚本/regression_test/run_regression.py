#!/usr/bin/env python3
"""Regression smoke tests for run_stage2_merge_delay.tcl.

The tests run with plain tclsh plus a small mock of the PrimeTime collection
and direction APIs. They do not replace PrimeTime validation, but they keep
parsing, matching, reporting, and static SDC emission deterministic.
"""

from __future__ import print_function

import importlib.util
import os
import shutil
import subprocess
import sys


HERE = os.path.abspath(os.path.dirname(__file__))
TOOL = os.path.abspath(os.path.join(HERE, "..", "run_stage2_merge_delay.tcl"))
REPORT_TOOL = os.path.abspath(os.path.join(HERE, "..", "run_stage2_report.py"))
WORK = os.path.join(HERE, "work")


DEFAULT_PT_PRELUDE = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg/Q out
    u_src_reg/CP in
    u_up/data_o out
    u_up/u_reg/Q out
    u_h0/cfg_i in
    u_h0/async_i in
    u_h0/unused_i in
    u_h0/other_i in
    u_h0/u_reg/D in
    u_h0/u_cfg_reg/D in
    u_h0/u_mode_reg/D in
    u_h0/i_niu_rst_n in
    u_h0/o_niu_rst_n out
    top_rst_n out
    u_mid/in_i in
    u_mid/out_o out
}

proc current_design {} {
    return current_integration_top
}

proc sizeof_collection {coll} {
    return [llength $coll]
}

proc foreach_in_collection {var coll body} {
    upvar 1 $var item
    foreach item $coll {
        uplevel 1 $body
    }
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set name [lindex $args end]
    return [list $name]
}

proc get_ports {args} {
    set name [lindex $args end]
    return [list $name]
}

proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "full_name"} {
        return $name
    }
    if {$attr eq "direction" && [info exists ::PT_MOCK_DIRECTIONS($name)]} {
        return $::PT_MOCK_DIRECTIONS($name)
    }
    return ""
}
'''


def write_file(path, text):
    directory = os.path.dirname(path)
    if directory and not os.path.isdir(directory):
        os.makedirs(directory)
    with open(path, "w") as fout:
        fout.write(text)


def read_file(path):
    with open(path, "r") as fin:
        return fin.read()


def run_case(case_name, top_sdc, harden_sdc, extra_build_args=None, extra_hardens=None, prelude=""):
    case_dir = os.path.join(WORK, case_name)
    if os.path.isdir(case_dir):
        shutil.rmtree(case_dir)
    os.makedirs(case_dir)
    top_path = os.path.join(case_dir, "top.sdc")
    harden_path = os.path.join(case_dir, "harden.sdc")
    harden_list = os.path.join(case_dir, "harden_list.csv")
    out_sdc = os.path.join(case_dir, "generated_e2e_delay.sdc")
    out_report = os.path.join(case_dir, "integration_delay_merge.rpt")
    out_removed = os.path.join(case_dir, "merged_delay_removed.sdc")
    out_review = os.path.join(case_dir, "unmerged_delay_review.rpt")
    out_final = os.path.join(case_dir, "top_flatten.sdc")
    out_summary = os.path.join(case_dir, "delay_path_summary")
    driver = os.path.join(case_dir, "run.tcl")
    write_file(top_path, top_sdc)
    write_file(harden_path, harden_sdc)
    rows = [
        "harden_name,inst_path,clean_sdc,delay_candidate_file,netlist,module",
        "h0,u_h0,%s,,h0.v,harden0" % harden_path.replace("\\", "/"),
    ]
    for item in extra_hardens or []:
        clean_sdc = ""
        module = item[2] if len(item) > 2 else item[0]
        if len(item) > 3:
            clean_sdc = os.path.join(case_dir, "%s_clean.sdc" % item[0])
            write_file(clean_sdc, item[3])
            clean_sdc = clean_sdc.replace("\\", "/")
        rows.append("%s,%s,%s,,,%s" % (item[0], item[1], clean_sdc, module))
    write_file(harden_list, "\n".join(rows) + "\n")
    args = [
        "-top_sdc", top_path.replace("\\", "/"),
        "-harden_list", harden_list.replace("\\", "/"),
        "-out_e2e_sdc", out_sdc.replace("\\", "/"),
        "-out_report", out_report.replace("\\", "/"),
        "-out_removed_sdc", out_removed.replace("\\", "/"),
        "-out_review_rpt", out_review.replace("\\", "/"),
    ]
    if extra_build_args:
        args.extend(extra_build_args)
    complete_prelude = DEFAULT_PT_PRELUDE + "\n" + prelude
    write_file(
        driver,
        '%s\nset ::STAGE2_AUTO_RUN false\nsource "%s"\nstage2_delay::build %s\n' % (
            complete_prelude,
            TOOL.replace("\\", "/"),
            " ".join('"%s"' % arg for arg in args),
        ),
    )
    proc = subprocess.Popen(
        ["tclsh", driver],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=case_dir,
    )
    stdout, stderr = proc.communicate()
    return {
        "code": proc.returncode,
        "stdout": stdout.decode("utf-8", "replace"),
        "stderr": stderr.decode("utf-8", "replace"),
        "case_dir": case_dir,
        "out_sdc": out_sdc,
        "report": out_report,
        "removed": out_removed,
        "review": out_review,
        "final": out_final,
        "summary": out_summary,
        "driver": driver,
    }


def assert_contains(path, needle):
    text = read_file(path)
    if needle not in text:
        raise AssertionError("Expected %r in %s\n--- file ---\n%s" % (needle, path, text))


def assert_exists(path):
    if not os.path.exists(path):
        raise AssertionError("Expected path to exist: %s" % path)


def assert_not_contains(path, needle):
    text = read_file(path)
    if needle in text:
        raise AssertionError("Did not expect %r in %s\n--- file ---\n%s" % (needle, path, text))


def assert_text_contains(text, needle):
    if needle not in text:
        raise AssertionError("Expected %r in text\n--- text ---\n%s" % (needle, text))


def assert_generated_delays_have_from(path):
    text = read_file(path)
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("set_max_delay ") or stripped.startswith("set_min_delay "):
            if " -from " not in stripped:
                raise AssertionError("Generated delay missing -from in %s:\n%s" % (path, stripped))


def require_ok(result):
    if result["code"] != 0:
        raise AssertionError(
            "case failed\nstdout=%s\nstderr=%s\ndriver=%s"
            % (result["stdout"], result["stderr"], read_file(result["driver"]))
        )
    if os.path.exists(result["out_sdc"]):
        assert_generated_delays_have_from(result["out_sdc"])


def test_release_identity_is_reconstructed_without_plaintext_constant():
    expected = "".join(chr(code) for code in (72, 111, 119, 97, 114, 100))
    assert_not_contains(TOOL, expected)
    assert_not_contains(REPORT_TOOL, expected)

    module_spec = importlib.util.spec_from_file_location("stage2_report_module", REPORT_TOOL)
    report_module = importlib.util.module_from_spec(module_spec)
    module_spec.loader.exec_module(report_module)
    if report_module.release_identity() != expected:
        raise AssertionError("Unexpected reconstructed Python release identity")

    driver = os.path.join(WORK, "release_identity.tcl")
    write_file(
        driver,
        'set ::STAGE2_AUTO_RUN false\nsource "%s"\nputs [stage2_delay::release_identity]\n'
        % TOOL.replace("\\", "/"),
    )
    proc = subprocess.Popen(
        ["tclsh", driver],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=WORK,
    )
    stdout, stderr = proc.communicate()
    if proc.returncode != 0:
        raise AssertionError(
            "identity reconstruction failed\nstdout=%s\nstderr=%s"
            % (stdout.decode("utf-8", "replace"), stderr.decode("utf-8", "replace"))
        )
    if stdout.decode("utf-8", "replace").strip() != expected:
        raise AssertionError("Unexpected reconstructed release identity")


def test_complete_complete_merge():
    result = run_case(
        "complete_complete",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "E2E_DELAY_MERGE_VERSION")
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["removed"], "set_max_delay 2.0")
    assert_contains(result["removed"], "set_max_delay 5.0")
    assert_contains(result["report"], "Merged constraints              : 1")


def test_top_open_from_infers_static_startpoint():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/cfg_i"} {
        return [list u_src_reg/Q]
    }
    return {}
}
'''
    result = run_case(
        "top_open_from",
        "set_min_delay 0.2 -to [get_pins u_h0/cfg_i]\n",
        "set_min_delay 0.8 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_min_delay 1 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")


def test_legacy_top_open_from_mode_still_emits_from_when_pt_knows_startpoint():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/cfg_i"} {
        return [list u_src_reg/Q]
    }
    return {}
}

proc all_fanout {args} {
    set from [lindex $args end]
    set name [lindex $from 0]
    if {$name eq "u_h0/cfg_i"} {
        return [list u_h0/u_reg/D]
    }
    return {}
}
'''
    result = run_case(
        "top_open_from_legacy_mode_late_from",
        "set_max_delay 0.5 -to [get_pins u_h0/cfg_i]\n",
        "",
        extra_build_args=["-top_open_from_mode", "through"],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 0.5 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")


def test_harden_open_from_with_explicit_through():
    result = run_case(
        "harden_open_from",
        "set_max_delay 1.5 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 4.5 -through [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 6 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")


def test_multi_hop_review():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_up/data_o"} {
        return [list u_up/u_reg/Q]
    }
    return {}
}
'''
    result = run_case(
        "multi_hop_review",
        "set_max_delay 2.0 -from [get_pins u_up/data_o] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[("up", "u_up", "upstream")],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_up/u_reg/Q}] -through [get_pins {u_up/data_o}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO harden=u_up from=u_up/u_reg/Q to=u_up/data_o")


def test_review_top_open_from_summary_infers_startpoint():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/cfg_i"} {
        return [list u_src_reg/CP]
    }
    return {}
}
'''
    result = run_case(
        "review_top_open_from_summary",
        "set_max_delay 0.5 -to [get_pins u_h0/cfg_i]\n",
        "",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["review"], "MISSING_HARDEN_SDC_ENDPOINT_NOT_FOUND")
    assert_contains(result["report"], "REVIEW_TOP_OPEN_FROM_STARTPOINT_INFERRED")
    assert_contains(os.path.join(result["summary"], "top.csv"), "u_src_reg/CP")
    assert_not_contains(os.path.join(result["summary"], "top.csv"), '"NOT FOUND","u_h0/cfg_i","0.5"')


def test_recursive_harden_output_to_harden_input_chain():
    result = run_case(
        "recursive_harden_output_to_input",
        "set_max_delay 2.0 -from [get_pins u_up/data_o] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            (
                "up",
                "u_up",
                "upstream",
                "set_max_delay 1.0 -from [get_pins u_up/u_reg/Q] -to [get_pins u_up/data_o]\n",
            )
        ],
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "# MERGED id=E2E000001")
    assert_contains(result["out_sdc"], "set_max_delay 8 -from [get_pins {u_up/u_reg/Q}] -through [get_pins {u_up/data_o}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "RECURSIVE_MERGED")
    assert_exists(os.path.join(result["summary"], "00_index.csv"))
    assert_exists(os.path.join(result["summary"], "top.csv"))
    assert_exists(os.path.join(result["summary"], "u_up.csv"))
    assert_exists(os.path.join(result["summary"], "u_h0.csv"))
    assert_contains(os.path.join(result["summary"], "00_index.csv"), "u_up.csv")
    assert_contains(os.path.join(result["summary"], "top.csv"), "e2e_id")
    assert_contains(os.path.join(result["summary"], "top.csv"), "E2E000001")
    assert_contains(os.path.join(result["summary"], "top.csv"), "through_1")
    assert_contains(os.path.join(result["summary"], "top.csv"), "Start Point")
    assert_contains(os.path.join(result["summary"], "top.csv"), "End Point")
    assert_contains(os.path.join(result["summary"], "top.csv"), "start_sdc_delay")
    assert_contains(os.path.join(result["summary"], "top.csv"), "start_from")
    assert_contains(os.path.join(result["summary"], "top.csv"), "start_to")
    assert_contains(os.path.join(result["summary"], "top.csv"), "end_sdc_delay")
    assert_contains(os.path.join(result["summary"], "top.csv"), "end_from")
    assert_contains(os.path.join(result["summary"], "top.csv"), "end_to")
    assert_contains(os.path.join(result["summary"], "top.csv"), "stage_1_sdc_delay")
    assert_contains(os.path.join(result["summary"], "top.csv"), "stage_1_from")
    assert_contains(os.path.join(result["summary"], "top.csv"), "stage_1_to")
    assert_contains(os.path.join(result["summary"], "top.csv"), '"8","u_up/u_reg/Q","1","u_up/u_reg/Q","u_up/data_o","1","u_up/u_reg/Q","u_up/data_o","u_up/data_o","2","u_up/data_o","u_h0/cfg_i","u_h0/cfg_i","5","u_h0/cfg_i","u_h0/u_reg/D","u_h0/u_reg/D","5","u_h0/cfg_i","u_h0/u_reg/D"')
    assert_contains(os.path.join(result["summary"], "00_index.csv"), "max_delay_used")
    assert_contains(os.path.join(result["summary"], "00_index.csv"), '"top.csv","1","1","0","0","1","1","1/1","0"')
    assert_contains(os.path.join(result["summary"], "u_up.csv"), "set_max_delay 8 -from [get_pins {u_up/u_reg/Q}] -through [get_pins {u_up/data_o}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(os.path.join(result["summary"], "u_h0.csv"), "u_h0/u_reg/D")
    xlsx = os.path.join(result["case_dir"], "top.xlsx")
    proc = subprocess.Popen(
        [sys.executable, REPORT_TOOL, result["summary"]],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=result["case_dir"],
    )
    stdout, stderr = proc.communicate()
    if proc.returncode != 0:
        raise AssertionError("report failed\nstdout=%s\nstderr=%s" % (stdout.decode("utf-8", "replace"), stderr.decode("utf-8", "replace")))
    assert_exists(xlsx)
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx)
    if workbook.sheetnames != ["top", "u_h0", "u_up"]:
        raise AssertionError("Unexpected workbook sheets: %s" % workbook.sheetnames)
    ws_top = workbook["top"]
    if [ws_top["A1"].value, ws_top["B1"].value, ws_top["E1"].value, ws_top["H1"].value] != ["E2E ID\nMax Delay Used: 1/1", "Start Point", "through_1", "End Point"]:
        raise AssertionError("Unexpected report headers: %s" % [ws_top["A1"].value, ws_top["B1"].value, ws_top["E1"].value, ws_top["H1"].value])
    if ws_top["A3"].value != "E2E000001":
        raise AssertionError("Unexpected E2E ID cell: %s" % ws_top["A3"].value)
    if [ws_top["B3"].value, ws_top["C3"].value, ws_top["D3"].value] != ["u_up/u_reg/Q", "u_up/data_o", "1"]:
        raise AssertionError("Unexpected start point row: %s" % [ws_top["B3"].value, ws_top["C3"].value, ws_top["D3"].value])
    if ws_top["E3"].fill.fgColor.rgb != "00FFF2CC":
        raise AssertionError("Expected top sheet through stage to be highlighted")


def test_missing_harden_sdc_stage_assumes_zero_and_reports_not_found():
    result = run_case(
        "missing_harden_sdc_stage",
        "\n".join(
            [
                "set_max_delay 2.0 -from [get_pins u_up/data_o] -to [get_pins u_mid/in_i]",
                "set_max_delay 3.0 -from [get_pins u_mid/out_o] -to [get_pins u_h0/cfg_i]",
                "",
            ]
        ),
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            (
                "up",
                "u_up",
                "upstream",
                "set_max_delay 1.0 -from [get_pins u_up/u_reg/Q] -to [get_pins u_up/data_o]\n",
            ),
            ("mid", "u_mid", "middle"),
        ],
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 11 -from [get_pins {u_up/u_reg/Q}] -through [get_pins {u_up/data_o}] -through [get_pins {u_mid/in_i}] -through [get_pins {u_mid/out_o}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO harden=u_mid from=u_mid/in_i to=u_mid/out_o")
    assert_contains(os.path.join(result["summary"], "u_mid.csv"), "MISSING_SDC")
    assert_contains(os.path.join(result["summary"], "00_index.csv"), '"u_mid.csv","1","1","0","0","0","0","0/0","1"')

    xlsx = os.path.join(result["case_dir"], "top.xlsx")
    proc = subprocess.Popen(
        [sys.executable, REPORT_TOOL, result["summary"]],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=result["case_dir"],
    )
    stdout, stderr = proc.communicate()
    if proc.returncode != 0:
        raise AssertionError("report failed\nstdout=%s\nstderr=%s" % (stdout.decode("utf-8", "replace"), stderr.decode("utf-8", "replace")))
    assert_exists(xlsx)

    from openpyxl import load_workbook

    workbook = load_workbook(xlsx)
    ws_mid = workbook["u_mid"]
    expected_header = "E2E ID\nMax Delay Used: N/A\nNative max_delay: 0\nMissing SDC Stage: 1"
    if ws_mid["A1"].value != expected_header:
        raise AssertionError("Unexpected missing SDC usage header: %s" % ws_mid["A1"].value)
    values = [ws_mid.cell(3, col).value for col in range(1, ws_mid.max_column + 1)]
    if "NOT FOUND" not in values:
        raise AssertionError("Expected NOT FOUND in missing harden stage row: %s" % values)
    red_cells = [ws_mid.cell(3, col).coordinate for col in range(1, ws_mid.max_column + 1) if ws_mid.cell(3, col).fill.fgColor.rgb == "00F4CCCC"]
    if not red_cells:
        raise AssertionError("Expected red NOT FOUND cell in u_mid report sheet")


def test_recursive_terminal_missing_harden_sdc_uses_pt_endpoint():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_h0/u_src_reg/Q out
    u_mid/u_reg/D in
    u_mid/U26/I in
}

proc all_fanout {args} {
    set from [lindex $args end]
    set name [lindex $from 0]
    if {$name eq "u_mid/in_i"} {
        if {[lsearch -exact $args "-endpoints_only"] >= 0} {
            return [list u_mid/u_reg/D]
        }
        return [list u_mid/U26/I u_mid/u_reg/D]
    }
    return {}
}
'''
    result = run_case(
        "recursive_terminal_missing_harden_sdc_endpoint",
        "set_max_delay 2.0 -from [get_pins u_h0/o_niu_rst_n] -to [get_pins u_mid/in_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/u_src_reg/Q] -to [get_ports o_niu_rst_n]\n",
        extra_hardens=[("mid", "u_mid", "middle")],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_h0/u_src_reg/Q}] -through [get_pins {u_h0/o_niu_rst_n}] -through [get_pins {u_mid/in_i}] -to [get_pins {u_mid/u_reg/D}]")
    assert_not_contains(result["out_sdc"], "u_mid/U26/I")
    assert_not_contains(result["out_sdc"], "-to [get_pins {u_mid/in_i}]")
    assert_contains(result["report"], "Merged constraints              : 1")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO harden=u_mid from=u_mid/in_i to=u_mid/u_reg/D")
    assert_contains(os.path.join(result["summary"], "u_mid.csv"), "MISSING_SDC")


def test_recursive_missing_top_and_terminal_harden_sdc_use_pt_graph():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_h0/u_src_reg/Q out
    u_mid/u_reg/D in
}

proc all_fanout {args} {
    set from [lindex $args end]
    set name [lindex $from 0]
    if {$name eq "u_h0/o_niu_rst_n"} {
        return [list u_mid/in_i]
    }
    if {$name eq "u_mid/in_i"} {
        return [list u_mid/u_reg/D]
    }
    return {}
}
'''
    result = run_case(
        "recursive_missing_top_and_terminal_harden_sdc",
        "",
        "set_max_delay 5.0 -from [get_pins u_h0/u_src_reg/Q] -to [get_ports o_niu_rst_n]\n",
        extra_hardens=[("mid", "u_mid", "middle")],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 5 -from [get_pins {u_h0/u_src_reg/Q}] -through [get_pins {u_h0/o_niu_rst_n}] -through [get_pins {u_mid/in_i}] -to [get_pins {u_mid/u_reg/D}]")
    assert_not_contains(result["out_sdc"], "-to [get_pins {u_mid/in_i}]")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO source=top from=u_h0/o_niu_rst_n to=u_mid/in_i")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO harden=u_mid from=u_mid/in_i to=u_mid/u_reg/D")
    assert_contains(os.path.join(result["summary"], "top.csv"), "MISSING_TOP_SDC")
    assert_contains(os.path.join(result["summary"], "u_mid.csv"), "MISSING_SDC")


def test_edge_specific_review():
    result = run_case(
        "edge_specific_review",
        "set_max_delay 2.0 -rise_from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
    )
    require_ok(result)
    assert_contains(result["review"], "EDGE_SPECIFIC_OPTION")


def test_multi_object_lists_expand_and_rewrite_remaining():
    result = run_case(
        "multi_object_lists",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [list [get_pins u_h0/cfg_i] [get_pins u_h0/unused_i]]\n",
        "set_max_delay 5.0 -from [list [get_pins u_h0/cfg_i] [get_pins u_h0/other_i]] -to [get_pins u_h0/u_reg/D]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_not_contains(result["out_sdc"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}] -to [get_pins {u_h0/unused_i}]")
    assert_contains(result["report"], "Merged constraints              : 1")
    assert_contains(result["report"], "Review required constraints     : 3")
    assert_contains(result["report"], "MISSING_HARDEN_SDC_ENDPOINT_NOT_FOUND")
    assert_contains(result["removed"], "split=1/2")
    assert_contains(result["removed"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}] -to [get_pins {u_h0/cfg_i}]")
    assert_not_contains(result["removed"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}] -to [get_pins {u_h0/unused_i}]")
    assert_contains(result["final"], "STAGE2_REWRITTEN CMD000001")
    assert_contains(result["final"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}] -to [get_pins {u_h0/unused_i}]")
    assert_contains(result["final"], "STAGE2_REWRITTEN CMD000002")
    assert_contains(result["final"], "set_max_delay 5 -from [get_pins {u_h0/other_i}] -to [get_pins {u_h0/u_reg/D}]")


def test_one_top_boundary_reused_for_multiple_harden_endpoints():
    result = run_case(
        "reuse_top_boundary_for_multi_to",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [list [get_pins u_h0/u_cfg_reg/D] [get_pins u_h0/u_mode_reg/D]]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_cfg_reg/D}]")
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_mode_reg/D}]")
    assert_contains(result["report"], "Merged constraints              : 2")


def test_top_port_maps_to_connected_harden_input():
    prelude = r'''
proc current_design {} {
    return current_integration_top
}

proc sizeof_collection {coll} {
    return [llength $coll]
}

proc foreach_in_collection {var coll body} {
    upvar 1 $var item
    foreach item $coll {
        uplevel 1 $body
    }
}

proc get_ports {args} {
    set name [lindex $args end]
    if {$name eq "cfg_top"} {
        return [list cfg_top]
    }
    return {}
}

proc get_nets {args} {
    set obj [lindex $args end]
    if {$obj eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        set obj [lindex $args end]
        if {$obj eq "cfg_net"} {
            return [list u_h0/cfg_i]
        }
        return {}
    }
    set name [lindex $args end]
    if {$name in {u_src_reg/Q u_h0/cfg_i u_h0/u_reg/D}} {
        return [list $name]
    }
    return {}
}

proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "full_name"} {
        return $name
    }
    if {$attr eq "direction"} {
        if {$name eq "cfg_top"} {
            return out
        }
        if {$name eq "u_src_reg/Q"} {
            return out
        }
        if {$name in {u_h0/cfg_i u_h0/u_reg/D}} {
            return in
        }
    }
    return ""
}
'''
    result = run_case(
        "top_port_boundary_map",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_ports cfg_top]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-verbose_pt_query", "true"],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "TOP_PORT_BOUNDARY_MAP")
    assert_contains(result["report"], "Top port boundary map mode      : connectivity")
    assert_contains(result["report"], "Verbose PT query                : true")
    assert_contains(result["final"], "STAGE2_CONSUMED CMD000001")
    assert_text_contains(result["stdout"], "PT_QUERY: get_ports -quiet {cfg_top}")


def test_harden_input_to_output_boundary_merges():
    result = run_case(
        "harden_input_to_output_boundary_merge",
        "set_max_delay 1.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/i_niu_rst_n]\n",
        "set_max_delay 5.0 -from [get_ports i_niu_rst_n] -to [get_ports o_niu_rst_n]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 6 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/i_niu_rst_n}] -to [get_pins {u_h0/o_niu_rst_n}]")
    assert_contains(result["report"], "Merged constraints              : 1")
    assert_not_contains(result["report"], "OUTPUT_DIRECTION_NOT_SUPPORTED")


def test_harden_feedthrough_missing_upstream_top_uses_pt_startpoint():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg/Q out
    u_h0/i_niu_rst_n in
    u_h0/o_niu_rst_n out
    u_mid/in_i in
    u_mid/u_reg/D in
}

proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/i_niu_rst_n"} {
        return [list u_src_reg/Q]
    }
    return {}
}
'''
    result = run_case(
        "harden_feedthrough_missing_upstream_top",
        "set_max_delay 2.0 -from [get_pins u_h0/o_niu_rst_n] -to [get_pins u_mid/in_i]\n",
        "set_max_delay 5.0 -from [get_ports i_niu_rst_n] -to [get_ports o_niu_rst_n]\n",
        extra_hardens=[
            (
                "mid",
                "u_mid",
                "middle",
                "set_max_delay 4.0 -from [get_pins u_mid/in_i] -to [get_pins u_mid/u_reg/D]\n",
            )
        ],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 11 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/i_niu_rst_n}] -through [get_pins {u_h0/o_niu_rst_n}] -through [get_pins {u_mid/in_i}] -to [get_pins {u_mid/u_reg/D}]")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO source=top from=u_src_reg/Q to=u_h0/i_niu_rst_n")
    assert_contains(os.path.join(result["summary"], "top.csv"), "MISSING_TOP_SDC_u_src_reg_Q_TO_u_h0_i_niu_rst_n")


def test_harden_feedthrough_to_top_output_terminal():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg/Q out
    u_h0/i_niu_rst_n in
    u_h0/o_niu_rst_n out
    top_rst_n out
}

proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/i_niu_rst_n"} {
        return [list u_src_reg/Q]
    }
    return {}
}

proc all_fanout {args} {
    set from [lindex $args end]
    set name [lindex $from 0]
    if {$name eq "u_h0/o_niu_rst_n"} {
        return [list top_rst_n]
    }
    return {}
}
'''
    result = run_case(
        "harden_feedthrough_to_top_output_terminal",
        "",
        "set_max_delay 5.0 -from [get_ports i_niu_rst_n] -to [get_ports o_niu_rst_n]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 5 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/i_niu_rst_n}] -through [get_pins {u_h0/o_niu_rst_n}] -to [get_ports {top_rst_n}]")
    assert_contains(result["report"], "RECURSIVE_MERGED_TERMINAL")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO source=top from=u_src_reg/Q to=u_h0/i_niu_rst_n")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO source=top from=u_h0/o_niu_rst_n to=top_rst_n")
    assert_not_contains(result["review"], "NO_TOP_SEGMENT_MATCHED")


def main():
    if os.path.isdir(WORK):
        shutil.rmtree(WORK)
    os.makedirs(WORK)
    tests = [
        test_release_identity_is_reconstructed_without_plaintext_constant,
        test_complete_complete_merge,
        test_top_open_from_infers_static_startpoint,
        test_legacy_top_open_from_mode_still_emits_from_when_pt_knows_startpoint,
        test_harden_open_from_with_explicit_through,
        test_multi_hop_review,
        test_review_top_open_from_summary_infers_startpoint,
        test_recursive_harden_output_to_harden_input_chain,
        test_missing_harden_sdc_stage_assumes_zero_and_reports_not_found,
        test_recursive_terminal_missing_harden_sdc_uses_pt_endpoint,
        test_recursive_missing_top_and_terminal_harden_sdc_use_pt_graph,
        test_edge_specific_review,
        test_multi_object_lists_expand_and_rewrite_remaining,
        test_one_top_boundary_reused_for_multiple_harden_endpoints,
        test_top_port_maps_to_connected_harden_input,
        test_harden_input_to_output_boundary_merges,
        test_harden_feedthrough_missing_upstream_top_uses_pt_startpoint,
        test_harden_feedthrough_to_top_output_terminal,
    ]
    for test in tests:
        test()
        print("PASS", test.__name__)
    shutil.rmtree(WORK)
    print("All Stage 2 regression tests passed.")


if __name__ == "__main__":
    main()
