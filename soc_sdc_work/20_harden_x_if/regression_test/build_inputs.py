#!/usr/bin/env python3
"""Build minimal 20 harden interface regression inputs in the current directory."""

from pathlib import Path

from openpyxl import Workbook


def main() -> None:
    root = Path.cwd()

    wb = Workbook()
    ws = wb.active
    ws.title = "info"
    ws.append(["inst_name", "module_name", "sdc_file"])
    ws.append(["u_a", "harden_a", "u_a.sdc"])
    ws.append(["u_b", "harden_b", "u_b.sdc"])
    wb.save(root / "info_all.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "u_a"
    ws.append(["Input", "From Whom", "Output", "To Top"])
    ws.append(["clk_i", "top.sys_clk", "data_o", "fabric_bus"])
    ws = wb.create_sheet("u_b")
    ws.append(["Input", "From Whom", "Output", "To Top"])
    ws.append(["data_i", "u_a/data_o", "", ""])
    wb.save(root / "ports.xlsx")

    (root / "u_a.sdc").write_text(
        "set_output_delay -max 1.5 -min -0.1 -clock [get_clocks {clk_a}] [get_ports {data_o}]\n",
        encoding="utf-8",
    )
    (root / "u_b.sdc").write_text(
        "set_input_delay -max 1.2 -clock [get_clocks {clk_b}] [get_ports {data_i}]\n",
        encoding="utf-8",
    )
    (root / "clock_inventory.csv").write_text(
        "clock_name,direct_source,producer_object,final_action\n"
        "clk,[get_ports {sys_clk}],,emit_top_clock\n",
        encoding="utf-8",
    )
    inv = root / "00_harden_port_inventory"
    (inv / "pending").mkdir(parents=True, exist_ok=True)
    (inv / "pending" / "u_a.ports").write_text(
        "input clk_i\n"
        "output data_o\n",
        encoding="utf-8",
    )
    (inv / "pending" / "u_b.ports").write_text(
        "input data_i\n",
        encoding="utf-8",
    )
    (inv / "connection_inventory.csv").write_text(
        "connection_id,connection_type,src_instance,src_direction,src_port,src_bit_index,src_endpoint_key,src_soc_object,"
        "dst_instance,dst_direction,dst_port,dst_bit_index,dst_endpoint_key,dst_soc_object,validation_status,note\n"
        "CONN_u_a_data_o__u_b_data_i,harden_to_harden,u_a,output,data_o,,u_a:output:data_o,u_a/data_o,"
        "u_b,input,data_i,,u_b:input:data_i,u_b/data_i,matched,\n"
        "CONN_u_a_data_o__fabric_bus,harden_to_fabric,u_a,output,data_o,,u_a:output:data_o,u_a/data_o,"
        "fabric,,fabric_bus,,fabric::fabric_bus,,matched,\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
