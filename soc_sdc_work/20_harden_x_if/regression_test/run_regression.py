#!/usr/bin/env python3
"""Complex regression for 20_extract_harden_x_if.py."""

from __future__ import print_function

import csv
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parent
SOC = BASE.parent.parent
EX20 = SOC / "20_harden_x_if" / "20_extract_harden_x_if.py"
WORK = BASE / "work_complex"
TARGET_FT_CHANNEL = "CH_u_a_data_o_bit0__u_c_data_i_bit0"

REQ = [
    "Input",
    "Input Width",
    "Input Used Width",
    "From Whom",
    "Output",
    "Output Width",
    "Output Used Width",
    "To Top",
    "Inout",
    "Inout Width",
    "Inout Connectivity",
    "Inout Name",
]


def clean_dir(path):
    if path.exists():
        shutil.rmtree(str(path))
    path.mkdir(parents=True)


def sh(args, cwd):
    return subprocess.run(
        [sys.executable] + [str(arg) for arg in args],
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def append_port_row(ws, **values):
    ws.append([values.get(col, "") for col in REQ])


def write_info(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "info"
    ws.append(["inst_name", "module_name", "owner", "sdc_path"])
    ws.append(["u_a", "harden_a", "alice", "u_a.sdc"])
    ws.append(["u_b", "harden_b", "bob", "u_b.sdc"])
    ws.append(["u_c", "harden_c", "carol", "u_c.sdc"])
    wb.save(str(path))


def write_ports(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "u_a"
    ws.append(REQ)
    append_port_row(ws, Output="data_o", **{"Output Width": 2})
    ws = wb.create_sheet("u_b")
    ws.append(REQ)
    append_port_row(ws, Input="fti_0_a2c_data", **{"Input Width": 2}, Output="fto_0_a2c_data", **{"Output Width": 2})
    ws = wb.create_sheet("u_c")
    ws.append(REQ)
    append_port_row(ws, Input="data_i", **{"Input Width": 2})
    wb.save(str(path))


def write_connection_inventory(path, noncanonical=False):
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "connection_id",
        "connection_type",
        "src_instance",
        "src_direction",
        "src_port",
        "src_bit_index",
        "src_endpoint_key",
        "src_soc_object",
        "dst_instance",
        "dst_direction",
        "dst_port",
        "dst_bit_index",
        "dst_endpoint_key",
        "dst_soc_object",
        "validation_status",
        "note",
    ]
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers)
        writer.writeheader()

        def add(cid, src_i, src_d, src_p, dst_i, dst_d, dst_p):
            src_bit = src_p.split("[")[-1].rstrip("]") if "[" in src_p else ""
            dst_bit = dst_p.split("[")[-1].rstrip("]") if "[" in dst_p else ""
            writer.writerow({
                "connection_id": cid,
                "connection_type": "harden_to_harden",
                "src_instance": src_i,
                "src_direction": src_d,
                "src_port": src_p,
                "src_bit_index": src_bit,
                "src_endpoint_key": "%s:%s:%s" % (src_i, src_d, src_p),
                "src_soc_object": "%s/%s" % (src_i, src_p),
                "dst_instance": dst_i,
                "dst_direction": dst_d,
                "dst_port": dst_p,
                "dst_bit_index": dst_bit,
                "dst_endpoint_key": "%s:%s:%s" % (dst_i, dst_d, dst_p),
                "dst_soc_object": "%s/%s" % (dst_i, dst_p),
                "validation_status": "matched",
                "note": "",
            })

        if noncanonical:
            add("CONN_BAD_RANGE", "u_a", "output", "data_o[1:0]", "u_c", "input", "data_i[1:0]")
        else:
            add("CONN_u_a_data_o_bit0__u_b_fti_bit0", "u_a", "output", "data_o[0]", "u_b", "input", "fti_0_a2c_data[0]")
            add("CONN_u_b_fto_bit0__u_c_data_i_bit0", "u_b", "output", "fto_0_a2c_data[0]", "u_c", "input", "data_i[0]")


def write_feedthrough_inventory(path):
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=[
            "feedthrough_id",
            "scenario",
            "feedthrough_instance",
            "feedthrough_module",
            "hop_index",
            "base",
            "src_name",
            "dst_name",
            "signal_name",
            "fti_port",
            "fto_port",
            "fti_endpoint",
            "fto_endpoint",
            "bit_index",
            "chain_id",
            "hop_order",
            "upstream_endpoint",
            "downstream_endpoint",
            "validation_status",
            "basis",
            "note",
        ])
        writer.writeheader()
        writer.writerow({
            "feedthrough_id": "FT_u_b_0_a2c_data_bit0",
            "scenario": "common",
            "feedthrough_instance": "u_b",
            "feedthrough_module": "harden_b",
            "hop_index": "0",
            "base": "a2c_data",
            "src_name": "a",
            "dst_name": "c",
            "signal_name": "data",
            "fti_port": "fti_0_a2c_data[0]",
            "fto_port": "fto_0_a2c_data[0]",
            "fti_endpoint": "[get_pins {u_b/fti_0_a2c_data[0]}]",
            "fto_endpoint": "[get_pins {u_b/fto_0_a2c_data[0]}]",
            "bit_index": "0",
            "chain_id": "CHAIN_a2c_data",
            "hop_order": "0",
            "upstream_endpoint": "[get_pins {u_a/data_o[0]}]",
            "downstream_endpoint": "[get_pins {u_c/data_i[0]}]",
            "validation_status": "matched",
            "basis": "paired_by_fti_fto_name",
            "note": "",
        })


def write_sdc_files(d):
    (d / "u_a.sdc").write_text(
        "set_output_delay -max 2.0 -clock [get_clocks {clk_a}] [get_ports {data_o[0]}]\n",
        encoding="utf-8",
    )
    (d / "u_b.sdc").write_text("", encoding="utf-8")
    (d / "u_c.sdc").write_text(
        "set_input_delay -max 1.5 -clock [get_clocks {clk_c}] [get_ports {data_i[0]}]\n",
        encoding="utf-8",
    )
    (d / "clock_inventory.csv").write_text(
        "clock_name,direct_source,producer_object,final_action\n",
        encoding="utf-8",
    )


def write_pending(d):
    pending = d / "00_harden_port_inventory" / "pending"
    pending.mkdir(parents=True, exist_ok=True)
    (pending / "u_a.ports").write_text("output data_o[0]\noutput data_o[1]\n", encoding="utf-8")
    (pending / "u_b.ports").write_text(
        "input fti_0_a2c_data[0]\noutput fto_0_a2c_data[0]\n",
        encoding="utf-8",
    )
    (pending / "u_c.ports").write_text("input data_i[0]\ninput data_i[1]\n", encoding="utf-8")


def build_positive(d):
    clean_dir(d)
    write_info(d / "info_all.xlsx")
    write_ports(d / "ports_harden.xlsx")
    write_connection_inventory(d / "00_harden_port_inventory" / "connection_inventory.csv")
    write_feedthrough_inventory(d / "feedthrough_inventory.csv")
    write_sdc_files(d)
    write_pending(d)


def header_map(ws):
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def approve_channel(form, channel_id):
    wb = load_workbook(str(form))
    ws = wb["interface_budget"]
    col = header_map(ws)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, col["channel_id"]).value != channel_id:
            continue
        updates = {
            "timing_model": "lib_blackbox",
            "budget_required": "yes",
            "budget_model": "interconnect_budget",
            "converted_max": "",
            "max_source": "",
            "derivation_basis": "",
            "tool_surface": "sta",
            "datapath_only": "yes",
            "budget_basis": "interconnect budget from block owners",
            "apply": "yes",
            "emit_max": "yes",
            "emit_min": "no",
            "review_status": "approved",
        }
        for key, value in updates.items():
            ws.cell(row, col[key], value)
        wb.save(str(form))
        return
    raise AssertionError("target channel not found: " + channel_id)


def run_feedthrough_bit_case():
    d = WORK / "feedthrough_bit"
    build_positive(d)
    first = sh([EX20, "-input", "clock_inventory.csv"], d)
    require(first.returncode == 1, "first 20 run should sync workbook")
    approve_channel(d / "20_harden_x_if.xlsx", TARGET_FT_CHANNEL)
    second = sh([EX20, "-input", "clock_inventory.csv"], d)
    require(second.returncode == 0, "20 feedthrough bit generation failed:\n%s\n%s" % (second.stdout, second.stderr))

    sdc = (d / "common" / "20_harden_x_if.sdc").read_text(encoding="utf-8")
    report = (d / "harden_x_if_check_report_common_all_all.txt").read_text(encoding="utf-8")
    pending_a = (d / "00_harden_port_inventory" / "pending" / "u_a.ports").read_text(encoding="utf-8")
    pending_b = (d / "00_harden_port_inventory" / "pending" / "u_b.ports").read_text(encoding="utf-8")
    pending_c = (d / "00_harden_port_inventory" / "pending" / "u_c.ports").read_text(encoding="utf-8")
    removed = (d / "00_harden_port_inventory" / "removed_log" / "20_harden_x_if.removed").read_text(encoding="utf-8")

    require("set_max_delay 1.5 -datapath_only -from [get_pins {u_a/data_o[0]}] -to [get_pins {u_c/data_i[0]}]" in sdc, "stitched bit SDC missing")
    require("auto-resolved converted_max=1.5" in report, "auto-resolve missing for stitched channel")
    require("removed 2 harden interface endpoint(s) from pending" in report, "pending removal summary missing")
    require("output data_o[0]" not in pending_a and "output data_o[1]" in pending_a, "source bit pending removal wrong")
    require("input data_i[0]" not in pending_c and "input data_i[1]" in pending_c, "destination bit pending removal wrong")
    require("fti_0_a2c_data[0]" in pending_b and "fto_0_a2c_data[0]" in pending_b, "20 must not consume feedthrough ports")
    require("u_a output data_o[0] covered_by=20_harden_x_if" in removed, "removed log missing source bit")
    require("u_c input data_i[0] covered_by=20_harden_x_if" in removed, "removed log missing destination bit")

    rerun = sh([EX20, "-input", "clock_inventory.csv"], d)
    require(rerun.returncode == 0, "20 pending update should be idempotent:\n%s\n%s" % (rerun.stdout, rerun.stderr))
    rerun_report = (d / "harden_x_if_check_report_common_all_all.txt").read_text(encoding="utf-8")
    require("no previous_removed record exists" not in rerun_report, "20 pending update is not idempotent")


def run_noncanonical_connection_case():
    d = WORK / "noncanonical_connection"
    clean_dir(d)
    write_info(d / "info_all.xlsx")
    write_ports(d / "ports_harden.xlsx")
    write_connection_inventory(d / "00_harden_port_inventory" / "connection_inventory.csv", noncanonical=True)
    write_sdc_files(d)
    result = sh([EX20, "-input", "clock_inventory.csv", "--no-update-pending"], d)
    require(result.returncode == 1, "noncanonical connection_inventory should fail")
    report = (d / "harden_x_if_check_report_common_all_all.txt").read_text(encoding="utf-8")
    require("not a canonical scalar/bit key" in report, "noncanonical connection error missing")


def main():
    clean_dir(WORK)
    run_feedthrough_bit_case()
    run_noncanonical_connection_case()
    print("20 harden_x_if complex regression: PASS")
    print("  cases: feedthrough_bit, noncanonical_connection")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
