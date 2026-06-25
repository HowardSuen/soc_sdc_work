#!/usr/bin/env python3
"""
Generate 30 harden-to-harden path exception SDC from bit-level integration
inventories and a reviewed exception workbook.

Current scope:
  * 00 connection_inventory.csv is the source of direct bit-to-bit pairing
  * 10 feedthrough_inventory.csv supplies bit-level feedthrough IDs
  * 20_harden_x_if.xlsx is read when present to detect active normal budgets
  * first run synchronizes candidates/rules, then stops for review
  * only apply=yes + review_status=approved rows are emitted
"""

import argparse
import csv
import hashlib
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

try:
    from dataclasses import dataclass, field
except ImportError:  # pragma: no cover - Python 3.6 compatibility path
    class _CompatField:
        def __init__(self, default_factory=None):
            self.default_factory = default_factory

    def field(default_factory=None):
        return _CompatField(default_factory=default_factory)

    def dataclass(_cls=None, frozen=False):
        def wrap(cls):
            annotations = getattr(cls, "__annotations__", {})
            names = list(annotations.keys())
            defaults = {}
            factories = {}
            for name in names:
                if hasattr(cls, name):
                    value = getattr(cls, name)
                    if isinstance(value, _CompatField):
                        factories[name] = value.default_factory
                        delattr(cls, name)
                    else:
                        defaults[name] = value

            def __init__(self, *args, **kwargs):
                if len(args) > len(names):
                    raise TypeError("too many positional arguments")
                values = {}
                for name, value in zip(names, args):
                    values[name] = value
                for name in names[len(args):]:
                    if name in kwargs:
                        values[name] = kwargs.pop(name)
                    elif name in factories:
                        values[name] = factories[name]()
                    elif name in defaults:
                        values[name] = defaults[name]
                    else:
                        raise TypeError("missing required argument: " + name)
                if kwargs:
                    raise TypeError("unexpected argument: " + sorted(kwargs)[0])
                for name in names:
                    object.__setattr__(self, name, values[name])

            def __eq__(self, other):
                if other.__class__ is not cls:
                    return False
                return all(getattr(self, name) == getattr(other, name) for name in names)

            cls.__init__ = __init__
            cls.__eq__ = __eq__
            if frozen:
                def __hash__(self):
                    return hash(tuple(getattr(self, name) for name in names))
                cls.__hash__ = __hash__
            else:
                cls.__hash__ = None
            return cls

        if _cls is None:
            return wrap
        return wrap(_cls)

try:
    import pandas as pd
except ImportError as exc:  # pragma: no cover - user environment guard
    print("ERROR: pandas is required to read integration xlsx files.", file=sys.stderr)
    raise SystemExit(2) from exc

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - user environment guard
    print("ERROR: openpyxl is required to read/write 30 exception xlsx files.", file=sys.stderr)
    raise SystemExit(2) from exc


SCENARIOS = {"common", "func", "scan", "mbist", "gpio_in", "gpio_out"}
STAGES = {"all", "synth", "prects", "postcts", "postroute"}
MATCHED_STATUSES = {"", "matched", "ok", "valid"}
YES_NO = {"", "yes", "no"}
REVIEW_STATUS_VALUES = {"", "pending", "approved", "rejected", "needs_review"}
EXCEPTION_TYPES = {
    "false_path",
    "multicycle_path",
    "max_delay_override",
    "min_delay_override",
    "max_min_delay_override",
    "needs_review",
    "not_exception",
}
EMITTED_EXCEPTION_TYPES = {
    "false_path",
    "multicycle_path",
    "max_delay_override",
    "min_delay_override",
    "max_min_delay_override",
}
PATH_CATEGORIES = {
    "",
    "data",
    "control",
    "config",
    "status",
    "reset",
    "test",
    "debug",
    "handshake",
    "interrupt",
    "cdc_sync",
    "rdc_sync",
    "static",
    "unknown",
}
TIMING_CONTRACT_STATUSES = {
    "",
    "both_sides_timed",
    "src_timed_only",
    "dst_timed_only",
    "no_port_timing",
    "clock_relative_only",
    "interconnect_budget",
    "unknown",
}
SOURCE_TYPES = {
    "",
    "extracted_harden_exception",
    "missing_timing_candidate",
    "integration_tag",
    "clock_relation_candidate",
    "manual_entry",
    "from_20_exception_path",
}
CLOCK_RELATION_CANONICAL = {"synchronous", "asynchronous", "logically_exclusive", "physically_exclusive", "unknown"}
CLOCK_RELATION_ALIASES = {
    "": "",
    "sync": "synchronous",
    "synchronous": "synchronous",
    "async": "asynchronous",
    "asynchronous": "asynchronous",
    "logically_exclusive": "logically_exclusive",
    "logically exclusive": "logically_exclusive",
    "physically_exclusive": "physically_exclusive",
    "physically exclusive": "physically_exclusive",
    "unknown": "unknown",
}
ASYNC_RELATIONS = {"asynchronous", "logically_exclusive", "physically_exclusive"}
CHECK_TYPES = {"", "both", "setup", "hold"}
MCP_REFERENCES = {"", "start", "end", "same_clock_default"}
CROSS_CLOCK_MCP_REVIEW = {"", "approved", "not_applicable", "pending", "reviewed", "yes"}
DATAPATH_VALUES = {"", "yes", "no"}
TOOL_SURFACES = {"", "sta", "dc", "both"}
CLOCK_CONTEXT_STATUSES = {"", "matched", "remapped_equivalent", "mismatch", "unknown", "not_applicable"}
RISK_LEVELS = {"", "low", "medium", "high"}
CDC_WINDOW_CATEGORIES = {"handshake", "cdc_sync", "status", "control", "data"}
MODE_SPECIFIC_TOKENS = ("scan", "mbist", "bist", "test", "gpio", "jtag", "bypass", "lp_", "low_power")
RESET_BASIS_TOKENS = (
    "recovery",
    "removal",
    "rdc",
    "waiver",
    "sync",
    "synchronizer",
    "reset synchronizer",
    "reset_sync",
    "恢复",
    "移除",
    "恢复移除",
    "复位同步",
    "复位同步器",
    "同步器",
)
PORT_BIT_RE = re.compile(r"^[^\s\[\]]+(?:\[\d+\])?$")
PORT_RANGE_RE = re.compile(r"^(.+)\[(\d+)\s*:\s*(\d+)\]$")
PORT_EXACT_BIT_RE = re.compile(r"^(.+)\[(\d+)\]$")

EXCEPTION_CANDIDATE_HEADERS = [
    "candidate_id",
    "scenario",
    "stage",
    "corner",
    "channel_id",
    "source_type",
    "path_category",
    "timing_contract_status",
    "src_instance",
    "src_module",
    "src_port",
    "src_bit_index",
    "src_endpoint",
    "dst_instance",
    "dst_module",
    "dst_port",
    "dst_bit_index",
    "dst_endpoint",
    "src_clock",
    "dst_clock",
    "clock_relation",
    "has_src_output_delay",
    "has_dst_input_delay",
    "related_20_channel_id",
    "related_20_status",
    "related_10_feedthrough_id",
    "harden_clock_context_status",
    "source_sdc_file",
    "source_line",
    "source_command",
    "source_digest",
    "extraction_time",
    "candidate_reason",
    "recommended_action",
    "note",
]

EXCEPTION_RULE_HEADERS = [
    "exception_id",
    "scenario",
    "stage",
    "corner",
    "apply",
    "review_status",
    "owner",
    "exception_type",
    "path_category",
    "channel_id",
    "related_20_channel_id",
    "related_10_feedthrough_id",
    "src_bit_index",
    "src_endpoint",
    "dst_bit_index",
    "dst_endpoint",
    "from_collection",
    "to_collection",
    "through_collection",
    "src_clock",
    "dst_clock",
    "clock_relation",
    "timing_contract_status",
    "harden_clock_context_status",
    "check_type",
    "max_value",
    "min_value",
    "setup_cycles",
    "hold_cycles",
    "mcp_reference",
    "cross_clock_mcp_review",
    "datapath_only",
    "tool_surface",
    "case_condition",
    "source_type",
    "source_sdc_file",
    "source_line",
    "source_command",
    "source_digest",
    "cdc_rdc_ref",
    "sta_waiver_ref",
    "protocol_ref",
    "basis",
    "risk_level",
    "expiry_or_review_date",
    "note",
]

EXTRACTION_LOG_HEADERS = [
    "source_sdc_file",
    "source_line",
    "instance",
    "command_type",
    "mapped_channel_ids",
    "from_collection",
    "to_collection",
    "through_collection",
    "source_digest",
    "extraction_time",
    "source_command",
    "message",
]

HEADER_FILL = PatternFill("solid", fgColor="674EA7")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C6CC"),
    right=Side(style="thin", color="B8C6CC"),
    top=Side(style="thin", color="B8C6CC"),
    bottom=Side(style="thin", color="B8C6CC"),
)


@dataclass
class PortInfo:
    name: str
    width: str = ""
    from_whom: str = ""
    to_top: str = ""


@dataclass
class InstInfo:
    module_name: str
    inst_name: str
    owner: str = ""
    file_path: str = ""
    sdc_hint: str = ""
    sdc_path: Optional[Path] = None
    inputs: Dict[str, PortInfo] = field(default_factory=dict)
    outputs: Dict[str, PortInfo] = field(default_factory=dict)
    inouts: Dict[str, PortInfo] = field(default_factory=dict)


@dataclass
class TclCommand:
    raw: str
    line_no: int


@dataclass(frozen=True)
class PortKey:
    inst_name: str
    direction: str
    port_name: str


@dataclass
class ConnectionEdge:
    connection_id: str
    connection_type: str
    src_instance: str
    src_direction: str
    src_port: str
    src_bit_index: str = ""
    src_endpoint_key: str = ""
    src_soc_object: str = ""
    dst_instance: str = ""
    dst_direction: str = ""
    dst_port: str = ""
    dst_bit_index: str = ""
    dst_endpoint_key: str = ""
    dst_soc_object: str = ""
    validation_status: str = ""
    note: str = ""


@dataclass
class ConnectionIndex:
    edges: List[ConnectionEdge] = field(default_factory=list)
    by_dst: Dict[Tuple[str, str], List[ConnectionEdge]] = field(default_factory=lambda: defaultdict(list))
    by_src: Dict[Tuple[str, str], List[ConnectionEdge]] = field(default_factory=lambda: defaultdict(list))


@dataclass
class FeedthroughRecord:
    feedthrough_id: str
    scenario: str
    feedthrough_instance: str
    hop_index: str
    base: str
    fti_port: str
    fto_port: str
    fti_endpoint: str
    fto_endpoint: str
    bit_index: str
    chain_id: str
    hop_order: str
    validation_status: str


@dataclass
class ChannelRecord:
    channel_id: str
    scenario: str
    stage: str
    corner: str
    channel_type: str
    connection_id: str
    src_instance: str
    src_module: str
    src_direction: str
    src_port: str
    src_bit_index: str
    src_endpoint: str
    dst_instance: str
    dst_module: str
    dst_direction: str
    dst_port: str
    dst_bit_index: str
    dst_endpoint: str
    related_10_feedthrough_id: str = ""
    through_collection: str = ""
    is_pad_related: str = "no"
    is_clock_related: str = "no"
    is_feedthrough_path: str = "no"
    note: str = ""


@dataclass
class DelayEvidence:
    inst_name: str
    port_name: str
    direction: str
    constraint_type: str
    clock_name: str


@dataclass
class ExceptionEvidence:
    inst_name: str
    module_name: str
    owner: str
    command_type: str
    exception_type: str
    from_ports: List[str]
    to_ports: List[str]
    from_collection: str
    to_collection: str
    through_collection: str
    check_type: str
    max_value: str
    min_value: str
    setup_cycles: str
    hold_cycles: str
    mcp_reference: str
    source_sdc_file: str
    source_line: str
    source_command: str
    source_digest: str
    extraction_time: str
    message: str = ""


@dataclass
class Active20Budget:
    channel_id: str
    max_active: bool = False
    min_active: bool = False
    rows: List[str] = field(default_factory=list)


@dataclass
class CandidateSeed:
    values: Dict[str, str]
    related_evidence: List[ExceptionEvidence] = field(default_factory=list)


@dataclass
class RuleSeed:
    values: Dict[str, str]


@dataclass
class FormRow:
    row_idx: int
    values: Dict[str, object]


class Report:
    def __init__(self) -> None:
        self.lines: List[str] = []
        self.warning_count = 0
        self.error_count = 0
        self.sync_changed = False

    def info(self, msg: str) -> None:
        self.lines.append("INFO: " + msg)

    def warn(self, msg: str) -> None:
        self.warning_count += 1
        self.lines.append("WARNING: " + msg)

    def error(self, msg: str) -> None:
        self.error_count += 1
        self.lines.append("ERROR: " + msg)


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def normalize_key(value) -> str:
    return clean_cell(value).strip().lower()


def canonical_clock_relation(value) -> str:
    text = normalize_key(value)
    if not text:
        return ""
    text = re.sub(r"[\s-]+", " ", text).replace("_", " ").strip()
    if text in CLOCK_RELATION_ALIASES:
        return CLOCK_RELATION_ALIASES[text]
    underscored = text.replace(" ", "_")
    if underscored in CLOCK_RELATION_CANONICAL:
        return underscored
    return ""


def normalize_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())


def column_map(df: pd.DataFrame) -> Dict[str, str]:
    return {normalize_col(col): col for col in df.columns}


def get_col(df: pd.DataFrame, aliases: Sequence[str]) -> Optional[str]:
    cmap = column_map(df)
    for alias in aliases:
        key = normalize_col(alias)
        if key in cmap:
            return cmap[key]
    return None


def read_excel_file(path: Path, sheet_name=0):
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except Exception as exc:
        raise RuntimeError("failed to read %s: %s" % (path, exc))


def read_text(path: Path) -> str:
    last_error: Optional[Exception] = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    assert last_error is not None
    raise last_error


def digest_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def safe_filename_token(value: str) -> str:
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-")
    token = "".join(char if char in allowed else "_" for char in clean_cell(value))
    return token or "unknown"


def sanitize_id(value: str) -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", clean_cell(value)).strip("_")
    return token or "unknown"


def format_number(value) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    return "%.12g" % number


def parse_number(value) -> Optional[float]:
    text = clean_cell(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def join_unique(values: Iterable[str], sep: str = "; ") -> str:
    result: List[str] = []
    for value in values:
        text = clean_cell(value)
        if text and text not in result:
            result.append(text)
    return sep.join(result)


def brace_list(names: Sequence[str]) -> str:
    return "{" + " ".join(clean_cell(name) for name in names if clean_cell(name)) + "}"


def get_collection(kind: str, objects: Sequence[str]) -> str:
    return "[%s %s]" % (kind, brace_list(objects))


def split_port_key(port: str) -> Tuple[str, str, bool, bool]:
    port = clean_cell(port)
    range_match = PORT_RANGE_RE.fullmatch(port)
    if range_match:
        return range_match.group(1), "", True, False
    bit_match = PORT_EXACT_BIT_RE.fullmatch(port)
    if bit_match:
        return bit_match.group(1), bit_match.group(2), False, False
    if "[" in port or "]" in port:
        return port, "", False, True
    return port, "", False, False


def port_base(port: str) -> str:
    base, _, _, _ = split_port_key(port)
    return base


def inferred_bit_index(port: str, explicit: str = "") -> str:
    if clean_cell(explicit):
        return clean_cell(explicit)
    _, bit_index, _, _ = split_port_key(port)
    return bit_index


def is_canonical_port_key(port: str) -> bool:
    return bool(PORT_BIT_RE.fullmatch(clean_cell(port)))


def is_feedthrough_port(port: str) -> bool:
    return port_base(port).startswith(("fti_", "fto_"))


def build_channel_id(src_inst: str, src_port: str, dst_inst: str, dst_port: str) -> str:
    def port_token(port: str) -> str:
        return re.sub(r"\[(\d+)\]", r"_bit\1", clean_cell(port))

    return "CH_%s__%s" % (
        sanitize_id(src_inst + "_" + port_token(src_port)),
        sanitize_id(dst_inst + "_" + port_token(dst_port)),
    )


def parse_endpoint_key(value: str) -> Tuple[str, str, str]:
    value = clean_cell(value)
    if not value:
        return "", "", ""
    parts = value.split(":", 2)
    if len(parts) == 3:
        return parts[0], parts[1], parts[2]
    return "", "", ""


def endpoint_from_soc_object(inst_name: str, port_name: str, soc_object: str, endpoint_key: str) -> str:
    obj = clean_cell(soc_object)
    inst = clean_cell(inst_name)
    port = clean_cell(port_name)
    if obj:
        if obj.startswith("["):
            return obj
        if "/" in obj:
            return get_collection("get_pins", [obj])
        if normalize_key(inst) == "top":
            return get_collection("get_ports", [obj])
        if normalize_key(inst) in {"fabric", "unknown", "constant", "const"}:
            return obj
        return get_collection("get_pins", [inst + "/" + obj])
    key_inst, _, key_port = parse_endpoint_key(endpoint_key)
    inst = inst or key_inst
    port = port or key_port
    if not inst or not port:
        return ""
    if normalize_key(inst) == "top":
        return get_collection("get_ports", [port])
    if normalize_key(inst) in {"fabric", "unknown", "constant", "const"}:
        return ""
    return get_collection("get_pins", [inst + "/" + port])


def read_info_all(path: Path, report: Report) -> Dict[str, InstInfo]:
    df = read_excel_file(path)
    module_col = get_col(df, ["module_name", "module name", "module"])
    inst_col = get_col(df, ["inst_name", "inst name", "instance", "instance_name"])
    owner_col = get_col(df, ["owner"])
    file_col = get_col(df, ["file_path", "file path", "empty_path", "verilog", "v_path"])
    sdc_col = get_col(df, ["sdc_path", "sdc file", "sdc_file", "sdc"])
    if not inst_col:
        raise RuntimeError("%s must contain an inst_name column" % path)

    instances: Dict[str, InstInfo] = {}
    for row_idx, row in df.iterrows():
        inst_name = clean_cell(row.get(inst_col))
        if not inst_name:
            continue
        module_name = clean_cell(row.get(module_col)) if module_col else ""
        file_path = clean_cell(row.get(file_col)) if file_col else ""
        if not module_name and file_path:
            module_name = Path(file_path).stem.replace("_empty", "")
        if not module_name:
            module_name = inst_name
            report.warn("%s row %d: module_name is empty; using inst_name" % (path.name, row_idx + 2))
        if inst_name in instances:
            report.warn("duplicate inst_name %s in %s; keeping first row" % (inst_name, path.name))
            continue
        instances[inst_name] = InstInfo(
            module_name=module_name,
            inst_name=inst_name,
            owner=clean_cell(row.get(owner_col)) if owner_col else "",
            file_path=file_path,
            sdc_hint=clean_cell(row.get(sdc_col)) if sdc_col else "",
        )
    report.info("loaded %d instance(s) from %s" % (len(instances), path.name))
    return instances


def parse_port_sheet(df: pd.DataFrame) -> Dict[str, Dict[str, PortInfo]]:
    input_col = get_col(df, ["Input"])
    input_width_col = get_col(df, ["Input Width"])
    from_col = get_col(df, ["From Whom"])
    output_col = get_col(df, ["Output"])
    output_width_col = get_col(df, ["Output Width"])
    to_top_col = get_col(df, ["To Top", "To Whom", "To"])
    inout_col = get_col(df, ["Inout"])
    inout_width_col = get_col(df, ["Inout Width"])

    inputs: Dict[str, PortInfo] = {}
    outputs: Dict[str, PortInfo] = {}
    inouts: Dict[str, PortInfo] = {}

    for _, row in df.iterrows():
        if input_col:
            name = clean_cell(row.get(input_col))
            if name:
                inputs[name] = PortInfo(name, clean_cell(row.get(input_width_col)) if input_width_col else "", clean_cell(row.get(from_col)) if from_col else "", "")
        if output_col:
            name = clean_cell(row.get(output_col))
            if name:
                outputs[name] = PortInfo(name, clean_cell(row.get(output_width_col)) if output_width_col else "", "", clean_cell(row.get(to_top_col)) if to_top_col else "")
        if inout_col:
            name = clean_cell(row.get(inout_col))
            if name:
                inouts[name] = PortInfo(name, clean_cell(row.get(inout_width_col)) if inout_width_col else "", "", "")
    return {"inputs": inputs, "outputs": outputs, "inouts": inouts}


def default_port_workbooks(cwd: Path, info_name: str, form_name: str, report: Report) -> List[Path]:
    excluded = {info_name, form_name}
    candidates: List[Path] = []
    skipped: List[str] = []
    for path in sorted(cwd.glob("*.xlsx")):
        if path.name in excluded or path.name.startswith("~$"):
            continue
        if path.name == "ports.xlsx" or path.name.startswith(("port_", "ports_")):
            candidates.append(path)
        else:
            skipped.append(path.name)
    if skipped:
        report.warn("ignored non-port workbook(s) in 30 input directory: " + ", ".join(skipped[:10]))
    return candidates


def read_port_workbooks(paths: Sequence[Path], report: Report) -> Dict[str, Dict[str, Dict[str, PortInfo]]]:
    sheets: Dict[str, Dict[str, Dict[str, PortInfo]]] = {}
    for path in paths:
        try:
            book = pd.ExcelFile(path)
        except Exception as exc:
            report.error("failed to open port workbook %s: %s" % (path.name, exc))
            continue
        for sheet_name in book.sheet_names:
            if sheet_name in sheets:
                report.warn("duplicate port sheet %s; keeping first occurrence" % sheet_name)
                continue
            try:
                sheets[sheet_name] = parse_port_sheet(read_excel_file(path, sheet_name))
            except Exception as exc:
                report.error("failed to read %s:%s: %s" % (path.name, sheet_name, exc))
    report.info("loaded %d instance port sheet(s) from %d workbook(s)" % (len(sheets), len(paths)))
    return sheets


def attach_port_data(instances: Dict[str, InstInfo], sheets: Dict[str, Dict[str, Dict[str, PortInfo]]], report: Report) -> None:
    sheet_names_by_norm: Dict[str, List[str]] = {}
    for sheet_name in sheets:
        sheet_names_by_norm.setdefault(sheet_name.strip().lower(), []).append(sheet_name)
    claimed_sheets: Set[str] = set()
    for inst in instances.values():
        data = sheets.get(inst.inst_name)
        matched_sheet = inst.inst_name if data else ""
        if not data:
            candidates = sheet_names_by_norm.get(inst.inst_name.strip().lower(), [])
            if len(candidates) == 1:
                matched_sheet = candidates[0]
                data = sheets[matched_sheet]
                report.warn(
                    "port sheet %r matched instance %r by case/space-insensitive fallback"
                    % (matched_sheet, inst.inst_name)
                )
            elif len(candidates) > 1:
                report.error(
                    "multiple port sheets match instance %r by case/space-insensitive fallback: %s"
                    % (inst.inst_name, ", ".join(repr(name) for name in candidates))
                )
        if not data:
            report.warn("no owner port sheet found for instance %s" % inst.inst_name)
            continue
        claimed_sheets.add(matched_sheet)
        inst.inputs = data["inputs"]
        inst.outputs = data["outputs"]
        inst.inouts = data["inouts"]
    for sheet_name in sorted(set(sheets) - claimed_sheets):
        report.warn("port workbook sheet %r does not match any inst_name; ignored" % sheet_name)


def resolve_sdc_paths(instances: Dict[str, InstInfo], cwd: Path, report: Report) -> None:
    all_sdcs = sorted(path for path in cwd.glob("*.sdc") if path.is_file())
    by_name = {path.name: path for path in all_sdcs}
    by_lower = {path.name.lower(): path for path in all_sdcs}
    for inst in instances.values():
        candidates: List[str] = []
        if inst.sdc_hint:
            candidates.append(Path(inst.sdc_hint).name)
        candidates.append(inst.inst_name + ".sdc")
        candidates.append(inst.module_name + ".sdc")
        if inst.file_path:
            stem = Path(inst.file_path).stem
            candidates.append(stem + ".sdc")
            if stem.endswith("_empty"):
                candidates.append(stem[:-6] + ".sdc")
        matches: List[Path] = []
        for name in dict.fromkeys(candidates):
            if not name:
                continue
            if name in by_name:
                matches.append(by_name[name])
            elif name.lower() in by_lower:
                matches.append(by_lower[name.lower()])
        unique: List[Path] = []
        for path in matches:
            if path not in unique:
                unique.append(path)
        if len(unique) == 1:
            inst.sdc_path = unique[0]
        elif len(unique) > 1:
            inst.sdc_path = unique[0]
            report.warn("multiple SDC candidates for %s: %s; using %s" % (inst.inst_name, ", ".join(path.name for path in unique), inst.sdc_path.name))
        else:
            report.warn("no SDC found for %s; tried: %s" % (inst.inst_name, ", ".join(candidates)))


def collect_current_sdc_digests(instances: Dict[str, InstInfo]) -> Dict[str, str]:
    digests: Dict[str, str] = {}
    for inst in instances.values():
        if inst.sdc_path and inst.sdc_path.is_file():
            try:
                digests[inst.sdc_path.name] = digest_file(inst.sdc_path)
            except OSError:
                continue
    return digests


def read_connection_inventory(path: Path, report: Report) -> ConnectionIndex:
    index = ConnectionIndex()
    if not path.is_file():
        report.error("connection inventory not found: %s" % path)
        return index
    seen_ids: Set[str] = set()
    seen_pairs: Set[Tuple[str, str, str, str]] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        if not reader.fieldnames:
            report.error("%s: connection_inventory.csv has no header" % path)
            return index
        for row_idx, row in enumerate(reader, start=2):
            src_inst = clean_cell(row.get("src_instance"))
            src_dir = clean_cell(row.get("src_direction"))
            src_port = clean_cell(row.get("src_port"))
            dst_inst = clean_cell(row.get("dst_instance"))
            dst_dir = clean_cell(row.get("dst_direction"))
            dst_port = clean_cell(row.get("dst_port"))
            if not (src_inst and src_port):
                key_inst, key_dir, key_port = parse_endpoint_key(clean_cell(row.get("src_endpoint_key")))
                src_inst = src_inst or key_inst
                src_dir = src_dir or key_dir
                src_port = src_port or key_port
            if not (dst_inst and dst_port):
                key_inst, key_dir, key_port = parse_endpoint_key(clean_cell(row.get("dst_endpoint_key")))
                dst_inst = dst_inst or key_inst
                dst_dir = dst_dir or key_dir
                dst_port = dst_port or key_port
            if not (src_inst and src_port and dst_inst and dst_port):
                report.warn("%s row %d: skipped connection edge with incomplete src/dst endpoint" % (path.name, row_idx))
                continue
            connection_id = clean_cell(row.get("connection_id")) or build_channel_id(src_inst, src_port, dst_inst, dst_port).replace("CH_", "CONN_", 1)
            if connection_id in seen_ids:
                report.error("%s row %d: duplicate connection_id %s" % (path.name, row_idx, connection_id))
            seen_ids.add(connection_id)
            pair_key = (src_inst, src_port, dst_inst, dst_port)
            if pair_key in seen_pairs:
                report.error("%s row %d: duplicate direct bit pair %s/%s -> %s/%s" % (path.name, row_idx, src_inst, src_port, dst_inst, dst_port))
            seen_pairs.add(pair_key)
            for role, port in (("src", src_port), ("dst", dst_port)):
                if port and not is_canonical_port_key(port):
                    report.error(
                        "%s row %d: %s_port %s is not a canonical scalar/bit key; 00 must expand bus/range connections before 30"
                        % (path.name, row_idx, role, port)
                    )
            edge = ConnectionEdge(
                connection_id=connection_id,
                connection_type=clean_cell(row.get("connection_type")),
                src_instance=src_inst,
                src_direction=src_dir,
                src_port=src_port,
                src_bit_index=inferred_bit_index(src_port, clean_cell(row.get("src_bit_index"))),
                src_endpoint_key=clean_cell(row.get("src_endpoint_key")),
                src_soc_object=clean_cell(row.get("src_soc_object")),
                dst_instance=dst_inst,
                dst_direction=dst_dir,
                dst_port=dst_port,
                dst_bit_index=inferred_bit_index(dst_port, clean_cell(row.get("dst_bit_index"))),
                dst_endpoint_key=clean_cell(row.get("dst_endpoint_key")),
                dst_soc_object=clean_cell(row.get("dst_soc_object")),
                validation_status=clean_cell(row.get("validation_status")),
                note=clean_cell(row.get("note")),
            )
            index.edges.append(edge)
            index.by_src[(edge.src_instance, edge.src_port)].append(edge)
            index.by_dst[(edge.dst_instance, edge.dst_port)].append(edge)
    report.info("loaded %d connection edge(s) from %s" % (len(index.edges), path))
    return index


def read_feedthrough_inventory(path: Path, report: Report) -> List[FeedthroughRecord]:
    if not path.is_file():
        report.info("feedthrough inventory not found: %s; feedthrough exception candidates will require manual review" % path)
        return []
    records: List[FeedthroughRecord] = []
    seen: Set[str] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        if not reader.fieldnames:
            report.warn("%s: feedthrough_inventory.csv has no header" % path)
            return records
        for row_idx, row in enumerate(reader, start=2):
            ft_id = clean_cell(row.get("feedthrough_id"))
            inst = clean_cell(row.get("feedthrough_instance"))
            fti = clean_cell(row.get("fti_port"))
            fto = clean_cell(row.get("fto_port"))
            if not (ft_id and inst and fti and fto):
                report.warn("%s row %d: skipped incomplete feedthrough record" % (path.name, row_idx))
                continue
            if ft_id in seen:
                report.error("%s row %d: duplicate feedthrough_id %s" % (path.name, row_idx, ft_id))
            seen.add(ft_id)
            for role, port in (("fti", fti), ("fto", fto)):
                if port and not is_canonical_port_key(port):
                    report.error(
                        "%s row %d: %s_port %s is not a bit-level canonical key; 10 must emit bit feedthrough ids for 30"
                        % (path.name, row_idx, role, port)
                    )
            records.append(
                FeedthroughRecord(
                    feedthrough_id=ft_id,
                    scenario=normalize_key(row.get("scenario")) or "common",
                    feedthrough_instance=inst,
                    hop_index=clean_cell(row.get("hop_index")),
                    base=clean_cell(row.get("base")),
                    fti_port=fti,
                    fto_port=fto,
                    fti_endpoint=clean_cell(row.get("fti_endpoint")),
                    fto_endpoint=clean_cell(row.get("fto_endpoint")),
                    bit_index=clean_cell(row.get("bit_index")),
                    chain_id=clean_cell(row.get("chain_id")) or clean_cell(row.get("base")),
                    hop_order=clean_cell(row.get("hop_order")),
                    validation_status=clean_cell(row.get("validation_status")),
                )
            )
    report.info("loaded %d feedthrough segment(s) from %s" % (len(records), path))
    return records


def edge_status_ok(edge: ConnectionEdge) -> bool:
    return normalize_key(edge.validation_status) in MATCHED_STATUSES


def feedthrough_status_ok(record: FeedthroughRecord) -> bool:
    return normalize_key(record.validation_status) in MATCHED_STATUSES


def instance_module(instances: Dict[str, InstInfo], inst_name: str) -> str:
    inst = instances.get(inst_name)
    if inst:
        return inst.module_name
    if normalize_key(inst_name) == "top":
        return "top"
    if normalize_key(inst_name) == "fabric":
        return "fabric"
    return ""


def classify_edge_type(edge: ConnectionEdge, instances: Dict[str, InstInfo]) -> str:
    raw = normalize_key(edge.connection_type)
    if raw:
        if "feedthrough" in raw:
            return "feedthrough"
        if raw in {"harden_to_harden", "fabric_to_harden", "harden_to_fabric", "top_pad_to_harden", "harden_to_top_pad", "pad_to_pad", "clock_connection", "exception_path"}:
            return raw
    if normalize_key(edge.src_instance) == "top" or normalize_key(edge.dst_instance) == "top":
        if normalize_key(edge.src_instance) == "top" and normalize_key(edge.dst_instance) == "top":
            return "pad_to_pad"
        return "top_pad_to_harden" if normalize_key(edge.src_instance) == "top" else "harden_to_top_pad"
    if is_feedthrough_port(edge.src_port) or is_feedthrough_port(edge.dst_port):
        return "feedthrough"
    if edge.src_instance in instances and edge.dst_instance in instances:
        return "harden_to_harden"
    if edge.src_instance in instances:
        return "harden_to_fabric"
    if edge.dst_instance in instances:
        return "fabric_to_harden"
    return "unknown"


def edge_to_channel(
    edge: ConnectionEdge,
    instances: Dict[str, InstInfo],
    force_type: str = "",
    related_10_ids: str = "",
    through_collection: str = "",
    note: str = "",
) -> ChannelRecord:
    channel_type = force_type or classify_edge_type(edge, instances)
    is_pad_related = "yes" if channel_type in {"top_pad_to_harden", "harden_to_top_pad", "pad_to_pad"} else "no"
    is_feedthrough_path = "yes" if related_10_ids else "no"
    messages = []
    if note:
        messages.append(note)
    if edge.validation_status and not edge_status_ok(edge):
        messages.append("00 connection validation_status=" + edge.validation_status)
    return ChannelRecord(
        channel_id=build_channel_id(edge.src_instance, edge.src_port, edge.dst_instance, edge.dst_port),
        scenario="common",
        stage="all",
        corner="all",
        channel_type=channel_type,
        connection_id=edge.connection_id,
        src_instance=edge.src_instance,
        src_module=instance_module(instances, edge.src_instance),
        src_direction=edge.src_direction,
        src_port=edge.src_port,
        src_bit_index=edge.src_bit_index,
        src_endpoint=endpoint_from_soc_object(edge.src_instance, edge.src_port, edge.src_soc_object, edge.src_endpoint_key),
        dst_instance=edge.dst_instance,
        dst_module=instance_module(instances, edge.dst_instance),
        dst_direction=edge.dst_direction,
        dst_port=edge.dst_port,
        dst_bit_index=edge.dst_bit_index,
        dst_endpoint=endpoint_from_soc_object(edge.dst_instance, edge.dst_port, edge.dst_soc_object, edge.dst_endpoint_key),
        related_10_feedthrough_id=related_10_ids,
        through_collection=through_collection,
        is_pad_related=is_pad_related,
        is_clock_related="yes" if channel_type == "clock_connection" else "no",
        is_feedthrough_path=is_feedthrough_path,
        note="; ".join(dict.fromkeys(messages)),
    )


def endpoint_for_ft(record: FeedthroughRecord, port: str) -> str:
    if port == record.fti_port and record.fti_endpoint:
        return record.fti_endpoint
    if port == record.fto_port and record.fto_endpoint:
        return record.fto_endpoint
    return get_collection("get_pins", [record.feedthrough_instance + "/" + port])


def through_collection_for_chain(records: Sequence[FeedthroughRecord]) -> str:
    parts: List[str] = []
    for record in sorted(records, key=feedthrough_sort_key):
        parts.append(endpoint_for_ft(record, record.fti_port))
        parts.append(endpoint_for_ft(record, record.fto_port))
    unique: List[str] = []
    for item in parts:
        if item and item not in unique:
            unique.append(item)
    return " ; ".join(unique)


def feedthrough_sort_key(record: FeedthroughRecord) -> Tuple[int, str]:
    token = record.hop_order or record.hop_index or "0"
    return (int(token) if token.isdigit() else 0, record.feedthrough_id)


def synthetic_edge_from_feedthrough_chain(
    records: Sequence[FeedthroughRecord],
    connections: ConnectionIndex,
    report: Report,
) -> Optional[Tuple[ConnectionEdge, str, str]]:
    ordered = sorted(records, key=feedthrough_sort_key)
    first = ordered[0]
    last = ordered[-1]
    incoming = connections.by_dst.get((first.feedthrough_instance, first.fti_port), [])
    outgoing = connections.by_src.get((last.feedthrough_instance, last.fto_port), [])
    label = "%s/%s" % (first.chain_id or first.base, first.bit_index or "scalar")
    if len(incoming) != 1 or len(outgoing) != 1:
        report.warn(
            "%s: cannot stitch 30 feedthrough path; incoming_edges=%d outgoing_edges=%d"
            % (label, len(incoming), len(outgoing))
        )
        return None
    if not all(feedthrough_status_ok(record) for record in ordered):
        report.warn("%s: feedthrough chain has non-matched segment; 30 stitched candidate skipped" % label)
        return None
    ft_ids = [record.feedthrough_id for record in ordered]
    src = incoming[0]
    dst = outgoing[0]
    edge = ConnectionEdge(
        connection_id=";".join([src.connection_id] + ft_ids + [dst.connection_id]),
        connection_type="harden_to_harden",
        src_instance=src.src_instance,
        src_direction=src.src_direction,
        src_port=src.src_port,
        src_bit_index=src.src_bit_index,
        src_endpoint_key=src.src_endpoint_key,
        src_soc_object=src.src_soc_object,
        dst_instance=dst.dst_instance,
        dst_direction=dst.dst_direction,
        dst_port=dst.dst_port,
        dst_bit_index=dst.dst_bit_index,
        dst_endpoint_key=dst.dst_endpoint_key,
        dst_soc_object=dst.dst_soc_object,
        validation_status="matched",
        note="via feedthrough " + ",".join(ft_ids),
    )
    return edge, ",".join(ft_ids), through_collection_for_chain(ordered)


def build_channels_from_inventories(
    instances: Dict[str, InstInfo],
    connections: ConnectionIndex,
    feedthroughs: Sequence[FeedthroughRecord],
    report: Report,
) -> List[ChannelRecord]:
    channels: List[ChannelRecord] = []
    seen: Set[str] = set()

    def add(record: ChannelRecord) -> None:
        if record.channel_id in seen:
            report.warn("duplicate channel_id %s; keeping first 30 candidate channel" % record.channel_id)
            return
        seen.add(record.channel_id)
        channels.append(record)

    for edge in connections.edges:
        channel = edge_to_channel(edge, instances)
        add(channel)

    groups: Dict[Tuple[str, str], List[FeedthroughRecord]] = defaultdict(list)
    for record in feedthroughs:
        groups[(record.chain_id or record.base, record.bit_index)].append(record)
    for _, records in sorted(groups.items()):
        stitched = synthetic_edge_from_feedthrough_chain(records, connections, report)
        if not stitched:
            continue
        edge, ft_ids, through = stitched
        add(edge_to_channel(edge, instances, force_type="harden_to_harden", related_10_ids=ft_ids, through_collection=through, note=edge.note))

    report.info("built %d 30 candidate channel(s) from 00/10 inventories" % len(channels))
    return channels


def is_escaped(text: str, index: int) -> bool:
    count = 0
    pos = index - 1
    while pos >= 0 and text[pos] == "\\":
        count += 1
        pos -= 1
    return count % 2 == 1


def strip_inline_comment(line: str) -> str:
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(line):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(line, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == "#" and brace_depth == 0 and bracket_depth == 0:
                if idx == 0 or line[idx - 1].isspace() or line[idx - 1] == ";":
                    return line[:idx].rstrip()
    return line.rstrip()


def split_semicolon_commands(text: str) -> List[str]:
    parts: List[str] = []
    start = 0
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(text):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(text, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == ";" and brace_depth == 0 and bracket_depth == 0:
                parts.append(text[start:idx])
                start = idx + 1
    parts.append(text[start:])
    return parts


def iter_tcl_commands_with_line(text: str) -> Iterable[TclCommand]:
    buf = ""
    start_line = 0
    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = strip_inline_comment(raw_line)
        if not line.strip():
            continue
        if not buf:
            start_line = line_no
        stripped = line.rstrip()
        continued = stripped.endswith("\\") and not is_escaped(stripped, len(stripped) - 1)
        if continued:
            stripped = stripped[:-1].rstrip()
            buf += stripped + " "
            continue
        buf += stripped
        for cmd in split_semicolon_commands(buf):
            cleaned = cmd.strip().rstrip(";").strip()
            if cleaned:
                yield TclCommand(cleaned, start_line)
        buf = ""
        start_line = 0
    if buf.strip():
        for cmd in split_semicolon_commands(buf):
            cleaned = cmd.strip().rstrip(";").strip()
            if cleaned:
                yield TclCommand(cleaned, start_line)


def find_matching(text: str, start: int, open_char: str, close_char: str) -> int:
    depth = 0
    quote = False
    idx = start
    while idx < len(text):
        char = text[idx]
        if char == "\\":
            idx += 2
            continue
        if char == '"' and open_char != '"' and not is_escaped(text, idx):
            quote = not quote
        elif not quote:
            if char == open_char:
                depth += 1
            elif char == close_char:
                depth -= 1
                if depth == 0:
                    return idx
        idx += 1
    return -1


def tokenize_tcl_words(text: str) -> List[str]:
    tokens: List[str] = []
    idx = 0
    while idx < len(text):
        while idx < len(text) and text[idx].isspace():
            idx += 1
        if idx >= len(text):
            break
        start = idx
        if text[idx] == "{":
            end = find_matching(text, idx, "{", "}")
            if end < 0:
                tokens.append(text[start:])
                break
            tokens.append(text[start : end + 1])
            idx = end + 1
            continue
        if text[idx] == '"':
            idx += 1
            while idx < len(text):
                if text[idx] == "\\":
                    idx += 2
                    continue
                if text[idx] == '"':
                    idx += 1
                    break
                idx += 1
            tokens.append(text[start:idx])
            continue
        pieces: List[str] = []
        while idx < len(text) and not text[idx].isspace():
            char = text[idx]
            if char == "\\":
                pieces.append(text[idx : idx + 2])
                idx += 2
            elif char == "[":
                end = find_matching(text, idx, "[", "]")
                if end < 0:
                    pieces.append(text[idx:])
                    idx = len(text)
                else:
                    pieces.append(text[idx : end + 1])
                    idx = end + 1
            else:
                pieces.append(char)
                idx += 1
        tokens.append("".join(pieces))
    return tokens


def strip_braces(text: str) -> str:
    text = clean_cell(text)
    if len(text) >= 2 and text[0] == "{" and text[-1] == "}":
        return text[1:-1].strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        return text[1:-1].strip()
    return text


def split_object_list(text: str) -> List[str]:
    text = strip_braces(text)
    return [part for part in re.split(r"[\s,;]+", text) if part]


COLLECTION_KINDS = {"get_ports", "get_pins", "get_nets", "get_clocks"}


def parse_collection(token: str) -> Optional[Tuple[str, List[str]]]:
    text = clean_cell(token)
    if not (text.startswith("[") and text.endswith("]")):
        return None
    end = find_matching(text, 0, "[", "]")
    if end != len(text) - 1:
        return None
    words = tokenize_tcl_words(text[1:-1].strip())
    if not words or words[0] not in COLLECTION_KINDS:
        return None
    kind = words[0]
    objects: List[str] = []
    idx = 1
    while idx < len(words):
        word = words[idx]
        if word.startswith("-"):
            idx += 1
            continue
        objects.extend(split_object_list(word))
        idx += 1
    return kind, objects


def option_value(tokens: Sequence[str], option: str) -> str:
    for idx, token in enumerate(tokens):
        if token == option and idx + 1 < len(tokens):
            return tokens[idx + 1]
    return ""


def has_option(tokens: Sequence[str], option: str) -> bool:
    return option in tokens


def positional_tokens(tokens: Sequence[str], options_with_value: Set[str], options_no_value: Set[str]) -> List[str]:
    result: List[str] = []
    skip_next = False
    for token in tokens[1:]:
        if skip_next:
            skip_next = False
            continue
        if token in options_with_value:
            skip_next = True
            continue
        if token in options_no_value:
            continue
        if token.startswith("-"):
            continue
        if parse_collection(token):
            continue
        result.append(strip_braces(token))
    return result


DELAY_OPTIONS_WITH_VALUE = {"-clock", "-min", "-max"}
DELAY_OPTIONS_NO_VALUE = {"-add_delay", "-clock_fall", "-rise", "-fall"}
EXC_OPTIONS_WITH_VALUE = {"-from", "-to", "-through"}
EXC_OPTIONS_NO_VALUE = {"-setup", "-hold", "-start", "-end", "-datapath_only"}


def extract_clock_name(tokens: Sequence[str]) -> str:
    value = option_value(tokens, "-clock")
    parsed = parse_collection(value)
    if parsed and parsed[0] == "get_clocks" and parsed[1]:
        return parsed[1][0]
    return strip_braces(value)


def last_non_clock_collection(tokens: Sequence[str]) -> Optional[Tuple[str, List[str], str]]:
    found: Optional[Tuple[str, List[str], str]] = None
    for token in tokens:
        parsed = parse_collection(token)
        if parsed and parsed[0] != "get_clocks":
            found = (parsed[0], parsed[1], token)
    return found


def normalize_sdc_port(kind: str, obj: str) -> str:
    obj = strip_braces(obj)
    if kind == "get_pins" and "/" in obj:
        return obj.split("/")[-1]
    return obj


def lookup_port_direction(inst: InstInfo, port_name: str) -> str:
    if port_name in inst.inputs:
        return "input"
    if port_name in inst.outputs:
        return "output"
    if port_name in inst.inouts:
        return "inout"
    base = port_base(port_name)
    if base in inst.inputs:
        return "input"
    if base in inst.outputs:
        return "output"
    if base in inst.inouts:
        return "inout"
    return "unknown"


def parse_delay_evidence(inst: InstInfo, cmd: TclCommand) -> Optional[DelayEvidence]:
    tokens = tokenize_tcl_words(cmd.raw)
    if not tokens or tokens[0] not in {"set_input_delay", "set_output_delay"}:
        return None
    ctype = "input_delay" if tokens[0] == "set_input_delay" else "output_delay"
    target = last_non_clock_collection(tokens)
    if not target:
        return None
    kind, objects, _ = target
    if len(objects) != 1:
        return None
    port_name = normalize_sdc_port(kind, objects[0])
    return DelayEvidence(
        inst.inst_name,
        port_name,
        lookup_port_direction(inst, port_name),
        ctype,
        extract_clock_name(tokens),
    )


def ports_from_collection_token(token: str) -> Tuple[List[str], str]:
    parsed = parse_collection(token)
    if not parsed:
        return [], ""
    kind, objects = parsed
    if kind == "get_clocks":
        return [], "clock_collection"
    if kind not in {"get_ports", "get_pins"}:
        return [], kind
    return [normalize_sdc_port(kind, obj) for obj in objects], kind


def parse_exception_evidence(inst: InstInfo, cmd: TclCommand, digest: str, now: str) -> Optional[ExceptionEvidence]:
    tokens = tokenize_tcl_words(cmd.raw)
    if not tokens or tokens[0] not in {"set_false_path", "set_multicycle_path", "set_max_delay", "set_min_delay"}:
        return None

    command = tokens[0]
    if command == "set_false_path":
        exception_type = "false_path"
    elif command == "set_multicycle_path":
        exception_type = "multicycle_path"
    elif command == "set_max_delay":
        exception_type = "max_delay_override"
    else:
        exception_type = "min_delay_override"

    from_token = option_value(tokens, "-from")
    to_token = option_value(tokens, "-to")
    through_token = option_value(tokens, "-through")
    from_ports, from_kind = ports_from_collection_token(from_token)
    to_ports, to_kind = ports_from_collection_token(to_token)
    message_parts: List[str] = []
    if from_kind == "clock_collection" or to_kind == "clock_collection":
        message_parts.append("clock-level from/to collection is not allowed in 30")
    if through_token:
        parsed = parse_collection(through_token)
        if parsed and parsed[0] == "get_clocks":
            message_parts.append("clock-level through collection is not allowed in 30")

    check_type = "both"
    if has_option(tokens, "-setup") and not has_option(tokens, "-hold"):
        check_type = "setup"
    elif has_option(tokens, "-hold") and not has_option(tokens, "-setup"):
        check_type = "hold"
    mcp_reference = ""
    if has_option(tokens, "-start"):
        mcp_reference = "start"
    elif has_option(tokens, "-end"):
        mcp_reference = "end"

    positionals = positional_tokens(tokens, EXC_OPTIONS_WITH_VALUE, EXC_OPTIONS_NO_VALUE)
    max_value = positionals[0] if command == "set_max_delay" and positionals else ""
    min_value = positionals[0] if command == "set_min_delay" and positionals else ""
    setup_cycles = ""
    hold_cycles = ""
    if command == "set_multicycle_path" and positionals:
        if check_type in {"both", "setup"}:
            setup_cycles = positionals[0]
        if check_type in {"both", "hold"}:
            hold_cycles = positionals[0]

    return ExceptionEvidence(
        inst_name=inst.inst_name,
        module_name=inst.module_name,
        owner=inst.owner,
        command_type=command,
        exception_type=exception_type,
        from_ports=from_ports,
        to_ports=to_ports,
        from_collection=from_token,
        to_collection=to_token,
        through_collection=through_token,
        check_type=check_type,
        max_value=max_value,
        min_value=min_value,
        setup_cycles=setup_cycles,
        hold_cycles=hold_cycles,
        mcp_reference=mcp_reference,
        source_sdc_file=inst.sdc_path.name if inst.sdc_path else "",
        source_line=str(cmd.line_no),
        source_command=cmd.raw,
        source_digest=digest,
        extraction_time=now,
        message="; ".join(message_parts),
    )


def extract_sdc_evidence(instances: Dict[str, InstInfo], report: Report) -> Tuple[List[DelayEvidence], List[ExceptionEvidence]]:
    delays: List[DelayEvidence] = []
    exceptions: List[ExceptionEvidence] = []
    now = datetime.now().isoformat(timespec="seconds")
    for inst in instances.values():
        if not inst.sdc_path:
            continue
        try:
            text = read_text(inst.sdc_path)
            digest = digest_file(inst.sdc_path)
        except Exception as exc:
            report.error("failed to read %s: %s" % (inst.sdc_path, exc))
            continue
        delay_count = 0
        exception_count = 0
        for cmd in iter_tcl_commands_with_line(text):
            delay = parse_delay_evidence(inst, cmd)
            if delay:
                delays.append(delay)
                delay_count += 1
            evidence = parse_exception_evidence(inst, cmd, digest, now)
            if evidence:
                exceptions.append(evidence)
                exception_count += 1
        report.info("extracted %d timing delay and %d exception candidate command(s) from %s" % (delay_count, exception_count, inst.sdc_path.name))
    return delays, exceptions


def channel_port_indexes(channels: Sequence[ChannelRecord]) -> Tuple[Dict[Tuple[str, str], List[ChannelRecord]], Dict[Tuple[str, str], List[ChannelRecord]]]:
    by_src: Dict[Tuple[str, str], List[ChannelRecord]] = defaultdict(list)
    by_dst: Dict[Tuple[str, str], List[ChannelRecord]] = defaultdict(list)
    for ch in channels:
        for port in (ch.src_port, port_base(ch.src_port)):
            by_src[(ch.src_instance, port)].append(ch)
        for port in (ch.dst_port, port_base(ch.dst_port)):
            by_dst[(ch.dst_instance, port)].append(ch)
    return by_src, by_dst


def channels_for_port(
    inst_name: str,
    port: str,
    by_src: Dict[Tuple[str, str], List[ChannelRecord]],
    by_dst: Dict[Tuple[str, str], List[ChannelRecord]],
) -> List[ChannelRecord]:
    result: List[ChannelRecord] = []
    seen: Set[str] = set()
    for key in ((inst_name, port), (inst_name, port_base(port))):
        for ch in by_src.get(key, []) + by_dst.get(key, []):
            if ch.channel_id not in seen:
                seen.add(ch.channel_id)
                result.append(ch)
    return result


def map_exception_to_channels(evidence: ExceptionEvidence, channels: Sequence[ChannelRecord]) -> List[ChannelRecord]:
    by_src, by_dst = channel_port_indexes(channels)
    from_matches: Set[str] = set()
    to_matches: Set[str] = set()
    by_id = {ch.channel_id: ch for ch in channels}
    for port in evidence.from_ports:
        for ch in channels_for_port(evidence.inst_name, port, by_src, by_dst):
            from_matches.add(ch.channel_id)
    for port in evidence.to_ports:
        for ch in channels_for_port(evidence.inst_name, port, by_src, by_dst):
            to_matches.add(ch.channel_id)
    if from_matches and to_matches:
        ids = from_matches & to_matches
    elif from_matches:
        ids = from_matches
    else:
        ids = to_matches
    return [by_id[ch_id] for ch_id in sorted(ids)]


def port_refers_to_channel_port(ports: Sequence[str], channel_port: str) -> bool:
    wanted = clean_cell(channel_port)
    wanted_base = port_base(wanted)
    for port in ports:
        text = clean_cell(port)
        if text == wanted or port_base(text) == wanted_base:
            return True
    return False


def evidence_endpoint_collection(evidence: Optional[ExceptionEvidence], channel: ChannelRecord, side: str) -> str:
    if side == "from":
        endpoint = channel.src_endpoint
        if evidence and evidence.inst_name == channel.src_instance:
            if port_refers_to_channel_port(evidence.from_ports, channel.src_port):
                return endpoint
        if evidence and evidence.from_collection:
            return evidence.from_collection
        return endpoint
    endpoint = channel.dst_endpoint
    if evidence and evidence.inst_name == channel.dst_instance:
        if port_refers_to_channel_port(evidence.to_ports, channel.dst_port):
            return endpoint
    if evidence and evidence.to_collection:
        return evidence.to_collection
    return endpoint


def timing_contract_for_channel(channel: ChannelRecord, delays: Sequence[DelayEvidence]) -> Tuple[str, str, str, str, str]:
    src_delays = []
    dst_delays = []
    src_clocks = []
    dst_clocks = []
    for delay in delays:
        if delay.inst_name == channel.src_instance and delay.constraint_type == "output_delay":
            if delay.port_name in {channel.src_port, port_base(channel.src_port)}:
                src_delays.append(delay)
                if delay.clock_name:
                    src_clocks.append(delay.clock_name)
        if delay.inst_name == channel.dst_instance and delay.constraint_type == "input_delay":
            if delay.port_name in {channel.dst_port, port_base(channel.dst_port)}:
                dst_delays.append(delay)
                if delay.clock_name:
                    dst_clocks.append(delay.clock_name)
    has_src = "yes" if src_delays else "no"
    has_dst = "yes" if dst_delays else "no"
    if src_delays and dst_delays:
        status = "both_sides_timed"
    elif src_delays:
        status = "src_timed_only"
    elif dst_delays:
        status = "dst_timed_only"
    else:
        status = "no_port_timing"
    return status, has_src, has_dst, join_unique(src_clocks), join_unique(dst_clocks)


def row_header_map(ws) -> Dict[str, int]:
    return {clean_cell(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean_cell(cell.value)}


def workbook_row_values(ws, row_idx: int, headers: Sequence[str]) -> Dict[str, object]:
    hmap = row_header_map(ws)
    return {header: ws.cell(row=row_idx, column=hmap[header]).value if header in hmap else "" for header in headers}


def row_scenario_value(values: Dict[str, object]) -> str:
    return normalize_key(values.get("scenario")) or "common"


def row_stage_value(values: Dict[str, object]) -> str:
    return normalize_key(values.get("stage")) or "all"


def row_corner_value(values: Dict[str, object]) -> str:
    return clean_cell(values.get("corner")) or "all"


def values_selected_for_assembled(values: Dict[str, object], scenario: str, stage: str, corner: str) -> bool:
    row_sc = row_scenario_value(values)
    if row_sc not in {"common", scenario}:
        return False
    row_st = row_stage_value(values)
    row_co = row_corner_value(values)
    return (row_st == "all" and row_co == "all") or (row_st == stage and row_co == corner)


def read_20_active_budgets(path: Path, scenario: str, stage: str, corner: str, report: Report) -> Dict[str, Active20Budget]:
    active: Dict[str, Active20Budget] = {}
    if not path.is_file():
        report.warn("20 workbook not found: %s; active 20 budget overlap checks are limited" % path)
        return active
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        report.error("failed to open 20 workbook %s: %s" % (path, exc))
        return active
    if "interface_budget" not in wb.sheetnames:
        report.warn("%s has no interface_budget sheet; active 20 budget overlap checks are limited" % path)
        return active
    ws = wb["interface_budget"]
    hmap = row_header_map(ws)
    required = {"channel_id", "scenario", "stage", "corner", "apply", "review_status", "emit_max", "emit_min", "converted_max", "converted_min"}
    missing = required - set(hmap)
    if missing:
        report.warn("%s interface_budget missing column(s): %s; active 20 check is limited" % (path.name, ", ".join(sorted(missing))))
        return active
    for row_idx in range(2, ws.max_row + 1):
        values = {header: ws.cell(row=row_idx, column=hmap[header]).value if header in hmap else "" for header in hmap}
        if not values_selected_for_assembled(values, scenario, stage, corner):
            continue
        if normalize_key(values.get("apply")) != "yes" or normalize_key(values.get("review_status")) != "approved":
            continue
        channel_id = clean_cell(values.get("channel_id"))
        if not channel_id:
            continue
        max_active = normalize_key(values.get("emit_max")) == "yes" and bool(clean_cell(values.get("converted_max")))
        min_active = normalize_key(values.get("emit_min")) == "yes" and bool(clean_cell(values.get("converted_min")))
        if not (max_active or min_active):
            continue
        item = active.setdefault(channel_id, Active20Budget(channel_id))
        item.max_active = item.max_active or max_active
        item.min_active = item.min_active or min_active
        item.rows.append(str(row_idx))
    report.info("loaded %d active 20 budget channel(s) from %s" % (len(active), path))
    return active


def create_candidate_and_rule_seeds(
    channels: Sequence[ChannelRecord],
    delays: Sequence[DelayEvidence],
    exceptions: Sequence[ExceptionEvidence],
    active20: Dict[str, Active20Budget],
    report: Report,
) -> Tuple[List[CandidateSeed], List[RuleSeed], List[ExceptionEvidence]]:
    seeds: List[CandidateSeed] = []
    rule_seeds: List[RuleSeed] = []
    evidence_log: List[ExceptionEvidence] = []

    channel_by_id = {ch.channel_id: ch for ch in channels}
    evidence_by_channel: Dict[str, List[ExceptionEvidence]] = defaultdict(list)
    for evidence in exceptions:
        mapped = map_exception_to_channels(evidence, channels)
        if not mapped:
            report.warn(
                "%s:%s %s: extracted exception did not map to any 00/10/20 harden channel"
                % (evidence.source_sdc_file, evidence.source_line, evidence.command_type)
            )
        for ch in mapped:
            evidence_by_channel[ch.channel_id].append(evidence)
        evidence_log.append(evidence)

    for ch in channels:
        if ch.channel_type != "harden_to_harden":
            continue
        timing_status, has_src, has_dst, src_clock, dst_clock = timing_contract_for_channel(ch, delays)
        active = active20.get(ch.channel_id)
        related_20_status = ""
        if active:
            related_20_status = "active_max=%s; active_min=%s; rows=%s" % (
                "yes" if active.max_active else "no",
                "yes" if active.min_active else "no",
                ",".join(active.rows),
            )
        base_source_type = "missing_timing_candidate" if timing_status == "no_port_timing" and not active else "manual_entry"
        base_reason = "no ordinary port timing found" if base_source_type == "missing_timing_candidate" else "harden-to-harden channel for review"
        if active:
            base_reason = "ordinary 20 budget is active; 30 only allowed for non-overlapping exception"
        values = {header: "" for header in EXCEPTION_CANDIDATE_HEADERS}
        values.update(
            {
                "candidate_id": "CAND_" + ch.channel_id,
                "scenario": ch.scenario,
                "stage": ch.stage,
                "corner": ch.corner,
                "channel_id": ch.channel_id,
                "source_type": base_source_type,
                "path_category": "unknown",
                "timing_contract_status": timing_status,
                "src_instance": ch.src_instance,
                "src_module": ch.src_module,
                "src_port": ch.src_port,
                "src_bit_index": ch.src_bit_index,
                "src_endpoint": ch.src_endpoint,
                "dst_instance": ch.dst_instance,
                "dst_module": ch.dst_module,
                "dst_port": ch.dst_port,
                "dst_bit_index": ch.dst_bit_index,
                "dst_endpoint": ch.dst_endpoint,
                "src_clock": src_clock,
                "dst_clock": dst_clock,
                "clock_relation": "unknown",
                "has_src_output_delay": has_src,
                "has_dst_input_delay": has_dst,
                "related_20_channel_id": ch.channel_id,
                "related_20_status": related_20_status,
                "related_10_feedthrough_id": ch.related_10_feedthrough_id,
                "harden_clock_context_status": "not_applicable",
                "candidate_reason": base_reason,
                "recommended_action": "review_exception_basis" if base_source_type == "missing_timing_candidate" else "review_if_30_needed",
                "note": ch.note,
            }
        )
        seeds.append(CandidateSeed(values, []))
        rule_seeds.append(seed_rule_from_candidate(values, ch, None))

        for idx, evidence in enumerate(evidence_by_channel.get(ch.channel_id, []), start=1):
            e_values = dict(values)
            e_values.update(
                {
                    "candidate_id": "CAND_%s_SDC_%s_%s" % (ch.channel_id, sanitize_id(evidence.source_sdc_file), evidence.source_line),
                    "source_type": "extracted_harden_exception",
                    "path_category": "unknown",
                    "harden_clock_context_status": "unknown",
                    "source_sdc_file": evidence.source_sdc_file,
                    "source_line": evidence.source_line,
                    "source_command": evidence.source_command,
                    "source_digest": evidence.source_digest,
                    "extraction_time": evidence.extraction_time,
                    "candidate_reason": "harden SDC boundary exception candidate",
                    "recommended_action": "confirm_clock_context_and_review_rule",
                    "note": "; ".join(part for part in [ch.note, evidence.message] if part),
                }
            )
            seeds.append(CandidateSeed(e_values, [evidence]))
            rule_seeds.append(seed_rule_from_candidate(e_values, ch, evidence))

    return seeds, rule_seeds, evidence_log


def seed_rule_from_candidate(values: Dict[str, str], channel: ChannelRecord, evidence: Optional[ExceptionEvidence]) -> RuleSeed:
    rule = {header: "" for header in EXCEPTION_RULE_HEADERS}
    exception_type = evidence.exception_type if evidence else "needs_review"
    rule.update(
        {
            "exception_id": clean_cell(values.get("candidate_id")).replace("CAND_", "EXC_", 1),
            "scenario": clean_cell(values.get("scenario")) or "common",
            "stage": clean_cell(values.get("stage")) or "all",
            "corner": clean_cell(values.get("corner")) or "all",
            "apply": "no",
            "review_status": "pending",
            "owner": evidence.owner if evidence else "",
            "exception_type": exception_type,
            "path_category": clean_cell(values.get("path_category")) or "unknown",
            "channel_id": channel.channel_id,
            "related_20_channel_id": channel.channel_id,
            "related_10_feedthrough_id": channel.related_10_feedthrough_id,
            "src_bit_index": channel.src_bit_index,
            "src_endpoint": channel.src_endpoint,
            "dst_bit_index": channel.dst_bit_index,
            "dst_endpoint": channel.dst_endpoint,
            "from_collection": evidence_endpoint_collection(evidence, channel, "from"),
            "to_collection": evidence_endpoint_collection(evidence, channel, "to"),
            "through_collection": evidence.through_collection if evidence and evidence.through_collection else channel.through_collection,
            "src_clock": clean_cell(values.get("src_clock")),
            "dst_clock": clean_cell(values.get("dst_clock")),
            "clock_relation": clean_cell(values.get("clock_relation")) or "unknown",
            "timing_contract_status": clean_cell(values.get("timing_contract_status")),
            "harden_clock_context_status": "unknown" if evidence else "not_applicable",
            "check_type": evidence.check_type if evidence else "both",
            "max_value": evidence.max_value if evidence else "",
            "min_value": evidence.min_value if evidence else "",
            "setup_cycles": evidence.setup_cycles if evidence else "",
            "hold_cycles": evidence.hold_cycles if evidence else "",
            "mcp_reference": evidence.mcp_reference if evidence and evidence.mcp_reference else "",
            "datapath_only": "no",
            "tool_surface": "sta",
            "source_type": clean_cell(values.get("source_type")),
            "source_sdc_file": evidence.source_sdc_file if evidence else "",
            "source_line": evidence.source_line if evidence else "",
            "source_command": evidence.source_command if evidence else "",
            "source_digest": evidence.source_digest if evidence else "",
            "risk_level": "medium" if evidence else "",
            "note": clean_cell(values.get("note")),
        }
    )
    return RuleSeed(rule)


def create_or_load_workbook(path: Path) -> Tuple[Workbook, bool]:
    if path.is_file():
        return load_workbook(path), False
    wb = Workbook()
    ws = wb.active
    ws.title = "exception_rule"
    return wb, True


def ensure_sheet(wb: Workbook, name: str, headers: Sequence[str]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if ws.max_row == 1 and all(ws.cell(row=1, column=col).value is None for col in range(1, len(headers) + 1)):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    existing = [clean_cell(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
    for header in headers:
        if header not in existing:
            ws.cell(row=1, column=len(existing) + 1, value=header)
            existing.append(header)
    style_sheet(ws)


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for column_cells in ws.columns:
        max_len = 8
        for cell in column_cells:
            max_len = max(max_len, len(clean_cell(cell.value)))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_len + 2, 48)
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1 and not ws.tables:
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:25] or "table"
        ref = "A1:%s%d" % (get_column_letter(ws.max_column), max(ws.max_row, 2))
        tab = Table(displayName=table_name + "_tbl", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium5", showRowStripes=True)
        try:
            ws.add_table(tab)
        except ValueError:
            pass


def header_map(ws) -> Dict[str, int]:
    return {clean_cell(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean_cell(cell.value)}


def row_values(ws, row_idx: int, headers: Sequence[str]) -> Dict[str, object]:
    hmap = header_map(ws)
    return {header: ws.cell(row=row_idx, column=hmap[header]).value if header in hmap else "" for header in headers}


def append_dict(ws, headers: Sequence[str], values: Dict[str, object], fill: Optional[PatternFill] = None) -> None:
    hmap = header_map(ws)
    row_idx = ws.max_row + 1
    for header in headers:
        col = hmap[header]
        cell = ws.cell(row=row_idx, column=col, value=values.get(header, ""))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def sync_workbook(path: Path, candidates: Sequence[CandidateSeed], rules: Sequence[RuleSeed], log_items: Sequence[ExceptionEvidence], report: Report) -> None:
    wb, created = create_or_load_workbook(path)
    ensure_sheet(wb, "exception_candidate", EXCEPTION_CANDIDATE_HEADERS)
    ensure_sheet(wb, "exception_rule", EXCEPTION_RULE_HEADERS)
    ensure_sheet(wb, "extraction_log", EXTRACTION_LOG_HEADERS)

    ws_cand = wb["exception_candidate"]
    existing_candidates = {
        clean_cell(row_values(ws_cand, row_idx, EXCEPTION_CANDIDATE_HEADERS).get("candidate_id"))
        for row_idx in range(2, ws_cand.max_row + 1)
        if clean_cell(ws_cand.cell(row=row_idx, column=1).value)
    }
    for seed in candidates:
        candidate_id = clean_cell(seed.values.get("candidate_id"))
        if candidate_id not in existing_candidates:
            append_dict(ws_cand, EXCEPTION_CANDIDATE_HEADERS, seed.values, NEW_FILL)
            existing_candidates.add(candidate_id)
            report.sync_changed = True

    ws_rule = wb["exception_rule"]
    existing_rules = {
        clean_cell(row_values(ws_rule, row_idx, EXCEPTION_RULE_HEADERS).get("exception_id"))
        for row_idx in range(2, ws_rule.max_row + 1)
        if clean_cell(ws_rule.cell(row=row_idx, column=1).value)
    }
    for seed in rules:
        exception_id = clean_cell(seed.values.get("exception_id"))
        if exception_id not in existing_rules:
            append_dict(ws_rule, EXCEPTION_RULE_HEADERS, seed.values, NEW_FILL)
            existing_rules.add(exception_id)
            report.sync_changed = True

    ws_log = wb["extraction_log"]
    if ws_log.max_row > 1:
        ws_log.delete_rows(2, ws_log.max_row - 1)
    for evidence in log_items:
        append_dict(
            ws_log,
            EXTRACTION_LOG_HEADERS,
            {
                "source_sdc_file": evidence.source_sdc_file,
                "source_line": evidence.source_line,
                "instance": evidence.inst_name,
                "command_type": evidence.command_type,
                "mapped_channel_ids": "",
                "from_collection": evidence.from_collection,
                "to_collection": evidence.to_collection,
                "through_collection": evidence.through_collection,
                "source_digest": evidence.source_digest,
                "extraction_time": evidence.extraction_time,
                "source_command": evidence.source_command,
                "message": evidence.message,
            },
        )

    add_validations(wb)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(path)
    if created or report.sync_changed:
        report.sync_changed = True
        report.info("synchronized workbook %s; review new rows before generation" % path.name)


def add_validations(wb: Workbook) -> None:
    if "exception_rule" not in wb.sheetnames:
        return
    ws = wb["exception_rule"]
    hmap = header_map(ws)

    def add_list(header: str, values: Sequence[str]) -> None:
        if header not in hmap:
            return
        col = get_column_letter(hmap[header])
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("%s2:%s1048576" % (col, col))

    add_list("scenario", sorted(SCENARIOS))
    add_list("stage", sorted(STAGES))
    add_list("apply", ["yes", "no"])
    add_list("review_status", sorted(REVIEW_STATUS_VALUES - {""}))
    add_list("exception_type", sorted(EXCEPTION_TYPES - {""}))
    add_list("path_category", sorted(PATH_CATEGORIES - {""}))
    add_list("clock_relation", sorted(CLOCK_RELATION_CANONICAL))
    add_list("timing_contract_status", sorted(TIMING_CONTRACT_STATUSES - {""}))
    add_list("harden_clock_context_status", sorted(CLOCK_CONTEXT_STATUSES - {""}))
    add_list("check_type", sorted(CHECK_TYPES - {""}))
    add_list("mcp_reference", sorted(MCP_REFERENCES - {""}))
    add_list("cross_clock_mcp_review", sorted(CROSS_CLOCK_MCP_REVIEW - {""}))
    add_list("datapath_only", ["yes", "no"])
    add_list("tool_surface", sorted(TOOL_SURFACES - {""}))
    add_list("source_type", sorted(SOURCE_TYPES - {""}))
    add_list("risk_level", sorted(RISK_LEVELS - {""}))


def read_rule_rows(path: Path) -> List[FormRow]:
    wb = load_workbook(path, data_only=False)
    if "exception_rule" not in wb.sheetnames:
        raise RuntimeError("%s missing exception_rule sheet" % path)
    ws = wb["exception_rule"]
    rows: List[FormRow] = []
    for row_idx in range(2, ws.max_row + 1):
        values = row_values(ws, row_idx, EXCEPTION_RULE_HEADERS)
        if not any(clean_cell(value) for value in values.values()):
            continue
        rows.append(FormRow(row_idx=row_idx, values=values))
    return rows


def row_scenario(row: FormRow) -> str:
    return normalize_key(row.values.get("scenario")) or "common"


def row_stage(row: FormRow) -> str:
    return normalize_key(row.values.get("stage")) or "all"


def row_corner(row: FormRow) -> str:
    return clean_cell(row.values.get("corner")) or "all"


def row_selected_for_output(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    return row_scenario(row) == scenario and row_stage(row) == stage and row_corner(row) == corner


def row_selected_for_assembled(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    row_sc = row_scenario(row)
    if row_sc not in {"common", scenario}:
        return False
    return (row_stage(row) == "all" and row_corner(row) == "all") or (row_stage(row) == stage and row_corner(row) == corner)


def is_apply_approved(row: FormRow) -> bool:
    return normalize_key(row.values.get("apply")) == "yes" and normalize_key(row.values.get("review_status")) == "approved"


def today_date() -> date:
    return datetime.now().date()


def parse_review_date(value: str) -> Optional[date]:
    text = clean_cell(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def collection_has_get_clocks(value: str) -> bool:
    text = clean_cell(value)
    return bool(re.search(r"\[\s*get_clocks\b", text))


def collection_has_get_nets(value: str) -> bool:
    text = clean_cell(value)
    return bool(re.search(r"\[\s*get_nets\b", text))


def collection_has_pattern(value: str) -> bool:
    text = clean_cell(value)
    return "*" in text or "?" in text


def collection_has_range_or_bus(value: str) -> bool:
    text = clean_cell(value)
    if PORT_RANGE_RE.search(text):
        return True
    for parsed in re.findall(r"\{([^}]+)\}", text):
        for obj in split_object_list(parsed):
            port = obj.split("/")[-1]
            if not is_canonical_port_key(port):
                return True
    return False


def endpoint_present(row: FormRow) -> Tuple[str, str]:
    values = row.values
    src = clean_cell(values.get("from_collection")) or clean_cell(values.get("src_endpoint"))
    dst = clean_cell(values.get("to_collection")) or clean_cell(values.get("dst_endpoint"))
    return src, dst


def check_source_digest(row: FormRow, current_digests: Dict[str, str], report: Report) -> None:
    sources = [part.strip() for part in clean_cell(row.values.get("source_sdc_file")).split(";") if part.strip()]
    digests = [part.strip() for part in clean_cell(row.values.get("source_digest")).split(";") if part.strip()]
    for idx, source in enumerate(sources):
        stored = digests[idx] if idx < len(digests) else ""
        current = current_digests.get(source)
        if current and stored and current != stored:
            report.warn(
                "exception_rule row %d %s: source_digest mismatch for %s; row may be stale"
                % (row.row_idx, clean_cell(row.values.get("exception_id")), source)
            )


def reset_false_path_basis_ok(values: Dict[str, object]) -> bool:
    text = " ".join(
        clean_cell(values.get(field))
        for field in ("basis", "cdc_rdc_ref", "sta_waiver_ref", "protocol_ref", "note")
    )
    folded = text.lower()
    return any(token in folded for token in RESET_BASIS_TOKENS)


def row_exception_type(row: FormRow) -> str:
    return normalize_key(row.values.get("exception_type"))


def check_dimensions_for_exception(etype: str, check_type: str) -> Tuple[bool, bool]:
    if etype == "false_path":
        if check_type == "hold":
            return False, True
        if check_type == "setup":
            return True, False
        return True, True
    if etype in {"max_min_delay_override", "multicycle_path"}:
        return True, True
    if etype == "max_delay_override":
        return True, False
    if etype == "min_delay_override":
        return False, True
    return False, False


def validate_rows(
    rows: Sequence[FormRow],
    channels: Sequence[ChannelRecord],
    active20: Dict[str, Active20Budget],
    feedthroughs: Sequence[FeedthroughRecord],
    scenario: str,
    stage: str,
    corner: str,
    current_digests: Dict[str, str],
    report: Report,
) -> None:
    channel_by_id = {ch.channel_id: ch for ch in channels}
    ft_ids = {ft.feedthrough_id for ft in feedthroughs}
    assembled = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner) and is_apply_approved(row)]
    by_path: Dict[Tuple[str, str, str], List[FormRow]] = defaultdict(list)

    for row in rows:
        values = row.values
        exception_id = clean_cell(values.get("exception_id")) or "row_%d" % row.row_idx
        apply_value = normalize_key(values.get("apply"))
        review_status = normalize_key(values.get("review_status"))
        etype = normalize_key(values.get("exception_type"))
        path_category = normalize_key(values.get("path_category"))
        source_type = normalize_key(values.get("source_type"))
        channel_id = clean_cell(values.get("channel_id"))
        channel = channel_by_id.get(channel_id)
        relation_raw = normalize_key(values.get("clock_relation"))
        clock_relation = canonical_clock_relation(values.get("clock_relation")) or ("unknown" if not relation_raw else "")
        check_type = normalize_key(values.get("check_type")) or "both"
        tool_surface = normalize_key(values.get("tool_surface"))
        datapath_only = normalize_key(values.get("datapath_only"))
        timing_status = normalize_key(values.get("timing_contract_status"))
        clock_context = normalize_key(values.get("harden_clock_context_status"))
        src, dst = endpoint_present(row)
        through = clean_cell(values.get("through_collection"))

        if apply_value and apply_value not in YES_NO:
            report.error("exception_rule row %d: apply must be yes/no" % row.row_idx)
        if review_status and review_status not in REVIEW_STATUS_VALUES:
            report.error("exception_rule row %d: invalid review_status %s" % (row.row_idx, review_status))
        if etype and etype not in EXCEPTION_TYPES:
            report.error("exception_rule row %d %s: invalid exception_type %s" % (row.row_idx, exception_id, etype))
        if path_category and path_category not in PATH_CATEGORIES:
            report.error("exception_rule row %d %s: invalid path_category %s" % (row.row_idx, exception_id, path_category))
        if source_type and source_type not in SOURCE_TYPES:
            report.error("exception_rule row %d %s: invalid source_type %s" % (row.row_idx, exception_id, source_type))
        if relation_raw and not canonical_clock_relation(values.get("clock_relation")):
            report.error("exception_rule row %d %s: invalid clock_relation %s" % (row.row_idx, exception_id, relation_raw))
        if check_type and check_type not in CHECK_TYPES:
            report.error("exception_rule row %d %s: invalid check_type %s" % (row.row_idx, exception_id, check_type))
        if normalize_key(values.get("mcp_reference")) and normalize_key(values.get("mcp_reference")) not in MCP_REFERENCES:
            report.error("exception_rule row %d %s: invalid mcp_reference %s" % (row.row_idx, exception_id, normalize_key(values.get("mcp_reference"))))
        if normalize_key(values.get("cross_clock_mcp_review")) and normalize_key(values.get("cross_clock_mcp_review")) not in CROSS_CLOCK_MCP_REVIEW:
            report.error("exception_rule row %d %s: invalid cross_clock_mcp_review %s" % (row.row_idx, exception_id, normalize_key(values.get("cross_clock_mcp_review"))))
        if datapath_only and datapath_only not in DATAPATH_VALUES:
            report.error("exception_rule row %d %s: datapath_only must be yes/no" % (row.row_idx, exception_id))
        if tool_surface and tool_surface not in TOOL_SURFACES:
            report.error("exception_rule row %d %s: invalid tool_surface %s" % (row.row_idx, exception_id, tool_surface))
        if timing_status and timing_status not in TIMING_CONTRACT_STATUSES:
            report.error("exception_rule row %d %s: invalid timing_contract_status %s" % (row.row_idx, exception_id, timing_status))
        if clock_context and clock_context not in CLOCK_CONTEXT_STATUSES:
            report.error("exception_rule row %d %s: invalid harden_clock_context_status %s" % (row.row_idx, exception_id, clock_context))

        if apply_value == "yes" and review_status != "approved":
            report.error("exception_rule row %d %s: apply=yes requires review_status=approved" % (row.row_idx, exception_id))

        if apply_value == "yes":
            if not etype or etype not in EMITTED_EXCEPTION_TYPES:
                report.error("exception_rule row %d %s: exception_type must be an emitted 30 type" % (row.row_idx, exception_id))
            if not clean_cell(values.get("owner")):
                report.error("exception_rule row %d %s: owner is required" % (row.row_idx, exception_id))
            if not clean_cell(values.get("basis")):
                report.error("exception_rule row %d %s: basis is required" % (row.row_idx, exception_id))
            if not path_category or path_category == "unknown":
                report.error("exception_rule row %d %s: path_category must be reviewed before emit" % (row.row_idx, exception_id))
            if source_type == "missing_timing_candidate" and timing_status == "no_port_timing":
                report.error("exception_rule row %d %s: no_port_timing/missing_timing_candidate alone cannot generate 30" % (row.row_idx, exception_id))
            if not (src or dst):
                report.error("exception_rule row %d %s: from/to endpoint is required" % (row.row_idx, exception_id))
            if src and collection_has_get_clocks(src) or dst and collection_has_get_clocks(dst) or through and collection_has_get_clocks(through):
                report.error("exception_rule row %d %s: 30 must be object-level; get_clocks collection is not allowed" % (row.row_idx, exception_id))
            if collection_has_range_or_bus(src) or collection_has_range_or_bus(dst) or collection_has_range_or_bus(through):
                report.error("exception_rule row %d %s: endpoint collection contains bus/range or non-canonical object" % (row.row_idx, exception_id))
            if channel and channel.is_pad_related == "yes":
                report.error("exception_rule row %d %s: pad-related path belongs to 04, not 30" % (row.row_idx, exception_id))
            if channel and channel.is_clock_related == "yes":
                report.error("exception_rule row %d %s: clock relationship belongs to 03, not 30" % (row.row_idx, exception_id))
            if channel and channel.related_10_feedthrough_id and not clean_cell(values.get("related_10_feedthrough_id")):
                report.error("exception_rule row %d %s: path passes feedthrough but related_10_feedthrough_id is blank" % (row.row_idx, exception_id))
            related_ft_ids = [part.strip() for part in clean_cell(values.get("related_10_feedthrough_id")).split(",") if part.strip()]
            if related_ft_ids:
                for ft_id in related_ft_ids:
                    if ft_id not in ft_ids:
                        report.error("exception_rule row %d %s: related_10_feedthrough_id %s not found in feedthrough_inventory" % (row.row_idx, exception_id, ft_id))
                if not through:
                    report.error("exception_rule row %d %s: feedthrough exception requires through_collection anchor" % (row.row_idx, exception_id))
            if etype in {"false_path", "multicycle_path"} and datapath_only == "yes":
                report.error("exception_rule row %d %s: datapath_only only applies to set_max_delay/set_min_delay" % (row.row_idx, exception_id))
            if etype == "max_delay_override" and not clean_cell(values.get("max_value")):
                report.error("exception_rule row %d %s: max_delay_override requires max_value" % (row.row_idx, exception_id))
            if etype == "min_delay_override" and not clean_cell(values.get("min_value")):
                report.error("exception_rule row %d %s: min_delay_override requires min_value" % (row.row_idx, exception_id))
            if etype == "max_min_delay_override" and (not clean_cell(values.get("max_value")) or not clean_cell(values.get("min_value"))):
                report.error("exception_rule row %d %s: max_min_delay_override requires both max_value and min_value" % (row.row_idx, exception_id))
            if etype == "multicycle_path":
                validate_multicycle_row(row, exception_id, check_type, report)
            if etype in {"max_delay_override", "min_delay_override", "max_min_delay_override"}:
                validate_datapath_strategy(row, exception_id, clock_relation, path_category, datapath_only, report)
            if path_category == "reset" and etype == "false_path":
                if not reset_false_path_basis_ok(values):
                    report.error("exception_rule row %d %s: reset false_path requires recovery/removal/RDC/waiver basis" % (row.row_idx, exception_id))
            if source_type == "extracted_harden_exception" and clock_context not in {"matched", "remapped_equivalent"}:
                report.error("exception_rule row %d %s: extracted harden exception requires matched/remapped_equivalent clock context" % (row.row_idx, exception_id))
            expiry = parse_review_date(clean_cell(values.get("expiry_or_review_date")))
            if expiry is not None and expiry < today_date():
                report.error("exception_rule row %d %s: expiry_or_review_date %s is expired" % (row.row_idx, exception_id, expiry.isoformat()))
            if tool_surface in {"dc", "both"} and etype in {"max_delay_override", "min_delay_override", "max_min_delay_override"} and datapath_only == "yes":
                basis = normalize_key(values.get("basis")) + " " + normalize_key(values.get("note"))
                if "dc" not in basis and "synthesis" not in basis and "tool" not in basis:
                    report.error("exception_rule row %d %s: DC/both datapath_only use requires tool support basis" % (row.row_idx, exception_id))
            if not commands_for_row(row):
                report.error("exception_rule row %d %s: approved row cannot generate any SDC command" % (row.row_idx, exception_id))
            check_source_digest(row, current_digests, report)

        if row in assembled:
            path_key = (
                clean_cell(values.get("channel_id")) or src,
                src,
                dst,
            )
            by_path[path_key].append(row)
            warn_row_shape(row, clock_relation, report)
            warn_active20_overlap(row, active20.get(clean_cell(values.get("related_20_channel_id")) or clean_cell(values.get("channel_id"))), report)

    for path_key, group in by_path.items():
        check_assembled_conflict(path_key, group, active20, report)


def validate_multicycle_row(row: FormRow, exception_id: str, check_type: str, report: Report) -> None:
    values = row.values
    setup = clean_cell(values.get("setup_cycles"))
    hold = clean_cell(values.get("hold_cycles"))
    if check_type == "both":
        if not setup or not hold:
            report.error("exception_rule row %d %s: multicycle check_type=both requires setup_cycles and hold_cycles" % (row.row_idx, exception_id))
    elif check_type == "setup":
        if not setup:
            report.error("exception_rule row %d %s: multicycle check_type=setup requires setup_cycles" % (row.row_idx, exception_id))
    elif check_type == "hold":
        if not hold:
            report.error("exception_rule row %d %s: multicycle check_type=hold requires hold_cycles" % (row.row_idx, exception_id))
    src_clk = clean_cell(values.get("src_clock"))
    dst_clk = clean_cell(values.get("dst_clock"))
    if src_clk and dst_clk and src_clk != dst_clk:
        if normalize_key(values.get("mcp_reference")) not in {"start", "end"}:
            report.error("exception_rule row %d %s: cross-clock multicycle requires mcp_reference=start/end" % (row.row_idx, exception_id))
        if normalize_key(values.get("cross_clock_mcp_review")) not in {"approved", "reviewed", "yes"}:
            report.error("exception_rule row %d %s: cross-clock multicycle requires cross_clock_mcp_review=approved" % (row.row_idx, exception_id))
    if check_type == "setup":
        report.warn("exception_rule row %d %s: setup-only multicycle must justify missing hold MCP" % (row.row_idx, exception_id))


def validate_datapath_strategy(row: FormRow, exception_id: str, clock_relation: str, path_category: str, datapath_only: str, report: Report) -> None:
    etype = row_exception_type(row)
    if clock_relation in ASYNC_RELATIONS and path_category in CDC_WINDOW_CATEGORIES:
        if datapath_only != "yes":
            report.error("exception_rule row %d %s: CDC/async max/min override requires datapath_only=yes" % (row.row_idx, exception_id))
        if etype in {"max_delay_override", "min_delay_override"}:
            report.warn("exception_rule row %d %s: CDC window usually needs max_min_delay_override pair; justify single-sided override" % (row.row_idx, exception_id))
    elif clock_relation in {"synchronous", "unknown", ""} and datapath_only == "yes":
        report.warn("exception_rule row %d %s: synchronous/unknown max/min override uses datapath_only=yes; confirm skew/latency should be excluded" % (row.row_idx, exception_id))


def warn_row_shape(row: FormRow, clock_relation: str, report: Report) -> None:
    values = row.values
    exception_id = clean_cell(values.get("exception_id")) or "row_%d" % row.row_idx
    src, dst = endpoint_present(row)
    through = clean_cell(values.get("through_collection"))
    etype = row_exception_type(row)
    if bool(src) != bool(dst):
        report.warn("exception_rule row %d %s: single-sided endpoint may be broader than intended" % (row.row_idx, exception_id))
    if collection_has_get_nets(through) or collection_has_pattern(through):
        report.warn("exception_rule row %d %s: through_collection uses nets/pattern; verify object matching across stages" % (row.row_idx, exception_id))
    text = " ".join(clean_cell(values.get(field)) for field in ("exception_id", "channel_id", "from_collection", "to_collection", "through_collection", "case_condition", "note"))
    if row_scenario(row) == "common" and any(token in normalize_key(text) for token in MODE_SPECIFIC_TOKENS):
        report.warn("exception_rule row %d %s: common exception looks mode-specific; confirm scenario placement" % (row.row_idx, exception_id))
    if clock_relation in ASYNC_RELATIONS and etype == "false_path":
        report.warn("exception_rule row %d %s: false_path may be redundant with 03 async/exclusive relation" % (row.row_idx, exception_id))
    if clock_relation not in ASYNC_RELATIONS and etype in {"max_delay_override", "min_delay_override", "max_min_delay_override"}:
        basis = normalize_key(values.get("basis")) + " " + normalize_key(values.get("cdc_rdc_ref")) + " " + normalize_key(values.get("protocol_ref"))
        if "cdc" in basis or "async" in basis:
            report.warn("exception_rule row %d %s: CDC/async basis without async/exclusive clock_relation; check 03" % (row.row_idx, exception_id))
    if normalize_key(values.get("path_category")) == "reset":
        report.warn("exception_rule row %d %s: reset path exception requires recovery/removal/RDC review" % (row.row_idx, exception_id))


def warn_active20_overlap(row: FormRow, active: Optional[Active20Budget], report: Report) -> None:
    if not active:
        return
    etype = row_exception_type(row)
    exception_id = clean_cell(row.values.get("exception_id")) or "row_%d" % row.row_idx
    if active.max_active and etype == "min_delay_override":
        report.warn("exception_rule row %d %s: active 20 max coexists with 30 min override; basis must explain hold/min source" % (row.row_idx, exception_id))
    if active.min_active and etype == "max_delay_override":
        report.warn("exception_rule row %d %s: active 20 min coexists with 30 max override; basis must explain setup/max source" % (row.row_idx, exception_id))


def check_assembled_conflict(path_key: Tuple[str, str, str], group: Sequence[FormRow], active20: Dict[str, Active20Budget], report: Report) -> None:
    if len(group) <= 1:
        for row in group:
            check_active20_conflict(row, active20, report)
        return
    rows_text = ", ".join(str(row.row_idx) for row in group)
    types = [row_exception_type(row) for row in group]
    if "false_path" in types and any(t in {"multicycle_path", "max_delay_override", "min_delay_override", "max_min_delay_override"} for t in types):
        report.error("assembled view conflict for path %s: false_path overlaps other exception rows %s" % (path_key[0], rows_text))
    max_values = {clean_cell(row.values.get("max_value")) for row in group if row_exception_type(row) in {"max_delay_override", "max_min_delay_override"}}
    min_values = {clean_cell(row.values.get("min_value")) for row in group if row_exception_type(row) in {"min_delay_override", "max_min_delay_override"}}
    if len(max_values - {""}) > 1:
        report.error("assembled view conflict for path %s: conflicting max overrides in rows %s" % (path_key[0], rows_text))
    if len(min_values - {""}) > 1:
        report.error("assembled view conflict for path %s: conflicting min overrides in rows %s" % (path_key[0], rows_text))
    if "multicycle_path" in types and any(t in {"max_delay_override", "min_delay_override", "max_min_delay_override"} for t in types):
        report.error("assembled view conflict for path %s: multicycle overlaps max/min override rows %s" % (path_key[0], rows_text))
    for row in group:
        check_active20_conflict(row, active20, report)


def check_active20_conflict(row: FormRow, active20: Dict[str, Active20Budget], report: Report) -> None:
    values = row.values
    channel_id = clean_cell(values.get("related_20_channel_id")) or clean_cell(values.get("channel_id"))
    active = active20.get(channel_id)
    if not active:
        return
    etype = row_exception_type(row)
    check_type = normalize_key(values.get("check_type")) or "both"
    setup_dim, hold_dim = check_dimensions_for_exception(etype, check_type)
    exception_id = clean_cell(values.get("exception_id")) or "row_%d" % row.row_idx
    basis = clean_cell(values.get("basis"))
    if active.max_active and setup_dim:
        report.error("exception_rule row %d %s: active 20 max budget overlaps 30 setup/max exception for channel %s" % (row.row_idx, exception_id, channel_id))
    if active.min_active and hold_dim:
        report.error("exception_rule row %d %s: active 20 min budget overlaps 30 hold/min exception for channel %s" % (row.row_idx, exception_id, channel_id))
    if etype == "max_delay_override" and active.max_active and not basis:
        report.error("exception_rule row %d %s: active 20 max and 30 max override need explicit adjudication basis" % (row.row_idx, exception_id))
    if etype == "min_delay_override" and active.min_active and not basis:
        report.error("exception_rule row %d %s: active 20 min and 30 min override need explicit adjudication basis" % (row.row_idx, exception_id))


def output_sdc_path(cwd: Path, scenario: str, stage: str, corner: str) -> Path:
    if scenario == "common":
        if stage == "all" and corner == "all":
            return cwd / "common/30_harden_to_harden_exception.sdc"
        return cwd / ("common/30_harden_to_harden_exception_%s_%s.sdc" % (stage, safe_filename_token(corner)))
    if stage == "all" and corner == "all":
        return cwd / ("scenarios/%s_exceptions.sdc" % scenario)
    return cwd / ("scenarios/%s_exceptions_%s_%s.sdc" % (scenario, stage, safe_filename_token(corner)))


def report_path(cwd: Path, scenario: str, stage: str, corner: str) -> Path:
    return cwd / ("harden_to_harden_exception_check_report_%s_%s_%s.txt" % (scenario, stage, safe_filename_token(corner)))


def format_source_refs(source_files, source_lines) -> str:
    files = [part.strip() for part in clean_cell(source_files).split(";") if part.strip()]
    lines = [part.strip() for part in clean_cell(source_lines).split(";") if part.strip()]
    if not files:
        return ""
    if len(files) == len(lines):
        return "; ".join("%s:%s" % (source, line) if line else source for source, line in zip(files, lines))
    if lines:
        return "%s lines %s" % ("; ".join(files), "; ".join(lines))
    return "; ".join(files)


def through_clauses(value: str) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    if "-through" in text:
        return text
    parts = [part.strip() for part in re.split(r"\s*;\s*|\s*\|\s*", text) if part.strip()]
    return " ".join("-through " + part for part in parts)


def mcp_reference_flag(value: str) -> str:
    text = normalize_key(value)
    if text in {"start", "end"}:
        return "-" + text
    return ""


def commands_for_row(row: FormRow) -> List[str]:
    values = row.values
    etype = row_exception_type(row)
    src, dst = endpoint_present(row)
    through = through_clauses(clean_cell(values.get("through_collection")))
    from_clause = ("-from " + src) if src else ""
    to_clause = ("-to " + dst) if dst else ""
    middle = " ".join(part for part in (from_clause, through, to_clause) if part)
    if not middle:
        return []
    check_type = normalize_key(values.get("check_type")) or "both"
    commands: List[str] = []
    if etype == "false_path":
        check = ""
        if check_type in {"setup", "hold"}:
            check = "-" + check_type + " "
        commands.append("set_false_path %s%s" % (check, middle))
    elif etype == "multicycle_path":
        ref = mcp_reference_flag(clean_cell(values.get("mcp_reference")))
        ref_text = (" " + ref) if ref else ""
        setup = clean_cell(values.get("setup_cycles"))
        hold = clean_cell(values.get("hold_cycles"))
        if check_type in {"both", "setup"} and setup:
            commands.append("set_multicycle_path %s -setup%s %s" % (format_number(setup), ref_text, middle))
        if check_type in {"both", "hold"} and hold:
            commands.append("set_multicycle_path %s -hold%s %s" % (format_number(hold), ref_text, middle))
    elif etype in {"max_delay_override", "max_min_delay_override"}:
        value = clean_cell(values.get("max_value"))
        if value:
            datapath = " -datapath_only" if normalize_key(values.get("datapath_only")) == "yes" else ""
            commands.append("set_max_delay %s%s %s" % (format_number(value), datapath, middle))
        if etype == "max_min_delay_override":
            value = clean_cell(values.get("min_value"))
            if value:
                datapath = " -datapath_only" if normalize_key(values.get("datapath_only")) == "yes" else ""
                commands.append("set_min_delay %s%s %s" % (format_number(value), datapath, middle))
    elif etype == "min_delay_override":
        value = clean_cell(values.get("min_value"))
        if value:
            datapath = " -datapath_only" if normalize_key(values.get("datapath_only")) == "yes" else ""
            commands.append("set_min_delay %s%s %s" % (format_number(value), datapath, middle))
    return commands


def generate_sdc(rows: Sequence[FormRow], scenario: str, stage: str, corner: str) -> List[str]:
    selected = [row for row in rows if row_selected_for_output(row, scenario, stage, corner) and is_apply_approved(row)]
    lines = [
        "################################################################################",
        "# Auto-generated SoC harden-to-harden exception constraints for scenario: %s, stage: %s, corner: %s" % (scenario, stage, corner),
        "# Source: 30_harden_to_harden_exception.xlsx exception_rule sheet",
        "# Only apply=yes and review_status=approved rows are emitted.",
        "################################################################################",
        "",
    ]
    emitted = 0
    for row in selected:
        commands = commands_for_row(row)
        if not commands:
            continue
        values = row.values
        lines.append("# row %d: %s" % (row.row_idx, clean_cell(values.get("exception_id"))))
        if clean_cell(values.get("basis")):
            lines.append("# Basis: " + clean_cell(values.get("basis")))
        for ref_header, label in (("cdc_rdc_ref", "CDC/RDC"), ("sta_waiver_ref", "STA waiver"), ("protocol_ref", "Protocol")):
            if clean_cell(values.get(ref_header)):
                lines.append("# %s: %s" % (label, clean_cell(values.get(ref_header))))
        source_refs = format_source_refs(values.get("source_sdc_file"), values.get("source_line"))
        if source_refs:
            lines.append("# Source SDC: " + source_refs)
        lines.extend(commands)
        lines.append("")
        emitted += len(commands)
    if emitted == 0:
        lines.append("# No harden-to-harden exception commands emitted for selected scenario/stage/corner.")
    return lines


def build_coverage_lines(rows: Sequence[FormRow], channels: Sequence[ChannelRecord], active20: Dict[str, Active20Budget], scenario: str, stage: str, corner: str) -> List[str]:
    selected = [row for row in rows if row_selected_for_output(row, scenario, stage, corner) and is_apply_approved(row)]
    assembled = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner) and is_apply_approved(row)]
    type_counts: Dict[str, int] = defaultdict(int)
    for row in assembled:
        type_counts[row_exception_type(row)] += 1
    no_decision = []
    approved_channels = {clean_cell(row.values.get("channel_id")) for row in assembled}
    for ch in channels:
        if ch.channel_type == "harden_to_harden" and ch.channel_id not in approved_channels and ch.channel_id not in active20:
            no_decision.append(ch.channel_id)
    lines = [
        "",
        "Coverage:",
        "  inventory channels      : %d" % len(channels),
        "  active 20 budget channel: %d" % len(active20),
        "  assembled approved row  : %d" % len(assembled),
        "  emitted approved row    : %d" % len(selected),
        "  exception type counts   : %s" % (", ".join("%s=%d" % item for item in sorted(type_counts.items())) or "none"),
        "  no 20/30 decision channel count: %d" % len(no_decision),
    ]
    if no_decision:
        lines.append("  no 20/30 decision samples:")
        lines.extend("    - " + item for item in no_decision[:20])
        if len(no_decision) > 20:
            lines.append("    ... %d more" % (len(no_decision) - 20))
    return lines


def write_report(
    path: Path,
    report: Report,
    scenario: str,
    stage: str,
    corner: str,
    form_path: Path,
    output_path: Path,
    coverage_lines: Sequence[str],
) -> None:
    lines = [
        "30_harden_to_harden_exception extraction report",
        "================================================",
        "",
        "Scenario: " + scenario,
        "Stage   : " + stage,
        "Corner  : " + corner,
        "Form    : " + str(form_path),
        "Output  : " + str(output_path),
        "Warnings: %d" % report.warning_count,
        "Errors  : %d" % report.error_count,
        "Sync changed: %s" % ("yes" if report.sync_changed else "no"),
        "",
        "Messages:",
    ]
    lines.extend(report.lines or ["INFO: no messages"])
    lines.extend(coverage_lines)
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def pending_line_key(line: str) -> Optional[Tuple[str, str]]:
    stripped = line.strip()
    if not stripped or stripped.startswith("#"):
        return None
    parts = stripped.split()
    if len(parts) < 2:
        return None
    direction, port = parts[0], parts[1]
    if direction not in {"input", "output", "inout"}:
        return None
    return direction, port


def removed_line_key(line: str) -> Optional[PortKey]:
    stripped = line.strip()
    if not stripped or stripped.startswith("#"):
        return None
    parts = stripped.split()
    if len(parts) < 3:
        return None
    inst_name, direction, port = parts[:3]
    if direction not in {"input", "output", "inout"}:
        return None
    return PortKey(inst_name, direction, port)


def read_removed_keys(log_dir: Path) -> Set[PortKey]:
    keys: Set[PortKey] = set()
    if not log_dir.is_dir():
        return keys
    for path in sorted(log_dir.glob("*.removed")):
        try:
            lines = path.read_text(encoding="utf-8").splitlines()
        except OSError:
            continue
        for line in lines:
            key = removed_line_key(line)
            if key is not None:
                keys.add(key)
    return keys


def harden_pending_key(inst_name: str, direction: str, port_name: str) -> Optional[PortKey]:
    if not inst_name or normalize_key(inst_name) in {"top", "fabric", "unknown", "constant", "const"}:
        return None
    if direction not in {"input", "output", "inout"}:
        return None
    if not port_name:
        return None
    return PortKey(inst_name, direction, port_name)


def emitted_rows_for_pending(rows: Sequence[FormRow], scenario: str, stage: str, corner: str) -> List[FormRow]:
    result = []
    for row in rows:
        if not row_selected_for_output(row, scenario, stage, corner) or not is_apply_approved(row):
            continue
        if commands_for_row(row):
            result.append(row)
    return result


def removed_log_line_30(row: FormRow, ch: ChannelRecord, key: PortKey) -> str:
    values = row.values
    return " ".join(
        [
            key.inst_name,
            key.direction,
            key.port_name,
            "covered_by=30_harden_to_harden_exception",
            "reason=path_exception",
            "exception_id=%s" % clean_cell(values.get("exception_id")),
            "exception_type=%s" % normalize_key(values.get("exception_type")),
            "channel=%s" % ch.channel_id,
            "scenario=%s" % row_scenario(row),
            "stage=%s" % row_stage(row),
            "corner=%s" % row_corner(row),
        ]
    )


def update_pending_for_30(
    cwd: Path,
    pending_root: Path,
    rows: Sequence[FormRow],
    channels: Sequence[ChannelRecord],
    scenario: str,
    stage: str,
    corner: str,
    report: Report,
) -> None:
    pending_dir = pending_root / "pending"
    if not pending_dir.exists():
        return
    if not pending_dir.is_dir():
        report.error("%s: pending path exists but is not a directory" % pending_dir)
        return
    channel_by_id = {ch.channel_id: ch for ch in channels}
    removals: List[Tuple[FormRow, ChannelRecord, PortKey]] = []
    for row in emitted_rows_for_pending(rows, scenario, stage, corner):
        ch = channel_by_id.get(clean_cell(row.values.get("channel_id")))
        if not ch or ch.channel_type != "harden_to_harden":
            continue
        src_key = harden_pending_key(ch.src_instance, ch.src_direction or "output", ch.src_port)
        dst_key = harden_pending_key(ch.dst_instance, ch.dst_direction or "input", ch.dst_port)
        for key in (src_key, dst_key):
            if key is not None:
                removals.append((row, ch, key))
    if not removals:
        return

    log_dir = pending_root / "removed_log"
    previous_removed = read_removed_keys(log_dir)
    by_inst: Dict[str, List[Tuple[FormRow, ChannelRecord, PortKey]]] = defaultdict(list)
    for item in removals:
        by_inst[item[2].inst_name].append(item)

    removed_items: List[Tuple[FormRow, ChannelRecord, PortKey]] = []
    for inst_name, inst_items in sorted(by_inst.items()):
        pending_file = pending_dir / (inst_name + ".ports")
        if not pending_file.is_file():
            for row, ch, key in inst_items:
                if key in previous_removed:
                    continue
                report.error("%s: missing pending file for 30 channel endpoint %s/%s" % (pending_file, key.inst_name, key.port_name))
            continue
        lines = pending_file.read_text(encoding="utf-8").splitlines()
        index: Dict[Tuple[str, str], int] = {}
        duplicate_keys: Set[Tuple[str, str]] = set()
        for idx, line in enumerate(lines):
            key = pending_line_key(line)
            if key is None:
                continue
            if key in index:
                duplicate_keys.add(key)
            else:
                index[key] = idx
        for direction, port in sorted(duplicate_keys):
            report.error("%s: duplicate pending port line %s %s" % (pending_file, direction, port))

        remove_line_indices: Set[int] = set()
        for row, ch, port_key in inst_items:
            key = (port_key.direction, port_key.port_name)
            if key not in index:
                if port_key in previous_removed:
                    continue
                report.error(
                    "%s: 30 wants to remove %s %s, but it is not present in pending and no previous_removed record exists"
                    % (pending_file, port_key.direction, port_key.port_name)
                )
                continue
            remove_line_indices.add(index[key])
            removed_items.append((row, ch, port_key))
        if remove_line_indices:
            kept = [line for idx, line in enumerate(lines) if idx not in remove_line_indices]
            pending_file.write_text("\n".join(kept).rstrip() + ("\n" if kept else ""), encoding="utf-8")

    if removed_items:
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / "30_harden_to_harden_exception.removed"
        existing_lines = log_path.read_text(encoding="utf-8").splitlines() if log_path.is_file() else []
        existing_keys = {key for key in (removed_line_key(line) for line in existing_lines) if key is not None}
        new_lines = []
        for row, ch, key in sorted(removed_items, key=lambda item: (item[2].inst_name, item[2].direction, item[2].port_name)):
            if key not in existing_keys:
                new_lines.append(removed_log_line_30(row, ch, key))
                existing_keys.add(key)
        if new_lines:
            log_lines = [line for line in existing_lines if line.strip()] + new_lines
            log_path.write_text("\n".join(log_lines).rstrip() + "\n", encoding="utf-8")
        try:
            display_path = log_path.relative_to(cwd)
        except ValueError:
            display_path = log_path
        report.info("removed %d harden exception endpoint(s) from pending; log=%s" % (len(removed_items), display_path))


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate 30 SoC harden-to-harden exception SDC.")
    parser.add_argument("-scenario", "--scenario", default="common", choices=sorted(SCENARIOS), help="target scenario")
    parser.add_argument("-stage", "--stage", default="all", choices=sorted(STAGES), help="target stage/view")
    parser.add_argument("-corner", "--corner", default="all", help="target corner/view")
    parser.add_argument("--info-all", default="info_all.xlsx", help="integration summary xlsx")
    parser.add_argument("--form", default="30_harden_to_harden_exception.xlsx", help="30 exception workbook")
    parser.add_argument("--connection-inventory", default="00_harden_port_inventory/connection_inventory.csv", help="00 bit-to-bit connection inventory")
    parser.add_argument("--feedthrough-inventory", default="feedthrough_inventory.csv", help="10 feedthrough inventory CSV")
    parser.add_argument("--twenty-form", default="20_harden_x_if.xlsx", help="20 workbook used for active budget overlap checks")
    parser.add_argument("--pending-root", default="00_harden_port_inventory", help="00 pending inventory root")
    parser.add_argument("--no-update-pending", action="store_true", help="do not remove emitted exception endpoints from pending")
    parser.add_argument("--force-generate-after-sync", action="store_true", help="generate SDC even if workbook was synchronized")
    parser.add_argument("--report", help="output report path")
    return parser.parse_args(argv)


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)
    cwd = Path.cwd()
    report = Report()

    info_all = cwd / args.info_all
    if not info_all.is_file():
        raise RuntimeError("integration file not found: %s" % info_all)

    instances = read_info_all(info_all, report)
    port_paths = default_port_workbooks(cwd, Path(args.info_all).name, Path(args.form).name, report)
    port_sheets = read_port_workbooks(port_paths, report)
    attach_port_data(instances, port_sheets, report)
    resolve_sdc_paths(instances, cwd, report)
    current_digests = collect_current_sdc_digests(instances)

    connections = read_connection_inventory(cwd / args.connection_inventory, report)
    feedthroughs = read_feedthrough_inventory(cwd / args.feedthrough_inventory, report)
    channels = build_channels_from_inventories(instances, connections, feedthroughs, report)
    active20 = read_20_active_budgets(cwd / args.twenty_form, args.scenario, args.stage, args.corner, report)
    delays, exceptions = extract_sdc_evidence(instances, report)
    candidate_seeds, rule_seeds, log_items = create_candidate_and_rule_seeds(channels, delays, exceptions, active20, report)

    form_path = cwd / args.form
    sync_workbook(form_path, candidate_seeds, rule_seeds, log_items, report)

    rows = read_rule_rows(form_path)
    validate_rows(rows, channels, active20, feedthroughs, args.scenario, args.stage, args.corner, current_digests, report)

    output_path = output_sdc_path(cwd, args.scenario, args.stage, args.corner)
    rpt_path = Path(args.report) if args.report else report_path(cwd, args.scenario, args.stage, args.corner)

    generated = False
    if report.sync_changed and not args.force_generate_after_sync:
        report.warn("workbook changed during sync; review 30_harden_to_harden_exception.xlsx before SDC generation")
    elif report.error_count == 0:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("\n".join(generate_sdc(rows, args.scenario, args.stage, args.corner)).rstrip() + "\n", encoding="utf-8")
        report.info("wrote %s" % output_path)
        generated = True
    else:
        report.warn("SDC generation skipped because errors were reported")

    if generated and not args.no_update_pending and args.pending_root:
        update_pending_for_30(cwd, cwd / args.pending_root, rows, channels, args.scenario, args.stage, args.corner, report)

    coverage_lines = build_coverage_lines(rows, channels, active20, args.scenario, args.stage, args.corner)
    write_report(rpt_path, report, args.scenario, args.stage, args.corner, form_path, output_path, coverage_lines)
    print("Report: %s" % rpt_path)
    print("Warnings: %d  Errors: %d  Sync changed: %s" % (report.warning_count, report.error_count, report.sync_changed))
    if report.error_count:
        return 1
    if report.sync_changed and not args.force_generate_after_sync:
        return 1
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except RuntimeError as exc:
        print("ERROR: %s" % exc, file=sys.stderr)
        raise SystemExit(2)
