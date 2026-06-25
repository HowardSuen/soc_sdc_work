#!/usr/bin/env python3
"""
Build and run a complex 00 -> 30 SoC SDC flow case.

The case has five hardens. Each harden starts with at least twenty pending
canonical scalar/bit ports. The script creates 00 pending/connection inventory
directly, then runs 01/02/03/04/10/20/30 in order and approves a controlled set
of rows so every stage emits representative artifacts.
"""

import csv
import json
import shutil
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
SOC = BASE.parent
WORK = BASE / "work" / "full_chain_00_to_30_complex"

EX01 = SOC / "01_soc_clocks" / "01_extract_soc_clocks.py"
EX02 = SOC / "02_soc_clock_timing" / "02_extract_soc_clock_timing.py"
EX03 = SOC / "03_soc_clock_groups" / "03_extract_soc_clock_groups.py"
EX04 = SOC / "04_soc_io_pads" / "04_extract_soc_io_pads.py"
EX10 = SOC / "10_feedthrough" / "10_extract_feedthrough.py"
EX20 = SOC / "20_harden_x_if" / "20_extract_harden_x_if.py"
EX30 = SOC / "30_harden_to_harden_exception" / "30_extract_harden_to_harden_exception.py"

REQ = [
    "Parameter",
    "Inout",
    "Inout Width",
    "Inout Connectivity",
    "Inout Name",
    "Input",
    "Input Width",
    "Input Used Width",
    "From Whom",
    "Output",
    "Output Width",
    "Output Used Width",
    "To Top",
]

INFO = [
    {"module_name": "clkgen", "inst_name": "u_clkgen", "owner": "clock_owner"},
    {"module_name": "dma", "inst_name": "u_dma", "owner": "dma_owner"},
    {"module_name": "dpg", "inst_name": "u_dpg", "owner": "dpg_owner"},
    {"module_name": "periph", "inst_name": "u_periph", "owner": "periph_owner"},
    {"module_name": "ctrl", "inst_name": "u_ctrl", "owner": "ctrl_owner"},
]


def sh(cmd, cwd):
    result = subprocess.run(
        [sys.executable] + [str(part) for part in cmd],
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )
    return result


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def run_ok(cmd, cwd):
    result = sh(cmd, cwd)
    require(result.returncode == 0, "command failed: %s\nSTDOUT:\n%s\nSTDERR:\n%s" % (" ".join(map(str, cmd)), result.stdout, result.stderr))
    return result


def run_sync(cmd, cwd):
    result = sh(cmd, cwd)
    require(result.returncode == 1, "sync/review command should return 1: %s\nSTDOUT:\n%s\nSTDERR:\n%s" % (" ".join(map(str, cmd)), result.stdout, result.stderr))
    return result


def port_sheet(rows):
    df = pd.DataFrame(rows)
    for col in REQ:
        if col not in df.columns:
            df[col] = ""
    return df[REQ].fillna("")


def row(Input="", Inout="", Output="", **kwargs):
    item = {"Input": Input, "Inout": Inout, "Output": Output}
    item.update(kwargs)
    return item


def build_ports():
    data = {}

    data["u_clkgen"] = [
        row(Input="ref_clk_i", **{"Input Width": 1, "From Whom": "top.ref_clk_pad"}),
        row(Input="scan_clk_i", **{"Input Width": 1, "From Whom": "top.scan_clk_pad"}),
        row(Input="peri_ref_i", **{"Input Width": 1, "From Whom": "top.peri_ref_pad"}),
        row(Input="pll_bypass_i", **{"Input Width": 1, "From Whom": "top.pll_bypass_pad"}),
        row(Input="cfg_boot_i", **{"Input Width": 1, "From Whom": "u_ctrl.boot_mode_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Output="core_clk_o", **{"Output Width": 1}),
        row(Output="bus_clk_o", **{"Output Width": 1}),
        row(Output="peri_clk_o", **{"Output Width": 1}),
        row(Output="scan_clk_o", **{"Output Width": 1}),
        row(Output="dbg_clk_o", **{"Output Width": 1}),
    ]
    for bit in range(8):
        data["u_clkgen"].append(row(Input="cfg_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_ctrl.clk_cfg_o[%d]" % bit}))
    for bit in range(8):
        data["u_clkgen"].append(row(Output="status_o[%d]" % bit, **{"Output Width": 1}))

    data["u_dma"] = [
        row(Input="clk_i", **{"Input Width": 1, "From Whom": "u_clkgen.core_clk_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Input="irq_ack_i", **{"Input Width": 1, "From Whom": "u_ctrl.irq_ack_o"}),
        row(Input="test_mode_i", **{"Input Width": 1, "From Whom": "top.test_mode_pad"}),
        row(Output="irq_o", **{"Output Width": 1, "To Top": "irq_pad"}),
        row(Output="done_o", **{"Output Width": 1}),
        row(Output="req_valid_o", **{"Output Width": 1}),
    ]
    for bit in range(4):
        data["u_dma"].append(row(Output="data_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dma"].append(row(Output="cfg_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dma"].append(row(Output="async_req_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dma"].append(row(Input="resp_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_periph.resp_o[%d]" % bit}))

    data["u_dpg"] = [
        row(Input="clk_i", **{"Input Width": 1, "From Whom": "u_clkgen.bus_clk_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Input="debug_sel_i", **{"Input Width": 1, "From Whom": "u_ctrl.debug_sel_o"}),
        row(Output="dpg_status_o", **{"Output Width": 1}),
    ]
    for bit in range(4):
        data["u_dpg"].append(row(Input="fti_0_dma2periph_data[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dma.data_o[%d]" % bit}))
    for bit in range(4):
        data["u_dpg"].append(row(Output="fto_0_dma2periph_data[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dpg"].append(row(Input="fti_0_dma2periph_async_req[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dma.async_req_o[%d]" % bit}))
    for bit in range(4):
        data["u_dpg"].append(row(Output="fto_0_dma2periph_async_req[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dpg"].append(row(Input="cfg_shadow_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_ctrl.ctrl_cfg_o[%d]" % bit}))

    data["u_periph"] = [
        row(Input="clk_i", **{"Input Width": 1, "From Whom": "u_clkgen.peri_clk_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Input="enable_i", **{"Input Width": 1, "From Whom": "u_ctrl.periph_enable_o"}),
        row(Output="ready_o", **{"Output Width": 1}),
        row(Output="event_o", **{"Output Width": 1, "To Top": "event_pad"}),
    ]
    for bit in range(4):
        data["u_periph"].append(row(Input="data_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dpg.fto_0_dma2periph_data[%d]" % bit}))
    for bit in range(4):
        data["u_periph"].append(row(Input="async_req_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dpg.fto_0_dma2periph_async_req[%d]" % bit}))
    for bit in range(4):
        data["u_periph"].append(row(Output="resp_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_periph"].append(row(Input="cfg_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dma.cfg_o[%d]" % bit}))

    data["u_ctrl"] = [
        row(Input="clk_i", **{"Input Width": 1, "From Whom": "u_clkgen.bus_clk_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Input="gpio_mode_i", **{"Input Width": 1, "From Whom": "top.gpio_mode_pad"}),
        row(Input="scan_en_i", **{"Input Width": 1, "From Whom": "top.scan_en_pad"}),
        row(Output="boot_mode_o", **{"Output Width": 1}),
        row(Output="irq_ack_o", **{"Output Width": 1}),
        row(Output="debug_sel_o", **{"Output Width": 1}),
        row(Output="periph_enable_o", **{"Output Width": 1}),
    ]
    for bit in range(8):
        data["u_ctrl"].append(row(Output="clk_cfg_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_ctrl"].append(row(Output="ctrl_cfg_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_ctrl"].append(row(Input="status_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_clkgen.status_o[%d]" % bit}))
    return data


def write_inputs(d):
    d.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(INFO).to_excel(d / "info_all.xlsx", index=False)
    ports = build_ports()
    with pd.ExcelWriter(d / "ports.xlsx", engine="xlsxwriter") as writer:
        for inst, rows in ports.items():
            port_sheet(rows).to_excel(writer, sheet_name=inst, index=False)

    (d / "virtual_clocks.csv").write_text(
        "clock_name,period,waveform,note\n"
        "v_ext_cfg,10.000,,external config virtual clock\n"
        "v_jtag,50.000,,JTAG virtual clock\n",
        encoding="utf-8",
    )

    (d / "clkgen.sdc").write_text(
        "create_clock -name ref_clk -period 10.000 [get_ports ref_clk_i]\n"
        "create_clock -name scan_clk -period 50.000 [get_ports scan_clk_i]\n"
        "create_clock -name peri_ref_clk -period 20.000 [get_ports peri_ref_i]\n"
        "create_generated_clock -name core_clk -source [get_ports ref_clk_i] -multiply_by 2 [get_ports core_clk_o]\n"
        "create_generated_clock -name bus_clk -source [get_ports ref_clk_i] -divide_by 1 [get_ports bus_clk_o]\n"
        "create_generated_clock -name peri_clk -source [get_ports peri_ref_i] -divide_by 2 [get_ports peri_clk_o]\n"
        "create_generated_clock -name scan_out_clk -source [get_ports scan_clk_i] -combinational [get_ports scan_clk_o]\n",
        encoding="utf-8",
    )
    (d / "dma.sdc").write_text(
        "create_clock -name dma_clk -period 5.000 [get_ports clk_i]\n"
        "set_output_delay -max 1.20 -clock dma_clk [get_ports data_o]\n"
        "set_output_delay -min 0.20 -clock dma_clk [get_ports data_o]\n"
        "set_output_delay -max 1.50 -clock dma_clk [get_ports cfg_o]\n"
        "set_false_path -from [get_ports cfg_o]\n",
        encoding="utf-8",
    )
    (d / "dpg.sdc").write_text(
        "create_clock -name dpg_clk -period 10.000 [get_ports clk_i]\n",
        encoding="utf-8",
    )
    (d / "periph.sdc").write_text(
        "create_clock -name periph_clk -period 20.000 [get_ports clk_i]\n"
        "set_input_delay -max 1.00 -clock periph_clk [get_ports data_i]\n"
        "set_input_delay -min 0.10 -clock periph_clk [get_ports data_i]\n"
        "set_input_delay -max 1.10 -clock periph_clk [get_ports cfg_i]\n"
        "set_load 0.35 [get_ports event_o]\n",
        encoding="utf-8",
    )
    (d / "ctrl.sdc").write_text(
        "create_clock -name ctrl_clk -period 10.000 [get_ports clk_i]\n"
        "set_input_transition 0.08 [get_ports gpio_mode_i]\n",
        encoding="utf-8",
    )


def endpoint_key(inst, direction, port):
    return "%s:%s:%s" % (inst, direction, port)


def soc_object(inst, port):
    if inst == "top":
        return port
    return "%s/%s" % (inst, port)


def bit(port):
    return port.split("[")[-1].rstrip("]") if "[" in port else ""


def conn_id(src_i, src_p, dst_i, dst_p):
    def tok(value):
        return value.replace("[", "_bit").replace("]", "")
    return "CONN_%s_%s__%s_%s" % (src_i, tok(src_p), dst_i, tok(dst_p))


def add_edge(rows, src_i, src_d, src_p, dst_i, dst_d, dst_p, ctype="harden_to_harden", note=""):
    rows.append(
        {
            "connection_id": conn_id(src_i, src_p, dst_i, dst_p),
            "connection_type": ctype,
            "src_instance": src_i,
            "src_direction": src_d,
            "src_port": src_p,
            "src_bit_index": bit(src_p),
            "src_endpoint_key": endpoint_key(src_i, src_d, src_p),
            "src_soc_object": soc_object(src_i, src_p),
            "dst_instance": dst_i,
            "dst_direction": dst_d,
            "dst_port": dst_p,
            "dst_bit_index": bit(dst_p),
            "dst_endpoint_key": endpoint_key(dst_i, dst_d, dst_p),
            "dst_soc_object": soc_object(dst_i, dst_p),
            "validation_status": "matched",
            "note": note,
        }
    )


def write_00_inventory(d):
    root = d / "00_harden_port_inventory"
    pending = root / "pending"
    pending.mkdir(parents=True, exist_ok=True)
    port_data = build_ports()
    for inst, rows in port_data.items():
        lines = []
        for item in rows:
            if item.get("Input"):
                lines.append("input %s" % item["Input"])
            if item.get("Output"):
                lines.append("output %s" % item["Output"])
            if item.get("Inout"):
                lines.append("inout %s" % item["Inout"])
        require(len(lines) >= 20, "%s has only %d pending ports" % (inst, len(lines)))
        (pending / ("%s.ports" % inst)).write_text("\n".join(lines) + "\n", encoding="utf-8")

    rows = []
    add_edge(rows, "top", "output", "ref_clk_pad", "u_clkgen", "input", "ref_clk_i", "clock_connection")
    add_edge(rows, "top", "output", "scan_clk_pad", "u_clkgen", "input", "scan_clk_i", "clock_connection")
    add_edge(rows, "top", "output", "peri_ref_pad", "u_clkgen", "input", "peri_ref_i", "clock_connection")
    add_edge(rows, "top", "output", "pll_bypass_pad", "u_clkgen", "input", "pll_bypass_i", "top_pad_to_harden")
    add_edge(rows, "top", "output", "rst_n_pad", "u_clkgen", "input", "rst_n_i", "top_pad_to_harden")
    add_edge(rows, "top", "output", "rst_n_pad", "u_dma", "input", "rst_n_i", "top_pad_to_harden")
    add_edge(rows, "top", "output", "rst_n_pad", "u_dpg", "input", "rst_n_i", "top_pad_to_harden")
    add_edge(rows, "top", "output", "rst_n_pad", "u_periph", "input", "rst_n_i", "top_pad_to_harden")
    add_edge(rows, "top", "output", "rst_n_pad", "u_ctrl", "input", "rst_n_i", "top_pad_to_harden")
    add_edge(rows, "top", "output", "test_mode_pad", "u_dma", "input", "test_mode_i", "top_pad_to_harden")
    add_edge(rows, "top", "output", "gpio_mode_pad", "u_ctrl", "input", "gpio_mode_i", "top_pad_to_harden")
    add_edge(rows, "top", "output", "scan_en_pad", "u_ctrl", "input", "scan_en_i", "top_pad_to_harden")
    add_edge(rows, "u_dma", "output", "irq_o", "top", "input", "irq_pad", "harden_to_top_pad")
    add_edge(rows, "u_periph", "output", "event_o", "top", "input", "event_pad", "harden_to_top_pad")

    add_edge(rows, "u_clkgen", "output", "core_clk_o", "u_dma", "input", "clk_i", "clock_connection")
    add_edge(rows, "u_clkgen", "output", "bus_clk_o", "u_dpg", "input", "clk_i", "clock_connection")
    add_edge(rows, "u_clkgen", "output", "peri_clk_o", "u_periph", "input", "clk_i", "clock_connection")
    add_edge(rows, "u_clkgen", "output", "bus_clk_o", "u_ctrl", "input", "clk_i", "clock_connection")

    for i in range(8):
        add_edge(rows, "u_ctrl", "output", "clk_cfg_o[%d]" % i, "u_clkgen", "input", "cfg_i[%d]" % i)
    add_edge(rows, "u_ctrl", "output", "boot_mode_o", "u_clkgen", "input", "cfg_boot_i")
    add_edge(rows, "u_ctrl", "output", "irq_ack_o", "u_dma", "input", "irq_ack_i")
    add_edge(rows, "u_ctrl", "output", "debug_sel_o", "u_dpg", "input", "debug_sel_i")
    add_edge(rows, "u_ctrl", "output", "periph_enable_o", "u_periph", "input", "enable_i")
    for i in range(4):
        add_edge(rows, "u_ctrl", "output", "ctrl_cfg_o[%d]" % i, "u_dpg", "input", "cfg_shadow_i[%d]" % i)
    for i in range(4):
        add_edge(rows, "u_dma", "output", "cfg_o[%d]" % i, "u_periph", "input", "cfg_i[%d]" % i)
    for i in range(4):
        add_edge(rows, "u_dma", "output", "data_o[%d]" % i, "u_dpg", "input", "fti_0_dma2periph_data[%d]" % i, "feedthrough")
        add_edge(rows, "u_dpg", "output", "fto_0_dma2periph_data[%d]" % i, "u_periph", "input", "data_i[%d]" % i, "feedthrough")
        add_edge(rows, "u_dma", "output", "async_req_o[%d]" % i, "u_dpg", "input", "fti_0_dma2periph_async_req[%d]" % i, "feedthrough")
        add_edge(rows, "u_dpg", "output", "fto_0_dma2periph_async_req[%d]" % i, "u_periph", "input", "async_req_i[%d]" % i, "feedthrough")
        add_edge(rows, "u_periph", "output", "resp_o[%d]" % i, "u_dma", "input", "resp_i[%d]" % i)
    for i in range(4):
        add_edge(rows, "u_clkgen", "output", "status_o[%d]" % i, "u_ctrl", "input", "status_i[%d]" % i)

    headers = [
        "connection_id",
        "connection_type",
        "src_instance",
        "src_direction",
        "src_port",
        "src_bit_index",
        "src_endpoint_key",
        "src_soc_object",
        "dst_instance",
        "dst_direction",
        "dst_port",
        "dst_bit_index",
        "dst_endpoint_key",
        "dst_soc_object",
        "validation_status",
        "note",
    ]
    with (root / "connection_inventory.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers)
        writer.writeheader()
        for item in rows:
            writer.writerow(item)


def set_cell(ws, headers, row_idx, header, value):
    if header in headers:
        ws.cell(row=row_idx, column=headers[header], value=value)


def approve_02(d, stage, corner):
    wb = load_workbook(str(d / ("02_soc_clock_timing_budget_%s.xlsx" % stage)))
    ws = wb["clock_budget"]
    headers = {}
    header_row = 0
    for r in range(1, 20):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if "clock_name" in values:
            header_row = r
            headers = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
            break
    require(header_row, "02 header row not found")
    for r in range(header_row + 1, ws.max_row + 1):
        clock = ws.cell(r, headers["clock_name"]).value
        if not clock:
            continue
        set_cell(ws, headers, r, "sync_status", "OK")
        set_cell(ws, headers, r, "apply", "yes")
        set_cell(ws, headers, r, "setup_uncertainty", "0.05")
        set_cell(ws, headers, r, "hold_uncertainty", "0.02")
        if "virtual" not in str(clock):
            set_cell(ws, headers, r, "transition_max", "0.20")
            set_cell(ws, headers, r, "propagated", "no")
        set_cell(ws, headers, r, "note", "full chain regression budget")
    wb.save(str(d / ("02_soc_clock_timing_budget_%s.xlsx" % stage)))


def approve_03(d):
    wb = load_workbook(str(d / "03_soc_clock_groups.xlsx"))
    ws = wb["clock_group_rules"]
    headers = {}
    header_row = 0
    for r in range(1, 20):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if "group_id" in values:
            header_row = r
            headers = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
            break
    require(header_row, "03 header row not found")
    r = header_row + 1
    values = {
        "scenario": "common",
        "group_id": "CG_CORE_PERI_ASYNC",
        "relation_type": "asynchronous",
        "group_1_clocks": "top_ref_clk_pad u_clkgen_core_clk_o u_clkgen_bus_clk_o",
        "group_2_clocks": "top_peri_ref_pad",
        "analysis_style": "normal",
        "apply": "yes",
        "review_status": "approved",
        "owner": "sta_owner",
        "basis": "peripheral clock domain is CDC-reviewed against core/bus domain",
        "cdc_required": "yes",
        "note": "full chain regression async relation",
    }
    for header, value in values.items():
        set_cell(ws, headers, r, header, value)
    wb.save(str(d / "03_soc_clock_groups.xlsx"))


def approve_04(d):
    wb = load_workbook(str(d / "04_soc_io_pads.xlsx"))
    ws = wb["io_constraints"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    for r in range(2, ws.max_row + 1):
        pad = ws.cell(r, headers["pad_name"]).value
        ctype = ws.cell(r, headers["constraint_type"]).value
        if not pad:
            continue
        set_cell(ws, headers, r, "apply", "yes")
        set_cell(ws, headers, r, "review_status", "approved")
        set_cell(ws, headers, r, "owner", "io_owner")
        set_cell(ws, headers, r, "basis", "board/pad environment reviewed for full chain regression")
        if ctype in ("input_delay", "output_delay"):
            set_cell(ws, headers, r, "clock_name", "top_ref_clk_pad")
            if not ws.cell(r, headers["max_value"]).value and not ws.cell(r, headers["value"]).value:
                set_cell(ws, headers, r, "max_value", "1.0")
        elif ctype in ("input_transition", "max_transition"):
            if not ws.cell(r, headers["value"]).value:
                set_cell(ws, headers, r, "value", "0.10")
        elif ctype == "load":
            if not ws.cell(r, headers["value"]).value:
                set_cell(ws, headers, r, "value", "0.35")
    wb.save(str(d / "04_soc_io_pads.xlsx"))


def approve_20(d):
    wb = load_workbook(str(d / "20_harden_x_if.xlsx"))
    ws = wb["interface_budget"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    approved = 0
    for r in range(2, ws.max_row + 1):
        channel = ws.cell(r, headers["channel_id"]).value or ""
        if "u_dma_data_o_bit0__u_periph_data_i_bit0" not in channel:
            continue
        for header, value in {
            "scenario": "common",
            "stage": "all",
            "corner": "all",
            "timing_model": "lib_blackbox",
            "budget_required": "yes",
            "clock_relation": "synchronous",
            "budget_model": "interconnect_budget",
            "converted_max": "1.0",
            "converted_min": "0.1",
            "max_source": "reviewed_min_two_sides",
            "min_source": "reviewed_hold_budget",
            "derivation_basis": "min(src_output_delay_max,dst_input_delay_max)",
            "tool_surface": "sta",
            "datapath_only": "yes",
            "min_sign_review": "approved",
            "budget_basis": "owners define this data[0] path as SoC interconnect budget",
            "source_type": "extracted",
            "apply": "yes",
            "emit_max": "yes",
            "emit_min": "yes",
            "review_status": "approved",
            "owner": "sta_owner",
        }.items():
            set_cell(ws, headers, r, header, value)
        approved += 1
    require(approved == 1, "expected one 20 budget approval, got %d" % approved)
    wb.save(str(d / "20_harden_x_if.xlsx"))


def approve_30(d):
    wb = load_workbook(str(d / "30_harden_to_harden_exception.xlsx"))
    ws = wb["exception_rule"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    approved = 0
    for r in range(2, ws.max_row + 1):
        channel = ws.cell(r, headers["channel_id"]).value or ""
        exception_type = ws.cell(r, headers["exception_type"]).value or ""
        if channel == "CH_u_dma_cfg_o_bit0__u_periph_cfg_i_bit0" and exception_type == "false_path":
            values = {
                "scenario": "func",
                "stage": "all",
                "corner": "all",
                "apply": "yes",
                "review_status": "approved",
                "owner": "fw_owner",
                "exception_type": "false_path",
                "path_category": "static",
                "source_type": "extracted_harden_exception",
                "harden_clock_context_status": "remapped_equivalent",
                "basis": "cfg bit0 is boot-static in func scenario after firmware lock",
                "risk_level": "medium",
            }
        elif channel == "CH_u_dma_async_req_o_bit1__u_periph_async_req_i_bit1" and exception_type == "needs_review":
            values = {
                "scenario": "common",
                "stage": "prects",
                "corner": "ss_125",
                "apply": "yes",
                "review_status": "approved",
                "owner": "cdc_owner",
                "exception_type": "max_min_delay_override",
                "path_category": "handshake",
                "clock_relation": "asynchronous",
                "max_value": "12.0",
                "min_value": "1.0",
                "datapath_only": "yes",
                "source_type": "manual_entry",
                "harden_clock_context_status": "not_applicable",
                "cdc_rdc_ref": "CDC-HS-001",
                "protocol_ref": "dma2periph async req handshake window",
                "basis": "CDC handshake skew window, tool report confirms path max/min not shadowed",
                "risk_level": "high",
            }
        elif channel == "CH_u_dma_cfg_o_bit1__u_periph_cfg_i_bit1" and exception_type == "false_path":
            values = {
                "scenario": "func",
                "stage": "all",
                "corner": "all",
                "apply": "yes",
                "review_status": "approved",
                "owner": "rdc_owner",
                "exception_type": "false_path",
                "path_category": "reset",
                "source_type": "extracted_harden_exception",
                "harden_clock_context_status": "remapped_equivalent",
                "basis": "复位同步器保证恢复/移除检查由RDC签核覆盖",
                "cdc_rdc_ref": "RDC-RESET-001",
                "risk_level": "high",
            }
        else:
            continue
        for header, value in values.items():
            set_cell(ws, headers, r, header, value)
        approved += 1
    require(approved == 3, "expected three 30 approvals, got %d" % approved)
    wb.save(str(d / "30_harden_to_harden_exception.xlsx"))


def copy_for_stage(d, stage_name):
    stage_dir = d / "stage_snapshots" / stage_name
    stage_dir.mkdir(parents=True, exist_ok=True)
    for pattern in ("*.txt", "*.csv", "*.sdc", "*.xlsx"):
        for path in d.glob(pattern):
            if path.is_file():
                shutil.copy2(str(path), str(stage_dir / path.name))
    if (d / "common").is_dir():
        copytree_replace(d / "common", stage_dir / "common")
    if (d / "scenarios").is_dir():
        copytree_replace(d / "scenarios", stage_dir / "scenarios")
    if (d / "00_harden_port_inventory").is_dir():
        copytree_replace(d / "00_harden_port_inventory", stage_dir / "00_harden_port_inventory")


def copytree_replace(src, dst):
    if dst.exists():
        shutil.rmtree(str(dst))
    shutil.copytree(str(src), str(dst))


def pending_summary(d):
    result = {}
    pending_dir = d / "00_harden_port_inventory" / "pending"
    for path in sorted(pending_dir.glob("*.ports")):
        lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
        result[path.name] = {"count": len(lines), "ports": lines}
    return result


def removed_summary(d):
    result = {}
    log_dir = d / "00_harden_port_inventory" / "removed_log"
    for path in sorted(log_dir.glob("*.removed")):
        lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
        result[path.name] = {"count": len(lines), "sample": lines[:8]}
    return result


def assert_outputs(d):
    expect_files = [
        "common/01_soc_clocks.sdc",
        "common/02_soc_clock_timing_prects_ss_125.sdc",
        "common/03_soc_clock_groups.sdc",
        "common/04_soc_io_pads.sdc",
        "common/10_feedthrough.sdc",
        "common/20_harden_x_if.sdc",
        "scenarios/func_exceptions.sdc",
        "common/30_harden_to_harden_exception_prects_ss_125.sdc",
    ]
    for rel in expect_files:
        require((d / rel).is_file(), "expected artifact missing: %s" % rel)

    s20 = (d / "common/20_harden_x_if.sdc").read_text(encoding="utf-8")
    require("set_max_delay 1 -datapath_only -from [get_pins {u_dma/data_o[0]}] -to [get_pins {u_periph/data_i[0]}]" in s20, "20 data[0] max budget missing")
    require("set_min_delay 0.1 -datapath_only -from [get_pins {u_dma/data_o[0]}] -to [get_pins {u_periph/data_i[0]}]" in s20, "20 data[0] min budget missing")

    s30_func = (d / "scenarios/func_exceptions.sdc").read_text(encoding="utf-8")
    require("set_false_path -from [get_pins {u_dma/cfg_o[0]}] -to [get_pins {u_periph/cfg_i[0]}]" in s30_func, "30 func cfg false_path missing")
    require("复位同步器" in s30_func, "30 reset Chinese basis missing")

    s30_view = (d / "common/30_harden_to_harden_exception_prects_ss_125.sdc").read_text(encoding="utf-8")
    require("set_max_delay 12 -datapath_only -from [get_pins {u_dma/async_req_o[1]}]" in s30_view, "30 async max override missing")
    require("-through [get_pins {u_dpg/fti_0_dma2periph_async_req[1]}]" in s30_view, "30 async through fti missing")
    require("-through [get_pins {u_dpg/fto_0_dma2periph_async_req[1]}]" in s30_view, "30 async through fto missing")
    require("set_min_delay 1 -datapath_only" in s30_view, "30 async min override missing")

    summary = pending_summary(d)
    require(summary["u_clkgen.ports"]["count"] >= 1, "u_clkgen pending should keep non-covered leftovers")
    require("output data_o[0]" not in "\n".join(summary["u_dma.ports"]["ports"]), "u_dma data_o[0] should be consumed")
    require("output cfg_o[0]" not in "\n".join(summary["u_dma.ports"]["ports"]), "u_dma cfg_o[0] should be consumed")
    require("input fti_0_dma2periph_async_req[1]" not in "\n".join(summary["u_dpg.ports"]["ports"]), "u_dpg fti async[1] should be consumed by 10")
    require("input cfg_i[0]" not in "\n".join(summary["u_periph.ports"]["ports"]), "u_periph cfg_i[0] should be consumed by 30")


def main():
    if WORK.exists():
        shutil.rmtree(str(WORK))
    write_inputs(WORK)
    write_00_inventory(WORK)

    run_ok([EX01, "--port-files", "ports.xlsx"], WORK)
    copy_for_stage(WORK, "01")

    run_sync([EX02, "-scenario", "common", "-stage", "prects", "-corner", "ss_125", "-input", "clock_inventory.csv"], WORK)
    approve_02(WORK, "prects", "ss_125")
    run_ok([EX02, "-scenario", "common", "-stage", "prects", "-corner", "ss_125", "-input", "clock_inventory.csv"], WORK)
    copy_for_stage(WORK, "02")

    run_sync([EX03, "-scenario", "common", "-input", "clock_inventory.csv"], WORK)
    approve_03(WORK)
    run_ok([EX03, "-scenario", "common", "-input", "clock_inventory.csv"], WORK)
    copy_for_stage(WORK, "03")

    run_sync([EX04, "-scenario", "common", "-input", "clock_inventory.csv"], WORK)
    approve_04(WORK)
    run_ok([EX04, "-scenario", "common", "-input", "clock_inventory.csv"], WORK)
    copy_for_stage(WORK, "04")

    run_ok([EX10, "--port-workbook", "ports.xlsx"], WORK)
    copy_for_stage(WORK, "10")

    run_sync([EX20, "-input", "clock_inventory.csv"], WORK)
    approve_20(WORK)
    run_ok([EX20, "-input", "clock_inventory.csv"], WORK)
    copy_for_stage(WORK, "20")

    run_sync([EX30], WORK)
    approve_30(WORK)
    run_ok([EX30, "-scenario", "func"], WORK)
    run_ok([EX30, "-stage", "prects", "-corner", "ss_125"], WORK)
    copy_for_stage(WORK, "30")

    assert_outputs(WORK)
    summary = {
        "case_dir": str(WORK),
        "pending": pending_summary(WORK),
        "removed": removed_summary(WORK),
        "key_artifacts": [
            "common/01_soc_clocks.sdc",
            "common/02_soc_clock_timing_prects_ss_125.sdc",
            "common/03_soc_clock_groups.sdc",
            "common/04_soc_io_pads.sdc",
            "common/10_feedthrough.sdc",
            "common/20_harden_x_if.sdc",
            "scenarios/func_exceptions.sdc",
            "common/30_harden_to_harden_exception_prects_ss_125.sdc",
        ],
    }
    (WORK / "full_chain_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print("00->30 full-chain complex case: PASS")
    print("Case dir:", WORK)
    print("Summary :", WORK / "full_chain_summary.json")


if __name__ == "__main__":
    main()
