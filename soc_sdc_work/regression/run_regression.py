#!/usr/bin/env python3
"""
One-shot regression for the SoC SDC 01 -> 02 -> 03 -> 04 -> 20 chain.

Layout (recreated under work/ each run):
  work/01_soc_clocks/        run 01, produce clock_inventory.csv + 01 sdc
  work/02_soc_clock_timing/  fill budgets, run 02 across stage/scenario/corner
  work/03_soc_clock_groups/  fill rules, run 03 + coverage
  work/04_soc_io_pads/       extract/review IO pad constraints and generate 04
  work/20_harden_x_if/       extract/review interface budgets and generate 20

It collects deterministic TEXT artifacts (sdc / csv / normalized reports /
coverage extracts) into work/artifacts and diffs them against expected/.

Usage:
  python3 run_regression.py            # compare against expected/ (fail on diff)
  python3 run_regression.py --update   # (re)write expected/ baseline
"""
import argparse
import csv
import io
import shutil
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
SOC = BASE.parent
EX01 = SOC / "01_soc_clocks" / "01_extract_soc_clocks.py"
EX02 = SOC / "02_soc_clock_timing" / "02_extract_soc_clock_timing.py"
EX03 = SOC / "03_soc_clock_groups" / "03_extract_soc_clock_groups.py"
EX04 = SOC / "04_soc_io_pads" / "04_extract_soc_io_pads.py"
EX20 = SOC / "20_harden_x_if" / "20_extract_harden_x_if.py"
WORK = BASE / "work"
EXP = BASE / "expected"
ART = WORK / "artifacts"

REQ = ["Parameter", "Inout", "Inout Width", "Inout Connectivity", "Inout Name",
       "Input", "Input Width", "Input Used Width", "From Whom",
       "Output", "Output Width", "Output Used Width", "To Top"]


def sh(cmd, cwd):
    return subprocess.run([sys.executable, *cmd], cwd=str(cwd),
                          stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                          universal_newlines=True)


def port_sheet(rows):
    df = pd.DataFrame(rows)
    for c in REQ:
        if c not in df.columns:
            df[c] = ""
    return df[REQ].fillna("")


# ----------------------------------------------------------------------------
# 01: build the demo2 complex topology and run extraction
# ----------------------------------------------------------------------------
def run_01(d: Path):
    d.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {"module_name": "pll_top", "inst_name": "u_pll",   "owner": "alice", "file_path": ""},
        {"module_name": "fab",     "inst_name": "u_fab0",  "owner": "alice", "file_path": ""},
        {"module_name": "fab",     "inst_name": "u_fab1",  "owner": "alice", "file_path": ""},
        {"module_name": "periph",  "inst_name": "u_periph", "owner": "bob",  "file_path": ""},
    ]).to_excel(d / "info_all.xlsx", index=False)

    with pd.ExcelWriter(d / "port_alice.xlsx", engine="xlsxwriter") as w:
        port_sheet([
            {"Input": "ref_clk_in", "Input Width": 1, "From Whom": "top.sys_clk_pad"},
            {"Output": "core_clk_o", "Output Width": 1},
            {"Output": "bus_clk_o", "Output Width": 1},
        ]).to_excel(w, sheet_name="u_pll", index=False)
        for inst in ("u_fab0", "u_fab1"):
            port_sheet([
                {"Input": "fab_clk_i", "Input Width": 1, "From Whom": "u_pll.core_clk_o"},
                {"Output": "fab_clk_o", "Output Width": 1},
            ]).to_excel(w, sheet_name=inst, index=False)
    with pd.ExcelWriter(d / "port_bob.xlsx", engine="xlsxwriter") as w:
        port_sheet([
            {"Input": "clk_i", "Input Width": 1, "From Whom": "u_fab0.fab_clk_o"},
            {"Input": "ref2_i", "Input Width": 1, "From Whom": "top.aux_clk_pad"},
            {"Input": "scan_mode_clk", "Input Width": 1, "From Whom": "top.scan_clk_pad"},
            {"Output": "clk_o", "Output Width": 1},
        ]).to_excel(w, sheet_name="u_periph", index=False)

    (d / "pll_top.sdc").write_text(
        "create_clock -name pll_ref -period 20.000 [get_ports ref_clk_in]\n"
        "create_generated_clock -name pll_core -source [get_ports ref_clk_in] -multiply_by 4 [get_ports core_clk_o]\n"
        "create_generated_clock -name pll_bus  -source [get_ports ref_clk_in] -multiply_by 2 [get_ports bus_clk_o]\n"
        "set_false_path -from [get_ports ref_clk_in]\n")
    (d / "fab.sdc").write_text(
        "create_clock -name fab_in -period 5.000 [get_ports fab_clk_i]\n"
        "create_generated_clock -name fab_out -source [get_ports fab_clk_i] -combinational [get_ports fab_clk_o]\n")
    (d / "periph.sdc").write_text(
        "create_clock -name periph_in   -period 5.000  [get_ports clk_i]\n"
        "create_clock -name periph_ref2 -period 40.000 [get_ports ref2_i]\n"
        "create_clock -name scan_clk    -period 50.000 [get_ports scan_mode_clk]\n"
        "create_generated_clock -name periph_out -source [get_ports clk_i] -combinational [get_ports clk_o]\n")
    (d / "virtual_clocks.csv").write_text(
        "clock_name,period,waveform,note\n"
        "v_ddr_ref,2.500,,DDR external reference\n"
        "v_pcie_ref,10.000,{0 5},PCIe external reference\n"
        "v_uart_rx,20.000,,UART RX board reference\n"
        "v_uart_tx,20.000,,UART TX board reference\n"
        "dqs_clk,2.500,,DDR DQS source-sync reference\n")

    pending = d / "00_harden_port_inventory/pending"
    pending.mkdir(parents=True, exist_ok=True)
    (pending / "u_pll.ports").write_text(
        "input ref_clk_in\n"
        "output bus_clk_o\n"
        "output core_clk_o\n"
        "input cfg_i\n",
        encoding="utf-8",
    )
    (pending / "u_fab0.ports").write_text(
        "input fab_clk_i\n"
        "output fab_clk_o\n"
        "output data_o\n",
        encoding="utf-8",
    )
    (pending / "u_fab1.ports").write_text(
        "input fab_clk_i\n"
        "output fab_clk_o\n"
        "output data_o\n",
        encoding="utf-8",
    )
    (pending / "u_periph.ports").write_text(
        "input clk_i\n"
        "input ref2_i\n"
        "input scan_mode_clk\n"
        "output clk_o\n"
        "input cfg_i\n",
        encoding="utf-8",
    )

    r = sh([str(EX01)], cwd=d)
    assert r.returncode == 0, f"01 failed:\n{r.stdout}\n{r.stderr}"


def run_01_edge_checks(d: Path):
    d.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {"module_name": "edge_mod", "inst_name": "u_edge", "owner": "eve", "file_path": ""},
    ]).to_excel(d / "info_all.xlsx", index=False)

    with pd.ExcelWriter(d / "port_eve.xlsx", engine="xlsxwriter") as w:
        port_sheet([
            {"Input": "ref_i", "Input Width": 1, "From Whom": "top.ref_pad"},
            {"Input": "no_period_i", "Input Width": 1, "From Whom": "top.no_period_pad"},
            {"Output": "gen_o", "Output Width": 1},
            {"Output": "add_o", "Output Width": 1},
        ]).to_excel(w, sheet_name="U_EDGE", index=False)
        port_sheet([
            {"Input": "unused_i", "Input Width": 1, "From Whom": "top.unused_pad"},
        ]).to_excel(w, sheet_name="u_orphan", index=False)

    (d / "edge_mod.sdc").write_text(
        "create_clock -name local_ref -period 10.000 [get_ports ref_i]\n"
        "create_generated_clock -name gen_clk -source [get_clocks local_ref] -divide_by 2 [get_ports gen_o]\n"
        "create_generated_clock -add -name add_clk -source [get_ports ref_i] -divide_by 4 [get_ports add_o]\n"
        "create_clock -name no_period [get_ports no_period_i]\n",
        encoding="utf-8",
    )

    pending = d / "00_harden_port_inventory/pending"
    pending.mkdir(parents=True, exist_ok=True)
    (pending / "u_edge.ports").write_text(
        "input ref_i\n"
        "input no_period_i\n"
        "output gen_o\n"
        "output add_o\n",
        encoding="utf-8",
    )

    r = sh([str(EX01)], cwd=d)
    assert r.returncode == 0, f"01 edge checks failed:\n{r.stdout}\n{r.stderr}"
    rerun = sh([str(EX01)], cwd=d)
    assert rerun.returncode == 0, f"01 pending update should be idempotent:\n{rerun.stdout}\n{rerun.stderr}"
    sdc = (d / "common/01_soc_clocks.sdc").read_text()
    report = (d / "clock_check_report.txt").read_text()
    assert "-source [get_clocks {top_ref_pad}]" in sdc, "01 did not remap -source [get_clocks local_ref]"
    assert "matched instance 'u_edge' by case/space-insensitive fallback" in report
    assert "port workbook sheet 'u_orphan' does not match any inst_name; ignored" in report
    assert "CLOCK_CREATE_CLOCK_MISSING_PERIOD" in report
    assert "CLOCK_ADD_OPTION_OUT_OF_SCOPE" in report
    assert "not present in pending and no previous_removed" not in report


def run_01_bad_source_checks(d: Path):
    d.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {"module_name": "bad_src_mod", "inst_name": "u_bad_src", "owner": "eve", "file_path": ""},
    ]).to_excel(d / "info_all.xlsx", index=False)

    with pd.ExcelWriter(d / "port_eve.xlsx", engine="xlsxwriter") as w:
        port_sheet([
            {"Input": "ref_i", "Input Width": 1, "From Whom": "top.ref_pad"},
            {"Output": "gen_o", "Output Width": 1},
        ]).to_excel(w, sheet_name="u_bad_src", index=False)

    (d / "bad_src_mod.sdc").write_text(
        "create_clock -name local_ref -period 10.000 [get_ports ref_i]\n"
        "create_generated_clock -name bad_gen -source [get_ports ref_typo] -divide_by 2 [get_ports gen_o]\n",
        encoding="utf-8",
    )

    r = sh([str(EX01), "--no-update-pending"], cwd=d)
    assert r.returncode == 1, "01 should fail when generated clock -source get_ports is not in owner sheet"
    report = (d / "clock_check_report.txt").read_text()
    assert "CLOCK_GENERATED_SOURCE_NOT_IN_OWNER_SHEET" in report


def run_01_multi_target_checks(d: Path):
    d.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {"module_name": "multi_mod", "inst_name": "u_multi", "owner": "eve", "file_path": ""},
    ]).to_excel(d / "info_all.xlsx", index=False)

    with pd.ExcelWriter(d / "port_eve.xlsx", engine="xlsxwriter") as w:
        port_sheet([
            {"Input": "clk_a", "Input Width": 1, "From Whom": "top.clk_a_pad"},
            {"Input": "clk_b", "Input Width": 1, "From Whom": "top.clk_b_pad"},
        ]).to_excel(w, sheet_name="u_multi", index=False)

    (d / "multi_mod.sdc").write_text(
        "create_clock -period 10.000 [get_ports {clk_a clk_b}]\n",
        encoding="utf-8",
    )

    r = sh([str(EX01), "--no-update-pending"], cwd=d)
    assert r.returncode == 1, "01 should fail fast on multi-target clock commands"
    report = (d / "clock_check_report.txt").read_text()
    inv = (d / "clock_inventory.csv").read_text()
    assert "CLOCK_MULTI_TARGET_NOT_SUPPORTED" in report
    assert "clk_a" in inv and "clk_b" in inv and "skipped" in inv


def active_clocks(inv_csv: Path):
    out = {}
    for row in csv.DictReader(inv_csv.open(encoding="utf-8-sig")):
        if row["final_action"].startswith("emit_"):
            out[row["clock_name"]] = row["clock_kind"]
    return out


# ----------------------------------------------------------------------------
# 02: fill budgets and run the stage/scenario/corner matrix
# ----------------------------------------------------------------------------
def budget_values(kind, corner):
    k = 1.0 if corner == "ss_125" else 0.6
    if "virtual" in kind:
        return dict(setup_uncertainty=round(0.05 * k, 3), hold_uncertainty=0.02)
    if "generated_combinational" in kind:
        return dict(setup_uncertainty=round(0.11 * k, 3), hold_uncertainty=round(0.035 * k, 3),
                    network_latency_early=round(0.25 * k, 3), network_latency_late=round(0.60 * k, 3))
    if "generated" in kind:
        return dict(setup_uncertainty=round(0.12 * k, 3), hold_uncertainty=round(0.04 * k, 3),
                    network_latency_early=round(0.30 * k, 3), network_latency_late=round(0.70 * k, 3),
                    transition_min=0.03, transition_max=round(0.12 * k, 3))
    return dict(setup_uncertainty=round(0.10 * k, 3), hold_uncertainty=round(0.03 * k, 3),
                source_latency_early=round(0.10 * k, 3), source_latency_late=round(0.24 * k, 3),
                network_latency_early=round(0.28 * k, 3), network_latency_late=round(0.65 * k, 3),
                transition_min=0.03, transition_max=round(0.11 * k, 3))


def fill_budget(form: Path, stage: str, clocks: dict, propagated: bool):
    wb = load_workbook(form)
    ws = wb["clock_budget"]
    hdr, col = None, {}
    for r in range(1, 8):
        names = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
        if "clock_name" in names:
            hdr, col = r, names
            break

    def setc(r, **kv):
        for k, v in kv.items():
            ws.cell(r, col[k], v)

    # fill auto-created common/ss_125 rows
    for r in range(hdr + 1, ws.max_row + 1):
        cn = ws.cell(r, col["clock_name"]).value
        if cn in clocks:
            if propagated:
                setc(r, apply="yes", sync_status="OK", propagated="yes",
                     setup_uncertainty=0.05, hold_uncertainty=0.02)
            else:
                setc(r, apply="yes", sync_status="OK", **budget_values(clocks[cn], "ss_125"))

    def last():
        lr = hdr
        for r in range(hdr + 1, ws.max_row + 1):
            if any(ws.cell(r, c).value not in (None, "") for c in col.values()):
                lr = r
        return lr

    if not propagated:
        # func override on the PLL core clock
        r = last() + 1
        setc(r, scenario="func", stage=stage, corner="ss_125", clock_name="u_pll_core_clk_o",
             setup_uncertainty=0.15, hold_uncertainty=0.05,
             network_latency_early=0.30, network_latency_late=0.70,
             transition_min=0.03, transition_max=0.12, apply="yes", sync_status="OK")
    wb.save(form)


def run_02(d: Path, inv: Path, clocks: dict):
    d.mkdir(parents=True, exist_ok=True)
    # prects: gate -> fill -> generate common + func
    sh([str(EX02), "-scenario", "common", "-stage", "prects", "-corner", "ss_125"], cwd=d)
    fill_budget(d / "02_soc_clock_timing_budget_prects.xlsx", "prects", clocks, propagated=False)
    for scen in ("common", "func"):
        r = sh([str(EX02), "-scenario", scen, "-stage", "prects", "-corner", "ss_125"], cwd=d)
        assert r.returncode == 0, f"02 prects {scen} failed:\n{r.stdout}\n{r.stderr}"
    # postcts: gate -> fill (propagated) -> generate common
    sh([str(EX02), "-scenario", "common", "-stage", "postcts", "-corner", "ss_125"], cwd=d)
    fill_budget(d / "02_soc_clock_timing_budget_postcts.xlsx", "postcts", clocks, propagated=True)
    r = sh([str(EX02), "-scenario", "common", "-stage", "postcts", "-corner", "ss_125"], cwd=d)
    assert r.returncode == 0, f"02 postcts failed:\n{r.stdout}\n{r.stderr}"


# ----------------------------------------------------------------------------
# 03: fill clock-group rules and run common
# ----------------------------------------------------------------------------
def run_03(d: Path):
    d.mkdir(parents=True, exist_ok=True)
    sh([str(EX03), "-scenario", "common"], cwd=d)  # gate creates form
    form = d / "03_soc_clock_groups.xlsx"
    wb = load_workbook(form)
    ws = wb["clock_group_rules"]
    hdr, col = None, {}
    for r in range(1, 8):
        names = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
        if "group_id" in names:
            hdr, col = r, names
            break

    def add(row, **kv):
        for k, v in kv.items():
            ws.cell(row, col[k], v)

    add(hdr + 1, scenario="common", group_id="CG_ASYNC_CORE_AUX", relation_type="asynchronous",
        group_1_clocks="u_pll_core_clk_o", group_2_clocks="top_aux_clk_pad",
        apply="yes", review_status="approved", cdc_required="yes",
        basis="CDC: core domain async to aux")
    add(hdr + 2, scenario="common", group_id="CG_ASYNC_BUS_AUX", relation_type="asynchronous",
        group_1_clocks="u_pll_bus_clk_o", group_2_clocks="top_aux_clk_pad",
        apply="yes", review_status="approved", cdc_required="yes",
        basis="CDC: bus async to aux")
    wb.save(form)
    r = sh([str(EX03), "-scenario", "common"], cwd=d)
    assert r.returncode == 0, f"03 failed:\n{r.stdout}\n{r.stderr}"


# ----------------------------------------------------------------------------
# 04: IO/pad extraction + reviewed common/scenario/view-specific generation
# ----------------------------------------------------------------------------
def build_04_inputs(d: Path):
    d.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {"inst_name": "u_io", "module_name": "io_ring", "owner": "carol", "sdc_path": "u_io.sdc"},
    ]).to_excel(d / "info_all.xlsx", index=False)

    with pd.ExcelWriter(d / "ports_u_io.xlsx", engine="xlsxwriter") as w:
        port_sheet([
            {
                "Input": "uart0_sin", "Input Width": 1, "From Whom": "top.pad_uart0_sin",
                "Output": "uart0_sout", "Output Width": 1, "To Top": "top.pad_uart0_sout",
                "Inout": "gpio0", "Inout Width": 1, "Inout Connectivity": "top.pad_gpio0",
            },
            {"Output": "ddr_dqs", "Output Width": 1, "To Top": "top.pad_ddr_dqs"},
        ]).to_excel(w, sheet_name="u_io", index=False)

    (d / "u_io.sdc").write_text(
        "# lower-level io_ring SDC (block signoff env)\n"
        "set_input_delay -clock [get_clocks v_uart_rx] -max 5.0 [get_ports uart0_sin]\n"
        "set_input_transition 0.2 [get_ports uart0_sin]\n"
        "set_driving_cell -lib_cell BUFX2 -pin Y [get_ports uart0_sin]\n"
        "set_output_delay -clock [get_clocks v_uart_tx] -max 4.0 [get_ports uart0_sout]\n"
        "set_load 0.05 [get_ports uart0_sout]\n"
        "set_input_delay -clock [get_clocks v_uart_rx] -max 3.0 [get_ports gpio0]\n",
        encoding="utf-8",
    )


def fill_04_form(form: Path):
    wb = load_workbook(form)
    ws = wb["io_constraints"]
    col = {cell.value: cell.column for cell in ws[1]}

    def setc(row, **kv):
        for k, v in kv.items():
            ws.cell(row=row, column=col[k], value=v)

    for r in range(2, ws.max_row + 1):
        pad = ws.cell(r, col["pad_name"]).value
        ctype = ws.cell(r, col["constraint_type"]).value
        if not pad or not ctype:
            continue
        if pad == "pad_uart0_sin" and ctype == "input_delay":
            setc(r, apply="yes", review_status="approved", timing_class="async", basis="UART RX board budget")
        elif pad == "pad_uart0_sin" and ctype == "input_transition":
            setc(r, apply="yes", review_status="approved", basis="IO spec input slew")
        elif pad == "pad_uart0_sin" and ctype == "driving_cell":
            setc(r, apply="no", review_status="rejected", note="rejected: input_transition used instead")
        elif pad == "pad_uart0_sout" and ctype == "output_delay":
            setc(r, apply="yes", review_status="approved", timing_class="async", basis="UART TX board budget")
        elif pad == "pad_uart0_sout" and ctype == "load":
            setc(r, apply="yes", review_status="approved", basis="package + PCB load")
        elif pad == "pad_gpio0" and ctype == "input_delay":
            setc(r, scenario="gpio_in", apply="yes", review_status="approved",
                 timing_class="timed", basis="GPIO input direction budget")

    # Add one view-specific electrical constraint on a pad that has no all/all load.
    new_row = ws.max_row + 1
    manual = {
        "scenario": "common",
        "stage": "prects",
        "corner": "ss_125",
        "pad_name": "pad_ddr_dqs",
        "soc_object": "pad_ddr_dqs",
        "subsys_instance": "u_io",
        "subsys_port": "ddr_dqs",
        "direction": "output",
        "timing_class": "timed",
        "constraint_type": "load",
        "value": "0.03",
        "object_granularity": "single_port",
        "unit_cap": "pF",
        "source_type": "manual",
        "apply": "yes",
        "review_status": "approved",
        "owner": "carol",
        "basis": "pre-CTS board/package estimate",
    }
    for key, value in manual.items():
        ws.cell(new_row, col[key], value)
    wb.save(form)


def run_04(d: Path):
    build_04_inputs(d)
    first = sh([str(EX04), "-scenario", "common", "-input", "../01_soc_clocks/clock_inventory.csv"], cwd=d)
    assert first.returncode == 1, f"04 first run should stop for review:\n{first.stdout}\n{first.stderr}"
    fill_04_form(d / "04_soc_io_pads.xlsx")
    for args in (
        ["-scenario", "common"],
        ["-scenario", "gpio_in"],
        ["-scenario", "common", "-stage", "prects", "-corner", "ss_125"],
    ):
        r = sh([str(EX04), *args, "-input", "../01_soc_clocks/clock_inventory.csv"], cwd=d)
        assert r.returncode == 0, f"04 {' '.join(args)} failed:\n{r.stdout}\n{r.stderr}"


# ----------------------------------------------------------------------------
# 20: harden/subsys interface budget extraction + auto-resolve + blocking checks
# ----------------------------------------------------------------------------
def build_20_inputs(d: Path):
    d.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {"inst_name": "u_a", "module_name": "harden_a", "owner": "dave", "sdc_path": "u_a.sdc"},
        {"inst_name": "u_b", "module_name": "harden_b", "owner": "erin", "sdc_path": "u_b.sdc"},
    ]).to_excel(d / "info_all.xlsx", index=False)

    with pd.ExcelWriter(d / "ports.xlsx", engine="xlsxwriter") as w:
        port_sheet([
            {"Input": "clk_i", "From Whom": "top.sys_clk_pad", "Output": "data_o", "To Top": "fabric_bus"},
        ]).to_excel(w, sheet_name="u_a", index=False)
        port_sheet([
            {"Input": "data_i", "From Whom": "u_a.data_o"},
        ]).to_excel(w, sheet_name="u_b", index=False)

    (d / "u_a.sdc").write_text(
        "set_output_delay -max 1.5 -min -0.1 -clock [get_clocks u_pll_core_clk_o] [get_ports data_o]\n",
        encoding="utf-8",
    )
    (d / "u_b.sdc").write_text(
        "set_input_delay -max 1.2 -clock [get_clocks u_pll_core_clk_o] [get_ports data_i]\n",
        encoding="utf-8",
    )
    inv = d / "00_harden_port_inventory"
    pending = inv / "pending"
    pending.mkdir(parents=True, exist_ok=True)
    (pending / "u_a.ports").write_text(
        "input clk_i\n"
        "output data_o\n",
        encoding="utf-8",
    )
    (pending / "u_b.ports").write_text(
        "input data_i\n",
        encoding="utf-8",
    )
    (inv / "connection_inventory.csv").write_text(
        "connection_id,connection_type,src_instance,src_direction,src_port,src_bit_index,src_endpoint_key,src_soc_object,"
        "dst_instance,dst_direction,dst_port,dst_bit_index,dst_endpoint_key,dst_soc_object,validation_status,note\n"
        "CONN_u_a_data_o__u_b_data_i,harden_to_harden,u_a,output,data_o,,u_a:output:data_o,u_a/data_o,"
        "u_b,input,data_i,,u_b:input:data_i,u_b/data_i,matched,\n"
        "CONN_u_a_data_o__fabric_bus,harden_to_fabric,u_a,output,data_o,,u_a:output:data_o,u_a/data_o,"
        "fabric,,fabric_bus,,fabric::fabric_bus,,matched,\n",
        encoding="utf-8",
    )


def approve_20(form: Path, async_relation: bool = False):
    wb = load_workbook(form)
    ws = wb["interface_budget"]
    col = {cell.value: cell.column for cell in ws[1]}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col["channel_id"]).value != "CH_u_a_data_o__u_b_data_i":
            continue
        updates = {
            "timing_model": "lib_blackbox",
            "budget_required": "yes",
            "budget_model": "interconnect_budget",
            "converted_max": "",
            "max_source": "",
            "derivation_basis": "",
            "tool_surface": "sta",
            "datapath_only": "yes",
            "budget_basis": "interconnect budget from block owners",
            "apply": "yes",
            "emit_max": "yes",
            "emit_min": "no",
            "review_status": "approved",
            "clock_relation": "async" if async_relation else "",
            "relationship_override_basis": "",
        }
        for key, value in updates.items():
            ws.cell(r, col[key], value)
        break
    else:
        raise AssertionError("20 target channel not found")
    wb.save(form)


def run_20(d: Path):
    build_20_inputs(d)
    first = sh([str(EX20), "-input", "../01_soc_clocks/clock_inventory.csv"], cwd=d)
    assert first.returncode == 1, f"20 first run should stop for review:\n{first.stdout}\n{first.stderr}"
    approve_20(d / "20_harden_x_if.xlsx")
    r = sh([str(EX20), "-input", "../01_soc_clocks/clock_inventory.csv", "--max-diff-threshold", "0.1"], cwd=d)
    assert r.returncode == 0, f"20 normal generation failed:\n{r.stdout}\n{r.stderr}"

    # Confirm async/exclusive gating in the same complex case without keeping its errored report.
    normal_report = (d / "harden_x_if_check_report_common_all_all.txt").read_text()
    normal_sdc = (d / "common/20_harden_x_if.sdc").read_text()
    approve_20(d / "20_harden_x_if.xlsx", async_relation=True)
    bad = sh([str(EX20), "-input", "../01_soc_clocks/clock_inventory.csv"], cwd=d)
    assert bad.returncode == 1, "20 async relation should block generation"
    assert "clock_relation=asynchronous blocks normal 20 budget" in (d / "harden_x_if_check_report_common_all_all.txt").read_text()

    # Restore the successful artifacts for deterministic collection.
    (d / "harden_x_if_check_report_common_all_all.txt").write_text(normal_report, encoding="utf-8")
    (d / "common/20_harden_x_if.sdc").write_text(normal_sdc, encoding="utf-8")


# ----------------------------------------------------------------------------
# artifact collection + normalization
# ----------------------------------------------------------------------------
def norm(text: str) -> str:
    return text.replace(str(WORK), "<WORK>")


def xlsx_sheet_to_csv(path: Path, sheet: str, cols, start_row=5) -> str:
    ws = load_workbook(path)[sheet]
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(cols)
    rows = []
    for r in range(start_row, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(cols) + 1)]
        if all(v in (None, "") for v in vals):
            continue
        rows.append(["" if v is None else str(v).replace("\n", " ") for v in vals])
    rows.sort()
    for row in rows:
        w.writerow(row)
    return buf.getvalue()


def collect(name: str, text: str):
    p = ART / name
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(text, encoding="utf-8")


def collect_artifacts(w01, w02, w03, w04, w20):
    if ART.exists():
        shutil.rmtree(ART)
    ART.mkdir(parents=True)
    # 01
    collect("01/01_soc_clocks.sdc", (w01 / "common/01_soc_clocks.sdc").read_text())
    collect("01/clock_inventory.csv", (w01 / "clock_inventory.csv").read_text())
    collect("01/clock_check_report.txt", norm((w01 / "clock_check_report.txt").read_text()))
    collect("01/removed_log/01_soc_clocks.removed",
            (w01 / "00_harden_port_inventory/removed_log/01_soc_clocks.removed").read_text())
    for path in sorted((w01 / "00_harden_port_inventory/pending").glob("*.ports")):
        collect(f"01/pending/{path.name}", path.read_text())
    # 02
    for rel in ["common/02_soc_clock_timing_prects_ss_125.sdc",
                "scenarios/func_clock_timing_prects_ss_125.sdc",
                "common/02_soc_clock_timing_postcts_ss_125.sdc"]:
        collect(f"02/{rel}", (w02 / rel).read_text())
    for rep in ["clock_timing_check_report_common_prects_ss_125.txt",
                "clock_timing_check_report_func_prects_ss_125.txt",
                "clock_timing_check_report_common_postcts_ss_125.txt"]:
        collect(f"02/{rep}", norm((w02 / rep).read_text()))
    # 03
    collect("03/03_soc_clock_groups.sdc", (w03 / "common/03_soc_clock_groups.sdc").read_text())
    collect("03/clock_group_check_report_common.txt",
            norm((w03 / "clock_group_check_report_common.txt").read_text()))
    cov = w03 / "clock_group_coverage_report_common.xlsx"
    collect("03/cov_uncovered.csv", xlsx_sheet_to_csv(
        cov, "uncovered_cross_root_pairs",
        ["clock_a", "clock_b", "tree_root_a", "tree_root_b", "root_source_a", "root_source_b", "clock_kind_a"]))
    collect("03/cov_rule_effective_groups.csv", xlsx_sheet_to_csv(
        cov, "rule_effective_groups",
        ["scenario", "group_id", "relation_type", "group_index", "explicit_clocks",
         "auto_added_descendants", "excluded_descendants", "effective_clocks", "review_note"]))
    collect("03/cov_participation.csv", xlsx_sheet_to_csv(
        cov, "clock_participation",
        ["clock_name", "clock_kind", "tree_root", "root_source", "direct_source", "final_action", "group_count"]))
    # 04
    for rel in [
        "common/04_soc_io_pads.sdc",
        "common/04_soc_io_pads_prects_ss_125.sdc",
        "scenarios/gpio_in_io_pads.sdc",
    ]:
        collect(f"04/{rel}", (w04 / rel).read_text())
    for rep in [
        "io_pad_check_report_common_all_all.txt",
        "io_pad_check_report_common_prects_ss_125.txt",
        "io_pad_check_report_gpio_in_all_all.txt",
    ]:
        collect(f"04/{rep}", norm((w04 / rep).read_text()))
    # 20
    collect("20/common/20_harden_x_if.sdc", (w20 / "common/20_harden_x_if.sdc").read_text())
    collect("20/harden_x_if_check_report_common_all_all.txt",
            norm((w20 / "harden_x_if_check_report_common_all_all.txt").read_text()))


# ----------------------------------------------------------------------------
# diff
# ----------------------------------------------------------------------------
def compare():
    art_files = sorted(p.relative_to(ART).as_posix() for p in ART.rglob("*") if p.is_file())
    exp_files = sorted(p.relative_to(EXP).as_posix() for p in EXP.rglob("*") if p.is_file()) if EXP.exists() else []
    fails = []
    for rel in art_files:
        a = (ART / rel).read_text()
        e_path = EXP / rel
        if not e_path.is_file():
            fails.append((rel, "MISSING in expected/"))
            continue
        if a != e_path.read_text():
            import difflib
            diff = list(difflib.unified_diff(
                e_path.read_text().splitlines(), a.splitlines(),
                fromfile=f"expected/{rel}", tofile=f"work/{rel}", lineterm=""))
            fails.append((rel, "\n".join(diff[:24])))
    for rel in exp_files:
        if rel not in art_files:
            fails.append((rel, "MISSING in work artifacts"))
    return art_files, fails


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--update", action="store_true", help="(re)write expected/ baseline")
    args = ap.parse_args()

    if WORK.exists():
        shutil.rmtree(WORK)
    w01 = WORK / "01_soc_clocks"
    w02 = WORK / "02_soc_clock_timing"
    w03 = WORK / "03_soc_clock_groups"
    w04 = WORK / "04_soc_io_pads"
    w20 = WORK / "20_harden_x_if"

    run_01(w01)
    run_01_edge_checks(WORK / "01_soc_clocks_edge_checks")
    run_01_bad_source_checks(WORK / "01_soc_clocks_bad_source_checks")
    run_01_multi_target_checks(WORK / "01_soc_clocks_multi_target_checks")
    clocks = active_clocks(w01 / "clock_inventory.csv")
    run_02(w02, w01 / "clock_inventory.csv", clocks)
    run_03(w03)
    run_04(w04)
    run_20(w20)
    collect_artifacts(w01, w02, w03, w04, w20)

    art_files, _ = compare()

    if args.update or not EXP.exists():
        if EXP.exists():
            shutil.rmtree(EXP)
        shutil.copytree(ART, EXP)
        print(f"baseline written: {len(art_files)} artifact(s) -> {EXP}")
        for f in art_files:
            print(f"  + {f}")
        return 0

    _, fails = compare()
    print(f"01->02->03->04->20 regression: {len(art_files)} artifact(s) checked")
    if not fails:
        print("RESULT: PASS (all artifacts match expected/)")
        return 0
    print(f"RESULT: FAIL ({len(fails)} mismatch)")
    for rel, detail in fails:
        print(f"\n--- {rel} ---\n{detail}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
