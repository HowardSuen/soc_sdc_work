#!/usr/bin/env python3
"""Regression smoke tests for run_stage2_merge_delay.tcl.

The tests run with plain tclsh plus a small mock of the PrimeTime collection
and direction APIs. They do not replace PrimeTime validation, but they keep
parsing, matching, reporting, and static SDC emission deterministic.
"""

from __future__ import print_function

import contextlib
import importlib.util
import io
import os
import re
import shutil
import subprocess
import sys


HERE = os.path.abspath(os.path.dirname(__file__))
TOOL = os.path.abspath(os.path.join(HERE, "..", "run_stage2_merge_delay.tcl"))
REPORT_TOOL = os.path.abspath(os.path.join(HERE, "..", "run_stage2_report.py"))
WORK = os.path.join(HERE, "work")


DEFAULT_PT_PRELUDE = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg/Q out
    u_src_reg/CP in
    u_up/data_o out
    u_up/u_reg/Q out
    u_h0/cfg_i in
    u_h0/async_i in
    u_h0/unused_i in
    u_h0/other_i in
    u_h0/u_reg/D in
    u_h0/u_cfg_reg/D in
    u_h0/u_mode_reg/D in
    u_h0/i_niu_rst_n in
    u_h0/o_niu_rst_n out
    top_rst_n out
    u_mid/in_i in
    u_mid/out_o out
}

proc current_design {} {
    return current_integration_top
}

proc sizeof_collection {coll} {
    return [llength $coll]
}

proc foreach_in_collection {var coll body} {
    upvar 1 $var item
    foreach item $coll {
        uplevel 1 $body
    }
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    return [lindex $args end]
}

proc get_ports {args} {
    return [lindex $args end]
}

proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "full_name"} {
        return $name
    }
    if {$attr eq "direction" && [info exists ::PT_MOCK_DIRECTIONS($name)]} {
        return $::PT_MOCK_DIRECTIONS($name)
    }
    return ""
}
'''


def write_file(path, text):
    directory = os.path.dirname(path)
    if directory and not os.path.isdir(directory):
        os.makedirs(directory)
    with open(path, "w") as fout:
        fout.write(text)


def read_file(path):
    with open(path, "r") as fin:
        return fin.read()


def run_case(case_name, top_sdc, harden_sdc, extra_build_args=None, extra_hardens=None, prelude="", post_build_tcl=""):
    case_dir = os.path.join(WORK, case_name)
    if os.path.isdir(case_dir):
        shutil.rmtree(case_dir)
    os.makedirs(case_dir)
    top_path = os.path.join(case_dir, "top.sdc")
    harden_path = os.path.join(case_dir, "harden.sdc")
    harden_list = os.path.join(case_dir, "harden_list.csv")
    out_sdc = os.path.join(case_dir, "generated_e2e_delay.sdc")
    out_report = os.path.join(case_dir, "integration_delay_merge.rpt")
    out_removed = os.path.join(case_dir, "merged_delay_removed.sdc")
    out_review = os.path.join(case_dir, "unmerged_delay_review.rpt")
    out_final = os.path.join(case_dir, "top_flatten.sdc")
    out_summary = os.path.join(case_dir, "delay_path_summary")
    out_trace = os.path.join(case_dir, "stage2_live.log")
    driver = os.path.join(case_dir, "run.tcl")
    write_file(top_path, top_sdc)
    write_file(harden_path, harden_sdc)
    rows = [
        "harden_name,inst_path,clean_sdc,delay_candidate_file,netlist,module",
        "h0,u_h0,%s,,h0.v,harden0" % harden_path.replace("\\", "/"),
    ]
    for item in extra_hardens or []:
        clean_sdc = ""
        module = item[2] if len(item) > 2 else item[0]
        if len(item) > 3:
            clean_sdc = os.path.join(case_dir, "%s_clean.sdc" % item[0])
            write_file(clean_sdc, item[3])
            clean_sdc = clean_sdc.replace("\\", "/")
        rows.append("%s,%s,%s,,,%s" % (item[0], item[1], clean_sdc, module))
    write_file(harden_list, "\n".join(rows) + "\n")
    args = [
        "-top_sdc", top_path.replace("\\", "/"),
        "-harden_list", harden_list.replace("\\", "/"),
        "-out_e2e_sdc", out_sdc.replace("\\", "/"),
        "-out_report", out_report.replace("\\", "/"),
        "-out_removed_sdc", out_removed.replace("\\", "/"),
        "-out_review_rpt", out_review.replace("\\", "/"),
        "-out_trace_file", out_trace.replace("\\", "/"),
    ]
    if extra_build_args:
        args.extend(extra_build_args)
    complete_prelude = DEFAULT_PT_PRELUDE + "\n" + prelude
    driver_text = '%s\nset ::STAGE2_AUTO_RUN false\nsource "%s"\nstage2_delay::build %s\n' % (
        complete_prelude,
        TOOL.replace("\\", "/"),
        " ".join('"%s"' % arg for arg in args),
    )
    if post_build_tcl:
        driver_text += post_build_tcl + "\n"
    write_file(driver, driver_text)
    proc = subprocess.Popen(
        ["tclsh", driver],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=case_dir,
    )
    stdout, stderr = proc.communicate()
    return {
        "code": proc.returncode,
        "stdout": stdout.decode("utf-8", "replace"),
        "stderr": stderr.decode("utf-8", "replace"),
        "case_dir": case_dir,
        "out_sdc": out_sdc,
        "report": out_report,
        "removed": out_removed,
        "review": out_review,
        "final": out_final,
        "summary": out_summary,
        "trace": out_trace,
        "driver": driver,
    }


def assert_contains(path, needle):
    text = read_file(path)
    if needle not in text:
        raise AssertionError("Expected %r in %s\n--- file ---\n%s" % (needle, path, text))


def assert_exists(path):
    if not os.path.exists(path):
        raise AssertionError("Expected path to exist: %s" % path)


def assert_not_contains(path, needle):
    text = read_file(path)
    if needle in text:
        raise AssertionError("Did not expect %r in %s\n--- file ---\n%s" % (needle, path, text))


def assert_text_contains(text, needle):
    if needle not in text:
        raise AssertionError("Expected %r in text\n--- text ---\n%s" % (needle, text))


def delay_command_lines(path):
    return [
        line.strip()
        for line in read_file(path).splitlines()
        if line.strip().startswith(("set_max_delay ", "set_min_delay "))
    ]


def stat_value(text, name):
    match = re.search(r"(?:^|[,: ])%s=([0-9]+)" % re.escape(name), text)
    if not match:
        raise AssertionError("Statistic %s not found in text:\n%s" % (name, text))
    return int(match.group(1))


def get_pins_list(prefix, indices):
    return "[list %s]" % " ".join(
        "[get_pins {%s[%d]}]" % (prefix, index) for index in indices
    )


def get_ports_list(prefix, indices):
    return "[list %s]" % " ".join(
        "[get_ports {%s[%d]}]" % (prefix, index) for index in indices
    )


def validate_static_sdc(path):
    """Source generated SDC with Tcl collection stubs and validate its shape."""
    text = read_file(path)
    for forbidden in ("all_fanin", "all_fanout", "foreach_in_collection"):
        if forbidden in text:
            raise AssertionError("Runtime PT query leaked into generated SDC %s: %s" % (path, forbidden))
    validator = os.path.join(os.path.dirname(path), "validate_static_sdc.tcl")
    write_file(
        validator,
        r'''
set ::SDC_DELAY_COUNT 0
proc get_pins {args} { return [lindex $args end] }
proc get_ports {args} { return [lindex $args end] }
proc get_cells {args} { return [lindex $args end] }
proc get_nets {args} { return [lindex $args end] }
proc set_max_delay {args} {
    if {[lsearch -exact $args -from] < 0 || [lsearch -exact $args -to] < 0} {
        error "generated set_max_delay must have explicit -from and -to"
    }
    incr ::SDC_DELAY_COUNT
}
proc set_min_delay {args} {
    if {[lsearch -exact $args -from] < 0 || [lsearch -exact $args -to] < 0} {
        error "generated set_min_delay must have explicit -from and -to"
    }
    incr ::SDC_DELAY_COUNT
}
source [lindex $argv 0]
if {$::SDC_DELAY_COUNT == 0} {
    error "generated SDC contains no delay command"
}
puts "STATIC_SDC_VALID commands=$::SDC_DELAY_COUNT"
''',
    )
    proc = subprocess.Popen(
        ["tclsh", validator, path],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=os.path.dirname(path),
    )
    stdout, stderr = proc.communicate()
    output = stdout.decode("utf-8", "replace")
    errors = stderr.decode("utf-8", "replace")
    if proc.returncode != 0:
        raise AssertionError(
            "Generated SDC failed Tcl source validation: %s\nstdout=%s\nstderr=%s"
            % (path, output, errors)
        )
    assert_text_contains(output, "STATIC_SDC_VALID commands=")


def assert_generated_delays_have_explicit_endpoints(path):
    text = read_file(path)
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("set_max_delay ") or stripped.startswith("set_min_delay "):
            if " -from " not in stripped:
                raise AssertionError("Generated delay missing -from in %s:\n%s" % (path, stripped))
            if " -to " not in stripped:
                raise AssertionError("Generated delay missing -to in %s:\n%s" % (path, stripped))


def require_ok(result):
    if result["code"] != 0:
        raise AssertionError(
            "case failed\nstdout=%s\nstderr=%s\ndriver=%s"
            % (result["stdout"], result["stderr"], read_file(result["driver"]))
    )
    if os.path.exists(result["out_sdc"]):
        assert_generated_delays_have_explicit_endpoints(result["out_sdc"])


def test_release_identity_is_reconstructed_without_plaintext_constant():
    expected = "".join(chr(code) for code in (72, 111, 119, 97, 114, 100))
    tamper_message = "Who is your daddy?"
    assert_not_contains(TOOL, expected)
    assert_not_contains(REPORT_TOOL, expected)

    module_spec = importlib.util.spec_from_file_location("stage2_report_module", REPORT_TOOL)
    report_module = importlib.util.module_from_spec(module_spec)
    module_spec.loader.exec_module(report_module)
    if report_module.release_identity() != expected:
        raise AssertionError("Unexpected reconstructed Python release identity")
    if report_module.guarded_release_identity() != expected:
        raise AssertionError("Unexpected guarded Python release identity")
    original_release_identity = report_module.release_identity
    report_module.release_identity = lambda: "Somebody"
    if report_module.guarded_release_identity() != tamper_message:
        raise AssertionError("Python identity guard did not detect tampering")
    captured_banner = io.StringIO()
    with contextlib.redirect_stdout(captured_banner):
        report_module.print_author_banner()
    if "  Author  : %s" % tamper_message not in captured_banner.getvalue():
        raise AssertionError("Python final author banner did not expose tampering")
    report_module.release_identity = original_release_identity

    driver = os.path.join(WORK, "release_identity.tcl")
    write_file(
        driver,
        'set ::STAGE2_AUTO_RUN false\n'
        'source "%s"\n'
        'puts [stage2_delay::release_identity]\n'
        'puts [stage2_delay::guarded_release_identity]\n'
        'rename stage2_delay::release_identity stage2_delay::original_release_identity\n'
        'proc stage2_delay::release_identity {} {return Somebody}\n'
        'puts [lindex [stage2_delay::author_banner_lines] 3]\n'
        % TOOL.replace("\\", "/"),
    )
    proc = subprocess.Popen(
        ["tclsh", driver],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=WORK,
    )
    stdout, stderr = proc.communicate()
    if proc.returncode != 0:
        raise AssertionError(
            "identity reconstruction failed\nstdout=%s\nstderr=%s"
            % (stdout.decode("utf-8", "replace"), stderr.decode("utf-8", "replace"))
        )
    identity_lines = stdout.decode("utf-8", "replace").strip().splitlines()
    if identity_lines != [expected, expected, "  Author  : %s" % tamper_message]:
        raise AssertionError("Unexpected Tcl identity guard output: %s" % identity_lines)


def test_complete_complete_merge():
    result = run_case(
        "complete_complete",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "E2E_DELAY_MERGE_VERSION")
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["removed"], "set_max_delay 2.0")
    assert_contains(result["removed"], "set_max_delay 5.0")
    assert_contains(result["report"], "Merged constraints              : 1")
    assert_contains(result["review"], "Overall result    : PASS")
    assert_contains(result["review"], "Highest severity  : NONE")
    assert_contains(result["review"], "Total reviews     : 0")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_live_trace_records_invalid_startpoint_object():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/u_reg/D"} {
        return [list u_src_reg/CP]
    }
    return {}
}
'''
    result = run_case(
        "live_trace_invalid_startpoint",
        "set_max_delay 2.0 -from [get_pins u_h0/async_i] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    trace = read_file(result["trace"])
    assert_text_contains(trace, "BUILD_START")
    assert_text_contains(trace, "PHASE match_delay_graph mode=auto")
    assert_text_contains(trace, "INVALID_STARTPOINT")
    assert_text_contains(trace, "top_id=CMD000001 harden_id=CMD000002")
    assert_text_contains(trace, "name=u_h0/async_i,direction=in,owner=u_h0")
    assert_text_contains(trace, "name=u_h0/u_reg/D,direction=in,owner=u_h0")
    assert_text_contains(trace, "BUILD_COMPLETE generated=0")
    review = read_file(result["review"])
    if review.index("[RUN_CONCLUSION]") > review.index("[DETAIL]"):
        raise AssertionError("Review conclusion must precede review details")
    assert_text_contains(review, "Overall result    : REVIEW_REQUIRED")
    assert_text_contains(review, "Highest severity  : ERROR")
    assert_text_contains(review, "[REASON_SUMMARY]")
    assert_text_contains(review, "ERROR     INVALID_STARTPOINT")
    assert_text_contains(review, "[DETAIL]")
    assert_text_contains(review, "[ERROR] reason=INVALID_STARTPOINT")


def test_pt_proven_input_clock_pin_is_accepted_as_startpoint():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/u_reg/D"} {
        return [list u_src_reg/CP]
    }
    return {}
}
'''
    result = run_case(
        "pt_proven_input_clock_startpoint",
        "set_max_delay 2.0 -from [get_pins u_src_reg/CP] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/CP}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["trace"], "STARTPOINT_PT_CONFIRMED")
    assert_contains(result["trace"], "name=u_src_reg/CP,direction=in,owner=,pt_startpoint=true")
    assert_not_contains(result["trace"], "INVALID_STARTPOINT")
    assert_not_contains(result["review"], "NO_TOP_SEGMENT_MATCHED")
    validate_static_sdc(result["out_sdc"])


def test_recursive_pt_proven_input_clock_pin_is_accepted():
    prelude = r'''
proc get_cells {args} {
    return [list [lindex $args end]]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return [list u_h0/cfg_i]
    }
    return [list [lindex $args end]]
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target eq "u_h0/u_reg/D"} {
        if {[lsearch -exact $args "-startpoints_only"] >= 0} {
            return [list u_src_reg/CP]
        }
        return [list u_h0/cfg_i u_src_reg/CP]
    }
    return {}
}
'''
    result = run_case(
        "recursive_pt_proven_input_clock_startpoint",
        "set_max_delay 2.0 -from [get_pins u_src_reg/CP] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/CP}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "RECURSIVE_MERGED")
    assert_contains(result["trace"], "STARTPOINT_PT_CONFIRMED")
    assert_not_contains(result["trace"], "INVALID_STARTPOINT")
    assert_not_contains(result["review"], "NO_TOP_SEGMENT_MATCHED")
    validate_static_sdc(result["out_sdc"])


def test_matrix_clock_pairs_skip_pt_disconnected_cross_pairs():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg_1/CP in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}

rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} {
        return true
    }
    return [stage2_default_get_attribute $obj $attr]
}

proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} {
        return [list u_src_reg/CP]
    }
    if {$target in {u_h1/cfg_i u_h1/u_reg/D}} {
        return [list u_src_reg_1/CP]
    }
    return {}
}
'''
    result = run_case(
        "matrix_clock_pairs_skip_disconnected_cross_pairs",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] [get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h1/cfg_i]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            (
                "h1",
                "u_h1",
                "harden1",
                "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n",
            )
        ],
        prelude=prelude,
    )
    require_ok(result)
    generated = read_file(result["out_sdc"])
    assert_text_contains(generated, "set_max_delay 7 -from [get_pins {u_src_reg/CP}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_text_contains(generated, "set_max_delay 8 -from [get_pins {u_src_reg_1/CP}] -through [get_pins {u_h1/cfg_i}] -to [get_pins {u_h1/u_reg/D}]")
    if generated.count("set_max_delay ") != 2:
        raise AssertionError("Expected only the two PT-connected matrix pairs:\n%s" % generated)
    trace = read_file(result["trace"])
    if trace.count("NO_PT_CONNECTIVITY_PAIR") != 2:
        raise AssertionError("Expected two skipped cross pairs:\n%s" % trace)
    if "INVALID_STARTPOINT" in trace:
        raise AssertionError("Disconnected matrix pairs must not become INVALID_STARTPOINT:\n%s" % trace)
    assert_text_contains(trace, "SPARSE_MATRIX_PLAN")
    assert_text_contains(trace, "product=4 pruned=2 retained=2")
    report = read_file(result["report"])
    if stat_value(report, "sparse_matrix_pairs_pruned") != 2:
        raise AssertionError("Expected two pairs pruned before expansion:\n%s" % report)
    assert_not_contains(result["review"], "NO_HARDEN_SEGMENT_MATCHED")
    assert_not_contains(result["review"], "NO_TOP_SEGMENT_MATCHED")
    assert_contains(result["final"], "# STAGE2_CONSUMED CMD000001")
    assert_not_contains(result["final"], "STAGE2_REWRITTEN CMD000001")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_sparse_matrix_numeric_bus_clock_bits_remain_exact_pins():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    {u_clk/CP[0]} in
    {u_clk/CP[1]} in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}

rename get_attribute stage2_numeric_bus_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {{u_clk/CP[0]} {u_clk/CP[1]}}} {
        return true
    }
    return [stage2_numeric_bus_get_attribute $obj $attr]
}

proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} {
        return [list {u_clk/CP[0]}]
    }
    if {$target in {u_h1/cfg_i u_h1/u_reg/D}} {
        return [list {u_clk/CP[1]}]
    }
    return {}
}
'''
    result = run_case(
        "sparse_numeric_bus_clock_bits",
        "set_max_delay 2.0 -from [list [get_pins -quiet -exact {u_clk/CP[0]}] [get_pins -exact {u_clk/CP[1]}]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h1/cfg_i]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            (
                "h1",
                "u_h1",
                "harden1",
                "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n",
            )
        ],
        prelude=prelude,
    )
    require_ok(result)
    trace = read_file(result["trace"])
    assert_text_contains(trace, "product=4 pruned=2 retained=2")
    generated = delay_command_lines(result["out_sdc"])
    if len(generated) != 2:
        raise AssertionError("Numeric bus clock bits were not treated as exact pins:\n%s" % read_file(result["out_sdc"]))
    assert_contains(result["out_sdc"], "-from [get_pins {u_clk/CP[0]}]")
    assert_contains(result["out_sdc"], "-from [get_pins {u_clk/CP[1]}]")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_matrix_pair_query_failure_does_not_silently_skip():
    prelude = r'''
rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name eq "u_src_reg/CP"} {
        return true
    }
    return [stage2_default_get_attribute $obj $attr]
}
'''
    result = run_case(
        "matrix_pair_query_failure_keeps_review",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h0/other_i]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    trace = read_file(result["trace"])
    if "NO_PT_CONNECTIVITY_PAIR" in trace:
        raise AssertionError("Unavailable PT query must not prove disconnection:\n%s" % trace)
    assert_text_contains(trace, "INVALID_STARTPOINT")
    assert_contains(result["final"], "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP]]")
    assert_contains(result["review"], "INVALID_STARTPOINT")


def test_sparse_matrix_missing_collection_iterator_is_unknown():
    original = (
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] "
        "[get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] "
        "[get_pins u_h1/cfg_i]]"
    )
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg_1/CP in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}
rename get_attribute stage2_missing_iterator_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} {
        return true
    }
    return [stage2_missing_iterator_get_attribute $obj $attr]
}
rename foreach_in_collection {}
'''
    result = run_case(
        "sparse_missing_collection_iterator",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=[
            "-generate_clock_group_review", "false",
            "-write_path_summary", "false",
        ],
        extra_hardens=[
            (
                "h1",
                "u_h1",
                "harden1",
                "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n",
            )
        ],
        prelude=prelude,
    )
    require_ok(result)
    trace = read_file(result["trace"])
    assert_text_contains(trace, "SPARSE_MATRIX_CLOCK_BATCH_FALLBACK")
    assert_text_contains(trace, "missing_collection_command:get_pins")
    if "NO_PT_CONNECTIVITY_PAIR" in trace or "SPARSE_MATRIX_PLAN" in trace:
        raise AssertionError("Missing collection iterator was treated as disconnection")
    if original not in delay_command_lines(result["final"]):
        raise AssertionError("Unknown collection query did not preserve the original matrix")


def test_pt_collection_runtime_iterator_failure_returns_unavailable():
    post_build_tcl = r'''
set rec [stage2_delay::object_record pin u_src_reg/CP in ""]
rename foreach_in_collection stage2_runtime_default_iterator
proc foreach_in_collection {var coll body} {
    error "mock iterator failure"
}
array set result [stage2_delay::pt_collection_for_records [list $rec] runtime-iterator]
puts "RUNTIME_ITERATOR_OK=$result(ok)"
puts "RUNTIME_ITERATOR_REASON=$result(reason)"
'''
    result = run_case(
        "pt_collection_runtime_iterator_failure",
        "",
        "",
        extra_build_args=[
            "-generate_clock_group_review", "false",
            "-write_path_summary", "false",
        ],
        post_build_tcl=post_build_tcl,
    )
    require_ok(result)
    assert_text_contains(result["stdout"], "RUNTIME_ITERATOR_OK=false")
    assert_text_contains(
        result["stdout"],
        "RUNTIME_ITERATOR_REASON=batch_collection_iteration_failed:mock iterator failure",
    )


def test_sparse_matrix_unresolved_fanin_object_is_unknown():
    original = (
        "set_max_delay 2.0 -from [get_pins u_src_reg/CP] "
        "-to [list [get_pins u_h0/other_i] [get_pins u_h0/unused_i]]"
    )
    prelude = r'''
rename get_attribute stage2_unresolved_fanin_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$name eq "opaque_startpoint" && $attr eq "full_name"} {
        error "mock full_name unavailable"
    }
    if {$attr eq "is_clock_pin" && $name eq "u_src_reg/CP"} {
        return true
    }
    return [stage2_unresolved_fanin_get_attribute $obj $attr]
}

proc get_object_name {obj} {
    if {[lindex $obj 0] eq "opaque_startpoint"} {
        error "mock object name unavailable"
    }
    return [lindex $obj 0]
}

proc all_fanin {args} {
    return [list opaque_startpoint]
}
'''
    result = run_case(
        "sparse_unresolved_fanin_object",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    trace = read_file(result["trace"])
    assert_text_contains(trace, "unable to resolve collection object full_name")
    if "NO_PT_CONNECTIVITY_PAIR" in trace or "SPARSE_MATRIX_PLAN" in trace:
        raise AssertionError("An unresolved collection handle was treated as disconnection")
    if original not in delay_command_lines(result["final"]):
        raise AssertionError("An unresolved fanin object did not preserve the original matrix")


def test_sparse_matrix_prune_can_be_disabled_for_legacy_diagnosis():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg_1/CP in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}
rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} {
        return true
    }
    return [stage2_default_get_attribute $obj $attr]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} { return [list u_src_reg/CP] }
    if {$target in {u_h1/cfg_i u_h1/u_reg/D}} { return [list u_src_reg_1/CP] }
    return {}
}
'''
    result = run_case(
        "sparse_matrix_prune_disabled_legacy",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] [get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h1/cfg_i]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-sparse_matrix_prune", "false"],
        extra_hardens=[
            ("h1", "u_h1", "harden1", "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n")
        ],
        prelude=prelude,
    )
    require_ok(result)
    trace = read_file(result["trace"])
    assert_text_contains(trace, "action=EXPAND")
    if "SPARSE_MATRIX_PLAN" in trace:
        raise AssertionError("Disabled sparse pruning must use the legacy expansion path:\n%s" % trace)
    if trace.count("NO_PT_CONNECTIVITY_PAIR") != 2:
        raise AssertionError("Legacy diagnosis path must still prune two disconnected pairs later:\n%s" % trace)


def test_sparse_matrix_clock_batch_failure_falls_back_without_loss():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg_1/CP in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}
rename get_pins stage2_default_get_pins
proc get_pins {args} {
    set patterns [lindex $args end]
    if {[llength $patterns] > 1} {
        error "mock rejects multi-pattern get_pins"
    }
    return [stage2_default_get_pins {*}$args]
}
rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} {
        return true
    }
    return [stage2_default_get_attribute $obj $attr]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} { return [list u_src_reg/CP] }
    if {$target in {u_h1/cfg_i u_h1/u_reg/D}} { return [list u_src_reg_1/CP] }
    return {}
}
'''
    result = run_case(
        "sparse_matrix_clock_batch_fallback",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] [get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h1/cfg_i]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            ("h1", "u_h1", "harden1", "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n")
        ],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["trace"], "SPARSE_MATRIX_CLOCK_BATCH_FALLBACK")
    assert_contains(result["trace"], "product=4 pruned=2 retained=2")
    if len(delay_command_lines(result["out_sdc"])) != 2:
        raise AssertionError("Clock batch fallback changed sparse E2E output")


def test_sparse_matrix_all_disconnected_uses_compact_command_bookkeeping():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg_1/CP in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}
rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} {
        return true
    }
    return [stage2_default_get_attribute $obj $attr]
}
proc all_fanin {args} { return {} }
'''
    original = "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] [get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h1/cfg_i]]"
    result = run_case(
        "sparse_matrix_all_disconnected",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            ("h1", "u_h1", "harden1", "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n")
        ],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["trace"], "product=4 pruned=4 retained=0")
    assert_contains(result["final"], "# STAGE2_CONSUMED CMD000001")
    assert_not_contains(result["final"], original)
    assert_contains(result["removed"], "PT_DISCONNECTED_MATRIX_PAIRS")
    assert_contains(result["removed"], "pruned=4 retained=0 product=4")
    assert_contains(
        os.path.join(result["summary"], "00_index.csv"),
        '"top.csv","0","0","0","0","0","1","0/1","0"',
    )


def test_sparse_matrix_partial_rewrite_excludes_pruned_cross_pair():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS { u_src_reg_1/CP in }
rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} {
        return true
    }
    return [stage2_default_get_attribute $obj $attr]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} { return [list u_src_reg/CP] }
    return {}
}
'''
    result = run_case(
        "sparse_matrix_partial_rewrite",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] [get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h0/u_reg/D]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["trace"], "product=4 pruned=1 retained=3")
    final = read_file(result["final"])
    assert_text_contains(final, "# STAGE2_REWRITTEN CMD000001")
    assert_text_contains(final, "-from [get_pins {u_src_reg/CP}] -to [get_pins {u_h0/u_reg/D}]")
    assert_text_contains(final, "-from [get_pins {u_src_reg_1/CP}] -to [get_pins {u_h0/u_reg/D}]")
    if "-from [get_pins {u_src_reg_1/CP}] -to [get_pins {u_h0/cfg_i}]" in final:
        raise AssertionError("PT-disconnected cross pair leaked back into final rewrite:\n%s" % final)


def test_sparse_matrix_prunes_before_pair_limit_but_falls_back_if_retained_exceeds_limit():
    diagonal_prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg_1/CP in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}
rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} { return true }
    return [stage2_default_get_attribute $obj $attr]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} { return [list u_src_reg/CP] }
    if {$target in {u_h1/cfg_i u_h1/u_reg/D}} { return [list u_src_reg_1/CP] }
    return {}
}
'''
    result = run_case(
        "sparse_matrix_under_pair_limit",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] [get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h1/cfg_i]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-max_segment_pairs", "2"],
        extra_hardens=[
            ("h1", "u_h1", "harden1", "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n")
        ],
        prelude=diagonal_prelude,
    )
    require_ok(result)
    assert_contains(result["trace"], "product=4 pruned=2 retained=2")
    assert_not_contains(result["review"], "MATRIX_EXPANSION_LIMIT")

    fallback_prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg_1/CP in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}
rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} { return true }
    return [stage2_default_get_attribute $obj $attr]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target eq "u_h0/cfg_i"} { return [list u_src_reg/CP u_src_reg_1/CP] }
    if {$target eq "u_h1/cfg_i"} { return [list u_src_reg/CP] }
    if {$target eq "u_h0/u_reg/D"} { return [list u_src_reg/CP] }
    if {$target eq "u_h1/u_reg/D"} { return [list u_src_reg_1/CP] }
    return {}
}
'''
    fallback = run_case(
        "sparse_matrix_retained_over_limit_fallback",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] [get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h1/cfg_i]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-max_segment_pairs", "2"],
        extra_hardens=[
            ("h1", "u_h1", "harden1", "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n")
        ],
        prelude=fallback_prelude,
    )
    require_ok(fallback)
    assert_contains(fallback["trace"], "SPARSE_MATRIX_FALLBACK")
    assert_contains(fallback["review"], "MATRIX_EXPANSION_LIMIT")
    assert_contains(fallback["final"], "set_max_delay 2.0 -from [list")


def test_sparse_matrix_scale_200x200_materializes_only_connected_diagonal():
    size = 200
    from_expr = "[list %s]" % " ".join(
        "[get_pins {u_src_%d/CP}]" % index for index in range(size)
    )
    to_expr = "[list %s]" % " ".join(
        "[get_pins {u_h%d/cfg_i}]" % index for index in range(size)
    )
    extra_hardens = [
        (
            "h%d" % index,
            "u_h%d" % index,
            "harden%d" % index,
            "set_max_delay 5.0 -from [get_pins u_h%d/cfg_i] -to [get_pins u_h%d/u_reg/D]\n"
            % (index, index),
        )
        for index in range(1, size)
    ]
    prelude = r'''
rename get_attribute stage2_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && [regexp {^u_src_[0-9]+/CP$} $name]} {
        return true
    }
    if {$attr eq "direction" &&
        ([regexp {^u_src_[0-9]+/CP$} $name] ||
         [regexp {^u_h[0-9]+/(cfg_i|u_reg/D)$} $name])} {
        return in
    }
    return [stage2_default_get_attribute $obj $attr]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {[regexp {^u_h([0-9]+)/(cfg_i|u_reg/D)$} $target -> index]} {
        return [list u_src_${index}/CP]
    }
    return {}
}
'''
    result = run_case(
        "sparse_matrix_scale_200x200",
        "set_max_delay 2.0 -from %s -to %s\n" % (from_expr, to_expr),
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=[
            "-max_segment_pairs", "500",
            "-generate_clock_group_review", "false",
            "-write_path_summary", "false",
        ],
        extra_hardens=extra_hardens,
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["trace"], "product=40000 pruned=39800 retained=200")
    assert_not_contains(result["review"], "MATRIX_EXPANSION_LIMIT")
    generated = delay_command_lines(result["out_sdc"])
    if len(generated) != size:
        raise AssertionError("Expected %d connected E2E commands, got %d" % (size, len(generated)))
    report = read_file(result["report"])
    if stat_value(report, "sparse_matrix_pairs_retained") != size:
        raise AssertionError("Sparse scale case retained more than the connected diagonal:\n%s" % report)
    if stat_value(report, "sparse_matrix_pairs_pruned") != size * size - size:
        raise AssertionError("Sparse scale case pruned count mismatch:\n%s" % report)


def test_sparse_matrix_non_exact_clock_selector_is_never_pruned():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS { u_src_0/CP in }

rename get_pins stage2_non_exact_get_pins
proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set out {}
    foreach pattern [lindex $args end] {
        if {$pattern eq {u_src_*/CP} ||
            $pattern eq {u_src_[0-9]/CP} ||
            $pattern eq {$clock_pin}} {
            lappend out u_src_0/CP
        } else {
            lappend out $pattern
        }
    }
    return $out
}

rename get_attribute stage2_non_exact_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name eq "u_src_0/CP"} {
        return true
    }
    return [stage2_non_exact_get_attribute $obj $attr]
}

proc all_fanin {args} {
    return [list u_src_0/CP]
}
'''
    selectors = (
        ("star", "[get_pins {u_src_*/CP}]"),
        ("bracket", "[get_pins {u_src_[0-9]/CP}]"),
        ("variable", "[get_pins {$clock_pin}]"),
        ("regexp", "[get_pins -regexp {u_src_0/CP}]"),
        ("hierarchical", "[get_pins -hierarchical {u_src_0/CP}]"),
        ("hier", "[get_pins -hier {u_src_0/CP}]"),
        ("nocase", "[get_pins -nocase {u_src_0/CP}]"),
        ("filter", "[get_pins -filter {direction == in} {u_src_0/CP}]"),
        ("of_objects", "[get_pins -of_objects [get_cells u_src_0]]"),
    )
    for token, from_expr in selectors:
        original = (
            "set_max_delay 2.0 -from %s "
            "-to [list [get_pins u_h0/cfg_i] [get_pins u_h0/other_i]]"
            % from_expr
        )
        for enabled in ("true", "false"):
            result = run_case(
                "sparse_non_exact_clock_selector_%s_%s" % (token, enabled),
                original + "\n",
                "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
                extra_build_args=["-sparse_matrix_prune", enabled],
                prelude=prelude,
            )
            require_ok(result)
            trace = read_file(result["trace"])
            if "NO_PT_CONNECTIVITY_PAIR" in trace or "SPARSE_MATRIX_PLAN" in trace:
                raise AssertionError(
                    "A non-exact selector was treated as a concrete clock pin:\n%s" % trace
                )
            if original not in delay_command_lines(result["final"]):
                raise AssertionError("Non-exact selector constraint was not preserved")


def test_non_exact_to_and_through_selectors_are_preserved():
    cases = (
        (
            "to_regexp",
            "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins -regexp {u_h0/cfg_i}]",
        ),
        (
            "to_wildcard",
            "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins {u_h0/*_i}]",
        ),
        (
            "port_nocase",
            "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_ports -nocase cfg_top]",
        ),
        (
            "through_hierarchical",
            "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -through [get_pins -hierarchical {u_h0/cfg_i}] -to [get_pins u_h0/u_reg/D]",
        ),
        (
            "through_dynamic",
            "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -through [get_pins [all_registers]] -to [get_pins u_h0/u_reg/D]",
        ),
    )
    for token, original in cases:
        for enabled in ("true", "false"):
            result = run_case(
                "non_exact_to_through_%s_%s" % (token, enabled),
                original + "\n",
                "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
                extra_build_args=["-sparse_matrix_prune", enabled],
            )
            require_ok(result)
            if original not in delay_command_lines(result["final"]):
                raise AssertionError("A non-exact to/through selector was rewritten")
            assert_not_contains(result["final"], "STAGE2_REWRITTEN CMD000001")
            assert_contains(result["review"], "CLOCK_OR_UNKNOWN_OBJECT")
            assert_not_contains(result["trace"], "SPARSE_MATRIX_PLAN")


def test_sparse_matrix_compact_clock_record_is_conservative():
    prelude = r'''
rename get_pins stage2_compact_clock_get_pins
proc get_pins {args} {
    set out {}
    foreach pattern [lindex $args end] {
        if {$pattern eq {clk[*]}} {
            foreach idx {0 1 2 3} {
                lappend out [format {clk[%d]} $idx]
            }
        } else {
            lappend out $pattern
        }
    }
    return $out
}
rename get_attribute stage2_compact_clock_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && [regexp {^clk\[[0-9]+\]$} $name]} {
        return true
    }
    return [stage2_compact_clock_get_attribute $obj $attr]
}
'''
    post_build_tcl = r'''
set members {}
set names {}
foreach idx {0 1 2 3} {
    set name [format {clk[%d]} $idx]
    lappend names $name
    lappend members [stage2_delay::object_record pin $name in ""]
}
set compact [stage2_delay::make_compact_bus_record $members {clk[*]} $names]
puts "COMPACT_CLOCK_FLAGS=[stage2_delay::pt_clock_pin_flags [list $compact]]"
'''
    result = run_case(
        "sparse_compact_clock_record",
        "",
        "",
        extra_build_args=["-generate_clock_group_review", "false"],
        prelude=prelude,
        post_build_tcl=post_build_tcl,
    )
    require_ok(result)
    assert_text_contains(result["stdout"], "COMPACT_CLOCK_FLAGS=0")


def test_sparse_matrix_clock_metadata_batch_respects_disable_switch():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg_1/CP in
    u_h1/cfg_i in
    u_h1/u_reg/D in
}
set ::CLOCK_METADATA_MULTI_CALLS 0
rename get_pins stage2_clock_batch_disabled_get_pins
proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set patterns [lindex $args end]
    if {[llength $patterns] > 1} {
        incr ::CLOCK_METADATA_MULTI_CALLS
        error "metadata batching must be disabled"
    }
    return [stage2_clock_batch_disabled_get_pins {*}$args]
}
rename get_attribute stage2_clock_batch_disabled_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {u_src_reg/CP u_src_reg_1/CP}} {
        return true
    }
    return [stage2_clock_batch_disabled_get_attribute $obj $attr]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} { return [list u_src_reg/CP] }
    if {$target in {u_h1/cfg_i u_h1/u_reg/D}} { return [list u_src_reg_1/CP] }
    return {}
}
'''
    result = run_case(
        "sparse_clock_metadata_batch_disabled",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/CP] [get_pins u_src_reg_1/CP]] -to [list [get_pins u_h0/cfg_i] [get_pins u_h1/cfg_i]]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-metadata_batch_enabled", "false"],
        extra_hardens=[
            ("h1", "u_h1", "harden1", "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n")
        ],
        prelude=prelude,
        post_build_tcl='puts "CLOCK_METADATA_MULTI_CALLS=$::CLOCK_METADATA_MULTI_CALLS"',
    )
    require_ok(result)
    assert_text_contains(result["stdout"], "CLOCK_METADATA_MULTI_CALLS=0")
    assert_contains(result["trace"], "SPARSE_MATRIX_CLOCK_BATCH_DISABLED records=2 mode=individual")
    assert_contains(result["trace"], "product=4 pruned=2 retained=2")
    report = read_file(result["report"])
    if stat_value(report, "sparse_matrix_clock_batches") != 0:
        raise AssertionError("Disabled clock metadata path issued a batch query")
    if len(delay_command_lines(result["out_sdc"])) != 2:
        raise AssertionError("Disabled clock metadata batching changed E2E output")


def test_sparse_matrix_top_port_limit_rolls_back_to_original_command():
    original = (
        "set_max_delay 2.0 -from [list [get_pins src0/CP] [get_pins src1/CP]] "
        "-to [list [get_pins u_h0/cfg_i] [get_ports cfg_top]]"
    )
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    src0/CP in
    src1/CP in
    cfg_top out
    u_h1/cfg_i in
    u_h2/cfg_i in
}
rename get_attribute stage2_sparse_port_limit_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {src0/CP src1/CP}} {
        return true
    }
    return [stage2_sparse_port_limit_get_attribute $obj $attr]
}
proc get_nets {args} {
    if {[lindex $args end] eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}
rename get_pins stage2_sparse_port_limit_get_pins
proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        if {[lindex $args end] eq "cfg_net"} {
            return [list u_h0/cfg_i u_h1/cfg_i u_h2/cfg_i]
        }
        return {}
    }
    return [stage2_sparse_port_limit_get_pins {*}$args]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} {
        return [list src0/CP]
    }
    return {}
}
'''
    result = run_case(
        "sparse_then_top_port_matrix_limit",
        original + "\n",
        "",
        extra_build_args=[
            "-max_segment_pairs", "3",
            "-generate_clock_group_review", "false",
        ],
        extra_hardens=[
            ("h1", "u_h1", "harden1"),
            ("h2", "u_h2", "harden2"),
        ],
        prelude=prelude,
    )
    require_ok(result)
    trace = read_file(result["trace"])
    assert_text_contains(trace, "product=4 pruned=1 retained=3")
    assert_text_contains(trace, "SPARSE_MATRIX_ROLLBACK")
    assert_text_contains(trace, "phase=TOP_PORT_MAP original=preserved")
    if delay_command_lines(result["final"]) != [original]:
        raise AssertionError("Top-port limit did not preserve the exact active command")
    assert_not_contains(result["final"], "STAGE2_REWRITTEN CMD000001")
    assert_not_contains(result["removed"], "PT_DISCONNECTED_MATRIX_PAIRS")


def test_sparse_matrix_top_port_under_limit_partial_and_complete_rewrite():
    original = (
        "set_max_delay 2.0 -from [list [get_pins src0/CP] [get_pins src1/CP]] "
        "-to [list [get_pins u_h0/cfg_i] [get_ports cfg_top]]"
    )
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    src0/CP in
    src1/CP in
    cfg_top out
    u_h1/cfg_i in
}
rename get_attribute stage2_sparse_port_under_limit_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name in {src0/CP src1/CP}} {
        return true
    }
    return [stage2_sparse_port_under_limit_get_attribute $obj $attr]
}
proc get_nets {args} {
    if {[lindex $args end] eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}
rename get_pins stage2_sparse_port_under_limit_get_pins
proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        if {[lindex $args end] eq "cfg_net"} {
            return [list u_h0/cfg_i u_h1/cfg_i]
        }
        return {}
    }
    return [stage2_sparse_port_under_limit_get_pins {*}$args]
}
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target in {u_h0/cfg_i u_h0/u_reg/D}} {
        return [list src0/CP]
    }
    return {}
}
'''
    common_args = [
        "-max_segment_pairs", "10",
        "-generate_clock_group_review", "false",
    ]
    partial = run_case(
        "sparse_then_top_port_under_limit_partial",
        original + "\n",
        "",
        extra_build_args=common_args,
        extra_hardens=[("h1", "u_h1", "harden1")],
        prelude=prelude,
    )
    require_ok(partial)
    assert_contains(partial["trace"], "product=4 pruned=1 retained=3")
    assert_not_contains(partial["trace"], "SPARSE_MATRIX_ROLLBACK")
    assert_contains(partial["report"], "TOP_PORT_BOUNDARY_MAP_KEEP_ORIGINAL")
    expected_partial = [
        "set_max_delay 2 -from [get_pins {src0/CP}] -to [get_pins {u_h0/cfg_i}]",
        "set_max_delay 2 -from [get_pins {src0/CP}] -to [get_ports {cfg_top}]",
    ]
    if delay_command_lines(partial["final"]) != expected_partial:
        raise AssertionError(
            "Sparse/top-port partial rewrite lost or retained the wrong pairs:\n%s"
            % read_file(partial["final"])
        )

    complete = run_case(
        "sparse_then_top_port_under_limit_complete",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=common_args,
        extra_hardens=[("h1", "u_h1", "harden1")],
        prelude=prelude,
    )
    require_ok(complete)
    assert_contains(complete["trace"], "product=4 pruned=1 retained=3")
    assert_contains(complete["report"], "TOP_PORT_BOUNDARY_MAP_CONSUMED")
    if len(delay_command_lines(complete["out_sdc"])) != 1:
        raise AssertionError("Sparse/top-port complete rewrite changed generated E2E cardinality")
    if original in delay_command_lines(complete["final"]):
        raise AssertionError("Sparse/top-port complete group kept the original command")
    assert_contains(complete["final"], "STAGE2_CONSUMED CMD000001")

    failed_emit_prelude = prelude + r'''
rename all_fanin stage2_sparse_port_connected_all_fanin
proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target eq "u_h0/cfg_i"} {
        return [list src0/CP]
    }
    return {}
}
'''
    failed_emit = run_case(
        "sparse_then_top_port_duplicate_emit_failure",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=common_args,
        extra_hardens=[("h1", "u_h1", "harden1")],
        prelude=failed_emit_prelude,
    )
    failed_emit_legacy = run_case(
        "sparse_then_top_port_duplicate_emit_failure_legacy",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=common_args + ["-sparse_matrix_prune", "false"],
        extra_hardens=[("h1", "u_h1", "harden1")],
        prelude=failed_emit_prelude,
    )
    for failed_case in (failed_emit, failed_emit_legacy):
        require_ok(failed_case)
        if delay_command_lines(failed_case["out_sdc"]):
            raise AssertionError("A failed duplicate path unexpectedly emitted E2E")
        failed_lines = delay_command_lines(failed_case["final"])
        for expected in expected_partial + [
            "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]"
        ]:
            if expected not in failed_lines:
                raise AssertionError(
                    "A failed first duplicate path consumed an un-emitted constraint:\n%s"
                    % read_file(failed_case["final"])
                )
        assert_not_contains(failed_case["final"], "STAGE2_CONSUMED CMD000002")
    if delay_command_lines(failed_emit["final"]) != delay_command_lines(failed_emit_legacy["final"]):
        raise AssertionError("Sparse on/off changed failed-duplicate final constraints")


def test_structural_passthrough_skips_getters_and_matrix_expansion():
    top_from = get_pins_list("u_h0/u_src/Q", range(4896))
    top_to = get_pins_list("u_h0/u_dst/D", range(2156))
    top_sdc = (
        "set_max_delay 2.0 -from %s -to %s\n"
        "set_min_delay 0.0 -from %s -to %s\n"
        % (top_from, top_to, top_from, top_to)
    )
    harden_sdc = (
        "set_max_delay 5.0 -from %s -to %s\n"
        % (
            get_pins_list("u_h0/u_src/Q", range(2)),
            get_pins_list("u_h0/u_dst/D", range(2)),
        )
    )
    prelude = r'''
set ::STRUCTURAL_GETTER_CALLS 0

proc get_pins {args} {
    incr ::STRUCTURAL_GETTER_CALLS
    error "structural passthrough must not query get_pins"
}
'''
    result = run_case(
        "structural_passthrough_zero_getters",
        top_sdc,
        harden_sdc,
        extra_build_args=[
            "-max_segment_pairs", "2",
            "-generate_clock_group_review", "false",
        ],
        prelude=prelude,
        post_build_tcl='puts "STRUCTURAL_GETTER_CALLS=$::STRUCTURAL_GETTER_CALLS"',
    )
    require_ok(result)
    assert_text_contains(result["stdout"], "STRUCTURAL_GETTER_CALLS=0")
    report = read_file(result["report"])
    expected_stats = {
        "structural_passthrough_commands": 3,
        "structural_passthrough_objects": 14108,
        "matrix_pairs_avoided": 21111556,
        "matrix_expansion_limited": 0,
        "matrix_pairs_expanded": 0,
        "matrix_expand_elapsed_ms": 0,
        "metadata_batch_queries": 0,
        "metadata_individual_queries": 0,
    }
    for name, expected in expected_stats.items():
        actual = stat_value(report, name)
        if actual != expected:
            raise AssertionError(
                "Unexpected structural passthrough statistic %s=%d, expected %d:\n%s"
                % (name, actual, expected, report)
            )
    assert_contains(result["report"], "Passthrough constraints         : 3")
    assert_contains(result["report"], "Review required constraints     : 0")
    assert_contains(result["report"], "Max segment pairs               : 2")
    trace = read_file(result["trace"])
    if trace.count("action=STRUCTURAL_PASSTHROUGH") != 3:
        raise AssertionError("Expected one structural plan per source command:\n%s" % trace)
    if delay_command_lines(result["out_sdc"]):
        raise AssertionError("Structural passthrough unexpectedly generated E2E delays")
    assert_contains(result["final"], top_sdc.strip())
    assert_contains(result["final"], harden_sdc.strip())
    assert_not_contains(result["final"], "STAGE2_CONSUMED")
    assert_not_contains(result["final"], "STAGE2_REWRITTEN")
    validate_static_sdc(result["final"])


def test_matrix_segment_pair_limit_and_inclusive_boundary():
    top_sdc = (
        "set_max_delay 2.0 -from %s -to %s\n"
        % (get_pins_list("src", range(3)), get_pins_list("u_h0/cfg", range(2)))
    )
    prelude = r'''
for {set idx 0} {$idx < 3} {incr idx} {
    set ::PT_MOCK_DIRECTIONS([format {src[%d]} $idx]) out
}
for {set idx 0} {$idx < 2} {incr idx} {
    set ::PT_MOCK_DIRECTIONS([format {u_h0/cfg[%d]} $idx]) in
}
'''
    limited = run_case(
        "matrix_segment_pair_limit_5",
        top_sdc,
        "",
        extra_build_args=[
            "-max_segment_pairs", "5",
            "-generate_clock_group_review", "false",
        ],
        prelude=prelude,
    )
    require_ok(limited)
    limited_report = read_file(limited["report"])
    for name, expected in {
        "structural_passthrough_commands": 0,
        "matrix_expansion_limited": 1,
        "matrix_pairs_avoided": 6,
        "matrix_pairs_expanded": 0,
    }.items():
        actual = stat_value(limited_report, name)
        if actual != expected:
            raise AssertionError(
                "Unexpected limited matrix statistic %s=%d, expected %d:\n%s"
                % (name, actual, expected, limited_report)
            )
    assert_contains(limited["report"], "Max segment pairs               : 5")
    assert_contains(limited["report"], "MATRIX_EXPANSION_LIMIT")
    assert_contains(limited["review"], "reason=MATRIX_EXPANSION_LIMIT")
    if read_file(limited["review"]).count("reason=MATRIX_EXPANSION_LIMIT") != 1:
        raise AssertionError("Limited matrix must create exactly one detailed review item")
    assert_contains(limited["review"], "matrix_from_count=3")
    assert_contains(limited["review"], "matrix_to_count=2")
    assert_contains(limited["review"], "matrix_pair_count=6")
    assert_contains(limited["review"], "matrix_limit=5")
    assert_contains(limited["trace"], "product=6 action=MATRIX_EXPANSION_LIMIT limit=5 original=preserved")
    assert_contains(limited["final"], top_sdc.strip())
    assert_not_contains(limited["final"], "STAGE2_REWRITTEN CMD000001")
    if delay_command_lines(limited["out_sdc"]):
        raise AssertionError("Limited matrix unexpectedly generated E2E delays")

    inclusive = run_case(
        "matrix_segment_pair_inclusive_6",
        top_sdc,
        "",
        extra_build_args=[
            "-max_segment_pairs", "6",
            "-generate_clock_group_review", "false",
        ],
        prelude=prelude,
    )
    require_ok(inclusive)
    inclusive_report = read_file(inclusive["report"])
    for name, expected in {
        "structural_passthrough_commands": 0,
        "matrix_expansion_limited": 0,
        "matrix_pairs_avoided": 0,
        "matrix_pairs_expanded": 6,
    }.items():
        actual = stat_value(inclusive_report, name)
        if actual != expected:
            raise AssertionError(
                "Unexpected inclusive matrix statistic %s=%d, expected %d:\n%s"
                % (name, actual, expected, inclusive_report)
            )
    if stat_value(inclusive_report, "matrix_expand_elapsed_ms") < 0:
        raise AssertionError("Matrix expansion elapsed time must not be negative")
    assert_contains(inclusive["report"], "Max segment pairs               : 6")
    assert_not_contains(inclusive["review"], "MATRIX_EXPANSION_LIMIT")
    assert_contains(inclusive["trace"], "SEGMENT_EXPAND_END")
    assert_contains(inclusive["trace"], "expanded=6 product=6 elapsed_ms=")


def test_top_port_mapping_cannot_bypass_matrix_pair_limit():
    top_sdc = "set_max_delay 2.0 -from %s -to [get_ports cfg_top]\n" % get_pins_list(
        "src", range(2)
    )
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    src[0] out
    src[1] out
    cfg_top out
    u_h0/cfg_i in
    u_h1/cfg_i in
    u_h2/cfg_i in
}

proc get_nets {args} {
    if {[lindex $args end] eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        if {[lindex $args end] eq "cfg_net"} {
            return [list u_h0/cfg_i u_h1/cfg_i u_h2/cfg_i]
        }
        return {}
    }
    return [lindex $args end]
}
'''
    result = run_case(
        "top_port_mapping_matrix_limit",
        top_sdc,
        "",
        extra_build_args=[
            "-max_segment_pairs", "2",
            "-generate_clock_group_review", "false",
        ],
        extra_hardens=[
            ("h1", "u_h1", "harden1"),
            ("h2", "u_h2", "harden2"),
        ],
        prelude=prelude,
    )
    require_ok(result)
    report = read_file(result["report"])
    for name, expected in {
        "matrix_expansion_limited": 1,
        "matrix_pairs_avoided": 6,
        "matrix_pairs_expanded": 2,
    }.items():
        actual = stat_value(report, name)
        if actual != expected:
            raise AssertionError(
                "Unexpected mapped matrix statistic %s=%d, expected %d:\n%s"
                % (name, actual, expected, report)
            )
    assert_contains(result["review"], "reason=MATRIX_EXPANSION_LIMIT")
    assert_contains(result["review"], "matrix_from_count=2")
    assert_contains(result["review"], "matrix_to_count=3")
    assert_contains(result["review"], "matrix_pair_count=6")
    if read_file(result["review"]).count("reason=MATRIX_EXPANSION_LIMIT") != 1:
        raise AssertionError("Mapped matrix limit must create one command-level review")
    assert_contains(
        result["trace"],
        "product=6 action=MATRIX_EXPANSION_LIMIT limit=2 phase=TOP_PORT_MAP original=preserved",
    )
    assert_not_contains(result["report"], "TOP_PORT_BOUNDARY_MAP")
    assert_contains(result["final"], top_sdc.strip())
    assert_not_contains(result["final"], "STAGE2_REWRITTEN CMD000001")
    if delay_command_lines(result["out_sdc"]):
        raise AssertionError("Mapped over-limit matrix unexpectedly generated E2E delays")


def test_final_rewrite_preserves_structural_and_limited_commands():
    merged_top = "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]"
    structural_top = "set_max_delay 1.0 -from %s -to %s" % (
        get_pins_list("u_local/src", range(2)),
        get_pins_list("u_local/dst", range(2)),
    )
    limited_top = "set_max_delay 3.0 -from %s -to %s" % (
        get_pins_list("src", range(3)),
        get_pins_list("u_h0/cfg", range(2)),
    )
    top_sdc = "\n".join([merged_top, structural_top, limited_top]) + "\n"
    prelude = r'''
for {set idx 0} {$idx < 3} {incr idx} {
    set ::PT_MOCK_DIRECTIONS([format {src[%d]} $idx]) out
}
for {set idx 0} {$idx < 2} {incr idx} {
    set ::PT_MOCK_DIRECTIONS([format {u_h0/cfg[%d]} $idx]) in
}
'''
    result = run_case(
        "final_rewrite_preserves_structural_and_limited",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=[
            "-max_segment_pairs", "5",
            "-generate_clock_group_review", "false",
        ],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7")
    assert_contains(result["final"], "STAGE2_CONSUMED CMD000001")
    assert_contains(result["final"], structural_top)
    assert_contains(result["final"], limited_top)
    assert_not_contains(result["final"], "STAGE2_REWRITTEN CMD000002")
    assert_not_contains(result["final"], "STAGE2_REWRITTEN CMD000003")
    assert_contains(result["review"], "reason=MATRIX_EXPANSION_LIMIT")
    if read_file(result["review"]).count("reason=MATRIX_EXPANSION_LIMIT") != 1:
        raise AssertionError("Final rewrite case must keep one limited command review")
    report = read_file(result["report"])
    if stat_value(report, "structural_passthrough_commands") != 1:
        raise AssertionError("Expected one structural command in mixed final rewrite case")
    if stat_value(report, "matrix_expansion_limited") != 1:
        raise AssertionError("Expected one limited command in mixed final rewrite case")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_structural_passthrough_immediate_boundary_uses_legacy_hydration():
    prelude = r'''
set ::BOUNDARY_METADATA_LOG [file join [pwd] boundary_metadata_calls.log]

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set patterns [lindex $args end]
    set fout [open $::BOUNDARY_METADATA_LOG a]
    puts $fout [join $patterns ,]
    close $fout
    return $patterns
}
'''
    result = run_case(
        "structural_immediate_boundary_fallback",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-generate_clock_group_review", "false"],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(
        result["out_sdc"],
        "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]",
    )
    report = read_file(result["report"])
    if stat_value(report, "structural_passthrough_commands") != 0:
        raise AssertionError("Immediate harden boundaries must bypass structural passthrough")
    if stat_value(report, "structural_passthrough_objects") != 0:
        raise AssertionError("Immediate boundary records must use legacy hydration")
    if stat_value(report, "metadata_individual_queries") != 3:
        raise AssertionError("Expected three cached individual metadata queries:\n%s" % report)
    calls = read_file(os.path.join(result["case_dir"], "boundary_metadata_calls.log")).splitlines()
    if calls != ["u_src_reg/Q", "u_h0/cfg_i", "u_h0/u_reg/D"]:
        raise AssertionError("Unexpected immediate-boundary metadata calls: %r" % calls)
    assert_not_contains(result["trace"], "action=STRUCTURAL_PASSTHROUGH")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_structural_passthrough_unsafe_object_shapes_use_legacy_flow():
    prelude = r'''
set ::LEGACY_GET_PINS_CALLS 0
set ::LEGACY_GET_PORTS_CALLS 0

proc get_pins {args} {
    incr ::LEGACY_GET_PINS_CALLS
    return [lindex $args end]
}

proc get_ports {args} {
    incr ::LEGACY_GET_PORTS_CALLS
    return [lindex $args end]
}

rename get_attribute stage2_shape_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "direction"} {
        if {$name in {u_local/src u_local/src_* src_i}} {
            return out
        }
        if {$name in {u_local/dst u_local/dst_* dst_o}} {
            return in
        }
    }
    return [stage2_shape_default_get_attribute $obj $attr]
}
'''
    cases = [
        (
            "unknown",
            "set_max_delay 2.0 -from [get_foo u_local/src] -to [get_pins u_local/dst]\n",
            1,
            0,
            1,
            "CLOCK_OR_UNKNOWN_OBJECT",
        ),
        (
            "clock",
            "set_max_delay 2.0 -from [get_clocks CLK] -to [get_pins u_local/dst]\n",
            1,
            0,
            1,
            "CLOCK_OR_UNKNOWN_OBJECT",
        ),
        (
            "port",
            "set_max_delay 2.0 -from [get_ports src_i] -to [get_ports dst_o]\n",
            0,
            2,
            2,
            None,
        ),
        (
            "wildcard",
            "set_max_delay 2.0 -from [get_pins {u_local/src_*}] -to [get_pins {u_local/dst_*}]\n",
            0,
            0,
            0,
            "CLOCK_OR_UNKNOWN_OBJECT",
        ),
        (
            "dynamic",
            "set_max_delay 2.0 -from [get_pins [all_registers]] -to [get_pins u_local/dst]\n",
            1,
            0,
            1,
            "CLOCK_OR_UNKNOWN_OBJECT",
        ),
    ]
    for token, top_sdc, pin_calls, port_calls, individual_queries, review_reason in cases:
        result = run_case(
            "structural_shape_fallback_%s" % token,
            top_sdc,
            "",
            extra_build_args=[
                "-top_port_boundary_map_mode", "off",
                "-generate_clock_group_review", "false",
            ],
            prelude=prelude,
            post_build_tcl=(
                'puts "LEGACY_CALLS pins=$::LEGACY_GET_PINS_CALLS ports=$::LEGACY_GET_PORTS_CALLS"'
            ),
        )
        require_ok(result)
        assert_text_contains(
            result["stdout"],
            "LEGACY_CALLS pins=%d ports=%d" % (pin_calls, port_calls),
        )
        report = read_file(result["report"])
        if stat_value(report, "structural_passthrough_commands") != 0:
            raise AssertionError("%s objects must use legacy classification:\n%s" % (token, report))
        if stat_value(report, "structural_passthrough_objects") != 0:
            raise AssertionError("%s objects were incorrectly counted as structural" % token)
        if stat_value(report, "metadata_individual_queries") != individual_queries:
            raise AssertionError("Unexpected %s individual query count:\n%s" % (token, report))
        assert_contains(result["final"], top_sdc.strip())
        assert_not_contains(result["final"], "STAGE2_CONSUMED")
        if review_reason is not None:
            assert_contains(result["review"], review_reason)
            assert_contains(result["report"], "Review required constraints     : 1")
        else:
            assert_contains(result["report"], "Passthrough constraints         : 1")
            assert_contains(result["report"], "Review required constraints     : 0")


def test_top_open_from_infers_static_startpoint():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/cfg_i"} {
        return [list u_src_reg/Q]
    }
    return {}
}
'''
    result = run_case(
        "top_open_from",
        "set_min_delay 0.2 -to [get_pins u_h0/cfg_i]\n",
        "set_min_delay 0.8 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_min_delay 1 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_top_open_to_multi_from_through_and_endpoint_expansion():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_h1/cfg_i in
    u_h1/u_reg/D in
}

proc get_cells {args} {
    return [list [lindex $args end]]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        set owner [lindex $args end]
        if {$owner eq "u_h0"} {
            return [list u_h0/cfg_i]
        }
        if {$owner eq "u_h1"} {
            return [list u_h1/cfg_i]
        }
        return {}
    }
    return [list [lindex $args end]]
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target eq "u_h0/u_reg/D"} {
        return [list u_h0/cfg_i]
    }
    if {$target eq "u_h1/u_reg/D"} {
        return [list u_h1/cfg_i]
    }
    return {}
}

proc all_fanout {args} {
    set seed [lindex [lindex $args end] 0]
    if {$seed in {u_mid/out_o u_up/data_o}} {
        return [list u_h0/u_reg/D u_h1/u_reg/D]
    }
    return {}
}
'''
    result = run_case(
        "top_open_to_multi_matrix",
        "set_max_delay 2.0 -from [list [get_pins u_src_reg/Q] [get_pins u_up/u_reg/Q]] -through [list [get_pins u_mid/out_o] [get_pins u_up/data_o]] -ignore_clock_latency\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D] -ignore_clock_latency\n",
        extra_hardens=[
            (
                "h1",
                "u_h1",
                "harden1",
                "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D] -ignore_clock_latency\n",
            )
        ],
        prelude=prelude,
    )
    require_ok(result)
    through_group = "-through [list [get_pins {u_mid/out_o}] [get_pins {u_up/data_o}]]"
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] %s -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}] -ignore_clock_latency" % through_group)
    assert_contains(result["out_sdc"], "set_max_delay 8 -from [get_pins {u_up/u_reg/Q}] %s -through [get_pins {u_h1/cfg_i}] -to [get_pins {u_h1/u_reg/D}] -ignore_clock_latency" % through_group)
    assert_contains(result["report"], "Merged constraints              : 4")
    assert_contains(result["report"], "OPEN_TO_ENDPOINT_INFERRED")
    assert_not_contains(result["review"], "NO_TO_OBJECT")
    assert_contains(result["final"], "STAGE2_CONSUMED CMD000001")


def test_object_metadata_batches_explicit_pin_list():
    indices = list(range(64))
    top_sdc = (
        "set_max_delay 2.0 -from %s -to [get_pins u_h0/cfg_i]\n"
        % get_pins_list("src", indices)
    )
    prelude = r'''
for {set idx 0} {$idx < 64} {incr idx} {
    set name [format {src[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($name) out
}
set ::METADATA_GETTER_LOG [file join [pwd] metadata_getter_calls.log]

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set patterns [lindex $args end]
    set fout [open $::METADATA_GETTER_LOG a]
    puts $fout "[llength $patterns]|[join $patterns ,]"
    close $fout
    return $patterns
}
'''
    result = run_case(
        "metadata_batch_explicit_pin_list",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    generated = read_file(result["out_sdc"])
    if generated.count("set_max_delay 7 ") != len(indices):
        raise AssertionError("Metadata batch changed list expansion semantics:\n%s" % generated)
    calls = read_file(os.path.join(result["case_dir"], "metadata_getter_calls.log")).splitlines()
    batch_calls = [line for line in calls if line.startswith("64|")]
    if len(batch_calls) != 1:
        raise AssertionError("Expected one 64-object metadata getter call: %r" % calls)
    report = read_file(result["report"])
    if stat_value(report, "metadata_batch_queries") != 1:
        raise AssertionError("Expected one metadata batch query:\n%s" % report)
    if stat_value(report, "metadata_batch_records") != len(indices):
        raise AssertionError("Unexpected metadata batch record count:\n%s" % report)
    if stat_value(report, "metadata_batch_fallbacks") != 0:
        raise AssertionError("Unexpected metadata batch fallback:\n%s" % report)
    top_rows = len(read_file(os.path.join(result["summary"], "top.csv")).splitlines()) - 1
    harden_rows = len(read_file(os.path.join(result["summary"], "u_h0.csv")).splitlines()) - 1
    if top_rows != len(indices) or harden_rows != len(indices):
        raise AssertionError(
            "Path summary lost rows: top=%d harden=%d expected=%d"
            % (top_rows, harden_rows, len(indices))
        )
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_object_metadata_batch_failure_falls_back_without_loss():
    indices = [0, 1, 2, 3]
    top_sdc = (
        "set_max_delay 2.0 -from %s -to [get_pins u_h0/cfg_i]\n"
        % get_pins_list("src", indices)
    )
    prelude = r'''
foreach idx {0 1 2 3} {
    set name [format {src[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($name) out
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set patterns [lindex $args end]
    if {[llength $patterns] > 1} {
        error "mock PT rejects metadata multi-pattern getter"
    }
    return $patterns
}
'''
    result = run_case(
        "metadata_batch_fallback",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    generated = read_file(result["out_sdc"])
    if generated.count("set_max_delay 7 ") != len(indices):
        raise AssertionError("Metadata fallback lost an object:\n%s" % generated)
    report = read_file(result["report"])
    if stat_value(report, "metadata_batch_fallbacks") != 1:
        raise AssertionError("Expected metadata batch fallback:\n%s" % report)
    if stat_value(report, "metadata_individual_queries") < len(indices):
        raise AssertionError("Expected per-object metadata fallback queries:\n%s" % report)
    assert_text_contains(result["stdout"], "object metadata batch fallback")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_object_metadata_chunk_sizes_preserve_output_and_trace():
    indices = list(range(10))
    top_sdc = (
        "set_max_delay 2.0 -from %s -to [get_pins u_h0/cfg_i]\n"
        % get_pins_list("src", indices)
    )
    prelude = r'''
for {set idx 0} {$idx < 10} {incr idx} {
    set name [format {src[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($name) out
}
set ::METADATA_CHUNK_LOG [file join [pwd] metadata_chunk_calls.log]

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set patterns [lindex $args end]
    set fout [open $::METADATA_CHUNK_LOG a]
    puts $fout "[llength $patterns]|[join $patterns ,]"
    close $fout
    return [lreverse $patterns]
}
'''
    commands_by_size = {}
    results = {}
    for size in (1, 4, 128):
        result = run_case(
            "metadata_chunk_size_%d" % size,
            top_sdc,
            "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
            extra_build_args=["-metadata_batch_size", str(size)],
            prelude=prelude,
        )
        require_ok(result)
        commands_by_size[size] = delay_command_lines(result["out_sdc"])
        results[size] = result
        if len(commands_by_size[size]) != len(indices):
            raise AssertionError(
                "Metadata chunk size %d changed generated command count: %s"
                % (size, commands_by_size[size])
            )

    if commands_by_size[1] != commands_by_size[4] or commands_by_size[4] != commands_by_size[128]:
        raise AssertionError("Metadata chunk size changed generated SDC semantics: %r" % commands_by_size)

    functional_outputs = {}
    for size, result in results.items():
        output_paths = [
            result["out_sdc"],
            result["final"],
            result["removed"],
            result["review"],
            os.path.join(result["summary"], "00_index.csv"),
            os.path.join(result["summary"], "top.csv"),
            os.path.join(result["summary"], "u_h0.csv"),
        ]
        case_dir = result["case_dir"].replace("\\", "/")
        functional_outputs[size] = [
            read_file(path).replace(case_dir, "<CASE_DIR>")
            for path in output_paths
        ]
    if functional_outputs[1] != functional_outputs[4] or functional_outputs[4] != functional_outputs[128]:
        raise AssertionError("Metadata chunk size changed normalized functional outputs")

    chunked = results[4]
    calls = read_file(os.path.join(chunked["case_dir"], "metadata_chunk_calls.log")).splitlines()
    source_batch_sizes = [
        int(line.split("|", 1)[0])
        for line in calls
        if "|src[" in line
    ]
    if source_batch_sizes != [4, 4, 2]:
        raise AssertionError("Expected metadata chunks 4/4/2, got %r from %r" % (source_batch_sizes, calls))
    report = read_file(chunked["report"])
    if stat_value(report, "metadata_batch_queries") != 3:
        raise AssertionError("Expected three metadata chunk queries:\n%s" % report)
    if stat_value(report, "metadata_batch_successes") != 3:
        raise AssertionError("Expected three successful metadata chunks:\n%s" % report)
    if stat_value(report, "metadata_batch_records") != len(indices):
        raise AssertionError("Unexpected metadata chunk record count:\n%s" % report)
    assert_contains(chunked["report"], "Metadata batch enabled          : true")
    assert_contains(chunked["report"], "Metadata batch size             : 4")
    assert_contains(chunked["trace"], "METADATA_BATCH_BEGIN class=pin chunk=1/3 getter=get_pins patterns=4")
    assert_contains(chunked["trace"], "METADATA_BATCH_END class=pin chunk=3/3 status=OK patterns=2 returned=2")
    assert_contains(chunked["trace"], "elapsed_ms=")
    validate_static_sdc(chunked["out_sdc"])
    validate_static_sdc(chunked["final"])


def test_object_metadata_failed_middle_chunk_only_falls_back():
    indices = list(range(10))
    top_sdc = (
        "set_max_delay 2.0 -from %s -to [get_pins u_h0/cfg_i]\n"
        % get_pins_list("src", indices)
    )
    prelude = r'''
for {set idx 0} {$idx < 10} {incr idx} {
    set name [format {src[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($name) out
}
set ::METADATA_FAILURE_LOG [file join [pwd] metadata_failure_calls.log]

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set patterns [lindex $args end]
    set mode [expr {[llength $patterns] > 1 ? "BATCH" : "SINGLE"}]
    set fout [open $::METADATA_FAILURE_LOG a]
    puts $fout "$mode|[join $patterns ,]"
    close $fout
    return $patterns
}

proc foreach_in_collection {var coll body} {
    upvar 1 $var item
    set visited 0
    foreach item $coll {
        uplevel 1 $body
        incr visited
        if {[llength $coll] == 4 &&
            [lsearch -exact $coll {src[4]}] >= 0 &&
            $visited == 2} {
            error "mock failure after partial middle metadata chunk"
        }
    }
}
'''
    result = run_case(
        "metadata_failed_middle_chunk",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-metadata_batch_size", "4"],
        prelude=prelude,
    )
    require_ok(result)
    if len(delay_command_lines(result["out_sdc"])) != len(indices):
        raise AssertionError("Failed metadata chunk lost generated constraints")
    calls = read_file(os.path.join(result["case_dir"], "metadata_failure_calls.log")).splitlines()
    source_singles = sorted(
        line.split("|", 1)[1]
        for line in calls
        if line.startswith("SINGLE|src[")
    )
    expected_singles = ["src[%d]" % idx for idx in range(4, 8)]
    if source_singles != expected_singles:
        raise AssertionError("Only the failed middle chunk should fall back: %r" % calls)
    report = read_file(result["report"])
    if stat_value(report, "metadata_batch_queries") != 3:
        raise AssertionError("Expected all three chunks to run:\n%s" % report)
    if stat_value(report, "metadata_batch_successes") != 2:
        raise AssertionError("Expected two successful chunks:\n%s" % report)
    if stat_value(report, "metadata_batch_fallbacks") != 1:
        raise AssertionError("Expected one failed-chunk fallback:\n%s" % report)
    if stat_value(report, "metadata_batch_returned_records") != 8:
        raise AssertionError("Expected partial failed-chunk returns in statistics:\n%s" % report)
    assert_contains(result["trace"], "METADATA_BATCH_END class=pin chunk=2/3 status=ERROR patterns=4 returned=2")
    assert_contains(result["trace"], "METADATA_BATCH_FALLBACK class=pin chunk=2/3 patterns=4")
    assert_contains(result["trace"], "METADATA_BATCH_END class=pin chunk=3/3 status=OK")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_object_metadata_mismatched_middle_chunk_only_falls_back():
    indices = list(range(10))
    top_sdc = (
        "set_max_delay 2.0 -from %s -to [get_pins u_h0/cfg_i]\n"
        % get_pins_list("src", indices)
    )
    prelude = r'''
for {set idx 0} {$idx < 10} {incr idx} {
    set name [format {src[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($name) out
}
set ::METADATA_MISMATCH_LOG [file join [pwd] metadata_mismatch_calls.log]

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set patterns [lindex $args end]
    set mode [expr {[llength $patterns] > 1 ? "BATCH" : "SINGLE"}]
    set fout [open $::METADATA_MISMATCH_LOG a]
    puts $fout "$mode|[join $patterns ,]"
    close $fout
    if {$mode eq "BATCH" && [lsearch -exact $patterns {src[4]}] >= 0} {
        return [lrange $patterns 0 end-1]
    }
    return $patterns
}
'''
    result = run_case(
        "metadata_mismatched_middle_chunk",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-metadata_batch_size", "4"],
        prelude=prelude,
    )
    require_ok(result)
    if len(delay_command_lines(result["out_sdc"])) != len(indices):
        raise AssertionError("Mismatched metadata chunk lost generated constraints")
    calls = read_file(os.path.join(result["case_dir"], "metadata_mismatch_calls.log")).splitlines()
    source_singles = sorted(
        line.split("|", 1)[1]
        for line in calls
        if line.startswith("SINGLE|src[")
    )
    expected_singles = ["src[%d]" % idx for idx in range(4, 8)]
    if source_singles != expected_singles:
        raise AssertionError("Only the mismatched middle chunk should fall back: %r" % calls)
    report = read_file(result["report"])
    if stat_value(report, "metadata_batch_successes") != 2:
        raise AssertionError("Expected two successful chunks around mismatch:\n%s" % report)
    if stat_value(report, "metadata_batch_fallbacks") != 1:
        raise AssertionError("Expected one mismatch fallback:\n%s" % report)
    assert_contains(result["trace"], "METADATA_BATCH_END class=pin chunk=2/3 status=MISMATCH patterns=4 returned=3")
    assert_contains(result["trace"], "METADATA_BATCH_FALLBACK class=pin chunk=2/3 patterns=4")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_object_metadata_batch_can_be_disabled_without_skipping_queries():
    indices = list(range(8))
    top_sdc = (
        "set_max_delay 2.0 -from %s -to [get_pins u_h0/cfg_i]\n"
        % get_pins_list("src", indices)
    )
    prelude = r'''
for {set idx 0} {$idx < 8} {incr idx} {
    set name [format {src[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($name) out
}
set ::METADATA_DISABLED_LOG [file join [pwd] metadata_disabled_calls.log]

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    set patterns [lindex $args end]
    if {[llength $patterns] > 1} {
        error "metadata batching must be disabled"
    }
    set fout [open $::METADATA_DISABLED_LOG a]
    puts $fout [lindex $patterns 0]
    close $fout
    return $patterns
}
'''
    result = run_case(
        "metadata_batch_disabled",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-metadata_batch_enabled", "false"],
        prelude=prelude,
    )
    require_ok(result)
    if len(delay_command_lines(result["out_sdc"])) != len(indices):
        raise AssertionError("Disabled metadata batch skipped direction queries")
    calls = read_file(os.path.join(result["case_dir"], "metadata_disabled_calls.log")).splitlines()
    source_calls = sorted(name for name in calls if name.startswith("src["))
    expected_calls = ["src[%d]" % idx for idx in indices]
    if source_calls != expected_calls:
        raise AssertionError("Expected every source pin to be queried individually: %r" % calls)
    report = read_file(result["report"])
    if stat_value(report, "metadata_batch_queries") != 0:
        raise AssertionError("Disabled metadata batch issued a batch query:\n%s" % report)
    if stat_value(report, "metadata_batch_disabled_groups") != 1:
        raise AssertionError("Expected one disabled metadata group:\n%s" % report)
    if stat_value(report, "metadata_individual_queries") < len(indices):
        raise AssertionError("Expected individual metadata queries:\n%s" % report)
    assert_contains(result["report"], "Metadata batch enabled          : false")
    assert_contains(result["trace"], "METADATA_BATCH_DISABLED class=pin records=8 mode=individual")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_object_metadata_batch_size_validation():
    for value, token in (("0", "zero"), ("-2", "negative"), ("abc", "text")):
        result = run_case(
            "metadata_batch_invalid_size_%s" % token,
            "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
            "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
            extra_build_args=["-metadata_batch_size", value],
        )
        if result["code"] == 0:
            raise AssertionError("metadata_batch_size=%s must be rejected" % value)
        assert_text_contains(result["stderr"], "-metadata_batch_size must be an integer >= 1")


def test_max_segment_pairs_validation():
    for value, token in (("0", "zero"), ("-2", "negative"), ("abc", "text")):
        result = run_case(
            "max_segment_pairs_invalid_%s" % token,
            "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
            "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
            extra_build_args=["-max_segment_pairs", value],
        )
        if result["code"] == 0:
            raise AssertionError("max_segment_pairs=%s must be rejected" % value)
        assert_text_contains(result["stderr"], "-max_segment_pairs must be an integer >= 1")


def test_startpoint_and_boundary_queries_are_cached():
    prelude = r'''
set ::FANIN_QUERY_LOG [file join [pwd] fanin_query_calls.log]

proc get_cells {args} {
    return [lindex $args end]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return [list u_h0/cfg_i]
    }
    return [lindex $args end]
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    set fout [open $::FANIN_QUERY_LOG a]
    puts $fout $target
    close $fout
    if {$target eq "u_h0/u_reg/D"} {
        return [list u_h0/cfg_i]
    }
    if {$target eq "u_h0/cfg_i"} {
        return [list u_src_reg/Q]
    }
    return {}
}
'''
    post_build_tcl = r'''
set boundary [stage2_delay::object_record pin u_h0/cfg_i in u_h0]
stage2_delay::pt_startpoints_to_boundary $boundary
stage2_delay::pt_startpoints_to_boundary $boundary
stage2_delay::cached_boundary_inputs_to_endpoint u_h0 u_h0/u_reg/D
stage2_delay::cached_boundary_inputs_to_endpoint u_h0 u_h0/u_reg/D
puts "PERF_AFTER_CACHE=[stage2_delay::performance_stats_summary]"
'''
    result = run_case(
        "startpoint_boundary_query_cache",
        "set_max_delay 2.0 -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
        post_build_tcl=post_build_tcl,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}]")
    calls = read_file(os.path.join(result["case_dir"], "fanin_query_calls.log")).splitlines()
    if calls.count("u_h0/u_reg/D") != 1 or calls.count("u_h0/cfg_i") != 1:
        raise AssertionError("Expected one PT fanin query per logical key: %r" % calls)
    if stat_value(result["stdout"], "startpoint_cache_hits") < 2:
        raise AssertionError("Expected startpoint cache hits:\n%s" % result["stdout"])
    if stat_value(result["stdout"], "boundary_cache_hits") < 2:
        raise AssertionError("Expected boundary cache hits:\n%s" % result["stdout"])


def test_missing_path_fanout_queries_are_cached():
    prelude = r'''
set ::FANOUT_QUERY_LOG [file join [pwd] fanout_query_calls.log]

proc all_fanout {args} {
    set seed [lindex [lindex $args end] 0]
    set fout [open $::FANOUT_QUERY_LOG a]
    puts $fout "$seed|[expr {[lsearch -exact $args -endpoints_only] >= 0}]"
    close $fout
    if {$seed eq "u_h0/cfg_i"} {
        if {[lsearch -exact $args -endpoints_only] >= 0} {
            return [list u_h0/u_reg/D]
        }
        return [list u_h0/o_niu_rst_n]
    }
    if {$seed eq "u_h0/o_niu_rst_n"} {
        return [list top_rst_n]
    }
    return {}
}
'''
    post_build_tcl = r'''
set input_boundary [stage2_delay::object_record pin u_h0/cfg_i in u_h0]
set output_boundary [stage2_delay::object_record pin u_h0/o_niu_rst_n out u_h0]
stage2_delay::pt_harden_fanout_targets_from_boundary $input_boundary
stage2_delay::pt_harden_fanout_targets_from_boundary $input_boundary
stage2_delay::pt_top_fanout_targets_from_harden_output_boundary $output_boundary
stage2_delay::pt_top_fanout_targets_from_harden_output_boundary $output_boundary
puts "PERF_AFTER_FANOUT_CACHE=[stage2_delay::performance_stats_summary]"
'''
    result = run_case(
        "missing_path_fanout_cache",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
        post_build_tcl=post_build_tcl,
    )
    require_ok(result)
    calls = read_file(os.path.join(result["case_dir"], "fanout_query_calls.log")).splitlines()
    if len(calls) != 4:
        raise AssertionError("Expected cached fanout helpers to issue four total PT queries: %r" % calls)
    if stat_value(result["stdout"], "missing_harden_cache_hits") != 1:
        raise AssertionError("Expected missing-harden target cache hit:\n%s" % result["stdout"])
    if stat_value(result["stdout"], "missing_top_cache_hits") != 1:
        raise AssertionError("Expected missing-top target cache hit:\n%s" % result["stdout"])


def test_final_rewrite_reuses_parsed_segments_and_skips_untouched_sdc():
    result = run_case(
        "final_rewrite_index_reuse",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            (
                "h1",
                "u_h1",
                "harden1",
                "# untouched harden constraint\nset_false_path -from [get_ports aux_i]\n",
            )
        ],
    )
    require_ok(result)
    assert_contains(result["final"], "set_false_path -from [get_ports aux_i]")
    report = read_file(result["report"])
    if stat_value(report, "final_rewrite_index_hits") != 2:
        raise AssertionError("Expected indexed rewrite lookup for top and h0:\n%s" % report)
    if stat_value(report, "parsed_segment_reuse_hits") != 2:
        raise AssertionError("Expected parsed segment reuse for top and h0:\n%s" % report)
    if stat_value(report, "final_rewrite_skipped_files") != 1:
        raise AssertionError("Expected untouched h1 SDC rewrite skip:\n%s" % report)


def test_open_to_complete_buses_compact_and_batch_once():
    indices = list(range(8))
    top_sdc = (
        "set_max_delay 2.0 -from %s -through %s\n"
        % (get_pins_list("src", indices), get_pins_list("mid", indices))
    )
    prelude = r'''
foreach idx {0 1 2 3 4 5 6 7} {
    set src_name [format {src[%d]} $idx]
    set mid_name [format {mid[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($src_name) out
    set ::PT_MOCK_DIRECTIONS($mid_name) out
}

set ::OPEN_TO_FANOUT_LOG [file join [pwd] open_to_fanout_calls.log]

proc get_cells {args} {
    return [lindex $args end]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        set owner [lindex $args end]
        if {$owner eq "u_h0"} {
            return [list u_h0/cfg_i]
        }
        return {}
    }
    set patterns [lindex $args end]
    set out {}
    foreach pattern $patterns {
        if {$pattern eq {src[*]}} {
            foreach idx {0 1 2 3 4 5 6 7} {
                lappend out [format {src[%d]} $idx]
            }
        } elseif {$pattern eq {mid[*]}} {
            foreach idx {0 1 2 3 4 5 6 7} {
                lappend out [format {mid[%d]} $idx]
            }
        } else {
            lappend out $pattern
        }
    }
    return $out
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target eq "u_h0/u_reg/D"} {
        return [list u_h0/cfg_i]
    }
    return {}
}

proc all_fanout {args} {
    set fout [open $::OPEN_TO_FANOUT_LOG a]
    puts $fout [join $args " "]
    close $fout
    return [list u_h0/u_reg/D]
}
'''
    result = run_case(
        "open_to_complete_bus_batch",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(
        result["out_sdc"],
        "set_max_delay 7 -from [get_pins {src[*]}] -through [get_pins {mid[*]}] "
        "-through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]",
    )
    assert_contains(result["report"], "compact_applied=2")
    assert_contains(result["report"], "compact_members_saved=14")
    assert_contains(result["report"], "batch_groups=1")
    assert_contains(result["report"], "batch_endpoint_queries=1")
    assert_contains(result["report"], "batch_fallbacks=0")
    fanout_log = os.path.join(result["case_dir"], "open_to_fanout_calls.log")
    calls = [line for line in read_file(fanout_log).splitlines() if line.strip()]
    if len(calls) != 1:
        raise AssertionError("Expected one batched all_fanout call, got %d: %r" % (len(calls), calls))
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_open_to_bus_with_missing_bit_is_not_compacted():
    indices = [0, 2, 3, 4]
    top_sdc = "set_max_delay 2.0 -from %s\n" % get_pins_list("src", indices)
    prelude = r'''
foreach idx {0 2 3 4} {
    set name [format {src[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($name) out
}

proc get_cells {args} {
    return [lindex $args end]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return [list u_h0/cfg_i]
    }
    return [lindex $args end]
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    return [list u_h0/cfg_i]
}

proc all_fanout {args} {
    return [list u_h0/u_reg/D]
}
'''
    result = run_case(
        "open_to_bus_missing_bit",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    generated = read_file(result["out_sdc"])
    if generated.count("set_max_delay 7 ") != len(indices):
        raise AssertionError("Expected one constraint per uncompressed bit:\n%s" % generated)
    assert_not_contains(result["out_sdc"], "src[*]")
    assert_contains(result["report"], "compact_rejected=1")
    assert_contains(result["report"], "batch_endpoint_queries=1")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_open_to_bus_wildcard_overmatch_is_not_compacted():
    indices = [0, 1, 2, 3]
    top_sdc = "set_max_delay 2.0 -from %s\n" % get_pins_list("src", indices)
    prelude = r'''
foreach idx {0 1 2 3 4} {
    set name [format {src[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($name) out
}

proc get_cells {args} {
    return [lindex $args end]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return [list u_h0/cfg_i]
    }
    set patterns [lindex $args end]
    if {[llength $patterns] == 1 && [lindex $patterns 0] eq {src[*]}} {
        set out {}
        foreach idx {0 1 2 3 4} {
            lappend out [format {src[%d]} $idx]
        }
        return $out
    }
    return $patterns
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    return [list u_h0/cfg_i]
}

proc all_fanout {args} {
    return [list u_h0/u_reg/D]
}
'''
    result = run_case(
        "open_to_bus_wildcard_overmatch",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    generated = read_file(result["out_sdc"])
    if generated.count("set_max_delay 7 ") != len(indices):
        raise AssertionError("Expected exact members after wildcard rejection:\n%s" % generated)
    assert_not_contains(result["out_sdc"], "src[*]")
    assert_contains(result["report"], "compact_rejected=1")
    assert_text_contains(result["stdout"], "reason=pt_set_mismatch")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_open_to_batch_getter_failure_falls_back_without_loss():
    top_sdc = (
        "set_max_delay 2.0 -from "
        "[list [get_pins u_src_reg/Q] [get_pins u_aux_reg/Q]]\n"
    )
    prelude = r'''
set ::PT_MOCK_DIRECTIONS(u_aux_reg/Q) out

proc get_cells {args} {
    return [lindex $args end]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return [list u_h0/cfg_i]
    }
    set patterns [lindex $args end]
    if {[llength $patterns] > 1} {
        error "mock PT rejects multi-pattern get_pins"
    }
    return $patterns
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    return [list u_h0/cfg_i]
}

proc all_fanout {args} {
    return [list u_h0/u_reg/D]
}
'''
    result = run_case(
        "open_to_batch_getter_fallback",
        top_sdc,
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    generated = read_file(result["out_sdc"])
    if generated.count("set_max_delay 7 ") != 2:
        raise AssertionError("Fallback lost a source object:\n%s" % generated)
    assert_contains(result["out_sdc"], "-from [get_pins {u_src_reg/Q}]")
    assert_contains(result["out_sdc"], "-from [get_pins {u_aux_reg/Q}]")
    assert_contains(result["report"], "batch_fallbacks=1")
    assert_contains(result["report"], "batch_endpoint_queries=2")
    assert_text_contains(result["stdout"], "open-to batch fallback")
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_harden_open_to_infers_endpoint_and_merges():
    prelude = r'''
proc all_fanout {args} {
    set seed [lindex [lindex $args end] 0]
    if {$seed eq "u_h0/cfg_i"} {
        return [list u_h0/u_reg/D]
    }
    return {}
}
'''
    result = run_case(
        "harden_open_to",
        "set_min_delay 0.25 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_min_delay 0.75 -from [get_ports cfg_i]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_min_delay 1 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["final"], "STAGE2_CONSUMED CMD000002")
    assert_not_contains(result["review"], "NO_TO_OBJECT")


def test_harden_open_to_batches_endpoint_and_full_fanout_queries():
    indices = [0, 1, 2, 3]
    top_sdc = (
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to %s\n"
        % get_pins_list("u_h0/cfg", indices)
    )
    harden_sdc = (
        "set_max_delay 5.0 -from %s\n"
        % get_ports_list("cfg", indices)
    )
    prelude = r'''
foreach idx {0 1 2 3} {
    set port_name [format {cfg[%d]} $idx]
    set pin_name [format {u_h0/cfg[%d]} $idx]
    set ::PT_MOCK_DIRECTIONS($port_name) in
    set ::PT_MOCK_DIRECTIONS($pin_name) in
}
set ::OPEN_TO_FANOUT_LOG [file join [pwd] harden_open_to_fanout_calls.log]

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return {}
    }
    return [lindex $args end]
}

proc get_ports {args} {
    return [lindex $args end]
}

proc all_fanout {args} {
    set fout [open $::OPEN_TO_FANOUT_LOG a]
    puts $fout [join $args " "]
    close $fout
    return [list u_h0/u_reg/D]
}
'''
    result = run_case(
        "harden_open_to_batch_fanout",
        top_sdc,
        harden_sdc,
        prelude=prelude,
    )
    require_ok(result)
    generated = read_file(result["out_sdc"])
    if generated.count("set_max_delay 7 ") != len(indices):
        raise AssertionError("Expected one merged constraint per harden input:\n%s" % generated)
    assert_contains(result["report"], "batch_groups=1")
    assert_contains(result["report"], "batch_endpoint_queries=1")
    assert_contains(result["report"], "batch_full_fanout_queries=1")
    assert_contains(result["report"], "batch_fallbacks=0")
    fanout_log = os.path.join(result["case_dir"], "harden_open_to_fanout_calls.log")
    calls = [line for line in read_file(fanout_log).splitlines() if line.strip()]
    if len(calls) != 2:
        raise AssertionError("Expected one endpoint and one full fanout call, got %d: %r" % (len(calls), calls))
    validate_static_sdc(result["out_sdc"])
    validate_static_sdc(result["final"])


def test_open_to_inference_failure_keeps_original_for_review():
    result = run_case(
        "open_to_inference_failure",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -through [get_pins u_mid/out_o]\n",
        "",
    )
    require_ok(result)
    assert_contains(result["review"], "OPEN_TO_ENDPOINT_NOT_INFERRED")
    assert_contains(result["final"], "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -through [get_pins u_mid/out_o]")
    assert_not_contains(result["final"], "STAGE2_CONSUMED CMD000001")


def test_open_to_delay_option_mismatch_is_not_consumed():
    prelude = r'''
proc get_cells {args} {
    return [list [lindex $args end]]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return [list u_h0/cfg_i]
    }
    return [list [lindex $args end]]
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    return [list u_h0/cfg_i]
}

proc all_fanout {args} {
    return [list u_h0/u_reg/D]
}
'''
    result = run_case(
        "open_to_delay_option_mismatch",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -ignore_clock_latency\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["review"], "DELAY_OPTION_MISMATCH")
    assert_contains(result["final"], "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -ignore_clock_latency")
    assert_not_contains(result["final"], "STAGE2_CONSUMED CMD000001")


def test_open_to_multiple_endpoints_same_boundary_are_fully_consumed_once():
    prelude = r'''
proc get_cells {args} {
    return [list [lindex $args end]]
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        return [list u_h0/cfg_i]
    }
    return [list [lindex $args end]]
}

proc filter_collection {coll expression} {
    return $coll
}

proc all_fanin {args} {
    return [list u_h0/cfg_i]
}

proc all_fanout {args} {
    return [list u_h0/u_cfg_reg/D u_h0/u_mode_reg/D]
}
'''
    result = run_case(
        "open_to_same_boundary_multi_endpoint",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [list [get_pins u_h0/u_cfg_reg/D] [get_pins u_h0/u_mode_reg/D]]\n",
        prelude=prelude,
    )
    require_ok(result)
    generated = read_file(result["out_sdc"])
    if generated.count("set_max_delay 7 ") != 2:
        raise AssertionError("Expected exactly two unique generated endpoints:\n%s" % generated)
    assert_contains(result["out_sdc"], "-to [get_pins {u_h0/u_cfg_reg/D}]")
    assert_contains(result["out_sdc"], "-to [get_pins {u_h0/u_mode_reg/D}]")
    assert_contains(result["final"], "STAGE2_CONSUMED CMD000001")
    assert_not_contains(result["final"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}]")
    assert_not_contains(result["review"], "NO_HARDEN_SEGMENT_MATCHED")


def test_legacy_top_open_from_mode_still_emits_from_when_pt_knows_startpoint():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/cfg_i"} {
        return [list u_src_reg/Q]
    }
    return {}
}

proc all_fanout {args} {
    set from [lindex $args end]
    set name [lindex $from 0]
    if {$name eq "u_h0/cfg_i"} {
        return [list u_h0/u_reg/D]
    }
    return {}
}
'''
    result = run_case(
        "top_open_from_legacy_mode_late_from",
        "set_max_delay 0.5 -to [get_pins u_h0/cfg_i]\n",
        "",
        extra_build_args=["-top_open_from_mode", "through"],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 0.5 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")


def test_harden_open_from_with_explicit_through():
    result = run_case(
        "harden_open_from",
        "set_max_delay 1.5 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 4.5 -through [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 6 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")


def test_multi_hop_review():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_up/data_o"} {
        return [list u_up/u_reg/Q]
    }
    return {}
}
'''
    result = run_case(
        "multi_hop_review",
        "set_max_delay 2.0 -from [get_pins u_up/data_o] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[("up", "u_up", "upstream")],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_up/u_reg/Q}] -through [get_pins {u_up/data_o}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO harden=u_up from=u_up/u_reg/Q to=u_up/data_o")


def test_review_top_open_from_summary_infers_startpoint():
    prelude = r'''
proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/cfg_i"} {
        return [list u_src_reg/CP]
    }
    return {}
}
'''
    result = run_case(
        "review_top_open_from_summary",
        "set_max_delay 0.5 -to [get_pins u_h0/cfg_i]\n",
        "",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["review"], "MISSING_HARDEN_SDC_ENDPOINT_NOT_FOUND")
    assert_contains(result["report"], "REVIEW_TOP_OPEN_FROM_STARTPOINT_INFERRED")
    assert_contains(os.path.join(result["summary"], "top.csv"), "u_src_reg/CP")
    assert_not_contains(os.path.join(result["summary"], "top.csv"), '"NOT FOUND","u_h0/cfg_i","0.5"')


def test_recursive_harden_output_to_harden_input_chain():
    result = run_case(
        "recursive_harden_output_to_input",
        "set_max_delay 2.0 -from [get_pins u_up/data_o] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            (
                "up",
                "u_up",
                "upstream",
                "set_max_delay 1.0 -from [get_pins u_up/u_reg/Q] -to [get_pins u_up/data_o]\n",
            )
        ],
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "# MERGED id=E2E000001")
    assert_contains(result["out_sdc"], "set_max_delay 8 -from [get_pins {u_up/u_reg/Q}] -through [get_pins {u_up/data_o}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "RECURSIVE_MERGED")
    assert_exists(os.path.join(result["summary"], "00_index.csv"))
    assert_exists(os.path.join(result["summary"], "top.csv"))
    assert_exists(os.path.join(result["summary"], "u_up.csv"))
    assert_exists(os.path.join(result["summary"], "u_h0.csv"))
    assert_contains(os.path.join(result["summary"], "00_index.csv"), "u_up.csv")
    assert_contains(os.path.join(result["summary"], "top.csv"), "e2e_id")
    assert_contains(os.path.join(result["summary"], "top.csv"), "E2E000001")
    assert_contains(os.path.join(result["summary"], "top.csv"), "through_1")
    assert_contains(os.path.join(result["summary"], "top.csv"), "Start Point")
    assert_contains(os.path.join(result["summary"], "top.csv"), "End Point")
    assert_contains(os.path.join(result["summary"], "top.csv"), "start_sdc_delay")
    assert_contains(os.path.join(result["summary"], "top.csv"), "start_from")
    assert_contains(os.path.join(result["summary"], "top.csv"), "start_to")
    assert_contains(os.path.join(result["summary"], "top.csv"), "end_sdc_delay")
    assert_contains(os.path.join(result["summary"], "top.csv"), "end_from")
    assert_contains(os.path.join(result["summary"], "top.csv"), "end_to")
    assert_contains(os.path.join(result["summary"], "top.csv"), "stage_1_sdc_delay")
    assert_contains(os.path.join(result["summary"], "top.csv"), "stage_1_from")
    assert_contains(os.path.join(result["summary"], "top.csv"), "stage_1_to")
    assert_contains(os.path.join(result["summary"], "top.csv"), '"8","u_up/u_reg/Q","1","u_up/u_reg/Q","u_up/data_o","1","u_up/u_reg/Q","u_up/data_o","u_up/data_o","2","u_up/data_o","u_h0/cfg_i","u_h0/cfg_i","5","u_h0/cfg_i","u_h0/u_reg/D","u_h0/u_reg/D","5","u_h0/cfg_i","u_h0/u_reg/D"')
    assert_contains(os.path.join(result["summary"], "00_index.csv"), "max_delay_used")
    assert_contains(os.path.join(result["summary"], "00_index.csv"), '"top.csv","1","1","0","0","1","1","1/1","0"')
    assert_contains(os.path.join(result["summary"], "u_up.csv"), "set_max_delay 8 -from [get_pins {u_up/u_reg/Q}] -through [get_pins {u_up/data_o}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(os.path.join(result["summary"], "u_h0.csv"), "u_h0/u_reg/D")
    xlsx = os.path.join(result["case_dir"], "top.xlsx")
    proc = subprocess.Popen(
        [sys.executable, REPORT_TOOL, result["summary"]],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=result["case_dir"],
    )
    stdout, stderr = proc.communicate()
    if proc.returncode != 0:
        raise AssertionError("report failed\nstdout=%s\nstderr=%s" % (stdout.decode("utf-8", "replace"), stderr.decode("utf-8", "replace")))
    assert_exists(xlsx)
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx)
    if workbook.sheetnames != ["top", "u_h0", "u_up"]:
        raise AssertionError("Unexpected workbook sheets: %s" % workbook.sheetnames)
    ws_top = workbook["top"]
    if [ws_top["A1"].value, ws_top["B1"].value, ws_top["E1"].value, ws_top["H1"].value] != ["E2E ID\nMax Delay Used: 1/1", "Start Point", "through_1", "End Point"]:
        raise AssertionError("Unexpected report headers: %s" % [ws_top["A1"].value, ws_top["B1"].value, ws_top["E1"].value, ws_top["H1"].value])
    if ws_top["A3"].value != "E2E000001":
        raise AssertionError("Unexpected E2E ID cell: %s" % ws_top["A3"].value)
    if [ws_top["B3"].value, ws_top["C3"].value, ws_top["D3"].value] != ["u_up/u_reg/Q", "u_up/data_o", "1"]:
        raise AssertionError("Unexpected start point row: %s" % [ws_top["B3"].value, ws_top["C3"].value, ws_top["D3"].value])
    if ws_top["E3"].fill.fgColor.rgb != "00FFF2CC":
        raise AssertionError("Expected top sheet through stage to be highlighted")


def test_recursive_chain_top_port_partial_mapping_keeps_original_command():
    chain_command = (
        "set_max_delay 2.0 -from [get_pins u_h0/o_niu_rst_n] -to [get_ports cfg_top]"
    )
    top_sdc = (
        "set_max_delay 1.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n"
        + chain_command
        + "\n"
    )
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    cfg_top out
    u_h1/cfg_i in
    u_h1/u_reg/D in
    u_h2/cfg_i in
    u_h2/u_reg/D in
}
proc get_nets {args} {
    if {[lindex $args end] eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}
rename get_pins stage2_recursive_chain_port_get_pins
proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        if {[lindex $args end] eq "cfg_net"} {
            return [list u_h1/cfg_i u_h2/cfg_i]
        }
        return {}
    }
    return [stage2_recursive_chain_port_get_pins {*}$args]
}
'''
    result = run_case(
        "recursive_chain_top_port_partial_group",
        top_sdc,
        "set_max_delay 3.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/o_niu_rst_n]\n",
        extra_hardens=[
            (
                "h1",
                "u_h1",
                "harden1",
                "set_max_delay 4.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n",
            ),
            ("h2", "u_h2", "harden2"),
        ],
        prelude=prelude,
    )
    require_ok(result)
    if chain_command not in delay_command_lines(result["final"]):
        raise AssertionError("Partially matched recursive chain port group was consumed")
    assert_contains(result["report"], "TOP_PORT_BOUNDARY_MAP_KEEP_ORIGINAL")
    assert_contains(result["report"], "matched=1 total=2 mode=recursive")

    complete = run_case(
        "recursive_chain_top_port_complete_group",
        top_sdc,
        "set_max_delay 3.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/o_niu_rst_n]\n",
        extra_hardens=[
            (
                "h1",
                "u_h1",
                "harden1",
                "set_max_delay 4.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n",
            ),
            (
                "h2",
                "u_h2",
                "harden2",
                "set_max_delay 5.0 -from [get_pins u_h2/cfg_i] -to [get_pins u_h2/u_reg/D]\n",
            ),
        ],
        prelude=prelude,
    )
    require_ok(complete)
    if chain_command in delay_command_lines(complete["final"]):
        raise AssertionError("Fully matched recursive chain port group kept its original command")
    assert_contains(complete["final"], "STAGE2_CONSUMED CMD000002")
    assert_contains(complete["report"], "TOP_PORT_BOUNDARY_MAP_CONSUMED")
    assert_contains(complete["report"], "matched=2 total=2 mode=recursive")


def test_missing_harden_sdc_stage_assumes_zero_and_reports_not_found():
    result = run_case(
        "missing_harden_sdc_stage",
        "\n".join(
            [
                "set_max_delay 2.0 -from [get_pins u_up/data_o] -to [get_pins u_mid/in_i]",
                "set_max_delay 3.0 -from [get_pins u_mid/out_o] -to [get_pins u_h0/cfg_i]",
                "",
            ]
        ),
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            (
                "up",
                "u_up",
                "upstream",
                "set_max_delay 1.0 -from [get_pins u_up/u_reg/Q] -to [get_pins u_up/data_o]\n",
            ),
            ("mid", "u_mid", "middle"),
        ],
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 11 -from [get_pins {u_up/u_reg/Q}] -through [get_pins {u_up/data_o}] -through [get_pins {u_mid/in_i}] -through [get_pins {u_mid/out_o}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO harden=u_mid from=u_mid/in_i to=u_mid/out_o")
    assert_contains(os.path.join(result["summary"], "u_mid.csv"), "MISSING_SDC")
    assert_contains(os.path.join(result["summary"], "00_index.csv"), '"u_mid.csv","1","1","0","0","0","0","0/0","1"')

    xlsx = os.path.join(result["case_dir"], "top.xlsx")
    proc = subprocess.Popen(
        [sys.executable, REPORT_TOOL, result["summary"]],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=result["case_dir"],
    )
    stdout, stderr = proc.communicate()
    if proc.returncode != 0:
        raise AssertionError("report failed\nstdout=%s\nstderr=%s" % (stdout.decode("utf-8", "replace"), stderr.decode("utf-8", "replace")))
    assert_exists(xlsx)

    from openpyxl import load_workbook

    workbook = load_workbook(xlsx)
    ws_mid = workbook["u_mid"]
    expected_header = "E2E ID\nMax Delay Used: N/A\nNative max_delay: 0\nMissing SDC Stage: 1"
    if ws_mid["A1"].value != expected_header:
        raise AssertionError("Unexpected missing SDC usage header: %s" % ws_mid["A1"].value)
    values = [ws_mid.cell(3, col).value for col in range(1, ws_mid.max_column + 1)]
    if "NOT FOUND" not in values:
        raise AssertionError("Expected NOT FOUND in missing harden stage row: %s" % values)
    red_cells = [ws_mid.cell(3, col).coordinate for col in range(1, ws_mid.max_column + 1) if ws_mid.cell(3, col).fill.fgColor.rgb == "00F4CCCC"]
    if not red_cells:
        raise AssertionError("Expected red NOT FOUND cell in u_mid report sheet")


def test_recursive_terminal_missing_harden_sdc_uses_pt_endpoint():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_h0/u_src_reg/Q out
    u_mid/u_reg/D in
    u_mid/U26/I in
}

proc all_fanout {args} {
    set from [lindex $args end]
    set name [lindex $from 0]
    if {$name eq "u_mid/in_i"} {
        if {[lsearch -exact $args "-endpoints_only"] >= 0} {
            return [list u_mid/u_reg/D]
        }
        return [list u_mid/U26/I u_mid/u_reg/D]
    }
    return {}
}
'''
    result = run_case(
        "recursive_terminal_missing_harden_sdc_endpoint",
        "set_max_delay 2.0 -from [get_pins u_h0/o_niu_rst_n] -to [get_pins u_mid/in_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/u_src_reg/Q] -to [get_ports o_niu_rst_n]\n",
        extra_hardens=[("mid", "u_mid", "middle")],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_h0/u_src_reg/Q}] -through [get_pins {u_h0/o_niu_rst_n}] -through [get_pins {u_mid/in_i}] -to [get_pins {u_mid/u_reg/D}]")
    assert_not_contains(result["out_sdc"], "u_mid/U26/I")
    assert_not_contains(result["out_sdc"], "-to [get_pins {u_mid/in_i}]")
    assert_contains(result["report"], "Merged constraints              : 1")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO harden=u_mid from=u_mid/in_i to=u_mid/u_reg/D")
    assert_contains(os.path.join(result["summary"], "u_mid.csv"), "MISSING_SDC")


def test_recursive_missing_top_and_terminal_harden_sdc_use_pt_graph():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_h0/u_src_reg/Q out
    u_mid/u_reg/D in
}

proc all_fanout {args} {
    set from [lindex $args end]
    set name [lindex $from 0]
    if {$name eq "u_h0/o_niu_rst_n"} {
        return [list u_mid/in_i]
    }
    if {$name eq "u_mid/in_i"} {
        return [list u_mid/u_reg/D]
    }
    return {}
}
'''
    result = run_case(
        "recursive_missing_top_and_terminal_harden_sdc",
        "",
        "set_max_delay 5.0 -from [get_pins u_h0/u_src_reg/Q] -to [get_ports o_niu_rst_n]\n",
        extra_hardens=[("mid", "u_mid", "middle")],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 5 -from [get_pins {u_h0/u_src_reg/Q}] -through [get_pins {u_h0/o_niu_rst_n}] -through [get_pins {u_mid/in_i}] -to [get_pins {u_mid/u_reg/D}]")
    assert_not_contains(result["out_sdc"], "-to [get_pins {u_mid/in_i}]")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO source=top from=u_h0/o_niu_rst_n to=u_mid/in_i")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO harden=u_mid from=u_mid/in_i to=u_mid/u_reg/D")
    assert_contains(os.path.join(result["summary"], "top.csv"), "MISSING_TOP_SDC")
    assert_contains(os.path.join(result["summary"], "u_mid.csv"), "MISSING_SDC")


def test_edge_specific_review():
    result = run_case(
        "edge_specific_review",
        "set_max_delay 2.0 -rise_from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
    )
    require_ok(result)
    assert_contains(result["review"], "EDGE_SPECIFIC_OPTION")


def test_multi_object_lists_expand_and_rewrite_remaining():
    result = run_case(
        "multi_object_lists",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [list [get_pins u_h0/cfg_i] [get_pins u_h0/unused_i]]\n",
        "set_max_delay 5.0 -from [list [get_pins u_h0/cfg_i] [get_pins u_h0/other_i]] -to [get_pins u_h0/u_reg/D]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_not_contains(result["out_sdc"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}] -to [get_pins {u_h0/unused_i}]")
    assert_contains(result["report"], "Merged constraints              : 1")
    assert_contains(result["report"], "Review required constraints     : 3")
    assert_contains(result["report"], "MISSING_HARDEN_SDC_ENDPOINT_NOT_FOUND")
    assert_contains(result["removed"], "split=1/2")
    assert_contains(result["removed"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}] -to [get_pins {u_h0/cfg_i}]")
    assert_not_contains(result["removed"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}] -to [get_pins {u_h0/unused_i}]")
    assert_contains(result["final"], "STAGE2_REWRITTEN CMD000001")
    assert_contains(result["final"], "set_max_delay 2 -from [get_pins {u_src_reg/Q}] -to [get_pins {u_h0/unused_i}]")
    assert_contains(result["final"], "STAGE2_REWRITTEN CMD000002")
    assert_contains(result["final"], "set_max_delay 5 -from [get_pins {u_h0/other_i}] -to [get_pins {u_h0/u_reg/D}]")
    if stat_value(read_file(result["report"]), "final_rewrite_signature_lookups") < 1:
        raise AssertionError("Expected indexed final rewrite signature lookups")


def test_one_top_boundary_reused_for_multiple_harden_endpoints():
    result = run_case(
        "reuse_top_boundary_for_multi_to",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [list [get_pins u_h0/u_cfg_reg/D] [get_pins u_h0/u_mode_reg/D]]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_cfg_reg/D}]")
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_mode_reg/D}]")
    assert_contains(result["report"], "Merged constraints              : 2")


def test_top_port_maps_to_connected_harden_input():
    prelude = r'''
proc current_design {} {
    return current_integration_top
}

proc sizeof_collection {coll} {
    return [llength $coll]
}

proc foreach_in_collection {var coll body} {
    upvar 1 $var item
    foreach item $coll {
        uplevel 1 $body
    }
}

proc get_ports {args} {
    set name [lindex $args end]
    if {$name eq "cfg_top"} {
        return [list cfg_top]
    }
    return {}
}

proc get_nets {args} {
    set obj [lindex $args end]
    if {$obj eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}

proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        set obj [lindex $args end]
        if {$obj eq "cfg_net"} {
            return [list u_h0/cfg_i]
        }
        return {}
    }
    set name [lindex $args end]
    if {$name in {u_src_reg/Q u_h0/cfg_i u_h0/u_reg/D}} {
        return [list $name]
    }
    return {}
}

proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "full_name"} {
        return $name
    }
    if {$attr eq "direction"} {
        if {$name eq "cfg_top"} {
            return out
        }
        if {$name eq "u_src_reg/Q"} {
            return out
        }
        if {$name in {u_h0/cfg_i u_h0/u_reg/D}} {
            return in
        }
    }
    return ""
}
'''
    result = run_case(
        "top_port_boundary_map",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_ports cfg_top]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-verbose_pt_query", "true"],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 7 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/cfg_i}] -to [get_pins {u_h0/u_reg/D}]")
    assert_contains(result["report"], "TOP_PORT_BOUNDARY_MAP")
    assert_contains(result["report"], "Top port boundary map mode      : connectivity")
    assert_contains(result["report"], "Verbose PT query                : true")
    assert_contains(result["final"], "STAGE2_CONSUMED CMD000001")
    assert_text_contains(result["stdout"], "PT_QUERY: get_ports -quiet {cfg_top}")


def test_top_port_mixed_known_and_unknown_boundaries_preserves_original():
    original = "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_ports cfg_top]"
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    cfg_top out
}

proc get_nets {args} {
    if {[lindex $args end] eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}

rename get_pins stage2_mixed_direction_port_get_pins
proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        if {[lindex $args end] eq "cfg_net"} {
            return [list u_h0/cfg_i u_h1/mystery_i]
        }
        return {}
    }
    return [stage2_mixed_direction_port_get_pins {*}$args]
}
'''
    result = run_case(
        "top_port_mixed_known_unknown_direction",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[("h1", "u_h1", "harden1")],
        prelude=prelude,
    )
    require_ok(result)
    if delay_command_lines(result["out_sdc"]):
        raise AssertionError("A partially known top-port fanout was merged")
    if original not in delay_command_lines(result["final"]):
        raise AssertionError("A partially known top-port fanout lost its original constraint")
    assert_contains(
        result["report"],
        "TOP_PORT_CONNECTED_TO_HARDEN_BOUNDARY_WITH_UNKNOWN_DIRECTION",
    )
    assert_contains(result["trace"], "unknown_direction_pins=1")


def test_direct_top_port_partial_group_with_disconnected_branch_keeps_original():
    original = (
        "set_max_delay 2.0 -from [list [get_pins u_clk/CP] "
        "[get_pins u_src_reg/Q]] -to [get_ports cfg_top]"
    )
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_clk/CP in
    cfg_top out
    u_h1/cfg_i in
}

proc get_nets {args} {
    if {[lindex $args end] eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}

rename get_pins stage2_direct_port_group_get_pins
proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        if {[lindex $args end] eq "cfg_net"} {
            return [list u_h0/cfg_i u_h1/cfg_i]
        }
        return {}
    }
    return [stage2_direct_port_group_get_pins {*}$args]
}

rename get_attribute stage2_direct_port_group_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "is_clock_pin" && $name eq "u_clk/CP"} {
        return true
    }
    return [stage2_direct_port_group_get_attribute $obj $attr]
}

proc all_fanin {args} {
    set target [lindex [lindex $args end] 0]
    if {$target eq "u_h0/cfg_i"} {
        return [list u_src_reg/Q]
    }
    if {$target eq "u_h1/cfg_i"} {
        return [list u_clk/CP]
    }
    return {}
}
'''
    result = run_case(
        "direct_top_port_partial_group_disconnected_branch",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-recursive_chain_mode", "off"],
        extra_hardens=[("h1", "u_h1", "harden1")],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["trace"], "NO_PT_CONNECTIVITY_PAIR")
    assert_contains(result["report"], "TOP_PORT_BOUNDARY_MAP_KEEP_ORIGINAL")
    if original not in delay_command_lines(result["final"]):
        raise AssertionError(
            "A disconnected mapped child partially consumed the original port pair:\n%s"
            % read_file(result["final"])
        )
    assert_not_contains(result["final"], "STAGE2_REWRITTEN CMD000001")


def test_recursive_top_port_partial_mapping_keeps_original_command():
    original = "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_ports cfg_top]"
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    cfg_top out
    u_h1/cfg_i in
    u_h1/u_reg/D in
}
proc get_nets {args} {
    if {[lindex $args end] eq "cfg_top"} {
        return [list cfg_net]
    }
    return {}
}
rename get_pins stage2_recursive_port_group_get_pins
proc get_pins {args} {
    if {[lsearch -exact $args "-of_objects"] >= 0} {
        if {[lindex $args end] eq "cfg_net"} {
            return [list u_h0/cfg_i u_h1/cfg_i]
        }
        return {}
    }
    return [stage2_recursive_port_group_get_pins {*}$args]
}
'''
    result = run_case(
        "recursive_top_port_partial_group",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[("h1", "u_h1", "harden1")],
        prelude=prelude,
    )
    require_ok(result)
    generated = delay_command_lines(result["out_sdc"])
    if len(generated) != 1 or "u_h0/u_reg/D" not in generated[0]:
        raise AssertionError("Expected only the mapped h0 branch to generate E2E")
    if original not in delay_command_lines(result["final"]):
        raise AssertionError("Partially matched recursive port group lost its original command")
    assert_not_contains(result["final"], "STAGE2_CONSUMED CMD000001")
    assert_not_contains(result["final"], "STAGE2_REWRITTEN CMD000001")
    assert_contains(
        result["report"],
        "TOP_PORT_BOUNDARY_MAP_KEEP_ORIGINAL",
    )
    assert_contains(result["report"], "matched=1 total=2 mode=recursive")

    complete = run_case(
        "recursive_top_port_complete_group",
        original + "\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_hardens=[
            (
                "h1",
                "u_h1",
                "harden1",
                "set_max_delay 6.0 -from [get_pins u_h1/cfg_i] -to [get_pins u_h1/u_reg/D]\n",
            )
        ],
        prelude=prelude,
    )
    require_ok(complete)
    if len(delay_command_lines(complete["out_sdc"])) != 2:
        raise AssertionError("A fully matched recursive port group lost a generated branch")
    if original in delay_command_lines(complete["final"]):
        raise AssertionError("A fully matched recursive port group kept the original command")
    assert_contains(complete["final"], "STAGE2_CONSUMED CMD000001")
    assert_contains(complete["report"], "TOP_PORT_BOUNDARY_MAP_CONSUMED")
    assert_contains(complete["report"], "matched=2 total=2 mode=recursive")


def test_harden_input_to_output_boundary_merges():
    result = run_case(
        "harden_input_to_output_boundary_merge",
        "set_max_delay 1.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/i_niu_rst_n]\n",
        "set_max_delay 5.0 -from [get_ports i_niu_rst_n] -to [get_ports o_niu_rst_n]\n",
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 6 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/i_niu_rst_n}] -to [get_pins {u_h0/o_niu_rst_n}]")
    assert_contains(result["report"], "Merged constraints              : 1")
    assert_not_contains(result["report"], "OUTPUT_DIRECTION_NOT_SUPPORTED")


def test_harden_feedthrough_missing_upstream_top_uses_pt_startpoint():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg/Q out
    u_h0/i_niu_rst_n in
    u_h0/o_niu_rst_n out
    u_mid/in_i in
    u_mid/u_reg/D in
}

proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/i_niu_rst_n"} {
        return [list u_src_reg/Q]
    }
    return {}
}
'''
    result = run_case(
        "harden_feedthrough_missing_upstream_top",
        "set_max_delay 2.0 -from [get_pins u_h0/o_niu_rst_n] -to [get_pins u_mid/in_i]\n",
        "set_max_delay 5.0 -from [get_ports i_niu_rst_n] -to [get_ports o_niu_rst_n]\n",
        extra_hardens=[
            (
                "mid",
                "u_mid",
                "middle",
                "set_max_delay 4.0 -from [get_pins u_mid/in_i] -to [get_pins u_mid/u_reg/D]\n",
            )
        ],
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 11 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/i_niu_rst_n}] -through [get_pins {u_h0/o_niu_rst_n}] -through [get_pins {u_mid/in_i}] -to [get_pins {u_mid/u_reg/D}]")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO source=top from=u_src_reg/Q to=u_h0/i_niu_rst_n")
    assert_contains(os.path.join(result["summary"], "top.csv"), "MISSING_TOP_SDC_u_src_reg_Q_TO_u_h0_i_niu_rst_n")


def test_harden_feedthrough_to_top_output_terminal():
    prelude = r'''
array set ::PT_MOCK_DIRECTIONS {
    u_src_reg/Q out
    u_h0/i_niu_rst_n in
    u_h0/o_niu_rst_n out
    top_rst_n out
}

proc all_fanin {args} {
    set target [lindex $args end]
    set name [lindex $target 0]
    if {$name eq "u_h0/i_niu_rst_n"} {
        return [list u_src_reg/Q]
    }
    return {}
}

proc all_fanout {args} {
    set from [lindex $args end]
    set name [lindex $from 0]
    if {$name eq "u_h0/o_niu_rst_n"} {
        return [list top_rst_n]
    }
    return {}
}
'''
    result = run_case(
        "harden_feedthrough_to_top_output_terminal",
        "",
        "set_max_delay 5.0 -from [get_ports i_niu_rst_n] -to [get_ports o_niu_rst_n]\n",
        prelude=prelude,
    )
    require_ok(result)
    assert_contains(result["out_sdc"], "set_max_delay 5 -from [get_pins {u_src_reg/Q}] -through [get_pins {u_h0/i_niu_rst_n}] -through [get_pins {u_h0/o_niu_rst_n}] -to [get_ports {top_rst_n}]")
    assert_contains(result["report"], "RECURSIVE_MERGED_TERMINAL")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO source=top from=u_src_reg/Q to=u_h0/i_niu_rst_n")
    assert_contains(result["report"], "MISSING_SDC_ASSUMED_ZERO source=top from=u_h0/o_niu_rst_n to=top_rst_n")
    assert_not_contains(result["review"], "NO_TOP_SEGMENT_MATCHED")


def test_clock_inventory_and_review_template_are_generated_without_active_groups():
    prelude = r'''
set ::PT_MOCK_CLOCKS {clk_b clk_a clk_a_div2}

proc get_clocks {args} {
    set pattern [lindex $args end]
    if {$pattern eq "*"} {
        return $::PT_MOCK_CLOCKS
    }
    if {$pattern in $::PT_MOCK_CLOCKS} {
        return [list $pattern]
    }
    return {}
}

rename get_attribute stage2_clock_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "period"} {
        if {$name eq "clk_a_div2"} { return 4.0 }
        return 2.0
    }
    if {$attr eq "is_generated"} {
        return [expr {$name eq "clk_a_div2"}]
    }
    if {$attr eq "master_clock" && $name eq "clk_a_div2"} {
        return [list clk_a]
    }
    if {$attr eq "sources"} {
        return [list ${name}_src]
    }
    return [stage2_clock_default_get_attribute $obj $attr]
}

proc report_clock {args} {
    if {[lsearch -exact $args "-groups"] >= 0} {
        return "Report : clock_groups\nTotal logically exclusive groups: 0\nTotal asynchronous groups: 1\n-group {clk_a} -group {clk_b}"
    }
    return "Report : clock\nClock clk_a period 2.0\nClock clk_a_div2 period 4.0\nClock clk_b period 2.0"
}
'''
    result = run_case(
        "clock_inventory_and_review_template",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    case_dir = result["case_dir"]
    inventory = os.path.join(case_dir, "top_clock_inventory.rpt")
    groups = os.path.join(case_dir, "top_clock_groups_existing.rpt")
    review_sdc = os.path.join(case_dir, "top_clock_groups_review.sdc")
    for path in (inventory, groups, review_sdc):
        assert_exists(path)
    assert_contains(inventory, "clock_name=clk_a")
    assert_contains(inventory, "clock_name=clk_a_div2")
    assert_contains(inventory, "master_clock=clk_a")
    assert_contains(groups, "Total asynchronous groups: 1")
    review = read_file(review_sdc)
    if "REVIEW ONLY" not in review:
        raise AssertionError("Expected review-only marker:\n%s" % review)
    if "# set_clock_groups -asynchronous" not in review:
        raise AssertionError("Expected commented clock group template:\n%s" % review)
    if re.search(r"(?m)^(?!#).*set_clock_groups", review):
        raise AssertionError("Clock review SDC must not contain active set_clock_groups:\n%s" % review)
    assert_contains(result["report"], "Clock review enabled            : true")
    assert_contains(result["report"], "Active clock groups added       : 0")
    assert_not_contains(result["final"], "set_clock_groups")


def test_clock_review_can_be_disabled():
    result = run_case(
        "clock_review_disabled",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-generate_clock_group_review", "false"],
    )
    require_ok(result)
    for name in ("top_clock_inventory.rpt", "top_clock_groups_existing.rpt", "top_clock_groups_review.sdc"):
        if os.path.exists(os.path.join(result["case_dir"], name)):
            raise AssertionError("Clock output should not exist when review is disabled: %s" % name)
    assert_contains(result["report"], "Clock review enabled            : false")


def test_clock_review_complex_generated_groups_and_redirect_capture():
    prelude = r'''
set ::PT_MOCK_CLOCKS [list clk_z {u_pll/clk_out} {clk_bus[0]} clk_root clk_root_div2 clk_root_div4 scan_clk test_clk]

proc get_clocks {args} {
    set pattern [lindex $args end]
    if {$pattern eq "*"} {
        return $::PT_MOCK_CLOCKS
    }
    if {[lsearch -exact $::PT_MOCK_CLOCKS $pattern] >= 0} {
        return [list $pattern]
    }
    return {}
}

rename get_attribute stage2_complex_clock_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "period"} {
        array set periods {
            clk_z 10.0
            u_pll/clk_out 1.25
            {clk_bus[0]} 5.0
            clk_root 2.0
            clk_root_div2 4.0
            clk_root_div4 8.0
            scan_clk 20.0
            test_clk 25.0
        }
        return $periods($name)
    }
    if {$attr eq "is_generated"} {
        return [expr {$name in {u_pll/clk_out clk_root_div2 clk_root_div4}}]
    }
    if {$attr eq "master_clock"} {
        if {$name in {clk_root_div2 clk_root_div4}} { return [list clk_root] }
        if {$name eq "u_pll/clk_out"} { return [list ref_clk] }
        return ""
    }
    if {$attr eq "sources"} {
        if {$name eq "clk_root"} { return [list top/clk_i top/clk_mux/Z] }
        return [list ${name}_source]
    }
    return [stage2_complex_clock_default_get_attribute $obj $attr]
}

proc report_clock {args} {
    if {[lsearch -exact $args "-groups"] >= 0} {
        return "Report : clock_groups\nTotal logically exclusive groups: 1\nNAME : func_test_mux\n-group {clk_root clk_root_div2 clk_root_div4}\n-group {test_clk}\nTotal asynchronous groups: 2\nNAME : async_scan_func (allow_paths: true)\n-group {scan_clk}\n-group {clk_root}\nTotal physically exclusive groups: 1\n-group {u_pll/clk_out}\n-group {clk_z}"
    }
    return "Report : clock\nClock count: 8\nGenerated clocks: u_pll/clk_out clk_root_div2 clk_root_div4"
}

proc redirect {args} {
    set file_idx [lsearch -exact $args "-file"]
    if {$file_idx < 0} { error "missing -file" }
    set path [lindex $args [expr {$file_idx + 1}]]
    set command [lindex $args end]
    set text [uplevel #0 $command]
    set fout [open $path w]
    puts -nonewline $fout $text
    close $fout
}
'''
    result = run_case(
        "clock_review_complex_generated_groups_redirect",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    inventory = os.path.join(result["case_dir"], "top_clock_inventory.rpt")
    groups = os.path.join(result["case_dir"], "top_clock_groups_existing.rpt")
    review_sdc = os.path.join(result["case_dir"], "top_clock_groups_review.sdc")
    assert_contains(inventory, "Clock count      : 8")
    assert_contains(inventory, "clock_name=clk_root_div4 period=8.0 is_generated=1 master_clock=clk_root")
    assert_contains(inventory, "clock_name=u_pll/clk_out period=1.25 is_generated=1 master_clock=ref_clk")
    assert_contains(inventory, "sources=top/clk_i,top/clk_mux/Z")
    assert_contains(inventory, "Generated clocks: u_pll/clk_out clk_root_div2 clk_root_div4")
    assert_contains(groups, "Total logically exclusive groups: 1")
    assert_contains(groups, "Total asynchronous groups: 2")
    assert_contains(groups, "Total physically exclusive groups: 1")
    assert_contains(groups, "allow_paths: true")
    review = read_file(review_sdc)
    if review.count("#     -group [get_clocks ") != 8:
        raise AssertionError("Expected one commented group per clock:\n%s" % review)
    expected_order = ["clk_bus[0]", "clk_root", "clk_root_div2", "clk_root_div4", "clk_z", "scan_clk", "test_clk", "u_pll/clk_out"]
    offsets = [review.index("[get_clocks {%s}]" % name) for name in expected_order]
    if offsets != sorted(offsets):
        raise AssertionError("Clock template is not dictionary-sorted:\n%s" % review)
    if re.search(r"(?m)^(?!#).*set_clock_groups", review):
        raise AssertionError("Complex review emitted an active clock group:\n%s" % review)
    assert_contains(result["trace"], "redirect -file")
    assert_not_contains(result["final"], "set_clock_groups")


def test_clock_review_optional_attributes_unsupported_is_nonfatal():
    prelude = r'''
set ::PT_MOCK_CLOCKS {clk_a clk_b}

proc get_clocks {args} {
    set pattern [lindex $args end]
    if {$pattern eq "*"} { return $::PT_MOCK_CLOCKS }
    if {[lsearch -exact $::PT_MOCK_CLOCKS $pattern] >= 0} { return [list $pattern] }
    return {}
}

rename get_attribute stage2_unsupported_clock_default_get_attribute
proc get_attribute {obj attr} {
    if {$attr in {period is_generated master_clock master sources source}} {
        error "attribute $attr is unavailable in this PT version"
    }
    return [stage2_unsupported_clock_default_get_attribute $obj $attr]
}

proc report_clock {args} {
    if {[lsearch -exact $args "-groups"] >= 0} { return "Report : clock_groups\nTotal asynchronous groups: 0" }
    return "Report : clock\nclk_a clk_b"
}
'''
    result = run_case(
        "clock_review_optional_attributes_unsupported",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    inventory = os.path.join(result["case_dir"], "top_clock_inventory.rpt")
    assert_contains(inventory, "clock_name=clk_a period=- is_generated=- master_clock=- sources=-")
    assert_contains(inventory, "clock_name=clk_b period=- is_generated=- master_clock=- sources=-")
    assert_contains(os.path.join(result["case_dir"], "top_clock_groups_existing.rpt"), "Total asynchronous groups: 0")
    assert_not_contains(result["final"], "set_clock_groups")


def test_clock_review_zero_and_single_clock_boundaries_are_safe():
    zero_prelude = r'''
proc get_clocks {args} { return {} }
proc report_clock {args} {
    if {[lsearch -exact $args "-groups"] >= 0} { return "Report : clock_groups\nNo clock groups." }
    return "Report : clock\nNo clocks."
}
'''
    zero = run_case(
        "clock_review_zero_clocks",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=zero_prelude,
    )
    require_ok(zero)
    zero_inventory = os.path.join(zero["case_dir"], "top_clock_inventory.rpt")
    zero_review = os.path.join(zero["case_dir"], "top_clock_groups_review.sdc")
    assert_contains(zero_inventory, "Clock count      : 0")
    assert_contains(zero_inventory, "No clocks found in the linked PrimeTime design.")
    assert_contains(zero_review, "fewer than two clocks")
    assert_not_contains(zero_review, "# set_clock_groups -asynchronous")

    one_prelude = r'''
proc get_clocks {args} {
    set pattern [lindex $args end]
    if {$pattern eq "*" || $pattern eq "only_clk"} { return [list only_clk] }
    return {}
}
proc report_clock {args} {
    if {[lsearch -exact $args "-groups"] >= 0} { return "Report : clock_groups\nNo clock groups." }
    return "Report : clock\nClock only_clk"
}
'''
    one = run_case(
        "clock_review_single_clock",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=one_prelude,
    )
    require_ok(one)
    one_review = os.path.join(one["case_dir"], "top_clock_groups_review.sdc")
    assert_contains(os.path.join(one["case_dir"], "top_clock_inventory.rpt"), "Clock count      : 1")
    assert_contains(one_review, "#   only_clk")
    assert_contains(one_review, "fewer than two clocks")
    assert_not_contains(one_review, "# set_clock_groups -asynchronous")


def test_clock_review_large_clock_set_is_complete_and_deterministic():
    prelude = r'''
set ::PT_MOCK_CLOCKS {}
for {set idx 127} {$idx >= 0} {incr idx -1} {
    lappend ::PT_MOCK_CLOCKS [format "clk_%03d" $idx]
}

proc get_clocks {args} {
    set pattern [lindex $args end]
    if {$pattern eq "*"} { return $::PT_MOCK_CLOCKS }
    if {[lsearch -exact $::PT_MOCK_CLOCKS $pattern] >= 0} { return [list $pattern] }
    return {}
}

rename get_attribute stage2_large_clock_default_get_attribute
proc get_attribute {obj attr} {
    set name [lindex $obj 0]
    if {$attr eq "period"} { return 2.5 }
    if {$attr eq "is_generated"} { return false }
    if {$attr eq "sources"} { return [list ${name}_src] }
    return [stage2_large_clock_default_get_attribute $obj $attr]
}

proc report_clock {args} {
    if {[lsearch -exact $args "-groups"] >= 0} { return "Report : clock_groups\nTotal asynchronous groups: 0" }
    return "Report : clock\nClock count: 128"
}
'''
    result = run_case(
        "clock_review_large_128_clock_set",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        prelude=prelude,
    )
    require_ok(result)
    inventory = read_file(os.path.join(result["case_dir"], "top_clock_inventory.rpt"))
    review = read_file(os.path.join(result["case_dir"], "top_clock_groups_review.sdc"))
    if inventory.count("clock_name=clk_") != 128:
        raise AssertionError("Large inventory did not preserve all clocks")
    if review.count("#     -group [get_clocks {clk_") != 128:
        raise AssertionError("Large review did not preserve all singleton groups")
    if review.index("[get_clocks {clk_000}]") > review.index("[get_clocks {clk_127}]"):
        raise AssertionError("Large review clock order is unstable")
    if re.search(r"(?m)^(?!#).*set_clock_groups", review):
        raise AssertionError("Large review emitted an active clock group")
    assert_not_contains(result["final"], "set_clock_groups")


def test_clock_review_disabled_performs_zero_clock_queries():
    prelude = r'''
set ::CLOCK_GET_CALLS 0
set ::CLOCK_REPORT_CALLS 0
proc get_clocks {args} {
    incr ::CLOCK_GET_CALLS
    return [list should_not_be_queried]
}
proc report_clock {args} {
    incr ::CLOCK_REPORT_CALLS
    return "should not be queried"
}
'''
    result = run_case(
        "clock_review_disabled_zero_queries",
        "set_max_delay 2.0 -from [get_pins u_src_reg/Q] -to [get_pins u_h0/cfg_i]\n",
        "set_max_delay 5.0 -from [get_pins u_h0/cfg_i] -to [get_pins u_h0/u_reg/D]\n",
        extra_build_args=["-generate_clock_group_review", "false"],
        prelude=prelude,
        post_build_tcl='puts "CLOCK_QUERY_COUNTS get=$::CLOCK_GET_CALLS report=$::CLOCK_REPORT_CALLS"',
    )
    require_ok(result)
    assert_text_contains(result["stdout"], "CLOCK_QUERY_COUNTS get=0 report=0")
    assert_contains(result["report"], "Clock review enabled            : false")


def main():
    if os.path.isdir(WORK):
        shutil.rmtree(WORK)
    os.makedirs(WORK)
    tests = [
        test_release_identity_is_reconstructed_without_plaintext_constant,
        test_complete_complete_merge,
        test_live_trace_records_invalid_startpoint_object,
        test_pt_proven_input_clock_pin_is_accepted_as_startpoint,
        test_recursive_pt_proven_input_clock_pin_is_accepted,
        test_matrix_clock_pairs_skip_pt_disconnected_cross_pairs,
        test_sparse_matrix_numeric_bus_clock_bits_remain_exact_pins,
        test_matrix_pair_query_failure_does_not_silently_skip,
        test_sparse_matrix_missing_collection_iterator_is_unknown,
        test_pt_collection_runtime_iterator_failure_returns_unavailable,
        test_sparse_matrix_unresolved_fanin_object_is_unknown,
        test_sparse_matrix_prune_can_be_disabled_for_legacy_diagnosis,
        test_sparse_matrix_clock_batch_failure_falls_back_without_loss,
        test_sparse_matrix_all_disconnected_uses_compact_command_bookkeeping,
        test_sparse_matrix_partial_rewrite_excludes_pruned_cross_pair,
        test_sparse_matrix_prunes_before_pair_limit_but_falls_back_if_retained_exceeds_limit,
        test_sparse_matrix_scale_200x200_materializes_only_connected_diagonal,
        test_sparse_matrix_non_exact_clock_selector_is_never_pruned,
        test_non_exact_to_and_through_selectors_are_preserved,
        test_sparse_matrix_compact_clock_record_is_conservative,
        test_sparse_matrix_clock_metadata_batch_respects_disable_switch,
        test_sparse_matrix_top_port_limit_rolls_back_to_original_command,
        test_sparse_matrix_top_port_under_limit_partial_and_complete_rewrite,
        test_structural_passthrough_skips_getters_and_matrix_expansion,
        test_matrix_segment_pair_limit_and_inclusive_boundary,
        test_top_port_mapping_cannot_bypass_matrix_pair_limit,
        test_final_rewrite_preserves_structural_and_limited_commands,
        test_structural_passthrough_immediate_boundary_uses_legacy_hydration,
        test_structural_passthrough_unsafe_object_shapes_use_legacy_flow,
        test_top_open_from_infers_static_startpoint,
        test_top_open_to_multi_from_through_and_endpoint_expansion,
        test_object_metadata_batches_explicit_pin_list,
        test_object_metadata_batch_failure_falls_back_without_loss,
        test_object_metadata_chunk_sizes_preserve_output_and_trace,
        test_object_metadata_failed_middle_chunk_only_falls_back,
        test_object_metadata_mismatched_middle_chunk_only_falls_back,
        test_object_metadata_batch_can_be_disabled_without_skipping_queries,
        test_object_metadata_batch_size_validation,
        test_max_segment_pairs_validation,
        test_startpoint_and_boundary_queries_are_cached,
        test_missing_path_fanout_queries_are_cached,
        test_final_rewrite_reuses_parsed_segments_and_skips_untouched_sdc,
        test_open_to_complete_buses_compact_and_batch_once,
        test_open_to_bus_with_missing_bit_is_not_compacted,
        test_open_to_bus_wildcard_overmatch_is_not_compacted,
        test_open_to_batch_getter_failure_falls_back_without_loss,
        test_harden_open_to_infers_endpoint_and_merges,
        test_harden_open_to_batches_endpoint_and_full_fanout_queries,
        test_open_to_inference_failure_keeps_original_for_review,
        test_open_to_delay_option_mismatch_is_not_consumed,
        test_open_to_multiple_endpoints_same_boundary_are_fully_consumed_once,
        test_legacy_top_open_from_mode_still_emits_from_when_pt_knows_startpoint,
        test_harden_open_from_with_explicit_through,
        test_multi_hop_review,
        test_review_top_open_from_summary_infers_startpoint,
        test_recursive_harden_output_to_harden_input_chain,
        test_recursive_chain_top_port_partial_mapping_keeps_original_command,
        test_missing_harden_sdc_stage_assumes_zero_and_reports_not_found,
        test_recursive_terminal_missing_harden_sdc_uses_pt_endpoint,
        test_recursive_missing_top_and_terminal_harden_sdc_use_pt_graph,
        test_edge_specific_review,
        test_multi_object_lists_expand_and_rewrite_remaining,
        test_one_top_boundary_reused_for_multiple_harden_endpoints,
        test_top_port_maps_to_connected_harden_input,
        test_top_port_mixed_known_and_unknown_boundaries_preserves_original,
        test_direct_top_port_partial_group_with_disconnected_branch_keeps_original,
        test_recursive_top_port_partial_mapping_keeps_original_command,
        test_harden_input_to_output_boundary_merges,
        test_harden_feedthrough_missing_upstream_top_uses_pt_startpoint,
        test_harden_feedthrough_to_top_output_terminal,
        test_clock_inventory_and_review_template_are_generated_without_active_groups,
        test_clock_review_can_be_disabled,
        test_clock_review_complex_generated_groups_and_redirect_capture,
        test_clock_review_optional_attributes_unsupported_is_nonfatal,
        test_clock_review_zero_and_single_clock_boundaries_are_safe,
        test_clock_review_large_clock_set_is_complete_and_deterministic,
        test_clock_review_disabled_performs_zero_clock_queries,
    ]
    for test in tests:
        test()
        print("PASS", test.__name__)
    shutil.rmtree(WORK)
    print("All Stage 2 regression tests passed.")


if __name__ == "__main__":
    main()
