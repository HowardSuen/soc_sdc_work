#!/usr/bin/env python3
"""Approve the main harden-to-harden row for 10 regression."""

import argparse
from pathlib import Path

from openpyxl import load_workbook


TARGET_CHANNEL = "CH_u_a_data_o__u_b_data_i"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--async-relation", action="store_true")
    args = parser.parse_args()

    path = Path("10_harden_x_if.xlsx")
    wb = load_workbook(path)
    ws = wb["interface_budget"]
    headers = {cell.value: cell.column for cell in ws[1]}

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=headers["channel_id"]).value != TARGET_CHANNEL:
            continue
        updates = {
            "timing_model": "lib_blackbox",
            "budget_required": "yes",
            "budget_model": "interconnect_budget",
            "converted_max": "",
            "max_source": "",
            "derivation_basis": "",
            "tool_surface": "sta",
            "datapath_only": "yes",
            "budget_basis": "interconnect budget from block owners",
            "apply": "yes",
            "emit_max": "yes",
            "emit_min": "no",
            "review_status": "approved",
            "clock_relation": "async" if args.async_relation else "",
            "relationship_override_basis": "",
        }
        for key, value in updates.items():
            ws.cell(row=row_idx, column=headers[key], value=value)
        break
    else:
        raise SystemExit(f"target channel not found: {TARGET_CHANNEL}")

    wb.save(path)


if __name__ == "__main__":
    main()
