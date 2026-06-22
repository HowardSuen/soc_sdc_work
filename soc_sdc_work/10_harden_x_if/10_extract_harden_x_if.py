#!/usr/bin/env python3
"""
Generate 10 SoC harden/subsys interface budget SDC from integration
spreadsheets, harden/subsys SDC files, and a reviewed channel workbook.

Current scope:
  * all xlsx and SDC inputs live in the command execution directory
  * first run creates/synchronizes 10_harden_x_if.xlsx, then stops for review
  * constraints are grouped by integration-table channel, not by raw SDC line
  * only apply=yes + review_status=approved rows are emitted
  * scenario/stage/corner-specific generation is supported
"""

import argparse
import csv
import hashlib
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

try:
    import pandas as pd
except ImportError as exc:  # pragma: no cover - user environment guard
    print("ERROR: pandas is required to read integration xlsx files.", file=sys.stderr)
    raise SystemExit(2) from exc

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - user environment guard
    print("ERROR: openpyxl is required to read/write 10 harden IF xlsx files.", file=sys.stderr)
    raise SystemExit(2) from exc


SCENARIOS = {"common", "func", "scan", "mbist", "gpio_in", "gpio_out"}
STAGES = {"all", "synth", "prects", "postcts", "postroute"}
CHANNEL_TYPES_10 = {"harden_to_harden", "fabric_to_harden", "harden_to_fabric"}
NON_10_CHANNEL_TYPES = {
    "top_pad_to_harden",
    "harden_to_top_pad",
    "pad_to_pad",
    "clock_connection",
    "feedthrough",
    "exception_path",
    "unknown",
}
TIMING_MODELS = {"", "visible_netlist", "lib_blackbox", "abstract_model", "unknown"}
BUDGET_MODELS = {"", "interconnect_budget", "clock_relative_io_delay", "manual_budget", "unknown"}
TOOL_SURFACES = {"", "sta", "dc", "both"}
YES_NO = {"", "yes", "no"}
REVIEW_STATUS_VALUES = {"", "pending", "approved", "rejected"}
SOURCE_TYPES = {"", "extracted", "manual", "na"}
MIN_REVIEW_VALUES = {"", "approved", "reviewed", "yes", "waived"}
RELATION_BLOCKING = {"async", "asynchronous", "logically_exclusive", "physically_exclusive"}
ACTIVE_01_ACTIONS = {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}

CHANNEL_HEADERS = [
    "channel_id",
    "scenario",
    "stage",
    "corner",
    "channel_type",
    "src_instance",
    "src_module",
    "src_port",
    "src_endpoint",
    "dst_instance",
    "dst_module",
    "dst_port",
    "dst_endpoint",
    "connection_source",
    "is_pad_related",
    "is_clock_related",
    "is_feedthrough",
    "timing_model",
    "budget_required",
    "clock_relation",
    "note",
]

BUDGET_HEADERS = [
    "channel_id",
    "scenario",
    "stage",
    "corner",
    "channel_type",
    "is_pad_related",
    "is_clock_related",
    "is_feedthrough",
    "src_endpoint",
    "dst_endpoint",
    "timing_model",
    "budget_required",
    "clock_relation",
    "budget_model",
    "src_output_delay_max",
    "src_output_delay_min",
    "dst_input_delay_max",
    "dst_input_delay_min",
    "converted_max",
    "converted_min",
    "max_source",
    "min_source",
    "derivation_basis",
    "original_src_clock",
    "original_dst_clock",
    "soc_clock",
    "complex_options",
    "tool_surface",
    "datapath_only",
    "min_sign_review",
    "budget_basis",
    "relationship_override_basis",
    "source_type",
    "source_sdc_file",
    "source_line",
    "source_digest",
    "extraction_time",
    "source_command",
    "apply",
    "emit_max",
    "emit_min",
    "review_status",
    "owner",
    "note",
]

LOG_HEADERS = [
    "source_sdc_file",
    "source_line",
    "instance",
    "port",
    "direction",
    "constraint_type",
    "clock_name",
    "min_value",
    "max_value",
    "parse_status",
    "channel_ids",
    "source_digest",
    "extraction_time",
    "original_command",
    "message",
]

HEADER_FILL = PatternFill("solid", fgColor="215967")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C6CC"),
    right=Side(style="thin", color="B8C6CC"),
    top=Side(style="thin", color="B8C6CC"),
    bottom=Side(style="thin", color="B8C6CC"),
)


@dataclass
class PortInfo:
    name: str
    width: str = ""
    used_width: str = ""
    from_whom: str = ""
    to_top: str = ""
    connectivity: str = ""
    inout_name: str = ""


@dataclass
class InstInfo:
    module_name: str
    inst_name: str
    owner: str = ""
    file_path: str = ""
    sdc_hint: str = ""
    sdc_path: Optional[Path] = None
    inputs: Dict[str, PortInfo] = field(default_factory=dict)
    outputs: Dict[str, PortInfo] = field(default_factory=dict)
    inouts: Dict[str, PortInfo] = field(default_factory=dict)


@dataclass
class TclCommand:
    raw: str
    line_no: int


@dataclass
class ClockInfo:
    clock_name: str
    direct_source: str = ""
    producer_object: str = ""
    final_action: str = ""


@dataclass
class ChannelRecord:
    channel_id: str
    scenario: str
    stage: str
    corner: str
    channel_type: str
    src_instance: str
    src_module: str
    src_port: str
    src_endpoint: str
    dst_instance: str
    dst_module: str
    dst_port: str
    dst_endpoint: str
    connection_source: str
    is_pad_related: str = "no"
    is_clock_related: str = "no"
    is_feedthrough: str = "no"
    timing_model: str = "unknown"
    budget_required: str = ""
    clock_relation: str = ""
    note: str = ""


@dataclass
class DelayCandidate:
    inst_name: str
    module_name: str
    owner: str
    port_name: str
    direction: str
    constraint_type: str
    clock_name: str
    min_value: str
    max_value: str
    bare_value: str
    complex_options: str
    source_sdc_file: str
    source_line: str
    source_digest: str
    extraction_time: str
    original_command: str
    parse_status: str
    message: str = ""


@dataclass
class BudgetSeed:
    values: Dict[str, str]
    parse_status: str
    channel: ChannelRecord
    candidates: List[DelayCandidate]


@dataclass
class FormRow:
    row_idx: int
    values: Dict[str, object]
    autofilled_fields: Set[str] = field(default_factory=set)


class Report:
    def __init__(self) -> None:
        self.lines: List[str] = []
        self.warning_count = 0
        self.error_count = 0
        self.sync_changed = False

    def info(self, msg: str) -> None:
        self.lines.append(f"INFO: {msg}")

    def warn(self, msg: str) -> None:
        self.warning_count += 1
        self.lines.append(f"WARNING: {msg}")

    def error(self, msg: str) -> None:
        self.error_count += 1
        self.lines.append(f"ERROR: {msg}")


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def normalize_key(value) -> str:
    return clean_cell(value).strip().lower()


def normalize_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())


def column_map(df: pd.DataFrame) -> Dict[str, str]:
    return {normalize_col(col): col for col in df.columns}


def get_col(df: pd.DataFrame, aliases: Sequence[str]) -> Optional[str]:
    cmap = column_map(df)
    for alias in aliases:
        key = normalize_col(alias)
        if key in cmap:
            return cmap[key]
    return None


def read_excel_file(path: Path, sheet_name=0):
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except Exception as exc:
        raise RuntimeError(f"failed to read {path}: {exc}") from exc


def safe_filename_token(value: str) -> str:
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-")
    token = "".join(char if char in allowed else "_" for char in clean_cell(value))
    return token or "unknown"


def sanitize_id(value: str) -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", clean_cell(value)).strip("_")
    return token or "unknown"


def brace_list(names: Sequence[str]) -> str:
    return "{" + " ".join(clean_cell(name) for name in names if clean_cell(name)) + "}"


def get_collection(kind: str, objects: Sequence[str]) -> str:
    return f"[{kind} {brace_list(objects)}]"


def parse_number(value) -> Optional[float]:
    text = clean_cell(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_number(value) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    return f"{number:.12g}"


def read_text(path: Path) -> str:
    last_error: Optional[Exception] = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    assert last_error is not None
    raise last_error


def digest_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def read_info_all(path: Path, report: Report) -> Dict[str, InstInfo]:
    df = read_excel_file(path)
    module_col = get_col(df, ["module_name", "module name", "module"])
    inst_col = get_col(df, ["inst_name", "inst name", "instance", "instance_name"])
    owner_col = get_col(df, ["owner"])
    file_col = get_col(df, ["file_path", "file path", "empty_path", "verilog", "v_path"])
    sdc_col = get_col(df, ["sdc_path", "sdc file", "sdc_file", "sdc"])

    if not inst_col:
        raise RuntimeError(f"{path} must contain an inst_name column")

    instances: Dict[str, InstInfo] = {}
    for row_idx, row in df.iterrows():
        inst_name = clean_cell(row.get(inst_col))
        if not inst_name:
            continue
        module_name = clean_cell(row.get(module_col)) if module_col else ""
        file_path = clean_cell(row.get(file_col)) if file_col else ""
        if not module_name and file_path:
            module_name = Path(file_path).stem.replace("_empty", "")
        if not module_name:
            module_name = inst_name
            report.warn(f"{path.name} row {row_idx + 2}: module_name is empty; using inst_name")
        if inst_name in instances:
            report.warn(f"duplicate inst_name {inst_name} in {path.name}; keeping first row")
            continue
        instances[inst_name] = InstInfo(
            module_name=module_name,
            inst_name=inst_name,
            owner=clean_cell(row.get(owner_col)) if owner_col else "",
            file_path=file_path,
            sdc_hint=clean_cell(row.get(sdc_col)) if sdc_col else "",
        )
    report.info(f"loaded {len(instances)} instance(s) from {path.name}")
    return instances


def parse_port_sheet(df: pd.DataFrame) -> Dict[str, Dict[str, PortInfo]]:
    input_col = get_col(df, ["Input"])
    input_width_col = get_col(df, ["Input Width"])
    input_used_col = get_col(df, ["Input Used Width"])
    from_col = get_col(df, ["From Whom"])

    output_col = get_col(df, ["Output"])
    output_width_col = get_col(df, ["Output Width"])
    output_used_col = get_col(df, ["Output Used Width"])
    to_top_col = get_col(df, ["To Top", "To Whom", "To"])

    inout_col = get_col(df, ["Inout"])
    inout_width_col = get_col(df, ["Inout Width"])
    inout_conn_col = get_col(df, ["Inout Connectivity"])
    inout_name_col = get_col(df, ["Inout Name"])

    inputs: Dict[str, PortInfo] = {}
    outputs: Dict[str, PortInfo] = {}
    inouts: Dict[str, PortInfo] = {}

    for _, row in df.iterrows():
        if input_col:
            name = clean_cell(row.get(input_col))
            if name:
                inputs[name] = PortInfo(
                    name=name,
                    width=clean_cell(row.get(input_width_col)) if input_width_col else "",
                    used_width=clean_cell(row.get(input_used_col)) if input_used_col else "",
                    from_whom=clean_cell(row.get(from_col)) if from_col else "",
                )
        if output_col:
            name = clean_cell(row.get(output_col))
            if name:
                outputs[name] = PortInfo(
                    name=name,
                    width=clean_cell(row.get(output_width_col)) if output_width_col else "",
                    used_width=clean_cell(row.get(output_used_col)) if output_used_col else "",
                    to_top=clean_cell(row.get(to_top_col)) if to_top_col else "",
                )
        if inout_col:
            name = clean_cell(row.get(inout_col))
            if name:
                inouts[name] = PortInfo(
                    name=name,
                    width=clean_cell(row.get(inout_width_col)) if inout_width_col else "",
                    connectivity=clean_cell(row.get(inout_conn_col)) if inout_conn_col else "",
                    inout_name=clean_cell(row.get(inout_name_col)) if inout_name_col else "",
                )
    return {"inputs": inputs, "outputs": outputs, "inouts": inouts}


def read_port_workbooks(paths: Sequence[Path], report: Report) -> Dict[str, Dict[str, Dict[str, PortInfo]]]:
    sheets: Dict[str, Dict[str, Dict[str, PortInfo]]] = {}
    for path in paths:
        try:
            book = pd.ExcelFile(path)
        except Exception as exc:
            report.error(f"failed to open port workbook {path.name}: {exc}")
            continue
        for sheet_name in book.sheet_names:
            if sheet_name in sheets:
                report.warn(f"duplicate port sheet {sheet_name}; keeping first occurrence")
                continue
            try:
                sheets[sheet_name] = parse_port_sheet(read_excel_file(path, sheet_name))
            except Exception as exc:
                report.error(f"failed to read {path.name}:{sheet_name}: {exc}")
    report.info(f"loaded {len(sheets)} instance port sheet(s) from {len(paths)} workbook(s)")
    return sheets


def attach_port_data(instances: Dict[str, InstInfo], sheets: Dict[str, Dict[str, Dict[str, PortInfo]]], report: Report) -> None:
    for inst in instances.values():
        data = sheets.get(inst.inst_name)
        if not data:
            report.warn(f"no port sheet found for instance {inst.inst_name}")
            continue
        inst.inputs = data["inputs"]
        inst.outputs = data["outputs"]
        inst.inouts = data["inouts"]


def resolve_sdc_paths(instances: Dict[str, InstInfo], cwd: Path, report: Report) -> None:
    all_sdcs = sorted(path for path in cwd.glob("*.sdc") if path.is_file())
    by_name = {path.name: path for path in all_sdcs}
    by_lower = {path.name.lower(): path for path in all_sdcs}

    for inst in instances.values():
        candidates: List[str] = []
        if inst.sdc_hint:
            candidates.append(Path(inst.sdc_hint).name)
        candidates.append(f"{inst.inst_name}.sdc")
        candidates.append(f"{inst.module_name}.sdc")
        if inst.file_path:
            stem = Path(inst.file_path).stem
            candidates.append(f"{stem}.sdc")
            if stem.endswith("_empty"):
                candidates.append(f"{stem[:-6]}.sdc")

        matches: List[Path] = []
        for name in dict.fromkeys(candidates):
            if not name:
                continue
            if name in by_name:
                matches.append(by_name[name])
            elif name.lower() in by_lower:
                matches.append(by_lower[name.lower()])
        unique = []
        for path in matches:
            if path not in unique:
                unique.append(path)
        if len(unique) == 1:
            inst.sdc_path = unique[0]
        elif len(unique) > 1:
            inst.sdc_path = unique[0]
            report.warn(
                f"multiple SDC candidates for {inst.inst_name}: "
                f"{', '.join(path.name for path in unique)}; using {inst.sdc_path.name}"
            )
        else:
            report.warn(f"no SDC found for {inst.inst_name}; tried: {', '.join(candidates)}")


def collect_current_sdc_digests(instances: Dict[str, InstInfo]) -> Dict[str, str]:
    digests: Dict[str, str] = {}
    for inst in instances.values():
        if inst.sdc_path and inst.sdc_path.is_file():
            try:
                digests[inst.sdc_path.name] = digest_file(inst.sdc_path)
            except OSError:
                continue
    return digests


def parse_connections(value: str) -> List[str]:
    text = clean_cell(value)
    if not text:
        return []
    text = text.replace("\n", ";")
    text = re.sub(r"\band\b", ";", text, flags=re.IGNORECASE)
    parts = [part.strip() for part in re.split(r"[;,]+", text) if part.strip()]
    return parts or [text]


def parse_endpoint_ref(value: str) -> Tuple[str, str, str]:
    text = clean_cell(value)
    lowered = text.lower()
    for prefix in ("top.", "top/", "top:"):
        if lowered.startswith(prefix):
            return "top", "", text[len(prefix) :]
    if lowered.startswith("[get_ports"):
        objs = re.findall(r"\{([^}]+)\}", text)
        if objs:
            ports = split_object_list(objs[-1])
            if ports:
                return "top", "", ports[0]
    match = re.match(r"^([^./:\s]+)[./:]([^./:\s]+)$", text)
    if match:
        return "inst", match.group(1), match.group(2)
    return "unknown", "", text


def endpoint_collection(inst_name: str, port_name: str) -> str:
    if not inst_name or not port_name:
        return ""
    return get_collection("get_pins", [f"{inst_name}/{port_name}"])


def build_channel_id(src_inst: str, src_port: str, dst_inst: str, dst_port: str) -> str:
    return "CH_" + "__".join(
        [
            f"{sanitize_id(src_inst)}_{sanitize_id(src_port)}",
            f"{sanitize_id(dst_inst)}_{sanitize_id(dst_port)}",
        ]
    )


def build_channels(
    instances: Dict[str, InstInfo],
    clock_objects: Set[str],
    report: Report,
) -> List[ChannelRecord]:
    channels: List[ChannelRecord] = []
    seen: Set[str] = set()

    def add(record: ChannelRecord) -> None:
        if record.channel_id in seen:
            return
        seen.add(record.channel_id)
        if endpoint_hits_clock(record.src_endpoint, clock_objects) or endpoint_hits_clock(record.dst_endpoint, clock_objects):
            record.is_clock_related = "yes"
            if record.channel_type in CHANNEL_TYPES_10:
                record.channel_type = "clock_connection"
        channels.append(record)

    for dst in instances.values():
        for port in dst.inputs.values():
            refs = parse_connections(port.from_whom)
            if not refs:
                channel_id = build_channel_id("unknown", "unknown", dst.inst_name, port.name)
                add(
                    ChannelRecord(
                        channel_id=channel_id,
                        scenario="common",
                        stage="all",
                        corner="all",
                        channel_type="unknown",
                        src_instance="",
                        src_module="",
                        src_port="",
                        src_endpoint="",
                        dst_instance=dst.inst_name,
                        dst_module=dst.module_name,
                        dst_port=port.name,
                        dst_endpoint=endpoint_collection(dst.inst_name, port.name),
                        connection_source="",
                        note="input source is blank in integration table",
                    )
                )
                continue
            for ref in refs:
                kind, src_inst, src_port = parse_endpoint_ref(ref)
                if kind == "inst" and src_inst in instances:
                    src = instances[src_inst]
                    channel_type = "harden_to_harden"
                    src_endpoint = endpoint_collection(src_inst, src_port)
                    src_module = src.module_name
                elif kind == "top":
                    channel_type = "top_pad_to_harden"
                    src_endpoint = get_collection("get_ports", [src_port])
                    src_module = "top"
                    src_inst = "top"
                else:
                    channel_type = "fabric_to_harden" if kind == "unknown" else "unknown"
                    src_endpoint = ""
                    src_module = ""
                channel_id = build_channel_id(src_inst or "fabric", src_port or "unknown", dst.inst_name, port.name)
                add(
                    ChannelRecord(
                        channel_id=channel_id,
                        scenario="common",
                        stage="all",
                        corner="all",
                        channel_type=channel_type,
                        src_instance=src_inst,
                        src_module=src_module,
                        src_port=src_port,
                        src_endpoint=src_endpoint,
                        dst_instance=dst.inst_name,
                        dst_module=dst.module_name,
                        dst_port=port.name,
                        dst_endpoint=endpoint_collection(dst.inst_name, port.name),
                        connection_source=ref,
                        is_pad_related="yes" if channel_type == "top_pad_to_harden" else "no",
                        timing_model="unknown",
                        note="" if channel_type in CHANNEL_TYPES_10 else "not a 10 channel by default",
                    )
                )

    for src in instances.values():
        for port in src.outputs.values():
            dst_refs = parse_connections(port.to_top)
            if not dst_refs:
                continue
            for ref in dst_refs:
                kind, dst_inst, dst_port = parse_endpoint_ref(ref)
                if kind == "top":
                    channel_type = "harden_to_top_pad"
                    dst_instance = "top"
                    dst_module = "top"
                    dst_endpoint = get_collection("get_ports", [dst_port])
                    is_pad_related = "yes"
                    note = "pad-related output channel; handled by 04"
                elif kind == "inst" and dst_inst in instances:
                    # harden_to_harden is primarily built from the destination input.
                    continue
                else:
                    channel_type = "harden_to_fabric"
                    dst_instance = "fabric"
                    dst_module = "fabric"
                    dst_endpoint = ""
                    is_pad_related = "no"
                    note = "fabric sink endpoint cannot be fully resolved from output-side table; review dst_endpoint"
                channel_id = build_channel_id(src.inst_name, port.name, dst_instance, dst_port or "unknown")
                add(
                    ChannelRecord(
                        channel_id=channel_id,
                        scenario="common",
                        stage="all",
                        corner="all",
                        channel_type=channel_type,
                        src_instance=src.inst_name,
                        src_module=src.module_name,
                        src_port=port.name,
                        src_endpoint=endpoint_collection(src.inst_name, port.name),
                        dst_instance=dst_instance,
                        dst_module=dst_module,
                        dst_port=dst_port,
                        dst_endpoint=dst_endpoint,
                        connection_source=ref,
                        is_pad_related=is_pad_related,
                        timing_model="unknown",
                        note=note,
                    )
                )
    report.info(f"built {len(channels)} integration channel(s)")
    return channels


def is_escaped(text: str, index: int) -> bool:
    count = 0
    pos = index - 1
    while pos >= 0 and text[pos] == "\\":
        count += 1
        pos -= 1
    return count % 2 == 1


def strip_inline_comment(line: str) -> str:
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(line):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(line, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == "#" and brace_depth == 0 and bracket_depth == 0:
                if idx == 0 or line[idx - 1].isspace() or line[idx - 1] == ";":
                    return line[:idx].rstrip()
    return line.rstrip()


def split_semicolon_commands(text: str) -> List[str]:
    parts: List[str] = []
    start = 0
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(text):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(text, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == ";" and brace_depth == 0 and bracket_depth == 0:
                parts.append(text[start:idx])
                start = idx + 1
    parts.append(text[start:])
    return parts


def iter_tcl_commands_with_line(text: str) -> Iterable[TclCommand]:
    buf = ""
    start_line = 0
    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = strip_inline_comment(raw_line)
        if not line.strip():
            continue
        if not buf:
            start_line = line_no
        stripped = line.rstrip()
        continued = stripped.endswith("\\") and not is_escaped(stripped, len(stripped) - 1)
        if continued:
            stripped = stripped[:-1].rstrip()
            buf += stripped + " "
            continue
        buf += stripped
        for cmd in split_semicolon_commands(buf):
            cleaned = cmd.strip().rstrip(";").strip()
            if cleaned:
                yield TclCommand(cleaned, start_line)
        buf = ""
        start_line = 0
    if buf.strip():
        for cmd in split_semicolon_commands(buf):
            cleaned = cmd.strip().rstrip(";").strip()
            if cleaned:
                yield TclCommand(cleaned, start_line)


def find_matching(text: str, start: int, open_char: str, close_char: str) -> int:
    depth = 0
    quote = False
    idx = start
    while idx < len(text):
        char = text[idx]
        if char == "\\":
            idx += 2
            continue
        if char == '"' and open_char != '"' and not is_escaped(text, idx):
            quote = not quote
        elif not quote:
            if char == open_char:
                depth += 1
            elif char == close_char:
                depth -= 1
                if depth == 0:
                    return idx
        idx += 1
    return -1


def tokenize_tcl_words(text: str) -> List[str]:
    tokens: List[str] = []
    idx = 0
    while idx < len(text):
        while idx < len(text) and text[idx].isspace():
            idx += 1
        if idx >= len(text):
            break
        start = idx
        if text[idx] == "{":
            end = find_matching(text, idx, "{", "}")
            if end < 0:
                tokens.append(text[start:])
                break
            tokens.append(text[start : end + 1])
            idx = end + 1
            continue
        if text[idx] == '"':
            idx += 1
            while idx < len(text):
                if text[idx] == "\\":
                    idx += 2
                    continue
                if text[idx] == '"':
                    idx += 1
                    break
                idx += 1
            tokens.append(text[start:idx])
            continue
        pieces: List[str] = []
        while idx < len(text) and not text[idx].isspace():
            char = text[idx]
            if char == "\\":
                pieces.append(text[idx : idx + 2])
                idx += 2
            elif char == "[":
                end = find_matching(text, idx, "[", "]")
                if end < 0:
                    pieces.append(text[idx:])
                    idx = len(text)
                else:
                    pieces.append(text[idx : end + 1])
                    idx = end + 1
            else:
                pieces.append(char)
                idx += 1
        tokens.append("".join(pieces))
    return tokens


def strip_braces(text: str) -> str:
    text = clean_cell(text)
    if len(text) >= 2 and text[0] == "{" and text[-1] == "}":
        return text[1:-1].strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        return text[1:-1].strip()
    return text


def split_object_list(text: str) -> List[str]:
    text = strip_braces(text)
    return [part for part in re.split(r"[\s,;]+", text) if part]


COLLECTION_RE = re.compile(r"\[(get_ports|get_pins|get_nets|get_clocks)\s+([^\]]+)\]")


def parse_collection(token: str) -> Optional[Tuple[str, List[str]]]:
    match = COLLECTION_RE.fullmatch(clean_cell(token))
    if not match:
        return None
    return match.group(1), split_object_list(match.group(2))


OPTIONS_WITH_VALUE = {"-clock", "-min", "-max"}
OPTIONS_NO_VALUE = {"-add_delay", "-clock_fall", "-rise", "-fall"}


def option_value(tokens: Sequence[str], option: str) -> str:
    for idx, token in enumerate(tokens):
        if token == option and idx + 1 < len(tokens):
            return strip_braces(tokens[idx + 1])
    return ""


def has_option(tokens: Sequence[str], option: str) -> bool:
    return option in tokens


def positional_tokens(tokens: Sequence[str]) -> List[str]:
    result: List[str] = []
    skip_next = False
    for token in tokens[1:]:
        if skip_next:
            skip_next = False
            continue
        if token in OPTIONS_WITH_VALUE:
            skip_next = True
            continue
        if token in OPTIONS_NO_VALUE:
            continue
        if token.startswith("-"):
            continue
        if parse_collection(token):
            continue
        result.append(strip_braces(token))
    return result


def extract_clock_name(tokens: Sequence[str]) -> str:
    value = option_value(tokens, "-clock")
    parsed = parse_collection(value)
    if parsed and parsed[0] == "get_clocks" and parsed[1]:
        return parsed[1][0]
    return strip_braces(value)


def last_non_clock_collection(tokens: Sequence[str]) -> Optional[Tuple[str, List[str], str]]:
    found: Optional[Tuple[str, List[str], str]] = None
    for token in tokens:
        parsed = parse_collection(token)
        if parsed and parsed[0] != "get_clocks":
            found = (parsed[0], parsed[1], token)
    return found


def normalize_sdc_port(kind: str, obj: str, inst: InstInfo) -> str:
    obj = strip_braces(obj)
    if kind == "get_pins" and "/" in obj:
        return obj.split("/")[-1]
    return obj


def parse_delay_candidate(inst: InstInfo, cmd: TclCommand, digest: str, now: str) -> Optional[DelayCandidate]:
    tokens = tokenize_tcl_words(cmd.raw)
    if not tokens or tokens[0] not in {"set_input_delay", "set_output_delay"}:
        return None
    ctype = "input_delay" if tokens[0] == "set_input_delay" else "output_delay"
    target = last_non_clock_collection(tokens)
    if not target:
        return DelayCandidate(
            inst.inst_name,
            inst.module_name,
            inst.owner,
            "",
            "unknown",
            ctype,
            extract_clock_name(tokens),
            option_value(tokens, "-min"),
            option_value(tokens, "-max"),
            positional_tokens(tokens)[0] if positional_tokens(tokens) else "",
            "",
            inst.sdc_path.name if inst.sdc_path else "",
            str(cmd.line_no),
            digest,
            now,
            cmd.raw,
            "needs_review",
            "no target collection found",
        )
    kind, objects, _ = target
    if len(objects) != 1:
        port_name = " ".join(objects)
        status = "needs_review"
        message = "target collection contains multiple objects"
    else:
        port_name = normalize_sdc_port(kind, objects[0], inst)
        status = "ok"
        message = ""
    direction = "input" if port_name in inst.inputs else "output" if port_name in inst.outputs else "inout" if port_name in inst.inouts else "unknown"
    if ctype == "input_delay" and direction != "input":
        status = "needs_review"
        message = "; ".join(filter(None, [message, "set_input_delay target is not an input port in integration table"]))
    if ctype == "output_delay" and direction != "output":
        status = "needs_review"
        message = "; ".join(filter(None, [message, "set_output_delay target is not an output port in integration table"]))
    complex_opts = []
    for opt in sorted(OPTIONS_NO_VALUE):
        if has_option(tokens, opt):
            complex_opts.append(opt)
    if len(objects) != 1:
        complex_opts.append("multi_object")
    return DelayCandidate(
        inst_name=inst.inst_name,
        module_name=inst.module_name,
        owner=inst.owner,
        port_name=port_name,
        direction=direction,
        constraint_type=ctype,
        clock_name=extract_clock_name(tokens),
        min_value=option_value(tokens, "-min"),
        max_value=option_value(tokens, "-max"),
        bare_value=positional_tokens(tokens)[0] if positional_tokens(tokens) else "",
        complex_options=",".join(complex_opts),
        source_sdc_file=inst.sdc_path.name if inst.sdc_path else "",
        source_line=str(cmd.line_no),
        source_digest=digest,
        extraction_time=now,
        original_command=cmd.raw,
        parse_status=status,
        message=message,
    )


def extract_delay_candidates(instances: Dict[str, InstInfo], report: Report) -> List[DelayCandidate]:
    results: List[DelayCandidate] = []
    now = datetime.now().isoformat(timespec="seconds")
    for inst in instances.values():
        if not inst.sdc_path:
            continue
        try:
            text = read_text(inst.sdc_path)
            digest = digest_file(inst.sdc_path)
        except Exception as exc:
            report.error(f"failed to read {inst.sdc_path}: {exc}")
            continue
        count = 0
        for cmd in iter_tcl_commands_with_line(text):
            cand = parse_delay_candidate(inst, cmd, digest, now)
            if cand:
                results.append(cand)
                count += 1
        report.info(f"extracted {count} interface delay candidate(s) from {inst.sdc_path.name}")
    return results


def read_clock_inventory(path: Path, report: Report) -> Dict[str, ClockInfo]:
    clocks: Dict[str, ClockInfo] = {}
    if not path.is_file():
        report.warn(f"clock inventory not found: {path}")
        return clocks
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        if not reader.fieldnames or "clock_name" not in reader.fieldnames:
            report.error(f"{path} does not contain clock_name column")
            return clocks
        for row in reader:
            action = clean_cell(row.get("final_action"))
            if action and action not in ACTIVE_01_ACTIONS:
                continue
            clock_name = clean_cell(row.get("clock_name"))
            if not clock_name:
                continue
            clocks[clock_name] = ClockInfo(
                clock_name=clock_name,
                direct_source=clean_cell(row.get("direct_source")),
                producer_object=clean_cell(row.get("producer_object")),
                final_action=action,
            )
    report.info(f"loaded {len(clocks)} clock(s) from {path}")
    return clocks


def clock_object_set(clocks: Dict[str, ClockInfo]) -> Set[str]:
    objects: Set[str] = set()
    for clock in clocks.values():
        for obj in (clock.direct_source, clock.producer_object):
            text = clean_cell(obj)
            if not text:
                continue
            for item in re.findall(r"\{([^}]+)\}", text):
                objects.update(split_object_list(item))
            if "/" in text and not text.startswith("["):
                objects.add(text)
    return objects


def endpoint_objects(endpoint: str) -> Set[str]:
    objects: Set[str] = set()
    text = clean_cell(endpoint)
    if not text:
        return objects
    for item in re.findall(r"\{([^}]+)\}", text):
        objects.update(split_object_list(item))
    if "/" in text and not text.startswith("["):
        objects.add(text)
    return objects


def endpoint_hits_clock(endpoint: str, clock_objects: Set[str]) -> bool:
    if clean_cell(endpoint) in clock_objects:
        return True
    return bool(endpoint_objects(endpoint) & clock_objects)


def create_budget_seeds(channels: Sequence[ChannelRecord], candidates: Sequence[DelayCandidate], report: Report) -> List[BudgetSeed]:
    src_by_key: Dict[Tuple[str, str], List[DelayCandidate]] = defaultdict(list)
    dst_by_key: Dict[Tuple[str, str], List[DelayCandidate]] = defaultdict(list)
    for cand in candidates:
        if cand.constraint_type == "output_delay":
            src_by_key[(cand.inst_name, cand.port_name)].append(cand)
        elif cand.constraint_type == "input_delay":
            dst_by_key[(cand.inst_name, cand.port_name)].append(cand)

    seeds: List[BudgetSeed] = []
    matched_candidate_ids: Set[int] = set()
    for ch in channels:
        source_cands = src_by_key.get((ch.src_instance, ch.src_port), [])
        dest_cands = dst_by_key.get((ch.dst_instance, ch.dst_port), [])
        related = source_cands + dest_cands
        matched_candidate_ids.update(id(cand) for cand in related)
        values = {header: "" for header in BUDGET_HEADERS}
        values.update(
            {
                "channel_id": ch.channel_id,
                "scenario": ch.scenario,
                "stage": ch.stage,
                "corner": ch.corner,
                "channel_type": ch.channel_type,
                "is_pad_related": ch.is_pad_related,
                "is_clock_related": ch.is_clock_related,
                "is_feedthrough": ch.is_feedthrough,
                "src_endpoint": ch.src_endpoint,
                "dst_endpoint": ch.dst_endpoint,
                "timing_model": ch.timing_model,
                "budget_required": ch.budget_required,
                "clock_relation": ch.clock_relation,
                "budget_model": "unknown",
                "src_output_delay_max": join_unique(c.max_value or c.bare_value for c in source_cands),
                "src_output_delay_min": join_unique(c.min_value for c in source_cands),
                "dst_input_delay_max": join_unique(c.max_value or c.bare_value for c in dest_cands),
                "dst_input_delay_min": join_unique(c.min_value for c in dest_cands),
                "max_source": "",
                "min_source": "",
                "original_src_clock": join_unique(c.clock_name for c in source_cands),
                "original_dst_clock": join_unique(c.clock_name for c in dest_cands),
                "complex_options": join_unique(c.complex_options for c in related),
                "tool_surface": "sta",
                "datapath_only": "yes",
                "source_type": "extracted" if related else "manual",
                "source_sdc_file": join_present(c.source_sdc_file for c in related),
                "source_line": join_present(c.source_line for c in related),
                "source_digest": join_present(c.source_digest for c in related),
                "extraction_time": join_present(c.extraction_time for c in related),
                "source_command": join_unique((c.original_command for c in related), sep=" || "),
                "apply": "no",
                "emit_max": "no",
                "emit_min": "no",
                "review_status": "pending",
                "owner": join_unique(c.owner for c in related),
            }
        )
        messages = [c.message for c in related if c.message]
        if ch.channel_type not in CHANNEL_TYPES_10:
            messages.append("channel_type is not emitted by 10")
        if not related and ch.channel_type in CHANNEL_TYPES_10:
            messages.append("no input/output delay candidate found on either side")
        if any(c.complex_options for c in related):
            messages.append("complex delay options require review")
        values["note"] = "; ".join(dict.fromkeys(messages))
        status = "ok" if related and not messages else "needs_review"
        seeds.append(BudgetSeed(values, status, ch, related))
    for cand in candidates:
        if id(cand) not in matched_candidate_ids and cand.direction in {"input", "output"}:
            report.warn(
                f"{cand.source_sdc_file}:{cand.source_line} {cand.inst_name}/{cand.port_name}: "
                f"{cand.constraint_type} candidate has no matching 10 channel in integration table"
            )
    return seeds


def join_unique(values: Iterable[str], sep: str = "; ") -> str:
    result: List[str] = []
    for value in values:
        text = clean_cell(value)
        if text and text not in result:
            result.append(text)
    return sep.join(result)


def join_present(values: Iterable[str], sep: str = "; ") -> str:
    return sep.join(clean_cell(value) for value in values if clean_cell(value))


def create_or_load_workbook(path: Path) -> Tuple[Workbook, bool]:
    if path.is_file():
        return load_workbook(path), False
    wb = Workbook()
    ws = wb.active
    ws.title = "interface_budget"
    return wb, True


def ensure_sheet(wb: Workbook, name: str, headers: Sequence[str]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if ws.max_row == 1 and all(ws.cell(row=1, column=col).value is None for col in range(1, len(headers) + 1)):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    existing = [clean_cell(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
    for header in headers:
        if header not in existing:
            ws.cell(row=1, column=len(existing) + 1, value=header)
            existing.append(header)
    style_sheet(ws)


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for column_cells in ws.columns:
        max_len = 8
        for cell in column_cells:
            max_len = max(max_len, len(clean_cell(cell.value)))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_len + 2, 45)
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:25] or "table"
        if not ws.tables:
            ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"
            tab = Table(displayName=f"{table_name}_tbl", ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            try:
                ws.add_table(tab)
            except ValueError:
                pass


def header_map(ws) -> Dict[str, int]:
    return {clean_cell(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean_cell(cell.value)}


def row_values(ws, row_idx: int, headers: Sequence[str]) -> Dict[str, object]:
    hmap = header_map(ws)
    return {header: ws.cell(row=row_idx, column=hmap[header]).value if header in hmap else "" for header in headers}


def append_dict(ws, headers: Sequence[str], values: Dict[str, object], fill: Optional[PatternFill] = None) -> None:
    hmap = header_map(ws)
    row_idx = ws.max_row + 1
    for header in headers:
        col = hmap[header]
        cell = ws.cell(row=row_idx, column=col, value=values.get(header, ""))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def channel_key(values: Dict[str, object]) -> str:
    return clean_cell(values.get("channel_id"))


def budget_key(values: Dict[str, object]) -> Tuple[str, str, str, str]:
    return (
        clean_cell(values.get("channel_id")),
        normalize_key(values.get("scenario")) or "common",
        normalize_key(values.get("stage")) or "all",
        clean_cell(values.get("corner")) or "all",
    )


def sync_workbook(path: Path, channels: Sequence[ChannelRecord], seeds: Sequence[BudgetSeed], report: Report) -> None:
    wb, created = create_or_load_workbook(path)
    ensure_sheet(wb, "interface_budget", BUDGET_HEADERS)
    ensure_sheet(wb, "channel_inventory", CHANNEL_HEADERS)
    ensure_sheet(wb, "extraction_log", LOG_HEADERS)

    ws_ch = wb["channel_inventory"]
    existing_channels = {
        channel_key(row_values(ws_ch, row_idx, CHANNEL_HEADERS))
        for row_idx in range(2, ws_ch.max_row + 1)
        if clean_cell(ws_ch.cell(row=row_idx, column=1).value)
    }
    for ch in channels:
        values = {header: getattr(ch, header) for header in CHANNEL_HEADERS}
        if ch.channel_id not in existing_channels:
            append_dict(ws_ch, CHANNEL_HEADERS, values, NEW_FILL)
            existing_channels.add(ch.channel_id)
            report.sync_changed = True

    ws_budget = wb["interface_budget"]
    existing_budget = {
        budget_key(row_values(ws_budget, row_idx, BUDGET_HEADERS))
        for row_idx in range(2, ws_budget.max_row + 1)
        if clean_cell(ws_budget.cell(row=row_idx, column=1).value)
    }
    for seed in seeds:
        key = budget_key(seed.values)
        if key not in existing_budget:
            append_dict(ws_budget, BUDGET_HEADERS, seed.values, NEW_FILL)
            existing_budget.add(key)
            report.sync_changed = True

    ws_log = wb["extraction_log"]
    if ws_log.max_row > 1:
        ws_log.delete_rows(2, ws_log.max_row - 1)
    for seed in seeds:
        for cand in seed.candidates:
            values = {
                "source_sdc_file": cand.source_sdc_file,
                "source_line": cand.source_line,
                "instance": cand.inst_name,
                "port": cand.port_name,
                "direction": cand.direction,
                "constraint_type": cand.constraint_type,
                "clock_name": cand.clock_name,
                "min_value": cand.min_value,
                "max_value": cand.max_value or cand.bare_value,
                "parse_status": cand.parse_status,
                "channel_ids": seed.channel.channel_id,
                "source_digest": cand.source_digest,
                "extraction_time": cand.extraction_time,
                "original_command": cand.original_command,
                "message": cand.message,
            }
            append_dict(ws_log, LOG_HEADERS, values)

    add_validations(wb)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(path)
    if created or report.sync_changed:
        report.sync_changed = True
        report.info(f"synchronized workbook {path.name}; review new rows before generation")


def add_validations(wb: Workbook) -> None:
    if "interface_budget" not in wb.sheetnames:
        return
    ws = wb["interface_budget"]
    hmap = header_map(ws)

    def add_list(header: str, values: Sequence[str]) -> None:
        if header not in hmap:
            return
        col = get_column_letter(hmap[header])
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1048576")

    add_list("scenario", sorted(SCENARIOS))
    add_list("stage", sorted(STAGES))
    add_list("channel_type", sorted(CHANNEL_TYPES_10 | NON_10_CHANNEL_TYPES))
    add_list("timing_model", sorted(TIMING_MODELS - {""}))
    add_list("budget_required", ["yes", "no"])
    add_list("budget_model", sorted(BUDGET_MODELS - {""}))
    add_list("tool_surface", sorted(TOOL_SURFACES - {""}))
    add_list("datapath_only", ["yes", "no"])
    add_list("min_sign_review", sorted(MIN_REVIEW_VALUES - {""}))
    add_list("source_type", sorted(SOURCE_TYPES - {""}))
    add_list("apply", ["yes", "no"])
    add_list("emit_max", ["yes", "no"])
    add_list("emit_min", ["yes", "no"])
    add_list("review_status", sorted(REVIEW_STATUS_VALUES - {""}))


def read_form_rows(path: Path) -> List[FormRow]:
    wb = load_workbook(path, data_only=False)
    if "interface_budget" not in wb.sheetnames:
        raise RuntimeError(f"{path} missing interface_budget sheet")
    ws = wb["interface_budget"]
    rows: List[FormRow] = []
    for row_idx in range(2, ws.max_row + 1):
        values = row_values(ws, row_idx, BUDGET_HEADERS)
        if not any(clean_cell(value) for value in values.values()):
            continue
        rows.append(FormRow(row_idx=row_idx, values=values))
    return rows


def write_autofilled_fields(path: Path, rows: Sequence[FormRow], report: Report) -> None:
    changed_rows = [row for row in rows if row.autofilled_fields]
    if not changed_rows:
        return
    wb = load_workbook(path)
    if "interface_budget" not in wb.sheetnames:
        return
    ws = wb["interface_budget"]
    hmap = header_map(ws)
    changed = 0
    for row in changed_rows:
        for field_name in sorted(row.autofilled_fields):
            if field_name not in hmap:
                continue
            new_value = clean_cell(row.values.get(field_name))
            cell = ws.cell(row=row.row_idx, column=hmap[field_name])
            if clean_cell(cell.value) != new_value:
                cell.value = new_value
                cell.fill = NEW_FILL
                changed += 1
    if changed:
        wb.save(path)
        report.info(f"wrote {changed} auto-resolved field(s) back to {path.name}")


def row_scenario(row: FormRow) -> str:
    return normalize_key(row.values.get("scenario")) or "common"


def row_stage(row: FormRow) -> str:
    return normalize_key(row.values.get("stage")) or "all"


def row_corner(row: FormRow) -> str:
    return clean_cell(row.values.get("corner")) or "all"


def is_apply_approved(row: FormRow) -> bool:
    return (
        normalize_key(row.values.get("apply")) == "yes"
        and normalize_key(row.values.get("review_status")) == "approved"
        and normalize_key(row.values.get("source_type")) != "na"
    )


def row_selected_for_output(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    return row_scenario(row) == scenario and row_stage(row) == stage and row_corner(row) == corner


def row_selected_for_assembled(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    row_sc = row_scenario(row)
    if row_sc not in {"common", scenario}:
        return False
    row_st = row_stage(row)
    row_co = row_corner(row)
    return (row_st == "all" and row_co == "all") or (row_st == stage and row_co == corner)


def validate_rows(
    rows: Sequence[FormRow],
    channels: Sequence[ChannelRecord],
    scenario: str,
    stage: str,
    corner: str,
    current_digests: Dict[str, str],
    max_diff_threshold: Optional[float],
    report: Report,
) -> None:
    channel_ids = {ch.channel_id for ch in channels}
    channel_by_id = {ch.channel_id: ch for ch in channels}
    assembled = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner) and is_apply_approved(row)]
    emitted_by_channel: Dict[str, List[FormRow]] = defaultdict(list)

    for row in rows:
        values = row.values
        apply_value = normalize_key(values.get("apply"))
        review_status = normalize_key(values.get("review_status"))
        source_type = normalize_key(values.get("source_type"))
        channel_id = clean_cell(values.get("channel_id"))
        channel_type = normalize_key(values.get("channel_type"))
        timing_model = normalize_key(values.get("timing_model"))
        budget_model = normalize_key(values.get("budget_model"))
        budget_required = normalize_key(values.get("budget_required"))
        tool_surface = normalize_key(values.get("tool_surface"))
        datapath_only = normalize_key(values.get("datapath_only"))
        clock_relation = normalize_key(values.get("clock_relation"))
        relation_basis = clean_cell(values.get("relationship_override_basis"))
        channel = channel_by_id.get(channel_id)
        is_pad_related = normalize_key(values.get("is_pad_related")) or (channel.is_pad_related if channel else "")
        is_clock_related = normalize_key(values.get("is_clock_related")) or (channel.is_clock_related if channel else "")
        is_feedthrough = normalize_key(values.get("is_feedthrough")) or (channel.is_feedthrough if channel else "")

        if apply_value and apply_value not in YES_NO:
            report.error(f"interface_budget row {row.row_idx}: apply must be yes/no")
        if review_status and review_status not in REVIEW_STATUS_VALUES:
            report.error(f"interface_budget row {row.row_idx}: invalid review_status {review_status}")
        if source_type and source_type not in SOURCE_TYPES:
            report.error(f"interface_budget row {row.row_idx}: invalid source_type {source_type}")
        if tool_surface and tool_surface not in TOOL_SURFACES:
            report.error(f"interface_budget row {row.row_idx}: invalid tool_surface {tool_surface}")
        if timing_model and timing_model not in TIMING_MODELS:
            report.error(f"interface_budget row {row.row_idx}: invalid timing_model {timing_model}")
        if budget_model and budget_model not in BUDGET_MODELS:
            report.error(f"interface_budget row {row.row_idx}: invalid budget_model {budget_model}")
        for flag_name, flag_value in (
            ("is_pad_related", is_pad_related),
            ("is_clock_related", is_clock_related),
            ("is_feedthrough", is_feedthrough),
        ):
            if flag_value and flag_value not in YES_NO:
                report.error(f"interface_budget row {row.row_idx}: {flag_name} must be yes/no")

        if apply_value == "yes":
            if not channel_id:
                report.error(f"interface_budget row {row.row_idx}: apply=yes but channel_id is blank")
            if channel_id and channel_id not in channel_ids:
                report.error(f"interface_budget row {row.row_idx}: channel_id {channel_id} not found in channel_inventory")
            if not clean_cell(values.get("scenario")):
                report.error(f"interface_budget row {row.row_idx}: apply=yes but scenario is blank")
            if not clean_cell(values.get("stage")):
                report.error(f"interface_budget row {row.row_idx}: apply=yes but stage is blank")
            if not clean_cell(values.get("corner")):
                report.error(f"interface_budget row {row.row_idx}: apply=yes but corner is blank")
            if not clean_cell(values.get("src_endpoint")) or not clean_cell(values.get("dst_endpoint")):
                report.error(f"interface_budget row {row.row_idx} {channel_id}: src/dst endpoint is required")
            if is_pad_related == "yes":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: pad-related channel must be handled by 04")
            if is_clock_related == "yes":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: clock-related channel must be handled by 01/02/03")
            if is_feedthrough == "yes":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: feedthrough channel must be handled by 30")
            if channel_type not in CHANNEL_TYPES_10:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: channel_type={channel_type} is not generated by 10")
            if timing_model == "visible_netlist" and budget_required != "yes":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: visible_netlist path needs budget_required=yes to emit 10")
            if not timing_model or timing_model == "unknown":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: timing_model must be reviewed before emit")
            if not budget_model or budget_model == "unknown":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: budget_model must be reviewed before emit")
            if budget_model == "clock_relative_io_delay" and not clean_cell(values.get("derivation_basis")):
                report.error(
                    f"interface_budget row {row.row_idx} {channel_id}: clock_relative_io_delay requires derivation_basis"
                )
            auto_resolve_interconnect_max(row, report)
            if relation_blocks_10(clock_relation) and not relation_basis:
                report.error(
                    f"interface_budget row {row.row_idx} {channel_id}: clock_relation={clock_relation} blocks normal 10 budget"
                )
            if normalize_key(values.get("emit_max")) == "yes" and not clean_cell(values.get("converted_max")):
                report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_max=yes but converted_max is blank")
            if normalize_key(values.get("emit_min")) == "yes":
                if not clean_cell(values.get("converted_min")):
                    report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_min=yes but converted_min is blank")
                if normalize_key(values.get("min_sign_review")) not in MIN_REVIEW_VALUES - {""}:
                    report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_min=yes requires min_sign_review")
            if tool_surface in {"dc", "both"} and not datapath_only:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: DC/both tool surface requires datapath_only strategy")
            if not clean_cell(values.get("budget_basis")):
                report.error(f"interface_budget row {row.row_idx} {channel_id}: budget_basis is required")
            check_source_digest(row, current_digests, report)

        if row in assembled and is_apply_approved(row):
            emitted_by_channel[channel_id].append(row)
            warn_budget_diff(row, max_diff_threshold, report)
            warn_budget_shape(row, report)
            if normalize_key(values.get("budget_model")) == "interconnect_budget" and "interconnect" not in normalize_key(values.get("budget_basis")):
                report.warn(
                    f"interface_budget row {row.row_idx} {channel_id}: interconnect_budget should cite owner definition in budget_basis"
                )
            if normalize_key(values.get("budget_model")) == "manual_budget" and not clean_cell(values.get("derivation_basis")):
                report.warn(f"interface_budget row {row.row_idx} {channel_id}: manual_budget should include derivation_basis")

    for channel_id, group in emitted_by_channel.items():
        max_values = {clean_cell(row.values.get("converted_max")) for row in group if normalize_key(row.values.get("emit_max")) == "yes"}
        min_values = {clean_cell(row.values.get("converted_min")) for row in group if normalize_key(row.values.get("emit_min")) == "yes"}
        if len(max_values - {""}) > 1 or len(min_values - {""}) > 1:
            rows_text = ", ".join(str(row.row_idx) for row in group)
            report.error(f"assembled view conflict for channel {channel_id}: rows {rows_text}")
    warn_manual_fanout_reuse(assembled, report)


def relation_blocks_10(value: str) -> bool:
    text = normalize_key(value)
    if not text:
        return False
    return any(item in text for item in RELATION_BLOCKING)


def warn_manual_fanout_reuse(rows: Sequence[FormRow], report: Report) -> None:
    by_src_budget: Dict[Tuple[str, str], List[FormRow]] = defaultdict(list)
    for row in rows:
        values = row.values
        if normalize_key(values.get("budget_model")) != "manual_budget":
            continue
        src = clean_cell(values.get("src_endpoint"))
        budget = clean_cell(values.get("converted_max"))
        if src and budget:
            by_src_budget[(src, budget)].append(row)
    for (src, budget), group in by_src_budget.items():
        dsts = {clean_cell(row.values.get("dst_endpoint")) for row in group}
        if len(dsts) > 1:
            rows_text = ", ".join(str(row.row_idx) for row in group)
            report.warn(
                f"manual_budget fanout reuse: src {src} uses converted_max={budget} "
                f"for {len(dsts)} sinks in rows {rows_text}; confirm shared budget is intentional"
            )


def numeric_parts(value) -> List[float]:
    numbers: List[float] = []
    for part in clean_cell(value).split(";"):
        number = parse_number(part)
        if number is not None:
            numbers.append(number)
    return numbers


def auto_resolve_interconnect_max(row: FormRow, report: Report) -> None:
    values = row.values
    if normalize_key(values.get("budget_model")) != "interconnect_budget":
        return
    if normalize_key(values.get("emit_max")) != "yes":
        return
    if clean_cell(values.get("converted_max")):
        return
    if clean_cell(values.get("complex_options")):
        report.warn(
            f"interface_budget row {row.row_idx} {clean_cell(values.get('channel_id'))}: "
            "complex delay options present; converted_max is not auto-resolved"
        )
        return
    src_nums = numeric_parts(values.get("src_output_delay_max"))
    dst_nums = numeric_parts(values.get("dst_input_delay_max"))
    if not src_nums or not dst_nums:
        return
    converted = min(src_nums + dst_nums)
    values["converted_max"] = format_number(converted)
    row.autofilled_fields.add("converted_max")
    if not clean_cell(values.get("max_source")):
        values["max_source"] = "auto_min_interconnect_budget"
        row.autofilled_fields.add("max_source")
    if not clean_cell(values.get("derivation_basis")):
        values["derivation_basis"] = "min(src_output_delay_max,dst_input_delay_max)"
        row.autofilled_fields.add("derivation_basis")
    report.info(
        f"interface_budget row {row.row_idx} {clean_cell(values.get('channel_id'))}: "
        f"auto-resolved converted_max={format_number(converted)} from interconnect budget candidates"
    )


def warn_budget_diff(row: FormRow, threshold: Optional[float], report: Report) -> None:
    if threshold is None:
        return
    values = row.values
    nums = numeric_parts(values.get("src_output_delay_max")) + numeric_parts(values.get("dst_input_delay_max"))
    if len(nums) >= 2 and max(nums) - min(nums) > threshold:
        report.warn(
            f"interface_budget row {row.row_idx} {clean_cell(values.get('channel_id'))}: "
            f"max candidates differ by {max(nums) - min(nums):.12g}"
        )


def warn_budget_shape(row: FormRow, report: Report) -> None:
    values = row.values
    channel_id = clean_cell(values.get("channel_id"))
    src_max = numeric_parts(values.get("src_output_delay_max"))
    dst_max = numeric_parts(values.get("dst_input_delay_max"))
    if normalize_key(values.get("budget_model")) == "interconnect_budget":
        if bool(src_max) != bool(dst_max):
            report.warn(
                f"interface_budget row {row.row_idx} {channel_id}: "
                "interconnect_budget has max candidate from only one side"
            )
    if any(number < 0 for number in numeric_parts(values.get("src_output_delay_min")) + numeric_parts(values.get("dst_input_delay_min"))):
        report.warn(
            f"interface_budget row {row.row_idx} {channel_id}: original min delay contains negative value; "
            "review set_min_delay sign convention carefully"
        )
    complex_options = clean_cell(values.get("complex_options"))
    if complex_options:
        report.warn(
            f"interface_budget row {row.row_idx} {channel_id}: approved row contains complex delay option(s): {complex_options}"
        )
    src_clock = clean_cell(values.get("original_src_clock"))
    dst_clock = clean_cell(values.get("original_dst_clock"))
    if src_clock and dst_clock and src_clock != dst_clock:
        report.warn(
            f"interface_budget row {row.row_idx} {channel_id}: source/destination original clocks differ "
            f"({src_clock} vs {dst_clock})"
        )
    if normalize_key(values.get("min_sign_review")) == "waived" and not clean_cell(values.get("note")):
        report.warn(
            f"interface_budget row {row.row_idx} {channel_id}: min_sign_review=waived should include note/basis"
        )


def check_source_digest(row: FormRow, current_digests: Dict[str, str], report: Report) -> None:
    sources = [part.strip() for part in clean_cell(row.values.get("source_sdc_file")).split(";") if part.strip()]
    digests = [part.strip() for part in clean_cell(row.values.get("source_digest")).split(";") if part.strip()]
    for idx, source in enumerate(sources):
        stored = digests[idx] if idx < len(digests) else ""
        current = current_digests.get(source)
        if current and stored and current != stored:
            report.warn(
                f"interface_budget row {row.row_idx} {clean_cell(row.values.get('channel_id'))}: "
                f"source_digest mismatch for {source}; row may be stale"
            )


def output_sdc_path(cwd: Path, scenario: str, stage: str, corner: str) -> Path:
    if scenario == "common":
        if stage == "all" and corner == "all":
            return cwd / "common/10_harden_x_if.sdc"
        return cwd / f"common/10_harden_x_if_{stage}_{safe_filename_token(corner)}.sdc"
    if stage == "all" and corner == "all":
        return cwd / f"scenarios/{scenario}_harden_x_if.sdc"
    return cwd / f"scenarios/{scenario}_harden_x_if_{stage}_{safe_filename_token(corner)}.sdc"


def report_path(cwd: Path, scenario: str, stage: str, corner: str) -> Path:
    return cwd / f"harden_x_if_check_report_{scenario}_{stage}_{safe_filename_token(corner)}.txt"


def generate_sdc(rows: Sequence[FormRow], scenario: str, stage: str, corner: str) -> List[str]:
    selected = [row for row in rows if row_selected_for_output(row, scenario, stage, corner) and is_apply_approved(row)]
    lines = [
        "################################################################################",
        (
            "# Auto-generated SoC harden/subsys interface budget constraints for "
            f"scenario: {scenario}, stage: {stage}, corner: {corner}"
        ),
        "# Source: 10_harden_x_if.xlsx interface_budget sheet",
        "# Only apply=yes and review_status=approved rows are emitted.",
        "################################################################################",
        "",
    ]
    emitted = 0
    for row in selected:
        values = row.values
        commands = commands_for_row(row)
        if not commands:
            continue
        lines.append(f"# row {row.row_idx}: {clean_cell(values.get('channel_id'))}")
        if clean_cell(values.get("budget_basis")):
            lines.append(f"# Budget basis: {clean_cell(values.get('budget_basis'))}")
        if clean_cell(values.get("derivation_basis")):
            lines.append(f"# Derivation: {clean_cell(values.get('derivation_basis'))}")
        source_refs = format_source_refs(values.get("source_sdc_file"), values.get("source_line"))
        if source_refs:
            lines.append(f"# Source SDC: {source_refs}")
        lines.extend(commands)
        lines.append("")
        emitted += len(commands)
    if emitted == 0:
        lines.append("# No harden/subsys interface budget commands emitted for selected scenario/stage/corner.")
    return lines


def format_source_refs(source_files, source_lines) -> str:
    files = [part.strip() for part in clean_cell(source_files).split(";") if part.strip()]
    lines = [part.strip() for part in clean_cell(source_lines).split(";") if part.strip()]
    if not files:
        return ""
    if len(files) == len(lines):
        return "; ".join(f"{source}:{line}" if line else source for source, line in zip(files, lines))
    if lines:
        return f"{'; '.join(files)} lines {'; '.join(lines)}"
    return "; ".join(files)


def commands_for_row(row: FormRow) -> List[str]:
    values = row.values
    src = clean_cell(values.get("src_endpoint"))
    dst = clean_cell(values.get("dst_endpoint"))
    if not src or not dst:
        return []
    datapath = " -datapath_only" if normalize_key(values.get("datapath_only")) != "no" else ""
    commands: List[str] = []
    if normalize_key(values.get("emit_max")) == "yes" and clean_cell(values.get("converted_max")):
        commands.append(f"set_max_delay {format_number(values.get('converted_max'))}{datapath} -from {src} -to {dst}")
    if normalize_key(values.get("emit_min")) == "yes" and clean_cell(values.get("converted_min")):
        commands.append(f"set_min_delay {format_number(values.get('converted_min'))}{datapath} -from {src} -to {dst}")
    return commands


def build_coverage_lines(rows: Sequence[FormRow], channels: Sequence[ChannelRecord], scenario: str, stage: str, corner: str) -> List[str]:
    selected = [row for row in rows if row_selected_for_output(row, scenario, stage, corner) and is_apply_approved(row)]
    assembled = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner) and is_apply_approved(row)]
    by_channel: Dict[str, List[FormRow]] = defaultdict(list)
    for row in assembled:
        by_channel[clean_cell(row.values.get("channel_id"))].append(row)

    lines = [
        "",
        "Coverage:",
        f"  inventory channels     : {len(channels)}",
        f"  assembled approved row : {len(assembled)}",
        f"  emitted approved row   : {len(selected)}",
        "",
        "  Per-channel assembled status:",
    ]
    if not channels:
        lines.append("    <no channel_inventory records>")
    for ch in channels:
        group = by_channel.get(ch.channel_id, [])
        if group:
            statuses = []
            for row in group:
                values = row.values
                emit = []
                if normalize_key(values.get("emit_max")) == "yes":
                    emit.append("max")
                if normalize_key(values.get("emit_min")) == "yes":
                    emit.append("min")
                statuses.append(
                    "row={row} timing={timing} budget={budget} emit={emit}".format(
                        row=row.row_idx,
                        timing=normalize_key(values.get("timing_model")) or "-",
                        budget=normalize_key(values.get("budget_model")) or "-",
                        emit="+".join(emit) or "none",
                    )
                )
            status = "; ".join(statuses)
        else:
            status = "NO_APPROVED_BUDGET"
        lines.append(
            f"    {ch.channel_id}: type={ch.channel_type} inventory_timing={ch.timing_model} "
            f"pad={ch.is_pad_related} clock={ch.is_clock_related} status={status}"
        )

    not_emitted = [
        row
        for row in rows
        if row_selected_for_output(row, scenario, stage, corner)
        and not is_apply_approved(row)
        and normalize_key(row.values.get("source_type")) != "na"
    ]
    lines.extend(["", "  Rows not emitted for this output:"])
    if not not_emitted:
        lines.append("    <none>")
    else:
        for row in not_emitted[:100]:
            values = row.values
            reasons = []
            if normalize_key(values.get("apply")) != "yes":
                reasons.append("apply!=yes")
            if normalize_key(values.get("review_status")) != "approved":
                reasons.append("review_status!=approved")
            if normalize_key(values.get("budget_model")) in {"", "unknown", "clock_relative_io_delay"}:
                reasons.append(f"budget_model={normalize_key(values.get('budget_model')) or 'blank'}")
            lines.append(
                f"    row {row.row_idx}: {clean_cell(values.get('channel_id')) or '-'} "
                f"reason={','.join(reasons) or 'not approved'}"
            )
        if len(not_emitted) > 100:
            lines.append(f"    ... truncated {len(not_emitted) - 100} additional row(s)")

    clock_relative = [
        row
        for row in rows
        if row_selected_for_output(row, scenario, stage, corner)
        and normalize_key(row.values.get("budget_model")) == "clock_relative_io_delay"
    ]
    lines.extend(["", "  Clock-relative extracted rows not directly convertible:"])
    if not clock_relative:
        lines.append("    <none>")
    else:
        for row in clock_relative[:100]:
            values = row.values
            lines.append(
                f"    row {row.row_idx}: {clean_cell(values.get('channel_id')) or '-'} "
                f"src_max={clean_cell(values.get('src_output_delay_max')) or '-'} "
                f"dst_max={clean_cell(values.get('dst_input_delay_max')) or '-'}"
            )
        if len(clock_relative) > 100:
            lines.append(f"    ... truncated {len(clock_relative) - 100} additional row(s)")

    diff_rows = []
    for row in rows:
        if not row_selected_for_output(row, scenario, stage, corner):
            continue
        values = row.values
        nums = numeric_parts(values.get("src_output_delay_max")) + numeric_parts(values.get("dst_input_delay_max"))
        if len(nums) >= 2 and max(nums) != min(nums):
            diff_rows.append((row, max(nums) - min(nums)))
    lines.extend(["", "  Channels with differing max candidates:"])
    if not diff_rows:
        lines.append("    <none>")
    else:
        for row, diff in diff_rows[:100]:
            lines.append(
                f"    row {row.row_idx}: {clean_cell(row.values.get('channel_id')) or '-'} diff={diff:.12g}"
            )
        if len(diff_rows) > 100:
            lines.append(f"    ... truncated {len(diff_rows) - 100} additional row(s)")

    rows_by_channel: Dict[str, List[FormRow]] = defaultdict(list)
    for row in rows:
        if row_selected_for_assembled(row, scenario, stage, corner):
            rows_by_channel[clean_cell(row.values.get("channel_id"))].append(row)
    missing_budget_channels = []
    for ch in channels:
        if ch.channel_type not in CHANNEL_TYPES_10:
            continue
        group = rows_by_channel.get(ch.channel_id, [])
        if not group:
            missing_budget_channels.append((ch, "no interface_budget row"))
            continue
        if not any(is_apply_approved(row) for row in group):
            has_candidate = any(
                clean_cell(row.values.get("src_output_delay_max"))
                or clean_cell(row.values.get("dst_input_delay_max"))
                or clean_cell(row.values.get("src_output_delay_min"))
                or clean_cell(row.values.get("dst_input_delay_min"))
                for row in group
            )
            missing_budget_channels.append((ch, "candidate exists but no approved budget" if has_candidate else "no extracted budget candidate"))
    lines.extend(["", "  10 channels without approved budget:"])
    if not missing_budget_channels:
        lines.append("    <none>")
    else:
        for ch, reason in missing_budget_channels[:100]:
            lines.append(f"    {ch.channel_id}: type={ch.channel_type} reason={reason}")
        if len(missing_budget_channels) > 100:
            lines.append(f"    ... truncated {len(missing_budget_channels) - 100} additional channel(s)")
    return lines


def write_report(
    path: Path,
    report: Report,
    scenario: str,
    stage: str,
    corner: str,
    form_path: Path,
    output_path: Path,
    coverage_lines: Sequence[str],
) -> None:
    lines = [
        "10_harden_x_if extraction report",
        "================================",
        "",
        f"Scenario: {scenario}",
        f"Stage   : {stage}",
        f"Corner  : {corner}",
        f"Form    : {form_path}",
        f"Output  : {output_path}",
        f"Warnings: {report.warning_count}",
        f"Errors  : {report.error_count}",
        f"Sync changed: {'yes' if report.sync_changed else 'no'}",
        "",
        "Messages:",
    ]
    lines.extend(report.lines or ["INFO: no messages"])
    lines.extend(coverage_lines)
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate 10 SoC harden/subsys interface budget SDC.")
    parser.add_argument("-scenario", "--scenario", default="common", choices=sorted(SCENARIOS), help="target scenario")
    parser.add_argument("-stage", "--stage", default="all", choices=sorted(STAGES), help="target stage/view")
    parser.add_argument("-corner", "--corner", default="all", help="target corner/view")
    parser.add_argument("-input", "--input", default="../01_soc_clocks/clock_inventory.csv", help="common 01 clock inventory CSV")
    parser.add_argument("--info-all", default="info_all.xlsx", help="integration summary xlsx")
    parser.add_argument("--form", default="10_harden_x_if.xlsx", help="harden interface budget workbook")
    parser.add_argument("--max-diff-threshold", type=float, help="warn when source/destination max budget differs by more than this")
    parser.add_argument("--force-generate-after-sync", action="store_true", help="generate SDC even if workbook was synchronized")
    parser.add_argument("--report", help="output report path")
    return parser.parse_args(argv)


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)
    cwd = Path.cwd()
    report = Report()

    info_all = cwd / args.info_all
    if not info_all.is_file():
        raise RuntimeError(f"integration file not found: {info_all}")

    clocks = read_clock_inventory((cwd / args.input).resolve(), report)
    clock_objects = clock_object_set(clocks)

    instances = read_info_all(info_all, report)
    port_paths = sorted(path for path in cwd.glob("*.xlsx") if path.name not in {Path(args.info_all).name, Path(args.form).name})
    port_sheets = read_port_workbooks(port_paths, report)
    attach_port_data(instances, port_sheets, report)
    resolve_sdc_paths(instances, cwd, report)
    current_digests = collect_current_sdc_digests(instances)

    channels = build_channels(instances, clock_objects, report)
    candidates = extract_delay_candidates(instances, report)
    seeds = create_budget_seeds(channels, candidates, report)

    form_path = cwd / args.form
    sync_workbook(form_path, channels, seeds, report)

    rows = read_form_rows(form_path)
    validate_rows(rows, channels, args.scenario, args.stage, args.corner, current_digests, args.max_diff_threshold, report)
    write_autofilled_fields(form_path, rows, report)

    output_path = output_sdc_path(cwd, args.scenario, args.stage, args.corner)
    rpt_path = Path(args.report) if args.report else report_path(cwd, args.scenario, args.stage, args.corner)

    if report.sync_changed and not args.force_generate_after_sync:
        report.warn("workbook changed during sync; review 10_harden_x_if.xlsx before SDC generation")
    elif report.error_count == 0:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            "\n".join(generate_sdc(rows, args.scenario, args.stage, args.corner)).rstrip() + "\n",
            encoding="utf-8",
        )
        report.info(f"wrote {output_path}")
    else:
        report.warn("SDC generation skipped because errors were reported")

    coverage_lines = build_coverage_lines(rows, channels, args.scenario, args.stage, args.corner)
    write_report(rpt_path, report, args.scenario, args.stage, args.corner, form_path, output_path, coverage_lines)
    print(f"Report: {rpt_path}")
    print(f"Warnings: {report.warning_count}  Errors: {report.error_count}  Sync changed: {report.sync_changed}")
    if report.error_count:
        return 1
    if report.sync_changed and not args.force_generate_after_sync:
        return 1
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(2)
