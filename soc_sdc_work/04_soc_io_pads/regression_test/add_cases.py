#!/usr/bin/env python3
"""Append targeted cases to 04_soc_io_pads.xlsx:
   - Demo A: source-synchronous multi-edge delays on pad_dq0 (single-row min/max + single-row rise/fall)
   - Demo B1: clean view-specific (prects/ss_125) load on pad_ddr_dqs
Also extends clock_inventory.csv with dqs_clk.
"""
import csv
from openpyxl import load_workbook

FORM = "04_soc_io_pads.xlsx"
CSV = "clock_inventory.csv"


def ensure_clock():
    names = set()
    with open(CSV, newline="") as fh:
        rows = list(csv.reader(fh))
    for r in rows[1:]:
        if r:
            names.add(r[0])
    if "dqs_clk" not in names:
        with open(CSV, "a", newline="") as fh:
            csv.writer(fh).writerow(["dqs_clk", "", "", "emit_virtual_clock", "01"])


def append_rows():
    wb = load_workbook(FORM)
    ws = wb["io_constraints"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(hdr)}

    def add(d):
        r = ws.max_row + 1
        for k, v in d.items():
            ws.cell(row=r, column=idx[k], value=v)
        return r

    common = dict(scenario="common", source_type="manual", apply="yes", review_status="approved")
    dq = dict(common, pad_name="pad_dq0", soc_object="[get_ports {pad_dq0}]",
              direction="input", timing_class="timed", constraint_type="input_delay",
              clock_name="dqs_clk")
    # Demo A row 1: single row carrying BOTH min and max -> two commands, second must get -add_delay
    add(dict(dq, stage="all", corner="all", min_value="0.10", max_value="0.80",
             add_delay="no", basis="DDR read min/max (single row)"))
    # Demo A row 2: same group, rise+fall -> all -add_delay (group already emitted)
    add(dict(dq, stage="all", corner="all", rise_value="0.05", fall_value="0.06",
             add_delay="yes", basis="DDR read rise/fall (second row)"))

    # Demo B1: clean view-specific electrical, prects/ss_125 only (no all/all counterpart yet)
    add(dict(common, stage="prects", corner="ss_125", pad_name="pad_ddr_dqs",
             soc_object="[get_ports {pad_ddr_dqs}]", direction="output",
             constraint_type="load", value="0.03", basis="DDR dqs view-specific load"))
    wb.save(FORM)
    print("appended Demo A (2 rows) + Demo B1 (1 row)")


if __name__ == "__main__":
    ensure_clock()
    append_rows()
