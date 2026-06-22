#!/usr/bin/env python3
"""Simulate human review on 04_soc_io_pads.xlsx io_constraints sheet."""
import sys
from openpyxl import load_workbook

FORM = "04_soc_io_pads.xlsx"


def main() -> int:
    wb = load_workbook(FORM)
    ws = wb["io_constraints"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(hdr)}

    def setv(row, col, val):
        ws.cell(row=row, column=idx[col], value=val)

    for r in range(2, ws.max_row + 1):
        ct = ws.cell(row=r, column=idx["constraint_type"]).value
        pad = ws.cell(row=r, column=idx["pad_name"]).value
        if not ct:
            continue
        if pad == "pad_uart0_sin" and ct == "input_delay":
            setv(r, "apply", "yes"); setv(r, "review_status", "approved")
            setv(r, "timing_class", "async"); setv(r, "basis", "UART RX board budget")
        elif pad == "pad_uart0_sin" and ct == "input_transition":
            setv(r, "apply", "yes"); setv(r, "review_status", "approved")
            setv(r, "basis", "IO spec input slew")
        elif pad == "pad_uart0_sin" and ct == "driving_cell":
            # driving_cell vs input_transition: pick input_transition, reject this
            setv(r, "apply", "no"); setv(r, "review_status", "rejected")
            setv(r, "note", "rejected: input_transition used instead")
        elif pad == "pad_uart0_sout" and ct == "output_delay":
            setv(r, "apply", "yes"); setv(r, "review_status", "approved")
            setv(r, "timing_class", "async"); setv(r, "basis", "UART TX board budget")
        elif pad == "pad_uart0_sout" and ct == "load":
            setv(r, "apply", "yes"); setv(r, "review_status", "approved")
            setv(r, "basis", "package + PCB load")
        elif pad == "pad_gpio0" and ct == "input_delay":
            # direction-specific: move to gpio_in scenario
            setv(r, "scenario", "gpio_in")
            setv(r, "apply", "yes"); setv(r, "review_status", "approved")
            setv(r, "timing_class", "timed"); setv(r, "basis", "gpio input mode budget")

    wb.save(FORM)
    print("review applied")
    return 0


if __name__ == "__main__":
    sys.exit(main())
