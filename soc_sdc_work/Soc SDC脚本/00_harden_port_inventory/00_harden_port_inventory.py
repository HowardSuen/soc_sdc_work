#!/usr/bin/env python3
"""Initialize one SoC SDC run root and its in-place port accounting state.

00 validates the integration inputs, applies only explicit structural terminal
dispositions (NC/open/tie-off), publishes the harden SDC manifest and input
snapshot, and starts the accounting digest chain. It does not generate SDC,
connection inventory, pending files, or removed logs.
"""

import argparse
import csv
import datetime
import hashlib
import io
import json
import os
import re
import shutil
import socket
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - deployment guard
    print("ERROR: openpyxl is required to read and update port workbooks.", file=sys.stderr)
    raise SystemExit(2) from exc


SCRIPT_NAME = "00_harden_port_inventory.py"
SCRIPT_VERSION = "2.0.0"
STAGE_NAME = "00_harden_port_inventory"
SCHEMA_VERSION = "1.0"
REQUIRED_PORT_COLUMNS = (
    "Input",
    "Input Width",
    "Input Used Width",
    "From Whom",
    "Output",
    "Output Width",
    "Output Used Width",
    "To Top",
    "Inout",
    "Inout Width",
    "Inout Connectivity",
    "Inout Name",
)
DIRECTION_COLUMNS = {
    "input": ("Input", "Input Width", "Input Used Width", "From Whom"),
    "output": ("Output", "Output Width", "Output Used Width", "To Top"),
    "inout": ("Inout", "Inout Width", "Inout Name", "Inout Connectivity"),
}
DELTA_HEADERS = (
    "schema_version",
    "run_id",
    "mode_label",
    "stage_name",
    "transaction_id",
    "view_id",
    "stage",
    "corner",
    "structure_digest",
    "accounting_digest_before",
    "accounting_digest_after",
    "workbook",
    "sheet",
    "row",
    "direction",
    "port",
    "legal_bits",
    "added_bits",
    "final_used_bits",
    "owner_object_id",
    "reason",
    "evidence_status",
)
MANIFEST_HEADERS = (
    "inst_name",
    "module_name",
    "sdc_path",
    "availability_status",
    "sdc_digest",
    "note",
)
SIGNAL_RE = re.compile(
    r"^(?P<base>[A-Za-z_][A-Za-z0-9_$]*)(?:\[(?P<left>-?\d+)(?::(?P<right>-?\d+))?\])?$"
)
VERILOG_LITERAL_RE = re.compile(
    r"^(?:(?P<width>\d+)\s*)?'(?P<signed>[sS]?)(?P<base>[bBoOdDhH])(?P<digits>[0-9a-fA-F_xXzZ]+)$"
)
INTERMEDIATE_USED_RE = re.compile(r"^-?\d+(?:\s*,\s*-?\d+)*$")
FINAL_USED_RE = re.compile(r"^(?:ALL\s+USED|USED:.*;\s*UNUSED:.*)$", re.IGNORECASE)
STRUCTURAL_NC = {"NC", "N/C", "NO_CONNECT", "UNCONNECTED"}
STRUCTURAL_OPEN = {"OPEN"}
STRUCTURAL_TIE = {"TIE0", "TIE1"}
YES_NO = {"yes", "no"}
STAGE_DELTA_DIRS = ("00_middle", "01_middle", "04_middle", "10_middle", "20_middle", "30_middle")


def _author_part_a():
    return chr(72) + chr(111)


def _author_part_b():
    return chr(119) + chr(97)


def _author_part_c():
    return chr(114) + chr(100)


def author_name():
    return _author_part_a() + _author_part_b() + _author_part_c()


def utc_timestamp():
    return datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ")


def compact_timestamp():
    return datetime.datetime.utcnow().strftime("%Y%m%dT%H%M%S%fZ")


class Report:
    def __init__(self):
        self.lines = []
        self.warning_count = 0
        self.error_count = 0
        self.message_limit = 1500

    def _add(self, level, message):
        if len(self.lines) < self.message_limit:
            self.lines.append("%s: %s" % (level, message))
        elif len(self.lines) == self.message_limit:
            self.lines.append("WARNING: further messages suppressed")

    def info(self, message):
        self._add("INFO", message)

    def warn(self, message):
        self.warning_count += 1
        self._add("WARNING", message)

    def error(self, message):
        self.error_count += 1
        self._add("ERROR", message)


class PortShape:
    def __init__(self, raw, base, width, scalar, left=None, right=None, explicit=False):
        self.raw = raw
        self.base = base
        self.width = width
        self.scalar = scalar
        self.left = left
        self.right = right
        self.explicit = explicit

    @property
    def count(self):
        return 1 if self.scalar else abs(self.left - self.right) + 1

    @property
    def implicit_range(self):
        return not self.scalar and not self.explicit

    def bits(self):
        if self.scalar:
            return (0,)
        low = min(self.left, self.right)
        high = max(self.left, self.right)
        return tuple(range(low, high + 1))

    def contains(self, bit):
        if self.scalar:
            return bit == 0
        return min(self.left, self.right) <= bit <= max(self.left, self.right)

    def descriptor(self):
        if self.scalar:
            return [self.base, "scalar", 0, 0]
        return [self.base, "range", self.left, self.right]


class PortRecord:
    def __init__(
        self,
        model,
        sheet,
        row,
        inst_name,
        direction,
        shape,
        port_col,
        width_col,
        used_col,
        connection_col,
        port_value,
        width_value,
        used_value,
        connection_value,
    ):
        self.model = model
        self.workbook = model.relative_name
        self.sheet = sheet
        self.row = row
        self.inst_name = inst_name
        self.direction = direction
        self.shape = shape
        self.port_col = port_col
        self.width_col = width_col
        self.used_col = used_col
        self.connection_col = connection_col
        self.port_value = port_value
        self.width_value = width_value
        self.connection_value = connection_value
        self.used_value = used_value
        self.used_bits = set()
        self.added_bits = set()
        self.structural_reason = ""
        self.modified = False

    def key(self):
        return (self.workbook, self.sheet, self.row, self.direction, self.shape.raw)

    def location(self):
        return "%s:%s row %d %s %s" % (
            self.workbook,
            self.sheet,
            self.row,
            self.direction,
            self.shape.raw,
        )


class WorkbookModel:
    def __init__(self, path, run_root, workbook, digest):
        self.path = path
        self.relative_name = relative_path(path, run_root)
        self.workbook = workbook
        self.digest_before = digest
        self.digest_after = digest
        self.records = []
        self.modified = False


class AccountingLock:
    def __init__(self, path, report):
        self.path = path
        self.report = report
        self.owned = False

    def __enter__(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "pid": os.getpid(),
            "host": socket.gethostname(),
            "created_at": utc_timestamp(),
            "stage": STAGE_NAME,
        }
        try:
            descriptor = os.open(str(self.path), os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o644)
        except FileExistsError:
            stale = False
            try:
                old = json.loads(self.path.read_text(encoding="utf-8"))
                old_pid = int(old.get("pid", 0))
                old_host = old.get("host", "")
                if old_host == socket.gethostname() and old_pid > 0:
                    try:
                        os.kill(old_pid, 0)
                    except OSError:
                        stale = True
            except Exception:
                stale = False
            if not stale:
                raise RuntimeError("port accounting lock already exists: %s" % self.path)
            self.report.warn("removed stale port accounting lock %s" % self.path)
            self.path.unlink()
            descriptor = os.open(str(self.path), os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o644)
        with os.fdopen(descriptor, "w") as file_obj:
            json.dump(payload, file_obj, sort_keys=True, separators=(",", ":"))
            file_obj.write("\n")
        self.owned = True
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        if self.owned and self.path.exists():
            self.path.unlink()


def clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", clean_cell(value).lower())


def sha256_bytes(data):
    return hashlib.sha256(data).hexdigest()


def sha256_file(path):
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json_bytes(value):
    text = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return text.encode("utf-8")


def canonical_digest(value):
    return sha256_bytes(canonical_json_bytes(value))


def relative_path(path, run_root):
    try:
        return str(path.resolve().relative_to(run_root.resolve()))
    except ValueError:
        return str(path.resolve())


def atomic_write_bytes(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(".%s.tmp.%s" % (path.name, os.getpid()))
    try:
        with tmp.open("wb") as file_obj:
            file_obj.write(data)
            file_obj.flush()
            os.fsync(file_obj.fileno())
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            tmp.unlink()


def atomic_write_text(path, text):
    atomic_write_bytes(path, text.encode("utf-8"))


def atomic_write_json(path, value):
    text = json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n"
    atomic_write_text(path, text)


def csv_text(headers, rows, include_header=True):
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=list(headers), extrasaction="ignore", lineterminator="\n")
    if include_header:
        writer.writeheader()
    for row in rows:
        writer.writerow({header: clean_cell(row.get(header)) for header in headers})
    return buffer.getvalue()


def read_json(path, description):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise RuntimeError("failed to read %s %s: %s" % (description, path, exc))


def normalized_csv_header(fieldnames, path, report):
    mapping = {}
    for name in fieldnames or []:
        key = normalize_header(name)
        if not key:
            continue
        if key in mapping:
            report.error("%s has duplicate normalized column %s" % (path, name))
        else:
            mapping[key] = name
    return mapping


def read_run_context(path, report):
    if not path.is_file():
        report.error("required run context is missing: %s" % path)
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        mapping = normalized_csv_header(reader.fieldnames, path, report)
        for required in ("runid", "modelabel", "designrevision", "note"):
            if required not in mapping:
                report.error("%s is missing required column %s" % (path, required))
        rows = [row for row in reader if any(clean_cell(value) for value in row.values())]
    if len(rows) != 1:
        report.error("%s must contain exactly one non-empty data row" % path)
        return {}
    row = rows[0]
    result = {
        "run_id": clean_cell(row.get(mapping.get("runid"))),
        "mode_label": clean_cell(row.get(mapping.get("modelabel"))),
        "design_revision": clean_cell(row.get(mapping.get("designrevision"))),
        "note": clean_cell(row.get(mapping.get("note"))),
    }
    if not result["run_id"] or not result["mode_label"]:
        report.error("%s requires non-empty run_id and mode_label" % path)
    if result["run_id"] and not re.fullmatch(r"[A-Za-z0-9_.-]+", result["run_id"]):
        report.error("run_id contains unsupported characters: %s" % result["run_id"])
    return result


def read_required_views(path, report):
    if not path.is_file():
        report.error("required view definition is missing: %s" % path)
        return []
    required = (
        "viewid",
        "stage",
        "corner",
        "require02",
        "require04",
        "require20",
        "require30",
        "note",
    )
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        mapping = normalized_csv_header(reader.fieldnames, path, report)
        for name in required:
            if name not in mapping:
                report.error("%s is missing required column %s" % (path, name))
        for row_idx, source in enumerate(reader, start=2):
            if not any(clean_cell(value) for value in source.values()):
                continue
            row = {
                "view_id": clean_cell(source.get(mapping.get("viewid"))),
                "stage": clean_cell(source.get(mapping.get("stage"))),
                "corner": clean_cell(source.get(mapping.get("corner"))),
                "require_02": clean_cell(source.get(mapping.get("require02"))).lower(),
                "require_04": clean_cell(source.get(mapping.get("require04"))).lower(),
                "require_20": clean_cell(source.get(mapping.get("require20"))).lower(),
                "require_30": clean_cell(source.get(mapping.get("require30"))).lower(),
                "note": clean_cell(source.get(mapping.get("note"))),
                "source_row": row_idx,
            }
            if not row["view_id"] or not row["stage"] or not row["corner"]:
                report.error("%s row %d requires view_id, stage and corner" % (path, row_idx))
            for flag in ("require_02", "require_04", "require_20", "require_30"):
                if row[flag] not in YES_NO:
                    report.error("%s row %d has invalid %s=%r; expected yes/no" % (path, row_idx, flag, row[flag]))
            rows.append(row)
    seen_ids = set()
    seen_views = set()
    for row in rows:
        if row["view_id"] in seen_ids:
            report.error("duplicate required view_id %s" % row["view_id"])
        seen_ids.add(row["view_id"])
        key = (row["stage"], row["corner"])
        if key in seen_views:
            report.error("duplicate required stage/corner %s/%s" % key)
        seen_views.add(key)
    return sorted(rows, key=lambda item: (item["view_id"], item["stage"], item["corner"]))


def workbook_header_map(sheet, required, location, report):
    mapping = {}
    for cell in sheet[1]:
        key = normalize_header(cell.value)
        if not key:
            continue
        if key in mapping:
            report.error("%s has duplicate normalized column %s" % (location, clean_cell(cell.value)))
        else:
            mapping[key] = cell.column
    for name in required:
        key = normalize_header(name)
        if key not in mapping:
            report.error("%s is missing required column %s" % (location, name))
    return mapping


def read_info_all(path, report):
    if not path.is_file():
        report.error("required integration workbook is missing: %s" % path)
        return {}, []
    try:
        workbook = load_workbook(str(path), read_only=True, data_only=False)
    except Exception as exc:
        report.error("failed to read %s: %s" % (path, exc))
        return {}, []
    sheet = workbook[workbook.sheetnames[0]]
    mapping = workbook_header_map(sheet, ("module_name", "inst_name", "owner"), str(path), report)
    semantic_rows = []
    instances = {}
    normalized_columns = []
    for cell in sheet[1]:
        key = normalize_header(cell.value)
        if key:
            normalized_columns.append((key, cell.column))
    for row_idx in range(2, sheet.max_row + 1):
        values = {key: clean_cell(sheet.cell(row_idx, column).value) for key, column in normalized_columns}
        if not any(values.values()):
            continue
        inst_name = values.get("instname", "")
        module_name = values.get("modulename", "")
        owner = values.get("owner", "")
        if not inst_name or not module_name or not owner:
            report.error("%s row %d requires module_name, inst_name and owner" % (path, row_idx))
            continue
        if inst_name in instances:
            report.error("%s row %d duplicates inst_name %s" % (path, row_idx, inst_name))
            continue
        entry = {
            "inst_name": inst_name,
            "module_name": module_name,
            "owner": owner,
            "sdc_path": values.get("sdcpath", "") or values.get("sdcfile", "") or values.get("sdc", ""),
            "sdc_status": (values.get("availabilitystatus", "") or values.get("sdcstatus", "")).lower(),
            "sdc_digest": values.get("sdcdigest", "").lower(),
            "sdc_note": values.get("sdcnote", "") or values.get("note", ""),
            "source_row": row_idx,
            "semantic": values,
        }
        instances[inst_name] = entry
        semantic_rows.append(values)
    workbook.close()
    report.info("loaded %d instance(s) from %s" % (len(instances), path.name))
    return instances, sorted(semantic_rows, key=lambda item: canonical_json_bytes(item))


def parse_port_shape(port_value, width_value, location):
    port = clean_cell(port_value)
    width_text = clean_cell(width_value)
    if not width_text or not re.fullmatch(r"\d+", width_text):
        raise ValueError("%s: width must be a positive integer, got %r" % (location, width_text))
    width = int(width_text)
    if width <= 0:
        raise ValueError("%s: width must be positive" % location)
    match = SIGNAL_RE.fullmatch(port)
    if not match:
        raise ValueError("%s: invalid port/range expression %r" % (location, port))
    base = match.group("base")
    left = match.group("left")
    right = match.group("right")
    if left is None:
        if width == 1:
            return PortShape(port, base, width, True)
        return PortShape(port, base, width, False, width - 1, 0, explicit=False)
    left_value = int(left)
    right_value = int(right) if right is not None else left_value
    selector_width = abs(left_value - right_value) + 1
    if right is None and width != 1:
        raise ValueError("%s: explicit bit selector requires width=1" % location)
    if selector_width != width:
        raise ValueError(
            "%s: explicit range has %d bit(s), but width is %d" % (location, selector_width, width)
        )
    return PortShape(port, base, width, False, left_value, right_value, explicit=True)


def looks_like_connection(value):
    text = clean_cell(value)
    return bool(re.fullmatch(r"[A-Za-z_][A-Za-z0-9_$]*\..+", text))


def parse_used_state(value, shape, resume, location, report):
    text = clean_cell(value)
    if not text:
        return set()
    if FINAL_USED_RE.fullmatch(text):
        report.error("%s contains final accounting token from a completed run: %s" % (location, text))
        return set()
    if not resume:
        report.error("%s must be blank in a fresh run, got %r" % (location, text))
        return set()
    if not INTERMEDIATE_USED_RE.fullmatch(text):
        report.error("%s has invalid intermediate Used state %r" % (location, text))
        return set()
    raw_bits = [int(item.strip()) for item in text.split(",")]
    bits = set(raw_bits)
    if len(bits) != len(raw_bits):
        report.warn("%s contains duplicate Used bit(s); semantic state is unchanged" % location)
    for bit in sorted(bits):
        if not shape.contains(bit):
            report.error("%s contains out-of-range Used bit %d" % (location, bit))
    return {bit for bit in bits if shape.contains(bit)}


def discover_port_workbooks(input_root, report):
    paths = sorted(
        path.resolve()
        for path in input_root.glob("port_*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    )
    if not paths:
        report.error("no required inputs/port_*.xlsx workbook was found")
    return paths


def read_port_workbooks(paths, run_root, instances, resume, report):
    models = []
    records = []
    sheet_owner = {}
    seen_bits = {}
    base_directions = {}
    implicit_count = 0
    for path in paths:
        digest = sha256_file(path)
        try:
            workbook = load_workbook(str(path), data_only=False)
        except Exception as exc:
            report.error("failed to open port workbook %s: %s" % (path, exc))
            continue
        model = WorkbookModel(path, run_root, workbook, digest)
        models.append(model)
        for sheet_name in workbook.sheetnames:
            location = "%s:%s" % (model.relative_name, sheet_name)
            if sheet_name in sheet_owner:
                report.error("instance sheet %s appears in both %s and %s" % (sheet_name, sheet_owner[sheet_name], model.relative_name))
                continue
            sheet_owner[sheet_name] = model.relative_name
            if sheet_name not in instances:
                report.error("%s does not map to an info_all inst_name" % location)
                continue
            sheet = workbook[sheet_name]
            mapping = workbook_header_map(sheet, REQUIRED_PORT_COLUMNS, location, report)
            if any(normalize_header(name) not in mapping for name in REQUIRED_PORT_COLUMNS):
                continue
            for row_idx in range(2, sheet.max_row + 1):
                for direction, names in DIRECTION_COLUMNS.items():
                    port_name, width_name, used_name, connection_name = names
                    port_col = mapping[normalize_header(port_name)]
                    width_col = mapping[normalize_header(width_name)]
                    used_col = mapping[normalize_header(used_name)]
                    connection_col = mapping[normalize_header(connection_name)]
                    port_value = sheet.cell(row_idx, port_col).value
                    if not clean_cell(port_value):
                        continue
                    width_value = sheet.cell(row_idx, width_col).value
                    used_value = sheet.cell(row_idx, used_col).value
                    connection_value = sheet.cell(row_idx, connection_col).value
                    row_location = "%s row %d %s" % (location, row_idx, direction)
                    try:
                        shape = parse_port_shape(port_value, width_value, row_location)
                    except ValueError as exc:
                        report.error(str(exc))
                        continue
                    if shape.implicit_range:
                        implicit_count += 1
                        report.info("%s: implicit range interpreted as [%d:0]" % (row_location, shape.width - 1))
                    record = PortRecord(
                        model,
                        sheet_name,
                        row_idx,
                        sheet_name,
                        direction,
                        shape,
                        port_col,
                        width_col,
                        used_col,
                        connection_col,
                        port_value,
                        width_value,
                        used_value,
                        connection_value,
                    )
                    if direction == "inout" and looks_like_connection(used_value):
                        report.error("%s: Inout Name still contains a connection; move it to Inout Connectivity" % record.location())
                    record.used_bits = parse_used_state(
                        used_value, shape, resume, record.location(), report
                    )
                    base_key = (sheet_name, shape.base)
                    old_direction = base_directions.get(base_key)
                    if old_direction and old_direction != direction:
                        report.error("%s: port base %s has conflicting direction %s" % (record.location(), shape.base, old_direction))
                    else:
                        base_directions[base_key] = direction
                    for bit in shape.bits():
                        bit_key = (sheet_name, direction, shape.base, bit)
                        if bit_key in seen_bits:
                            report.error("%s duplicates canonical bit declared at %s" % (record.location(), seen_bits[bit_key]))
                        else:
                            seen_bits[bit_key] = record.location()
                    model.records.append(record)
                    records.append(record)
    for inst_name in sorted(instances):
        if inst_name not in sheet_owner:
            report.error("info_all instance %s has no unique port workbook sheet" % inst_name)
    report.info(
        "loaded %d port row/direction record(s), %d canonical bit(s), %d implicit range(s)"
        % (len(records), sum(record.shape.count for record in records), implicit_count)
    )
    return models, records


def structural_reason(value, shape, location, report):
    text = clean_cell(value)
    upper = text.upper().replace(" ", "")
    if upper in STRUCTURAL_NC:
        return "structural_nc"
    if upper in STRUCTURAL_OPEN:
        return "structural_open"
    if upper in STRUCTURAL_TIE:
        return "structural_tie_off"
    literal = VERILOG_LITERAL_RE.fullmatch(text.replace(" ", ""))
    if literal:
        literal_width = int(literal.group("width")) if literal.group("width") else None
        if literal_width is not None and literal_width != shape.count:
            report.error(
                "%s: sized Verilog literal width %d does not match port width %d"
                % (location, literal_width, shape.count)
            )
            return ""
        return "structural_tie_off"
    if re.fullmatch(r"\d+", text):
        return "structural_tie_off"
    return ""


def parse_endpoint_signal(value, location):
    match = SIGNAL_RE.fullmatch(clean_cell(value))
    if not match:
        raise ValueError("%s: invalid endpoint port/range %r" % (location, clean_cell(value)))
    base = match.group("base")
    left = match.group("left")
    right = match.group("right")
    if left is None:
        return base, None
    left_value = int(left)
    right_value = int(right) if right is not None else left_value
    low = min(left_value, right_value)
    high = max(left_value, right_value)
    return base, set(range(low, high + 1))


def validate_connections(records, instances, report):
    index = {}
    directions = {}
    for record in records:
        key = (record.inst_name, record.shape.base)
        directions[key] = record.direction
        index.setdefault(key, {})
        for bit in record.shape.bits():
            index[key][bit] = record

    for record in records:
        connection = clean_cell(record.connection_value)
        if not connection:
            continue
        if structural_reason(connection, record.shape, record.location(), report):
            continue
        if record.direction == "output":
            if connection.lower() in ("y", "yes", "true"):
                report.warn("%s: legacy To Top token %r interpreted as top.%s" % (record.location(), connection, record.shape.raw))
                continue
            endpoint = connection[4:] if connection.startswith("top.") else connection
            if "." in endpoint:
                report.error("%s: To Top must name a top port, not %r" % (record.location(), connection))
                continue
            try:
                _, bits = parse_endpoint_signal(endpoint, record.location())
            except ValueError as exc:
                report.error(str(exc))
                continue
            if bits is not None and len(bits) != record.shape.count:
                report.error("%s: To Top range width does not match output width" % record.location())
            continue

        if "." not in connection:
            report.error("%s: connection must be <inst>.<port/range> or an explicit structural token" % record.location())
            continue
        peer_inst, peer_signal = connection.split(".", 1)
        try:
            peer_base, peer_bits = parse_endpoint_signal(peer_signal, record.location())
        except ValueError as exc:
            report.error(str(exc))
            continue
        if peer_inst == "top":
            if peer_bits is not None and len(peer_bits) != record.shape.count:
                report.error("%s: top endpoint range width does not match destination width" % record.location())
            continue
        if peer_inst not in instances:
            report.warn("%s: connection references unknown instance %s" % (record.location(), peer_inst))
            continue
        peer_key = (peer_inst, peer_base)
        if peer_key not in index:
            report.warn("%s: connection references unknown port %s.%s" % (record.location(), peer_inst, peer_base))
            continue
        peer_direction = directions.get(peer_key)
        expected = {"inout"} if record.direction == "inout" else {"output", "inout"}
        if peer_direction not in expected:
            report.error("%s: driver %s.%s has illegal direction %s" % (record.location(), peer_inst, peer_base, peer_direction))
        selected = set(index[peer_key]) if peer_bits is None else peer_bits
        missing = sorted(selected - set(index[peer_key]))
        if missing:
            report.error("%s: connection selects undeclared bit(s) %s.%s[%s]" % (record.location(), peer_inst, peer_base, ",".join(str(bit) for bit in missing)))
            continue
        if len(selected) != record.shape.count:
            report.error("%s: source width %d does not match destination width %d" % (record.location(), len(selected), record.shape.count))


def format_bits(bits):
    return ",".join(str(bit) for bit in sorted(bits))


def accounting_digest(records, empty=False):
    rows = []
    for record in records:
        bits = [] if empty else sorted(record.used_bits)
        rows.append([
            record.workbook,
            record.sheet,
            record.row,
            record.direction,
            record.shape.raw,
            bits,
        ])
    rows.sort(key=lambda item: canonical_json_bytes(item))
    return canonical_digest([SCHEMA_VERSION, rows])


def structure_digest(run_context, required_views, info_semantic, records):
    port_rows = []
    for record in records:
        port_rows.append([
            record.workbook,
            record.sheet,
            record.inst_name,
            record.direction,
            record.shape.raw,
            record.shape.descriptor(),
            clean_cell(record.connection_value),
        ])
    port_rows.sort(key=lambda item: canonical_json_bytes(item))
    view_rows = [
        {key: value for key, value in row.items() if key != "source_row"}
        for row in required_views
    ]
    return canonical_digest([
        SCHEMA_VERSION,
        run_context,
        view_rows,
        info_semantic,
        port_rows,
    ])


def structural_id(record, bit, reason):
    payload = [SCHEMA_VERSION, record.inst_name, record.direction, record.shape.base, int(bit), reason]
    return "STRUCT_" + sha256_bytes(canonical_json_bytes(payload))


def apply_structural_defaults(records, report):
    terminals = []
    for record in records:
        reason = structural_reason(record.connection_value, record.shape, record.location(), report)
        if not reason:
            continue
        legal = set(record.shape.bits())
        added = legal - record.used_bits
        record.structural_reason = reason
        record.added_bits = added
        record.used_bits |= legal
        if added:
            text = format_bits(record.used_bits)
            if len(text) > 32767:
                report.error("%s: Used state exceeds the Excel cell text limit" % record.location())
                continue
            sheet = record.model.workbook[record.sheet]
            cell = sheet.cell(record.row, record.used_col)
            cell.value = text
            cell.number_format = "@"
            record.modified = True
            record.model.modified = True
        terminals.append(record)
    report.info(
        "structural terminals=%d added_bits=%d"
        % (len(terminals), sum(len(record.added_bits) for record in terminals))
    )
    return terminals


def discover_sdc_files(input_root):
    return sorted(path.resolve() for path in input_root.glob("*.sdc") if path.is_file())


def manifest_path_value(path, run_root):
    return relative_path(path, run_root)


def choose_sdc(instance, input_root, run_root, sdc_files, report):
    status_hint = instance.get("sdc_status", "")
    if status_hint and status_hint not in ("available", "missing", "not_required"):
        report.error("info_all row %d has invalid SDC status %r" % (instance["source_row"], status_hint))
    if status_hint == "not_required":
        if not instance.get("sdc_note"):
            report.error("%s is not_required but has no explicit note/basis" % instance["inst_name"])
        return "", "not_required", "", instance.get("sdc_note")

    hint = instance.get("sdc_path", "")
    selected = None
    note = instance.get("sdc_note", "")
    if hint:
        selected = Path(hint).expanduser()
        if not selected.is_absolute():
            selected = input_root / selected
        selected = selected.resolve()
        if not selected.is_file():
            if status_hint == "available":
                report.error("mapped available SDC is missing for %s: %s" % (instance["inst_name"], selected))
            return manifest_path_value(selected, run_root), "missing", "", note or "mapped SDC not delivered"
    else:
        by_inst = [path for path in sdc_files if path.stem == instance["inst_name"]]
        by_module = [path for path in sdc_files if path.stem == instance["module_name"]]
        candidates = by_inst if by_inst else by_module
        if len(candidates) > 1:
            report.error("multiple SDC candidates for %s: %s" % (instance["inst_name"], ", ".join(str(path) for path in candidates)))
            return "", "missing", "", "conflicting exact SDC candidates"
        if candidates:
            selected = candidates[0]
            note = note or ("matched_by=inst_name" if by_inst else "matched_by=module_name")
        else:
            return "", "missing", "", note or "no exact inst_name/module_name SDC match"
    try:
        digest = sha256_file(selected)
    except OSError as exc:
        report.error("SDC is not readable for %s: %s" % (instance["inst_name"], exc))
        return manifest_path_value(selected, run_root), "missing", "", "SDC unreadable"
    expected = instance.get("sdc_digest", "")
    if expected and expected != digest:
        report.error("SDC digest mismatch for %s: expected %s, got %s" % (instance["inst_name"], expected, digest))
    return manifest_path_value(selected, run_root), "available", digest, note


def build_manifest(instances, input_root, run_root, report):
    sdc_files = discover_sdc_files(input_root)
    rows = []
    counts = {"available": 0, "missing": 0, "not_required": 0}
    for inst_name in sorted(instances):
        instance = instances[inst_name]
        path, status, digest, note = choose_sdc(instance, input_root, run_root, sdc_files, report)
        counts[status] += 1
        if status == "missing":
            report.warn("harden SDC missing for %s" % inst_name)
        rows.append({
            "inst_name": inst_name,
            "module_name": instance["module_name"],
            "sdc_path": path,
            "availability_status": status,
            "sdc_digest": digest,
            "note": note,
        })
    return rows, counts


def record_index(records):
    return {record.key(): record for record in records}


def parse_bits_field(value):
    text = clean_cell(value)
    if not text:
        return set()
    if not INTERMEDIATE_USED_RE.fullmatch(text):
        raise ValueError("invalid bit list %r" % text)
    return {int(item.strip()) for item in text.split(",")}


def load_delta_rows(path, report):
    if not path.is_file():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        missing = [name for name in DELTA_HEADERS if name not in (reader.fieldnames or [])]
        if missing:
            report.error("%s is missing delta column(s): %s" % (path, ", ".join(missing)))
        return [dict(row) for row in reader]


def delta_rows_digest(rows):
    normalized = [[clean_cell(row.get(name)) for name in DELTA_HEADERS] for row in rows]
    return canonical_digest([SCHEMA_VERSION, normalized])


def validate_resume_evidence(run_root, records, run_context, structure, current_digest, empty_digest, report):
    index = record_index(records)
    explained = {key: set() for key in index}
    transactions = []
    any_delta = False
    for directory in STAGE_DELTA_DIRS:
        delta_path = run_root / directory / "port_accounting_delta.csv"
        meta_path = run_root / directory / "port_accounting_delta.meta"
        if delta_path.exists() != meta_path.exists():
            report.error("resume requires paired delta/meta: %s and %s" % (delta_path, meta_path))
            continue
        if not delta_path.exists():
            continue
        any_delta = True
        try:
            meta = read_json(meta_path, "accounting delta meta")
        except RuntimeError as exc:
            report.error(str(exc))
            continue
        if meta.get("run_id") != run_context.get("run_id") or meta.get("mode_label") != run_context.get("mode_label"):
            report.error("%s run provenance does not match run_context.csv" % meta_path)
        if meta.get("structure_digest") != structure:
            report.error("%s structure_digest is stale" % meta_path)
        if meta.get("delta_csv_digest") != sha256_file(delta_path):
            report.error("%s delta_csv_digest does not match %s" % (meta_path, delta_path))
        rows = load_delta_rows(delta_path, report)
        grouped = {}
        for row in rows:
            if clean_cell(row.get("run_id")) != run_context.get("run_id"):
                report.error("%s contains a delta row with mismatched run_id" % delta_path)
            if clean_cell(row.get("mode_label")) != run_context.get("mode_label"):
                report.error("%s contains a delta row with mismatched mode_label" % delta_path)
            if clean_cell(row.get("structure_digest")) != structure:
                report.error("%s contains a delta row with stale structure_digest" % delta_path)
            if not clean_cell(row.get("owner_object_id")) or not clean_cell(row.get("evidence_status")):
                report.error("%s contains a delta row without owner/evidence" % delta_path)
            key = (
                clean_cell(row.get("workbook")),
                clean_cell(row.get("sheet")),
                int(clean_cell(row.get("row")) or 0),
                clean_cell(row.get("direction")),
                clean_cell(row.get("port")),
            )
            if key not in index:
                report.error("%s contains delta for unknown port row %r" % (delta_path, key))
                continue
            try:
                added = parse_bits_field(row.get("added_bits"))
            except ValueError as exc:
                report.error("%s: %s" % (delta_path, exc))
                continue
            for bit in added:
                if not index[key].shape.contains(bit):
                    report.error("%s contains out-of-range delta bit %d for %r" % (delta_path, bit, key))
            explained[key] |= added
            txn_id = clean_cell(row.get("transaction_id"))
            grouped.setdefault(txn_id, []).append(row)
        meta_transaction_ids = set()
        for txn in meta.get("transactions", []):
            txn_id = clean_cell(txn.get("transaction_id"))
            meta_transaction_ids.add(txn_id)
            if txn.get("structure_digest") != structure:
                report.error("transaction %s has stale structure_digest" % txn_id)
            if txn.get("delta_rows_digest") != delta_rows_digest(grouped.get(txn_id, [])):
                report.error("transaction %s delta row digest mismatch" % txn_id)
            transactions.append(txn)
        if not set(grouped).issubset(meta_transaction_ids):
            report.error("%s contains transaction IDs absent from %s" % (delta_path, meta_path))
    current_state = {key: set(record.used_bits) for key, record in index.items()}
    if any(current_state.values()) and not any_delta:
        report.error("resume has pre-existing Used bits but no committed accounting delta evidence")
    for key in current_state:
        if explained[key] != current_state[key]:
            report.error("resume delta union does not explain Used state for %r" % (key,))
    if transactions:
        transactions.sort(key=lambda item: (clean_cell(item.get("committed_at")), clean_cell(item.get("transaction_id"))))
        expected_before = empty_digest
        seen = set()
        for txn in transactions:
            txn_id = clean_cell(txn.get("transaction_id"))
            if not txn_id or txn_id in seen:
                report.error("duplicate/empty accounting transaction id %r" % txn_id)
            seen.add(txn_id)
            if txn.get("accounting_digest_before") != expected_before:
                report.error("accounting digest chain breaks before transaction %s" % txn_id)
            expected_before = txn.get("accounting_digest_after")
        if expected_before != current_digest:
            report.error("accounting digest chain does not end at current workbook state")
    elif any_delta:
        report.error("resume delta meta contains no transaction chain")


def transaction_manifest_path(transaction_dir):
    return transaction_dir / "transaction.json"


def write_transaction_manifest(transaction_dir, manifest):
    atomic_write_json(transaction_manifest_path(transaction_dir), manifest)


def within_root(path, root):
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def restore_transaction(run_root, transaction_dir, manifest, report):
    for item in manifest.get("workbooks", []):
        target = Path(item["target"])
        backup = Path(item["original"])
        if not backup.is_file():
            raise RuntimeError("transaction recovery backup is missing: %s" % backup)
        if sha256_file(backup) != item.get("digest_before"):
            raise RuntimeError("transaction recovery backup digest mismatch: %s" % backup)
        tmp = target.with_name(".%s.restore.%s" % (target.name, os.getpid()))
        shutil.copy2(str(backup), str(tmp))
        os.replace(str(tmp), str(target))
        if sha256_file(target) != item.get("digest_before"):
            raise RuntimeError("transaction recovery failed digest verification: %s" % target)
    for item in manifest.get("artifacts", []):
        path = Path(item.get("path", ""))
        if not path or not within_root(path, run_root):
            raise RuntimeError("transaction recovery artifact path escapes run root: %s" % path)
        if item.get("existed_before"):
            backup = Path(item.get("backup", ""))
            if not backup.is_file() or sha256_file(backup) != item.get("digest_before"):
                raise RuntimeError("transaction recovery artifact backup mismatch: %s" % backup)
            tmp = path.with_name(".%s.restore.%s" % (path.name, os.getpid()))
            path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(backup), str(tmp))
            os.replace(str(tmp), str(path))
        elif path.exists():
            path.unlink()
    manifest["status"] = "ROLLED_BACK"
    manifest["rolled_back_at"] = utc_timestamp()
    write_transaction_manifest(transaction_dir, manifest)
    atomic_write_text(transaction_dir / "ROLLED_BACK", manifest["rolled_back_at"] + "\n")
    for name in ("original", "candidate", "artifact_original"):
        path = transaction_dir / name
        if path.exists():
            shutil.rmtree(str(path))
    report.warn("rolled back incomplete accounting transaction %s" % manifest.get("transaction_id"))


def recover_transactions(run_root, report):
    root = run_root / ".accounting_txn"
    if not root.is_dir():
        return
    for transaction_dir in sorted(path for path in root.iterdir() if path.is_dir()):
        manifest_path = transaction_manifest_path(transaction_dir)
        if not manifest_path.is_file():
            raise RuntimeError("accounting transaction has no transaction.json: %s" % transaction_dir)
        manifest = read_json(manifest_path, "transaction manifest")
        status = manifest.get("status")
        committed_marker = transaction_dir / "COMMITTED"
        if status in ("PREPARED", "APPLYING") and not committed_marker.exists():
            restore_transaction(run_root, transaction_dir, manifest, report)
            continue
        if status == "COMMITTED" or committed_marker.exists():
            for item in manifest.get("workbooks", []):
                target = Path(item["target"])
                if not target.is_file() or sha256_file(target) != item.get("digest_after"):
                    raise RuntimeError("committed transaction target digest mismatch: %s" % target)
            for artifact in manifest.get("artifacts", []):
                path = Path(artifact.get("path", ""))
                digest = artifact.get("digest_after", "")
                if digest and (not path.is_file() or sha256_file(path) != digest):
                    raise RuntimeError("committed transaction artifact digest mismatch: %s" % path)
            shutil.rmtree(str(transaction_dir))
            continue
        if status == "ROLLED_BACK":
            continue
        raise RuntimeError("unsupported accounting transaction status %r in %s" % (status, transaction_dir))


def cleanup_stale_previews(middle_root, report):
    if not middle_root.is_dir():
        return
    for path in middle_root.glob(".00_preview_*"):
        if path.is_dir():
            shutil.rmtree(str(path))
            report.warn("removed stale 00 preview directory %s" % path)
        elif path.exists():
            path.unlink()
            report.warn("removed stale 00 preview file %s" % path)


def make_transaction_id(run_id):
    seed = "%s|%s|%s|%s" % (run_id, compact_timestamp(), os.getpid(), socket.gethostname())
    return "00_%s_%s" % (compact_timestamp(), hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12])


def candidate_cell_validation(model, candidate_path, report):
    try:
        candidate = load_workbook(str(candidate_path), data_only=False)
    except Exception as exc:
        report.error("failed to reopen candidate workbook %s: %s" % (candidate_path, exc))
        return
    if candidate.sheetnames != model.workbook.sheetnames:
        report.error("candidate workbook changed sheet order: %s" % model.relative_name)
    for record in model.records:
        sheet = candidate[record.sheet]
        if sheet.cell(record.row, record.port_col).value != record.port_value:
            report.error("candidate changed non-accounting port cell for %s" % record.location())
        if sheet.cell(record.row, record.width_col).value != record.width_value:
            report.error("candidate changed non-accounting width cell for %s" % record.location())
        if sheet.cell(record.row, record.connection_col).value != record.connection_value:
            report.error("candidate changed connectivity cell for %s" % record.location())
        used_cell = sheet.cell(record.row, record.used_col)
        if record.modified:
            if clean_cell(used_cell.value) != format_bits(record.used_bits):
                report.error("candidate Used state mismatch for %s" % record.location())
            if used_cell.number_format != "@":
                report.error("candidate Used state is not stored as text for %s" % record.location())
        elif used_cell.value != record.used_value:
            report.error("candidate changed an unrelated Used state for %s" % record.location())
    candidate.close()


def build_delta_rows(old_rows, terminals, run_context, transaction_id, structure, before, after):
    new_rows = []
    for record in terminals:
        legal_text = format_bits(set(record.shape.bits()))
        final_text = format_bits(record.used_bits)
        for bit in sorted(record.shape.bits()):
            new_rows.append({
                "schema_version": SCHEMA_VERSION,
                "run_id": run_context["run_id"],
                "mode_label": run_context["mode_label"],
                "stage_name": STAGE_NAME,
                "transaction_id": transaction_id,
                "view_id": "",
                "stage": "",
                "corner": "",
                "structure_digest": structure,
                "accounting_digest_before": before,
                "accounting_digest_after": after,
                "workbook": record.workbook,
                "sheet": record.sheet,
                "row": record.row,
                "direction": record.direction,
                "port": record.shape.raw,
                "legal_bits": legal_text,
                "added_bits": str(bit) if bit in record.added_bits else "",
                "final_used_bits": final_text,
                "owner_object_id": structural_id(record, bit, record.structural_reason),
                "reason": record.structural_reason,
                "evidence_status": "approved",
            })
    return old_rows + new_rows, new_rows


def render_environment_report(
    run_root,
    run_context,
    required_views,
    instances,
    records,
    manifest_rows,
    manifest_counts,
    structure,
    before,
    after,
    workbook_before,
    workbook_after,
    transaction_id,
    mode,
    report,
):
    terminals = [record for record in records if record.structural_reason]
    lines = [
        "SoC SDC 00 Environment Report",
        "Author: %s" % author_name(),
        "Stage: %s" % STAGE_NAME,
        "Script: %s" % SCRIPT_NAME,
        "Script version: %s" % SCRIPT_VERSION,
        "Run root: %s" % run_root,
        "Run ID: %s" % run_context.get("run_id", ""),
        "Mode label: %s" % run_context.get("mode_label", ""),
        "Design revision: %s" % run_context.get("design_revision", ""),
        "Initialization mode: %s" % mode,
        "Port accounting: enabled; structural defaults only",
        "Transaction ID: %s" % transaction_id,
        "",
        "Summary",
        "Instances: %d" % len(instances),
        "Instance sheets: %d" % len({record.sheet for record in records}),
        "Port row/direction records: %d" % len(records),
        "Canonical port bits: %d" % sum(record.shape.count for record in records),
        "Required views: %d" % len(required_views),
        "Available harden SDC: %d" % manifest_counts.get("available", 0),
        "Missing harden SDC: %d" % manifest_counts.get("missing", 0),
        "Not-required harden SDC: %d" % manifest_counts.get("not_required", 0),
        "Structural terminal rows: %d" % len(terminals),
        "Structural added bits: %d" % sum(len(record.added_bits) for record in terminals),
        "Warnings: %d" % report.warning_count,
        "Errors: %d" % report.error_count,
        "",
        "Required Views",
    ]
    for item in required_views:
        lines.append(
            "%s stage=%s corner=%s require_02=%s require_04=%s require_20=%s require_30=%s"
            % (
                item["view_id"],
                item["stage"],
                item["corner"],
                item["require_02"],
                item["require_04"],
                item["require_20"],
                item["require_30"],
            )
        )
    lines.extend(["", "Harden SDC Manifest"])
    for item in manifest_rows:
        lines.append(
            "%s module=%s status=%s path=%s digest=%s note=%s"
            % (
                item["inst_name"],
                item["module_name"],
                item["availability_status"],
                item["sdc_path"],
                item["sdc_digest"],
                item["note"],
            )
        )
    lines.extend(["", "Structural Accounting"])
    for record in terminals:
        lines.append(
            "%s:%s:%d %s %s legal=%s added=%s final=%s reason=%s"
            % (
                record.workbook,
                record.sheet,
                record.row,
                record.direction,
                record.shape.raw,
                format_bits(set(record.shape.bits())),
                format_bits(record.added_bits),
                format_bits(record.used_bits),
                record.structural_reason,
            )
        )
    lines.extend([
        "",
        "Digests",
        "structure_digest=%s" % structure,
        "accounting_digest_before=%s" % before,
        "accounting_digest_after=%s" % after,
    ])
    for name in sorted(workbook_before):
        lines.append(
            "workbook=%s before=%s after=%s"
            % (name, workbook_before[name], workbook_after.get(name, ""))
        )
    lines.extend(["", "Messages"])
    lines.extend(report.lines or ["INFO: no messages"])
    lines.append("")
    return "\n".join(lines) + "\n"


def execute_transaction(
    run_root,
    models,
    prepared_candidates,
    artifact_payloads,
    transaction_id,
    run_context,
    structure,
    before,
    after,
    report,
):
    transaction_dir = run_root / ".accounting_txn" / (STAGE_NAME + "_" + transaction_id)
    original_dir = transaction_dir / "original"
    candidate_dir = transaction_dir / "candidate"
    artifact_original_dir = transaction_dir / "artifact_original"
    original_dir.mkdir(parents=True, exist_ok=False)
    candidate_dir.mkdir(parents=True, exist_ok=False)
    artifact_original_dir.mkdir(parents=True, exist_ok=False)
    manifest = {
        "schema_version": SCHEMA_VERSION,
        "stage_name": STAGE_NAME,
        "transaction_id": transaction_id,
        "run_id": run_context["run_id"],
        "mode_label": run_context["mode_label"],
        "structure_digest": structure,
        "accounting_digest_before": before,
        "accounting_digest_after": after,
        "created_at": utc_timestamp(),
        "status": "BUILDING",
        "workbooks": [],
        "artifacts": [],
    }
    try:
        for model in models:
            if not model.modified:
                continue
            original = original_dir / model.path.name
            candidate = candidate_dir / model.path.name
            shutil.copy2(str(model.path), str(original))
            prepared = prepared_candidates.get(model.relative_name)
            if prepared is None or not prepared.is_file():
                raise RuntimeError("prepared candidate is missing for %s" % model.relative_name)
            shutil.copy2(str(prepared), str(candidate))
            candidate_cell_validation(model, candidate, report)
            model.digest_after = sha256_file(candidate)
            manifest["workbooks"].append({
                "target": str(model.path),
                "original": str(original),
                "candidate": str(candidate),
                "digest_before": model.digest_before,
                "digest_after": model.digest_after,
                "applied": False,
            })
        for index, (path, _) in enumerate(artifact_payloads):
            existed = path.is_file()
            backup = artifact_original_dir / ("%02d_%s" % (index, path.name))
            digest_before = ""
            if existed:
                shutil.copy2(str(path), str(backup))
                digest_before = sha256_file(backup)
            manifest["artifacts"].append({
                "path": str(path),
                "existed_before": existed,
                "backup": str(backup) if existed else "",
                "digest_before": digest_before,
                "digest_after": "",
                "published": False,
            })
        if report.error_count:
            raise RuntimeError("candidate workbook validation failed")
        manifest["status"] = "PREPARED"
        write_transaction_manifest(transaction_dir, manifest)
        manifest["status"] = "APPLYING"
        write_transaction_manifest(transaction_dir, manifest)
        for item in manifest["workbooks"]:
            target = Path(item["target"])
            if sha256_file(target) != item["digest_before"]:
                raise RuntimeError("concurrent workbook modification detected: %s" % target)
            os.replace(item["candidate"], item["target"])
            if sha256_file(target) != item["digest_after"]:
                raise RuntimeError("applied workbook digest mismatch: %s" % target)
            item["applied"] = True
            write_transaction_manifest(transaction_dir, manifest)
        for index, (path, payload) in enumerate(artifact_payloads):
            atomic_write_bytes(path, payload)
            manifest["artifacts"][index]["digest_after"] = sha256_file(path)
            manifest["artifacts"][index]["published"] = True
            write_transaction_manifest(transaction_dir, manifest)
        committed_at = utc_timestamp()
        atomic_write_text(transaction_dir / "COMMITTED", committed_at + "\n")
        manifest["status"] = "COMMITTED"
        manifest["committed_at"] = committed_at
        write_transaction_manifest(transaction_dir, manifest)
        shutil.rmtree(str(transaction_dir))
        return committed_at
    except Exception:
        if manifest.get("status") in ("PREPARED", "APPLYING"):
            restore_transaction(run_root, transaction_dir, manifest, report)
        elif transaction_dir.exists():
            shutil.rmtree(str(transaction_dir))
        raise


def parse_args(argv):
    parser = argparse.ArgumentParser(description="Initialize one SoC SDC run root and port accounting chain.")
    parser.add_argument("--run-root", required=True, help="single-scenario target runtime root")
    parser.add_argument(
        "--resume-accounting",
        action="store_true",
        help="validate and resume the same run's committed accounting chain",
    )
    parser.add_argument(
        "--require-complete-harden-sdc",
        action="store_true",
        help="treat any missing required harden SDC as an error",
    )
    parser.add_argument("--debug", action="store_true", help="print detailed validation messages")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(sys.argv[1:] if argv is None else argv)
    print("Author: %s" % author_name())
    report = Report()
    run_root = Path(args.run_root).expanduser().resolve()
    input_root = run_root / "inputs"
    middle_root = run_root / "00_middle"
    report_path = run_root / "00_result" / "reports" / "environment_report.txt"
    run_context_path = input_root / "run_context.csv"
    required_views_path = input_root / "required_views.csv"
    info_path = input_root / "info_all.xlsx"
    manifest_path = middle_root / "harden_sdc_manifest.csv"
    snapshot_path = middle_root / "input_snapshot.meta"
    delta_path = middle_root / "port_accounting_delta.csv"
    delta_meta_path = middle_root / "port_accounting_delta.meta"
    completion_path = middle_root / "stage_completion.meta"
    lock_path = input_root / ".port_accounting.lock"
    mode = "resume" if args.resume_accounting else "fresh"
    run_context = {}
    required_views = []
    instances = {}
    info_semantic = []
    models = []
    records = []
    manifest_rows = []
    manifest_counts = {"available": 0, "missing": 0, "not_required": 0}
    structure = ""
    before = ""
    after = ""
    transaction_id = ""
    workbook_before = {}
    workbook_after = {}
    preview_dir = None
    success = False
    try:
        if not input_root.is_dir():
            report.error("required inputs directory is missing: %s" % input_root)
        else:
            with AccountingLock(lock_path, report):
                recover_transactions(run_root, report)
                cleanup_stale_previews(middle_root, report)
                run_context = read_run_context(run_context_path, report)
                required_views = read_required_views(required_views_path, report)
                instances, info_semantic = read_info_all(info_path, report)
                port_paths = discover_port_workbooks(input_root, report)
                models, records = read_port_workbooks(
                    port_paths, run_root, instances, args.resume_accounting, report
                )
                validate_connections(records, instances, report)
                structure = structure_digest(run_context, required_views, info_semantic, records)
                before = accounting_digest(records)
                empty_digest = accounting_digest(records, empty=True)
                workbook_before = {model.relative_name: model.digest_before for model in models}
                if not args.resume_accounting:
                    stale = [path for path in (delta_path, delta_meta_path, completion_path) if path.exists()]
                    if stale:
                        report.error("fresh run found existing accounting artifact(s): %s" % ", ".join(str(path) for path in stale))
                else:
                    validate_resume_evidence(
                        run_root,
                        records,
                        run_context,
                        structure,
                        before,
                        empty_digest,
                        report,
                    )
                manifest_rows, manifest_counts = build_manifest(instances, input_root, run_root, report)
                if args.require_complete_harden_sdc and manifest_counts.get("missing", 0):
                    report.error(
                        "--require-complete-harden-sdc: %d harden SDC file(s) are missing"
                        % manifest_counts["missing"]
                    )
                terminals = apply_structural_defaults(records, report)
                after = accounting_digest(records)
                if report.error_count:
                    raise RuntimeError("input/environment validation failed")

                transaction_id = make_transaction_id(run_context["run_id"])
                old_delta_rows = load_delta_rows(delta_path, report) if args.resume_accounting else []
                old_delta_text = delta_path.read_text(encoding="utf-8") if args.resume_accounting and delta_path.is_file() else ""
                old_meta = read_json(delta_meta_path, "00 delta meta") if args.resume_accounting and delta_meta_path.is_file() else {}
                all_delta_rows, new_delta_rows = build_delta_rows(
                    old_delta_rows,
                    terminals,
                    run_context,
                    transaction_id,
                    structure,
                    before,
                    after,
                )
                if old_delta_text:
                    delta_payload_text = old_delta_text
                    if not delta_payload_text.endswith("\n"):
                        delta_payload_text += "\n"
                    delta_payload_text += csv_text(DELTA_HEADERS, new_delta_rows, include_header=False)
                else:
                    delta_payload_text = csv_text(DELTA_HEADERS, all_delta_rows)
                delta_payload = delta_payload_text.encode("utf-8")
                transaction_entry = {
                    "transaction_id": transaction_id,
                    "committed_at": utc_timestamp(),
                    "structure_digest": structure,
                    "accounting_digest_before": before,
                    "accounting_digest_after": after,
                    "delta_rows_digest": delta_rows_digest(new_delta_rows),
                }
                transactions = list(old_meta.get("transactions", [])) + [transaction_entry]
                delta_meta = {
                    "schema_version": SCHEMA_VERSION,
                    "run_id": run_context["run_id"],
                    "mode_label": run_context["mode_label"],
                    "stage_name": STAGE_NAME,
                    "completion_status": "complete",
                    "structure_digest": structure,
                    "accounting_digest_before": before,
                    "accounting_digest_after": after,
                    "delta_csv_digest": sha256_bytes(delta_payload),
                    "transactions": transactions,
                }
                for model in models:
                    if not model.modified:
                        model.digest_after = model.digest_before
                # Modified workbook digests are filled after candidate save. Build static artifacts later.
                manifest_payload = csv_text(MANIFEST_HEADERS, manifest_rows).encode("utf-8")
                middle_root.mkdir(parents=True, exist_ok=True)

                # Save candidates once to obtain exact after-file digests, then the transaction routine
                # performs the authoritative save/validation. The workbook objects are deterministic.
                preview_dir = middle_root / (".00_preview_%s" % os.getpid())
                preview_dir.mkdir(parents=True, exist_ok=False)
                prepared_candidates = {}
                for model in models:
                    if model.modified:
                        preview = preview_dir / model.path.name
                        model.workbook.save(str(preview))
                        model.digest_after = sha256_file(preview)
                        prepared_candidates[model.relative_name] = preview
                workbook_after = {model.relative_name: model.digest_after for model in models}

                sdc_snapshot = [
                    {
                        "inst_name": row["inst_name"],
                        "module_name": row["module_name"],
                        "availability_status": row["availability_status"],
                        "sdc_path": row["sdc_path"],
                        "sdc_digest": row["sdc_digest"],
                    }
                    for row in manifest_rows
                ]
                snapshot = {
                    "schema_version": SCHEMA_VERSION,
                    "author": author_name(),
                    "stage_name": STAGE_NAME,
                    "script_version": SCRIPT_VERSION,
                    "timestamp": utc_timestamp(),
                    "run_root": str(run_root),
                    "run_id": run_context["run_id"],
                    "mode_label": run_context["mode_label"],
                    "design_revision": run_context.get("design_revision", ""),
                    "initialization_mode": mode,
                    "transaction_id": transaction_id,
                    "run_context_digest": sha256_file(run_context_path),
                    "required_views_digest": sha256_file(required_views_path),
                    "info_all_digest": sha256_file(info_path),
                    "structure_digest": structure,
                    "accounting_digest_before": before,
                    "accounting_digest_after": after,
                    "workbook_file_digest_before": workbook_before,
                    "workbook_file_digest_after": workbook_after,
                    "harden_sdc": sdc_snapshot,
                }
                report_text = render_environment_report(
                    run_root,
                    run_context,
                    required_views,
                    instances,
                    records,
                    manifest_rows,
                    manifest_counts,
                    structure,
                    before,
                    after,
                    workbook_before,
                    workbook_after,
                    transaction_id,
                    mode,
                    report,
                )
                completion = {
                    "schema_version": SCHEMA_VERSION,
                    "author": author_name(),
                    "run_id": run_context["run_id"],
                    "mode_label": run_context["mode_label"],
                    "stage_name": STAGE_NAME,
                    "stage": "",
                    "corner": "",
                    "completion_status": "complete",
                    "error_count": 0,
                    "sync_changed": "no",
                    "structure_digest": structure,
                    "accounting_digest_before": before,
                    "accounting_digest_after": after,
                    "upstream_artifact_digests": {
                        "run_context.csv": snapshot["run_context_digest"],
                        "required_views.csv": snapshot["required_views_digest"],
                        "info_all.xlsx": snapshot["info_all_digest"],
                    },
                    "output_sdc_digest": "",
                    "accounting_delta_digest": sha256_bytes(delta_payload),
                    "transaction_id": transaction_id,
                }
                artifact_payloads = [
                    (manifest_path, manifest_payload),
                    (snapshot_path, (json.dumps(snapshot, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")),
                    (delta_path, delta_payload),
                    (delta_meta_path, (json.dumps(delta_meta, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")),
                    (completion_path, (json.dumps(completion, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")),
                    (report_path, report_text.encode("utf-8")),
                ]
                execute_transaction(
                    run_root,
                    models,
                    prepared_candidates,
                    artifact_payloads,
                    transaction_id,
                    run_context,
                    structure,
                    before,
                    after,
                    report,
                )
                success = True
    except Exception as exc:
        if not report.error_count:
            report.error(str(exc))
        failure_context = run_context or {"run_id": "", "mode_label": "", "design_revision": ""}
        failure_text = render_environment_report(
            run_root,
            failure_context,
            required_views,
            instances,
            records,
            manifest_rows,
            manifest_counts,
            structure,
            before,
            after,
            workbook_before,
            workbook_after,
            transaction_id,
            mode,
            report,
        )
        atomic_write_text(report_path, failure_text)
    finally:
        if preview_dir is not None and preview_dir.exists():
            shutil.rmtree(str(preview_dir))
        for model in models:
            try:
                model.workbook.close()
            except Exception:
                pass

    if args.debug:
        for line in report.lines:
            print(line)
    for line in report.lines:
        if line.startswith("ERROR:"):
            print(line, file=sys.stderr)
    print(
        "Run ID: %s | mode: %s | warnings: %d | errors: %d"
        % (run_context.get("run_id", ""), run_context.get("mode_label", ""), report.warning_count, report.error_count)
    )
    if not success or report.error_count:
        print("Report: %s" % report_path, file=sys.stderr)
        return 2
    print("Manifest: %s" % manifest_path)
    print("Input snapshot: %s" % snapshot_path)
    print("Accounting delta: %s" % delta_path)
    print("Stage completion: %s" % completion_path)
    print("Report: %s" % report_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
