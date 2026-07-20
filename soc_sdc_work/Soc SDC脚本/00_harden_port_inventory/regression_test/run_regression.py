#!/usr/bin/env python3
"""Regression tests for the current 00 environment initialization runtime."""

import argparse
import csv
import hashlib
import json
import os
import shutil
import socket
import subprocess
import sys
import tempfile
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


BASE = Path(__file__).resolve().parent
EX00 = BASE.parent / "00_harden_port_inventory.py"
PORT_COLUMNS = [
    "Input",
    "Input Width",
    "Input Used Width",
    "From Whom",
    "Output",
    "Output Width",
    "Output Used Width",
    "To Top",
    "Inout",
    "Inout Width",
    "Inout Connectivity",
    "Inout Name",
]


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def run_00(root, *extra):
    command = [sys.executable, str(EX00), "--run-root", str(root)] + list(extra)
    return subprocess.run(
        command,
        cwd=str(BASE.parent),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )


def read_csv(path):
    with path.open("r", encoding="utf-8", newline="") as file_obj:
        return list(csv.DictReader(file_obj))


def write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def append_port_row(sheet, **values):
    row = [values.get(column, "") for column in PORT_COLUMNS]
    sheet.append(row)


def new_port_workbook(sheet_names):
    workbook = Workbook()
    first = True
    for sheet_name in sheet_names:
        if first:
            sheet = workbook.active
            sheet.title = sheet_name
            first = False
        else:
            sheet = workbook.create_sheet(sheet_name)
        sheet.append(PORT_COLUMNS + ["Audit Formula"])
    return workbook


def write_context(inputs, invalid_flag=False):
    write_csv(
        inputs / "run_context.csv",
        ["run_id", "mode_label", "design_revision", "note"],
        [{"run_id": "RUN_FUNC_001", "mode_label": "func", "design_revision": "revA", "note": "00 regression"}],
    )
    write_csv(
        inputs / "required_views.csv",
        ["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"],
        [
            {
                "view_id": "prects_ss",
                "stage": "prects",
                "corner": "ss_125",
                "require_02": "maybe" if invalid_flag else "yes",
                "require_04": "yes",
                "require_20": "yes",
                "require_30": "yes",
                "note": "required signoff view",
            }
        ],
    )


def write_info(inputs):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["module_name", "inst_name", "owner", "sdc_status", "sdc_note"])
    sheet.append(["a", "u_a", "owner_a", "", ""])
    sheet.append(["b", "u_b", "owner_b", "", ""])
    sheet.append(["io", "u_io", "owner_io", "not_required", "black-box IO shell"])
    workbook.save(str(inputs / "info_all.xlsx"))


def write_ports(inputs):
    alpha = new_port_workbook(["u_a"])
    sheet = alpha["u_a"]
    append_port_row(sheet, Output="data_o[7:4]", **{"Output Width": 4})
    append_port_row(sheet, Output="open_o", **{"Output Width": 1, "To Top": "OPEN"})
    append_port_row(sheet, Input="top_i", **{"Input Width": 1, "From Whom": "top.pad_i"})
    sheet["E3"].fill = PatternFill(fill_type="solid", fgColor="00FFFF00")
    sheet["M3"] = "=1+1"
    alpha.save(str(inputs / "port_alpha.xlsx"))

    beta = new_port_workbook(["u_b", "u_io"])
    sheet = beta["u_b"]
    append_port_row(sheet, Input="data_i[3:0]", **{"Input Width": 4, "From Whom": "u_a.data_o[7:4]"})
    append_port_row(sheet, Input="nc_i[1:0]", **{"Input Width": 2, "From Whom": "NC"})
    append_port_row(sheet, Input="tie_i[3:0]", **{"Input Width": 4, "From Whom": "4'b1010"})
    sheet = beta["u_io"]
    append_port_row(
        sheet,
        Inout="pad_io[1:0]",
        **{"Inout Width": 2, "Inout Connectivity": "top.pad_io[1:0]"}
    )
    append_port_row(sheet, Inout="open_io", **{"Inout Width": 1, "Inout Connectivity": "OPEN"})
    beta.save(str(inputs / "port_beta.xlsx"))


def write_case(root, invalid_flag=False):
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_context(inputs, invalid_flag=invalid_flag)
    write_info(inputs)
    write_ports(inputs)
    (inputs / "a.sdc").write_text("# u_a harden SDC\n", encoding="utf-8")


def used_value(path, sheet_name, cell):
    workbook = load_workbook(str(path), data_only=False)
    value = workbook[sheet_name][cell].value
    workbook.close()
    return value


def test_fresh_and_resume(work):
    root = work / "main"
    write_case(root)
    alpha = root / "inputs/port_alpha.xlsx"
    beta = root / "inputs/port_beta.xlsx"
    first = run_00(root)
    require(first.returncode == 0, "fresh run failed:\n%s\n%s" % (first.stdout, first.stderr))
    require(first.stdout.count("Author: Howard") == 1, "stdout author marker missing/duplicated")

    middle = root / "00_middle"
    report_path = root / "00_result/reports/environment_report.txt"
    for path in (
        middle / "harden_sdc_manifest.csv",
        middle / "input_snapshot.meta",
        middle / "port_accounting_delta.csv",
        middle / "port_accounting_delta.meta",
        middle / "stage_completion.meta",
        report_path,
    ):
        require(path.is_file(), "required 00 artifact missing: %s" % path)
    require(not (middle / "connection_inventory.csv").exists(), "obsolete connection inventory was generated")
    require(not list(middle.glob("**/pending")), "obsolete pending directory was generated")

    require(used_value(alpha, "u_a", "G3") == "0", "OPEN output was not structurally accounted")
    require(used_value(beta, "u_b", "C3") == "0,1", "NC input bits were not accounted")
    require(used_value(beta, "u_b", "C4") == "0,1,2,3", "constant input bits were not accounted")
    require(used_value(beta, "u_io", "L3") == "0", "OPEN inout was not accounted")

    styled = load_workbook(str(alpha), data_only=False)
    require(styled["u_a"]["E3"].fill.fgColor.rgb.endswith("FFFF00"), "non-accounting fill was not preserved")
    require(styled["u_a"]["M3"].value == "=1+1", "formula was not preserved")
    styled.close()

    delta = read_csv(middle / "port_accounting_delta.csv")
    require(len(delta) == 8, "expected 8 structural bit delta rows, got %d" % len(delta))
    require(all(row["owner_object_id"].startswith("STRUCT_") and len(row["owner_object_id"]) == 71 for row in delta), "STRUCT ID format mismatch")
    require({row["reason"] for row in delta} == {"structural_nc", "structural_open", "structural_tie_off"}, "structural reasons mismatch")

    manifest = {row["inst_name"]: row for row in read_csv(middle / "harden_sdc_manifest.csv")}
    require(manifest["u_a"]["availability_status"] == "available", "available SDC mapping failed")
    require(manifest["u_a"]["sdc_digest"] == sha256(root / "inputs/a.sdc"), "SDC digest missing/mismatched")
    require(manifest["u_b"]["availability_status"] == "missing", "missing SDC status failed")
    require(manifest["u_io"]["availability_status"] == "not_required", "not_required SDC status failed")

    snapshot_first = json.loads((middle / "input_snapshot.meta").read_text(encoding="utf-8"))
    completion = json.loads((middle / "stage_completion.meta").read_text(encoding="utf-8"))
    meta_first = json.loads((middle / "port_accounting_delta.meta").read_text(encoding="utf-8"))
    require(completion["completion_status"] == "complete" and completion["error_count"] == 0, "completion meta invalid")
    require(snapshot_first["structure_digest"] == completion["structure_digest"], "structure digest mismatch")
    require(snapshot_first["accounting_digest_after"] == completion["accounting_digest_after"], "accounting digest mismatch")
    require(len(meta_first["transactions"]) == 1, "fresh transaction chain length mismatch")
    require(not any((root / ".accounting_txn").iterdir()), "committed transaction scratch was not cleaned")

    fresh_again = run_00(root)
    require(fresh_again.returncode == 2, "fresh rerun should reject non-empty Used state")
    resume = run_00(root, "--resume-accounting")
    require(resume.returncode == 0, "resume failed:\n%s\n%s" % (resume.stdout, resume.stderr))
    meta_resume = json.loads((middle / "port_accounting_delta.meta").read_text(encoding="utf-8"))
    delta_resume = read_csv(middle / "port_accounting_delta.csv")
    snapshot_resume = json.loads((middle / "input_snapshot.meta").read_text(encoding="utf-8"))
    require(len(meta_resume["transactions"]) == 2, "resume transaction was not appended")
    require(len(delta_resume) == 16, "resume cumulative delta row count mismatch")
    require(all(not row["added_bits"] for row in delta_resume[8:]), "idempotent resume re-added structural bits")
    require(snapshot_first["structure_digest"] == snapshot_resume["structure_digest"], "Used write changed structure digest")
    require(snapshot_resume["accounting_digest_before"] == snapshot_resume["accounting_digest_after"], "idempotent resume changed accounting state")
    report = report_path.read_text(encoding="utf-8")
    require("Initialization mode: resume" in report, "resume report metadata missing")
    require("Missing harden SDC: 1" in report, "missing SDC report count missing")


def test_strict_missing(work):
    root = work / "strict"
    write_case(root)
    before = sha256(root / "inputs/port_beta.xlsx")
    result = run_00(root, "--require-complete-harden-sdc")
    require(result.returncode == 2, "strict missing SDC should fail")
    require("--require-complete-harden-sdc" in result.stderr, "strict failure reason missing")
    require(sha256(root / "inputs/port_beta.xlsx") == before, "strict failure modified workbook")
    require(not (root / "00_middle/stage_completion.meta").exists(), "strict failure published completion")


def test_invalid_inputs(work):
    root = work / "invalid_used"
    write_case(root)
    beta = root / "inputs/port_beta.xlsx"
    workbook = load_workbook(str(beta))
    workbook["u_b"]["C2"] = "0"
    workbook.save(str(beta))
    result = run_00(root)
    require(result.returncode == 2 and "must be blank in a fresh run" in result.stderr, "fresh Used state was not rejected")

    root = work / "invalid_inout"
    write_case(root)
    beta = root / "inputs/port_beta.xlsx"
    workbook = load_workbook(str(beta))
    workbook["u_io"]["L2"] = "top.legacy_pad"
    workbook.save(str(beta))
    result = run_00(root)
    require(result.returncode == 2 and "Inout Name still contains a connection" in result.stderr, "legacy Inout Name connection was not rejected")

    root = work / "invalid_view"
    write_case(root, invalid_flag=True)
    result = run_00(root)
    require(result.returncode == 2 and "expected yes/no" in result.stderr, "invalid required view flag was not rejected")

    root = work / "width_mismatch"
    write_case(root)
    beta = root / "inputs/port_beta.xlsx"
    workbook = load_workbook(str(beta))
    workbook["u_b"]["D2"] = "u_a.data_o[7:5]"
    workbook.save(str(beta))
    result = run_00(root)
    require(result.returncode == 2 and "source width 3 does not match destination width 4" in result.stderr, "range width mismatch was not rejected")

    root = work / "literal_width_mismatch"
    write_case(root)
    beta = root / "inputs/port_beta.xlsx"
    workbook = load_workbook(str(beta))
    append_port_row(
        workbook["u_b"],
        Input="bad_literal_i[3:0]",
        **{"Input Width": 4, "From Whom": "2'b10"}
    )
    workbook.save(str(beta))
    workbook.close()
    result = run_00(root)
    width_message = "sized Verilog literal width 2 does not match port width 4"
    generic_message = "connection must be <inst>.<port/range> or an explicit structural token"
    report = (root / "00_result/reports/environment_report.txt").read_text(encoding="utf-8")
    require(result.returncode == 2, "sized literal width mismatch was not rejected")
    require(report.count(width_message) == 1, "sized literal width mismatch was not reported exactly once")
    require(generic_message not in report, "invalid sized literal was misclassified as a generic connection")
    require(used_value(beta, "u_b", "C5") in (None, ""), "failed literal validation modified Used state")
    require(not (root / "00_middle/stage_completion.meta").exists(), "failed literal validation published completion")


def test_lock(work):
    root = work / "lock"
    write_case(root)
    lock = root / "inputs/.port_accounting.lock"
    lock.write_text(
        json.dumps({"pid": os.getpid(), "host": socket.gethostname(), "stage": "test"}) + "\n",
        encoding="utf-8",
    )
    result = run_00(root)
    require(result.returncode == 2 and "lock already exists" in result.stderr, "active accounting lock was not enforced")


def test_transaction_recovery(work):
    root = work / "recovery"
    write_case(root)
    first = run_00(root)
    require(first.returncode == 0, "recovery setup run failed")
    target = root / "inputs/port_alpha.xlsx"
    committed_digest = sha256(target)
    txn = root / ".accounting_txn/fake_apply"
    original = txn / "original"
    candidate = txn / "candidate"
    artifact_original = txn / "artifact_original"
    original.mkdir(parents=True)
    candidate.mkdir()
    artifact_original.mkdir()
    backup = original / target.name
    shutil.copy2(str(target), str(backup))
    workbook = load_workbook(str(target))
    workbook["u_a"]["G3"] = "99"
    workbook.save(str(target))
    manifest = {
        "schema_version": "1.0",
        "stage_name": "00_harden_port_inventory",
        "transaction_id": "fake_apply",
        "run_id": "RUN_FUNC_001",
        "mode_label": "func",
        "status": "APPLYING",
        "workbooks": [
            {
                "target": str(target),
                "original": str(backup),
                "candidate": str(candidate / target.name),
                "digest_before": committed_digest,
                "digest_after": "unused",
                "applied": True,
            }
        ],
        "artifacts": [],
    }
    (txn / "transaction.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    recovered = run_00(root, "--resume-accounting")
    require(recovered.returncode == 0, "transaction recovery/resume failed: %s" % recovered.stderr)
    require(used_value(target, "u_a", "G3") == "0", "incomplete transaction was not rolled back")
    rolled = json.loads((txn / "transaction.json").read_text(encoding="utf-8"))
    require(rolled["status"] == "ROLLED_BACK", "transaction was not marked rolled back")


def write_scale_case(root, bit_count):
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_csv(
        inputs / "run_context.csv",
        ["run_id", "mode_label", "design_revision", "note"],
        [{"run_id": "RUN_SCALE", "mode_label": "func", "design_revision": "scale", "note": ""}],
    )
    write_csv(
        inputs / "required_views.csv",
        ["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"],
        [],
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["module_name", "inst_name", "owner"])
    sheet.append(["a", "u_a", "a"])
    sheet.append(["b", "u_b", "b"])
    workbook.save(str(inputs / "info_all.xlsx"))
    ports = new_port_workbook(["u_a", "u_b"])
    append_port_row(ports["u_a"], Output="data_o", **{"Output Width": bit_count})
    append_port_row(ports["u_b"], Input="data_i", **{"Input Width": bit_count, "From Whom": "u_a.data_o"})
    ports.save(str(inputs / "port_scale.xlsx"))


def test_scale(work, bit_count):
    if bit_count <= 0:
        return
    root = work / "scale"
    write_scale_case(root, bit_count)
    started = time.time()
    result = run_00(root)
    elapsed = time.time() - started
    require(result.returncode == 0, "scale run failed: %s" % result.stderr)
    resumed = run_00(root, "--resume-accounting")
    require(resumed.returncode == 0, "zero-delta scale resume failed: %s" % resumed.stderr)
    report = (root / "00_result/reports/environment_report.txt").read_text(encoding="utf-8")
    require("Canonical port bits: %d" % (bit_count * 2) in report, "scale bit count mismatch")
    print("scale: %d port bits validated in %.2fs" % (bit_count * 2, elapsed))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--scale-bits", type=int, default=0)
    args = parser.parse_args()
    require(EX00.is_file(), "00 script not found in 00 folder: %s" % EX00)
    work = Path(tempfile.mkdtemp(prefix="soc_sdc_00_regression_"))
    try:
        test_fresh_and_resume(work)
        test_strict_missing(work)
        test_invalid_inputs(work)
        test_lock(work)
        test_transaction_recovery(work)
        test_scale(work, args.scale_bits)
    finally:
        shutil.rmtree(str(work))
    print("00 environment initialization regression: PASS")


if __name__ == "__main__":
    main()
