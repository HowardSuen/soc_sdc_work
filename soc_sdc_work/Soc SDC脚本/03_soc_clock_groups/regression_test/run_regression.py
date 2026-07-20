#!/usr/bin/env python3
"""Complex regression for 03_extract_soc_clock_groups.py.

The test builds fresh inputs under work_complex/ and checks:
  * first-run workbook creation gate
  * bit-level clock_name handling from 01 clock_inventory.csv
  * domain closure expansion through a bit-level source object
  * relation_type alias normalization to the canonical enum
"""
from __future__ import print_function

import csv
import hashlib
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
SOC = BASE.parent.parent
EX03 = SOC / "03_soc_clock_groups" / "03_extract_soc_clock_groups.py"
WORK = BASE / "work_complex"


def clean_dir(path):
    if path.exists():
        shutil.rmtree(str(path))
    path.mkdir(parents=True)


def sh(args, cwd):
    return subprocess.run(
        [sys.executable] + [str(arg) for arg in args],
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def write_inventory(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "inst_name",
        "module_name",
        "port_name",
        "direction",
        "clock_name",
        "clock_kind",
        "period",
        "waveform",
        "direct_source",
        "root_source",
        "from_whom",
        "original_sdc",
        "source_line",
        "original_clock_name",
        "original_command",
        "final_action",
        "note",
    ]
    rows = [
        {
            "inst_name": "u_busclk",
            "port_name": "ref_clk_i[1]",
            "direction": "input",
            "clock_name": "top_ref_clk_pad_bit1",
            "clock_kind": "create_clock",
            "direct_source": "top/ref_clk_pad[1]",
            "root_source": "top/ref_clk_pad[1]",
            "final_action": "emit_top_clock",
        },
        {
            "inst_name": "u_busclk",
            "port_name": "clk_o[1]",
            "direction": "output",
            "clock_name": "u_busclk_clk_o_bit1",
            "clock_kind": "create_generated_clock",
            "direct_source": "u_busclk/ref_clk_i[1]",
            "root_source": "top/ref_clk_pad[1]",
            "final_action": "emit_output_clock",
        },
        {
            "inst_name": "u_aux",
            "port_name": "aux_clk_i",
            "direction": "input",
            "clock_name": "top_aux_clk_pad",
            "clock_kind": "create_clock",
            "direct_source": "top/aux_clk_pad",
            "root_source": "top/aux_clk_pad",
            "final_action": "emit_top_clock",
        },
    ]
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            full = dict((field, "") for field in fields)
            full.update(row)
            writer.writerow(full)


def sha256_file(path):
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        while True:
            chunk = file_obj.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def sdc_header_value(path, field):
    prefix = "# %s: " % field
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.startswith(prefix):
            return line[len(prefix):].strip()
    return ""


def write_target_inventory(root, scenario, rows, completeness="complete", missing_instances=None):
    missing_instances = sorted(missing_instances or [])
    run_id = root.name
    mode_label = scenario
    inputs = root / "inputs"
    inputs.mkdir(parents=True, exist_ok=True)
    (inputs / "run_context.csv").write_text(
        "run_id,mode_label,design_revision,note\n%s,%s,rev_a,03 regression\n"
        % (run_id, mode_label),
        encoding="utf-8",
    )
    (inputs / "required_views.csv").write_text(
        "view_id,stage,corner,require_02,require_04,require_20,require_30,note\n"
        "prects_ss,prects,ss_125,yes,yes,no,no,regression view\n",
        encoding="utf-8",
    )
    inventory_path = root / "01_middle" / "clock_inventory.csv"
    inventory_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "schema_version", "run_id", "mode_label", "design_revision",
        "inst_name", "module_name", "port_name", "direction", "clock_name", "clock_kind",
        "period", "waveform", "direct_source", "root_source", "from_whom", "original_sdc",
        "source_line", "original_clock_name", "original_command", "final_action", "source_type",
        "source_file", "target_object", "final_sdc_digest", "run_completeness",
        "missing_instances", "structure_digest", "note",
    ]
    active_names = sorted(
        row["clock_name"]
        for row in rows
        if row.get("final_action") in {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}
    )
    structure_digest = hashlib.sha256(
        json.dumps([run_id, mode_label, active_names], separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    with inventory_path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            full = dict((field, "") for field in fields)
            full.update(row)
            full["schema_version"] = "1.0"
            full["run_id"] = run_id
            full["mode_label"] = mode_label
            full["design_revision"] = "rev_a"
            full["run_completeness"] = completeness
            full["missing_instances"] = ";".join(missing_instances)
            full["structure_digest"] = structure_digest
            writer.writerow(full)

    sdc_path = root / "01_result" / "01_soc_clocks.sdc"
    sdc_path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["# target regression clock SDC"]
    for row in rows:
        if row.get("final_action") not in {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}:
            continue
        name = row["clock_name"]
        if "generated" in row.get("clock_kind", ""):
            lines.append("create_generated_clock -name %s [get_pins {%s}]" % (name, row.get("port_name") or name))
        else:
            lines.append("create_clock -name %s -period 10 [get_ports {%s}]" % (name, row.get("port_name") or name))
    sdc_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    meta = {
        "schema_version": "1.0",
        "author": "Howard",
        "stage_name": "01_soc_clocks",
        "run_id": run_id,
        "mode_label": mode_label,
        "design_revision": "rev_a",
        "completion_status": "complete",
        "error_count": 0,
        "structure_digest": structure_digest,
        "final_sdc_path": str(sdc_path.resolve()),
        "final_sdc_digest": sha256_file(sdc_path),
        "inventory_path": str(inventory_path.resolve()),
        "inventory_digest": sha256_file(inventory_path),
        "clock_set_digest": hashlib.sha256("\n".join(active_names).encode("utf-8")).hexdigest(),
        "clock_count": len(active_names),
        "run_completeness": completeness,
        "available_harden_count": 1 if completeness == "complete" else 1,
        "missing_harden_count": len(missing_instances),
        "not_required_harden_count": 0,
        "missing_instances": missing_instances,
    }
    meta_path = inventory_path.with_suffix(".meta")
    meta_path.write_text(json.dumps(meta, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    completion = {
        "schema_version": "1.0",
        "author": "Howard",
        "stage_name": "01_soc_clocks",
        "run_id": run_id,
        "mode_label": mode_label,
        "design_revision": "rev_a",
        "completion_status": "complete",
        "error_count": 0,
        "sync_changed": "no",
        "structure_digest": structure_digest,
        "clock_inventory_digest": sha256_file(inventory_path),
        "output_sdc_digest": sha256_file(sdc_path),
    }
    completion_path = root / "01_middle" / "stage_completion.meta"
    completion_path.write_text(
        json.dumps(completion, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return inventory_path, meta_path, sdc_path


def base_target_rows():
    return [
        {
            "inst_name": "top", "port_name": "cpu_clk", "clock_name": "cpu_clk",
            "clock_kind": "create_clock", "direct_source": "top/cpu_clk",
            "root_source": "top/cpu_clk", "final_action": "emit_top_clock",
        },
        {
            "inst_name": "u_cpu", "port_name": "cpu_div", "clock_name": "cpu_div",
            "clock_kind": "create_generated_clock", "direct_source": "top/cpu_clk",
            "root_source": "top/cpu_clk", "final_action": "emit_output_clock",
        },
        {
            "inst_name": "top", "port_name": "cpu_manual", "clock_name": "cpu_manual",
            "clock_kind": "create_clock", "direct_source": "top/cpu_manual",
            "root_source": "top/cpu_manual", "final_action": "emit_top_clock",
        },
        {
            "inst_name": "top", "port_name": "aon_clk", "clock_name": "aon_clk",
            "clock_kind": "create_clock", "direct_source": "top/aon_clk",
            "root_source": "top/aon_clk", "final_action": "emit_top_clock",
        },
    ]


def sheet_header_map(ws, required):
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row_idx, col_idx).value
            if value:
                mapping[str(value).strip()] = col_idx
        if required in mapping:
            return row_idx, mapping
    raise AssertionError("%s header not found" % required)


def header_map(ws):
    return sheet_header_map(ws, "group_id")


def set_row(ws, row_idx, mapping, values):
    for key, value in values.items():
        if key in mapping:
            ws.cell(row_idx, mapping[key], value)


def run_bit_closure_case():
    d = WORK / "bit_closure"
    clean_dir(d)
    write_inventory(d / "clock_inventory.csv")

    first = sh([EX03, "-scenario", "common", "-input", "clock_inventory.csv"], d)
    require(first.returncode == 1, "first 03 run should create workbook and stop")
    require((d / "03_soc_clock_groups.xlsx").is_file(), "03 workbook was not created")

    wb = load_workbook(str(d / "03_soc_clock_groups.xlsx"))
    ws = wb["clock_group_rules"]
    header_row, mapping = header_map(ws)
    set_row(ws, header_row + 1, mapping, {
        "scenario": "common",
        "group_id": "CG_ASYNC_BIT_AUX",
        "relation_type": "async",
        "group_1_clocks": "top_ref_clk_pad_bit1",
        "group_2_clocks": "top_aux_clk_pad",
        "analysis_style": "normal",
        "apply": "yes",
        "review_status": "approved",
        "owner": "sta",
        "basis": "CDC bit clock domain async to aux",
        "cdc_required": "yes",
    })
    wb.save(str(d / "03_soc_clock_groups.xlsx"))

    result = sh([EX03, "-scenario", "common", "-input", "clock_inventory.csv"], d)
    require(result.returncode == 0, "03 bit closure generation failed:\n%s\n%s" % (result.stdout, result.stderr))
    sdc = (d / "common" / "03_soc_clock_groups.sdc").read_text(encoding="utf-8")
    report = (d / "clock_group_check_report_common.txt").read_text(encoding="utf-8")
    require("set_clock_groups -asynchronous" in sdc, "relation_type alias was not canonicalized")
    require(
        "-group [get_clocks {top_ref_clk_pad_bit1 u_busclk_clk_o_bit1}]" in sdc,
        "bit-level descendant was not auto-added to the effective group",
    )
    require("group_1 auto-added descendant clock(s): u_busclk_clk_o_bit1" in report, "auto-added bit descendant not reported")
    require("relation_type async normalized to asynchronous" in report, "canonical relation_type rewrite was not reported")

    rewritten = load_workbook(str(d / "03_soc_clock_groups.xlsx"))["clock_group_rules"]
    _, rewritten_mapping = header_map(rewritten)
    require(
        rewritten.cell(header_row + 1, rewritten_mapping["relation_type"]).value == "asynchronous",
        "relation_type alias was not written back to the workbook",
    )
    require(result.stdout.count("Author: Howard") == 1, "legacy stdout author marker missing or duplicated")
    require((d / "relation_map" / "common.csv").is_file(), "legacy complete relation map missing")


def fill_domain_workbook(root, domain_rows, rule_rows):
    form = root / "03_middle" / "03_soc_clock_groups.xlsx"
    wb = load_workbook(str(form))
    domain_ws = wb["clock_domain_membership"]
    domain_header, domain_mapping = sheet_header_map(domain_ws, "domain_id")
    existing_clock_rows = {}
    for row_idx in range(domain_header + 1, domain_ws.max_row + 1):
        clock_name = domain_ws.cell(row_idx, domain_mapping["clock_name"]).value
        if clock_name:
            existing_clock_rows[str(clock_name).strip()] = row_idx
    touched_rows = set()
    for values in domain_rows:
        clock_name = str(values.get("clock_name", "")).strip()
        row_idx = existing_clock_rows.get(clock_name, domain_ws.max_row + 1)
        set_row(domain_ws, row_idx, domain_mapping, values)
        touched_rows.add(row_idx)
    for row_idx in range(domain_header + 1, domain_ws.max_row + 1):
        if row_idx in touched_rows:
            continue
        clock_name = domain_ws.cell(row_idx, domain_mapping["clock_name"]).value
        if not clock_name:
            continue
        set_row(domain_ws, row_idx, domain_mapping, {
            "apply": "no",
            "review_status": "reviewed",
            "note": "Reviewed: no explicit domain membership required for this clock.",
        })
    rule_ws = wb["clock_group_rules"]
    rule_header, rule_mapping = header_map(rule_ws)
    for offset, values in enumerate(rule_rows, start=1):
        set_row(rule_ws, rule_header + offset, rule_mapping, values)
    wb.save(str(form))


def run_target_after_review(root, extra_args=None):
    args = [EX03, "--run-root", root] + list(extra_args or [])
    sync = sh(args, BASE)
    require(
        sync.returncode == 1,
        "first post-review target run must synchronize NEW_FROM_01 rows:\n%s\n%s"
        % (sync.stdout, sync.stderr),
    )
    report = (root / "03_result" / "reports" / "clock_group_check_report.txt").read_text(
        encoding="utf-8"
    )
    require("Sync changed: yes" in report, "post-review synchronization gate was not reported")
    return sh(args, BASE)


def run_target_domain_relation_case():
    root = WORK / "target_domain"
    clean_dir(root)
    write_target_inventory(root, "common", base_target_rows())

    first = sh([EX03, "--run-root", root], BASE)
    require(first.returncode == 1, "target first run should create workbook")
    require(first.stdout.count("Author: Howard") == 1, "target stdout author marker missing or duplicated")
    fill_domain_workbook(
        root,
        [
            {
                "domain_id": "DOM_CPU", "clock_name": "cpu_clk",
                "membership_type": "seed", "include_descendants": "yes", "apply": "yes",
                "review_status": "approved", "owner": "sta", "basis": "CPU root domain",
            },
            {
                "domain_id": "DOM_CPU", "clock_name": "cpu_manual",
                "membership_type": "explicit_member", "include_descendants": "no", "apply": "yes",
                "review_status": "approved", "owner": "sta",
                "basis": "Manual clock confirmed synchronous with CPU domain",
            },
            {
                "domain_id": "DOM_AON", "clock_name": "aon_clk",
                "membership_type": "seed", "include_descendants": "yes", "apply": "yes",
                "review_status": "approved", "owner": "sta", "basis": "AON domain",
            },
        ],
        [
            {
                "group_id": "CG_CPU_AON", "relation_type": "asynchronous",
                "group_1_domains": "DOM_CPU", "group_2_domains": "DOM_AON",
                "analysis_style": "normal", "apply": "yes", "review_status": "approved",
                "owner": "sta", "basis": "CDC architecture",
            },
        ],
    )
    result = run_target_after_review(root)
    require(result.returncode == 0, "target domain generation failed:\n%s\n%s" % (result.stdout, result.stderr))
    require(result.stdout.count("Author: Howard") == 1, "target stdout author marker missing or duplicated")
    sdc_path = root / "03_result" / "03_soc_clock_groups.sdc"
    relation_path = root / "03_middle" / "relation_map.csv"
    meta_path = root / "03_middle" / "relation_map.meta"
    completion_path = root / "03_middle" / "stage_completion.meta"
    report_path = root / "03_result" / "reports" / "clock_group_check_report.txt"
    sdc = sdc_path.read_text(encoding="utf-8")
    require("Author: Howard" in sdc, "target SDC author metadata missing")
    require("Run completeness: complete" in sdc, "target SDC completeness metadata missing")
    require(
        "-group [get_clocks {cpu_clk cpu_div cpu_manual}]" in sdc,
        "domain seed closure and explicit manual member were not resolved",
    )
    with relation_path.open("r", encoding="utf-8", newline="") as file_obj:
        relation_rows = list(csv.DictReader(file_obj))
    require(len(relation_rows) == 6, "complete relation map should contain C(4,2)=6 pairs")
    lookup = dict(((row["clock_a"], row["clock_b"]), row) for row in relation_rows)
    require(
        lookup[("aon_clk", "cpu_manual")]["relation_type"] == "asynchronous",
        "manual domain member relation missing",
    )
    require(
        lookup[("cpu_clk", "cpu_manual")]["relation_type"] == "synchronous"
        and lookup[("cpu_clk", "cpu_manual")]["relation_source"] == "default_synchronous",
        "same-side clocks must remain default synchronous in the complete map",
    )
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    require(meta["clock_universe_digest"], "relation meta clock universe digest missing")
    require(meta["active_rule_ids"] == ["CG_CPU_AON"], "relation meta active rules incorrect")
    completion = json.loads(completion_path.read_text(encoding="utf-8"))
    require(completion["completion_status"] == "complete", "03 completion was not published")
    require(completion["error_count"] == 0 and completion["sync_changed"] == "no", "03 completion gate fields incorrect")
    require(completion["port_accounting_summary"] == "Port accounting: not_applicable; added_bits=0", "03 port accounting contract missing")
    require("Author: Howard" in report_path.read_text(encoding="utf-8"), "report author metadata missing")


def run_sparse_group_domain_exclusion_case():
    root = WORK / "sparse_group_domain_exclusion"
    clean_dir(root)
    rows = [base_target_rows()[0], base_target_rows()[1], base_target_rows()[3]]
    write_target_inventory(root, "common", rows)

    first = sh([EX03, "--run-root", root], BASE)
    require(first.returncode == 1, "sparse-group first run should create workbook")
    fill_domain_workbook(
        root,
        [
            {
                "domain_id": "DOM_CPU", "clock_name": "cpu_clk",
                "membership_type": "seed", "include_descendants": "yes", "apply": "yes",
                "review_status": "approved", "owner": "sta", "basis": "CPU root domain",
            },
            {
                "domain_id": "DOM_AON", "clock_name": "aon_clk",
                "membership_type": "seed", "include_descendants": "yes", "apply": "yes",
                "review_status": "approved", "owner": "sta", "basis": "AON domain",
            },
        ],
        [
            {
                "group_id": "CG_SPARSE_GROUP",
                "relation_type": "asynchronous", "group_1_domains": "DOM_CPU",
                "group_3_domains": "DOM_AON", "exclude_descendant_clocks": "cpu_div",
                "analysis_style": "normal", "apply": "yes", "review_status": "approved",
                "owner": "sta", "basis": "Exclude generated CPU branch for this relation",
            },
        ],
    )
    result = run_target_after_review(root)
    require(result.returncode == 0, "sparse-group generation failed:\n%s\n%s" % (result.stdout, result.stderr))
    sdc = (root / "03_result" / "03_soc_clock_groups.sdc").read_text(encoding="utf-8")
    report = (root / "03_result" / "reports" / "clock_group_check_report.txt").read_text(encoding="utf-8")
    require("cpu_div" not in sdc, "domain descendant was not excluded from the rule")
    require("excluded clock is not a descendant" not in report, "valid domain descendant exclusion was warned")

    coverage = load_workbook(str(root / "03_result" / "reports" / "clock_group_coverage_report.xlsx"))
    pair_ws = coverage["explicit_relation_coverage"]
    pair_header, pair_mapping = sheet_header_map(pair_ws, "clock_a")
    require(pair_ws.cell(pair_header + 1, pair_mapping["group_a"]).value == "group_1", "group_1 provenance changed")
    require(pair_ws.cell(pair_header + 1, pair_mapping["group_b"]).value == "group_3", "sparse group_3 was renumbered")
    complete_pair_ws = coverage["pair_relation_map"]
    complete_pair_header, _ = sheet_header_map(complete_pair_ws, "clock_a")
    require(
        complete_pair_ws.max_row - complete_pair_header == 3,
        "coverage pair_relation_map is not the complete C(3,2) relation map",
    )
    participation_ws = coverage["clock_participation"]
    participation_header, participation_mapping = sheet_header_map(participation_ws, "clock_name")
    aon_label = ""
    for row_idx in range(participation_header + 1, participation_ws.max_row + 1):
        if participation_ws.cell(row_idx, participation_mapping["clock_name"]).value == "aon_clk":
            aon_label = participation_ws.cell(row_idx, participation_mapping["groups"]).value or ""
            break
    require(":group_3:" in aon_label, "clock participation lost the original group_3 number")


def run_partial_blocked_case():
    root = WORK / "partial_blocked"
    clean_dir(root)
    rows = [base_target_rows()[0], base_target_rows()[3]]
    write_target_inventory(root, "common", rows, completeness="partial", missing_instances=["u_missing"])
    first = sh([EX03, "--run-root", root], BASE)
    require(first.returncode == 1, "partial first run should create workbook")
    fill_domain_workbook(
        root,
        [
            {
                "domain_id": "DOM_CPU", "clock_name": "cpu_clk",
                "membership_type": "seed", "include_descendants": "yes", "apply": "yes",
                "review_status": "approved", "basis": "available CPU domain",
            },
            {
                "domain_id": "DOM_AON", "clock_name": "aon_clk",
                "membership_type": "seed", "include_descendants": "yes", "apply": "yes",
                "review_status": "approved", "basis": "available AON domain",
            },
            {
                "domain_id": "DOM_MISSING", "clock_name": "missing_clk",
                "membership_type": "explicit_member", "include_descendants": "no",
                "source_instance": "u_missing", "apply": "yes", "review_status": "approved",
                "basis": "Expected clock from missing harden SDC",
            },
        ],
        [
            {
                "group_id": "CG_AVAILABLE", "relation_type": "asynchronous",
                "group_1_domains": "DOM_CPU", "group_2_domains": "DOM_AON",
                "analysis_style": "normal", "apply": "yes", "review_status": "approved",
                "owner": "sta", "basis": "available CDC relation",
            },
            {
                "group_id": "CG_BLOCKED", "relation_type": "asynchronous",
                "group_1_domains": "DOM_MISSING", "group_2_domains": "DOM_AON",
                "analysis_style": "normal", "apply": "yes", "review_status": "approved",
                "owner": "sta", "basis": "blocked CDC relation",
            },
        ],
    )
    result = run_target_after_review(root)
    require(result.returncode == 0, "partial generation should continue for available rules")
    sdc = (root / "03_result" / "03_soc_clock_groups.sdc").read_text(encoding="utf-8")
    report = (root / "03_result" / "reports" / "clock_group_check_report.txt").read_text(encoding="utf-8")
    meta = json.loads((root / "03_middle" / "relation_map.meta").read_text(encoding="utf-8"))
    require("CG_AVAILABLE" in sdc and "CG_BLOCKED" not in sdc, "blocked rule emission behavior incorrect")
    require("Run completeness: partial" in sdc, "partial completeness missing from SDC")
    require("blocked_by_missing_sdc" in report and "u_missing" in report, "blocked rule not reported")
    require(meta["blocked_rule_ids"] == ["CG_BLOCKED"], "blocked rule missing from relation meta")
    strict = sh([
        EX03, "--run-root", root, "--require-complete-harden-sdc",
    ], BASE)
    require(strict.returncode == 1, "strict completeness must block partial inventory")


def run_merged_basis_gate_case():
    root = WORK / "merged_basis_gate"
    clean_dir(root)
    rows = [base_target_rows()[0], base_target_rows()[3]]
    write_target_inventory(root, "common", rows)
    first = sh([EX03, "--run-root", root], BASE)
    require(first.returncode == 1, "merged gate first run should create workbook")
    fill_domain_workbook(
        root,
        [],
        [
            {
                "group_id": "CG_BAD_MERGED",
                "relation_type": "logically_exclusive", "group_1_clocks": "cpu_clk",
                "group_2_clocks": "aon_clk", "analysis_style": "merged_exclusive",
                "apply": "yes", "review_status": "approved",
                "owner": "sta", "basis": "mux select is case-fixed; no exception",
            },
        ],
    )
    sync = sh([EX03, "--run-root", root], BASE)
    require(sync.returncode == 1, "merged gate review sync should require one rerun")
    result = sh([EX03, "--run-root", root], BASE)
    require(result.returncode == 1, "case-fixed merged-exclusive basis must block generation")
    report = (root / "03_result" / "reports" / "clock_group_check_report.txt").read_text(encoding="utf-8")
    require("no_case_analysis" in report, "merged-exclusive evidence error missing")


def run_target_cli_contract_case():
    root = WORK / "target_cli_contract"
    clean_dir(root)
    write_target_inventory(root, "func", [base_target_rows()[0], base_target_rows()[3]])
    rejected = sh([EX03, "--run-root", root, "--scenario", "common"], BASE)
    require(rejected.returncode == 2, "target mode must reject --scenario")
    require("accepts no override" in rejected.stderr, "target scenario rejection reason missing")

    first = sh([EX03, "--run-root", root], BASE)
    require(first.returncode == 1, "diagnose contract first run should create workbook")
    fill_domain_workbook(root, [], [])
    sync = sh([EX03, "--run-root", root, "--diagnose-only"], BASE)
    require(sync.returncode == 1, "diagnose sync must still require review rerun")
    diagnostic = sh([EX03, "--run-root", root, "--diagnose-only"], BASE)
    require(diagnostic.returncode == 0, "reviewed diagnose-only validation should pass")
    require(not (root / "03_middle" / "stage_completion.meta").exists(), "diagnose-only published completion")
    require(not (root / "03_result" / "03_soc_clock_groups.sdc").exists(), "diagnose-only published formal SDC")
    require(not (root / "03_middle" / "relation_map.csv").exists(), "diagnose-only published relation map")
    report = (root / "03_result" / "reports" / "clock_group_check_report.txt").read_text(encoding="utf-8")
    require("Port accounting mode: diagnostic/read-only" in report, "diagnostic port-accounting marker missing")
    require("Accounting closure: not evaluated" in report, "diagnostic accounting-closure marker missing")


def run_target_stale_review_case():
    root = WORK / "target_stale_review"
    clean_dir(root)
    rows = [base_target_rows()[0], base_target_rows()[2], base_target_rows()[3]]
    write_target_inventory(root, "func", rows)
    first = sh([EX03, "--run-root", root], BASE)
    require(first.returncode == 1, "stale review first run should create workbook")
    fill_domain_workbook(root, [], [])
    formal = run_target_after_review(root)
    require(formal.returncode == 0, "stale review baseline formal generation failed")
    completion = root / "03_middle" / "stage_completion.meta"
    require(completion.is_file(), "stale review baseline completion missing")

    write_target_inventory(root, "func", [base_target_rows()[0], base_target_rows()[2]])
    stale = sh([EX03, "--run-root", root], BASE)
    require(stale.returncode == 1, "removed 01 clock must trigger review gate")
    require(not completion.exists(), "stale workbook retained a complete stage marker")
    coverage = load_workbook(str(root / "03_result" / "reports" / "clock_group_coverage_report.xlsx"))
    stale_ws = coverage["stale_invalid_rules"]
    stale_header, stale_mapping = sheet_header_map(stale_ws, "object_id")
    stale_objects = {
        stale_ws.cell(row_idx, stale_mapping["object_id"]).value
        for row_idx in range(stale_header + 1, stale_ws.max_row + 1)
    }
    require("aon_clk" in stale_objects, "removed 01 clock was not shown in stale/invalid coverage")
    unknown_ws = coverage["unknown_incomplete_pairs"]
    unknown_header, _ = sheet_header_map(unknown_ws, "clock_a")
    require(unknown_ws.max_row - unknown_header == 1, "review-required active pair was not marked unknown")


def run_per_run_case_gate_case():
    root = WORK / "per_run_case_gate"
    clean_dir(root)
    write_target_inventory(root, "func", [base_target_rows()[0], base_target_rows()[3]])
    first = sh([EX03, "--run-root", root], BASE)
    require(first.returncode == 1, "per_run_case first run should create workbook")
    fill_domain_workbook(
        root,
        [],
        [{
            "group_id": "CG_PER_RUN_CASE",
            "relation_type": "logically_exclusive",
            "group_1_clocks": "cpu_clk",
            "group_2_clocks": "aon_clk",
            "analysis_style": "per_run_case",
            "apply": "yes",
            "review_status": "approved",
            "owner": "sta",
            "basis": "mux select is fixed by pre-setup case analysis",
        }],
    )
    sync = sh([EX03, "--run-root", root], BASE)
    require(sync.returncode == 1, "per_run_case review sync should require one rerun")
    blocked = sh([EX03, "--run-root", root], BASE)
    require(blocked.returncode == 1, "per_run_case apply=yes must block generation")
    report = (root / "03_result" / "reports" / "clock_group_check_report.txt").read_text(encoding="utf-8")
    require("per_run_case" in report and "must not apply=yes" in report, "per_run_case gate error missing")


def run_semantic_digest_boundary_case():
    root = WORK / "semantic_digest_boundary"
    clean_dir(root)
    write_target_inventory(root, "func", [base_target_rows()[0], base_target_rows()[3]])
    first = sh([EX03, "--run-root", root], BASE)
    require(first.returncode == 1, "semantic digest first run should create workbook")
    fill_domain_workbook(
        root,
        [],
        [{
            "group_id": "CG_DIGEST_BOUNDARY",
            "relation_type": "asynchronous",
            "group_1_clocks": "cpu_clk",
            "group_2_clocks": "aon_clk",
            "analysis_style": "normal",
            "apply": "yes",
            "review_status": "approved",
            "owner": "sta",
            "basis": "independent clock roots are CDC-reviewed asynchronous",
        }],
    )
    baseline = run_target_after_review(root)
    require(baseline.returncode == 0, "semantic digest baseline generation failed")

    sdc_path = root / "03_result/03_soc_clock_groups.sdc"
    relation_path = root / "03_middle/relation_map.csv"
    completion_path = root / "01_middle/stage_completion.meta"
    baseline_sdc = sdc_path.read_bytes()
    baseline_relation = relation_path.read_bytes()
    baseline_semantic_digest = sdc_header_value(sdc_path, "Workbook semantic digest")
    baseline_completion_digest = sha256_file(completion_path)
    require(baseline_semantic_digest, "baseline workbook semantic digest missing")

    upstream_completion = json.loads(completion_path.read_text(encoding="utf-8"))
    upstream_completion["transaction_id"] = "01_dynamic_provenance_only_change"
    completion_path.write_text(
        json.dumps(upstream_completion, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    changed_completion_digest = sha256_file(completion_path)
    require(
        changed_completion_digest != baseline_completion_digest,
        "test did not change the raw 01 completion file digest",
    )

    provenance_sync = sh([EX03, "--run-root", root], BASE)
    require(provenance_sync.returncode == 1, "changed upstream provenance must synchronize run_metadata")
    provenance_formal = sh([EX03, "--run-root", root], BASE)
    require(provenance_formal.returncode == 0, "provenance-only formal rerun failed")
    require(sdc_path.read_bytes() == baseline_sdc, "provenance-only change altered formal 03 SDC bytes")
    require(relation_path.read_bytes() == baseline_relation, "provenance-only change altered relation_map.csv")
    require(
        sdc_header_value(sdc_path, "Workbook semantic digest") == baseline_semantic_digest,
        "provenance-only change altered workbook semantic digest",
    )
    relation_meta = json.loads((root / "03_middle/relation_map.meta").read_text(encoding="utf-8"))
    stage_completion = json.loads((root / "03_middle/stage_completion.meta").read_text(encoding="utf-8"))
    require(
        relation_meta["upstream_01_completion_digest"] == changed_completion_digest,
        "relation meta did not retain changed raw 01 completion provenance",
    )
    require(
        stage_completion["upstream_01_completion_digest"] == changed_completion_digest,
        "03 completion did not retain changed raw 01 completion provenance",
    )

    form_path = root / "03_middle/03_soc_clock_groups.xlsx"
    wb = load_workbook(str(form_path))
    ws = wb["clock_group_rules"]
    rule_header, mapping = header_map(ws)
    ws.cell(rule_header + 1, mapping["relation_type"], "physically_exclusive")
    ws.cell(
        rule_header + 1,
        mapping["basis"],
        "package source strap makes these clock roots physically exclusive",
    )
    wb.save(str(form_path))
    semantic_change = sh([EX03, "--run-root", root], BASE)
    require(semantic_change.returncode == 0, "real clock-group semantic change failed")
    require(sdc_path.read_bytes() != baseline_sdc, "real semantic change did not alter formal 03 SDC")
    require(
        sdc_header_value(sdc_path, "Workbook semantic digest") != baseline_semantic_digest,
        "real semantic change did not alter workbook semantic digest",
    )
    require("set_clock_groups -physically_exclusive" in sdc_path.read_text(encoding="utf-8"), "changed relation was not emitted")


def main():
    clean_dir(WORK)
    run_bit_closure_case()
    run_target_domain_relation_case()
    run_sparse_group_domain_exclusion_case()
    run_partial_blocked_case()
    run_merged_basis_gate_case()
    run_target_cli_contract_case()
    run_target_stale_review_case()
    run_per_run_case_gate_case()
    run_semantic_digest_boundary_case()
    print("03 complex regression: PASS")
    print("  cases: bit_closure, target_domain, sparse_group_domain_exclusion, partial_blocked, merged_basis_gate, target_cli_contract, target_stale_review, per_run_case_gate, semantic_digest_boundary")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
