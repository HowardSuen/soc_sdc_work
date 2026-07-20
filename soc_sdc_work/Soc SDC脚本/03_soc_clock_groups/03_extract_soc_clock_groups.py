#!/usr/bin/env python3
"""
Generate 03 SoC clock group SDC and review reports from a clock group workbook.

The script follows the current SoC SDC policy:
  - default synchronous + explicit async/exclusive groups
  - common + scenario assembled view checks
  - domain closure expansion from 01 clock_inventory.csv
"""

import argparse
import csv
import hashlib
import itertools
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

try:
    from dataclasses import dataclass, field
except ImportError:  # pragma: no cover - Python 3.6 compatibility path
    class _CompatField:
        def __init__(self, default_factory=None):
            self.default_factory = default_factory

    def field(default_factory=None):
        return _CompatField(default_factory=default_factory)

    def dataclass(_cls=None, frozen=False):
        def wrap(cls):
            annotations = getattr(cls, "__annotations__", {})
            names = list(annotations.keys())
            defaults = {}
            factories = {}
            for name in names:
                if hasattr(cls, name):
                    value = getattr(cls, name)
                    if isinstance(value, _CompatField):
                        factories[name] = value.default_factory
                        delattr(cls, name)
                    else:
                        defaults[name] = value

            def __init__(self, *args, **kwargs):
                if len(args) > len(names):
                    raise TypeError("too many positional arguments")
                values = {}
                for name, value in zip(names, args):
                    values[name] = value
                for name in names[len(args):]:
                    if name in kwargs:
                        values[name] = kwargs.pop(name)
                    elif name in factories:
                        values[name] = factories[name]()
                    elif name in defaults:
                        values[name] = defaults[name]
                    else:
                        raise TypeError("missing required argument: " + name)
                if kwargs:
                    raise TypeError("unexpected argument: " + sorted(kwargs)[0])
                for name in names:
                    object.__setattr__(self, name, values[name])

            def __eq__(self, other):
                if other.__class__ is not cls:
                    return False
                return all(getattr(self, name) == getattr(other, name) for name in names)

            cls.__init__ = __init__
            cls.__eq__ = __eq__
            if frozen:
                def __hash__(self):
                    return hash(tuple(getattr(self, name) for name in names))
                cls.__hash__ = __hash__
            else:
                cls.__hash__ = None
            return cls

        if _cls is None:
            return wrap
        return wrap(_cls)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - user environment guard
    print("ERROR: openpyxl is required to read/write 03 clock group xlsx files.", file=sys.stderr)
    raise SystemExit(2) from exc


SCENARIOS = {"common", "func", "scan", "mbist", "gpio_in", "gpio_out"}
RELATION_TYPES = {"asynchronous", "logically_exclusive", "physically_exclusive"}
RELATION_TYPE_ALIASES = {
    "async": "asynchronous",
    "asynchronous": "asynchronous",
    "logical_exclusive": "logically_exclusive",
    "logically_exclusive": "logically_exclusive",
    "logically exclusive": "logically_exclusive",
    "physical_exclusive": "physically_exclusive",
    "physically_exclusive": "physically_exclusive",
    "physically exclusive": "physically_exclusive",
}
ANALYSIS_STYLES = {"", "normal", "merged_exclusive", "per_scenario_case", "per_run_case"}
APPLY_VALUES = {"yes", "no"}
REVIEW_STATUS_VALUES = {"draft", "reviewed", "approved", "rejected"}
YES_NO = {"", "yes", "no"}

ACTIVE_01_ACTIONS = {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}
CHECK_ONLY_ACTION = "check_only"

DEFAULT_GROUP_COLUMNS = 8
MAX_NON_CLIQUE_WARNINGS = 50

RULE_SHEET = "clock_group_rules"
DOMAIN_SHEET = "clock_domain_membership"
CANDIDATE_SHEET = "clock_group_candidates"

SCHEMA_VERSION = "1"
DOMAIN_MEMBERSHIP_TYPES = {"seed", "explicit_member", "exclude_descendant"}
TARGET_SYNC_VALUES = {"", "OK", "NEW_FROM_01", "STALE_NOT_IN_01", "BLOCKED_BY_MISSING_SDC"}
TARGET_ANALYSIS_STYLES = {"", "normal", "merged_exclusive", "per_run_case"}

BASE_RULE_HEADERS = [
    "scenario",
    "group_id",
    "relation_type",
]
DOMAIN_HEADERS = [
    "scenario",
    "domain_id",
    "clock_name",
    "membership_type",
    "include_descendants",
    "source_instance",
    "apply",
    "review_status",
    "owner",
    "basis",
    "note",
]
TARGET_DOMAIN_HEADERS = [
    "domain_id",
    "clock_name",
    "membership_type",
    "root_clock",
    "include_descendants",
    "source_instance",
    "apply",
    "review_status",
    "owner",
    "basis",
    "note",
    "sync_status",
    "clock_kind",
    "root_source",
    "source_inventory_digest",
]
TARGET_RULE_HEADERS = [
    "group_id",
    "relation_type",
] + [
    value
    for idx in range(1, DEFAULT_GROUP_COLUMNS + 1)
    for value in (f"group_{idx}_domains", f"group_{idx}_clocks")
] + [
    "exclude_descendant_clocks",
    "analysis_style",
    "apply",
    "review_status",
    "owner",
    "basis",
    "note",
]
TARGET_CANDIDATE_HEADERS = [
    "candidate_id", "candidate_type", "clock_a", "clock_b", "tree_root_a",
    "tree_root_b", "reason", "evidence", "recommended_action", "decision", "note",
]
TRAILING_RULE_HEADERS = [
    "exclude_descendant_clocks",
    "analysis_style",
    "apply",
    "review_status",
    "owner",
    "basis",
    "cdc_required",
    "note",
]
CANDIDATE_HEADERS = [
    "candidate_id",
    "candidate_type",
    "clock_a",
    "clock_b",
    "tree_root_a",
    "tree_root_b",
    "root_source_a",
    "root_source_b",
    "evidence",
    "suggested_relation",
    "decision",
    "target_group_id",
    "note",
]

HEADER_FILL = PatternFill("solid", fgColor="215967")
TITLE_FILL = PatternFill("solid", fgColor="335C81")
SUBTITLE_FILL = PatternFill("solid", fgColor="EAF3F6")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
BLOCKED_FILL = PatternFill("solid", fgColor="FCE5CD")
NEW_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C6CC"),
    right=Side(style="thin", color="B8C6CC"),
    top=Side(style="thin", color="B8C6CC"),
    bottom=Side(style="thin", color="B8C6CC"),
)


@dataclass
class ClockInfo:
    clock_name: str
    clock_kind: str = ""
    root_source: str = ""
    direct_source: str = ""
    final_action: str = ""
    inst_name: str = ""
    port_name: str = ""
    direction: str = ""
    from_whom: str = ""


@dataclass
class Inventory:
    active: Dict[str, ClockInfo] = field(default_factory=dict)
    all_records: List[ClockInfo] = field(default_factory=list)
    object_to_clock: Dict[str, str] = field(default_factory=dict)
    check_by_object: Dict[str, ClockInfo] = field(default_factory=dict)
    children: Dict[str, Set[str]] = field(default_factory=lambda: defaultdict(set))
    parent: Dict[str, str] = field(default_factory=dict)
    tree_root: Dict[str, str] = field(default_factory=dict)


@dataclass
class InventoryContext:
    inventory: Inventory
    path: Path
    scenario: str = ""
    completeness: str = "complete"
    missing_instances: Tuple[str, ...] = ()
    available_harden_count: int = 0
    missing_harden_count: int = 0
    not_required_harden_count: int = 0
    meta_path: Optional[Path] = None
    final_sdc_path: Optional[Path] = None
    final_sdc_digest: str = ""
    inventory_digest: str = ""
    clock_universe_digest: str = ""
    run_id: str = ""
    mode_label: str = ""
    design_revision: str = ""
    structure_digest: str = ""
    completion_path: Optional[Path] = None
    completion_digest: str = ""


@dataclass
class TargetRunContext:
    run_id: str
    mode_label: str
    design_revision: str = ""
    note: str = ""


@dataclass
class DomainMemberRow:
    row_idx: int
    scenario: str
    domain_id: str
    clock_name: str
    membership_type: str
    include_descendants: str
    source_instance: str
    apply: str
    review_status: str
    owner: str
    basis: str
    note: str


@dataclass
class ClockDomain:
    domain_id: str
    members: List[str] = field(default_factory=list)
    closure_members: Set[str] = field(default_factory=set)
    blocked_instances: Set[str] = field(default_factory=set)
    source_rows: List[int] = field(default_factory=list)


@dataclass
class ParsedRule:
    row_idx: int
    scenario: str
    group_id: str
    relation_type: str
    analysis_style: str
    apply: str
    review_status: str
    owner: str
    basis: str
    cdc_required: str
    note: str
    explicit_groups: List[List[str]]
    explicit_domain_groups: List[List[str]]
    group_indices: List[int]
    excluded: Set[str]
    effective_groups: List[List[str]] = field(default_factory=list)
    auto_added_by_group: List[List[str]] = field(default_factory=list)
    blocked_instances: Set[str] = field(default_factory=set)


@dataclass(frozen=True)
class PairOccurrence:
    relation_type: str
    group_id: str
    scenario: str
    group_a: str
    group_b: str


class Report:
    def __init__(self) -> None:
        self.lines: List[str] = []
        self.warning_count = 0
        self.error_count = 0
        self.sync_changed = False

    def info(self, msg: str) -> None:
        self.lines.append(f"INFO: {msg}")

    def warn(self, msg: str) -> None:
        self.warning_count += 1
        self.lines.append(f"WARNING: {msg}")

    def error(self, msg: str) -> None:
        self.error_count += 1
        self.lines.append(f"ERROR: {msg}")


def _author_part_a() -> str:
    return chr(72) + chr(111)


def _author_part_b() -> str:
    return chr(119) + chr(97)


def _author_part_c() -> str:
    return chr(114) + chr(100)


def author_name() -> str:
    return _author_part_a() + _author_part_b() + _author_part_c()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        while True:
            chunk = file_obj.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def sha256_clock_set(clock_names: Iterable[str]) -> str:
    payload = "\n".join(sorted(clock_names)).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def sha256_payload(payload) -> str:
    text = json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def atomic_write_text(path: Path, content: str, encoding: str = "utf-8") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(".%s.tmp.%s" % (path.name, os.getpid()))
    try:
        tmp_path.write_text(content, encoding=encoding)
        os.replace(str(tmp_path), str(path))
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def atomic_save_workbook(wb, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(".%s.tmp.%s.xlsx" % (path.stem, os.getpid()))
    try:
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(path))
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def atomic_write_csv(path: Path, fieldnames: Sequence[str], rows: Sequence[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(".%s.tmp.%s" % (path.name, os.getpid()))
    try:
        with tmp_path.open("w", encoding="utf-8", newline="") as file_obj:
            writer = csv.DictWriter(file_obj, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        os.replace(str(tmp_path), str(path))
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def clean_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def normalize_key(value) -> str:
    return clean_cell(value).strip().lower()


def canonical_relation_type(value) -> str:
    key = normalize_key(value).replace("-", "_")
    return RELATION_TYPE_ALIASES.get(key, key)


def safe_filename_token(value: str) -> str:
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-")
    token = "".join(char if char in allowed else "_" for char in clean_cell(value))
    return token or "unknown"


def tcl_obj_list(names: Sequence[str]) -> str:
    return "{" + " ".join(names) + "}"


def get_clocks(names: Sequence[str]) -> str:
    return f"[get_clocks {tcl_obj_list(names)}]"


def output_sdc_path(cwd: Path, scenario: str) -> Path:
    if scenario == "common":
        return cwd / "common/03_soc_clock_groups.sdc"
    return cwd / f"scenarios/{scenario}_clock_groups.sdc"


def parse_clock_list(value) -> List[str]:
    text = clean_cell(value)
    if not text:
        return []
    text = text.replace("{", " ").replace("}", " ")
    tokens = [token.strip() for token in re.split(r"[\s,;]+", text) if token.strip()]
    result: List[str] = []
    seen: Set[str] = set()
    for token in tokens:
        if token not in seen:
            result.append(token)
            seen.add(token)
    return result


def clock_object(info: ClockInfo) -> str:
    if info.inst_name and info.port_name:
        return f"{info.inst_name}/{info.port_name}"
    return info.direct_source


def from_whom_to_object(value: str) -> str:
    text = clean_cell(value)
    if not text or text.startswith("top."):
        return text.replace(".", "/", 1) if text else ""
    if "." in text:
        inst, port = text.split(".", 1)
        return f"{inst}/{port}"
    return text


def read_clock_inventory(path: Path, report: Report) -> Inventory:
    if not path.is_file():
        raise RuntimeError(f"01 clock inventory not found: {path}")

    inventory = Inventory()
    duplicates: List[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        if not reader.fieldnames or "clock_name" not in reader.fieldnames:
            raise RuntimeError(f"{path} does not contain a clock_name column")
        for row_idx, row in enumerate(reader, start=2):
            clock_name = clean_cell(row.get("clock_name"))
            final_action = clean_cell(row.get("final_action"))
            if not clock_name:
                report.warn(f"{path.name} row {row_idx}: record has empty clock_name")
                continue
            info = ClockInfo(
                clock_name=clock_name,
                clock_kind=clean_cell(row.get("clock_kind")),
                root_source=clean_cell(row.get("root_source")),
                direct_source=clean_cell(row.get("direct_source")),
                final_action=final_action,
                inst_name=clean_cell(row.get("inst_name")),
                port_name=clean_cell(row.get("port_name")),
                direction=clean_cell(row.get("direction")),
                from_whom=clean_cell(row.get("from_whom")),
            )
            inventory.all_records.append(info)
            obj = clock_object(info)
            if final_action == CHECK_ONLY_ACTION and obj:
                inventory.check_by_object[obj] = info
            if final_action not in ACTIVE_01_ACTIONS:
                continue
            if clock_name in inventory.active:
                duplicates.append(clock_name)
                continue
            inventory.active[clock_name] = info
            for alias in active_clock_aliases(info):
                inventory.object_to_clock.setdefault(alias, clock_name)

    if duplicates:
        report.error("duplicate active clock_name in 01 inventory: " + ", ".join(sorted(set(duplicates))))
    build_clock_genealogy(inventory)
    report.info(f"loaded {len(inventory.active)} active clock(s) from {path}")
    return inventory


def read_active_inventory_digests(path: Path) -> Set[str]:
    digests: Set[str] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        for row in reader:
            if clean_cell(row.get("final_action")) not in ACTIVE_01_ACTIONS:
                continue
            value = clean_cell(row.get("final_sdc_digest"))
            if value:
                digests.add(value)
    return digests


def extract_clock_names_from_sdc(path: Path) -> Set[str]:
    logical_lines: List[str] = []
    pending = ""
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.split("#", 1)[0].rstrip()
        if not line and not pending:
            continue
        if line.endswith("\\"):
            pending += line[:-1] + " "
            continue
        logical_lines.append(pending + line)
        pending = ""
    if pending:
        logical_lines.append(pending)

    names: Set[str] = set()
    name_pattern = re.compile(r'(?:^|\s)-name\s+(?:\{([^}]*)\}|"([^"]*)"|(\S+))')
    for line in logical_lines:
        stripped = line.lstrip()
        if not (stripped.startswith("create_clock ") or stripped.startswith("create_generated_clock ")):
            continue
        match = name_pattern.search(stripped)
        if match:
            names.add(next(group for group in match.groups() if group is not None))
    return names


def load_inventory_context(
    inventory_path: Path,
    report: Report,
    expected_scenario: str,
    meta_path: Optional[Path] = None,
    clock_sdc_path: Optional[Path] = None,
    require_meta: bool = False,
) -> InventoryContext:
    inventory = read_clock_inventory(inventory_path, report)
    context = InventoryContext(
        inventory=inventory,
        path=inventory_path,
        scenario=expected_scenario,
        inventory_digest=sha256_file(inventory_path),
        clock_universe_digest=sha256_clock_set(inventory.active),
    )

    if meta_path is not None and meta_path.is_file():
        try:
            payload = json.loads(meta_path.read_text(encoding="utf-8"))
        except (ValueError, OSError) as exc:
            raise RuntimeError(f"invalid 01 clock inventory meta {meta_path}: {exc}")
        context.meta_path = meta_path
        context.scenario = clean_cell(payload.get("scenario"))
        context.completeness = normalize_key(payload.get("run_completeness")) or "complete"
        context.missing_instances = tuple(
            sorted(clean_cell(item) for item in payload.get("missing_instances", []) if clean_cell(item))
        )
        context.available_harden_count = int(payload.get("available_harden_count", 0) or 0)
        context.missing_harden_count = int(payload.get("missing_harden_count", 0) or 0)
        context.not_required_harden_count = int(payload.get("not_required_harden_count", 0) or 0)
        final_path_text = clean_cell(payload.get("final_sdc_path"))
        if final_path_text:
            context.final_sdc_path = Path(final_path_text)
        context.final_sdc_digest = clean_cell(payload.get("final_sdc_digest"))

        if context.scenario != expected_scenario:
            report.error(
                f"01 inventory meta scenario mismatch: expected {expected_scenario}, "
                f"got {context.scenario or '<blank>'}"
            )
        expected_inventory_digest = clean_cell(payload.get("inventory_digest"))
        if expected_inventory_digest != context.inventory_digest:
            report.error(
                f"stale 01 inventory: meta digest {expected_inventory_digest or '<blank>'} "
                f"does not match {context.inventory_digest}"
            )
        expected_clock_digest = clean_cell(payload.get("clock_set_digest"))
        if expected_clock_digest != context.clock_universe_digest:
            report.error(
                f"stale 01 inventory: clock set digest {expected_clock_digest or '<blank>'} "
                f"does not match {context.clock_universe_digest}"
            )
        expected_count = payload.get("clock_count")
        if expected_count is not None:
            try:
                clock_count = int(expected_count)
            except (TypeError, ValueError):
                report.error(f"01 inventory meta has invalid clock_count: {expected_count}")
            else:
                if clock_count != len(inventory.active):
                    report.error(f"01 inventory meta clock_count={clock_count}, actual={len(inventory.active)}")
    elif require_meta:
        raise RuntimeError(f"01 assembled clock inventory meta not found: {meta_path}")
    elif meta_path is not None:
        report.warn(f"01 inventory meta not found; legacy validation only: {meta_path}")

    effective_sdc_path = clock_sdc_path or context.final_sdc_path
    if effective_sdc_path is not None:
        if not effective_sdc_path.is_file():
            raise RuntimeError(f"01 final clock SDC not found: {effective_sdc_path}")
        actual_sdc_digest = sha256_file(effective_sdc_path)
        if context.final_sdc_digest and context.final_sdc_digest != actual_sdc_digest:
            report.error(
                f"stale 01 final SDC: meta digest {context.final_sdc_digest} does not match {actual_sdc_digest}"
            )
        inventory_digests = read_active_inventory_digests(inventory_path)
        if inventory_digests and inventory_digests != {actual_sdc_digest}:
            report.error(
                "stale 01 inventory: active final_sdc_digest value(s) do not match final SDC digest: "
                + ", ".join(sorted(inventory_digests))
            )
        sdc_clock_names = extract_clock_names_from_sdc(effective_sdc_path)
        if sdc_clock_names != set(inventory.active):
            missing = sorted(set(inventory.active) - sdc_clock_names)
            extra = sorted(sdc_clock_names - set(inventory.active))
            report.error(
                f"01 final SDC clock set differs from inventory; missing_in_sdc={missing}; extra_in_sdc={extra}"
            )
        context.final_sdc_path = effective_sdc_path
        context.final_sdc_digest = actual_sdc_digest
        report.info(
            f"verified 01 final SDC digest and {len(sdc_clock_names)} clock name(s): {effective_sdc_path}"
        )
    elif require_meta:
        report.error("01 assembled meta does not provide final_sdc_path")
    else:
        report.warn("01 final clock SDC was not found; legacy inventory authenticity check was skipped")

    if context.completeness not in {"complete", "partial"}:
        report.error(f"01 inventory meta has invalid run_completeness: {context.completeness}")
    report.info(
        f"01 run completeness: {context.completeness}; missing instances: "
        f"{', '.join(context.missing_instances) or '<none>'}"
    )
    return context


def active_clock_aliases(info: ClockInfo) -> List[str]:
    aliases: List[str] = [info.clock_name]
    obj = clock_object(info)
    if obj:
        aliases.append(obj)

    # For top/virtual clocks, the direct_source is the clock object or an alias
    # of the SoC top source. For generated output clocks, direct_source is the
    # parent source pin, so it must not alias to the generated output clock.
    if info.final_action in {"emit_top_clock", "emit_virtual_clock"} and info.direct_source:
        aliases.append(info.direct_source)
    if info.final_action in {"emit_top_clock", "emit_virtual_clock"} and info.root_source == info.direct_source:
        aliases.append(info.root_source)
    return [alias for idx, alias in enumerate(aliases) if alias and alias not in aliases[:idx]]


def resolve_source_clock(source_obj: str, inventory: Inventory, seen: Optional[Set[str]] = None) -> str:
    if not source_obj:
        return ""
    if seen is None:
        seen = set()
    if source_obj in seen:
        return ""
    seen.add(source_obj)

    direct = inventory.object_to_clock.get(source_obj)
    if direct:
        return direct

    check = inventory.check_by_object.get(source_obj)
    if not check:
        return ""

    upstream_obj = from_whom_to_object(check.from_whom)
    if upstream_obj:
        upstream_clock = resolve_source_clock(upstream_obj, inventory, seen)
        if upstream_clock:
            return upstream_clock

    if check.root_source:
        root_clock = resolve_source_clock(check.root_source, inventory, seen)
        if root_clock:
            return root_clock
    return ""


def build_clock_genealogy(inventory: Inventory) -> None:
    for clock_name, info in inventory.active.items():
        if "generated" not in normalize_key(info.clock_kind):
            continue
        parent = resolve_source_clock(info.direct_source, inventory)
        if parent and parent != clock_name:
            inventory.parent[clock_name] = parent
            inventory.children[parent].add(clock_name)
    for clock_name in inventory.active:
        inventory.tree_root[clock_name] = tree_root_of(clock_name, inventory)


def tree_root_of(clock_name: str, inventory: Inventory) -> str:
    if not clock_name:
        return ""
    seen: Set[str] = set()
    current = clock_name
    while current and current not in seen:
        seen.add(current)
        parent = inventory.parent.get(current)
        if not parent:
            return current
        current = parent
    return current or clock_name


def descendants_of(clock_name: str, inventory: Inventory) -> List[str]:
    result: List[str] = []
    seen = {clock_name}
    stack = sorted(inventory.children.get(clock_name, set()))
    while stack:
        child = stack.pop(0)
        if child in seen:
            continue
        seen.add(child)
        result.append(child)
        for grandchild in sorted(inventory.children.get(child, set())):
            if grandchild not in seen:
                stack.append(grandchild)
    return result


def style_title(ws, title: str, subtitle: str, width_cols: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width_cols)
    ws.cell(1, 1, title)
    ws.cell(1, 1).fill = TITLE_FILL
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=14)
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width_cols)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).fill = SUBTITLE_FILL
    ws.cell(2, 1).font = Font(color="5C6670")
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 36


def write_header(ws, headers: Sequence[str], row: int = 4) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN_BORDER
    ws.freeze_panes = f"A{row + 1}"


def set_widths(ws, widths: Sequence[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_validation(ws, cell_range: str, values: Sequence[str]) -> None:
    quoted = ",".join(values)
    dv = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def ensure_table(ws, name: str, ref: str) -> None:
    for table in ws.tables.values():
        if table.name == name:
            table.ref = ref
            return
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_body_row(ws, row_idx: int, col_count: int, fill: Optional[PatternFill] = None) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row_idx, col_idx)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill is not None:
            cell.fill = fill


def default_rule_headers() -> List[str]:
    groups: List[str] = []
    for idx in range(1, DEFAULT_GROUP_COLUMNS + 1):
        groups.extend([f"group_{idx}_domains", f"group_{idx}_clocks"])
    return BASE_RULE_HEADERS + groups + TRAILING_RULE_HEADERS


def setup_domain_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(DOMAIN_SHEET, 0)
    style_title(
        ws,
        "03 SoC Clock Groups - clock_domain_membership",
        (
            "Authoritative human-reviewed clock-to-domain membership. "
            "Manual clocks use explicit_member; no relation is inferred from names."
        ),
        len(DOMAIN_HEADERS),
    )
    write_header(ws, DOMAIN_HEADERS)
    set_widths(ws, [12, 24, 30, 22, 20, 24, 10, 16, 18, 48, 42])
    add_validation(ws, "A5:A5000", sorted(SCENARIOS))
    add_validation(ws, "D5:D5000", sorted(DOMAIN_MEMBERSHIP_TYPES))
    add_validation(ws, "E5:E5000", ["yes", "no"])
    add_validation(ws, "G5:G5000", ["yes", "no"])
    add_validation(ws, "H5:H5000", ["draft", "reviewed", "approved", "rejected"])
    ensure_table(ws, "ClockDomainMembership", f"A4:{get_column_letter(len(DOMAIN_HEADERS))}5")


def setup_rules_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(RULE_SHEET)
    headers = default_rule_headers()
    style_title(
        ws,
        "03 SoC Clock Groups - clock_group_rules",
        (
            "Default policy: synchronous unless explicitly listed here. "
            "Only apply=yes and review_status=approved rows generate SDC."
        ),
        len(headers),
    )
    write_header(ws, headers)
    widths = [12, 28, 22] + [22, 28] * DEFAULT_GROUP_COLUMNS + [28, 20, 10, 16, 18, 42, 14, 42]
    set_widths(ws, widths)
    add_validation(ws, "A5:A5000", sorted(SCENARIOS))
    add_validation(ws, "C5:C5000", sorted(RELATION_TYPES))
    analysis_col = get_column_letter(headers.index("analysis_style") + 1)
    apply_col = get_column_letter(headers.index("apply") + 1)
    review_col = get_column_letter(headers.index("review_status") + 1)
    cdc_col = get_column_letter(headers.index("cdc_required") + 1)
    add_validation(ws, f"{analysis_col}5:{analysis_col}5000", ["normal", "merged_exclusive", "per_scenario_case"])
    add_validation(ws, f"{apply_col}5:{apply_col}5000", ["yes", "no"])
    add_validation(ws, f"{review_col}5:{review_col}5000", ["draft", "reviewed", "approved", "rejected"])
    add_validation(ws, f"{cdc_col}5:{cdc_col}5000", ["yes", "no"])
    ensure_table(ws, "ClockGroupRules", f"A4:{get_column_letter(len(headers))}5")


def setup_candidates_sheet(wb: Workbook, inventory: Inventory, max_pairs: int) -> None:
    ws = wb.create_sheet(CANDIDATE_SHEET)
    style_title(
        ws,
        "03 SoC Clock Groups - clock_group_candidates",
        "Candidate rows are for review only. They never generate SDC until moved into clock_group_rules.",
        len(CANDIDATE_HEADERS),
    )
    write_header(ws, CANDIDATE_HEADERS)
    set_widths(ws, [18, 24, 24, 24, 28, 28, 28, 28, 46, 22, 16, 22, 42])
    active = sorted(inventory.active.values(), key=lambda item: item.clock_name)
    row_idx = 5
    candidate_idx = 1
    for a, b in itertools.combinations(active, 2):
        tree_a = inventory.tree_root.get(a.clock_name, a.clock_name)
        tree_b = inventory.tree_root.get(b.clock_name, b.clock_name)
        if not tree_a or not tree_b or tree_a == tree_b:
            continue
        ws.cell(row_idx, 1, f"CG_CAND_{candidate_idx:04d}")
        ws.cell(row_idx, 2, "cross_tree_pair")
        ws.cell(row_idx, 3, a.clock_name)
        ws.cell(row_idx, 4, b.clock_name)
        ws.cell(row_idx, 5, tree_a)
        ws.cell(row_idx, 6, tree_b)
        ws.cell(row_idx, 7, a.root_source)
        ws.cell(row_idx, 8, b.root_source)
        ws.cell(row_idx, 9, "genealogy tree_root differs; review CDC/STA architecture before deciding")
        ws.cell(row_idx, 10, "")
        ws.cell(row_idx, 11, "draft")
        style_body_row(ws, row_idx, len(CANDIDATE_HEADERS))
        row_idx += 1
        candidate_idx += 1
        if candidate_idx > max_pairs:
            break
    if candidate_idx > max_pairs:
        ws.cell(row_idx, 1, "INFO")
        ws.cell(row_idx, len(CANDIDATE_HEADERS), f"candidate list capped at {max_pairs} cross-tree pairs")
        style_body_row(ws, row_idx, len(CANDIDATE_HEADERS), WARNING_FILL)
        row_idx += 1
    last_row = max(row_idx - 1, 5)
    ensure_table(ws, "ClockGroupCandidates", f"A4:{get_column_letter(len(CANDIDATE_HEADERS))}{last_row}")


def create_new_workbook(form_path: Path, inventory: Inventory, max_candidate_pairs: int) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    setup_domain_sheet(wb)
    setup_rules_sheet(wb)
    setup_candidates_sheet(wb, inventory, max_candidate_pairs)
    atomic_save_workbook(wb, form_path)


def find_header_row(ws, required: str) -> Tuple[int, Dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        mapping: Dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            value = clean_cell(ws.cell(row_idx, col_idx).value)
            if value:
                mapping[value] = col_idx
        if required in mapping:
            return row_idx, mapping
    raise RuntimeError(f"sheet {ws.title} does not contain a {required} header")


def group_indices(mapping: Dict[str, int]) -> List[int]:
    groups: Set[int] = set()
    for header in mapping:
        match = re.fullmatch(r"group_(\d+)_(?:domains|clocks)", header)
        if match:
            groups.add(int(match.group(1)))
    return sorted(groups)


def ensure_rule_headers(ws) -> Tuple[int, Dict[str, int], bool]:
    header_row, mapping = find_header_row(ws, "group_id")
    changed = False
    if not group_indices(mapping):
        for idx in range(1, DEFAULT_GROUP_COLUMNS + 1):
            for suffix in ("domains", "clocks"):
                header = f"group_{idx}_{suffix}"
                ws.cell(header_row, ws.max_column + 1, header)
                changed = True
        header_row, mapping = find_header_row(ws, "group_id")
    for idx in group_indices(mapping):
        domain_header = f"group_{idx}_domains"
        clock_header = f"group_{idx}_clocks"
        if domain_header not in mapping:
            ws.cell(header_row, ws.max_column + 1, domain_header)
            changed = True
        if clock_header not in mapping:
            ws.cell(header_row, ws.max_column + 1, clock_header)
            changed = True
    for header in BASE_RULE_HEADERS + TRAILING_RULE_HEADERS:
        if header not in mapping:
            ws.cell(header_row, ws.max_column + 1, header)
            changed = True
    header_row, mapping = find_header_row(ws, "group_id")
    headers = [clean_cell(ws.cell(header_row, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
    write_header(ws, headers, header_row)
    return header_row, mapping, changed


def ensure_domain_sheet(wb: Workbook) -> bool:
    if DOMAIN_SHEET not in wb.sheetnames:
        setup_domain_sheet(wb)
        return True
    ws = wb[DOMAIN_SHEET]
    header_row, mapping = find_header_row(ws, "domain_id")
    changed = False
    for header in DOMAIN_HEADERS:
        if header not in mapping:
            ws.cell(header_row, ws.max_column + 1, header)
            changed = True
    header_row, _ = find_header_row(ws, "domain_id")
    headers = [clean_cell(ws.cell(header_row, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
    write_header(ws, headers, header_row)
    return changed


def rewrite_relation_type_cells(ws, header_row: int, mapping: Dict[str, int], report: Report) -> bool:
    relation_col = mapping.get("relation_type")
    if not relation_col:
        return False
    changed = False
    for row_idx in range(header_row + 1, ws.max_row + 1):
        original = clean_cell(ws.cell(row_idx, relation_col).value)
        if not original:
            continue
        canonical = canonical_relation_type(original)
        if canonical not in RELATION_TYPES or original == canonical:
            continue
        ws.cell(row_idx, relation_col, canonical)
        changed = True
        report.info(
            f"clock_group_rules row {row_idx}: relation_type {original} normalized to {canonical}"
        )
    return changed


def read_rule_rows(ws, header_row: int, mapping: Dict[str, int]) -> List[ParsedRule]:
    groups = group_indices(mapping)
    rows: List[ParsedRule] = []
    relevant_cols = list(mapping.values())
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if not any(clean_cell(ws.cell(row_idx, col_idx).value) for col_idx in relevant_cols):
            continue
        explicit_groups: List[List[str]] = []
        explicit_domain_groups: List[List[str]] = []
        used_group_indices: List[int] = []
        for group_idx in groups:
            clock_header = f"group_{group_idx}_clocks"
            domain_header = f"group_{group_idx}_domains"
            clocks = parse_clock_list(ws.cell(row_idx, mapping.get(clock_header, 0)).value)
            domains = parse_clock_list(ws.cell(row_idx, mapping.get(domain_header, 0)).value)
            if not clocks and not domains:
                continue
            explicit_groups.append(clocks)
            explicit_domain_groups.append(domains)
            used_group_indices.append(group_idx)
        rows.append(
            ParsedRule(
                row_idx=row_idx,
                scenario=normalize_key(ws.cell(row_idx, mapping.get("scenario", 0)).value),
                group_id=clean_cell(ws.cell(row_idx, mapping.get("group_id", 0)).value),
                relation_type=canonical_relation_type(ws.cell(row_idx, mapping.get("relation_type", 0)).value),
                analysis_style=normalize_key(ws.cell(row_idx, mapping.get("analysis_style", 0)).value),
                apply=normalize_key(ws.cell(row_idx, mapping.get("apply", 0)).value),
                review_status=normalize_key(ws.cell(row_idx, mapping.get("review_status", 0)).value),
                owner=clean_cell(ws.cell(row_idx, mapping.get("owner", 0)).value),
                basis=clean_cell(ws.cell(row_idx, mapping.get("basis", 0)).value),
                cdc_required=normalize_key(ws.cell(row_idx, mapping.get("cdc_required", 0)).value),
                note=clean_cell(ws.cell(row_idx, mapping.get("note", 0)).value),
                explicit_groups=explicit_groups,
                explicit_domain_groups=explicit_domain_groups,
                group_indices=used_group_indices,
                excluded=set(parse_clock_list(ws.cell(row_idx, mapping.get("exclude_descendant_clocks", 0)).value)),
            )
        )
    return rows


def read_domain_rows(ws) -> List[DomainMemberRow]:
    header_row, mapping = find_header_row(ws, "domain_id")
    rows: List[DomainMemberRow] = []
    relevant_cols = list(mapping.values())
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if not any(clean_cell(ws.cell(row_idx, col_idx).value) for col_idx in relevant_cols):
            continue
        rows.append(
            DomainMemberRow(
                row_idx=row_idx,
                scenario=normalize_key(ws.cell(row_idx, mapping.get("scenario", 0)).value),
                domain_id=clean_cell(ws.cell(row_idx, mapping.get("domain_id", 0)).value),
                clock_name=clean_cell(ws.cell(row_idx, mapping.get("clock_name", 0)).value),
                membership_type=normalize_key(ws.cell(row_idx, mapping.get("membership_type", 0)).value),
                include_descendants=normalize_key(ws.cell(row_idx, mapping.get("include_descendants", 0)).value),
                source_instance=clean_cell(ws.cell(row_idx, mapping.get("source_instance", 0)).value),
                apply=normalize_key(ws.cell(row_idx, mapping.get("apply", 0)).value),
                review_status=normalize_key(ws.cell(row_idx, mapping.get("review_status", 0)).value),
                owner=clean_cell(ws.cell(row_idx, mapping.get("owner", 0)).value),
                basis=clean_cell(ws.cell(row_idx, mapping.get("basis", 0)).value),
                note=clean_cell(ws.cell(row_idx, mapping.get("note", 0)).value),
            )
        )
    return rows


def validate_and_build_domains(
    rows: Sequence[DomainMemberRow],
    scenario: Optional[str],
    inventory_context: InventoryContext,
    report: Report,
) -> Dict[str, ClockDomain]:
    inventory = inventory_context.inventory
    active_clock_names = set(inventory.active)
    missing_instances = set(inventory_context.missing_instances)
    active_rows: List[DomainMemberRow] = []

    for row in rows:
        if row.scenario and row.scenario not in SCENARIOS:
            report.error(f"{DOMAIN_SHEET} row {row.row_idx}: invalid scenario {row.scenario}")
        if row.membership_type and row.membership_type not in DOMAIN_MEMBERSHIP_TYPES:
            report.error(
                f"{DOMAIN_SHEET} row {row.row_idx}: invalid membership_type {row.membership_type}"
            )
        if row.include_descendants and row.include_descendants not in {"yes", "no"}:
            report.error(
                f"{DOMAIN_SHEET} row {row.row_idx}: include_descendants must be yes/no, "
                f"got {row.include_descendants}"
            )
        if row.apply and row.apply not in APPLY_VALUES:
            report.error(f"{DOMAIN_SHEET} row {row.row_idx}: apply must be yes/no, got {row.apply}")
        if row.review_status and row.review_status not in REVIEW_STATUS_VALUES:
            report.error(
                f"{DOMAIN_SHEET} row {row.row_idx}: invalid review_status {row.review_status}"
            )
        if row.apply == "yes" and row.review_status != "approved":
            report.error(
                f"{DOMAIN_SHEET} row {row.row_idx}: apply=yes requires review_status=approved"
            )
        if not (row.apply == "yes" and row.review_status == "approved"):
            continue
        if row.scenario not in {"common", scenario}:
            continue
        active_rows.append(row)
        if not row.domain_id:
            report.error(f"{DOMAIN_SHEET} row {row.row_idx}: active membership has empty domain_id")
        if not row.clock_name:
            report.error(f"{DOMAIN_SHEET} row {row.row_idx}: active membership has empty clock_name")
        if not row.membership_type:
            report.error(f"{DOMAIN_SHEET} row {row.row_idx}: active membership has empty membership_type")
        if not row.include_descendants:
            report.error(f"{DOMAIN_SHEET} row {row.row_idx}: active membership requires include_descendants")
        if not (row.basis or row.note):
            report.error(f"{DOMAIN_SHEET} row {row.row_idx}: active membership requires basis or note")
        if row.membership_type == "exclude_descendant" and row.include_descendants == "yes":
            report.error(
                f"{DOMAIN_SHEET} row {row.row_idx}: exclude_descendant must set include_descendants=no"
            )

    domains: Dict[str, ClockDomain] = {}
    include_rows: Dict[str, List[DomainMemberRow]] = defaultdict(list)
    exclude_rows: Dict[str, List[DomainMemberRow]] = defaultdict(list)
    membership_rows_by_clock: Dict[str, int] = {}
    for row in active_rows:
        previous_row = membership_rows_by_clock.get(row.clock_name)
        if previous_row is not None:
            report.error(
                f"{DOMAIN_SHEET} duplicate active membership for clock {row.clock_name} "
                f"at rows {previous_row} and {row.row_idx}"
            )
        else:
            membership_rows_by_clock[row.clock_name] = row.row_idx
        domain = domains.setdefault(row.domain_id, ClockDomain(domain_id=row.domain_id))
        domain.source_rows.append(row.row_idx)
        if row.clock_name not in active_clock_names:
            if (
                inventory_context.completeness == "partial"
                and row.source_instance
                and row.source_instance in missing_instances
            ):
                domain.blocked_instances.add(row.source_instance)
                report.warn(
                    f"{DOMAIN_SHEET} row {row.row_idx} {row.domain_id}: clock {row.clock_name} "
                    f"blocked_by_missing_sdc for {row.source_instance}"
                )
                continue
            report.error(
                f"{DOMAIN_SHEET} row {row.row_idx} {row.domain_id}: clock not found in active 01 "
                f"clock inventory: {row.clock_name}"
            )
            continue
        if row.membership_type == "exclude_descendant":
            exclude_rows[row.domain_id].append(row)
        else:
            include_rows[row.domain_id].append(row)

    for domain_id, domain in domains.items():
        members: List[str] = []
        closure_pool: Set[str] = set()
        for row in include_rows.get(domain_id, []):
            if row.clock_name not in members:
                members.append(row.clock_name)
            if row.include_descendants == "yes":
                for descendant in descendants_of(row.clock_name, inventory):
                    closure_pool.add(descendant)
                    if descendant not in members:
                        members.append(descendant)
        for row in exclude_rows.get(domain_id, []):
            if row.clock_name not in closure_pool:
                report.warn(
                    f"{DOMAIN_SHEET} row {row.row_idx} {domain_id}: excluded clock is not in domain closure: "
                    f"{row.clock_name}"
                )
            members = [clock for clock in members if clock != row.clock_name]
        domain.members = members
        domain.closure_members = closure_pool
        if not domain.members and not domain.blocked_instances:
            report.error(f"domain {domain_id} has no effective clock members")

    clock_to_domain: Dict[str, str] = {}
    for domain_id, domain in sorted(domains.items()):
        for clock_name in domain.members:
            previous = clock_to_domain.get(clock_name)
            if previous and previous != domain_id:
                report.error(
                    f"clock {clock_name} belongs to multiple active domains in the assembled run: "
                    f"{previous}, {domain_id}"
                )
            clock_to_domain[clock_name] = domain_id

    report.info(
        f"loaded {len(rows)} clock_domain_membership row(s); "
        f"resolved {len(domains)} active domain(s)"
    )
    return domains


def active_rule(rule: ParsedRule) -> bool:
    return rule.apply == "yes" and rule.review_status == "approved"


def assembled_rules(rows: Sequence[ParsedRule], scenario: str) -> List[ParsedRule]:
    result = []
    for rule in rows:
        if not active_rule(rule):
            continue
        if rule.scenario == "common" or rule.scenario == scenario:
            result.append(rule)
    return result


def output_rules(rows: Sequence[ParsedRule], scenario: str) -> List[ParsedRule]:
    return [
        rule
        for rule in rows
        if active_rule(rule) and rule.scenario == scenario and not rule.blocked_instances
    ]


def validate_and_expand_rules(
    rows: Sequence[ParsedRule],
    scenario: Optional[str],
    inventory: Inventory,
    domains: Dict[str, ClockDomain],
    report: Report,
) -> List[ParsedRule]:
    seen_ids: Set[str] = set()
    active_rows: List[ParsedRule] = []
    active_clock_names = set(inventory.active)

    for rule in rows:
        if rule.group_id:
            if rule.group_id in seen_ids:
                report.error(f"clock_group_rules row {rule.row_idx}: duplicate group_id {rule.group_id}")
            seen_ids.add(rule.group_id)

        if rule.scenario and rule.scenario not in SCENARIOS:
            report.error(f"clock_group_rules row {rule.row_idx}: invalid scenario {rule.scenario}")
        if rule.relation_type and rule.relation_type not in RELATION_TYPES:
            report.error(f"clock_group_rules row {rule.row_idx}: invalid relation_type {rule.relation_type}")
        if rule.analysis_style not in ANALYSIS_STYLES:
            report.error(f"clock_group_rules row {rule.row_idx}: invalid analysis_style {rule.analysis_style}")
        if rule.apply and rule.apply not in APPLY_VALUES:
            report.error(f"clock_group_rules row {rule.row_idx}: apply must be yes/no, got {rule.apply}")
        if rule.review_status and rule.review_status not in REVIEW_STATUS_VALUES:
            report.error(
                f"clock_group_rules row {rule.row_idx}: invalid review_status {rule.review_status}"
            )
        if rule.cdc_required not in YES_NO:
            report.error(f"clock_group_rules row {rule.row_idx}: cdc_required must be yes/no, got {rule.cdc_required}")

        if not active_rule(rule):
            if rule.apply == "yes" and rule.review_status != "approved":
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: "
                    "apply=yes requires review_status=approved"
                )
            continue

        if rule.scenario not in {"common", scenario}:
            continue
        active_rows.append(rule)

        if not rule.group_id:
            report.error(f"clock_group_rules row {rule.row_idx}: active rule has empty group_id")
        if not rule.scenario:
            report.error(f"clock_group_rules row {rule.row_idx} {rule.group_id}: active rule has empty scenario")
        if not rule.relation_type:
            report.error(f"clock_group_rules row {rule.row_idx} {rule.group_id}: active rule has empty relation_type")
        if rule.scenario == "common" and not rule.basis:
            report.error(f"clock_group_rules row {rule.row_idx} {rule.group_id}: common rule basis is required")
        if len(rule.explicit_groups) < 2:
            report.error(f"clock_group_rules row {rule.row_idx} {rule.group_id}: at least two non-empty groups required")
        if rule.analysis_style == "per_run_case":
            report.error(
                f"clock_group_rules row {rule.row_idx} {rule.group_id}: "
                "analysis_style=per_run_case records a pre-setup case selection and must not apply=yes"
            )
        if rule.relation_type == "asynchronous" and rule.cdc_required != "yes":
            report.warn(
                f"clock_group_rules row {rule.row_idx} {rule.group_id}: asynchronous rule should set cdc_required=yes"
            )
        if rule.relation_type == "logically_exclusive":
            if not rule.analysis_style:
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: "
                    "logically_exclusive requires explicit analysis_style"
                )
            basis_lower = rule.basis.lower()
            if not any(token in basis_lower for token in ["mux", "select", "mode", "merged", "exclusive"]):
                report.warn(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: logically_exclusive basis should "
                    "mention mux/select/mode evidence"
                )
            if rule.analysis_style == "per_scenario_case":
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: "
                    f"logically_exclusive with analysis_style={rule.analysis_style} must not apply=yes"
                )
            if rule.analysis_style == "merged_exclusive" and not basis_says_no_single_leg_case(rule.basis):
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: merged_exclusive basis should state "
                    "that the view does not single-leg case the mux select; use an explicit phrase such as "
                    "no_case_analysis or without single-leg case analysis"
                )

        expand_rule_groups(rule, active_clock_names, inventory, domains, report)

    report.info(f"loaded {len(rows)} clock_group_rules row(s); {len(active_rows)} active row(s) in assembled view")
    return active_rows


def basis_says_no_single_leg_case(basis: str) -> bool:
    text = " ".join(basis.lower().replace("-", " ").replace("_", " ").split())
    accepted = [
        "no case analysis",
        "without case analysis",
        "without single leg case analysis",
        "not case fixed",
        "mux select is not case fixed",
        "未做 case analysis",
        "未做单腿 case",
        "不做 case analysis",
        "无 case analysis",
    ]
    return any(phrase in text for phrase in accepted)


def expand_rule_groups(
    rule: ParsedRule,
    active_clock_names: Set[str],
    inventory: Inventory,
    domains: Dict[str, ClockDomain],
    report: Report,
) -> None:
    rule.effective_groups = []
    rule.auto_added_by_group = []
    all_explicit: Set[str] = set()
    descendant_pool: Set[str] = set()

    for list_idx, clocks in enumerate(rule.explicit_groups):
        group_idx = rule.group_indices[list_idx]
        domain_ids = rule.explicit_domain_groups[list_idx]
        explicit_seen: Set[str] = set()
        for clock in clocks:
            if clock in explicit_seen:
                report.warn(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: duplicate clock {clock} "
                    f"in group_{group_idx}"
                )
            explicit_seen.add(clock)
            if clock not in active_clock_names:
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: clock not found in active 01 "
                    f"clock inventory: {clock}"
                )
            if clock in all_explicit:
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: clock appears in multiple groups: {clock}"
                )
            all_explicit.add(clock)

        effective: List[str] = []
        auto_added: List[str] = []
        for domain_id in domain_ids:
            domain = domains.get(domain_id)
            if domain is None:
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: unknown domain_id "
                    f"in group_{group_idx}: {domain_id}"
                )
                continue
            if domain.blocked_instances:
                rule.blocked_instances.update(domain.blocked_instances)
            descendant_pool.update(domain.closure_members)
            for member in domain.members:
                if member not in effective:
                    effective.append(member)
        for clock in clocks:
            if clock not in effective:
                effective.append(clock)
            for descendant in descendants_of(clock, inventory):
                descendant_pool.add(descendant)
                if descendant in rule.excluded:
                    continue
                if descendant not in effective:
                    effective.append(descendant)
                    auto_added.append(descendant)

        for excluded in rule.excluded:
            if excluded in clocks:
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: explicit clock cannot be excluded: {excluded}"
                )
            if excluded in active_clock_names and excluded in effective:
                effective = [clock for clock in effective if clock != excluded]

        if not effective and not rule.blocked_instances:
            report.error(
                f"clock_group_rules row {rule.row_idx} {rule.group_id}: group_{group_idx} is empty after exclusions"
            )
        if auto_added:
            report.info(
                f"clock_group_rules row {rule.row_idx} {rule.group_id}: group_{group_idx} auto-added "
                f"descendant clock(s): {', '.join(auto_added)}"
            )
            if rule.relation_type == "logically_exclusive":
                report.warn(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: logically_exclusive group_{group_idx} "
                    f"auto-added descendant clock(s) requiring mux-merge review: {', '.join(auto_added)}"
                )
        rule.effective_groups.append(effective)
        rule.auto_added_by_group.append(auto_added)

    if rule.blocked_instances:
        report.warn(
            f"clock_group_rules row {rule.row_idx} {rule.group_id}: blocked_by_missing_sdc; "
            f"missing instances: {', '.join(sorted(rule.blocked_instances))}"
        )

    for excluded in sorted(rule.excluded):
        if excluded not in active_clock_names:
            report.error(
                f"clock_group_rules row {rule.row_idx} {rule.group_id}: excluded clock not found in active 01 "
                f"clock inventory: {excluded}"
            )
        elif excluded not in descendant_pool:
            report.warn(
                f"clock_group_rules row {rule.row_idx} {rule.group_id}: excluded clock is not a descendant of "
                f"any explicit group clock: {excluded}"
            )
        if excluded and not (rule.basis or rule.note):
            report.error(
                f"clock_group_rules row {rule.row_idx} {rule.group_id}: exclude_descendant_clocks requires "
                "basis or note"
            )

    seen_effective: Dict[str, int] = {}
    for list_idx, group in enumerate(rule.effective_groups):
        group_idx = rule.group_indices[list_idx]
        for clock in group:
            previous_group = seen_effective.get(clock)
            if previous_group is not None and previous_group != group_idx:
                report.error(
                    f"clock_group_rules row {rule.row_idx} {rule.group_id}: effective clock {clock} "
                    f"appears in both group_{previous_group} and group_{group_idx}"
                )
            seen_effective[clock] = group_idx


def pair_key(clock_a: str, clock_b: str) -> Tuple[str, str]:
    return tuple(sorted((clock_a, clock_b)))


def build_pair_maps(
    rules: Sequence[ParsedRule],
) -> Tuple[Dict[Tuple[str, str], List[PairOccurrence]], Dict[Tuple[str, str], List[str]], Dict[str, List[str]]]:
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]] = defaultdict(list)
    same_group_pairs: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    participation: Dict[str, List[str]] = defaultdict(list)

    for rule in rules:
        if rule.blocked_instances:
            continue
        for list_idx, group in enumerate(rule.effective_groups):
            group_idx = rule.group_indices[list_idx]
            label = f"{rule.group_id}:group_{group_idx}:{rule.relation_type}:{rule.scenario}"
            for clock in group:
                participation[clock].append(label)
            for a, b in itertools.combinations(group, 2):
                same_group_pairs[pair_key(a, b)].append(label)

        for idx_a in range(len(rule.effective_groups)):
            for idx_b in range(idx_a + 1, len(rule.effective_groups)):
                group_a = rule.effective_groups[idx_a]
                group_b = rule.effective_groups[idx_b]
                occurrence = PairOccurrence(
                    relation_type=rule.relation_type,
                    group_id=rule.group_id,
                    scenario=rule.scenario,
                    group_a=f"group_{rule.group_indices[idx_a]}",
                    group_b=f"group_{rule.group_indices[idx_b]}",
                )
                for clock_a in group_a:
                    for clock_b in group_b:
                        if clock_a == clock_b:
                            continue
                        relation_pairs[pair_key(clock_a, clock_b)].append(occurrence)
    return relation_pairs, same_group_pairs, participation


def check_assembled_view(
    rules: Sequence[ParsedRule],
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]],
    same_group_pairs: Dict[Tuple[str, str], List[str]],
    inventory: Inventory,
    report: Report,
) -> None:
    for pair, occurrences in sorted(relation_pairs.items()):
        relation_types = {occ.relation_type for occ in occurrences}
        scenarios = {occ.scenario for occ in occurrences}
        labels = [f"{occ.group_id}/{occ.relation_type}/{occ.scenario}" for occ in occurrences]
        if len(relation_types) > 1:
            report.error(
                f"assembled view conflict for pair {pair[0]} <-> {pair[1]}: "
                f"multiple relation_type values: {', '.join(labels)}"
            )
        elif len(occurrences) > 1:
            if "common" in scenarios and len(scenarios) > 1:
                report.warn(
                    f"assembled view duplicate common+scenario relation for pair {pair[0]} <-> {pair[1]}: "
                    f"{', '.join(labels)}"
                )
            else:
                report.warn(
                    f"assembled view duplicate relation for pair {pair[0]} <-> {pair[1]}: {', '.join(labels)}"
                )
        if pair in same_group_pairs:
            report.error(
                f"assembled view conflict for pair {pair[0]} <-> {pair[1]}: same effective group in "
                f"{'; '.join(same_group_pairs[pair])}, but also related by {', '.join(labels)}"
            )

    warn_non_clique_relations(relation_pairs, same_group_pairs, inventory, report)


def warn_non_clique_relations(
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]],
    same_group_pairs: Dict[Tuple[str, str], List[str]],
    inventory: Inventory,
    report: Report,
) -> None:
    warned = 0
    for relation_type in sorted(RELATION_TYPES):
        edges = {pair for pair, occs in relation_pairs.items() if any(occ.relation_type == relation_type for occ in occs)}
        neighbors: Dict[str, Set[str]] = defaultdict(set)
        for a, b in edges:
            neighbors[a].add(b)
            neighbors[b].add(a)
        for a in sorted(neighbors):
            neigh = sorted(neighbors[a])
            for b, c in itertools.combinations(neigh, 2):
                b_c_pair = pair_key(b, c)
                if b_c_pair in same_group_pairs:
                    continue
                if same_tree(b, c, inventory):
                    continue
                if b_c_pair not in edges:
                    report.warn(
                        f"assembled view {relation_type} relation has {a}<->{b} and {a}<->{c}, "
                        f"but {b}<->{c} is uncovered; if intended pairwise, merge into one multi-group rule "
                        "or add the missing pair"
                    )
                    warned += 1
                    if warned >= MAX_NON_CLIQUE_WARNINGS:
                        report.warn(
                            f"assembled view non-clique warning limit reached ({MAX_NON_CLIQUE_WARNINGS}); "
                            "see coverage report for remaining uncovered pairs"
                        )
                        return


def same_tree(clock_a: str, clock_b: str, inventory: Inventory) -> bool:
    root_a = inventory.tree_root.get(clock_a, clock_a)
    root_b = inventory.tree_root.get(clock_b, clock_b)
    return bool(root_a and root_b and root_a == root_b)


def generate_sdc(
    rules: Sequence[ParsedRule],
    scenario: str,
    inventory_context: InventoryContext,
    form_digest: str,
) -> List[str]:
    lines = [
        "################################################################################",
        f"# Auto-generated SoC clock group constraints for scenario: {scenario}",
        f"# Author: {author_name()}",
        "# Stage: 03_soc_clock_groups",
        "# Script: 03_extract_soc_clock_groups.py",
        f"# Run completeness: {inventory_context.completeness}",
        f"# Harden SDC available: {inventory_context.available_harden_count}",
        f"# Harden SDC missing: {inventory_context.missing_harden_count}",
        f"# Harden SDC not_required: {inventory_context.not_required_harden_count}",
        f"# Missing instances: {','.join(inventory_context.missing_instances) or '<none>'}",
        f"# Clock universe digest: {inventory_context.clock_universe_digest}",
        f"# Form digest: {form_digest}",
        "# Source: 03_soc_clock_groups.xlsx clock_group_rules sheet",
        "# Policy: default synchronous + explicit async/exclusive groups",
        "################################################################################",
        "",
    ]
    emitted = 0
    for rule in rules:
        if not rule.effective_groups:
            continue
        analysis_style = rule.analysis_style or "normal"
        lines.append(f"# row {rule.row_idx}: {rule.group_id} ({rule.relation_type}, {analysis_style})")
        if rule.basis:
            lines.append(f"# Basis: {rule.basis}")
        lines.append(f"set_clock_groups -{rule.relation_type} \\")
        for group_idx, group in enumerate(rule.effective_groups):
            suffix = " \\" if group_idx != len(rule.effective_groups) - 1 else ""
            lines.append(f"  -group {get_clocks(group)}{suffix}")
        lines.append("")
        emitted += 1
    if emitted == 0:
        lines.append("# No clock group commands emitted for selected scenario.")
    return lines


def assembled_view_digest(
    scenario: str,
    rules: Sequence[ParsedRule],
    domains: Dict[str, ClockDomain],
    clock_universe_digest: str,
) -> str:
    payload = {
        "schema_version": SCHEMA_VERSION,
        "scenario": scenario,
        "clock_universe_digest": clock_universe_digest,
        "domains": [
            {
                "domain_id": domain_id,
                "members": sorted(domain.members),
                "blocked_instances": sorted(domain.blocked_instances),
            }
            for domain_id, domain in sorted(domains.items())
        ],
        "rules": [
            {
                "group_id": rule.group_id,
                "scenario": rule.scenario,
                "relation_type": rule.relation_type,
                "effective_groups": [sorted(group) for group in rule.effective_groups],
                "blocked_instances": sorted(rule.blocked_instances),
            }
            for rule in sorted(rules, key=lambda item: (item.scenario, item.group_id, item.row_idx))
        ],
    }
    return sha256_payload(payload)


def workbook_semantic_digest(
    domain_rows: Sequence[DomainMemberRow],
    rules: Sequence[ParsedRule],
) -> str:
    payload = {
        "schema_version": SCHEMA_VERSION,
        "domain_rows": [
            {
                "scenario": row.scenario,
                "domain_id": row.domain_id,
                "clock_name": row.clock_name,
                "membership_type": row.membership_type,
                "include_descendants": row.include_descendants,
                "source_instance": row.source_instance,
                "apply": row.apply,
                "review_status": row.review_status,
                "owner": row.owner,
                "basis": row.basis,
                "note": row.note,
            }
            for row in domain_rows
        ],
        "rules": [
            {
                "scenario": rule.scenario,
                "group_id": rule.group_id,
                "relation_type": rule.relation_type,
                "analysis_style": rule.analysis_style,
                "apply": rule.apply,
                "review_status": rule.review_status,
                "owner": rule.owner,
                "basis": rule.basis,
                "cdc_required": rule.cdc_required,
                "note": rule.note,
                "group_indices": list(rule.group_indices),
                "explicit_domain_groups": [list(group) for group in rule.explicit_domain_groups],
                "explicit_groups": [list(group) for group in rule.explicit_groups],
                "excluded": sorted(rule.excluded),
            }
            for rule in rules
        ],
    }
    return sha256_payload(payload)


def complete_relation_rows(
    scenario: str,
    inventory_context: InventoryContext,
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]],
    view_digest: str,
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for clock_a, clock_b in itertools.combinations(sorted(inventory_context.inventory.active), 2):
        occurrences = relation_pairs.get(pair_key(clock_a, clock_b), [])
        if occurrences:
            relation_types = sorted(set(occ.relation_type for occ in occurrences))
            if len(relation_types) != 1:
                raise RuntimeError(
                    f"cannot write relation map conflict for {clock_a}<->{clock_b}: {relation_types}"
                )
            relation_type = relation_types[0]
            relation_source = "explicit_rule"
            source_rule_ids = ";".join(sorted(set(occ.group_id for occ in occurrences)))
        else:
            relation_type = "synchronous"
            relation_source = "default_synchronous"
            source_rule_ids = ""
        rows.append({
            "schema_version": SCHEMA_VERSION,
            "scenario": scenario,
            "clock_a": clock_a,
            "clock_b": clock_b,
            "relation_type": relation_type,
            "relation_source": relation_source,
            "source_rule_ids": source_rule_ids,
            "clock_universe_digest": inventory_context.clock_universe_digest,
            "assembled_view_digest": view_digest,
        })
    return rows


def sdc_header_value(path: Path, key: str) -> str:
    prefix = "# " + key + ":"
    for line in path.read_text(encoding="utf-8").splitlines()[:40]:
        if line.startswith(prefix):
            return line[len(prefix):].strip()
    return ""


def write_relation_map_meta(
    path: Path,
    scenario: str,
    inventory_context: InventoryContext,
    form_path: Path,
    relation_map_path: Path,
    output_path: Path,
    common_sdc_path: Path,
    view_digest: str,
    form_digest: str,
    active_rules: Sequence[ParsedRule],
) -> None:
    payload = {
        "schema_version": SCHEMA_VERSION,
        "author": author_name(),
        "stage": "03_soc_clock_groups",
        "script": "03_extract_soc_clock_groups.py",
        "scenario": scenario,
        "run_completeness": inventory_context.completeness,
        "available_harden_count": inventory_context.available_harden_count,
        "missing_harden_count": inventory_context.missing_harden_count,
        "not_required_harden_count": inventory_context.not_required_harden_count,
        "missing_instances": list(inventory_context.missing_instances),
        "inventory_path": str(inventory_context.path.resolve()),
        "inventory_digest": inventory_context.inventory_digest,
        "inventory_meta_path": (
            str(inventory_context.meta_path.resolve()) if inventory_context.meta_path else ""
        ),
        "inventory_meta_digest": (
            sha256_file(inventory_context.meta_path) if inventory_context.meta_path else ""
        ),
        "clock_universe_digest": inventory_context.clock_universe_digest,
        "final_clock_sdc_path": (
            str(inventory_context.final_sdc_path.resolve()) if inventory_context.final_sdc_path else ""
        ),
        "final_clock_sdc_digest": inventory_context.final_sdc_digest,
        "form_path": str(form_path.resolve()),
        "form_digest": form_digest,
        "form_file_digest": sha256_file(form_path),
        "common_sdc_path": str(common_sdc_path.resolve()),
        "common_sdc_digest": sha256_file(common_sdc_path),
        "scenario_sdc_path": str(output_path.resolve()) if scenario != "common" else "",
        "scenario_sdc_digest": sha256_file(output_path) if scenario != "common" else "",
        "relation_map_path": str(relation_map_path.resolve()),
        "relation_map_digest": sha256_file(relation_map_path),
        "assembled_view_digest": view_digest,
        "active_rule_ids": sorted(
            rule.group_id for rule in active_rules if not rule.blocked_instances
        ),
        "blocked_rule_ids": sorted(
            rule.group_id for rule in active_rules if rule.blocked_instances
        ),
    }
    atomic_write_text(path, json.dumps(payload, indent=2, sort_keys=True) + "\n")


def write_check_report(
    path: Path,
    report: Report,
    scenario: str,
    form_path: Path,
    inventory_path: Path,
    output_path: Path,
    coverage_path: Path,
    inventory_context: Optional[InventoryContext] = None,
    relation_map_path: Optional[Path] = None,
    relation_meta_path: Optional[Path] = None,
) -> None:
    lines = [
        "03_soc_clock_groups extraction report",
        "=====================================",
        "",
        f"Author  : {author_name()}",
        "Stage ID: 03_soc_clock_groups",
        "Script  : 03_extract_soc_clock_groups.py",
        f"Scenario: {scenario}",
        f"Form    : {form_path}",
        f"Input   : {inventory_path}",
        f"Output  : {output_path}",
        f"Coverage: {coverage_path}",
        f"Relation map: {relation_map_path or '<none>'}",
        f"Relation meta: {relation_meta_path or '<none>'}",
        f"Warnings: {report.warning_count}",
        f"Errors  : {report.error_count}",
        f"Sync changed: {'yes' if report.sync_changed else 'no'}",
    ]
    if inventory_context is not None:
        lines.extend([
            f"Run completeness: {inventory_context.completeness}",
            f"Harden SDC available: {inventory_context.available_harden_count}",
            f"Harden SDC missing: {inventory_context.missing_harden_count}",
            f"Harden SDC not_required: {inventory_context.not_required_harden_count}",
            f"Missing instances: {', '.join(inventory_context.missing_instances) or '<none>'}",
            f"Inventory meta: {inventory_context.meta_path.resolve() if inventory_context.meta_path else '<none>'}",
            f"Clock universe digest: {inventory_context.clock_universe_digest}",
        ])
    lines.extend(["", "Messages:"])
    lines.extend(report.lines or ["INFO: no messages"])
    atomic_write_text(path, "\n".join(lines).rstrip() + "\n")


def write_coverage_report(
    path: Path,
    inventory_context: InventoryContext,
    rules: Sequence[ParsedRule],
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]],
    participation: Dict[str, List[str]],
    domains: Dict[str, ClockDomain],
    scenario: Optional[str],
    assembled_view_digest: str,
    complete_rows: Optional[Sequence[Dict[str, object]]] = None,
    stale_invalid_rows: Optional[Sequence[Dict[str, object]]] = None,
) -> None:
    inventory = inventory_context.inventory
    wb = Workbook()
    ws = wb.active
    ws.title = "clock_participation"
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - clock_participation",
        "Lists every active 01 clock and the clock groups it participates in.",
        8,
    )
    headers = [
        "clock_name",
        "clock_kind",
        "tree_root",
        "root_source",
        "direct_source",
        "final_action",
        "group_count",
        "groups",
    ]
    write_header(ws, headers)
    row_idx = 5
    for clock_name in sorted(inventory.active):
        info = inventory.active[clock_name]
        groups = participation.get(clock_name, [])
        ws.cell(row_idx, 1, clock_name)
        ws.cell(row_idx, 2, info.clock_kind)
        ws.cell(row_idx, 3, inventory.tree_root.get(clock_name, clock_name))
        ws.cell(row_idx, 4, info.root_source)
        ws.cell(row_idx, 5, info.direct_source)
        ws.cell(row_idx, 6, info.final_action)
        ws.cell(row_idx, 7, len(groups))
        ws.cell(row_idx, 8, "\n".join(groups))
        style_body_row(ws, row_idx, len(headers), WARNING_FILL if not groups else None)
        row_idx += 1
    set_widths(ws, [28, 24, 28, 32, 32, 18, 12, 70])
    ensure_table(ws, "ClockParticipation", f"A4:H{max(row_idx - 1, 5)}")

    if complete_rows is None:
        write_pair_relation_sheet(wb, relation_pairs)
    else:
        write_explicit_relation_coverage_sheet(wb, relation_pairs)
        write_complete_pair_relation_sheet(wb, complete_rows)
    incomplete_review = bool(
        complete_rows is not None
        and any(clean_cell(row.get("relation_type")) == "unknown" for row in complete_rows)
    )
    write_uncovered_pairs_sheet(wb, inventory, relation_pairs, incomplete_review)
    write_root_pair_summary_sheet(wb, inventory, relation_pairs)
    write_rule_effective_groups_sheet(wb, rules, include_scenario=complete_rows is None)
    write_domain_membership_sheet(wb, domains)
    write_coverage_metadata_sheet(wb, inventory_context, scenario, assembled_view_digest)
    if complete_rows is not None:
        write_relation_subset_sheet(
            wb,
            "default_synchronous_pairs",
            complete_rows,
            lambda row: clean_cell(row.get("relation_type")) == "synchronous",
        )
        write_relation_subset_sheet(
            wb,
            "unknown_incomplete_pairs",
            complete_rows,
            lambda row: clean_cell(row.get("relation_type")) == "unknown",
        )
        write_stale_invalid_sheet(wb, stale_invalid_rows or [])

    atomic_save_workbook(wb, path)


def write_relation_subset_sheet(wb, name, rows, predicate) -> None:
    ws = wb.create_sheet(name)
    headers = [
        "clock_a", "clock_b", "relation_type", "relation_source",
        "source_rule_ids", "clock_universe_digest", "assembled_view_digest",
    ]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - " + name,
        "Subset of the same complete pair map used to publish relation_map.csv.",
        len(headers),
    )
    write_header(ws, headers)
    row_idx = 5
    for row in rows:
        if not predicate(row):
            continue
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row.get(header, ""))
        style_body_row(ws, row_idx, len(headers), WARNING_FILL if name.startswith("unknown") else None)
        row_idx += 1
    set_widths(ws, [28, 28, 24, 24, 36, 70, 70])
    ensure_table(ws, safe_filename_token(name).replace("-", "_")[:25] + "Table", f"A4:G{max(row_idx - 1, 5)}")


def write_stale_invalid_sheet(wb, rows: Sequence[Dict[str, object]]) -> None:
    ws = wb.create_sheet("stale_invalid_rules")
    headers = ["sheet", "row", "object_id", "status", "reason"]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - stale_invalid_rules",
        "Workbook rows that are stale, blocked, or invalid in the current run.",
        len(headers),
    )
    write_header(ws, headers)
    row_idx = 5
    for row in rows:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row.get(header, ""))
        style_body_row(ws, row_idx, len(headers), ERROR_FILL)
        row_idx += 1
    set_widths(ws, [28, 12, 34, 28, 80])
    ensure_table(ws, "StaleInvalidRules", f"A4:E{max(row_idx - 1, 5)}")


def write_complete_pair_relation_sheet(wb, rows: Sequence[Dict[str, object]]) -> None:
    ws = wb.create_sheet("pair_relation_map")
    headers = [
        "clock_a", "clock_b", "relation_type", "relation_source",
        "source_rule_ids", "clock_universe_digest", "assembled_view_digest",
    ]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - pair_relation_map",
        "Review view of the same complete canonical pair map published as relation_map.csv.",
        len(headers),
    )
    write_header(ws, headers)
    row_idx = 5
    for row in rows:
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row.get(header, ""))
        style_body_row(ws, row_idx, len(headers))
        row_idx += 1
    set_widths(ws, [28, 28, 24, 24, 36, 70, 70])
    ensure_table(ws, "CompletePairRelationMap", f"A4:G{max(row_idx - 1, 5)}")


def write_explicit_relation_coverage_sheet(
    wb: Workbook,
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]],
) -> None:
    ws = wb.create_sheet("explicit_relation_coverage")
    headers = ["clock_a", "clock_b", "relation_type", "group_id", "group_a", "group_b"]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - explicit_relation_coverage",
        "Every canonical clock pair covered by an approved explicit clock-group rule.",
        len(headers),
    )
    write_header(ws, headers)
    row_idx = 5
    for (clock_a, clock_b), occurrences in sorted(relation_pairs.items()):
        for occ in occurrences:
            values = [clock_a, clock_b, occ.relation_type, occ.group_id, occ.group_a, occ.group_b]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row_idx, col_idx, value)
            style_body_row(ws, row_idx, len(headers))
            row_idx += 1
    set_widths(ws, [28, 28, 22, 30, 12, 12])
    ensure_table(ws, "ExplicitRelationCoverage", f"A4:F{max(row_idx - 1, 5)}")


def write_domain_membership_sheet(wb: Workbook, domains: Dict[str, ClockDomain]) -> None:
    ws = wb.create_sheet("resolved_clock_domains")
    headers = ["domain_id", "clock_name", "blocked_instances", "source_rows"]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - resolved_clock_domains",
        "Resolved human-reviewed clock domain members after genealogy closure.",
        len(headers),
    )
    write_header(ws, headers)
    row_idx = 5
    for domain_id, domain in sorted(domains.items()):
        members = domain.members or [""]
        for clock_name in members:
            values = [
                domain_id,
                clock_name,
                ", ".join(sorted(domain.blocked_instances)),
                ", ".join(str(item) for item in sorted(domain.source_rows)),
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row_idx, col_idx, value)
            style_body_row(ws, row_idx, len(headers), BLOCKED_FILL if domain.blocked_instances else None)
            row_idx += 1
    set_widths(ws, [28, 32, 34, 24])
    ensure_table(ws, "ResolvedClockDomains", f"A4:D{max(row_idx - 1, 5)}")


def write_coverage_metadata_sheet(
    wb: Workbook,
    inventory_context: InventoryContext,
    scenario: Optional[str],
    assembled_view_digest: str,
) -> None:
    ws = wb.create_sheet("metadata", 0)
    headers = ["key", "value"]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - metadata",
        "Machine-readable provenance for this review workbook.",
        len(headers),
    )
    write_header(ws, headers)
    values = [
        ("author", author_name()),
        ("stage", "03_soc_clock_groups"),
        ("script", "03_extract_soc_clock_groups.py"),
        ("run_id", inventory_context.run_id),
        ("mode_label", inventory_context.mode_label),
        ("design_revision", inventory_context.design_revision),
        ("run_completeness", inventory_context.completeness),
        ("missing_instances", ",".join(inventory_context.missing_instances)),
        ("structure_digest", inventory_context.structure_digest),
        ("clock_universe_digest", inventory_context.clock_universe_digest),
        ("assembled_view_digest", assembled_view_digest),
    ]
    if scenario is not None:
        values.insert(6, ("scenario", scenario))
    for row_idx, (key, value) in enumerate(values, start=5):
        ws.cell(row_idx, 1, key)
        ws.cell(row_idx, 2, value)
        style_body_row(ws, row_idx, 2)
    set_widths(ws, [30, 90])
    ensure_table(ws, "CoverageMetadata", f"A4:B{max(len(values) + 4, 5)}")


def write_coverage_sheet_title(ws, title: str, subtitle: str, width_cols: int) -> None:
    style_title(ws, title, subtitle, width_cols)


def write_pair_relation_sheet(wb: Workbook, relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]]) -> None:
    ws = wb.create_sheet("pair_relation_map")
    headers = ["clock_a", "clock_b", "relation_type", "group_id", "scenario", "group_a", "group_b"]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - pair_relation_map",
        "Every clock pair covered by active set_clock_groups in the assembled view.",
        len(headers),
    )
    write_header(ws, headers)
    row_idx = 5
    for (clock_a, clock_b), occurrences in sorted(relation_pairs.items()):
        for occ in occurrences:
            values = [clock_a, clock_b, occ.relation_type, occ.group_id, occ.scenario, occ.group_a, occ.group_b]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row_idx, col_idx, value)
            style_body_row(ws, row_idx, len(headers))
            row_idx += 1
    set_widths(ws, [28, 28, 22, 30, 14, 12, 12])
    ensure_table(ws, "PairRelationMap", f"A4:G{max(row_idx - 1, 5)}")


def write_uncovered_pairs_sheet(
    wb: Workbook,
    inventory: Inventory,
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]],
    incomplete_review: bool = False,
) -> None:
    ws = wb.create_sheet("uncovered_cross_root_pairs")
    headers = [
        "clock_a",
        "clock_b",
        "tree_root_a",
        "tree_root_b",
        "root_source_a",
        "root_source_b",
        "clock_kind_a",
        "clock_kind_b",
        "note",
    ]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - uncovered_cross_root_pairs",
        "Cross genealogy-tree active clock pairs not covered by active clock groups. They remain default synchronous.",
        len(headers),
    )
    write_header(ws, headers)
    row_idx = 5
    active = sorted(inventory.active.values(), key=lambda item: item.clock_name)
    for a, b in itertools.combinations(active, 2):
        tree_a = inventory.tree_root.get(a.clock_name, a.clock_name)
        tree_b = inventory.tree_root.get(b.clock_name, b.clock_name)
        if not tree_a or not tree_b or tree_a == tree_b:
            continue
        if pair_key(a.clock_name, b.clock_name) in relation_pairs:
            continue
        values = [
            a.clock_name,
            b.clock_name,
            tree_a,
            tree_b,
            a.root_source,
            b.root_source,
            a.clock_kind,
            b.clock_kind,
            (
                "unknown; workbook synchronization/review is incomplete"
                if incomplete_review
                else "default synchronous; coverage uses genealogy tree_root and does not model case analysis"
            ),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
        style_body_row(ws, row_idx, len(headers), WARNING_FILL)
        row_idx += 1
    set_widths(ws, [28, 28, 28, 28, 32, 32, 24, 24, 70])
    ensure_table(ws, "UncoveredCrossRootPairs", f"A4:I{max(row_idx - 1, 5)}")


def write_root_pair_summary_sheet(
    wb: Workbook,
    inventory: Inventory,
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]],
) -> None:
    ws = wb.create_sheet("root_pair_summary")
    headers = [
        "tree_root_a",
        "tree_root_b",
        "covered_pairs",
        "uncovered_pairs",
        "relation_types",
        "sample_uncovered_pairs",
    ]
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - root_pair_summary",
        "Aggregated view of covered/uncovered clock pairs by genealogy tree_root pair.",
        len(headers),
    )
    write_header(ws, headers)
    summary: Dict[Tuple[str, str], Dict[str, object]] = {}
    active = sorted(inventory.active.values(), key=lambda item: item.clock_name)
    for a, b in itertools.combinations(active, 2):
        tree_a = inventory.tree_root.get(a.clock_name, a.clock_name)
        tree_b = inventory.tree_root.get(b.clock_name, b.clock_name)
        if not tree_a or not tree_b or tree_a == tree_b:
            continue
        roots = pair_key(tree_a, tree_b)
        item = summary.setdefault(
            roots,
            {"covered": 0, "uncovered": 0, "relation_types": set(), "samples": []},
        )
        pair = pair_key(a.clock_name, b.clock_name)
        occurrences = relation_pairs.get(pair, [])
        if occurrences:
            item["covered"] = int(item["covered"]) + 1
            item["relation_types"].update(occ.relation_type for occ in occurrences)
        else:
            item["uncovered"] = int(item["uncovered"]) + 1
            samples = item["samples"]
            if len(samples) < 5:
                samples.append(f"{a.clock_name}<->{b.clock_name}")
    row_idx = 5
    for (root_a, root_b), item in sorted(summary.items()):
        values = [
            root_a,
            root_b,
            item["covered"],
            item["uncovered"],
            ", ".join(sorted(item["relation_types"])),
            "\n".join(item["samples"]),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
        fill = WARNING_FILL if item["uncovered"] else None
        style_body_row(ws, row_idx, len(headers), fill)
        row_idx += 1
    set_widths(ws, [32, 32, 15, 17, 28, 70])
    ensure_table(ws, "RootPairSummary", f"A4:F{max(row_idx - 1, 5)}")


def write_rule_effective_groups_sheet(
    wb: Workbook,
    rules: Sequence[ParsedRule],
    include_scenario: bool = True,
) -> None:
    ws = wb.create_sheet("rule_effective_groups")
    headers = [
        "group_id",
        "relation_type",
        "group_index",
        "explicit_clocks",
        "auto_added_descendants",
        "excluded_descendants",
        "effective_clocks",
        "review_note",
        "explicit_domains",
        "blocked_instances",
    ]
    if include_scenario:
        headers.insert(0, "scenario")
    write_coverage_sheet_title(
        ws,
        "03 Clock Group Coverage - rule_effective_groups",
        "Shows explicit groups, domain-closure additions, exclusions, and final effective groups.",
        len(headers),
    )
    write_header(ws, headers)
    row_idx = 5
    for rule in rules:
        for list_idx, effective in enumerate(rule.effective_groups):
            group_idx = rule.group_indices[list_idx]
            explicit = rule.explicit_groups[list_idx] if list_idx < len(rule.explicit_groups) else []
            explicit_domains = (
                rule.explicit_domain_groups[list_idx]
                if list_idx < len(rule.explicit_domain_groups)
                else []
            )
            auto_added = rule.auto_added_by_group[list_idx] if list_idx < len(rule.auto_added_by_group) else []
            review_note = ""
            fill = BLOCKED_FILL if rule.blocked_instances else None
            if rule.relation_type == "logically_exclusive" and auto_added:
                review_note = "review auto-added descendants for mux merge/shared downstream risk"
                fill = WARNING_FILL
            if rule.blocked_instances:
                review_note = "blocked_by_missing_sdc"
            values = [
                rule.group_id,
                rule.relation_type,
                group_idx,
                "\n".join(explicit),
                "\n".join(auto_added),
                "\n".join(sorted(rule.excluded)),
                "\n".join(effective),
                review_note,
                "\n".join(explicit_domains),
                "\n".join(sorted(rule.blocked_instances)),
            ]
            if include_scenario:
                values.insert(0, rule.scenario)
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row_idx, col_idx, value)
            style_body_row(ws, row_idx, len(headers), fill)
            row_idx += 1
    widths = [30, 22, 12, 40, 40, 34, 50, 55, 34, 34]
    if include_scenario:
        widths.insert(0, 12)
    set_widths(ws, widths)
    end_col = get_column_letter(len(headers))
    ensure_table(ws, "RuleEffectiveGroups", f"A4:{end_col}{max(row_idx - 1, 5)}")


def target_read_csv(path: Path, label: str, report: Report) -> Tuple[List[str], List[Dict[str, str]]]:
    if not path.is_file():
        report.error(f"required {label} is missing: {path}")
        return [], []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.DictReader(file_obj)
            headers = [clean_cell(item) for item in (reader.fieldnames or [])]
            if len(headers) != len(set(headers)):
                report.error(f"{label} has duplicate headers: {path}")
            return headers, [dict(row) for row in reader]
    except (OSError, csv.Error) as exc:
        report.error(f"cannot read {label} {path}: {exc}")
        return [], []


def target_read_json(path: Path, label: str, report: Report) -> Dict[str, object]:
    if not path.is_file():
        report.error(f"required {label} is missing: {path}")
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        report.error(f"cannot read {label} {path}: {exc}")
        return {}
    if not isinstance(payload, dict):
        report.error(f"{label} must contain a JSON object: {path}")
        return {}
    return payload


def read_target_run_context(path: Path, report: Report) -> TargetRunContext:
    headers, rows = target_read_csv(path, "run context", report)
    expected = ["run_id", "mode_label", "design_revision", "note"]
    if headers != expected:
        report.error("run_context.csv headers must be exactly: " + ",".join(expected))
    if len(rows) != 1:
        report.error("run_context.csv must contain exactly one data row")
    row = rows[0] if rows else {}
    run = TargetRunContext(
        clean_cell(row.get("run_id")),
        clean_cell(row.get("mode_label")),
        clean_cell(row.get("design_revision")),
        clean_cell(row.get("note")),
    )
    if not run.run_id:
        report.error("run_context.csv run_id is required")
    if not run.mode_label:
        report.error("run_context.csv mode_label is required")
    return run


def validate_target_required_views(path: Path, report: Report) -> str:
    headers, rows = target_read_csv(path, "required views", report)
    expected = [
        "view_id", "stage", "corner", "require_02", "require_04",
        "require_20", "require_30", "note",
    ]
    if headers != expected:
        report.error("required_views.csv headers must be exactly: " + ",".join(expected))
    seen_ids: Set[str] = set()
    seen_views: Set[Tuple[str, str]] = set()
    for row_idx, row in enumerate(rows, start=2):
        view_id = clean_cell(row.get("view_id"))
        stage = clean_cell(row.get("stage"))
        corner = clean_cell(row.get("corner"))
        if not view_id or not stage or not corner:
            report.error(f"required_views.csv row {row_idx} has blank view_id/stage/corner")
        if view_id in seen_ids:
            report.error(f"required_views.csv duplicate view_id {view_id}")
        if (stage, corner) in seen_views:
            report.error(f"required_views.csv duplicate stage/corner {stage}/{corner}")
        seen_ids.add(view_id)
        seen_views.add((stage, corner))
        for flag in ("require_02", "require_04", "require_20", "require_30"):
            if normalize_key(row.get(flag)) not in {"yes", "no"}:
                report.error(f"required_views.csv row {row_idx} {flag} must be yes/no")
    return sha256_file(path) if path.is_file() else ""


def target_validate_provenance(
    payload: Dict[str, object],
    label: str,
    run: TargetRunContext,
    report: Report,
) -> None:
    for field in ("run_id", "mode_label", "design_revision"):
        expected = clean_cell(getattr(run, field))
        actual = clean_cell(payload.get(field))
        if actual != expected:
            report.error(
                f"{label} {field} mismatch: expected {expected or '<blank>'}, "
                f"got {actual or '<blank>'}"
            )


def load_target_inventory_context(
    run_root: Path,
    run: TargetRunContext,
    report: Report,
) -> Optional[InventoryContext]:
    inventory_path = run_root / "01_middle/clock_inventory.csv"
    meta_path = run_root / "01_middle/clock_inventory.meta"
    completion_path = run_root / "01_middle/stage_completion.meta"
    final_sdc_path = run_root / "01_result/01_soc_clocks.sdc"
    meta = target_read_json(meta_path, "01 clock inventory meta", report)
    completion = target_read_json(completion_path, "01 stage completion", report)
    if not inventory_path.is_file() or not meta or not completion:
        if not inventory_path.is_file():
            report.error(f"required 01 clock inventory is missing: {inventory_path}")
        return None
    try:
        inventory = read_clock_inventory(inventory_path, report)
    except Exception as exc:
        report.error(str(exc))
        return None

    target_validate_provenance(meta, "01 clock inventory meta", run, report)
    target_validate_provenance(completion, "01 stage completion", run, report)
    if clean_cell(meta.get("stage_name")) != "01_soc_clocks":
        report.error("01 clock inventory meta stage_name must be 01_soc_clocks")
    if clean_cell(completion.get("stage_name")) != "01_soc_clocks":
        report.error("01 stage completion stage_name must be 01_soc_clocks")
    if clean_cell(meta.get("completion_status")) != "complete":
        report.error("01 clock inventory meta is not complete")
    if clean_cell(completion.get("completion_status")) != "complete":
        report.error("01 stage completion is not complete")
    if clean_cell(completion.get("sync_changed")).lower() != "no":
        report.error("01 stage completion sync_changed must be no")
    try:
        if int(completion.get("error_count", 1)) != 0:
            report.error("01 stage completion error_count must be zero")
    except (TypeError, ValueError):
        report.error("01 stage completion error_count is invalid")

    inventory_digest = sha256_file(inventory_path)
    if clean_cell(meta.get("inventory_digest")) != inventory_digest:
        report.error("01 inventory digest does not match clock_inventory.meta")
    if clean_cell(completion.get("clock_inventory_digest")) != inventory_digest:
        report.error("01 inventory digest does not match stage_completion.meta")
    structure_digest = clean_cell(meta.get("structure_digest"))
    if not structure_digest or clean_cell(completion.get("structure_digest")) != structure_digest:
        report.error("01 structure_digest is blank or inconsistent")
    clock_digest = sha256_clock_set(inventory.active)
    if clean_cell(meta.get("clock_set_digest")) != clock_digest:
        report.error("01 clock_set_digest does not match active inventory clocks")
    try:
        if int(meta.get("clock_count", -1)) != len(inventory.active):
            report.error("01 clock_count does not match active inventory clocks")
    except (TypeError, ValueError):
        report.error("01 clock_count is invalid")

    final_sdc_digest = ""
    if not final_sdc_path.is_file():
        report.error(f"required 01 final SDC is missing: {final_sdc_path}")
    else:
        final_sdc_digest = sha256_file(final_sdc_path)
        if clean_cell(meta.get("final_sdc_digest")) != final_sdc_digest:
            report.error("01 final SDC digest does not match clock_inventory.meta")
        if clean_cell(completion.get("output_sdc_digest")) != final_sdc_digest:
            report.error("01 final SDC digest does not match stage_completion.meta")
        sdc_names = extract_clock_names_from_sdc(final_sdc_path)
        if sdc_names != set(inventory.active):
            report.error(
                "01 final SDC clock set differs from inventory; "
                f"missing_in_sdc={sorted(set(inventory.active) - sdc_names)}; "
                f"extra_in_sdc={sorted(sdc_names - set(inventory.active))}"
            )

    completeness = normalize_key(meta.get("run_completeness")) or "complete"
    missing_raw = meta.get("missing_instances", [])
    missing_instances = tuple(
        sorted(clean_cell(item) for item in missing_raw if clean_cell(item))
    ) if isinstance(missing_raw, list) else ()
    if completeness == "partial":
        report.warn(
            "01 run completeness is partial; candidate evidence may be incomplete: "
            + (", ".join(missing_instances) or "<unspecified>")
        )
    elif completeness != "complete":
        report.error(f"01 run completeness must be complete or partial, got {completeness}")
    return InventoryContext(
        inventory=inventory,
        path=inventory_path,
        scenario="common",
        completeness=completeness,
        missing_instances=missing_instances,
        meta_path=meta_path,
        final_sdc_path=final_sdc_path,
        final_sdc_digest=final_sdc_digest,
        inventory_digest=inventory_digest,
        clock_universe_digest=clock_digest,
        run_id=run.run_id,
        mode_label=run.mode_label,
        design_revision=run.design_revision,
        structure_digest=structure_digest,
        completion_path=completion_path,
        completion_digest=sha256_file(completion_path),
    )


def target_style_header(ws, headers: Sequence[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(13, min(34, len(header) + 4))
    ws.freeze_panes = "A2"


def target_setup_sheet(wb: Workbook, name: str, headers: Sequence[str]):
    ws = wb.create_sheet(name)
    target_style_header(ws, headers)
    return ws


def target_metadata_rows(
    run: TargetRunContext,
    inventory_context: InventoryContext,
    required_views_digest: str,
) -> List[Tuple[str, str]]:
    return [
        ("Author", author_name()),
        ("run_id", run.run_id),
        ("mode_label", run.mode_label),
        ("design_revision", run.design_revision),
        ("structure_digest", inventory_context.structure_digest),
        ("required_views_digest", required_views_digest),
        ("01_inventory_path", str(inventory_context.path.resolve())),
        ("01_inventory_digest", inventory_context.inventory_digest),
        ("01_completion_path", str(inventory_context.completion_path.resolve())),
        ("01_completion_digest", inventory_context.completion_digest),
        ("run_completeness", inventory_context.completeness),
        ("Port accounting", "not_applicable; added_bits=0"),
    ]


def update_target_metadata(
    wb: Workbook,
    run: TargetRunContext,
    inventory_context: InventoryContext,
    required_views_digest: str,
) -> bool:
    expected = target_metadata_rows(run, inventory_context, required_views_digest)
    if "run_metadata" in wb.sheetnames:
        ws = wb["run_metadata"]
        current = [
            (clean_cell(ws.cell(row, 1).value), clean_cell(ws.cell(row, 2).value))
            for row in range(2, ws.max_row + 1)
        ]
        if current == expected:
            return False
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("run_metadata", 0)
    target_style_header(ws, ["Field", "Value"])
    for row_idx, (field_name, value) in enumerate(expected, start=2):
        ws.cell(row_idx, 1, field_name).font = Font(bold=True)
        ws.cell(row_idx, 2, value)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    return True


def append_target_domain_row(
    ws,
    clock: ClockInfo,
    inventory_digest: str,
    sync_status: str,
    is_root: bool,
) -> None:
    mapping = {header: idx + 1 for idx, header in enumerate(TARGET_DOMAIN_HEADERS)}
    row_idx = ws.max_row + 1
    values = {
        "clock_name": clock.clock_name,
        "root_clock": "yes" if is_root else "no",
        "apply": "",
        "review_status": "draft",
        "sync_status": sync_status,
        "clock_kind": clock.clock_kind,
        "root_source": clock.root_source,
        "source_inventory_digest": inventory_digest,
    }
    for header, value in values.items():
        ws.cell(row_idx, mapping[header], value)
    fill = NEW_FILL if sync_status == "NEW_FROM_01" else None
    style_body_row(ws, row_idx, len(TARGET_DOMAIN_HEADERS), fill)


def target_candidate_id(clock_a: str, clock_b: str) -> str:
    payload = "\n".join(pair_key(clock_a, clock_b)).encode("utf-8")
    return "CG_CAND_" + hashlib.sha256(payload).hexdigest()[:16]


def append_target_candidate_rows(ws, inventory: Inventory, existing: Set[str], max_pairs: int) -> int:
    mapping = {header: idx + 1 for idx, header in enumerate(TARGET_CANDIDATE_HEADERS)}
    added = 0
    active = sorted(inventory.active)
    for clock_a, clock_b in itertools.combinations(active, 2):
        if len(existing) >= max_pairs:
            break
        if same_tree(clock_a, clock_b, inventory):
            continue
        candidate_id = target_candidate_id(clock_a, clock_b)
        if candidate_id in existing:
            continue
        values = {
            "candidate_id": candidate_id,
            "candidate_type": "cross_tree_pair",
            "clock_a": clock_a,
            "clock_b": clock_b,
            "tree_root_a": inventory.tree_root.get(clock_a, clock_a),
            "tree_root_b": inventory.tree_root.get(clock_b, clock_b),
            "reason": "active clocks belong to different 01 genealogy roots",
            "evidence": "genealogy tree_root differs; human CDC/STA review required",
            "recommended_action": "review_only",
            "decision": "draft",
        }
        row_idx = ws.max_row + 1
        for header, value in values.items():
            ws.cell(row_idx, mapping[header], value)
        style_body_row(ws, row_idx, len(TARGET_CANDIDATE_HEADERS))
        existing.add(candidate_id)
        added += 1
    return added


def create_target_workbook(
    path: Path,
    run: TargetRunContext,
    inventory_context: InventoryContext,
    required_views_digest: str,
    max_candidate_pairs: int,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    domain_ws = target_setup_sheet(wb, DOMAIN_SHEET, TARGET_DOMAIN_HEADERS)
    target_setup_sheet(wb, RULE_SHEET, TARGET_RULE_HEADERS)
    candidate_ws = target_setup_sheet(wb, CANDIDATE_SHEET, TARGET_CANDIDATE_HEADERS)
    update_target_metadata(wb, run, inventory_context, required_views_digest)
    for clock_name in sorted(inventory_context.inventory.active):
        append_target_domain_row(
            domain_ws,
            inventory_context.inventory.active[clock_name],
            inventory_context.inventory_digest,
            "NEW_FROM_01",
            inventory_context.inventory.tree_root.get(clock_name, clock_name) == clock_name,
        )
    append_target_candidate_rows(
        candidate_ws, inventory_context.inventory, set(), max_candidate_pairs
    )
    atomic_save_workbook(wb, path)
    return wb


def target_sheet_mapping(ws, expected: Sequence[str], label: str, report: Report) -> Dict[str, int]:
    headers = [clean_cell(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    duplicate = sorted(item for item in set(headers) if item and headers.count(item) > 1)
    if duplicate:
        report.error(f"{label} has duplicate headers: {','.join(duplicate)}")
    for header in expected:
        if header not in headers:
            ws.cell(1, ws.max_column + 1, header)
            headers.append(header)
            report.warn(f"{label} added missing header {header}")
            report.sync_changed = True
    target_style_header(ws, headers)
    return {header: headers.index(header) + 1 for header in expected}


def target_domain_row_has_intent(values: Dict[str, object]) -> bool:
    apply_value = normalize_key(values.get("apply"))
    review = normalize_key(values.get("review_status"))
    if apply_value == "yes" and review == "approved":
        return bool(clean_cell(values.get("domain_id")) and clean_cell(values.get("membership_type")))
    if apply_value == "no" and review in {"reviewed", "approved", "rejected"}:
        return bool(clean_cell(values.get("note")) or clean_cell(values.get("basis")))
    return False


def sync_target_workbook(
    wb: Workbook,
    form_path: Path,
    run: TargetRunContext,
    inventory_context: InventoryContext,
    required_views_digest: str,
    max_candidate_pairs: int,
    report: Report,
) -> Tuple[object, Dict[str, int]]:
    for sheet_name, headers in (
        (DOMAIN_SHEET, TARGET_DOMAIN_HEADERS),
        (RULE_SHEET, TARGET_RULE_HEADERS),
        (CANDIDATE_SHEET, TARGET_CANDIDATE_HEADERS),
    ):
        if sheet_name not in wb.sheetnames:
            target_setup_sheet(wb, sheet_name, headers)
            report.warn(f"workbook added missing sheet {sheet_name}")
            report.sync_changed = True
    domain_ws = wb[DOMAIN_SHEET]
    domain_mapping = target_sheet_mapping(domain_ws, TARGET_DOMAIN_HEADERS, DOMAIN_SHEET, report)
    target_sheet_mapping(wb[RULE_SHEET], TARGET_RULE_HEADERS, RULE_SHEET, report)
    candidate_ws = wb[CANDIDATE_SHEET]
    candidate_mapping = target_sheet_mapping(
        candidate_ws, TARGET_CANDIDATE_HEADERS, CANDIDATE_SHEET, report
    )

    rows: List[Tuple[int, Dict[str, object]]] = []
    by_clock: Dict[str, List[Tuple[int, Dict[str, object]]]] = defaultdict(list)
    for row_idx in range(2, domain_ws.max_row + 1):
        values = {
            header: domain_ws.cell(row_idx, col_idx).value
            for header, col_idx in domain_mapping.items()
        }
        if not any(clean_cell(value) for value in values.values()):
            continue
        rows.append((row_idx, values))
        by_clock[clean_cell(values.get("clock_name"))].append((row_idx, values))
    for clock_name in sorted(inventory_context.inventory.active):
        if clock_name not in by_clock:
            append_target_domain_row(
                domain_ws,
                inventory_context.inventory.active[clock_name],
                inventory_context.inventory_digest,
                "NEW_FROM_01",
                inventory_context.inventory.tree_root.get(clock_name, clock_name) == clock_name,
            )
            report.warn(f"added new 01 clock to domain membership review: {clock_name}")
            report.sync_changed = True

    for row_idx, values in rows:
        clock_name = clean_cell(values.get("clock_name"))
        status = clean_cell(values.get("sync_status"))
        if clock_name not in inventory_context.inventory.active:
            desired = (
                "BLOCKED_BY_MISSING_SDC"
                if inventory_context.completeness == "partial"
                else "STALE_NOT_IN_01"
            )
            if status != desired:
                domain_ws.cell(row_idx, domain_mapping["sync_status"], desired)
                report.sync_changed = True
            style_body_row(domain_ws, row_idx, domain_ws.max_column, BLOCKED_FILL if desired.startswith("BLOCKED") else ERROR_FILL)
            continue
        clock = inventory_context.inventory.active[clock_name]
        expected_machine = {
            "clock_kind": clock.clock_kind,
            "root_source": clock.root_source,
            "source_inventory_digest": inventory_context.inventory_digest,
        }
        machine_changed = False
        for column, expected in expected_machine.items():
            if clean_cell(values.get(column)) != expected:
                domain_ws.cell(row_idx, domain_mapping[column], expected)
                machine_changed = True
        if machine_changed:
            domain_ws.cell(row_idx, domain_mapping["sync_status"], "NEW_FROM_01")
            report.warn(f"synchronized machine context for clock {clock_name}")
            report.sync_changed = True
        elif status in {"NEW_FROM_01", "STALE_NOT_IN_01", "BLOCKED_BY_MISSING_SDC"} and target_domain_row_has_intent(values):
            domain_ws.cell(row_idx, domain_mapping["sync_status"], "OK")
            report.info(f"reset reviewed sync_status to OK for clock {clock_name}")
            report.sync_changed = True

    existing_candidates = set()
    for row_idx in range(2, candidate_ws.max_row + 1):
        candidate_id = clean_cell(candidate_ws.cell(row_idx, candidate_mapping["candidate_id"]).value)
        if candidate_id:
            existing_candidates.add(candidate_id)
    added_candidates = append_target_candidate_rows(
        candidate_ws,
        inventory_context.inventory,
        existing_candidates,
        max_candidate_pairs,
    )
    if added_candidates:
        report.warn(f"added {added_candidates} new clock-group candidate(s)")
        report.sync_changed = True
    if update_target_metadata(wb, run, inventory_context, required_views_digest):
        report.warn("updated run_metadata machine context")
        report.sync_changed = True
    if report.sync_changed:
        report.info(f"workbook synchronization changed {form_path}")
    return domain_ws, domain_mapping


def read_target_domain_rows(
    ws,
    mapping: Dict[str, int],
    inventory_context: InventoryContext,
    report: Report,
) -> List[DomainMemberRow]:
    rows: List[DomainMemberRow] = []
    for row_idx in range(2, ws.max_row + 1):
        values = {header: ws.cell(row_idx, col_idx).value for header, col_idx in mapping.items()}
        if not any(clean_cell(value) for value in values.values()):
            continue
        sync_status = clean_cell(values.get("sync_status"))
        if sync_status not in TARGET_SYNC_VALUES:
            report.error(f"{DOMAIN_SHEET} row {row_idx}: invalid sync_status {sync_status}")
        source_instance = clean_cell(values.get("source_instance"))
        allowed_blocked = (
            sync_status == "BLOCKED_BY_MISSING_SDC"
            and inventory_context.completeness == "partial"
            and source_instance in set(inventory_context.missing_instances)
        )
        if sync_status != "OK" and not allowed_blocked:
            report.error(
                f"{DOMAIN_SHEET} row {row_idx}: clock {clean_cell(values.get('clock_name')) or '<blank>'} "
                f"has blocking sync_status {sync_status or '<blank>'}"
            )
        root_clock = normalize_key(values.get("root_clock"))
        if root_clock not in {"", "yes", "no"}:
            report.error(f"{DOMAIN_SHEET} row {row_idx}: root_clock must be yes/no/blank")
        membership_type = normalize_key(values.get("membership_type"))
        include_descendants = normalize_key(values.get("include_descendants"))
        if not include_descendants:
            include_descendants = "yes" if membership_type == "seed" or root_clock == "yes" else "no"
        rows.append(
            DomainMemberRow(
                row_idx=row_idx,
                scenario="common",
                domain_id=clean_cell(values.get("domain_id")),
                clock_name=clean_cell(values.get("clock_name")),
                membership_type=membership_type,
                include_descendants=include_descendants,
                source_instance=source_instance,
                apply=normalize_key(values.get("apply")),
                review_status=normalize_key(values.get("review_status")),
                owner=clean_cell(values.get("owner")),
                basis=clean_cell(values.get("basis")),
                note=clean_cell(values.get("note")),
            )
        )
    return rows


def read_target_rule_rows(ws, report: Report) -> List[ParsedRule]:
    mapping = target_sheet_mapping(ws, TARGET_RULE_HEADERS, RULE_SHEET, report)
    numbered = group_indices(mapping)
    rows: List[ParsedRule] = []
    for row_idx in range(2, ws.max_row + 1):
        values = {header: ws.cell(row_idx, col_idx).value for header, col_idx in mapping.items()}
        if not any(clean_cell(value) for value in values.values()):
            continue
        explicit_groups: List[List[str]] = []
        explicit_domains: List[List[str]] = []
        used_indices: List[int] = []
        for group_idx in numbered:
            clocks = parse_clock_list(values.get(f"group_{group_idx}_clocks"))
            domains = parse_clock_list(values.get(f"group_{group_idx}_domains"))
            if clocks or domains:
                explicit_groups.append(clocks)
                explicit_domains.append(domains)
                used_indices.append(group_idx)
        apply_value = normalize_key(values.get("apply"))
        review = normalize_key(values.get("review_status"))
        analysis_style = normalize_key(values.get("analysis_style"))
        if analysis_style not in TARGET_ANALYSIS_STYLES:
            report.error(
                f"{RULE_SHEET} row {row_idx}: invalid target analysis_style {analysis_style}"
            )
        owner = clean_cell(values.get("owner"))
        basis = clean_cell(values.get("basis"))
        if apply_value == "yes" and review == "approved":
            if not owner:
                report.error(f"{RULE_SHEET} row {row_idx}: active rule requires owner")
            if not basis:
                report.error(f"{RULE_SHEET} row {row_idx}: active rule requires basis")
        rows.append(
            ParsedRule(
                row_idx=row_idx,
                scenario="common",
                group_id=clean_cell(values.get("group_id")),
                relation_type=canonical_relation_type(values.get("relation_type")),
                analysis_style=analysis_style,
                apply=apply_value,
                review_status=review,
                owner=owner,
                basis=basis,
                cdc_required="yes",
                note=clean_cell(values.get("note")),
                explicit_groups=explicit_groups,
                explicit_domain_groups=explicit_domains,
                group_indices=used_indices,
                excluded=set(parse_clock_list(values.get("exclude_descendant_clocks"))),
            )
        )
    return rows


def collect_target_stale_invalid_rows(wb: Workbook) -> List[Dict[str, object]]:
    result: List[Dict[str, object]] = []
    if DOMAIN_SHEET in wb.sheetnames:
        ws = wb[DOMAIN_SHEET]
        headers = [clean_cell(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        mapping = {header: headers.index(header) + 1 for header in headers if header}
        for row_idx in range(2, ws.max_row + 1):
            clock_name = clean_cell(ws.cell(row_idx, mapping.get("clock_name", 0)).value) if mapping.get("clock_name") else ""
            if not clock_name:
                continue
            status = clean_cell(ws.cell(row_idx, mapping.get("sync_status", 0)).value) if mapping.get("sync_status") else ""
            apply_value = normalize_key(ws.cell(row_idx, mapping.get("apply", 0)).value) if mapping.get("apply") else ""
            domain_id = clean_cell(ws.cell(row_idx, mapping.get("domain_id", 0)).value) if mapping.get("domain_id") else ""
            reason = ""
            if status != "OK":
                reason = "blocking sync_status " + (status or "<blank>")
            elif apply_value == "yes" and not domain_id:
                reason = "active membership has blank domain_id"
            if reason:
                result.append({
                    "sheet": DOMAIN_SHEET, "row": row_idx, "object_id": clock_name,
                    "status": status, "reason": reason,
                })
    if RULE_SHEET in wb.sheetnames:
        ws = wb[RULE_SHEET]
        headers = [clean_cell(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        mapping = {header: headers.index(header) + 1 for header in headers if header}
        for row_idx in range(2, ws.max_row + 1):
            group_id = clean_cell(ws.cell(row_idx, mapping.get("group_id", 0)).value) if mapping.get("group_id") else ""
            if not group_id:
                continue
            apply_value = normalize_key(ws.cell(row_idx, mapping.get("apply", 0)).value) if mapping.get("apply") else ""
            review = normalize_key(ws.cell(row_idx, mapping.get("review_status", 0)).value) if mapping.get("review_status") else ""
            owner = clean_cell(ws.cell(row_idx, mapping.get("owner", 0)).value) if mapping.get("owner") else ""
            basis = clean_cell(ws.cell(row_idx, mapping.get("basis", 0)).value) if mapping.get("basis") else ""
            reason = ""
            if apply_value == "yes" and review != "approved":
                reason = "apply=yes requires review_status=approved"
            elif apply_value == "yes" and not owner:
                reason = "active rule requires owner"
            elif apply_value == "yes" and not basis:
                reason = "active rule requires basis"
            if reason:
                result.append({
                    "sheet": RULE_SHEET, "row": row_idx, "object_id": group_id,
                    "status": review, "reason": reason,
                })
    return result


def target_workbook_snapshot_digest(wb: Workbook) -> str:
    """Hash the complete target workbook for review-state diagnostics only."""
    payload: Dict[str, object] = {}
    for sheet_name in (DOMAIN_SHEET, RULE_SHEET, CANDIDATE_SHEET, "run_metadata"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        payload[sheet_name] = [
            [clean_cell(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
            for row in range(1, ws.max_row + 1)
        ]
    return sha256_payload(payload)


def target_workbook_semantic_digest(
    domain_rows: Sequence[DomainMemberRow],
    rules: Sequence[ParsedRule],
) -> str:
    """Hash reviewed constraint decisions, excluding machine provenance/context."""
    return workbook_semantic_digest(domain_rows, rules)


def generate_target_sdc(
    rules: Sequence[ParsedRule],
    run: TargetRunContext,
    inventory_context: InventoryContext,
    workbook_digest: str,
) -> str:
    lines = [
        "# Auto-generated by 03_extract_soc_clock_groups.py",
        f"# Author: {author_name()}",
        f"# Run ID: {run.run_id}",
        f"# Mode label: {run.mode_label}",
        f"# Design revision: {run.design_revision}",
        f"# Structure digest: {inventory_context.structure_digest}",
        f"# 01 inventory digest: {inventory_context.inventory_digest}",
        f"# Run completeness: {inventory_context.completeness}",
        f"# Missing harden SDC instances: {', '.join(inventory_context.missing_instances) or '<none>'}",
        f"# Workbook semantic digest: {workbook_digest}",
        "# Port accounting: not_applicable; added_bits=0",
        "",
    ]
    emitted = 0
    for rule in rules:
        if not active_rule(rule) or rule.blocked_instances or not rule.effective_groups:
            continue
        lines.append(f"# clock_group_rules row {rule.row_idx}: {rule.group_id}")
        lines.append(f"# Basis: {rule.basis}")
        lines.append(f"set_clock_groups -{rule.relation_type} \\")
        for list_idx, group in enumerate(rule.effective_groups):
            suffix = " \\" if list_idx != len(rule.effective_groups) - 1 else ""
            lines.append(f"  -group {get_clocks(group)}{suffix}")
        lines.append("")
        emitted += 1
    if not emitted:
        lines.append("# No approved clock-group commands emitted for this run.")
    return "\n".join(lines).rstrip() + "\n"


def target_relation_rows(
    inventory_context: InventoryContext,
    relation_pairs: Dict[Tuple[str, str], List[PairOccurrence]],
    view_digest: str,
) -> List[Dict[str, object]]:
    rows = []
    for row in complete_relation_rows("common", inventory_context, relation_pairs, view_digest):
        rows.append({key: value for key, value in row.items() if key != "scenario"})
    return rows


def target_review_required_relation_rows(
    inventory_context: InventoryContext,
    view_digest: str,
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for clock_a, clock_b in itertools.combinations(sorted(inventory_context.inventory.active), 2):
        rows.append({
            "schema_version": SCHEMA_VERSION,
            "clock_a": clock_a,
            "clock_b": clock_b,
            "relation_type": "unknown",
            "relation_source": "workbook_review_incomplete",
            "source_rule_ids": "",
            "clock_universe_digest": inventory_context.clock_universe_digest,
            "assembled_view_digest": view_digest,
        })
    return rows


def write_target_review_required_coverage(
    path: Path,
    wb: Workbook,
    inventory_context: InventoryContext,
    report: Report,
) -> None:
    view_digest = sha256_payload({
        "state": "review_required",
        "clock_universe_digest": inventory_context.clock_universe_digest,
        "workbook_snapshot_digest": target_workbook_snapshot_digest(wb),
    })
    relation_rows = target_review_required_relation_rows(inventory_context, view_digest)
    stale_invalid_rows = collect_target_stale_invalid_rows(wb)
    write_coverage_report(
        path,
        inventory_context,
        [],
        {},
        {},
        {},
        None,
        view_digest,
        relation_rows,
        stale_invalid_rows,
    )
    report.warn(
        f"coverage marks {len(relation_rows)} active clock pair(s) unknown while workbook review is incomplete"
    )


def write_target_report(
    path: Path,
    report: Report,
    run: TargetRunContext,
    inventory_context: Optional[InventoryContext],
    form_path: Path,
    output_path: Path,
    relation_path: Path,
    completion_path: Path,
    completion_status: str,
) -> None:
    lines = [
        "03_soc_clock_groups extraction report",
        "=====================================",
        "",
        f"Author: {author_name()}",
        "Stage: 03_soc_clock_groups",
        "Script: 03_extract_soc_clock_groups.py",
        f"Run ID: {run.run_id}",
        f"Mode label: {run.mode_label}",
        f"Design revision: {run.design_revision}",
        f"Completion status: {completion_status}",
        "Port accounting: not_applicable; added_bits=0",
        f"Form: {form_path}",
        f"Output: {output_path}",
        f"Relation map: {relation_path}",
        f"Completion: {completion_path}",
        f"Warnings: {report.warning_count}",
        f"Errors: {report.error_count}",
        f"Sync changed: {'yes' if report.sync_changed else 'no'}",
    ]
    if completion_status == "diagnostic":
        lines.extend([
            "Port accounting mode: diagnostic/read-only",
            "Accounting closure: not evaluated",
        ])
    if inventory_context is not None:
        lines.extend([
            f"Run completeness: {inventory_context.completeness}",
            f"Structure digest: {inventory_context.structure_digest}",
            f"01 inventory digest: {inventory_context.inventory_digest}",
            f"Missing instances: {', '.join(inventory_context.missing_instances) or '<none>'}",
        ])
    lines.extend(["", "Messages:"])
    lines.extend(report.lines or ["INFO: no messages"])
    atomic_write_text(path, "\n".join(lines).rstrip() + "\n")


def run_target_single_run_03(args: argparse.Namespace) -> int:
    print(f"Author: {author_name()}")
    report = Report()
    run_root = Path(args.run_root).expanduser().resolve()
    input_root = run_root / "inputs"
    form_path = run_root / "03_middle/03_soc_clock_groups.xlsx"
    relation_path = run_root / "03_middle/relation_map.csv"
    relation_meta_path = run_root / "03_middle/relation_map.meta"
    completion_path = run_root / "03_middle/stage_completion.meta"
    output_path = run_root / "03_result/03_soc_clock_groups.sdc"
    report_path = run_root / "03_result/reports/clock_group_check_report.txt"
    coverage_path = run_root / "03_result/reports/clock_group_coverage_report.xlsx"
    run = read_target_run_context(input_root / "run_context.csv", report)
    required_views_digest = validate_target_required_views(
        input_root / "required_views.csv", report
    )
    inventory_context = load_target_inventory_context(run_root, run, report)
    print(f"Run root: {run_root}")
    print(f"Run ID: {run.run_id}")
    print(f"Mode label: {run.mode_label}")
    print("Port accounting: not_applicable; added_bits=0")

    if not args.diagnose_only and completion_path.exists():
        completion_path.unlink()
        report.info("invalidated prior 03 completion meta before formal rerun")
    if report.error_count or inventory_context is None:
        write_target_report(
            report_path, report, run, inventory_context, form_path, output_path,
            relation_path, completion_path, "failed",
        )
        print(f"Report: {report_path}")
        return 1
    if args.require_complete_harden_sdc and inventory_context.completeness != "complete":
        report.error("01 inventory is partial and strict completeness was requested")
        write_target_report(
            report_path, report, run, inventory_context, form_path, output_path,
            relation_path, completion_path, "failed",
        )
        return 1
    if not form_path.is_file():
        wb = create_target_workbook(
            form_path, run, inventory_context, required_views_digest,
            args.max_candidate_pairs,
        )
        report.sync_changed = True
        report.warn("created 03 review workbook with NEW_FROM_01 membership rows")
        write_target_review_required_coverage(
            coverage_path, wb, inventory_context, report
        )
        write_target_report(
            report_path, report, run, inventory_context, form_path, output_path,
            relation_path, completion_path, "review_required",
        )
        print(f"Workbook: {form_path}")
        print(f"Report: {report_path}")
        return 1
    try:
        wb = load_workbook(str(form_path))
    except Exception as exc:
        report.error(f"cannot load workbook {form_path}: {exc}")
        write_target_report(
            report_path, report, run, inventory_context, form_path, output_path,
            relation_path, completion_path, "failed",
        )
        return 1

    domain_ws, domain_mapping = sync_target_workbook(
        wb, form_path, run, inventory_context, required_views_digest,
        args.max_candidate_pairs, report,
    )
    if report.sync_changed:
        atomic_save_workbook(wb, form_path)
        report.warn("workbook changed during synchronization; review and rerun")
        write_target_review_required_coverage(
            coverage_path, wb, inventory_context, report
        )
        write_target_report(
            report_path, report, run, inventory_context, form_path, output_path,
            relation_path, completion_path, "review_required",
        )
        print(f"Workbook: {form_path}")
        print(f"Report: {report_path}")
        return 1

    domain_rows = read_target_domain_rows(
        domain_ws, domain_mapping, inventory_context, report
    )
    domains = validate_and_build_domains(domain_rows, "common", inventory_context, report)
    rules = read_target_rule_rows(wb[RULE_SHEET], report)
    stale_invalid_rows = collect_target_stale_invalid_rows(wb)
    assembled = validate_and_expand_rules(
        rules, "common", inventory_context.inventory, domains, report
    )
    relation_pairs, same_group_pairs, participation = build_pair_maps(assembled)
    check_assembled_view(
        assembled, relation_pairs, same_group_pairs, inventory_context.inventory, report
    )
    view_digest = assembled_view_digest(
        "single_run", assembled, domains, inventory_context.clock_universe_digest
    )
    relation_rows = target_relation_rows(inventory_context, relation_pairs, view_digest)
    write_coverage_report(
        coverage_path, inventory_context, assembled, relation_pairs, participation,
        domains, None, view_digest, relation_rows, stale_invalid_rows,
    )
    report.info(f"wrote coverage report: {coverage_path}")
    if report.error_count:
        write_target_report(
            report_path, report, run, inventory_context, form_path, output_path,
            relation_path, completion_path, "failed",
        )
        return 1

    workbook_digest = target_workbook_semantic_digest(domain_rows, rules)
    if args.diagnose_only:
        report.info("diagnose-only completed; formal SDC/relation/completion were not published")
        write_target_report(
            report_path, report, run, inventory_context, form_path, output_path,
            relation_path, completion_path, "diagnostic",
        )
        print("Diagnostic validation passed.")
        return 0

    sdc_text = generate_target_sdc(assembled, run, inventory_context, workbook_digest)
    atomic_write_text(output_path, sdc_text)
    relation_fields = [
        "schema_version", "clock_a", "clock_b", "relation_type",
        "relation_source", "source_rule_ids", "clock_universe_digest",
        "assembled_view_digest",
    ]
    atomic_write_csv(relation_path, relation_fields, relation_rows)
    relation_meta = {
        "schema_version": SCHEMA_VERSION,
        "author": author_name(),
        "stage_name": "03_soc_clock_groups",
        "run_id": run.run_id,
        "mode_label": run.mode_label,
        "design_revision": run.design_revision,
        "completion_status": "complete",
        "structure_digest": inventory_context.structure_digest,
        "required_views_digest": required_views_digest,
        "upstream_01_inventory_path": str(inventory_context.path.resolve()),
        "upstream_01_inventory_digest": inventory_context.inventory_digest,
        "upstream_01_inventory_meta_path": str(inventory_context.meta_path.resolve()),
        "upstream_01_inventory_meta_digest": sha256_file(inventory_context.meta_path),
        "upstream_01_completion_path": str(inventory_context.completion_path.resolve()),
        "upstream_01_completion_digest": inventory_context.completion_digest,
        "upstream_artifact_digests": {
            "01_clock_inventory": inventory_context.inventory_digest,
            "01_clock_inventory_meta": sha256_file(inventory_context.meta_path),
            "01_stage_completion": inventory_context.completion_digest,
        },
        "workbook_path": str(form_path.resolve()),
        "workbook_semantic_digest": workbook_digest,
        "workbook_file_digest": sha256_file(form_path),
        "output_sdc_path": str(output_path.resolve()),
        "output_sdc_digest": sha256_file(output_path),
        "relation_map_path": str(relation_path.resolve()),
        "relation_map_digest": sha256_file(relation_path),
        "clock_universe_digest": inventory_context.clock_universe_digest,
        "assembled_view_digest": view_digest,
        "run_completeness": inventory_context.completeness,
        "missing_instances": list(inventory_context.missing_instances),
        "active_rule_ids": sorted(
            rule.group_id for rule in assembled if not rule.blocked_instances
        ),
        "blocked_rule_ids": sorted(
            rule.group_id for rule in assembled if rule.blocked_instances
        ),
    }
    atomic_write_text(
        relation_meta_path,
        json.dumps(relation_meta, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
    )
    completion = {
        "schema_version": "1.0",
        "author": author_name(),
        "stage_name": "03_soc_clock_groups",
        "stage": "all",
        "corner": "all",
        "run_id": run.run_id,
        "mode_label": run.mode_label,
        "design_revision": run.design_revision,
        "completion_status": "complete",
        "error_count": 0,
        "sync_changed": "no",
        "structure_digest": inventory_context.structure_digest,
        "required_views_digest": required_views_digest,
        "upstream_01_inventory_path": str(inventory_context.path.resolve()),
        "upstream_01_inventory_digest": inventory_context.inventory_digest,
        "upstream_01_inventory_meta_path": str(inventory_context.meta_path.resolve()),
        "upstream_01_inventory_meta_digest": sha256_file(inventory_context.meta_path),
        "upstream_01_completion_path": str(inventory_context.completion_path.resolve()),
        "upstream_01_completion_digest": inventory_context.completion_digest,
        "upstream_artifact_digests": {
            "01_clock_inventory": inventory_context.inventory_digest,
            "01_clock_inventory_meta": sha256_file(inventory_context.meta_path),
            "01_stage_completion": inventory_context.completion_digest,
        },
        "workbook_path": str(form_path.resolve()),
        "workbook_semantic_digest": workbook_digest,
        "workbook_file_digest": sha256_file(form_path),
        "relation_map_path": str(relation_path.resolve()),
        "relation_map_digest": sha256_file(relation_path),
        "relation_meta_path": str(relation_meta_path.resolve()),
        "relation_meta_digest": sha256_file(relation_meta_path),
        "output_sdc_path": str(output_path.resolve()),
        "output_sdc_digest": sha256_file(output_path),
        "run_completeness": inventory_context.completeness,
        "port_accounting": "not_applicable",
        "added_bits": 0,
        "accounting_digest_before": "not_applicable",
        "accounting_digest_after": "not_applicable",
        "accounting_delta_digest": "not_applicable",
        "port_accounting_summary": "Port accounting: not_applicable; added_bits=0",
    }
    atomic_write_text(
        completion_path,
        json.dumps(completion, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
    )
    report.info("published formal SDC, relation map/meta, and stage completion")
    write_target_report(
        report_path, report, run, inventory_context, form_path, output_path,
        relation_path, completion_path, "complete",
    )
    print(f"SDC: {output_path}")
    print(f"Relation map: {relation_path}")
    print(f"Completion: {completion_path}")
    print(f"Report: {report_path}")
    return 0


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate 03 clock group SDC and coverage reports from 03_soc_clock_groups.xlsx."
    )
    parser.add_argument(
        "--run-root",
        help="target runtime root; reads flat 01_middle artifacts and writes flat 03_middle/03_result artifacts",
    )
    parser.add_argument(
        "-scenario", "--scenario", choices=sorted(SCENARIOS),
        help="legacy mode only; target --run-root mode accepts no scenario selector",
    )
    parser.add_argument(
        "-input",
        "--input",
        default=None,
        help="01 clock inventory CSV path",
    )
    parser.add_argument("--inventory-meta", help="01 assembled clock inventory meta JSON path")
    parser.add_argument("--clock-sdc", help="final 01 clock SDC path used for digest and clock-set validation")
    parser.add_argument("--form", help="03 clock group workbook path")
    parser.add_argument(
        "--report",
        help="output check report path; default: clock_group_check_report_<scenario>.txt",
    )
    parser.add_argument("--coverage", help="coverage workbook output path")
    parser.add_argument("--relation-map", help="complete clock-pair relation CSV output path")
    parser.add_argument("--relation-meta", help="relation-map provenance meta output path")
    parser.add_argument(
        "--max-candidate-pairs",
        type=int,
        default=500,
        help="maximum cross-root candidate pairs written when creating a new workbook",
    )
    parser.add_argument(
        "--require-complete-harden-sdc",
        action="store_true",
        help="block when the 01 assembled inventory reports partial harden SDC availability",
    )
    parser.add_argument(
        "--diagnose-only",
        action="store_true",
        help="validate target inputs/workbook without publishing formal SDC or completion meta",
    )
    return parser.parse_args(argv)


def resolve_path(base: Path, value: Optional[str], default: str) -> Path:
    candidate = Path(value) if value else Path(default)
    if candidate.is_absolute():
        return candidate
    return base / candidate


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)
    if args.run_root:
        target_overrides = [
            name
            for name, value in (
                ("--scenario", args.scenario),
                ("--input", args.input),
                ("--inventory-meta", args.inventory_meta),
                ("--clock-sdc", args.clock_sdc),
                ("--form", args.form),
                ("--report", args.report),
                ("--coverage", args.coverage),
                ("--relation-map", args.relation_map),
                ("--relation-meta", args.relation_meta),
            )
            if value
        ]
        if target_overrides:
            print(f"Author: {author_name()}")
            print(
                "ERROR: target single-run mode uses fixed paths and accepts no override(s): "
                + ", ".join(target_overrides),
                file=sys.stderr,
            )
            return 2
        return run_target_single_run_03(args)
    if not args.scenario:
        print("ERROR: legacy mode requires --scenario; target mode requires --run-root", file=sys.stderr)
        return 2
    print(f"Author: {author_name()}")
    scenario = args.scenario
    cwd = Path.cwd()
    target_layout = bool(args.run_root)
    run_root = Path(args.run_root).expanduser().resolve() if target_layout else cwd
    scenario_token = safe_filename_token(scenario)
    if target_layout:
        form_path = resolve_path(run_root, args.form, "03_middle/03_soc_clock_groups.xlsx")
        inventory_path = resolve_path(
            run_root,
            args.input,
            f"01_middle/assembled/{scenario}/clock_inventory.csv",
        )
        meta_path: Optional[Path] = resolve_path(
            run_root,
            args.inventory_meta,
            f"01_middle/assembled/{scenario}/clock_inventory.meta",
        )
        clock_sdc_path = resolve_path(run_root, args.clock_sdc, args.clock_sdc) if args.clock_sdc else None
        result_root = run_root / "03_result"
        output_path = output_sdc_path(result_root, scenario)
        common_sdc_path = output_sdc_path(result_root, "common")
        report_path = resolve_path(
            run_root,
            args.report,
            f"03_result/reports/clock_group_check_report_{scenario_token}.txt",
        )
        coverage_path = resolve_path(
            run_root,
            args.coverage,
            f"03_result/reports/clock_group_coverage_report_{scenario_token}.xlsx",
        )
        relation_map_path = resolve_path(
            run_root,
            args.relation_map,
            f"03_middle/relation_map/{scenario}.csv",
        )
        relation_meta_path = resolve_path(
            run_root,
            args.relation_meta,
            f"03_middle/relation_map/{scenario}.meta",
        )
    else:
        form_path = resolve_path(cwd, args.form, "03_soc_clock_groups.xlsx")
        inventory_path = resolve_path(cwd, args.input, "../01_soc_clocks/clock_inventory.csv")
        if args.inventory_meta:
            meta_path = resolve_path(cwd, args.inventory_meta, args.inventory_meta)
        else:
            candidate_meta = inventory_path.with_suffix(".meta")
            meta_path = candidate_meta if candidate_meta.is_file() else None
        if args.clock_sdc:
            clock_sdc_path = resolve_path(cwd, args.clock_sdc, args.clock_sdc)
        else:
            candidate_sdc = cwd / "../01_soc_clocks/common/01_soc_clocks.sdc"
            clock_sdc_path = candidate_sdc if candidate_sdc.is_file() else None
        output_path = output_sdc_path(cwd, scenario)
        common_sdc_path = output_sdc_path(cwd, "common")
        report_path = resolve_path(
            cwd,
            args.report,
            f"clock_group_check_report_{scenario_token}.txt",
        )
        coverage_path = resolve_path(
            cwd,
            args.coverage,
            f"clock_group_coverage_report_{scenario_token}.xlsx",
        )
        relation_map_path = resolve_path(
            cwd,
            args.relation_map,
            f"relation_map/{scenario}.csv",
        )
        relation_meta_path = resolve_path(
            cwd,
            args.relation_meta,
            f"relation_map/{scenario}.meta",
        )
    report = Report()

    if args.max_candidate_pairs < 1:
        print("ERROR: --max-candidate-pairs must be at least 1", file=sys.stderr)
        return 2

    report.info(f"resolved run root: {run_root.resolve()}")
    report.info(f"resolved inventory: {inventory_path.resolve()}")
    report.info(f"resolved inventory meta: {meta_path.resolve() if meta_path else '<none>'}")
    report.info(f"resolved clock SDC: {clock_sdc_path.resolve() if clock_sdc_path else '<from meta or unavailable>'}")
    report.info(f"resolved form: {form_path.resolve()}")
    report.info(f"resolved output SDC: {output_path.resolve()}")

    try:
        inventory_context = load_inventory_context(
            inventory_path,
            report,
            expected_scenario=scenario,
            meta_path=meta_path,
            clock_sdc_path=clock_sdc_path,
            require_meta=target_layout,
        )
    except Exception as exc:
        report.error(str(exc))
        write_check_report(
            report_path,
            report,
            scenario,
            form_path,
            inventory_path,
            output_path,
            coverage_path,
            relation_map_path=relation_map_path,
            relation_meta_path=relation_meta_path,
        )
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    if report.error_count:
        write_check_report(
            report_path, report, scenario, form_path, inventory_path, output_path, coverage_path,
            inventory_context, relation_map_path, relation_meta_path,
        )
        print(f"ERROR: failed to load clean 01 clock inventory. Report: {report_path}", file=sys.stderr)
        return 1
    if args.require_complete_harden_sdc and inventory_context.completeness != "complete":
        report.error("01 assembled inventory is partial and --require-complete-harden-sdc was requested")
        write_check_report(
            report_path, report, scenario, form_path, inventory_path, output_path, coverage_path,
            inventory_context, relation_map_path, relation_meta_path,
        )
        print(f"ERROR: incomplete harden SDC availability. Report: {report_path}", file=sys.stderr)
        return 1

    if not form_path.is_file():
        create_new_workbook(form_path, inventory_context.inventory, args.max_candidate_pairs)
        report.sync_changed = True
        report.warn(f"created new workbook: {form_path}")
        report.warn("fill clock_domain_membership and clock_group_rules, then approve rows before generating SDC")
        write_check_report(
            report_path, report, scenario, form_path, inventory_path, output_path, coverage_path,
            inventory_context, relation_map_path, relation_meta_path,
        )
        print(f"Created new workbook: {form_path}")
        print(f"Report: {report_path}")
        print("SDC was not generated because the workbook needs review.")
        return 1

    try:
        wb = load_workbook(form_path)
        if RULE_SHEET not in wb.sheetnames:
            raise RuntimeError(f"{form_path.name} does not contain sheet {RULE_SHEET}")
        domain_changed = ensure_domain_sheet(wb)
        ws = wb[RULE_SHEET]
        header_row, mapping, header_changed = ensure_rule_headers(ws)
        if header_changed or domain_changed:
            report.sync_changed = True
            report.warn(f"{form_path.name}: added missing 03 workbook sheet/header(s)")
            atomic_save_workbook(wb, form_path)
            write_check_report(
                report_path, report, scenario, form_path, inventory_path, output_path, coverage_path,
                inventory_context, relation_map_path, relation_meta_path,
            )
            print(f"Updated workbook: {form_path}")
            print(f"Report: {report_path}")
            print("SDC was not generated because workbook headers changed.")
            return 1

        if rewrite_relation_type_cells(ws, header_row, mapping, report):
            report.sync_changed = True
            atomic_save_workbook(wb, form_path)

        domain_rows = read_domain_rows(wb[DOMAIN_SHEET])
        domains = validate_and_build_domains(domain_rows, scenario, inventory_context, report)
        rows = read_rule_rows(ws, header_row, mapping)
        form_digest = workbook_semantic_digest(domain_rows, rows)
        assembled = validate_and_expand_rules(
            rows, scenario, inventory_context.inventory, domains, report
        )
        relation_pairs, same_group_pairs, participation = build_pair_maps(assembled)
        check_assembled_view(
            assembled, relation_pairs, same_group_pairs, inventory_context.inventory, report
        )
        view_digest = assembled_view_digest(
            scenario, assembled, domains, inventory_context.clock_universe_digest
        )
        write_coverage_report(
            coverage_path,
            inventory_context,
            assembled,
            relation_pairs,
            participation,
            domains,
            scenario,
            view_digest,
        )
        report.info(f"wrote coverage report: {coverage_path}")

        if report.error_count:
            write_check_report(
                report_path, report, scenario, form_path, inventory_path, output_path, coverage_path,
                inventory_context, relation_map_path, relation_meta_path,
            )
            print(f"ERROR: validation failed. Report: {report_path}", file=sys.stderr)
            print(f"Coverage: {coverage_path}", file=sys.stderr)
            return 1

        if scenario != "common" and target_layout:
            if not common_sdc_path.is_file():
                raise RuntimeError(
                    f"common 03 SDC must be generated before scenario {scenario}: {common_sdc_path}"
                )
            common_form_digest = sdc_header_value(common_sdc_path, "Form digest")
            if common_form_digest != form_digest:
                raise RuntimeError(
                    "common 03 SDC is stale relative to the current workbook; rerun -scenario common first"
                )
        lines = generate_sdc(output_rules(rows, scenario), scenario, inventory_context, form_digest)
        atomic_write_text(output_path, "\n".join(lines).rstrip() + "\n")
        report.info(f"wrote SDC: {output_path}")
        relation_rows = complete_relation_rows(
            scenario, inventory_context, relation_pairs, view_digest
        )
        relation_fields = [
            "schema_version", "scenario", "clock_a", "clock_b", "relation_type",
            "relation_source", "source_rule_ids", "clock_universe_digest",
            "assembled_view_digest",
        ]
        atomic_write_csv(relation_map_path, relation_fields, relation_rows)
        report.info(f"wrote complete relation map: {relation_map_path}")
        effective_common_sdc = output_path if scenario == "common" else common_sdc_path
        write_relation_map_meta(
            relation_meta_path,
            scenario,
            inventory_context,
            form_path,
            relation_map_path,
            output_path,
            effective_common_sdc,
            view_digest,
            form_digest,
            assembled,
        )
        report.info(f"wrote relation map meta: {relation_meta_path}")
        write_check_report(
            report_path, report, scenario, form_path, inventory_path, output_path, coverage_path,
            inventory_context, relation_map_path, relation_meta_path,
        )
    except Exception as exc:
        report.error(str(exc))
        write_check_report(
            report_path, report, scenario, form_path, inventory_path, output_path, coverage_path,
            inventory_context, relation_map_path, relation_meta_path,
        )
        print(f"ERROR: {exc}", file=sys.stderr)
        print(f"Report: {report_path}", file=sys.stderr)
        return 2

    print(f"Scenario: {scenario}")
    print(f"Form    : {form_path}")
    print(f"SDC     : {output_path}")
    print(f"Report  : {report_path}")
    print(f"Coverage: {coverage_path}")
    print(f"Relation: {relation_map_path}")
    print(f"Meta    : {relation_meta_path}")
    print(f"Warnings: {report.warning_count}")
    print(f"Errors  : {report.error_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
