#!/usr/bin/env python3
"""Generate reviewed SoC feedthrough direct-edge constraints."""

from __future__ import print_function

import argparse
import csv
import datetime
import hashlib
import importlib.util
import io
import itertools
import json
import math
import os
import re
import shutil
import socket
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

try:
    from dataclasses import dataclass, field
except ImportError:  # pragma: no cover - Python 3.6 compatibility path
    class _CompatField:
        def __init__(self, default_factory=None):
            self.default_factory = default_factory

    def field(default_factory=None):
        return _CompatField(default_factory=default_factory)

    def dataclass(_cls=None, frozen=False):
        def wrap(cls):
            annotations = getattr(cls, "__annotations__", {})
            names = list(annotations.keys())
            defaults = {}
            factories = {}
            for name in names:
                if hasattr(cls, name):
                    value = getattr(cls, name)
                    if isinstance(value, _CompatField):
                        factories[name] = value.default_factory
                        delattr(cls, name)
                    else:
                        defaults[name] = value

            def __init__(self, *args, **kwargs):
                if len(args) > len(names):
                    raise TypeError("too many positional arguments")
                values = dict(zip(names, args))
                for name in names[len(args):]:
                    if name in kwargs:
                        values[name] = kwargs.pop(name)
                    elif name in factories:
                        values[name] = factories[name]()
                    elif name in defaults:
                        values[name] = defaults[name]
                    else:
                        raise TypeError("missing required argument: " + name)
                if kwargs:
                    raise TypeError("unexpected argument: " + sorted(kwargs)[0])
                for name in names:
                    object.__setattr__(self, name, values[name])

            def __eq__(self, other):
                return other.__class__ is cls and all(
                    getattr(self, name) == getattr(other, name) for name in names
                )

            cls.__init__ = __init__
            cls.__eq__ = __eq__
            if frozen:
                cls.__hash__ = lambda self: hash(tuple(getattr(self, name) for name in names))
            else:
                cls.__hash__ = None
            return cls

        if _cls is None:
            return wrap
        return wrap(_cls)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - environment guard
    print("ERROR: openpyxl is required to read/write the 10 review workbook.", file=sys.stderr)
    raise SystemExit(2) from exc


SCHEMA_VERSION = "1"
TARGET_SCHEMA_VERSION = "1.0"
SCENARIOS = {"common", "func", "scan", "mbist", "gpio_in", "gpio_out"}
STAGES = {"all", "synth", "prects", "postcts", "postroute"}
TOOLS = {"dc", "sta"}
MATCHED_STATUSES = {"matched", "ok", "valid"}
CANONICAL_CONNECTION_TYPES = {
    "harden_to_harden",
    "fabric_to_harden",
    "harden_to_fabric",
    "top_pad_to_harden",
    "harden_to_top_pad",
    "pad_to_pad",
    "clock_connection",
    "feedthrough_candidate",
    "constant_tie",
    "no_connect",
    "unknown",
}
EXCLUDED_CONNECTION_TYPES = {
    "top_pad_to_harden",
    "harden_to_top_pad",
    "pad_to_pad",
    "clock_connection",
    "constant_tie",
    "no_connect",
}
DISPOSITIONS = {"emit_budget", "no_soc_budget_required", "route_to_30", "not_applicable", "pending"}
BUDGET_MODELS = {
    "", "interconnect_budget", "manual_budget", "reviewed_two_side_budget",
    "clock_relative_io_delay", "unknown",
}
REVIEW_STATUSES = {"", "pending", "approved", "rejected"}
YES_NO = {"", "yes", "no"}
TOOL_SURFACES = {"", "dc", "sta", "both"}
ACTIVE_CLOCK_ACTIONS = {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}
CANONICAL_RELATIONS = {
    "synchronous",
    "asynchronous",
    "logically_exclusive",
    "physically_exclusive",
}
CANONICAL_RELATION_SOURCES = {"explicit_rule", "default_synchronous"}
PSEUDO_INSTANCES = {"top", "fabric", "unknown", "constant", "const", "nc", "no_connect"}
VERSIONED_POLICY_RE = re.compile(r"(?:^|[^A-Za-z0-9])(?:v|version)[._-]?\d+(?:[._-]\d+)*", re.IGNORECASE)

PORT_BIT_RE = re.compile(r"^[^\s\[\]]+(?:\[-?\d+\])?$")
PORT_RANGE_RE = re.compile(r"^(.+)\[(-?\d+)\s*:\s*(-?\d+)\]$")
PORT_EXACT_BIT_RE = re.compile(r"^(.+)\[(-?\d+)\]$")
FEEDTHROUGH_RE = re.compile(r"^(fti|fto)_(?:(\d+)_)?(.+)$")

EDGE_HEADERS = [
    "schema_version",
    "run_completeness",
    "available_harden_count",
    "missing_harden_count",
    "not_required_harden_count",
    "missing_instances",
    "port_accounting",
    "connection_inventory_path",
    "harden_sdc_manifest_path",
    "scenario",
    "stage",
    "corner",
    "feedthrough_edge_id",
    "connection_id",
    "scenario_scope",
    "connection_type",
    "src_instance",
    "src_direction",
    "src_port",
    "src_bit_index",
    "src_endpoint_key",
    "src_soc_object",
    "dst_instance",
    "dst_direction",
    "dst_port",
    "dst_bit_index",
    "dst_endpoint_key",
    "dst_soc_object",
    "feedthrough_instance",
    "feedthrough_port",
    "feedthrough_side",
    "channel_disposition",
    "budget_scope",
    "budget_model",
    "budget_required",
    "src_output_delay_max",
    "src_output_delay_min",
    "dst_input_delay_max",
    "dst_input_delay_min",
    "converted_max",
    "converted_min",
    "emit_max",
    "emit_min",
    "datapath_only",
    "tool_surface",
    "src_sdc_status",
    "dst_sdc_status",
    "evidence_status",
    "src_clock",
    "dst_clock",
    "clock_relation",
    "source_sdc_file",
    "source_line",
    "source_digest",
    "source_command",
    "machine_digest",
    "apply",
    "review_status",
    "owner",
    "reviewer",
    "review_date",
    "approved_machine_digest",
    "disposition_basis",
    "sdc_independent_basis",
    "relationship_override_basis",
    "min_sign_review",
    "related_pair_id",
    "related_chain_id",
    "validation_status",
    "machine_note",
    "note",
]

TARGET_EDGE_HEADERS = [
    "schema_version",
    "run_id",
    "mode_label",
    "design_revision",
    "run_completeness",
    "structure_digest",
    "accounting_digest_before",
    "accounting_digest_after",
    "feedthrough_edge_id",
    "connection_id",
    "edge_role",
    "src_instance",
    "src_direction",
    "src_port",
    "src_bit_index",
    "src_endpoint",
    "src_soc_object",
    "dst_instance",
    "dst_direction",
    "dst_port",
    "dst_bit_index",
    "dst_endpoint",
    "dst_soc_object",
    "source_workbook",
    "source_sheet",
    "source_row",
    "feedthrough_instance",
    "feedthrough_port",
    "feedthrough_side",
    "classification_source",
    "classification_status",
    "feedthrough_confirmed",
    "src_sdc_status",
    "dst_sdc_status",
    "evidence_status",
    "src_output_delay_max",
    "src_output_delay_min",
    "dst_input_delay_max",
    "dst_input_delay_min",
    "src_clock",
    "dst_clock",
    "clock_relation",
    "source_sdc_file",
    "source_line",
    "source_digest",
    "source_command",
    "machine_digest",
    "channel_disposition",
    "budget_model",
    "budget_required",
    "converted_max",
    "converted_min",
    "emit_max",
    "emit_min",
    "datapath_only",
    "tool_surface",
    "apply",
    "review_status",
    "owner",
    "reviewer",
    "review_date",
    "approved_machine_digest",
    "basis",
    "sdc_independent_basis",
    "relationship_override_basis",
    "min_sign_review",
    "validation_status",
    "machine_note",
    "note",
]

TARGET_MACHINE_HEADERS = {
    "schema_version", "run_id", "mode_label", "design_revision",
    "run_completeness", "structure_digest", "accounting_digest_before",
    "accounting_digest_after", "feedthrough_edge_id", "connection_id",
    "edge_role", "src_instance", "src_direction", "src_port",
    "src_bit_index", "src_endpoint", "src_soc_object", "dst_instance",
    "dst_direction", "dst_port", "dst_bit_index", "dst_endpoint",
    "dst_soc_object", "source_workbook", "source_sheet", "source_row",
    "feedthrough_instance", "feedthrough_port", "feedthrough_side",
    "classification_source", "classification_status", "src_sdc_status",
    "dst_sdc_status", "evidence_status", "src_output_delay_max",
    "src_output_delay_min", "dst_input_delay_max", "dst_input_delay_min",
    "src_clock", "dst_clock", "clock_relation", "source_sdc_file",
    "source_line", "source_digest", "source_command", "machine_digest",
    "validation_status", "machine_note",
}

TARGET_REVIEW_INVALIDATING_HEADERS = {
    "run_id", "mode_label", "design_revision", "structure_digest",
    "feedthrough_edge_id", "connection_id", "edge_role", "src_instance",
    "src_direction", "src_port", "src_bit_index", "src_endpoint",
    "src_soc_object", "dst_instance", "dst_direction", "dst_port",
    "dst_bit_index", "dst_endpoint", "dst_soc_object", "source_workbook",
    "source_sheet", "source_row", "feedthrough_instance", "feedthrough_port",
    "feedthrough_side", "classification_source", "classification_status",
    "src_sdc_status", "dst_sdc_status", "evidence_status",
    "src_output_delay_max", "src_output_delay_min", "dst_input_delay_max",
    "dst_input_delay_min", "source_sdc_file", "source_line", "source_digest",
    "source_command", "validation_status",
}

TARGET_TERMINAL_DISPOSITIONS = {
    "emit_budget", "no_soc_budget_required", "not_applicable"
}

LOG_HEADERS = [
    "instance",
    "port",
    "constraint_type",
    "clock_name",
    "min_value",
    "max_value",
    "bare_value",
    "complex_options",
    "source_sdc_file",
    "source_line",
    "source_digest",
    "parse_status",
    "original_command",
    "message",
]

MACHINE_HEADERS = {
    "schema_version",
    "run_completeness",
    "available_harden_count",
    "missing_harden_count",
    "not_required_harden_count",
    "missing_instances",
    "port_accounting",
    "connection_inventory_path",
    "harden_sdc_manifest_path",
    "scenario",
    "stage",
    "corner",
    "feedthrough_edge_id",
    "connection_id",
    "scenario_scope",
    "connection_type",
    "src_instance",
    "src_direction",
    "src_port",
    "src_bit_index",
    "src_endpoint_key",
    "src_soc_object",
    "dst_instance",
    "dst_direction",
    "dst_port",
    "dst_bit_index",
    "dst_endpoint_key",
    "dst_soc_object",
    "feedthrough_instance",
    "feedthrough_port",
    "feedthrough_side",
    "budget_scope",
    "src_output_delay_max",
    "src_output_delay_min",
    "dst_input_delay_max",
    "dst_input_delay_min",
    "src_sdc_status",
    "dst_sdc_status",
    "evidence_status",
    "src_clock",
    "dst_clock",
    "clock_relation",
    "source_sdc_file",
    "source_line",
    "source_digest",
    "source_command",
    "machine_digest",
    "related_pair_id",
    "related_chain_id",
    "validation_status",
    "machine_note",
}

REVIEW_INVALIDATING_MACHINE_HEADERS = {
    "scenario_scope",
    "connection_type",
    "src_instance",
    "src_direction",
    "src_port",
    "src_bit_index",
    "src_endpoint_key",
    "src_soc_object",
    "dst_instance",
    "dst_direction",
    "dst_port",
    "dst_bit_index",
    "dst_endpoint_key",
    "dst_soc_object",
    "feedthrough_instance",
    "feedthrough_port",
    "feedthrough_side",
    "budget_scope",
    "src_output_delay_max",
    "src_output_delay_min",
    "dst_input_delay_max",
    "dst_input_delay_min",
    "src_sdc_status",
    "dst_sdc_status",
    "evidence_status",
    "source_sdc_file",
    "source_line",
    "source_digest",
    "source_command",
    "validation_status",
}

HEADER_FILL = PatternFill("solid", fgColor="215967")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C6CC"),
    right=Side(style="thin", color="B8C6CC"),
    top=Side(style="thin", color="B8C6CC"),
    bottom=Side(style="thin", color="B8C6CC"),
)


@dataclass
class HardenSdc:
    scenario: str
    inst_name: str
    module_name: str
    sdc_path: str
    availability_status: str
    note: str
    source_row: int
    resolved_path: Optional[Path] = None


@dataclass
class RunCompleteness:
    status: str
    available_instances: List[str] = field(default_factory=list)
    missing_instances: List[str] = field(default_factory=list)
    not_required_instances: List[str] = field(default_factory=list)
    manifest_path: str = ""

    @property
    def available_count(self) -> int:
        return len(self.available_instances)

    @property
    def missing_count(self) -> int:
        return len(self.missing_instances)

    @property
    def not_required_count(self) -> int:
        return len(self.not_required_instances)


@dataclass
class ClockInventory:
    active_names: Set[str] = field(default_factory=set)
    aliases: Dict[Tuple[str, str], str] = field(default_factory=dict)
    unresolved_aliases: Set[Tuple[str, str]] = field(default_factory=set)
    status: str = "unavailable"
    inventory_digest: str = ""
    meta_digest: str = ""
    clock_set_digest: str = ""
    final_sdc_digest: str = ""


@dataclass
class RelationMap:
    relations: Dict[Tuple[str, str], str] = field(default_factory=dict)
    status: str = "unavailable"
    assembled_view_digest: str = ""


@dataclass
class ConnectionEdge:
    schema_version: str
    connection_id: str
    scenario_scope: str
    connection_type: str
    src_instance: str
    src_direction: str
    src_port: str
    src_bit_index: str
    src_endpoint_key: str
    src_soc_object: str
    dst_instance: str
    dst_direction: str
    dst_port: str
    dst_bit_index: str
    dst_endpoint_key: str
    dst_soc_object: str
    validation_status: str
    note: str
    source_row: int


@dataclass
class FeedthroughEdge:
    edge: ConnectionEdge
    feedthrough_edge_id: str
    feedthrough_instance: str
    feedthrough_port: str
    feedthrough_side: str
    validation_status: str
    machine_note: str


@dataclass
class TargetFeedthroughEdge:
    edge: ConnectionEdge
    connection_hash: str
    edge_role: str
    feedthrough_instance: str
    feedthrough_port: str
    feedthrough_side: str
    classification_source: str
    classification_status: str
    source_record: Any
    src_record: Any
    dst_record: Any


@dataclass
class TargetHardenSdc:
    inst_name: str
    module_name: str
    sdc_path: str
    availability_status: str
    sdc_digest: str
    note: str
    source_row: int
    resolved_path: Optional[Path] = None


@dataclass
class TargetDirectContext:
    edge: ConnectionEdge
    source_record: Any
    src_record: Any
    dst_record: Any


@dataclass
class TclCommand:
    raw: str
    line_no: int


@dataclass
class DelayEvidence:
    inst_name: str
    port_name: str
    constraint_type: str
    clock_name: str
    min_value: str
    max_value: str
    bare_value: str
    complex_options: str
    source_sdc_file: str
    source_line: str
    source_digest: str
    parse_status: str
    original_command: str
    message: str


@dataclass
class FormRow:
    row_idx: int
    values: Dict[str, object]


@dataclass(frozen=True)
class PortKey:
    inst_name: str
    direction: str
    port_name: str


@dataclass
class PendingPlan:
    pending_updates: Dict[Path, List[str]] = field(default_factory=dict)
    removed_log_lines: List[str] = field(default_factory=list)
    removed_count: int = 0


@dataclass
class EmissionBlock:
    rows: List[FormRow]
    commands: List[str]
    merged: bool = False


class Report:
    def __init__(self) -> None:
        self.lines: List[str] = []
        self.warning_count = 0
        self.error_count = 0
        self.sync_changed = False

    def info(self, msg: str) -> None:
        self.lines.append("INFO: " + msg)

    def warn(self, msg: str) -> None:
        self.warning_count += 1
        self.lines.append("WARNING: " + msg)

    def error(self, msg: str) -> None:
        self.error_count += 1
        self.lines.append("ERROR: " + msg)


def _author_part_a() -> str:
    return chr(72) + chr(111)


def _author_part_b() -> str:
    return chr(119) + chr(97)


def _author_part_c() -> str:
    return chr(114) + chr(100)


def author_name() -> str:
    return _author_part_a() + _author_part_b() + _author_part_c()


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text


def normalize_key(value) -> str:
    return clean_cell(value).lower()


def parse_scenario_scope(value) -> Optional[List[str]]:
    raw = clean_cell(value)
    if not raw:
        return None
    tokens = [normalize_key(token) for token in re.split(r"[,;|\s]+", raw) if clean_cell(token)]
    if not tokens or any(token not in SCENARIOS for token in tokens):
        return None
    if len(tokens) != len(set(tokens)) or tokens != sorted(tokens):
        return None
    return tokens


def has_versioned_policy_basis(value) -> bool:
    return bool(VERSIONED_POLICY_RE.search(clean_cell(value)))


def safe_token(value) -> str:
    token = re.sub(r"[^A-Za-z0-9_.-]+", "_", clean_cell(value)).strip("_")
    return token or "unknown"


def parse_number(value) -> Optional[float]:
    text = clean_cell(value)
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def format_number(value) -> str:
    number = parse_number(value)
    return "" if number is None else "{0:.12g}".format(number)


def join_unique(values: Iterable[str], separator: str = "; ") -> str:
    result: List[str] = []
    for value in values:
        text = clean_cell(value)
        if text and text not in result:
            result.append(text)
    return separator.join(result)


def comment_text(value) -> str:
    text = clean_cell(value)
    text = re.sub(r"[\x00-\x1f\x7f]+", " ", text)
    return text.replace("\\", "/").strip()


def append_chunked_comment(
    lines: List[str], label: str, values: Sequence[str], chunk_size: int = 8
) -> None:
    cleaned = [comment_text(value) for value in values if comment_text(value)]
    for offset in range(0, len(cleaned), chunk_size):
        suffix = "" if offset == 0 else " (cont.)"
        lines.append("# {0}{1}={2}".format(label, suffix, ",".join(cleaned[offset:offset + chunk_size])))


def machine_digest(values: Dict[str, object]) -> str:
    payload = "\n".join(
        "{0}={1}".format(header, clean_cell(values.get(header)))
        for header in sorted(REVIEW_INVALIDATING_MACHINE_HEADERS)
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def atomic_write_text(path: Path, text: str, encoding: str = "utf-8") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(".{0}.tmp.{1}".format(path.name, os.getpid()))
    try:
        tmp.write_text(text, encoding=encoding)
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            tmp.unlink()


def atomic_write_csv(path: Path, headers: Sequence[str], rows: Sequence[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(".{0}.tmp.{1}".format(path.name, os.getpid()))
    try:
        with tmp.open("w", encoding="utf-8", newline="") as file_obj:
            writer = csv.DictWriter(file_obj, fieldnames=list(headers), extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({header: clean_cell(row.get(header)) for header in headers})
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            tmp.unlink()


def atomic_save_workbook(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(".{0}.tmp.{1}{2}".format(path.stem, os.getpid(), path.suffix))
    try:
        workbook.save(str(tmp))
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            tmp.unlink()


def digest_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_path(base: Path, value: Optional[str], default: str) -> Path:
    path = Path(value).expanduser() if value else Path(default)
    return path if path.is_absolute() else base / path


def port_base(port: str) -> str:
    text = clean_cell(port)
    match = PORT_EXACT_BIT_RE.fullmatch(text) or PORT_RANGE_RE.fullmatch(text)
    return match.group(1) if match else text


def port_bit(port: str, explicit: str = "") -> str:
    if clean_cell(explicit):
        return clean_cell(explicit)
    match = PORT_EXACT_BIT_RE.fullmatch(clean_cell(port))
    return match.group(2) if match else ""


def is_canonical_port(port: str) -> bool:
    text = clean_cell(port)
    return bool(text and not any(char in text for char in "*?") and PORT_BIT_RE.fullmatch(text))


def parse_feedthrough_port(port: str) -> Optional[Tuple[str, str, str]]:
    match = FEEDTHROUGH_RE.fullmatch(port_base(port))
    if not match:
        return None
    prefix, index, remainder = match.groups()
    return prefix, index or "single", remainder


def canonical_soc_object(inst_name: str, port_name: str) -> str:
    inst_key = normalize_key(inst_name)
    if inst_key == "top":
        return clean_cell(port_name)
    if inst_key in {"unknown", "constant", "const"}:
        return ""
    return "{0}/{1}".format(clean_cell(inst_name), clean_cell(port_name))


def validate_soc_object(
    connection_id: str,
    side: str,
    inst_name: str,
    port_name: str,
    soc_object: str,
    report: Report,
) -> None:
    provided = clean_cell(soc_object)
    expected = canonical_soc_object(inst_name, port_name)
    if expected and not provided:
        report.error(
            "{0}: {1}_soc_object is missing; expected canonical direct endpoint {2}".format(
                connection_id, side, expected
            )
        )
    elif provided and expected and provided != expected:
        report.error(
            "{0}: {1}_soc_object={2} does not match canonical direct endpoint {3}".format(
                connection_id, side, provided, expected
            )
        )
    elif provided and not expected:
        report.warn(
            "{0}: {1}_soc_object is ignored for non-addressable endpoint {2}".format(
                connection_id, side, inst_name
            )
        )


def endpoint_collection(inst_name: str, port_name: str, soc_object: str) -> str:
    del soc_object
    obj = canonical_soc_object(inst_name, port_name)
    if not obj:
        return ""
    if normalize_key(inst_name) == "top":
        return "[get_ports {{{0}}}]".format(obj)
    return "[get_pins {{{0}}}]".format(obj)


def read_harden_sdc_manifest(
    path: Path,
    run_root: Path,
    scenario: str,
    require_complete: bool,
    report: Report,
) -> Tuple[Dict[str, HardenSdc], RunCompleteness]:
    errors_before = report.error_count
    if not path.is_file():
        report.error("{0}: HARDEN_SDC_MANIFEST_MISSING".format(path))
        return {}, RunCompleteness(status="invalid", manifest_path=str(path))

    required = {"scenario", "inst_name", "module_name", "sdc_path", "availability_status", "note"}
    entries: Dict[str, HardenSdc] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = set(reader.fieldnames or [])
        missing_fields = sorted(required - fields)
        if missing_fields:
            report.error(
                "{0}: HARDEN_SDC_MANIFEST_SCHEMA_ERROR missing {1}".format(
                    path, ",".join(missing_fields)
                )
            )
        for row_idx, row in enumerate(reader, start=2):
            row_scenario = clean_cell(row.get("scenario"))
            inst_name = clean_cell(row.get("inst_name"))
            if not inst_name:
                report.error("{0} row {1}: empty inst_name".format(path.name, row_idx))
                continue
            if row_scenario != scenario:
                report.error(
                    "{0} row {1}: scenario={2} does not match requested {3}".format(
                        path.name, row_idx, row_scenario or "<empty>", scenario
                    )
                )
                continue
            if inst_name in entries:
                report.error("{0} row {1}: duplicate inst_name {2}".format(path.name, row_idx, inst_name))
                continue
            status = normalize_key(row.get("availability_status"))
            sdc_value = clean_cell(row.get("sdc_path"))
            entry = HardenSdc(
                scenario=row_scenario,
                inst_name=inst_name,
                module_name=clean_cell(row.get("module_name")),
                sdc_path=sdc_value,
                availability_status=status,
                note=clean_cell(row.get("note")),
                source_row=row_idx,
            )
            entries[inst_name] = entry

    available: List[str] = []
    missing: List[str] = []
    not_required: List[str] = []
    resolved_owners: Dict[str, str] = {}
    for inst_name in sorted(entries):
        entry = entries[inst_name]
        if entry.availability_status == "available":
            if not entry.sdc_path:
                report.error("{0} row {1}: available {2} has empty sdc_path".format(path.name, entry.source_row, inst_name))
                continue
            resolved = resolve_path(run_root, entry.sdc_path, entry.sdc_path)
            if not resolved.is_file():
                report.error(
                    "{0} row {1}: available SDC missing for {2}: {3}".format(
                        path.name, entry.source_row, inst_name, resolved
                    )
                )
                continue
            try:
                with resolved.open("rb") as file_obj:
                    file_obj.read(1)
            except OSError as exc:
                report.error(
                    "{0} row {1}: available SDC is not readable for {2}: {3}".format(
                        path.name, entry.source_row, inst_name, exc
                    )
                )
                continue
            entry.resolved_path = resolved.resolve()
            resolved_key = str(entry.resolved_path)
            previous_owner = resolved_owners.get(resolved_key)
            if previous_owner and previous_owner != inst_name:
                report.error(
                    "{0}: available SDC path is shared by {1} and {2}: {3}".format(
                        path.name, previous_owner, inst_name, entry.resolved_path
                    )
                )
                continue
            resolved_owners[resolved_key] = inst_name
            available.append(inst_name)
        elif entry.availability_status == "missing":
            missing.append(inst_name)
            report.warn(
                "{0} row {1}: HARDEN_SDC_MISSING {2}: {3}".format(
                    path.name, entry.source_row, inst_name, entry.note or "<no note>"
                )
            )
        elif entry.availability_status == "not_required":
            if entry.sdc_path:
                report.error(
                    "{0} row {1}: not_required {2} must have empty sdc_path".format(
                        path.name, entry.source_row, inst_name
                    )
                )
            if not entry.note:
                report.error(
                    "{0} row {1}: not_required {2} requires an explicit note/basis".format(
                        path.name, entry.source_row, inst_name
                    )
                )
            not_required.append(inst_name)
        else:
            report.error(
                "{0} row {1}: invalid availability_status {2} for {3}".format(
                    path.name, entry.source_row, entry.availability_status or "<empty>", inst_name
                )
            )

    status = "partial" if missing else "complete"
    if report.error_count > errors_before:
        status = "invalid"
    if require_complete and missing:
        report.error("HARDEN_SDC_COMPLETENESS_REQUIRED: " + ",".join(missing))
    completeness = RunCompleteness(
        status=status,
        available_instances=available,
        missing_instances=missing,
        not_required_instances=not_required,
        manifest_path=str(path.resolve()),
    )
    report.info(
        "harden SDC completeness={0} available={1} missing={2} not_required={3}".format(
            completeness.status,
            completeness.available_count,
            completeness.missing_count,
            completeness.not_required_count,
        )
    )
    return entries, completeness


def read_connection_inventory(
    path: Path,
    report: Report,
    scenario: str,
    require_target_schema: bool = False,
) -> List[ConnectionEdge]:
    if not path.is_file():
        report.error("connection inventory not found: {0}".format(path))
        return []
    required = {
        "connection_id",
        "connection_type",
        "src_instance",
        "src_direction",
        "src_port",
        "dst_instance",
        "dst_direction",
        "dst_port",
        "validation_status",
    }
    if require_target_schema:
        required.update(
            {
                "schema_version", "scenario_scope",
                "src_bit_index", "src_endpoint_key", "src_soc_object",
                "dst_bit_index", "dst_endpoint_key", "dst_soc_object",
            }
        )
    edges: List[ConnectionEdge] = []
    seen: Set[str] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = set(reader.fieldnames or [])
        missing = sorted(required - fields)
        if missing:
            report.error("{0}: missing field(s): {1}".format(path, ",".join(missing)))
            return edges
        has_scenario_scope = "scenario_scope" in fields
        has_schema_version = "schema_version" in fields
        for row_idx, row in enumerate(reader, start=2):
            scope_tokens = (
                parse_scenario_scope(row.get("scenario_scope"))
                if has_scenario_scope else ["common"]
            )
            if scope_tokens is None:
                report.error(
                    "{0} row {1}: invalid or unstable scenario_scope={2}".format(
                        path.name, row_idx, clean_cell(row.get("scenario_scope")) or "<empty>"
                    )
                )
                continue
            if "common" not in scope_tokens and scenario not in scope_tokens:
                continue
            schema_version = clean_cell(row.get("schema_version")) if has_schema_version else SCHEMA_VERSION
            if schema_version != SCHEMA_VERSION:
                report.error(
                    "{0} row {1}: unsupported schema_version={2}".format(
                        path.name, row_idx, schema_version or "<empty>"
                    )
                )
                continue
            connection_id = clean_cell(row.get("connection_id"))
            if not connection_id:
                report.error("{0} row {1}: empty connection_id".format(path.name, row_idx))
                continue
            if connection_id in seen:
                report.error("{0} row {1}: duplicate connection_id {2}".format(path.name, row_idx, connection_id))
                continue
            seen.add(connection_id)
            connection_type = normalize_key(row.get("connection_type"))
            if connection_type not in CANONICAL_CONNECTION_TYPES:
                report.error(
                    "{0} row {1}: connection_type={2} is outside the canonical 00 enum".format(
                        path.name, row_idx, connection_type or "<empty>"
                    )
                )
                continue
            src_port = clean_cell(row.get("src_port"))
            dst_port = clean_cell(row.get("dst_port"))
            src_instance = clean_cell(row.get("src_instance"))
            src_direction = normalize_key(row.get("src_direction"))
            dst_instance = clean_cell(row.get("dst_instance"))
            dst_direction = normalize_key(row.get("dst_direction"))
            src_endpoint_key = clean_cell(row.get("src_endpoint_key"))
            dst_endpoint_key = clean_cell(row.get("dst_endpoint_key"))
            expected_src_key = "{0}:{1}:{2}".format(src_instance, src_direction, src_port)
            expected_dst_key = "{0}:{1}:{2}".format(dst_instance, dst_direction, dst_port)
            if require_target_schema and not src_endpoint_key:
                report.error("{0} row {1}: empty src_endpoint_key".format(path.name, row_idx))
            elif src_endpoint_key and src_endpoint_key != expected_src_key:
                report.error(
                    "{0} row {1}: src_endpoint_key={2} does not match {3}".format(
                        path.name, row_idx, src_endpoint_key, expected_src_key
                    )
                )
            if require_target_schema and not dst_endpoint_key:
                report.error("{0} row {1}: empty dst_endpoint_key".format(path.name, row_idx))
            elif dst_endpoint_key and dst_endpoint_key != expected_dst_key:
                report.error(
                    "{0} row {1}: dst_endpoint_key={2} does not match {3}".format(
                        path.name, row_idx, dst_endpoint_key, expected_dst_key
                    )
                )
            src_explicit_bit = clean_cell(row.get("src_bit_index"))
            dst_explicit_bit = clean_cell(row.get("dst_bit_index"))
            src_inferred_bit = port_bit(src_port)
            dst_inferred_bit = port_bit(dst_port)
            if src_explicit_bit and src_explicit_bit != src_inferred_bit:
                report.error(
                    "{0} row {1}: src_bit_index={2} disagrees with {3}".format(
                        path.name, row_idx, src_explicit_bit, src_port
                    )
                )
            if dst_explicit_bit and dst_explicit_bit != dst_inferred_bit:
                report.error(
                    "{0} row {1}: dst_bit_index={2} disagrees with {3}".format(
                        path.name, row_idx, dst_explicit_bit, dst_port
                    )
                )
            edge = ConnectionEdge(
                schema_version=schema_version,
                connection_id=connection_id,
                scenario_scope=",".join(scope_tokens),
                connection_type=connection_type,
                src_instance=src_instance,
                src_direction=src_direction,
                src_port=src_port,
                src_bit_index=src_explicit_bit or src_inferred_bit,
                src_endpoint_key=src_endpoint_key,
                src_soc_object=clean_cell(row.get("src_soc_object")),
                dst_instance=dst_instance,
                dst_direction=dst_direction,
                dst_port=dst_port,
                dst_bit_index=dst_explicit_bit or dst_inferred_bit,
                dst_endpoint_key=dst_endpoint_key,
                dst_soc_object=clean_cell(row.get("dst_soc_object")),
                validation_status=normalize_key(row.get("validation_status")),
                note=clean_cell(row.get("note")),
                source_row=row_idx,
            )
            edges.append(edge)
    report.info(
        "loaded {0} effective direct connection edge(s) for scenario={1} from {2}".format(
            len(edges), scenario, path
        )
    )
    return edges


def validate_feedthrough_direction(inst: str, direction: str, port: str, report: Report) -> None:
    parsed = parse_feedthrough_port(port)
    if not parsed:
        return
    prefix = parsed[0]
    expected = "input" if prefix == "fti" else "output"
    if direction != expected:
        report.error(
            "{0}/{1}: {2}_* feedthrough boundary must be {3}, got {4}".format(
                inst, port, prefix, expected, direction or "<empty>"
            )
        )


def classify_feedthrough_edges(
    edges: Sequence[ConnectionEdge],
    manifest: Dict[str, HardenSdc],
    report: Report,
) -> List[FeedthroughEdge]:
    owned: List[FeedthroughEdge] = []
    for edge in edges:
        if edge.connection_type in EXCLUDED_CONNECTION_TYPES:
            continue

        src_ft = parse_feedthrough_port(edge.src_port)
        dst_ft = parse_feedthrough_port(edge.dst_port)
        if edge.connection_type == "feedthrough_candidate" and not (src_ft or dst_ft):
            report.warn(
                "{0}: feedthrough_candidate edge has no fti_/fto_ boundary prefix".format(
                    edge.connection_id
                )
            )
        src_is_fto = bool(src_ft and src_ft[0] == "fto")
        dst_is_fti = bool(dst_ft and dst_ft[0] == "fti")
        if not (src_is_fto or dst_is_fti):
            continue

        validate_feedthrough_direction(edge.src_instance, edge.src_direction, edge.src_port, report)
        validate_feedthrough_direction(edge.dst_instance, edge.dst_direction, edge.dst_port, report)

        boundary_instances: List[Tuple[str, str]] = []
        if src_is_fto:
            boundary_instances.append((edge.src_instance, edge.src_port))
        if dst_is_fti:
            boundary_instances.append((edge.dst_instance, edge.dst_port))
        for inst_name, port_name in boundary_instances:
            if normalize_key(inst_name) in PSEUDO_INSTANCES or inst_name not in manifest:
                report.error(
                    "{0}: feedthrough boundary {1}/{2} is not owned by a manifest harden".format(
                        edge.connection_id, inst_name or "<empty>", port_name
                    )
                )

        validate_soc_object(
            edge.connection_id, "src", edge.src_instance, edge.src_port,
            edge.src_soc_object, report,
        )
        validate_soc_object(
            edge.connection_id, "dst", edge.dst_instance, edge.dst_port,
            edge.dst_soc_object, report,
        )

        if not is_canonical_port(edge.src_port) or not is_canonical_port(edge.dst_port):
            report.error(
                "{0}: owned edge endpoints must be canonical scalar/bit keys: {1} -> {2}".format(
                    edge.connection_id, edge.src_port, edge.dst_port
                )
            )
        for inst_name in (edge.src_instance, edge.dst_instance):
            if normalize_key(inst_name) in PSEUDO_INSTANCES:
                continue
            if inst_name not in manifest:
                report.error("{0}: endpoint instance {1} missing from harden SDC manifest".format(edge.connection_id, inst_name))

        if src_is_fto and dst_is_fti:
            side = "between_feedthroughs"
            ft_instance = edge.src_instance + "," + edge.dst_instance
            ft_port = edge.src_port + "," + edge.dst_port
        elif dst_is_fti:
            side = "ingress"
            ft_instance = edge.dst_instance
            ft_port = edge.dst_port
        else:
            side = "egress"
            ft_instance = edge.src_instance
            ft_port = edge.src_port

        validation_status = (
            "matched"
            if edge.validation_status in MATCHED_STATUSES and edge.connection_type != "unknown"
            else "needs_review"
        )
        notes: List[str] = []
        if validation_status != "matched":
            notes.append("00 validation_status={0}".format(edge.validation_status or "<empty>"))
        if edge.connection_type == "unknown":
            notes.append("00 connection_type=unknown; 01/04 ownership exclusion is not proven")
            report.warn(
                "{0}: connection_type=unknown requires ownership review before terminal disposition".format(
                    edge.connection_id
                )
            )
        if edge.note:
            notes.append(edge.note)
        feedthrough_edge_id = "FTE_" + edge.connection_id
        owned.append(
            FeedthroughEdge(
                edge=edge,
                feedthrough_edge_id=feedthrough_edge_id,
                feedthrough_instance=ft_instance,
                feedthrough_port=ft_port,
                feedthrough_side=side,
                validation_status=validation_status,
                machine_note="; ".join(notes),
            )
        )

        if edge.src_bit_index and edge.dst_bit_index and edge.src_bit_index != edge.dst_bit_index:
            report.warn(
                "{0}: direct edge renumbers bit {1} -> {2}; retained as explicit bit connection".format(
                    edge.connection_id, edge.src_bit_index, edge.dst_bit_index
                )
            )

    boundary_edges: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    fti_keys: Set[Tuple[str, str, str, str]] = set()
    fto_keys: Set[Tuple[str, str, str, str]] = set()
    for item in owned:
        edge = item.edge
        for inst_name, port_name in (
            (edge.src_instance, edge.src_port),
            (edge.dst_instance, edge.dst_port),
        ):
            parsed = parse_feedthrough_port(port_name)
            if not parsed:
                continue
            prefix, index, remainder = parsed
            boundary_edges[(inst_name, port_name)].append(edge.connection_id)
            key = (inst_name, index, remainder, port_bit(port_name))
            if prefix == "fti":
                fti_keys.add(key)
            else:
                fto_keys.add(key)
    for (inst_name, port_name), connection_ids in sorted(boundary_edges.items()):
        if len(connection_ids) > 1:
            report.warn(
                "{0}/{1}: feedthrough boundary participates in multiple direct edges: {2}".format(
                    inst_name, port_name, ",".join(sorted(connection_ids))
                )
            )
    for key in sorted(fti_keys - fto_keys):
        report.warn(
            "{0}: optional fti/fto report pair missing fto for index={1} remainder={2} bit={3}".format(
                key[0], key[1], key[2], key[3] or "scalar"
            )
        )
    for key in sorted(fto_keys - fti_keys):
        report.warn(
            "{0}: optional fti/fto report pair missing fti for index={1} remainder={2} bit={3}".format(
                key[0], key[1], key[2], key[3] or "scalar"
            )
        )
    owned.sort(key=lambda item: item.edge.connection_id)
    report.info("classified {0} feedthrough-owned direct edge(s)".format(len(owned)))
    return owned


def is_escaped(text: str, index: int) -> bool:
    count = 0
    pos = index - 1
    while pos >= 0 and text[pos] == "\\":
        count += 1
        pos -= 1
    return count % 2 == 1


def strip_inline_comment(line: str) -> str:
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(line):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(line, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == "#" and brace_depth == 0 and bracket_depth == 0:
                if idx == 0 or line[idx - 1].isspace() or line[idx - 1] == ";":
                    return line[:idx].rstrip()
    return line.rstrip()


def split_semicolon_commands(text: str) -> List[str]:
    parts: List[str] = []
    start = 0
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(text):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(text, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == ";" and brace_depth == 0 and bracket_depth == 0:
                parts.append(text[start:idx])
                start = idx + 1
    parts.append(text[start:])
    return parts


def tcl_command_complete(text: str) -> bool:
    quote = False
    brace_depth = 0
    bracket_depth = 0
    idx = 0
    while idx < len(text):
        char = text[idx]
        if char == "\\":
            idx += 2
            continue
        if char == '"' and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}":
                if brace_depth == 0:
                    raise ValueError("unexpected closing brace")
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]":
                if bracket_depth == 0:
                    raise ValueError("unexpected closing bracket")
                bracket_depth -= 1
        idx += 1
    return not quote and brace_depth == 0 and bracket_depth == 0


def iter_tcl_commands(text: str) -> Iterable[TclCommand]:
    buffer = ""
    start_line = 0
    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = strip_inline_comment(raw_line)
        if not line.strip():
            continue
        if not buffer:
            start_line = line_no
        stripped = line.rstrip()
        continued = stripped.endswith("\\") and not is_escaped(stripped, len(stripped) - 1)
        if continued:
            buffer += stripped[:-1].rstrip() + " "
            continue
        buffer += stripped + " "
        if not tcl_command_complete(buffer):
            continue
        for command in split_semicolon_commands(buffer):
            cleaned = command.strip().rstrip(";").strip()
            if cleaned:
                yield TclCommand(cleaned, start_line)
        buffer = ""
        start_line = 0
    if buffer.strip():
        if not tcl_command_complete(buffer):
            raise ValueError("unterminated Tcl command starting at line {0}".format(start_line))
        for command in split_semicolon_commands(buffer):
            cleaned = command.strip().rstrip(";").strip()
            if cleaned:
                yield TclCommand(cleaned, start_line)


def find_matching(text: str, start: int, open_char: str, close_char: str) -> int:
    depth = 0
    quote = False
    idx = start
    while idx < len(text):
        char = text[idx]
        if char == "\\":
            idx += 2
            continue
        if char == '"' and not is_escaped(text, idx):
            quote = not quote
        elif not quote:
            if char == open_char:
                depth += 1
            elif char == close_char:
                depth -= 1
                if depth == 0:
                    return idx
        idx += 1
    return -1


def tokenize_tcl_words(text: str) -> List[str]:
    tokens: List[str] = []
    idx = 0
    while idx < len(text):
        while idx < len(text) and text[idx].isspace():
            idx += 1
        if idx >= len(text):
            break
        start = idx
        if text[idx] == "{":
            end = find_matching(text, idx, "{", "}")
            if end < 0:
                tokens.append(text[start:])
                break
            tokens.append(text[start:end + 1])
            idx = end + 1
            continue
        if text[idx] == '"':
            idx += 1
            while idx < len(text):
                if text[idx] == "\\":
                    idx += 2
                    continue
                if text[idx] == '"':
                    idx += 1
                    break
                idx += 1
            tokens.append(text[start:idx])
            continue
        pieces: List[str] = []
        while idx < len(text) and not text[idx].isspace():
            if text[idx] == "\\":
                pieces.append(text[idx:idx + 2])
                idx += 2
            elif text[idx] == "[":
                end = find_matching(text, idx, "[", "]")
                if end < 0:
                    pieces.append(text[idx:])
                    idx = len(text)
                else:
                    pieces.append(text[idx:end + 1])
                    idx = end + 1
            else:
                pieces.append(text[idx])
                idx += 1
        tokens.append("".join(pieces))
    return tokens


def strip_braces(text: str) -> str:
    value = clean_cell(text)
    if len(value) >= 2 and value[0] == "{" and value[-1] == "}":
        return value[1:-1].strip()
    if len(value) >= 2 and value[0] == '"' and value[-1] == '"':
        return value[1:-1].strip()
    return value


def split_object_list(text: str) -> List[str]:
    return [part for part in re.split(r"[\s,;]+", strip_braces(text)) if part]


def parse_collection(token: str) -> Optional[Tuple[str, List[str]]]:
    text = clean_cell(token)
    if not (text.startswith("[") and text.endswith("]")):
        return None
    if find_matching(text, 0, "[", "]") != len(text) - 1:
        return None
    words = tokenize_tcl_words(text[1:-1].strip())
    if not words or words[0] not in {"get_ports", "get_pins"}:
        return None
    objects: List[str] = []
    for word in words[1:]:
        if not word.startswith("-"):
            objects.extend(split_object_list(word))
    return words[0], objects


def option_value(tokens: Sequence[str], option: str) -> str:
    for idx, token in enumerate(tokens):
        if token == option and idx + 1 < len(tokens):
            return strip_braces(tokens[idx + 1])
    return ""


def positional_value(tokens: Sequence[str]) -> str:
    skip = False
    for token in tokens[1:]:
        if skip:
            skip = False
            continue
        if token in {"-clock", "-min", "-max"}:
            skip = True
            continue
        if token.startswith("-") or parse_collection(token):
            continue
        return strip_braces(token)
    return ""


def extract_clock_name(tokens: Sequence[str]) -> str:
    value = option_value(tokens, "-clock")
    if value.startswith("["):
        words = tokenize_tcl_words(value[1:-1])
        if words and words[0] == "get_clocks" and len(words) > 1:
            return strip_braces(words[-1])
    return strip_braces(value)


def target_collection(tokens: Sequence[str]) -> Optional[Tuple[str, List[str]]]:
    found = None
    for token in tokens:
        parsed = parse_collection(token)
        if parsed:
            found = parsed
    return found


def expand_port_expression(value: str) -> Optional[List[str]]:
    text = clean_cell(value)
    if any(char in text for char in "*?"):
        return None
    match = PORT_RANGE_RE.fullmatch(text)
    if not match:
        return [text]
    base = match.group(1)
    left = int(match.group(2))
    right = int(match.group(3))
    step = 1 if right >= left else -1
    return ["{0}[{1}]".format(base, bit) for bit in range(left, right + step, step)]


def parse_delay_evidence(
    inst_name: str,
    command: TclCommand,
    source_path: Path,
    source_digest: str,
) -> List[DelayEvidence]:
    tokens = tokenize_tcl_words(command.raw)
    if not tokens or tokens[0] not in {"set_input_delay", "set_output_delay"}:
        return []
    constraint_type = "input_delay" if tokens[0] == "set_input_delay" else "output_delay"
    collection = target_collection(tokens)
    if not collection:
        return [
            DelayEvidence(
                inst_name, "", constraint_type, extract_clock_name(tokens),
                option_value(tokens, "-min"), option_value(tokens, "-max"), positional_value(tokens),
                "", str(source_path), str(command.line_no), source_digest, "needs_review",
                command.raw, "no get_ports/get_pins target collection",
            )
        ]
    kind, objects = collection
    results: List[DelayEvidence] = []
    complex_options = ",".join(
        option for option in ("-add_delay", "-clock_fall", "-rise", "-fall") if option in tokens
    )
    for obj in objects:
        port_expr = strip_braces(obj).split("/")[-1] if kind == "get_pins" else strip_braces(obj)
        expanded = expand_port_expression(port_expr)
        if expanded is None:
            results.append(
                DelayEvidence(
                    inst_name, port_expr, constraint_type, extract_clock_name(tokens),
                    option_value(tokens, "-min"), option_value(tokens, "-max"), positional_value(tokens),
                    complex_options, str(source_path), str(command.line_no), source_digest,
                    "needs_review", command.raw, "wildcard target cannot be mapped to canonical bit keys",
                )
            )
            continue
        for port_name in expanded:
            results.append(
                DelayEvidence(
                    inst_name, port_name, constraint_type, extract_clock_name(tokens),
                    option_value(tokens, "-min"), option_value(tokens, "-max"), positional_value(tokens),
                    complex_options, str(source_path), str(command.line_no), source_digest,
                    "ok" if not complex_options else "needs_review", command.raw,
                    "" if not complex_options else "complex delay options require review",
                )
            )
    return results


def extract_delay_evidence(manifest: Dict[str, HardenSdc], report: Report) -> List[DelayEvidence]:
    evidence: List[DelayEvidence] = []
    for inst_name in sorted(manifest):
        entry = manifest[inst_name]
        if entry.availability_status != "available" or entry.resolved_path is None:
            continue
        try:
            text = entry.resolved_path.read_text(encoding="utf-8")
            source_digest = digest_file(entry.resolved_path)
        except Exception as exc:
            report.error("failed to read {0}: {1}".format(entry.resolved_path, exc))
            continue
        count = 0
        try:
            for command in iter_tcl_commands(text):
                parsed = parse_delay_evidence(inst_name, command, entry.resolved_path, source_digest)
                evidence.extend(parsed)
                count += len(parsed)
        except ValueError as exc:
            report.error("failed to parse {0}: {1}".format(entry.resolved_path, exc))
        report.info("extracted {0} delay evidence record(s) from {1}".format(count, entry.resolved_path))
    return evidence


def normalize_clock_object(value: str) -> str:
    text = clean_cell(value)
    if text.startswith("clock:") or "/" in text or "." not in text:
        return text
    inst_name, port_name = text.split(".", 1)
    return inst_name + "/" + port_name


def read_diagnostic_meta(path: Path, label: str, report: Report) -> Optional[Dict[str, object]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        report.warn("{0} diagnostic meta invalid: {1}: {2}".format(label, path, exc))
        return None
    if not isinstance(payload, dict):
        report.warn("{0} diagnostic meta must contain a JSON object: {1}".format(label, path))
        return None
    return payload


def read_clock_inventory(
    path: Path,
    report: Report,
    require_active_action: bool = False,
    meta_path: Optional[Path] = None,
    expected_scenario: str = "",
) -> ClockInventory:
    inventory = ClockInventory()
    paired_meta = meta_path or path.with_suffix(".meta")
    if not path.is_file() and not paired_meta.is_file():
        report.warn("optional clock diagnostic unavailable: {0}".format(path))
        return inventory
    if not path.is_file() or not paired_meta.is_file():
        report.warn(
            "optional clock diagnostic requires paired CSV/meta: {0}, {1}".format(
                path, paired_meta
            )
        )
        inventory.status = "invalid"
        return inventory

    meta = read_diagnostic_meta(paired_meta, "clock", report)
    if meta is None:
        inventory.status = "invalid"
        return inventory
    try:
        inventory.inventory_digest = digest_file(path)
        inventory.meta_digest = digest_file(paired_meta)
    except OSError as exc:
        report.warn("optional clock diagnostic is not readable: {0}".format(exc))
        inventory.status = "invalid"
        return inventory
    invalid_problems: List[str] = []
    stale_problems: List[str] = []
    if expected_scenario and clean_cell(meta.get("scenario")) != expected_scenario:
        stale_problems.append(
            "meta scenario mismatch: {0} != requested {1}".format(
                clean_cell(meta.get("scenario")) or "<empty>", expected_scenario
            )
        )
    if clean_cell(meta.get("inventory_digest")) != inventory.inventory_digest:
        stale_problems.append("CSV digest does not match meta inventory_digest")
    inventory.final_sdc_digest = clean_cell(meta.get("final_sdc_digest"))
    final_sdc_value = clean_cell(meta.get("final_sdc_path"))
    if not inventory.final_sdc_digest or not final_sdc_value:
        stale_problems.append("meta final_sdc_path/final_sdc_digest is incomplete")
    else:
        final_sdc_path = Path(final_sdc_value).expanduser()
        if not final_sdc_path.is_absolute():
            final_sdc_path = paired_meta.parent / final_sdc_path
        if not final_sdc_path.is_file():
            stale_problems.append("final clock SDC is missing: {0}".format(final_sdc_path))
        else:
            try:
                actual_final_digest = digest_file(final_sdc_path)
            except OSError as exc:
                stale_problems.append("final clock SDC is not readable: {0}".format(exc))
            else:
                if actual_final_digest != inventory.final_sdc_digest:
                    stale_problems.append("final clock SDC digest does not match meta")

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.DictReader(file_obj)
            fields = set(reader.fieldnames or [])
            if "clock_name" not in fields:
                invalid_problems.append("missing clock_name column")
            if require_active_action and "final_action" not in fields:
                invalid_problems.append("missing final_action column")
            rows = [(row_idx, dict(row)) for row_idx, row in enumerate(reader, start=2)]
    except (OSError, csv.Error) as exc:
        invalid_problems.append("could not parse CSV: {0}".format(exc))
        rows = []
        fields = set()

    active_rows: List[Tuple[int, Dict[str, str]]] = []
    check_rows: List[Tuple[int, Dict[str, str]]] = []
    object_to_clock: Dict[str, str] = {}

    def add_alias(inst_name: str, alias: str, final_name: str, row_idx: int) -> None:
        if not alias:
            return
        key = (inst_name, alias)
        previous = inventory.aliases.get(key)
        if previous and previous != final_name:
            invalid_problems.append(
                "row {0}: clock alias {1}/{2} maps to both {3} and {4}".format(
                    row_idx, inst_name or "<global>", alias, previous, final_name
                )
            )
            return
        inventory.aliases[key] = final_name

    for row_idx, row in rows:
        name = clean_cell(row.get("clock_name"))
        if not name:
            continue
        action = normalize_key(row.get("final_action"))
        if "final_action" not in fields or action in ACTIVE_CLOCK_ACTIONS:
            active_rows.append((row_idx, row))
        elif action == "check_only":
            check_rows.append((row_idx, row))

    for row_idx, row in active_rows:
        name = clean_cell(row.get("clock_name"))
        action = normalize_key(row.get("final_action"))
        if name in inventory.active_names:
            invalid_problems.append("row {0}: duplicate active clock_name {1}".format(row_idx, name))
            continue
        inventory.active_names.add(name)
        inst_name = clean_cell(row.get("inst_name"))
        add_alias(inst_name, name, name, row_idx)
        add_alias(inst_name, clean_cell(row.get("original_clock_name")), name, row_idx)
        objects = {
            normalize_clock_object(row.get("target_object")),
            normalize_clock_object(
                "{0}/{1}".format(inst_name, clean_cell(row.get("port_name")))
                if inst_name and clean_cell(row.get("port_name")) else ""
            ),
        }
        if action in {"emit_top_clock", "emit_virtual_clock"}:
            objects.add(normalize_clock_object(row.get("direct_source")))
            if clean_cell(row.get("root_source")) == clean_cell(row.get("direct_source")):
                objects.add(normalize_clock_object(row.get("root_source")))
        for obj in (item for item in objects if item):
            previous = object_to_clock.get(obj)
            if previous and previous != name:
                invalid_problems.append(
                    "row {0}: clock object {1} maps to both {2} and {3}".format(
                        row_idx, obj, previous, name
                    )
                )
            object_to_clock[obj] = name

    unresolved = list(check_rows)
    for _ in range(len(unresolved) + 1):
        if not unresolved:
            break
        next_unresolved: List[Tuple[int, Dict[str, str]]] = []
        progress = False
        for row_idx, row in unresolved:
            upstream_values = [
                normalize_clock_object(row.get("from_whom")),
                normalize_clock_object(row.get("root_source")),
            ]
            resolved_names: Set[str] = set()
            for upstream in upstream_values:
                if upstream.startswith("clock:") and upstream[6:] in inventory.active_names:
                    resolved_names.add(upstream[6:])
                elif upstream in inventory.active_names:
                    resolved_names.add(upstream)
                elif upstream in object_to_clock:
                    resolved_names.add(object_to_clock[upstream])
            if len(resolved_names) > 1:
                invalid_problems.append(
                    "row {0}: check_only upstreams resolve to multiple active clocks: {1}".format(
                        row_idx, ",".join(sorted(resolved_names))
                    )
                )
                progress = True
                continue
            if not resolved_names:
                next_unresolved.append((row_idx, row))
                continue
            final_name = next(iter(resolved_names))
            inst_name = clean_cell(row.get("inst_name"))
            add_alias(inst_name, clean_cell(row.get("clock_name")), final_name, row_idx)
            add_alias(inst_name, clean_cell(row.get("original_clock_name")), final_name, row_idx)
            target_obj = normalize_clock_object(row.get("target_object"))
            if not target_obj and inst_name and clean_cell(row.get("port_name")):
                target_obj = normalize_clock_object(
                    "{0}/{1}".format(inst_name, clean_cell(row.get("port_name")))
                )
            if target_obj:
                previous = object_to_clock.get(target_obj)
                if previous and previous != final_name:
                    invalid_problems.append(
                        "row {0}: check_only target {1} maps to both {2} and {3}".format(
                            row_idx, target_obj, previous, final_name
                        )
                    )
                else:
                    object_to_clock[target_obj] = final_name
            progress = True
        unresolved = next_unresolved
        if not progress:
            break

    inventory.clock_set_digest = hashlib.sha256(
        "\n".join(sorted(inventory.active_names)).encode("utf-8")
    ).hexdigest()
    if clean_cell(meta.get("clock_set_digest")) != inventory.clock_set_digest:
        stale_problems.append("active clock set digest does not match meta")
    if clean_cell(meta.get("clock_count")) and clean_cell(meta.get("clock_count")) != str(len(inventory.active_names)):
        stale_problems.append("active clock count does not match meta")
    completeness_status = normalize_key(meta.get("run_completeness"))
    if completeness_status not in {"complete", "partial"}:
        invalid_problems.append(
            "invalid run_completeness {0}".format(completeness_status or "<empty>")
        )

    if invalid_problems:
        for problem in invalid_problems:
            report.warn("clock diagnostic invalid: {0}: {1}".format(path, problem))
        return ClockInventory(status="invalid")
    if stale_problems:
        for problem in stale_problems:
            report.warn("clock diagnostic stale: {0}: {1}".format(path, problem))
        return ClockInventory(status="stale")

    for row_idx, row in unresolved:
        inst_name = clean_cell(row.get("inst_name"))
        for alias in {
            clean_cell(row.get("clock_name")),
            clean_cell(row.get("original_clock_name")),
        }:
            if alias:
                inventory.unresolved_aliases.add((inst_name, alias))
        report.warn(
            "clock diagnostic unresolved: {0} row {1}: {2}/{3}".format(
                path.name, row_idx, inst_name or "<global>",
                clean_cell(row.get("original_clock_name")) or clean_cell(row.get("clock_name")),
            )
        )
    inventory.status = "complete" if completeness_status == "complete" else "incomplete"
    report.info(
        "loaded {0} active assembled clock(s) from {1}; diagnostic_status={2}".format(
            len(inventory.active_names), path, inventory.status
        )
    )
    return inventory


def read_relation_map(
    path: Path,
    report: Report,
    expected_scenario: str = "",
    meta_path: Optional[Path] = None,
    clock_inventory: Optional[ClockInventory] = None,
) -> RelationMap:
    result = RelationMap()
    paired_meta = meta_path or path.with_suffix(".meta")
    if not path.is_file() and not paired_meta.is_file():
        report.warn("optional relation diagnostic unavailable: {0}".format(path))
        return result
    if not path.is_file() or not paired_meta.is_file():
        report.warn(
            "optional relation diagnostic requires paired CSV/meta: {0}, {1}".format(
                path, paired_meta
            )
        )
        result.status = "invalid"
        return result
    if clock_inventory is None or clock_inventory.status not in {"complete", "incomplete"}:
        report.warn("relation diagnostic unavailable because clock diagnostic is not valid")
        result.status = (
            clock_inventory.status if clock_inventory is not None else "unavailable"
        )
        return result

    meta = read_diagnostic_meta(paired_meta, "relation", report)
    if meta is None:
        result.status = "invalid"
        return result
    invalid_problems: List[str] = []
    stale_problems: List[str] = []
    try:
        relation_digest = digest_file(path)
    except OSError as exc:
        report.warn("optional relation diagnostic is not readable: {0}".format(exc))
        result.status = "invalid"
        return result
    result.assembled_view_digest = clean_cell(meta.get("assembled_view_digest"))
    if expected_scenario and clean_cell(meta.get("scenario")) != expected_scenario:
        stale_problems.append(
            "meta scenario mismatch: {0} != requested {1}".format(
                clean_cell(meta.get("scenario")) or "<empty>", expected_scenario
            )
        )
    if clean_cell(meta.get("relation_map_digest")) != relation_digest:
        stale_problems.append("CSV digest does not match meta relation_map_digest")
    if clean_cell(meta.get("inventory_digest")) != clock_inventory.inventory_digest:
        stale_problems.append("01 inventory digest does not match relation meta")
    if clean_cell(meta.get("inventory_meta_digest")) != clock_inventory.meta_digest:
        stale_problems.append("01 inventory meta digest does not match relation meta")
    if clean_cell(meta.get("clock_universe_digest")) != clock_inventory.clock_set_digest:
        stale_problems.append("clock universe digest does not match 01 active clock set")
    if clean_cell(meta.get("final_clock_sdc_digest")) != clock_inventory.final_sdc_digest:
        stale_problems.append("final clock SDC digest does not match 01 meta")
    if not result.assembled_view_digest:
        invalid_problems.append("meta assembled_view_digest is empty")

    required = {
        "schema_version", "scenario", "clock_a", "clock_b", "relation_type",
        "relation_source", "source_rule_ids", "clock_universe_digest",
        "assembled_view_digest",
    }
    rows: List[Tuple[int, Dict[str, str]]] = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.DictReader(file_obj)
            fields = set(reader.fieldnames or [])
            if not required.issubset(fields):
                invalid_problems.append("missing column(s): {0}".format(",".join(sorted(required - fields))))
            rows = [(row_idx, dict(row)) for row_idx, row in enumerate(reader, start=2)]
    except (OSError, csv.Error) as exc:
        invalid_problems.append("could not parse CSV: {0}".format(exc))

    relations: Dict[Tuple[str, str], str] = {}
    for row_idx, row in rows:
        row_scenario = clean_cell(row.get("scenario"))
        clock_a = clean_cell(row.get("clock_a"))
        clock_b = clean_cell(row.get("clock_b"))
        relation = normalize_key(row.get("relation_type"))
        relation_source = normalize_key(row.get("relation_source"))
        source_rule_ids = clean_cell(row.get("source_rule_ids"))
        if expected_scenario and row_scenario != expected_scenario:
            stale_problems.append(
                "row {0}: scenario mismatch: {1} != requested {2}".format(
                    row_idx, row_scenario or "<empty>", expected_scenario
                )
            )
        if not clock_a or not clock_b or clock_a == clock_b:
            invalid_problems.append("row {0}: invalid clock pair".format(row_idx))
            continue
        if relation not in CANONICAL_RELATIONS:
            invalid_problems.append("row {0}: invalid relation_type {1}".format(row_idx, relation or "<empty>"))
        if relation_source not in CANONICAL_RELATION_SOURCES:
            invalid_problems.append("row {0}: invalid relation_source {1}".format(row_idx, relation_source or "<empty>"))
        elif relation_source == "default_synchronous":
            if relation != "synchronous" or source_rule_ids:
                invalid_problems.append(
                    "row {0}: default_synchronous must be synchronous with empty source_rule_ids".format(
                        row_idx
                    )
                )
        elif relation_source == "explicit_rule" and not source_rule_ids:
            invalid_problems.append(
                "row {0}: explicit_rule requires source_rule_ids".format(row_idx)
            )
        if clock_a not in clock_inventory.active_names or clock_b not in clock_inventory.active_names:
            stale_problems.append("row {0}: clock pair is outside the active 01 universe".format(row_idx))
        if clean_cell(row.get("clock_universe_digest")) != clock_inventory.clock_set_digest:
            stale_problems.append("row {0}: clock_universe_digest mismatch".format(row_idx))
        if clean_cell(row.get("assembled_view_digest")) != result.assembled_view_digest:
            stale_problems.append("row {0}: assembled_view_digest mismatch".format(row_idx))
        key = tuple(sorted((clock_a, clock_b)))
        if key in relations:
            invalid_problems.append("row {0}: duplicate relation pair {1}/{2}".format(row_idx, clock_a, clock_b))
        relations[key] = relation

    completeness_status = normalize_key(meta.get("run_completeness"))
    if completeness_status not in {"complete", "partial"}:
        invalid_problems.append(
            "invalid run_completeness {0}".format(completeness_status or "<empty>")
        )
    if invalid_problems:
        for problem in invalid_problems:
            report.warn("relation diagnostic invalid: {0}: {1}".format(path, problem))
        result.status = "invalid"
        return result
    if stale_problems:
        for problem in stale_problems:
            report.warn("relation diagnostic stale: {0}: {1}".format(path, problem))
        result.status = "stale"
        return result

    expected_pairs = set(itertools.combinations(sorted(clock_inventory.active_names), 2))
    actual_pairs = set(relations)
    incomplete_pairs = expected_pairs - actual_pairs
    if incomplete_pairs:
        report.warn(
            "relation diagnostic incomplete: {0} missing {1} active clock pair(s)".format(
                path, len(incomplete_pairs)
            )
        )
    result.relations = relations
    result.status = (
        "complete"
        if not incomplete_pairs
        and clock_inventory.status == "complete"
        and completeness_status == "complete"
        else "incomplete"
    )
    report.info(
        "loaded {0} clock relation pair(s) from {1}; diagnostic_status={2}".format(
            len(relations), path, result.status
        )
    )
    return result


def endpoint_sdc_status(inst_name: str, manifest: Dict[str, HardenSdc]) -> str:
    if normalize_key(inst_name) in PSEUDO_INSTANCES:
        return "not_required"
    entry = manifest.get(inst_name)
    return entry.availability_status if entry else "missing"


def evidence_index(evidence: Sequence[DelayEvidence]):
    index: Dict[Tuple[str, str, str], List[DelayEvidence]] = defaultdict(list)
    for item in evidence:
        index[(item.inst_name, item.port_name, item.constraint_type)].append(item)
    return index


def matching_evidence(
    index: Dict[Tuple[str, str, str], List[DelayEvidence]],
    inst_name: str,
    port_name: str,
    constraint_type: str,
) -> List[DelayEvidence]:
    result: List[DelayEvidence] = []
    seen: Set[int] = set()
    for candidate_port in (port_name, port_base(port_name)):
        for item in index.get((inst_name, candidate_port, constraint_type), []):
            if id(item) not in seen:
                seen.add(id(item))
                result.append(item)
    return result


def resolve_clock_relation(
    src_clocks: Sequence[str],
    dst_clocks: Sequence[str],
    clock_inventory: ClockInventory,
    relation_map: RelationMap,
) -> str:
    src = sorted(set(clock for clock in src_clocks if clock))
    dst = sorted(set(clock for clock in dst_clocks if clock))
    if len(src) != 1 or len(dst) != 1:
        return "unknown"
    if src[0] not in clock_inventory.active_names or dst[0] not in clock_inventory.active_names:
        return "unknown"
    if src[0] == dst[0]:
        return "synchronous"
    return relation_map.relations.get(tuple(sorted((src[0], dst[0]))), "unknown")


def resolve_evidence_clock(
    inst_name: str, clock_name: str, clock_inventory: ClockInventory
) -> str:
    local_name = clean_cell(clock_name)
    if not local_name:
        return ""
    mapped = clock_inventory.aliases.get((clean_cell(inst_name), local_name))
    if mapped:
        return mapped
    return "unresolved:" + local_name


def edge_seed(
    item: FeedthroughEdge,
    manifest: Dict[str, HardenSdc],
    completeness: RunCompleteness,
    evidence_by_key: Dict[Tuple[str, str, str], List[DelayEvidence]],
    clock_inventory: ClockInventory,
    relation_map: RelationMap,
    scenario: str,
    stage: str,
    corner: str,
    port_accounting: str,
    connection_inventory_path: Path,
) -> Dict[str, object]:
    edge = item.edge
    src_evidence = matching_evidence(evidence_by_key, edge.src_instance, edge.src_port, "output_delay")
    dst_evidence = matching_evidence(evidence_by_key, edge.dst_instance, edge.dst_port, "input_delay")
    related = src_evidence + dst_evidence
    src_status = endpoint_sdc_status(edge.src_instance, manifest)
    dst_status = endpoint_sdc_status(edge.dst_instance, manifest)

    if "missing" in {src_status, dst_status}:
        evidence_status = "incomplete_missing_sdc"
    elif any(e.parse_status != "ok" for e in related):
        evidence_status = "needs_review"
    elif related:
        evidence_status = "complete"
    else:
        evidence_status = "complete_no_delay_candidate"

    src_clocks = [
        resolve_evidence_clock(edge.src_instance, e.clock_name, clock_inventory)
        for e in src_evidence if e.clock_name
    ]
    dst_clocks = [
        resolve_evidence_clock(edge.dst_instance, e.clock_name, clock_inventory)
        for e in dst_evidence if e.clock_name
    ]
    relation = resolve_clock_relation(src_clocks, dst_clocks, clock_inventory, relation_map)
    notes: List[str] = [item.machine_note] if item.machine_note else []
    if clock_inventory.status != "complete" or relation_map.status != "complete":
        notes.append(
            "clock diagnostic status=clock:{0},relation:{1}".format(
                clock_inventory.status, relation_map.status
            )
        )
    for evidence_item in related:
        if evidence_item.message:
            notes.append(evidence_item.message)
    if not related:
        notes.append("no input/output delay evidence on direct-edge endpoints")

    values: Dict[str, object] = {header: "" for header in EDGE_HEADERS}
    values.update(
        {
            "schema_version": SCHEMA_VERSION,
            "run_completeness": completeness.status,
            "available_harden_count": completeness.available_count,
            "missing_harden_count": completeness.missing_count,
            "not_required_harden_count": completeness.not_required_count,
            "missing_instances": ",".join(completeness.missing_instances),
            "port_accounting": port_accounting,
            "connection_inventory_path": str(connection_inventory_path.resolve()),
            "harden_sdc_manifest_path": completeness.manifest_path,
            "scenario": scenario,
            "stage": stage,
            "corner": corner,
            "feedthrough_edge_id": item.feedthrough_edge_id,
            "connection_id": edge.connection_id,
            "scenario_scope": edge.scenario_scope,
            "connection_type": edge.connection_type,
            "src_instance": edge.src_instance,
            "src_direction": edge.src_direction,
            "src_port": edge.src_port,
            "src_bit_index": edge.src_bit_index,
            "src_endpoint_key": edge.src_endpoint_key,
            "src_soc_object": endpoint_collection(edge.src_instance, edge.src_port, edge.src_soc_object),
            "dst_instance": edge.dst_instance,
            "dst_direction": edge.dst_direction,
            "dst_port": edge.dst_port,
            "dst_bit_index": edge.dst_bit_index,
            "dst_endpoint_key": edge.dst_endpoint_key,
            "dst_soc_object": endpoint_collection(edge.dst_instance, edge.dst_port, edge.dst_soc_object),
            "feedthrough_instance": item.feedthrough_instance,
            "feedthrough_port": item.feedthrough_port,
            "feedthrough_side": item.feedthrough_side,
            "channel_disposition": "pending",
            "budget_scope": "direct_edge",
            "budget_model": "unknown",
            "budget_required": "",
            "src_output_delay_max": join_unique(e.max_value or e.bare_value for e in src_evidence),
            "src_output_delay_min": join_unique(e.min_value for e in src_evidence),
            "dst_input_delay_max": join_unique(e.max_value or e.bare_value for e in dst_evidence),
            "dst_input_delay_min": join_unique(e.min_value for e in dst_evidence),
            "emit_max": "no",
            "emit_min": "no",
            "datapath_only": "yes",
            "tool_surface": "dc",
            "src_sdc_status": src_status,
            "dst_sdc_status": dst_status,
            "evidence_status": evidence_status,
            "src_clock": join_unique(src_clocks),
            "dst_clock": join_unique(dst_clocks),
            "clock_relation": relation,
            "source_sdc_file": join_unique(e.source_sdc_file for e in related),
            "source_line": join_unique(e.source_line for e in related),
            "source_digest": join_unique(e.source_digest for e in related),
            "source_command": join_unique((e.original_command for e in related), " || "),
            "apply": "no",
            "review_status": "pending",
            "validation_status": item.validation_status,
            "machine_note": "; ".join(dict.fromkeys(note for note in notes if note)),
        }
    )
    values["machine_digest"] = machine_digest(values)
    return values


def build_edge_seeds(
    owned: Sequence[FeedthroughEdge],
    manifest: Dict[str, HardenSdc],
    completeness: RunCompleteness,
    evidence: Sequence[DelayEvidence],
    clock_inventory: ClockInventory,
    relation_map: RelationMap,
    scenario: str,
    stage: str,
    corner: str,
    port_accounting: str,
    connection_inventory_path: Path,
) -> List[Dict[str, object]]:
    index = evidence_index(evidence)
    return [
        edge_seed(
            item, manifest, completeness, index, clock_inventory, relation_map,
            scenario, stage, corner, port_accounting, connection_inventory_path,
        )
        for item in owned
    ]


def create_or_load_workbook(path: Path) -> Tuple[Workbook, bool]:
    if path.is_file():
        return load_workbook(str(path)), False
    workbook = Workbook()
    workbook.active.title = "feedthrough_edges"
    return workbook, True


def ensure_sheet(workbook: Workbook, name: str, headers: Sequence[str]) -> None:
    sheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)
    if sheet.max_row == 1 and all(sheet.cell(1, col).value is None for col in range(1, len(headers) + 1)):
        for col, header in enumerate(headers, start=1):
            sheet.cell(1, col, header)
    existing = [clean_cell(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    for header in headers:
        if header not in existing:
            sheet.cell(1, len(existing) + 1, header)
            existing.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = max(8, max(len(clean_cell(cell.value)) for cell in column_cells))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 44)
    ref = "A1:{0}{1}".format(get_column_letter(sheet.max_column), max(sheet.max_row, 2))
    if not sheet.tables:
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", name)[:24] + "_tbl"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        try:
            sheet.add_table(table)
        except ValueError:
            pass
    else:
        for table_name in list(sheet.tables):
            sheet.tables[table_name].ref = ref


def header_map(sheet) -> Dict[str, int]:
    return {
        clean_cell(cell.value): idx
        for idx, cell in enumerate(sheet[1], start=1)
        if clean_cell(cell.value)
    }


def sheet_row_values(sheet, row_idx: int, headers: Sequence[str]) -> Dict[str, object]:
    mapping = header_map(sheet)
    return {
        header: sheet.cell(row_idx, mapping[header]).value if header in mapping else ""
        for header in headers
    }


def row_key(values: Dict[str, object]) -> Tuple[str, str, str, str]:
    return (
        clean_cell(values.get("scenario")),
        clean_cell(values.get("stage")),
        clean_cell(values.get("corner")),
        clean_cell(values.get("connection_id")),
    )


def append_row(sheet, headers: Sequence[str], values: Dict[str, object], fill=None) -> None:
    mapping = header_map(sheet)
    row_idx = sheet.max_row + 1
    for header in headers:
        cell = sheet.cell(row_idx, mapping[header], values.get(header, ""))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill is not None:
            cell.fill = fill


def add_workbook_validations(workbook: Workbook) -> None:
    sheet = workbook["feedthrough_edges"]
    mapping = header_map(sheet)
    sheet.data_validations.dataValidation = []
    validations = {
        "channel_disposition": sorted(DISPOSITIONS),
        "budget_model": sorted(BUDGET_MODELS),
        "budget_required": ["yes", "no"],
        "emit_max": ["yes", "no"],
        "emit_min": ["yes", "no"],
        "datapath_only": ["yes", "no"],
        "tool_surface": ["dc", "sta", "both"],
        "apply": ["yes", "no"],
        "review_status": ["pending", "approved", "rejected"],
    }
    for header, values in validations.items():
        if header not in mapping:
            continue
        formula = '"{0}"'.format(",".join(values))
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        letter = get_column_letter(mapping[header])
        validation.add("{0}2:{0}1048576".format(letter))


def evidence_log_rows(evidence: Sequence[DelayEvidence]) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for item in evidence:
        rows.append(
            {
                "instance": item.inst_name,
                "port": item.port_name,
                "constraint_type": item.constraint_type,
                "clock_name": item.clock_name,
                "min_value": item.min_value,
                "max_value": item.max_value,
                "bare_value": item.bare_value,
                "complex_options": item.complex_options,
                "source_sdc_file": item.source_sdc_file,
                "source_line": item.source_line,
                "source_digest": item.source_digest,
                "parse_status": item.parse_status,
                "original_command": item.original_command,
                "message": item.message,
            }
        )
    return rows


def sync_workbook(
    path: Path,
    seeds: Sequence[Dict[str, object]],
    evidence: Sequence[DelayEvidence],
    report: Report,
) -> None:
    workbook, created = create_or_load_workbook(path)
    ensure_sheet(workbook, "feedthrough_edges", EDGE_HEADERS)
    ensure_sheet(workbook, "extraction_log", LOG_HEADERS)
    sheet = workbook["feedthrough_edges"]
    mapping = header_map(sheet)

    existing: Dict[Tuple[str, str, str, str], int] = {}
    for row_idx in range(2, sheet.max_row + 1):
        values = sheet_row_values(sheet, row_idx, EDGE_HEADERS)
        key = row_key(values)
        if key[3]:
            if key in existing:
                report.error("review workbook duplicate key {0}".format(key))
            else:
                existing[key] = row_idx

    review_changed = created
    seed_keys: Set[Tuple[str, str, str, str]] = set()
    for seed in seeds:
        key = row_key(seed)
        seed_keys.add(key)
        row_idx = existing.get(key)
        if row_idx is None:
            append_row(sheet, EDGE_HEADERS, seed, NEW_FILL)
            existing[key] = sheet.max_row
            report.info("added review row for {0}".format(seed.get("feedthrough_edge_id")))
            review_changed = True
            continue
        material_changed = False
        for header in sorted(MACHINE_HEADERS):
            old_value = clean_cell(sheet.cell(row_idx, mapping[header]).value)
            new_value = clean_cell(seed.get(header))
            if old_value != new_value:
                sheet.cell(row_idx, mapping[header], seed.get(header, ""))
                if header in REVIEW_INVALIDATING_MACHINE_HEADERS:
                    material_changed = True
                report.info(
                    "refreshed {0}.{1}".format(seed.get("feedthrough_edge_id"), header)
                )
        if material_changed:
            review_changed = True
            sheet.cell(row_idx, mapping["apply"], "no").fill = NEW_FILL
            sheet.cell(row_idx, mapping["review_status"], "pending").fill = NEW_FILL
            report.warn(
                "{0}: material machine evidence changed; apply/review_status reset".format(
                    seed.get("feedthrough_edge_id")
                )
            )

        if (
            normalize_key(sheet.cell(row_idx, mapping["apply"]).value) == "yes"
            and normalize_key(sheet.cell(row_idx, mapping["review_status"]).value) == "approved"
        ):
            current_digest = clean_cell(seed.get("machine_digest"))
            approved_cell = sheet.cell(row_idx, mapping["approved_machine_digest"])
            if clean_cell(approved_cell.value) != current_digest:
                approved_cell.value = current_digest
                report.info(
                    "recorded approved machine digest for {0}".format(
                        seed.get("feedthrough_edge_id")
                    )
                )

    current_view_keys = seed_keys
    current_view = next(iter(current_view_keys))[:3] if current_view_keys else None
    for key, row_idx in existing.items():
        if current_view is not None and key[:3] == current_view and key not in seed_keys:
            report.warn("review workbook contains obsolete current-view connection row: {0}".format(key[3]))

    log_sheet = workbook["extraction_log"]
    if log_sheet.max_row > 1:
        log_sheet.delete_rows(2, log_sheet.max_row - 1)
    for values in evidence_log_rows(evidence):
        append_row(log_sheet, LOG_HEADERS, values)

    add_workbook_validations(workbook)
    ensure_sheet(workbook, "feedthrough_edges", EDGE_HEADERS)
    ensure_sheet(workbook, "extraction_log", LOG_HEADERS)
    atomic_save_workbook(workbook, path)
    report.sync_changed = review_changed
    if review_changed:
        report.warn("review workbook synchronized; review new/refreshed 10 rows before generation")


def read_form_rows(path: Path) -> List[FormRow]:
    workbook = load_workbook(str(path), data_only=False)
    if "feedthrough_edges" not in workbook.sheetnames:
        raise RuntimeError("{0}: missing feedthrough_edges sheet".format(path))
    sheet = workbook["feedthrough_edges"]
    return [
        FormRow(row_idx, sheet_row_values(sheet, row_idx, EDGE_HEADERS))
        for row_idx in range(2, sheet.max_row + 1)
        if clean_cell(sheet.cell(row_idx, header_map(sheet).get("connection_id", 1)).value)
    ]


def row_selected(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    values = row.values
    row_stage = clean_cell(values.get("stage")) or "all"
    row_corner = clean_cell(values.get("corner")) or "all"
    return (
        clean_cell(values.get("scenario")) == scenario
        and row_stage == stage
        and row_corner == corner
    )


def is_approved_applied(row: FormRow) -> bool:
    return (
        normalize_key(row.values.get("apply")) == "yes"
        and normalize_key(row.values.get("review_status")) == "approved"
    )


def terminal_disposition(row: FormRow) -> bool:
    disposition = normalize_key(row.values.get("channel_disposition"))
    return is_approved_applied(row) and disposition in {
        "emit_budget", "no_soc_budget_required", "not_applicable"
    }


def row_tool_matches(row: FormRow, tool: str) -> bool:
    surface = normalize_key(row.values.get("tool_surface"))
    if tool == "all":
        return surface in {"dc", "sta", "both"}
    return surface in {tool, "both"}


def commands_for_row(row: FormRow, tool: str) -> List[str]:
    values = row.values
    if not terminal_disposition(row):
        return []
    if normalize_key(values.get("channel_disposition")) != "emit_budget":
        return []
    if not row_tool_matches(row, tool):
        return []
    source = clean_cell(values.get("src_soc_object"))
    destination = clean_cell(values.get("dst_soc_object"))
    datapath = " -datapath_only" if normalize_key(values.get("datapath_only")) == "yes" else ""
    commands: List[str] = []
    if normalize_key(values.get("emit_max")) == "yes":
        commands.append(
            "set_max_delay {0}{1} -from {2} -to {3}".format(
                format_number(values.get("converted_max")), datapath, source, destination
            )
        )
    if normalize_key(values.get("emit_min")) == "yes":
        commands.append(
            "set_min_delay {0}{1} -from {2} -to {3}".format(
                format_number(values.get("converted_min")), datapath, source, destination
            )
        )
    return commands


def exact_bit_endpoint(values: Dict[str, object], side: str):
    port_name = clean_cell(values.get(side + "_port"))
    match = PORT_EXACT_BIT_RE.fullmatch(port_name)
    if not match:
        return None
    explicit_bit = clean_cell(values.get(side + "_bit_index"))
    if explicit_bit and explicit_bit != match.group(2):
        return None
    return (
        clean_cell(values.get(side + "_instance")),
        normalize_key(values.get(side + "_direction")),
        match.group(1),
        int(match.group(2)),
        port_name,
    )


def compaction_cohort_key(row: FormRow):
    src = exact_bit_endpoint(row.values, "src")
    dst = exact_bit_endpoint(row.values, "dst")
    if src is None or dst is None:
        return None
    return (
        src[0], src[1], src[2],
        dst[0], dst[1], dst[2],
        clean_cell(row.values.get("feedthrough_side")),
    )


def compaction_policy_key(row: FormRow):
    values = row.values
    return (
        normalize_key(values.get("channel_disposition")),
        normalize_key(values.get("budget_scope")),
        normalize_key(values.get("budget_model")),
        normalize_key(values.get("budget_required")),
        normalize_key(values.get("emit_max")),
        normalize_key(values.get("emit_min")),
        format_number(values.get("converted_max")),
        format_number(values.get("converted_min")),
        normalize_key(values.get("datapath_only")),
        normalize_key(values.get("tool_surface")),
        normalize_key(values.get("apply")),
        normalize_key(values.get("review_status")),
        normalize_key(values.get("validation_status")),
        normalize_key(values.get("classification_status")),
        normalize_key(values.get("feedthrough_confirmed")),
        clean_cell(values.get("basis")) or clean_cell(values.get("disposition_basis")),
        clean_cell(values.get("sdc_independent_basis")),
        clean_cell(values.get("relationship_override_basis")),
        normalize_key(values.get("min_sign_review")),
        clean_cell(values.get("owner")),
        clean_cell(values.get("reviewer")),
        clean_cell(values.get("review_date")),
    )


def explicit_endpoint_collection(rows: Sequence[FormRow], side: str) -> str:
    values = rows[0].values
    inst_name = clean_cell(values.get(side + "_instance"))
    objects = [
        canonical_soc_object(
            clean_cell(row.values.get(side + "_instance")),
            clean_cell(row.values.get(side + "_port")),
        )
        for row in rows
    ]
    command = "get_ports" if normalize_key(inst_name) == "top" else "get_pins"
    return "[{0} {{{1}}}]".format(command, " ".join(objects))


def compacted_emission_block(
    cohort: Sequence[FormRow], all_edges: Sequence[ConnectionEdge], tool: str
) -> Optional[EmissionBlock]:
    if len(cohort) < 2:
        return None
    if any(not commands_for_row(row, tool) for row in cohort):
        return None
    policies = {compaction_policy_key(row) for row in cohort}
    if len(policies) != 1:
        return None

    parsed = []
    for row in cohort:
        src = exact_bit_endpoint(row.values, "src")
        dst = exact_bit_endpoint(row.values, "dst")
        if src is None or dst is None:
            return None
        parsed.append((src, dst, row))
    parsed.sort(key=lambda item: item[0][3])
    src_bits = [item[0][3] for item in parsed]
    dst_bits = [item[1][3] for item in parsed]
    if len(set(src_bits)) != len(src_bits) or len(set(dst_bits)) != len(dst_bits):
        return None
    if src_bits != list(range(src_bits[0], src_bits[0] + len(src_bits))):
        return None
    if dst_bits != list(range(dst_bits[0], dst_bits[0] + len(dst_bits))):
        return None
    offsets = {dst_bit - src_bit for src_bit, dst_bit in zip(src_bits, dst_bits)}
    if len(offsets) != 1:
        return None

    ordered_rows = [item[2] for item in parsed]
    expected_ids = {clean_cell(row.values.get("connection_id")) for row in ordered_rows}
    edge_by_id = {edge.connection_id: edge for edge in all_edges}
    outgoing: Dict[Tuple[str, str], Set[str]] = defaultdict(set)
    incoming: Dict[Tuple[str, str], Set[str]] = defaultdict(set)
    for edge in all_edges:
        outgoing[(edge.src_instance, edge.src_port)].add(edge.connection_id)
        incoming[(edge.dst_instance, edge.dst_port)].add(edge.connection_id)

    expected_pairs = set()
    source_endpoints = set()
    destination_endpoints = set()
    for row in ordered_rows:
        values = row.values
        connection_id = clean_cell(values.get("connection_id"))
        edge = edge_by_id.get(connection_id)
        if edge is None:
            return None
        src_endpoint = (
            clean_cell(values.get("src_instance")), clean_cell(values.get("src_port"))
        )
        dst_endpoint = (
            clean_cell(values.get("dst_instance")), clean_cell(values.get("dst_port"))
        )
        if (edge.src_instance, edge.src_port) != src_endpoint:
            return None
        if (edge.dst_instance, edge.dst_port) != dst_endpoint:
            return None
        if outgoing[src_endpoint] != {connection_id} or incoming[dst_endpoint] != {connection_id}:
            return None
        source_endpoints.add(src_endpoint)
        destination_endpoints.add(dst_endpoint)
        expected_pairs.add((src_endpoint, dst_endpoint, connection_id))

    actual_pairs = {
        ((edge.src_instance, edge.src_port), (edge.dst_instance, edge.dst_port), edge.connection_id)
        for edge in all_edges
        if (edge.src_instance, edge.src_port) in source_endpoints
        and (edge.dst_instance, edge.dst_port) in destination_endpoints
    }
    if actual_pairs != expected_pairs or {item[2] for item in actual_pairs} != expected_ids:
        return None

    values = ordered_rows[0].values
    source = explicit_endpoint_collection(ordered_rows, "src")
    destination = explicit_endpoint_collection(ordered_rows, "dst")
    datapath = " -datapath_only" if normalize_key(values.get("datapath_only")) == "yes" else ""
    commands: List[str] = []
    if normalize_key(values.get("emit_max")) == "yes":
        commands.append(
            "set_max_delay {0}{1} -from {2} -to {3}".format(
                format_number(values.get("converted_max")), datapath, source, destination
            )
        )
    if normalize_key(values.get("emit_min")) == "yes":
        commands.append(
            "set_min_delay {0}{1} -from {2} -to {3}".format(
                format_number(values.get("converted_min")), datapath, source, destination
            )
        )
    return EmissionBlock(rows=ordered_rows, commands=commands, merged=True)


def validate_assembled_view(
    rows: Sequence[FormRow],
    seeds: Sequence[Dict[str, object]],
    scenario: str,
    stage: str,
    corner: str,
    tool: str,
    report: Report,
) -> Set[int]:
    suppressed_scenario_rows: Set[int] = set()
    if scenario == "common":
        return suppressed_scenario_rows

    seed_by_connection = {
        clean_cell(seed.get("connection_id")): seed for seed in seeds
    }
    common_rows = {
        clean_cell(row.values.get("connection_id")): row
        for row in rows
        if row_selected(row, "common", stage, corner)
    }
    scenario_rows = {
        clean_cell(row.values.get("connection_id")): row
        for row in rows
        if row_selected(row, scenario, stage, corner)
    }
    direct_headers = (
        "feedthrough_edge_id", "connection_id", "src_instance", "src_direction",
        "src_port", "src_soc_object", "dst_instance", "dst_direction", "dst_port",
        "dst_soc_object", "feedthrough_side", "budget_scope",
    )
    for connection_id, common_row in sorted(common_rows.items()):
        if not is_approved_applied(common_row):
            continue
        approved_digest = clean_cell(common_row.values.get("approved_machine_digest"))
        current_digest = clean_cell(common_row.values.get("machine_digest"))
        if not approved_digest or approved_digest != current_digest:
            report.error(
                "feedthrough_edges row {0}: active common approval is stale in assembled scenario {1}".format(
                    common_row.row_idx, scenario
                )
            )
        seed = seed_by_connection.get(connection_id)
        if seed is None:
            report.error(
                "feedthrough_edges row {0}: active common connection {1} is absent from scenario {2}".format(
                    common_row.row_idx, connection_id, scenario
                )
            )
            continue
        for header in direct_headers:
            if clean_cell(common_row.values.get(header)) != clean_cell(seed.get(header)):
                report.error(
                    "feedthrough_edges row {0}: active common {1} conflicts with scenario {2} direct edge".format(
                        common_row.row_idx, header, scenario
                    )
                )

        scenario_row = scenario_rows.get(connection_id)
        if scenario_row is None or not is_approved_applied(scenario_row):
            continue
        common_disposition = normalize_key(common_row.values.get("channel_disposition"))
        scenario_disposition = normalize_key(scenario_row.values.get("channel_disposition"))
        common_commands = commands_for_row(common_row, tool)
        scenario_commands = commands_for_row(scenario_row, tool)
        if (
            common_disposition == scenario_disposition
            and common_commands == scenario_commands
        ):
            if scenario_commands:
                suppressed_scenario_rows.add(scenario_row.row_idx)
                report.info(
                    "connection {0}: identical common/scenario command omitted from {1} overlay".format(
                        connection_id, scenario
                    )
                )
            continue
        report.error(
            "connection {0}: approved common/scenario terminal effects conflict in assembled {1} view".format(
                connection_id, scenario
            )
        )
    return suppressed_scenario_rows


def build_emission_blocks(
    rows: Sequence[FormRow],
    scenario: str,
    stage: str,
    corner: str,
    tool: str,
    all_edges: Sequence[ConnectionEdge],
    suppressed_row_indices: Optional[Set[int]] = None,
) -> List[EmissionBlock]:
    suppressed = suppressed_row_indices or set()
    view_rows = [
        row for row in form_rows_for_view(rows, scenario, stage, corner)
        if row.row_idx not in suppressed
    ]
    cohorts: Dict[Tuple[str, ...], List[FormRow]] = defaultdict(list)
    standalone: List[FormRow] = []
    for row in view_rows:
        key = compaction_cohort_key(row)
        if key is None:
            standalone.append(row)
        else:
            cohorts[key].append(row)

    blocks: List[EmissionBlock] = []
    for cohort in cohorts.values():
        merged = compacted_emission_block(cohort, all_edges, tool)
        if merged is not None:
            blocks.append(merged)
            continue
        for row in cohort:
            commands = commands_for_row(row, tool)
            if commands:
                blocks.append(EmissionBlock(rows=[row], commands=commands))
    for row in standalone:
        commands = commands_for_row(row, tool)
        if commands:
            blocks.append(EmissionBlock(rows=[row], commands=commands))
    return sorted(blocks, key=lambda block: min(row.row_idx for row in block.rows))


def validate_rows(
    rows: Sequence[FormRow],
    seeds: Sequence[Dict[str, object]],
    scenario: str,
    stage: str,
    corner: str,
    report: Report,
) -> None:
    seed_by_key = {row_key(seed): seed for seed in seeds}
    seen_keys: Set[Tuple[str, str, str, str]] = set()
    for row in rows:
        if not row_selected(row, scenario, stage, corner):
            continue
        values = row.values
        key = row_key(values)
        if key in seen_keys:
            report.error("feedthrough_edges row {0}: duplicate view/connection key {1}".format(row.row_idx, key))
            continue
        seen_keys.add(key)
        seed = seed_by_key.get(key)
        if seed is None:
            report.error("feedthrough_edges row {0}: connection is not owned by current 10 view".format(row.row_idx))
            continue
        for header in (
            "feedthrough_edge_id", "connection_id", "src_instance", "src_direction", "src_port",
            "src_soc_object", "dst_instance", "dst_direction", "dst_port", "dst_soc_object",
            "feedthrough_side", "budget_scope", "machine_digest",
        ):
            if clean_cell(values.get(header)) != clean_cell(seed.get(header)):
                report.error(
                    "feedthrough_edges row {0}: machine field {1} does not match current direct edge".format(
                        row.row_idx, header
                    )
                )

        disposition = normalize_key(values.get("channel_disposition"))
        budget_model = normalize_key(values.get("budget_model"))
        budget_required = normalize_key(values.get("budget_required"))
        emit_max = normalize_key(values.get("emit_max"))
        emit_min = normalize_key(values.get("emit_min"))
        review_status = normalize_key(values.get("review_status"))
        apply_value = normalize_key(values.get("apply"))
        tool_surface = normalize_key(values.get("tool_surface"))
        datapath_only = normalize_key(values.get("datapath_only"))

        if disposition not in DISPOSITIONS:
            report.error("feedthrough_edges row {0}: invalid channel_disposition {1}".format(row.row_idx, disposition))
        if budget_model not in BUDGET_MODELS:
            report.error("feedthrough_edges row {0}: invalid budget_model {1}".format(row.row_idx, budget_model))
        if budget_required not in YES_NO:
            report.error("feedthrough_edges row {0}: budget_required must be yes/no".format(row.row_idx))
        if emit_max not in YES_NO or emit_min not in YES_NO:
            report.error("feedthrough_edges row {0}: emit_max/emit_min must be yes/no".format(row.row_idx))
        if review_status not in REVIEW_STATUSES:
            report.error("feedthrough_edges row {0}: invalid review_status".format(row.row_idx))
        if apply_value not in YES_NO:
            report.error("feedthrough_edges row {0}: apply must be yes/no".format(row.row_idx))
        if tool_surface not in TOOL_SURFACES:
            report.error("feedthrough_edges row {0}: invalid tool_surface".format(row.row_idx))
        if datapath_only not in YES_NO:
            report.error("feedthrough_edges row {0}: datapath_only must be yes/no".format(row.row_idx))
        if disposition != "emit_budget" and (emit_max == "yes" or emit_min == "yes"):
            report.error(
                "feedthrough_edges row {0}: only emit_budget may set emit_max/emit_min=yes".format(row.row_idx)
            )

        current_machine_digest = clean_cell(values.get("machine_digest"))
        approved_machine_digest = clean_cell(values.get("approved_machine_digest"))
        if approved_machine_digest and approved_machine_digest != current_machine_digest:
            report.error(
                "feedthrough_edges row {0}: machine evidence changed after approval; re-review required".format(
                    row.row_idx
                )
            )

        if not is_approved_applied(row):
            continue
        if not approved_machine_digest:
            report.error(
                "feedthrough_edges row {0}: approved terminal row requires approved_machine_digest".format(
                    row.row_idx
                )
            )
        if normalize_key(values.get("validation_status")) != "matched":
            report.error("feedthrough_edges row {0}: terminal/approved row requires validation_status=matched".format(row.row_idx))
        missing_evidence = "missing" in {
            normalize_key(values.get("src_sdc_status")),
            normalize_key(values.get("dst_sdc_status")),
        }
        if missing_evidence and disposition in {
            "emit_budget", "no_soc_budget_required", "not_applicable"
        } and not clean_cell(values.get("sdc_independent_basis")):
            report.error(
                "feedthrough_edges row {0}: missing harden SDC requires approved sdc_independent_basis".format(
                    row.row_idx
                )
            )

        if disposition == "emit_budget":
            if budget_required != "yes":
                report.error("feedthrough_edges row {0}: emit_budget requires budget_required=yes".format(row.row_idx))
            if budget_model not in {"interconnect_budget", "manual_budget"}:
                report.error("feedthrough_edges row {0}: emit_budget requires reviewed interconnect/manual model".format(row.row_idx))
            if emit_max != "yes" and emit_min != "yes":
                report.error("feedthrough_edges row {0}: emit_budget must emit max and/or min".format(row.row_idx))
            if datapath_only != "yes":
                report.error("feedthrough_edges row {0}: emit_budget requires datapath_only=yes".format(row.row_idx))
            if tool_surface not in {"dc", "sta", "both"}:
                report.error("feedthrough_edges row {0}: emit_budget requires a non-empty tool_surface".format(row.row_idx))
            if not clean_cell(values.get("src_soc_object")) or not clean_cell(values.get("dst_soc_object")):
                report.error("feedthrough_edges row {0}: emit_budget requires two resolved direct endpoints".format(row.row_idx))
            if emit_max == "yes" and parse_number(values.get("converted_max")) is None:
                report.error("feedthrough_edges row {0}: emit_max requires finite converted_max".format(row.row_idx))
            if emit_min == "yes":
                if parse_number(values.get("converted_min")) is None:
                    report.error("feedthrough_edges row {0}: emit_min requires finite converted_min".format(row.row_idx))
                if normalize_key(values.get("min_sign_review")) not in {"approved", "reviewed", "yes"}:
                    report.error("feedthrough_edges row {0}: emit_min requires min_sign_review".format(row.row_idx))
            if not clean_cell(values.get("disposition_basis")):
                report.error("feedthrough_edges row {0}: emit_budget requires disposition_basis".format(row.row_idx))
        elif disposition == "no_soc_budget_required":
            if budget_required != "no":
                report.error("feedthrough_edges row {0}: no-budget disposition requires budget_required=no".format(row.row_idx))
            if emit_max != "no" or emit_min != "no":
                report.error(
                    "feedthrough_edges row {0}: no-budget disposition requires emit_max=no and emit_min=no".format(
                        row.row_idx
                    )
                )
            missing_fields = [
                name for name in ("owner", "reviewer", "review_date", "disposition_basis")
                if not clean_cell(values.get(name))
            ]
            if missing_fields:
                report.error(
                    "feedthrough_edges row {0}: no-budget policy approval missing {1}".format(
                        row.row_idx, ",".join(missing_fields)
                    )
                )
            elif not has_versioned_policy_basis(values.get("disposition_basis")):
                report.error(
                    "feedthrough_edges row {0}: no-budget disposition requires a versioned policy/basis".format(
                        row.row_idx
                    )
                )
        elif disposition == "not_applicable":
            missing_fields = [
                name for name in ("owner", "reviewer", "disposition_basis")
                if not clean_cell(values.get(name))
            ]
            if missing_fields:
                report.error(
                    "feedthrough_edges row {0}: not_applicable missing {1}".format(
                        row.row_idx, ",".join(missing_fields)
                    )
                )
        elif disposition == "route_to_30":
            missing_fields = [
                name for name in ("owner", "reviewer", "review_date", "disposition_basis")
                if not clean_cell(values.get(name))
            ]
            if missing_fields:
                report.error(
                    "feedthrough_edges row {0}: route_to_30 missing reviewed exception evidence: {1}".format(
                        row.row_idx, ",".join(missing_fields)
                    )
                )
        elif disposition == "pending":
            report.error(
                "feedthrough_edges row {0}: pending disposition cannot be approved/applied".format(
                    row.row_idx
                )
            )

    missing_rows = sorted(set(seed_by_key) - seen_keys)
    for key in missing_rows:
        report.error("review workbook missing current 10 edge row {0}".format(key[3]))


def form_rows_for_view(
    rows: Sequence[FormRow], scenario: str, stage: str, corner: str
) -> List[FormRow]:
    return sorted(
        (row for row in rows if row_selected(row, scenario, stage, corner)),
        key=lambda row: clean_cell(row.values.get("connection_id")),
    )


def write_inventory_upsert(
    path: Path,
    rows: Sequence[FormRow],
    scenario: str,
    stage: str,
    corner: str,
) -> None:
    preserved: List[Dict[str, object]] = []
    if path.is_file():
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            for row in csv.DictReader(file_obj):
                if not (
                    clean_cell(row.get("scenario")) == scenario
                    and clean_cell(row.get("stage")) == stage
                    and clean_cell(row.get("corner")) == corner
                ):
                    preserved.append(dict(row))
    current = [dict(row.values) for row in form_rows_for_view(rows, scenario, stage, corner)]
    combined = preserved + current
    combined.sort(
        key=lambda row: (
            clean_cell(row.get("scenario")), clean_cell(row.get("stage")),
            clean_cell(row.get("corner")), clean_cell(row.get("connection_id")),
        )
    )
    atomic_write_csv(path, EDGE_HEADERS, combined)


def validate_pending_directory(
    pending_dir: Path,
    manifest: Dict[str, HardenSdc],
    report: Report,
) -> None:
    if not pending_dir.exists():
        report.error("pending directory not found: {0}".format(pending_dir))
        return
    if not pending_dir.is_dir():
        report.error("pending path is not a directory: {0}".format(pending_dir))
        return

    for inst_name in sorted(manifest):
        pending_file = pending_dir / (inst_name + ".ports")
        if not pending_file.is_file():
            report.error("pending file missing for manifest harden {0}: {1}".format(inst_name, pending_file))
            continue
        try:
            lines = pending_file.read_text(encoding="utf-8").splitlines()
        except (OSError, UnicodeError) as exc:
            report.error("pending file is not readable: {0}: {1}".format(pending_file, exc))
            continue
        seen: Set[Tuple[str, str]] = set()
        for line_idx, line in enumerate(lines, start=1):
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            parts = stripped.split()
            if len(parts) != 2 or parts[0] not in {"input", "output", "inout"}:
                report.error(
                    "{0} row {1}: malformed pending line {2!r}".format(
                        pending_file, line_idx, line
                    )
                )
                continue
            direction, port_name = parts
            if not is_canonical_port(port_name):
                report.error(
                    "{0} row {1}: pending port is not an exact canonical scalar/bit key: {2}".format(
                        pending_file, line_idx, port_name
                    )
                )
                continue
            key = (direction, port_name)
            if key in seen:
                report.error(
                    "{0} row {1}: duplicate pending line {2} {3}".format(
                        pending_file, line_idx, direction, port_name
                    )
                )
                continue
            seen.add(key)


def pending_line_key(line: str) -> Optional[Tuple[str, str]]:
    stripped = line.strip()
    if not stripped or stripped.startswith("#"):
        return None
    parts = stripped.split()
    if (
        len(parts) != 2
        or parts[0] not in {"input", "output", "inout"}
        or not is_canonical_port(parts[1])
    ):
        return None
    return parts[0], parts[1]


def removed_line_key(line: str) -> Optional[PortKey]:
    parts = line.strip().split()
    if len(parts) < 3 or parts[1] not in {"input", "output", "inout"}:
        return None
    return PortKey(parts[0], parts[1], parts[2])


def read_removed_key_sources(
    paths: Sequence[Path], report: Report
) -> Dict[PortKey, Path]:
    sources: Dict[PortKey, Path] = {}
    for path in paths:
        if not path.exists():
            continue
        if not path.is_file():
            report.error("removed log path is not a file: {0}".format(path))
            continue
        try:
            lines = path.read_text(encoding="utf-8").splitlines()
        except (OSError, UnicodeError) as exc:
            report.error("removed log is not readable: {0}: {1}".format(path, exc))
            continue
        seen_in_file: Set[PortKey] = set()
        for line_idx, line in enumerate(lines, start=1):
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            key = removed_line_key(line)
            if key is None or not is_canonical_port(key.port_name):
                report.error(
                    "{0} row {1}: malformed removed-log line".format(path, line_idx)
                )
                continue
            if key in seen_in_file:
                report.error(
                    "{0} row {1}: duplicate removed key {2}/{3}".format(
                        path, line_idx, key.inst_name, key.port_name
                    )
                )
                continue
            seen_in_file.add(key)
            previous = sources.get(key)
            if previous is not None and previous != path:
                report.error(
                    "port-level removal has multiple owners for {0} {1} {2}: {3}, {4}".format(
                        key.inst_name, key.direction, key.port_name, previous, path
                    )
                )
                continue
            sources[key] = path
    return sources


def edge_port_keys(edge: ConnectionEdge) -> List[PortKey]:
    result: List[PortKey] = []
    for inst_name, direction, port_name in (
        (edge.src_instance, edge.src_direction, edge.src_port),
        (edge.dst_instance, edge.dst_direction, edge.dst_port),
    ):
        if normalize_key(inst_name) in PSEUDO_INSTANCES:
            continue
        if direction in {"input", "output", "inout"} and port_name:
            result.append(PortKey(inst_name, direction, port_name))
    return result


def terminal_for_pending(row: FormRow, tool: str) -> bool:
    if not terminal_disposition(row):
        return False
    if normalize_key(row.values.get("channel_disposition")) == "emit_budget":
        return bool(commands_for_row(row, tool))
    return True


def removed_log_line(
    key: PortKey,
    rows: Sequence[FormRow],
    scenario: str,
    stage: str,
    corner: str,
) -> str:
    edge_ids = [clean_cell(row.values.get("feedthrough_edge_id")) for row in rows]
    connection_ids = [clean_cell(row.values.get("connection_id")) for row in rows]
    dispositions = [normalize_key(row.values.get("channel_disposition")) for row in rows]
    return " ".join(
        [
            key.inst_name,
            key.direction,
            key.port_name,
            "covered_by=10_feedthrough",
            "reason=feedthrough_direct_edge_terminal",
            "feedthrough_edge_id=" + ",".join(edge_ids),
            "connection_id=" + ",".join(connection_ids),
            "channel_disposition=" + ",".join(dispositions),
            "scenario=" + scenario,
            "stage=" + stage,
            "corner=" + corner,
        ]
    )


def prepare_pending_plan(
    pending_dir: Path,
    removed_log_path: Path,
    previous_removed_paths: Sequence[Path],
    all_edges: Sequence[ConnectionEdge],
    owned: Sequence[FeedthroughEdge],
    rows: Sequence[FormRow],
    scenario: str,
    stage: str,
    corner: str,
    tool: str,
    report: Report,
) -> PendingPlan:
    plan = PendingPlan()
    if not pending_dir.exists():
        report.error("pending directory not found: {0}".format(pending_dir))
        return plan
    if not pending_dir.is_dir():
        report.error("pending path is not a directory: {0}".format(pending_dir))
        return plan

    owned_by_connection = {item.edge.connection_id: item for item in owned}
    current_rows = {
        clean_cell(row.values.get("connection_id")): row
        for row in rows
        if row_selected(row, scenario, stage, corner)
    }
    terminal_connections = {
        connection_id
        for connection_id, row in current_rows.items()
        if terminal_for_pending(row, tool)
    }

    incident: Dict[PortKey, Set[str]] = defaultdict(set)
    for edge in all_edges:
        for key in edge_port_keys(edge):
            incident[key].add(edge.connection_id)

    candidate_keys: Set[PortKey] = set()
    rows_by_key: Dict[PortKey, List[FormRow]] = defaultdict(list)
    for connection_id in terminal_connections:
        item = owned_by_connection.get(connection_id)
        if item is None:
            continue
        row = current_rows[connection_id]
        for key in edge_port_keys(item.edge):
            candidate_keys.add(key)
            rows_by_key[key].append(row)

    removable: Set[PortKey] = set()
    for key in sorted(candidate_keys, key=lambda item: (item.inst_name, item.direction, item.port_name)):
        connection_ids = incident.get(key, set())
        if not connection_ids.issubset(set(owned_by_connection)):
            report.warn(
                "{0}/{1}: port also participates in non-10 connection(s); pending key retained".format(
                    key.inst_name, key.port_name
                )
            )
            continue
        if not connection_ids.issubset(terminal_connections):
            report.info(
                "{0}/{1}: not all 10-owned direct edges are terminal; pending key retained".format(
                    key.inst_name, key.port_name
                )
            )
            continue
        removable.add(key)

    previous_sources = read_removed_key_sources(previous_removed_paths, report)
    current_sources = read_removed_key_sources([removed_log_path], report)
    for key in sorted(
        set(previous_sources).intersection(current_sources),
        key=lambda item: (item.inst_name, item.direction, item.port_name),
    ):
        report.error(
            "10 removal duplicates previous owner for {0} {1} {2}: {3}".format(
                key.inst_name, key.direction, key.port_name, previous_sources[key]
            )
        )
    try:
        existing_log_lines = (
            removed_log_path.read_text(encoding="utf-8").splitlines()
            if removed_log_path.is_file() else []
        )
    except (OSError, UnicodeError) as exc:
        report.error("removed log is not readable: {0}: {1}".format(removed_log_path, exc))
        existing_log_lines = []
    previous_removed = set(current_sources)
    new_log_lines: List[str] = []
    by_instance: Dict[str, List[PortKey]] = defaultdict(list)
    for key in removable:
        if key in previous_sources:
            report.info(
                "{0}/{1}: retained previous port-level removal owner {2}".format(
                    key.inst_name, key.port_name, previous_sources[key]
                )
            )
            continue
        by_instance[key.inst_name].append(key)

    for inst_name, keys in sorted(by_instance.items()):
        pending_file = pending_dir / (inst_name + ".ports")
        if not pending_file.is_file():
            for key in keys:
                if key not in previous_removed:
                    report.error("{0}: missing pending file for {1}/{2}".format(pending_file, inst_name, key.port_name))
            continue
        lines = pending_file.read_text(encoding="utf-8").splitlines()
        index: Dict[Tuple[str, str], int] = {}
        duplicates: Set[Tuple[str, str]] = set()
        for line_idx, line in enumerate(lines):
            parsed = pending_line_key(line)
            if parsed is None:
                continue
            if parsed in index:
                duplicates.add(parsed)
            else:
                index[parsed] = line_idx
        for direction, port_name in sorted(duplicates):
            report.error("{0}: duplicate pending line {1} {2}".format(pending_file, direction, port_name))
        remove_indices: Set[int] = set()
        for key in keys:
            pending_key = (key.direction, key.port_name)
            if pending_key not in index:
                if key in previous_removed:
                    continue
                report.error(
                    "{0}: 10 wants to remove {1} {2}, but no previous 10 removal exists".format(
                        pending_file, key.direction, key.port_name
                    )
                )
                continue
            remove_indices.add(index[pending_key])
            if key not in previous_removed:
                new_log_lines.append(removed_log_line(key, rows_by_key[key], scenario, stage, corner))
            plan.removed_count += 1
        if remove_indices:
            kept = [line for idx, line in enumerate(lines) if idx not in remove_indices]
            plan.pending_updates[pending_file] = kept

    plan.removed_log_lines = [line for line in existing_log_lines if line.strip()] + new_log_lines
    return plan


def apply_pending_plan(plan: PendingPlan, removed_log_path: Path) -> None:
    for path, lines in plan.pending_updates.items():
        text = "\n".join(lines).rstrip()
        atomic_write_text(path, text + ("\n" if text else ""))
    if plan.removed_log_lines:
        atomic_write_text(removed_log_path, "\n".join(plan.removed_log_lines).rstrip() + "\n")


def output_sdc_path(base: Path, scenario: str, stage: str, corner: str) -> Path:
    suffix = "" if stage == "all" and corner == "all" else "_{0}_{1}".format(stage, safe_token(corner))
    if scenario == "common":
        return base / "common" / ("10_feedthrough{0}.sdc".format(suffix))
    return base / "scenarios" / ("{0}_feedthrough{1}.sdc".format(scenario, suffix))


def generate_sdc(
    emission_blocks: Sequence[EmissionBlock],
    scenario: str,
    stage: str,
    corner: str,
    tool: str,
    completeness: RunCompleteness,
    port_accounting: str,
    connection_path: Path,
    manifest_path: Path,
) -> List[str]:
    lines = [
        "################################################################################",
        "# Auto-generated SoC feedthrough direct-edge constraints",
        "# Author: {0}".format(author_name()),
        "# Stage: 10_feedthrough",
        "# Script: 10_extract_feedthrough.py",
        "# Scenario: {0}".format(scenario),
        "# View: stage={0}, corner={1}, tool={2}".format(stage, corner, tool),
        "# Run completeness: {0}".format(completeness.status),
        "# Port accounting: {0}".format(port_accounting),
        "# Connection inventory: {0}".format(connection_path.resolve()),
        "# Harden SDC manifest: {0}".format(manifest_path.resolve()),
        "# Harden SDC available: {0}".format(completeness.available_count),
        "# Harden SDC missing: {0}".format(completeness.missing_count),
        "# Harden SDC not_required: {0}".format(completeness.not_required_count),
        "# Missing instances: {0}".format(",".join(completeness.missing_instances) or "<none>"),
        "# Direct edges only; no harden-internal or end-to-end stitching is emitted.",
        "################################################################################",
        "",
    ]
    emitted = 0
    for block in emission_blocks:
        values = block.rows[0].values
        if block.merged:
            append_chunked_comment(
                lines,
                "Merged feedthrough_edge_ids",
                [clean_cell(row.values.get("feedthrough_edge_id")) for row in block.rows],
            )
            append_chunked_comment(
                lines,
                "connection_ids",
                [clean_cell(row.values.get("connection_id")) for row in block.rows],
            )
            lines.append("# side=" + comment_text(values.get("feedthrough_side")))
        else:
            lines.append(
                "# {0} connection_id={1} side={2}".format(
                    comment_text(values.get("feedthrough_edge_id")),
                    comment_text(values.get("connection_id")),
                    comment_text(values.get("feedthrough_side")),
                )
            )
        if clean_cell(values.get("disposition_basis")):
            lines.append("# Basis: " + comment_text(values.get("disposition_basis")))
        lines.extend(block.commands)
        lines.append("")
        emitted += len(block.commands)
    if emitted == 0:
        lines.append("# No feedthrough direct-edge timing commands emitted for this view.")
    return lines


def build_coverage_lines(
    rows: Sequence[FormRow],
    owned: Sequence[FeedthroughEdge],
    scenario: str,
    stage: str,
    corner: str,
    tool: str,
    emission_blocks: Sequence[EmissionBlock],
    generated: bool,
) -> List[str]:
    view_rows = form_rows_for_view(rows, scenario, stage, corner)
    disposition_counts: Dict[str, int] = defaultdict(int)
    logical_commands = 0
    terminal_rows = 0
    for row in view_rows:
        disposition_counts[normalize_key(row.values.get("channel_disposition")) or "<blank>"] += 1
        logical_commands += len(commands_for_row(row, tool))
        if terminal_for_pending(row, tool):
            terminal_rows += 1
    lines = [
        "",
        "Coverage:",
        "  10-owned direct bit edges: {0}".format(len(owned)),
        "  current-view workbook rows: {0}".format(len(view_rows)),
        "  terminal rows: {0}".format(terminal_rows),
        "  logical bit commands: {0}".format(logical_commands),
        "  planned Tcl commands: {0}".format(sum(len(block.commands) for block in emission_blocks)),
        "  actual emitted Tcl commands: {0}".format(
            sum(len(block.commands) for block in emission_blocks) if generated else 0
        ),
        "  planned compacted groups: {0}".format(sum(1 for block in emission_blocks if block.merged)),
        "  planned compacted bit rows: {0}".format(
            sum(len(block.rows) for block in emission_blocks if block.merged)
        ),
        "  formal SDC written: {0}".format("yes" if generated else "no"),
        "  dispositions:",
    ]
    for disposition in sorted(disposition_counts):
        lines.append("    {0}: {1}".format(disposition, disposition_counts[disposition]))
    merged_blocks = [block for block in emission_blocks if block.merged]
    if merged_blocks:
        lines.append("")
        lines.append("  Compaction groups:")
        for group_idx, block in enumerate(merged_blocks, start=1):
            command_types = [command.split(None, 1)[0] for command in block.commands]
            lines.append(
                "    group {0}: commands={1} bits={2}".format(
                    group_idx, ",".join(command_types), len(block.rows)
                )
            )
            fte_ids = [comment_text(row.values.get("feedthrough_edge_id")) for row in block.rows]
            connection_ids = [comment_text(row.values.get("connection_id")) for row in block.rows]
            for offset in range(0, len(fte_ids), 8):
                label = "FTEs" if offset == 0 else "FTEs (cont.)"
                lines.append("      {0}: {1}".format(label, ",".join(fte_ids[offset:offset + 8])))
            for offset in range(0, len(connection_ids), 8):
                label = "connections" if offset == 0 else "connections (cont.)"
                lines.append(
                    "      {0}: {1}".format(label, ",".join(connection_ids[offset:offset + 8]))
                )
    lines.append("")
    lines.append("  Per-edge status:")
    for row in view_rows:
        values = row.values
        lines.append(
            "    {0}: connection={1} side={2} disposition={3} review={4} validation={5} evidence={6}".format(
                clean_cell(values.get("feedthrough_edge_id")),
                clean_cell(values.get("connection_id")),
                clean_cell(values.get("feedthrough_side")),
                normalize_key(values.get("channel_disposition")) or "-",
                normalize_key(values.get("review_status")) or "-",
                normalize_key(values.get("validation_status")) or "-",
                normalize_key(values.get("evidence_status")) or "-",
            )
        )
    return lines


def write_report(
    path: Path,
    report: Report,
    scenario: str,
    stage: str,
    corner: str,
    tool: str,
    completeness: RunCompleteness,
    form_path: Path,
    inventory_path: Path,
    output_path: Path,
    connection_path: Path,
    manifest_path: Path,
    port_accounting: str,
    coverage: Sequence[str],
) -> None:
    lines = [
        "10_feedthrough direct-edge report",
        "=================================",
        "",
        "Author: {0}".format(author_name()),
        "Stage: 10_feedthrough",
        "Script: 10_extract_feedthrough.py",
        "Scenario: {0}".format(scenario),
        "View: stage={0}, corner={1}, tool={2}".format(stage, corner, tool),
        "Form: {0}".format(form_path),
        "Inventory: {0}".format(inventory_path),
        "Output: {0}".format(output_path),
        "Connection inventory: {0}".format(connection_path.resolve()),
        "Harden SDC manifest: {0}".format(manifest_path.resolve()),
        "Run completeness: {0}".format(completeness.status),
        "Port accounting: {0}".format(port_accounting),
        "Available harden SDC: {0}".format(completeness.available_count),
        "Missing harden SDC: {0}".format(completeness.missing_count),
        "Not-required harden SDC: {0}".format(completeness.not_required_count),
        "Missing instances: {0}".format(",".join(completeness.missing_instances) or "<none>"),
        "Warnings: {0}".format(report.warning_count),
        "Errors: {0}".format(report.error_count),
        "Sync changed: {0}".format("yes" if report.sync_changed else "no"),
        "",
        "Messages:",
    ]
    lines.extend(report.lines or ["INFO: no messages"])
    lines.extend(coverage)
    atomic_write_text(path, "\n".join(lines).rstrip() + "\n")


TARGET_SIGNAL_RE = re.compile(
    r"^(?P<base>[A-Za-z_][A-Za-z0-9_$]*)(?:\[(?P<left>-?\d+)(?::(?P<right>-?\d+))?\])?$"
)


def load_accounting_runtime():
    """Load the shared workbook/accounting implementation hosted by stage 00."""
    here = Path(__file__).resolve()
    candidates = [
        here.parent.parent / "00_harden_port_inventory" / "00_harden_port_inventory.py",
        here.with_name("00_harden_port_inventory.py"),
    ]
    runtime_path = next((path for path in candidates if path.is_file()), None)
    if runtime_path is None:
        raise RuntimeError("shared accounting runtime 00_harden_port_inventory.py was not found")
    spec = importlib.util.spec_from_file_location(
        "soc_sdc_accounting_runtime_for_10", str(runtime_path)
    )
    if spec is None or spec.loader is None:
        raise RuntimeError("failed to load shared accounting runtime from {0}".format(runtime_path))
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    module.STAGE_NAME = "10_feedthrough"
    return module


def target_json_digest(value) -> str:
    payload = json.dumps(
        value, ensure_ascii=False, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def target_ordered_bits(shape) -> List[int]:
    if shape.scalar:
        return [0]
    step = 1 if shape.right >= shape.left else -1
    return list(range(shape.left, shape.right + step, step))


def target_exact_port(shape, bit: int) -> str:
    return shape.base if shape.scalar else "{0}[{1}]".format(shape.base, bit)


def target_endpoint_key(inst_name: str, direction: str, port_name: str) -> str:
    return "{0}:{1}:{2}".format(inst_name, direction, port_name)


def target_endpoint_collection(inst_name: str, port_name: str) -> str:
    if normalize_key(inst_name) == "top":
        return "[get_ports {{{0}}}]".format(port_name)
    return "[get_pins {{{0}/{1}}}]".format(inst_name, port_name)


def target_parse_signal(value: str, location: str) -> Tuple[str, Optional[List[int]]]:
    text = clean_cell(value)
    match = TARGET_SIGNAL_RE.fullmatch(text)
    if not match:
        raise ValueError("{0}: invalid endpoint signal {1!r}".format(location, text))
    base = match.group("base")
    left = match.group("left")
    right = match.group("right")
    if left is None:
        return base, None
    left_value = int(left)
    right_value = int(right) if right is not None else left_value
    step = 1 if right_value >= left_value else -1
    return base, list(range(left_value, right_value + step, step))


def target_connection_ids(
    src_instance: str,
    src_direction: str,
    src_base: str,
    src_bit: int,
    dst_instance: str,
    dst_direction: str,
    dst_base: str,
    dst_bit: int,
) -> Tuple[str, str, str]:
    canonical_tuple = [
        TARGET_SCHEMA_VERSION,
        clean_cell(src_instance),
        normalize_key(src_direction),
        clean_cell(src_base),
        int(src_bit),
        clean_cell(dst_instance),
        normalize_key(dst_direction),
        clean_cell(dst_base),
        int(dst_bit),
    ]
    digest = target_json_digest(canonical_tuple)
    return "CONN_" + digest, "FTE_" + digest, digest


def target_source_records(port_records) -> Dict[Tuple[str, str], List[Any]]:
    result: Dict[Tuple[str, str], List[Any]] = defaultdict(list)
    for record in port_records:
        if record.direction in {"output", "inout"}:
            result[(record.inst_name, record.shape.base)].append(record)
    for key in result:
        result[key].sort(key=lambda item: (item.workbook, item.sheet, item.row))
    return result


def target_select_source_bits(
    source_records: Sequence[Any],
    explicit_bits: Optional[List[int]],
    destination_count: int,
    location: str,
) -> List[Tuple[Any, int]]:
    if explicit_bits is None:
        if len(source_records) != 1:
            raise ValueError(
                "{0}: source without selector maps to {1} port rows; use an exact bit/range".format(
                    location, len(source_records)
                )
            )
        selected_bits = target_ordered_bits(source_records[0].shape)
    else:
        selected_bits = list(explicit_bits)
    if len(selected_bits) != destination_count:
        raise ValueError(
            "{0}: source width {1} does not match destination width {2}".format(
                location, len(selected_bits), destination_count
            )
        )
    selected: List[Tuple[Any, int]] = []
    for bit in selected_bits:
        matches = [record for record in source_records if record.shape.contains(bit)]
        if len(matches) != 1:
            raise ValueError(
                "{0}: source bit {1} matches {2} declared port rows".format(
                    location, bit, len(matches)
                )
            )
        selected.append((matches[0], bit))
    return selected


def build_target_direct_edges(runtime, port_records, report: Report) -> List[TargetDirectContext]:
    source_index = target_source_records(port_records)
    contexts: List[TargetDirectContext] = []
    seen_destinations: Dict[Tuple[str, str, str], str] = {}

    for record in port_records:
        if record.direction not in {"input", "inout"}:
            continue
        connection = clean_cell(record.connection_value)
        structural = runtime.structural_reason(
            connection, record.shape, record.location(), report
        )
        if structural:
            continue
        if not connection:
            report.error("{0}: unresolved destination has empty From Whom/Inout Connectivity".format(record.location()))
            continue
        if "." not in connection:
            report.error(
                "{0}: direct driver must be <inst>.<port/range>, got {1!r}".format(
                    record.location(), connection
                )
            )
            continue
        source_instance, source_signal = connection.split(".", 1)
        try:
            source_base, explicit_bits = target_parse_signal(
                source_signal, record.location()
            )
        except ValueError as exc:
            report.error(str(exc))
            continue
        destination_bits = target_ordered_bits(record.shape)
        if source_instance == "top":
            source_bits = explicit_bits if explicit_bits is not None else list(destination_bits)
            if len(source_bits) != len(destination_bits):
                report.error(
                    "{0}: top source width {1} does not match destination width {2}".format(
                        record.location(), len(source_bits), len(destination_bits)
                    )
                )
                continue
            selected_sources = [(None, bit) for bit in source_bits]
        else:
            source_rows = source_index.get((source_instance, source_base), [])
            if not source_rows:
                report.error(
                    "{0}: unknown direct source {1}.{2}".format(
                        record.location(), source_instance, source_base
                    )
                )
                continue
            try:
                selected_sources = target_select_source_bits(
                    source_rows, explicit_bits, len(destination_bits), record.location()
                )
            except ValueError as exc:
                report.error(str(exc))
                continue

        for (source_record, source_bit), destination_bit in zip(
            selected_sources, destination_bits
        ):
            src_direction = source_record.direction if source_record is not None else "input"
            src_port = (
                target_exact_port(source_record.shape, source_bit)
                if source_record is not None
                else (source_base if len(selected_sources) == 1 and source_bit == 0 else "{0}[{1}]".format(source_base, source_bit))
            )
            dst_port = target_exact_port(record.shape, destination_bit)
            destination_key = (record.inst_name, record.direction, dst_port)
            if destination_key in seen_destinations:
                report.error(
                    "{0}: destination bit has multiple drivers; previous={1}".format(
                        record.location(), seen_destinations[destination_key]
                    )
                )
                continue
            connection_id, _, _ = target_connection_ids(
                source_instance,
                src_direction,
                source_base,
                source_bit,
                record.inst_name,
                record.direction,
                record.shape.base,
                destination_bit,
            )
            seen_destinations[destination_key] = connection_id
            contexts.append(
                TargetDirectContext(
                    edge=ConnectionEdge(
                        schema_version=TARGET_SCHEMA_VERSION,
                        connection_id=connection_id,
                        scenario_scope="",
                        connection_type="port_workbook_direct",
                        src_instance=source_instance,
                        src_direction=src_direction,
                        src_port=src_port,
                        src_bit_index=str(source_bit),
                        src_endpoint_key=target_endpoint_key(source_instance, src_direction, src_port),
                        src_soc_object=(src_port if source_instance == "top" else "{0}/{1}".format(source_instance, src_port)),
                        dst_instance=record.inst_name,
                        dst_direction=record.direction,
                        dst_port=dst_port,
                        dst_bit_index=str(destination_bit),
                        dst_endpoint_key=target_endpoint_key(record.inst_name, record.direction, dst_port),
                        dst_soc_object="{0}/{1}".format(record.inst_name, dst_port),
                        validation_status="matched",
                        note="",
                        source_row=record.row,
                    ),
                    source_record=record,
                    src_record=source_record,
                    dst_record=record,
                )
            )

    for record in port_records:
        if record.direction != "output":
            continue
        connection = clean_cell(record.connection_value)
        if not connection or runtime.structural_reason(
            connection, record.shape, record.location(), report
        ):
            continue
        top_signal = record.shape.raw if connection.lower() in {"y", "yes", "true"} else connection
        if top_signal.startswith("top."):
            top_signal = top_signal[4:]
        if "." in top_signal:
            report.error("{0}: To Top must name a top port, got {1!r}".format(record.location(), connection))
            continue
        try:
            top_base, explicit_bits = target_parse_signal(top_signal, record.location())
        except ValueError as exc:
            report.error(str(exc))
            continue
        source_bits = target_ordered_bits(record.shape)
        top_bits = explicit_bits if explicit_bits is not None else list(source_bits)
        if len(top_bits) != len(source_bits):
            report.error("{0}: To Top width mismatch".format(record.location()))
            continue
        for source_bit, top_bit in zip(source_bits, top_bits):
            src_port = target_exact_port(record.shape, source_bit)
            dst_port = top_base if len(top_bits) == 1 and top_bit == 0 else "{0}[{1}]".format(top_base, top_bit)
            connection_id, _, _ = target_connection_ids(
                record.inst_name, "output", record.shape.base, source_bit,
                "top", "output", top_base, top_bit,
            )
            contexts.append(
                TargetDirectContext(
                    edge=ConnectionEdge(
                        schema_version=TARGET_SCHEMA_VERSION,
                        connection_id=connection_id,
                        scenario_scope="",
                        connection_type="port_workbook_direct",
                        src_instance=record.inst_name,
                        src_direction="output",
                        src_port=src_port,
                        src_bit_index=str(source_bit),
                        src_endpoint_key=target_endpoint_key(record.inst_name, "output", src_port),
                        src_soc_object="{0}/{1}".format(record.inst_name, src_port),
                        dst_instance="top",
                        dst_direction="output",
                        dst_port=dst_port,
                        dst_bit_index=str(top_bit),
                        dst_endpoint_key=target_endpoint_key("top", "output", dst_port),
                        dst_soc_object=dst_port,
                        validation_status="matched",
                        note="",
                        source_row=record.row,
                    ),
                    source_record=record,
                    src_record=record,
                    dst_record=None,
                )
            )
    contexts.sort(key=lambda item: item.edge.connection_id)
    report.info("built {0} exact direct bit edge(s) from port workbooks".format(len(contexts)))
    return contexts


def read_target_harden_manifest(
    path: Path,
    run_root: Path,
    instances: Dict[str, Dict[str, object]],
    require_complete: bool,
    report: Report,
) -> Tuple[Dict[str, TargetHardenSdc], RunCompleteness]:
    if not path.is_file():
        report.error("required harden SDC manifest is missing: {0}".format(path))
        return {}, RunCompleteness(status="invalid", manifest_path=str(path.resolve()))
    required = {
        "inst_name", "module_name", "sdc_path", "availability_status",
        "sdc_digest", "note",
    }
    entries: Dict[str, TargetHardenSdc] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        missing_headers = sorted(required - set(reader.fieldnames or []))
        if missing_headers:
            report.error("{0}: missing manifest field(s): {1}".format(path, ",".join(missing_headers)))
        for row_idx, row in enumerate(reader, start=2):
            inst_name = clean_cell(row.get("inst_name"))
            if not inst_name:
                report.error("{0} row {1}: empty inst_name".format(path.name, row_idx))
                continue
            if inst_name in entries:
                report.error("{0} row {1}: duplicate inst_name {2}".format(path.name, row_idx, inst_name))
                continue
            entries[inst_name] = TargetHardenSdc(
                inst_name=inst_name,
                module_name=clean_cell(row.get("module_name")),
                sdc_path=clean_cell(row.get("sdc_path")),
                availability_status=normalize_key(row.get("availability_status")),
                sdc_digest=normalize_key(row.get("sdc_digest")),
                note=clean_cell(row.get("note")),
                source_row=row_idx,
            )
    available: List[str] = []
    missing: List[str] = []
    not_required: List[str] = []
    resolved_paths: Dict[str, str] = {}
    for inst_name in sorted(instances):
        info = instances[inst_name]
        entry = entries.get(inst_name)
        if entry is None:
            report.error("{0}: manifest has no row for {1}".format(path.name, inst_name))
            continue
        if entry.module_name != clean_cell(info.get("module_name")):
            report.error("{0} row {1}: module mismatch for {2}".format(path.name, entry.source_row, inst_name))
        status = entry.availability_status
        if status == "available":
            if not entry.sdc_path or not entry.sdc_digest:
                report.error("{0} row {1}: available SDC requires path and digest".format(path.name, entry.source_row))
                continue
            resolved = resolve_path(run_root, entry.sdc_path, entry.sdc_path).resolve()
            if not resolved.is_file():
                report.error("{0} row {1}: available SDC is missing: {2}".format(path.name, entry.source_row, resolved))
                continue
            actual_digest = digest_file(resolved)
            if actual_digest != entry.sdc_digest:
                report.error("{0} row {1}: available SDC digest mismatch for {2}".format(path.name, entry.source_row, inst_name))
                continue
            previous = resolved_paths.get(str(resolved))
            if previous and previous != inst_name:
                report.error("available SDC path is shared by {0} and {1}: {2}".format(previous, inst_name, resolved))
                continue
            resolved_paths[str(resolved)] = inst_name
            entry.resolved_path = resolved
            available.append(inst_name)
        elif status == "missing":
            missing.append(inst_name)
            report.warn("missing harden SDC for {0}: {1}".format(inst_name, entry.note or "<no note>"))
        elif status == "not_required":
            if not entry.note:
                report.error("{0} row {1}: not_required requires a basis in note".format(path.name, entry.source_row))
            not_required.append(inst_name)
        else:
            report.error("{0} row {1}: invalid availability_status {2}".format(path.name, entry.source_row, status or "<empty>"))
    extra = sorted(set(entries) - set(instances))
    for inst_name in extra:
        report.error("{0}: manifest contains unknown instance {1}".format(path.name, inst_name))
    if require_complete and missing:
        report.error("HARDEN_SDC_COMPLETENESS_REQUIRED: " + ",".join(missing))
    status = "partial" if missing else "complete"
    if report.error_count:
        status = "invalid"
    completeness = RunCompleteness(
        status=status,
        available_instances=available,
        missing_instances=missing,
        not_required_instances=not_required,
        manifest_path=str(path.resolve()),
    )
    return entries, completeness


def target_read_json(path: Path, label: str, report: Report) -> Dict[str, object]:
    if not path.is_file():
        report.error("required {0} is missing: {1}".format(label, path))
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        report.error("invalid {0} {1}: {2}".format(label, path, exc))
        return {}
    if not isinstance(payload, dict):
        report.error("{0} must contain a JSON object: {1}".format(label, path))
        return {}
    return payload


def target_validate_provenance(
    payload: Dict[str, object],
    path: Path,
    run_context: Dict[str, str],
    structure: str,
    report: Report,
    require_complete: bool = True,
    require_gate_fields: bool = False,
    expected_stage_name: str = "",
) -> None:
    if not payload:
        return
    if clean_cell(payload.get("run_id")) != clean_cell(run_context.get("run_id")):
        report.error("{0}: run_id does not match inputs/run_context.csv".format(path))
    if clean_cell(payload.get("mode_label")) != clean_cell(run_context.get("mode_label")):
        report.error("{0}: mode_label does not match inputs/run_context.csv".format(path))
    design = clean_cell(payload.get("design_revision"))
    if design and design != clean_cell(run_context.get("design_revision")):
        report.error("{0}: design_revision does not match inputs/run_context.csv".format(path))
    if expected_stage_name and clean_cell(payload.get("stage_name")) != expected_stage_name:
        report.error(
            "{0}: stage_name must be {1}".format(path, expected_stage_name)
        )
    if clean_cell(payload.get("structure_digest")) != structure:
        report.error("{0}: structure_digest is stale".format(path))
    if require_complete:
        if normalize_key(payload.get("completion_status")) != "complete":
            report.error("{0}: completion_status is not complete".format(path))
        if require_gate_fields and "sync_changed" not in payload:
            report.error("{0}: sync_changed is required".format(path))
        elif "sync_changed" in payload and normalize_key(payload.get("sync_changed")) != "no":
            report.error("{0}: sync_changed must be no for a complete artifact".format(path))
        if require_gate_fields and "error_count" not in payload:
            report.error("{0}: error_count is required".format(path))
        try:
            error_count = int(clean_cell(payload.get("error_count")) or "0")
        except ValueError:
            error_count = -1
        if error_count != 0:
            report.error("{0}: error_count is not zero".format(path))


def target_read_csv_rows(path: Path, label: str, report: Report) -> List[Dict[str, str]]:
    if not path.is_file():
        report.error("required {0} is missing: {1}".format(label, path))
        return []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.DictReader(file_obj)
            if not reader.fieldnames:
                report.error("{0} has no CSV header: {1}".format(label, path))
                return []
            return [dict(row) for row in reader]
    except (OSError, csv.Error) as exc:
        report.error("failed to read {0} {1}: {2}".format(label, path, exc))
        return []


def target_validate_output_digest(
    payload: Dict[str, object],
    meta_path: Path,
    default_path: Path,
    report: Report,
    digest_key: str = "output_sdc_digest",
    path_key: str = "output_sdc_path",
    path_base: Optional[Path] = None,
) -> None:
    if not payload:
        return
    expected = clean_cell(payload.get(digest_key))
    path_value = clean_cell(payload.get(path_key))
    output_path = default_path
    if path_value:
        declared_path = Path(path_value).expanduser()
        if not declared_path.is_absolute():
            declared_path = (
                path_base if path_base is not None else meta_path.parent
            ) / declared_path
        if declared_path.resolve() != default_path.resolve():
            report.error(
                "{0}: {1} must identify canonical output {2}".format(
                    meta_path, path_key, default_path
                )
            )
    if not expected:
        report.error("{0}: {1} is required".format(meta_path, digest_key))
        return
    if not output_path.is_file():
        report.error("{0}: declared output is missing: {1}".format(meta_path, output_path))
        return
    if digest_file(output_path) != expected:
        report.error("{0}: {1} does not match {2}".format(meta_path, digest_key, output_path))


def target_validate_artifact_digest(
    payload: Dict[str, object],
    meta_path: Path,
    artifact_path: Path,
    digest_key: str,
    report: Report,
) -> None:
    if not payload:
        return
    expected = clean_cell(payload.get(digest_key))
    if not expected:
        report.error("{0}: {1} is required".format(meta_path, digest_key))
        return
    if not artifact_path.is_file():
        report.error("{0}: required artifact is missing: {1}".format(meta_path, artifact_path))
        return
    if digest_file(artifact_path) != expected:
        report.error("{0}: {1} does not match {2}".format(meta_path, digest_key, artifact_path))


def target_validate_inventory_pair(
    path: Path,
    meta_path: Path,
    label: str,
    run_context: Dict[str, str],
    structure: str,
    report: Report,
    digest_keys: Sequence[str],
) -> Tuple[List[Dict[str, str]], Dict[str, object]]:
    rows = target_read_csv_rows(path, label, report)
    meta = target_read_json(meta_path, label + " meta", report)
    target_validate_provenance(meta, meta_path, run_context, structure, report)
    if path.is_file() and meta:
        actual = digest_file(path)
        expected = ""
        for key in digest_keys:
            expected = clean_cell(meta.get(key))
            if expected:
                break
        if not expected:
            report.error("{0}: no inventory digest field was found".format(meta_path))
        elif expected != actual:
            report.error("{0}: inventory digest is stale".format(meta_path))
    for row_idx, row in enumerate(rows, start=2):
        row_run_id = clean_cell(row.get("run_id"))
        row_mode = clean_cell(row.get("mode_label"))
        row_structure = clean_cell(row.get("structure_digest"))
        if row_run_id and row_run_id != clean_cell(run_context.get("run_id")):
            report.error("{0} row {1}: run_id is stale".format(path, row_idx))
        if row_mode and row_mode != clean_cell(run_context.get("mode_label")):
            report.error("{0} row {1}: mode_label is stale".format(path, row_idx))
        if row_structure and row_structure != structure:
            report.error("{0} row {1}: structure_digest is stale".format(path, row_idx))
    return rows, meta


def target_port_record_lookup(port_records) -> Dict[Tuple[str, str, str], List[Any]]:
    result: Dict[Tuple[str, str, str], List[Any]] = defaultdict(list)
    for record in port_records:
        result[(record.inst_name, record.direction, record.shape.base)].append(record)
    return result


def target_resolve_inventory_endpoint(
    inst_name: str,
    direction: str,
    port_value: str,
    bit_value: str,
    lookup: Dict[Tuple[str, str, str], List[Any]],
) -> List[Tuple[str, str, str, int]]:
    inst = clean_cell(inst_name)
    direction_key = normalize_key(direction)
    port_text = clean_cell(port_value)
    if not inst or not direction_key or not port_text:
        return []
    try:
        base, explicit = target_parse_signal(port_text, "machine inventory endpoint")
    except ValueError:
        return []
    explicit_bit = clean_cell(bit_value)
    if explicit_bit:
        try:
            bits = [int(explicit_bit)]
        except ValueError:
            return []
    elif explicit is not None:
        bits = explicit
    else:
        candidates = lookup.get((inst, direction_key, base), [])
        if len(candidates) != 1:
            return []
        bits = target_ordered_bits(candidates[0].shape)
    return [(inst, direction_key, base, bit) for bit in bits]


def target_delta_owner_map(runtime, run_root: Path, port_records, report: Report):
    record_by_key = runtime.record_index(port_records)
    owners: Dict[Tuple[str, str, str, int], List[Tuple[str, str]]] = defaultdict(list)
    for directory in ("00_middle", "01_middle", "04_middle", "10_middle"):
        path = run_root / directory / "port_accounting_delta.csv"
        if not path.is_file():
            continue
        for row in runtime.load_delta_rows(path, report):
            key = (
                clean_cell(row.get("workbook")),
                clean_cell(row.get("sheet")),
                int(clean_cell(row.get("row")) or "0"),
                normalize_key(row.get("direction")),
                clean_cell(row.get("port")),
            )
            record = record_by_key.get(key)
            if record is None:
                continue
            try:
                added = runtime.parse_bits_field(row.get("added_bits"))
            except ValueError:
                continue
            for bit in added:
                owners[(record.inst_name, record.direction, record.shape.base, bit)].append(
                    (clean_cell(row.get("stage_name")), clean_cell(row.get("owner_object_id")))
                )
    return owners


def validate_target_upstreams(
    runtime,
    run_root: Path,
    run_context: Dict[str, str],
    required_views: Sequence[Dict[str, str]],
    structure: str,
    accounting_before: str,
    port_records,
    report: Report,
) -> Dict[str, object]:
    upstream_digests: Dict[str, str] = {}
    delta_meta: Dict[str, Dict[str, object]] = {}
    stage_completions: Dict[str, Dict[str, object]] = {}
    for directory, label in (
        ("00_middle", "00"),
        ("01_middle", "01"),
        ("04_middle", "04"),
    ):
        expected_stage_name = {
            "00": "00_harden_port_inventory",
            "01": "01_soc_clocks",
            "04": "04_soc_io_pads",
        }[label]
        middle = run_root / directory
        delta_path = middle / "port_accounting_delta.csv"
        delta_meta_path = middle / "port_accounting_delta.meta"
        completion_path = middle / "stage_completion.meta"
        meta = target_read_json(delta_meta_path, label + " accounting delta meta", report)
        completion = target_read_json(completion_path, label + " stage completion", report)
        target_validate_provenance(meta, delta_meta_path, run_context, structure, report)
        target_validate_provenance(
            completion,
            completion_path,
            run_context,
            structure,
            report,
            require_gate_fields=True,
            expected_stage_name=expected_stage_name,
        )
        if label == "01":
            target_validate_output_digest(
                completion,
                completion_path,
                run_root / "01_result" / "01_soc_clocks.sdc",
                report,
                path_base=run_root,
            )
        if not delta_path.is_file():
            report.error("required {0} accounting delta is missing: {1}".format(label, delta_path))
        elif meta and clean_cell(meta.get("delta_csv_digest")) != digest_file(delta_path):
            report.error("{0}: delta_csv_digest is stale".format(delta_meta_path))
        target_validate_artifact_digest(
            completion,
            completion_path,
            delta_path,
            "accounting_delta_digest",
            report,
        )
        if meta and completion:
            if clean_cell(completion.get("accounting_digest_before")) != clean_cell(meta.get("accounting_digest_before")):
                report.error("{0}: completion/delta before-digest mismatch".format(label))
            if clean_cell(completion.get("accounting_digest_after")) != clean_cell(meta.get("accounting_digest_after")):
                report.error("{0}: completion/delta after-digest mismatch".format(label))
        delta_meta[label] = meta
        stage_completions[label] = completion
        for path in (delta_path, delta_meta_path, completion_path):
            upstream_digests["{0}:{1}".format(label, path.name)] = (
                digest_file(path) if path.is_file() else ""
            )

    snapshot_path = run_root / "00_middle" / "input_snapshot.meta"
    snapshot = target_read_json(snapshot_path, "00 input snapshot", report)
    target_validate_provenance(
        snapshot, snapshot_path, run_context, structure, report,
        require_complete=False,
    )
    upstream_digests["00:input_snapshot.meta"] = (
        digest_file(snapshot_path) if snapshot_path.is_file() else ""
    )

    after_00 = clean_cell(delta_meta.get("00", {}).get("accounting_digest_after"))
    before_01 = clean_cell(delta_meta.get("01", {}).get("accounting_digest_before"))
    after_01 = clean_cell(delta_meta.get("01", {}).get("accounting_digest_after"))
    before_04 = clean_cell(delta_meta.get("04", {}).get("accounting_digest_before"))
    after_04 = clean_cell(delta_meta.get("04", {}).get("accounting_digest_after"))
    if after_00 and before_01 != after_00:
        report.error("00->01 accounting digest chain is discontinuous")
    if after_01 and before_04 != after_01:
        report.error("01->04 accounting digest chain is discontinuous")

    empty_digest = runtime.accounting_digest(port_records, empty=True)
    runtime.validate_resume_evidence(
        run_root,
        port_records,
        run_context,
        structure,
        accounting_before,
        empty_digest,
        report,
    )
    existing_10_meta_path = run_root / "10_middle" / "port_accounting_delta.meta"
    existing_10_meta = (
        target_read_json(existing_10_meta_path, "10 accounting delta meta", report)
        if existing_10_meta_path.is_file() else {}
    )
    if existing_10_meta:
        target_validate_provenance(
            existing_10_meta, existing_10_meta_path, run_context, structure, report
        )
    expected_current = clean_cell(
        existing_10_meta.get("accounting_digest_after") if existing_10_meta else after_04
    )
    if expected_current and expected_current != accounting_before:
        report.error(
            "accounting digest chain mismatch before 10: expected={0}, current={1}".format(
                expected_current, accounting_before
            )
        )

    clock_path = run_root / "01_middle" / "clock_inventory.csv"
    clock_meta_path = run_root / "01_middle" / "clock_inventory.meta"
    clock_rows, clock_meta = target_validate_inventory_pair(
        clock_path,
        clock_meta_path,
        "01 clock inventory",
        run_context,
        structure,
        report,
        ("inventory_digest", "clock_inventory_digest"),
    )
    upstream_digests["01:clock_inventory.csv"] = digest_file(clock_path) if clock_path.is_file() else ""
    upstream_digests["01:clock_inventory.meta"] = digest_file(clock_meta_path) if clock_meta_path.is_file() else ""

    relation_path = run_root / "03_middle" / "relation_map.csv"
    relation_meta_path = run_root / "03_middle" / "relation_map.meta"
    relation_completion_path = run_root / "03_middle" / "stage_completion.meta"
    relation_completion = target_read_json(
        relation_completion_path, "03 stage completion", report
    )
    target_validate_provenance(
        relation_completion,
        relation_completion_path,
        run_context,
        structure,
        report,
        require_gate_fields=True,
        expected_stage_name="03_soc_clock_groups",
    )
    target_validate_output_digest(
        relation_completion,
        relation_completion_path,
        run_root / "03_result" / "03_soc_clock_groups.sdc",
        report,
        path_base=run_root,
    )
    relation_rows: List[Dict[str, str]] = []
    relation_meta: Dict[str, object] = {}
    if relation_path.is_file() or relation_meta_path.is_file():
        if not relation_path.is_file() or not relation_meta_path.is_file():
            report.warn("03 relation diagnostic requires paired CSV/meta")
        else:
            try:
                with relation_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
                    reader = csv.DictReader(file_obj)
                    relation_rows = [dict(row) for row in reader] if reader.fieldnames else []
                relation_meta_value = json.loads(relation_meta_path.read_text(encoding="utf-8"))
                relation_meta = relation_meta_value if isinstance(relation_meta_value, dict) else {}
            except (OSError, ValueError, csv.Error) as exc:
                report.warn("03 relation diagnostic is invalid and will be ignored: {0}".format(exc))
                relation_rows = []
                relation_meta = {}
            relation_valid = True
            if clean_cell(relation_meta.get("run_id")) != clean_cell(run_context.get("run_id")):
                relation_valid = False
            if clean_cell(relation_meta.get("mode_label")) != clean_cell(run_context.get("mode_label")):
                relation_valid = False
            if clean_cell(relation_meta.get("structure_digest")) != structure:
                relation_valid = False
            if normalize_key(relation_meta.get("completion_status")) not in {"", "complete"}:
                relation_valid = False
            if not relation_valid:
                report.warn("03 relation map provenance is invalid; relation diagnostic marked unavailable")
                relation_rows = []
            expected_relation_digest = clean_cell(
                relation_meta.get("relation_map_digest") or relation_meta.get("inventory_digest")
            )
            if not expected_relation_digest:
                report.warn("03 relation map digest is missing; relation diagnostic marked unavailable")
                relation_rows = []
            elif expected_relation_digest != digest_file(relation_path):
                report.warn("03 relation map digest is stale; relation diagnostic marked unavailable")
                relation_rows = []
            expected_01_inventory = clean_cell(
                relation_meta.get("upstream_01_inventory_digest")
            )
            if (
                not expected_01_inventory
                or not clock_path.is_file()
                or expected_01_inventory != digest_file(clock_path)
            ):
                report.warn(
                    "03 relation map does not authenticate the current 01 clock inventory; "
                    "relation diagnostic marked unavailable"
                )
                relation_rows = []
            expected_01_completion = clean_cell(
                relation_meta.get("upstream_01_completion_digest")
            )
            completion_01_path = run_root / "01_middle" / "stage_completion.meta"
            if (
                not expected_01_completion
                or not completion_01_path.is_file()
                or expected_01_completion != digest_file(completion_01_path)
            ):
                report.warn(
                    "03 relation map does not authenticate the current 01 completion; "
                    "relation diagnostic marked unavailable"
                )
                relation_rows = []
    else:
        report.warn("optional 03 relation diagnostic is unavailable")
    for path in (relation_path, relation_meta_path, relation_completion_path):
        upstream_digests["03:{0}".format(path.name)] = digest_file(path) if path.is_file() else ""

    pad_path = run_root / "04_middle" / "pad_inventory.csv"
    pad_meta_path = run_root / "04_middle" / "pad_inventory.meta"
    pad_rows, pad_meta = target_validate_inventory_pair(
        pad_path,
        pad_meta_path,
        "04 pad inventory",
        run_context,
        structure,
        report,
        ("inventory_digest", "pad_inventory_digest"),
    )
    upstream_digests["04:pad_inventory.csv"] = digest_file(pad_path) if pad_path.is_file() else ""
    upstream_digests["04:pad_inventory.meta"] = digest_file(pad_meta_path) if pad_meta_path.is_file() else ""
    completion_04 = stage_completions.get("04", {})
    target_validate_artifact_digest(
        completion_04,
        run_root / "04_middle" / "stage_completion.meta",
        pad_path,
        "pad_inventory_digest",
        report,
    )
    required_04_ids = {
        clean_cell(view.get("view_id"))
        for view in required_views
        if normalize_key(view.get("require_04")) == "yes"
    }
    parent_view_digests = completion_04.get("required_view_completions", {})
    if required_04_ids:
        if not isinstance(parent_view_digests, dict):
            report.error(
                "04 run-wide completion requires required_view_completions mapping"
            )
            parent_view_digests = {}
        parent_ids = set(clean_cell(value) for value in parent_view_digests)
        if parent_ids != required_04_ids:
            report.error(
                "04 run-wide required_view_completions does not exactly match "
                "required_views.csv: expected={0}, actual={1}".format(
                    ",".join(sorted(required_04_ids)),
                    ",".join(sorted(parent_ids)),
                )
            )
    for view in required_views:
        if normalize_key(view.get("require_04")) != "yes":
            continue
        completion_path = (
            run_root / "04_middle" / "completion" /
            (safe_token(view.get("stage")) + "_" + safe_token(view.get("corner")) + ".meta")
        )
        completion = target_read_json(completion_path, "04 required view completion", report)
        target_validate_provenance(
            completion,
            completion_path,
            run_context,
            structure,
            report,
            require_gate_fields=True,
            expected_stage_name="04_soc_io_pads",
        )
        stage_token = safe_token(view.get("stage"))
        corner_token = safe_token(view.get("corner"))
        if normalize_key(view.get("stage")) == "all" and normalize_key(view.get("corner")) == "all":
            view_sdc_path = run_root / "04_result" / "04_soc_io_pads.sdc"
        else:
            view_sdc_path = (
                run_root / "04_result" /
                "04_soc_io_pads_{0}_{1}.sdc".format(stage_token, corner_token)
            )
        target_validate_output_digest(
            completion,
            completion_path,
            view_sdc_path,
            report,
            path_base=run_root,
        )
        if completion and clean_cell(completion.get("view_id")) != clean_cell(view.get("view_id")):
            report.error("{0}: view_id does not match required_views.csv".format(completion_path))
        if completion and clean_cell(completion.get("stage")) != clean_cell(view.get("stage")):
            report.error("{0}: stage does not match required_views.csv".format(completion_path))
        if completion and clean_cell(completion.get("corner")) != clean_cell(view.get("corner")):
            report.error("{0}: corner does not match required_views.csv".format(completion_path))
        actual_completion_digest = (
            digest_file(completion_path) if completion_path.is_file() else ""
        )
        expected_completion_digest = clean_cell(
            parent_view_digests.get(clean_cell(view.get("view_id")))
            if isinstance(parent_view_digests, dict) else ""
        )
        if not expected_completion_digest:
            report.error(
                "{0}: run-wide 04 completion does not authenticate this view".format(
                    completion_path
                )
            )
        elif expected_completion_digest != actual_completion_digest:
            report.error(
                "{0}: run-wide 04 required-view digest is stale".format(
                    completion_path
                )
            )
        upstream_digests["04:view:{0}".format(clean_cell(view.get("view_id")))] = (
            actual_completion_digest
        )

    lookup = target_port_record_lookup(port_records)
    clock_endpoints: Set[Tuple[str, str, str, int]] = set()
    clock_aliases: Dict[Tuple[str, str], str] = {}
    alias_candidates: Dict[Tuple[str, str], Set[str]] = defaultdict(set)
    active_clocks: Set[str] = set()
    known_clock_names: Set[str] = set()
    trusted_clock_actions = ACTIVE_CLOCK_ACTIONS | {"check_only", "alias_only"}
    for row in clock_rows:
        action = normalize_key(row.get("final_action"))
        clock_name = clean_cell(row.get("clock_name"))
        if action in ACTIVE_CLOCK_ACTIONS or action == "check_only":
            clock_endpoints.update(
                target_resolve_inventory_endpoint(
                    row.get("inst_name", ""), row.get("direction", ""),
                    row.get("port_name", ""), row.get("bit_index", ""), lookup,
                )
            )
        if action in ACTIVE_CLOCK_ACTIONS and clock_name:
            active_clocks.add(clock_name)
        if action in trusted_clock_actions and clock_name:
            known_clock_names.add(clock_name)
        inst_name = clean_cell(row.get("inst_name"))
        if action in trusted_clock_actions:
            for alias in (clean_cell(row.get("original_clock_name")), clock_name):
                if inst_name and alias and clock_name:
                    alias_candidates[(inst_name, alias)].add(clock_name)
    for key, mapped_names in sorted(alias_candidates.items()):
        if len(mapped_names) == 1:
            clock_aliases[key] = next(iter(mapped_names))
        else:
            report.warn(
                "01 clock inventory has conflicting alias mapping for {0}/{1}; "
                "timing relation will remain unknown".format(key[0], key[1])
            )

    pad_endpoints: Set[Tuple[str, str, str, int]] = set()
    pad_connection_ids: Set[str] = set()
    for row in pad_rows:
        connection_id = clean_cell(row.get("connection_id"))
        if connection_id:
            pad_connection_ids.add(connection_id)
        inst_name = (
            clean_cell(row.get("harden_instance"))
            or clean_cell(row.get("subsys_instance"))
            or clean_cell(row.get("inst_name"))
        )
        port_name = (
            clean_cell(row.get("harden_port"))
            or clean_cell(row.get("subsys_port"))
            or clean_cell(row.get("port"))
            or clean_cell(row.get("port_name"))
        )
        pad_endpoints.update(
            target_resolve_inventory_endpoint(
                inst_name, row.get("direction", ""), port_name,
                row.get("bit_index", ""), lookup,
            )
        )

    relations: Dict[Tuple[str, str], str] = {}
    ambiguous_relations: Set[Tuple[str, str]] = set()
    clock_universe_digest = clean_cell(clock_meta.get("clock_set_digest"))
    for row_idx, row in enumerate(relation_rows, start=2):
        clock_a = clean_cell(row.get("clock_a"))
        clock_b = clean_cell(row.get("clock_b"))
        relation = normalize_key(row.get("relation_type"))
        row_universe = clean_cell(row.get("clock_universe_digest"))
        if not clock_universe_digest or row_universe != clock_universe_digest:
            report.warn(
                "03 relation diagnostic row {0} has stale clock universe and was ignored".format(
                    row_idx
                )
            )
            continue
        if clock_a not in known_clock_names or clock_b not in known_clock_names:
            report.warn(
                "03 relation diagnostic row {0} references clock outside 01 inventory and was ignored".format(
                    row_idx
                )
            )
            continue
        if clock_a and clock_b and relation in CANONICAL_RELATIONS:
            key = tuple(sorted((clock_a, clock_b)))
            if key in ambiguous_relations:
                continue
            previous = relations.get(key)
            if previous is not None and previous != relation:
                relations.pop(key, None)
                ambiguous_relations.add(key)
                report.warn(
                    "03 relation diagnostic has conflicting relations for {0}/{1}; "
                    "timing relation will remain unknown".format(key[0], key[1])
                )
            else:
                relations[key] = relation
        elif clock_a or clock_b or relation:
            report.warn(
                "03 relation diagnostic row {0} is incomplete or has invalid relation_type".format(
                    row_idx
                )
            )

    return {
        "upstream_digests": upstream_digests,
        "upstream_accounting_after": after_04,
        "existing_10_meta": existing_10_meta,
        "clock_endpoints": clock_endpoints,
        "clock_aliases": clock_aliases,
        "active_clocks": active_clocks,
        "relations": relations,
        "pad_endpoints": pad_endpoints,
        "pad_connection_ids": pad_connection_ids,
        "owner_map": target_delta_owner_map(runtime, run_root, port_records, report),
        "clock_meta": clock_meta,
        "relation_meta": relation_meta,
        "pad_meta": pad_meta,
    }


def target_instance_feedthrough_confirmed(instance: Dict[str, object]) -> bool:
    semantic = instance.get("semantic", {}) if isinstance(instance, dict) else {}
    if not isinstance(semantic, dict):
        return False
    for key, value in semantic.items():
        normalized = normalize_key(value)
        if any(token in normalize_key(key) for token in ("feedthrough", "blocktype", "classification", "hardenclass")):
            if normalized in {"yes", "true", "1", "feedthrough", "ft", "dpg"} or "feedthrough" in normalized:
                return True
    return False


def target_edge_endpoint_tuple(edge: ConnectionEdge, side: str) -> Tuple[str, str, str, int]:
    inst = clean_cell(getattr(edge, side + "_instance"))
    direction = normalize_key(getattr(edge, side + "_direction"))
    port = clean_cell(getattr(edge, side + "_port"))
    bit = int(clean_cell(getattr(edge, side + "_bit_index")) or "0")
    try:
        base, _ = target_parse_signal(port, "direct edge endpoint")
    except ValueError:
        base = port
    return inst, direction, base, bit


def classify_target_feedthrough_edges(
    contexts: Sequence[TargetDirectContext],
    manifest: Dict[str, TargetHardenSdc],
    instances: Dict[str, Dict[str, object]],
    upstream: Dict[str, object],
    report: Report,
) -> List[TargetFeedthroughEdge]:
    result: List[TargetFeedthroughEdge] = []
    clock_endpoints = upstream.get("clock_endpoints", set())
    pad_endpoints = upstream.get("pad_endpoints", set())
    pad_connection_ids = upstream.get("pad_connection_ids", set())
    for context in contexts:
        edge = context.edge
        src_key = target_edge_endpoint_tuple(edge, "src")
        dst_key = target_edge_endpoint_tuple(edge, "dst")
        if src_key in clock_endpoints or dst_key in clock_endpoints:
            continue
        if normalize_key(edge.src_instance) == "top" or normalize_key(edge.dst_instance) == "top":
            harden_key = dst_key if normalize_key(edge.src_instance) == "top" else src_key
            if harden_key not in pad_endpoints and edge.connection_id not in pad_connection_ids:
                report.error(
                    "{0}: top/pad direct edge is absent from 04 pad inventory".format(
                        edge.connection_id
                    )
                )
            continue
        src_ft = parse_feedthrough_port(edge.src_port)
        dst_ft = parse_feedthrough_port(edge.dst_port)
        src_is_fto = bool(src_ft and src_ft[0] == "fto")
        dst_is_fti = bool(dst_ft and dst_ft[0] == "fti")
        if not (src_is_fto or dst_is_fti):
            continue
        if edge.src_instance == edge.dst_instance:
            report.error("{0}: harden-internal feedthrough edge is not SoC-visible".format(edge.connection_id))
            continue
        if src_is_fto and edge.src_direction != "output":
            report.error("{0}: fto boundary must be output".format(edge.connection_id))
        if dst_is_fti and edge.dst_direction != "input":
            report.error("{0}: fti boundary must be input".format(edge.connection_id))
        boundary_instances = []
        if src_is_fto:
            boundary_instances.append(edge.src_instance)
        if dst_is_fti:
            boundary_instances.append(edge.dst_instance)
        if any(inst not in manifest or inst not in instances for inst in boundary_instances):
            report.error("{0}: feedthrough boundary is not a manifest harden".format(edge.connection_id))
            continue
        confirmed = all(
            target_instance_feedthrough_confirmed(instances[inst])
            for inst in boundary_instances
        )
        classification_source = "info_all" if confirmed else "name_candidate"
        classification_status = "confirmed" if confirmed else "needs_review"
        if not confirmed:
            report.warn(
                "{0}: feedthrough classification comes only from fti_/fto_ naming; explicit review confirmation is required".format(
                    edge.connection_id
                )
            )
        if src_is_fto and dst_is_fti:
            edge_role = "between_feedthroughs"
            ft_instance = edge.src_instance + "," + edge.dst_instance
            ft_port = edge.src_port + "," + edge.dst_port
        elif dst_is_fti:
            edge_role = "ingress"
            ft_instance = edge.dst_instance
            ft_port = edge.dst_port
        else:
            edge_role = "egress"
            ft_instance = edge.src_instance
            ft_port = edge.src_port
        _, _, digest = target_connection_ids(
            edge.src_instance,
            edge.src_direction,
            src_key[2],
            src_key[3],
            edge.dst_instance,
            edge.dst_direction,
            dst_key[2],
            dst_key[3],
        )
        result.append(
            TargetFeedthroughEdge(
                edge=edge,
                connection_hash=digest,
                edge_role=edge_role,
                feedthrough_instance=ft_instance,
                feedthrough_port=ft_port,
                feedthrough_side=edge_role,
                classification_source=classification_source,
                classification_status=classification_status,
                source_record=context.source_record,
                src_record=context.src_record,
                dst_record=context.dst_record,
            )
        )
    result.sort(key=lambda item: item.edge.connection_id)
    report.info("classified {0} feedthrough-adjacent direct edge(s)".format(len(result)))
    return result


def target_machine_digest(values: Dict[str, object]) -> str:
    payload = "\n".join(
        "{0}={1}".format(header, clean_cell(values.get(header)))
        for header in sorted(TARGET_REVIEW_INVALIDATING_HEADERS)
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def target_sdc_status(inst_name: str, manifest: Dict[str, TargetHardenSdc]) -> str:
    if normalize_key(inst_name) == "top":
        return "not_required"
    entry = manifest.get(inst_name)
    return entry.availability_status if entry is not None else "missing"


def target_resolve_evidence_clock(
    inst_name: str, clock_name: str, upstream: Dict[str, object]
) -> str:
    local = clean_cell(clock_name)
    if not local:
        return ""
    mapped = upstream.get("clock_aliases", {}).get((inst_name, local))
    relation_clocks: Set[str] = set()
    for pair in upstream.get("relations", {}):
        relation_clocks.update(pair)
    if mapped and (
        mapped in upstream.get("active_clocks", set()) or mapped in relation_clocks
    ):
        return mapped
    return "unresolved:" + local


def target_clock_relation(
    src_clocks: Sequence[str],
    dst_clocks: Sequence[str],
    upstream: Dict[str, object],
) -> str:
    if any(
        item and item.startswith("unresolved:")
        for item in list(src_clocks) + list(dst_clocks)
    ):
        return "unknown"
    src = sorted(set(item for item in src_clocks if item and not item.startswith("unresolved:")))
    dst = sorted(set(item for item in dst_clocks if item and not item.startswith("unresolved:")))
    if len(src) != 1 or len(dst) != 1:
        return "unknown"
    if src[0] == dst[0]:
        return "synchronous"
    return upstream.get("relations", {}).get(tuple(sorted((src[0], dst[0]))), "unknown")


def target_edge_seed(
    item: TargetFeedthroughEdge,
    manifest: Dict[str, TargetHardenSdc],
    completeness: RunCompleteness,
    evidence_by_key: Dict[Tuple[str, str, str], List[DelayEvidence]],
    run_context: Dict[str, str],
    structure: str,
    accounting_before: str,
    upstream: Dict[str, object],
) -> Dict[str, object]:
    edge = item.edge
    src_evidence = matching_evidence(
        evidence_by_key, edge.src_instance, edge.src_port, "output_delay"
    )
    dst_evidence = matching_evidence(
        evidence_by_key, edge.dst_instance, edge.dst_port, "input_delay"
    )
    related = src_evidence + dst_evidence
    src_status = target_sdc_status(edge.src_instance, manifest)
    dst_status = target_sdc_status(edge.dst_instance, manifest)
    if "missing" in {src_status, dst_status}:
        evidence_status = "incomplete_missing_sdc"
    elif any(entry.parse_status != "ok" for entry in related):
        evidence_status = "needs_review"
    elif related:
        evidence_status = "complete"
    else:
        evidence_status = "complete_no_delay_candidate"
    src_clocks = [
        target_resolve_evidence_clock(edge.src_instance, entry.clock_name, upstream)
        for entry in src_evidence if entry.clock_name
    ]
    dst_clocks = [
        target_resolve_evidence_clock(edge.dst_instance, entry.clock_name, upstream)
        for entry in dst_evidence if entry.clock_name
    ]
    relation = target_clock_relation(src_clocks, dst_clocks, upstream)
    values: Dict[str, object] = {header: "" for header in TARGET_EDGE_HEADERS}
    values.update(
        {
            "schema_version": TARGET_SCHEMA_VERSION,
            "run_id": run_context.get("run_id", ""),
            "mode_label": run_context.get("mode_label", ""),
            "design_revision": run_context.get("design_revision", ""),
            "run_completeness": completeness.status,
            "structure_digest": structure,
            "accounting_digest_before": accounting_before,
            "accounting_digest_after": accounting_before,
            "feedthrough_edge_id": "FTE_" + item.connection_hash,
            "connection_id": edge.connection_id,
            "edge_role": item.edge_role,
            "src_instance": edge.src_instance,
            "src_direction": edge.src_direction,
            "src_port": edge.src_port,
            "src_bit_index": edge.src_bit_index,
            "src_endpoint": edge.src_endpoint_key,
            "src_soc_object": target_endpoint_collection(edge.src_instance, edge.src_port),
            "dst_instance": edge.dst_instance,
            "dst_direction": edge.dst_direction,
            "dst_port": edge.dst_port,
            "dst_bit_index": edge.dst_bit_index,
            "dst_endpoint": edge.dst_endpoint_key,
            "dst_soc_object": target_endpoint_collection(edge.dst_instance, edge.dst_port),
            "source_workbook": item.source_record.workbook,
            "source_sheet": item.source_record.sheet,
            "source_row": item.source_record.row,
            "feedthrough_instance": item.feedthrough_instance,
            "feedthrough_port": item.feedthrough_port,
            "feedthrough_side": item.feedthrough_side,
            "classification_source": item.classification_source,
            "classification_status": item.classification_status,
            "feedthrough_confirmed": "yes" if item.classification_status == "confirmed" else "no",
            "src_sdc_status": src_status,
            "dst_sdc_status": dst_status,
            "evidence_status": evidence_status,
            "src_output_delay_max": join_unique(entry.max_value or entry.bare_value for entry in src_evidence),
            "src_output_delay_min": join_unique(entry.min_value for entry in src_evidence),
            "dst_input_delay_max": join_unique(entry.max_value or entry.bare_value for entry in dst_evidence),
            "dst_input_delay_min": join_unique(entry.min_value for entry in dst_evidence),
            "src_clock": join_unique(src_clocks),
            "dst_clock": join_unique(dst_clocks),
            "clock_relation": relation,
            "source_sdc_file": join_unique(entry.source_sdc_file for entry in related),
            "source_line": join_unique(entry.source_line for entry in related),
            "source_digest": join_unique(entry.source_digest for entry in related),
            "source_command": join_unique((entry.original_command for entry in related), " || "),
            "channel_disposition": "pending",
            "budget_model": "unknown",
            "budget_required": "",
            "emit_max": "no",
            "emit_min": "no",
            "datapath_only": "yes",
            "tool_surface": "dc",
            "apply": "no",
            "review_status": "pending",
            "validation_status": "matched",
            "machine_note": (
                "feedthrough classification requires explicit review confirmation"
                if item.classification_status != "confirmed" else ""
            ),
        }
    )
    if not related:
        values["machine_note"] = join_unique(
            [values.get("machine_note", ""), "no boundary input/output delay evidence"]
        )
    values["machine_digest"] = target_machine_digest(values)
    return values


def build_target_seeds(
    owned: Sequence[TargetFeedthroughEdge],
    manifest: Dict[str, TargetHardenSdc],
    completeness: RunCompleteness,
    evidence: Sequence[DelayEvidence],
    run_context: Dict[str, str],
    structure: str,
    accounting_before: str,
    upstream: Dict[str, object],
) -> List[Dict[str, object]]:
    index = evidence_index(evidence)
    return [
        target_edge_seed(
            item, manifest, completeness, index, run_context, structure,
            accounting_before, upstream,
        )
        for item in owned
    ]


def target_workbook_row_values(sheet, row_idx: int) -> Dict[str, object]:
    return sheet_row_values(sheet, row_idx, TARGET_EDGE_HEADERS)


def target_run_metadata_rows(
    run_context: Dict[str, str], structure: str, accounting_before: str
) -> List[Tuple[str, str]]:
    return [
        ("schema_version", TARGET_SCHEMA_VERSION),
        ("author", author_name()),
        ("stage_name", "10_feedthrough"),
        ("run_id", run_context.get("run_id", "")),
        ("mode_label", run_context.get("mode_label", "")),
        ("design_revision", run_context.get("design_revision", "")),
        ("structure_digest", structure),
        ("accounting_digest_before", accounting_before),
    ]


def sync_target_workbook(
    path: Path,
    seeds: Sequence[Dict[str, object]],
    evidence: Sequence[DelayEvidence],
    run_context: Dict[str, str],
    structure: str,
    accounting_before: str,
    report: Report,
) -> None:
    workbook, created = create_or_load_workbook(path)
    ensure_sheet(workbook, "feedthrough_edges", TARGET_EDGE_HEADERS)
    ensure_sheet(workbook, "extraction_log", LOG_HEADERS)
    ensure_sheet(workbook, "run_metadata", ["key", "value"])
    sheet = workbook["feedthrough_edges"]
    mapping = header_map(sheet)
    existing: Dict[str, int] = {}
    for row_idx in range(2, sheet.max_row + 1):
        connection_id = clean_cell(sheet.cell(row_idx, mapping["connection_id"]).value)
        if not connection_id:
            continue
        if connection_id in existing:
            report.error("review workbook duplicates connection_id {0}".format(connection_id))
        else:
            existing[connection_id] = row_idx
    changed = created
    seed_ids: Set[str] = set()
    for seed in seeds:
        connection_id = clean_cell(seed.get("connection_id"))
        seed_ids.add(connection_id)
        row_idx = existing.get(connection_id)
        if row_idx is None:
            append_row(sheet, TARGET_EDGE_HEADERS, seed, NEW_FILL)
            existing[connection_id] = sheet.max_row
            changed = True
            report.info("added target review row {0}".format(connection_id))
            continue
        material_changed = False
        for header in sorted(TARGET_MACHINE_HEADERS):
            old_value = clean_cell(sheet.cell(row_idx, mapping[header]).value)
            new_value = clean_cell(seed.get(header))
            if old_value == new_value:
                continue
            sheet.cell(row_idx, mapping[header], seed.get(header, ""))
            if header in TARGET_REVIEW_INVALIDATING_HEADERS:
                material_changed = True
        if material_changed:
            sheet.cell(row_idx, mapping["apply"], "no").fill = NEW_FILL
            sheet.cell(row_idx, mapping["review_status"], "pending").fill = NEW_FILL
            sheet.cell(row_idx, mapping["approved_machine_digest"], "")
            changed = True
            report.warn(
                "{0}: material machine evidence changed; approval reset".format(
                    connection_id
                )
            )
        if (
            normalize_key(sheet.cell(row_idx, mapping["apply"]).value) == "yes"
            and normalize_key(sheet.cell(row_idx, mapping["review_status"]).value) == "approved"
        ):
            approved_cell = sheet.cell(row_idx, mapping["approved_machine_digest"])
            digest = clean_cell(seed.get("machine_digest"))
            if clean_cell(approved_cell.value) != digest:
                approved_cell.value = digest

    for connection_id, row_idx in existing.items():
        if connection_id not in seed_ids:
            report.error(
                "review workbook contains obsolete target connection row {0} at row {1}".format(
                    connection_id, row_idx
                )
            )

    log_sheet = workbook["extraction_log"]
    if log_sheet.max_row > 1:
        log_sheet.delete_rows(2, log_sheet.max_row - 1)
    for values in evidence_log_rows(evidence):
        append_row(log_sheet, LOG_HEADERS, values)

    metadata_sheet = workbook["run_metadata"]
    if metadata_sheet.max_row > 1:
        metadata_sheet.delete_rows(2, metadata_sheet.max_row - 1)
    for key, value in target_run_metadata_rows(
        run_context, structure, accounting_before
    ):
        append_row(metadata_sheet, ["key", "value"], {"key": key, "value": value})

    add_target_workbook_validations(workbook)
    ensure_sheet(workbook, "feedthrough_edges", TARGET_EDGE_HEADERS)
    ensure_sheet(workbook, "extraction_log", LOG_HEADERS)
    ensure_sheet(workbook, "run_metadata", ["key", "value"])
    atomic_save_workbook(workbook, path)
    report.sync_changed = changed
    if changed:
        report.warn("review workbook synchronized; formal SDC/accounting requires re-review")


def add_target_workbook_validations(workbook: Workbook) -> None:
    sheet = workbook["feedthrough_edges"]
    mapping = header_map(sheet)
    sheet.data_validations.dataValidation = []
    choices = {
        "channel_disposition": sorted(DISPOSITIONS),
        "budget_model": sorted(item for item in BUDGET_MODELS if item),
        "budget_required": ["yes", "no"],
        "emit_max": ["yes", "no"],
        "emit_min": ["yes", "no"],
        "datapath_only": ["yes", "no"],
        "tool_surface": ["dc", "sta", "both"],
        "feedthrough_confirmed": ["yes", "no"],
        "apply": ["yes", "no"],
        "review_status": ["pending", "approved", "rejected"],
    }
    for header, values in choices.items():
        if header not in mapping:
            continue
        validation = DataValidation(
            type="list", formula1='"{0}"'.format(",".join(values)), allow_blank=True
        )
        sheet.add_data_validation(validation)
        letter = get_column_letter(mapping[header])
        validation.add("{0}2:{0}1048576".format(letter))


def read_target_form_rows(path: Path) -> List[FormRow]:
    workbook = load_workbook(str(path), data_only=False)
    if "feedthrough_edges" not in workbook.sheetnames:
        raise RuntimeError("{0}: missing feedthrough_edges sheet".format(path))
    sheet = workbook["feedthrough_edges"]
    mapping = header_map(sheet)
    rows = [
        FormRow(row_idx, target_workbook_row_values(sheet, row_idx))
        for row_idx in range(2, sheet.max_row + 1)
        if clean_cell(sheet.cell(row_idx, mapping.get("connection_id", 1)).value)
    ]
    workbook.close()
    return rows


def validate_target_rows(
    rows: Sequence[FormRow],
    seeds: Sequence[Dict[str, object]],
    report: Report,
) -> None:
    seed_by_id = {clean_cell(seed.get("connection_id")): seed for seed in seeds}
    seen: Set[str] = set()
    for row in rows:
        values = row.values
        connection_id = clean_cell(values.get("connection_id"))
        if connection_id in seen:
            report.error("feedthrough_edges row {0}: duplicate connection_id".format(row.row_idx))
            continue
        seen.add(connection_id)
        seed = seed_by_id.get(connection_id)
        if seed is None:
            report.error("feedthrough_edges row {0}: connection is not owned by current 10 run".format(row.row_idx))
            continue
        for header in TARGET_REVIEW_INVALIDATING_HEADERS.union({"machine_digest"}):
            if clean_cell(values.get(header)) != clean_cell(seed.get(header)):
                report.error(
                    "feedthrough_edges row {0}: machine field {1} is stale".format(
                        row.row_idx, header
                    )
                )
        disposition = normalize_key(values.get("channel_disposition"))
        apply_value = normalize_key(values.get("apply"))
        review_status = normalize_key(values.get("review_status"))
        emit_max = normalize_key(values.get("emit_max"))
        emit_min = normalize_key(values.get("emit_min"))
        budget_required = normalize_key(values.get("budget_required"))
        budget_model = normalize_key(values.get("budget_model"))
        if disposition not in DISPOSITIONS:
            report.error("feedthrough_edges row {0}: invalid channel_disposition".format(row.row_idx))
        if apply_value not in YES_NO or review_status not in REVIEW_STATUSES:
            report.error("feedthrough_edges row {0}: invalid apply/review_status".format(row.row_idx))
        if emit_max not in YES_NO or emit_min not in YES_NO:
            report.error("feedthrough_edges row {0}: emit flags must be yes/no".format(row.row_idx))
        if disposition != "emit_budget" and (emit_max == "yes" or emit_min == "yes"):
            report.error("feedthrough_edges row {0}: only emit_budget may emit commands".format(row.row_idx))
        approved = apply_value == "yes" and review_status == "approved"
        if not approved:
            if disposition == "route_to_30" and not clean_cell(values.get("basis")):
                report.warn("feedthrough_edges row {0}: route_to_30 lacks follow-up basis".format(row.row_idx))
            continue
        if disposition == "pending":
            report.error("feedthrough_edges row {0}: pending cannot be approved/applied".format(row.row_idx))
            continue
        missing_review = [
            name for name in ("owner", "reviewer", "review_date", "basis")
            if not clean_cell(values.get(name))
        ]
        if missing_review:
            report.error(
                "feedthrough_edges row {0}: terminal review missing {1}".format(
                    row.row_idx, ",".join(missing_review)
                )
            )
        approved_digest = clean_cell(values.get("approved_machine_digest"))
        if not approved_digest or approved_digest != clean_cell(values.get("machine_digest")):
            report.error("feedthrough_edges row {0}: approved machine digest is stale".format(row.row_idx))
        if (
            normalize_key(values.get("classification_status")) != "confirmed"
            and normalize_key(values.get("feedthrough_confirmed")) != "yes"
        ):
            report.error("feedthrough_edges row {0}: name-only feedthrough candidate requires explicit confirmation".format(row.row_idx))
        if disposition == "emit_budget":
            if budget_required != "yes":
                report.error("feedthrough_edges row {0}: emit_budget requires budget_required=yes".format(row.row_idx))
            if budget_model not in {"manual_budget", "interconnect_budget", "reviewed_two_side_budget"}:
                report.error("feedthrough_edges row {0}: emit_budget requires a reviewed budget model".format(row.row_idx))
            if emit_max != "yes" and emit_min != "yes":
                report.error("feedthrough_edges row {0}: emit_budget must emit max and/or min".format(row.row_idx))
            if normalize_key(values.get("datapath_only")) != "yes":
                report.error("feedthrough_edges row {0}: emit_budget requires datapath_only=yes".format(row.row_idx))
            if normalize_key(values.get("tool_surface")) not in {"dc", "sta", "both"}:
                report.error("feedthrough_edges row {0}: emit_budget requires tool_surface".format(row.row_idx))
            if emit_max == "yes" and parse_number(values.get("converted_max")) is None:
                report.error("feedthrough_edges row {0}: emit_max requires finite converted_max".format(row.row_idx))
            if emit_min == "yes":
                if parse_number(values.get("converted_min")) is None:
                    report.error("feedthrough_edges row {0}: emit_min requires finite converted_min".format(row.row_idx))
                if normalize_key(values.get("min_sign_review")) not in {"yes", "approved", "reviewed"}:
                    report.error("feedthrough_edges row {0}: emit_min requires min_sign_review".format(row.row_idx))
            if "missing" in {
                normalize_key(values.get("src_sdc_status")),
                normalize_key(values.get("dst_sdc_status")),
            } and not clean_cell(values.get("sdc_independent_basis")):
                report.error("feedthrough_edges row {0}: emit with missing SDC requires sdc_independent_basis".format(row.row_idx))
            if normalize_key(values.get("clock_relation")) in {
                "asynchronous", "logically_exclusive", "physically_exclusive"
            } and not clean_cell(values.get("relationship_override_basis")):
                report.error("feedthrough_edges row {0}: async/exclusive normal budget requires override basis".format(row.row_idx))
            if normalize_key(values.get("clock_relation")) == "unknown":
                report.warn("feedthrough_edges row {0}: clock relation is unknown; budget basis must not assume synchronous timing".format(row.row_idx))
        elif disposition in {"no_soc_budget_required", "not_applicable"}:
            if emit_max != "no" or emit_min != "no":
                report.error("feedthrough_edges row {0}: non-emitting terminal requires emit_max=no and emit_min=no".format(row.row_idx))
            if disposition == "no_soc_budget_required" and budget_required != "no":
                report.error("feedthrough_edges row {0}: no-budget terminal requires budget_required=no".format(row.row_idx))
        elif disposition == "route_to_30":
            if emit_max == "yes" or emit_min == "yes":
                report.error("feedthrough_edges row {0}: route_to_30 cannot emit normal budget".format(row.row_idx))
    for connection_id in sorted(set(seed_by_id) - seen):
        report.error("review workbook is missing current edge {0}".format(connection_id))
    bus_groups: Dict[Tuple[str, ...], List[FormRow]] = defaultdict(list)
    for row in rows:
        key = compaction_cohort_key(row)
        if key is not None:
            bus_groups[key].append(row)
    for key, group in bus_groups.items():
        if len(group) < 2:
            continue
        terminal_count = sum(1 for row in group if terminal_disposition(row))
        if 0 < terminal_count < len(group):
            report.warn(
                "feedthrough bus cohort is only partially terminal: {0}/{1} bit edge(s)".format(
                    terminal_count, len(group)
                )
            )


def target_form_rows(rows: Sequence[FormRow]) -> List[FormRow]:
    return sorted(rows, key=lambda row: clean_cell(row.values.get("connection_id")))


def build_target_emission_blocks(
    rows: Sequence[FormRow], all_edges: Sequence[ConnectionEdge]
) -> List[EmissionBlock]:
    cohorts: Dict[Tuple[str, ...], List[FormRow]] = defaultdict(list)
    standalone: List[FormRow] = []
    for row in target_form_rows(rows):
        key = compaction_cohort_key(row)
        if key is None:
            standalone.append(row)
        else:
            cohorts[key].append(row)
    blocks: List[EmissionBlock] = []
    for cohort in cohorts.values():
        merged = compacted_emission_block(cohort, all_edges, "all")
        if merged is not None:
            blocks.append(merged)
        else:
            for row in cohort:
                commands = commands_for_row(row, "all")
                if commands:
                    blocks.append(EmissionBlock(rows=[row], commands=commands))
    for row in standalone:
        commands = commands_for_row(row, "all")
        if commands:
            blocks.append(EmissionBlock(rows=[row], commands=commands))
    return sorted(blocks, key=lambda block: min(row.row_idx for row in block.rows))


def target_record_for_endpoint(
    edge: ConnectionEdge,
    side: str,
    lookup: Dict[Tuple[str, str, str], List[Any]],
    report: Report,
) -> Optional[Tuple[Any, int]]:
    inst, direction, base, bit = target_edge_endpoint_tuple(edge, side)
    if normalize_key(inst) == "top":
        return None
    matches = [
        record for record in lookup.get((inst, direction, base), [])
        if record.shape.contains(bit)
    ]
    if len(matches) != 1:
        report.error(
            "{0}: accounting endpoint {1} matched {2} workbook row(s)".format(
                edge.connection_id, target_endpoint_key(inst, direction, target_exact_port(matches[0].shape, bit)) if matches else "{0}:{1}:{2}[{3}]".format(inst, direction, base, bit), len(matches)
            )
        )
        return None
    return matches[0], bit


def plan_target_feedthrough_accounting(
    rows: Sequence[FormRow],
    owned: Sequence[TargetFeedthroughEdge],
    port_records,
    upstream: Dict[str, object],
    runtime,
    report: Report,
) -> List[Tuple[FormRow, TargetFeedthroughEdge, Any, int, str, bool]]:
    lookup = target_port_record_lookup(port_records)
    owned_by_connection = {
        item.edge.connection_id: item for item in owned
    }
    owner_map = upstream.get("owner_map", {})
    planned: List[Tuple[FormRow, TargetFeedthroughEdge, Any, int, str, bool]] = []
    for row in target_form_rows(rows):
        if not terminal_disposition(row):
            continue
        connection_id = clean_cell(row.values.get("connection_id"))
        item = owned_by_connection.get(connection_id)
        if item is None:
            continue
        for side in ("src", "dst"):
            resolved = target_record_for_endpoint(
                item.edge, side, lookup, report
            )
            if resolved is None:
                continue
            record, bit = resolved
            owner_key = (record.inst_name, record.direction, record.shape.base, bit)
            early_owners = [
                owner for owner in owner_map.get(owner_key, [])
                if owner[0] in {"00_harden_port_inventory", "01_soc_clocks", "04_soc_io_pads"}
            ]
            if early_owners:
                report.info(
                    "{0}: {1} endpoint already accounted by earlier owner(s) {2}; 10 keeps timing ownership without another Used claim".format(
                        connection_id,
                        side,
                        ",".join("{0}:{1}".format(stage, owner_id) for stage, owner_id in early_owners),
                    )
                )
                continue
            added = bit not in record.used_bits
            if added:
                record.used_bits.add(bit)
                record.added_bits.add(bit)
                record.modified = True
                record.model.modified = True
            planned.append((row, item, record, bit, side, added))
            report.info(
                "accounting {0}:{1} row {2} {3} {4}: added_bits={5} final_used_bits={6}".format(
                    record.workbook,
                    record.sheet,
                    record.row,
                    record.direction,
                    record.shape.raw,
                    str(bit) if added else "<none>",
                    runtime.format_bits(record.used_bits),
                )
            )
    for record in port_records:
        if record.modified:
            cell = record.model.workbook[record.sheet].cell(record.row, record.used_col)
            cell.value = runtime.format_bits(record.used_bits)
            cell.number_format = "@"
    return planned


def make_target_10_transaction_id(run_id: str) -> str:
    timestamp = datetime.datetime.utcnow().strftime("%Y%m%dT%H%M%S%fZ")
    seed = "10|{0}|{1}|{2}|{3}".format(
        run_id, timestamp, os.getpid(), socket.gethostname()
    )
    return "10_{0}_{1}".format(
        timestamp, hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12]
    )


def build_target_10_delta_rows(
    planned: Sequence[Tuple[FormRow, TargetFeedthroughEdge, Any, int, str, bool]],
    run_context: Dict[str, str],
    transaction_id: str,
    structure: str,
    before: str,
    after: str,
    runtime,
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for form_row, item, record, bit, side, added in planned:
        rows.append(
            {
                "schema_version": TARGET_SCHEMA_VERSION,
                "run_id": run_context.get("run_id", ""),
                "mode_label": run_context.get("mode_label", ""),
                "stage_name": "10_feedthrough",
                "transaction_id": transaction_id,
                "view_id": "",
                "stage": "",
                "corner": "",
                "structure_digest": structure,
                "accounting_digest_before": before,
                "accounting_digest_after": after,
                "workbook": record.workbook,
                "sheet": record.sheet,
                "row": record.row,
                "direction": record.direction,
                "port": record.shape.raw,
                "legal_bits": runtime.format_bits(set(record.shape.bits())),
                "added_bits": str(bit) if added else "",
                "final_used_bits": runtime.format_bits(record.used_bits),
                "owner_object_id": clean_cell(form_row.values.get("feedthrough_edge_id")),
                "reason": "feedthrough_{0}_{1}".format(
                    normalize_key(form_row.values.get("channel_disposition")), side
                ),
                "evidence_status": "approved",
            }
        )
    return rows


def target_inventory_text(rows: Sequence[FormRow]) -> str:
    stream = io.StringIO()
    writer = csv.DictWriter(
        stream, fieldnames=TARGET_EDGE_HEADERS, extrasaction="ignore", lineterminator="\n"
    )
    writer.writeheader()
    for row in target_form_rows(rows):
        writer.writerow(
            {header: clean_cell(row.values.get(header)) for header in TARGET_EDGE_HEADERS}
        )
    return stream.getvalue()


def render_target_10_sdc(
    blocks: Sequence[EmissionBlock],
    run_context: Dict[str, str],
    completeness: RunCompleteness,
    structure: str,
    accounting_before: str,
    accounting_after: str,
    manifest_path: Path,
) -> str:
    lines = [
        "################################################################################",
        "# Auto-generated SoC feedthrough direct-edge constraints",
        "# Author: {0}".format(author_name()),
        "# Stage: 10_feedthrough",
        "# Script: 10_extract_feedthrough.py",
        "# Run ID: {0}".format(run_context.get("run_id", "")),
        "# Mode label: {0}".format(run_context.get("mode_label", "")),
        "# Design revision: {0}".format(run_context.get("design_revision", "")),
        "# Port accounting: enabled",
        "# Structure digest: {0}".format(structure),
        "# Accounting digest before: {0}".format(accounting_before),
        "# Accounting digest after: {0}".format(accounting_after),
        "# Run completeness: {0}".format(completeness.status),
        "# Harden SDC manifest: {0}".format(manifest_path.resolve()),
        "# Direct edges only; no harden-internal traversal or stitched end-to-end path is emitted.",
        "################################################################################",
        "",
    ]
    emitted = 0
    for block in blocks:
        first = block.rows[0].values
        if block.merged:
            append_chunked_comment(
                lines,
                "Merged feedthrough_edge_ids",
                [clean_cell(row.values.get("feedthrough_edge_id")) for row in block.rows],
            )
            append_chunked_comment(
                lines,
                "connection_ids",
                [clean_cell(row.values.get("connection_id")) for row in block.rows],
            )
        else:
            lines.append(
                "# {0} connection_id={1} role={2}".format(
                    comment_text(first.get("feedthrough_edge_id")),
                    comment_text(first.get("connection_id")),
                    comment_text(first.get("edge_role")),
                )
            )
        if clean_cell(first.get("basis")):
            lines.append("# Basis: " + comment_text(first.get("basis")))
        for command in block.commands:
            if " -through " in " " + command + " ":
                raise RuntimeError("target 10 attempted to emit a -through command")
            lines.append(command)
            emitted += 1
        lines.append("")
    if emitted == 0:
        lines.append("# No feedthrough direct-edge timing commands emitted for this run.")
    return "\n".join(lines).rstrip() + "\n"


def target_inventory_meta_payload(
    run_context: Dict[str, str],
    completion_status: str,
    structure: str,
    before: str,
    after: str,
    inventory_path: Path,
    inventory_digest: str,
    form_path: Path,
    output_path: Path,
    output_digest: str,
    completeness: RunCompleteness,
    workbook_before: Dict[str, str],
    workbook_after: Dict[str, str],
    upstream_digests: Dict[str, str],
) -> Dict[str, object]:
    return {
        "schema_version": TARGET_SCHEMA_VERSION,
        "author": author_name(),
        "stage_name": "10_feedthrough",
        "run_id": run_context.get("run_id", ""),
        "mode_label": run_context.get("mode_label", ""),
        "design_revision": run_context.get("design_revision", ""),
        "completion_status": completion_status,
        "structure_digest": structure,
        "accounting_digest_before": before,
        "accounting_digest_after": after,
        "inventory_path": str(inventory_path.resolve()),
        "inventory_digest": inventory_digest,
        "review_workbook_path": str(form_path.resolve()),
        "review_workbook_digest": digest_file(form_path) if form_path.is_file() else "",
        "output_sdc_path": str(output_path.resolve()) if output_digest else "",
        "output_sdc_digest": output_digest,
        "run_completeness": completeness.status,
        "missing_instances": completeness.missing_instances,
        "workbook_file_digest_before": workbook_before,
        "workbook_file_digest_after": workbook_after,
        "upstream_artifact_digests": upstream_digests,
    }


def render_target_10_report(
    report: Report,
    rows: Sequence[FormRow],
    run_context: Dict[str, str],
    completeness: RunCompleteness,
    structure: str,
    accounting_before: str,
    accounting_after: str,
    completion_status: str,
    diagnostic: bool,
    planned: Sequence[Tuple[FormRow, TargetFeedthroughEdge, Any, int, str, bool]],
) -> str:
    dispositions: Dict[str, int] = defaultdict(int)
    for row in rows:
        dispositions[normalize_key(row.values.get("channel_disposition")) or "<blank>"] += 1
    lines = [
        "10_feedthrough direct-edge report",
        "=================================",
        "",
        "Author: {0}".format(author_name()),
        "Stage: 10_feedthrough",
        "Script: 10_extract_feedthrough.py",
        "Run ID: {0}".format(run_context.get("run_id", "")),
        "Mode label: {0}".format(run_context.get("mode_label", "")),
        "Design revision: {0}".format(run_context.get("design_revision", "")),
        "Completion status: {0}".format(completion_status),
        (
            "Port accounting: diagnostic/read-only"
            if diagnostic else
            ("Port accounting: enabled" if completion_status == "complete" else "Port accounting: not committed")
        ),
        (
            "Accounting closure: not evaluated"
            if diagnostic else "Accounting closure: stage terminal bits {0}".format(
                "committed" if completion_status == "complete" else "not committed"
            )
        ),
        "Run completeness: {0}".format(completeness.status),
        "Structure digest: {0}".format(structure or "<unavailable>"),
        "Accounting digest before: {0}".format(accounting_before or "<unavailable>"),
        "Accounting digest after: {0}".format(accounting_after or "<unavailable>"),
        "Warnings: {0}".format(report.warning_count),
        "Errors: {0}".format(report.error_count),
        "Sync changed: {0}".format("yes" if report.sync_changed else "no"),
        "",
        "Coverage:",
        "  feedthrough direct bit edges: {0}".format(len(rows)),
        "  terminal rows: {0}".format(sum(1 for row in rows if terminal_disposition(row))),
        "  accounting endpoint claims: {0}".format(len(planned)),
        "  newly added bits: {0}".format(sum(1 for item in planned if item[5])),
        "  dispositions:",
    ]
    for disposition in sorted(dispositions):
        lines.append("    {0}: {1}".format(disposition, dispositions[disposition]))
    lines.extend(["", "Accounting details:"])
    for form_row, item, record, bit, side, added in planned:
        lines.append(
            "  {0} {1} {2}:{3}:{4} bit={5} added={6} final={7}".format(
                clean_cell(form_row.values.get("feedthrough_edge_id")),
                side,
                record.workbook,
                record.sheet,
                record.row,
                bit,
                "yes" if added else "no",
                ",".join(str(value) for value in sorted(record.used_bits)),
            )
        )
    lines.extend(["", "Messages:"])
    lines.extend(report.lines or ["INFO: no messages"])
    return "\n".join(lines).rstrip() + "\n"


def invalidate_prior_target_completion(
    runtime,
    input_root: Path,
    completion_path: Path,
    prior_completion_digest: str,
    run_context: Dict[str, str],
    structure: str,
    accounting_before: str,
    models: Sequence[Any],
    upstream: Dict[str, object],
    form_path: Path,
    report_path: Path,
    completeness: RunCompleteness,
    report: Report,
) -> None:
    if not prior_completion_digest or not structure:
        return
    try:
        with runtime.AccountingLock(input_root / ".port_accounting.lock", report):
            runtime.recover_transactions(completion_path.parents[1], report)
            if not completion_path.is_file():
                return
            if digest_file(completion_path) != prior_completion_digest:
                report.warn(
                    "10 completion changed while the failed run was active; "
                    "the newer marker was not invalidated"
                )
                return
            try:
                prior = json.loads(completion_path.read_text(encoding="utf-8"))
            except (OSError, ValueError) as exc:
                report.warn("could not inspect prior 10 completion marker: {0}".format(exc))
                return
            if not isinstance(prior, dict):
                return
            if normalize_key(prior.get("completion_status")) != "complete":
                return
            workbook_digests: Dict[str, str] = {}
            workbook_changed = False
            for model in models:
                if not model.path.is_file():
                    workbook_changed = True
                    continue
                current_digest = digest_file(model.path)
                workbook_digests[model.relative_name] = current_digest
                if current_digest != model.digest_before:
                    workbook_changed = True
            if workbook_changed:
                completion_path.unlink()
                report.warn(
                    "port workbook state changed during the failed run; the prior complete "
                    "marker was removed instead of publishing an unverifiable failed marker"
                )
                return
            failed_completion = {
                "schema_version": TARGET_SCHEMA_VERSION,
                "author": author_name(),
                "stage_name": "10_feedthrough",
                "stage": "",
                "corner": "",
                "run_id": run_context.get("run_id", ""),
                "mode_label": run_context.get("mode_label", ""),
                "design_revision": run_context.get("design_revision", ""),
                "completion_status": "failed",
                "error_count": max(1, report.error_count),
                "sync_changed": "yes" if report.sync_changed else "no",
                "structure_digest": structure,
                "accounting_digest_before": accounting_before,
                "accounting_digest_after": accounting_before,
                "port_accounting": "not_committed",
                "added_bits": 0,
                "workbook_file_digest_before": workbook_digests,
                "workbook_file_digest_after": workbook_digests,
                "upstream_artifact_digests": upstream.get("upstream_digests", {}),
                "output_sdc_digest": "",
                "feedthrough_inventory_digest": "",
                "accounting_delta_digest": "",
                "review_workbook_digest": (
                    digest_file(form_path) if form_path.is_file() else ""
                ),
                "transaction_id": "",
                "run_completeness": completeness.status,
                "missing_instances": completeness.missing_instances,
                "failed_at": runtime.utc_timestamp(),
                "failure_report_path": str(report_path.resolve()),
                "invalidated_completion_digest": prior_completion_digest,
            }
            atomic_write_text(
                completion_path,
                json.dumps(
                    failed_completion,
                    ensure_ascii=False,
                    sort_keys=True,
                    indent=2,
                ) + "\n",
            )
    except Exception as exc:
        report.warn(
            "could not invalidate prior complete 10 marker after failure: {0}".format(exc)
        )


def run_target_single_run(args: argparse.Namespace, argv: Sequence[str]) -> int:
    report = Report()
    runtime = load_accounting_runtime()
    run_root = Path(args.run_root).expanduser().resolve()
    input_root = run_root / "inputs"
    middle_root = run_root / "10_middle"
    result_root = run_root / "10_result"
    form_path = middle_root / "10_feedthrough.xlsx"
    inventory_path = middle_root / "feedthrough_edge_inventory.csv"
    inventory_meta_path = middle_root / "feedthrough_edge_inventory.meta"
    delta_path = middle_root / "port_accounting_delta.csv"
    delta_meta_path = middle_root / "port_accounting_delta.meta"
    completion_path = middle_root / "stage_completion.meta"
    output_path = result_root / "10_feedthrough.sdc"
    report_path = result_root / "reports" / "feedthrough_check_report.txt"
    manifest_path = run_root / "00_middle" / "harden_sdc_manifest.csv"
    run_context_path = input_root / "run_context.csv"
    required_views_path = input_root / "required_views.csv"
    info_path = input_root / "info_all.xlsx"
    diagnostic = bool(args.diagnose_only)

    run_context: Dict[str, str] = {}
    required_views: List[Dict[str, str]] = []
    instances: Dict[str, Dict[str, object]] = {}
    models: List[Any] = []
    port_records: List[Any] = []
    owned: List[TargetFeedthroughEdge] = []
    rows: List[FormRow] = []
    planned: List[Tuple[FormRow, TargetFeedthroughEdge, Any, int, str, bool]] = []
    completeness = RunCompleteness(status="invalid")
    structure = ""
    accounting_before = ""
    accounting_after = ""
    upstream: Dict[str, object] = {}
    preview_dir: Optional[Path] = None
    completion_status = "failed"
    prior_completion_digest = ""
    try:
        if not input_root.is_dir():
            report.error("required inputs directory is missing: {0}".format(input_root))
            raise RuntimeError("target input layout validation failed")
        with runtime.AccountingLock(input_root / ".port_accounting.lock", report):
            runtime.recover_transactions(run_root, report)
            prior_completion_digest = (
                digest_file(completion_path) if completion_path.is_file() else ""
            )
            run_context = runtime.read_run_context(run_context_path, report)
            required_views = runtime.read_required_views(required_views_path, report)
            instances, info_semantic = runtime.read_info_all(info_path, report)
            port_paths = runtime.discover_port_workbooks(input_root, report)
            models, port_records = runtime.read_port_workbooks(
                port_paths, run_root, instances, True, report
            )
            runtime.validate_connections(port_records, instances, report)
            structure = runtime.structure_digest(
                run_context, required_views, info_semantic, port_records
            )
            accounting_before = runtime.accounting_digest(port_records)
            accounting_after = accounting_before
            upstream = validate_target_upstreams(
                runtime,
                run_root,
                run_context,
                required_views,
                structure,
                accounting_before,
                port_records,
                report,
            )
            manifest, completeness = read_target_harden_manifest(
                manifest_path,
                run_root,
                instances,
                args.require_complete_harden_sdc,
                report,
            )
            direct_contexts = build_target_direct_edges(runtime, port_records, report)
            owned = classify_target_feedthrough_edges(
                direct_contexts, manifest, instances, upstream, report
            )
            evidence = extract_delay_evidence(manifest, report)
            seeds = build_target_seeds(
                owned,
                manifest,
                completeness,
                evidence,
                run_context,
                structure,
                accounting_before,
                upstream,
            )

            if report.error_count:
                raise RuntimeError("target 10 input validation failed")

            if diagnostic:
                rows = [FormRow(index + 2, dict(seed)) for index, seed in enumerate(seeds)]
                inventory_text = target_inventory_text(rows)
                inventory_digest = hashlib.sha256(inventory_text.encode("utf-8")).hexdigest()
                workbook_digests = {model.relative_name: model.digest_before for model in models}
                diagnostic_inventory_path = (
                    middle_root / "diagnostic" / "feedthrough_edge_inventory.csv"
                )
                diagnostic_inventory_meta_path = (
                    middle_root / "diagnostic" / "feedthrough_edge_inventory.meta"
                )
                meta = target_inventory_meta_payload(
                    run_context,
                    "diagnostic",
                    structure,
                    accounting_before,
                    accounting_before,
                    diagnostic_inventory_path,
                    inventory_digest,
                    form_path,
                    output_path,
                    "",
                    completeness,
                    workbook_digests,
                    workbook_digests,
                    upstream.get("upstream_digests", {}),
                )
                report_text = render_target_10_report(
                    report,
                    rows,
                    run_context,
                    completeness,
                    structure,
                    accounting_before,
                    accounting_before,
                    "diagnostic",
                    True,
                    [],
                )
                atomic_write_text(diagnostic_inventory_path, inventory_text)
                atomic_write_text(
                    diagnostic_inventory_meta_path,
                    json.dumps(meta, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
                )
                atomic_write_text(report_path, report_text)
                completion_status = "diagnostic"
            else:
                sync_target_workbook(
                    form_path,
                    seeds,
                    evidence,
                    run_context,
                    structure,
                    accounting_before,
                    report,
                )
                rows = read_target_form_rows(form_path)
                if report.sync_changed:
                    inventory_text = target_inventory_text(rows)
                    inventory_digest = hashlib.sha256(inventory_text.encode("utf-8")).hexdigest()
                    workbook_digests = {model.relative_name: model.digest_before for model in models}
                    meta = target_inventory_meta_payload(
                        run_context,
                        "review_required",
                        structure,
                        accounting_before,
                        accounting_before,
                        inventory_path,
                        inventory_digest,
                        form_path,
                        output_path,
                        "",
                        completeness,
                        workbook_digests,
                        workbook_digests,
                        upstream.get("upstream_digests", {}),
                    )
                    report_text = render_target_10_report(
                        report,
                        rows,
                        run_context,
                        completeness,
                        structure,
                        accounting_before,
                        accounting_before,
                        "review_required",
                        False,
                        [],
                    )
                    atomic_write_text(inventory_path, inventory_text)
                    atomic_write_text(
                        inventory_meta_path,
                        json.dumps(meta, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
                    )
                    atomic_write_text(report_path, report_text)
                    if completion_path.exists():
                        completion_path.unlink()
                    completion_status = "review_required"
                else:
                    validate_target_rows(rows, seeds, report)
                    if report.error_count:
                        raise RuntimeError("target review validation failed")
                    all_edges = [context.edge for context in direct_contexts]
                    emission_blocks = build_target_emission_blocks(rows, all_edges)
                    planned = plan_target_feedthrough_accounting(
                        rows, owned, port_records, upstream, runtime, report
                    )
                    accounting_after = runtime.accounting_digest(port_records)
                    for row in rows:
                        row.values["accounting_digest_after"] = accounting_after
                    if report.error_count:
                        raise RuntimeError("target accounting plan validation failed")
                    sdc_text = render_target_10_sdc(
                        emission_blocks,
                        run_context,
                        completeness,
                        structure,
                        accounting_before,
                        accounting_after,
                        manifest_path,
                    )
                    sdc_payload = sdc_text.encode("utf-8")
                    sdc_digest = hashlib.sha256(sdc_payload).hexdigest()
                    preview_dir = middle_root / (".10_preview_{0}".format(os.getpid()))
                    if preview_dir.exists():
                        shutil.rmtree(str(preview_dir))
                    preview_dir.mkdir(parents=True, exist_ok=False)
                    prepared_candidates: Dict[str, Path] = {}
                    workbook_before = {model.relative_name: model.digest_before for model in models}
                    for model in models:
                        if model.modified:
                            candidate = preview_dir / model.path.name
                            model.workbook.save(str(candidate))
                            model.digest_after = digest_file(candidate)
                            prepared_candidates[model.relative_name] = candidate
                        else:
                            model.digest_after = model.digest_before
                    workbook_after = {model.relative_name: model.digest_after for model in models}

                    transaction_id = make_target_10_transaction_id(
                        run_context.get("run_id", "")
                    )
                    old_delta_rows = (
                        runtime.load_delta_rows(delta_path, report)
                        if delta_path.is_file() else []
                    )
                    old_delta_meta = upstream.get("existing_10_meta", {})
                    new_delta_rows = build_target_10_delta_rows(
                        planned,
                        run_context,
                        transaction_id,
                        structure,
                        accounting_before,
                        accounting_after,
                        runtime,
                    )
                    if delta_path.is_file():
                        delta_text = delta_path.read_text(encoding="utf-8")
                        if delta_text and not delta_text.endswith("\n"):
                            delta_text += "\n"
                        if new_delta_rows:
                            delta_text += runtime.csv_text(
                                runtime.DELTA_HEADERS,
                                new_delta_rows,
                                include_header=False,
                            )
                    else:
                        delta_text = runtime.csv_text(
                            runtime.DELTA_HEADERS, old_delta_rows + new_delta_rows
                        )
                    delta_payload = delta_text.encode("utf-8")
                    transaction_entry = {
                        "transaction_id": transaction_id,
                        "committed_at": runtime.utc_timestamp(),
                        "structure_digest": structure,
                        "accounting_digest_before": accounting_before,
                        "accounting_digest_after": accounting_after,
                        "delta_rows_digest": runtime.delta_rows_digest(new_delta_rows),
                    }
                    delta_meta = {
                        "schema_version": TARGET_SCHEMA_VERSION,
                        "run_id": run_context.get("run_id", ""),
                        "mode_label": run_context.get("mode_label", ""),
                        "design_revision": run_context.get("design_revision", ""),
                        "stage_name": "10_feedthrough",
                        "completion_status": "complete",
                        "structure_digest": structure,
                        "accounting_digest_before": accounting_before,
                        "accounting_digest_after": accounting_after,
                        "workbook_file_digest_before": workbook_before,
                        "workbook_file_digest_after": workbook_after,
                        "delta_csv_digest": hashlib.sha256(delta_payload).hexdigest(),
                        "transactions": list(old_delta_meta.get("transactions", [])) + [transaction_entry],
                    }
                    inventory_text = target_inventory_text(rows)
                    inventory_payload = inventory_text.encode("utf-8")
                    inventory_digest = hashlib.sha256(inventory_payload).hexdigest()
                    inventory_meta = target_inventory_meta_payload(
                        run_context,
                        "complete",
                        structure,
                        accounting_before,
                        accounting_after,
                        inventory_path,
                        inventory_digest,
                        form_path,
                        output_path,
                        sdc_digest,
                        completeness,
                        workbook_before,
                        workbook_after,
                        upstream.get("upstream_digests", {}),
                    )
                    completion = {
                        "schema_version": TARGET_SCHEMA_VERSION,
                        "author": author_name(),
                        "stage_name": "10_feedthrough",
                        "stage": "",
                        "corner": "",
                        "run_id": run_context.get("run_id", ""),
                        "mode_label": run_context.get("mode_label", ""),
                        "design_revision": run_context.get("design_revision", ""),
                        "completion_status": "complete",
                        "error_count": 0,
                        "sync_changed": "no",
                        "structure_digest": structure,
                        "accounting_digest_before": accounting_before,
                        "accounting_digest_after": accounting_after,
                        "port_accounting": "enabled",
                        "added_bits": sum(1 for item in planned if item[5]),
                        "workbook_file_digest_before": workbook_before,
                        "workbook_file_digest_after": workbook_after,
                        "upstream_artifact_digests": upstream.get("upstream_digests", {}),
                        "output_sdc_digest": sdc_digest,
                        "feedthrough_inventory_digest": inventory_digest,
                        "accounting_delta_digest": hashlib.sha256(delta_payload).hexdigest(),
                        "review_workbook_digest": digest_file(form_path),
                        "transaction_id": transaction_id,
                        "run_completeness": completeness.status,
                        "missing_instances": completeness.missing_instances,
                    }
                    report_text = render_target_10_report(
                        report,
                        rows,
                        run_context,
                        completeness,
                        structure,
                        accounting_before,
                        accounting_after,
                        "complete",
                        False,
                        planned,
                    )
                    artifact_payloads = [
                        (output_path, sdc_payload),
                        (inventory_path, inventory_payload),
                        (
                            inventory_meta_path,
                            (json.dumps(inventory_meta, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8"),
                        ),
                        (delta_path, delta_payload),
                        (
                            delta_meta_path,
                            (json.dumps(delta_meta, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8"),
                        ),
                        (
                            completion_path,
                            (json.dumps(completion, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8"),
                        ),
                        (report_path, report_text.encode("utf-8")),
                    ]
                    for model in models:
                        if (
                            not model.path.is_file()
                            or digest_file(model.path) != model.digest_before
                        ):
                            report.error(
                                "concurrent workbook modification detected before 10 commit: {0}".format(
                                    model.path
                                )
                            )
                    if report.error_count:
                        raise RuntimeError(
                            "target workbook snapshot changed before transaction commit"
                        )
                    runtime.execute_transaction(
                        run_root,
                        models,
                        prepared_candidates,
                        artifact_payloads,
                        transaction_id,
                        run_context,
                        structure,
                        accounting_before,
                        accounting_after,
                        report,
                    )
                    completion_status = "complete"
    except Exception as exc:
        if not report.error_count:
            report.error(str(exc))
        if not diagnostic:
            invalidate_prior_target_completion(
                runtime,
                input_root,
                completion_path,
                prior_completion_digest,
                run_context,
                structure,
                accounting_before,
                models,
                upstream,
                form_path,
                report_path,
                completeness,
                report,
            )
        failure_text = render_target_10_report(
            report,
            rows,
            run_context,
            completeness,
            structure,
            accounting_before,
            accounting_after or accounting_before,
            "failed",
            diagnostic,
            planned,
        )
        atomic_write_text(report_path, failure_text)
        completion_status = "failed"
    finally:
        for model in models:
            try:
                model.workbook.close()
            except Exception:
                pass
        if preview_dir is not None and preview_dir.exists():
            shutil.rmtree(str(preview_dir))

    print("Author: {0}".format(author_name()))
    print("Run ID: {0}".format(run_context.get("run_id", "")))
    print("Mode label: {0}".format(run_context.get("mode_label", "")))
    print("Design revision: {0}".format(run_context.get("design_revision", "")))
    print(
        "Port accounting: {0}".format(
            "diagnostic/read-only" if diagnostic else (
                "enabled" if completion_status == "complete" else "not committed"
            )
        )
    )
    print("Completion status: {0}".format(completion_status))
    print("Report: {0}".format(report_path))
    print("Warnings: {0}  Errors: {1}".format(report.warning_count, report.error_count))
    if report.error_count:
        return 1
    if completion_status == "review_required":
        return 1
    return 0


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate one run's SoC feedthrough direct-edge SDC and accounting artifacts."
    )
    parser.add_argument("--run-root", help="single-run target runtime root")
    parser.add_argument("-scenario", "--scenario", default="common", choices=sorted(SCENARIOS))
    parser.add_argument("-stage", "--stage", default="all", choices=sorted(STAGES))
    parser.add_argument("-corner", "--corner", default="all")
    parser.add_argument("--tool", default="dc", choices=sorted(TOOLS))
    parser.add_argument("--connection-inventory")
    parser.add_argument("--harden-sdc-manifest")
    parser.add_argument("--require-complete-harden-sdc", action="store_true")
    parser.add_argument("--clock-inventory")
    parser.add_argument("--clock-inventory-meta")
    parser.add_argument("--relation-map")
    parser.add_argument("--relation-map-meta")
    parser.add_argument("--form")
    parser.add_argument("--inventory")
    parser.add_argument("--output")
    parser.add_argument("--report")
    parser.add_argument("--pending-root", help="directory containing scenario pending .ports files")
    parser.add_argument("--no-update-pending", action="store_true")
    parser.add_argument(
        "--diagnose-only",
        action="store_true",
        help="validate and publish diagnostic inventory/report without SDC or xlsx accounting writes",
    )
    return parser.parse_args(argv)


def cli_option_present(argv: Sequence[str], *names: str) -> bool:
    for token in argv:
        for name in names:
            if token == name or token.startswith(name + "="):
                return True
    return False


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)
    if args.run_root is None:
        print(
            "ERROR: --run-root is required by the target single-run runtime",
            file=sys.stderr,
        )
        return 2
    if args.run_root is not None:
        forbidden = []
        for label, names in (
            ("--scenario", ("--scenario", "-scenario")),
            ("--stage", ("--stage", "-stage")),
            ("--corner", ("--corner", "-corner")),
            ("--tool", ("--tool",)),
            ("--connection-inventory", ("--connection-inventory",)),
            ("--harden-sdc-manifest", ("--harden-sdc-manifest",)),
            ("--clock-inventory", ("--clock-inventory",)),
            ("--clock-inventory-meta", ("--clock-inventory-meta",)),
            ("--relation-map", ("--relation-map",)),
            ("--relation-map-meta", ("--relation-map-meta",)),
            ("--form", ("--form",)),
            ("--inventory", ("--inventory",)),
            ("--output", ("--output",)),
            ("--report", ("--report",)),
            ("--pending-root", ("--pending-root",)),
            ("--no-update-pending", ("--no-update-pending",)),
        ):
            if cli_option_present(argv, *names):
                forbidden.append(label)
        if forbidden:
            print(
                "ERROR: target single-run mode rejects scenario/view/path override(s): {0}".format(
                    ",".join(forbidden)
                ),
                file=sys.stderr,
            )
            return 2
        return run_target_single_run(args, argv)
    cwd = Path.cwd()
    target_layout = args.run_root is not None
    run_root = Path(args.run_root).expanduser().resolve() if target_layout else cwd
    report = Report()
    port_accounting = (
        "disabled by explicit option" if args.no_update_pending else "enabled"
    )

    if target_layout:
        target_overrides = [
            flag
            for flag, value in (
                ("--connection-inventory", args.connection_inventory),
                ("--harden-sdc-manifest", args.harden_sdc_manifest),
                ("--clock-inventory", args.clock_inventory),
                ("--clock-inventory-meta", args.clock_inventory_meta),
                ("--relation-map", args.relation_map),
                ("--relation-map-meta", args.relation_map_meta),
                ("--form", args.form),
                ("--inventory", args.inventory),
                ("--output", args.output),
                ("--report", args.report),
                ("--pending-root", args.pending_root),
            )
            if value is not None
        ]
        if target_overrides:
            report.error(
                "target mode uses fixed artifact paths; path override(s) are not allowed: {0}".format(
                    ",".join(target_overrides)
                )
            )

        connection_path = run_root / "00_middle" / "connection_inventory.csv"
        manifest_path = (
            run_root / "00_middle" / "scenario" / args.scenario / "harden_sdc_manifest.csv"
        )
        clock_path = (
            run_root / "01_middle" / "assembled" / args.scenario / "clock_inventory.csv"
        )
        clock_meta_path = (
            run_root / "01_middle" / "assembled" / args.scenario / "clock_inventory.meta"
        )
        relation_path = run_root / "03_middle" / "relation_map" / (args.scenario + ".csv")
        relation_meta_path = run_root / "03_middle" / "relation_map" / (args.scenario + ".meta")
        form_path = run_root / "10_middle" / "10_feedthrough.xlsx"
        inventory_path = (
            run_root / "10_middle" / "scenario" / args.scenario / "feedthrough_edge_inventory.csv"
        )
        output_path = output_sdc_path(
            run_root / "10_result", args.scenario, args.stage, args.corner
        )
        report_path = (
            run_root / "10_result" / "reports" /
            "feedthrough_check_report_{0}.txt".format(args.scenario)
        )
        pending_dir = run_root / "00_middle" / "scenario" / args.scenario / "pending"
        removed_log_path = run_root / "10_middle" / "scenario" / args.scenario / "removed_log" / "10_feedthrough.removed"
        previous_removed_paths = [
            run_root / "00_middle" / "scenario" / args.scenario / "removed_log" / "00_disposition.removed",
            run_root / "01_middle" / "scenario" / args.scenario / "removed_log" / "01_soc_clocks.removed",
            run_root / "04_middle" / "scenario" / args.scenario / "removed_log" / "04_soc_io_pads.removed",
        ]
    else:
        connection_path = resolve_path(
            cwd, args.connection_inventory, "00_harden_port_inventory/connection_inventory.csv"
        )
        manifest_path = resolve_path(
            cwd,
            args.harden_sdc_manifest,
            "00_harden_port_inventory/harden_sdc_manifest.csv",
        )
        clock_path = resolve_path(cwd, args.clock_inventory, "../01_soc_clocks/clock_inventory.csv")
        clock_meta_path = (
            resolve_path(cwd, args.clock_inventory_meta, args.clock_inventory_meta)
            if args.clock_inventory_meta
            else clock_path.with_suffix(".meta")
        )
        relation_path = resolve_path(cwd, args.relation_map, "../03_soc_clock_groups/relation_map_common.csv")
        relation_meta_path = (
            resolve_path(cwd, args.relation_map_meta, args.relation_map_meta)
            if args.relation_map_meta
            else relation_path.with_suffix(".meta")
        )
        form_path = resolve_path(cwd, args.form, "10_feedthrough.xlsx")
        inventory_path = resolve_path(cwd, args.inventory, "feedthrough_edge_inventory.csv")
        output_path = resolve_path(
            cwd, args.output, str(output_sdc_path(Path("."), args.scenario, args.stage, args.corner))
        )
        report_path = resolve_path(
            cwd, args.report, "feedthrough_check_report_{0}.txt".format(args.scenario)
        )
        pending_dir = resolve_path(
            cwd, args.pending_root, "00_harden_port_inventory/pending"
        )
        removed_log_path = cwd / "00_harden_port_inventory" / "removed_log" / "10_feedthrough.removed"
        previous_removed_root = cwd / "00_harden_port_inventory" / "removed_log"
        previous_removed_paths = [
            previous_removed_root / "00_disposition.removed",
            previous_removed_root / "01_soc_clocks.removed",
            previous_removed_root / "04_soc_io_pads.removed",
        ]

    report.info("resolved run root: {0}".format(run_root))
    report.info("resolved connection inventory: {0}".format(connection_path))
    report.info("resolved harden SDC manifest: {0}".format(manifest_path))
    report.info("resolved assembled clock inventory: {0}".format(clock_path))
    report.info("resolved assembled clock inventory meta: {0}".format(clock_meta_path))
    report.info("resolved relation map: {0}".format(relation_path))
    report.info("resolved relation map meta: {0}".format(relation_meta_path))
    report.info("resolved pending directory: {0}".format(pending_dir))
    report.info("resolved form: {0}".format(form_path))
    report.info("resolved inventory: {0}".format(inventory_path))
    report.info("resolved output: {0}".format(output_path))
    all_edges = read_connection_inventory(
        connection_path,
        report,
        args.scenario,
        require_target_schema=target_layout,
    )
    manifest, completeness = read_harden_sdc_manifest(
        manifest_path,
        run_root,
        args.scenario,
        args.require_complete_harden_sdc,
        report,
    )
    if target_layout and not args.no_update_pending:
        validate_pending_directory(pending_dir, manifest, report)

    owned = classify_feedthrough_edges(all_edges, manifest, report)

    clocks = read_clock_inventory(
        clock_path,
        report,
        require_active_action=target_layout,
        meta_path=clock_meta_path,
        expected_scenario=args.scenario,
    )
    relations = read_relation_map(
        relation_path,
        report,
        args.scenario,
        meta_path=relation_meta_path,
        clock_inventory=clocks,
    )

    evidence = extract_delay_evidence(manifest, report)
    seeds = build_edge_seeds(
        owned,
        manifest,
        completeness,
        evidence,
        clocks,
        relations,
        args.scenario,
        args.stage,
        args.corner,
        port_accounting,
        connection_path,
    )

    if report.error_count == 0:
        sync_workbook(form_path, seeds, evidence, report)
    rows = read_form_rows(form_path) if form_path.is_file() else []
    if rows and report.error_count == 0:
        validate_rows(rows, seeds, args.scenario, args.stage, args.corner, report)
    suppressed_scenario_rows: Set[int] = set()
    if rows and report.error_count == 0:
        suppressed_scenario_rows = validate_assembled_view(
            rows,
            seeds,
            args.scenario,
            args.stage,
            args.corner,
            args.tool,
            report,
        )
    emission_blocks = build_emission_blocks(
        rows,
        args.scenario,
        args.stage,
        args.corner,
        args.tool,
        all_edges,
        suppressed_scenario_rows,
    )

    generation_allowed = not report.sync_changed
    pending_plan = PendingPlan()
    if report.error_count == 0 and generation_allowed and not args.no_update_pending:
        pending_plan = prepare_pending_plan(
            pending_dir,
            removed_log_path,
            previous_removed_paths,
            all_edges,
            owned,
            rows,
            args.scenario,
            args.stage,
            args.corner,
            args.tool,
            report,
        )

    if report.error_count == 0:
        write_inventory_upsert(
            inventory_path, rows, args.scenario, args.stage, args.corner
        )
        report.info("wrote bit-level feedthrough edge inventory: {0}".format(inventory_path))

    generated = False
    if report.error_count == 0 and generation_allowed:
        sdc_text = "\n".join(
            generate_sdc(
                emission_blocks,
                args.scenario,
                args.stage,
                args.corner,
                args.tool,
                completeness,
                port_accounting,
                connection_path,
                manifest_path,
            )
        ).rstrip() + "\n"
        atomic_write_text(output_path, sdc_text)
        report.info("wrote direct-edge SDC: {0}".format(output_path))
        generated = True
        if not args.no_update_pending:
            apply_pending_plan(pending_plan, removed_log_path)
            if pending_plan.removed_count:
                report.info(
                    "removed {0} terminal direct-edge endpoint(s); log={1}".format(
                        pending_plan.removed_count, removed_log_path
                    )
                )
    elif report.sync_changed:
        report.warn("SDC/pending update skipped until synchronized workbook is reviewed")
    elif report.error_count:
        report.warn("formal SDC/pending update skipped because errors were reported")

    coverage = build_coverage_lines(
        rows, owned, args.scenario, args.stage, args.corner, args.tool,
        emission_blocks, generated,
    )
    write_report(
        report_path,
        report,
        args.scenario,
        args.stage,
        args.corner,
        args.tool,
        completeness,
        form_path,
        inventory_path,
        output_path,
        connection_path,
        manifest_path,
        port_accounting,
        coverage,
    )
    print("Author: {0}".format(author_name()))
    print("Scenario: {0}".format(args.scenario))
    print("Run completeness: {0}".format(completeness.status))
    print("Port accounting: {0}".format(port_accounting))
    print("Connection inventory: {0}".format(connection_path.resolve()))
    print("Harden SDC manifest: {0}".format(manifest_path.resolve()))
    print("Report: {0}".format(report_path))
    print(
        "Warnings: {0}  Errors: {1}  Sync changed: {2}".format(
            report.warning_count, report.error_count, report.sync_changed
        )
    )
    if report.error_count:
        return 1
    if report.sync_changed:
        return 1
    return 0 if generated else 1


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except RuntimeError as exc:
        print("ERROR: {0}".format(exc), file=sys.stderr)
        raise SystemExit(2)
