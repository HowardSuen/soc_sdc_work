#!/usr/bin/env python3
"""Regression for the flat, workbook-centric 10 feedthrough runtime."""

from __future__ import print_function

import csv
import hashlib
import json
import os
import re
import shutil
import socket
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


BASE = Path(__file__).resolve().parent
STAGE10 = BASE.parent
SCRIPTS = STAGE10.parent
EX10 = STAGE10 / "10_extract_feedthrough.py"
EX00 = SCRIPTS / "00_harden_port_inventory" / "00_harden_port_inventory.py"
WORK = BASE / "work_latest"

SCHEMA_VERSION = "1.0"
RUN_ID = "RUN_10_FEEDTHROUGH_REGRESSION"
MODE_LABEL = "func"
DESIGN_REVISION = "rev_ft_10"

PORT_COLUMNS = [
    "Input",
    "Input Width",
    "Input Used Width",
    "From Whom",
    "Output",
    "Output Width",
    "Output Used Width",
    "To Top",
    "Inout",
    "Inout Width",
    "Inout Connectivity",
    "Inout Name",
]

DELTA_HEADERS = [
    "schema_version",
    "run_id",
    "mode_label",
    "stage_name",
    "transaction_id",
    "view_id",
    "stage",
    "corner",
    "structure_digest",
    "accounting_digest_before",
    "accounting_digest_after",
    "workbook",
    "sheet",
    "row",
    "direction",
    "port",
    "legal_bits",
    "added_bits",
    "final_used_bits",
    "owner_object_id",
    "reason",
    "evidence_status",
]

DIRECTION_COLUMNS = {
    "input": ("Input", "Input Width", "Input Used Width", "From Whom"),
    "output": ("Output", "Output Width", "Output Used Width", "To Top"),
    "inout": ("Inout", "Inout Width", "Inout Name", "Inout Connectivity"),
}


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def clean_dir(path):
    if path.exists():
        shutil.rmtree(str(path))
    path.mkdir(parents=True)


def run_script(script, args, cwd):
    return subprocess.run(
        [sys.executable, str(script)] + [str(arg) for arg in args],
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )


def run_00(root):
    return run_script(EX00, ["--run-root", root], root)


def run_10(root, *extra):
    return run_script(EX10, ["--run-root", root] + list(extra), root)


def sha256_bytes(value):
    return hashlib.sha256(value).hexdigest()


def sha256_file(path):
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_bytes(value):
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")


def canonical_digest(value):
    return sha256_bytes(canonical_bytes(value))


def stable_edge_ids(src_inst, src_direction, src_base, src_bit,
                    dst_inst, dst_direction, dst_base, dst_bit):
    payload = [
        SCHEMA_VERSION,
        src_inst,
        src_direction,
        src_base,
        int(src_bit),
        dst_inst,
        dst_direction,
        dst_base,
        int(dst_bit),
    ]
    digest = canonical_digest(payload)
    return "CONN_" + digest, "FTE_" + digest


def write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
        encoding="utf-8",
    )


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        return list(csv.DictReader(file_obj))


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def workbook_headers(sheet):
    return {
        normalize_header(cell.value): cell.column
        for cell in sheet[1]
        if normalize_header(cell.value)
    }


def add_port_row(sheet, **values):
    row = [values.get(column, "") for column in PORT_COLUMNS]
    row.append(values.get("Audit Formula", ""))
    sheet.append(row)


def new_port_workbook(sheet_names):
    workbook = Workbook()
    for index, sheet_name in enumerate(sheet_names):
        if index == 0:
            sheet = workbook.active
            sheet.title = sheet_name
        else:
            sheet = workbook.create_sheet(sheet_name)
        sheet.append(PORT_COLUMNS + ["Audit Formula"])
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["M"].width = 18
    return workbook


def harden_sdc(inst_name):
    if inst_name == "u_src":
        return (
            "set_output_delay -max 2.0 -clock [get_clocks {clk_a}] [get_ports {data_o[1:0]}]\n"
            "set_output_delay -min -0.2 -clock [get_clocks {clk_a}] [get_ports {data_o[1:0]}]\n"
            "set_output_delay -max 3.0 -clock [get_clocks {clk_a}] [get_ports {normal_o}]\n"
            "create_clock -name clk_sys -period 10 [get_ports {clk_o}]\n"
        )
    if inst_name == "u_ft1":
        return (
            "set_input_delay -max 1.8 -clock [get_clocks {clk_a}] [get_ports {fti_local[5:4]}]\n"
            "set_output_delay -max 1.5 -clock [get_clocks {clk_a}] [get_ports {fto_local[7:6]}]\n"
            "set_input_delay -max 0.5 -clock [get_clocks {clk_sys}] [get_ports {fti_clk}]\n"
        )
    if inst_name == "u_ft2":
        return (
            "set_input_delay -max 1.3 -clock [get_clocks {clk_b}] [get_ports {fti_transport[11:10]}]\n"
            "set_output_delay -max 1.1 -clock [get_clocks {clk_b}] [get_ports {fto_transport[11:10]}]\n"
        )
    return (
        "set_input_delay -max 0.9 -clock [get_clocks {clk_b}] [get_ports {data_i[3:2]}]\n"
        "set_input_delay -max 2.5 -clock [get_clocks {clk_a}] [get_ports {normal_i}]\n"
    )


def write_inputs(root, missing_sdcs=None, feedthrough_classification=True):
    missing_sdcs = set(missing_sdcs or [])
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_csv(
        inputs / "run_context.csv",
        ["run_id", "mode_label", "design_revision", "note"],
        [{
            "run_id": RUN_ID,
            "mode_label": MODE_LABEL,
            "design_revision": DESIGN_REVISION,
            "note": "10 feedthrough latest-runtime regression",
        }],
    )
    write_csv(
        inputs / "required_views.csv",
        [
            "view_id", "stage", "corner", "require_02", "require_04",
            "require_20", "require_30", "note",
        ],
        [{
            "view_id": "synth_ss",
            "stage": "synth",
            "corner": "ss",
            "require_02": "no",
            "require_04": "yes",
            "require_20": "yes",
            "require_30": "yes",
            "note": "required regression view",
        }],
    )

    info = Workbook()
    sheet = info.active
    sheet.title = "integration"
    sheet.append([
        "module_name", "inst_name", "owner", "sdc_path", "sdc_status",
        "sdc_note", "harden_class", "feedthrough_classification",
    ])
    for inst_name in ("u_src", "u_ft1", "u_ft2", "u_dst"):
        is_missing = inst_name in missing_sdcs
        is_feedthrough = inst_name in ("u_ft1", "u_ft2")
        sheet.append([
            "mod_" + inst_name[2:],
            inst_name,
            "owner_" + inst_name[2:],
            "" if is_missing else "%s.sdc" % inst_name,
            "missing" if is_missing else "available",
            "not delivered" if is_missing else "",
            "feedthrough" if is_feedthrough and feedthrough_classification else "functional",
            "confirmed" if is_feedthrough and feedthrough_classification else "",
        ])
        if not is_missing:
            (inputs / (inst_name + ".sdc")).write_text(
                harden_sdc(inst_name), encoding="utf-8"
            )
    info.save(str(inputs / "info_all.xlsx"))

    source = new_port_workbook(["u_src", "u_ft1"])
    sheet = source["u_src"]
    add_port_row(
        sheet,
        Output="data_o[1:0]",
        **{"Output Width": 2, "Audit Formula": "=1+1"}
    )
    add_port_row(sheet, Output="normal_o", **{"Output Width": 1})
    add_port_row(sheet, Output="clk_o", **{"Output Width": 1})
    sheet["E2"].fill = PatternFill(fill_type="solid", fgColor="00FFFF00")
    sheet["E2"].comment = Comment("preserve this comment", "regression")
    sheet = source["u_ft1"]
    add_port_row(
        sheet,
        Input="fti_local[5:4]",
        **{"Input Width": 2, "From Whom": "u_src.data_o[1:0]"}
    )
    add_port_row(
        sheet,
        Input="fti_clk",
        **{"Input Width": 1, "From Whom": "u_src.clk_o"}
    )
    add_port_row(
        sheet,
        Input="fti_pad",
        **{"Input Width": 1, "From Whom": "top.pad_in"}
    )
    add_port_row(sheet, Output="fto_local[7:6]", **{"Output Width": 2})
    source.save(str(inputs / "port_source.xlsx"))

    sink = new_port_workbook(["u_ft2", "u_dst"])
    sheet = sink["u_ft2"]
    add_port_row(
        sheet,
        Input="fti_transport[11:10]",
        **{"Input Width": 2, "From Whom": "u_ft1.fto_local[7:6]"}
    )
    add_port_row(sheet, Output="fto_transport[11:10]", **{"Output Width": 2})
    sheet = sink["u_dst"]
    add_port_row(
        sheet,
        Input="data_i[3:2]",
        **{"Input Width": 2, "From Whom": "u_ft2.fto_transport[11:10]"}
    )
    add_port_row(
        sheet,
        Input="normal_i",
        **{"Input Width": 1, "From Whom": "u_src.normal_o"}
    )
    add_port_row(sheet, Input="nc_i", **{"Input Width": 1, "From Whom": "NC"})
    sink.save(str(inputs / "port_sink.xlsx"))


def used_state(root, sheet_name, direction, port_name):
    port_column, _, used_column, _ = DIRECTION_COLUMNS[direction]
    for path in sorted((root / "inputs").glob("port_*.xlsx")):
        workbook = load_workbook(str(path), data_only=False)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            continue
        sheet = workbook[sheet_name]
        headers = workbook_headers(sheet)
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row_idx, headers[normalize_header(port_column)]).value or "") == port_name:
                value = sheet.cell(row_idx, headers[normalize_header(used_column)]).value
                workbook.close()
                return "" if value is None else str(value)
        workbook.close()
    raise AssertionError("port row not found: %s/%s/%s" % (sheet_name, direction, port_name))


def set_used_state(root, sheet_name, direction, port_name, value):
    port_column, _, used_column, _ = DIRECTION_COLUMNS[direction]
    for path in sorted((root / "inputs").glob("port_*.xlsx")):
        workbook = load_workbook(str(path), data_only=False)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            continue
        sheet = workbook[sheet_name]
        headers = workbook_headers(sheet)
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row_idx, headers[normalize_header(port_column)]).value or "") == port_name:
                sheet.cell(row_idx, headers[normalize_header(used_column)], value)
                workbook.save(str(path))
                workbook.close()
                return path, row_idx
        workbook.close()
    raise AssertionError("port row not found: %s/%s/%s" % (sheet_name, direction, port_name))


def port_record(root, sheet_name, direction, port_name):
    port_column, width_column, used_column, _ = DIRECTION_COLUMNS[direction]
    for path in sorted((root / "inputs").glob("port_*.xlsx")):
        workbook = load_workbook(str(path), data_only=False)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            continue
        sheet = workbook[sheet_name]
        headers = workbook_headers(sheet)
        for row_idx in range(2, sheet.max_row + 1):
            raw = str(sheet.cell(row_idx, headers[normalize_header(port_column)]).value or "")
            if raw == port_name:
                width = int(sheet.cell(row_idx, headers[normalize_header(width_column)]).value)
                used = sheet.cell(row_idx, headers[normalize_header(used_column)]).value
                workbook.close()
                return {
                    "workbook": str(path.resolve().relative_to(root.resolve())),
                    "sheet": sheet_name,
                    "row": row_idx,
                    "direction": direction,
                    "port": raw,
                    "width": width,
                    "used": "" if used is None else str(used),
                }
        workbook.close()
    raise AssertionError("port row not found: %s/%s/%s" % (sheet_name, direction, port_name))


def accounting_digest(root):
    rows = []
    for path in sorted((root / "inputs").glob("port_*.xlsx")):
        workbook = load_workbook(str(path), data_only=False)
        relative = str(path.resolve().relative_to(root.resolve()))
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers = workbook_headers(sheet)
            for row_idx in range(2, sheet.max_row + 1):
                for direction in ("input", "output", "inout"):
                    port_column, _, used_column, _ = DIRECTION_COLUMNS[direction]
                    port_key = normalize_header(port_column)
                    used_key = normalize_header(used_column)
                    if port_key not in headers or used_key not in headers:
                        continue
                    port = str(sheet.cell(row_idx, headers[port_key]).value or "").strip()
                    if not port:
                        continue
                    raw_used = str(sheet.cell(row_idx, headers[used_key]).value or "").strip()
                    bits = []
                    if raw_used and re.fullmatch(r"\d+(?:\s*,\s*\d+)*", raw_used):
                        bits = sorted(set(int(item.strip()) for item in raw_used.split(",")))
                    rows.append([relative, sheet_name, row_idx, direction, port, bits])
        workbook.close()
    rows.sort(key=canonical_bytes)
    return canonical_digest([SCHEMA_VERSION, rows])


def legal_bits(port, width):
    match = re.fullmatch(r"(.+)\[(\d+)\s*:\s*(\d+)\]", port)
    if match:
        left = int(match.group(2))
        right = int(match.group(3))
        return sorted(range(min(left, right), max(left, right) + 1))
    match = re.fullmatch(r"(.+)\[(\d+)\]", port)
    if match:
        return [int(match.group(2))]
    return [0] if width == 1 else list(range(width))


def format_bits(bits):
    return ",".join(str(bit) for bit in sorted(bits))


def workbook_digests(root):
    return {
        str(path.resolve().relative_to(root.resolve())): sha256_file(path)
        for path in sorted((root / "inputs").glob("port_*.xlsx"))
    }


def make_delta_row(root, stage_name, transaction_id, structure,
                   accounting_before, accounting_after, sheet_name,
                   direction, port_name, added_bits, owner_id, reason):
    record = port_record(root, sheet_name, direction, port_name)
    final_used = []
    if record["used"]:
        final_used = [int(item) for item in record["used"].split(",")]
    return {
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "stage_name": stage_name,
        "transaction_id": transaction_id,
        "view_id": "",
        "stage": "",
        "corner": "",
        "structure_digest": structure,
        "accounting_digest_before": accounting_before,
        "accounting_digest_after": accounting_after,
        "workbook": record["workbook"],
        "sheet": record["sheet"],
        "row": record["row"],
        "direction": direction,
        "port": record["port"],
        "legal_bits": format_bits(legal_bits(record["port"], record["width"])),
        "added_bits": format_bits(added_bits),
        "final_used_bits": format_bits(final_used),
        "owner_object_id": owner_id,
        "reason": reason,
        "evidence_status": "complete",
    }


def delta_rows_digest(rows):
    normalized = [
        [str(row.get(header, "") or "").strip() for header in DELTA_HEADERS]
        for row in rows
    ]
    return canonical_digest([SCHEMA_VERSION, normalized])


def publish_delta(root, middle, stage_name, transaction_id, structure,
                  accounting_before, accounting_after, rows,
                  workbook_before, workbook_after, committed_at):
    delta_path = root / middle / "port_accounting_delta.csv"
    write_csv(delta_path, DELTA_HEADERS, rows)
    delta_digest = sha256_file(delta_path)
    meta = {
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "design_revision": DESIGN_REVISION,
        "stage_name": stage_name,
        "completion_status": "complete",
        "structure_digest": structure,
        "accounting_digest_before": accounting_before,
        "accounting_digest_after": accounting_after,
        "workbook_file_digest_before": workbook_before,
        "workbook_file_digest_after": workbook_after,
        "delta_csv_digest": delta_digest,
        "transactions": [{
            "transaction_id": transaction_id,
            "structure_digest": structure,
            "accounting_digest_before": accounting_before,
            "accounting_digest_after": accounting_after,
            "delta_rows_digest": delta_rows_digest(rows),
            "committed_at": committed_at,
        }],
    }
    write_json(root / middle / "port_accounting_delta.meta", meta)
    return delta_path, delta_digest


def publish_upstream_artifacts(root):
    snapshot = read_json(root / "00_middle" / "input_snapshot.meta")
    structure = snapshot["structure_digest"]
    accounting_00 = snapshot["accounting_digest_after"]
    meta_00 = read_json(root / "00_middle" / "port_accounting_delta.meta")
    committed_at = meta_00["transactions"][-1]["committed_at"]
    require(accounting_00 == accounting_digest(root), "fixture accounting digest differs from 00")

    before_01_files = workbook_digests(root)
    set_used_state(root, "u_src", "output", "clk_o", "0")
    set_used_state(root, "u_ft1", "input", "fti_clk", "0")
    accounting_01 = accounting_digest(root)
    after_01_files = workbook_digests(root)
    clk_id = "CLK_" + canonical_digest([SCHEMA_VERSION, "u_src", "output", "clk_o", 0])
    rows_01 = [
        make_delta_row(
            root, "01_soc_clocks", "01_fixture_txn", structure,
            accounting_00, accounting_01, "u_src", "output", "clk_o",
            [0], clk_id, "soc_visible_clock",
        ),
        make_delta_row(
            root, "01_soc_clocks", "01_fixture_txn", structure,
            accounting_00, accounting_01, "u_ft1", "input", "fti_clk",
            [0], clk_id, "soc_visible_clock",
        ),
    ]
    delta_01, delta_01_digest = publish_delta(
        root, "01_middle", "01_soc_clocks", "01_fixture_txn", structure,
        accounting_00, accounting_01, rows_01, before_01_files, after_01_files,
        committed_at,
    )
    clock_sdc = root / "01_result" / "01_soc_clocks.sdc"
    clock_sdc.parent.mkdir(parents=True, exist_ok=True)
    clock_sdc.write_text("# fixture SoC clocks\n", encoding="utf-8")
    clock_rows = [
        {
            "schema_version": SCHEMA_VERSION,
            "run_id": RUN_ID,
            "mode_label": MODE_LABEL,
            "design_revision": DESIGN_REVISION,
            "clock_record_id": clk_id,
            "inst_name": "u_src",
            "module_name": "mod_src",
            "port_name": "clk_o",
            "direction": "output",
            "clock_name": "clk_sys",
            "clock_kind": "primary",
            "period": "10",
            "final_action": "emit_output_clock",
            "target_object": "u_src/clk_o",
            "structure_digest": structure,
            "accounting_digest_before": accounting_00,
            "accounting_digest_after": accounting_01,
            "accounting_action": "accounted",
            "accounting_added_bits": "0",
        },
        {
            "schema_version": SCHEMA_VERSION,
            "run_id": RUN_ID,
            "mode_label": MODE_LABEL,
            "design_revision": DESIGN_REVISION,
            "clock_record_id": clk_id,
            "inst_name": "u_ft1",
            "module_name": "mod_ft1",
            "port_name": "fti_clk",
            "direction": "input",
            "clock_name": "clk_sys",
            "clock_kind": "propagated",
            "period": "10",
            "final_action": "check_only",
            "target_object": "u_ft1/fti_clk",
            "structure_digest": structure,
            "accounting_digest_before": accounting_00,
            "accounting_digest_after": accounting_01,
            "accounting_action": "accounted",
            "accounting_added_bits": "0",
        },
        {
            "schema_version": SCHEMA_VERSION,
            "run_id": RUN_ID,
            "mode_label": MODE_LABEL,
            "design_revision": DESIGN_REVISION,
            "clock_record_id": "CLK_ALIAS_SRC_A",
            "inst_name": "u_src",
            "module_name": "mod_src",
            "port_name": "",
            "direction": "input",
            "clock_name": "clk_a",
            "original_clock_name": "clk_a",
            "clock_kind": "alias",
            "final_action": "alias_only",
            "structure_digest": structure,
            "accounting_digest_before": accounting_00,
            "accounting_digest_after": accounting_01,
        },
        {
            "schema_version": SCHEMA_VERSION,
            "run_id": RUN_ID,
            "mode_label": MODE_LABEL,
            "design_revision": DESIGN_REVISION,
            "clock_record_id": "CLK_ALIAS_FT1_A",
            "inst_name": "u_ft1",
            "module_name": "mod_ft1",
            "port_name": "",
            "direction": "input",
            "clock_name": "clk_a",
            "original_clock_name": "clk_a",
            "clock_kind": "alias",
            "final_action": "alias_only",
            "structure_digest": structure,
            "accounting_digest_before": accounting_00,
            "accounting_digest_after": accounting_01,
        },
        {
            "schema_version": SCHEMA_VERSION,
            "run_id": RUN_ID,
            "mode_label": MODE_LABEL,
            "design_revision": DESIGN_REVISION,
            "clock_record_id": "CLK_ALIAS_FT2_B",
            "inst_name": "u_ft2",
            "module_name": "mod_ft2",
            "port_name": "",
            "direction": "input",
            "clock_name": "clk_b",
            "original_clock_name": "clk_b",
            "clock_kind": "alias",
            "final_action": "alias_only",
            "structure_digest": structure,
            "accounting_digest_before": accounting_00,
            "accounting_digest_after": accounting_01,
        },
        {
            "schema_version": SCHEMA_VERSION,
            "run_id": RUN_ID,
            "mode_label": MODE_LABEL,
            "design_revision": DESIGN_REVISION,
            "clock_record_id": "CLK_ALIAS_DST_B",
            "inst_name": "u_dst",
            "module_name": "mod_dst",
            "port_name": "",
            "direction": "input",
            "clock_name": "clk_b",
            "original_clock_name": "clk_b",
            "clock_kind": "alias",
            "final_action": "alias_only",
            "structure_digest": structure,
            "accounting_digest_before": accounting_00,
            "accounting_digest_after": accounting_01,
        },
    ]
    clock_fields = sorted(set(key for row in clock_rows for key in row))
    clock_inventory = root / "01_middle" / "clock_inventory.csv"
    write_csv(clock_inventory, clock_fields, clock_rows)
    clock_meta = {
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "design_revision": DESIGN_REVISION,
        "stage_name": "01_soc_clocks",
        "completion_status": "complete",
        "structure_digest": structure,
        "accounting_digest_before": accounting_00,
        "accounting_digest_after": accounting_01,
        "inventory_path": str(clock_inventory.resolve()),
        "inventory_digest": sha256_file(clock_inventory),
        "clock_inventory_digest": sha256_file(clock_inventory),
        "final_sdc_path": str(clock_sdc.resolve()),
        "final_sdc_digest": sha256_file(clock_sdc),
        "clock_set_digest": canonical_digest(["clk_a", "clk_b", "clk_sys"]),
    }
    write_json(root / "01_middle" / "clock_inventory.meta", clock_meta)
    completion_01 = {
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "design_revision": DESIGN_REVISION,
        "stage_name": "01_soc_clocks",
        "stage": "",
        "corner": "",
        "completion_status": "complete",
        "error_count": 0,
        "sync_changed": "no",
        "structure_digest": structure,
        "accounting_digest_before": accounting_00,
        "accounting_digest_after": accounting_01,
        "upstream_artifact_digests": {
            "00_stage_completion": sha256_file(root / "00_middle" / "stage_completion.meta"),
            "00_port_accounting_delta": sha256_file(root / "00_middle" / "port_accounting_delta.csv"),
        },
        "output_sdc_digest": sha256_file(clock_sdc),
        "clock_inventory_digest": sha256_file(clock_inventory),
        "accounting_delta_digest": delta_01_digest,
        "transaction_id": "01_fixture_txn",
    }
    write_json(root / "01_middle" / "stage_completion.meta", completion_01)

    relation_rows = [
        {
            "schema_version": SCHEMA_VERSION,
            "clock_a": "clk_a",
            "clock_b": "clk_b",
            "relation_type": "asynchronous",
            "relation_source": "explicit_rule",
            "source_rule_ids": "CG_ASYNC_A_B",
            "clock_universe_digest": clock_meta["clock_set_digest"],
            "assembled_view_digest": canonical_digest(["clk_a", "clk_b", "asynchronous"]),
        }
    ]
    relation_path = root / "03_middle" / "relation_map.csv"
    write_csv(relation_path, list(relation_rows[0]), relation_rows)
    relation_meta = {
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "design_revision": DESIGN_REVISION,
        "stage_name": "03_soc_clock_groups",
        "completion_status": "complete",
        "structure_digest": structure,
        "relation_map_path": str(relation_path.resolve()),
        "relation_map_digest": sha256_file(relation_path),
        "upstream_01_inventory_digest": sha256_file(clock_inventory),
        "upstream_01_completion_digest": sha256_file(root / "01_middle" / "stage_completion.meta"),
    }
    write_json(root / "03_middle" / "relation_map.meta", relation_meta)
    clock_group_sdc = root / "03_result" / "03_soc_clock_groups.sdc"
    clock_group_sdc.parent.mkdir(parents=True, exist_ok=True)
    clock_group_sdc.write_text("# fixture asynchronous clk_a/clk_b\n", encoding="utf-8")
    completion_03 = {
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "design_revision": DESIGN_REVISION,
        "stage_name": "03_soc_clock_groups",
        "stage": "",
        "corner": "",
        "completion_status": "complete",
        "error_count": 0,
        "sync_changed": "no",
        "structure_digest": structure,
        "accounting_digest_before": "not_applicable",
        "accounting_digest_after": "not_applicable",
        "port_accounting": "not_applicable",
        "added_bits": 0,
        "upstream_artifact_digests": {
            "01_clock_inventory": sha256_file(clock_inventory),
            "01_stage_completion": sha256_file(root / "01_middle" / "stage_completion.meta"),
        },
        "output_sdc_digest": sha256_file(clock_group_sdc),
        "relation_map_digest": sha256_file(relation_path),
        "accounting_delta_digest": "not_applicable",
    }
    write_json(root / "03_middle" / "stage_completion.meta", completion_03)

    before_04_files = workbook_digests(root)
    set_used_state(root, "u_ft1", "input", "fti_pad", "0")
    accounting_04 = accounting_digest(root)
    after_04_files = workbook_digests(root)
    pad_digest = canonical_digest([
        SCHEMA_VERSION, "pad_in", 0, "input", "u_ft1", "input", "fti_pad", 0,
    ])
    pad_id = "PAD_" + pad_digest
    rows_04 = [
        make_delta_row(
            root, "04_soc_io_pads", "04_fixture_txn", structure,
            accounting_01, accounting_04, "u_ft1", "input", "fti_pad",
            [0], pad_id, "soc_top_pad",
        )
    ]
    delta_04, delta_04_digest = publish_delta(
        root, "04_middle", "04_soc_io_pads", "04_fixture_txn", structure,
        accounting_01, accounting_04, rows_04, before_04_files, after_04_files,
        committed_at,
    )
    pad_rows = [{
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "design_revision": DESIGN_REVISION,
        "pad_id": pad_id,
        "view_id": "synth_ss",
        "stage": "synth",
        "corner": "ss",
        "top_port": "pad_in",
        "direction": "input",
        "source_workbook": "inputs/port_source.xlsx",
        "source_sheet": "u_ft1",
        "source_row": "4",
        "harden_instance": "u_ft1",
        "harden_direction": "input",
        "harden_port": "fti_pad",
        "harden_bit_index": "0",
        "connection_status": "matched",
        "sdc_status": "available",
        "coverage_status": "covered",
        "pad_disposition": "constrained",
        "apply": "yes",
        "review_status": "approved",
        "structure_digest": structure,
        "note": "fixture pad owner",
    }]
    pad_inventory = root / "04_middle" / "pad_inventory.csv"
    write_csv(pad_inventory, list(pad_rows[0]), pad_rows)
    pad_meta = {
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "design_revision": DESIGN_REVISION,
        "stage_name": "04_soc_io_pads",
        "completion_status": "complete",
        "structure_digest": structure,
        "accounting_digest_before": accounting_01,
        "accounting_digest_after": accounting_04,
        "pad_inventory_path": str(pad_inventory.resolve()),
        "pad_inventory_digest": sha256_file(pad_inventory),
    }
    write_json(root / "04_middle" / "pad_inventory.meta", pad_meta)
    pad_sdc = root / "04_result" / "04_soc_io_pads_synth_ss.sdc"
    pad_sdc.parent.mkdir(parents=True, exist_ok=True)
    pad_sdc.write_text("# fixture pad timing\n", encoding="utf-8")
    view_completion = {
        "schema_version": SCHEMA_VERSION,
        "run_id": RUN_ID,
        "mode_label": MODE_LABEL,
        "design_revision": DESIGN_REVISION,
        "stage_name": "04_soc_io_pads",
        "view_id": "synth_ss",
        "stage": "synth",
        "corner": "ss",
        "completion_status": "complete",
        "error_count": 0,
        "sync_changed": "no",
        "structure_digest": structure,
        "accounting_digest_before": accounting_01,
        "accounting_digest_after": accounting_04,
        "output_sdc_digest": sha256_file(pad_sdc),
        "pad_inventory_digest": sha256_file(pad_inventory),
        "accounting_delta_digest": delta_04_digest,
    }
    view_completion_path = root / "04_middle" / "completion" / "synth_ss.meta"
    write_json(view_completion_path, view_completion)
    completion_04 = dict(view_completion)
    completion_04.update({
        "view_id": "",
        "stage": "",
        "corner": "",
        "required_view_completions": {
            "synth_ss": sha256_file(view_completion_path),
        },
        "upstream_artifact_digests": {
            "03_stage_completion": sha256_file(root / "03_middle" / "stage_completion.meta"),
            "04_pad_inventory": sha256_file(pad_inventory),
        },
        "transaction_id": "04_fixture_txn",
    })
    write_json(root / "04_middle" / "stage_completion.meta", completion_04)
    return {
        "structure_digest": structure,
        "accounting_00": accounting_00,
        "accounting_01": accounting_01,
        "accounting_04": accounting_04,
        "clock_id": clk_id,
        "pad_id": pad_id,
        "delta_01": delta_01,
        "delta_04": delta_04,
    }


def build_root(root, missing_sdcs=None, feedthrough_classification=True):
    clean_dir(root)
    write_inputs(
        root,
        missing_sdcs=missing_sdcs,
        feedthrough_classification=feedthrough_classification,
    )
    initialized = run_00(root)
    require(
        initialized.returncode == 0,
        "00 fixture initialization failed:\n%s\n%s" % (initialized.stdout, initialized.stderr),
    )
    context = publish_upstream_artifacts(root)
    require(
        context["accounting_04"] == accounting_digest(root),
        "fixture 04 accounting digest does not match workbooks",
    )
    return context


def report_text(root):
    path = root / "10_result" / "reports" / "feedthrough_check_report.txt"
    require(path.is_file(), "10 report missing: %s" % path)
    return path.read_text(encoding="utf-8")


def port_bytes(root):
    return {
        path.name: path.read_bytes()
        for path in sorted((root / "inputs").glob("port_*.xlsx"))
    }


def inventory_rows(root):
    return read_csv(root / "10_middle" / "feedthrough_edge_inventory.csv")


def edge_key(row):
    return (
        row.get("src_instance", ""),
        int(row.get("src_bit_index", "0")),
        row.get("dst_instance", ""),
        int(row.get("dst_bit_index", "0")),
    )


def review_sheet(form):
    workbook = load_workbook(str(form), data_only=False)
    require("feedthrough_edges" in workbook.sheetnames, "review workbook feedthrough_edges sheet missing")
    sheet = workbook["feedthrough_edges"]
    return workbook, sheet, workbook_headers(sheet)


def header_column(headers, *names):
    for name in names:
        key = normalize_header(name)
        if key in headers:
            return headers[key]
    return None


def set_review_value(sheet, headers, row_idx, value, *names):
    column = header_column(headers, *names)
    if column is not None:
        sheet.cell(row_idx, column, value)
        return True
    return False


def approve_edges(form, specifications):
    workbook, sheet, headers = review_sheet(form)
    required_identity = {
        "src_instance": header_column(headers, "src_instance"),
        "src_bit_index": header_column(headers, "src_bit_index"),
        "dst_instance": header_column(headers, "dst_instance"),
        "dst_bit_index": header_column(headers, "dst_bit_index"),
    }
    require(all(required_identity.values()), "review workbook exact endpoint identity columns missing")
    found = set()
    for row_idx in range(2, sheet.max_row + 1):
        key = (
            str(sheet.cell(row_idx, required_identity["src_instance"]).value or ""),
            int(sheet.cell(row_idx, required_identity["src_bit_index"]).value or 0),
            str(sheet.cell(row_idx, required_identity["dst_instance"]).value or ""),
            int(sheet.cell(row_idx, required_identity["dst_bit_index"]).value or 0),
        )
        values = specifications.get(key)
        if values is None:
            continue
        found.add(key)
        aliases = {
            "basis": ("basis", "disposition_basis"),
            "reviewer": ("reviewer",),
            "review_date": ("review_date",),
            "sdc_independent_basis": ("sdc_independent_basis",),
            "relationship_override_basis": ("relationship_override_basis",),
            "min_sign_review": ("min_sign_review",),
        }
        for name, value in values.items():
            candidates = aliases.get(name, (name,))
            written = set_review_value(sheet, headers, row_idx, value, *candidates)
            if name in (
                "channel_disposition", "apply", "review_status", "owner", "basis",
            ):
                require(written, "review workbook missing required review column %s" % name)
    workbook.save(str(form))
    workbook.close()
    require(found == set(specifications), "review edge(s) not found: %s" % sorted(set(specifications) - found))


def approved(disposition, value="", override="", independent=""):
    result = {
        "channel_disposition": disposition,
        "budget_model": "manual_budget" if disposition == "emit_budget" else "",
        "budget_required": "yes" if disposition == "emit_budget" else "no",
        "converted_max": value,
        "emit_max": "yes" if disposition == "emit_budget" else "no",
        "emit_min": "no",
        "datapath_only": "yes",
        "tool_surface": "dc",
        "apply": "yes",
        "review_status": "approved",
        "owner": "soc_timing_owner",
        "reviewer": "soc_timing_reviewer",
        "review_date": "2026-07-16",
        "basis": "reviewed exact direct-edge disposition",
        "sdc_independent_basis": independent,
        "relationship_override_basis": override,
        "min_sign_review": "not_emitted",
    }
    return result


def delay_commands(text):
    return [
        line.strip()
        for line in text.splitlines()
        if line.strip().startswith(("set_max_delay ", "set_min_delay "))
    ]


def assert_no_obsolete_layout(root):
    require(not (root / "00_middle" / "connection_inventory.csv").exists(), "obsolete connection inventory exists")
    require(not list(root.glob("**/pending")), "obsolete pending directory exists")
    require(not list(root.glob("**/removed_log")), "obsolete removed-log directory exists")
    require(not (root / "10_middle" / "scenario").exists(), "obsolete 10 scenario directory exists")
    require(not (root / "10_result" / "common").exists(), "obsolete 10 common result directory exists")


EXPECTED_EDGE_KEYS = {
    ("u_src", 1, "u_ft1", 5),
    ("u_src", 0, "u_ft1", 4),
    ("u_ft1", 7, "u_ft2", 11),
    ("u_ft1", 6, "u_ft2", 10),
    ("u_ft2", 11, "u_dst", 3),
    ("u_ft2", 10, "u_dst", 2),
}


def expected_ids(key):
    return stable_edge_ids(
        key[0], "output", {
            "u_src": "data_o",
            "u_ft1": "fto_local",
            "u_ft2": "fto_transport",
        }[key[0]], key[1],
        key[2], "input", {
            "u_ft1": "fti_local",
            "u_ft2": "fti_transport",
            "u_dst": "data_i",
        }[key[2]], key[3],
    )


def run_lifecycle_contract():
    root = WORK / "lifecycle"
    context = build_root(root)
    before_ports = port_bytes(root)
    first = run_10(root)
    require(
        first.returncode != 0,
        "first synchronization must require review:\n%s\n%s" % (first.stdout, first.stderr),
    )
    form = root / "10_middle" / "10_feedthrough.xlsx"
    inventory = root / "10_middle" / "feedthrough_edge_inventory.csv"
    report = root / "10_result" / "reports" / "feedthrough_check_report.txt"
    require(form.is_file(), "first synchronization did not create review workbook")
    require(inventory.is_file(), "first synchronization did not publish structural inventory")
    require(report.is_file(), "first synchronization did not publish report")
    require(port_bytes(root) == before_ports, "first synchronization modified port workbooks")
    require(not (root / "10_result" / "10_feedthrough.sdc").exists(), "review sync published formal SDC")
    require(not (root / "10_middle" / "port_accounting_delta.csv").exists(), "review sync published accounting delta")
    require(not (root / "10_middle" / "stage_completion.meta").exists(), "review sync published completion")
    assert_no_obsolete_layout(root)

    rows = inventory_rows(root)
    by_key = {edge_key(row): row for row in rows}
    require(set(by_key) == EXPECTED_EDGE_KEYS, "workbook direct-edge mapping is wrong: %s" % sorted(by_key))
    for key, row in by_key.items():
        connection_id, feedthrough_id = expected_ids(key)
        require(row.get("schema_version") == SCHEMA_VERSION, "inventory schema_version mismatch")
        require(row.get("connection_id") == connection_id, "canonical connection_id mismatch for %s" % (key,))
        require(row.get("feedthrough_edge_id") == feedthrough_id, "canonical feedthrough_edge_id mismatch for %s" % (key,))
        require(re.fullmatch(r"CONN_[0-9a-f]{64}", connection_id), "connection ID is not full SHA-256")
        require(re.fullmatch(r"FTE_[0-9a-f]{64}", feedthrough_id), "feedthrough ID is not full SHA-256")
        require(row.get("run_id") == RUN_ID, "inventory run_id missing")
        require(row.get("mode_label") == MODE_LABEL, "inventory mode_label missing")
        require(row.get("design_revision") == DESIGN_REVISION, "inventory design_revision missing")
        require(row.get("structure_digest") == context["structure_digest"], "inventory structure digest mismatch")
        require(row.get("source_workbook"), "inventory source workbook missing")
        require(row.get("source_sheet"), "inventory source sheet missing")
        require(row.get("source_row"), "inventory source row missing")
    inventory_text = inventory.read_text(encoding="utf-8")
    for excluded in ("fti_clk", "fti_pad", "normal_i", "normal_o", "nc_i"):
        require(excluded not in inventory_text, "excluded clock/pad/normal/NC edge leaked into 10: %s" % excluded)

    specifications = {
        ("u_src", 1, "u_ft1", 5): approved("emit_budget", "1.2"),
        ("u_src", 0, "u_ft1", 4): approved("route_to_30"),
        ("u_ft1", 7, "u_ft2", 11): approved("no_soc_budget_required"),
        # u_ft1[6] -> u_ft2[10] intentionally remains pending.
        ("u_ft2", 11, "u_dst", 3): approved("not_applicable"),
        ("u_ft2", 10, "u_dst", 2): approved("emit_budget", "0.8"),
    }
    approve_edges(form, specifications)
    second = run_10(root)
    require(
        second.returncode == 0,
        "approved lifecycle failed:\n%s\n%s" % (second.stdout, second.stderr),
    )

    output = root / "10_result" / "10_feedthrough.sdc"
    require(output.is_file(), "approved lifecycle did not publish flat SDC")
    sdc = output.read_text(encoding="utf-8")
    commands = delay_commands(sdc)
    require(len(commands) == 2, "unexpected emitted delay command count: %s" % commands)
    require("-through" not in sdc, "feedthrough SDC crossed harden internal logic with -through")
    require(
        any("u_src/data_o[1]" in line and "u_ft1/fti_local[5]" in line for line in commands),
        "ingress exact direct-edge command missing",
    )
    require(
        any("u_ft2/fto_transport[10]" in line and "u_dst/data_i[2]" in line for line in commands),
        "egress exact direct-edge command missing",
    )
    require(
        not any("u_src/data_o" in line and "u_dst/data_i" in line for line in commands),
        "synthetic end-to-end feedthrough chain was emitted",
    )
    for marker in (RUN_ID, MODE_LABEL, DESIGN_REVISION):
        require(marker in sdc, "SDC provenance missing: %s" % marker)

    require(used_state(root, "u_src", "output", "data_o[1:0]") == "1", "emit source bit was not accounted")
    require(used_state(root, "u_ft1", "input", "fti_local[5:4]") == "5", "emit destination bit was not accounted")
    require(used_state(root, "u_ft1", "output", "fto_local[7:6]") == "7", "no-budget source bit was not accounted")
    require(used_state(root, "u_ft2", "input", "fti_transport[11:10]") == "11", "no-budget destination bit was not accounted")
    require(used_state(root, "u_ft2", "output", "fto_transport[11:10]") == "10,11", "egress terminal bits were not unioned")
    require(used_state(root, "u_dst", "input", "data_i[3:2]") == "2,3", "egress destination bits were not unioned")
    require("0" not in used_state(root, "u_src", "output", "data_o[1:0]"), "route_to_30 source bit was accounted")
    require("4" not in used_state(root, "u_ft1", "input", "fti_local[5:4]"), "route_to_30 destination bit was accounted")
    require("6" not in used_state(root, "u_ft1", "output", "fto_local[7:6]"), "pending source bit was accounted")
    require("10" not in used_state(root, "u_ft2", "input", "fti_transport[11:10]"), "pending destination bit was accounted")
    require(used_state(root, "u_src", "output", "clk_o") == "0", "01 owner state was removed")
    require(used_state(root, "u_ft1", "input", "fti_pad") == "0", "04 owner state was removed")
    require(used_state(root, "u_dst", "input", "nc_i") == "0", "00 structural state was removed")

    delta_path = root / "10_middle" / "port_accounting_delta.csv"
    delta_meta_path = root / "10_middle" / "port_accounting_delta.meta"
    completion_path = root / "10_middle" / "stage_completion.meta"
    inventory_meta_path = root / "10_middle" / "feedthrough_edge_inventory.meta"
    for path in (delta_path, delta_meta_path, completion_path, inventory_meta_path):
        require(path.is_file(), "approved lifecycle artifact missing: %s" % path)
    delta_rows = read_csv(delta_path)
    require(delta_rows, "10 accounting delta is empty")
    require(
        all(not row["owner_object_id"] or re.fullmatch(r"FTE_[0-9a-f]{64}", row["owner_object_id"])
            for row in delta_rows),
        "10 delta owner_object_id is not a feedthrough edge ID",
    )
    require(
        all(row["structure_digest"] == context["structure_digest"] for row in delta_rows),
        "10 delta structure digest mismatch",
    )
    completion = read_json(completion_path)
    require(completion.get("completion_status") == "complete", "10 completion is not complete")
    require(completion.get("error_count") == 0 and completion.get("sync_changed") == "no", "10 completion gates are invalid")
    require(completion.get("run_id") == RUN_ID and completion.get("mode_label") == MODE_LABEL, "10 completion provenance missing")
    require(completion.get("structure_digest") == context["structure_digest"], "10 completion structure digest mismatch")
    require(completion.get("accounting_digest_before") == context["accounting_04"], "10 completion accounting chain is discontinuous")
    require(completion.get("accounting_digest_after") == accounting_digest(root), "10 completion after digest mismatch")
    require(completion.get("output_sdc_digest") == sha256_file(output), "10 completion SDC digest mismatch")
    require(completion.get("accounting_delta_digest") == sha256_file(delta_path), "10 completion delta digest mismatch")
    inventory_meta = read_json(inventory_meta_path)
    require(
        inventory_meta.get("inventory_digest", inventory_meta.get("feedthrough_edge_inventory_digest"))
        == sha256_file(root / "10_middle" / "feedthrough_edge_inventory.csv"),
        "10 inventory meta digest mismatch",
    )
    text = report_text(root)
    require("added" in text.lower() and "final" in text.lower(), "report lacks added/final Used-bit evidence")
    assert_no_obsolete_layout(root)

    styled = load_workbook(str(root / "inputs" / "port_source.xlsx"), data_only=False)
    require(styled["u_src"]["M2"].value == "=1+1", "accounting write lost formula")
    require(styled["u_src"]["E2"].fill.fgColor.rgb.endswith("FFFF00"), "accounting write lost style")
    require(styled["u_src"]["E2"].comment is not None, "accounting write lost comment")
    require(styled["u_src"].column_dimensions["A"].width == 24, "accounting write lost column width")
    styled.close()
    transaction_root = root / ".accounting_txn"
    require(not transaction_root.exists() or not any(transaction_root.iterdir()), "committed transaction scratch was not cleaned")

    meta_before = read_json(delta_meta_path)
    accounting_before_rerun = accounting_digest(root)
    rerun = run_10(root)
    require(rerun.returncode == 0, "idempotent 10 rerun failed")
    meta_after = read_json(delta_meta_path)
    require(
        len(meta_after.get("transactions", [])) == len(meta_before.get("transactions", [])) + 1,
        "idempotent rerun did not append a transaction",
    )
    require(meta_after.get("accounting_digest_before") == meta_after.get("accounting_digest_after"), "idempotent rerun changed accounting state")
    require(accounting_digest(root) == accounting_before_rerun, "idempotent rerun changed Used state")

    accounting_before_failure = accounting_digest(root)
    ports_before_failure = port_bytes(root)
    require(
        read_json(completion_path).get("completion_status") == "complete",
        "failure-rerun precondition lacks complete marker",
    )
    invalid = approved("emit_budget", "1.2")
    invalid["owner"] = ""
    approve_edges(form, {("u_src", 1, "u_ft1", 5): invalid})
    failed_rerun = run_10(root)
    require(failed_rerun.returncode != 0, "invalid rerun after success was accepted")
    require(port_bytes(root) == ports_before_failure, "failed rerun modified port accounting")
    require(accounting_digest(root) == accounting_before_failure, "failed rerun changed accounting digest")
    require(completion_path.is_file(), "failed rerun removed completion without publishing failure state")
    failed_completion = read_json(completion_path)
    require(
        failed_completion.get("completion_status") == "failed",
        "failed rerun left an active complete marker",
    )
    require(
        int(failed_completion.get("error_count") or 0) > 0,
        "failed completion marker has no error count",
    )
    require(
        failed_completion.get("run_id") == RUN_ID
        and failed_completion.get("structure_digest") == context["structure_digest"],
        "failed completion marker provenance is invalid",
    )
    require(
        failed_completion.get("accounting_digest_before") == accounting_before_failure
        and failed_completion.get("accounting_digest_after") == accounting_before_failure,
        "failed completion marker does not preserve the committed accounting state",
    )


def run_cli_and_diagnostic_contract():
    missing_root = run_script(EX10, [], BASE)
    require(missing_root.returncode != 0, "target runtime accepted invocation without --run-root")
    require(
        "run-root" in (missing_root.stdout + missing_root.stderr).lower(),
        "missing --run-root diagnostic absent",
    )

    rejected_root = WORK / "cli_reject_scenario"
    build_root(rejected_root)
    rejected = run_10(rejected_root, "--scenario", "common")
    require(rejected.returncode != 0, "target runtime accepted obsolete --scenario")
    require("scenario" in (rejected.stdout + rejected.stderr).lower(), "obsolete scenario diagnostic missing")
    require(not (rejected_root / "10_middle" / "10_feedthrough.xlsx").exists(), "rejected CLI mutated target artifacts")

    root = WORK / "diagnose_only"
    build_root(root)
    before = port_bytes(root)
    result = run_10(root, "--diagnose-only")
    require(
        result.returncode == 0,
        "diagnose-only failed:\n%s\n%s" % (result.stdout, result.stderr),
    )
    require(port_bytes(root) == before, "diagnose-only modified port workbooks")
    require(not (root / "10_result" / "10_feedthrough.sdc").exists(), "diagnose-only published formal SDC")
    require(not (root / "10_middle" / "port_accounting_delta.csv").exists(), "diagnose-only published delta")
    require(not (root / "10_middle" / "stage_completion.meta").exists(), "diagnose-only published completion")
    report = report_text(root).lower()
    require("port accounting: diagnostic/read-only" in report, "diagnose-only accounting metadata missing")
    require("accounting closure: not evaluated" in report, "diagnose-only closure metadata missing")


def run_upstream_gate_contract():
    missing_view_root = WORK / "upstream_missing_04_view"
    build_root(missing_view_root)
    (missing_view_root / "04_middle" / "completion" / "synth_ss.meta").unlink()
    missing = run_10(missing_view_root)
    require(missing.returncode != 0, "missing required 04 view completion was accepted")
    require("04" in report_text(missing_view_root), "missing 04 view diagnostic absent")
    require(not (missing_view_root / "10_result" / "10_feedthrough.sdc").exists(), "missing 04 view wrote formal SDC")

    stale_root = WORK / "upstream_accounting_stale"
    build_root(stale_root)
    completion_path = stale_root / "04_middle" / "stage_completion.meta"
    completion = read_json(completion_path)
    completion["accounting_digest_after"] = "0" * 64
    write_json(completion_path, completion)
    stale = run_10(stale_root)
    require(stale.returncode != 0, "stale upstream accounting digest was accepted")
    stale_report = report_text(stale_root).lower()
    require("accounting" in stale_report and ("digest" in stale_report or "chain" in stale_report), "accounting-chain diagnostic absent")
    require(not (stale_root / "10_result" / "10_feedthrough.sdc").exists(), "stale chain wrote formal SDC")

    sync_root = WORK / "upstream_sync_changed"
    build_root(sync_root)
    view_path = sync_root / "04_middle" / "completion" / "synth_ss.meta"
    view_completion = read_json(view_path)
    view_completion["sync_changed"] = "yes"
    write_json(view_path, view_completion)
    run_completion_path = sync_root / "04_middle" / "stage_completion.meta"
    run_completion = read_json(run_completion_path)
    run_completion["required_view_completions"]["synth_ss"] = sha256_file(view_path)
    write_json(run_completion_path, run_completion)
    sync_blocked = run_10(sync_root)
    require(sync_blocked.returncode != 0, "sync_changed=yes upstream completion was accepted")
    sync_report = report_text(sync_root).lower()
    require(
        "sync_changed" in sync_report and "must be no" in sync_report,
        "sync_changed upstream diagnostic absent",
    )
    require(
        not (sync_root / "10_result" / "10_feedthrough.sdc").exists(),
        "sync_changed upstream emitted formal SDC",
    )

    parent_root = WORK / "upstream_04_parent_digest"
    build_root(parent_root)
    child_path = parent_root / "04_middle" / "completion" / "synth_ss.meta"
    child_completion = read_json(child_path)
    child_completion["note"] = "changed after run-wide completion"
    write_json(child_path, child_completion)
    parent_blocked = run_10(parent_root)
    require(parent_blocked.returncode != 0, "stale 04 parent/child completion link was accepted")
    parent_report = report_text(parent_root).lower()
    require(
        "run-wide 04 required-view digest is stale" in parent_report,
        "04 parent/child digest diagnostic absent",
    )
    require(
        not (parent_root / "10_result" / "10_feedthrough.sdc").exists(),
        "stale 04 parent/child completion emitted formal 10 SDC",
    )

    digest_root = WORK / "upstream_04_sdc_digest"
    build_root(digest_root)
    pad_sdc = digest_root / "04_result" / "04_soc_io_pads_synth_ss.sdc"
    pad_sdc.write_text(
        pad_sdc.read_text(encoding="utf-8") + "# tampered after completion\n",
        encoding="utf-8",
    )
    digest_blocked = run_10(digest_root)
    require(digest_blocked.returncode != 0, "mutated required 04 SDC was accepted")
    digest_report = report_text(digest_root).lower()
    require(
        "output_sdc_digest" in digest_report
        and "does not match" in digest_report
        and "04_soc_io_pads_synth_ss.sdc" in digest_report,
        "required 04 SDC digest diagnostic absent",
    )
    require(
        not (digest_root / "10_result" / "10_feedthrough.sdc").exists(),
        "stale required 04 SDC emitted formal 10 SDC",
    )


def run_review_gate_contract():
    root = WORK / "terminal_owner_gate"
    build_root(root)
    first = run_10(root)
    require(first.returncode != 0, "review gate first sync did not stop")
    form = root / "10_middle" / "10_feedthrough.xlsx"
    values = approved("emit_budget", "1.0")
    values["owner"] = ""
    approve_edges(form, {("u_src", 1, "u_ft1", 5): values})
    result = run_10(root)
    require(result.returncode != 0, "terminal row without owner was accepted")
    report = report_text(root).lower()
    require("owner" in report and ("terminal" in report or "approved" in report), "terminal-owner diagnostic missing")
    require(not (root / "10_result" / "10_feedthrough.sdc").exists(), "invalid terminal review wrote SDC")


def run_async_override_contract():
    root = WORK / "async_override"
    build_root(root)
    first = run_10(root)
    require(first.returncode != 0, "async override first sync did not stop")
    form = root / "10_middle" / "10_feedthrough.xlsx"
    key = ("u_ft1", 7, "u_ft2", 11)
    approve_edges(form, {key: approved("emit_budget", "1.0")})
    blocked = run_10(root)
    require(blocked.returncode != 0, "async normal budget without override basis was accepted")
    report = report_text(root).lower()
    require("async" in report and "override" in report, "async override diagnostic missing")
    require(not (root / "10_result" / "10_feedthrough.sdc").exists(), "async override failure wrote SDC")

    approve_edges(
        form,
        {key: approved(
            "emit_budget",
            "1.0",
            override="architecture review permits an explicit physical datapath budget",
        )},
    )
    allowed = run_10(root)
    require(
        allowed.returncode == 0,
        "approved async override failed:\n%s\n%s" % (allowed.stdout, allowed.stderr),
    )
    sdc = (root / "10_result" / "10_feedthrough.sdc").read_text(encoding="utf-8")
    require("u_ft1/fto_local[7]" in sdc and "u_ft2/fti_transport[11]" in sdc, "async direct-edge command missing")


def run_missing_sdc_contract():
    root = WORK / "missing_sdc_basis"
    build_root(root, missing_sdcs={"u_dst"})
    first = run_10(root)
    require(first.returncode != 0, "partial-SDC first sync did not stop")
    rows = {edge_key(row): row for row in inventory_rows(root)}
    key = ("u_ft2", 10, "u_dst", 2)
    require(key in rows, "missing-SDC edge disappeared from inventory")
    require(
        rows[key].get("evidence_status") == "incomplete_missing_sdc"
        or rows[key].get("dst_sdc_status") == "missing",
        "missing-SDC evidence status absent",
    )
    form = root / "10_middle" / "10_feedthrough.xlsx"
    approve_edges(form, {key: approved("emit_budget", "0.7")})
    blocked = run_10(root)
    require(blocked.returncode != 0, "missing-SDC emit without independent basis was accepted")
    report = report_text(root).lower()
    require("missing" in report and ("independent" in report or "basis" in report), "missing-SDC basis diagnostic absent")

    approve_edges(
        form,
        {key: approved(
            "emit_budget",
            "0.7",
            independent="integration owner supplied a reviewed SDC-independent interconnect budget",
        )},
    )
    allowed = run_10(root)
    require(
        allowed.returncode == 0,
        "SDC-independent emit failed:\n%s\n%s" % (allowed.stdout, allowed.stderr),
    )
    require("u_dst/data_i[2]" in (root / "10_result" / "10_feedthrough.sdc").read_text(encoding="utf-8"), "independent emit command missing")

    strict_root = WORK / "missing_sdc_strict"
    build_root(strict_root, missing_sdcs={"u_dst"})
    strict = run_10(strict_root, "--require-complete-harden-sdc")
    require(strict.returncode != 0, "strict missing-SDC run was accepted")
    strict_report = report_text(strict_root).lower()
    require("missing" in strict_report and ("strict" in strict_report or "require-complete" in strict_report), "strict missing-SDC diagnostic absent")
    require(not (strict_root / "10_result" / "10_feedthrough.sdc").exists(), "strict failure wrote formal SDC")


def run_lock_and_final_token_contract():
    root = WORK / "accounting_lock"
    build_root(root)
    first = run_10(root)
    require(first.returncode != 0, "lock case first sync did not stop")
    form = root / "10_middle" / "10_feedthrough.xlsx"
    approve_edges(form, {("u_src", 1, "u_ft1", 5): approved("emit_budget", "1.0")})
    before = port_bytes(root)
    lock = root / "inputs" / ".port_accounting.lock"
    lock.write_text(
        json.dumps({
            "pid": os.getpid(),
            "host": socket.gethostname(),
            "stage": "regression_lock_holder",
        }) + "\n",
        encoding="utf-8",
    )
    blocked = run_10(root)
    require(blocked.returncode != 0, "active accounting lock was ignored")
    require(port_bytes(root) == before, "lock failure modified port workbooks")
    lock.unlink()
    allowed = run_10(root)
    require(allowed.returncode == 0, "run after lock release failed")

    final_root = WORK / "final_token_block"
    build_root(final_root)
    set_used_state(final_root, "u_src", "output", "data_o[1:0]", "ALL USED")
    final = run_10(final_root)
    require(final.returncode != 0, "10 accepted a 30 final accounting token")
    final_report = report_text(final_root).lower()
    require("all used" in final_report or "final accounting" in final_report, "final-token diagnostic missing")
    require(not (final_root / "10_result" / "10_feedthrough.sdc").exists(), "final-token failure wrote SDC")


def run_bus_compaction_contract():
    root = WORK / "bus_compaction"
    build_root(root)
    first = run_10(root)
    require(first.returncode != 0, "bus compaction first sync did not stop")
    form = root / "10_middle" / "10_feedthrough.xlsx"
    keys = {
        ("u_src", 1, "u_ft1", 5): approved("emit_budget", "1.1"),
        ("u_src", 0, "u_ft1", 4): approved("emit_budget", "1.1"),
    }
    approve_edges(form, keys)
    compact = run_10(root)
    require(
        compact.returncode == 0,
        "homogeneous bus budget failed:\n%s\n%s" % (compact.stdout, compact.stderr),
    )
    output = root / "10_result" / "10_feedthrough.sdc"
    commands = delay_commands(output.read_text(encoding="utf-8"))
    require(len(commands) == 1, "homogeneous bus bits were not compacted exactly once: %s" % commands)
    command = commands[0]
    require("*" not in command, "bus compaction used a wildcard collection")
    require(
        ("data_o[1:0]" in command or ("data_o[1]" in command and "data_o[0]" in command))
        and ("fti_local[5:4]" in command or ("fti_local[5]" in command and "fti_local[4]" in command)),
        "compacted command does not exactly identify both intended buses",
    )
    require(used_state(root, "u_src", "output", "data_o[1:0]") == "0,1", "bus accounting did not union exact source bits")
    require(used_state(root, "u_ft1", "input", "fti_local[5:4]") == "4,5", "bus accounting did not union exact destination bits")
    require(len(inventory_rows(root)) == 6, "command compaction changed bit-level inventory")

    split_values = {
        ("u_src", 1, "u_ft1", 5): approved("emit_budget", "1.1"),
        ("u_src", 0, "u_ft1", 4): approved("emit_budget", "1.3"),
    }
    approve_edges(form, split_values)
    split = run_10(root)
    require(split.returncode == 0, "heterogeneous bus fallback failed")
    commands = delay_commands(output.read_text(encoding="utf-8"))
    require(len(commands) == 2, "different per-bit budgets were incorrectly compacted")


def main():
    clean_dir(WORK)
    run_lifecycle_contract()
    run_cli_and_diagnostic_contract()
    run_upstream_gate_contract()
    run_review_gate_contract()
    run_async_override_contract()
    run_missing_sdc_contract()
    run_lock_and_final_token_contract()
    run_bus_compaction_contract()
    print("10 feedthrough latest-runtime regression: PASS")
    print(
        "  cases: lifecycle, cli_diagnostic, upstream_gate, review_gate, "
        "async_override, missing_sdc, lock_final_token, bus_compaction"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
