#!/usr/bin/env python3
"""Run a target-layout 01 -> 02 -> 03 -> 04 -> 10 integration case."""

from __future__ import print_function

import csv
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from build_case import RUN_ROOT, build_case
from fill_reviews import fill_02, fill_03, fill_04


BASE = Path(__file__).resolve().parent
SCRIPTS = BASE.parent.parent
EX01 = SCRIPTS / "01_soc_clocks/01_extract_soc_clocks.py"
EX02 = SCRIPTS / "02_soc_clock_timing/02_extract_soc_clock_timing.py"
EX03 = SCRIPTS / "03_soc_clock_groups/03_extract_soc_clock_groups.py"
EX04 = SCRIPTS / "04_soc_io_pads/04_extract_soc_io_pads.py"
EX10 = SCRIPTS / "10_feedthrough/10_extract_feedthrough.py"

INGRESS_ID = "CONN_DATA_OUT__FTI_PAYLOAD"
EGRESS_ID = "CONN_FTO_PAYLOAD__DATA_IN"


def require(condition, message):
    if not condition:
        raise RuntimeError(message)


def run(command):
    print("Running: %s" % " ".join(str(item) for item in command), flush=True)
    return subprocess.run([str(item) for item in command], cwd=str(BASE))


def run_ok(command, stage):
    result = run(command)
    require(result.returncode == 0, "%s failed with exit code %s" % (stage, result.returncode))


def run_review_gate(command, stage):
    result = run(command)
    require(result.returncode == 1, "%s first run should stop for workbook review" % stage)


def append_port_row(sheet, values):
    columns = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }
    row_idx = sheet.max_row + 1
    for name, value in values.items():
        sheet.cell(row_idx, columns[name], value)


def extend_case_for_10():
    port_path = RUN_ROOT / "inputs/port_demo.xlsx"
    workbook = load_workbook(str(port_path))
    append_port_row(
        workbook["u_harden_a"],
        {"Output": "data_o[0]", "Output Width": 1},
    )
    append_port_row(
        workbook["u_harden_a"],
        {
            "Input": "data_i[0]",
            "Input Width": 1,
            "From Whom": "u_harden_b.fto_payload[3]",
        },
    )
    append_port_row(
        workbook["u_harden_b"],
        {
            "Input": "fti_payload[3]",
            "Input Width": 1,
            "From Whom": "u_harden_a.data_o[0]",
        },
    )
    append_port_row(
        workbook["u_harden_b"],
        {"Output": "fto_payload[3]", "Output Width": 1},
    )
    workbook.save(str(port_path))

    harden_a = RUN_ROOT / "inputs/harden_a.sdc"
    harden_a.write_text(
        harden_a.read_text(encoding="utf-8")
        + "set_output_delay -clock [get_clocks pll_clk] -max 1.20 [get_ports {data_o[0]}]\n"
        + "set_input_delay -clock [get_clocks pll_clk] -max 1.10 [get_ports {data_i[0]}]\n",
        encoding="utf-8",
    )
    harden_b = RUN_ROOT / "inputs/harden_b.sdc"
    harden_b.write_text(
        harden_b.read_text(encoding="utf-8")
        + "set_input_delay -clock [get_clocks b_clk_in] -max 1.00 [get_ports {fti_payload[3]}]\n"
        + "set_output_delay -clock [get_clocks b_clk_in] -max 0.90 [get_ports {fto_payload[3]}]\n",
        encoding="utf-8",
    )

    connection_path = RUN_ROOT / "00_middle/connection_inventory.csv"
    with connection_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = list(reader.fieldnames or [])
        rows = list(reader)
    rows.extend(
        [
            {
                "schema_version": "1",
                "connection_id": INGRESS_ID,
                "scenario_scope": "common",
                "connection_type": "feedthrough_candidate",
                "src_instance": "u_harden_a",
                "src_direction": "output",
                "src_port": "data_o[0]",
                "src_bit_index": "0",
                "src_endpoint_key": "u_harden_a:output:data_o[0]",
                "src_soc_object": "u_harden_a/data_o[0]",
                "dst_instance": "u_harden_b",
                "dst_direction": "input",
                "dst_port": "fti_payload[3]",
                "dst_bit_index": "3",
                "dst_endpoint_key": "u_harden_b:input:fti_payload[3]",
                "dst_soc_object": "u_harden_b/fti_payload[3]",
                "fanout_index": "0",
                "range_source_expr": "data_o[0]",
                "range_sink_expr": "fti_payload[3]",
                "bit_pair_order": "explicit_map",
                "source_workbook": "demo",
                "source_sheet": "connections",
                "source_row": str(len(rows) + 2),
                "validation_status": "matched",
                "owner_hint": "",
                "note": "01-to-10 ingress feedthrough demo edge",
            },
            {
                "schema_version": "1",
                "connection_id": EGRESS_ID,
                "scenario_scope": "common",
                "connection_type": "feedthrough_candidate",
                "src_instance": "u_harden_b",
                "src_direction": "output",
                "src_port": "fto_payload[3]",
                "src_bit_index": "3",
                "src_endpoint_key": "u_harden_b:output:fto_payload[3]",
                "src_soc_object": "u_harden_b/fto_payload[3]",
                "dst_instance": "u_harden_a",
                "dst_direction": "input",
                "dst_port": "data_i[0]",
                "dst_bit_index": "0",
                "dst_endpoint_key": "u_harden_a:input:data_i[0]",
                "dst_soc_object": "u_harden_a/data_i[0]",
                "fanout_index": "0",
                "range_source_expr": "fto_payload[3]",
                "range_sink_expr": "data_i[0]",
                "bit_pair_order": "explicit_map",
                "source_workbook": "demo",
                "source_sheet": "connections",
                "source_row": str(len(rows) + 3),
                "validation_status": "matched",
                "owner_hint": "",
                "note": "01-to-10 egress feedthrough demo edge",
            },
        ]
    )
    with connection_path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    pending = RUN_ROOT / "00_middle/scenario/common/pending"
    pending_a = pending / "u_harden_a.ports"
    pending_a.write_text(
        pending_a.read_text(encoding="utf-8") + "output data_o[0]\ninput data_i[0]\n",
        encoding="utf-8",
    )
    pending_b = pending / "u_harden_b.ports"
    pending_b.write_text(
        pending_b.read_text(encoding="utf-8")
        + "input fti_payload[3]\noutput fto_payload[3]\n",
        encoding="utf-8",
    )


def set_values(sheet, row_idx, columns, values):
    for name, value in values.items():
        if name in columns:
            sheet.cell(row_idx, columns[name], value)


def fill_10():
    path = RUN_ROOT / "10_middle/10_feedthrough.xlsx"
    workbook = load_workbook(str(path))
    sheet = workbook["feedthrough_edges"]
    columns = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }
    approved = set()
    for row_idx in range(2, sheet.max_row + 1):
        connection_id = sheet.cell(row_idx, columns["connection_id"]).value
        if connection_id == INGRESS_ID:
            values = {
                "channel_disposition": "emit_budget",
                "budget_model": "manual_budget",
                "budget_required": "yes",
                "converted_max": 0.80,
                "converted_min": 0.05,
                "emit_max": "yes",
                "emit_min": "yes",
                "min_sign_review": "approved",
                "datapath_only": "yes",
                "tool_surface": "dc",
                "apply": "yes",
                "review_status": "approved",
                "owner": "soc_demo",
                "reviewer": "sta_demo",
                "review_date": "2026-07-13",
                "disposition_basis": "reviewed ingress direct-edge interconnect budget",
                "relationship_override_basis": "integration owner confirmed the direct-edge clock context",
            }
        elif connection_id == EGRESS_ID:
            values = {
                "channel_disposition": "emit_budget",
                "budget_model": "manual_budget",
                "budget_required": "yes",
                "converted_max": 0.90,
                "emit_max": "yes",
                "emit_min": "no",
                "datapath_only": "yes",
                "tool_surface": "dc",
                "apply": "yes",
                "review_status": "approved",
                "owner": "soc_demo",
                "reviewer": "sta_demo",
                "review_date": "2026-07-13",
                "disposition_basis": "reviewed egress direct-edge interconnect budget",
                "relationship_override_basis": "integration owner confirmed the direct-edge clock context",
            }
        else:
            continue
        set_values(sheet, row_idx, columns, values)
        approved.add(connection_id)
    require(approved == {INGRESS_ID, EGRESS_ID}, "10 expected feedthrough edges are missing")
    workbook.save(str(path))
    print("10 review filled: %s" % path, flush=True)


def read_csv_rows(path):
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        return list(csv.DictReader(file_obj))


def nonempty_lines(path):
    if not path.is_file():
        return []
    return [line for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_assembled_preview(paths):
    path = RUN_ROOT / "assembled/common_prects_ss_125_through_10.sdc"
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["# 01 -> 10 assembled source preview", ""]
    for item in paths:
        lines.append("source %s" % item.relative_to(RUN_ROOT).as_posix())
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def main():
    build_case()
    extend_case_for_10()

    run_ok([sys.executable, EX01, "--run-root", RUN_ROOT, "--scenario", "common"], "01")

    command02 = [
        sys.executable, EX02, "--run-root", RUN_ROOT,
        "--scenario", "common", "--stage", "prects", "--corner", "ss_125",
    ]
    run_review_gate(command02, "02")
    fill_02()
    run_ok(command02, "02")

    command03 = [sys.executable, EX03, "--run-root", RUN_ROOT, "--scenario", "common"]
    run_review_gate(command03, "03")
    fill_03()
    run_ok(command03, "03")

    command04 = [sys.executable, EX04, "--run-root", RUN_ROOT, "--scenario", "common"]
    run_review_gate(command04, "04")
    fill_04()
    run_ok(command04, "04")

    command10 = [sys.executable, EX10, "--run-root", RUN_ROOT, "--scenario", "common"]
    run_review_gate(command10, "10")
    fill_10()
    run_ok(command10, "10")

    clock_sdc = RUN_ROOT / "01_result/common/01_soc_clocks.sdc"
    timing_sdc = RUN_ROOT / "02_result/common/02_soc_clock_timing_prects_ss_125.sdc"
    group_sdc = RUN_ROOT / "03_result/common/03_soc_clock_groups.sdc"
    io_sdc = RUN_ROOT / "04_result/common/04_soc_io_pads.sdc"
    feedthrough_sdc = RUN_ROOT / "10_result/common/10_feedthrough.sdc"
    inventory_path = RUN_ROOT / "10_middle/scenario/common/feedthrough_edge_inventory.csv"
    removed_path = RUN_ROOT / "10_middle/scenario/common/removed_log/10_feedthrough.removed"
    final_paths = [clock_sdc, timing_sdc, group_sdc, io_sdc, feedthrough_sdc]
    for path in final_paths + [inventory_path, removed_path]:
        require(path.is_file(), "expected artifact missing: %s" % path)

    inventory = {row["connection_id"]: row for row in read_csv_rows(inventory_path)}
    require(set(inventory) == {INGRESS_ID, EGRESS_ID}, "10 inventory ownership set is wrong")
    require(inventory[INGRESS_ID]["feedthrough_side"] == "ingress", "ingress classification missing")
    require(inventory[EGRESS_ID]["feedthrough_side"] == "egress", "egress classification missing")

    feedthrough_text = feedthrough_sdc.read_text(encoding="utf-8")
    require(
        "set_max_delay 0.8 -datapath_only -from [get_pins {u_harden_a/data_o[0]}] "
        "-to [get_pins {u_harden_b/fti_payload[3]}]" in feedthrough_text,
        "10 ingress max delay missing",
    )
    require(
        "set_min_delay 0.05 -datapath_only -from [get_pins {u_harden_a/data_o[0]}] "
        "-to [get_pins {u_harden_b/fti_payload[3]}]" in feedthrough_text,
        "10 ingress min delay missing",
    )
    require(
        "set_max_delay 0.9 -datapath_only -from [get_pins {u_harden_b/fto_payload[3]}] "
        "-to [get_pins {u_harden_a/data_i[0]}]" in feedthrough_text,
        "10 egress max delay missing",
    )
    require("-from [get_pins {u_harden_b/fti_payload[3]}]" not in feedthrough_text, "internal fti->fto command emitted")
    require(
        "-from [get_pins {u_harden_a/data_o[0]}] -to [get_pins {u_harden_a/data_i[0]}]"
        not in feedthrough_text,
        "synthetic end-to-end command emitted",
    )

    removed_text = removed_path.read_text(encoding="utf-8")
    require("connection_id=" + INGRESS_ID in removed_text, "ingress removed-log evidence missing")
    require("connection_id=" + EGRESS_ID in removed_text, "egress removed-log evidence missing")
    pending = RUN_ROOT / "00_middle/scenario/common/pending"
    require(not nonempty_lines(pending / "u_harden_a.ports"), "u_harden_a pending is not empty")
    require(not nonempty_lines(pending / "u_harden_b.ports"), "u_harden_b pending is not empty")

    preview = write_assembled_preview(final_paths)
    summary = RUN_ROOT / "chain_summary_01_to_10.txt"
    summary.write_text(
        "01 -> 02 -> 03 -> 04 -> 10 target runtime case: PASS\n"
        "10 edges: %s, %s\n"
        "10 commands: ingress max/min, egress max\n"
        "Pending u_harden_a: empty\n"
        "Pending u_harden_b: empty\n"
        "Assembled preview: %s\n"
        % (INGRESS_ID, EGRESS_ID, preview.relative_to(RUN_ROOT).as_posix()),
        encoding="utf-8",
    )

    print("\n01 -> 10 chain completed")
    print("Run root          : %s" % RUN_ROOT)
    print("10 inventory      : %s" % inventory_path)
    print("10 SDC            : %s" % feedthrough_sdc)
    print("10 removed log    : %s" % removed_path)
    print("Assembled preview : %s" % preview)
    print("Pending ports     : all consumed")
    print("Summary           : %s" % summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
