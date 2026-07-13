#!/usr/bin/env python3
"""Fill the demo review workbooks created by stages 02, 03 and 04."""

from __future__ import print_function

from openpyxl import load_workbook

from build_case import RUN_ROOT


def find_header(ws, required):
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row_idx, col_idx).value
            if value not in (None, ""):
                mapping[str(value).strip()] = col_idx
        if required in mapping:
            return row_idx, mapping
    raise RuntimeError("%s header not found in %s" % (required, ws.title))


def set_values(ws, row_idx, columns, values):
    for key, value in values.items():
        if key in columns:
            ws.cell(row_idx, columns[key], value)


def fill_02():
    path = RUN_ROOT / "02_middle/02_soc_clock_timing_budget_prects.xlsx"
    wb = load_workbook(str(path))
    ws = wb["clock_budget"]
    header_row, columns = find_header(ws, "clock_name")
    budgets = {
        "top_clk_ref_pad": {
            "setup_uncertainty": 0.10,
            "hold_uncertainty": 0.03,
            "source_latency_early": 0.08,
            "source_latency_late": 0.20,
            "network_latency_early": 0.25,
            "network_latency_late": 0.60,
            "transition_min": 0.03,
            "transition_max": 0.11,
        },
        "u_harden_a_clk_pll_o": {
            "setup_uncertainty": 0.12,
            "hold_uncertainty": 0.04,
            "network_latency_early": 0.30,
            "network_latency_late": 0.70,
            "transition_min": 0.03,
            "transition_max": 0.12,
        },
        "u_harden_b_clk_o": {
            "setup_uncertainty": 0.11,
            "hold_uncertainty": 0.035,
            "network_latency_early": 0.24,
            "network_latency_late": 0.62,
            "transition_min": 0.03,
            "transition_max": 0.12,
        },
        "v_pcie_ref_clk": {"setup_uncertainty": 0.05, "hold_uncertainty": 0.02},
        "v_gpio_ref_clk": {"setup_uncertainty": 0.06, "hold_uncertainty": 0.02},
    }
    seen = set()
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row_idx, columns["clock_name"]).value
        if name not in budgets:
            continue
        seen.add(name)
        values = dict(budgets[name])
        values.update({"apply": "yes", "note": "01-04 demo reviewed clock budget"})
        set_values(ws, row_idx, columns, values)
    missing = sorted(set(budgets) - seen)
    if missing:
        raise RuntimeError("02 expected clock(s) missing: %s" % ", ".join(missing))
    wb.save(str(path))
    print("02 review filled: %s" % path, flush=True)


def fill_03():
    path = RUN_ROOT / "03_middle/03_soc_clock_groups.xlsx"
    wb = load_workbook(str(path))
    ws = wb["clock_group_rules"]
    header_row, columns = find_header(ws, "group_id")
    values = {
        "scenario": "common",
        "group_id": "CG_SYSTEM_VS_GPIO",
        "relation_type": "asynchronous",
        "group_1_clocks": "top_clk_ref_pad",
        "group_2_clocks": "v_gpio_ref_clk",
        "analysis_style": "normal",
        "apply": "yes",
        "review_status": "approved",
        "owner": "sta_demo",
        "basis": "UART virtual board clock is asynchronous to the on-chip system clock tree",
        "cdc_required": "yes",
    }
    set_values(ws, header_row + 1, columns, values)
    wb.save(str(path))
    print("03 review filled: %s" % path, flush=True)


def fill_04():
    path = RUN_ROOT / "04_middle/04_soc_io_pads.xlsx"
    wb = load_workbook(str(path))
    ws = wb["io_constraints"]
    header_row, columns = find_header(ws, "constraint_type")
    approved = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        pad = ws.cell(row_idx, columns["pad_name"]).value
        ctype = ws.cell(row_idx, columns["constraint_type"]).value
        if pad not in {"uart_rx_pad", "uart_tx_pad"}:
            continue
        if ctype not in {"input_delay", "output_delay", "input_transition", "load"}:
            continue
        basis = "UART board timing budget" if ctype in {"input_delay", "output_delay"} else "UART board electrical model"
        add_delay = ""
        if ctype in {"input_delay", "output_delay"}:
            add_delay = "yes" if ws.cell(row_idx, columns["max_value"]).value not in (None, "") else "no"
        set_values(
            ws,
            row_idx,
            columns,
            {
                "apply": "yes",
                "review_status": "approved",
                "timing_class": "timed",
                "add_delay": add_delay,
                "owner": "io_demo",
                "basis": basis,
            },
        )
        approved += 1
    if approved != 6:
        raise RuntimeError("04 expected 6 UART rows, approved %d" % approved)
    wb.save(str(path))
    print("04 review filled: %s" % path, flush=True)


if __name__ == "__main__":
    fill_02()
    fill_03()
    fill_04()
