#!/usr/bin/env python3
"""Demo2: fill prects budget form across corners/scenarios."""
from openpyxl import load_workbook

FORM = "02_soc_clock_timing_budget_prects.xlsx"
wb = load_workbook(FORM)
ws = wb["clock_budget"]

hdr_row, col = None, {}
for r in range(1, 8):
    names = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
    if "clock_name" in names:
        hdr_row, col = r, names
        break

def set_cells(r, **kv):
    for k, v in kv.items():
        ws.cell(r, col[k], v)

# clock_kind groups (from 01 inventory)
TOP = ["top_sys_clk_pad", "top_aux_clk_pad", "top_scan_clk_pad"]
GEN = ["u_pll_core_clk_o", "u_pll_bus_clk_o"]
FWD = ["u_fab0_fab_clk_o", "u_fab1_fab_clk_o", "u_periph_clk_o"]
VIRT = ["v_ddr_ref", "v_pcie_ref"]
ALL = TOP + GEN + FWD + VIRT

def vals_for(clock, corner):
    # tighter at ff_m40 to make corners visibly different
    k = 1.0 if corner == "ss_125" else 0.6
    if clock in TOP:
        return dict(setup_uncertainty=round(0.10*k,3), hold_uncertainty=round(0.03*k,3),
                    source_latency_early=round(0.10*k,3), source_latency_late=round(0.24*k,3),
                    network_latency_early=round(0.28*k,3), network_latency_late=round(0.65*k,3),
                    transition_min=0.03, transition_max=round(0.11*k,3))
    if clock in GEN:
        return dict(setup_uncertainty=round(0.12*k,3), hold_uncertainty=round(0.04*k,3),
                    network_latency_early=round(0.30*k,3), network_latency_late=round(0.70*k,3),
                    transition_min=0.03, transition_max=round(0.12*k,3))
    if clock in FWD:
        return dict(setup_uncertainty=round(0.11*k,3), hold_uncertainty=round(0.035*k,3),
                    network_latency_early=round(0.25*k,3), network_latency_late=round(0.60*k,3))
    # virtual: uncertainty only
    return dict(setup_uncertainty=round(0.05*k,3), hold_uncertainty=0.02)

# 1) fill auto-created common/ss_125 rows
for r in range(hdr_row + 1, ws.max_row + 1):
    cn = ws.cell(r, col["clock_name"]).value
    if cn in ALL:
        set_cells(r, apply="yes", sync_status="OK", **vals_for(cn, "ss_125"))

def last_row():
    lr = hdr_row
    for r in range(hdr_row + 1, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in col.values()):
            lr = r
    return lr

def append(scenario, corner, clock, **extra):
    r = last_row() + 1
    base = dict(scenario=scenario, stage="prects", corner=corner, clock_name=clock,
                apply="yes", sync_status="OK")
    base.update(extra)
    set_cells(r, **base)
    return r

# 2) common/ff_m40 : all 10 clocks, tighter values; v_ddr_ref gets a DELIBERATE
#    network_latency (virtual clock has no network) -> should warn on ff_m40 run
for cn in ALL:
    extra = vals_for(cn, "ff_m40")
    if cn == "v_ddr_ref":
        extra["network_latency_late"] = 0.40   # <- deliberate clock_kind warning
    append("common", "ff_m40", cn, **extra)

# 3) func/ss_125 : override core clk uncertainty only (other 9 fall back to common)
append("func", "ss_125", "u_pll_core_clk_o",
       setup_uncertainty=0.15, hold_uncertainty=0.05,
       network_latency_early=0.30, network_latency_late=0.70,
       transition_min=0.03, transition_max=0.12)

# 4) scan/ss_125 : suppress core clk (apply=no winner), bump scan clock uncertainty
append("scan", "ss_125", "u_pll_core_clk_o", apply="no")              # suppression
append("scan", "ss_125", "top_scan_clk_pad",
       setup_uncertainty=0.20, hold_uncertainty=0.06,
       source_latency_early=0.10, source_latency_late=0.30,
       network_latency_early=0.30, network_latency_late=0.80,
       transition_min=0.03, transition_max=0.15)

wb.save(FORM)
print("prects form filled:", last_row() - hdr_row, "data rows")
