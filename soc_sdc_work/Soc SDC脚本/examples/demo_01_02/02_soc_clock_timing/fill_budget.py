#!/usr/bin/env python3
"""Demo: fill the auto-created prects budget workbook."""
from openpyxl import load_workbook

FORM = "02_soc_clock_timing_budget_prects.xlsx"
wb = load_workbook(FORM)
ws = wb["clock_budget"]

# locate header row + name->col
hdr_row = None
col = {}
for r in range(1, 8):
    names = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
    if "clock_name" in names:
        hdr_row, col = r, names
        break

def set_row(r, **kv):
    for k, v in kv.items():
        ws.cell(r, col[k], v)

# fill the 5 common rows by clock_name
common_vals = {
    "top_clk_ref_pad":      dict(setup_uncertainty=0.10, hold_uncertainty=0.03,
                                 source_latency_early=0.10, source_latency_late=0.24,
                                 network_latency_early=0.28, network_latency_late=0.65,
                                 transition_min=0.03, transition_max=0.11),
    "u_harden_a_clk_pll_o": dict(setup_uncertainty=0.12, hold_uncertainty=0.04,
                                 network_latency_early=0.30, network_latency_late=0.70,
                                 transition_min=0.03, transition_max=0.12),  # generated: no source_latency
    "u_harden_b_clk_o":     dict(setup_uncertainty=0.11, hold_uncertainty=0.035,
                                 network_latency_early=0.25, network_latency_late=0.60),
    "v_pcie_ref_clk":       dict(setup_uncertainty=0.05, hold_uncertainty=0.02),  # virtual: uncertainty only
    "v_gpio_ref_clk":       dict(setup_uncertainty=0.06, hold_uncertainty=0.02),
}

last_row = hdr_row
for r in range(hdr_row + 1, ws.max_row + 1):
    cn = ws.cell(r, col["clock_name"]).value
    if not cn:
        continue
    last_row = r
    if cn in common_vals:
        set_row(r, apply="yes", sync_status="OK", **common_vals[cn])

# append a func override row for the generated clock; intentionally include a
# source_latency value to trigger the generated-clock clock_kind warning (non-blocking)
fr = last_row + 1
set_row(fr, scenario="func", stage="prects", corner="ss_125",
        clock_name="u_harden_a_clk_pll_o",
        setup_uncertainty=0.15, hold_uncertainty=0.05,
        source_latency_early=0.10,                 # <- deliberate clock_kind warning
        network_latency_early=0.30, network_latency_late=0.70,
        transition_min=0.03, transition_max=0.12,
        apply="yes", sync_status="OK")

wb.save(FORM)
print("budget form filled")
