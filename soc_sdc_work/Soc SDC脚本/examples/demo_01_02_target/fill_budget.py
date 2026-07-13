#!/usr/bin/env python3
"""Fill the 02 pre-CTS workbook created from the 01 assembled inventory."""

from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
FORM = BASE / "run/02_middle/02_soc_clock_timing_budget_prects.xlsx"


def fill_budget():
    wb = load_workbook(str(FORM))
    ws = wb["clock_budget"]

    header_row = None
    columns = {}
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        mapping = {
            str(ws.cell(row_idx, col_idx).value).strip(): col_idx
            for col_idx in range(1, ws.max_column + 1)
            if ws.cell(row_idx, col_idx).value
        }
        if "clock_name" in mapping:
            header_row = row_idx
            columns = mapping
            break
    if header_row is None:
        raise RuntimeError("clock_budget header not found")

    budgets = {
        "top_clk_ref_pad": {
            "setup_uncertainty": 0.10,
            "hold_uncertainty": 0.03,
            "source_latency_early": 0.08,
            "source_latency_late": 0.20,
            "network_latency_early": 0.25,
            "network_latency_late": 0.60,
            "transition_min": 0.03,
            "transition_max": 0.11,
        },
        "u_harden_a_clk_pll_o": {
            "setup_uncertainty": 0.12,
            "hold_uncertainty": 0.04,
            "network_latency_early": 0.30,
            "network_latency_late": 0.70,
            "transition_min": 0.03,
            "transition_max": 0.12,
        },
        "u_harden_b_clk_o": {
            "setup_uncertainty": 0.11,
            "hold_uncertainty": 0.035,
            "network_latency_early": 0.24,
            "network_latency_late": 0.62,
            "transition_min": 0.03,
            "transition_max": 0.12,
        },
        "v_pcie_ref_clk": {
            "setup_uncertainty": 0.05,
            "hold_uncertainty": 0.02,
        },
        "v_gpio_ref_clk": {
            "setup_uncertainty": 0.06,
            "hold_uncertainty": 0.02,
        },
    }

    seen = set()
    for row_idx in range(header_row + 1, ws.max_row + 1):
        clock_name = ws.cell(row_idx, columns["clock_name"]).value
        if clock_name not in budgets:
            continue
        seen.add(clock_name)
        ws.cell(row_idx, columns["apply"], "yes")
        for field, value in budgets[clock_name].items():
            ws.cell(row_idx, columns[field], value)

    missing = sorted(set(budgets) - seen)
    if missing:
        raise RuntimeError("expected clock(s) missing from 02 form: %s" % ", ".join(missing))
    wb.save(str(FORM))
    print("02 timing budget filled: %s" % FORM, flush=True)


if __name__ == "__main__":
    fill_budget()
