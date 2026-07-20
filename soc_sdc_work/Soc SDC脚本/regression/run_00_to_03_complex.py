#!/usr/bin/env python3
"""Run a current-contract complex 00 -> 03 SoC SDC integration case."""

from __future__ import print_function

import csv
import hashlib
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from run_full_chain_complex import write_inputs


BASE = Path(__file__).resolve().parent
SOC = BASE.parent
WORK = BASE / "work" / "full_chain_00_to_03_complex"

EX00 = SOC / "00_harden_port_inventory" / "00_harden_port_inventory.py"
EX01 = SOC / "01_soc_clocks" / "01_extract_soc_clocks.py"
EX02 = SOC / "02_soc_clock_timing" / "02_extract_soc_clock_timing.py"
EX03 = SOC / "03_soc_clock_groups" / "03_extract_soc_clock_groups.py"


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def sha256_file(path):
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def run(command):
    result = subprocess.run(
        [sys.executable] + [str(item) for item in command],
        cwd=str(SOC),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )
    print("Running: %s" % " ".join(str(item) for item in command), flush=True)
    if result.stdout.strip():
        print(result.stdout.rstrip())
    if result.stderr.strip():
        print(result.stderr.rstrip(), file=sys.stderr)
    return result


def run_ok(command, label):
    result = run(command)
    require(
        result.returncode == 0,
        "%s failed with exit code %d" % (label, result.returncode),
    )
    return result


def run_review(command, label):
    result = run(command)
    require(
        result.returncode == 1,
        "%s should stop at the review/synchronization gate, got %d"
        % (label, result.returncode),
    )
    return result


def write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def build_case():
    if WORK.exists():
        shutil.rmtree(str(WORK))
    WORK.mkdir(parents=True)
    write_inputs(WORK)
    inputs = WORK / "inputs"
    old_port = inputs / "ports.xlsx"
    new_port = inputs / "port_complex.xlsx"
    require(old_port.is_file(), "complex input builder did not create ports.xlsx")
    old_port.rename(new_port)
    normalize_legacy_bit_rows(new_port)
    write_csv(
        inputs / "run_context.csv",
        ["run_id", "mode_label", "design_revision", "note"],
        [{
            "run_id": "RUN_00_03_COMPLEX",
            "mode_label": "func",
            "design_revision": "rev_complex_a",
            "note": "five-harden complex 00 to 03 integration regression",
        }],
    )
    write_csv(
        inputs / "required_views.csv",
        [
            "view_id", "stage", "corner", "require_02", "require_04",
            "require_20", "require_30", "note",
        ],
        [{
            "view_id": "prects_ss_125",
            "stage": "prects",
            "corner": "ss_125",
            "require_02": "yes",
            "require_04": "no",
            "require_20": "no",
            "require_30": "no",
            "note": "00-03 complex required timing view",
        }],
    )


def indexed_name(value):
    match = re.match(r"^(.*)\[(\d+)\]$", str(value or "").strip())
    if not match:
        return None
    return match.group(1), int(match.group(2))


def collapse_indexed_values(values, indices):
    parsed = [indexed_name(value) for value in values]
    if not values or any(item is None for item in parsed):
        return values[0] if values else ""
    bases = {item[0] for item in parsed}
    value_indices = [item[1] for item in parsed]
    if len(bases) != 1 or sorted(value_indices) != sorted(indices):
        return values[0]
    base = next(iter(bases))
    return "%s[%d:%d]" % (base, max(indices), min(indices))


def normalize_legacy_bit_rows(path):
    workbook = load_workbook(str(path))
    for sheet in workbook.worksheets:
        headers = [str(sheet.cell(1, col).value or "") for col in range(1, sheet.max_column + 1)]
        rows = []
        for row_idx in range(2, sheet.max_row + 1):
            rows.append({
                header: sheet.cell(row_idx, col_idx).value
                for col_idx, header in enumerate(headers, start=1)
                if header
            })
        groups = {}
        for item in rows:
            for direction in ("Input", "Output", "Inout"):
                parsed = indexed_name(item.get(direction))
                if parsed is None:
                    continue
                key = (direction, parsed[0])
                groups.setdefault(key, []).append((parsed[1], item))
        emitted = set()
        normalized = []
        for item in rows:
            matched = None
            for direction in ("Input", "Output", "Inout"):
                parsed = indexed_name(item.get(direction))
                if parsed is not None:
                    matched = (direction, parsed[0])
                    break
            if matched is None or len(groups[matched]) == 1:
                normalized.append(item)
                continue
            if matched in emitted:
                continue
            emitted.add(matched)
            direction, base = matched
            members = sorted(groups[matched], key=lambda pair: pair[0])
            indices = [pair[0] for pair in members]
            require(
                indices == list(range(min(indices), max(indices) + 1)),
                "%s %s has non-contiguous legacy bit rows" % (sheet.title, base),
            )
            merged = dict(members[0][1])
            merged[direction] = "%s[%d:%d]" % (base, max(indices), min(indices))
            merged[direction + " Width"] = len(indices)
            if direction == "Input":
                merged["From Whom"] = collapse_indexed_values(
                    [pair[1].get("From Whom") for pair in members], indices
                )
            elif direction == "Output":
                to_top = [pair[1].get("To Top") for pair in members if pair[1].get("To Top")]
                if to_top:
                    merged["To Top"] = collapse_indexed_values(to_top, indices)
            elif direction == "Inout":
                merged["Inout Connectivity"] = collapse_indexed_values(
                    [pair[1].get("Inout Connectivity") for pair in members], indices
                )
            normalized.append(merged)
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        for item in normalized:
            sheet.append([item.get(header, "") for header in headers])
    workbook.save(str(path))


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        return list(csv.DictReader(file_obj))


def header_map(ws):
    return {
        str(ws.cell(1, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value
    }


def review_02():
    path = WORK / "02_middle/02_soc_clock_timing_budget_prects.xlsx"
    wb = load_workbook(str(path))
    ws = wb["clock_budget"]
    mapping = header_map(ws)
    reviewed = 0
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row_idx, mapping["stage"]).value or "") != "prects":
            continue
        if str(ws.cell(row_idx, mapping["corner"]).value or "") != "ss_125":
            continue
        clock_name = str(ws.cell(row_idx, mapping["clock_name"]).value or "")
        if not clock_name:
            continue
        ws.cell(row_idx, mapping["setup_uncertainty"], 0.08)
        ws.cell(row_idx, mapping["hold_uncertainty"], 0.03)
        ws.cell(row_idx, mapping["transition_max"], 0.20)
        ws.cell(row_idx, mapping["propagated"], "no")
        ws.cell(row_idx, mapping["apply"], "yes")
        ws.cell(row_idx, mapping["note"], "reviewed complex 00-03 clock budget")
        reviewed += 1
    require(reviewed >= 6, "expected a complex 02 clock universe, got %d rows" % reviewed)
    wb.save(str(path))


def choose_domain_roots(inventory_rows):
    active = [
        row for row in inventory_rows
        if row.get("final_action") in {
            "emit_top_clock", "emit_output_clock", "emit_virtual_clock",
        }
    ]
    by_root = {}
    for row in active:
        root_source = row.get("root_source") or row.get("direct_source") or row["clock_name"]
        by_root.setdefault(root_source, []).append(row)
    selected = {}
    for root_source, rows in sorted(by_root.items()):
        rows = sorted(
            rows,
            key=lambda row: (
                0 if row.get("clock_kind") in {"create_clock", "virtual_clock"} else 1,
                row["clock_name"],
            ),
        )
        selected[root_source] = rows[0]["clock_name"]
    require(3 <= len(selected) <= 8, "expected 3-8 independent clock roots, got %d" % len(selected))
    return active, selected, by_root


def review_03():
    inventory_rows = read_csv(WORK / "01_middle/clock_inventory.csv")
    active, selected, by_root = choose_domain_roots(inventory_rows)
    path = WORK / "03_middle/03_soc_clock_groups.xlsx"
    wb = load_workbook(str(path))
    membership = wb["clock_domain_membership"]
    mapping = header_map(membership)
    row_by_clock = {}
    for row_idx in range(2, membership.max_row + 1):
        clock_name = str(membership.cell(row_idx, mapping["clock_name"]).value or "")
        if clock_name:
            row_by_clock[clock_name] = row_idx

    domain_ids = []
    selected_clocks = set(selected.values())
    for index, (_, clock_name) in enumerate(sorted(selected.items()), start=1):
        domain_id = "DOM_ROOT_%02d" % index
        domain_ids.append(domain_id)
        row_idx = row_by_clock[clock_name]
        membership.cell(row_idx, mapping["domain_id"], domain_id)
        membership.cell(row_idx, mapping["membership_type"], "seed")
        membership.cell(row_idx, mapping["include_descendants"], "yes")
        membership.cell(row_idx, mapping["apply"], "yes")
        membership.cell(row_idx, mapping["review_status"], "approved")
        membership.cell(row_idx, mapping["owner"], "sta_owner")
        membership.cell(row_idx, mapping["basis"], "independent 01 genealogy root")

    for row in active:
        clock_name = row["clock_name"]
        if clock_name in selected_clocks:
            continue
        row_idx = row_by_clock[clock_name]
        membership.cell(row_idx, mapping["apply"], "no")
        membership.cell(row_idx, mapping["review_status"], "reviewed")
        membership.cell(
            row_idx,
            mapping["note"],
            "covered through the selected root domain closure",
        )

    rules = wb["clock_group_rules"]
    rule_map = header_map(rules)
    row_idx = 2
    rules.cell(row_idx, rule_map["group_id"], "CG_ALL_ROOTS_ASYNC")
    rules.cell(row_idx, rule_map["relation_type"], "asynchronous")
    for index, domain_id in enumerate(domain_ids, start=1):
        rules.cell(row_idx, rule_map["group_%d_domains" % index], domain_id)
    rules.cell(row_idx, rule_map["analysis_style"], "normal")
    rules.cell(row_idx, rule_map["apply"], "yes")
    rules.cell(row_idx, rule_map["review_status"], "approved")
    rules.cell(row_idx, rule_map["owner"], "sta_owner")
    rules.cell(
        row_idx,
        rule_map["basis"],
        "complex regression treats independently rooted clock trees as CDC-reviewed asynchronous domains",
    )
    wb.save(str(path))
    return len(active), len(domain_ids), sorted(len(rows) for rows in by_root.values())


def port_workbook_digests():
    return {
        path.name: sha256_file(path)
        for path in sorted((WORK / "inputs").glob("port_*.xlsx"))
    }


def validate_outputs(active_clock_count, domain_count, domain_sizes, port_digests_before_03):
    expected = [
        WORK / "00_middle/stage_completion.meta",
        WORK / "01_middle/clock_inventory.csv",
        WORK / "01_middle/clock_inventory.meta",
        WORK / "01_middle/stage_completion.meta",
        WORK / "01_result/01_soc_clocks.sdc",
        WORK / "02_middle/completion/prects_ss_125.meta",
        WORK / "02_result/02_soc_clock_timing_prects_ss_125.sdc",
        WORK / "03_middle/relation_map.csv",
        WORK / "03_middle/relation_map.meta",
        WORK / "03_middle/stage_completion.meta",
        WORK / "03_result/03_soc_clock_groups.sdc",
        WORK / "03_result/reports/clock_group_coverage_report.xlsx",
    ]
    for path in expected:
        require(path.is_file(), "expected artifact missing: %s" % path)

    completions = {}
    for stage, path in {
        "00": WORK / "00_middle/stage_completion.meta",
        "01": WORK / "01_middle/stage_completion.meta",
        "02": WORK / "02_middle/completion/prects_ss_125.meta",
        "03": WORK / "03_middle/stage_completion.meta",
    }.items():
        payload = json.loads(path.read_text(encoding="utf-8"))
        require(payload["completion_status"] == "complete", "%s completion is not complete" % stage)
        require(int(payload["error_count"]) == 0, "%s completion has errors" % stage)
        require(payload["run_id"] == "RUN_00_03_COMPLEX", "%s run provenance mismatch" % stage)
        completions[stage] = payload

    relation_rows = read_csv(WORK / "03_middle/relation_map.csv")
    expected_pairs = active_clock_count * (active_clock_count - 1) // 2
    require(len(relation_rows) == expected_pairs, "relation map is not complete")
    async_count = sum(row["relation_type"] == "asynchronous" for row in relation_rows)
    synchronous_count = sum(row["relation_type"] == "synchronous" for row in relation_rows)
    expected_synchronous = sum(size * (size - 1) // 2 for size in domain_sizes)
    expected_async = expected_pairs - expected_synchronous
    require(async_count == expected_async, "cross-domain asynchronous pair count mismatch")
    require(synchronous_count == expected_synchronous, "same-domain synchronous pair count mismatch")
    group_sdc = (WORK / "03_result/03_soc_clock_groups.sdc").read_text(encoding="utf-8")
    require("set_clock_groups -asynchronous" in group_sdc, "03 async command missing")
    require(group_sdc.count("-group [get_clocks") == domain_count, "03 domain group count mismatch")
    require(port_workbook_digests() == port_digests_before_03, "03 modified port accounting workbooks")
    require(
        completions["03"]["port_accounting_summary"]
        == "Port accounting: not_applicable; added_bits=0",
        "03 completion port accounting contract mismatch",
    )
    return completions, relation_rows, async_count, synchronous_count


def main():
    build_case()
    run_ok([EX00, "--run-root", WORK], "00")
    run_ok([EX01, "--run-root", WORK], "01")

    command_02 = [EX02, "--run-root", WORK, "--stage", "prects", "--corner", "ss_125"]
    run_review(command_02, "02 workbook creation")
    review_02()
    run_review(command_02, "02 workbook synchronization")
    run_ok(command_02, "02 formal generation")

    command_03 = [EX03, "--run-root", WORK]
    run_review(command_03, "03 workbook creation")
    active_clock_count, domain_count, domain_sizes = review_03()
    run_review(command_03, "03 workbook synchronization")
    port_digests_before_03 = port_workbook_digests()
    run_ok(command_03, "03 formal generation")

    completions, relation_rows, async_count, synchronous_count = validate_outputs(
        active_clock_count, domain_count, domain_sizes, port_digests_before_03
    )
    summary = {
        "status": "PASS",
        "run_root": str(WORK.resolve()),
        "harden_count": 5,
        "active_clock_count": active_clock_count,
        "clock_domain_count": domain_count,
        "relation_pair_count": len(relation_rows),
        "asynchronous_pair_count": async_count,
        "synchronous_pair_count": synchronous_count,
        "stages": {
            stage: {
                "completion_status": payload["completion_status"],
                "structure_digest": payload["structure_digest"],
            }
            for stage, payload in completions.items()
        },
    }
    summary_path = WORK / "00_to_03_complex_summary.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
        encoding="utf-8",
    )
    print("\n00 -> 03 complex chain: PASS")
    print("Run root           : %s" % WORK)
    print("Hardens            : 5")
    print("Active clocks      : %d" % active_clock_count)
    print("Clock domains      : %d" % domain_count)
    print("Relation pairs     : %d" % len(relation_rows))
    print("  asynchronous     : %d" % async_count)
    print("  synchronous      : %d" % synchronous_count)
    print("Summary            : %s" % summary_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
