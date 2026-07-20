#!/usr/bin/env python3
"""Release-preflight matrix for the current 00/01 target runtimes.

All generated fixtures and artifacts stay below::

    regression/work/preflight_00_01_matrix/

The 00-30 production implementations and rule documents are treated as
read-only subjects.  Their hashes are checked before and after the matrix.
"""

from __future__ import print_function

import csv
import hashlib
import json
import re
import shutil
import sys
import traceback
from pathlib import Path

from openpyxl import load_workbook

import run_full_chain_complex as full


BASE = Path(__file__).resolve().parent
SOC = BASE.parent
WORK = BASE / "work" / "preflight_00_01_matrix"

EX00 = full.EX00
EX01 = full.EX01

STAGE_DIRS = [
    "00_harden_port_inventory",
    "01_soc_clocks",
    "02_soc_clock_timing",
    "03_soc_clock_groups",
    "04_soc_io_pads",
    "10_feedthrough",
    "20_harden_x_if",
    "30_harden_to_harden_exception",
]


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def clean_dir(path):
    if path.exists():
        shutil.rmtree(str(path))
    path.mkdir(parents=True)


def sha256_file(path):
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def subject_digests():
    """Hash 00-30 production source/rule files, excluding stage-local tests."""
    result = {}
    for directory in STAGE_DIRS:
        root = SOC / directory
        for path in sorted(root.rglob("*")):
            if not path.is_file():
                continue
            if "regression_test" in path.parts or "__pycache__" in path.parts:
                continue
            if path.suffix not in {".py", ".md"}:
                continue
            result[str(path.relative_to(SOC))] = sha256_file(path)
    return result


def log_result(root, label, result):
    log_dir = root / "preflight_logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    token = re.sub(r"[^A-Za-z0-9_.-]+", "_", label)
    (log_dir / (token + ".stdout.txt")).write_text(
        result.stdout or "", encoding="utf-8"
    )
    (log_dir / (token + ".stderr.txt")).write_text(
        result.stderr or "", encoding="utf-8"
    )


def run_command(root, command, label, expected=None):
    result = full.sh(command, root)
    log_result(root, label, result)
    print("[%s] rc=%d" % (label, result.returncode), flush=True)
    if result.stdout.strip():
        print(result.stdout.rstrip())
    if result.stderr.strip():
        print(result.stderr.rstrip(), file=sys.stderr)
    if expected is not None:
        require(
            result.returncode == expected,
            "%s expected rc=%d, got rc=%d\nSTDOUT:\n%s\nSTDERR:\n%s"
            % (label, expected, result.returncode, result.stdout, result.stderr),
        )
    return result


def prepare_case(root):
    clean_dir(root)
    full.prepare_current_inputs(root)


def find_headers(ws):
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value not in (None, "")
    }


def append_port_row(root, sheet_name, values):
    path = root / "inputs" / "port_complex.xlsx"
    wb = load_workbook(str(path))
    ws = wb[sheet_name]
    headers = find_headers(ws)
    row_idx = ws.max_row + 1
    for header, value in values.items():
        require(header in headers, "port workbook missing column %s" % header)
        ws.cell(row_idx, headers[header], value)
    wb.save(str(path))
    wb.close()
    return row_idx


def read_used_value(root, sheet_name, port_column, port_name, used_column):
    path = root / "inputs" / "port_complex.xlsx"
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = find_headers(ws)
    result = None
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row_idx, headers[port_column]).value or "") == port_name:
            result = str(ws.cell(row_idx, headers[used_column]).value or "")
            break
    wb.close()
    require(result is not None, "%s/%s not found" % (sheet_name, port_name))
    return result


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        return list(csv.DictReader(file_obj))


def assert_contains(path, snippets):
    require(path.is_file(), "missing expected artifact: %s" % path)
    text = path.read_text(encoding="utf-8")
    for snippet in snippets:
        require(snippet in text, "%s missing expected text: %s" % (path, snippet))
    return text


def append_structural_positive_rows(root):
    append_port_row(
        root,
        "u_ctrl",
        {
            "Input": "spare_nc_i[1:0]",
            "Input Width": 2,
            "From Whom": "NC",
        },
    )
    append_port_row(
        root,
        "u_ctrl",
        {
            "Output": "spare_open_o",
            "Output Width": 1,
            "To Top": "OPEN",
        },
    )
    append_port_row(
        root,
        "u_ctrl",
        {
            "Input": "strap_tie_i",
            "Input Width": 1,
            "From Whom": "TIE1",
        },
    )
    append_port_row(
        root,
        "u_ctrl",
        {
            "Inout": "mode_literal_io[3:0]",
            "Inout Width": 4,
            "Inout Connectivity": "4'b1010",
        },
    )


def case_00_structural_terminals():
    root = WORK / "00_structural_terminals"
    prepare_case(root)
    append_structural_positive_rows(root)
    run_command(root, [EX00, "--run-root", root], "00_structural", expected=0)

    used = {
        "NC": read_used_value(
            root, "u_ctrl", "Input", "spare_nc_i[1:0]", "Input Used Width"
        ),
        "OPEN": read_used_value(
            root, "u_ctrl", "Output", "spare_open_o", "Output Used Width"
        ),
        "TIE1": read_used_value(
            root, "u_ctrl", "Input", "strap_tie_i", "Input Used Width"
        ),
        "4'b1010": read_used_value(
            root, "u_ctrl", "Inout", "mode_literal_io[3:0]", "Inout Name"
        ),
    }
    require(used["NC"] == "0,1", "NC vector was not fully accounted: %s" % used)
    require(used["OPEN"] == "0", "OPEN scalar was not accounted: %s" % used)
    require(used["TIE1"] == "0", "TIE1 scalar was not accounted: %s" % used)
    require(
        used["4'b1010"] == "0,1,2,3",
        "sized literal vector was not fully accounted: %s" % used,
    )

    delta_path = root / "00_middle" / "port_accounting_delta.csv"
    rows = read_csv(delta_path)
    selected = [
        row
        for row in rows
        if row["port"]
        in {"spare_nc_i[1:0]", "spare_open_o", "strap_tie_i", "mode_literal_io[3:0]"}
    ]
    require(len(selected) == 8, "expected eight structural bit deltas, got %d" % len(selected))
    reasons = {row["port"]: row["reason"] for row in selected}
    require(reasons["spare_nc_i[1:0]"] == "structural_nc", "NC reason mismatch")
    require(reasons["spare_open_o"] == "structural_open", "OPEN reason mismatch")
    require(reasons["strap_tie_i"] == "structural_tie_off", "TIE reason mismatch")
    require(
        reasons["mode_literal_io[3:0]"] == "structural_tie_off",
        "literal reason mismatch",
    )
    require(
        all(row["owner_object_id"].startswith("STRUCT_") for row in selected),
        "structural owner IDs are not stable STRUCT_* IDs",
    )
    require(
        (root / "00_middle" / "stage_completion.meta").is_file(),
        "00 completion meta missing",
    )
    return {
        "status": "PASS",
        "covered": ["NC", "OPEN", "TIE1", "sized_verilog_literal"],
        "used_state": used,
        "delta_rows": len(selected),
        "artifact": str(delta_path),
    }


def case_00_literal_width_rejected():
    root = WORK / "00_literal_width_rejected"
    prepare_case(root)
    append_port_row(
        root,
        "u_ctrl",
        {
            "Input": "bad_literal_i[3:0]",
            "Input Width": 4,
            "From Whom": "2'b10",
        },
    )
    result = run_command(root, [EX00, "--run-root", root], "00_expected_rejection", expected=2)
    report = root / "00_result" / "reports" / "environment_report.txt"
    assert_contains(report, ["sized Verilog literal width 2 does not match port width 4"])
    require(
        read_used_value(root, "u_ctrl", "Input", "bad_literal_i[3:0]", "Input Used Width") == "",
        "failed 00 validation must not commit Used state",
    )
    require(
        not (root / "00_middle" / "stage_completion.meta").exists(),
        "failed 00 validation unexpectedly published completion meta",
    )
    return {
        "status": "PASS",
        "expected_return_code": 2,
        "actual_return_code": result.returncode,
        "rejection": "sized literal width mismatch",
        "report": str(report),
    }


def case_00_fresh_resume():
    root = WORK / "00_fresh_resume"
    prepare_case(root)
    append_port_row(
        root,
        "u_ctrl",
        {
            "Input": "resume_nc_i",
            "Input Width": 1,
            "From Whom": "NO_CONNECT",
        },
    )
    run_command(root, [EX00, "--run-root", root], "00_fresh", expected=0)
    fresh_retry = run_command(
        root, [EX00, "--run-root", root], "00_fresh_retry_rejected", expected=2
    )
    resume = run_command(
        root,
        [EX00, "--run-root", root, "--resume-accounting"],
        "00_resume",
        expected=0,
    )
    meta_path = root / "00_middle" / "port_accounting_delta.meta"
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    transactions = meta.get("transactions", [])
    require(len(transactions) == 2, "fresh+resume should commit two transactions")
    require(
        transactions[1]["accounting_digest_before"]
        == transactions[1]["accounting_digest_after"],
        "resume without new terminals should preserve accounting digest",
    )
    delta_rows = read_csv(root / "00_middle" / "port_accounting_delta.csv")
    second_tx = transactions[1]["transaction_id"]
    resumed_rows = [row for row in delta_rows if row["transaction_id"] == second_tx]
    require(resumed_rows, "resume transaction has no evidence rows")
    require(
        all(row["added_bits"] == "" for row in resumed_rows),
        "resume should not claim newly added bits",
    )
    report = root / "00_result" / "reports" / "environment_report.txt"
    assert_contains(report, ["Initialization mode: resume"])
    return {
        "status": "PASS",
        "fresh_retry_return_code": fresh_retry.returncode,
        "resume_return_code": resume.returncode,
        "transaction_count": len(transactions),
        "report": str(report),
    }


def remove_sdc(root, filename):
    path = root / "inputs" / filename
    require(path.is_file(), "fixture SDC missing before mutation: %s" % path)
    path.unlink()


def manifest_status(root, inst_name):
    rows = read_csv(root / "00_middle" / "harden_sdc_manifest.csv")
    matches = [row for row in rows if row["inst_name"] == inst_name]
    require(len(matches) == 1, "manifest row count mismatch for %s" % inst_name)
    return matches[0]


def case_00_missing_sdc_strictness():
    default_root = WORK / "00_missing_sdc_default"
    prepare_case(default_root)
    remove_sdc(default_root, "dpg.sdc")
    run_command(default_root, [EX00, "--run-root", default_root], "00_default", expected=0)
    row = manifest_status(default_root, "u_dpg")
    require(row["availability_status"] == "missing", "default manifest did not record missing SDC")

    strict_root = WORK / "00_missing_sdc_strict"
    prepare_case(strict_root)
    remove_sdc(strict_root, "dpg.sdc")
    strict = run_command(
        strict_root,
        [EX00, "--run-root", strict_root, "--require-complete-harden-sdc"],
        "00_strict_expected_rejection",
        expected=2,
    )
    report = strict_root / "00_result" / "reports" / "environment_report.txt"
    assert_contains(report, ["--require-complete-harden-sdc: 1 harden SDC file(s) are missing"])
    require(
        not (strict_root / "00_middle" / "stage_completion.meta").exists(),
        "strict missing-SDC failure unexpectedly published completion meta",
    )
    return {
        "status": "PASS",
        "default_status": row["availability_status"],
        "strict_return_code": strict.returncode,
        "strict_report": str(report),
    }


def write_clkgen_sdc(root, lines):
    path = root / "inputs" / "clkgen.sdc"
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return path


def primary_clock_lines():
    return [
        "create_clock -name ref_clk -period 10.000 -waveform {0 5} [get_ports ref_clk_i]",
        "create_clock -name scan_clk -period 50.000 [get_ports scan_clk_i]",
        "create_clock -name peri_ref_clk -period 20.000 [get_ports peri_ref_i]",
    ]


def sdc_command_for_clock(text, clock_name):
    marker = "-name %s" % clock_name
    matches = [line.strip() for line in text.splitlines() if marker in line and line.strip().startswith("create_")]
    require(len(matches) == 1, "expected one command for %s, got %d" % (clock_name, len(matches)))
    return matches[0]


def case_01_generated_clock_options():
    root = WORK / "01_generated_clock_options"
    prepare_case(root)
    append_port_row(
        root,
        "u_clkgen",
        {
            "Output": "legacy_clk_o[7]",
            "Output Width": 1,
        },
    )
    write_clkgen_sdc(
        root,
        primary_clock_lines()
        + [
            # Source before target.
            "create_generated_clock -name core_clk -source [get_ports ref_clk_i] -multiply_by 2 -waveform {0 2.5} [get_ports core_clk_o]",
            # Positional target before -source: the target must not be confused with source.
            "create_generated_clock -name bus_clk [get_ports bus_clk_o] -source [get_ports ref_clk_i] -divide_by 2",
            # Historical alias plus edge-list options.
            "create_generate_clock -name peri_clk -source [get_ports peri_ref_i] -edges {1 3 5} -edge_shift {0.0 0.1 0.2} [get_ports peri_clk_o]",
            "create_generated_clock -name scan_out_clk -source [get_ports scan_clk_i] -combinational [get_ports scan_clk_o]",
            "create_generated_clock -name dbg_clk -source [get_ports ref_clk_i] -master_clock ref_clk -divide_by 4 [get_ports dbg_clk_o]",
            "create_generated_clock -name legacy_bit -source [get_ports ref_clk_i] -divide_by 2 [get_ports legacy_clk_o[7]]",
        ],
    )
    run_command(root, [EX00, "--run-root", root], "00", expected=0)
    run_command(root, [EX01, "--run-root", root], "01", expected=0)

    output = root / "01_result" / "01_soc_clocks.sdc"
    text = output.read_text(encoding="utf-8")
    core = sdc_command_for_clock(text, "u_clkgen_core_clk_o")
    bus = sdc_command_for_clock(text, "u_clkgen_bus_clk_o")
    peri = sdc_command_for_clock(text, "u_clkgen_peri_clk_o")
    scan = sdc_command_for_clock(text, "u_clkgen_scan_clk_o")
    dbg = sdc_command_for_clock(text, "u_clkgen_dbg_clk_o")
    legacy_bit = sdc_command_for_clock(text, "u_clkgen_legacy_clk_o_bit7")

    require("-source [get_pins {u_clkgen/ref_clk_i}]" in core, "source-before-target rewrite failed")
    require("-multiply_by 2" in core and "-waveform {0 2.5}" in core, "multiply/waveform lost")
    require(
        core.endswith("[get_pins {u_clkgen/core_clk_o}]") and "ref_clk_i}]" in core,
        "core target/source rewrite mismatch",
    )
    require(
        "[get_pins {u_clkgen/bus_clk_o}] -source [get_pins {u_clkgen/ref_clk_i}] -divide_by 2"
        in bus,
        "target-before-source ordering was not preserved correctly",
    )
    require(peri.startswith("create_generated_clock "), "create_generate_clock alias was not normalized")
    require("create_generate_clock" not in text, "historical alias leaked into final SDC")
    require("-edges {1 3 5}" in peri, "edges option lost")
    require("-edge_shift {0.0 0.1 0.2}" in peri, "edge_shift option lost")
    require("-combinational" in scan, "combinational option lost")
    require("-master_clock top_ref_clk_pad" in dbg, "master_clock name was not remapped")
    require("-divide_by 4" in dbg, "divide_by option lost from master-clock case")
    require(
        legacy_bit.endswith("[get_pins {u_clkgen/legacy_clk_o[7]}]"),
        "declared nonzero exact-bit target was not emitted canonically",
    )

    top_ref = sdc_command_for_clock(text, "top_ref_clk_pad")
    require("-waveform {0 5}" in top_ref, "primary clock waveform option lost")

    inventory_path = root / "01_middle" / "clock_inventory.csv"
    active = [
        row
        for row in read_csv(inventory_path)
        if row.get("final_action") in {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}
    ]
    by_name = {row["clock_name"]: row for row in active}
    expected = {
        "u_clkgen_core_clk_o",
        "u_clkgen_bus_clk_o",
        "u_clkgen_peri_clk_o",
        "u_clkgen_scan_clk_o",
        "u_clkgen_dbg_clk_o",
        "u_clkgen_legacy_clk_o_bit7",
    }
    require(expected.issubset(by_name), "generated clock inventory coverage incomplete")
    require(
        by_name["u_clkgen_scan_clk_o"]["clock_kind"] == "generated_combinational",
        "combinational clock kind not retained in inventory",
    )
    require(
        "create_generate_clock" in by_name["u_clkgen_peri_clk_o"]["original_command"],
        "inventory lost alias provenance",
    )
    require(
        by_name["u_clkgen_legacy_clk_o_bit7"]["port_name"] == "legacy_clk_o[7]",
        "inventory lost declared nonzero exact-bit key",
    )
    require(
        read_used_value(
            root,
            "u_clkgen",
            "Output",
            "legacy_clk_o[7]",
            "Output Used Width",
        )
        == "7",
        "declared nonzero exact-bit target was not accounted with HDL index 7",
    )
    report = root / "01_result" / "reports" / "clock_check_report.txt"
    report_text = assert_contains(report, ["Errors  : 0"])
    require(
        "CLOCK_TARGET_BIT_OUT_OF_RANGE" not in report_text,
        "declared nonzero exact-bit target was misclassified as out of range",
    )
    return {
        "status": "PASS",
        "artifact": str(output),
        "inventory": str(inventory_path),
        "covered": [
            "source_before_target",
            "target_before_source",
            "create_generate_clock_alias",
            "multiply_by",
            "divide_by",
            "combinational",
            "edges",
            "edge_shift",
            "master_clock",
            "waveform",
            "legacy_nonzero_exact_bit",
        ],
    }


def case_01_missing_source_rejected():
    root = WORK / "01_missing_source_rejected"
    prepare_case(root)
    write_clkgen_sdc(
        root,
        primary_clock_lines()
        + [
            "create_generated_clock -name bad_missing_source -divide_by 2 [get_ports dbg_clk_o]",
        ],
    )
    run_command(root, [EX00, "--run-root", root], "00", expected=0)
    result = run_command(root, [EX01, "--run-root", root], "01_expected_rejection", expected=1)
    report = root / "01_result" / "reports" / "clock_check_report.txt"
    assert_contains(report, ["CLOCK_GENERATED_MISSING_SOURCE"])
    return {
        "status": "PASS",
        "expected_return_code": 1,
        "actual_return_code": result.returncode,
        "rejection": "generated clock missing -source",
        "report": str(report),
    }


def case_01_range_target_rejected():
    root = WORK / "01_range_target_rejected"
    prepare_case(root)
    append_port_row(
        root,
        "u_clkgen",
        {
            "Output": "window_clk_o[11:8]",
            "Output Width": 4,
        },
    )
    write_clkgen_sdc(
        root,
        primary_clock_lines()
        + [
            "create_generated_clock -name bad_range -source [get_ports ref_clk_i] -divide_by 2 [get_ports status_o[7:0]]",
            "create_generated_clock -name bad_nonzero_range -source [get_ports ref_clk_i] -divide_by 2 [get_ports window_clk_o[11:8]]",
        ],
    )
    run_command(root, [EX00, "--run-root", root], "00", expected=0)
    port_workbook = root / "inputs" / "port_complex.xlsx"
    workbook_digest_before = sha256_file(port_workbook)
    result = run_command(root, [EX01, "--run-root", root], "01_expected_rejection", expected=1)
    workbook_digest_after = sha256_file(port_workbook)
    report = root / "01_result" / "reports" / "clock_check_report.txt"
    report_text = assert_contains(
        report,
        [
            "CLOCK_TARGET_RANGE_NOT_SUPPORTED",
            "status_o[0]",
            "window_clk_o[8]",
        ],
    )
    require(
        "CLOCK_TARGET_VECTOR_REQUIRES_BIT" not in report_text,
        "range target was misclassified as a vector base",
    )
    require(
        "status_o[7:0][0]" not in report_text,
        "range target suggestion used invalid double indexing",
    )
    require(
        "window_clk_o[0]" not in report_text
        and "window_clk_o[11:8][0]" not in report_text,
        "nonzero range target suggestion did not use a legal HDL bit",
    )
    require(
        workbook_digest_before == workbook_digest_after,
        "range target failure modified the port workbook",
    )
    forbidden_artifacts = [
        root / "01_result" / "01_soc_clocks.sdc",
        root / "01_middle" / "clock_inventory.csv",
        root / "01_middle" / "clock_inventory.meta",
        root / "01_middle" / "port_accounting_delta.csv",
        root / "01_middle" / "port_accounting_delta.meta",
        root / "01_middle" / "stage_completion.meta",
    ]
    for artifact in forbidden_artifacts:
        require(
            not artifact.exists(),
            "range target failure published formal artifact: %s" % artifact,
        )
    return {
        "status": "PASS",
        "expected_return_code": 1,
        "actual_return_code": result.returncode,
        "rejection": "generated clock vector/range target",
        "observed_diagnostic": "CLOCK_TARGET_RANGE_NOT_SUPPORTED",
        "suggested_bits": ["status_o[0]", "window_clk_o[8]"],
        "workbook_unchanged": True,
        "formal_artifacts_published": False,
        "report": str(report),
    }


def detect_active_findings():
    """Return low-risk subject findings only while their exact evidence remains."""
    findings = []

    report_00 = (
        WORK
        / "00_literal_width_rejected"
        / "00_result"
        / "reports"
        / "environment_report.txt"
    )
    if report_00.is_file():
        text_00 = report_00.read_text(encoding="utf-8")
        width_message = "sized Verilog literal width 2 does not match port width 4"
        generic_message = (
            "connection must be <inst>.<port/range> or an explicit structural token"
        )
        if text_00.count(width_message) >= 2 and generic_message in text_00:
            findings.append(
                {
                    "finding_id": "00-F001",
                    "stage": "00",
                    "severity": "low",
                    "type": "diagnostic_quality",
                    "description_zh": "非法 sized literal 的宽度错误被重复报告，并附带误导性的连接语法错误；核心拒绝与不提交行为正确。",
                    "observed_evidence": [
                        str(report_00),
                        str(
                            WORK
                            / "00_literal_width_rejected"
                            / "preflight_logs"
                            / "00_expected_rejection.stderr.txt"
                        ),
                    ],
                }
            )

    report_01 = (
        WORK
        / "01_range_target_rejected"
        / "01_result"
        / "reports"
        / "clock_check_report.txt"
    )
    if report_01.is_file():
        text_01 = report_01.read_text(encoding="utf-8")
        if (
            "CLOCK_TARGET_VECTOR_REQUIRES_BIT" in text_01
            and "status_o[7:0][0]" in text_01
        ):
            findings.append(
                {
                    "finding_id": "01-F001",
                    "stage": "01",
                    "severity": "low",
                    "type": "diagnostic_quality",
                    "description_zh": "整段 range target 被正确拒绝，但 rule-id 分类为 VECTOR_REQUIRES_BIT，且修复示例被拼成非法的 status_o[7:0][0]。",
                    "observed_evidence": [
                        str(report_01),
                        str(
                            WORK
                            / "01_range_target_rejected"
                            / "preflight_logs"
                            / "01_expected_rejection.stdout.txt"
                        ),
                    ],
                }
            )

    return findings


def main():
    before = subject_digests()
    clean_dir(WORK)
    cases = [
        ("00_structural_terminals", case_00_structural_terminals),
        ("00_literal_width_rejected", case_00_literal_width_rejected),
        ("00_fresh_resume", case_00_fresh_resume),
        ("00_missing_sdc_strictness", case_00_missing_sdc_strictness),
        ("01_generated_clock_options", case_01_generated_clock_options),
        ("01_missing_source_rejected", case_01_missing_source_rejected),
        ("01_range_target_rejected", case_01_range_target_rejected),
    ]
    summary = {
        "work_root": str(WORK),
        "subject_scope": "00-30 stage source/rule files are read-only",
        "cases": {},
    }
    failures = []
    for name, function in cases:
        print("\n=== %s ===" % name, flush=True)
        try:
            summary["cases"][name] = function()
        except Exception as exc:
            failures.append(name)
            summary["cases"][name] = {
                "status": "FAIL",
                "error": str(exc),
                "traceback": traceback.format_exc(),
            }
            print("CASE FAILED %s: %s" % (name, exc), file=sys.stderr)

    summary["active_findings"] = detect_active_findings()
    after = subject_digests()
    summary["subject_files_checked"] = len(before)
    summary["subject_files_unchanged"] = before == after
    if before != after:
        changed = sorted(set(before) | set(after))
        changed = [name for name in changed if before.get(name) != after.get(name)]
        summary["subject_files_changed"] = changed
        failures.append("subject_immutability")
    summary["overall_status"] = "PASS" if not failures else "FAIL"
    summary["failed_cases"] = failures
    summary_path = WORK / "summary.json"
    summary_path.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    print("\n00-01 preflight matrix: %s" % summary["overall_status"])
    print("Work root:", WORK)
    print("Summary  :", summary_path)
    print("Subject files unchanged:", summary["subject_files_unchanged"])
    if failures:
        print("Failed cases:", ", ".join(failures), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
