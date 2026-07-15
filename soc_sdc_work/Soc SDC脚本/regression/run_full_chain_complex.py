#!/usr/bin/env python3
"""Build and run a complex 00 -> 30 SoC top synthesis SDC flow case.

The case has five hardens. Each harden starts with at least twenty pending
canonical scalar/bit ports. The script runs 00/01/02/03/04/10/20/30 in order
and approves a controlled set of rows so every stage emits representative
artifacts through the fixed middle/result runtime contract.
"""

import csv
import json
import shutil
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
SOC = BASE.parent
WORK = BASE / "work" / "full_chain_00_to_30_complex"

EX00 = SOC / "00_harden_port_inventory.py"
EX01 = SOC / "01_soc_clocks" / "01_extract_soc_clocks.py"
EX02 = SOC / "02_soc_clock_timing" / "02_extract_soc_clock_timing.py"
EX03 = SOC / "03_soc_clock_groups" / "03_extract_soc_clock_groups.py"
EX04 = SOC / "04_soc_io_pads" / "04_extract_soc_io_pads.py"
EX10 = SOC / "10_feedthrough" / "10_extract_feedthrough.py"
EX20 = SOC / "20_harden_x_if" / "20_extract_harden_x_if.py"
EX30 = SOC / "30_harden_to_harden_exception" / "30_extract_harden_to_harden_exception.py"

REQ = [
    "Parameter",
    "Inout",
    "Inout Width",
    "Inout Connectivity",
    "Inout Name",
    "Input",
    "Input Width",
    "Input Used Width",
    "From Whom",
    "Output",
    "Output Width",
    "Output Used Width",
    "To Top",
]

INFO = [
    {"module_name": "clkgen", "inst_name": "u_clkgen", "owner": "clock_owner"},
    {"module_name": "dma", "inst_name": "u_dma", "owner": "dma_owner"},
    {"module_name": "dpg", "inst_name": "u_dpg", "owner": "dpg_owner"},
    {"module_name": "periph", "inst_name": "u_periph", "owner": "periph_owner"},
    {"module_name": "ctrl", "inst_name": "u_ctrl", "owner": "ctrl_owner"},
]


def sh(cmd, cwd):
    result = subprocess.run(
        [sys.executable] + [str(part) for part in cmd],
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )
    return result


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def run_ok(cmd, cwd):
    result = sh(cmd, cwd)
    require(result.returncode == 0, "command failed: %s\nSTDOUT:\n%s\nSTDERR:\n%s" % (" ".join(map(str, cmd)), result.stdout, result.stderr))
    return result


def run_sync(cmd, cwd):
    result = sh(cmd, cwd)
    require(result.returncode == 1, "sync/review command should return 1: %s\nSTDOUT:\n%s\nSTDERR:\n%s" % (" ".join(map(str, cmd)), result.stdout, result.stderr))
    return result


def port_sheet(rows):
    df = pd.DataFrame(rows)
    for col in REQ:
        if col not in df.columns:
            df[col] = ""
    return df[REQ].fillna("")


def row(Input="", Inout="", Output="", **kwargs):
    item = {"Input": Input, "Inout": Inout, "Output": Output}
    item.update(kwargs)
    return item


def conn_id(src_i, src_p, dst_i, dst_p):
    def tok(value):
        return value.replace("[", "_bit").replace("]", "")

    return "CONN_%s_%s__%s_%s" % (src_i, tok(src_p), dst_i, tok(dst_p))


def build_ports():
    data = {}

    data["u_clkgen"] = [
        row(Input="ref_clk_i", **{"Input Width": 1, "From Whom": "top.ref_clk_pad"}),
        row(Input="scan_clk_i", **{"Input Width": 1, "From Whom": "top.scan_clk_pad"}),
        row(Input="peri_ref_i", **{"Input Width": 1, "From Whom": "top.peri_ref_pad"}),
        row(Input="pll_bypass_i", **{"Input Width": 1, "From Whom": "top.pll_bypass_pad"}),
        row(Input="cfg_boot_i", **{"Input Width": 1, "From Whom": "u_ctrl.boot_mode_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Output="core_clk_o", **{"Output Width": 1}),
        row(Output="bus_clk_o", **{"Output Width": 1}),
        row(Output="peri_clk_o", **{"Output Width": 1}),
        row(Output="scan_clk_o", **{"Output Width": 1}),
        row(Output="dbg_clk_o", **{"Output Width": 1}),
    ]
    for bit in range(8):
        data["u_clkgen"].append(row(Input="cfg_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_ctrl.clk_cfg_o[%d]" % bit}))
    for bit in range(8):
        data["u_clkgen"].append(row(Output="status_o[%d]" % bit, **{"Output Width": 1}))

    data["u_dma"] = [
        row(Input="clk_i", **{"Input Width": 1, "From Whom": "u_clkgen.core_clk_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Input="irq_ack_i", **{"Input Width": 1, "From Whom": "u_ctrl.irq_ack_o"}),
        row(Input="test_mode_i", **{"Input Width": 1, "From Whom": "top.test_mode_pad"}),
        row(Output="irq_o", **{"Output Width": 1, "To Top": "irq_pad"}),
        row(Output="done_o", **{"Output Width": 1}),
        row(Output="req_valid_o", **{"Output Width": 1}),
    ]
    for bit in range(4):
        data["u_dma"].append(row(Output="data_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dma"].append(row(Output="cfg_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dma"].append(row(Output="async_req_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dma"].append(row(Input="resp_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_periph.resp_o[%d]" % bit}))

    data["u_dpg"] = [
        row(Input="clk_i", **{"Input Width": 1, "From Whom": "u_clkgen.bus_clk_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Input="debug_sel_i", **{"Input Width": 1, "From Whom": "u_ctrl.debug_sel_o"}),
        row(Output="dpg_status_o", **{"Output Width": 1}),
    ]
    for bit in range(4):
        data["u_dpg"].append(row(Input="fti_0_dma2periph_data[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dma.data_o[%d]" % bit}))
    for bit in range(4):
        data["u_dpg"].append(row(Output="fto_0_dma2periph_data[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dpg"].append(row(Input="fti_0_dma2periph_async_req[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dma.async_req_o[%d]" % bit}))
    for bit in range(4):
        data["u_dpg"].append(row(Output="fto_0_dma2periph_async_req[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_dpg"].append(row(Input="cfg_shadow_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_ctrl.ctrl_cfg_o[%d]" % bit}))

    data["u_periph"] = [
        row(Input="clk_i", **{"Input Width": 1, "From Whom": "u_clkgen.peri_clk_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Input="enable_i", **{"Input Width": 1, "From Whom": "u_ctrl.periph_enable_o"}),
        row(Output="ready_o", **{"Output Width": 1}),
        row(Output="event_o", **{"Output Width": 1, "To Top": "event_pad"}),
    ]
    for bit in range(4):
        data["u_periph"].append(row(Input="data_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dpg.fto_0_dma2periph_data[%d]" % bit}))
    for bit in range(4):
        data["u_periph"].append(row(Input="async_req_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dpg.fto_0_dma2periph_async_req[%d]" % bit}))
    for bit in range(4):
        data["u_periph"].append(row(Output="resp_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_periph"].append(row(Input="cfg_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_dma.cfg_o[%d]" % bit}))

    data["u_ctrl"] = [
        row(Input="clk_i", **{"Input Width": 1, "From Whom": "u_clkgen.bus_clk_o"}),
        row(Input="rst_n_i", **{"Input Width": 1, "From Whom": "top.rst_n_pad"}),
        row(Input="gpio_mode_i", **{"Input Width": 1, "From Whom": "top.gpio_mode_pad"}),
        row(Input="scan_en_i", **{"Input Width": 1, "From Whom": "top.scan_en_pad"}),
        row(Output="boot_mode_o", **{"Output Width": 1}),
        row(Output="irq_ack_o", **{"Output Width": 1}),
        row(Output="debug_sel_o", **{"Output Width": 1}),
        row(Output="periph_enable_o", **{"Output Width": 1}),
    ]
    for bit in range(8):
        data["u_ctrl"].append(row(Output="clk_cfg_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_ctrl"].append(row(Output="ctrl_cfg_o[%d]" % bit, **{"Output Width": 1}))
    for bit in range(4):
        data["u_ctrl"].append(row(Input="status_i[%d]" % bit, **{"Input Width": 1, "From Whom": "u_clkgen.status_o[%d]" % bit}))
    return data


def write_inputs(d):
    d = d / "inputs"
    d.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(INFO).to_excel(d / "info_all.xlsx", index=False)
    ports = build_ports()
    with pd.ExcelWriter(d / "ports.xlsx", engine="xlsxwriter") as writer:
        for inst, rows in ports.items():
            port_sheet(rows).to_excel(writer, sheet_name=inst, index=False)

    (d / "virtual_clocks.csv").write_text(
        "clock_name,period,waveform,note\n"
        "v_ext_cfg,10.000,,external config virtual clock\n"
        "v_jtag,50.000,,JTAG virtual clock\n",
        encoding="utf-8",
    )

    (d / "clkgen.sdc").write_text(
        "create_clock -name ref_clk -period 10.000 [get_ports ref_clk_i]\n"
        "create_clock -name scan_clk -period 50.000 [get_ports scan_clk_i]\n"
        "create_clock -name peri_ref_clk -period 20.000 [get_ports peri_ref_i]\n"
        "create_generated_clock -name core_clk -source [get_ports ref_clk_i] -multiply_by 2 [get_ports core_clk_o]\n"
        "create_generated_clock -name bus_clk -source [get_ports ref_clk_i] -divide_by 1 [get_ports bus_clk_o]\n"
        "create_generated_clock -name peri_clk -source [get_ports peri_ref_i] -divide_by 2 [get_ports peri_clk_o]\n"
        "create_generated_clock -name scan_out_clk -source [get_ports scan_clk_i] -combinational [get_ports scan_clk_o]\n",
        encoding="utf-8",
    )
    (d / "dma.sdc").write_text(
        "create_clock -name dma_clk -period 5.000 [get_ports clk_i]\n"
        "set_output_delay -max 1.20 -clock dma_clk [get_ports data_o]\n"
        "set_output_delay -min 0.20 -clock dma_clk [get_ports data_o]\n"
        "set_false_path -from [get_ports cfg_o]\n",
        encoding="utf-8",
    )
    (d / "dpg.sdc").write_text(
        "create_clock -name dpg_clk -period 10.000 [get_ports clk_i]\n",
        encoding="utf-8",
    )
    (d / "periph.sdc").write_text(
        "create_clock -name periph_clk -period 20.000 [get_ports clk_i]\n"
        "set_input_delay -max 1.00 -clock periph_clk [get_ports data_i]\n"
        "set_input_delay -min 0.10 -clock periph_clk [get_ports data_i]\n"
        "set_load 0.35 [get_ports event_o]\n",
        encoding="utf-8",
    )
    (d / "ctrl.sdc").write_text(
        "create_clock -name ctrl_clk -period 10.000 [get_ports clk_i]\n"
        "set_input_transition 0.08 [get_ports gpio_mode_i]\n",
        encoding="utf-8",
    )


def set_cell(ws, headers, row_idx, header, value):
    if header in headers:
        ws.cell(row=row_idx, column=headers[header], value=value)


def approve_02(d, stage, corner):
    wb = load_workbook(str(d / "02_middle" / ("02_soc_clock_timing_budget_%s.xlsx" % stage)))
    ws = wb["clock_budget"]
    headers = {}
    header_row = 0
    for r in range(1, 20):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if "clock_name" in values:
            header_row = r
            headers = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
            break
    require(header_row, "02 header row not found")
    for r in range(header_row + 1, ws.max_row + 1):
        clock = ws.cell(r, headers["clock_name"]).value
        if not clock:
            continue
        set_cell(ws, headers, r, "sync_status", "OK")
        set_cell(ws, headers, r, "apply", "yes")
        set_cell(ws, headers, r, "setup_uncertainty", "0.05")
        set_cell(ws, headers, r, "hold_uncertainty", "0.02")
        if "virtual" not in str(clock):
            set_cell(ws, headers, r, "transition_max", "0.20")
            set_cell(ws, headers, r, "propagated", "no")
        set_cell(ws, headers, r, "note", "full chain regression budget")
    wb.save(str(d / "02_middle" / ("02_soc_clock_timing_budget_%s.xlsx" % stage)))


def approve_03(d):
    wb = load_workbook(str(d / "03_middle" / "03_soc_clock_groups.xlsx"))
    ws = wb["clock_group_rules"]
    headers = {}
    header_row = 0
    for r in range(1, 20):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if "group_id" in values:
            header_row = r
            headers = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
            break
    require(header_row, "03 header row not found")
    r = header_row + 1
    values = {
        "scenario": "common",
        "group_id": "CG_CORE_PERI_ASYNC",
        "relation_type": "asynchronous",
        "group_1_clocks": "top_ref_clk_pad u_clkgen_core_clk_o u_clkgen_bus_clk_o",
        "group_2_clocks": "top_peri_ref_pad",
        "analysis_style": "normal",
        "apply": "yes",
        "review_status": "approved",
        "owner": "sta_owner",
        "basis": "peripheral clock domain is CDC-reviewed against core/bus domain",
        "cdc_required": "yes",
        "note": "full chain regression async relation",
    }
    for header, value in values.items():
        set_cell(ws, headers, r, header, value)
    wb.save(str(d / "03_middle" / "03_soc_clock_groups.xlsx"))


def approve_04(d):
    wb = load_workbook(str(d / "04_middle" / "04_soc_io_pads.xlsx"))
    ws = wb["io_constraints"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    for r in range(2, ws.max_row + 1):
        pad = ws.cell(r, headers["pad_name"]).value
        ctype = ws.cell(r, headers["constraint_type"]).value
        if not pad:
            continue
        set_cell(ws, headers, r, "apply", "yes")
        set_cell(ws, headers, r, "review_status", "approved")
        set_cell(ws, headers, r, "owner", "io_owner")
        set_cell(ws, headers, r, "basis", "board/pad environment reviewed for full chain regression")
        if ctype in ("input_delay", "output_delay"):
            set_cell(ws, headers, r, "clock_name", "top_ref_clk_pad")
            if not ws.cell(r, headers["max_value"]).value and not ws.cell(r, headers["value"]).value:
                set_cell(ws, headers, r, "max_value", "1.0")
        elif ctype in ("input_transition", "max_transition"):
            if not ws.cell(r, headers["value"]).value:
                set_cell(ws, headers, r, "value", "0.10")
        elif ctype == "load":
            if not ws.cell(r, headers["value"]).value:
                set_cell(ws, headers, r, "value", "0.35")
    wb.save(str(d / "04_middle" / "04_soc_io_pads.xlsx"))


def approve_10(d):
    path = d / "10_middle" / "10_feedthrough.xlsx"
    wb = load_workbook(str(path))
    ws = wb["feedthrough_edges"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    approved = 0
    routed = 0
    for r in range(2, ws.max_row + 1):
        connection_id = ws.cell(r, headers["connection_id"]).value or ""
        is_data_bit0 = (
            connection_id == conn_id("u_dma", "data_o[0]", "u_dpg", "fti_0_dma2periph_data[0]")
            or connection_id == conn_id("u_dpg", "fto_0_dma2periph_data[0]", "u_periph", "data_i[0]")
        )
        is_async_route = connection_id == conn_id(
            "u_dpg", "fto_0_dma2periph_async_req[1]", "u_periph", "async_req_i[1]"
        )
        if is_data_bit0:
            values = {
                "channel_disposition": "emit_budget",
                "budget_model": "manual_budget",
                "budget_required": "yes",
                "converted_max": "0.8" if "data_o_bit0" in connection_id else "0.9",
                "converted_min": "0.05",
                "emit_max": "yes",
                "emit_min": "yes",
                "min_sign_review": "approved",
                "datapath_only": "yes",
                "tool_surface": "dc",
                "apply": "yes",
                "review_status": "approved",
                "owner": "soc_timing_owner",
                "reviewer": "full_chain_reviewer",
                "review_date": "2026-07-14",
                "disposition_basis": "reviewed feedthrough-adjacent direct-edge budget",
            }
            approved += 1
        elif is_async_route:
            values = {
                "channel_disposition": "route_to_30",
                "budget_required": "no",
                "emit_max": "no",
                "emit_min": "no",
                "apply": "yes",
                "review_status": "approved",
                "owner": "cdc_owner",
                "reviewer": "full_chain_reviewer",
                "review_date": "2026-07-14",
                "disposition_basis": "CDC propagation window is owned by stage 30",
            }
            routed += 1
        else:
            values = {
                "channel_disposition": "no_soc_budget_required",
                "budget_required": "no",
                "emit_max": "no",
                "emit_min": "no",
                "apply": "yes",
                "review_status": "approved",
                "owner": "soc_policy_owner",
                "reviewer": "full_chain_reviewer",
                "review_date": "2026-07-14",
                "disposition_basis": "project_feedthrough_no_soc_budget_v1",
            }
        for header, value in values.items():
            set_cell(ws, headers, r, header, value)
    require(approved == 2, "expected two 10 emit-budget edges, got %d" % approved)
    require(routed == 1, "expected one 10 route-to-30 edge, got %d" % routed)
    wb.save(str(path))


def approve_20(d):
    form = d / "20_middle" / "20_harden_x_if.xlsx"
    wb = load_workbook(str(form))
    ws = wb["interface_budget"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    approved = 0
    routed = 0
    for r in range(2, ws.max_row + 1):
        channel = ws.cell(r, headers["channel_id"]).value or ""
        if "u_periph_resp_o_bit0__u_dma_resp_i_bit0" in channel:
            values = {
                "scenario": "common",
                "stage": "all",
                "corner": "all",
                "timing_model": "lib_blackbox",
                "channel_disposition": "emit_budget",
                "budget_required": "yes",
                "clock_relation": "synchronous",
                "budget_model": "interconnect_budget",
                "converted_max": "1.4",
                "converted_min": "0.15",
                "max_source": "manual_arch_budget",
                "min_source": "reviewed_hold_budget",
                "derivation_basis": "manual direct-edge architecture budget",
                "tool_surface": "sta",
                "datapath_only": "yes",
                "min_sign_review": "approved",
                "budget_basis": "owners define resp[0] as a normal SoC interconnect budget",
                "source_type": "manual",
                "apply": "yes",
                "emit_max": "yes",
                "emit_min": "yes",
                "review_status": "approved",
                "owner": "sta_owner",
                "reviewer": "full_chain_reviewer",
                "review_date": "2026-07-15",
            }
            approved += 1
        elif channel in {
            "CH_u_dma_cfg_o_bit0__u_periph_cfg_i_bit0",
            "CH_u_ctrl_ctrl_cfg_o_bit0__u_dpg_cfg_shadow_i_bit0",
        }:
            values = {
                "channel_disposition": "route_to_30",
                "budget_required": "no",
                "apply": "yes",
                "emit_max": "no",
                "emit_min": "no",
                "review_status": "approved",
                "owner": "exception_owner",
                "reviewer": "full_chain_reviewer",
                "review_date": "2026-07-15",
                "disposition_basis": "reviewed exception intent is owned by stage 30",
            }
            routed += 1
        else:
            continue
        for header, value in values.items():
            set_cell(ws, headers, r, header, value)
    require(approved == 1, "expected one 20 budget approval, got %d" % approved)
    require(routed == 2, "expected two 20 route-to-30 approvals, got %d" % routed)
    wb.save(str(form))


def approve_30(d):
    wb = load_workbook(str(d / "30_middle" / "30_harden_to_harden_exception.xlsx"))
    ws = wb["exception_rule"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    approved = set()
    for r in range(2, ws.max_row + 1):
        channel = ws.cell(r, headers["channel_id"]).value or ""
        exception_type = ws.cell(r, headers["exception_type"]).value or ""
        if (
            channel == "CH_u_dma_cfg_o_bit0__u_periph_cfg_i_bit0"
            and "cfg_false" not in approved
            and exception_type in {"false_path", "needs_review"}
        ):
            values = {
                "scenario": "common",
                "stage": "all",
                "corner": "all",
                "apply": "yes",
                "review_status": "approved",
                "owner": "fw_owner",
                "exception_type": "false_path",
                "path_category": "static",
                "source_type": "extracted_harden_exception",
                "harden_clock_context_status": "remapped_equivalent",
                "basis": "cfg bit0 is boot-static after firmware lock in every assembled scenario",
                "risk_level": "medium",
            }
            approved.add("cfg_false")
        elif (
            channel == "CH_u_dpg_fto_0_dma2periph_async_req_bit1__u_periph_async_req_i_bit1"
            and "ft_window" not in approved
        ):
            values = {
                "scenario": "common",
                "stage": "all",
                "corner": "all",
                "apply": "yes",
                "review_status": "approved",
                "owner": "cdc_owner",
                "exception_type": "max_min_delay_override",
                "path_category": "handshake",
                "clock_relation": "asynchronous",
                "max_value": "12.0",
                "min_value": "1.0",
                "datapath_only": "yes",
                "source_type": "manual_entry",
                "harden_clock_context_status": "not_applicable",
                "cdc_rdc_ref": "CDC-HS-001",
                "protocol_ref": "dma2periph async req handshake window",
                "basis": "PrimeTime report_timing and exception report verified 03/30 exception priority keeps this CDC max/min active",
                "risk_level": "high",
            }
            approved.add("ft_window")
        elif (
            channel == "CH_u_ctrl_ctrl_cfg_o_bit0__u_dpg_cfg_shadow_i_bit0"
            and "ctrl_mcp" not in approved
        ):
            values = {
                "scenario": "common",
                "stage": "all",
                "corner": "all",
                "apply": "yes",
                "review_status": "approved",
                "owner": "ctrl_protocol_owner",
                "exception_type": "multicycle_path",
                "path_category": "config",
                "source_type": "manual_entry",
                "harden_clock_context_status": "not_applicable",
                "check_type": "both",
                "src_clock": "u_clkgen_bus_clk_o",
                "dst_clock": "u_clkgen_bus_clk_o",
                "clock_relation": "synchronous",
                "setup_cycles": "2",
                "hold_cycles": "1",
                "mcp_reference": "same_clock_default",
                "cross_clock_mcp_review": "not_applicable",
                "protocol_ref": "CTRL-CFG-2CYCLE",
                "basis": "ctrl_cfg is sampled by a documented two-cycle same-clock protocol",
                "risk_level": "medium",
            }
            approved.add("ctrl_mcp")
        else:
            continue
        for header, value in values.items():
            set_cell(ws, headers, r, header, value)
    require(
        approved == {"cfg_false", "ft_window", "ctrl_mcp"},
        "expected three 30 approvals, got %s" % sorted(approved),
    )
    wb.save(str(d / "30_middle" / "30_harden_to_harden_exception.xlsx"))


def copy_for_stage(d, stage_name):
    stage_dir = d / "stage_snapshots" / stage_name
    stage_dir.mkdir(parents=True, exist_ok=True)
    for name in ("00_middle", "%s_middle" % stage_name, "%s_result" % stage_name):
        source = d / name
        if source.is_dir():
            copytree_replace(source, stage_dir / name)


def copytree_replace(src, dst):
    if dst.exists():
        shutil.rmtree(str(dst))
    shutil.copytree(str(src), str(dst))


def pending_summary(d):
    result = {}
    pending_dir = d / "00_middle" / "scenario" / "common" / "pending"
    for path in sorted(pending_dir.glob("*.ports")):
        lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
        result[path.name] = {"count": len(lines), "ports": lines}
    return result


def removed_summary(d):
    result = {}
    for path in sorted(d.glob("*_middle/scenario/common/removed_log/*.removed")):
        lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
        result["%s/%s" % (path.parts[-5], path.name)] = {"count": len(lines), "sample": lines[:8]}
    return result


def assert_outputs(d):
    expect_files = [
        "01_result/common/01_soc_clocks.sdc",
        "02_result/common/02_soc_clock_timing_prects_ss_125.sdc",
        "03_result/common/03_soc_clock_groups.sdc",
        "04_result/common/04_soc_io_pads.sdc",
        "10_result/common/10_feedthrough.sdc",
        "20_result/common/20_harden_x_if.sdc",
        "30_result/common/30_harden_to_harden_exception.sdc",
    ]
    for rel in expect_files:
        require((d / rel).is_file(), "expected artifact missing: %s" % rel)

    s10 = (d / "10_result/common/10_feedthrough.sdc").read_text(encoding="utf-8")
    require(
        "set_max_delay 0.8 -datapath_only -from [get_pins {u_dma/data_o[0]}] -to [get_pins {u_dpg/fti_0_dma2periph_data[0]}]" in s10,
        "10 ingress data[0] max budget missing",
    )
    require(
        "set_max_delay 0.9 -datapath_only -from [get_pins {u_dpg/fto_0_dma2periph_data[0]}] -to [get_pins {u_periph/data_i[0]}]" in s10,
        "10 egress data[0] max budget missing",
    )
    require("u_dpg/fti_0_dma2periph_data[0]}] -to [get_pins {u_dpg/fto_0_dma2periph_data[0]" not in s10, "10 emitted internal fti->fto constraint")
    require("u_dma/data_o[0]}] -to [get_pins {u_periph/data_i[0]" not in s10, "10 emitted synthetic end-to-end constraint")

    s20 = (d / "20_result/common/20_harden_x_if.sdc").read_text(encoding="utf-8")
    require(
        "set_max_delay 1.4 -datapath_only -from [get_pins {u_periph/resp_o[0]}] -to [get_pins {u_dma/resp_i[0]}]" in s20,
        "20 normal resp[0] max budget missing",
    )
    require(
        "set_min_delay 0.15 -datapath_only -from [get_pins {u_periph/resp_o[0]}] -to [get_pins {u_dma/resp_i[0]}]" in s20,
        "20 normal resp[0] min budget missing",
    )
    require("fti_" not in s20 and "fto_" not in s20, "20 emitted a feedthrough-owned edge or stitched path")

    s30 = (d / "30_result/common/30_harden_to_harden_exception.sdc").read_text(encoding="utf-8")
    require(
        "set_false_path -from [get_pins {u_dma/cfg_o[0]}] -to [get_pins {u_periph/cfg_i[0]}]" in s30,
        "30 cfg false_path missing",
    )
    require("set_multicycle_path 2 -setup" in s30 and "set_multicycle_path 1 -hold" in s30, "30 two-cycle MCP pair missing")
    require(
        "set_max_delay 12 -datapath_only -from [get_pins {u_dpg/fto_0_dma2periph_async_req[1]}] -to [get_pins {u_periph/async_req_i[1]}]" in s30,
        "30 feedthrough egress max override missing",
    )
    require("set_min_delay 1 -datapath_only" in s30, "30 feedthrough egress min override missing")
    require("-through" not in s30, "30 used through_collection to cross a harden interior")

    with (d / "10_middle/scenario/common/feedthrough_edge_inventory.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as file_obj:
        ten_rows = list(csv.DictReader(file_obj))
    require(len(ten_rows) == 16, "10 should own exactly sixteen feedthrough-adjacent bit edges")
    require(
        all(row["feedthrough_edge_id"] == "FTE_" + row["connection_id"] for row in ten_rows),
        "10 feedthrough edge IDs are not stable",
    )
    with (d / "20_middle/scenario/common/channel_inventory.csv").open("r", encoding="utf-8-sig", newline="") as file_obj:
        twenty_rows = list(csv.DictReader(file_obj))
    require(
        all("fti_" not in row.get("channel_id", "") and "fto_" not in row.get("channel_id", "") for row in twenty_rows),
        "20 channel inventory contains a 10-owned feedthrough edge",
    )

    summary = pending_summary(d)
    require(summary["u_clkgen.ports"]["count"] >= 1, "u_clkgen pending should keep non-covered leftovers")
    require("output data_o[0]" not in "\n".join(summary["u_dma.ports"]["ports"]), "u_dma data_o[0] should be consumed by 10")
    require("output cfg_o[0]" not in "\n".join(summary["u_dma.ports"]["ports"]), "u_dma cfg_o[0] should be consumed")
    require("input resp_i[0]" not in "\n".join(summary["u_dma.ports"]["ports"]), "u_dma resp_i[0] should be consumed by 20")
    require("output fto_0_dma2periph_async_req[1]" not in "\n".join(summary["u_dpg.ports"]["ports"]), "u_dpg fto async[1] should be consumed by 30")
    require("input cfg_shadow_i[0]" not in "\n".join(summary["u_dpg.ports"]["ports"]), "u_dpg cfg_shadow_i[0] should be consumed by 30")
    require("input cfg_i[0]" not in "\n".join(summary["u_periph.ports"]["ports"]), "u_periph cfg_i[0] should be consumed by 30")


def main():
    if WORK.exists():
        shutil.rmtree(str(WORK))
    write_inputs(WORK)

    run_ok([EX00, "--run-root", WORK, "--scenario", "common"], WORK)

    run_ok([EX01, "--run-root", WORK, "--scenario", "common"], WORK)
    copy_for_stage(WORK, "01")

    command_02 = [
        EX02,
        "--run-root",
        WORK,
        "--scenario",
        "common",
        "--stage",
        "prects",
        "--corner",
        "ss_125",
    ]
    run_sync(command_02, WORK)
    approve_02(WORK, "prects", "ss_125")
    run_ok(command_02, WORK)
    copy_for_stage(WORK, "02")

    command_03 = [EX03, "--run-root", WORK, "--scenario", "common"]
    run_sync(command_03, WORK)
    approve_03(WORK)
    run_ok(command_03, WORK)
    copy_for_stage(WORK, "03")

    command_04 = [EX04, "--run-root", WORK, "--scenario", "common"]
    run_sync(command_04, WORK)
    approve_04(WORK)
    run_ok(command_04, WORK)
    copy_for_stage(WORK, "04")

    command_10 = [EX10, "--run-root", WORK, "--scenario", "common"]
    run_sync(command_10, WORK)
    approve_10(WORK)
    run_ok(command_10, WORK)
    copy_for_stage(WORK, "10")

    command_20 = [
        EX20,
        "--run-root",
        WORK,
        "--scenario",
        "common",
        "--mode",
        "budget_output",
        "--stage",
        "all",
        "--corner",
        "all",
    ]
    run_sync(command_20, WORK)
    approve_20(WORK)
    run_sync(command_20, WORK)
    run_ok(command_20, WORK)
    copy_for_stage(WORK, "20")

    command_30 = [EX30, "--run-root", WORK, "--scenario", "common"]
    run_sync(command_30, WORK)
    approve_30(WORK)
    run_ok(command_30, WORK)
    copy_for_stage(WORK, "30")

    assert_outputs(WORK)
    assembled = WORK / "assembled" / "common_prects_ss_125_01_to_30.sdc"
    assembled.parent.mkdir(parents=True, exist_ok=True)
    source_paths = [
        "01_result/common/01_soc_clocks.sdc",
        "02_result/common/02_soc_clock_timing_prects_ss_125.sdc",
        "03_result/common/03_soc_clock_groups.sdc",
        "04_result/common/04_soc_io_pads.sdc",
        "10_result/common/10_feedthrough.sdc",
        "20_result/common/20_harden_x_if.sdc",
        "30_result/common/30_harden_to_harden_exception.sdc",
    ]
    assembled.write_text(
        "# 01 -> 30 complex SoC top synthesis SDC source preview\n\n"
        + "\n".join("source %s" % path for path in source_paths)
        + "\n",
        encoding="utf-8",
    )
    summary = {
        "case_dir": str(WORK),
        "status": "PASS",
        "topology": "five hardens with clocks, pads, feedthrough direct edges, normal interfaces, and exceptions",
        "compatibility_note": "20 runs directly in target budget_output mode",
        "pending": pending_summary(WORK),
        "removed": removed_summary(WORK),
        "key_artifacts": source_paths,
        "assembled_preview": str(assembled.relative_to(WORK)),
    }
    (WORK / "full_chain_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print("00->30 full-chain complex case: PASS")
    print("Case dir:", WORK)
    print("Preview :", assembled)
    print("Summary :", WORK / "full_chain_summary.json")


if __name__ == "__main__":
    main()
