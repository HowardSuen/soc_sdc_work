#!/usr/bin/env python3
"""Regression for the workbook-centric single-run stage-20 runtime."""

from __future__ import print_function

import csv
import importlib.util
import json
import os
import re
import shutil
import socket
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
STAGE20 = BASE.parent
SCRIPTS = STAGE20.parent
EX20 = STAGE20 / "20_extract_harden_x_if.py"
T10_REGRESSION = SCRIPTS / "10_feedthrough" / "regression_test" / "run_regression.py"
WORK = BASE / "work_latest"


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def load_t10_fixture():
    spec = importlib.util.spec_from_file_location("stage10_fixture_for_20", str(T10_REGRESSION))
    require(spec is not None and spec.loader is not None, "cannot load stage-10 fixture")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


T10 = load_t10_fixture()


def run_20(root, *extra):
    return T10.run_script(EX20, ["--run-root", root] + list(extra), root)


def prepare_upstreams(root, missing_sdcs=None, require_20=True):
    if require_20:
        T10.build_root(root, missing_sdcs=missing_sdcs)
    else:
        T10.clean_dir(root)
        T10.write_inputs(root, missing_sdcs=missing_sdcs)
        required_path = root / "inputs" / "required_views.csv"
        required_rows = read_csv(required_path)
        require(required_rows, "required view fixture is empty")
        required_rows[0]["require_20"] = "no"
        T10.write_csv(required_path, list(required_rows[0]), required_rows)
        initialized = T10.run_00(root)
        require(initialized.returncode == 0, "00 no-required-20 fixture initialization failed")
        T10.publish_upstream_artifacts(root)
    first = T10.run_10(root)
    require(first.returncode != 0, "stage 10 first sync did not stop")
    form = root / "10_middle" / "10_feedthrough.xlsx"
    specifications = {}
    for row in T10.inventory_rows(root):
        specifications[T10.edge_key(row)] = T10.approved(
            "no_soc_budget_required",
            independent="regression upstream owner basis independent of partial SDC",
        )
    T10.approve_edges(form, specifications)
    second = T10.run_10(root)
    require(
        second.returncode == 0,
        "stage 10 fixture completion failed:\n%s\n%s" % (second.stdout, second.stderr),
    )


def prepare_upstreams_with_normal_exception(root):
    T10.clean_dir(root)
    T10.write_inputs(root)
    source_sdc = root / "inputs" / "u_src.sdc"
    source_sdc.write_text(
        source_sdc.read_text(encoding="utf-8")
        + "set_false_path -from [get_ports {normal_o}]\n",
        encoding="utf-8",
    )
    initialized = T10.run_00(root)
    require(initialized.returncode == 0, "00 exception fixture initialization failed")
    T10.publish_upstream_artifacts(root)
    first = T10.run_10(root)
    require(first.returncode != 0, "stage 10 exception fixture first sync did not stop")
    form = root / "10_middle" / "10_feedthrough.xlsx"
    T10.approve_edges(
        form,
        {
            T10.edge_key(row): T10.approved("no_soc_budget_required")
            for row in T10.inventory_rows(root)
        },
    )
    second = T10.run_10(root)
    require(second.returncode == 0, "stage 10 exception fixture failed")


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        return list(csv.DictReader(file_obj))


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def report_text(root, stage="synth", corner="ss"):
    suffix = "" if (stage, corner) == ("all", "all") else "_%s_%s" % (stage, corner)
    path = root / "20_result" / "reports" / ("harden_x_if_check_report%s.txt" % suffix)
    require(path.is_file(), "20 report missing: %s" % path)
    return path.read_text(encoding="utf-8")


def port_bytes(root):
    return {
        path.name: path.read_bytes()
        for path in sorted((root / "inputs").glob("port_*.xlsx"))
    }


def used_state(root, sheet_name, direction, port_name):
    return T10.used_state(root, sheet_name, direction, port_name)


def inventory_rows(root):
    return read_csv(root / "20_middle" / "channel_inventory.csv")


def current_review_rows(form, stage="synth", corner="ss"):
    workbook = load_workbook(str(form), data_only=False)
    sheet = workbook["interface_budget"]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    rows = [
        row_idx for row_idx in range(2, sheet.max_row + 1)
        if str(sheet.cell(row_idx, headers["stage"]).value or "") == stage
        and str(sheet.cell(row_idx, headers["corner"]).value or "") == corner
    ]
    return workbook, sheet, headers, rows


def approve_20(form, updates, stage="synth", corner="ss"):
    workbook, sheet, headers, rows = current_review_rows(form, stage, corner)
    require(rows, "20 review workbook has no current-view rows")
    for row_idx in rows:
        for name, value in updates.items():
            require(name in headers, "20 workbook missing review field %s" % name)
            sheet.cell(row_idx, headers[name], value)
        sheet.cell(
            row_idx,
            headers["approved_machine_digest"],
            sheet.cell(row_idx, headers["machine_digest"]).value,
        )
    workbook.save(str(form))
    workbook.close()


def approved_budget(independent=""):
    return {
        "channel_disposition": "emit_budget",
        "budget_required": "yes",
        "budget_model": "manual_budget",
        "converted_max": "1.2",
        "max_source": "architecture_review",
        "derivation_basis": "reviewed physical interconnect allocation",
        "tool_surface": "sta",
        "datapath_only": "yes",
        "apply": "yes",
        "emit_max": "yes",
        "emit_min": "no",
        "review_status": "approved",
        "owner": "soc_timing_owner",
        "reviewer": "soc_timing_reviewer",
        "review_date": "2026-07-17",
        "budget_basis": "approved normal interface interconnect budget",
        "relationship_override_basis": "reviewed budget is valid without synchronous clock inference",
        "sdc_independent_basis": independent,
    }


def run_lifecycle_contract():
    root = WORK / "lifecycle"
    prepare_upstreams(root)
    before = port_bytes(root)
    args = ("--mode", "audit_only", "--stage", "synth", "--corner", "ss")
    first = run_20(root, *args)
    require(first.returncode != 0, "first 20 sync did not require review publication")
    require((root / "20_middle" / "20_harden_x_if.xlsx").is_file(), "review workbook missing")
    require((root / "20_middle" / "channel_inventory.csv").is_file(), "review inventory missing")
    require(port_bytes(root) == before, "review synchronization changed port workbooks")
    require(not (root / "20_result" / "20_harden_x_if_synth_ss.sdc").exists(), "review sync wrote formal SDC")
    require(not (root / "20_middle" / "port_accounting_delta.csv").exists(), "review sync wrote delta")

    second = run_20(root, *args)
    require(second.returncode == 0, "approved audit run failed:\n%s\n%s" % (second.stdout, second.stderr))
    rows = inventory_rows(root)
    require(len(rows) == 1, "normal inventory should contain exactly one non-clock/pad/feedthrough channel")
    row = rows[0]
    require(re.fullmatch(r"CONN_[0-9a-f]{64}", row["connection_id"]), "connection ID is not full SHA-256")
    require(row["channel_id"] == "CH_" + row["connection_id"][5:], "channel ID does not share canonical hash")
    require(row["src_port"] == "normal_o" and row["dst_port"] == "normal_i", "wrong normal channel classified")
    require(row["channel_disposition"] == "no_soc_budget_required", "audit policy disposition missing")
    sdc = (root / "20_result" / "20_harden_x_if_synth_ss.sdc").read_text(encoding="utf-8")
    require("set_max_delay" not in sdc and "set_min_delay" not in sdc, "audit emitted timing command")
    require(used_state(root, "u_src", "output", "normal_o") == "0", "source bit was not accounted")
    require(used_state(root, "u_dst", "input", "normal_i") == "0", "destination bit was not accounted")
    delta = read_csv(root / "20_middle" / "port_accounting_delta.csv")
    require(delta and all(item["owner_object_id"] == row["channel_id"] for item in delta), "delta owner is not channel_id")
    view_completion = read_json(root / "20_middle" / "completion" / "synth_ss.meta")
    run_completion = read_json(root / "20_middle" / "stage_completion.meta")
    require(view_completion["completion_status"] == "complete", "view completion is not complete")
    require(run_completion["completion_status"] == "complete", "run-wide completion is not complete")
    require("Mode: audit_only" in second.stdout, "stdout mode metadata missing")
    require("accounting digest before" in report_text(root).lower(), "coverage/report digest metadata missing")

    third = run_20(root, *args)
    require(third.returncode == 0, "idempotent rerun failed")
    require(used_state(root, "u_src", "output", "normal_o") == "0", "idempotent rerun changed source state")

    protected_paths = [
        root / "20_result" / "20_harden_x_if_synth_ss.sdc",
        root / "20_middle" / "completion" / "synth_ss.meta",
        root / "20_middle" / "stage_completion.meta",
        root / "20_middle" / "20_harden_x_if.xlsx",
    ]
    protected_before = {str(path): path.read_bytes() for path in protected_paths}
    undeclared = run_20(
        root, "--mode", "audit_only", "--stage", "postcts", "--corner", "ff_0"
    )
    require(undeclared.returncode != 0, "undeclared formal 20 view was accepted")
    protected_after = {str(path): path.read_bytes() for path in protected_paths}
    require(
        protected_before == protected_after,
        "undeclared 20 view changed a completed required view or run-wide completion",
    )
    require(
        not (root / "20_middle" / "completion" / "postcts_ff_0.meta").exists(),
        "undeclared 20 view wrote a per-view completion",
    )


def run_budget_contract():
    root = WORK / "budget"
    prepare_upstreams(root)
    args = ("--mode", "budget_output", "--stage", "synth", "--corner", "ss")
    first = run_20(root, *args)
    require(first.returncode != 0, "budget first sync did not stop")
    form = root / "20_middle" / "20_harden_x_if.xlsx"
    unreviewed = run_20(root, *args)
    require(unreviewed.returncode != 0, "pending budget review was published as complete")
    require(
        not (root / "20_result" / "20_harden_x_if_synth_ss.sdc").exists(),
        "pending budget review wrote formal SDC",
    )
    require(
        not (root / "20_middle" / "port_accounting_delta.csv").exists(),
        "pending budget review wrote accounting delta",
    )
    approve_20(form, approved_budget())
    second = run_20(root, *args)
    require(second.returncode == 0, "approved budget failed:\n%s\n%s" % (second.stdout, second.stderr))
    output = root / "20_result" / "20_harden_x_if_synth_ss.sdc"
    commands = [line for line in output.read_text(encoding="utf-8").splitlines() if line.startswith("set_")]
    require(len(commands) == 1, "budget output command count mismatch: %s" % commands)
    require("u_src/normal_o" in commands[0] and "u_dst/normal_i" in commands[0], "budget command endpoints are wrong")
    require("-datapath_only" in commands[0], "datapath strategy was lost")

    stale = load_workbook(str(form), data_only=False)
    sheet = stale["interface_budget"]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    sheet.cell(2, headers["approved_machine_digest"], "0" * 64)
    stale.save(str(form))
    stale.close()
    before = port_bytes(root)
    blocked = run_20(root, *args)
    require(blocked.returncode != 0, "stale approved machine digest was accepted")
    require(port_bytes(root) == before, "failed review validation changed accounting")


def run_cli_diagnostic_contract():
    missing = T10.run_script(EX20, [], BASE)
    require(missing.returncode != 0, "20 accepted invocation without --run-root")
    root = WORK / "cli"
    prepare_upstreams(root)
    obsolete = run_20(root, "--scenario", "common")
    require(obsolete.returncode != 0, "20 accepted obsolete --scenario")
    before = port_bytes(root)
    diagnostic = run_20(
        root, "--mode", "audit_only", "--stage", "postroute",
        "--corner", "ff", "--diagnose-only",
    )
    require(diagnostic.returncode == 0, "non-required diagnostic run failed")
    require(port_bytes(root) == before, "diagnostic mode changed workbooks")
    require(not (root / "20_result" / "20_harden_x_if_postroute_ff.sdc").exists(), "diagnostic wrote formal SDC")
    text = report_text(root, "postroute", "ff").lower()
    require("port accounting: diagnostic/read-only" in text, "diagnostic accounting metadata missing")
    require("accounting closure: not evaluated" in text, "diagnostic closure metadata missing")


def run_upstream_and_lock_contract():
    root = WORK / "stale_10"
    prepare_upstreams(root)
    completion_path = root / "10_middle" / "stage_completion.meta"
    payload = read_json(completion_path)
    payload["accounting_digest_after"] = "0" * 64
    completion_path.write_text(json.dumps(payload, sort_keys=True, indent=2) + "\n", encoding="utf-8")
    blocked = run_20(root, "--mode", "audit_only", "--stage", "synth", "--corner", "ss")
    require(blocked.returncode != 0, "stale 10 completion was accepted")
    require(not (root / "20_result" / "20_harden_x_if_synth_ss.sdc").exists(), "stale upstream wrote SDC")

    lock_root = WORK / "lock"
    prepare_upstreams(lock_root)
    lock = lock_root / "inputs" / ".port_accounting.lock"
    lock.write_text(json.dumps({"pid": os.getpid(), "host": socket.gethostname(), "stage": "test"}) + "\n", encoding="utf-8")
    locked = run_20(lock_root, "--mode", "audit_only", "--stage", "synth", "--corner", "ss")
    require(locked.returncode != 0, "active accounting lock was ignored")
    lock.unlink()

    final_root = WORK / "final_token"
    prepare_upstreams(final_root)
    T10.set_used_state(final_root, "u_src", "output", "normal_o", "ALL USED")
    final = run_20(final_root, "--mode", "audit_only", "--stage", "synth", "--corner", "ss")
    require(final.returncode != 0, "20 accepted final accounting token")


def run_partial_sdc_contract():
    root = WORK / "partial"
    prepare_upstreams(root, missing_sdcs={"u_dst"})
    audit_args = ("--mode", "audit_only", "--stage", "synth", "--corner", "ss")
    require(run_20(root, *audit_args).returncode != 0, "partial audit first sync did not stop")
    allowed = run_20(root, *audit_args)
    require(allowed.returncode == 0, "SDC-independent audit policy failed")
    rows = inventory_rows(root)
    require(rows[0]["evidence_status"] == "incomplete_missing_sdc", "partial evidence status missing")
    require(rows[0]["sdc_independent_basis"], "audit independent basis missing")

    strict_root = WORK / "partial_strict"
    prepare_upstreams(strict_root, missing_sdcs={"u_dst"})
    strict = run_20(
        strict_root, "--mode", "budget_output", "--stage", "synth",
        "--corner", "ss", "--require-complete-harden-sdc",
    )
    require(strict.returncode != 0, "strict missing-SDC run was accepted")
    require(not (strict_root / "20_result" / "20_harden_x_if_synth_ss.sdc").exists(), "strict failure wrote SDC")


def run_exception_route_contract():
    root = WORK / "route_to_30"
    prepare_upstreams_with_normal_exception(root)
    args = ("--mode", "audit_only", "--stage", "synth", "--corner", "ss")
    require(run_20(root, *args).returncode != 0, "exception route first sync did not stop")
    rows = inventory_rows(root)
    require(len(rows) == 1, "exception route normal channel missing")
    require(rows[0]["channel_disposition"] == "route_to_30", "known exception was not routed to 30")
    require(rows[0]["review_status"] == "approved", "automatic exception evidence was not reviewed")
    second = run_20(root, *args)
    require(second.returncode == 0, "approved route_to_30 run failed")
    require(used_state(root, "u_src", "output", "normal_o") == "", "route_to_30 source was incorrectly accounted by 20")
    require(used_state(root, "u_dst", "input", "normal_i") == "", "route_to_30 destination was incorrectly accounted by 20")


def run_runwide_audit_contract():
    root = WORK / "runwide_audit"
    prepare_upstreams(root, require_20=False)
    args = ("--mode", "audit_only")
    first = run_20(root, *args)
    require(first.returncode != 0, "run-wide audit first sync did not stop")
    second = run_20(root, *args)
    require(second.returncode == 0, "run-wide audit completion failed")
    require((root / "20_middle" / "channel_inventory.csv").is_file(), "run-wide inventory missing")
    require((root / "20_middle" / "port_accounting_delta.csv").is_file(), "run-wide accounting delta missing")
    require(
        read_json(root / "20_middle" / "stage_completion.meta")["completion_status"] == "complete",
        "run-wide audit completion is not complete",
    )
    require(
        not (root / "20_result" / "20_harden_x_if.sdc").exists(),
        "run with no require_20=yes view wrote a formal SDC",
    )
    require(
        not (root / "20_middle" / "completion" / "all_all.meta").exists(),
        "run with no require_20=yes view wrote a view completion",
    )
    invalid_view = run_20(root, "--mode", "audit_only", "--stage", "synth", "--corner", "ss")
    require(invalid_view.returncode != 0, "non-diagnostic view was accepted without require_20=yes")


def main():
    if WORK.exists():
        shutil.rmtree(str(WORK))
    WORK.mkdir(parents=True)
    run_lifecycle_contract()
    run_budget_contract()
    run_cli_diagnostic_contract()
    run_upstream_and_lock_contract()
    run_partial_sdc_contract()
    run_exception_route_contract()
    run_runwide_audit_contract()
    print("20 harden-x-if latest-runtime regression: PASS")
    print("  cases: lifecycle, budget, cli_diagnostic, upstream_lock, partial_sdc, route_to_30, runwide_audit")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
