#!/usr/bin/env python3
"""Audit SoC harden interfaces and optionally emit reviewed channel budgets."""

import argparse
import csv
import hashlib
import importlib.util
import io
import json
import math
import os
import re
import shutil
import socket
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

try:
    from dataclasses import dataclass, field
except ImportError:  # pragma: no cover - Python 3.6 compatibility path
    class _CompatField:
        def __init__(self, default_factory=None):
            self.default_factory = default_factory

    def field(default_factory=None):
        return _CompatField(default_factory=default_factory)

    def dataclass(_cls=None, frozen=False):
        def wrap(cls):
            annotations = getattr(cls, "__annotations__", {})
            names = list(annotations.keys())
            defaults = {}
            factories = {}
            for name in names:
                if hasattr(cls, name):
                    value = getattr(cls, name)
                    if isinstance(value, _CompatField):
                        factories[name] = value.default_factory
                        delattr(cls, name)
                    else:
                        defaults[name] = value

            def __init__(self, *args, **kwargs):
                if len(args) > len(names):
                    raise TypeError("too many positional arguments")
                values = {}
                for name, value in zip(names, args):
                    values[name] = value
                for name in names[len(args):]:
                    if name in kwargs:
                        values[name] = kwargs.pop(name)
                    elif name in factories:
                        values[name] = factories[name]()
                    elif name in defaults:
                        values[name] = defaults[name]
                    else:
                        raise TypeError("missing required argument: " + name)
                if kwargs:
                    raise TypeError("unexpected argument: " + sorted(kwargs)[0])
                for name in names:
                    object.__setattr__(self, name, values[name])

            def __eq__(self, other):
                if other.__class__ is not cls:
                    return False
                return all(getattr(self, name) == getattr(other, name) for name in names)

            cls.__init__ = __init__
            cls.__eq__ = __eq__
            if frozen:
                def __hash__(self):
                    return hash(tuple(getattr(self, name) for name in names))
                cls.__hash__ = __hash__
            else:
                cls.__hash__ = None
            return cls

        if _cls is None:
            return wrap
        return wrap(_cls)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - user environment guard
    print("ERROR: openpyxl is required to read/write 20 harden IF xlsx files.", file=sys.stderr)
    raise SystemExit(2) from exc


SCHEMA_VERSION = "1"
TARGET_SCHEMA_VERSION = "1.0"
SCENARIOS = {"common", "func", "scan", "mbist", "gpio_in", "gpio_out"}
STAGES = {"all", "synth", "prects", "postcts", "postroute"}
MODES = {"audit_only", "budget_output"}
DISPOSITIONS = {"emit_budget", "no_soc_budget_required", "route_to_30", "not_applicable", "pending"}
POLICY_ID = "project_default_pr_adjacent_no_harden_if_delay_v1"
SDC_INDEPENDENT_POLICY = "project_pr_adjacent_policy_independent_of_block_sdc_v1"
POLICY_OWNER = "soc_sdc_methodology"
POLICY_REVIEWER = "policy_owner"
POLICY_REVIEW_DATE = "2026-07-15"
PSEUDO_INSTANCES = {"top", "fabric", "unknown", "constant", "const", "nc"}
CHANNEL_TYPES_20 = {"harden_to_harden", "fabric_to_harden", "harden_to_fabric"}
NON_20_CHANNEL_TYPES = {
    "top_pad_to_harden",
    "harden_to_top_pad",
    "pad_to_pad",
    "clock_connection",
    "feedthrough",
    "feedthrough_candidate",
    "exception_path",
    "constant_tie",
    "no_connect",
    "unknown",
}
CANONICAL_CONNECTION_TYPES = CHANNEL_TYPES_20 | NON_20_CHANNEL_TYPES
TIMING_MODELS = {"", "visible_netlist", "lib_blackbox", "abstract_model", "unknown"}
BUDGET_MODELS = {"", "interconnect_budget", "clock_relative_io_delay", "manual_budget", "unknown"}
TOOL_SURFACES = {"", "sta", "dc", "both"}
YES_NO = {"", "yes", "no"}
REVIEW_STATUS_VALUES = {"", "pending", "approved", "rejected"}
SOURCE_TYPES = {"", "extracted", "manual", "na"}
MIN_REVIEW_VALUES = {"", "approved", "reviewed", "yes", "waived"}
CLOCK_RELATION_CANONICAL = {"synchronous", "asynchronous", "logically_exclusive", "physically_exclusive", "unknown"}
CLOCK_RELATION_ALIASES = {
    "": "",
    "sync": "synchronous",
    "synchronous": "synchronous",
    "async": "asynchronous",
    "asynchronous": "asynchronous",
    "logically_exclusive": "logically_exclusive",
    "logically exclusive": "logically_exclusive",
    "physically_exclusive": "physically_exclusive",
    "physically exclusive": "physically_exclusive",
    "unknown": "unknown",
}
RELATION_BLOCKING = {"asynchronous", "logically_exclusive", "physically_exclusive"}
RELATION_SOURCES = {"explicit_rule", "default_synchronous"}
ACTIVE_01_ACTIONS = {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}
PORT_BIT_RE = re.compile(r"^[^\s\[\]]+(?:\[\d+\])?$")
PORT_RANGE_RE = re.compile(r"^(.+)\[(\d+)\s*:\s*(\d+)\]$")
PORT_EXACT_BIT_RE = re.compile(r"^(.+)\[(\d+)\]$")
MATCHED_STATUSES = {"", "matched", "ok", "valid"}

TARGET20_HEADERS = [
    "schema_version", "run_id", "mode_label", "design_revision", "mode",
    "run_completeness", "structure_digest", "accounting_digest_before",
    "accounting_digest_after", "channel_id", "connection_id", "stage",
    "corner", "channel_type", "owner_stage", "src_instance", "src_direction",
    "src_port", "src_bit_index", "src_endpoint", "src_sdc_status",
    "dst_instance", "dst_direction", "dst_port", "dst_bit_index",
    "dst_endpoint", "dst_sdc_status", "source_workbook", "source_sheet",
    "source_row", "is_pad_related", "is_clock_related", "is_feedthrough",
    "evidence_status", "original_src_clock", "original_dst_clock",
    "original_src_max", "original_src_min", "original_dst_max",
    "original_dst_min", "complex_options", "source_sdc_file", "source_line",
    "source_digest", "source_command", "clock_relation", "machine_digest",
    "approved_machine_digest", "timing_model", "channel_disposition",
    "budget_required", "budget_model", "converted_max", "converted_min",
    "max_source", "min_source", "derivation_basis", "tool_surface",
    "datapath_only", "min_sign_review", "budget_basis", "source_type",
    "apply", "emit_max", "emit_min", "review_status", "owner", "reviewer",
    "review_date", "disposition_basis", "sdc_independent_basis",
    "relationship_override_basis", "validation_status", "note",
]

TARGET20_MACHINE_FIELDS = {
    "schema_version", "run_id", "mode_label", "design_revision", "mode",
    "run_completeness", "structure_digest", "accounting_digest_before",
    "accounting_digest_after", "channel_id", "connection_id", "stage", "corner",
    "channel_type", "owner_stage", "src_instance", "src_direction", "src_port",
    "src_bit_index", "src_endpoint", "src_sdc_status", "dst_instance",
    "dst_direction", "dst_port", "dst_bit_index", "dst_endpoint",
    "dst_sdc_status", "source_workbook", "source_sheet", "source_row",
    "is_pad_related", "is_clock_related", "is_feedthrough", "evidence_status",
    "original_src_clock", "original_dst_clock", "original_src_max",
    "original_src_min", "original_dst_max", "original_dst_min", "complex_options",
    "source_sdc_file", "source_line", "source_digest", "source_command",
    "clock_relation", "machine_digest", "validation_status",
}

TARGET20_REVIEW_INVALIDATING_FIELDS = TARGET20_MACHINE_FIELDS - {
    "accounting_digest_before", "accounting_digest_after", "machine_digest",
}

TARGET20_TERMINAL_DISPOSITIONS = {
    "emit_budget", "no_soc_budget_required", "not_applicable",
}

CHANNEL_HEADERS = [
    "schema_version",
    "author",
    "mode",
    "sdc_consumption",
    "run_completeness",
    "available_harden_count",
    "missing_harden_count",
    "not_required_harden_count",
    "missing_instances",
    "connection_inventory_digest",
    "channel_id",
    "connection_id",
    "scenario",
    "stage",
    "corner",
    "channel_type",
    "owner_stage",
    "src_instance",
    "src_module",
    "src_direction",
    "src_port",
    "src_bit_index",
    "src_endpoint",
    "src_sdc_status",
    "dst_instance",
    "dst_module",
    "dst_direction",
    "dst_port",
    "dst_bit_index",
    "dst_endpoint",
    "dst_sdc_status",
    "connection_source",
    "is_pad_related",
    "is_clock_related",
    "is_feedthrough",
    "evidence_status",
    "budget_required",
    "clock_relation",
    "channel_disposition",
    "budget_type",
    "budget_model",
    "apply",
    "review_status",
    "emit_max",
    "emit_min",
    "converted_max",
    "converted_min",
    "disposition_basis",
    "sdc_independent_basis",
    "note",
]

BUDGET_HEADERS = [
    "channel_id",
    "connection_id",
    "scenario",
    "stage",
    "corner",
    "channel_type",
    "is_pad_related",
    "is_clock_related",
    "is_feedthrough",
    "src_endpoint",
    "dst_endpoint",
    "src_sdc_status",
    "dst_sdc_status",
    "evidence_status",
    "timing_model",
    "budget_required",
    "clock_relation",
    "channel_disposition",
    "budget_model",
    "src_output_delay_max",
    "src_output_delay_min",
    "dst_input_delay_max",
    "dst_input_delay_min",
    "converted_max",
    "converted_min",
    "max_source",
    "min_source",
    "derivation_basis",
    "original_src_clock",
    "original_dst_clock",
    "soc_clock",
    "complex_options",
    "tool_surface",
    "datapath_only",
    "min_sign_review",
    "budget_basis",
    "disposition_basis",
    "sdc_independent_basis",
    "relationship_override_basis",
    "source_type",
    "source_sdc_file",
    "source_line",
    "source_digest",
    "extraction_time",
    "source_command",
    "apply",
    "emit_max",
    "emit_min",
    "review_status",
    "owner",
    "reviewer",
    "review_date",
    "note",
]

BUDGET_REVIEW_FIELDS = {
    "channel_disposition",
    "budget_required",
    "budget_model",
    "converted_max",
    "converted_min",
    "max_source",
    "min_source",
    "derivation_basis",
    "tool_surface",
    "datapath_only",
    "min_sign_review",
    "budget_basis",
    "disposition_basis",
    "sdc_independent_basis",
    "relationship_override_basis",
    "source_type",
    "apply",
    "emit_max",
    "emit_min",
    "review_status",
    "owner",
    "reviewer",
    "review_date",
    "note",
}
BUDGET_MACHINE_FIELDS = [header for header in BUDGET_HEADERS if header not in BUDGET_REVIEW_FIELDS]
REVIEW_INVALIDATING_FIELDS = {
    "connection_id",
    "channel_type",
    "src_endpoint",
    "dst_endpoint",
    "src_sdc_status",
    "dst_sdc_status",
    "evidence_status",
    "source_digest",
    "source_command",
}

LOG_HEADERS = [
    "source_sdc_file",
    "source_line",
    "instance",
    "port",
    "direction",
    "constraint_type",
    "clock_name",
    "min_value",
    "max_value",
    "parse_status",
    "channel_ids",
    "source_digest",
    "extraction_time",
    "original_command",
    "message",
]

HEADER_FILL = PatternFill("solid", fgColor="215967")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C6CC"),
    right=Side(style="thin", color="B8C6CC"),
    top=Side(style="thin", color="B8C6CC"),
    bottom=Side(style="thin", color="B8C6CC"),
)


@dataclass
class PortInfo:
    name: str
    width: str = ""
    used_width: str = ""
    from_whom: str = ""
    to_top: str = ""
    connectivity: str = ""
    inout_name: str = ""


@dataclass
class InstInfo:
    module_name: str
    inst_name: str
    owner: str = ""
    file_path: str = ""
    sdc_hint: str = ""
    sdc_path: Optional[Path] = None
    sdc_status: str = "missing"
    sdc_note: str = ""
    inputs: Dict[str, PortInfo] = field(default_factory=dict)
    outputs: Dict[str, PortInfo] = field(default_factory=dict)
    inouts: Dict[str, PortInfo] = field(default_factory=dict)


@dataclass
class TclCommand:
    raw: str
    line_no: int
    parse_complete: bool = True


@dataclass
class ClockInfo:
    clock_name: str
    direct_source: str = ""
    producer_object: str = ""
    final_action: str = ""


@dataclass
class ClockContext:
    clocks: Dict[str, ClockInfo] = field(default_factory=dict)
    aliases: Dict[Tuple[str, str], str] = field(default_factory=dict)
    objects: Set[str] = field(default_factory=set)
    status: str = "unavailable"
    inventory_digest: str = ""
    meta_digest: str = ""
    clock_set_digest: str = ""


@dataclass
class RelationContext:
    relations: Dict[Tuple[str, str], str] = field(default_factory=dict)
    status: str = "unavailable"


@dataclass
class ChannelRecord:
    channel_id: str
    scenario: str
    stage: str
    corner: str
    channel_type: str
    connection_id: str
    src_instance: str
    src_module: str
    src_direction: str
    src_port: str
    src_bit_index: str
    src_endpoint: str
    dst_instance: str
    dst_module: str
    dst_direction: str
    dst_port: str
    dst_bit_index: str
    dst_endpoint: str
    connection_source: str
    is_pad_related: str = "no"
    is_clock_related: str = "no"
    is_feedthrough: str = "no"
    src_sdc_status: str = "not_required"
    dst_sdc_status: str = "not_required"
    evidence_status: str = "complete"
    timing_model: str = "unknown"
    budget_required: str = ""
    clock_relation: str = ""
    channel_disposition: str = "pending"
    note: str = ""


@dataclass(frozen=True)
class PortKey:
    inst_name: str
    direction: str
    port_name: str


@dataclass
class ConnectionEdge:
    connection_id: str
    connection_type: str
    src_instance: str
    src_direction: str
    src_port: str
    src_bit_index: str = ""
    src_endpoint_key: str = ""
    src_soc_object: str = ""
    dst_instance: str = ""
    dst_direction: str = ""
    dst_port: str = ""
    dst_bit_index: str = ""
    dst_endpoint_key: str = ""
    dst_soc_object: str = ""
    validation_status: str = ""
    scenario_scope: str = "common"
    note: str = ""


@dataclass
class ConnectionIndex:
    edges: List[ConnectionEdge] = field(default_factory=list)
    by_dst: Dict[Tuple[str, str], List[ConnectionEdge]] = field(default_factory=lambda: defaultdict(list))
    by_src: Dict[Tuple[str, str], List[ConnectionEdge]] = field(default_factory=lambda: defaultdict(list))


@dataclass
class RunCompleteness:
    status: str = "invalid"
    available_instances: List[str] = field(default_factory=list)
    missing_instances: List[str] = field(default_factory=list)
    not_required_instances: List[str] = field(default_factory=list)
    manifest_path: str = ""

    @property
    def available_count(self) -> int:
        return len(self.available_instances)

    @property
    def missing_count(self) -> int:
        return len(self.missing_instances)

    @property
    def not_required_count(self) -> int:
        return len(self.not_required_instances)


@dataclass
class ExceptionEvidence:
    inst_name: str
    port_name: str
    constraint_type: str
    source_sdc_file: str
    source_line: str
    source_digest: str
    original_command: str


@dataclass
class PendingPlan:
    pending_updates: Dict[Path, str] = field(default_factory=dict)
    removed_lines: List[str] = field(default_factory=list)
    removed_count: int = 0


@dataclass
class DelayCandidate:
    inst_name: str
    module_name: str
    owner: str
    port_name: str
    direction: str
    constraint_type: str
    clock_name: str
    min_value: str
    max_value: str
    bare_value: str
    complex_options: str
    source_sdc_file: str
    source_line: str
    source_digest: str
    extraction_time: str
    original_command: str
    parse_status: str
    message: str = ""


@dataclass
class BudgetSeed:
    values: Dict[str, str]
    parse_status: str
    channel: ChannelRecord
    candidates: List[DelayCandidate]


@dataclass
class FormRow:
    row_idx: int
    values: Dict[str, object]
    autofilled_fields: Set[str] = field(default_factory=set)


@dataclass
class Target20Channel:
    context: Any
    channel_id: str
    connection_hash: str


class Report:
    def __init__(self) -> None:
        self.lines: List[str] = []
        self.warning_count = 0
        self.error_count = 0
        self.sync_changed = False

    def info(self, msg: str) -> None:
        self.lines.append(f"INFO: {msg}")

    def warn(self, msg: str) -> None:
        self.warning_count += 1
        self.lines.append(f"WARNING: {msg}")

    def error(self, msg: str) -> None:
        self.error_count += 1
        self.lines.append(f"ERROR: {msg}")


def _author_part_a() -> str:
    return chr(72) + chr(111)


def _author_part_b() -> str:
    return chr(119) + chr(97)


def _author_part_c() -> str:
    return chr(114) + chr(100)


def author_name() -> str:
    return _author_part_a() + _author_part_b() + _author_part_c()


def atomic_write_text(path: Path, content: str, encoding: str = "utf-8") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.tmp.{os.getpid()}")
    try:
        tmp.write_text(content, encoding=encoding)
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            tmp.unlink()


def atomic_write_csv(path: Path, headers: Sequence[str], rows: Sequence[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.tmp.{os.getpid()}")
    try:
        with tmp.open("w", encoding="utf-8", newline="") as file_obj:
            writer = csv.DictWriter(
                file_obj,
                fieldnames=list(headers),
                extrasaction="ignore",
                lineterminator="\n",
            )
            writer.writeheader()
            for row in rows:
                writer.writerow({header: clean_cell(row.get(header)) for header in headers})
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            tmp.unlink()


def atomic_write_json(path: Path, payload: Dict[str, object]) -> None:
    atomic_write_text(path, json.dumps(payload, indent=2, sort_keys=True) + "\n")


def atomic_save_workbook(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.stem}.tmp.{os.getpid()}{path.suffix}")
    try:
        workbook.save(str(tmp))
        os.replace(str(tmp), str(path))
    finally:
        if tmp.exists():
            tmp.unlink()


def resolve_path(base: Path, value: Optional[str], default: str) -> Path:
    path = Path(value).expanduser() if value else Path(default)
    return path if path.is_absolute() else base / path


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def normalize_key(value) -> str:
    return clean_cell(value).strip().lower()


def canonical_clock_relation(value) -> str:
    text = normalize_key(value)
    if not text:
        return ""
    text = re.sub(r"[\s-]+", " ", text)
    text = text.replace("_", " ")
    alias_key = text.strip()
    if alias_key in CLOCK_RELATION_ALIASES:
        return CLOCK_RELATION_ALIASES[alias_key]
    underscore_key = alias_key.replace(" ", "_")
    if underscore_key in CLOCK_RELATION_CANONICAL:
        return underscore_key
    return ""


def safe_filename_token(value: str) -> str:
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-")
    token = "".join(char if char in allowed else "_" for char in clean_cell(value))
    return token or "unknown"


def sanitize_id(value: str) -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", clean_cell(value)).strip("_")
    return token or "unknown"


def brace_list(names: Sequence[str]) -> str:
    return "{" + " ".join(clean_cell(name) for name in names if clean_cell(name)) + "}"


def get_collection(kind: str, objects: Sequence[str]) -> str:
    return f"[{kind} {brace_list(objects)}]"


def parse_number(value) -> Optional[float]:
    text = clean_cell(value)
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def format_number(value) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    return f"{number:.12g}"


def read_text(path: Path) -> str:
    last_error: Optional[Exception] = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    assert last_error is not None
    raise last_error


def digest_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def read_harden_sdc_manifest(
    path: Path,
    run_root: Path,
    scenario: str,
    require_complete: bool,
    report: Report,
) -> Tuple[Dict[str, InstInfo], RunCompleteness]:
    errors_before = report.error_count
    if not path.is_file():
        report.error(f"{path}: HARDEN_SDC_MANIFEST_MISSING")
        return {}, RunCompleteness(status="invalid", manifest_path=str(path))

    required = {"scenario", "inst_name", "module_name", "availability_status"}
    instances: Dict[str, InstInfo] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = set(reader.fieldnames or [])
        missing_fields = sorted(required - fields)
        if missing_fields:
            report.error(f"{path}: manifest missing field(s): {','.join(missing_fields)}")
        if "sdc_path" not in fields and "resolved_sdc_path" not in fields:
            report.error(f"{path}: manifest missing sdc_path")
        for row_idx, row in enumerate(reader, start=2):
            row_scenario = clean_cell(row.get("scenario"))
            inst_name = clean_cell(row.get("inst_name"))
            module_name = clean_cell(row.get("module_name"))
            status = normalize_key(row.get("availability_status"))
            if row_scenario != scenario:
                report.error(
                    f"{path.name} row {row_idx}: scenario={row_scenario or '<empty>'} "
                    f"does not match requested {scenario}"
                )
                continue
            if not inst_name or not module_name:
                report.error(f"{path.name} row {row_idx}: inst_name/module_name is required")
                continue
            if inst_name in instances:
                report.error(f"{path.name} row {row_idx}: duplicate inst_name {inst_name}")
                continue
            sdc_value = clean_cell(row.get("sdc_path")) or clean_cell(row.get("resolved_sdc_path"))
            inst = InstInfo(
                module_name=module_name,
                inst_name=inst_name,
                owner=clean_cell(row.get("owner")),
                sdc_hint=sdc_value,
                sdc_status=status,
                sdc_note=clean_cell(row.get("note")),
            )
            if status == "available":
                if not sdc_value:
                    report.error(f"{path.name} row {row_idx}: available {inst_name} has empty sdc_path")
                else:
                    resolved = resolve_path(run_root, sdc_value, sdc_value)
                    if not resolved.is_file():
                        report.error(f"{path.name} row {row_idx}: available SDC missing for {inst_name}: {resolved}")
                    elif not os.access(str(resolved), os.R_OK):
                        report.error(f"{path.name} row {row_idx}: available SDC is not readable for {inst_name}: {resolved}")
                    else:
                        inst.sdc_path = resolved.resolve()
            elif status == "missing":
                report.warn(f"{path.name} row {row_idx}: HARDEN_SDC_MISSING {inst_name}: {inst.sdc_note or '<no note>'}")
            elif status == "not_required":
                if sdc_value:
                    report.error(f"{path.name} row {row_idx}: not_required {inst_name} must have empty sdc_path")
                if not inst.sdc_note:
                    report.error(f"{path.name} row {row_idx}: not_required {inst_name} requires an explicit note/basis")
            else:
                report.error(f"{path.name} row {row_idx}: invalid availability_status {status or '<empty>'}")
            instances[inst_name] = inst

    available = sorted(name for name, inst in instances.items() if inst.sdc_status == "available" and inst.sdc_path)
    missing = sorted(name for name, inst in instances.items() if inst.sdc_status == "missing")
    not_required = sorted(name for name, inst in instances.items() if inst.sdc_status == "not_required")
    status = "partial" if missing else "complete"
    if report.error_count > errors_before:
        status = "invalid"
    if require_complete and missing:
        report.error("HARDEN_SDC_COMPLETENESS_REQUIRED: " + ",".join(missing))
    completeness = RunCompleteness(
        status=status,
        available_instances=available,
        missing_instances=missing,
        not_required_instances=not_required,
        manifest_path=str(path.resolve()),
    )
    report.info(
        f"harden SDC completeness={completeness.status} available={completeness.available_count} "
        f"missing={completeness.missing_count} not_required={completeness.not_required_count}"
    )
    return instances, completeness


def infer_legacy_manifest(
    edges: Sequence[ConnectionEdge], cwd: Path, scenario: str, report: Report
) -> Tuple[Dict[str, InstInfo], RunCompleteness]:
    del scenario
    names = sorted(
        {
            name
            for edge in edges
            for name in (edge.src_instance, edge.dst_instance)
            if normalize_key(name) not in PSEUDO_INSTANCES
        }
    )
    instances: Dict[str, InstInfo] = {}
    available: List[str] = []
    missing: List[str] = []
    for name in names:
        path = cwd / f"{name}.sdc"
        status = "available" if path.is_file() else "missing"
        inst = InstInfo(module_name=name, inst_name=name, sdc_status=status, sdc_hint=str(path))
        if status == "available":
            inst.sdc_path = path.resolve()
            available.append(name)
        else:
            missing.append(name)
            report.warn(f"legacy mode: no SDC found for {name}")
        instances[name] = inst
    return instances, RunCompleteness(
        status="partial" if missing else "complete",
        available_instances=available,
        missing_instances=missing,
    )


def attach_connection_ports(instances: Dict[str, InstInfo], edges: Sequence[ConnectionEdge], report: Report) -> None:
    for edge in edges:
        for inst_name, direction, port in (
            (edge.src_instance, normalize_key(edge.src_direction), edge.src_port),
            (edge.dst_instance, normalize_key(edge.dst_direction), edge.dst_port),
        ):
            if normalize_key(inst_name) in PSEUDO_INSTANCES:
                continue
            inst = instances.get(inst_name)
            if inst is None:
                report.error(f"{edge.connection_id}: endpoint instance {inst_name} missing from harden SDC manifest")
                continue
            if direction == "input":
                target = inst.inputs
            elif direction == "output":
                target = inst.outputs
            elif direction == "inout":
                target = inst.inouts
            else:
                report.error(f"{edge.connection_id}: invalid direction {direction or '<empty>'} for {inst_name}/{port}")
                continue
            target.setdefault(port, PortInfo(name=port))
            target.setdefault(port_base(port), PortInfo(name=port_base(port)))


def scenario_scope_matches(value: str, scenario: str) -> bool:
    tokens = {
        normalize_key(token)
        for token in re.split(r"[\s,;|]+", clean_cell(value))
        if clean_cell(token)
    }
    if not tokens:
        tokens = {"common"}
    return "common" in tokens or scenario in tokens


def collect_current_sdc_digests(instances: Dict[str, InstInfo]) -> Dict[str, str]:
    digests: Dict[str, str] = {}
    for inst in instances.values():
        if inst.sdc_path and inst.sdc_path.is_file():
            try:
                digest = digest_file(inst.sdc_path)
                digests[str(inst.sdc_path)] = digest
                if inst.sdc_path.name not in digests:
                    digests[inst.sdc_path.name] = digest
            except OSError:
                continue
    return digests


def parse_endpoint_key(value: str) -> Tuple[str, str, str]:
    value = clean_cell(value)
    if not value:
        return "", "", ""
    parts = value.split(":", 2)
    if len(parts) == 3:
        return parts[0], parts[1], parts[2]
    return "", "", ""


def endpoint_collection(inst_name: str, port_name: str) -> str:
    if not inst_name or not port_name:
        return ""
    return get_collection("get_pins", [f"{inst_name}/{port_name}"])


def split_port_key(port: str) -> Tuple[str, str, bool, bool]:
    port = clean_cell(port)
    range_match = PORT_RANGE_RE.fullmatch(port)
    if range_match:
        return range_match.group(1), "", True, False
    bit_match = PORT_EXACT_BIT_RE.fullmatch(port)
    if bit_match:
        return bit_match.group(1), bit_match.group(2), False, False
    if "[" in port or "]" in port:
        return port, "", False, True
    return port, "", False, False


def port_base(port: str) -> str:
    base, _, _, _ = split_port_key(port)
    return base


def inferred_bit_index(port: str, explicit: str = "") -> str:
    if clean_cell(explicit):
        return clean_cell(explicit)
    _, bit_index, _, _ = split_port_key(port)
    return bit_index


def is_canonical_port_key(port: str) -> bool:
    return bool(PORT_BIT_RE.fullmatch(clean_cell(port)))


def is_feedthrough_port(port: str) -> bool:
    return port_base(port).startswith(("fti_", "fto_"))


def build_channel_id(src_inst: str, src_port: str, dst_inst: str, dst_port: str) -> str:
    def port_token(port: str) -> str:
        return re.sub(r"\[(\d+)\]", r"_bit\1", clean_cell(port))

    return "CH_" + "__".join(
        [
            f"{sanitize_id(src_inst)}_{sanitize_id(port_token(src_port))}",
            f"{sanitize_id(dst_inst)}_{sanitize_id(port_token(dst_port))}",
        ]
    )


def endpoint_from_soc_object(inst_name: str, port_name: str, soc_object: str, endpoint_key: str) -> str:
    obj = clean_cell(soc_object)
    inst = clean_cell(inst_name)
    port = clean_cell(port_name)
    if obj:
        if obj.startswith("["):
            return obj
        if "/" in obj:
            return get_collection("get_pins", [obj])
        if normalize_key(inst) == "top":
            return get_collection("get_ports", [obj])
        if normalize_key(inst) in {"fabric", "unknown", "constant", "const"}:
            return obj
        return get_collection("get_pins", [f"{inst}/{obj}"])
    key_inst, _, key_port = parse_endpoint_key(endpoint_key)
    inst = inst or key_inst
    port = port or key_port
    if not inst or not port:
        return ""
    if normalize_key(inst) == "top":
        return get_collection("get_ports", [port])
    if normalize_key(inst) in {"fabric", "unknown", "constant", "const"}:
        return ""
    return endpoint_collection(inst, port)


def read_connection_inventory(
    path: Path,
    report: Report,
    scenario: str = "common",
    target_layout: bool = False,
) -> ConnectionIndex:
    index = ConnectionIndex()
    if not path.is_file():
        report.error(f"connection inventory not found: {path}")
        return index
    seen_ids: Set[str] = set()
    seen_pairs: Set[Tuple[str, str, str, str]] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        if not reader.fieldnames:
            report.error(f"{path}: connection_inventory.csv has no header")
            return index
        fields = set(reader.fieldnames)
        if target_layout and "scenario_scope" not in fields:
            report.error(f"{path}: target connection inventory is missing scenario_scope")
            return index
        if target_layout:
            required_fields = {
                "schema_version", "connection_id", "connection_type", "src_instance", "src_direction",
                "src_port", "src_endpoint_key", "src_soc_object", "dst_instance",
                "dst_direction", "dst_port", "dst_endpoint_key", "dst_soc_object",
                "validation_status", "scenario_scope",
            }
            missing_fields = sorted(required_fields - fields)
            if missing_fields:
                report.error(f"{path}: target connection inventory missing field(s): {','.join(missing_fields)}")
                return index
        for row_idx, row in enumerate(reader, start=2):
            if target_layout and clean_cell(row.get("schema_version")) != SCHEMA_VERSION:
                report.error(f"{path.name} row {row_idx}: unsupported schema_version")
                continue
            scope = clean_cell(row.get("scenario_scope")) or "common"
            if not scenario_scope_matches(scope, scenario):
                continue
            src_inst = clean_cell(row.get("src_instance"))
            src_dir = clean_cell(row.get("src_direction"))
            src_port = clean_cell(row.get("src_port"))
            dst_inst = clean_cell(row.get("dst_instance"))
            dst_dir = clean_cell(row.get("dst_direction"))
            dst_port = clean_cell(row.get("dst_port"))
            if not (src_inst and src_port):
                key_inst, key_dir, key_port = parse_endpoint_key(clean_cell(row.get("src_endpoint_key")))
                src_inst = src_inst or key_inst
                src_dir = src_dir or key_dir
                src_port = src_port or key_port
            if not (dst_inst and dst_port):
                key_inst, key_dir, key_port = parse_endpoint_key(clean_cell(row.get("dst_endpoint_key")))
                dst_inst = dst_inst or key_inst
                dst_dir = dst_dir or key_dir
                dst_port = dst_port or key_port
            if not (src_inst and src_port and dst_inst and dst_port):
                report.warn(f"{path.name} row {row_idx}: skipped connection edge with incomplete src/dst endpoint")
                continue
            raw_connection_type = normalize_key(row.get("connection_type"))
            if target_layout and raw_connection_type not in CANONICAL_CONNECTION_TYPES:
                report.error(
                    f"{path.name} row {row_idx}: connection_type={raw_connection_type or '<empty>'} "
                    "is outside the canonical 00 enum"
                )
                continue
            connection_id = clean_cell(row.get("connection_id"))
            if not connection_id:
                if target_layout:
                    report.error(f"{path.name} row {row_idx}: target edge has empty connection_id")
                    continue
                connection_id = build_channel_id(src_inst, src_port, dst_inst, dst_port).replace("CH_", "CONN_", 1)
            if connection_id in seen_ids:
                report.error(f"{path.name} row {row_idx}: duplicate connection_id {connection_id}")
            seen_ids.add(connection_id)
            pair_key = (src_inst, src_port, dst_inst, dst_port)
            if pair_key in seen_pairs:
                report.error(f"{path.name} row {row_idx}: duplicate direct bit pair {src_inst}/{src_port} -> {dst_inst}/{dst_port}")
            seen_pairs.add(pair_key)
            for role, port in (("src", src_port), ("dst", dst_port)):
                if port and not is_canonical_port_key(port):
                    report.error(
                        f"{path.name} row {row_idx}: {role}_port {port} is not a canonical scalar/bit key; "
                        "00 must expand bus/range connections before 20"
                    )
            src_explicit_bit = clean_cell(row.get("src_bit_index"))
            dst_explicit_bit = clean_cell(row.get("dst_bit_index"))
            src_inferred_bit = inferred_bit_index(src_port)
            dst_inferred_bit = inferred_bit_index(dst_port)
            if src_explicit_bit and src_explicit_bit != src_inferred_bit:
                report.error(f"{path.name} row {row_idx}: src_bit_index disagrees with {src_port}")
            if dst_explicit_bit and dst_explicit_bit != dst_inferred_bit:
                report.error(f"{path.name} row {row_idx}: dst_bit_index disagrees with {dst_port}")
            if target_layout:
                expected_src_key = f"{src_inst}:{src_dir}:{src_port}"
                expected_dst_key = f"{dst_inst}:{dst_dir}:{dst_port}"
                if clean_cell(row.get("src_endpoint_key")) != expected_src_key:
                    report.error(f"{path.name} row {row_idx}: src_endpoint_key is not canonical")
                if clean_cell(row.get("dst_endpoint_key")) != expected_dst_key:
                    report.error(f"{path.name} row {row_idx}: dst_endpoint_key is not canonical")
                src_inst_key = normalize_key(src_inst)
                dst_inst_key = normalize_key(dst_inst)
                expected_src_obj = src_port if src_inst_key == "top" else f"{src_inst}/{src_port}"
                expected_dst_obj = dst_port if dst_inst_key == "top" else f"{dst_inst}/{dst_port}"
                if src_inst_key not in {"fabric", "unknown", "constant", "const", "nc"} and clean_cell(row.get("src_soc_object")) != expected_src_obj:
                    report.error(f"{path.name} row {row_idx}: src_soc_object is not canonical")
                if dst_inst_key not in {"fabric", "unknown", "constant", "const", "nc"} and clean_cell(row.get("dst_soc_object")) != expected_dst_obj:
                    report.error(f"{path.name} row {row_idx}: dst_soc_object is not canonical")
                if normalize_key(row.get("validation_status")) not in MATCHED_STATUSES - {""}:
                    report.error(f"{path.name} row {row_idx}: target validation_status must be matched/ok/valid")
            edge = ConnectionEdge(
                connection_id=connection_id,
                connection_type=raw_connection_type,
                src_instance=src_inst,
                src_direction=src_dir,
                src_port=src_port,
                src_bit_index=inferred_bit_index(src_port, src_explicit_bit),
                src_endpoint_key=clean_cell(row.get("src_endpoint_key")),
                src_soc_object=clean_cell(row.get("src_soc_object")),
                dst_instance=dst_inst,
                dst_direction=dst_dir,
                dst_port=dst_port,
                dst_bit_index=inferred_bit_index(dst_port, dst_explicit_bit),
                dst_endpoint_key=clean_cell(row.get("dst_endpoint_key")),
                dst_soc_object=clean_cell(row.get("dst_soc_object")),
                validation_status=clean_cell(row.get("validation_status")),
                scenario_scope=scope,
                note=clean_cell(row.get("note")),
            )
            index.edges.append(edge)
            index.by_src[(src_inst, src_port)].append(edge)
            index.by_dst[(dst_inst, dst_port)].append(edge)
    report.info(f"loaded {len(index.edges)} connection edge(s) from {path}")
    return index


def read_feedthrough_ownership(
    path: Path,
    scenario: str,
    connection_ids: Set[str],
    target_layout: bool,
    report: Report,
) -> Set[str]:
    if not path.is_file():
        if target_layout:
            report.error(f"feedthrough edge inventory not found: {path}")
        else:
            report.info(f"legacy feedthrough edge inventory not found: {path}")
        return set()
    owned: Set[str] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = set(reader.fieldnames or [])
        if "connection_id" not in fields:
            report.error(f"{path}: feedthrough edge inventory is missing connection_id")
            return owned
        if target_layout and not {"schema_version", "scenario"}.issubset(fields):
            report.error(f"{path}: target feedthrough edge inventory is missing schema_version/scenario")
            return owned
        for row_idx, row in enumerate(reader, start=2):
            if target_layout and clean_cell(row.get("schema_version")) != SCHEMA_VERSION:
                report.error(f"{path.name} row {row_idx}: unsupported schema_version {clean_cell(row.get('schema_version'))}")
                continue
            row_scenario = normalize_key(row.get("scenario")) or "common"
            if row_scenario not in {"common", scenario}:
                report.error(
                    f"{path.name} row {row_idx}: scenario={row_scenario} does not belong to requested {scenario}"
                )
                continue
            connection_id = clean_cell(row.get("connection_id"))
            if not connection_id:
                report.error(f"{path.name} row {row_idx}: empty connection_id")
                continue
            if connection_id not in connection_ids:
                report.error(f"{path.name} row {row_idx}: stale connection_id {connection_id} not found in 00 inventory")
                continue
            owned.add(connection_id)
    report.info(f"loaded {len(owned)} 10-owned direct edge(s) from {path}")
    return owned


def edge_status_ok(edge: ConnectionEdge) -> bool:
    return normalize_key(edge.validation_status) in MATCHED_STATUSES


def instance_module(instances: Dict[str, InstInfo], inst_name: str) -> str:
    inst = instances.get(inst_name)
    if inst:
        return inst.module_name
    if normalize_key(inst_name) == "top":
        return "top"
    if normalize_key(inst_name) == "fabric":
        return "fabric"
    return ""


def classify_edge_type(edge: ConnectionEdge, instances: Dict[str, InstInfo]) -> str:
    raw = normalize_key(edge.connection_type)
    if raw in CHANNEL_TYPES_20 or raw in NON_20_CHANNEL_TYPES:
        return raw
    if "feedthrough" in raw:
        return "feedthrough"
    if normalize_key(edge.src_instance) == "top" or normalize_key(edge.dst_instance) == "top":
        if normalize_key(edge.src_instance) == "top" and normalize_key(edge.dst_instance) == "top":
            return "pad_to_pad"
        return "top_pad_to_harden" if normalize_key(edge.src_instance) == "top" else "harden_to_top_pad"
    if edge.src_instance in instances and edge.dst_instance in instances:
        return "harden_to_harden"
    if edge.src_instance in instances:
        return "harden_to_fabric"
    if edge.dst_instance in instances:
        return "fabric_to_harden"
    return "unknown"


def edge_to_channel(
    edge: ConnectionEdge,
    instances: Dict[str, InstInfo],
    clock_objects: Set[str],
    report: Report,
    scenario: str = "common",
    force_type: str = "",
    connection_source: str = "",
    note: str = "",
) -> ChannelRecord:
    channel_type = force_type or classify_edge_type(edge, instances)
    if is_feedthrough_port(edge.src_port) or is_feedthrough_port(edge.dst_port):
        channel_type = "feedthrough"
    is_pad_related = "yes" if channel_type in {"top_pad_to_harden", "harden_to_top_pad", "pad_to_pad"} else "no"
    is_feedthrough = "yes" if channel_type == "feedthrough" else "no"
    src_endpoint = endpoint_from_soc_object(edge.src_instance, edge.src_port, edge.src_soc_object, edge.src_endpoint_key)
    dst_endpoint = endpoint_from_soc_object(edge.dst_instance, edge.dst_port, edge.dst_soc_object, edge.dst_endpoint_key)
    is_clock_related = "no"
    if endpoint_hits_clock(src_endpoint, clock_objects) or endpoint_hits_clock(dst_endpoint, clock_objects):
        is_clock_related = "yes"
        if channel_type in CHANNEL_TYPES_20:
            channel_type = "clock_connection"
    messages = []
    if note:
        messages.append(note)
    if edge.validation_status and not edge_status_ok(edge):
        messages.append(f"00 connection validation_status={edge.validation_status}")
    if channel_type not in CHANNEL_TYPES_20:
        messages.append("not a 20 channel by default")
    src_inst = instances.get(edge.src_instance)
    dst_inst = instances.get(edge.dst_instance)
    src_status = src_inst.sdc_status if src_inst else "not_required"
    dst_status = dst_inst.sdc_status if dst_inst else "not_required"
    evidence_status = "incomplete_missing_sdc" if "missing" in {src_status, dst_status} else "complete"
    return ChannelRecord(
        channel_id=build_channel_id(edge.src_instance, edge.src_port, edge.dst_instance, edge.dst_port),
        scenario=scenario,
        stage="all",
        corner="all",
        channel_type=channel_type,
        connection_id=edge.connection_id,
        src_instance=edge.src_instance,
        src_module=instance_module(instances, edge.src_instance),
        src_direction=edge.src_direction,
        src_port=edge.src_port,
        src_bit_index=edge.src_bit_index,
        src_endpoint=src_endpoint,
        dst_instance=edge.dst_instance,
        dst_module=instance_module(instances, edge.dst_instance),
        dst_direction=edge.dst_direction,
        dst_port=edge.dst_port,
        dst_bit_index=edge.dst_bit_index,
        dst_endpoint=dst_endpoint,
        connection_source=connection_source or edge.connection_id,
        is_pad_related=is_pad_related,
        is_clock_related=is_clock_related,
        is_feedthrough=is_feedthrough,
        src_sdc_status=src_status,
        dst_sdc_status=dst_status,
        evidence_status=evidence_status,
        timing_model="unknown",
        note="; ".join(dict.fromkeys(messages)),
    )


def build_channels_from_inventories(
    instances: Dict[str, InstInfo],
    connections: ConnectionIndex,
    feedthrough_connection_ids: Set[str],
    clock_objects: Set[str],
    report: Report,
    scenario: str = "common",
) -> List[ChannelRecord]:
    channels: List[ChannelRecord] = []
    seen: Set[str] = set()

    def add(record: ChannelRecord) -> None:
        if record.channel_id in seen:
            report.error(f"duplicate/colliding channel_id {record.channel_id}; 20 cannot drop a direct bit edge")
            return
        seen.add(record.channel_id)
        channels.append(record)

    for edge in connections.edges:
        if edge.connection_id in feedthrough_connection_ids:
            report.info(f"excluded 10-owned direct edge {edge.connection_id} from 20")
            continue
        if is_feedthrough_port(edge.src_port) or is_feedthrough_port(edge.dst_port):
            report.error(
                f"{edge.connection_id}: feedthrough-adjacent edge is absent from current-scenario 10 inventory"
            )
            continue
        if normalize_key(edge.connection_type) in {"feedthrough", "feedthrough_candidate"}:
            report.error(
                f"{edge.connection_id}: 00 feedthrough candidate is absent from current-scenario 10 inventory"
            )
            continue
        channel_type = classify_edge_type(edge, instances)
        if channel_type not in CHANNEL_TYPES_20:
            continue
        if edge.src_instance in instances and normalize_key(edge.src_direction) not in {"output", "inout"}:
            report.error(f"{edge.connection_id}: harden source direction must be output/inout")
            continue
        if edge.dst_instance in instances and normalize_key(edge.dst_direction) not in {"input", "inout"}:
            report.error(f"{edge.connection_id}: harden destination direction must be input/inout")
            continue
        if not edge_status_ok(edge):
            report.error(
                f"{edge.connection_id}: validation_status={edge.validation_status or '<empty>'} "
                "does not permit a 20 channel"
            )
            continue
        record = edge_to_channel(edge, instances, clock_objects, report, scenario=scenario)
        if record.channel_type not in CHANNEL_TYPES_20:
            report.info(f"excluded {record.channel_type} edge {edge.connection_id} from 20")
            continue
        add(record)

    channels.sort(key=lambda record: (record.connection_id, record.channel_id))
    report.info(f"built {len(channels)} direct 20 channel record(s) after 10 ownership exclusion")
    return channels


def is_escaped(text: str, index: int) -> bool:
    count = 0
    pos = index - 1
    while pos >= 0 and text[pos] == "\\":
        count += 1
        pos -= 1
    return count % 2 == 1


def strip_inline_comment(line: str) -> str:
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(line):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(line, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == "#" and brace_depth == 0 and bracket_depth == 0:
                if idx == 0 or line[idx - 1].isspace() or line[idx - 1] == ";":
                    return line[:idx].rstrip()
    return line.rstrip()


def split_semicolon_commands(text: str) -> List[str]:
    parts: List[str] = []
    start = 0
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(text):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(text, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == ";" and brace_depth == 0 and bracket_depth == 0:
                parts.append(text[start:idx])
                start = idx + 1
    parts.append(text[start:])
    return parts


def iter_tcl_commands_with_line(text: str) -> Iterable[TclCommand]:
    commands: List[TclCommand] = []
    buffer: List[str] = []
    line_no = 1
    start_line = 0
    quote = False
    brace_depth = 0
    bracket_depth = 0
    invalid_structure = False
    idx = 0

    def flush(complete: bool = True) -> None:
        nonlocal buffer, start_line
        cleaned = "".join(buffer).strip().rstrip(";").strip()
        if cleaned:
            commands.append(TclCommand(cleaned, start_line or line_no, complete))
        buffer = []
        start_line = 0

    while idx < len(text):
        char = text[idx]
        if char == "\\" and idx + 1 < len(text):
            if text[idx + 1] == "\n":
                buffer.append(" ")
                line_no += 1
                idx += 2
                continue
            if not start_line:
                start_line = line_no
            buffer.extend((char, text[idx + 1]))
            idx += 2
            continue
        if char == "#" and not quote and brace_depth == 0 and bracket_depth == 0:
            previous = buffer[-1] if buffer else ""
            if not "".join(buffer).strip() or previous.isspace() or previous == ";":
                while idx < len(text) and text[idx] != "\n":
                    idx += 1
                continue
        if char == '"' and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}":
                if brace_depth == 0:
                    invalid_structure = True
                else:
                    brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]":
                if bracket_depth == 0:
                    invalid_structure = True
                else:
                    bracket_depth -= 1
        if char == "\n":
            if quote or brace_depth or bracket_depth:
                buffer.append(" ")
            else:
                flush(not invalid_structure)
                invalid_structure = False
            line_no += 1
            idx += 1
            continue
        if char == ";" and not quote and brace_depth == 0 and bracket_depth == 0:
            flush(not invalid_structure)
            invalid_structure = False
            idx += 1
            continue
        if not start_line and not char.isspace():
            start_line = line_no
        buffer.append(char)
        idx += 1

    complete = not (quote or brace_depth or bracket_depth or invalid_structure)
    flush(complete)
    for command in commands:
        yield command


def find_matching(text: str, start: int, open_char: str, close_char: str) -> int:
    depth = 0
    quote = False
    idx = start
    while idx < len(text):
        char = text[idx]
        if char == "\\":
            idx += 2
            continue
        if char == '"' and open_char != '"' and not is_escaped(text, idx):
            quote = not quote
        elif not quote:
            if char == open_char:
                depth += 1
            elif char == close_char:
                depth -= 1
                if depth == 0:
                    return idx
        idx += 1
    return -1


def tokenize_tcl_words(text: str) -> List[str]:
    tokens: List[str] = []
    idx = 0
    while idx < len(text):
        while idx < len(text) and text[idx].isspace():
            idx += 1
        if idx >= len(text):
            break
        start = idx
        if text[idx] == "{":
            end = find_matching(text, idx, "{", "}")
            if end < 0:
                tokens.append(text[start:])
                break
            tokens.append(text[start : end + 1])
            idx = end + 1
            continue
        if text[idx] == '"':
            idx += 1
            while idx < len(text):
                if text[idx] == "\\":
                    idx += 2
                    continue
                if text[idx] == '"':
                    idx += 1
                    break
                idx += 1
            tokens.append(text[start:idx])
            continue
        pieces: List[str] = []
        while idx < len(text) and not text[idx].isspace():
            char = text[idx]
            if char == "\\":
                pieces.append(text[idx : idx + 2])
                idx += 2
            elif char == "[":
                end = find_matching(text, idx, "[", "]")
                if end < 0:
                    pieces.append(text[idx:])
                    idx = len(text)
                else:
                    pieces.append(text[idx : end + 1])
                    idx = end + 1
            else:
                pieces.append(char)
                idx += 1
        tokens.append("".join(pieces))
    return tokens


def strip_braces(text: str) -> str:
    text = clean_cell(text)
    if len(text) >= 2 and text[0] == "{" and text[-1] == "}":
        return text[1:-1].strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        return text[1:-1].strip()
    return text


def split_object_list(text: str) -> List[str]:
    text = strip_braces(text)
    return [part for part in re.split(r"[\s,;]+", text) if part]


COLLECTION_KINDS = {"get_ports", "get_pins", "get_nets", "get_clocks"}


def parse_collection(token: str) -> Optional[Tuple[str, List[str]]]:
    text = clean_cell(token)
    if not (text.startswith("[") and text.endswith("]")):
        return None
    end = find_matching(text, 0, "[", "]")
    if end != len(text) - 1:
        return None
    words = tokenize_tcl_words(text[1:-1].strip())
    if not words or words[0] not in COLLECTION_KINDS:
        return None
    kind = words[0]
    objects: List[str] = []
    idx = 1
    while idx < len(words):
        word = words[idx]
        if word.startswith("-"):
            idx += 1
            continue
        objects.extend(split_object_list(word))
        idx += 1
    return kind, objects


OPTIONS_WITH_VALUE = {"-clock", "-reference_pin"}
OPTIONS_NO_VALUE = {"-add_delay", "-clock_fall", "-rise", "-fall", "-min", "-max"}
COMPLEX_DELAY_FLAGS = {"-add_delay", "-clock_fall", "-rise", "-fall"}


def option_value(tokens: Sequence[str], option: str) -> str:
    for idx, token in enumerate(tokens):
        if token == option and idx + 1 < len(tokens):
            return strip_braces(tokens[idx + 1])
    return ""


def has_option(tokens: Sequence[str], option: str) -> bool:
    return option in tokens


def extract_clock_name(tokens: Sequence[str]) -> str:
    value = option_value(tokens, "-clock")
    parsed = parse_collection(value)
    if parsed and parsed[0] == "get_clocks" and parsed[1]:
        return parsed[1][0]
    return strip_braces(value)


def parse_delay_values(tokens: Sequence[str]) -> Tuple[str, str, str]:
    min_value = ""
    max_value = ""
    bare_values: List[str] = []
    qualifier = ""
    idx = 1
    while idx < len(tokens):
        token = tokens[idx]
        if token in {"-min", "-max"}:
            qualifier = token[1:]
            idx += 1
            continue
        if token in OPTIONS_WITH_VALUE:
            idx += 2
            continue
        if token in OPTIONS_NO_VALUE or (token.startswith("-") and parse_number(token) is None):
            idx += 1
            continue
        if parse_collection(token):
            idx += 1
            continue
        value = strip_braces(token)
        if parse_number(value) is not None:
            if qualifier == "min" and not min_value:
                min_value = value
                qualifier = ""
            elif qualifier == "max" and not max_value:
                max_value = value
                qualifier = ""
            else:
                bare_values.append(value)
        idx += 1
    bare = bare_values[0] if bare_values else ""
    if bare:
        if "-min" in tokens and not min_value:
            min_value = bare
        if "-max" in tokens and not max_value:
            max_value = bare
        if "-min" in tokens or "-max" in tokens:
            bare = ""
        else:
            min_value = min_value or bare
            max_value = max_value or bare
            bare = ""
    return min_value, max_value, bare


def last_non_clock_collection(tokens: Sequence[str]) -> Optional[Tuple[str, List[str], str]]:
    found: Optional[Tuple[str, List[str], str]] = None
    for token in tokens:
        parsed = parse_collection(token)
        if parsed and parsed[0] != "get_clocks":
            found = (parsed[0], parsed[1], token)
    return found


def normalize_sdc_port(kind: str, obj: str, inst: InstInfo) -> str:
    obj = strip_braces(obj)
    if kind == "get_pins" and "/" in obj:
        return obj.split("/")[-1]
    return obj


def lookup_port_direction(inst: InstInfo, port_name: str) -> str:
    if port_name in inst.inputs:
        return "input"
    if port_name in inst.outputs:
        return "output"
    if port_name in inst.inouts:
        return "inout"
    base = port_base(port_name)
    if base in inst.inputs:
        return "input"
    if base in inst.outputs:
        return "output"
    if base in inst.inouts:
        return "inout"
    return "unknown"


def parse_delay_candidate(inst: InstInfo, cmd: TclCommand, digest: str, now: str) -> Optional[DelayCandidate]:
    tokens = tokenize_tcl_words(cmd.raw)
    if not tokens or tokens[0] not in {"set_input_delay", "set_output_delay"}:
        return None
    ctype = "input_delay" if tokens[0] == "set_input_delay" else "output_delay"
    min_value, max_value, bare_value = parse_delay_values(tokens)
    target = last_non_clock_collection(tokens)
    if not target:
        return DelayCandidate(
            inst.inst_name,
            inst.module_name,
            inst.owner,
            "",
            "unknown",
            ctype,
            extract_clock_name(tokens),
            min_value,
            max_value,
            bare_value,
            "",
            str(inst.sdc_path) if inst.sdc_path else "",
            str(cmd.line_no),
            digest,
            now,
            cmd.raw,
            "needs_review",
            "no target collection found",
        )
    kind, objects, _ = target
    if len(objects) != 1:
        port_name = " ".join(objects)
        status = "needs_review"
        message = "target collection contains multiple objects"
    else:
        port_name = normalize_sdc_port(kind, objects[0], inst)
        status = "ok"
        message = ""
    direction = lookup_port_direction(inst, port_name)
    if ctype == "input_delay" and direction != "input":
        status = "needs_review"
        message = "; ".join(filter(None, [message, "set_input_delay target is not an input port in integration table"]))
    if ctype == "output_delay" and direction != "output":
        status = "needs_review"
        message = "; ".join(filter(None, [message, "set_output_delay target is not an output port in integration table"]))
    complex_opts = []
    for opt in sorted(COMPLEX_DELAY_FLAGS):
        if has_option(tokens, opt):
            complex_opts.append(opt)
    if len(objects) != 1:
        complex_opts.append("multi_object")
    if not any((min_value, max_value, bare_value)):
        status = "needs_review"
        message = "; ".join(filter(None, [message, "delay value is missing or non-numeric"]))
    return DelayCandidate(
        inst_name=inst.inst_name,
        module_name=inst.module_name,
        owner=inst.owner,
        port_name=port_name,
        direction=direction,
        constraint_type=ctype,
        clock_name=extract_clock_name(tokens),
        min_value=min_value,
        max_value=max_value,
        bare_value=bare_value,
        complex_options=",".join(complex_opts),
        source_sdc_file=str(inst.sdc_path) if inst.sdc_path else "",
        source_line=str(cmd.line_no),
        source_digest=digest,
        extraction_time=now,
        original_command=cmd.raw,
        parse_status=status,
        message=message,
    )


def expand_port_expression(value: str) -> List[str]:
    result: List[str] = []
    for item in split_object_list(value):
        match = PORT_RANGE_RE.fullmatch(item)
        if not match:
            result.append(item)
            continue
        base = match.group(1)
        left = int(match.group(2))
        right = int(match.group(3))
        step = -1 if left > right else 1
        result.extend(f"{base}[{index}]" for index in range(left, right + step, step))
    return result


def clone_delay_candidate(candidate: DelayCandidate, inst: InstInfo, port_name: str) -> DelayCandidate:
    return DelayCandidate(
        inst_name=candidate.inst_name,
        module_name=candidate.module_name,
        owner=candidate.owner,
        port_name=port_name,
        direction=lookup_port_direction(inst, port_name),
        constraint_type=candidate.constraint_type,
        clock_name=candidate.clock_name,
        min_value=candidate.min_value,
        max_value=candidate.max_value,
        bare_value=candidate.bare_value,
        complex_options=candidate.complex_options,
        source_sdc_file=candidate.source_sdc_file,
        source_line=candidate.source_line,
        source_digest=candidate.source_digest,
        extraction_time=candidate.extraction_time,
        original_command=candidate.original_command,
        parse_status=candidate.parse_status,
        message=candidate.message,
    )


def extract_delay_candidates(instances: Dict[str, InstInfo], report: Report) -> List[DelayCandidate]:
    results: List[DelayCandidate] = []
    for inst in sorted(instances.values(), key=lambda item: item.inst_name):
        if not inst.sdc_path:
            continue
        try:
            text = read_text(inst.sdc_path)
            digest = digest_file(inst.sdc_path)
            now = datetime.fromtimestamp(inst.sdc_path.stat().st_mtime).isoformat(timespec="seconds")
        except Exception as exc:
            report.error(f"failed to read {inst.sdc_path}: {exc}")
            continue
        count = 0
        for cmd in iter_tcl_commands_with_line(text):
            if not cmd.parse_complete:
                report.error(
                    f"{inst.sdc_path}:{cmd.line_no}: unbalanced Tcl brace/bracket/quote; "
                    "exception evidence is incomplete"
                )
                continue
            cand = parse_delay_candidate(inst, cmd, digest, now)
            if cand:
                expanded = expand_port_expression(cand.port_name) or [cand.port_name]
                results.extend(clone_delay_candidate(cand, inst, port_name) for port_name in expanded)
                count += len(expanded)
        report.info(f"extracted {count} interface delay evidence record(s) from {inst.sdc_path}")
    return results


EXCEPTION_COMMANDS = {
    "set_false_path",
    "set_multicycle_path",
    "set_max_delay",
    "set_min_delay",
}


def extract_exception_evidence(
    instances: Dict[str, InstInfo], report: Report
) -> List[ExceptionEvidence]:
    evidence: List[ExceptionEvidence] = []
    for inst in sorted(instances.values(), key=lambda item: item.inst_name):
        if not inst.sdc_path:
            continue
        try:
            text = read_text(inst.sdc_path)
            digest = digest_file(inst.sdc_path)
        except Exception as exc:
            report.error(f"failed to read exception evidence from {inst.sdc_path}: {exc}")
            continue
        count = 0
        for command in iter_tcl_commands_with_line(text):
            if not command.parse_complete:
                continue
            tokens = tokenize_tcl_words(command.raw)
            if not tokens or tokens[0] not in EXCEPTION_COMMANDS:
                continue
            ports: List[str] = []
            for token in tokens:
                collection = parse_collection(token)
                if not collection or collection[0] not in {"get_ports", "get_pins"}:
                    continue
                for obj in collection[1]:
                    port = normalize_sdc_port(collection[0], obj, inst)
                    ports.extend(expand_port_expression(port))
            for port in sorted(set(ports)):
                if lookup_port_direction(inst, port) == "unknown":
                    continue
                evidence.append(
                    ExceptionEvidence(
                        inst_name=inst.inst_name,
                        port_name=port,
                        constraint_type=tokens[0],
                        source_sdc_file=str(inst.sdc_path),
                        source_line=str(command.line_no),
                        source_digest=digest,
                        original_command=command.raw,
                    )
                )
                count += 1
        report.info(f"extracted {count} known exception evidence record(s) from {inst.sdc_path}")
    return evidence


def endpoint_objects(endpoint: str) -> Set[str]:
    objects: Set[str] = set()
    text = clean_cell(endpoint)
    if not text:
        return objects
    for item in re.findall(r"\{([^}]+)\}", text):
        objects.update(split_object_list(item))
    if "/" in text and not text.startswith("["):
        objects.add(text)
    return objects


def endpoint_hits_clock(endpoint: str, clock_objects: Set[str]) -> bool:
    if clean_cell(endpoint) in clock_objects:
        return True
    return bool(endpoint_objects(endpoint) & clock_objects)


def read_json_meta(path: Path, label: str, required: bool, report: Report) -> Optional[Dict[str, object]]:
    if not path.is_file():
        if required:
            report.error(f"{label} meta not found: {path}")
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        report.error(f"{label} meta invalid: {path}: {exc}")
        return None
    if not isinstance(payload, dict):
        report.error(f"{label} meta must contain a JSON object: {path}")
        return None
    return payload


def _clock_objects_from_row(row: Dict[str, str]) -> Set[str]:
    objects: Set[str] = set()
    for field_name in ("direct_source", "producer_object", "target_object", "root_source"):
        value = clean_cell(row.get(field_name))
        if not value:
            continue
        objects.update(endpoint_objects(value))
        if "/" in value and not value.startswith("["):
            objects.add(value)
    inst_name = clean_cell(row.get("inst_name"))
    port_name = clean_cell(row.get("port_name"))
    if inst_name and port_name:
        objects.add(f"{inst_name}/{port_name}")
    return objects


def read_clock_context(
    path: Path,
    meta_path: Path,
    scenario: str,
    target_layout: bool,
    report: Report,
) -> ClockContext:
    context = ClockContext()
    if not path.is_file():
        if target_layout:
            report.error(f"assembled clock inventory not found: {path}")
        else:
            report.warn(f"clock inventory not found: {path}")
        return context

    meta = read_json_meta(meta_path, "clock inventory", target_layout, report)
    if target_layout and meta is None:
        return context
    try:
        context.inventory_digest = digest_file(path)
        context.meta_digest = digest_file(meta_path) if meta_path.is_file() else ""
    except OSError as exc:
        report.error(f"clock inventory is not readable: {exc}")
        return context
    if meta is not None:
        if clean_cell(meta.get("scenario")) != scenario:
            report.error(
                f"clock inventory meta scenario={clean_cell(meta.get('scenario')) or '<empty>'} "
                f"does not match requested {scenario}"
            )
        if clean_cell(meta.get("inventory_digest")) != context.inventory_digest:
            report.error("clock inventory digest does not match meta")
        final_path_value = clean_cell(meta.get("final_sdc_path"))
        final_digest = clean_cell(meta.get("final_sdc_digest"))
        if target_layout and (not final_path_value or not final_digest):
            report.error("clock inventory meta final_sdc_path/final_sdc_digest is incomplete")
        elif final_path_value and final_digest:
            final_path = Path(final_path_value).expanduser()
            if not final_path.is_absolute():
                final_path = meta_path.parent / final_path
            if not final_path.is_file() or digest_file(final_path) != final_digest:
                report.error("clock inventory final SDC digest/path is stale")

    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = set(reader.fieldnames or [])
        if "clock_name" not in fields:
            report.error(f"{path}: missing clock_name column")
            return context
        if target_layout and "final_action" not in fields:
            report.error(f"{path}: target assembled clock inventory is missing final_action")
            return context
        for row_idx, raw_row in enumerate(reader, start=2):
            row = dict(raw_row)
            action = normalize_key(row.get("final_action"))
            if action and action not in ACTIVE_01_ACTIONS:
                continue
            clock_name = clean_cell(row.get("clock_name"))
            if not clock_name:
                continue
            if clock_name in context.clocks:
                report.error(f"{path.name} row {row_idx}: duplicate active clock_name {clock_name}")
                continue
            info = ClockInfo(
                clock_name=clock_name,
                direct_source=clean_cell(row.get("direct_source")),
                producer_object=clean_cell(row.get("producer_object")),
                final_action=action,
            )
            context.clocks[clock_name] = info
            inst_name = clean_cell(row.get("inst_name"))
            for alias in {
                clock_name,
                clean_cell(row.get("original_clock_name")),
            }:
                if alias:
                    context.aliases[(inst_name, alias)] = clock_name
                    context.aliases.setdefault(("", alias), clock_name)
            context.objects.update(_clock_objects_from_row(row))

    context.clock_set_digest = hashlib.sha256(
        "\n".join(sorted(context.clocks)).encode("utf-8")
    ).hexdigest()
    if meta is not None:
        expected_set_digest = clean_cell(meta.get("clock_set_digest"))
        if expected_set_digest and expected_set_digest != context.clock_set_digest:
            report.error("assembled active clock set digest does not match meta")
        if clean_cell(meta.get("clock_count")) and clean_cell(meta.get("clock_count")) != str(len(context.clocks)):
            report.error("assembled active clock count does not match meta")
        completeness = normalize_key(meta.get("run_completeness"))
        if completeness not in {"complete", "partial"}:
            report.error(f"clock inventory meta has invalid run_completeness {completeness or '<empty>'}")
            context.status = "invalid"
        else:
            context.status = completeness
    else:
        context.status = "complete"
    report.info(f"loaded {len(context.clocks)} assembled clock(s) from {path}")
    return context


def resolve_soc_clock(context: ClockContext, inst_name: str, clock_name: str) -> str:
    name = clean_cell(clock_name)
    if not name:
        return ""
    return context.aliases.get((inst_name, name), context.aliases.get(("", name), name if name in context.clocks else ""))


def read_relation_context(
    path: Path,
    meta_path: Path,
    scenario: str,
    clocks: ClockContext,
    required: bool,
    report: Report,
) -> RelationContext:
    result = RelationContext()
    if not path.is_file() or not meta_path.is_file():
        if required:
            report.error(f"relation map CSV/meta is required for budget output: {path}, {meta_path}")
        return result
    meta = read_json_meta(meta_path, "relation map", required, report)
    if meta is None:
        return result
    if clean_cell(meta.get("schema_version")) != SCHEMA_VERSION:
        report.error("relation map meta has unsupported schema_version")
    if clean_cell(meta.get("scenario")) != scenario:
        report.error(
            f"relation map meta scenario={clean_cell(meta.get('scenario')) or '<empty>'} "
            f"does not match requested {scenario}"
        )
    if clean_cell(meta.get("relation_map_digest")) != digest_file(path):
        report.error("relation map digest does not match meta")
    if clean_cell(meta.get("inventory_digest")) and clean_cell(meta.get("inventory_digest")) != clocks.inventory_digest:
        report.error("relation map references a different 01 inventory digest")
    if clean_cell(meta.get("inventory_meta_digest")) and clean_cell(meta.get("inventory_meta_digest")) != clocks.meta_digest:
        report.error("relation map references a different 01 inventory meta digest")
    if clean_cell(meta.get("clock_universe_digest")) and clean_cell(meta.get("clock_universe_digest")) != clocks.clock_set_digest:
        report.error("relation map clock universe digest does not match 01")

    required_fields = {
        "schema_version", "scenario", "clock_a", "clock_b", "relation_type",
        "relation_source", "source_rule_ids", "clock_universe_digest",
        "assembled_view_digest",
    }
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = set(reader.fieldnames or [])
        missing_fields = sorted(required_fields - fields)
        if missing_fields:
            report.error(f"{path}: relation map missing field(s): {','.join(missing_fields)}")
            return result
        for row_idx, row in enumerate(reader, start=2):
            if clean_cell(row.get("schema_version")) != SCHEMA_VERSION:
                report.error(f"{path.name} row {row_idx}: unsupported schema_version")
                continue
            row_scenario = clean_cell(row.get("scenario"))
            if row_scenario != scenario:
                report.error(f"{path.name} row {row_idx}: scenario mismatch {row_scenario}")
                continue
            clock_a = clean_cell(row.get("clock_a"))
            clock_b = clean_cell(row.get("clock_b"))
            relation = canonical_clock_relation(row.get("relation_type"))
            relation_source = normalize_key(row.get("relation_source"))
            if not clock_a or not clock_b or not relation:
                report.error(f"{path.name} row {row_idx}: invalid clock pair/relation")
                continue
            if relation_source not in RELATION_SOURCES:
                report.error(f"{path.name} row {row_idx}: invalid relation_source {relation_source or '<empty>'}")
            if clean_cell(row.get("clock_universe_digest")) != clocks.clock_set_digest:
                report.error(f"{path.name} row {row_idx}: clock_universe_digest mismatch")
            if clean_cell(row.get("assembled_view_digest")) != clean_cell(meta.get("assembled_view_digest")):
                report.error(f"{path.name} row {row_idx}: assembled_view_digest mismatch")
            key = tuple(sorted((clock_a, clock_b)))
            previous = result.relations.get(key)
            if previous and previous != relation:
                report.error(f"{path.name} row {row_idx}: conflicting relation for {clock_a}/{clock_b}")
            result.relations[key] = relation
    result.status = normalize_key(meta.get("run_completeness")) or "complete"
    if result.status not in {"complete", "partial"}:
        report.error(f"relation map meta has invalid run_completeness {result.status}")
        result.status = "invalid"
    elif clocks.status in {"complete", "partial"} and result.status != clocks.status:
        report.error(
            f"relation map completeness={result.status} does not match 01 assembled clocks {clocks.status}"
        )
    report.info(f"loaded {len(result.relations)} clock relation pair(s) from {path}")
    return result


def create_budget_seeds(
    channels: Sequence[ChannelRecord],
    candidates: Sequence[DelayCandidate],
    exceptions: Sequence[ExceptionEvidence],
    mode: str,
    report: Report,
) -> List[BudgetSeed]:
    src_by_key: Dict[Tuple[str, str], List[DelayCandidate]] = defaultdict(list)
    dst_by_key: Dict[Tuple[str, str], List[DelayCandidate]] = defaultdict(list)
    for cand in candidates:
        if cand.constraint_type == "output_delay":
            src_by_key[(cand.inst_name, cand.port_name)].append(cand)
        elif cand.constraint_type == "input_delay":
            dst_by_key[(cand.inst_name, cand.port_name)].append(cand)

    exception_by_key: Dict[Tuple[str, str], List[ExceptionEvidence]] = defaultdict(list)
    for item in exceptions:
        exception_by_key[(item.inst_name, item.port_name)].append(item)

    seeds: List[BudgetSeed] = []
    matched_candidate_ids: Set[int] = set()

    def candidate_list(index: Dict[Tuple[str, str], List[DelayCandidate]], inst_name: str, port_name: str) -> List[DelayCandidate]:
        result: List[DelayCandidate] = []
        seen: Set[int] = set()
        for key in ((inst_name, port_name), (inst_name, port_base(port_name))):
            for cand in index.get(key, []):
                if id(cand) not in seen:
                    seen.add(id(cand))
                    result.append(cand)
        return result

    for ch in channels:
        source_cands = candidate_list(src_by_key, ch.src_instance, ch.src_port)
        dest_cands = candidate_list(dst_by_key, ch.dst_instance, ch.dst_port)
        related = source_cands + dest_cands
        exception_items: List[ExceptionEvidence] = []
        for key in (
            (ch.src_instance, ch.src_port),
            (ch.src_instance, port_base(ch.src_port)),
            (ch.dst_instance, ch.dst_port),
            (ch.dst_instance, port_base(ch.dst_port)),
        ):
            for item in exception_by_key.get(key, []):
                if item not in exception_items:
                    exception_items.append(item)
        matched_candidate_ids.update(id(cand) for cand in related)
        values = {header: "" for header in BUDGET_HEADERS}
        values.update(
            {
                "channel_id": ch.channel_id,
                "connection_id": ch.connection_id,
                "scenario": ch.scenario,
                "stage": ch.stage,
                "corner": ch.corner,
                "channel_type": ch.channel_type,
                "is_pad_related": ch.is_pad_related,
                "is_clock_related": ch.is_clock_related,
                "is_feedthrough": ch.is_feedthrough,
                "src_endpoint": ch.src_endpoint,
                "dst_endpoint": ch.dst_endpoint,
                "src_sdc_status": ch.src_sdc_status,
                "dst_sdc_status": ch.dst_sdc_status,
                "evidence_status": ch.evidence_status,
                "timing_model": ch.timing_model,
                "budget_required": ch.budget_required,
                "clock_relation": ch.clock_relation,
                "channel_disposition": "pending",
                "budget_model": "unknown",
                "src_output_delay_max": join_unique(c.max_value or c.bare_value for c in source_cands),
                "src_output_delay_min": join_unique(c.min_value for c in source_cands),
                "dst_input_delay_max": join_unique(c.max_value or c.bare_value for c in dest_cands),
                "dst_input_delay_min": join_unique(c.min_value for c in dest_cands),
                "max_source": "",
                "min_source": "",
                "original_src_clock": join_unique(c.clock_name for c in source_cands),
                "original_dst_clock": join_unique(c.clock_name for c in dest_cands),
                "complex_options": join_unique(c.complex_options for c in related),
                "tool_surface": "sta",
                "datapath_only": "yes",
                "source_type": "extracted" if related else "manual",
                "source_sdc_file": join_present(c.source_sdc_file for c in related),
                "source_line": join_present(c.source_line for c in related),
                "source_digest": join_present(c.source_digest for c in related),
                "extraction_time": join_present(c.extraction_time for c in related),
                "source_command": join_unique((c.original_command for c in related), sep=" || "),
                "apply": "no",
                "emit_max": "no",
                "emit_min": "no",
                "review_status": "pending",
                "owner": join_unique(c.owner for c in related),
            }
        )
        messages = [c.message for c in related if c.message]
        if not related:
            messages.append("no input/output delay evidence found on either side")
        if any(c.complex_options for c in related):
            messages.append("complex delay options require review")
        if exception_items:
            commands = join_unique(
                (
                    f"{item.source_sdc_file}:{item.source_line} {item.constraint_type}"
                    for item in exception_items
                )
            )
            values.update(
                {
                    "channel_disposition": "route_to_30",
                    "budget_required": "no",
                    "apply": "yes",
                    "emit_max": "no",
                    "emit_min": "no",
                    "review_status": "approved",
                    "owner": "30_harden_to_harden_exception",
                    "reviewer": POLICY_REVIEWER,
                    "review_date": POLICY_REVIEW_DATE,
                    "disposition_basis": "known_harden_exception_evidence",
                    "source_sdc_file": join_present(
                        list(c.source_sdc_file for c in related)
                        + list(item.source_sdc_file for item in exception_items)
                    ),
                    "source_line": join_present(
                        list(c.source_line for c in related)
                        + list(item.source_line for item in exception_items)
                    ),
                    "source_digest": join_present(
                        list(c.source_digest for c in related)
                        + list(item.source_digest for item in exception_items)
                    ),
                    "source_command": join_unique(
                        list(c.original_command for c in related)
                        + list(item.original_command for item in exception_items),
                        sep=" || ",
                    ),
                }
            )
            messages.append("route_to_30 evidence: " + commands)
            ch.channel_disposition = "route_to_30"
        elif mode == "audit_only":
            values.update(
                {
                    "channel_disposition": "no_soc_budget_required",
                    "budget_required": "no",
                    "apply": "yes",
                    "emit_max": "no",
                    "emit_min": "no",
                    "review_status": "approved",
                    "owner": POLICY_OWNER,
                    "reviewer": POLICY_REVIEWER,
                    "review_date": POLICY_REVIEW_DATE,
                    "disposition_basis": POLICY_ID,
                    "sdc_independent_basis": (
                        SDC_INDEPENDENT_POLICY
                        if ch.evidence_status == "incomplete_missing_sdc"
                        else ""
                    ),
                }
            )
            ch.channel_disposition = "no_soc_budget_required"
        values["note"] = "; ".join(dict.fromkeys(messages))
        status = "ok" if values["channel_disposition"] != "pending" else "needs_review"
        seeds.append(BudgetSeed(values, status, ch, related))
    for cand in candidates:
        if id(cand) not in matched_candidate_ids and cand.direction in {"input", "output"}:
            report.warn(
                f"{cand.source_sdc_file}:{cand.source_line} {cand.inst_name}/{cand.port_name}: "
                f"{cand.constraint_type} candidate has no matching 20 channel in integration table"
            )
    return seeds


def join_unique(values: Iterable[str], sep: str = "; ") -> str:
    result: List[str] = []
    for value in values:
        text = clean_cell(value)
        if text and text not in result:
            result.append(text)
    return sep.join(result)


def join_present(values: Iterable[str], sep: str = "; ") -> str:
    return sep.join(clean_cell(value) for value in values if clean_cell(value))


def create_or_load_workbook(path: Path) -> Tuple[Workbook, bool]:
    if path.is_file():
        return load_workbook(path), False
    wb = Workbook()
    ws = wb.active
    ws.title = "interface_budget"
    return wb, True


def ensure_sheet(wb: Workbook, name: str, headers: Sequence[str]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if ws.max_row == 1 and all(ws.cell(row=1, column=col).value is None for col in range(1, len(headers) + 1)):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    existing = [clean_cell(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
    for header in headers:
        if header not in existing:
            ws.cell(row=1, column=len(existing) + 1, value=header)
            existing.append(header)
    style_sheet(ws)


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for column_cells in ws.columns:
        max_len = 8
        for cell in column_cells:
            max_len = max(max_len, len(clean_cell(cell.value)))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_len + 2, 45)
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:25] or "table"
        if not ws.tables:
            ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"
            tab = Table(displayName=f"{table_name}_tbl", ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            try:
                ws.add_table(tab)
            except ValueError:
                pass


def header_map(ws) -> Dict[str, int]:
    return {clean_cell(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean_cell(cell.value)}


def row_values(ws, row_idx: int, headers: Sequence[str]) -> Dict[str, object]:
    hmap = header_map(ws)
    return {header: ws.cell(row=row_idx, column=hmap[header]).value if header in hmap else "" for header in headers}


def append_dict(ws, headers: Sequence[str], values: Dict[str, object], fill: Optional[PatternFill] = None) -> None:
    hmap = header_map(ws)
    row_idx = ws.max_row + 1
    for header in headers:
        col = hmap[header]
        cell = ws.cell(row=row_idx, column=col, value=values.get(header, ""))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def budget_key(values: Dict[str, object]) -> Tuple[str, str, str, str]:
    return (
        clean_cell(values.get("channel_id")),
        normalize_key(values.get("scenario")) or "common",
        normalize_key(values.get("stage")) or "all",
        clean_cell(values.get("corner")) or "all",
    )


def sync_workbook(
    path: Path,
    channels: Sequence[ChannelRecord],
    seeds: Sequence[BudgetSeed],
    exceptions: Sequence[ExceptionEvidence],
    scenario: str,
    stage: str,
    corner: str,
    mode: str,
    metadata: Dict[str, object],
    report: Report,
) -> None:
    wb, created = create_or_load_workbook(path)
    ensure_sheet(wb, "interface_budget", BUDGET_HEADERS)
    ensure_sheet(wb, "channel_inventory", CHANNEL_HEADERS)
    ensure_sheet(wb, "extraction_log", LOG_HEADERS)
    if "run_metadata" not in wb.sheetnames:
        wb.create_sheet("run_metadata", 0)
    ws_meta = wb["run_metadata"]
    if ws_meta.max_row:
        ws_meta.delete_rows(1, ws_meta.max_row)
    ws_meta.append(["key", "value"])
    for key in sorted(metadata):
        value = metadata[key]
        if isinstance(value, (list, tuple, set)):
            value = ",".join(str(item) for item in value)
        ws_meta.append([key, value])

    ws_budget = wb["interface_budget"]
    existing_budget: Dict[Tuple[str, str, str, str], int] = {}
    for row_idx in range(2, ws_budget.max_row + 1):
        values = row_values(ws_budget, row_idx, BUDGET_HEADERS)
        if clean_cell(values.get("channel_id")):
            key = budget_key(values)
            if key in existing_budget:
                report.error(
                    f"interface_budget rows {existing_budget[key]} and {row_idx} have duplicate channel/view key {key}"
                )
            else:
                existing_budget[key] = row_idx
    current_seed_keys = {budget_key(seed.values) for seed in seeds}
    for seed in seeds:
        key = budget_key(seed.values)
        if key not in existing_budget:
            append_dict(ws_budget, BUDGET_HEADERS, seed.values, NEW_FILL)
            existing_budget[key] = ws_budget.max_row
            report.sync_changed = True
            continue

        row_idx = existing_budget[key]
        current = row_values(ws_budget, row_idx, BUDGET_HEADERS)
        hmap = header_map(ws_budget)
        invalidated = False
        for header in BUDGET_MACHINE_FIELDS:
            new_value = clean_cell(seed.values.get(header))
            if (
                header in {"src_endpoint", "dst_endpoint"}
                and not is_canonical_emit_endpoint(new_value)
                and is_canonical_emit_endpoint(current.get(header))
            ):
                continue
            if clean_cell(current.get(header)) == new_value:
                continue
            ws_budget.cell(row=row_idx, column=hmap[header], value=new_value)
            if header in REVIEW_INVALIDATING_FIELDS:
                invalidated = True
            report.sync_changed = True

        current_disposition = normalize_key(current.get("channel_disposition"))
        has_emit_intent = (
            current_disposition == "emit_budget"
            or normalize_key(current.get("budget_required")) == "yes"
            or normalize_key(current.get("emit_max")) == "yes"
            or normalize_key(current.get("emit_min")) == "yes"
        )
        seed_disposition = normalize_key(seed.values.get("channel_disposition"))
        if seed_disposition == "route_to_30":
            for header in sorted(BUDGET_REVIEW_FIELDS):
                if header not in hmap:
                    continue
                new_value = clean_cell(seed.values.get(header))
                if clean_cell(ws_budget.cell(row=row_idx, column=hmap[header]).value) != new_value:
                    ws_budget.cell(row=row_idx, column=hmap[header], value=new_value)
                    report.sync_changed = True
        elif mode == "audit_only" and not has_emit_intent and current_disposition in {"", "pending", "no_soc_budget_required"}:
            for header in (
                "channel_disposition", "budget_required", "apply", "emit_max", "emit_min",
                "review_status", "owner", "reviewer", "review_date", "disposition_basis",
                "sdc_independent_basis",
            ):
                new_value = clean_cell(seed.values.get(header))
                if clean_cell(ws_budget.cell(row=row_idx, column=hmap[header]).value) != new_value:
                    ws_budget.cell(row=row_idx, column=hmap[header], value=new_value)
                    report.sync_changed = True
        elif invalidated and normalize_key(current.get("review_status")) == "approved":
            for header, value in {
                "apply": "no",
                "emit_max": "no",
                "emit_min": "no",
                "review_status": "pending",
                "converted_max": "",
                "converted_min": "",
                "max_source": "",
                "min_source": "",
                "derivation_basis": "",
            }.items():
                ws_budget.cell(row=row_idx, column=hmap[header], value=value)
            report.warn(f"{key[0]}: machine evidence changed; prior approval reset to pending")
            report.sync_changed = True

    stale_rows: List[int] = []
    for key, row_idx in existing_budget.items():
        if key[1:] == (scenario, stage, corner) and key not in current_seed_keys:
            stale_rows.append(row_idx)
    for row_idx in sorted(stale_rows, reverse=True):
        ws_budget.delete_rows(row_idx, 1)
        report.sync_changed = True

    ws_log = wb["extraction_log"]
    if ws_log.max_row > 1:
        ws_log.delete_rows(2, ws_log.max_row - 1)
    for seed in seeds:
        for cand in seed.candidates:
            values = {
                "source_sdc_file": cand.source_sdc_file,
                "source_line": cand.source_line,
                "instance": cand.inst_name,
                "port": cand.port_name,
                "direction": cand.direction,
                "constraint_type": cand.constraint_type,
                "clock_name": cand.clock_name,
                "min_value": cand.min_value,
                "max_value": cand.max_value or cand.bare_value,
                "parse_status": cand.parse_status,
                "channel_ids": seed.channel.channel_id,
                "source_digest": cand.source_digest,
                "extraction_time": cand.extraction_time,
                "original_command": cand.original_command,
                "message": cand.message,
            }
            append_dict(ws_log, LOG_HEADERS, values)
    for item in exceptions:
        append_dict(
            ws_log,
            LOG_HEADERS,
            {
                "source_sdc_file": item.source_sdc_file,
                "source_line": item.source_line,
                "instance": item.inst_name,
                "port": item.port_name,
                "constraint_type": item.constraint_type,
                "parse_status": "route_to_30_candidate",
                "source_digest": item.source_digest,
                "original_command": item.original_command,
                "message": "known exception evidence",
            },
        )

    add_validations(wb)
    for ws in wb.worksheets:
        style_sheet(ws)
    atomic_save_workbook(wb, path)
    if created or report.sync_changed:
        report.sync_changed = True
        report.info(f"synchronized workbook {path}")


def add_validations(wb: Workbook) -> None:
    if "interface_budget" not in wb.sheetnames:
        return
    ws = wb["interface_budget"]
    ws.data_validations.dataValidation = []
    hmap = header_map(ws)

    def add_list(header: str, values: Sequence[str]) -> None:
        if header not in hmap:
            return
        col = get_column_letter(hmap[header])
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1048576")

    add_list("scenario", sorted(SCENARIOS))
    add_list("stage", sorted(STAGES))
    add_list("channel_type", sorted(CHANNEL_TYPES_20 | NON_20_CHANNEL_TYPES))
    add_list("timing_model", sorted(TIMING_MODELS - {""}))
    add_list("budget_required", ["yes", "no"])
    add_list("channel_disposition", sorted(DISPOSITIONS))
    add_list("budget_model", sorted(BUDGET_MODELS - {""}))
    add_list("tool_surface", sorted(TOOL_SURFACES - {""}))
    add_list("datapath_only", ["yes", "no"])
    add_list("min_sign_review", sorted(MIN_REVIEW_VALUES - {""}))
    add_list("source_type", sorted(SOURCE_TYPES - {""}))
    add_list("apply", ["yes", "no"])
    add_list("emit_max", ["yes", "no"])
    add_list("emit_min", ["yes", "no"])
    add_list("review_status", sorted(REVIEW_STATUS_VALUES - {""}))


def read_form_rows(path: Path) -> List[FormRow]:
    wb = load_workbook(path, data_only=False)
    if "interface_budget" not in wb.sheetnames:
        raise RuntimeError(f"{path} missing interface_budget sheet")
    ws = wb["interface_budget"]
    rows: List[FormRow] = []
    for row_idx in range(2, ws.max_row + 1):
        values = row_values(ws, row_idx, BUDGET_HEADERS)
        if not any(clean_cell(value) for value in values.values()):
            continue
        rows.append(FormRow(row_idx=row_idx, values=values))
    return rows


def write_autofilled_fields(path: Path, rows: Sequence[FormRow], report: Report) -> None:
    changed_rows = [row for row in rows if row.autofilled_fields]
    if not changed_rows:
        return
    wb = load_workbook(path)
    if "interface_budget" not in wb.sheetnames:
        return
    ws = wb["interface_budget"]
    hmap = header_map(ws)
    changed = 0
    for row in changed_rows:
        for field_name in sorted(row.autofilled_fields):
            if field_name not in hmap:
                continue
            new_value = clean_cell(row.values.get(field_name))
            cell = ws.cell(row=row.row_idx, column=hmap[field_name])
            if clean_cell(cell.value) != new_value:
                cell.value = new_value
                cell.fill = NEW_FILL
                changed += 1
    if changed:
        atomic_save_workbook(wb, path)
        report.info(f"wrote {changed} auto-resolved field(s) back to {path.name}")


def row_scenario(row: FormRow) -> str:
    return normalize_key(row.values.get("scenario")) or "common"


def row_stage(row: FormRow) -> str:
    return normalize_key(row.values.get("stage")) or "all"


def row_corner(row: FormRow) -> str:
    return clean_cell(row.values.get("corner")) or "all"


def is_apply_approved(row: FormRow) -> bool:
    return (
        normalize_key(row.values.get("apply")) == "yes"
        and normalize_key(row.values.get("review_status")) == "approved"
    )


def row_selected_for_output(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    return row_scenario(row) == scenario and row_stage(row) == stage and row_corner(row) == corner


def row_selected_for_assembled(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    row_sc = row_scenario(row)
    if row_sc not in {"common", scenario}:
        return False
    row_st = row_stage(row)
    row_co = row_corner(row)
    return (row_st == "all" and row_co == "all") or (row_st == stage and row_co == corner)


def inject_clock_relations(
    rows: Sequence[FormRow],
    scenario: str,
    stage: str,
    corner: str,
    clocks: ClockContext,
    relations: RelationContext,
    report: Report,
) -> None:
    for row in rows:
        if not row_selected_for_assembled(row, scenario, stage, corner):
            continue
        values = row.values
        if normalize_key(values.get("channel_disposition")) != "emit_budget":
            continue
        src_names = [part.strip() for part in clean_cell(values.get("original_src_clock")).split(";") if part.strip()]
        dst_names = [part.strip() for part in clean_cell(values.get("original_dst_clock")).split(";") if part.strip()]
        channel_id = clean_cell(values.get("channel_id"))
        if len(src_names) > 1 or len(dst_names) > 1:
            report.error(f"{channel_id}: multiple source/destination clocks require explicit review")
            continue
        src_inst = ""
        dst_inst = ""
        src_endpoint = clean_cell(values.get("src_endpoint"))
        dst_endpoint = clean_cell(values.get("dst_endpoint"))
        src_objs = endpoint_objects(src_endpoint)
        dst_objs = endpoint_objects(dst_endpoint)
        if src_objs:
            src_inst = sorted(src_objs)[0].split("/", 1)[0]
        if dst_objs:
            dst_inst = sorted(dst_objs)[0].split("/", 1)[0]
        src_clock = resolve_soc_clock(clocks, src_inst, src_names[0]) if src_names else ""
        dst_clock = resolve_soc_clock(clocks, dst_inst, dst_names[0]) if dst_names else ""
        relation = ""
        if src_clock and dst_clock:
            if src_clock == dst_clock:
                relation = "synchronous"
            else:
                relation = relations.relations.get(tuple(sorted((src_clock, dst_clock))), "unknown")
        elif src_names or dst_names:
            relation = "unknown"
        if relation and canonical_clock_relation(values.get("clock_relation")) != relation:
            values["clock_relation"] = relation
            row.autofilled_fields.add("clock_relation")


def validate_rows(
    rows: Sequence[FormRow],
    channels: Sequence[ChannelRecord],
    scenario: str,
    stage: str,
    corner: str,
    current_digests: Dict[str, str],
    max_diff_threshold: Optional[float],
    mode: str,
    report: Report,
) -> None:
    channel_by_id = {ch.channel_id: ch for ch in channels}
    assembled = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner)]
    emitted_by_channel: Dict[str, List[FormRow]] = defaultdict(list)

    for row in assembled:
        values = row.values
        apply_value = normalize_key(values.get("apply"))
        review_status = normalize_key(values.get("review_status"))
        source_type = normalize_key(values.get("source_type"))
        channel_id = clean_cell(values.get("channel_id"))
        channel_type = normalize_key(values.get("channel_type"))
        timing_model = normalize_key(values.get("timing_model"))
        disposition = normalize_key(values.get("channel_disposition"))
        budget_model = normalize_key(values.get("budget_model"))
        budget_required = normalize_key(values.get("budget_required"))
        tool_surface = normalize_key(values.get("tool_surface"))
        datapath_only = normalize_key(values.get("datapath_only"))
        clock_relation_raw = normalize_key(values.get("clock_relation"))
        clock_relation = canonical_clock_relation(values.get("clock_relation"))
        relation_basis = clean_cell(values.get("relationship_override_basis"))
        channel = channel_by_id.get(channel_id)
        is_pad_related = normalize_key(values.get("is_pad_related")) or (channel.is_pad_related if channel else "")
        is_clock_related = normalize_key(values.get("is_clock_related")) or (channel.is_clock_related if channel else "")
        is_feedthrough = normalize_key(values.get("is_feedthrough")) or (channel.is_feedthrough if channel else "")

        if apply_value and apply_value not in YES_NO:
            report.error(f"interface_budget row {row.row_idx}: apply must be yes/no")
        if review_status and review_status not in REVIEW_STATUS_VALUES:
            report.error(f"interface_budget row {row.row_idx}: invalid review_status {review_status}")
        if source_type and source_type not in SOURCE_TYPES:
            report.error(f"interface_budget row {row.row_idx}: invalid source_type {source_type}")
        if tool_surface and tool_surface not in TOOL_SURFACES:
            report.error(f"interface_budget row {row.row_idx}: invalid tool_surface {tool_surface}")
        if timing_model and timing_model not in TIMING_MODELS:
            report.error(f"interface_budget row {row.row_idx}: invalid timing_model {timing_model}")
        if disposition not in DISPOSITIONS:
            report.error(f"interface_budget row {row.row_idx}: invalid channel_disposition {disposition or '<empty>'}")
        if budget_model and budget_model not in BUDGET_MODELS:
            report.error(f"interface_budget row {row.row_idx}: invalid budget_model {budget_model}")
        if clock_relation_raw and not clock_relation:
            report.error(
                f"interface_budget row {row.row_idx}: invalid clock_relation {clock_relation_raw}; "
                "use synchronous/asynchronous/logically_exclusive/physically_exclusive/unknown"
            )
        for flag_name, flag_value in (
            ("is_pad_related", is_pad_related),
            ("is_clock_related", is_clock_related),
            ("is_feedthrough", is_feedthrough),
        ):
            if flag_value and flag_value not in YES_NO:
                report.error(f"interface_budget row {row.row_idx}: {flag_name} must be yes/no")

        if not channel_id or channel is None:
            report.error(f"interface_budget row {row.row_idx}: channel_id {channel_id or '<empty>'} not found in current channel inventory")
            continue
        if clean_cell(values.get("connection_id")) != channel.connection_id:
            report.error(f"interface_budget row {row.row_idx} {channel_id}: connection_id differs from 00 machine edge")
        src_endpoint = clean_cell(values.get("src_endpoint"))
        dst_endpoint = clean_cell(values.get("dst_endpoint"))
        if not src_endpoint or not dst_endpoint:
            report.error(f"interface_budget row {row.row_idx} {channel_id}: src/dst endpoint is required")
        src_override_ok = not is_canonical_emit_endpoint(channel.src_endpoint) and is_canonical_emit_endpoint(src_endpoint)
        dst_override_ok = not is_canonical_emit_endpoint(channel.dst_endpoint) and is_canonical_emit_endpoint(dst_endpoint)
        if (src_endpoint != channel.src_endpoint and not src_override_ok) or (dst_endpoint != channel.dst_endpoint and not dst_override_ok):
            report.error(f"interface_budget row {row.row_idx} {channel_id}: src/dst endpoint differs from 00 edge without a canonical fabric override")
        if channel_type != channel.channel_type or channel_type not in CHANNEL_TYPES_20:
            report.error(f"interface_budget row {row.row_idx} {channel_id}: invalid channel_type {channel_type}")
        if is_pad_related == "yes" or is_clock_related == "yes" or is_feedthrough == "yes":
            report.error(f"interface_budget row {row.row_idx} {channel_id}: non-20 ownership flag is set")
        if channel.channel_disposition == "route_to_30" and disposition != "route_to_30":
            report.error(f"interface_budget row {row.row_idx} {channel_id}: known exception evidence requires route_to_30")

        emit_max = normalize_key(values.get("emit_max"))
        emit_min = normalize_key(values.get("emit_min"))
        has_emit_intent = (
            disposition == "emit_budget"
            or budget_required == "yes"
            or emit_max == "yes"
            or emit_min == "yes"
        )
        if mode == "audit_only" and has_emit_intent:
            report.error(
                f"interface_budget row {row.row_idx} {channel_id}: audit_only forbids emit intent "
                f"(disposition={disposition}, budget_required={budget_required or '<empty>'}, "
                f"emit_max={emit_max or '<empty>'}, emit_min={emit_min or '<empty>'})"
            )

        if disposition == "no_soc_budget_required":
            missing_fields = []
            if apply_value != "yes":
                missing_fields.append("apply=yes")
            if review_status != "approved":
                missing_fields.append("review_status=approved")
            if budget_required != "no":
                missing_fields.append("budget_required=no")
            if emit_max != "no" or emit_min != "no":
                missing_fields.append("emit_max/emit_min=no")
            if not clean_cell(values.get("disposition_basis")):
                missing_fields.append("disposition_basis")
            if "missing" in {
                normalize_key(values.get("src_sdc_status")),
                normalize_key(values.get("dst_sdc_status")),
            } and not clean_cell(values.get("sdc_independent_basis")):
                missing_fields.append("sdc_independent_basis")
            if mode == "audit_only" and clean_cell(values.get("disposition_basis")) != POLICY_ID:
                missing_fields.append(f"disposition_basis={POLICY_ID}")
            if missing_fields:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: no-budget terminal row missing {', '.join(missing_fields)}")
        elif disposition == "not_applicable":
            required_values = {
                "apply": apply_value == "yes",
                "review_status": review_status == "approved",
                "owner": bool(clean_cell(values.get("owner"))),
                "reviewer": bool(clean_cell(values.get("reviewer"))),
                "review_date": bool(clean_cell(values.get("review_date"))),
                "disposition_basis": bool(clean_cell(values.get("disposition_basis"))),
            }
            if "missing" in {
                normalize_key(values.get("src_sdc_status")),
                normalize_key(values.get("dst_sdc_status")),
            }:
                required_values["sdc_independent_basis"] = bool(clean_cell(values.get("sdc_independent_basis")))
            missing = [name for name, ok in required_values.items() if not ok]
            if missing:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: not_applicable missing {', '.join(missing)}")
        elif disposition in {"route_to_30", "pending"}:
            if emit_max == "yes" or emit_min == "yes":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: {disposition} cannot emit")
        elif disposition == "emit_budget":
            if mode != "budget_output":
                continue
            if not is_canonical_emit_endpoint(values.get("src_endpoint")) or not is_canonical_emit_endpoint(values.get("dst_endpoint")):
                report.error(f"interface_budget row {row.row_idx} {channel_id}: emit endpoints must be exact get_pins/get_ports collections")
            if not is_apply_approved(row):
                report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_budget requires apply=yes/review_status=approved")
            if budget_required != "yes":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_budget requires budget_required=yes")
            if "missing" in {
                normalize_key(values.get("src_sdc_status")),
                normalize_key(values.get("dst_sdc_status")),
            } and not clean_cell(values.get("sdc_independent_basis")):
                report.error(f"interface_budget row {row.row_idx} {channel_id}: missing harden SDC requires sdc_independent_basis")
            if not budget_model or budget_model == "unknown":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_budget requires reviewed budget_model")
            if budget_model == "clock_relative_io_delay" and not clean_cell(values.get("derivation_basis")):
                report.error(f"interface_budget row {row.row_idx} {channel_id}: clock_relative_io_delay requires derivation_basis")
            auto_resolve_interconnect_max(row, report)
            if relation_blocks_20(clock_relation) and not relation_basis:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: clock_relation={clock_relation} blocks normal 20 budget")
            if (clean_cell(values.get("original_src_clock")) or clean_cell(values.get("original_dst_clock"))) and clock_relation in {"", "unknown"} and not relation_basis:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: clock relation is unknown/incomplete")
            if emit_max != "yes" and emit_min != "yes":
                report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_budget must emit max and/or min")
            if emit_max == "yes" and parse_number(values.get("converted_max")) is None:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_max requires finite converted_max")
            if emit_min == "yes":
                if parse_number(values.get("converted_min")) is None:
                    report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_min requires finite converted_min")
                if normalize_key(values.get("min_sign_review")) not in MIN_REVIEW_VALUES - {""}:
                    report.error(f"interface_budget row {row.row_idx} {channel_id}: emit_min requires min_sign_review")
            if tool_surface not in {"sta", "dc", "both"}:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: tool_surface is required")
            if datapath_only not in {"yes", "no"}:
                report.error(f"interface_budget row {row.row_idx} {channel_id}: datapath_only strategy is required")
            if not clean_cell(values.get("budget_basis")):
                report.error(f"interface_budget row {row.row_idx} {channel_id}: budget_basis is required")
            check_source_digest(row, current_digests, report, fatal=True)
            emitted_by_channel[channel_id].append(row)
            warn_budget_diff(row, max_diff_threshold, report)
            warn_budget_shape(row, report)
            if normalize_key(values.get("budget_model")) == "interconnect_budget" and "interconnect" not in normalize_key(values.get("budget_basis")):
                report.warn(
                    f"interface_budget row {row.row_idx} {channel_id}: interconnect_budget should cite owner definition in budget_basis"
                )
            if normalize_key(values.get("budget_model")) == "manual_budget" and not clean_cell(values.get("derivation_basis")):
                report.warn(f"interface_budget row {row.row_idx} {channel_id}: manual_budget should include derivation_basis")

    for channel_id, group in emitted_by_channel.items():
        max_values = {clean_cell(row.values.get("converted_max")) for row in group if normalize_key(row.values.get("emit_max")) == "yes"}
        min_values = {clean_cell(row.values.get("converted_min")) for row in group if normalize_key(row.values.get("emit_min")) == "yes"}
        if len(max_values - {""}) > 1 or len(min_values - {""}) > 1:
            rows_text = ", ".join(str(row.row_idx) for row in group)
            report.error(f"assembled view conflict for channel {channel_id}: rows {rows_text}")
    warn_manual_fanout_reuse([row for group in emitted_by_channel.values() for row in group], report)


def relation_blocks_20(value: str) -> bool:
    relation = canonical_clock_relation(value)
    if not relation:
        return False
    return relation in RELATION_BLOCKING


def is_canonical_emit_endpoint(value: str) -> bool:
    parsed = parse_collection(clean_cell(value))
    if not parsed or parsed[0] not in {"get_pins", "get_ports"} or len(parsed[1]) != 1:
        return False
    obj = parsed[1][0]
    if any(char in obj for char in "*?"):
        return False
    port = obj.rsplit("/", 1)[-1]
    return is_canonical_port_key(port)


def warn_manual_fanout_reuse(rows: Sequence[FormRow], report: Report) -> None:
    by_src_budget: Dict[Tuple[str, str], List[FormRow]] = defaultdict(list)
    for row in rows:
        values = row.values
        if normalize_key(values.get("budget_model")) != "manual_budget":
            continue
        src = clean_cell(values.get("src_endpoint"))
        budget = clean_cell(values.get("converted_max"))
        if src and budget:
            by_src_budget[(src, budget)].append(row)
    for (src, budget), group in by_src_budget.items():
        dsts = {clean_cell(row.values.get("dst_endpoint")) for row in group}
        if len(dsts) > 1:
            rows_text = ", ".join(str(row.row_idx) for row in group)
            report.warn(
                f"manual_budget fanout reuse: src {src} uses converted_max={budget} "
                f"for {len(dsts)} sinks in rows {rows_text}; confirm shared budget is intentional"
            )


def numeric_parts(value) -> List[float]:
    numbers: List[float] = []
    for part in clean_cell(value).split(";"):
        number = parse_number(part)
        if number is not None:
            numbers.append(number)
    return numbers


def auto_resolve_interconnect_max(row: FormRow, report: Report) -> None:
    values = row.values
    if normalize_key(values.get("budget_model")) != "interconnect_budget":
        return
    if normalize_key(values.get("emit_max")) != "yes":
        return
    if clean_cell(values.get("converted_max")):
        return
    if clean_cell(values.get("complex_options")):
        report.warn(
            f"interface_budget row {row.row_idx} {clean_cell(values.get('channel_id'))}: "
            "complex delay options present; converted_max is not auto-resolved"
        )
        return
    src_nums = numeric_parts(values.get("src_output_delay_max"))
    dst_nums = numeric_parts(values.get("dst_input_delay_max"))
    if not src_nums or not dst_nums:
        return
    converted = min(src_nums + dst_nums)
    values["converted_max"] = format_number(converted)
    row.autofilled_fields.add("converted_max")
    if not clean_cell(values.get("max_source")):
        values["max_source"] = "auto_min_interconnect_budget"
        row.autofilled_fields.add("max_source")
    if not clean_cell(values.get("derivation_basis")):
        values["derivation_basis"] = "min(src_output_delay_max,dst_input_delay_max)"
        row.autofilled_fields.add("derivation_basis")
    report.info(
        f"interface_budget row {row.row_idx} {clean_cell(values.get('channel_id'))}: "
        f"auto-resolved converted_max={format_number(converted)} from interconnect budget candidates"
    )


def warn_budget_diff(row: FormRow, threshold: Optional[float], report: Report) -> None:
    if threshold is None:
        return
    values = row.values
    nums = numeric_parts(values.get("src_output_delay_max")) + numeric_parts(values.get("dst_input_delay_max"))
    if len(nums) >= 2 and max(nums) - min(nums) > threshold:
        report.warn(
            f"interface_budget row {row.row_idx} {clean_cell(values.get('channel_id'))}: "
            f"max candidates differ by {max(nums) - min(nums):.12g}"
        )


def warn_budget_shape(row: FormRow, report: Report) -> None:
    values = row.values
    channel_id = clean_cell(values.get("channel_id"))
    src_max = numeric_parts(values.get("src_output_delay_max"))
    dst_max = numeric_parts(values.get("dst_input_delay_max"))
    if normalize_key(values.get("budget_model")) == "interconnect_budget":
        if bool(src_max) != bool(dst_max):
            report.warn(
                f"interface_budget row {row.row_idx} {channel_id}: "
                "interconnect_budget has max candidate from only one side"
            )
    if any(number < 0 for number in numeric_parts(values.get("src_output_delay_min")) + numeric_parts(values.get("dst_input_delay_min"))):
        report.warn(
            f"interface_budget row {row.row_idx} {channel_id}: original min delay contains negative value; "
            "review set_min_delay sign convention carefully"
        )
    complex_options = clean_cell(values.get("complex_options"))
    if complex_options:
        report.warn(
            f"interface_budget row {row.row_idx} {channel_id}: approved row contains complex delay option(s): {complex_options}"
        )
    src_clock = clean_cell(values.get("original_src_clock"))
    dst_clock = clean_cell(values.get("original_dst_clock"))
    if src_clock and dst_clock and src_clock != dst_clock:
        report.warn(
            f"interface_budget row {row.row_idx} {channel_id}: source/destination original clocks differ "
            f"({src_clock} vs {dst_clock})"
        )
    if normalize_key(values.get("min_sign_review")) == "waived" and not clean_cell(values.get("note")):
        report.warn(
            f"interface_budget row {row.row_idx} {channel_id}: min_sign_review=waived should include note/basis"
        )


def check_source_digest(
    row: FormRow,
    current_digests: Dict[str, str],
    report: Report,
    fatal: bool = False,
) -> None:
    sources = [part.strip() for part in clean_cell(row.values.get("source_sdc_file")).split(";") if part.strip()]
    digests = [part.strip() for part in clean_cell(row.values.get("source_digest")).split(";") if part.strip()]
    for idx, source in enumerate(sources):
        stored = digests[idx] if idx < len(digests) else ""
        current = current_digests.get(source)
        if not stored or not current or current != stored:
            message = (
                f"interface_budget row {row.row_idx} {clean_cell(row.values.get('channel_id'))}: "
                f"source_digest is missing/stale for {source}"
            )
            if fatal:
                report.error(message)
            else:
                report.warn(message)


def output_sdc_path(cwd: Path, scenario: str, stage: str, corner: str) -> Path:
    if scenario == "common":
        if stage == "all" and corner == "all":
            return cwd / "common/20_harden_x_if.sdc"
        return cwd / f"common/20_harden_x_if_{stage}_{safe_filename_token(corner)}.sdc"
    if stage == "all" and corner == "all":
        return cwd / f"scenarios/{scenario}_harden_x_if.sdc"
    return cwd / f"scenarios/{scenario}_harden_x_if_{stage}_{safe_filename_token(corner)}.sdc"


def generate_sdc(
    rows: Sequence[FormRow],
    scenario: str,
    stage: str,
    corner: str,
    mode: str,
    completeness: RunCompleteness,
    accounting_enabled: bool,
    connection_path: Path,
    manifest_path: Optional[Path],
) -> List[str]:
    selected = [
        row
        for row in rows
        if row_selected_for_output(row, scenario, stage, corner)
        and is_apply_approved(row)
        and normalize_key(row.values.get("channel_disposition")) == "emit_budget"
    ]
    lines = [
        "################################################################################",
        "# 20 SoC harden/subsys interface stage",
        f"# Author: {author_name()}",
        "# Stage: 20_harden_x_if",
        "# Script: 20_extract_harden_x_if.py",
        f"# Mode: {mode}",
        f"# Scenario: {scenario}",
        f"# View: stage={stage}, corner={corner}",
        f"# SDC consumption: {'disabled' if mode == 'audit_only' else 'enabled'}",
        f"# Run completeness: {completeness.status}",
        f"# Port accounting: {'enabled' if accounting_enabled else 'disabled by explicit option'}",
        f"# Connection inventory: {connection_path}",
        f"# Harden SDC manifest: {manifest_path or '<legacy inference>'}",
        f"# Policy: {POLICY_ID}",
        "# Source: 20_harden_x_if.xlsx interface_budget sheet",
        "################################################################################",
        "",
    ]
    if mode == "audit_only":
        lines.append("# No timing constraints emitted.")
        return lines
    emitted = 0
    for row in selected:
        values = row.values
        commands = commands_for_row(row)
        if not commands:
            continue
        lines.append(f"# row {row.row_idx}: {clean_cell(values.get('channel_id'))}")
        if clean_cell(values.get("budget_basis")):
            lines.append(f"# Budget basis: {clean_cell(values.get('budget_basis'))}")
        if clean_cell(values.get("derivation_basis")):
            lines.append(f"# Derivation: {clean_cell(values.get('derivation_basis'))}")
        source_refs = format_source_refs(values.get("source_sdc_file"), values.get("source_line"))
        if source_refs:
            lines.append(f"# Source SDC: {source_refs}")
        lines.extend(commands)
        lines.append("")
        emitted += len(commands)
    if emitted == 0:
        lines.append("# No harden/subsys interface budget commands emitted for selected scenario/stage/corner.")
    return lines


def format_source_refs(source_files, source_lines) -> str:
    files = [part.strip() for part in clean_cell(source_files).split(";") if part.strip()]
    lines = [part.strip() for part in clean_cell(source_lines).split(";") if part.strip()]
    if not files:
        return ""
    if len(files) == len(lines):
        return "; ".join(f"{source}:{line}" if line else source for source, line in zip(files, lines))
    if lines:
        return f"{'; '.join(files)} lines {'; '.join(lines)}"
    return "; ".join(files)


def commands_for_row(row: FormRow) -> List[str]:
    values = row.values
    if normalize_key(values.get("channel_disposition")) != "emit_budget" or not is_apply_approved(row):
        return []
    src = clean_cell(values.get("src_endpoint"))
    dst = clean_cell(values.get("dst_endpoint"))
    if not src or not dst:
        return []
    datapath = " -datapath_only" if normalize_key(values.get("datapath_only")) != "no" else ""
    commands: List[str] = []
    if normalize_key(values.get("emit_max")) == "yes" and parse_number(values.get("converted_max")) is not None:
        commands.append(f"set_max_delay {format_number(values.get('converted_max'))}{datapath} -from {src} -to {dst}")
    if normalize_key(values.get("emit_min")) == "yes" and parse_number(values.get("converted_min")) is not None:
        commands.append(f"set_min_delay {format_number(values.get('converted_min'))}{datapath} -from {src} -to {dst}")
    return commands


def _row_resolution_rank(row: FormRow, scenario: str, stage: str, corner: str) -> Tuple[int, int]:
    scenario_rank = 1 if row_scenario(row) == scenario else 0
    view_rank = 1 if row_stage(row) == stage and row_corner(row) == corner else 0
    return scenario_rank, view_rank


def build_resolved_channel_rows(
    channels: Sequence[ChannelRecord],
    rows: Sequence[FormRow],
    scenario: str,
    stage: str,
    corner: str,
    mode: str,
    completeness: RunCompleteness,
    connection_digest: str,
    report: Report,
) -> List[Dict[str, object]]:
    active_by_channel: Dict[str, List[FormRow]] = defaultdict(list)
    for row in rows:
        if row_selected_for_assembled(row, scenario, stage, corner):
            active_by_channel[clean_cell(row.values.get("channel_id"))].append(row)

    result: List[Dict[str, object]] = []
    for channel in sorted(channels, key=lambda item: (item.connection_id, item.channel_id)):
        candidates = active_by_channel.get(channel.channel_id, [])
        selected: Optional[FormRow] = None
        if candidates:
            ranked = sorted(
                candidates,
                key=lambda item: (_row_resolution_rank(item, scenario, stage, corner), -item.row_idx),
                reverse=True,
            )
            selected = ranked[0]
            if len(ranked) > 1 and _row_resolution_rank(ranked[0], scenario, stage, corner) == _row_resolution_rank(ranked[1], scenario, stage, corner):
                report.error(f"{channel.channel_id}: multiple equally specific active workbook rows")
        values = selected.values if selected else {}
        disposition = normalize_key(values.get("channel_disposition")) or "pending"
        machine = {header: "" for header in CHANNEL_HEADERS}
        machine.update(
            {
                "schema_version": SCHEMA_VERSION,
                "author": author_name(),
                "mode": mode,
                "sdc_consumption": "disabled" if mode == "audit_only" else "enabled",
                "run_completeness": completeness.status,
                "available_harden_count": str(completeness.available_count),
                "missing_harden_count": str(completeness.missing_count),
                "not_required_harden_count": str(completeness.not_required_count),
                "missing_instances": ",".join(completeness.missing_instances),
                "connection_inventory_digest": connection_digest,
                "channel_id": channel.channel_id,
                "connection_id": channel.connection_id,
                "scenario": scenario,
                "stage": stage,
                "corner": corner,
                "channel_type": channel.channel_type,
                "owner_stage": "20",
                "src_instance": channel.src_instance,
                "src_module": channel.src_module,
                "src_direction": channel.src_direction,
                "src_port": channel.src_port,
                "src_bit_index": channel.src_bit_index,
                "src_endpoint": clean_cell(values.get("src_endpoint")) or channel.src_endpoint,
                "src_sdc_status": channel.src_sdc_status,
                "dst_instance": channel.dst_instance,
                "dst_module": channel.dst_module,
                "dst_direction": channel.dst_direction,
                "dst_port": channel.dst_port,
                "dst_bit_index": channel.dst_bit_index,
                "dst_endpoint": clean_cell(values.get("dst_endpoint")) or channel.dst_endpoint,
                "dst_sdc_status": channel.dst_sdc_status,
                "connection_source": channel.connection_source,
                "is_pad_related": channel.is_pad_related,
                "is_clock_related": channel.is_clock_related,
                "is_feedthrough": channel.is_feedthrough,
                "evidence_status": clean_cell(values.get("evidence_status")) or channel.evidence_status,
                "budget_required": clean_cell(values.get("budget_required")),
                "clock_relation": canonical_clock_relation(values.get("clock_relation")),
                "channel_disposition": disposition,
                "budget_type": "channel_datapath_budget" if disposition == "emit_budget" else "none",
                "budget_model": clean_cell(values.get("budget_model")),
                "apply": clean_cell(values.get("apply")),
                "review_status": clean_cell(values.get("review_status")),
                "emit_max": clean_cell(values.get("emit_max")),
                "emit_min": clean_cell(values.get("emit_min")),
                "converted_max": clean_cell(values.get("converted_max")),
                "converted_min": clean_cell(values.get("converted_min")),
                "disposition_basis": clean_cell(values.get("disposition_basis")),
                "sdc_independent_basis": clean_cell(values.get("sdc_independent_basis")),
                "note": clean_cell(values.get("note")) or channel.note,
            }
        )
        result.append(machine)
    return result


def write_resolved_channel_artifacts(
    form_path: Path,
    inventory_path: Path,
    inventory_meta_path: Path,
    resolved_rows: Sequence[Dict[str, object]],
    metadata: Dict[str, object],
) -> None:
    ordered = sorted(
        resolved_rows,
        key=lambda values: (
            clean_cell(values.get("connection_id")),
            clean_cell(values.get("channel_id")),
        ),
    )
    workbook = load_workbook(form_path)
    ensure_sheet(workbook, "channel_inventory", CHANNEL_HEADERS)
    sheet = workbook["channel_inventory"]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for values in ordered:
        append_dict(sheet, CHANNEL_HEADERS, values)
    style_sheet(sheet)
    atomic_save_workbook(workbook, form_path)

    atomic_write_csv(inventory_path, CHANNEL_HEADERS, ordered)
    inventory_digest = digest_file(inventory_path)
    payload = dict(metadata)
    payload.update(
        {
            "schema_version": SCHEMA_VERSION,
            "author": author_name(),
            "stage": "20_harden_x_if",
            "script": "20_extract_harden_x_if.py",
            "channel_count": len(ordered),
            "channel_inventory_path": str(inventory_path.resolve()),
            "channel_inventory_digest": inventory_digest,
        }
    )
    atomic_write_json(inventory_meta_path, payload)


def build_coverage_lines(
    rows: Sequence[FormRow],
    channels: Sequence[ChannelRecord],
    scenario: str,
    stage: str,
    corner: str,
    mode: str,
    accounting_enabled: bool,
) -> List[str]:
    assembled = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner)]
    by_channel: Dict[str, List[FormRow]] = defaultdict(list)
    for row in assembled:
        by_channel[clean_cell(row.values.get("channel_id"))].append(row)

    disposition_counts: Dict[str, int] = defaultdict(int)
    timing_command_count = 0
    terminal_count = 0
    incomplete_count = 0
    for row in assembled:
        disposition = normalize_key(row.values.get("channel_disposition")) or "<blank>"
        disposition_counts[disposition] += 1
        timing_command_count += len(commands_for_row(row))
        if disposition in {"no_soc_budget_required", "not_applicable"} and is_apply_approved(row):
            terminal_count += 1
        elif disposition == "emit_budget" and commands_for_row(row):
            terminal_count += 1
        if normalize_key(row.values.get("evidence_status")) == "incomplete_missing_sdc":
            incomplete_count += 1

    lines = [
        "",
        "Coverage:",
        f"  mode: {mode}",
        f"  SDC consumption: {'disabled' if mode == 'audit_only' else 'enabled'}",
        f"  Port accounting: {'enabled' if accounting_enabled else 'disabled by explicit option'}",
        f"  inventory direct bit channels: {len(channels)}",
        f"  assembled workbook rows: {len(assembled)}",
        f"  terminal rows: {terminal_count}",
        f"  incomplete exception evidence rows: {incomplete_count}",
        f"  timing-command count: {timing_command_count}",
        "  dispositions:",
    ]
    for disposition in sorted(disposition_counts):
        lines.append(f"    {disposition}: {disposition_counts[disposition]}")
    lines.extend([
        "",
        "  Per-channel assembled status:",
    ])
    if not channels:
        lines.append("    <no channel_inventory records>")
    for ch in channels:
        group = by_channel.get(ch.channel_id, [])
        if group:
            statuses = []
            for row in group:
                values = row.values
                statuses.append(
                    "row={row} disposition={disposition} review={review} evidence={evidence} emit={emit}".format(
                        row=row.row_idx,
                        disposition=normalize_key(values.get("channel_disposition")) or "-",
                        review=normalize_key(values.get("review_status")) or "-",
                        evidence=normalize_key(values.get("evidence_status")) or "-",
                        emit="+".join(
                            name
                            for name, flag in (
                                ("max", values.get("emit_max")),
                                ("min", values.get("emit_min")),
                            )
                            if normalize_key(flag) == "yes"
                        ) or "none",
                    )
                )
            status = "; ".join(statuses)
        else:
            status = "NO_CURRENT_VIEW_ROW"
        lines.append(
            f"    {ch.channel_id}: connection={ch.connection_id} type={ch.channel_type} "
            f"src_sdc={ch.src_sdc_status} dst_sdc={ch.dst_sdc_status} status={status}"
        )
    return lines


def write_report(
    path: Path,
    report: Report,
    scenario: str,
    stage: str,
    corner: str,
    mode: str,
    completeness: RunCompleteness,
    accounting_enabled: bool,
    form_path: Path,
    inventory_path: Path,
    output_path: Path,
    connection_path: Path,
    manifest_path: Optional[Path],
    coverage_lines: Sequence[str],
) -> None:
    lines = [
        "20_harden_x_if extraction report",
        "================================",
        "",
        f"Author: {author_name()}",
        "Stage: 20_harden_x_if",
        "Script: 20_extract_harden_x_if.py",
        f"Scenario: {scenario}",
        f"View: stage={stage}, corner={corner}",
        f"Mode: {mode}",
        f"SDC consumption: {'disabled' if mode == 'audit_only' else 'enabled'}",
        f"Port accounting: {'enabled' if accounting_enabled else 'disabled by explicit option'}",
        f"Run completeness: {completeness.status}",
        f"Available harden SDC: {completeness.available_count}",
        f"Missing harden SDC: {completeness.missing_count}",
        f"Not-required harden SDC: {completeness.not_required_count}",
        f"Missing instances: {','.join(completeness.missing_instances) or '<none>'}",
        f"Connection inventory: {connection_path}",
        f"Harden SDC manifest: {manifest_path or '<legacy inference>'}",
        f"Form: {form_path}",
        f"Inventory: {inventory_path}",
        f"Output: {output_path}",
        f"Warnings: {report.warning_count}",
        f"Errors  : {report.error_count}",
        f"Sync changed: {'yes' if report.sync_changed else 'no'}",
        "",
        "Messages:",
    ]
    lines.extend(report.lines or ["INFO: no messages"])
    lines.extend(coverage_lines)
    atomic_write_text(path, "\n".join(lines).rstrip() + "\n")


def pending_line_key(line: str) -> Optional[Tuple[str, str]]:
    stripped = line.strip()
    if not stripped or stripped.startswith("#"):
        return None
    parts = stripped.split()
    if len(parts) < 2:
        return None
    direction, port = parts[0], parts[1]
    if direction not in {"input", "output", "inout"}:
        return None
    return direction, port


def removed_line_key(line: str) -> Optional[PortKey]:
    stripped = line.strip()
    if not stripped or stripped.startswith("#"):
        return None
    parts = stripped.split()
    if len(parts) < 3:
        return None
    inst_name, direction, port = parts[:3]
    if direction not in {"input", "output", "inout"}:
        return None
    return PortKey(inst_name, direction, port)


def read_removed_keys(paths: Sequence[Path]) -> Set[PortKey]:
    keys: Set[PortKey] = set()
    for path in paths:
        if not path.is_file():
            continue
        try:
            lines = path.read_text(encoding="utf-8").splitlines()
        except OSError:
            continue
        for line in lines:
            key = removed_line_key(line)
            if key is not None:
                keys.add(key)
    return keys


def harden_pending_key(inst_name: str, direction: str, port_name: str) -> Optional[PortKey]:
    if not inst_name or normalize_key(inst_name) in {"top", "fabric", "unknown", "constant", "const"}:
        return None
    if direction not in {"input", "output", "inout"}:
        return None
    if not port_name:
        return None
    return PortKey(inst_name, direction, port_name)


def terminal_rows_for_pending(rows: Sequence[FormRow], scenario: str, stage: str, corner: str) -> List[FormRow]:
    result = []
    for row in rows:
        if not row_selected_for_assembled(row, scenario, stage, corner) or not is_apply_approved(row):
            continue
        disposition = normalize_key(row.values.get("channel_disposition"))
        if disposition in {"no_soc_budget_required", "not_applicable"}:
            result.append(row)
        elif disposition == "emit_budget" and commands_for_row(row):
            result.append(row)
    return result


def removed_log_line_20(row: FormRow, ch: ChannelRecord, key: PortKey, mode: str) -> str:
    values = row.values
    emit = []
    if normalize_key(values.get("emit_max")) == "yes":
        emit.append("max")
    if normalize_key(values.get("emit_min")) == "yes":
        emit.append("min")
    return " ".join(
        [
            key.inst_name,
            key.direction,
            key.port_name,
            "covered_by=20_harden_x_if",
            f"reason={normalize_key(values.get('channel_disposition'))}",
            f"channel={ch.channel_id}",
            f"connection={ch.connection_id}",
            f"scenario={row_scenario(row)}",
            f"stage={row_stage(row)}",
            f"corner={row_corner(row)}",
            f"mode={mode}",
            f"basis={sanitize_id(clean_cell(values.get('disposition_basis')) or clean_cell(values.get('budget_basis')))}",
            f"emit={'+'.join(emit) or 'none'}",
        ]
    )


def prepare_pending_plan(
    pending_dir: Path,
    removed_log_path: Path,
    previous_removed_paths: Sequence[Path],
    rows: Sequence[FormRow],
    channels: Sequence[ChannelRecord],
    scenario: str,
    stage: str,
    corner: str,
    mode: str,
    report: Report,
) -> PendingPlan:
    plan = PendingPlan()
    if not pending_dir.exists():
        report.error(f"pending directory not found: {pending_dir}")
        return plan
    if not pending_dir.is_dir():
        report.error(f"pending path is not a directory: {pending_dir}")
        return plan

    channel_by_id = {ch.channel_id: ch for ch in channels}
    removals: List[Tuple[FormRow, ChannelRecord, PortKey]] = []
    for row in terminal_rows_for_pending(rows, scenario, stage, corner):
        ch = channel_by_id.get(clean_cell(row.values.get("channel_id")))
        if not ch:
            continue
        if ch.channel_type not in CHANNEL_TYPES_20:
            continue
        src_key = harden_pending_key(ch.src_instance, ch.src_direction, ch.src_port)
        dst_key = harden_pending_key(ch.dst_instance, ch.dst_direction, ch.dst_port)
        for key in (src_key, dst_key):
            if key is not None:
                removals.append((row, ch, key))
    previous_removed = read_removed_keys(list(previous_removed_paths) + [removed_log_path])
    by_inst: Dict[str, List[Tuple[FormRow, ChannelRecord, PortKey]]] = defaultdict(list)
    for item in removals:
        by_inst[item[2].inst_name].append(item)

    removed_items: Dict[PortKey, Tuple[FormRow, ChannelRecord, PortKey]] = {}
    for inst_name, inst_items in sorted(by_inst.items()):
        pending_file = pending_dir / f"{inst_name}.ports"
        if not pending_file.is_file():
            for row, ch, key in inst_items:
                if key in previous_removed:
                    continue
                report.error(f"{pending_file}: missing pending file for 20 channel endpoint {key.inst_name}/{key.port_name}")
            continue
        lines = pending_file.read_text(encoding="utf-8").splitlines()
        index: Dict[Tuple[str, str], int] = {}
        duplicate_keys: Set[Tuple[str, str]] = set()
        for idx, line in enumerate(lines):
            key = pending_line_key(line)
            if key is None:
                continue
            if key in index:
                duplicate_keys.add(key)
            else:
                index[key] = idx
        for direction, port in sorted(duplicate_keys):
            report.error(f"{pending_file}: duplicate pending port line {direction} {port}")

        remove_line_indices: Set[int] = set()
        for row, ch, port_key in inst_items:
            key = (port_key.direction, port_key.port_name)
            if key not in index:
                if port_key in previous_removed:
                    continue
                report.error(
                    f"{pending_file}: 20 wants to remove {port_key.direction} {port_key.port_name}, "
                    "but it is not present in pending and no previous_removed record exists"
                )
                continue
            remove_line_indices.add(index[key])
            removed_items.setdefault(port_key, (row, ch, port_key))

        if remove_line_indices:
            kept = [line for idx, line in enumerate(lines) if idx not in remove_line_indices]
            plan.pending_updates[pending_file] = "\n".join(kept).rstrip() + ("\n" if kept else "")

    for _, item in sorted(
        removed_items.items(),
        key=lambda pair: (pair[0].inst_name, pair[0].direction, pair[0].port_name),
    ):
        plan.removed_lines.append(removed_log_line_20(item[0], item[1], item[2], mode))
    plan.removed_count = len(removed_items)
    return plan


def apply_pending_plan(
    plan: PendingPlan,
    removed_log_path: Path,
    scenario: str,
    mode: str,
) -> None:
    for path in sorted(plan.pending_updates, key=lambda item: str(item)):
        atomic_write_text(path, plan.pending_updates[path])
    existing = removed_log_path.read_text(encoding="utf-8").splitlines() if removed_log_path.is_file() else []
    if not existing:
        existing = [
            f"# Author: {author_name()}",
            "# Stage: 20_harden_x_if",
            f"# Scenario: {scenario}",
            f"# Mode: {mode}",
        ]
    existing_keys = {key for key in (removed_line_key(line) for line in existing) if key is not None}
    for line in plan.removed_lines:
        key = removed_line_key(line)
        if key is not None and key not in existing_keys:
            existing.append(line)
            existing_keys.add(key)
    atomic_write_text(removed_log_path, "\n".join(line for line in existing if line.strip()).rstrip() + "\n")


# ---------------------------------------------------------------------------
# Single-run workbook-centric target runtime (latest shared runtime contract)
# ---------------------------------------------------------------------------


def load_stage10_runtime():
    """Load target helpers already shared by the migrated stage-10 runtime."""
    here = Path(__file__).resolve()
    path = here.parent.parent / "10_feedthrough" / "10_extract_feedthrough.py"
    if not path.is_file():
        raise RuntimeError("migrated stage-10 runtime was not found: {0}".format(path))
    spec = importlib.util.spec_from_file_location("soc_sdc_stage10_runtime_for_20", str(path))
    if spec is None or spec.loader is None:
        raise RuntimeError("failed to load migrated stage-10 runtime: {0}".format(path))
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def target20_json_bytes(value) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")


def target20_json_digest(value) -> str:
    return hashlib.sha256(target20_json_bytes(value)).hexdigest()


def target20_safe_token(value) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", clean_cell(value)).strip("_") or "all"


def target20_output_path(run_root: Path, stage: str, corner: str) -> Path:
    if stage == "all" and corner == "all":
        return run_root / "20_result" / "20_harden_x_if.sdc"
    return (
        run_root / "20_result" /
        "20_harden_x_if_{0}_{1}.sdc".format(
            target20_safe_token(stage), target20_safe_token(corner)
        )
    )


def target20_report_path(run_root: Path, stage: str, corner: str) -> Path:
    suffix = "" if stage == "all" and corner == "all" else "_{0}_{1}".format(
        target20_safe_token(stage), target20_safe_token(corner)
    )
    return (
        run_root / "20_result" / "reports" /
        "harden_x_if_check_report{0}.txt".format(suffix)
    )


def target20_completion_path(run_root: Path, stage: str, corner: str) -> Path:
    return (
        run_root / "20_middle" / "completion" /
        "{0}_{1}.meta".format(target20_safe_token(stage), target20_safe_token(corner))
    )


def target20_required_view(
    required_views: Sequence[Dict[str, str]], stage: str, corner: str
) -> Optional[Dict[str, str]]:
    matches = [
        row for row in required_views
        if clean_cell(row.get("stage")) == stage
        and clean_cell(row.get("corner")) == corner
        and normalize_key(row.get("require_20")) == "yes"
    ]
    return matches[0] if len(matches) == 1 else None


def target20_read_json(path: Path, label: str, report: Report) -> Dict[str, object]:
    if not path.is_file():
        report.error("required {0} is missing: {1}".format(label, path))
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        report.error("invalid {0} {1}: {2}".format(label, path, exc))
        return {}
    if not isinstance(payload, dict):
        report.error("{0} must contain a JSON object: {1}".format(label, path))
        return {}
    return payload


def target20_validate_completion(
    stage10,
    path: Path,
    stage_name: str,
    run_context: Dict[str, str],
    structure: str,
    report: Report,
) -> Dict[str, object]:
    payload = target20_read_json(path, stage_name + " completion", report)
    stage10.target_validate_provenance(
        payload,
        path,
        run_context,
        structure,
        report,
        require_gate_fields=True,
        expected_stage_name=stage_name,
    )
    return payload


def target20_validate_upstreams(
    runtime,
    stage10,
    run_root: Path,
    run_context: Dict[str, str],
    required_views: Sequence[Dict[str, str]],
    structure: str,
    accounting_before: str,
    port_records,
    mode: str,
    report: Report,
) -> Dict[str, object]:
    upstream_digests: Dict[str, str] = {}
    delta_meta: Dict[str, Dict[str, object]] = {}
    completions: Dict[str, Dict[str, object]] = {}
    stage_specs = (
        ("00", "00_middle", "00_harden_port_inventory"),
        ("01", "01_middle", "01_soc_clocks"),
        ("04", "04_middle", "04_soc_io_pads"),
        ("10", "10_middle", "10_feedthrough"),
    )
    previous_after = ""
    for label, directory, stage_name in stage_specs:
        middle = run_root / directory
        delta_path = middle / "port_accounting_delta.csv"
        meta_path = middle / "port_accounting_delta.meta"
        completion_path = middle / "stage_completion.meta"
        meta = target20_read_json(meta_path, label + " accounting delta meta", report)
        stage10.target_validate_provenance(
            meta, meta_path, run_context, structure, report,
            expected_stage_name=stage_name,
        )
        completion = target20_validate_completion(
            stage10, completion_path, stage_name, run_context, structure, report
        )
        if not delta_path.is_file():
            report.error("required {0} accounting delta is missing: {1}".format(label, delta_path))
        else:
            actual_delta_digest = digest_file(delta_path)
            if clean_cell(meta.get("delta_csv_digest")) != actual_delta_digest:
                report.error("{0}: delta_csv_digest is stale".format(meta_path))
            stage10.target_validate_artifact_digest(
                completion, completion_path, delta_path,
                "accounting_delta_digest", report,
            )
        if meta and completion:
            for field_name in ("accounting_digest_before", "accounting_digest_after"):
                if clean_cell(meta.get(field_name)) != clean_cell(completion.get(field_name)):
                    report.error("{0}: completion/delta {1} mismatch".format(label, field_name))
        before = clean_cell(meta.get("accounting_digest_before"))
        after = clean_cell(meta.get("accounting_digest_after"))
        if previous_after and before != previous_after:
            report.error("accounting digest chain is discontinuous before {0}".format(label))
        if after:
            previous_after = after
        delta_meta[label] = meta
        completions[label] = completion
        for path in (delta_path, meta_path, completion_path):
            upstream_digests["{0}:{1}".format(label, path.name)] = (
                digest_file(path) if path.is_file() else ""
            )

    stage10.target_validate_output_digest(
        completions.get("01", {}),
        run_root / "01_middle" / "stage_completion.meta",
        run_root / "01_result" / "01_soc_clocks.sdc",
        report,
        path_base=run_root,
    )
    stage10.target_validate_output_digest(
        completions.get("10", {}),
        run_root / "10_middle" / "stage_completion.meta",
        run_root / "10_result" / "10_feedthrough.sdc",
        report,
        path_base=run_root,
    )

    completion_03_path = run_root / "03_middle" / "stage_completion.meta"
    completion_03 = target20_validate_completion(
        stage10,
        completion_03_path,
        "03_soc_clock_groups",
        run_context,
        structure,
        report,
    )
    stage10.target_validate_output_digest(
        completion_03,
        completion_03_path,
        run_root / "03_result" / "03_soc_clock_groups.sdc",
        report,
        path_base=run_root,
    )
    upstream_digests["03:stage_completion.meta"] = (
        digest_file(completion_03_path) if completion_03_path.is_file() else ""
    )

    completion_04 = completions.get("04", {})
    required_04 = [row for row in required_views if normalize_key(row.get("require_04")) == "yes"]
    authenticated = completion_04.get("required_view_completions", {})
    if required_04 and not isinstance(authenticated, dict):
        report.error("04 run-wide completion requires required_view_completions")
        authenticated = {}
    for view in required_04:
        view_path = (
            run_root / "04_middle" / "completion" /
            "{0}_{1}.meta".format(
                target20_safe_token(view.get("stage")),
                target20_safe_token(view.get("corner")),
            )
        )
        payload = target20_validate_completion(
            stage10, view_path, "04_soc_io_pads", run_context, structure, report
        )
        if clean_cell(payload.get("view_id")) != clean_cell(view.get("view_id")):
            report.error("{0}: view_id does not match required_views.csv".format(view_path))
        if clean_cell(payload.get("stage")) != clean_cell(view.get("stage")):
            report.error("{0}: stage does not match required_views.csv".format(view_path))
        if clean_cell(payload.get("corner")) != clean_cell(view.get("corner")):
            report.error("{0}: corner does not match required_views.csv".format(view_path))
        actual = digest_file(view_path) if view_path.is_file() else ""
        expected = clean_cell(authenticated.get(clean_cell(view.get("view_id"))))
        if not expected or expected != actual:
            report.error("{0}: 04 run-wide completion does not authenticate this view".format(view_path))
        if clean_cell(view.get("stage")) == "all" and clean_cell(view.get("corner")) == "all":
            sdc_path = run_root / "04_result" / "04_soc_io_pads.sdc"
        else:
            sdc_path = (
                run_root / "04_result" /
                "04_soc_io_pads_{0}_{1}.sdc".format(
                    target20_safe_token(view.get("stage")),
                    target20_safe_token(view.get("corner")),
                )
            )
        stage10.target_validate_output_digest(
            payload, view_path, sdc_path, report, path_base=run_root
        )
        upstream_digests["04:view:{0}".format(clean_cell(view.get("view_id")))] = actual

    empty_digest = runtime.accounting_digest(port_records, empty=True)
    runtime.validate_resume_evidence(
        run_root,
        port_records,
        run_context,
        structure,
        accounting_before,
        empty_digest,
        report,
    )
    existing_20_meta_path = run_root / "20_middle" / "port_accounting_delta.meta"
    existing_20_meta = (
        target20_read_json(existing_20_meta_path, "20 accounting delta meta", report)
        if existing_20_meta_path.is_file() else {}
    )
    if existing_20_meta:
        stage10.target_validate_provenance(
            existing_20_meta,
            existing_20_meta_path,
            run_context,
            structure,
            report,
            expected_stage_name="20_harden_x_if",
        )
    expected_current = clean_cell(
        existing_20_meta.get("accounting_digest_after")
        if existing_20_meta else delta_meta.get("10", {}).get("accounting_digest_after")
    )
    if expected_current and expected_current != accounting_before:
        report.error(
            "accounting digest chain mismatch before 20: expected={0}, current={1}".format(
                expected_current, accounting_before
            )
        )

    clock_path = run_root / "01_middle" / "clock_inventory.csv"
    clock_meta_path = run_root / "01_middle" / "clock_inventory.meta"
    clock_rows, clock_meta = stage10.target_validate_inventory_pair(
        clock_path, clock_meta_path, "01 clock inventory", run_context,
        structure, report, ("inventory_digest", "clock_inventory_digest"),
    )
    pad_path = run_root / "04_middle" / "pad_inventory.csv"
    pad_meta_path = run_root / "04_middle" / "pad_inventory.meta"
    pad_rows, pad_meta = stage10.target_validate_inventory_pair(
        pad_path, pad_meta_path, "04 pad inventory", run_context,
        structure, report, ("inventory_digest", "pad_inventory_digest"),
    )
    feedthrough_path = run_root / "10_middle" / "feedthrough_edge_inventory.csv"
    feedthrough_meta_path = run_root / "10_middle" / "feedthrough_edge_inventory.meta"
    feedthrough_rows, feedthrough_meta = stage10.target_validate_inventory_pair(
        feedthrough_path, feedthrough_meta_path, "10 feedthrough inventory",
        run_context, structure, report,
        ("inventory_digest", "feedthrough_edge_inventory_digest", "feedthrough_inventory_digest"),
    )
    stage10.target_validate_artifact_digest(
        completions.get("10", {}),
        run_root / "10_middle" / "stage_completion.meta",
        feedthrough_path,
        "feedthrough_inventory_digest",
        report,
    )
    for label, path in (
        ("01:clock_inventory.csv", clock_path),
        ("01:clock_inventory.meta", clock_meta_path),
        ("04:pad_inventory.csv", pad_path),
        ("04:pad_inventory.meta", pad_meta_path),
        ("10:feedthrough_edge_inventory.csv", feedthrough_path),
        ("10:feedthrough_edge_inventory.meta", feedthrough_meta_path),
    ):
        upstream_digests[label] = digest_file(path) if path.is_file() else ""

    lookup = stage10.target_port_record_lookup(port_records)
    clock_endpoints: Set[Tuple[str, str, str, int]] = set()
    clock_aliases: Dict[Tuple[str, str], str] = {}
    active_clocks: Set[str] = set()
    for row in clock_rows:
        action = normalize_key(row.get("final_action"))
        if action in stage10.ACTIVE_CLOCK_ACTIONS or action == "check_only":
            clock_endpoints.update(stage10.target_resolve_inventory_endpoint(
                row.get("inst_name", ""), row.get("direction", ""),
                row.get("port_name", ""), row.get("bit_index", ""), lookup,
            ))
        clock_name = clean_cell(row.get("clock_name"))
        if action in stage10.ACTIVE_CLOCK_ACTIONS and clock_name:
            active_clocks.add(clock_name)
        inst_name = clean_cell(row.get("inst_name"))
        for alias in (clean_cell(row.get("original_clock_name")), clock_name):
            if inst_name and alias and clock_name:
                clock_aliases[(inst_name, alias)] = clock_name

    pad_connection_ids: Set[str] = set()
    pad_endpoints: Set[Tuple[str, str, str, int]] = set()
    for row in pad_rows:
        if clean_cell(row.get("connection_id")):
            pad_connection_ids.add(clean_cell(row.get("connection_id")))
        inst_name = clean_cell(
            row.get("harden_instance") or row.get("subsys_instance") or row.get("inst_name")
        )
        port_name = clean_cell(
            row.get("harden_port") or row.get("subsys_port") or row.get("port") or row.get("port_name")
        )
        pad_endpoints.update(stage10.target_resolve_inventory_endpoint(
            inst_name, row.get("direction", ""), port_name,
            row.get("bit_index", ""), lookup,
        ))

    feedthrough_connection_ids: Set[str] = set()
    feedthrough_endpoints: Set[Tuple[str, str, str, int]] = set()
    for row in feedthrough_rows:
        connection_id = clean_cell(row.get("connection_id"))
        if connection_id:
            feedthrough_connection_ids.add(connection_id)
        for prefix in ("src", "dst"):
            feedthrough_endpoints.update(stage10.target_resolve_inventory_endpoint(
                row.get(prefix + "_instance", ""),
                row.get(prefix + "_direction", ""),
                row.get(prefix + "_port", ""),
                row.get(prefix + "_bit_index", ""),
                lookup,
            ))

    relations: Dict[Tuple[str, str], str] = {}
    relation_path = run_root / "03_middle" / "relation_map.csv"
    relation_meta_path = run_root / "03_middle" / "relation_map.meta"
    if mode == "budget_output":
        relation_rows, relation_meta = stage10.target_validate_inventory_pair(
            relation_path, relation_meta_path, "03 relation map", run_context,
            structure, report, ("relation_map_digest", "inventory_digest"),
        )
        for row in relation_rows:
            clock_a = clean_cell(row.get("clock_a"))
            clock_b = clean_cell(row.get("clock_b"))
            relation = normalize_key(row.get("relation_type"))
            if clock_a and clock_b and relation in CLOCK_RELATION_CANONICAL:
                key = tuple(sorted((clock_a, clock_b)))
                previous = relations.get(key)
                if previous and previous != relation:
                    report.error("03 relation map has conflicting relation for {0}/{1}".format(*key))
                relations[key] = relation
        upstream_digests["03:relation_map.csv"] = digest_file(relation_path) if relation_path.is_file() else ""
        upstream_digests["03:relation_map.meta"] = digest_file(relation_meta_path) if relation_meta_path.is_file() else ""

    return {
        "upstream_digests": upstream_digests,
        "existing_20_meta": existing_20_meta,
        "clock_endpoints": clock_endpoints,
        "clock_aliases": clock_aliases,
        "active_clocks": active_clocks,
        "relations": relations,
        "pad_connection_ids": pad_connection_ids,
        "pad_endpoints": pad_endpoints,
        "feedthrough_connection_ids": feedthrough_connection_ids,
        "feedthrough_endpoints": feedthrough_endpoints,
        "owner_map": stage10.target_delta_owner_map(runtime, run_root, port_records, report),
    }


def target20_edge_tuple(stage10, edge, side: str) -> Tuple[str, str, str, int]:
    return stage10.target_edge_endpoint_tuple(edge, side)


def target20_build_channels(
    stage10,
    direct_contexts,
    upstream: Dict[str, object],
    report: Report,
) -> List[Target20Channel]:
    channels: List[Target20Channel] = []
    seen: Set[str] = set()
    for context in direct_contexts:
        edge = context.edge
        src_tuple = target20_edge_tuple(stage10, edge, "src")
        dst_tuple = target20_edge_tuple(stage10, edge, "dst")
        if normalize_key(edge.src_instance) == "top" or normalize_key(edge.dst_instance) == "top":
            continue
        if edge.connection_id in upstream.get("pad_connection_ids", set()):
            continue
        if src_tuple in upstream.get("pad_endpoints", set()) or dst_tuple in upstream.get("pad_endpoints", set()):
            continue
        if src_tuple in upstream.get("clock_endpoints", set()) or dst_tuple in upstream.get("clock_endpoints", set()):
            continue
        if edge.connection_id in upstream.get("feedthrough_connection_ids", set()):
            continue
        if src_tuple in upstream.get("feedthrough_endpoints", set()) or dst_tuple in upstream.get("feedthrough_endpoints", set()):
            report.error(
                "{0}: endpoint overlaps 10 feedthrough inventory but connection_id is absent".format(
                    edge.connection_id
                )
            )
            continue
        if is_feedthrough_port(edge.src_port) or is_feedthrough_port(edge.dst_port):
            report.error(
                "{0}: feedthrough-named edge is missing from 10 feedthrough inventory".format(
                    edge.connection_id
                )
            )
            continue
        if normalize_key(edge.src_direction) not in {"output", "inout"}:
            report.error("{0}: normal channel source direction is invalid".format(edge.connection_id))
            continue
        if normalize_key(edge.dst_direction) not in {"input", "inout"}:
            report.error("{0}: normal channel destination direction is invalid".format(edge.connection_id))
            continue
        if context.src_record is None or context.dst_record is None:
            report.error("{0}: normal channel must resolve both workbook endpoints".format(edge.connection_id))
            continue
        connection_hash = edge.connection_id[len("CONN_"):] if edge.connection_id.startswith("CONN_") else ""
        if not re.fullmatch(r"[0-9a-f]{64}", connection_hash):
            report.error("{0}: connection ID is not a full canonical SHA-256".format(edge.connection_id))
            continue
        channel_id = "CH_" + connection_hash
        if channel_id in seen:
            report.error("duplicate normal channel ID {0}".format(channel_id))
            continue
        seen.add(channel_id)
        channels.append(Target20Channel(context, channel_id, connection_hash))
    channels.sort(key=lambda item: item.channel_id)
    report.info("classified {0} normal functional direct channel(s)".format(len(channels)))
    return channels


def target20_manifest_instances(
    manifest: Dict[str, Any],
    instances: Dict[str, Dict[str, object]],
    port_records,
) -> Dict[str, InstInfo]:
    converted: Dict[str, InstInfo] = {}
    for inst_name, info in instances.items():
        entry = manifest.get(inst_name)
        inst = InstInfo(
            module_name=clean_cell(info.get("module_name")),
            inst_name=inst_name,
            owner=clean_cell(info.get("owner")),
            sdc_hint=clean_cell(entry.sdc_path) if entry is not None else "",
            sdc_status=normalize_key(entry.availability_status) if entry is not None else "missing",
            sdc_note=clean_cell(entry.note) if entry is not None else "",
        )
        if entry is not None and entry.resolved_path is not None:
            inst.sdc_path = entry.resolved_path
        converted[inst_name] = inst
    for record in port_records:
        inst = converted.get(record.inst_name)
        if inst is None:
            continue
        target = inst.inputs if record.direction == "input" else (
            inst.outputs if record.direction == "output" else inst.inouts
        )
        for bit in record.shape.bits():
            port_name = record.shape.base if record.shape.scalar else "{0}[{1}]".format(record.shape.base, bit)
            target.setdefault(port_name, PortInfo(name=port_name))
        target.setdefault(record.shape.base, PortInfo(name=record.shape.base))
    return converted


def target20_exception_index(exceptions: Sequence[ExceptionEvidence]):
    result: Dict[Tuple[str, str], List[ExceptionEvidence]] = defaultdict(list)
    for item in exceptions:
        result[(item.inst_name, item.port_name)].append(item)
        result[(item.inst_name, port_base(item.port_name))].append(item)
    return result


def target20_evidence_clock(
    inst_name: str, clock_name: str, upstream: Dict[str, object]
) -> str:
    name = clean_cell(clock_name)
    if not name:
        return ""
    mapped = upstream.get("clock_aliases", {}).get((inst_name, name))
    return mapped or "unresolved:" + name


def target20_relation(
    src_clocks: Sequence[str], dst_clocks: Sequence[str], upstream: Dict[str, object]
) -> str:
    if any(value.startswith("unresolved:") for value in list(src_clocks) + list(dst_clocks) if value):
        return "unknown"
    src = sorted(set(value for value in src_clocks if value))
    dst = sorted(set(value for value in dst_clocks if value))
    if len(src) != 1 or len(dst) != 1:
        return "unknown"
    if src[0] == dst[0]:
        return "synchronous"
    return upstream.get("relations", {}).get(tuple(sorted((src[0], dst[0]))), "unknown")


def target20_machine_digest(values: Dict[str, object]) -> str:
    payload = [
        [header, clean_cell(values.get(header))]
        for header in sorted(TARGET20_REVIEW_INVALIDATING_FIELDS)
    ]
    return target20_json_digest(payload)


def target20_build_seeds(
    stage10,
    channels: Sequence[Target20Channel],
    manifest: Dict[str, Any],
    completeness: RunCompleteness,
    evidence,
    exceptions: Sequence[ExceptionEvidence],
    run_context: Dict[str, str],
    structure: str,
    accounting_before: str,
    upstream: Dict[str, object],
    mode: str,
    stage: str,
    corner: str,
    report: Report,
) -> List[Dict[str, object]]:
    evidence_by_key = stage10.evidence_index(evidence)
    exceptions_by_key = target20_exception_index(exceptions)
    seeds: List[Dict[str, object]] = []
    for channel in channels:
        context = channel.context
        edge = context.edge
        src_evidence = stage10.matching_evidence(
            evidence_by_key, edge.src_instance, edge.src_port, "output_delay"
        ) if mode == "budget_output" else []
        dst_evidence = stage10.matching_evidence(
            evidence_by_key, edge.dst_instance, edge.dst_port, "input_delay"
        ) if mode == "budget_output" else []
        related = src_evidence + dst_evidence
        exception_items: List[ExceptionEvidence] = []
        for key in (
            (edge.src_instance, edge.src_port),
            (edge.src_instance, port_base(edge.src_port)),
            (edge.dst_instance, edge.dst_port),
            (edge.dst_instance, port_base(edge.dst_port)),
        ):
            for item in exceptions_by_key.get(key, []):
                if item not in exception_items:
                    exception_items.append(item)
        src_entry = manifest.get(edge.src_instance)
        dst_entry = manifest.get(edge.dst_instance)
        src_status = normalize_key(src_entry.availability_status) if src_entry is not None else "missing"
        dst_status = normalize_key(dst_entry.availability_status) if dst_entry is not None else "missing"
        if "missing" in {src_status, dst_status}:
            evidence_status = "incomplete_missing_sdc"
        elif mode == "audit_only":
            evidence_status = "audit_not_consumed"
        elif any(item.parse_status != "ok" for item in related):
            evidence_status = "needs_review"
        elif related:
            evidence_status = "complete"
        else:
            evidence_status = "complete_no_delay_candidate"
        src_clocks = [
            target20_evidence_clock(edge.src_instance, item.clock_name, upstream)
            for item in src_evidence if clean_cell(item.clock_name)
        ]
        dst_clocks = [
            target20_evidence_clock(edge.dst_instance, item.clock_name, upstream)
            for item in dst_evidence if clean_cell(item.clock_name)
        ]
        clock_relation = target20_relation(src_clocks, dst_clocks, upstream)
        if mode == "budget_output":
            if bool(src_evidence) != bool(dst_evidence):
                report.warn(
                    "{0}: delay evidence is present on only one channel side".format(
                        channel.channel_id
                    )
                )
            if related and clock_relation in {
                "unknown", "asynchronous", "logically_exclusive", "physically_exclusive",
            }:
                report.warn(
                    "{0}: clock relation {1} requires explicit budget review".format(
                        channel.channel_id, clock_relation
                    )
                )
        values: Dict[str, object] = {header: "" for header in TARGET20_HEADERS}
        values.update({
            "schema_version": TARGET_SCHEMA_VERSION,
            "run_id": run_context.get("run_id", ""),
            "mode_label": run_context.get("mode_label", ""),
            "design_revision": run_context.get("design_revision", ""),
            "mode": mode,
            "run_completeness": completeness.status,
            "structure_digest": structure,
            "accounting_digest_before": accounting_before,
            "accounting_digest_after": accounting_before,
            "channel_id": channel.channel_id,
            "connection_id": edge.connection_id,
            "stage": stage,
            "corner": corner,
            "channel_type": "harden_to_harden",
            "owner_stage": "20",
            "src_instance": edge.src_instance,
            "src_direction": edge.src_direction,
            "src_port": edge.src_port,
            "src_bit_index": edge.src_bit_index,
            "src_endpoint": stage10.target_endpoint_collection(edge.src_instance, edge.src_port),
            "src_sdc_status": src_status,
            "dst_instance": edge.dst_instance,
            "dst_direction": edge.dst_direction,
            "dst_port": edge.dst_port,
            "dst_bit_index": edge.dst_bit_index,
            "dst_endpoint": stage10.target_endpoint_collection(edge.dst_instance, edge.dst_port),
            "dst_sdc_status": dst_status,
            "source_workbook": context.source_record.workbook,
            "source_sheet": context.source_record.sheet,
            "source_row": context.source_record.row,
            "is_pad_related": "no",
            "is_clock_related": "no",
            "is_feedthrough": "no",
            "evidence_status": evidence_status,
            "original_src_clock": join_unique(src_clocks),
            "original_dst_clock": join_unique(dst_clocks),
            "original_src_max": join_unique(item.max_value or item.bare_value for item in src_evidence),
            "original_src_min": join_unique(item.min_value for item in src_evidence),
            "original_dst_max": join_unique(item.max_value or item.bare_value for item in dst_evidence),
            "original_dst_min": join_unique(item.min_value for item in dst_evidence),
            "complex_options": join_unique(item.complex_options for item in related),
            "source_sdc_file": join_unique(
                [item.source_sdc_file for item in related] +
                [item.source_sdc_file for item in exception_items]
            ),
            "source_line": join_unique(
                [item.source_line for item in related] +
                [item.source_line for item in exception_items]
            ),
            "source_digest": join_unique(
                [item.source_digest for item in related] +
                [item.source_digest for item in exception_items]
            ),
            "source_command": join_unique(
                [item.original_command for item in related] +
                [item.original_command for item in exception_items], " || "
            ),
            "clock_relation": clock_relation,
            "timing_model": "unknown",
            "channel_disposition": "pending",
            "budget_required": "",
            "budget_model": "unknown",
            "tool_surface": "sta",
            "datapath_only": "yes",
            "source_type": "extracted" if related or exception_items else "manual",
            "apply": "no",
            "emit_max": "no",
            "emit_min": "no",
            "review_status": "pending",
            "validation_status": "matched",
        })
        if exception_items:
            values.update({
                "channel_disposition": "route_to_30",
                "budget_required": "no",
                "apply": "yes",
                "review_status": "approved",
                "owner": "30_harden_to_harden_exception",
                "reviewer": POLICY_REVIEWER,
                "review_date": POLICY_REVIEW_DATE,
                "disposition_basis": "known_harden_exception_evidence",
                "note": "known harden exception evidence; normal 20 budget is blocked",
            })
        elif mode == "audit_only":
            values.update({
                "channel_disposition": "no_soc_budget_required",
                "budget_required": "no",
                "apply": "yes",
                "review_status": "approved",
                "owner": POLICY_OWNER,
                "reviewer": POLICY_REVIEWER,
                "review_date": POLICY_REVIEW_DATE,
                "disposition_basis": POLICY_ID,
                "sdc_independent_basis": SDC_INDEPENDENT_POLICY,
                "note": "audit-only project policy; no channel timing command emitted",
            })
        values["machine_digest"] = target20_machine_digest(values)
        if normalize_key(values.get("review_status")) == "approved":
            values["approved_machine_digest"] = values["machine_digest"]
        seeds.append(values)
    return seeds


def target20_row_key(values: Dict[str, object]) -> Tuple[str, str, str]:
    return (
        clean_cell(values.get("channel_id")),
        clean_cell(values.get("stage")) or "all",
        clean_cell(values.get("corner")) or "all",
    )


def target20_sync_workbook(
    path: Path,
    seeds: Sequence[Dict[str, object]],
    evidence,
    run_context: Dict[str, str],
    structure: str,
    accounting_before: str,
    mode: str,
    report: Report,
) -> None:
    workbook, created = create_or_load_workbook(path)
    ensure_sheet(workbook, "interface_budget", TARGET20_HEADERS)
    ensure_sheet(workbook, "channel_inventory", TARGET20_HEADERS)
    ensure_sheet(workbook, "extraction_log", LOG_HEADERS)
    ensure_sheet(workbook, "run_metadata", ["key", "value"])
    sheet = workbook["interface_budget"]
    mapping = header_map(sheet)
    existing: Dict[Tuple[str, str, str], int] = {}
    for row_idx in range(2, sheet.max_row + 1):
        values = row_values(sheet, row_idx, TARGET20_HEADERS)
        key = target20_row_key(values)
        if not key[0]:
            continue
        if key in existing:
            report.error("interface_budget duplicates channel/view key {0}".format(key))
        existing[key] = row_idx
    seed_keys = {target20_row_key(seed) for seed in seeds}
    changed = created
    for seed in seeds:
        key = target20_row_key(seed)
        row_idx = existing.get(key)
        if row_idx is None:
            append_dict(sheet, TARGET20_HEADERS, seed, NEW_FILL)
            existing[key] = sheet.max_row
            changed = True
            continue
        material_changed = False
        for header in sorted(TARGET20_MACHINE_FIELDS):
            old = clean_cell(sheet.cell(row_idx, mapping[header]).value)
            new = clean_cell(seed.get(header))
            if old == new:
                continue
            sheet.cell(row_idx, mapping[header], seed.get(header, ""))
            if header in TARGET20_REVIEW_INVALIDATING_FIELDS:
                material_changed = True
            if header not in {"accounting_digest_before", "accounting_digest_after"}:
                changed = True
        if material_changed:
            for header, value in {
                "apply": "no", "emit_max": "no", "emit_min": "no",
                "review_status": "pending", "approved_machine_digest": "",
                "converted_max": "", "converted_min": "", "max_source": "",
                "min_source": "", "derivation_basis": "",
            }.items():
                sheet.cell(row_idx, mapping[header], value)
            report.warn("{0}: machine evidence changed; review was reset".format(key[0]))
        elif mode == "audit_only":
            for header in (
                "channel_disposition", "budget_required", "apply", "emit_max",
                "emit_min", "review_status", "owner", "reviewer", "review_date",
                "disposition_basis", "sdc_independent_basis", "approved_machine_digest",
                "note",
            ):
                new = clean_cell(seed.get(header))
                if clean_cell(sheet.cell(row_idx, mapping[header]).value) != new:
                    sheet.cell(row_idx, mapping[header], seed.get(header, ""))
                    changed = True
    stale = [
        row_idx for key, row_idx in existing.items()
        if key[1:] == (
            clean_cell(seeds[0].get("stage")) if seeds else "all",
            clean_cell(seeds[0].get("corner")) if seeds else "all",
        ) and key not in seed_keys
    ]
    for row_idx in sorted(stale, reverse=True):
        sheet.delete_rows(row_idx, 1)
        changed = True

    inventory_sheet = workbook["channel_inventory"]
    if inventory_sheet.max_row > 1:
        inventory_sheet.delete_rows(2, inventory_sheet.max_row - 1)
    for seed in sorted(seeds, key=lambda item: clean_cell(item.get("channel_id"))):
        append_dict(inventory_sheet, TARGET20_HEADERS, seed)

    log_sheet = workbook["extraction_log"]
    if log_sheet.max_row > 1:
        log_sheet.delete_rows(2, log_sheet.max_row - 1)
    for item in evidence:
        append_dict(log_sheet, LOG_HEADERS, {
            "source_sdc_file": item.source_sdc_file,
            "source_line": item.source_line,
            "instance": item.inst_name,
            "port": item.port_name,
            "direction": "",
            "constraint_type": item.constraint_type,
            "clock_name": item.clock_name,
            "min_value": item.min_value,
            "max_value": item.max_value or item.bare_value,
            "parse_status": item.parse_status,
            "source_digest": item.source_digest,
            "original_command": item.original_command,
            "message": item.message,
        })

    metadata_sheet = workbook["run_metadata"]
    if metadata_sheet.max_row:
        metadata_sheet.delete_rows(1, metadata_sheet.max_row)
    metadata_sheet.append(["key", "value"])
    for key, value in (
        ("schema_version", TARGET_SCHEMA_VERSION),
        ("author", author_name()),
        ("stage_name", "20_harden_x_if"),
        ("run_id", run_context.get("run_id", "")),
        ("mode_label", run_context.get("mode_label", "")),
        ("design_revision", run_context.get("design_revision", "")),
        ("mode", mode),
        ("structure_digest", structure),
        ("accounting_digest_before", accounting_before),
    ):
        metadata_sheet.append([key, value])
    for worksheet in workbook.worksheets:
        style_sheet(worksheet)
    atomic_save_workbook(workbook, path)
    if changed:
        report.sync_changed = True
        report.info("synchronized target review workbook {0}".format(path))


def target20_read_rows(path: Path) -> List[FormRow]:
    workbook = load_workbook(str(path), data_only=False)
    if "interface_budget" not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError("{0}: missing interface_budget sheet".format(path))
    sheet = workbook["interface_budget"]
    mapping = header_map(sheet)
    rows = [
        FormRow(row_idx, row_values(sheet, row_idx, TARGET20_HEADERS))
        for row_idx in range(2, sheet.max_row + 1)
        if clean_cell(sheet.cell(row_idx, mapping.get("channel_id", 1)).value)
    ]
    workbook.close()
    return rows


def target20_current_rows(
    rows: Sequence[FormRow], stage: str, corner: str
) -> List[FormRow]:
    return [
        row for row in rows
        if clean_cell(row.values.get("stage")) == stage
        and clean_cell(row.values.get("corner")) == corner
    ]


def target20_is_approved(row: FormRow) -> bool:
    return (
        normalize_key(row.values.get("apply")) == "yes"
        and normalize_key(row.values.get("review_status")) == "approved"
    )


def target20_terminal(row: FormRow) -> bool:
    return (
        target20_is_approved(row)
        and normalize_key(row.values.get("channel_disposition"))
        in TARGET20_TERMINAL_DISPOSITIONS
    )


def target20_validate_rows(
    rows: Sequence[FormRow],
    seeds: Sequence[Dict[str, object]],
    mode: str,
    stage: str,
    corner: str,
    report: Report,
) -> None:
    current = target20_current_rows(rows, stage, corner)
    seed_by_id = {clean_cell(seed.get("channel_id")): seed for seed in seeds}
    seen: Set[str] = set()
    for row in current:
        values = row.values
        channel_id = clean_cell(values.get("channel_id"))
        if channel_id in seen:
            report.error("interface_budget row {0}: duplicate channel/view".format(row.row_idx))
            continue
        seen.add(channel_id)
        seed = seed_by_id.get(channel_id)
        if seed is None:
            report.error("interface_budget row {0}: channel is not active in current structure".format(row.row_idx))
            continue
        for header in TARGET20_REVIEW_INVALIDATING_FIELDS.union({"machine_digest"}):
            if clean_cell(values.get(header)) != clean_cell(seed.get(header)):
                report.error(
                    "interface_budget row {0}: machine field {1} is stale".format(
                        row.row_idx, header
                    )
                )
        disposition = normalize_key(values.get("channel_disposition"))
        apply_value = normalize_key(values.get("apply"))
        review_status = normalize_key(values.get("review_status"))
        emit_max = normalize_key(values.get("emit_max"))
        emit_min = normalize_key(values.get("emit_min"))
        budget_required = normalize_key(values.get("budget_required"))
        budget_model = normalize_key(values.get("budget_model"))
        if disposition not in DISPOSITIONS:
            report.error("interface_budget row {0}: invalid channel_disposition".format(row.row_idx))
        if apply_value not in YES_NO or review_status not in REVIEW_STATUS_VALUES:
            report.error("interface_budget row {0}: invalid apply/review status".format(row.row_idx))
        if emit_max not in YES_NO or emit_min not in YES_NO:
            report.error("interface_budget row {0}: emit flags must be yes/no".format(row.row_idx))
        if mode == "audit_only" and (
            disposition == "emit_budget" or budget_required == "yes"
            or emit_max == "yes" or emit_min == "yes"
        ):
            report.error("interface_budget row {0}: audit_only forbids budget emission".format(row.row_idx))
        known_exception = normalize_key(seed.get("channel_disposition")) == "route_to_30"
        if known_exception and disposition != "route_to_30":
            report.error("interface_budget row {0}: known exception evidence requires route_to_30".format(row.row_idx))
        if disposition != "emit_budget" and (emit_max == "yes" or emit_min == "yes"):
            report.error("interface_budget row {0}: only emit_budget may emit commands".format(row.row_idx))
        approved = target20_is_approved(row)
        if not approved:
            report.error(
                "interface_budget row {0}: unresolved review state blocks formal completion".format(
                    row.row_idx
                )
            )
            continue
        if disposition == "pending":
            report.error("interface_budget row {0}: pending cannot be approved/applied".format(row.row_idx))
            continue
        if disposition in TARGET20_TERMINAL_DISPOSITIONS.union({"route_to_30"}):
            missing = [
                name for name in ("owner", "reviewer", "review_date", "disposition_basis")
                if not clean_cell(values.get(name))
            ]
            if disposition == "emit_budget":
                missing = [name for name in missing if name != "disposition_basis"]
                if not clean_cell(values.get("budget_basis")):
                    missing.append("budget_basis")
            if missing:
                report.error(
                    "interface_budget row {0}: approved decision missing {1}".format(
                        row.row_idx, ",".join(missing)
                    )
                )
        if clean_cell(values.get("approved_machine_digest")) != clean_cell(values.get("machine_digest")):
            report.error("interface_budget row {0}: approved_machine_digest is stale".format(row.row_idx))
        if disposition == "route_to_30":
            if emit_max == "yes" or emit_min == "yes":
                report.error("interface_budget row {0}: route_to_30 cannot emit".format(row.row_idx))
            continue
        if disposition in {"no_soc_budget_required", "not_applicable"}:
            if emit_max != "no" or emit_min != "no":
                report.error("interface_budget row {0}: non-emitting terminal requires emit flags=no".format(row.row_idx))
            if disposition == "no_soc_budget_required" and budget_required != "no":
                report.error("interface_budget row {0}: no-budget decision requires budget_required=no".format(row.row_idx))
            if "missing" in {
                normalize_key(values.get("src_sdc_status")),
                normalize_key(values.get("dst_sdc_status")),
            } and not clean_cell(values.get("sdc_independent_basis")):
                report.error("interface_budget row {0}: missing SDC requires independent basis".format(row.row_idx))
            continue
        if disposition != "emit_budget":
            continue
        if mode != "budget_output":
            report.error("interface_budget row {0}: emit_budget requires budget_output mode".format(row.row_idx))
        if budget_required != "yes":
            report.error("interface_budget row {0}: emit_budget requires budget_required=yes".format(row.row_idx))
        if budget_model not in {
            "manual_budget", "interconnect_budget", "reviewed_two_side_budget",
            "clock_relative_io_delay",
        }:
            report.error("interface_budget row {0}: emit_budget requires reviewed budget_model".format(row.row_idx))
        if budget_model == "clock_relative_io_delay" and not clean_cell(values.get("derivation_basis")):
            report.error("interface_budget row {0}: clock-relative conversion requires derivation_basis".format(row.row_idx))
        if emit_max != "yes" and emit_min != "yes":
            report.error("interface_budget row {0}: emit_budget must emit max and/or min".format(row.row_idx))
        if emit_max == "yes" and parse_number(values.get("converted_max")) is None:
            report.error("interface_budget row {0}: emit_max requires finite converted_max".format(row.row_idx))
        if emit_min == "yes":
            if parse_number(values.get("converted_min")) is None:
                report.error("interface_budget row {0}: emit_min requires finite converted_min".format(row.row_idx))
            if normalize_key(values.get("min_sign_review")) not in {"yes", "approved", "reviewed", "waived"}:
                report.error("interface_budget row {0}: emit_min requires min_sign_review".format(row.row_idx))
        if normalize_key(values.get("tool_surface")) not in {"dc", "sta", "both"}:
            report.error("interface_budget row {0}: tool_surface is required".format(row.row_idx))
        if normalize_key(values.get("datapath_only")) not in {"yes", "no"}:
            report.error("interface_budget row {0}: datapath_only strategy is required".format(row.row_idx))
        if "missing" in {
            normalize_key(values.get("src_sdc_status")),
            normalize_key(values.get("dst_sdc_status")),
        } and not clean_cell(values.get("sdc_independent_basis")):
            report.error("interface_budget row {0}: emit with missing SDC requires independent basis".format(row.row_idx))
        relation = normalize_key(values.get("clock_relation"))
        if relation in {"asynchronous", "logically_exclusive", "physically_exclusive", "unknown"} and not clean_cell(values.get("relationship_override_basis")):
            report.error("interface_budget row {0}: non-synchronous/unknown relation requires override basis".format(row.row_idx))
        if clean_cell(values.get("complex_options")) and not clean_cell(values.get("derivation_basis")):
            report.error("interface_budget row {0}: complex SDC options require explicit derivation review".format(row.row_idx))
    for channel_id in sorted(set(seed_by_id) - seen):
        report.error("review workbook is missing current channel {0}".format(channel_id))


def target20_commands(row: FormRow) -> List[str]:
    if not target20_is_approved(row):
        return []
    values = row.values
    if normalize_key(values.get("channel_disposition")) != "emit_budget":
        return []
    src = clean_cell(values.get("src_endpoint"))
    dst = clean_cell(values.get("dst_endpoint"))
    datapath = " -datapath_only" if normalize_key(values.get("datapath_only")) == "yes" else ""
    commands: List[str] = []
    if normalize_key(values.get("emit_max")) == "yes" and parse_number(values.get("converted_max")) is not None:
        commands.append(
            "set_max_delay {0}{1} -from {2} -to {3}".format(
                format_number(values.get("converted_max")), datapath, src, dst
            )
        )
    if normalize_key(values.get("emit_min")) == "yes" and parse_number(values.get("converted_min")) is not None:
        commands.append(
            "set_min_delay {0}{1} -from {2} -to {3}".format(
                format_number(values.get("converted_min")), datapath, src, dst
            )
        )
    return commands


def target20_render_sdc(
    rows: Sequence[FormRow],
    run_context: Dict[str, str],
    mode: str,
    stage: str,
    corner: str,
    completeness: RunCompleteness,
    structure: str,
    before: str,
    after: str,
) -> str:
    lines = [
        "################################################################################",
        "# 20 SoC harden/subsys normal interface stage",
        "# Author: {0}".format(author_name()),
        "# Stage: 20_harden_x_if",
        "# Run ID: {0}".format(run_context.get("run_id", "")),
        "# Mode label: {0}".format(run_context.get("mode_label", "")),
        "# Design revision: {0}".format(run_context.get("design_revision", "")),
        "# Mode: {0}".format(mode),
        "# View: stage={0}, corner={1}".format(stage, corner),
        "# Run completeness: {0}".format(completeness.status),
        "# Structure digest: {0}".format(structure),
        "# Accounting digest before: {0}".format(before),
        "# Accounting digest after: {0}".format(after),
        "################################################################################",
        "",
    ]
    emitted = 0
    for row in sorted(target20_current_rows(rows, stage, corner), key=lambda item: clean_cell(item.values.get("channel_id"))):
        commands = target20_commands(row)
        if not commands:
            continue
        lines.append("# channel {0}".format(clean_cell(row.values.get("channel_id"))))
        lines.append("# budget basis: {0}".format(clean_cell(row.values.get("budget_basis"))))
        if clean_cell(row.values.get("derivation_basis")):
            lines.append("# derivation: {0}".format(clean_cell(row.values.get("derivation_basis"))))
        lines.extend(commands)
        lines.append("")
        emitted += len(commands)
    if emitted == 0:
        lines.append("# No timing constraints emitted.")
    return "\n".join(lines).rstrip() + "\n"


def target20_inventory_text(rows: Sequence[FormRow]) -> str:
    stream = io.StringIO()
    writer = csv.DictWriter(
        stream, fieldnames=TARGET20_HEADERS, extrasaction="ignore", lineterminator="\n"
    )
    writer.writeheader()
    for row in sorted(
        rows,
        key=lambda item: (
            clean_cell(item.values.get("channel_id")),
            clean_cell(item.values.get("stage")),
            clean_cell(item.values.get("corner")),
        ),
    ):
        writer.writerow({header: clean_cell(row.values.get(header)) for header in TARGET20_HEADERS})
    return stream.getvalue()


def target20_record_for_endpoint(stage10, edge, side: str, lookup, report: Report):
    return stage10.target_record_for_endpoint(edge, side, lookup, report)


def target20_plan_accounting(
    stage10,
    rows: Sequence[FormRow],
    channels: Sequence[Target20Channel],
    port_records,
    upstream: Dict[str, object],
    runtime,
    stage: str,
    corner: str,
    report: Report,
):
    lookup = stage10.target_port_record_lookup(port_records)
    by_id = {item.channel_id: item for item in channels}
    planned = []
    for row in target20_current_rows(rows, stage, corner):
        if not target20_terminal(row):
            continue
        item = by_id.get(clean_cell(row.values.get("channel_id")))
        if item is None:
            continue
        for side in ("src", "dst"):
            resolved = target20_record_for_endpoint(
                stage10, item.context.edge, side, lookup, report
            )
            if resolved is None:
                continue
            record, bit = resolved
            owner_key = (record.inst_name, record.direction, record.shape.base, bit)
            earlier = upstream.get("owner_map", {}).get(owner_key, [])
            if earlier:
                report.info(
                    "{0}: {1} endpoint already accounted by {2}; 20 keeps path ownership without duplicate Used claim".format(
                        item.channel_id,
                        side,
                        ",".join("{0}:{1}".format(stage_name, owner_id) for stage_name, owner_id in earlier),
                    )
                )
                continue
            added = bit not in record.used_bits
            if added:
                record.used_bits.add(bit)
                record.added_bits.add(bit)
                record.modified = True
                record.model.modified = True
            planned.append((row, item, record, bit, side, added))
            report.info(
                "accounting {0}:{1}:{2} {3} {4}: added={5} final={6}".format(
                    record.workbook, record.sheet, record.row,
                    record.direction, record.shape.raw,
                    str(bit) if added else "<none>",
                    runtime.format_bits(record.used_bits),
                )
            )
    for record in port_records:
        if record.modified:
            cell = record.model.workbook[record.sheet].cell(record.row, record.used_col)
            cell.value = runtime.format_bits(record.used_bits)
            cell.number_format = "@"
    return planned


def target20_transaction_id(run_id: str) -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%dT%H%M%S%fZ")
    seed = "20|{0}|{1}|{2}|{3}".format(run_id, timestamp, os.getpid(), socket.gethostname())
    return "20_{0}_{1}".format(timestamp, hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12])


def target20_delta_rows(
    planned,
    run_context: Dict[str, str],
    transaction_id: str,
    view_id: str,
    stage: str,
    corner: str,
    structure: str,
    before: str,
    after: str,
    runtime,
) -> List[Dict[str, object]]:
    result = []
    for form_row, item, record, bit, side, added in planned:
        result.append({
            "schema_version": TARGET_SCHEMA_VERSION,
            "run_id": run_context.get("run_id", ""),
            "mode_label": run_context.get("mode_label", ""),
            "stage_name": "20_harden_x_if",
            "transaction_id": transaction_id,
            "view_id": view_id,
            "stage": stage,
            "corner": corner,
            "structure_digest": structure,
            "accounting_digest_before": before,
            "accounting_digest_after": after,
            "workbook": record.workbook,
            "sheet": record.sheet,
            "row": record.row,
            "direction": record.direction,
            "port": record.shape.raw,
            "legal_bits": runtime.format_bits(set(record.shape.bits())),
            "added_bits": str(bit) if added else "",
            "final_used_bits": runtime.format_bits(record.used_bits),
            "owner_object_id": item.channel_id,
            "reason": "{0}_{1}".format(
                normalize_key(form_row.values.get("channel_disposition")), side
            ),
            "evidence_status": clean_cell(form_row.values.get("evidence_status")) or "approved",
        })
    return result


def target20_inventory_meta(
    run_context: Dict[str, str],
    mode: str,
    status: str,
    structure: str,
    before: str,
    after: str,
    inventory_path: Path,
    inventory_digest: str,
    form_path: Path,
    output_path: Path,
    output_digest: str,
    completeness: RunCompleteness,
    workbook_before: Dict[str, str],
    workbook_after: Dict[str, str],
    upstream_digests: Dict[str, str],
) -> Dict[str, object]:
    return {
        "schema_version": TARGET_SCHEMA_VERSION,
        "author": author_name(),
        "stage_name": "20_harden_x_if",
        "run_id": run_context.get("run_id", ""),
        "mode_label": run_context.get("mode_label", ""),
        "design_revision": run_context.get("design_revision", ""),
        "mode": mode,
        "completion_status": status,
        "structure_digest": structure,
        "accounting_digest_before": before,
        "accounting_digest_after": after,
        "inventory_path": str(inventory_path.resolve()),
        "inventory_digest": inventory_digest,
        "channel_inventory_digest": inventory_digest,
        "review_workbook_path": str(form_path.resolve()),
        "review_workbook_digest": digest_file(form_path) if form_path.is_file() else "",
        "output_sdc_path": str(output_path.resolve()) if output_digest else "",
        "output_sdc_digest": output_digest,
        "run_completeness": completeness.status,
        "missing_instances": completeness.missing_instances,
        "workbook_file_digest_before": workbook_before,
        "workbook_file_digest_after": workbook_after,
        "upstream_artifact_digests": upstream_digests,
    }


def target20_render_report(
    report: Report,
    rows: Sequence[FormRow],
    run_context: Dict[str, str],
    mode: str,
    stage: str,
    corner: str,
    completeness: RunCompleteness,
    structure: str,
    before: str,
    after: str,
    status: str,
    diagnostic: bool,
    planned,
) -> str:
    current = target20_current_rows(rows, stage, corner)
    lines = [
        "20_harden_x_if extraction report",
        "================================",
        "Author: {0}".format(author_name()),
        "Stage: 20_harden_x_if",
        "Run ID: {0}".format(run_context.get("run_id", "")),
        "Mode label: {0}".format(run_context.get("mode_label", "")),
        "Design revision: {0}".format(run_context.get("design_revision", "")),
        "Mode: {0}".format(mode),
        "View: stage={0}, corner={1}".format(stage, corner),
        "Completion status: {0}".format(status),
        "Port accounting: {0}".format("diagnostic/read-only" if diagnostic else ("enabled" if status == "complete" else "not committed")),
        "Accounting closure: {0}".format("not evaluated" if diagnostic else ("committed" if status == "complete" else "not complete")),
        "Run completeness: {0}".format(completeness.status),
        "Structure digest: {0}".format(structure),
        "Accounting digest before: {0}".format(before),
        "Accounting digest after: {0}".format(after),
        "Warnings: {0}".format(report.warning_count),
        "Errors: {0}".format(report.error_count),
        "Sync changed: {0}".format("yes" if report.sync_changed else "no"),
        "",
        "Coverage",
        "  channels: {0}".format(len(current)),
        "  terminal channels: {0}".format(sum(1 for row in current if target20_terminal(row))),
        "  route_to_30 channels: {0}".format(sum(1 for row in current if normalize_key(row.values.get("channel_disposition")) == "route_to_30")),
        "  pending channels: {0}".format(sum(1 for row in current if normalize_key(row.values.get("channel_disposition")) == "pending")),
        "  timing commands: {0}".format(sum(len(target20_commands(row)) for row in current)),
        "  accounting endpoint records: {0}".format(len(planned)),
        "  accounting added bits: {0}".format(sum(1 for item in planned if item[5])),
        "",
        "Per-channel status",
    ]
    if not current:
        lines.append("  <no normal functional channel>")
    for row in sorted(current, key=lambda item: clean_cell(item.values.get("channel_id"))):
        values = row.values
        lines.append(
            "  {0} connection={1} {2}/{3} -> {4}/{5} disposition={6} "
            "review={7} evidence={8} model={9} src_delay={10}/{11} "
            "dst_delay={12}/{13} converted={14}/{15} emit={16} basis={17}".format(
                clean_cell(values.get("channel_id")),
                clean_cell(values.get("connection_id")),
                clean_cell(values.get("src_instance")), clean_cell(values.get("src_port")),
                clean_cell(values.get("dst_instance")), clean_cell(values.get("dst_port")),
                clean_cell(values.get("channel_disposition")),
                clean_cell(values.get("review_status")),
                clean_cell(values.get("evidence_status")),
                clean_cell(values.get("budget_model")),
                clean_cell(values.get("original_src_max")) or "-",
                clean_cell(values.get("original_src_min")) or "-",
                clean_cell(values.get("original_dst_max")) or "-",
                clean_cell(values.get("original_dst_min")) or "-",
                clean_cell(values.get("converted_max")) or "-",
                clean_cell(values.get("converted_min")) or "-",
                "+".join(name for name, flag in (("max", values.get("emit_max")), ("min", values.get("emit_min"))) if normalize_key(flag) == "yes") or "none",
                clean_cell(values.get("disposition_basis")) or clean_cell(values.get("budget_basis")) or "-",
            )
        )
    lines.extend(["", "Accounting changes"])
    if not planned:
        lines.append("  <none>")
    for row, item, record, bit, side, added in planned:
        lines.append(
            "  {0}:{1}:{2} {3} {4} side={5} channel={6} added={7} final={8}".format(
                record.workbook, record.sheet, record.row, record.direction,
                record.shape.raw, side, item.channel_id,
                str(bit) if added else "<none>",
                ",".join(str(value) for value in sorted(record.used_bits)),
            )
        )
    lines.extend(["", "Messages"])
    lines.extend(report.lines or ["INFO: no messages"])
    return "\n".join(lines).rstrip() + "\n"


def target20_view_completion_payload(
    run_context: Dict[str, str],
    mode: str,
    view_id: str,
    stage: str,
    corner: str,
    structure: str,
    before: str,
    after: str,
    workbook_before: Dict[str, str],
    workbook_after: Dict[str, str],
    upstream_digests: Dict[str, str],
    output_digest: str,
    inventory_digest: str,
    delta_digest: str,
    form_path: Path,
    transaction_id: str,
    completeness: RunCompleteness,
    status: str = "complete",
    error_count: int = 0,
    sync_changed: str = "no",
) -> Dict[str, object]:
    return {
        "schema_version": TARGET_SCHEMA_VERSION,
        "author": author_name(),
        "stage_name": "20_harden_x_if",
        "view_id": view_id,
        "stage": stage,
        "corner": corner,
        "run_id": run_context.get("run_id", ""),
        "mode_label": run_context.get("mode_label", ""),
        "design_revision": run_context.get("design_revision", ""),
        "mode": mode,
        "completion_status": status,
        "error_count": error_count,
        "sync_changed": sync_changed,
        "structure_digest": structure,
        "accounting_digest_before": before,
        "accounting_digest_after": after,
        "port_accounting": "enabled" if status == "complete" else "not_committed",
        "workbook_file_digest_before": workbook_before,
        "workbook_file_digest_after": workbook_after,
        "upstream_artifact_digests": upstream_digests,
        "output_sdc_digest": output_digest,
        "channel_inventory_digest": inventory_digest,
        "accounting_delta_digest": delta_digest,
        "review_workbook_digest": digest_file(form_path) if form_path.is_file() else "",
        "transaction_id": transaction_id,
        "run_completeness": completeness.status,
        "missing_instances": completeness.missing_instances,
    }


def target20_runwide_completion(
    run_root: Path,
    required_views: Sequence[Dict[str, str]],
    current_view: Dict[str, str],
    current_completion_bytes: bytes,
    run_context: Dict[str, str],
    mode: str,
    structure: str,
    before: str,
    after: str,
    upstream_digests: Dict[str, str],
    inventory_digest: str,
    delta_digest: str,
    output_digest: str,
    report: Report,
) -> Dict[str, object]:
    required_20 = [row for row in required_views if normalize_key(row.get("require_20")) == "yes"]
    digests: Dict[str, str] = {}
    all_complete = True
    current_id = clean_cell(current_view.get("view_id"))
    for view in required_20:
        view_id = clean_cell(view.get("view_id"))
        if view_id == current_id:
            digests[view_id] = hashlib.sha256(current_completion_bytes).hexdigest()
            continue
        path = target20_completion_path(
            run_root, clean_cell(view.get("stage")), clean_cell(view.get("corner"))
        )
        if not path.is_file():
            all_complete = False
            continue
        payload = target20_read_json(path, "20 required view completion", report)
        if normalize_key(payload.get("completion_status")) != "complete":
            all_complete = False
            continue
        if clean_cell(payload.get("stage_name")) != "20_harden_x_if":
            all_complete = False
            report.error("{0}: stage_name is stale".format(path))
            continue
        if clean_cell(payload.get("view_id")) != view_id:
            all_complete = False
            report.error("{0}: view_id does not match required_views.csv".format(path))
            continue
        if clean_cell(payload.get("stage")) != clean_cell(view.get("stage")):
            all_complete = False
            report.error("{0}: stage does not match required_views.csv".format(path))
            continue
        if clean_cell(payload.get("corner")) != clean_cell(view.get("corner")):
            all_complete = False
            report.error("{0}: corner does not match required_views.csv".format(path))
            continue
        if clean_cell(payload.get("run_id")) != clean_cell(run_context.get("run_id")):
            all_complete = False
            report.error("{0}: run_id is stale".format(path))
            continue
        if clean_cell(payload.get("mode_label")) != clean_cell(run_context.get("mode_label")):
            all_complete = False
            report.error("{0}: mode_label is stale".format(path))
            continue
        design_revision = clean_cell(payload.get("design_revision"))
        if design_revision and design_revision != clean_cell(run_context.get("design_revision")):
            all_complete = False
            report.error("{0}: design_revision is stale".format(path))
            continue
        if clean_cell(payload.get("structure_digest")) != structure:
            all_complete = False
            report.error("{0}: structure_digest is stale".format(path))
            continue
        if normalize_key(payload.get("sync_changed")) != "no":
            all_complete = False
            report.error("{0}: sync_changed must be no".format(path))
            continue
        try:
            error_count = int(clean_cell(payload.get("error_count")) or "0")
        except ValueError:
            error_count = -1
        if error_count != 0:
            all_complete = False
            report.error("{0}: error_count is not zero".format(path))
            continue
        prior_output = target20_output_path(
            run_root, clean_cell(view.get("stage")), clean_cell(view.get("corner"))
        )
        prior_output_digest = clean_cell(payload.get("output_sdc_digest"))
        if (
            not prior_output_digest
            or not prior_output.is_file()
            or digest_file(prior_output) != prior_output_digest
        ):
            all_complete = False
            report.error("{0}: output_sdc_digest is stale".format(path))
            continue
        digests[view_id] = digest_file(path)
    if not required_20:
        all_complete = True
    return {
        "schema_version": TARGET_SCHEMA_VERSION,
        "author": author_name(),
        "stage_name": "20_harden_x_if",
        "stage": "",
        "corner": "",
        "run_id": run_context.get("run_id", ""),
        "mode_label": run_context.get("mode_label", ""),
        "design_revision": run_context.get("design_revision", ""),
        "mode": mode,
        "completion_status": "complete" if all_complete and not report.error_count else "incomplete",
        "error_count": report.error_count,
        "sync_changed": "no",
        "structure_digest": structure,
        "accounting_digest_before": before,
        "accounting_digest_after": after,
        "upstream_artifact_digests": upstream_digests,
        "output_sdc_digest": output_digest,
        "channel_inventory_digest": inventory_digest,
        "accounting_delta_digest": delta_digest,
        "required_view_completions": digests,
    }


def run_target20(args: argparse.Namespace, argv: Sequence[str]) -> int:
    report = Report()
    stage10 = load_stage10_runtime()
    runtime = stage10.load_accounting_runtime()
    runtime.STAGE_NAME = "20_harden_x_if"
    run_root = Path(args.run_root).expanduser().resolve()
    input_root = run_root / "inputs"
    middle_root = run_root / "20_middle"
    form_path = middle_root / "20_harden_x_if.xlsx"
    inventory_path = middle_root / "channel_inventory.csv"
    inventory_meta_path = middle_root / "channel_inventory.meta"
    delta_path = middle_root / "port_accounting_delta.csv"
    delta_meta_path = middle_root / "port_accounting_delta.meta"
    runwide_completion_path = middle_root / "stage_completion.meta"
    output_path = target20_output_path(run_root, args.stage, args.corner)
    report_path = target20_report_path(run_root, args.stage, args.corner)
    view_completion_path = target20_completion_path(run_root, args.stage, args.corner)
    diagnostic = bool(args.diagnose_only)
    run_context: Dict[str, str] = {}
    required_views: List[Dict[str, str]] = []
    current_view: Dict[str, str] = {}
    instances: Dict[str, Dict[str, object]] = {}
    models = []
    port_records = []
    rows: List[FormRow] = []
    channels: List[Target20Channel] = []
    planned = []
    completeness = RunCompleteness(status="invalid")
    structure = ""
    accounting_before = ""
    accounting_after = ""
    upstream: Dict[str, object] = {}
    completion_status = "failed"
    preview_dir: Optional[Path] = None
    classification_only = False
    completion_publication_authorized = False
    try:
        if not input_root.is_dir():
            report.error("required inputs directory is missing: {0}".format(input_root))
            raise RuntimeError("target input layout validation failed")
        with runtime.AccountingLock(input_root / ".port_accounting.lock", report):
            runtime.recover_transactions(run_root, report)
            run_context = runtime.read_run_context(input_root / "run_context.csv", report)
            required_views = runtime.read_required_views(input_root / "required_views.csv", report)
            required_20 = [row for row in required_views if normalize_key(row.get("require_20")) == "yes"]
            classification_only = not required_20 and not diagnostic
            current_view_value = target20_required_view(required_views, args.stage, args.corner)
            if current_view_value is None:
                current_view = {
                    "view_id": (
                        "run_wide_audit" if classification_only else
                        "diagnostic_{0}_{1}".format(
                            target20_safe_token(args.stage), target20_safe_token(args.corner)
                        )
                    ),
                    "stage": args.stage,
                    "corner": args.corner,
                    "require_20": "no",
                }
                if required_20 and not diagnostic:
                    report.error(
                        "stage={0}, corner={1} is not a required require_20=yes view; use --diagnose-only".format(
                            args.stage, args.corner
                        )
                    )
                    raise RuntimeError("undeclared formal 20 view was rejected")
                if classification_only and (args.stage != "all" or args.corner != "all"):
                    report.error(
                        "a run with no require_20=yes view supports only the run-wide all/all audit; "
                        "use --diagnose-only for another stage/corner"
                    )
                    raise RuntimeError("invalid run-wide 20 audit view was rejected")
                if classification_only and args.mode != "audit_only":
                    report.error("a run with no require_20=yes view supports audit_only classification only")
                    raise RuntimeError("invalid run-wide 20 audit mode was rejected")
                completion_publication_authorized = classification_only
            else:
                current_view = dict(current_view_value)
                completion_publication_authorized = not diagnostic
            if args.mode == "budget_output" and (
                not args.stage_explicit or not args.corner_explicit
            ):
                report.error("budget_output requires explicit --stage and --corner")
            instances, info_semantic = runtime.read_info_all(input_root / "info_all.xlsx", report)
            port_paths = runtime.discover_port_workbooks(input_root, report)
            models, port_records = runtime.read_port_workbooks(
                port_paths, run_root, instances, True, report
            )
            runtime.validate_connections(port_records, instances, report)
            structure = runtime.structure_digest(
                run_context, required_views, info_semantic, port_records
            )
            accounting_before = runtime.accounting_digest(port_records)
            accounting_after = accounting_before
            upstream = target20_validate_upstreams(
                runtime, stage10, run_root, run_context, required_views,
                structure, accounting_before, port_records, args.mode, report,
            )
            manifest, completeness = stage10.read_target_harden_manifest(
                run_root / "00_middle" / "harden_sdc_manifest.csv",
                run_root,
                instances,
                args.require_complete_harden_sdc,
                report,
            )
            direct_contexts = stage10.build_target_direct_edges(runtime, port_records, report)
            channels = target20_build_channels(stage10, direct_contexts, upstream, report)
            evidence = (
                stage10.extract_delay_evidence(manifest, report)
                if args.mode == "budget_output" else []
            )
            converted_instances = target20_manifest_instances(
                manifest, instances, port_records
            )
            exceptions = extract_exception_evidence(converted_instances, report)
            seeds = target20_build_seeds(
                stage10, channels, manifest, completeness, evidence, exceptions,
                run_context, structure, accounting_before, upstream, args.mode,
                args.stage, args.corner, report,
            )
            if report.error_count:
                raise RuntimeError("target 20 input validation failed")

            if diagnostic:
                rows = [FormRow(index + 2, dict(seed)) for index, seed in enumerate(seeds)]
                inventory_text = target20_inventory_text(rows)
                diagnostic_inventory = middle_root / "diagnostic" / "channel_inventory.csv"
                diagnostic_meta = middle_root / "diagnostic" / "channel_inventory.meta"
                workbook_digests = {model.relative_name: model.digest_before for model in models}
                meta = target20_inventory_meta(
                    run_context, args.mode, "diagnostic", structure,
                    accounting_before, accounting_before, diagnostic_inventory,
                    hashlib.sha256(inventory_text.encode("utf-8")).hexdigest(),
                    form_path, output_path, "", completeness, workbook_digests,
                    workbook_digests, upstream.get("upstream_digests", {}),
                )
                report_text = target20_render_report(
                    report, rows, run_context, args.mode, args.stage, args.corner,
                    completeness, structure, accounting_before, accounting_before,
                    "diagnostic", True, [],
                )
                atomic_write_text(diagnostic_inventory, inventory_text)
                atomic_write_json(diagnostic_meta, meta)
                atomic_write_text(report_path, report_text)
                completion_status = "diagnostic"
            else:
                target20_sync_workbook(
                    form_path, seeds, evidence, run_context, structure,
                    accounting_before, args.mode, report,
                )
                rows = target20_read_rows(form_path)
                inventory_text = target20_inventory_text(rows)
                inventory_digest = hashlib.sha256(inventory_text.encode("utf-8")).hexdigest()
                workbook_digests = {model.relative_name: model.digest_before for model in models}
                if report.sync_changed:
                    meta = target20_inventory_meta(
                        run_context, args.mode, "review_required", structure,
                        accounting_before, accounting_before, inventory_path,
                        inventory_digest, form_path, output_path, "", completeness,
                        workbook_digests, workbook_digests,
                        upstream.get("upstream_digests", {}),
                    )
                    report_text = target20_render_report(
                        report, rows, run_context, args.mode, args.stage, args.corner,
                        completeness, structure, accounting_before, accounting_before,
                        "review_required", False, [],
                    )
                    atomic_write_text(inventory_path, inventory_text)
                    atomic_write_json(inventory_meta_path, meta)
                    if not classification_only:
                        review_completion = target20_view_completion_payload(
                            run_context, args.mode, clean_cell(current_view.get("view_id")),
                            args.stage, args.corner, structure, accounting_before,
                            accounting_before, workbook_digests, workbook_digests,
                            upstream.get("upstream_digests", {}), "", inventory_digest,
                            "", form_path, "", completeness,
                            status="review_required", error_count=0, sync_changed="yes",
                        )
                        atomic_write_json(view_completion_path, review_completion)
                    atomic_write_json(runwide_completion_path, {
                        "schema_version": TARGET_SCHEMA_VERSION,
                        "stage_name": "20_harden_x_if",
                        "run_id": run_context.get("run_id", ""),
                        "mode_label": run_context.get("mode_label", ""),
                        "completion_status": "incomplete",
                        "error_count": 0,
                        "sync_changed": "yes",
                        "structure_digest": structure,
                        "accounting_digest_before": accounting_before,
                        "accounting_digest_after": accounting_before,
                    })
                    atomic_write_text(report_path, report_text)
                    completion_status = "review_required"
                else:
                    target20_validate_rows(
                        rows, seeds, args.mode, args.stage, args.corner, report
                    )
                    if report.error_count:
                        raise RuntimeError("target 20 review validation failed")
                    planned = target20_plan_accounting(
                        stage10, rows, channels, port_records, upstream, runtime,
                        args.stage, args.corner, report,
                    )
                    accounting_after = runtime.accounting_digest(port_records)
                    for row in rows:
                        row.values["accounting_digest_after"] = accounting_after
                    if report.error_count:
                        raise RuntimeError("target 20 accounting plan validation failed")
                    if classification_only:
                        sdc_payload = b""
                        sdc_digest = ""
                    else:
                        sdc_text = target20_render_sdc(
                            rows, run_context, args.mode, args.stage, args.corner,
                            completeness, structure, accounting_before, accounting_after,
                        )
                        sdc_payload = sdc_text.encode("utf-8")
                        sdc_digest = hashlib.sha256(sdc_payload).hexdigest()
                    inventory_text = target20_inventory_text(rows)
                    inventory_payload = inventory_text.encode("utf-8")
                    inventory_digest = hashlib.sha256(inventory_payload).hexdigest()
                    preview_dir = middle_root / ".20_preview_{0}".format(os.getpid())
                    if preview_dir.exists():
                        shutil.rmtree(str(preview_dir))
                    preview_dir.mkdir(parents=True, exist_ok=False)
                    prepared_candidates: Dict[str, Path] = {}
                    workbook_before = {model.relative_name: model.digest_before for model in models}
                    for model in models:
                        if model.modified:
                            candidate = preview_dir / model.path.name
                            model.workbook.save(str(candidate))
                            model.digest_after = digest_file(candidate)
                            prepared_candidates[model.relative_name] = candidate
                        else:
                            model.digest_after = model.digest_before
                    workbook_after = {model.relative_name: model.digest_after for model in models}
                    transaction_id = target20_transaction_id(run_context.get("run_id", ""))
                    old_delta_rows = runtime.load_delta_rows(delta_path, report) if delta_path.is_file() else []
                    old_meta = upstream.get("existing_20_meta", {})
                    new_delta_rows = target20_delta_rows(
                        planned, run_context, transaction_id,
                        clean_cell(current_view.get("view_id")), args.stage, args.corner,
                        structure, accounting_before, accounting_after, runtime,
                    )
                    if delta_path.is_file():
                        delta_text = delta_path.read_text(encoding="utf-8")
                        if delta_text and not delta_text.endswith("\n"):
                            delta_text += "\n"
                        if new_delta_rows:
                            delta_text += runtime.csv_text(
                                runtime.DELTA_HEADERS, new_delta_rows, include_header=False
                            )
                    else:
                        delta_text = runtime.csv_text(
                            runtime.DELTA_HEADERS, old_delta_rows + new_delta_rows
                        )
                    delta_payload = delta_text.encode("utf-8")
                    delta_digest = hashlib.sha256(delta_payload).hexdigest()
                    transaction_entry = {
                        "transaction_id": transaction_id,
                        "committed_at": runtime.utc_timestamp(),
                        "structure_digest": structure,
                        "accounting_digest_before": accounting_before,
                        "accounting_digest_after": accounting_after,
                        "delta_rows_digest": runtime.delta_rows_digest(new_delta_rows),
                    }
                    delta_meta = {
                        "schema_version": TARGET_SCHEMA_VERSION,
                        "run_id": run_context.get("run_id", ""),
                        "mode_label": run_context.get("mode_label", ""),
                        "design_revision": run_context.get("design_revision", ""),
                        "stage_name": "20_harden_x_if",
                        "completion_status": "complete",
                        "structure_digest": structure,
                        "accounting_digest_before": accounting_before,
                        "accounting_digest_after": accounting_after,
                        "workbook_file_digest_before": workbook_before,
                        "workbook_file_digest_after": workbook_after,
                        "delta_csv_digest": delta_digest,
                        "transactions": list(old_meta.get("transactions", [])) + [transaction_entry],
                    }
                    inventory_meta = target20_inventory_meta(
                        run_context, args.mode, "complete", structure,
                        accounting_before, accounting_after, inventory_path,
                        inventory_digest, form_path, output_path, sdc_digest,
                        completeness, workbook_before, workbook_after,
                        upstream.get("upstream_digests", {}),
                    )
                    if classification_only:
                        view_completion_bytes = b""
                    else:
                        view_completion = target20_view_completion_payload(
                            run_context, args.mode, clean_cell(current_view.get("view_id")),
                            args.stage, args.corner, structure, accounting_before,
                            accounting_after, workbook_before, workbook_after,
                            upstream.get("upstream_digests", {}), sdc_digest,
                            inventory_digest, delta_digest, form_path, transaction_id,
                            completeness,
                        )
                        view_completion_bytes = (
                            json.dumps(view_completion, ensure_ascii=False, sort_keys=True, indent=2) + "\n"
                        ).encode("utf-8")
                    runwide_completion = target20_runwide_completion(
                        run_root, required_views, current_view, view_completion_bytes,
                        run_context, args.mode, structure, accounting_before,
                        accounting_after, upstream.get("upstream_digests", {}),
                        inventory_digest, delta_digest, sdc_digest, report,
                    )
                    report_text = target20_render_report(
                        report, rows, run_context, args.mode, args.stage, args.corner,
                        completeness, structure, accounting_before, accounting_after,
                        "complete", False, planned,
                    )
                    artifact_payloads = [
                        (inventory_path, inventory_payload),
                        (inventory_meta_path, (json.dumps(inventory_meta, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")),
                        (delta_path, delta_payload),
                        (delta_meta_path, (json.dumps(delta_meta, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")),
                        (runwide_completion_path, (json.dumps(runwide_completion, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")),
                        (report_path, report_text.encode("utf-8")),
                    ]
                    if not classification_only:
                        artifact_payloads.extend([
                            (output_path, sdc_payload),
                            (view_completion_path, view_completion_bytes),
                        ])
                    for model in models:
                        if not model.path.is_file() or digest_file(model.path) != model.digest_before:
                            report.error("concurrent workbook modification detected before 20 commit: {0}".format(model.path))
                    if report.error_count:
                        raise RuntimeError("target workbook snapshot changed before 20 commit")
                    runtime.execute_transaction(
                        run_root, models, prepared_candidates, artifact_payloads,
                        transaction_id, run_context, structure, accounting_before,
                        accounting_after, report,
                    )
                    completion_status = "complete"
    except Exception as exc:
        if not report.error_count:
            report.error(str(exc))
        failure_text = target20_render_report(
            report, rows, run_context, args.mode, args.stage, args.corner,
            completeness, structure, accounting_before,
            accounting_after or accounting_before, "failed", diagnostic, planned,
        )
        atomic_write_text(report_path, failure_text)
        if completion_publication_authorized and structure:
            if not classification_only:
                failed_completion = target20_view_completion_payload(
                    run_context, args.mode, clean_cell(current_view.get("view_id")),
                    args.stage, args.corner, structure, accounting_before,
                    accounting_before, {}, {}, upstream.get("upstream_digests", {}),
                    "", "", "", form_path, "", completeness,
                    status="failed", error_count=max(1, report.error_count),
                    sync_changed="yes" if report.sync_changed else "no",
                )
                atomic_write_json(view_completion_path, failed_completion)
            atomic_write_json(runwide_completion_path, {
                "schema_version": TARGET_SCHEMA_VERSION,
                "stage_name": "20_harden_x_if",
                "run_id": run_context.get("run_id", ""),
                "mode_label": run_context.get("mode_label", ""),
                "completion_status": "failed",
                "error_count": max(1, report.error_count),
                "sync_changed": "yes" if report.sync_changed else "no",
                "structure_digest": structure,
                "accounting_digest_before": accounting_before,
                "accounting_digest_after": accounting_before,
            })
        completion_status = "failed"
    finally:
        for model in models:
            try:
                model.workbook.close()
            except Exception:
                pass
        if preview_dir is not None and preview_dir.exists():
            shutil.rmtree(str(preview_dir))

    print("Author: {0}".format(author_name()))
    print("Run ID: {0}".format(run_context.get("run_id", "")))
    print("Mode label: {0}".format(run_context.get("mode_label", "")))
    print("Design revision: {0}".format(run_context.get("design_revision", "")))
    print("Mode: {0}".format(args.mode))
    print("View: stage={0}, corner={1}".format(args.stage, args.corner))
    print("Run completeness: {0}".format(completeness.status))
    print("Port accounting: {0}".format(
        "diagnostic/read-only" if diagnostic else (
            "enabled" if completion_status == "complete" else "not committed"
        )
    ))
    print("Completion status: {0}".format(completion_status))
    print("Report: {0}".format(report_path))
    print("Warnings: {0}  Errors: {1}  Sync changed: {2}".format(
        report.warning_count, report.error_count, report.sync_changed
    ))
    if report.error_count:
        return 1
    if completion_status == "review_required":
        return 1
    return 0


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Classify one run's normal harden interfaces and optionally emit reviewed budgets."
    )
    parser.add_argument("--run-root", required=True, help="single-run target runtime root")
    parser.add_argument("--mode", default="audit_only", choices=sorted(MODES))
    parser.add_argument("--stage", default="all", help="required timing view stage")
    parser.add_argument("--corner", default="all", help="required timing view corner")
    parser.add_argument(
        "--require-complete-harden-sdc",
        action="store_true",
        help="treat any missing required harden SDC as an error",
    )
    parser.add_argument(
        "--diagnose-only",
        action="store_true",
        help="publish diagnostic inventory/report without formal SDC or workbook accounting writes",
    )
    args = parser.parse_args(argv)
    args.stage_explicit = any(token == "--stage" or token.startswith("--stage=") for token in argv)
    args.corner_explicit = any(token == "--corner" or token.startswith("--corner=") for token in argv)
    return args


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)
    return run_target20(args, argv)


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(2)
