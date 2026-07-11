#!/usr/bin/env python3
"""Inject a view-independent (all/all) load on pad_ddr_dqs that conflicts with
the view-specific (prects/ss_125) load, to exercise the assembled-view check."""
from openpyxl import load_workbook

FORM = "04_soc_io_pads.xlsx"

wb = load_workbook(FORM)
ws = wb["io_constraints"]
hdr = [c.value for c in ws[1]]
idx = {h: i + 1 for i, h in enumerate(hdr)}
r = ws.max_row + 1
vals = dict(
    scenario="common", stage="all", corner="all",
    pad_name="pad_ddr_dqs", soc_object="[get_ports {pad_ddr_dqs}]",
    direction="output", constraint_type="load", value="0.05",
    source_type="manual", apply="yes", review_status="approved",
    basis="view-independent load conflicting with prects/ss_125",
)
for k, v in vals.items():
    ws.cell(row=r, column=idx[k], value=v)
wb.save(FORM)
print(f"injected conflicting all/all load at row {r}")
