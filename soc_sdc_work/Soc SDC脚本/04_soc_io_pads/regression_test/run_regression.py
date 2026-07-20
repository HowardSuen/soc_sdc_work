#!/usr/bin/env python3
"""Complex regression for 04_extract_soc_io_pads.py.

The test builds fresh inputs under work_complex/ and checks:
  * bit-level get_ports parsing for names containing brackets
  * SoC pad mapping for canonical port keys such as dq_i[0]
  * pending/removed_log consumption after approved 04 generation
"""
from __future__ import print_function

import csv
import hashlib
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parent
SOC = BASE.parent.parent
EX00 = SOC / "00_harden_port_inventory" / "00_harden_port_inventory.py"
EX01 = SOC / "01_soc_clocks" / "01_extract_soc_clocks.py"
EX04 = SOC / "04_soc_io_pads" / "04_extract_soc_io_pads.py"
WORK = BASE / "work_complex"


def clean_dir(path):
    if path.exists():
        shutil.rmtree(str(path))
    path.mkdir(parents=True)


def sh(args, cwd):
    return subprocess.run(
        [sys.executable] + [str(arg) for arg in args],
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
    )


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def write_info_all(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "info_all"
    ws.append(["inst_name", "module_name", "owner", "sdc_path"])
    ws.append(["u_io", "io_ring", "alice", "u_io.sdc"])
    wb.save(str(path))


def write_ports(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "u_io"
    ws.append([
        "Input", "Input Width", "Input Used Width", "From Whom",
        "Output", "Output Width", "Output Used Width", "To Top",
        "Inout", "Inout Width", "Inout Connectivity", "Inout Name",
    ])
    ws.append(["dq_i[0]", 1, 1, "top.pad_dq[0]", "", "", "", "", "", "", "", ""])
    ws.append(["pad_fp[0]", 1, 1, "top.pad_fp[0]", "", "", "", "", "", "", "", ""])
    ws.append(["untouched_i", 1, 1, "fabric.debug", "", "", "", "", "", "", "", ""])
    wb.save(str(path))


def write_clock_inventory(path):
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["clock_name", "direct_source", "producer_object", "final_action", "source_file"])
        writer.writerow(["dqs_clk", "", "", "emit_virtual_clock", "01"])


def write_target_clock_inventory(root, scenario, clock_names):
    path = root / "01_middle" / "assembled" / scenario / "clock_inventory.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["clock_name", "direct_source", "producer_object", "final_action", "source_file"])
        for name in clock_names:
            writer.writerow([name, "", "", "emit_virtual_clock", "01"])


def write_target_manifest(root, scenario, rows):
    path = root / "00_middle" / "scenario" / scenario / "harden_sdc_manifest.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(
            file_obj,
            fieldnames=["scenario", "inst_name", "module_name", "sdc_path", "availability_status", "note"],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_target_connections(root, include_missing, scenario):
    path = root / "00_middle" / "connection_inventory.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "schema_version", "connection_id", "scenario_scope", "connection_type",
        "src_instance", "src_direction", "src_port", "src_bit_index",
        "src_endpoint_key", "src_soc_object", "dst_instance", "dst_direction",
        "dst_port", "dst_bit_index", "dst_endpoint_key", "dst_soc_object",
        "fanout_index", "range_source_expr", "range_sink_expr", "bit_pair_order",
        "source_workbook", "source_sheet", "source_row", "validation_status",
        "owner_hint", "note",
    ]

    def edge(connection_id, scope, top_port, inst_name, harden_port, source_row):
        return {
            "schema_version": "1.0",
            "connection_id": connection_id,
            "scenario_scope": scope,
            "connection_type": "top_pad_to_harden",
            "src_instance": "top",
            "src_direction": "input",
            "src_port": top_port,
            "src_bit_index": top_port.split("[")[-1].rstrip("]") if "[" in top_port else "",
            "src_endpoint_key": "top:input:%s" % top_port,
            "src_soc_object": top_port,
            "dst_instance": inst_name,
            "dst_direction": "input",
            "dst_port": harden_port,
            "dst_bit_index": harden_port.split("[")[-1].rstrip("]") if "[" in harden_port else "",
            "dst_endpoint_key": "%s:input:%s" % (inst_name, harden_port),
            "dst_soc_object": "%s/%s" % (inst_name, harden_port),
            "fanout_index": "0",
            "range_source_expr": top_port,
            "range_sink_expr": harden_port,
            "bit_pair_order": "explicit_map",
            "source_workbook": "integration.xlsx",
            "source_sheet": "pads",
            "source_row": str(source_row),
            "validation_status": "matched",
            "owner_hint": "io_owner",
            "note": "04 regression pad edge",
        }

    rows = [
        edge("CONN_PAD_DQ", scenario, "pad_dq[0]", "u_io", "dq_i[0]", 10),
        edge("CONN_PAD_FP", "common", "pad_fp[0]", "u_io", "pad_fp[0]", 11),
        edge("CONN_FOREIGN_SCAN", "scan", "pad_scan_only", "u_io", "scan_only_i", 12),
    ]
    if include_missing:
        rows.append(edge("CONN_PAD_MISSING", scenario, "pad_missing", "u_missing", "missing_i", 13))
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def write_missing_ports(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "u_missing"
    ws.append([
        "Input", "Input Width", "Input Used Width", "From Whom",
        "Output", "Output Width", "Output Used Width", "To Top",
        "Inout", "Inout Width", "Inout Connectivity", "Inout Name",
    ])
    ws.append(["missing_i", 1, 1, "top.pad_missing", "", "", "", "", "", "", "", ""])
    wb.save(str(path))


def build_target_inputs(root, scenario="common", clock_name="dqs_clk", include_missing=False):
    clean_dir(root)
    inputs = root / "inputs"
    inputs.mkdir(parents=True)

    (inputs / "u_io.sdc").write_text(
        "set_input_delay -clock [get_clocks %s] -max 1.25 [get_ports {dq_i[0]}]\n" % clock_name
        + "set_false_path -from [get_ports {pad_fp[0]}]\n"
        + "set_load 0.9 [get_ports {internal_non_pad_o}]\n",
        encoding="utf-8",
    )
    if scenario == "common":
        write_target_clock_inventory(root, "common", [clock_name])
    else:
        write_target_clock_inventory(root, "common", ["common_clk"])
        write_target_clock_inventory(root, scenario, ["common_clk", clock_name])

    manifest_rows = [
        {
            "scenario": scenario,
            "inst_name": "u_io",
            "module_name": "io_ring",
            "sdc_path": "inputs/u_io.sdc",
            "availability_status": "available",
            "note": "delivered",
        }
    ]
    if include_missing:
        manifest_rows.append(
            {
                "scenario": scenario,
                "inst_name": "u_missing",
                "module_name": "missing_ring",
                "sdc_path": "inputs/missing.sdc",
                "availability_status": "missing",
                "note": "owner delivery pending",
            }
        )
    write_target_manifest(root, scenario, manifest_rows)
    write_target_connections(root, include_missing, scenario)

    pending = root / "00_middle" / "scenario" / scenario / "pending"
    pending.mkdir(parents=True)
    (pending / "u_io.ports").write_text("input dq_i[0]\ninput pad_fp[0]\n", encoding="utf-8")
    if include_missing:
        (pending / "u_missing.ports").write_text("input missing_i\n", encoding="utf-8")


def build_inputs(d):
    clean_dir(d)
    write_info_all(d / "info_all.xlsx")
    write_ports(d / "ports_u_io.xlsx")
    write_clock_inventory(d / "clock_inventory.csv")
    (d / "u_io.sdc").write_text(
        "set_input_delay -clock [get_clocks dqs_clk] -max 1.25 [get_ports {dq_i[0]}]\n"
        "set_false_path -from [get_ports {pad_fp[0]}]\n",
        encoding="utf-8",
    )
    pending = d / "00_harden_port_inventory" / "pending"
    pending.mkdir(parents=True)
    (pending / "u_io.ports").write_text(
        "input dq_i[0]\n"
        "input pad_fp[0]\n"
        "input untouched_i\n",
        encoding="utf-8",
    )


def header_map(ws):
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def approve_bit_row(form):
    wb = load_workbook(str(form))
    ws = wb["io_constraints"]
    col = header_map(ws)
    matched_delay = False
    matched_false_path = False
    for row in range(2, ws.max_row + 1):
        pad = ws.cell(row, col["pad_name"]).value
        ctype = ws.cell(row, col["constraint_type"]).value
        if pad == "pad_dq[0]" and ctype == "input_delay":
            matched_delay = True
            ws.cell(row, col["apply"], "yes")
            ws.cell(row, col["review_status"], "approved")
            ws.cell(row, col["timing_class"], "timed")
            ws.cell(row, col["owner"], "io_owner")
            ws.cell(row, col["basis"], "bit-level DDR input budget")
        if pad == "pad_fp[0]" and ctype == "false_path":
            matched_false_path = True
            ws.cell(row, col["apply"], "yes")
            ws.cell(row, col["review_status"], "approved")
            ws.cell(row, col["timing_class"], "async")
            ws.cell(row, col["owner"], "legacy_exception_owner")
            ws.cell(row, col["basis"], "bit-level false path passthrough")
    require(matched_delay, "bit-level input_delay candidate was not extracted")
    require(matched_false_path, "bit-level false_path candidate was not extracted")
    wb.save(str(form))


def run_bit_pending_case():
    d = WORK / "bit_pending"
    build_inputs(d)

    first = sh([EX04, "-scenario", "common", "-input", "clock_inventory.csv"], d)
    require(first.returncode == 1, "first 04 run should create/sync workbook and stop")
    require((d / "04_soc_io_pads.xlsx").is_file(), "04 workbook was not created")

    approve_bit_row(d / "04_soc_io_pads.xlsx")
    second = sh([EX04, "-scenario", "common", "-input", "clock_inventory.csv"], d)
    require(second.returncode == 0, "04 bit generation failed:\n%s\n%s" % (second.stdout, second.stderr))

    sdc = (d / "common" / "04_soc_io_pads.sdc").read_text(encoding="utf-8")
    report = (d / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    pending = (d / "00_harden_port_inventory" / "pending" / "u_io.ports").read_text(encoding="utf-8")
    removed = (d / "00_harden_port_inventory" / "removed_log" / "04_soc_io_pads.removed").read_text(encoding="utf-8")

    require("set_input_delay -clock [get_clocks {dqs_clk}] -max 1.25 [get_ports {pad_dq[0]}]" in sdc, "bit-level SDC command missing")
    require("set_false_path -from [get_ports {pad_fp[0]}]" in sdc, "identity-rewritten false_path was not emitted")
    require("input dq_i[0]" not in pending, "covered bit-level pad port was not removed from pending")
    require("input pad_fp[0]" not in pending, "covered bit-level false_path port was not removed from pending")
    require("input untouched_i" in pending, "uncovered pending port was incorrectly removed")
    require("u_io input dq_i[0] covered_by=04_soc_io_pads" in removed, "removed log missing bit-level pad key")
    require("u_io input pad_fp[0] covered_by=04_soc_io_pads" in removed, "removed log missing false_path bit key")
    require("removed 2 harden pad port(s) from pending" in report, "pending removal was not reported")


def run_empty_command_error_case():
    d = WORK / "empty_command"
    build_inputs(d)
    first = sh([EX04, "-scenario", "common", "-input", "clock_inventory.csv"], d)
    require(first.returncode == 1, "empty-command first 04 run should sync workbook")
    wb = load_workbook(str(d / "04_soc_io_pads.xlsx"))
    ws = wb["io_constraints"]
    col = header_map(ws)
    row = ws.max_row + 1
    values = {
        "scenario": "common",
        "stage": "all",
        "corner": "all",
        "pad_name": "pad_dq[0]",
        "soc_object": "[get_ports {pad_dq[0]}]",
        "subsys_instance": "u_io",
        "subsys_port": "dq_i[0]",
        "direction": "input",
        "timing_class": "async",
        "constraint_type": "false_path",
        "source_type": "manual",
        "apply": "yes",
        "review_status": "approved",
        "basis": "manual row intentionally lacks rewritten/original command",
    }
    for key, value in values.items():
        ws.cell(row, col[key], value)
    wb.save(str(d / "04_soc_io_pads.xlsx"))
    result = sh([EX04, "-scenario", "common", "-input", "clock_inventory.csv"], d)
    require(result.returncode == 1, "approved empty false_path row should fail")
    report = (d / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    require("approved row cannot emit an SDC command" in report, "empty command error missing")


def run_noncanonical_port_error_case():
    d = WORK / "noncanonical_port"
    build_inputs(d)
    wb = load_workbook(str(d / "ports_u_io.xlsx"))
    ws = wb["u_io"]
    ws.append(["dq_bus", 2, 2, "top.pad_dq[1:0]", "", "", "", "", "", "", "", ""])
    wb.save(str(d / "ports_u_io.xlsx"))
    result = sh([EX04, "-scenario", "common", "-input", "clock_inventory.csv"], d)
    require(result.returncode == 1, "non-canonical bus/range input should fail")
    report = (d / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    require("dq_bus has width=2 used_width=2 but is not bit-expanded" in report, "bus scalar width error missing")
    require("non-canonical top pad pad_dq[1:0]" in report, "top range canonical error missing")


def run_target_complete_case():
    root = WORK / "target_complete"
    build_target_inputs(root)
    first = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(first.returncode == 1, "target first run should create/sync workbook")
    require(first.stdout.count("Author: Howard") == 1, "target stdout author marker missing or duplicated")
    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    require(form.is_file(), "target 04_middle workbook missing")
    wb = load_workbook(str(form), data_only=False)
    io_ws = wb["io_constraints"]
    io_col = header_map(io_ws)
    require(
        all(io_ws.cell(row, io_col["subsys_port"]).value != "internal_non_pad_o" for row in range(2, io_ws.max_row + 1)),
        "target non-pad command leaked into io_constraints",
    )
    pad_ws = wb["pad_inventory"]
    pad_col = header_map(pad_ws)
    pad_names = {pad_ws.cell(row, pad_col["pad_name"]).value for row in range(2, pad_ws.max_row + 1)}
    require("pad_scan_only" not in pad_names, "foreign-scenario edge leaked into target pad inventory")
    require(
        any(
            pad_ws.cell(row, pad_col["connection_id"]).value == "CONN_PAD_DQ"
            and pad_ws.cell(row, pad_col["source_workbook"]).value == "integration.xlsx"
            for row in range(2, pad_ws.max_row + 1)
        ),
        "00 connection provenance missing from pad inventory",
    )
    log_ws = wb["extraction_log"]
    log_col = header_map(log_ws)
    require(
        any(
            log_ws.cell(row, log_col["parse_status"]).value == "out_of_scope_non_pad"
            for row in range(2, log_ws.max_row + 1)
        ),
        "target non-pad command was not retained in extraction_log",
    )
    approve_bit_row(form)

    second = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(second.returncode == 0, "target complete generation failed:\n%s\n%s" % (second.stdout, second.stderr))
    sdc = (root / "04_result" / "common" / "04_soc_io_pads.sdc").read_text(encoding="utf-8")
    report = (root / "04_result" / "reports" / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    removed = root / "04_middle" / "scenario" / "common" / "removed_log" / "04_soc_io_pads.removed"
    require("# Author: Howard" in sdc, "target SDC author metadata missing")
    require("# Run completeness: complete" in sdc, "target complete metadata missing")
    require("Author  : Howard" in report, "target report author metadata missing")
    require("Run completeness: complete" in report, "target report completeness missing")
    require(removed.is_file(), "target 04 removed log missing")
    require((root / "00_middle" / "scenario" / "common" / "pending" / "u_io.ports").read_text(encoding="utf-8") == "", "target covered pending ports were not consumed")


def run_target_partial_strict_case():
    root = WORK / "target_partial"
    build_target_inputs(root, include_missing=True)
    first = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(first.returncode == 1, "partial target first run should sync workbook")
    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    approve_bit_row(form)

    partial = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(partial.returncode == 0, "partial target generation should continue for available SDC")
    sdc_path = root / "04_result" / "common" / "04_soc_io_pads.sdc"
    report_path = root / "04_result" / "reports" / "io_pad_check_report_common_all_all.txt"
    sdc = sdc_path.read_text(encoding="utf-8")
    report = report_path.read_text(encoding="utf-8")
    require("# Run completeness: partial" in sdc, "partial SDC completeness metadata missing")
    require("u_missing" in report and "HARDEN_SDC_MISSING" in report, "missing SDC evidence not reported")
    missing_pending = root / "00_middle" / "scenario" / "common" / "pending" / "u_missing.ports"
    require(missing_pending.read_text(encoding="utf-8") == "input missing_i\n", "missing SDC pending port was consumed")
    wb = load_workbook(str(form), data_only=False)
    pad_ws = wb["pad_inventory"]
    col = header_map(pad_ws)
    statuses = {
        pad_ws.cell(row, col["pad_name"]).value: pad_ws.cell(row, col["source_sdc_status"]).value
        for row in range(2, pad_ws.max_row + 1)
    }
    require(statuses.get("pad_missing") == "missing", "missing SDC pad status absent from pad inventory")

    before = sdc_path.read_text(encoding="utf-8")
    strict = sh([
        EX04, "--run-root", root, "-scenario", "common", "--require-complete-harden-sdc",
    ], BASE)
    require(strict.returncode == 1, "strict target mode must block partial harden SDC availability")
    strict_report = report_path.read_text(encoding="utf-8")
    require("HARDEN_SDC_COMPLETENESS_REQUIRED" in strict_report, "strict completeness error missing")
    require("Run completeness: partial" in strict_report, "strict report must preserve partial completeness state")
    require(sdc_path.read_text(encoding="utf-8") == before, "strict error run modified official SDC")
    require(missing_pending.read_text(encoding="utf-8") == "input missing_i\n", "strict error run modified missing pending")


def run_target_scenario_clock_case():
    root = WORK / "target_func"
    build_target_inputs(root, scenario="func", clock_name="func_io_clk")
    first = sh([EX04, "--run-root", root, "-scenario", "func"], BASE)
    require(first.returncode == 1, "scenario target first run should sync workbook")
    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    approve_bit_row(form)
    second = sh([EX04, "--run-root", root, "-scenario", "func"], BASE)
    require(second.returncode == 0, "scenario assembled clock generation failed:\n%s\n%s" % (second.stdout, second.stderr))
    sdc = (root / "04_result" / "scenarios" / "func_io_pads.sdc").read_text(encoding="utf-8")
    report = (root / "04_result" / "reports" / "io_pad_check_report_func_all_all.txt").read_text(encoding="utf-8")
    require("[get_clocks {func_io_clk}]" in sdc, "scenario-only assembled clock was not accepted")
    require("01_middle/assembled/func/clock_inventory.csv" in report, "scenario assembled inventory path not reported")


def run_target_approved_na_case():
    root = WORK / "target_approved_na"
    build_target_inputs(root)
    first = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(first.returncode == 1, "approved NA first run should sync workbook")
    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    wb = load_workbook(str(form))
    ws = wb["io_constraints"]
    col = header_map(ws)
    row = ws.max_row + 1
    values = {
        "scenario": "common",
        "stage": "all",
        "corner": "all",
        "pad_name": "pad_dq[0]",
        "subsys_instance": "u_io",
        "subsys_port": "dq_i[0]",
        "direction": "input",
        "timing_class": "untimed",
        "source_type": "na",
        "source_sdc_status": "not_required",
        "apply": "yes",
        "review_status": "approved",
        "owner": "io_owner",
        "basis": "demo pad is intentionally untimed and needs no 04 command",
        "reviewer": "04_regression",
        "review_date": "2026-07-20",
    }
    for key, value in values.items():
        ws.cell(row, col[key], value)
    wb.save(str(form))

    result = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(result.returncode == 0, "approved NA row should be a valid terminal disposition")
    report = (root / "04_result" / "reports" / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    pending = (root / "00_middle" / "scenario" / "common" / "pending" / "u_io.ports").read_text(encoding="utf-8")
    require("assembled approved NA : 1" in report, "approved NA coverage count missing")
    require("pad_dq[0]: dir=input class=untimed constraints=NA" in report, "approved NA pad status missing")
    require("input dq_i[0]" not in pending, "approved NA exact pad key was not consumed")
    require("input pad_fp[0]" in pending, "unreviewed pad was incorrectly consumed by NA")


def run_target_delay_conflict_case():
    root = WORK / "target_delay_conflict"
    build_target_inputs(root)
    first = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(first.returncode == 1, "delay conflict first run should sync workbook")
    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    wb = load_workbook(str(form))
    ws = wb["io_constraints"]
    col = header_map(ws)
    source_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, col["pad_name"]).value == "pad_dq[0]":
            source_row = row
            ws.cell(row, col["apply"], "yes")
            ws.cell(row, col["review_status"], "approved")
            ws.cell(row, col["timing_class"], "timed")
            ws.cell(row, col["owner"], "io_owner")
            ws.cell(row, col["basis"], "primary max delay")
            break
    require(source_row is not None, "source delay row missing")
    duplicate = ws.max_row + 1
    for header, column in col.items():
        ws.cell(duplicate, column, ws.cell(source_row, column).value)
    ws.cell(duplicate, col["source_type"], "manual")
    ws.cell(duplicate, col["source_sdc_file"], "")
    ws.cell(duplicate, col["source_line"], "")
    ws.cell(duplicate, col["source_digest"], "")
    ws.cell(duplicate, col["max_value"], "1.50")
    ws.cell(duplicate, col["basis"], "conflicting max delay")
    wb.save(str(form))

    result = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(result.returncode == 1, "duplicate delay semantic slot must block generation")
    report = (root / "04_result" / "reports" / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    require("duplicate/conflict for delay semantic slot" in report, "delay semantic conflict was not reported")


def run_target_connection_contract_case():
    root = WORK / "target_connection_contract"
    build_target_inputs(root)
    connection = root / "00_middle" / "connection_inventory.csv"
    with connection.open("r", encoding="utf-8", newline="") as file_obj:
        rows = list(csv.DictReader(file_obj))
        fields = list(rows[0].keys())
    rows[0]["src_port"] = "pad_dq[1:0]"
    with connection.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    result = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(result.returncode == 1, "target noncanonical 00 edge must fail")
    report = (root / "04_result" / "reports" / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    require("CONNECTION_PORT_NONCANONICAL" in report, "target canonical edge error missing")

    root = WORK / "target_connection_values"
    build_target_inputs(root)
    connection = root / "00_middle" / "connection_inventory.csv"
    with connection.open("r", encoding="utf-8", newline="") as file_obj:
        rows = list(csv.DictReader(file_obj))
        fields = list(rows[0].keys())
    rows[0]["scenario_scope"] = ""
    rows[1]["connection_type"] = "top_to_harden"
    with connection.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    result = sh([EX04, "--run-root", root, "-scenario", "common"], BASE)
    require(result.returncode == 1, "target invalid connection values must fail")
    report = (root / "04_result" / "reports" / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    require("CONNECTION_SCENARIO_SCOPE_EMPTY" in report, "empty scenario_scope error missing")
    require("CONNECTION_TYPE_INVALID" in report, "invalid connection_type error missing")

    root = WORK / "target_legacy_input_forbidden"
    build_target_inputs(root)
    result = sh([
        EX04, "--run-root", root, "-scenario", "common", "--info-all", "inputs/info_all.xlsx",
        "--no-update-pending",
    ], BASE)
    require(result.returncode == 1, "target legacy integration input must fail")
    report = (root / "04_result" / "reports" / "io_pad_check_report_common_all_all.txt").read_text(encoding="utf-8")
    require("TARGET_LEGACY_INTEGRATION_INPUT_FORBIDDEN" in report, "target legacy-input error missing")
    require("Pending update: disabled" in report, "no-update-pending status missing from report")


def run_flat_target_case():
    root = WORK / "flat_target"
    clean_dir(root)
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_info_all(inputs / "info_all.xlsx")
    write_ports(inputs / "port_u_io.xlsx")
    wb = load_workbook(str(inputs / "port_u_io.xlsx"))
    ws = wb["u_io"]
    columns = header_map(ws)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, columns["Input Used Width"], "")
    ws.delete_rows(4, 1)
    wb.save(str(inputs / "port_u_io.xlsx"))
    (inputs / "u_io.sdc").write_text(
        "set_input_delay -clock [get_clocks dqs_clk] -max 1.25 [get_ports {dq_i[0]}]\n"
        "set_false_path -from [get_ports {pad_fp[0]}]\n",
        encoding="utf-8",
    )
    with (inputs / "run_context.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["run_id", "mode_label", "design_revision", "note"])
        writer.writerow(["RUN_04_FLAT", "func", "rev_a", "04 flat target regression"])
    with (inputs / "required_views.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"])
        writer.writerow(["prects_ss_125", "prects", "ss_125", "no", "yes", "no", "no", "04 flat view"])
    with (inputs / "virtual_clocks.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["clock_name", "period", "waveform", "note"])
        writer.writerow(["dqs_clk", "10", "0 5", "04 IO reference"])

    stage_00 = sh([EX00, "--run-root", root], BASE)
    require(stage_00.returncode == 0, "flat 00 setup failed:\n%s\n%s" % (stage_00.stdout, stage_00.stderr))
    stage_01 = sh([EX01, "--run-root", root], BASE)
    require(stage_01.returncode == 0, "flat 01 setup failed:\n%s\n%s" % (stage_01.stdout, stage_01.stderr))
    first = sh([EX04, "--run-root", root], BASE)
    require(first.returncode == 1, "flat 04 first run should synchronize the review workbook")
    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    approved_delay = False
    for row in range(2, io_sheet.max_row + 1):
        if (
            io_sheet.cell(row, io_columns["pad_name"]).value == "pad_dq[0]"
            and io_sheet.cell(row, io_columns["constraint_type"]).value == "input_delay"
        ):
            approved_delay = True
            io_sheet.cell(row, io_columns["apply"], "yes")
            io_sheet.cell(row, io_columns["review_status"], "approved")
            io_sheet.cell(row, io_columns["timing_class"], "timed")
            io_sheet.cell(row, io_columns["owner"], "io_owner")
            io_sheet.cell(row, io_columns["basis"], "flat required-view input budget")
    require(approved_delay, "flat input_delay candidate missing")
    pad_sheet = book["pad_inventory"]
    pad_columns = header_map(pad_sheet)
    routed = False
    for row in range(2, pad_sheet.max_row + 1):
        if (
            pad_sheet.cell(row, pad_columns["pad_name"]).value == "pad_fp[0]"
            and pad_sheet.cell(row, pad_columns["view_id"]).value == "prects_ss_125"
        ):
            routed = True
            values = {
                "pad_disposition": "route_to_30", "apply": "yes",
                "review_status": "approved", "owner": "cdc_owner",
                "basis": "flat false_path is owned by stage 30",
                "related_exception_intent": "false_path for exact pad_fp[0] path",
                "reviewer": "04_regression", "review_date": "2026-07-20",
                "approved_machine_digest": pad_sheet.cell(row, pad_columns["machine_digest"]).value,
            }
            for name, value in values.items():
                pad_sheet.cell(row, pad_columns[name], value)
    require(routed, "flat route_to_30 pad row missing")
    book.save(str(form))
    second = sh([EX04, "--run-root", root], BASE)
    require(second.returncode == 0, "flat 04 generation failed:\n%s\n%s" % (second.stdout, second.stderr))
    flat_sdc_path = root / "04_result" / "04_soc_io_pads_prects_ss_125.sdc"
    require(flat_sdc_path.is_file(), "flat 04 SDC missing")
    flat_sdc = flat_sdc_path.read_text(encoding="utf-8")
    require("set_input_delay" in flat_sdc and "pad_dq[0]" in flat_sdc, "flat all/all input_delay was not assembled into the required view")
    require("set_false_path" not in flat_sdc and "pad_fp[0]" not in flat_sdc, "flat 04 emitted an exception owned by stage 30")
    require((root / "04_middle" / "pad_inventory.csv").is_file(), "flat pad inventory missing")
    require((root / "04_middle" / "port_accounting_delta.meta").is_file(), "flat accounting meta missing")
    require((root / "04_middle" / "completion" / "prects_ss_125.meta").is_file(), "flat view completion missing")
    require(not (root / "00_middle" / "scenario").exists(), "flat 04 unexpectedly required a scenario adapter")
    require(not (root / "00_middle" / "connection_inventory.csv").exists(), "flat 04 unexpectedly required connection_inventory.csv")

    for completion_path in (
        root / "00_middle" / "stage_completion.meta",
        root / "01_middle" / "stage_completion.meta",
    ):
        official_completion = completion_path.read_bytes()
        completion_path.unlink()
        missing_upstream = sh([EX04, "--run-root", root], BASE)
        require(missing_upstream.returncode == 1, "flat 04 completed with a missing upstream completion")
        missing_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
        require("required upstream completion is missing" in missing_report, "missing upstream completion diagnostic missing")
        require(flat_sdc_path.read_text(encoding="utf-8") == flat_sdc, "missing upstream completion changed official SDC")
        completion_path.write_bytes(official_completion)

    stage_01_completion_path = root / "01_middle" / "stage_completion.meta"
    official_stage_01 = stage_01_completion_path.read_bytes()
    incomplete_stage_01 = json.loads(official_stage_01.decode("utf-8"))
    incomplete_stage_01["completion_status"] = "review_required"
    stage_01_completion_path.write_text(json.dumps(incomplete_stage_01, indent=2) + "\n", encoding="utf-8")
    incomplete_upstream = sh([EX04, "--run-root", root], BASE)
    require(incomplete_upstream.returncode == 1, "flat 04 completed with an incomplete 01 stage")
    incomplete_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("upstream completion is stale or incomplete" in incomplete_report, "incomplete upstream completion diagnostic missing")
    require(flat_sdc_path.read_text(encoding="utf-8") == flat_sdc, "incomplete upstream completion changed official SDC")
    stage_01_completion_path.write_bytes(official_stage_01)

    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    delay_row = next(
        row for row in range(2, io_sheet.max_row + 1)
        if io_sheet.cell(row, io_columns["constraint_type"]).value == "input_delay"
    )
    io_sheet.cell(delay_row, io_columns["max_value"], "")
    book.save(str(form))
    commandless = sh([EX04, "--run-root", root], BASE)
    require(commandless.returncode == 1, "assembled all/all row without a value bypassed validation")
    commandless_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("approved row cannot emit an SDC command" in commandless_report, "assembled commandless-row error missing")
    require(flat_sdc_path.read_text(encoding="utf-8") == flat_sdc, "commandless all/all row changed official SDC")
    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    io_sheet.cell(delay_row, io_columns["max_value"], "1.25")
    io_sheet.cell(delay_row, io_columns["constraint_type"], "output_delay")
    io_sheet.cell(delay_row, io_columns["direction"], "output")
    book.save(str(form))
    direction_conflict = sh([EX04, "--run-root", root], BASE)
    require(direction_conflict.returncode == 1, "flat input pad accepted an output effective direction")
    direction_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("conflicts with integration direction input" in direction_report, "integration direction conflict diagnostic missing")
    require(flat_sdc_path.read_text(encoding="utf-8") == flat_sdc, "direction-conflict row changed official SDC")
    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    io_sheet.cell(delay_row, io_columns["constraint_type"], "input_delay")
    io_sheet.cell(delay_row, io_columns["direction"], "input")
    approved_false_path = False
    for row in range(2, io_sheet.max_row + 1):
        if io_sheet.cell(row, io_columns["constraint_type"]).value == "false_path":
            approved_false_path = True
            io_sheet.cell(row, io_columns["apply"], "yes")
            io_sheet.cell(row, io_columns["review_status"], "approved")
            io_sheet.cell(row, io_columns["timing_class"], "async")
            io_sheet.cell(row, io_columns["owner"], "legacy_exception_owner")
            io_sheet.cell(row, io_columns["basis"], "intentional flat ownership rejection test")
    require(approved_false_path, "flat false_path candidate missing")
    book.save(str(form))
    rejected = sh([EX04, "--run-root", root], BASE)
    require(rejected.returncode == 1, "flat 04 accepted an exception command")
    flat_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("flat 04 does not own false_path" in flat_report, "flat false_path ownership error missing")
    require(flat_sdc_path.read_text(encoding="utf-8") == flat_sdc, "rejected flat false_path changed official SDC")


def run_flat_empty_pad_multiview_case():
    root = WORK / "flat_empty_pad_multiview"
    clean_dir(root)
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_info_all(inputs / "info_all.xlsx")

    workbook = Workbook()
    port_sheet = workbook.active
    port_sheet.title = "u_io"
    port_sheet.append([
        "Input", "Input Width", "Input Used Width", "From Whom",
        "Output", "Output Width", "Output Used Width", "To Top",
        "Inout", "Inout Width", "Inout Connectivity", "Inout Name",
    ])
    port_sheet.append(["internal_i", 1, "", "", "", "", "", "", "", "", "", ""])
    workbook.save(str(inputs / "port_u_io.xlsx"))
    (inputs / "u_io.sdc").write_text("# no IO constraints for this harden\n", encoding="utf-8")

    with (inputs / "run_context.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["run_id", "mode_label", "design_revision", "note"])
        writer.writerow(["RUN_04_EMPTY_MULTI", "func", "rev_empty", "04 empty pad multi-view regression"])
    required_views = [
        ("prects_ss_125", "prects", "ss_125"),
        ("all_all", "all", "all"),
        ("postcts_ss_125", "postcts", "ss_125"),
    ]
    with (inputs / "required_views.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"])
        for view_id, stage, corner in required_views:
            writer.writerow([view_id, stage, corner, "no", "yes", "no", "no", "empty 04 inventory"])
    with (inputs / "virtual_clocks.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["clock_name", "period", "waveform", "note"])
        writer.writerow(["dqs_clk", "10", "0 5", "empty 04 inventory setup"])

    stage_00 = sh([EX00, "--run-root", root], BASE)
    require(stage_00.returncode == 0, "empty/multi 00 setup failed:\n%s\n%s" % (stage_00.stdout, stage_00.stderr))
    stage_01 = sh([EX01, "--run-root", root], BASE)
    require(stage_01.returncode == 0, "empty/multi 01 setup failed:\n%s\n%s" % (stage_01.stdout, stage_01.stderr))

    inventory_path = root / "04_middle" / "pad_inventory.csv"
    inventory_meta_path = root / "04_middle" / "pad_inventory.meta"
    stage_completion_path = root / "04_middle" / "stage_completion.meta"
    completed_views = set()
    for index, (view_id, stage, corner) in enumerate(required_views):
        result = sh([EX04, "--run-root", root, "--stage", stage, "--corner", corner], BASE)
        if index == 0:
            require(result.returncode == 1, "empty/multi first 04 run should create the review workbook")
            result = sh([EX04, "--run-root", root, "--stage", stage, "--corner", corner], BASE)
        require(result.returncode == 0, "empty/multi %s generation failed:\n%s\n%s" % (view_id, result.stdout, result.stderr))
        completed_views.add(view_id)

        with inventory_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            require(list(csv.DictReader(file_obj)) == [], "empty pad inventory unexpectedly published owner rows")
        inventory_meta = json.loads(inventory_meta_path.read_text(encoding="utf-8"))
        require(set(inventory_meta["view_ids"]) == completed_views, "empty pad inventory lost an authenticated prior view")
        stage_completion = json.loads(stage_completion_path.read_text(encoding="utf-8"))
        expected_status = "complete" if index == len(required_views) - 1 else "review_required"
        require(stage_completion["completion_status"] == expected_status, "empty multi-view completion status is wrong")
        require(set(stage_completion["required_view_completions"]) == completed_views, "empty multi-view completion map is incomplete")

        output_path = root / "04_result" / (
            "04_soc_io_pads.sdc"
            if stage == "all" and corner == "all"
            else "04_soc_io_pads_%s_%s.sdc" % (stage, corner)
        )
        output_text = output_path.read_text(encoding="utf-8")
        require("# No IO/pad commands emitted" in output_text, "empty view did not publish the explicit no-command marker")
        require(not any(line.strip().startswith("set_") for line in output_text.splitlines()), "empty view emitted an SDC command")

    inventory_digest = hashlib.sha256(inventory_path.read_bytes()).hexdigest()
    for view_id, stage, corner in required_views:
        completion_path = root / "04_middle" / "completion" / ("%s_%s.meta" % (stage, corner))
        completion = json.loads(completion_path.read_text(encoding="utf-8"))
        require(completion["view_id"] == view_id, "empty view completion identity is wrong")
        require(completion["pad_inventory_digest"] == inventory_digest, "empty view completion does not authenticate the header-only inventory")


def run_flat_fanout_case():
    root = WORK / "flat_fanout"
    clean_dir(root)
    inputs = root / "inputs"
    inputs.mkdir(parents=True)

    info_book = Workbook()
    info_sheet = info_book.active
    info_sheet.title = "info_all"
    info_sheet.append(["inst_name", "module_name", "owner", "sdc_path"])
    info_sheet.append(["u_fan_a", "fan_sink", "alice", "u_fan_a.sdc"])
    info_sheet.append(["u_fan_b", "fan_sink", "bob", "u_fan_b.sdc"])
    info_book.save(str(inputs / "info_all.xlsx"))

    for instance in ("u_fan_a", "u_fan_b"):
        port_book = Workbook()
        port_sheet = port_book.active
        port_sheet.title = instance
        port_sheet.append([
            "Input", "Input Width", "Input Used Width", "From Whom",
            "Output", "Output Width", "Output Used Width", "To Top",
            "Inout", "Inout Width", "Inout Connectivity", "Inout Name",
        ])
        port_sheet.append(["fan_i[0]", 1, "", "top.pad_fan[0]", "", "", "", "", "", "", "", ""])
        port_book.save(str(inputs / ("port_%s.xlsx" % instance)))
    (inputs / "u_fan_a.sdc").write_text(
        "set_input_transition 0.07 [get_ports {fan_i[0]}]\n"
        "set_input_delay -clock [get_clocks dqs_clk] -max 1.10 [get_ports {fan_i[0]}]\n",
        encoding="utf-8",
    )
    (inputs / "u_fan_b.sdc").write_text("# shared top pad command is delivered by u_fan_a\n", encoding="utf-8")
    with (inputs / "run_context.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["run_id", "mode_label", "design_revision", "note"])
        writer.writerow(["RUN_04_FANOUT", "func", "rev_fanout", "04 exact top-bit fanout regression"])
    with (inputs / "required_views.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"])
        writer.writerow(["prects_ss_125", "prects", "ss_125", "no", "yes", "no", "no", "fanout view"])
    with (inputs / "virtual_clocks.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["clock_name", "period", "waveform", "note"])
        writer.writerow(["dqs_clk", "10", "0 5", "fanout IO reference"])

    stage_00 = sh([EX00, "--run-root", root], BASE)
    require(stage_00.returncode == 0, "fanout 00 setup failed:\n%s\n%s" % (stage_00.stdout, stage_00.stderr))
    stage_01 = sh([EX01, "--run-root", root], BASE)
    require(stage_01.returncode == 0, "fanout 01 setup failed:\n%s\n%s" % (stage_01.stdout, stage_01.stderr))
    command = [EX04, "--run-root", root, "--stage", "prects", "--corner", "ss_125"]
    first = sh(command, BASE)
    require(first.returncode == 1, "fanout first 04 run should synchronize the review workbook")

    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    transition_row = next(
        row for row in range(2, io_sheet.max_row + 1)
        if io_sheet.cell(row, io_columns["constraint_type"]).value == "input_transition"
    )
    for name, value in (
        ("apply", "yes"), ("review_status", "approved"),
        ("timing_class", "async"), ("owner", "fanout_io_owner"),
        ("basis", "one exact top-port electrical command covers every fanout edge"),
    ):
        io_sheet.cell(transition_row, io_columns[name], value)
    pad_sheet = book["pad_inventory"]
    pad_columns = header_map(pad_sheet)
    route_row = next(
        row for row in range(2, pad_sheet.max_row + 1)
        if pad_sheet.cell(row, pad_columns["subsys_instance"]).value == "u_fan_b"
    )
    for name, value in {
        "pad_disposition": "route_to_30", "apply": "yes",
        "review_status": "approved", "owner": "fanout_exception_owner",
        "basis": "one fanout edge is exception-owned while electrical IO remains in 04",
        "related_exception_intent": "false_path for pad_fan[0] to u_fan_b/fan_i[0]",
        "reviewer": "04_regression", "review_date": "2026-07-20",
        "approved_machine_digest": pad_sheet.cell(route_row, pad_columns["machine_digest"]).value,
    }.items():
        pad_sheet.cell(route_row, pad_columns[name], value)
    book.save(str(form))

    electrical = sh(command, BASE)
    require(electrical.returncode == 0, "fanout electrical/route coexistence failed:\n%s\n%s" % (electrical.stdout, electrical.stderr))
    output_path = root / "04_result" / "04_soc_io_pads_prects_ss_125.sdc"
    electrical_sdc = output_path.read_text(encoding="utf-8")
    require("set_input_transition 0.07 [get_ports {pad_fan[0]}]" in electrical_sdc, "fanout electrical command was not emitted on the exact top bit")
    require("set_input_delay" not in electrical_sdc, "unapproved fanout input delay was emitted")
    inventory_path = root / "04_middle" / "pad_inventory.csv"
    with inventory_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        electrical_rows = [row for row in csv.DictReader(file_obj) if row.get("pad_name") == "pad_fan[0]"]
    require(len(electrical_rows) == 2, "fanout did not publish one canonical owner row per sink edge")
    by_instance = {row["subsys_instance"]: row for row in electrical_rows}
    require(by_instance["u_fan_a"]["pad_disposition"] == "constrained", "reviewed fanout edge was not constrained")
    require(by_instance["u_fan_b"]["pad_disposition"] == "route_to_30", "exception fanout edge lost route_to_30")

    def used_value(instance):
        used_book = load_workbook(str(inputs / ("port_%s.xlsx" % instance)), read_only=True, data_only=True)
        used_sheet = used_book[instance]
        used_columns = header_map(used_sheet)
        value = str(used_sheet.cell(2, used_columns["Input Used Width"]).value or "")
        used_book.close()
        return value

    require(used_value("u_fan_a") == "0", "constrained fanout edge was not accounted")
    require(used_value("u_fan_b") == "", "route_to_30 fanout edge was incorrectly accounted")

    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    delay_row = next(
        row for row in range(2, io_sheet.max_row + 1)
        if io_sheet.cell(row, io_columns["constraint_type"]).value == "input_delay"
    )
    for name, value in (
        ("apply", "yes"), ("review_status", "approved"),
        ("timing_class", "timed"), ("owner", "fanout_io_owner"),
        ("basis", "normal input budget applies to the exact top bit and every sink edge"),
    ):
        io_sheet.cell(delay_row, io_columns[name], value)
    book.save(str(form))
    conflict = sh(command, BASE)
    if conflict.returncode == 1:
        first_conflict_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
        if "route_to_30 conflicts with active 04 normal timing" not in first_conflict_report:
            conflict = sh(command, BASE)
    require(conflict.returncode == 1, "fanout route_to_30 bypassed active normal top-port timing")
    conflict_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("route_to_30 conflicts with active 04 normal timing" in conflict_report, "fanout normal-timing route conflict diagnostic missing")
    require(output_path.read_text(encoding="utf-8") == electrical_sdc, "rejected fanout route conflict changed official SDC")
    require(used_value("u_fan_b") == "", "rejected fanout route conflict changed accounting")

    book = load_workbook(str(form))
    pad_sheet = book["pad_inventory"]
    pad_columns = header_map(pad_sheet)
    route_row = next(
        row for row in range(2, pad_sheet.max_row + 1)
        if pad_sheet.cell(row, pad_columns["subsys_instance"]).value == "u_fan_b"
    )
    for name, value in (
        ("pad_disposition", ""), ("apply", "no"),
        ("review_status", "pending"), ("approved_machine_digest", ""),
    ):
        pad_sheet.cell(route_row, pad_columns[name], value)
    book.save(str(form))
    constrained = sh(command, BASE)
    require(constrained.returncode == 0, "fanout normal timing did not constrain every sink edge:\n%s\n%s" % (constrained.stdout, constrained.stderr))
    constrained_sdc = output_path.read_text(encoding="utf-8")
    require("set_input_delay" in constrained_sdc and "[get_ports {pad_fan[0]}]" in constrained_sdc, "fanout normal timing command was not emitted")
    with inventory_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        final_rows = [row for row in csv.DictReader(file_obj) if row.get("pad_name") == "pad_fan[0]"]
    require({row["pad_disposition"] for row in final_rows} == {"constrained"}, "normal top-port timing did not resolve every fanout edge as constrained")
    require(used_value("u_fan_a") == "0" and used_value("u_fan_b") == "0", "normal top-port timing did not account every fanout sink bit")
    delta_text = (root / "04_middle" / "port_accounting_delta.csv").read_text(encoding="utf-8")
    require(all(row["pad_id"] in delta_text for row in final_rows), "fanout accounting delta does not reference every canonical pad owner")


def run_flat_vector_expansion_case():
    root = WORK / "flat_vector_expansion"
    clean_dir(root)
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_info_all(inputs / "info_all.xlsx")

    port_book = Workbook()
    port_sheet = port_book.active
    port_sheet.title = "u_io"
    port_sheet.append([
        "Input", "Input Width", "Input Used Width", "From Whom",
        "Output", "Output Width", "Output Used Width", "To Top",
        "Inout", "Inout Width", "Inout Connectivity", "Inout Name",
    ])
    port_sheet.append(["vec_i[3:0]", 4, "", "top.pad_vec[7:4]", "", "", "", "", "", "", "", ""])
    port_sheet.append(["", "", "", "", "vec_o[1:0]", 2, "", "top.pad_out[4:5]", "", "", "", ""])
    port_book.save(str(inputs / "port_u_io.xlsx"))
    (inputs / "u_io.sdc").write_text(
        "set_input_transition 0.06 [get_ports {vec_i[3:0]}]\n"
        "set_input_delay -clock [get_clocks dqs_clk] -max 0.95 [get_ports {vec_i[3:0]}]\n"
        "set_max_transition 0.14 [get_ports {vec_i[3:2]}]\n"
        "set_load 0.22 [get_ports {vec_o[1:0]}]\n",
        encoding="utf-8",
    )
    with (inputs / "run_context.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["run_id", "mode_label", "design_revision", "note"])
        writer.writerow(["RUN_04_VECTOR", "func", "rev_vector", "04 range evidence exact-bit expansion"])
    with (inputs / "required_views.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"])
        writer.writerow(["prects_ss_125", "prects", "ss_125", "no", "yes", "no", "no", "vector view"])
    with (inputs / "virtual_clocks.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["clock_name", "period", "waveform", "note"])
        writer.writerow(["dqs_clk", "10", "0 5", "vector IO reference"])

    stage_00 = sh([EX00, "--run-root", root], BASE)
    require(stage_00.returncode == 0, "vector 00 setup failed:\n%s\n%s" % (stage_00.stdout, stage_00.stderr))
    stage_01 = sh([EX01, "--run-root", root], BASE)
    require(stage_01.returncode == 0, "vector 01 setup failed:\n%s\n%s" % (stage_01.stdout, stage_01.stderr))
    command = [EX04, "--run-root", root, "--stage", "prects", "--corner", "ss_125"]
    first = sh(command, BASE)
    require(first.returncode == 1, "vector first 04 run should synchronize exact-bit review rows")

    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    expected_input_pads = {"pad_vec[7]", "pad_vec[6]", "pad_vec[5]", "pad_vec[4]"}
    expected_output_pads = {"pad_out[4]", "pad_out[5]"}
    expected_pads = expected_input_pads | expected_output_pads
    expected_harden_ports = {"vec_i[3]", "vec_i[2]", "vec_i[1]", "vec_i[0]", "vec_o[1]", "vec_o[0]"}
    reviewed_rows = []
    for row in range(2, io_sheet.max_row + 1):
        ctype = io_sheet.cell(row, io_columns["constraint_type"]).value
        if ctype not in {"input_transition", "input_delay", "max_transition", "load"}:
            continue
        pad_name = io_sheet.cell(row, io_columns["pad_name"]).value
        require(pad_name in expected_pads, "range IO evidence was not expanded to an exact top bit")
        require(io_sheet.cell(row, io_columns["subsys_port"]).value in expected_harden_ports, "range IO evidence retained a non-exact harden port")
        require(io_sheet.cell(row, io_columns["soc_object"]).value == "[get_ports {%s}]" % pad_name, "range IO evidence retained a broad top collection")
        require(io_sheet.cell(row, io_columns["object_granularity"]).value == "single_pad", "expanded IO row is not marked single_pad")
        for name, value in (
            ("apply", "yes"), ("review_status", "approved"),
            ("timing_class", "timed" if ctype in {"input_delay", "load"} else "async"),
            ("owner", "vector_io_owner"),
            ("basis", "range evidence was reviewed as independent exact pad bits"),
        ):
            io_sheet.cell(row, io_columns[name], value)
        reviewed_rows.append(row)
    require(len(reviewed_rows) == 12, "input/output range and subrange commands did not expand into twelve exact review rows")
    book.save(str(form))

    generated = sh(command, BASE)
    require(generated.returncode == 0, "vector exact-bit generation failed:\n%s\n%s" % (generated.stdout, generated.stderr))
    output_path = root / "04_result" / "04_soc_io_pads_prects_ss_125.sdc"
    output_text = output_path.read_text(encoding="utf-8")
    require(output_text.count("set_input_transition") == 4, "vector transition did not emit one command per exact top bit")
    require(output_text.count("set_input_delay") == 4, "vector input delay did not emit one command per exact top bit")
    require(output_text.count("set_max_transition") == 2, "vector subrange did not emit one command per selected exact top bit")
    require(output_text.count("set_load") == 2, "vector output load did not emit one command per exact top bit")
    for pad_name in expected_pads:
        require("[get_ports {%s}]" % pad_name in output_text, "vector SDC is missing exact target %s" % pad_name)
    require("pad_vec[7:4]" not in output_text and "pad_out[4:5]" not in output_text, "vector SDC emitted an unreviewed range target")

    inventory_path = root / "04_middle" / "pad_inventory.csv"
    with inventory_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        inventory_rows = list(csv.DictReader(file_obj))
    require({row["pad_name"] for row in inventory_rows} == expected_pads, "vector canonical inventory bit set is wrong")
    require({row["pad_disposition"] for row in inventory_rows} == {"constrained"}, "vector exact bits were not resolved as constrained")
    used_book = load_workbook(str(inputs / "port_u_io.xlsx"), read_only=True, data_only=True)
    used_sheet = used_book["u_io"]
    used_columns = header_map(used_sheet)
    input_used_value = str(used_sheet.cell(2, used_columns["Input Used Width"]).value or "")
    output_used_value = str(used_sheet.cell(3, used_columns["Output Used Width"]).value or "")
    used_book.close()
    require(input_used_value == "0,1,2,3", "vector input accounting did not union all harden bits")
    require(output_used_value == "0,1", "vector output accounting did not preserve reversed top/harden bit order")
    delta_text = (root / "04_middle" / "port_accounting_delta.csv").read_text(encoding="utf-8")
    require(all(row["pad_id"] in delta_text for row in inventory_rows), "vector accounting delta lost an exact pad owner")

    source_sdc_path = inputs / "u_io.sdc"
    source_sdc = source_sdc_path.read_text(encoding="utf-8")
    source_sdc_path.write_text(
        source_sdc + "set_max_capacitance 0.20 [get_ports {vec_i[4:3]}]\n",
        encoding="utf-8",
    )
    partial = sh(command, BASE)
    require(partial.returncode == 1, "vector selector with an unmapped bit was partially accepted")
    partial_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("selector vec_i[4:3] bit 4 maps to 0 canonical pad edge(s)" in partial_report, "unmapped vector-bit fail-closed diagnostic missing")
    require(output_path.read_text(encoding="utf-8") == output_text, "unmapped vector selector changed official SDC")
    source_sdc_path.write_text(source_sdc, encoding="utf-8")

    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    bypass_row = reviewed_rows[0]
    io_sheet.cell(bypass_row, io_columns["object_granularity"], "pattern")
    io_sheet.cell(bypass_row, io_columns["soc_object"], "[get_ports {pad_vec[7:4]}]")
    book.save(str(form))
    bypass = sh(command, BASE)
    require(bypass.returncode == 1, "flat vector row bypassed exact-bit review with object_granularity=pattern")
    bypass_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("removed 1 obsolete auto-extracted broad/range IO row" in bypass_report, "flat range/pattern resynchronization diagnostic missing")
    require(output_path.read_text(encoding="utf-8") == output_text, "rejected vector pattern changed official SDC")


def run_flat_inout_direction_case():
    root = WORK / "flat_inout_direction"
    clean_dir(root)
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_info_all(inputs / "info_all.xlsx")

    port_book = Workbook()
    port_sheet = port_book.active
    port_sheet.title = "u_io"
    port_sheet.append([
        "Input", "Input Width", "Input Used Width", "From Whom",
        "Output", "Output Width", "Output Used Width", "To Top",
        "Inout", "Inout Width", "Inout Connectivity", "Inout Name",
    ])
    port_sheet.append(["", "", "", "", "", "", "", "", "gpio_io[0:1]", 2, "top.pad_gpio[3:2]", ""])
    port_book.save(str(inputs / "port_u_io.xlsx"))
    (inputs / "u_io.sdc").write_text(
        "set_input_transition 0.05 [get_ports {gpio_io[0:1]}]\n",
        encoding="utf-8",
    )
    with (inputs / "run_context.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["run_id", "mode_label", "design_revision", "note"])
        writer.writerow(["RUN_04_INOUT", "gpio_in", "rev_gpio", "04 inout effective-direction regression"])
    with (inputs / "required_views.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"])
        writer.writerow(["prects_ss_125", "prects", "ss_125", "no", "yes", "no", "no", "GPIO input view"])
    with (inputs / "virtual_clocks.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["clock_name", "period", "waveform", "note"])
        writer.writerow(["dqs_clk", "10", "0 5", "inout setup"])

    stage_00 = sh([EX00, "--run-root", root], BASE)
    require(stage_00.returncode == 0, "inout 00 setup failed:\n%s\n%s" % (stage_00.stdout, stage_00.stderr))
    stage_01 = sh([EX01, "--run-root", root], BASE)
    require(stage_01.returncode == 0, "inout 01 setup failed:\n%s\n%s" % (stage_01.stdout, stage_01.stderr))
    command = [EX04, "--run-root", root, "--stage", "prects", "--corner", "ss_125"]
    first = sh(command, BASE)
    require(first.returncode == 1, "inout first 04 run should synchronize the review workbook")

    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    rows = [
        row for row in range(2, io_sheet.max_row + 1)
        if io_sheet.cell(row, io_columns["constraint_type"]).value == "input_transition"
    ]
    require(len(rows) == 2, "inout range did not expand into two exact review rows")
    require(
        {io_sheet.cell(row, io_columns["pad_name"]).value for row in rows}
        == {"pad_gpio[3]", "pad_gpio[2]"},
        "inout range exact top-bit mapping is wrong",
    )
    for row in rows:
        require(io_sheet.cell(row, io_columns["direction"]).value == "inout", "inout extraction lost integration direction")
        for name, value in (
            ("apply", "yes"), ("review_status", "approved"),
            ("timing_class", "async"), ("owner", "gpio_owner"),
            ("basis", "GPIO is configured as input in this run"),
        ):
            io_sheet.cell(row, io_columns[name], value)
    book.save(str(form))
    unresolved = sh(command, BASE)
    require(unresolved.returncode == 1, "approved inout row emitted without an effective input/output direction")
    unresolved_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("inout pad requires an effective direction of input or output" in unresolved_report, "inout effective-direction diagnostic missing")

    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    for row in rows:
        io_sheet.cell(row, io_columns["direction"], "input")
    pad_sheet = book["pad_inventory"]
    pad_columns = header_map(pad_sheet)
    route_rows = [
        row for row in range(2, pad_sheet.max_row + 1)
        if pad_sheet.cell(row, pad_columns["pad_name"]).value in {"pad_gpio[3]", "pad_gpio[2]"}
    ]
    require(len(route_rows) == 2, "inout route review rows are missing")
    for route_row in route_rows:
        pad_name = pad_sheet.cell(route_row, pad_columns["pad_name"]).value
        for name, value in {
            "pad_disposition": "route_to_30", "effective_direction": "output",
            "apply": "yes", "review_status": "approved",
            "owner": "gpio_exception_owner",
            "basis": "exercise oriented inout exception handoff",
            "related_exception_intent": "false_path for %s GPIO edge" % pad_name,
            "reviewer": "04_regression", "review_date": "2026-07-20",
            "approved_machine_digest": pad_sheet.cell(route_row, pad_columns["machine_digest"]).value,
        }.items():
            pad_sheet.cell(route_row, pad_columns[name], value)
    book.save(str(form))
    unsupported_output_route = sh(command, BASE)
    require(unsupported_output_route.returncode == 1, "inout output route published a reverse 30 endpoint")
    unsupported_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("inout route_to_30 effective_direction=output is not supported" in unsupported_report, "inout output-route fail-closed diagnostic missing")

    book = load_workbook(str(form))
    pad_sheet = book["pad_inventory"]
    pad_columns = header_map(pad_sheet)
    for route_row in route_rows:
        pad_sheet.cell(route_row, pad_columns["effective_direction"], "input")
    book.save(str(form))
    routed = sh(command, BASE)
    require(routed.returncode == 0, "inout input route_to_30 generation failed:\n%s\n%s" % (routed.stdout, routed.stderr))
    output_path = root / "04_result" / "04_soc_io_pads_prects_ss_125.sdc"
    output_text = output_path.read_text(encoding="utf-8")
    require(output_text.count("set_input_transition") == 2, "inout input view did not emit one command per exact top bit")
    require("[get_ports {pad_gpio[3]}]" in output_text and "[get_ports {pad_gpio[2]}]" in output_text, "inout input view exact top-pad mapping is wrong")
    inventory_path = root / "04_middle" / "pad_inventory.csv"
    with inventory_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        routed_rows = list(csv.DictReader(file_obj))
    require({row["pad_disposition"] for row in routed_rows} == {"route_to_30"}, "inout input route was not published")
    require({row["effective_direction"] for row in routed_rows} == {"input"}, "inout input route lost effective direction")
    require(all(row["src_instance"] == "top" and row["dst_instance"] == "u_io" for row in routed_rows), "inout input route published reversed endpoints")
    used_book = load_workbook(str(inputs / "port_u_io.xlsx"), read_only=True, data_only=True)
    used_sheet = used_book["u_io"]
    used_columns = header_map(used_sheet)
    require(str(used_sheet.cell(2, used_columns["Inout Name"]).value or "") == "", "inout route_to_30 bits were incorrectly accounted")
    used_book.close()

    book = load_workbook(str(form))
    pad_sheet = book["pad_inventory"]
    pad_columns = header_map(pad_sheet)
    for route_row in route_rows:
        for name, value in (
            ("pad_disposition", ""), ("effective_direction", ""),
            ("apply", "no"), ("review_status", "pending"),
            ("approved_machine_digest", ""),
        ):
            pad_sheet.cell(route_row, pad_columns[name], value)
    book.save(str(form))
    constrained = sh(command, BASE)
    require(constrained.returncode == 0, "inout input-direction generation failed:\n%s\n%s" % (constrained.stdout, constrained.stderr))
    used_book = load_workbook(str(inputs / "port_u_io.xlsx"), read_only=True, data_only=True)
    used_sheet = used_book["u_io"]
    used_columns = header_map(used_sheet)
    require(str(used_sheet.cell(2, used_columns["Inout Name"]).value or "") == "0,1", "inout exact bits were not accounted in Inout Name")
    used_book.close()

    rerun = sh(command, BASE)
    if rerun.returncode == 1:
        rerun = sh(command, BASE)
    require(rerun.returncode == 0, "inout accounting state was mistaken for top connectivity on rerun")
    rerun_text = output_path.read_text(encoding="utf-8")
    require("[get_ports {pad_gpio[3]}]" in rerun_text and "[get_ports {pad_gpio[2]}]" in rerun_text and "[get_ports {0,1}]" not in rerun_text, "inout rerun used accounting state as a top pad name")


def run_flat_route_multiview_case():
    root = WORK / "flat_route_multiview"
    clean_dir(root)
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_info_all(inputs / "info_all.xlsx")
    write_ports(inputs / "port_u_io.xlsx")
    wb = load_workbook(str(inputs / "port_u_io.xlsx"))
    ws = wb["u_io"]
    columns = header_map(ws)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, columns["Input Used Width"], "")
    ws.delete_rows(4, 1)
    wb.save(str(inputs / "port_u_io.xlsx"))
    (inputs / "u_io.sdc").write_text(
        "set_input_delay -clock [get_clocks dqs_clk] -max 1.25 [get_ports {dq_i[0]}]\n"
        "set_input_transition 0.08 [get_ports {pad_fp[0]}]\n",
        encoding="utf-8",
    )
    with (inputs / "run_context.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["run_id", "mode_label", "design_revision", "note"])
        writer.writerow(["RUN_04_ROUTE_MULTI", "func", "rev_route", "04 route and multi-view regression"])
    with (inputs / "required_views.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"])
        writer.writerow(["prects_ss_125", "prects", "ss_125", "no", "yes", "no", "no", "04 first view"])
        writer.writerow(["postcts_ss_125", "postcts", "ss_125", "no", "yes", "no", "no", "04 second view"])
    with (inputs / "virtual_clocks.csv").open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(["clock_name", "period", "waveform", "note"])
        writer.writerow(["dqs_clk", "10", "0 5", "04 IO reference"])

    stage_00 = sh([EX00, "--run-root", root], BASE)
    require(stage_00.returncode == 0, "route/multi 00 setup failed:\n%s\n%s" % (stage_00.stdout, stage_00.stderr))
    stage_01 = sh([EX01, "--run-root", root], BASE)
    require(stage_01.returncode == 0, "route/multi 01 setup failed:\n%s\n%s" % (stage_01.stdout, stage_01.stderr))

    def command(stage):
        return [EX04, "--run-root", root, "--stage", stage, "--corner", "ss_125"]

    def route_pad(view_id):
        form = root / "04_middle" / "04_soc_io_pads.xlsx"
        book = load_workbook(str(form))
        pad_sheet = book["pad_inventory"]
        pad_columns = header_map(pad_sheet)
        matched = False
        for row in range(2, pad_sheet.max_row + 1):
            if (
                pad_sheet.cell(row, pad_columns["pad_name"]).value == "pad_fp[0]"
                and pad_sheet.cell(row, pad_columns["view_id"]).value == view_id
            ):
                matched = True
                values = {
                    "pad_disposition": "route_to_30",
                    "apply": "yes",
                    "review_status": "approved",
                    "owner": "cdc_owner",
                    "basis": "exact async pad path is owned by stage 30",
                    "related_exception_intent": "false_path for pad_fp[0] to u_io/pad_fp[0]",
                    "reviewer": "04_regression",
                    "review_date": "2026-07-20",
                    "approved_machine_digest": pad_sheet.cell(row, pad_columns["machine_digest"]).value,
                }
                for name, value in values.items():
                    pad_sheet.cell(row, pad_columns[name], value)
        require(matched, "route pad row missing for %s" % view_id)
        book.save(str(form))

    first = sh(command("prects"), BASE)
    require(first.returncode == 1, "route/multi prects first run should sync")
    form = root / "04_middle" / "04_soc_io_pads.xlsx"
    book = load_workbook(str(form))
    io_sheet = book["io_constraints"]
    io_columns = header_map(io_sheet)
    approved_delay = False
    approved_electrical = False
    for row in range(2, io_sheet.max_row + 1):
        pad_name = io_sheet.cell(row, io_columns["pad_name"]).value
        constraint_type = io_sheet.cell(row, io_columns["constraint_type"]).value
        if pad_name == "pad_dq[0]" and constraint_type == "input_delay":
            approved_delay = True
            io_sheet.cell(row, io_columns["apply"], "yes")
            io_sheet.cell(row, io_columns["review_status"], "approved")
            io_sheet.cell(row, io_columns["timing_class"], "timed")
            io_sheet.cell(row, io_columns["owner"], "io_owner")
            io_sheet.cell(row, io_columns["basis"], "reviewed input budget")
        elif pad_name == "pad_fp[0]" and constraint_type == "input_transition":
            approved_electrical = True
            io_sheet.cell(row, io_columns["apply"], "yes")
            io_sheet.cell(row, io_columns["review_status"], "approved")
            io_sheet.cell(row, io_columns["timing_class"], "async")
            io_sheet.cell(row, io_columns["owner"], "io_owner")
            io_sheet.cell(row, io_columns["basis"], "04 retains the pad electrical environment")
    require(approved_delay, "route/multi input_delay candidate missing")
    require(approved_electrical, "route/multi input_transition candidate missing")
    book.save(str(form))
    route_pad("prects_ss_125")

    prects = sh(command("prects"), BASE)
    require(prects.returncode == 0, "route/multi prects generation failed:\n%s\n%s" % (prects.stdout, prects.stderr))
    prects_sdc = (root / "04_result" / "04_soc_io_pads_prects_ss_125.sdc").read_text(encoding="utf-8")
    require("set_input_delay" in prects_sdc and "pad_dq[0]" in prects_sdc, "prects did not emit assembled all/all IO timing")
    require("set_input_transition 0.08 [get_ports {pad_fp[0]}]" in prects_sdc, "route_to_30 pad lost its 04 electrical constraint")
    require("set_false_path" not in prects_sdc, "route_to_30 exception leaked into prects SDC")
    inventory_path = root / "04_middle" / "pad_inventory.csv"
    with inventory_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        prects_rows = list(csv.DictReader(file_obj))
    prects_route = next(
        row for row in prects_rows
        if row.get("pad_name") == "pad_fp[0]" and row.get("view_id") == "prects_ss_125"
    )
    pad_id = prects_route["pad_id"]
    require(re.fullmatch(r"[0-9a-f]{64}", pad_id), "route pad_id is not full SHA-256")
    require(prects_route["connection_id"] == "CONN_" + pad_id, "route connection_id is not tied to pad_id")
    require(prects_route["src_endpoint"] == "[get_ports {pad_fp[0]}]", "route source endpoint is not exact")
    require(prects_route["dst_endpoint"] == "[get_pins {u_io/pad_fp[0]}]", "route destination endpoint is not exact")
    consumer_matches = [
        row for row in prects_rows
        if row.get("connection_id") == prects_route["connection_id"]
        and row.get("apply") == "yes"
        and row.get("review_status") == "approved"
        and row.get("pad_disposition") == "route_to_30"
    ]
    require(len(consumer_matches) == 1, "30-style route consumer did not find one approved 04 owner")
    related_04_pad_id = consumer_matches[0]["pad_id"]
    exception_command = "set_false_path -from %s -to %s" % (
        consumer_matches[0]["src_endpoint"], consumer_matches[0]["dst_endpoint"],
    )
    require(related_04_pad_id == pad_id, "30-style related_04_pad_id does not match canonical pad_id")
    require(
        exception_command == "set_false_path -from [get_ports {pad_fp[0]}] -to [get_pins {u_io/pad_fp[0]}]",
        "30-style exact false_path could not be emitted from 04 inventory",
    )

    used_book = load_workbook(str(inputs / "port_u_io.xlsx"), read_only=True, data_only=True)
    used_sheet = used_book["u_io"]
    used_columns = header_map(used_sheet)
    route_used = next(
        str(used_sheet.cell(row, used_columns["Input Used Width"]).value or "")
        for row in range(2, used_sheet.max_row + 1)
        if used_sheet.cell(row, used_columns["Input"]).value == "pad_fp[0]"
    )
    used_book.close()
    require(route_used == "", "route_to_30 pad was incorrectly accounted by 04 despite electrical emission")
    delta_text = (root / "04_middle" / "port_accounting_delta.csv").read_text(encoding="utf-8")
    require(pad_id not in delta_text and "pad_fp[0]" not in delta_text, "route_to_30 pad leaked into 04 accounting delta")

    prects_completion = root / "04_middle" / "completion" / "prects_ss_125.meta"
    prects_digest_before = hashlib.sha256(prects_completion.read_bytes()).hexdigest()
    postcts_first = sh(command("postcts"), BASE)
    require(postcts_first.returncode == 1, "route/multi postcts first run should sync")
    route_pad("postcts_ss_125")
    postcts = sh(command("postcts"), BASE)
    require(postcts.returncode == 0, "route/multi postcts generation failed:\n%s\n%s" % (postcts.stdout, postcts.stderr))
    postcts_sdc = (root / "04_result" / "04_soc_io_pads_postcts_ss_125.sdc").read_text(encoding="utf-8")
    require("set_input_delay" in postcts_sdc and "pad_dq[0]" in postcts_sdc, "postcts did not emit assembled all/all IO timing")
    require("set_input_transition 0.08 [get_ports {pad_fp[0]}]" in postcts_sdc, "postcts route pad lost its 04 electrical constraint")
    require("set_false_path" not in postcts_sdc, "route_to_30 exception leaked into postcts SDC")

    book = load_workbook(str(form))
    pad_sheet = book["pad_inventory"]
    pad_columns = header_map(pad_sheet)
    source_row = next(
        row for row in range(2, pad_sheet.max_row + 1)
        if pad_sheet.cell(row, pad_columns["pad_name"]).value == "pad_fp[0]"
        and pad_sheet.cell(row, pad_columns["view_id"]).value == "postcts_ss_125"
    )
    stale_row = pad_sheet.max_row + 1
    for column in range(1, pad_sheet.max_column + 1):
        pad_sheet.cell(stale_row, column, pad_sheet.cell(source_row, column).value)
    for name in ("pad_name", "soc_top_port", "top_port"):
        pad_sheet.cell(stale_row, pad_columns[name], "obsolete_pad")
    pad_sheet.cell(stale_row, pad_columns["pad_id"], "f" * 64)
    pad_sheet.cell(stale_row, pad_columns["connection_id"], "CONN_" + "f" * 64)
    book.save(str(form))
    stale_sync = sh(command("postcts"), BASE)
    require(stale_sync.returncode == 1, "obsolete current-view row did not trigger sync")
    settled = sh(command("postcts"), BASE)
    require(settled.returncode == 0, "postcts did not settle after stale row cleanup")

    with inventory_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        final_rows = list(csv.DictReader(file_obj))
    route_rows = [row for row in final_rows if row.get("pad_name") == "pad_fp[0]"]
    require({row["view_id"] for row in route_rows} == {"prects_ss_125", "postcts_ss_125"}, "multi-view route rows missing")
    require({row["pad_id"] for row in route_rows} == {pad_id}, "pad_id changed across views")
    require(not any(row.get("pad_name") == "obsolete_pad" for row in final_rows), "obsolete current-view row was published")

    inventory_digest = hashlib.sha256(inventory_path.read_bytes()).hexdigest()
    postcts_completion = root / "04_middle" / "completion" / "postcts_ss_125.meta"
    stage_completion = root / "04_middle" / "stage_completion.meta"
    prects_meta = json.loads(prects_completion.read_text(encoding="utf-8"))
    postcts_meta = json.loads(postcts_completion.read_text(encoding="utf-8"))
    stage_meta = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(hashlib.sha256(prects_completion.read_bytes()).hexdigest() != prects_digest_before, "prior view completion was not refreshed")
    require(prects_meta["pad_inventory_digest"] == inventory_digest, "prects completion does not authenticate final inventory")
    require(postcts_meta["pad_inventory_digest"] == inventory_digest, "postcts completion does not authenticate final inventory")
    require(stage_meta["pad_inventory_digest"] == inventory_digest, "run-wide completion does not authenticate final inventory")
    required_digests = stage_meta["required_view_completions"]
    require(set(required_digests) == {"prects_ss_125", "postcts_ss_125"}, "run-wide required view set is wrong")
    require(required_digests["prects_ss_125"] == hashlib.sha256(prects_completion.read_bytes()).hexdigest(), "prects completion digest is stale")
    require(required_digests["postcts_ss_125"] == hashlib.sha256(postcts_completion.read_bytes()).hexdigest(), "postcts completion digest is stale")

    stale_prects = dict(prects_meta)
    stale_prects["structure_digest"] = "0" * 64
    prects_completion.write_text(json.dumps(stale_prects, indent=2) + "\n", encoding="utf-8")
    recover_postcts = sh(command("postcts"), BASE)
    require(recover_postcts.returncode == 0, "stale prior structure blocked current-view recovery")
    partial_stage = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(partial_stage["completion_status"] == "review_required", "stale prior view did not make run-wide completion incomplete")
    require(set(partial_stage["required_view_completions"]) == {"postcts_ss_125"}, "stale prior view was incorrectly retained")
    recover_prects = sh(command("prects"), BASE)
    if recover_prects.returncode == 1:
        recover_prects = sh(command("prects"), BASE)
    require(recover_prects.returncode == 0, "stale prects view could not be rebuilt")
    recovered_stage = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(recovered_stage["completion_status"] == "complete", "multi-view completion did not recover after rebuilding stale view")
    require(set(recovered_stage["required_view_completions"]) == {"prects_ss_125", "postcts_ss_125"}, "recovered required view set is incomplete")

    official_inventory = inventory_path.read_bytes()
    official_stage_digest = hashlib.sha256(stage_completion.read_bytes()).hexdigest()
    inventory_meta_path = root / "04_middle" / "pad_inventory.meta"
    official_meta_digest = hashlib.sha256(inventory_meta_path.read_bytes()).hexdigest()
    tampered_inventory = official_inventory.replace(b"pad_fp[0]", b"pad_fx[0]", 1)
    require(tampered_inventory != official_inventory, "failed to construct prior inventory tamper")
    inventory_path.write_bytes(tampered_inventory)
    tampered_run = sh(command("postcts"), BASE)
    require(tampered_run.returncode == 1, "tampered prior pad inventory was re-signed")
    require(hashlib.sha256(stage_completion.read_bytes()).hexdigest() == official_stage_digest, "tampered prior inventory changed run-wide completion")
    require(hashlib.sha256(inventory_meta_path.read_bytes()).hexdigest() == official_meta_digest, "tampered prior inventory changed inventory metadata")
    inventory_path.write_bytes(official_inventory)
    restored_run = sh(command("postcts"), BASE)
    require(restored_run.returncode == 0, "restored prior pad inventory did not recover")

    legacy_meta = json.loads(inventory_meta_path.read_text(encoding="utf-8"))
    legacy_meta.pop("view_ids", None)
    inventory_meta_path.write_text(json.dumps(legacy_meta, indent=2) + "\n", encoding="utf-8")
    legacy_postcts = sh(command("postcts"), BASE)
    require(legacy_postcts.returncode == 0, "non-canonical prior inventory blocked current-view rebuild")
    legacy_partial = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(legacy_partial["completion_status"] == "review_required", "non-canonical prior inventory was silently certified")
    require(set(legacy_partial["required_view_completions"]) == {"postcts_ss_125"}, "non-canonical prior view completion was retained")
    legacy_prects = sh(command("prects"), BASE)
    if legacy_prects.returncode == 1:
        legacy_prects = sh(command("prects"), BASE)
    require(legacy_prects.returncode == 0, "prior view could not rebuild after canonical inventory migration")
    legacy_recovered = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(legacy_recovered["completion_status"] == "complete", "canonical inventory migration did not recover all views")

    clock_inventory_path = root / "01_middle" / "clock_inventory.csv"
    official_clock_inventory = clock_inventory_path.read_bytes()
    clock_inventory_path.write_bytes(official_clock_inventory + b"\r\n")
    current_clock_digest = hashlib.sha256(clock_inventory_path.read_bytes()).hexdigest()
    stage_01_completion_path = root / "01_middle" / "stage_completion.meta"
    stage_01_completion = json.loads(stage_01_completion_path.read_text(encoding="utf-8"))
    stage_01_completion["clock_inventory_digest"] = current_clock_digest
    stage_01_completion_path.write_text(json.dumps(stage_01_completion, indent=2) + "\n", encoding="utf-8")
    current_stage_01_digest = hashlib.sha256(stage_01_completion_path.read_bytes()).hexdigest()
    upstream_changed = sh(command("postcts"), BASE)
    require(upstream_changed.returncode == 0, "current view could not rebuild after upstream digest change")
    upstream_partial = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(upstream_partial["completion_status"] == "review_required", "stale upstream digest did not invalidate prior view")
    require(set(upstream_partial["required_view_completions"]) == {"postcts_ss_125"}, "prior view with stale upstream digest was retained")
    upstream_restored = sh(command("prects"), BASE)
    if upstream_restored.returncode == 1:
        upstream_restored = sh(command("prects"), BASE)
    require(upstream_restored.returncode == 0, "stale prior view could not rebuild against the new upstream digest")
    upstream_recovered = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(upstream_recovered["completion_status"] == "complete", "rebuilt upstream digest did not recover all views")
    require(set(upstream_recovered["required_view_completions"]) == {"prects_ss_125", "postcts_ss_125"}, "upstream digest recovery lost a required view")
    for completion_file in (prects_completion, postcts_completion):
        rebuilt_completion = json.loads(completion_file.read_text(encoding="utf-8"))
        require(rebuilt_completion["upstream_artifact_digests"]["01_clock_inventory"] == current_clock_digest, "view completion retained a stale 01 clock inventory digest")
        require(rebuilt_completion["upstream_artifact_digests"]["01_stage_completion"] == current_stage_01_digest, "view completion retained a stale 01 stage completion digest")

    stage_00_completion_path = root / "00_middle" / "stage_completion.meta"
    stage_00_completion_path.write_bytes(stage_00_completion_path.read_bytes() + b"\n")
    stage_00_changed = sh(command("postcts"), BASE)
    require(stage_00_changed.returncode == 0, "current view could not rebuild after 00 completion digest change")
    stage_00_partial = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(stage_00_partial["completion_status"] == "review_required", "stale 00 completion digest did not invalidate prior view")
    require(set(stage_00_partial["required_view_completions"]) == {"postcts_ss_125"}, "prior view with stale 00 completion digest was retained")
    stage_00_rebuilt = sh(command("prects"), BASE)
    if stage_00_rebuilt.returncode == 1:
        stage_00_rebuilt = sh(command("prects"), BASE)
    require(stage_00_rebuilt.returncode == 0, "stale prior view could not rebuild against the new 00 completion digest")
    stage_00_recovered = json.loads(stage_completion.read_text(encoding="utf-8"))
    require(stage_00_recovered["completion_status"] == "complete", "rebuilt 00 completion digest did not recover all views")
    current_stage_00_digest = hashlib.sha256(stage_00_completion_path.read_bytes()).hexdigest()
    for completion_file in (prects_completion, postcts_completion):
        rebuilt_completion = json.loads(completion_file.read_text(encoding="utf-8"))
        require(rebuilt_completion["upstream_artifact_digests"]["00_stage_completion"] == current_stage_00_digest, "view completion retained a stale 00 completion digest")

    book = load_workbook(str(form))
    pad_sheet = book["pad_inventory"]
    pad_columns = header_map(pad_sheet)
    prects_route_row = next(
        row for row in range(2, pad_sheet.max_row + 1)
        if pad_sheet.cell(row, pad_columns["pad_name"]).value == "pad_fp[0]"
        and pad_sheet.cell(row, pad_columns["view_id"]).value == "prects_ss_125"
    )
    for name, value in (
        ("pad_disposition", ""), ("apply", "no"),
        ("review_status", "pending"), ("approved_machine_digest", ""),
    ):
        pad_sheet.cell(prects_route_row, pad_columns[name], value)
    book.save(str(form))
    constrained_run = sh(command("prects"), BASE)
    require(constrained_run.returncode == 0, "electrical pad could not enter constrained 04 ownership")
    mixed_view_run = sh(command("postcts"), BASE)
    if mixed_view_run.returncode == 1:
        mixed_view_report = (root / "04_result" / "reports" / "io_pad_check_report_postcts_ss_125.txt").read_text(encoding="utf-8")
        require("already claimed by a prior 04 accounting transaction" not in mixed_view_report, "mixed-view route was rejected despite another required view retaining 04 ownership")
        mixed_view_run = sh(command("postcts"), BASE)
    require(mixed_view_run.returncode == 0, "mixed-view constrained/route ownership depended on execution order")
    route_pad("prects_ss_125")
    transfer_run = sh(command("prects"), BASE)
    if transfer_run.returncode == 1:
        transfer_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
        if "already claimed by a prior 04 accounting transaction" not in transfer_report:
            transfer_run = sh(command("prects"), BASE)
    require(transfer_run.returncode == 1, "constrained-to-route ownership transfer was accepted in-place")
    transfer_report = (root / "04_result" / "reports" / "io_pad_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("already claimed by a prior 04 accounting transaction" in transfer_report, "constrained-to-route fresh-run diagnostic missing")


def main():
    clean_dir(WORK)
    run_bit_pending_case()
    run_empty_command_error_case()
    run_noncanonical_port_error_case()
    run_target_complete_case()
    run_target_partial_strict_case()
    run_target_scenario_clock_case()
    run_target_approved_na_case()
    run_target_delay_conflict_case()
    run_target_connection_contract_case()
    run_flat_target_case()
    run_flat_empty_pad_multiview_case()
    run_flat_fanout_case()
    run_flat_vector_expansion_case()
    run_flat_inout_direction_case()
    run_flat_route_multiview_case()
    print("04 complex regression: PASS")
    print("  legacy cases: bit_pending, empty_command, noncanonical_port")
    print("  target cases: flat direct, empty/multi-view, exact fanout/vector, inout direction, canonical route owner, multi-view inventory, partial+strict, scenario, approved NA, conflict, contract gates")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
