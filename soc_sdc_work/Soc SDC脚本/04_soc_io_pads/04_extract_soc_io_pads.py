#!/usr/bin/env python3
"""
Generate 04 SoC IO/pad SDC and review reports from integration spreadsheets,
lower-level iobuffer/module SDC files, and a reviewed IO constraint workbook.

Current scope:
  * --run-root uses fixed inputs/00_middle/01_middle/04_middle/04_result paths
  * legacy cwd mode remains available for existing flows
  * first run creates/synchronizes 04_soc_io_pads.xlsx, then stops for review
  * only apply=yes + review_status=approved rows are emitted
  * scenario/stage/corner-specific generation is supported
"""

import argparse
import csv
import hashlib
import importlib.util
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

try:
    from dataclasses import dataclass, field
except ImportError:  # pragma: no cover - Python 3.6 compatibility path
    class _CompatField:
        def __init__(self, default_factory=None):
            self.default_factory = default_factory

    def field(default_factory=None):
        return _CompatField(default_factory=default_factory)

    def dataclass(_cls=None, frozen=False):
        def wrap(cls):
            annotations = getattr(cls, "__annotations__", {})
            names = list(annotations.keys())
            defaults = {}
            factories = {}
            for name in names:
                if hasattr(cls, name):
                    value = getattr(cls, name)
                    if isinstance(value, _CompatField):
                        factories[name] = value.default_factory
                        delattr(cls, name)
                    else:
                        defaults[name] = value

            def __init__(self, *args, **kwargs):
                if len(args) > len(names):
                    raise TypeError("too many positional arguments")
                values = {}
                for name, value in zip(names, args):
                    values[name] = value
                for name in names[len(args):]:
                    if name in kwargs:
                        values[name] = kwargs.pop(name)
                    elif name in factories:
                        values[name] = factories[name]()
                    elif name in defaults:
                        values[name] = defaults[name]
                    else:
                        raise TypeError("missing required argument: " + name)
                if kwargs:
                    raise TypeError("unexpected argument: " + sorted(kwargs)[0])
                for name in names:
                    object.__setattr__(self, name, values[name])

            def __eq__(self, other):
                if other.__class__ is not cls:
                    return False
                return all(getattr(self, name) == getattr(other, name) for name in names)

            cls.__init__ = __init__
            cls.__eq__ = __eq__
            if frozen:
                def __hash__(self):
                    return hash(tuple(getattr(self, name) for name in names))
                cls.__hash__ = __hash__
            else:
                cls.__hash__ = None
            return cls

        if _cls is None:
            return wrap
        return wrap(_cls)

try:
    import pandas as pd
except ImportError as exc:  # pragma: no cover - user environment guard
    print("ERROR: pandas is required to read integration xlsx files.", file=sys.stderr)
    raise SystemExit(2) from exc

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - user environment guard
    print("ERROR: openpyxl is required to read/write 04 IO pad xlsx files.", file=sys.stderr)
    raise SystemExit(2) from exc


SCENARIOS = {"common", "func", "scan", "mbist", "gpio_in", "gpio_out"}
STAGES = {"all", "synth", "prects", "postcts", "postroute"}
APPLY_VALUES = {"", "yes", "no"}
REVIEW_STATUS_VALUES = {"", "pending", "approved", "rejected"}
SOURCE_TYPES = {"", "extracted", "manual", "na"}
SOURCE_SDC_STATUS_VALUES = {"", "available", "missing", "not_required"}
TOOLS = {"sta", "synth", "both"}
ACTIVE_01_ACTIONS = {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}
PORT_BIT_RE = re.compile(r"^[^\s\[\]]+(?:\[\d+\])?$")

SUPPORTED_COMMANDS = {
    "set_input_delay": "input_delay",
    "set_output_delay": "output_delay",
    "set_load": "load",
    "set_driving_cell": "driving_cell",
    "set_drive": "drive",
    "set_input_transition": "input_transition",
    "set_false_path": "false_path",
    "set_dont_touch_network": "dont_touch_network",
    "set_max_transition": "max_transition",
    "set_max_capacitance": "max_capacitance",
}

DELAY_TYPES = {"input_delay", "output_delay"}
ELECTRICAL_TYPES = {
    "load",
    "driving_cell",
    "drive",
    "input_transition",
    "max_transition",
    "max_capacitance",
}

IO_HEADERS = [
    "scenario",
    "stage",
    "corner",
    "pad_name",
    "soc_object",
    "subsys_instance",
    "subsys_port",
    "direction",
    "timing_class",
    "constraint_type",
    "clock_name",
    "value",
    "min_value",
    "max_value",
    "rise_value",
    "fall_value",
    "delay_edge",
    "delay_polarity",
    "add_delay",
    "drive_lib_cell",
    "drive_pin",
    "drive_from_pin",
    "drive_input_transition_rise",
    "drive_input_transition_fall",
    "object_granularity",
    "unit_time",
    "unit_cap",
    "extra_options",
    "source_type",
    "source_sdc_status",
    "source_sdc_file",
    "source_line",
    "source_digest",
    "extraction_time",
    "original_command",
    "original_object",
    "apply",
    "review_status",
    "owner",
    "basis",
    "reviewer",
    "review_date",
    "note",
]

PAD_HEADERS = [
    "schema_version",
    "run_id",
    "mode_label",
    "design_revision",
    "run_completeness",
    "structure_digest",
    "accounting_digest_before",
    "accounting_digest_after",
    "pad_id",
    "view_id",
    "stage",
    "corner",
    "pad_name",
    "soc_top_port",
    "top_port",
    "top_bit_index",
    "top_endpoint",
    "subsys_instance",
    "harden_instance",
    "subsys_port",
    "harden_port",
    "harden_bit_index",
    "harden_endpoint",
    "direction_from_integration",
    "direction",
    "effective_direction",
    "src_instance",
    "src_direction",
    "src_port",
    "src_bit_index",
    "src_soc_object",
    "src_endpoint",
    "dst_instance",
    "dst_direction",
    "dst_port",
    "dst_bit_index",
    "dst_soc_object",
    "dst_endpoint",
    "is_gpio_or_inout",
    "related_scenarios",
    "source_sdc_status",
    "sdc_status",
    "connection_id",
    "connection_status",
    "scenario_scope",
    "connection_type",
    "pad_disposition",
    "timing_active",
    "coverage_status",
    "apply",
    "review_status",
    "owner",
    "basis",
    "related_exception_intent",
    "reviewer",
    "review_date",
    "source_workbook",
    "source_sheet",
    "source_row",
    "source_digest",
    "machine_digest",
    "approved_machine_digest",
    "note",
]

LOG_HEADERS = [
    "source_sdc_file",
    "source_sdc_status",
    "source_line",
    "command_type",
    "original_command",
    "parse_status",
    "mapped_soc_object",
    "source_digest",
    "extraction_time",
    "message",
]

HEADER_FILL = PatternFill("solid", fgColor="215967")
TITLE_FILL = PatternFill("solid", fgColor="335C81")
SUBTITLE_FILL = PatternFill("solid", fgColor="EAF3F6")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C6CC"),
    right=Side(style="thin", color="B8C6CC"),
    top=Side(style="thin", color="B8C6CC"),
    bottom=Side(style="thin", color="B8C6CC"),
)


@dataclass
class PortInfo:
    name: str
    width: str = ""
    used_width: str = ""
    from_whom: str = ""
    to_top: str = ""
    connectivity: str = ""
    inout_name: str = ""


@dataclass
class InstInfo:
    module_name: str
    inst_name: str
    owner: str = ""
    file_path: str = ""
    sdc_hint: str = ""
    sdc_path: Optional[Path] = None
    sdc_status: str = ""
    sdc_note: str = ""
    inputs: Dict[str, PortInfo] = field(default_factory=dict)
    outputs: Dict[str, PortInfo] = field(default_factory=dict)
    inouts: Dict[str, PortInfo] = field(default_factory=dict)


@dataclass
class TclCommand:
    raw: str
    line_no: int


@dataclass
class PadRecord:
    pad_name: str
    soc_top_port: str
    subsys_instance: str
    subsys_port: str
    direction: str
    is_gpio_or_inout: str = "no"
    related_scenarios: str = ""
    source_sdc_status: str = ""
    connection_id: str = ""
    scenario_scope: str = ""
    connection_type: str = ""
    source_workbook: str = ""
    source_sheet: str = ""
    source_row: str = ""
    note: str = ""


@dataclass
class ObjectMapping:
    kind: str
    objects: List[str]
    collection: str
    pad_names: List[str] = field(default_factory=list)
    unresolved: bool = False
    message: str = ""


@dataclass
class ExtractedConstraint:
    values: Dict[str, str]
    parse_status: str
    mapped_soc_object: str
    message: str = ""
    include_in_form: bool = True


@dataclass
class ClockInfo:
    clock_name: str
    direct_source: str = ""
    producer_object: str = ""
    final_action: str = ""
    source_file: str = ""


@dataclass
class HardenSdcManifestEntry:
    scenario: str
    inst_name: str
    module_name: str
    sdc_path: str
    availability_status: str
    note: str
    source_row: int


@dataclass
class RunCompleteness:
    status: str
    available_instances: List[str] = field(default_factory=list)
    missing_instances: List[str] = field(default_factory=list)
    not_required_instances: List[str] = field(default_factory=list)
    manifest_path: str = ""

    @property
    def available_count(self) -> int:
        return len(self.available_instances)

    @property
    def missing_count(self) -> int:
        return len(self.missing_instances)

    @property
    def not_required_count(self) -> int:
        return len(self.not_required_instances)


@dataclass(frozen=True)
class ConnectionEdge:
    connection_id: str
    scenario_scope: str
    connection_type: str
    src_instance: str
    src_direction: str
    src_port: str
    src_endpoint_key: str
    src_soc_object: str
    dst_instance: str
    dst_direction: str
    dst_port: str
    dst_endpoint_key: str
    dst_soc_object: str
    validation_status: str = ""
    owner_hint: str = ""
    source_workbook: str = ""
    source_sheet: str = ""
    source_row: str = ""
    note: str = ""


@dataclass
class FormRow:
    row_idx: int
    values: Dict[str, object]


@dataclass(frozen=True)
class PortKey:
    inst_name: str
    direction: str
    port: str


class Report:
    def __init__(self) -> None:
        self.lines: List[str] = []
        self.warning_count = 0
        self.error_count = 0
        self.sync_changed = False

    def info(self, msg: str) -> None:
        self.lines.append(f"INFO: {msg}")

    def warn(self, msg: str) -> None:
        self.warning_count += 1
        self.lines.append(f"WARNING: {msg}")

    def error(self, msg: str) -> None:
        self.error_count += 1
        self.lines.append(f"ERROR: {msg}")


def _author_part_a() -> str:
    return chr(72) + chr(111)


def _author_part_b() -> str:
    return chr(119) + chr(97)


def _author_part_c() -> str:
    return chr(114) + chr(100)


def author_name() -> str:
    return _author_part_a() + _author_part_b() + _author_part_c()


def atomic_write_text(path: Path, text: str, encoding: str = "utf-8") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.tmp.{os.getpid()}")
    try:
        tmp_path.write_text(text, encoding=encoding)
        os.replace(str(tmp_path), str(path))
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def atomic_save_workbook(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.stem}.tmp.{os.getpid()}{path.suffix}")
    try:
        wb.save(tmp_path)
        os.replace(str(tmp_path), str(path))
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def normalize_key(value) -> str:
    return clean_cell(value).strip().lower()


def normalize_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())


def column_map(df: pd.DataFrame) -> Dict[str, str]:
    return {normalize_col(col): col for col in df.columns}


def get_col(df: pd.DataFrame, aliases: Sequence[str]) -> Optional[str]:
    cmap = column_map(df)
    for alias in aliases:
        key = normalize_col(alias)
        if key in cmap:
            return cmap[key]
    return None


def read_excel_file(path: Path, sheet_name=0):
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except Exception as exc:
        raise RuntimeError(f"failed to read {path}: {exc}") from exc


def safe_filename_token(value: str) -> str:
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-")
    token = "".join(char if char in allowed else "_" for char in clean_cell(value))
    return token or "unknown"


def brace_list(names: Sequence[str]) -> str:
    return "{" + " ".join(clean_cell(name) for name in names if clean_cell(name)) + "}"


def get_collection(kind: str, objects: Sequence[str]) -> str:
    return f"[{kind} {brace_list(objects)}]"


def parse_number(value) -> Optional[float]:
    text = clean_cell(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_number(value) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    return f"{number:.12g}"


def compact_command(text: str, max_len: int = 260) -> str:
    cleaned = re.sub(r"\s+", " ", text).strip()
    if len(cleaned) <= max_len:
        return cleaned
    return cleaned[: max_len - 3] + "..."


def read_info_all(path: Path, report: Report) -> Dict[str, InstInfo]:
    df = read_excel_file(path)
    module_col = get_col(df, ["module_name", "module name", "module"])
    inst_col = get_col(df, ["inst_name", "inst name", "instance", "instance_name"])
    owner_col = get_col(df, ["owner"])
    file_col = get_col(df, ["file_path", "file path", "empty_path", "verilog", "v_path"])
    sdc_col = get_col(df, ["sdc_path", "sdc file", "sdc_file", "sdc"])

    if not inst_col:
        raise RuntimeError(f"{path} must contain an inst_name column")

    instances: Dict[str, InstInfo] = {}
    for row_idx, row in df.iterrows():
        inst_name = clean_cell(row.get(inst_col))
        if not inst_name:
            continue
        module_name = clean_cell(row.get(module_col)) if module_col else ""
        file_path = clean_cell(row.get(file_col)) if file_col else ""
        if not module_name and file_path:
            module_name = Path(file_path).stem.replace("_empty", "")
        if not module_name:
            module_name = inst_name
            report.warn(f"{path.name} row {row_idx + 2}: module_name is empty; using inst_name")
        if inst_name in instances:
            report.warn(f"duplicate inst_name {inst_name} in {path.name}; keeping first row")
            continue
        instances[inst_name] = InstInfo(
            module_name=module_name,
            inst_name=inst_name,
            owner=clean_cell(row.get(owner_col)) if owner_col else "",
            file_path=file_path,
            sdc_hint=clean_cell(row.get(sdc_col)) if sdc_col else "",
        )
    report.info(f"loaded {len(instances)} instance(s) from {path.name}")
    return instances


def parse_port_sheet(df: pd.DataFrame) -> Dict[str, Dict[str, PortInfo]]:
    input_col = get_col(df, ["Input"])
    input_width_col = get_col(df, ["Input Width"])
    input_used_col = get_col(df, ["Input Used Width"])
    from_col = get_col(df, ["From Whom"])

    output_col = get_col(df, ["Output"])
    output_width_col = get_col(df, ["Output Width"])
    output_used_col = get_col(df, ["Output Used Width"])
    to_top_col = get_col(df, ["To Top"])

    inout_col = get_col(df, ["Inout"])
    inout_width_col = get_col(df, ["Inout Width"])
    inout_conn_col = get_col(df, ["Inout Connectivity"])
    inout_name_col = get_col(df, ["Inout Name"])

    inputs: Dict[str, PortInfo] = {}
    outputs: Dict[str, PortInfo] = {}
    inouts: Dict[str, PortInfo] = {}

    for _, row in df.iterrows():
        if input_col:
            name = clean_cell(row.get(input_col))
            if name:
                inputs[name] = PortInfo(
                    name=name,
                    width=clean_cell(row.get(input_width_col)) if input_width_col else "",
                    used_width=clean_cell(row.get(input_used_col)) if input_used_col else "",
                    from_whom=clean_cell(row.get(from_col)) if from_col else "",
                )
        if output_col:
            name = clean_cell(row.get(output_col))
            if name:
                outputs[name] = PortInfo(
                    name=name,
                    width=clean_cell(row.get(output_width_col)) if output_width_col else "",
                    used_width=clean_cell(row.get(output_used_col)) if output_used_col else "",
                    to_top=clean_cell(row.get(to_top_col)) if to_top_col else "",
                )
        if inout_col:
            name = clean_cell(row.get(inout_col))
            if name:
                inouts[name] = PortInfo(
                    name=name,
                    width=clean_cell(row.get(inout_width_col)) if inout_width_col else "",
                    connectivity=clean_cell(row.get(inout_conn_col)) if inout_conn_col else "",
                    inout_name=clean_cell(row.get(inout_name_col)) if inout_name_col else "",
                )
    return {"inputs": inputs, "outputs": outputs, "inouts": inouts}


def read_port_workbooks(paths: Sequence[Path], report: Report) -> Dict[str, Dict[str, Dict[str, PortInfo]]]:
    sheets: Dict[str, Dict[str, Dict[str, PortInfo]]] = {}
    for path in paths:
        try:
            book = pd.ExcelFile(path)
        except Exception as exc:
            report.error(f"failed to open port workbook {path.name}: {exc}")
            continue
        for sheet_name in book.sheet_names:
            if sheet_name in sheets:
                report.warn(f"duplicate port sheet {sheet_name}; keeping first occurrence")
                continue
            try:
                sheets[sheet_name] = parse_port_sheet(read_excel_file(path, sheet_name))
            except Exception as exc:
                report.error(f"failed to read {path.name}:{sheet_name}: {exc}")
    report.info(f"loaded {len(sheets)} instance port sheet(s) from {len(paths)} workbook(s)")
    return sheets


def default_port_workbooks(cwd: Path, info_name: str, form_name: str, report: Report) -> List[Path]:
    excluded = {info_name, form_name}
    candidates: List[Path] = []
    skipped: List[str] = []
    for path in sorted(cwd.glob("*.xlsx")):
        if path.name in excluded or path.name.startswith("~$"):
            continue
        if path.name.startswith(("port_", "ports_")):
            candidates.append(path)
        else:
            skipped.append(path.name)
    if skipped:
        report.warn(
            "ignored non-port workbook(s) in 04 input directory: " + ", ".join(skipped[:10])
        )
    return candidates


def attach_port_data(instances: Dict[str, InstInfo], sheets: Dict[str, Dict[str, Dict[str, PortInfo]]], report: Report) -> None:
    for inst in instances.values():
        data = sheets.get(inst.inst_name)
        if not data:
            report.warn(f"no port sheet found for instance {inst.inst_name}")
            continue
        inst.inputs = data["inputs"]
        inst.outputs = data["outputs"]
        inst.inouts = data["inouts"]


def resolve_manifest_path(run_root: Path, value: str) -> Path:
    path = Path(value).expanduser()
    return path if path.is_absolute() else run_root / path


def apply_harden_sdc_manifest(
    instances: Dict[str, InstInfo],
    manifest_path: Path,
    run_root: Path,
    scenario: str,
    require_complete: bool,
    report: Report,
    create_instances: bool = False,
) -> RunCompleteness:
    error_count_before = report.error_count
    if not manifest_path.is_file():
        report.error(f"{manifest_path}: HARDEN_SDC_MANIFEST_MISSING: required target runtime manifest is absent")
        return RunCompleteness(status="invalid", manifest_path=str(manifest_path.resolve()))

    required_fields = {"scenario", "inst_name", "module_name", "availability_status"}
    entries: Dict[str, HardenSdcManifestEntry] = {}
    with manifest_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = set(reader.fieldnames or [])
        missing_fields = sorted(required_fields - fields)
        if missing_fields:
            report.error(
                f"{manifest_path}: HARDEN_SDC_MANIFEST_SCHEMA_ERROR: missing field(s): {', '.join(missing_fields)}"
            )
        if "sdc_path" not in fields and "resolved_sdc_path" not in fields:
            report.error(f"{manifest_path}: HARDEN_SDC_MANIFEST_SCHEMA_ERROR: missing sdc_path")
        for row_idx, row in enumerate(reader, start=2):
            row_scenario = clean_cell(row.get("scenario"))
            inst_name = clean_cell(row.get("inst_name"))
            if not inst_name:
                report.error(f"{manifest_path.name} row {row_idx}: HARDEN_SDC_MANIFEST_EMPTY_INSTANCE")
                continue
            if row_scenario != scenario:
                report.error(
                    f"{manifest_path.name} row {row_idx}: HARDEN_SDC_MANIFEST_SCENARIO_MISMATCH: "
                    f"row scenario={row_scenario or '<empty>'}, requested={scenario}"
                )
                continue
            if inst_name in entries:
                report.error(
                    f"{manifest_path.name} row {row_idx}: HARDEN_SDC_MANIFEST_DUPLICATE_INSTANCE: {inst_name}"
                )
                continue
            entries[inst_name] = HardenSdcManifestEntry(
                scenario=row_scenario,
                inst_name=inst_name,
                module_name=clean_cell(row.get("module_name")),
                sdc_path=clean_cell(row.get("sdc_path")) or clean_cell(row.get("resolved_sdc_path")),
                availability_status=normalize_key(row.get("availability_status")),
                note=clean_cell(row.get("note")),
                source_row=row_idx,
            )

    if create_instances:
        for inst_name, entry in sorted(entries.items()):
            if not entry.module_name:
                report.error(
                    f"{manifest_path.name} row {entry.source_row}: "
                    f"HARDEN_SDC_MANIFEST_MODULE_EMPTY: {inst_name}"
                )
            instances[inst_name] = InstInfo(
                module_name=entry.module_name or inst_name,
                inst_name=inst_name,
            )
    else:
        for inst_name in sorted(set(entries) - set(instances)):
            report.error(
                f"{manifest_path.name} row {entries[inst_name].source_row}: "
                f"HARDEN_SDC_MANIFEST_ORPHAN_INSTANCE: {inst_name}"
            )

    available: List[str] = []
    missing: List[str] = []
    not_required: List[str] = []
    for inst_name in sorted(instances):
        inst = instances[inst_name]
        entry = entries.get(inst_name)
        if entry is None:
            report.error(f"{manifest_path.name}: HARDEN_SDC_MANIFEST_INSTANCE_MISSING: {inst_name}")
            inst.sdc_status = "missing"
            missing.append(inst_name)
            continue
        if entry.module_name and entry.module_name != inst.module_name:
            report.error(
                f"{manifest_path.name} row {entry.source_row}: HARDEN_SDC_MANIFEST_MODULE_MISMATCH: "
                f"{inst_name} workbook={inst.module_name} manifest={entry.module_name}"
            )
        inst.sdc_status = entry.availability_status
        inst.sdc_note = entry.note
        if entry.availability_status == "available":
            if not entry.sdc_path:
                report.error(
                    f"{manifest_path.name} row {entry.source_row}: HARDEN_SDC_MANIFEST_AVAILABLE_PATH_EMPTY: {inst_name}"
                )
                continue
            sdc_path = resolve_manifest_path(run_root, entry.sdc_path)
            if not sdc_path.is_file():
                report.error(
                    f"{manifest_path.name} row {entry.source_row}: HARDEN_SDC_MANIFEST_AVAILABLE_FILE_MISSING: "
                    f"{inst_name} -> {sdc_path}"
                )
                continue
            inst.sdc_path = sdc_path
            available.append(inst_name)
            report.info(f"manifest selected {inst_name}: status=available path={sdc_path.resolve()}")
        elif entry.availability_status == "missing":
            missing.append(inst_name)
            report.warn(
                f"{manifest_path.name} row {entry.source_row}: HARDEN_SDC_MISSING: "
                f"{inst_name}: {entry.note or '<no note>'}"
            )
        elif entry.availability_status == "not_required":
            not_required.append(inst_name)
            report.info(f"manifest marks {inst_name} as not_required: {entry.note or '<no note>'}")
        else:
            report.error(
                f"{manifest_path.name} row {entry.source_row}: HARDEN_SDC_MANIFEST_STATUS_INVALID: "
                f"{inst_name} status={entry.availability_status or '<empty>'}"
            )

    strict_error = bool(require_complete and missing)
    if strict_error:
        report.error(
            "HARDEN_SDC_COMPLETENESS_REQUIRED: missing harden SDC instance(s): " + ", ".join(missing)
        )
    status = "partial" if missing else "complete"
    manifest_error_count = report.error_count - error_count_before - (1 if strict_error else 0)
    if manifest_error_count:
        status = "invalid"
    return RunCompleteness(
        status=status,
        available_instances=available,
        missing_instances=missing,
        not_required_instances=not_required,
        manifest_path=str(manifest_path.resolve()),
    )


def resolve_sdc_paths(instances: Dict[str, InstInfo], cwd: Path, report: Report) -> RunCompleteness:
    all_sdcs = sorted(path for path in cwd.glob("*.sdc") if path.is_file())
    by_name = {path.name: path for path in all_sdcs}
    by_lower = {path.name.lower(): path for path in all_sdcs}

    for inst in instances.values():
        candidates: List[str] = []
        if inst.sdc_hint:
            candidates.append(Path(inst.sdc_hint).name)
        candidates.append(f"{inst.inst_name}.sdc")
        candidates.append(f"{inst.module_name}.sdc")
        if inst.file_path:
            stem = Path(inst.file_path).stem
            candidates.append(f"{stem}.sdc")
            if stem.endswith("_empty"):
                candidates.append(f"{stem[:-6]}.sdc")

        matches: List[Path] = []
        for name in dict.fromkeys(candidates):
            if not name:
                continue
            if name in by_name:
                matches.append(by_name[name])
            elif name.lower() in by_lower:
                matches.append(by_lower[name.lower()])
        unique = []
        for path in matches:
            if path not in unique:
                unique.append(path)
        if len(unique) == 1:
            inst.sdc_path = unique[0]
            inst.sdc_status = "available"
        elif len(unique) > 1:
            inst.sdc_path = unique[0]
            inst.sdc_status = "available"
            report.warn(
                f"multiple SDC candidates for {inst.inst_name}: "
                f"{', '.join(path.name for path in unique)}; using {inst.sdc_path.name}"
            )
        else:
            inst.sdc_status = "missing"
            report.warn(f"no SDC found for {inst.inst_name}; tried: {', '.join(candidates)}")

    available = sorted(inst.inst_name for inst in instances.values() if inst.sdc_status == "available")
    missing = sorted(inst.inst_name for inst in instances.values() if inst.sdc_status == "missing")
    return RunCompleteness(
        status="partial" if missing else "complete",
        available_instances=available,
        missing_instances=missing,
    )


TARGET_CONNECTION_FIELDS = {
    "schema_version",
    "connection_id",
    "scenario_scope",
    "connection_type",
    "src_instance",
    "src_direction",
    "src_port",
    "src_bit_index",
    "src_endpoint_key",
    "src_soc_object",
    "dst_instance",
    "dst_direction",
    "dst_port",
    "dst_bit_index",
    "dst_endpoint_key",
    "dst_soc_object",
    "fanout_index",
    "range_source_expr",
    "range_sink_expr",
    "bit_pair_order",
    "source_workbook",
    "source_sheet",
    "source_row",
    "validation_status",
    "owner_hint",
    "note",
}
PAD_CONNECTION_TYPES = {"top_pad_to_harden", "harden_to_top_pad", "pad_to_pad"}
CONNECTION_TYPES = {
    "harden_to_harden",
    "fabric_to_harden",
    "harden_to_fabric",
    "top_pad_to_harden",
    "harden_to_top_pad",
    "pad_to_pad",
    "clock_connection",
    "feedthrough_candidate",
    "constant_tie",
    "no_connect",
    "unknown",
}


def scenario_scope_matches(value: str, scenario: str) -> bool:
    scopes = {
        normalize_key(item)
        for item in re.split(r"[,;|\s]+", clean_cell(value))
        if normalize_key(item)
    }
    if not scopes:
        scopes = {"common"}
    return "common" in scopes or normalize_key(scenario) in scopes


def endpoint_key(instance: str, direction: str, port: str) -> str:
    return f"{instance}:{direction}:{port}"


def read_connection_inventory(
    path: Path,
    report: Report,
    scenario: str = "common",
    strict_schema: bool = False,
) -> List[ConnectionEdge]:
    edges: List[ConnectionEdge] = []
    required = {
        "connection_id", "src_instance", "src_direction", "src_port",
        "dst_instance", "dst_direction", "dst_port",
    }
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        fields = set(reader.fieldnames or [])
        missing = sorted((TARGET_CONNECTION_FIELDS if strict_schema else required) - fields)
        if missing:
            report.error(f"{path}: CONNECTION_INVENTORY_SCHEMA_ERROR: missing field(s): {', '.join(missing)}")
            return edges
        seen_ids: Set[str] = set()
        for row_idx, row in enumerate(reader, start=2):
            connection_id = clean_cell(row.get("connection_id"))
            if not connection_id:
                report.error(f"{path.name} row {row_idx}: CONNECTION_ID_EMPTY")
                continue
            if connection_id in seen_ids:
                report.error(f"{path.name} row {row_idx}: CONNECTION_ID_DUPLICATE: {connection_id}")
                continue
            seen_ids.add(connection_id)
            scenario_scope = clean_cell(row.get("scenario_scope")) or "common"
            raw_scopes = [
                normalize_key(item)
                for item in re.split(r"[,;|\s]+", clean_cell(row.get("scenario_scope")))
                if normalize_key(item)
            ]
            if strict_schema and not raw_scopes:
                report.error(f"{path.name} row {row_idx}: CONNECTION_SCENARIO_SCOPE_EMPTY: {connection_id}")
            invalid_scopes = sorted(set(raw_scopes) - SCENARIOS)
            if invalid_scopes:
                report.error(
                    f"{path.name} row {row_idx}: CONNECTION_SCENARIO_SCOPE_INVALID: "
                    f"{connection_id} scope={','.join(invalid_scopes)}"
                )
            if not scenario_scope_matches(scenario_scope, scenario):
                continue
            edge = ConnectionEdge(
                connection_id=connection_id,
                scenario_scope=scenario_scope,
                connection_type=normalize_key(row.get("connection_type")),
                src_instance=clean_cell(row.get("src_instance")),
                src_direction=normalize_key(row.get("src_direction")),
                src_port=clean_cell(row.get("src_port")),
                src_endpoint_key=clean_cell(row.get("src_endpoint_key")),
                src_soc_object=clean_cell(row.get("src_soc_object")),
                dst_instance=clean_cell(row.get("dst_instance")),
                dst_direction=normalize_key(row.get("dst_direction")),
                dst_port=clean_cell(row.get("dst_port")),
                dst_endpoint_key=clean_cell(row.get("dst_endpoint_key")),
                dst_soc_object=clean_cell(row.get("dst_soc_object")),
                validation_status=normalize_key(row.get("validation_status")),
                owner_hint=clean_cell(row.get("owner_hint")),
                source_workbook=clean_cell(row.get("source_workbook")),
                source_sheet=clean_cell(row.get("source_sheet")),
                source_row=clean_cell(row.get("source_row")) or str(row_idx),
                note=clean_cell(row.get("note")),
            )
            if edge.validation_status != "matched":
                report.error(
                    f"{path.name} row {row_idx}: CONNECTION_NOT_MATCHED: "
                    f"{connection_id} status={edge.validation_status or '<empty>'}"
                )
            if edge.connection_type not in CONNECTION_TYPES:
                report.error(
                    f"{path.name} row {row_idx}: CONNECTION_TYPE_INVALID: "
                    f"{connection_id} type={edge.connection_type or '<empty>'}"
                )
            for side, instance, direction, port, key in (
                ("src", edge.src_instance, edge.src_direction, edge.src_port, edge.src_endpoint_key),
                ("dst", edge.dst_instance, edge.dst_direction, edge.dst_port, edge.dst_endpoint_key),
            ):
                if direction not in {"input", "output", "inout"}:
                    report.error(
                        f"{path.name} row {row_idx}: CONNECTION_DIRECTION_INVALID: "
                        f"{connection_id} {side}_direction={direction or '<empty>'}"
                    )
                if not PORT_BIT_RE.fullmatch(port):
                    report.error(
                        f"{path.name} row {row_idx}: CONNECTION_PORT_NONCANONICAL: "
                        f"{connection_id} {side}_port={port or '<empty>'}"
                    )
                expected_key = endpoint_key(instance, direction, port)
                if strict_schema and not key:
                    report.error(
                        f"{path.name} row {row_idx}: CONNECTION_ENDPOINT_KEY_EMPTY: {connection_id} {side}"
                    )
                elif key and key != expected_key:
                    report.error(
                        f"{path.name} row {row_idx}: CONNECTION_ENDPOINT_KEY_MISMATCH: "
                        f"{connection_id} {side}={key}, expected={expected_key}"
                    )
            edges.append(edge)
    report.info(f"loaded {len(edges)} connection edge(s) for scenario={scenario} from {path}")
    return edges


def attach_target_pad_edges(
    instances: Dict[str, InstInfo],
    edges: Sequence[ConnectionEdge],
    report: Report,
) -> List[PadRecord]:
    pads: List[PadRecord] = []
    seen: Set[Tuple[str, str, str, str]] = set()

    for edge in edges:
        if edge.connection_type not in PAD_CONNECTION_TYPES:
            continue
        src_top = normalize_key(edge.src_instance) == "top"
        dst_top = normalize_key(edge.dst_instance) == "top"
        if edge.connection_type == "top_pad_to_harden" and not (src_top and not dst_top):
            report.error(f"{edge.connection_id}: PAD_CONNECTION_TYPE_TOPOLOGY_MISMATCH: top_pad_to_harden")
            continue
        if edge.connection_type == "harden_to_top_pad" and not (dst_top and not src_top):
            report.error(f"{edge.connection_id}: PAD_CONNECTION_TYPE_TOPOLOGY_MISMATCH: harden_to_top_pad")
            continue
        if src_top == dst_top:
            report.warn(
                f"{edge.connection_id}: pad-related edge does not have exactly one top endpoint; "
                "kept out of harden pad inventory"
            )
            continue
        if src_top:
            top_port = edge.src_port
            harden_name = edge.dst_instance
            harden_direction = edge.dst_direction
            harden_port = edge.dst_port
        else:
            top_port = edge.dst_port
            harden_name = edge.src_instance
            harden_direction = edge.src_direction
            harden_port = edge.src_port

        inst = instances.get(harden_name)
        if inst is None:
            report.error(
                f"{edge.connection_id}: PAD_EDGE_INSTANCE_NOT_IN_MANIFEST: {harden_name}"
            )
            continue
        if edge.owner_hint and not inst.owner:
            inst.owner = edge.owner_hint
        port = PortInfo(name=harden_port)
        if harden_direction == "input":
            port.from_whom = f"top.{top_port}"
            existing = inst.inputs.get(harden_port)
            if existing and existing.from_whom != port.from_whom:
                report.error(
                    f"{edge.connection_id}: PAD_PORT_MAPPING_CONFLICT: "
                    f"{harden_name}/{harden_port} maps to both {existing.from_whom} and {port.from_whom}"
                )
                continue
            inst.inputs[harden_port] = port
        elif harden_direction == "output":
            port.to_top = f"top.{top_port}"
            existing = inst.outputs.get(harden_port)
            if existing and existing.to_top != port.to_top:
                report.error(
                    f"{edge.connection_id}: PAD_PORT_MAPPING_CONFLICT: "
                    f"{harden_name}/{harden_port} maps to both {existing.to_top} and {port.to_top}"
                )
                continue
            inst.outputs[harden_port] = port
        elif harden_direction == "inout":
            port.inout_name = top_port
            existing = inst.inouts.get(harden_port)
            if existing and existing.inout_name != port.inout_name:
                report.error(
                    f"{edge.connection_id}: PAD_PORT_MAPPING_CONFLICT: "
                    f"{harden_name}/{harden_port} maps to both {existing.inout_name} and {port.inout_name}"
                )
                continue
            inst.inouts[harden_port] = port
        else:
            continue

        key = (top_port, harden_name, harden_port, harden_direction)
        if key in seen:
            report.error(f"{edge.connection_id}: PAD_EDGE_DUPLICATE_MAPPING: {key}")
            continue
        seen.add(key)
        pads.append(
            PadRecord(
                pad_name=top_port,
                soc_top_port=top_port,
                subsys_instance=harden_name,
                subsys_port=harden_port,
                direction=harden_direction,
                is_gpio_or_inout="yes" if harden_direction == "inout" else "no",
                related_scenarios="gpio_in,gpio_out" if harden_direction == "inout" else "",
                source_sdc_status=inst.sdc_status,
                connection_id=edge.connection_id,
                scenario_scope=edge.scenario_scope,
                connection_type=edge.connection_type,
                source_workbook=edge.source_workbook,
                source_sheet=edge.source_sheet,
                source_row=edge.source_row,
                note=edge.note,
            )
        )
    report.info(f"derived {len(pads)} pad record(s) from 00 connection inventory")
    return pads


def validate_pad_connections(pads: Sequence[PadRecord], edges: Sequence[ConnectionEdge], report: Report) -> None:
    edge_keys = {
        (
            edge.src_instance, edge.src_direction, edge.src_port,
            edge.dst_instance, edge.dst_direction, edge.dst_port,
        )
        for edge in edges
    }
    for pad in pads:
        input_key = ("top", "input", pad.pad_name, pad.subsys_instance, "input", pad.subsys_port)
        output_key = (pad.subsys_instance, "output", pad.subsys_port, "top", "output", pad.pad_name)
        matched = input_key in edge_keys if pad.direction == "input" else output_key in edge_keys
        if pad.direction == "inout":
            matched = input_key in edge_keys or output_key in edge_keys
        if not matched:
            report.error(
                "PAD_CONNECTION_INVENTORY_MISSING: "
                f"{pad.pad_name} <-> {pad.subsys_instance}:{pad.direction}:{pad.subsys_port}"
            )


def read_text(path: Path) -> str:
    last_error: Optional[Exception] = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    assert last_error is not None
    raise last_error


def digest_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def collect_current_sdc_digests(instances: Dict[str, InstInfo]) -> Dict[str, str]:
    digests: Dict[str, str] = {}
    for inst in instances.values():
        if inst.sdc_path and inst.sdc_path.is_file():
            try:
                digests[inst.sdc_path.name] = digest_file(inst.sdc_path)
            except OSError:
                continue
    return digests


def is_escaped(text: str, index: int) -> bool:
    count = 0
    pos = index - 1
    while pos >= 0 and text[pos] == "\\":
        count += 1
        pos -= 1
    return count % 2 == 1


def strip_inline_comment(line: str) -> str:
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(line):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(line, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == "#" and brace_depth == 0 and bracket_depth == 0:
                if idx == 0 or line[idx - 1].isspace() or line[idx - 1] == ";":
                    return line[:idx].rstrip()
    return line.rstrip()


def split_semicolon_commands(text: str) -> List[str]:
    parts: List[str] = []
    start = 0
    quote = False
    brace_depth = 0
    bracket_depth = 0
    for idx, char in enumerate(text):
        if char == "\\":
            continue
        if char == '"' and not is_escaped(text, idx) and brace_depth == 0:
            quote = not quote
        elif not quote:
            if char == "{":
                brace_depth += 1
            elif char == "}" and brace_depth:
                brace_depth -= 1
            elif char == "[":
                bracket_depth += 1
            elif char == "]" and bracket_depth:
                bracket_depth -= 1
            elif char == ";" and brace_depth == 0 and bracket_depth == 0:
                parts.append(text[start:idx])
                start = idx + 1
    parts.append(text[start:])
    return parts


def iter_tcl_commands_with_line(text: str) -> Iterable[TclCommand]:
    buf = ""
    start_line = 0
    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = strip_inline_comment(raw_line)
        if not line.strip():
            continue
        if not buf:
            start_line = line_no
        stripped = line.rstrip()
        continued = stripped.endswith("\\") and not is_escaped(stripped, len(stripped) - 1)
        if continued:
            stripped = stripped[:-1].rstrip()
            buf += stripped + " "
            continue
        buf += stripped
        for cmd in split_semicolon_commands(buf):
            cleaned = cmd.strip().rstrip(";").strip()
            if cleaned:
                yield TclCommand(cleaned, start_line)
        buf = ""
        start_line = 0
    if buf.strip():
        for cmd in split_semicolon_commands(buf):
            cleaned = cmd.strip().rstrip(";").strip()
            if cleaned:
                yield TclCommand(cleaned, start_line)


def find_matching(text: str, start: int, open_char: str, close_char: str) -> int:
    depth = 0
    quote = False
    idx = start
    while idx < len(text):
        char = text[idx]
        if char == "\\":
            idx += 2
            continue
        if char == '"' and open_char != '"' and not is_escaped(text, idx):
            quote = not quote
        elif not quote:
            if char == open_char:
                depth += 1
            elif char == close_char:
                depth -= 1
                if depth == 0:
                    return idx
        idx += 1
    return -1


def tokenize_tcl_words(text: str) -> List[str]:
    tokens: List[str] = []
    idx = 0
    while idx < len(text):
        while idx < len(text) and text[idx].isspace():
            idx += 1
        if idx >= len(text):
            break
        start = idx
        if text[idx] == "{":
            end = find_matching(text, idx, "{", "}")
            if end < 0:
                tokens.append(text[start:])
                break
            tokens.append(text[start : end + 1])
            idx = end + 1
            continue
        if text[idx] == '"':
            idx += 1
            while idx < len(text):
                if text[idx] == "\\":
                    idx += 2
                    continue
                if text[idx] == '"':
                    idx += 1
                    break
                idx += 1
            tokens.append(text[start:idx])
            continue
        pieces: List[str] = []
        while idx < len(text) and not text[idx].isspace():
            char = text[idx]
            if char == "\\":
                pieces.append(text[idx : idx + 2])
                idx += 2
            elif char == "[":
                end = find_matching(text, idx, "[", "]")
                if end < 0:
                    pieces.append(text[idx:])
                    idx = len(text)
                else:
                    pieces.append(text[idx : end + 1])
                    idx = end + 1
            else:
                pieces.append(char)
                idx += 1
        tokens.append("".join(pieces))
    return tokens


def strip_braces(text: str) -> str:
    text = clean_cell(text)
    if len(text) >= 2 and text[0] == "{" and text[-1] == "}":
        return text[1:-1].strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        return text[1:-1].strip()
    return text


def split_object_list(text: str) -> List[str]:
    text = strip_braces(text)
    return [part for part in re.split(r"[\s,;]+", text) if part]


COLLECTION_KINDS = {"get_ports", "get_pins", "get_nets", "get_clocks"}


def parse_collection(token: str) -> Optional[Tuple[str, List[str]]]:
    text = clean_cell(token)
    if not (text.startswith("[") and text.endswith("]")):
        return None
    end = find_matching(text, 0, "[", "]")
    if end != len(text) - 1:
        return None
    words = tokenize_tcl_words(text[1:-1].strip())
    if not words or words[0] not in COLLECTION_KINDS:
        return None
    kind = words[0]
    objects: List[str] = []
    idx = 1
    while idx < len(words):
        word = words[idx]
        if word.startswith("-"):
            idx += 1
            continue
        objects.extend(split_object_list(word))
        idx += 1
    return kind, objects


def iter_collection_spans(text: str) -> Iterable[Tuple[int, int, str, List[str]]]:
    idx = 0
    while idx < len(text):
        start = text.find("[", idx)
        if start < 0:
            return
        end = find_matching(text, start, "[", "]")
        if end < 0:
            return
        token = text[start : end + 1]
        parsed = parse_collection(token)
        if parsed:
            kind, objects = parsed
            yield start, end + 1, kind, objects
            idx = end + 1
        else:
            idx = start + 1


def top_port_from_connection(value: str) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    lowered = text.lower()
    for prefix in ("top.", "top/", "top:"):
        if lowered.startswith(prefix):
            return text[len(prefix) :]
    if lowered.startswith("get_ports"):
        objs = re.findall(r"\{([^}]+)\}", text)
        if objs:
            parts = split_object_list(objs[-1])
            return parts[0] if parts else ""
    return ""


def clean_top_name(value: str) -> str:
    text = clean_cell(value)
    for marker in ("top.", "top/", "top:"):
        if text.lower().startswith(marker):
            return text[len(marker) :]
    return text


def build_pad_records(instances: Dict[str, InstInfo]) -> List[PadRecord]:
    records: List[PadRecord] = []
    seen: Set[Tuple[str, str, str, str]] = set()

    def add(record: PadRecord) -> None:
        key = (record.pad_name, record.subsys_instance, record.subsys_port, record.direction)
        if record.pad_name and key not in seen:
            seen.add(key)
            records.append(record)

    for inst in instances.values():
        for port in inst.inputs.values():
            top = top_port_from_connection(port.from_whom)
            if top:
                add(PadRecord(top, top, inst.inst_name, port.name, "input", source_sdc_status=inst.sdc_status))
        for port in inst.outputs.values():
            top = clean_top_name(port.to_top)
            if top:
                add(PadRecord(top, top, inst.inst_name, port.name, "output", source_sdc_status=inst.sdc_status))
        for port in inst.inouts.values():
            top = top_port_from_connection(port.connectivity) or clean_top_name(port.inout_name) or port.name
            add(
                PadRecord(
                    top,
                    top,
                    inst.inst_name,
                    port.name,
                    "inout",
                    "yes",
                    "gpio_in,gpio_out",
                    inst.sdc_status,
                )
            )
    return records


def is_multi_bit_width(value: str) -> bool:
    text = clean_cell(value)
    if not text:
        return False
    match = re.fullmatch(r"\d+", text)
    return bool(match and int(text) > 1)


def validate_canonical_port_key(inst_name: str, direction: str, port: PortInfo, report: Report) -> None:
    if not PORT_BIT_RE.fullmatch(port.name):
        report.error(
            f"{inst_name}: {direction} port {port.name} is not a canonical scalar/bit key; "
            "expand bus/range ports to per-bit keys like name[0]"
        )
    if "[" not in port.name and (is_multi_bit_width(port.width) or is_multi_bit_width(port.used_width)):
        report.error(
            f"{inst_name}: {direction} port {port.name} has width={port.width or '-'} "
            f"used_width={port.used_width or '-'} but is not bit-expanded"
        )


def validate_canonical_top_key(inst_name: str, direction: str, port_name: str, top_name: str, report: Report) -> None:
    if top_name and not PORT_BIT_RE.fullmatch(top_name):
        report.error(
            f"{inst_name}: {direction} port {port_name} maps to non-canonical top pad {top_name}; "
            "expand bus/range pads to per-bit keys like pad[0]"
        )


def validate_integration_port_keys(instances: Dict[str, InstInfo], report: Report) -> None:
    for inst in instances.values():
        for port in inst.inputs.values():
            validate_canonical_port_key(inst.inst_name, "input", port, report)
            validate_canonical_top_key(
                inst.inst_name,
                "input",
                port.name,
                top_port_from_connection(port.from_whom),
                report,
            )
        for port in inst.outputs.values():
            validate_canonical_port_key(inst.inst_name, "output", port, report)
            validate_canonical_top_key(inst.inst_name, "output", port.name, clean_top_name(port.to_top), report)
        for port in inst.inouts.values():
            validate_canonical_port_key(inst.inst_name, "inout", port, report)
            top = top_port_from_connection(port.connectivity) or clean_top_name(port.inout_name) or port.name
            validate_canonical_top_key(inst.inst_name, "inout", port.name, top, report)


def map_single_object(inst: InstInfo, kind: str, obj: str) -> ObjectMapping:
    obj = strip_braces(obj)
    if kind == "get_clocks":
        return ObjectMapping(kind, [obj], get_collection(kind, [obj]))
    if kind == "get_pins":
        pin = obj if "/" in obj else f"{inst.inst_name}/{obj}"
        return ObjectMapping("get_pins", [pin], get_collection("get_pins", [pin]))
    if kind == "get_nets":
        top = port_top_for(inst, obj)
        if top:
            return ObjectMapping("get_nets", [top], get_collection("get_nets", [top]), [top])
        return ObjectMapping(
            "get_nets",
            [obj],
            get_collection("get_nets", [obj]),
            [],
            unresolved=True,
            message="net name could not be proven stable at SoC level",
        )
    if kind != "get_ports":
        return ObjectMapping(kind, [obj], get_collection(kind, [obj]), unresolved=True, message=f"unsupported kind {kind}")

    top = port_top_for(inst, obj)
    if top:
        return ObjectMapping("get_ports", [top], get_collection("get_ports", [top]), [top])
    if obj in inst.inputs or obj in inst.outputs or obj in inst.inouts:
        pin = f"{inst.inst_name}/{obj}"
        return ObjectMapping("get_pins", [pin], get_collection("get_pins", [pin]))
    return ObjectMapping(
        "get_pins",
        [f"{inst.inst_name}/{obj}"],
        get_collection("get_pins", [f"{inst.inst_name}/{obj}"]),
        [],
        unresolved=True,
        message=f"port {obj} not found in integration table for {inst.inst_name}",
    )


def port_top_for(inst: InstInfo, port: str) -> str:
    if port in inst.inputs:
        return top_port_from_connection(inst.inputs[port].from_whom)
    if port in inst.outputs:
        return clean_top_name(inst.outputs[port].to_top)
    if port in inst.inouts:
        return top_port_from_connection(inst.inouts[port].connectivity) or clean_top_name(inst.inouts[port].inout_name) or inst.inouts[port].name
    return ""


def merge_mappings(mappings: Sequence[ObjectMapping]) -> ObjectMapping:
    if not mappings:
        return ObjectMapping("", [], "", unresolved=True, message="empty mapping")
    kinds = {mapping.kind for mapping in mappings}
    objects: List[str] = []
    pads: List[str] = []
    unresolved = False
    messages: List[str] = []
    for mapping in mappings:
        objects.extend(mapping.objects)
        pads.extend(mapping.pad_names)
        unresolved = unresolved or mapping.unresolved
        if mapping.message:
            messages.append(mapping.message)
    if len(kinds) == 1:
        kind = mappings[0].kind
        return ObjectMapping(kind, objects, get_collection(kind, objects), pads, unresolved, "; ".join(messages))
    return ObjectMapping(
        "mixed",
        objects,
        " ".join(mapping.collection for mapping in mappings),
        pads,
        True,
        "mixed object kinds after SoC mapping; review command rewrite",
    )


def rewrite_command(raw: str, inst: InstInfo) -> Tuple[str, List[str], bool, str]:
    pads: List[str] = []
    unresolved = False
    messages: List[str] = []
    replacements: List[Tuple[int, int, str]] = []

    for start, end, kind, objects in iter_collection_spans(raw):
        mappings = [map_single_object(inst, kind, obj) for obj in objects]
        merged = merge_mappings(mappings)
        pads.extend(merged.pad_names)
        unresolved = unresolved or merged.unresolved
        if merged.message:
            messages.append(merged.message)
        replacements.append((start, end, merged.collection or raw[start:end]))

    if not replacements:
        rewritten = raw
    else:
        chunks: List[str] = []
        cursor = 0
        for start, end, value in replacements:
            chunks.append(raw[cursor:start])
            chunks.append(value)
            cursor = end
        chunks.append(raw[cursor:])
        rewritten = "".join(chunks)
    return rewritten, sorted(set(pads)), unresolved, "; ".join(dict.fromkeys(messages))


def option_value(tokens: Sequence[str], option: str) -> str:
    for idx, token in enumerate(tokens):
        if token == option and idx + 1 < len(tokens):
            return strip_braces(tokens[idx + 1])
    return ""


def has_option(tokens: Sequence[str], option: str) -> bool:
    return option in tokens


OPTIONS_WITH_VALUE = {
    "-clock",
    "-lib_cell",
    "-pin",
    "-from_pin",
    "-max",
    "-min",
    "-input_transition_rise",
    "-input_transition_fall",
}
OPTIONS_NO_VALUE = {"-add_delay", "-clock_fall", "-rise", "-fall"}


def positional_tokens(tokens: Sequence[str]) -> List[str]:
    result: List[str] = []
    skip_next = False
    for token in tokens[1:]:
        if skip_next:
            skip_next = False
            continue
        if token in OPTIONS_WITH_VALUE:
            skip_next = True
            continue
        if token in OPTIONS_NO_VALUE:
            continue
        if token.startswith("-"):
            continue
        if parse_collection(token):
            continue
        result.append(strip_braces(token))
    return result


def extract_clock_name(tokens: Sequence[str]) -> str:
    value = option_value(tokens, "-clock")
    parsed = parse_collection(value)
    if parsed and parsed[0] == "get_clocks" and parsed[1]:
        return parsed[1][0]
    return strip_braces(value)


def last_non_clock_collection(tokens: Sequence[str]) -> Optional[Tuple[str, List[str], str]]:
    found: Optional[Tuple[str, List[str], str]] = None
    for token in tokens:
        parsed = parse_collection(token)
        if parsed and parsed[0] != "get_clocks":
            found = (parsed[0], parsed[1], token)
    return found


def all_non_clock_collections(tokens: Sequence[str]) -> List[Tuple[str, List[str], str]]:
    result: List[Tuple[str, List[str], str]] = []
    for token in tokens:
        parsed = parse_collection(token)
        if parsed and parsed[0] != "get_clocks":
            result.append((parsed[0], parsed[1], token))
    return result


def extract_constraints_from_instance(inst: InstInfo, scenario: str, report: Report) -> List[ExtractedConstraint]:
    if not inst.sdc_path:
        return []
    try:
        text = read_text(inst.sdc_path)
        digest = digest_file(inst.sdc_path)
    except Exception as exc:
        report.error(f"failed to read {inst.sdc_path}: {exc}")
        return []

    now = datetime.now().isoformat(timespec="seconds")
    results: List[ExtractedConstraint] = []
    for cmd in iter_tcl_commands_with_line(text):
        tokens = tokenize_tcl_words(cmd.raw)
        if not tokens:
            continue
        command = tokens[0]
        ctype = SUPPORTED_COMMANDS.get(command)
        if not ctype:
            continue
        rewritten, pads, unresolved, rewrite_msg = rewrite_command(cmd.raw, inst)
        status = "ok" if not unresolved else "needs_review"

        values = {header: "" for header in IO_HEADERS}
        values.update(
            {
                "scenario": scenario,
                "stage": "all",
                "corner": "all",
                "constraint_type": ctype,
                "source_type": "extracted",
                "source_sdc_status": "available",
                "source_sdc_file": inst.sdc_path.name,
                "source_line": str(cmd.line_no),
                "source_digest": digest,
                "extraction_time": now,
                "original_command": cmd.raw,
                "apply": "no",
                "review_status": "pending",
                "owner": inst.owner,
                "extra_options": f"rewritten_command={rewritten}" if rewritten != cmd.raw else "",
            }
        )

        target = last_non_clock_collection(tokens)
        mapped_obj = ""
        original_object = ""
        if target:
            kind, objects, original_token = target
            original_object = original_token
            mapped = merge_mappings([map_single_object(inst, kind, obj) for obj in objects])
            mapped_obj = mapped.collection
            values["soc_object"] = mapped.collection
            if mapped.pad_names:
                pads.extend(mapped.pad_names)
            if mapped.message:
                rewrite_msg = "; ".join(filter(None, [rewrite_msg, mapped.message]))
        elif ctype == "false_path":
            collections = all_non_clock_collections(tokens)
            mapped_collections = []
            for kind, objects, _ in collections:
                mapped = merge_mappings([map_single_object(inst, kind, obj) for obj in objects])
                mapped_collections.append(mapped.collection)
                pads.extend(mapped.pad_names)
            mapped_obj = " ".join(mapped_collections)
            values["soc_object"] = mapped_obj

        values["pad_name"] = pads[0] if pads else infer_pad_name(inst, target)
        values["subsys_instance"] = inst.inst_name
        values["subsys_port"] = infer_subsys_port(inst, target)
        values["direction"] = infer_direction(inst, values["subsys_port"], ctype)
        values["original_object"] = original_object
        values["note"] = rewrite_msg

        fill_command_values(values, tokens, ctype)

        if ctype == "dont_touch_network" and unresolved:
            status = "needs_review"
        include_in_form = bool(values["pad_name"])
        if not include_in_form:
            status = "out_of_scope_non_pad"
            rewrite_msg = "; ".join(
                filter(None, [rewrite_msg, "no 00 pad edge maps this constraint; kept in extraction_log only"])
            )

        values["note"] = rewrite_msg
        results.append(
            ExtractedConstraint(
                values=values,
                parse_status=status,
                mapped_soc_object=mapped_obj or values["soc_object"],
                message=rewrite_msg,
                include_in_form=include_in_form,
            )
        )
    form_count = sum(1 for item in results if item.include_in_form)
    report.info(
        f"extracted {form_count} IO/pad candidate(s) and logged "
        f"{len(results) - form_count} out-of-scope command(s) from {inst.sdc_path.name}"
    )
    return results


def infer_pad_name(inst: InstInfo, target: Optional[Tuple[str, List[str], str]]) -> str:
    if not target:
        return ""
    _, objects, _ = target
    for obj in objects:
        top = port_top_for(inst, obj)
        if top:
            return top
    return ""


def infer_subsys_port(inst: InstInfo, target: Optional[Tuple[str, List[str], str]]) -> str:
    if not target:
        return ""
    _, objects, _ = target
    for obj in objects:
        if obj in inst.inputs or obj in inst.outputs or obj in inst.inouts:
            return obj
    return objects[0] if objects else ""


def infer_direction(inst: InstInfo, port: str, ctype: str) -> str:
    if port in inst.inputs:
        return "input"
    if port in inst.outputs:
        return "output"
    if port in inst.inouts:
        return "inout"
    if ctype in {"input_delay", "input_transition", "driving_cell", "drive"}:
        return "input"
    if ctype in {"output_delay", "load"}:
        return "output"
    return "unknown"


def fill_command_values(values: Dict[str, str], tokens: Sequence[str], ctype: str) -> None:
    pos = positional_tokens(tokens)
    if ctype in DELAY_TYPES:
        values["clock_name"] = extract_clock_name(tokens)
        min_value = option_value(tokens, "-min")
        max_value = option_value(tokens, "-max")
        if min_value:
            values["min_value"] = min_value
        if max_value:
            values["max_value"] = max_value
        if not min_value and not max_value and pos:
            values["value"] = pos[0]
        if has_option(tokens, "-rise"):
            values["delay_edge"] = "data_rise"
        if has_option(tokens, "-fall"):
            values["delay_edge"] = "data_fall"
        if has_option(tokens, "-clock_fall"):
            values["delay_polarity"] = "clock_fall"
        values["add_delay"] = "yes" if has_option(tokens, "-add_delay") else "no"
    elif ctype in {"load", "drive", "input_transition", "max_transition", "max_capacitance"}:
        if pos:
            values["value"] = pos[0]
        if has_option(tokens, "-rise"):
            values["delay_edge"] = "data_rise"
        if has_option(tokens, "-fall"):
            values["delay_edge"] = "data_fall"
    elif ctype == "driving_cell":
        values["drive_lib_cell"] = option_value(tokens, "-lib_cell")
        values["drive_pin"] = option_value(tokens, "-pin")
        values["drive_from_pin"] = option_value(tokens, "-from_pin")
        values["drive_input_transition_rise"] = option_value(tokens, "-input_transition_rise")
        values["drive_input_transition_fall"] = option_value(tokens, "-input_transition_fall")


def read_clock_inventory(
    path: Path,
    report: Report,
    source_file: str = "common",
    required: bool = False,
) -> Dict[str, ClockInfo]:
    clocks: Dict[str, ClockInfo] = {}
    if not path.is_file():
        if required:
            report.error(f"TARGET_UPSTREAM_CLOCK_INVENTORY_MISSING: {path}")
        else:
            report.warn(f"clock inventory not found: {path}")
        return clocks
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        if not reader.fieldnames or "clock_name" not in reader.fieldnames:
            report.error(f"{path} does not contain clock_name column")
            return clocks
        for row in reader:
            action = clean_cell(row.get("final_action"))
            if action and action not in ACTIVE_01_ACTIONS:
                continue
            clock_name = clean_cell(row.get("clock_name"))
            if not clock_name:
                continue
            clocks[clock_name] = ClockInfo(
                clock_name=clock_name,
                direct_source=clean_cell(row.get("direct_source")),
                producer_object=clean_cell(row.get("producer_object")),
                final_action=action,
                source_file=source_file,
            )
    report.info(f"loaded {len(clocks)} clock(s) from {path}")
    return clocks


def create_or_load_workbook(path: Path) -> Tuple[Workbook, bool]:
    if path.is_file():
        return load_workbook(path), False
    wb = Workbook()
    ws = wb.active
    ws.title = "io_constraints"
    return wb, True


def ensure_sheet(wb: Workbook, name: str, headers: Sequence[str]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if ws.max_row == 1 and all(ws.cell(row=1, column=col).value is None for col in range(1, len(headers) + 1)):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    existing = [clean_cell(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
    for header in headers:
        if header not in existing:
            ws.cell(row=1, column=len(existing) + 1, value=header)
            existing.append(header)
    style_sheet(ws)


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for column_cells in ws.columns:
        max_len = 8
        for cell in column_cells:
            max_len = max(max_len, len(clean_cell(cell.value)))
        width = min(max_len + 2, 45)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:25] or "table"
        if not ws.tables:
            ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"
            tab = Table(displayName=f"{table_name}_tbl", ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            try:
                ws.add_table(tab)
            except ValueError:
                pass


def header_map(ws) -> Dict[str, int]:
    return {clean_cell(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if clean_cell(cell.value)}


def row_values(ws, row_idx: int, headers: Sequence[str]) -> Dict[str, object]:
    hmap = header_map(ws)
    return {header: ws.cell(row=row_idx, column=hmap[header]).value if header in hmap else "" for header in headers}


def append_dict(ws, headers: Sequence[str], values: Dict[str, object], fill: Optional[PatternFill] = None) -> None:
    hmap = header_map(ws)
    row_idx = ws.max_row + 1
    for header in headers:
        col = hmap[header]
        cell = ws.cell(row=row_idx, column=col, value=values.get(header, ""))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def form_row_key(values: Dict[str, object]) -> Tuple[str, ...]:
    return (
        clean_cell(values.get("source_sdc_file")),
        clean_cell(values.get("source_line")),
        clean_cell(values.get("constraint_type")),
        hashlib.sha1(clean_cell(values.get("original_command")).encode("utf-8")).hexdigest()[:12],
        clean_cell(values.get("pad_name")),
        clean_cell(values.get("soc_object")),
    )


def pad_key(values: Dict[str, object]) -> Tuple[str, str, str, str, str]:
    return (
        clean_cell(values.get("view_id")),
        clean_cell(values.get("pad_name")),
        clean_cell(values.get("subsys_instance")),
        clean_cell(values.get("subsys_port")),
        clean_cell(values.get("direction_from_integration")),
    )


def sync_workbook(
    path: Path,
    pads: Sequence[PadRecord],
    extracted: Sequence[ExtractedConstraint],
    report: Report,
    pad_context: Optional[Dict[str, object]] = None,
) -> None:
    wb, created = create_or_load_workbook(path)
    ensure_sheet(wb, "io_constraints", IO_HEADERS)
    ensure_sheet(wb, "pad_inventory", PAD_HEADERS)
    ensure_sheet(wb, "extraction_log", LOG_HEADERS)

    ws_io = wb["io_constraints"]
    if pad_context:
        desired_rows = [item.values for item in extracted if item.include_in_form]
        desired_keys = {form_row_key(values) for values in desired_rows}
        desired_sources = {form_row_key(values)[:4] for values in desired_rows}
        obsolete_rows: List[int] = []
        for row_idx in range(2, ws_io.max_row + 1):
            values = row_values(ws_io, row_idx, IO_HEADERS)
            if normalize_key(values.get("source_type")) != "extracted":
                continue
            key = form_row_key(values)
            if key in desired_keys or key[:4] not in desired_sources:
                continue
            pad_name = clean_cell(values.get("pad_name"))
            subsys_port = clean_cell(values.get("subsys_port"))
            soc_object = clean_cell(values.get("soc_object"))
            if (
                not PORT_BIT_RE.fullmatch(pad_name)
                or not PORT_BIT_RE.fullmatch(subsys_port)
                or any(token in soc_object for token in (":", "*", "?"))
            ):
                obsolete_rows.append(row_idx)
        for row_idx in reversed(obsolete_rows):
            ws_io.delete_rows(row_idx, 1)
        if obsolete_rows:
            report.warn(
                f"removed {len(obsolete_rows)} obsolete auto-extracted broad/range IO row(s); "
                "review the replacement exact-bit rows"
            )
            report.sync_changed = True
    existing_keys = {
        form_row_key(row_values(ws_io, row_idx, IO_HEADERS))
        for row_idx in range(2, ws_io.max_row + 1)
        if any(clean_cell(ws_io.cell(row=row_idx, column=col).value) for col in range(1, ws_io.max_column + 1))
    }
    for item in extracted:
        if not item.include_in_form:
            continue
        key = form_row_key(item.values)
        if key not in existing_keys:
            append_dict(ws_io, IO_HEADERS, item.values, NEW_FILL)
            existing_keys.add(key)
            report.sync_changed = True
    if created:
        report.sync_changed = True

    ws_pad = wb["pad_inventory"]
    pad_columns = header_map(ws_pad)
    existing_pads = {
        pad_key(row_values(ws_pad, row_idx, PAD_HEADERS)): row_idx
        for row_idx in range(2, ws_pad.max_row + 1)
        if clean_cell(ws_pad.cell(row=row_idx, column=pad_columns["pad_name"]).value)
    }
    for pad in pads:
        values = {
            "schema_version": "1.0",
            "pad_name": pad.pad_name,
            "soc_top_port": pad.soc_top_port,
            "subsys_instance": pad.subsys_instance,
            "subsys_port": pad.subsys_port,
            "direction_from_integration": pad.direction,
            "effective_direction": pad.direction if pad.direction in {"input", "output"} else "",
            "is_gpio_or_inout": pad.is_gpio_or_inout,
            "related_scenarios": pad.related_scenarios,
            "source_sdc_status": pad.source_sdc_status,
            "connection_id": pad.connection_id,
            "scenario_scope": pad.scenario_scope,
            "connection_type": pad.connection_type,
            "source_workbook": pad.source_workbook,
            "source_sheet": pad.source_sheet,
            "source_row": pad.source_row,
            "note": pad.note,
        }
        if pad_context:
            values.update(pad_context)
        key = pad_key(values)
        if key not in existing_pads and key[0]:
            legacy_key = ("",) + key[1:]
            legacy_row = existing_pads.get(legacy_key)
            if legacy_row is not None:
                ws_pad.cell(legacy_row, header_map(ws_pad)["view_id"], key[0])
                existing_pads[key] = legacy_row
                existing_pads.pop(legacy_key, None)
                report.sync_changed = True
        if key not in existing_pads:
            append_dict(ws_pad, PAD_HEADERS, values, NEW_FILL)
            existing_pads[key] = ws_pad.max_row
            report.sync_changed = True
        else:
            row_idx = existing_pads[key]
            status_col = header_map(ws_pad).get("source_sdc_status")
            if status_col and clean_cell(ws_pad.cell(row_idx, status_col).value) != pad.source_sdc_status:
                ws_pad.cell(row_idx, status_col, pad.source_sdc_status)
                report.sync_changed = True

    ws_log = wb["extraction_log"]
    if ws_log.max_row > 1:
        ws_log.delete_rows(2, ws_log.max_row - 1)
    for item in extracted:
        values = {
            "source_sdc_file": item.values.get("source_sdc_file", ""),
            "source_sdc_status": item.values.get("source_sdc_status", ""),
            "source_line": item.values.get("source_line", ""),
            "command_type": item.values.get("constraint_type", ""),
            "original_command": item.values.get("original_command", ""),
            "parse_status": item.parse_status,
            "mapped_soc_object": item.mapped_soc_object,
            "source_digest": item.values.get("source_digest", ""),
            "extraction_time": item.values.get("extraction_time", ""),
            "message": item.message,
        }
        append_dict(ws_log, LOG_HEADERS, values)
    for pad in pads:
        if pad.source_sdc_status not in {"missing", "not_required"}:
            continue
        append_dict(
            ws_log,
            LOG_HEADERS,
            {
                "source_sdc_status": pad.source_sdc_status,
                "parse_status": "incomplete_evidence" if pad.source_sdc_status == "missing" else "not_required",
                "mapped_soc_object": get_collection("get_ports", [pad.soc_top_port]),
                "message": (
                    f"{pad.subsys_instance}:{pad.subsys_port} lower-level SDC is missing; "
                    "candidate extraction remains incomplete"
                    if pad.source_sdc_status == "missing"
                    else f"{pad.subsys_instance}:{pad.subsys_port} lower-level SDC is not required"
                ),
            },
        )

    add_validations(wb)
    for ws in wb.worksheets:
        style_sheet(ws)
    atomic_save_workbook(wb, path)
    if report.sync_changed:
        report.info(f"synchronized workbook {path.name}; review new rows before generation")


def add_validations(wb: Workbook) -> None:
    if "io_constraints" not in wb.sheetnames:
        return
    ws = wb["io_constraints"]
    hmap = header_map(ws)

    def add_list(header: str, values: Sequence[str]) -> None:
        if header not in hmap:
            return
        col = get_column_letter(hmap[header])
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1048576")

    add_list("scenario", sorted(SCENARIOS))
    add_list("stage", sorted(STAGES))
    add_list("source_type", sorted(SOURCE_TYPES))
    add_list("source_sdc_status", sorted(SOURCE_SDC_STATUS_VALUES - {""}))
    add_list("apply", sorted(APPLY_VALUES - {""}))
    add_list("review_status", sorted(REVIEW_STATUS_VALUES - {""}))
    add_list("add_delay", ["yes", "no"])
    add_list("direction", ["input", "output", "inout", "unknown"])
    add_list("timing_class", ["timed", "async", "untimed", "config"])
    add_list("constraint_type", sorted(set(SUPPORTED_COMMANDS.values())))
    add_list("delay_edge", ["data_rise", "data_fall", "both"])
    add_list("delay_polarity", ["clock_rise", "clock_fall"])
    add_list("object_granularity", ["single_pad", "port_list", "pattern"])

    if "pad_inventory" in wb.sheetnames:
        pad_ws = wb["pad_inventory"]
        pad_map = header_map(pad_ws)
        for header, values in (
            ("pad_disposition", ["constrained", "not_applicable", "route_to_30", "pending"]),
            ("effective_direction", ["input", "output"]),
            ("apply", ["yes", "no"]),
            ("review_status", ["pending", "approved", "rejected"]),
        ):
            if header not in pad_map:
                continue
            col = get_column_letter(pad_map[header])
            validation = DataValidation(
                type="list", formula1='"' + ",".join(values) + '"', allow_blank=True
            )
            pad_ws.add_data_validation(validation)
            validation.add(f"{col}2:{col}1048576")


def read_form_rows(path: Path) -> List[FormRow]:
    wb = load_workbook(path, data_only=False)
    if "io_constraints" not in wb.sheetnames:
        raise RuntimeError(f"{path} missing io_constraints sheet")
    ws = wb["io_constraints"]
    rows: List[FormRow] = []
    for row_idx in range(2, ws.max_row + 1):
        values = row_values(ws, row_idx, IO_HEADERS)
        if not any(clean_cell(value) for value in values.values()):
            continue
        rows.append(FormRow(row_idx=row_idx, values=values))
    return rows


def row_scenario(row: FormRow) -> str:
    return normalize_key(row.values.get("scenario")) or "common"


def row_stage(row: FormRow) -> str:
    return normalize_key(row.values.get("stage")) or "all"


def row_corner(row: FormRow) -> str:
    return clean_cell(row.values.get("corner")) or "all"


def is_apply_approved(row: FormRow) -> bool:
    return (
        normalize_key(row.values.get("apply")) == "yes"
        and normalize_key(row.values.get("review_status")) == "approved"
        and normalize_key(row.values.get("source_type")) != "na"
    )


def row_selected_for_output(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    return row_scenario(row) == scenario and row_stage(row) == stage and row_corner(row) == corner


def row_selected_for_assembled(row: FormRow, scenario: str, stage: str, corner: str) -> bool:
    row_sc = row_scenario(row)
    if row_sc not in {"common", scenario}:
        return False
    row_st = row_stage(row)
    row_co = row_corner(row)
    return (row_st == "all" and row_co == "all") or (row_st == stage and row_co == corner)


def delay_semantic_slots(row: FormRow) -> List[Tuple[str, ...]]:
    values = row.values
    prefix = (
        clean_cell(values.get("pad_name")),
        normalize_key(values.get("constraint_type")),
        clean_cell(values.get("clock_name")),
        normalize_key(values.get("delay_edge")) or "both",
        normalize_key(values.get("delay_polarity")) or "clock_rise",
        normalize_key(values.get("add_delay")) or "no",
    )
    slots = []
    for field in ("value", "min_value", "max_value", "rise_value", "fall_value"):
        if clean_cell(values.get(field)):
            slots.append(prefix + (field,))
    return slots


def validate_rows(
    rows: Sequence[FormRow],
    pads: Sequence[PadRecord],
    scenario: str,
    stage: str,
    corner: str,
    common_clocks: Dict[str, ClockInfo],
    scenario_clocks: Dict[str, ClockInfo],
    current_digests: Dict[str, str],
    expected_time_unit: str,
    expected_cap_unit: str,
    tool: str,
    report: Report,
    allow_false_path: bool = True,
) -> None:
    assembled = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner) and is_apply_approved(row)]
    assembled_na = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner) and is_approved_na(row)]
    approved_by_key: Dict[Tuple[str, str], List[FormRow]] = defaultdict(list)
    delay_groups: Dict[Tuple[str, str, str], List[FormRow]] = defaultdict(list)
    delay_slots: Dict[Tuple[str, ...], List[FormRow]] = defaultdict(list)
    pad_rows: Dict[str, List[FormRow]] = defaultdict(list)
    na_rows_by_pad: Dict[str, List[FormRow]] = defaultdict(list)
    pad_identities = {
        (pad.pad_name, pad.subsys_instance, pad.subsys_port)
        for pad in pads
    }
    pad_directions: Dict[Tuple[str, str, str], Set[str]] = defaultdict(set)
    for pad in pads:
        pad_directions[(pad.pad_name, pad.subsys_instance, pad.subsys_port)].add(
            normalize_key(pad.direction)
        )
    for row in assembled_na:
        pad = clean_cell(row.values.get("pad_name"))
        if pad:
            na_rows_by_pad[pad].append(row)

    all_clocks = dict(common_clocks)
    all_clocks.update(scenario_clocks)

    for row in rows:
        values = row.values
        apply_value = normalize_key(values.get("apply"))
        review_status = normalize_key(values.get("review_status"))
        source_type = normalize_key(values.get("source_type"))
        source_sdc_status = normalize_key(values.get("source_sdc_status"))
        ctype = normalize_key(values.get("constraint_type"))
        pad = clean_cell(values.get("pad_name"))
        soc_object = clean_cell(values.get("soc_object"))
        clock_name = clean_cell(values.get("clock_name"))
        raw_scenario = clean_cell(values.get("scenario"))
        is_na = source_type == "na"

        if apply_value and apply_value not in APPLY_VALUES:
            report.error(f"io_constraints row {row.row_idx}: apply must be yes/no, got {apply_value}")
        if review_status and review_status not in REVIEW_STATUS_VALUES:
            report.error(f"io_constraints row {row.row_idx}: invalid review_status {review_status}")
        if source_type and source_type not in SOURCE_TYPES:
            report.error(f"io_constraints row {row.row_idx}: invalid source_type {source_type}")
        if source_sdc_status and source_sdc_status not in SOURCE_SDC_STATUS_VALUES:
            report.error(f"io_constraints row {row.row_idx}: invalid source_sdc_status {source_sdc_status}")

        if apply_value == "yes":
            if not raw_scenario:
                report.error(f"io_constraints row {row.row_idx}: apply=yes but scenario is blank")
            if not pad:
                report.error(f"io_constraints row {row.row_idx}: apply=yes but pad_name is blank")
            if not soc_object and not is_na:
                report.error(f"io_constraints row {row.row_idx}: apply=yes but soc_object is blank")
            if not ctype and not is_na:
                report.error(f"io_constraints row {row.row_idx}: apply=yes but constraint_type is blank")
            if not is_na and review_status == "approved":
                if not clean_cell(values.get("owner")):
                    report.error(
                        f"io_constraints row {row.row_idx}: approved row requires owner"
                    )
                if not clean_cell(values.get("basis")) and not clean_cell(values.get("note")):
                    report.error(
                        f"io_constraints row {row.row_idx}: approved row requires basis or note"
                    )
            if is_na and review_status == "approved":
                missing_review = [
                    name for name in ("owner", "basis", "reviewer", "review_date")
                    if not clean_cell(values.get(name))
                ]
                if missing_review:
                    report.error(
                        f"io_constraints row {row.row_idx}: approved NA row requires "
                        + ", ".join(missing_review)
                    )
                identity = (
                    pad,
                    clean_cell(values.get("subsys_instance")),
                    clean_cell(values.get("subsys_port")),
                )
                if identity not in pad_identities:
                    report.error(
                        f"io_constraints row {row.row_idx}: approved NA row does not match an exact pad inventory key"
                    )
            if (
                not is_na
                and review_status == "approved"
                and normalize_key(values.get("object_granularity")) not in {"port_list", "pattern"}
            ):
                identity = (
                    pad,
                    clean_cell(values.get("subsys_instance")),
                    clean_cell(values.get("subsys_port")),
                )
                if identity not in pad_identities:
                    report.error(
                        f"io_constraints row {row.row_idx}: approved row does not match an exact pad inventory key"
                    )
                else:
                    integration_directions = pad_directions.get(identity, set())
                    effective_direction = normalize_key(values.get("direction"))
                    if integration_directions == {"inout"}:
                        if effective_direction not in {"input", "output"}:
                            report.error(
                                f"io_constraints row {row.row_idx}: inout pad requires an effective "
                                "direction of input or output for this run"
                            )
                    elif effective_direction not in integration_directions:
                        report.error(
                            f"io_constraints row {row.row_idx}: direction {effective_direction or '<blank>'} "
                            "conflicts with integration direction "
                            f"{','.join(sorted(integration_directions)) or '<unknown>'}"
                        )
            if (
                not allow_false_path
                and not is_na
                and review_status == "approved"
                and normalize_key(values.get("object_granularity")) in {"port_list", "pattern"}
            ):
                report.error(
                    f"io_constraints row {row.row_idx}: flat 04 requires one exact top pad bit; "
                    "port_list/pattern approval is not allowed"
                )
            if ctype in DELAY_TYPES and not clock_name:
                report.error(f"io_constraints row {row.row_idx}: {ctype} requires clock_name")
            if ctype in DELAY_TYPES and clock_name:
                if row_scenario(row) == "common" and clock_name not in common_clocks:
                    report.error(
                        f"io_constraints row {row.row_idx}: common 04 delay references non-common clock {clock_name}"
                    )
                elif clock_name not in all_clocks:
                    report.error(
                        f"io_constraints row {row.row_idx}: clock_name {clock_name} not found in assembled inventory"
                    )
            if ctype in DELAY_TYPES:
                if clean_cell(values.get("value")) and (
                    clean_cell(values.get("min_value")) or clean_cell(values.get("max_value"))
                ):
                    report.error(
                        f"io_constraints row {row.row_idx} {pad}: value must not be filled together with "
                        "min_value/max_value"
                    )
                if (clean_cell(values.get("rise_value")) or clean_cell(values.get("fall_value"))) and (
                    clean_cell(values.get("value"))
                    or clean_cell(values.get("min_value"))
                    or clean_cell(values.get("max_value"))
                ):
                    report.error(
                        f"io_constraints row {row.row_idx} {pad}: rise_value/fall_value must not be filled "
                        "together with value/min_value/max_value"
                    )
                if clean_cell(values.get("value")) and not (
                    clean_cell(values.get("min_value"))
                    or clean_cell(values.get("max_value"))
                    or clean_cell(values.get("rise_value"))
                    or clean_cell(values.get("fall_value"))
                ):
                    report.warn(
                        f"io_constraints row {row.row_idx} {pad}: delay uses bare value without -min/-max; "
                        "confirm tool/methodology allows this"
                    )
            if ctype in DELAY_TYPES or ctype == "driving_cell":
                if has_structured_extra_option_overlap(values):
                    report.error(
                        f"io_constraints row {row.row_idx} {pad}: extra_options repeats structured delay/driving fields"
                    )
            if ctype in ELECTRICAL_TYPES and row_stage(row) == "all" and row_corner(row) == "all":
                report.warn(
                    f"io_constraints row {row.row_idx} {pad}: {ctype} is view-independent; confirm external "
                    "electrical environment is shared by all stage/corner views"
                )
            if ctype == "dont_touch_network" and normalize_key(values.get("note")).find("synthesis") < 0:
                report.warn(
                    f"io_constraints row {row.row_idx} {pad}: dont_touch_network is synthesis-only; "
                    "note should state synthesis usage"
                )
            check_source_digest(row, current_digests, report)
            check_units(row, expected_time_unit, expected_cap_unit, report)
            if normalize_key(values.get("object_granularity")) == "pattern":
                report.warn(
                    f"io_constraints row {row.row_idx} {pad}: object_granularity=pattern; review expanded objects"
                )

        if is_apply_approved(row) and row in assembled:
            approved_by_key[(pad, ctype)].append(row)
            if pad:
                pad_rows[pad].append(row)
            if ctype in DELAY_TYPES:
                delay_groups[(pad, ctype, clock_name)].append(row)
                for slot in delay_semantic_slots(row):
                    delay_slots[slot].append(row)
            if ctype in DELAY_TYPES and soc_object_references_clock_source(soc_object, all_clocks):
                report.warn(
                    f"io_constraints row {row.row_idx} {pad}: delay target appears to be a clock source/target; "
                    "review whether this should be modeled in 01 instead"
                )
            if ctype in DELAY_TYPES and not clean_cell(values.get("basis")) and not clean_cell(values.get("note")):
                if clean_cell(values.get("max_value")) and not clean_cell(values.get("min_value")):
                    report.warn(
                        f"io_constraints row {row.row_idx} {pad}: max delay exists without min delay/basis; "
                        "hold-side external constraint may be missing"
                    )
            if (
                ctype == "false_path"
                and normalize_key(values.get("timing_class")) == "timed"
                and not has_explanation(row)
            ):
                report.error(
                    f"io_constraints row {row.row_idx} {pad}: timed IO has approved false_path without basis/note"
                )
            if ctype == "false_path" and not allow_false_path:
                report.error(
                    f"io_constraints row {row.row_idx} {pad}: flat 04 does not own false_path; "
                    "approve the exact pad as route_to_30"
                )
            direction = normalize_key(values.get("direction"))
            if direction == "output" and ctype in {"input_delay", "input_transition", "driving_cell", "drive"}:
                report.error(
                    f"io_constraints row {row.row_idx} {pad}: {ctype} is incompatible with output direction"
                )
            if direction == "input" and ctype in {"output_delay", "load"}:
                report.error(
                    f"io_constraints row {row.row_idx} {pad}: {ctype} is incompatible with input direction"
                )
            if normalize_key(values.get("timing_class")) in {"async", "untimed"} and not has_explanation(row):
                report.warn(
                    f"io_constraints row {row.row_idx} {pad}: timing_class={normalize_key(values.get('timing_class'))} "
                    "requires basis or note"
                )
            if row_scenario(row) == "common" and normalize_key(values.get("direction")) == "inout" and ctype in DELAY_TYPES:
                report.warn(
                    f"io_constraints row {row.row_idx} {pad}: inout/GPIO direction-specific delay is in common"
                )

        if is_apply_approved(row) and row in assembled:
            if normalize_key(values.get("constraint_type")) == "dont_touch_network" and tool == "sta":
                continue
            if not commands_for_row(row, tool, True):
                report.error(
                    f"io_constraints row {row.row_idx} {pad}: approved row cannot emit an SDC command; "
                    "check required value fields or false_path rewritten/original command"
                )

    for key, group in approved_by_key.items():
        ctype = key[1]
        if ctype in DELAY_TYPES:
            continue
        if len(group) > 1:
            report.error(
                f"assembled view duplicate/conflict for pad={key[0]} constraint_type={ctype}: "
                f"rows {', '.join(str(row.row_idx) for row in group)}"
            )

    for slot, group in delay_slots.items():
        if len(group) > 1:
            report.error(
                "assembled view duplicate/conflict for delay semantic slot "
                f"{'/'.join(slot)}: rows {', '.join(str(row.row_idx) for row in group)}"
            )

    for pad, group in pad_rows.items():
        timing_classes = {normalize_key(row.values.get("timing_class")) for row in group if normalize_key(row.values.get("timing_class"))}
        has_delay = any(normalize_key(row.values.get("constraint_type")) in DELAY_TYPES for row in group)
        has_false_path = any(normalize_key(row.values.get("constraint_type")) == "false_path" for row in group)
        has_load = any(normalize_key(row.values.get("constraint_type")) == "load" for row in group)
        has_input_env = any(
            normalize_key(row.values.get("constraint_type")) in {"driving_cell", "drive", "input_transition"}
            for row in group
        )
        has_dc = any(normalize_key(row.values.get("constraint_type")) == "driving_cell" for row in group)
        has_it = any(normalize_key(row.values.get("constraint_type")) == "input_transition" for row in group)
        explanations = any(has_explanation(row) for row in group + na_rows_by_pad.get(pad, []))

        if "timed" in timing_classes and not has_delay and not explanations:
            report.error(f"assembled view pad {pad}: timing_class=timed but no input/output delay or basis/note")
        if "timed" in timing_classes and has_delay and has_false_path and not explanations:
            report.error(f"assembled view pad {pad}: timed delay and false_path both approved without basis/note")
        if has_dc and has_it and not explanations:
            report.warn(f"assembled view pad {pad}: driving_cell and input_transition both approved without basis/note")

        directions = {normalize_key(row.values.get("direction")) for row in group}
        if "output" in directions and not has_load and not explanations:
            report.warn(f"assembled view pad {pad}: output pad lacks set_load or NA/basis")
        if "input" in directions and not has_input_env and not explanations:
            report.warn(f"assembled view pad {pad}: input pad lacks driving_cell/drive/input_transition or NA/basis")

    for pad in pads:
        group = pad_rows.get(pad.pad_name, [])
        ctypes = {normalize_key(row.values.get("constraint_type")) for row in group}
        explanations = any(has_explanation(row) for row in group + na_rows_by_pad.get(pad.pad_name, []))
        direction = normalize_key(pad.direction)
        if direction == "output" and "load" not in ctypes and not explanations:
            report.warn(f"pad_inventory {pad.pad_name}: output pad lacks approved set_load or NA/basis")
        if direction == "input" and not ({"driving_cell", "drive", "input_transition"} & ctypes) and not explanations:
            report.warn(
                f"pad_inventory {pad.pad_name}: input pad lacks approved driving_cell/drive/input_transition or NA/basis"
            )

    for (pad, ctype, clock), group in delay_groups.items():
        sorted_group = sorted(group, key=lambda row: row.row_idx)
        no_add = [row for row in sorted_group if normalize_key(row.values.get("add_delay")) != "yes"]
        if not no_add:
            report.warn(
                f"delay group {pad}/{ctype}/{clock}: all emitted rows have add_delay=yes; first row will be emitted as base"
            )
        elif len(no_add) > 1:
            report.warn(
                f"delay group {pad}/{ctype}/{clock}: multiple base rows without add_delay: "
                f"{', '.join(str(row.row_idx) for row in no_add)}"
            )
        has_min = any(clean_cell(row.values.get("min_value")) for row in sorted_group)
        has_max = any(clean_cell(row.values.get("max_value")) for row in sorted_group)
        if has_max and not has_min and not any(clean_cell(row.values.get("basis")) or clean_cell(row.values.get("note")) for row in sorted_group):
            report.warn(f"delay group {pad}/{ctype}/{clock}: has -max but no -min or basis")


def soc_object_references_clock_source(soc_object: str, clocks: Dict[str, ClockInfo]) -> bool:
    text = clean_cell(soc_object)
    for clock in clocks.values():
        for obj in (clock.direct_source, clock.producer_object):
            obj = clean_cell(obj)
            if obj and obj in text:
                return True
    return False


def output_sdc_path(cwd: Path, scenario: str, stage: str, corner: str) -> Path:
    if scenario == "common":
        if stage == "all" and corner == "all":
            return cwd / "common/04_soc_io_pads.sdc"
        return cwd / f"common/04_soc_io_pads_{stage}_{safe_filename_token(corner)}.sdc"
    if stage == "all" and corner == "all":
        return cwd / f"scenarios/{scenario}_io_pads.sdc"
    return cwd / f"scenarios/{scenario}_io_pads_{stage}_{safe_filename_token(corner)}.sdc"


def report_path(cwd: Path, scenario: str, stage: str, corner: str) -> Path:
    return cwd / f"io_pad_check_report_{scenario}_{stage}_{safe_filename_token(corner)}.txt"


def generate_sdc(
    rows: Sequence[FormRow],
    scenario: str,
    stage: str,
    corner: str,
    tool: str,
    completeness: RunCompleteness,
    assembled: bool = False,
) -> List[str]:
    selector = row_selected_for_assembled if assembled else row_selected_for_output
    selected = [
        row for row in rows
        if selector(row, scenario, stage, corner) and is_apply_approved(row)
    ]
    emitted_delay_groups: Set[Tuple[str, str, str]] = set()

    lines = [
        "################################################################################",
        f"# Author: {author_name()}",
        "# Stage: 04_soc_io_pads",
        "# Script: 04_extract_soc_io_pads.py",
        f"# Run completeness: {completeness.status}",
        f"# Available harden SDC: {completeness.available_count}",
        f"# Missing harden SDC: {completeness.missing_count}",
        f"# Not-required harden SDC: {completeness.not_required_count}",
        f"# Missing instances: {','.join(completeness.missing_instances) or '<none>'}",
        (
            "# Auto-generated SoC IO/pad constraints for "
            f"scenario: {scenario}, stage: {stage}, corner: {corner}, tool: {tool}"
        ),
        "# Source: 04_soc_io_pads.xlsx io_constraints sheet",
        "# Only apply=yes and review_status=approved rows are emitted.",
        "################################################################################",
        "",
    ]
    emitted = 0
    for row in selected:
        delay_key = row_delay_group_key(row)
        is_first_delay_command = bool(delay_key and delay_key not in emitted_delay_groups)
        commands = commands_for_row(row, tool, is_first_delay_command)
        if not commands:
            continue
        if delay_key:
            emitted_delay_groups.add(delay_key)
        values = row.values
        lines.append(
            f"# row {row.row_idx}: {clean_cell(values.get('pad_name'))} "
            f"{clean_cell(values.get('constraint_type'))}"
        )
        if clean_cell(values.get("basis")):
            lines.append(f"# Basis: {clean_cell(values.get('basis'))}")
        if clean_cell(values.get("source_type")) == "extracted":
            source = f"{clean_cell(values.get('source_sdc_file'))}:{clean_cell(values.get('source_line'))}"
            lines.append(f"# Extracted from: {source}")
        lines.extend(commands)
        lines.append("")
        emitted += len(commands)
    if emitted == 0:
        lines.append("# No IO/pad commands emitted for selected scenario/stage/corner/tool.")
    return lines


def row_delay_group_key(row: FormRow) -> Optional[Tuple[str, str, str]]:
    ctype = normalize_key(row.values.get("constraint_type"))
    if ctype not in DELAY_TYPES:
        return None
    return (
        clean_cell(row.values.get("pad_name")),
        ctype,
        clean_cell(row.values.get("clock_name")),
    )


def commands_for_row(row: FormRow, tool: str, is_first_delay_command: bool) -> List[str]:
    values = row.values
    ctype = normalize_key(values.get("constraint_type"))
    obj = format_soc_object(clean_cell(values.get("soc_object")), ctype)
    if not obj:
        return []
    if ctype == "dont_touch_network" and tool == "sta":
        return []
    if ctype == "false_path":
        rewritten = extra_option_value(clean_cell(values.get("extra_options")), "rewritten_command")
        if rewritten:
            return [rewritten]
        original = clean_cell(values.get("original_command"))
        return [original] if original.startswith("set_false_path") else []
    if ctype in DELAY_TYPES:
        return delay_commands(values, ctype, obj, is_first_delay_command)
    if ctype == "load":
        return [f"set_load {format_number(values.get('value'))} {obj}"] if clean_cell(values.get("value")) else []
    if ctype == "input_transition":
        return transition_like_commands("set_input_transition", values, obj)
    if ctype == "drive":
        return [f"set_drive {format_number(values.get('value'))} {obj}"] if clean_cell(values.get("value")) else []
    if ctype == "driving_cell":
        return driving_cell_commands(values, obj)
    if ctype == "max_transition":
        return [f"set_max_transition {format_number(values.get('value'))} {obj}"] if clean_cell(values.get("value")) else []
    if ctype == "max_capacitance":
        return [f"set_max_capacitance {format_number(values.get('value'))} {obj}"] if clean_cell(values.get("value")) else []
    if ctype == "dont_touch_network":
        return [f"set_dont_touch_network {obj}"]
    return []


def format_soc_object(value: str, ctype: str) -> str:
    value = clean_cell(value)
    if not value:
        return ""
    if value.startswith("["):
        return value
    if ctype == "dont_touch_network":
        return get_collection("get_nets", split_object_list(value))
    if "/" in value:
        return get_collection("get_pins", split_object_list(value))
    return get_collection("get_ports", split_object_list(value))


def delay_commands(values: Dict[str, object], ctype: str, obj: str, is_first_delay_command: bool) -> List[str]:
    cmd = "set_input_delay" if ctype == "input_delay" else "set_output_delay"
    clock = clean_cell(values.get("clock_name"))
    if not clock:
        return []
    base_opts = [f"-clock [get_clocks {brace_list([clock])}]"]
    edge = normalize_key(values.get("delay_edge"))
    if edge == "data_rise":
        base_opts.append("-rise")
    elif edge == "data_fall":
        base_opts.append("-fall")
    if normalize_key(values.get("delay_polarity")) == "clock_fall":
        base_opts.append("-clock_fall")

    commands: List[str] = []
    min_value = clean_cell(values.get("min_value"))
    max_value = clean_cell(values.get("max_value"))
    value = clean_cell(values.get("value"))
    rise_value = clean_cell(values.get("rise_value"))
    fall_value = clean_cell(values.get("fall_value"))

    def add_command(option_parts: Sequence[str], number: str) -> None:
        opts = list(option_parts)
        if commands or not is_first_delay_command:
            opts.append("-add_delay")
        commands.append(f"{cmd} {' '.join(opts)} {format_number(number)} {obj}")

    if rise_value:
        opts = [opt for opt in base_opts if opt not in {"-rise", "-fall"}]
        opts.append("-rise")
        add_command(opts, rise_value)
    if fall_value:
        opts = [opt for opt in base_opts if opt not in {"-rise", "-fall"}]
        opts.append("-fall")
        add_command(opts, fall_value)
    if min_value:
        add_command([*base_opts, "-min"], min_value)
    if max_value:
        add_command([*base_opts, "-max"], max_value)
    if not commands and value:
        add_command(base_opts, value)
    return commands


def transition_like_commands(command: str, values: Dict[str, object], obj: str) -> List[str]:
    value = clean_cell(values.get("value"))
    rise_value = clean_cell(values.get("rise_value"))
    fall_value = clean_cell(values.get("fall_value"))
    commands: List[str] = []
    if rise_value:
        commands.append(f"{command} -rise {format_number(rise_value)} {obj}")
    if fall_value:
        commands.append(f"{command} -fall {format_number(fall_value)} {obj}")
    if commands:
        return commands
    if not value:
        return []
    edge = normalize_key(values.get("delay_edge"))
    opt = ""
    if edge == "data_rise":
        opt = " -rise"
    elif edge == "data_fall":
        opt = " -fall"
    return [f"{command}{opt} {format_number(value)} {obj}"]


def driving_cell_commands(values: Dict[str, object], obj: str) -> List[str]:
    opts: List[str] = []
    for column, option in (
        ("drive_lib_cell", "-lib_cell"),
        ("drive_pin", "-pin"),
        ("drive_from_pin", "-from_pin"),
        ("drive_input_transition_rise", "-input_transition_rise"),
        ("drive_input_transition_fall", "-input_transition_fall"),
    ):
        val = clean_cell(values.get(column))
        if val:
            opts.append(f"{option} {val}")
    if not opts:
        rewritten = extra_option_value(clean_cell(values.get("extra_options")), "rewritten_command")
        if rewritten:
            return [rewritten]
        note = clean_cell(values.get("note")) + " " + clean_cell(values.get("extra_options"))
        if "passthrough_required" in note:
            original = clean_cell(values.get("original_command"))
            return [original] if original else []
        return []
    return [f"set_driving_cell {' '.join(opts)} {obj}"]


def extra_option_value(text: str, key: str) -> str:
    prefix = f"{key}="
    for part in text.split(";"):
        part = part.strip()
        if part.startswith(prefix):
            return part[len(prefix) :].strip()
    if text.startswith(prefix):
        return text[len(prefix) :].strip()
    return ""


def build_coverage_lines(
    rows: Sequence[FormRow],
    pads: Sequence[PadRecord],
    scenario: str,
    stage: str,
    corner: str,
) -> List[str]:
    selected = [row for row in rows if row_selected_for_output(row, scenario, stage, corner) and is_apply_approved(row)]
    assembled = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner) and is_apply_approved(row)]
    assembled_na = [row for row in rows if row_selected_for_assembled(row, scenario, stage, corner) and is_approved_na(row)]
    by_pad: Dict[str, List[FormRow]] = defaultdict(list)
    for row in assembled + assembled_na:
        pad = clean_cell(row.values.get("pad_name")) or "<no-pad>"
        by_pad[pad].append(row)

    lines = [
        "",
        "Coverage:",
        f"  inventory pads        : {len(pads)}",
        f"  assembled approved row: {len(assembled)}",
        f"  assembled approved NA : {len(assembled_na)}",
        f"  emitted approved row  : {len(selected)}",
        "",
        "  Per-pad assembled status:",
    ]
    if not pads:
        lines.append("    <no pad_inventory records>")
    for pad in pads:
        group = by_pad.get(pad.pad_name, [])
        ctypes = sorted({normalize_key(row.values.get("constraint_type")) for row in group if normalize_key(row.values.get("constraint_type"))})
        timing_classes = sorted({normalize_key(row.values.get("timing_class")) for row in group if normalize_key(row.values.get("timing_class"))})
        origin = sorted({row_scenario(row) + "/" + row_stage(row) + "/" + row_corner(row) for row in group})
        has_na = any(normalize_key(row.values.get("source_type")) == "na" for row in group)
        status = ", ".join(ctypes) if ctypes else ("NA" if has_na else "NO_APPROVED_CONSTRAINT")
        cls = ",".join(timing_classes) if timing_classes else "class-unset"
        src = ",".join(origin) if origin else "-"
        lines.append(
            f"    {pad.pad_name}: dir={pad.direction} class={cls} constraints={status} sources={src}"
        )

    extracted_rows = [
        row
        for row in rows
        if normalize_key(row.values.get("source_type")) == "extracted"
        and row_selected_for_output(row, scenario, stage, corner)
        and not is_apply_approved(row)
    ]
    lines.extend(["", "  Extracted rows not emitted for this output:"])
    if not extracted_rows:
        lines.append("    <none>")
    else:
        for row in extracted_rows[:100]:
            values = row.values
            reason = []
            if normalize_key(values.get("apply")) != "yes":
                reason.append("apply!=yes")
            if normalize_key(values.get("review_status")) != "approved":
                reason.append("review_status!=approved")
            if normalize_key(values.get("source_type")) == "na":
                reason.append("source_type=na")
            lines.append(
                "    row {row}: {src}:{line} {ctype} pad={pad} reason={reason}".format(
                    row=row.row_idx,
                    src=clean_cell(values.get("source_sdc_file")) or "-",
                    line=clean_cell(values.get("source_line")) or "-",
                    ctype=clean_cell(values.get("constraint_type")) or "-",
                    pad=clean_cell(values.get("pad_name")) or "-",
                    reason=",".join(reason) or "not approved",
                )
            )
        if len(extracted_rows) > 100:
            lines.append(f"    ... truncated {len(extracted_rows) - 100} additional row(s)")
    return lines


def has_explanation(row: FormRow) -> bool:
    return bool(clean_cell(row.values.get("basis")) or clean_cell(row.values.get("note")))


def has_structured_extra_option_overlap(values: Dict[str, object]) -> bool:
    extra = clean_cell(values.get("extra_options"))
    if not extra:
        return False
    for key in (
        "delay_edge",
        "delay_polarity",
        "add_delay",
        "drive_lib_cell",
        "drive_pin",
        "drive_from_pin",
        "drive_input_transition_rise",
        "drive_input_transition_fall",
    ):
        if clean_cell(values.get(key)) and key in extra:
            return True
    return False


def check_source_digest(row: FormRow, current_digests: Dict[str, str], report: Report) -> None:
    source = clean_cell(row.values.get("source_sdc_file"))
    stored = clean_cell(row.values.get("source_digest"))
    if not source or not stored:
        return
    current = current_digests.get(source)
    if current and current != stored:
        report.warn(
            f"io_constraints row {row.row_idx} {clean_cell(row.values.get('pad_name'))}: "
            f"source_digest mismatch for {source}; row may be stale"
        )


def check_units(row: FormRow, expected_time_unit: str, expected_cap_unit: str, report: Report) -> None:
    unit_time = clean_cell(row.values.get("unit_time"))
    unit_cap = clean_cell(row.values.get("unit_cap"))
    pad = clean_cell(row.values.get("pad_name"))
    if expected_time_unit:
        if not unit_time:
            report.warn(f"io_constraints row {row.row_idx} {pad}: unit_time is blank; expected {expected_time_unit}")
        elif normalize_key(unit_time) != normalize_key(expected_time_unit):
            report.warn(
                f"io_constraints row {row.row_idx} {pad}: unit_time={unit_time} differs from expected {expected_time_unit}"
            )
    if expected_cap_unit:
        if not unit_cap:
            report.warn(f"io_constraints row {row.row_idx} {pad}: unit_cap is blank; expected {expected_cap_unit}")
        elif normalize_key(unit_cap) != normalize_key(expected_cap_unit):
            report.warn(
                f"io_constraints row {row.row_idx} {pad}: unit_cap={unit_cap} differs from expected {expected_cap_unit}"
            )


def pending_line_key(line: str) -> Optional[Tuple[str, str]]:
    stripped = line.strip()
    if not stripped or stripped.startswith("#"):
        return None
    parts = stripped.split()
    if len(parts) < 2:
        return None
    direction, port = parts[0], parts[1]
    if direction not in {"input", "output", "inout"}:
        return None
    return direction, port


def removed_line_key(line: str) -> Optional[PortKey]:
    stripped = line.strip()
    if not stripped or stripped.startswith("#"):
        return None
    parts = stripped.split()
    if len(parts) < 3:
        return None
    inst_name, direction, port = parts[:3]
    if direction not in {"input", "output", "inout"}:
        return None
    return PortKey(inst_name, direction, port)


def read_removed_keys(log_dir: Path) -> Set[PortKey]:
    keys: Set[PortKey] = set()
    if not log_dir.is_dir():
        return keys
    for path in sorted(log_dir.glob("*.removed")):
        try:
            lines = path.read_text(encoding="utf-8").splitlines()
        except OSError:
            continue
        for line in lines:
            key = removed_line_key(line)
            if key is not None:
                keys.add(key)
    return keys


def is_approved_na(row: FormRow) -> bool:
    return (
        normalize_key(row.values.get("apply")) == "yes"
        and normalize_key(row.values.get("review_status")) == "approved"
        and normalize_key(row.values.get("source_type")) == "na"
    )


def removable_pad_records_for_04(
    rows: Sequence[FormRow],
    pads: Sequence[PadRecord],
    scenario: str,
    stage: str,
    corner: str,
) -> List[PadRecord]:
    confirmed_pads: Set[str] = set()
    for row in rows:
        if not row_selected_for_assembled(row, scenario, stage, corner):
            continue
        if not (is_apply_approved(row) or is_approved_na(row)):
            continue
        pad = clean_cell(row.values.get("pad_name"))
        ctype = normalize_key(row.values.get("constraint_type"))
        if pad and (ctype in set(SUPPORTED_COMMANDS.values()) or is_approved_na(row)):
            confirmed_pads.add(pad)

    result: List[PadRecord] = []
    seen: Set[Tuple[str, str, str]] = set()
    for pad in pads:
        key = (pad.subsys_instance, pad.direction, pad.subsys_port)
        if pad.pad_name in confirmed_pads and key not in seen:
            seen.add(key)
            result.append(pad)
    return result


def removed_log_line_04(pad: PadRecord, scenario: str, stage: str, corner: str) -> str:
    return " ".join(
        [
            pad.subsys_instance,
            pad.direction,
            pad.subsys_port,
            "covered_by=04_soc_io_pads",
            "reason=pad_environment",
            f"pad={pad.pad_name}",
            f"scenario={scenario}",
            f"stage={stage}",
            f"corner={corner}",
        ]
    )


def update_pending_for_04(
    cwd: Path,
    pending_dir: Path,
    log_dir: Path,
    rows: Sequence[FormRow],
    pads: Sequence[PadRecord],
    scenario: str,
    stage: str,
    corner: str,
    report: Report,
) -> None:
    if not pending_dir.exists():
        return
    if not pending_dir.is_dir():
        report.error(f"{pending_dir}: pending path exists but is not a directory")
        return

    removable = removable_pad_records_for_04(rows, pads, scenario, stage, corner)
    if not removable:
        return

    previous_removed = read_removed_keys(log_dir)
    by_inst: Dict[str, List[PadRecord]] = defaultdict(list)
    for pad in removable:
        by_inst[pad.subsys_instance].append(pad)

    removed_records: List[PadRecord] = []
    for inst_name, inst_pads in sorted(by_inst.items()):
        pending_file = pending_dir / f"{inst_name}.ports"
        if not pending_file.is_file():
            for pad in inst_pads:
                key = PortKey(pad.subsys_instance, pad.direction, pad.subsys_port)
                if key in previous_removed:
                    continue
                report.error(
                    f"{pending_file}: missing pending file for 04 pad port "
                    f"{pad.subsys_instance}/{pad.subsys_port}"
                )
            continue

        lines = pending_file.read_text(encoding="utf-8").splitlines()
        index: Dict[Tuple[str, str], int] = {}
        duplicate_keys: Set[Tuple[str, str]] = set()
        for idx, line in enumerate(lines):
            key = pending_line_key(line)
            if key is None:
                continue
            if key in index:
                duplicate_keys.add(key)
            else:
                index[key] = idx
        for direction, port in sorted(duplicate_keys):
            report.error(f"{pending_file}: duplicate pending port line {direction} {port}")

        remove_line_indices: Set[int] = set()
        for pad in inst_pads:
            key = (pad.direction, pad.subsys_port)
            port_key = PortKey(pad.subsys_instance, pad.direction, pad.subsys_port)
            if key not in index:
                if port_key in previous_removed:
                    continue
                report.error(
                    f"{pending_file}: 04 wants to remove {pad.direction} {pad.subsys_port}, "
                    "but it is not present in pending and no previous_removed record exists"
                )
                continue
            remove_line_indices.add(index[key])
            removed_records.append(pad)

        if remove_line_indices:
            kept = [line for idx, line in enumerate(lines) if idx not in remove_line_indices]
            atomic_write_text(pending_file, "\n".join(kept).rstrip() + ("\n" if kept else ""))

    if removed_records:
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / "04_soc_io_pads.removed"
        existing_lines = log_path.read_text(encoding="utf-8").splitlines() if log_path.is_file() else []
        existing_keys = {
            key
            for key in (removed_line_key(line) for line in existing_lines)
            if key is not None
        }
        new_lines = []
        for pad in sorted(removed_records, key=lambda item: (item.subsys_instance, item.direction, item.subsys_port)):
            key = PortKey(pad.subsys_instance, pad.direction, pad.subsys_port)
            if key not in existing_keys:
                new_lines.append(removed_log_line_04(pad, scenario, stage, corner))
                existing_keys.add(key)
        if new_lines:
            log_lines = [line for line in existing_lines if line.strip()] + new_lines
            atomic_write_text(log_path, "\n".join(log_lines).rstrip() + "\n")
        try:
            display_path = log_path.relative_to(cwd)
        except ValueError:
            display_path = log_path
        report.info(f"removed {len(removed_records)} harden pad port(s) from pending; log={display_path}")


def write_report(
    path: Path,
    report: Report,
    scenario: str,
    stage: str,
    corner: str,
    tool: str,
    form_path: Path,
    output_path: Path,
    coverage_lines: Sequence[str],
    completeness: RunCompleteness,
    clock_inventory_paths: Sequence[Path],
    manifest_path: Optional[Path],
    connection_path: Optional[Path],
    pending_update_enabled: bool,
) -> None:
    lines = [
        "04_soc_io_pads extraction report",
        "================================",
        "",
        f"Author  : {author_name()}",
        "Stage   : 04_soc_io_pads",
        "Script  : 04_extract_soc_io_pads.py",
        f"Scenario: {scenario}",
        f"Stage   : {stage}",
        f"Corner  : {corner}",
        f"Tool    : {tool}",
        f"Form    : {form_path}",
        f"Output  : {output_path}",
        f"Clock inventory: {', '.join(str(path) for path in clock_inventory_paths)}",
        f"Harden SDC manifest: {manifest_path or '<legacy inference>'}",
        f"Connection inventory: {connection_path or '<not used>'}",
        f"Pending update: {'enabled' if pending_update_enabled else 'disabled'}",
        f"Run completeness: {completeness.status}",
        f"Available harden SDC: {completeness.available_count}",
        f"Missing harden SDC: {completeness.missing_count}",
        f"Not-required harden SDC: {completeness.not_required_count}",
        f"Missing instances: {','.join(completeness.missing_instances) or '<none>'}",
        f"Warnings: {report.warning_count}",
        f"Errors  : {report.error_count}",
        f"Sync changed: {'yes' if report.sync_changed else 'no'}",
        "",
        "Messages:",
    ]
    lines.extend(report.lines or ["INFO: no messages"])
    lines.extend(coverage_lines)
    atomic_write_text(path, "\n".join(lines).rstrip() + "\n")


FLAT_PAD_SCHEMA_VERSION = "1.0"
FLAT_SIGNAL_RE = re.compile(
    r"^(?P<base>[A-Za-z_][A-Za-z0-9_$]*)(?:\[(?P<left>-?\d+)(?::(?P<right>-?\d+))?\])?$"
)


def _flat_json_digest(value: object) -> str:
    payload = json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _flat_ordered_bits(shape: Any) -> List[int]:
    if shape.scalar:
        return [0]
    step = 1 if shape.right >= shape.left else -1
    return list(range(shape.left, shape.right + step, step))


def _flat_parse_signal(value: object, location: str) -> Tuple[str, Optional[List[int]]]:
    text = clean_cell(value)
    match = FLAT_SIGNAL_RE.fullmatch(text)
    if not match:
        raise ValueError(f"{location}: invalid endpoint signal {text!r}")
    base = match.group("base")
    left = match.group("left")
    right = match.group("right")
    if left is None:
        return base, None
    left_value = int(left)
    right_value = int(right) if right is not None else left_value
    step = 1 if right_value >= left_value else -1
    return base, list(range(left_value, right_value + step, step))


def _flat_exact_port(base: str, bit: int, scalar: bool = False) -> str:
    return base if scalar and bit == 0 else f"{base}[{bit}]"


def _flat_collection(instance: str, port: str) -> str:
    if normalize_key(instance) == "top":
        return f"[get_ports {{{port}}}]"
    return f"[get_pins {{{instance}/{port}}}]"


def _flat_connection_identity(
    src_instance: str,
    src_direction: str,
    src_base: str,
    src_bit: int,
    dst_instance: str,
    dst_direction: str,
    dst_base: str,
    dst_bit: int,
) -> Tuple[str, str]:
    canonical = [
        FLAT_PAD_SCHEMA_VERSION,
        clean_cell(src_instance),
        normalize_key(src_direction),
        clean_cell(src_base),
        int(src_bit),
        clean_cell(dst_instance),
        normalize_key(dst_direction),
        clean_cell(dst_base),
        int(dst_bit),
    ]
    digest = _flat_json_digest(canonical)
    return "CONN_" + digest, digest


def _flat_structural_reason(runtime: Any, record: Any, report: Report) -> str:
    reason = clean_cell(getattr(record, "structural_reason", ""))
    if reason or hasattr(record, "structural_status"):
        return reason
    legacy = getattr(runtime, "structural_reason", None)
    if callable(legacy):
        return clean_cell(
            legacy(
                clean_cell(record.connection_value),
                record.shape,
                record.location(),
                report,
            )
        )
    parser = getattr(runtime, "parse_structural_token", None)
    if callable(parser):
        _, reason, diagnostic = parser(record.connection_value, record.shape)
        if diagnostic:
            report.error(f"{record.location()}: {diagnostic}")
        return clean_cell(reason)
    return ""


def _build_flat_exact_pads(
    runtime: Any,
    records: Sequence[Any],
    instances: Dict[str, InstInfo],
    report: Report,
) -> List[PadRecord]:
    pads: List[PadRecord] = []
    seen: Set[Tuple[str, str, str, str]] = set()

    def append_pad(record: Any, top_port: str, harden_port: str, direction: str, connection_id: str, connection_type: str) -> None:
        key = (top_port, record.inst_name, harden_port, direction)
        if key in seen:
            report.error(f"duplicate exact pad mapping: {top_port} -> {record.inst_name}/{harden_port}")
            return
        seen.add(key)
        inst = instances.get(record.inst_name)
        pads.append(
            PadRecord(
                pad_name=top_port,
                soc_top_port=top_port,
                subsys_instance=record.inst_name,
                subsys_port=harden_port,
                direction=direction,
                is_gpio_or_inout="yes" if record.direction == "inout" else "no",
                source_sdc_status=inst.sdc_status if inst else "",
                connection_id=connection_id,
                connection_type=connection_type,
                source_workbook=record.workbook,
                source_sheet=record.sheet,
                source_row=str(record.row),
            )
        )

    for record in records:
        connection = clean_cell(record.connection_value)
        if not connection or _flat_structural_reason(runtime, record, report):
            continue
        if record.direction in {"input", "inout"}:
            if not connection.startswith("top."):
                continue
            try:
                top_base, selected_top_bits = _flat_parse_signal(connection[4:], record.location())
            except ValueError as exc:
                report.error(str(exc))
                continue
            harden_bits = _flat_ordered_bits(record.shape)
            top_bits = selected_top_bits if selected_top_bits is not None else list(harden_bits)
            if len(top_bits) != len(harden_bits):
                report.error(f"{record.location()}: top pad width does not match harden port width")
                continue
            for top_bit, harden_bit in zip(top_bits, harden_bits):
                top_port = _flat_exact_port(top_base, top_bit, len(top_bits) == 1 and selected_top_bits is None)
                harden_port = _flat_exact_port(record.shape.base, harden_bit, record.shape.scalar)
                connection_id, _ = _flat_connection_identity(
                    "top", "input", top_base, top_bit,
                    record.inst_name, record.direction, record.shape.base, harden_bit,
                )
                append_pad(record, top_port, harden_port, record.direction, connection_id, "top_pad_to_harden")
        elif record.direction == "output":
            top_signal = connection[4:] if connection.startswith("top.") else connection
            if connection.lower() in {"y", "yes", "true"}:
                top_signal = record.shape.raw
            if "." in top_signal:
                continue
            try:
                top_base, selected_top_bits = _flat_parse_signal(top_signal, record.location())
            except ValueError as exc:
                report.error(str(exc))
                continue
            harden_bits = _flat_ordered_bits(record.shape)
            top_bits = selected_top_bits if selected_top_bits is not None else list(harden_bits)
            if len(top_bits) != len(harden_bits):
                report.error(f"{record.location()}: top pad width does not match harden port width")
                continue
            for harden_bit, top_bit in zip(harden_bits, top_bits):
                harden_port = _flat_exact_port(record.shape.base, harden_bit, record.shape.scalar)
                top_port = _flat_exact_port(top_base, top_bit, len(top_bits) == 1 and selected_top_bits is None)
                connection_id, _ = _flat_connection_identity(
                    record.inst_name, "output", record.shape.base, harden_bit,
                    "top", "output", top_base, top_bit,
                )
                append_pad(record, top_port, harden_port, "output", connection_id, "harden_to_top_pad")
    pads.sort(key=lambda item: (item.pad_name, item.subsys_instance, item.subsys_port, item.direction))
    report.info(f"built {len(pads)} exact flat pad owner record(s) from port workbooks")
    return pads


def _without_rewritten_command(value: object) -> str:
    return "; ".join(
        part.strip()
        for part in clean_cell(value).split(";")
        if part.strip() and not part.strip().startswith("rewritten_command=")
    )


def _expand_flat_extracted_constraints(
    extracted: Sequence[ExtractedConstraint],
    pads: Sequence[PadRecord],
    report: Report,
) -> List[ExtractedConstraint]:
    exact_keys = {
        (pad.pad_name, pad.subsys_instance, pad.subsys_port)
        for pad in pads
    }
    expanded: List[ExtractedConstraint] = []
    expanded_sources = 0
    expanded_rows = 0
    for item in extracted:
        values = item.values
        key = (
            clean_cell(values.get("pad_name")),
            clean_cell(values.get("subsys_instance")),
            clean_cell(values.get("subsys_port")),
        )
        if key in exact_keys:
            expanded.append(item)
            continue
        try:
            harden_base, selected_harden_bits = _flat_parse_signal(
                key[2], "flat IO harden selector"
            )
        except ValueError:
            expanded.append(item)
            continue
        base_candidates: List[Tuple[PadRecord, Optional[int]]] = []
        for pad in pads:
            if pad.subsys_instance != key[1]:
                continue
            try:
                candidate_base, candidate_bits = _flat_parse_signal(
                    pad.subsys_port, "flat IO exact harden endpoint"
                )
            except ValueError:
                continue
            if candidate_base != harden_base:
                continue
            candidate_bit = candidate_bits[0] if candidate_bits else None
            base_candidates.append((pad, candidate_bit))
        if not base_candidates:
            expanded.append(item)
            continue

        invalid_reason = ""
        matches: List[PadRecord] = []
        if selected_harden_bits is None:
            scalar_matches = [pad for pad, bit in base_candidates if bit is None]
            if len(scalar_matches) == 1:
                matches = scalar_matches
            else:
                invalid_reason = (
                    f"bare vector selector {key[2]} is ambiguous; use an explicit exact bit/range"
                )
        else:
            for bit in selected_harden_bits:
                bit_matches = [pad for pad, candidate_bit in base_candidates if candidate_bit == bit]
                if len(bit_matches) != 1:
                    invalid_reason = (
                        f"selector {key[2]} bit {bit} maps to {len(bit_matches)} canonical pad edge(s)"
                    )
                    break
                matches.append(bit_matches[0])

        if not invalid_reason and key[0]:
            try:
                top_base, selected_top_bits = _flat_parse_signal(
                    key[0], "flat IO top selector"
                )
                mapped_top: List[Tuple[str, Optional[int]]] = []
                for pad in matches:
                    mapped_base, mapped_bits = _flat_parse_signal(
                        pad.pad_name, "flat IO exact top endpoint"
                    )
                    mapped_top.append(
                        (mapped_base, mapped_bits[0] if mapped_bits else None)
                    )
                if any(mapped_base != top_base for mapped_base, _ in mapped_top):
                    invalid_reason = f"top selector {key[0]} does not match canonical pad base"
                elif selected_top_bits is None:
                    if len(mapped_top) != 1 or mapped_top[0][1] is not None:
                        invalid_reason = (
                            f"bare vector top selector {key[0]} is ambiguous; use an explicit exact bit/range"
                        )
                elif [bit for _, bit in mapped_top] != selected_top_bits:
                    invalid_reason = (
                        f"top/harden range order mismatch for {key[0]} and {key[2]}"
                    )
            except ValueError as exc:
                invalid_reason = str(exc)

        if invalid_reason:
            message = f"flat exact-bit expansion failed: {invalid_reason}"
            report.error(message)
            invalid_values = dict(values)
            invalid_values["note"] = "; ".join(
                filter(None, [clean_cell(values.get("note")), message])
            )
            expanded.append(
                ExtractedConstraint(
                    values=invalid_values,
                    parse_status="invalid",
                    mapped_soc_object=item.mapped_soc_object,
                    message="; ".join(filter(None, [item.message, message])),
                    include_in_form=False,
                )
            )
            continue

        expanded_sources += 1
        seen_top_ports: Set[str] = set()
        for pad in matches:
            if pad.pad_name in seen_top_ports:
                continue
            seen_top_ports.add(pad.pad_name)
            exact_values = dict(values)
            exact_values.update(
                {
                    "pad_name": pad.pad_name,
                    "soc_object": _flat_collection("top", pad.pad_name),
                    "subsys_instance": pad.subsys_instance,
                    "subsys_port": pad.subsys_port,
                    "direction": pad.direction,
                    "object_granularity": "single_pad",
                    "extra_options": _without_rewritten_command(values.get("extra_options")),
                }
            )
            expansion_note = (
                f"exact-bit expansion from {key[0]} -> {key[1]}/{key[2]}"
            )
            exact_values["note"] = "; ".join(
                filter(None, [clean_cell(values.get("note")), expansion_note])
            )
            exact_object = _flat_collection("top", pad.pad_name)
            expanded.append(
                ExtractedConstraint(
                    values=exact_values,
                    parse_status=item.parse_status,
                    mapped_soc_object=exact_object,
                    message="; ".join(filter(None, [item.message, expansion_note])),
                    include_in_form=True,
                )
            )
            expanded_rows += 1
    if expanded_sources:
        report.info(
            f"expanded {expanded_sources} range IO evidence row(s) into "
            f"{expanded_rows} exact pad-bit review row(s)"
        )
    return expanded


def _flat_pad_machine_values(
    pad: PadRecord,
    run_context: Dict[str, str],
    completeness: RunCompleteness,
    structure: str,
    accounting_before: str,
    view_id: str,
    stage: str,
    corner: str,
    source_digest: str,
) -> Dict[str, object]:
    top_match = re.fullmatch(r"(?P<base>[^\[\]]+)(?:\[(?P<bit>-?\d+)\])?", pad.soc_top_port)
    harden_match = re.fullmatch(r"(?P<base>[^\[\]]+)(?:\[(?P<bit>-?\d+)\])?", pad.subsys_port)
    if not top_match or not harden_match:
        raise RuntimeError(f"non-exact pad identity: {pad.soc_top_port} -> {pad.subsys_instance}/{pad.subsys_port}")
    top_base = top_match.group("base")
    harden_base = harden_match.group("base")
    top_bit = int(top_match.group("bit") or "0")
    harden_bit = int(harden_match.group("bit") or "0")
    direction = normalize_key(pad.direction)
    if direction == "output":
        src_instance, src_direction, src_port, src_bit = pad.subsys_instance, "output", pad.subsys_port, harden_bit
        dst_instance, dst_direction, dst_port, dst_bit = "top", "output", pad.soc_top_port, top_bit
    else:
        src_instance, src_direction, src_port, src_bit = "top", "input", pad.soc_top_port, top_bit
        dst_instance, dst_direction, dst_port, dst_bit = pad.subsys_instance, direction, pad.subsys_port, harden_bit
    connection_id, pad_id = _flat_connection_identity(
        src_instance, src_direction, top_base if src_instance == "top" else harden_base, src_bit,
        dst_instance, dst_direction, top_base if dst_instance == "top" else harden_base, dst_bit,
    )
    values: Dict[str, object] = {
        "schema_version": FLAT_PAD_SCHEMA_VERSION,
        "run_id": run_context.get("run_id", ""),
        "mode_label": run_context.get("mode_label", ""),
        "design_revision": run_context.get("design_revision", ""),
        "run_completeness": completeness.status,
        "structure_digest": structure,
        "accounting_digest_before": accounting_before,
        "accounting_digest_after": accounting_before,
        "pad_id": pad_id,
        "view_id": view_id,
        "stage": stage,
        "corner": corner,
        "pad_name": pad.pad_name,
        "soc_top_port": pad.soc_top_port,
        "top_port": pad.soc_top_port,
        "top_bit_index": str(top_bit),
        "top_endpoint": _flat_collection("top", pad.soc_top_port),
        "subsys_instance": pad.subsys_instance,
        "harden_instance": pad.subsys_instance,
        "subsys_port": pad.subsys_port,
        "harden_port": pad.subsys_port,
        "harden_bit_index": str(harden_bit),
        "harden_endpoint": _flat_collection(pad.subsys_instance, pad.subsys_port),
        "direction_from_integration": pad.direction,
        "direction": pad.direction,
        "effective_direction": direction if direction in {"input", "output"} else "",
        "src_instance": src_instance,
        "src_direction": src_direction,
        "src_port": src_port,
        "src_bit_index": str(src_bit),
        "src_soc_object": src_port if src_instance == "top" else f"{src_instance}/{src_port}",
        "src_endpoint": _flat_collection(src_instance, src_port),
        "dst_instance": dst_instance,
        "dst_direction": dst_direction,
        "dst_port": dst_port,
        "dst_bit_index": str(dst_bit),
        "dst_soc_object": dst_port if dst_instance == "top" else f"{dst_instance}/{dst_port}",
        "dst_endpoint": _flat_collection(dst_instance, dst_port),
        "is_gpio_or_inout": pad.is_gpio_or_inout,
        "related_scenarios": pad.related_scenarios,
        "source_sdc_status": pad.source_sdc_status,
        "sdc_status": pad.source_sdc_status,
        "connection_id": connection_id,
        "connection_status": "matched",
        "scenario_scope": pad.scenario_scope,
        "connection_type": pad.connection_type,
        "source_workbook": pad.source_workbook,
        "source_sheet": pad.source_sheet,
        "source_row": pad.source_row,
        "source_digest": source_digest,
        "note": pad.note,
    }
    digest_fields = [
        name for name in PAD_HEADERS
        if name not in {
            "accounting_digest_before", "accounting_digest_after",
            "pad_disposition", "timing_active", "coverage_status",
            "apply", "review_status", "owner", "basis", "related_exception_intent",
            "reviewer", "review_date", "effective_direction",
            "machine_digest", "approved_machine_digest", "note",
        }
    ]
    values["machine_digest"] = _flat_json_digest([[name, clean_cell(values.get(name))] for name in digest_fields])
    return values


def _sync_flat_pad_machine_fields(
    path: Path,
    pads: Sequence[PadRecord],
    run_context: Dict[str, str],
    completeness: RunCompleteness,
    structure: str,
    accounting_before: str,
    view_id: str,
    stage: str,
    corner: str,
    report: Report,
) -> None:
    workbook = load_workbook(path)
    if "pad_inventory" not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError(f"{path} missing pad_inventory sheet")
    sheet = workbook["pad_inventory"]
    mapping = header_map(sheet)

    machine_rows: List[Dict[str, object]] = []
    for pad in pads:
        source_digest = _flat_json_digest(
            [
                FLAT_PAD_SCHEMA_VERSION,
                pad.source_workbook,
                pad.source_sheet,
                pad.source_row,
                pad.direction,
                pad.soc_top_port,
                pad.subsys_instance,
                pad.subsys_port,
                pad.connection_id,
            ]
        )
        machine_rows.append(
            _flat_pad_machine_values(
                pad, run_context, completeness, structure, accounting_before,
                view_id, stage, corner, source_digest,
            )
        )

    expected_keys = {pad_key(machine) for machine in machine_rows}
    stale_rows: List[int] = []
    for row_idx in range(2, sheet.max_row + 1):
        values = row_values(sheet, row_idx, PAD_HEADERS)
        if clean_cell(values.get("view_id")) != view_id:
            continue
        if clean_cell(values.get("pad_name")) and pad_key(values) not in expected_keys:
            stale_rows.append(row_idx)
    for row_idx in reversed(stale_rows):
        sheet.delete_rows(row_idx, 1)
    changed = bool(stale_rows)
    if stale_rows:
        report.warn(
            f"removed {len(stale_rows)} obsolete pad_inventory row(s) from current view {view_id}"
        )

    rows_by_key: Dict[Tuple[str, str, str, str, str], int] = {}
    for row_idx in range(2, sheet.max_row + 1):
        values = row_values(sheet, row_idx, PAD_HEADERS)
        if not clean_cell(values.get("pad_name")):
            continue
        key = pad_key(values)
        if key in rows_by_key:
            report.error(f"pad_inventory duplicates exact owner row {key}")
        else:
            rows_by_key[key] = row_idx

    review_headers = {
        "pad_disposition", "apply", "review_status", "owner", "basis",
        "related_exception_intent", "reviewer", "review_date", "approved_machine_digest",
        "effective_direction", "note",
    }
    machine_headers = [name for name in PAD_HEADERS if name not in review_headers]
    for machine in machine_rows:
        key = pad_key(machine)
        row_idx = rows_by_key.get(key)
        if row_idx is None:
            report.error(f"pad_inventory machine row missing after synchronization: {key}")
            continue
        old_digest = clean_cell(sheet.cell(row_idx, mapping["machine_digest"]).value)
        machine_changed = bool(old_digest and old_digest != clean_cell(machine.get("machine_digest")))
        for name in machine_headers:
            old_value = clean_cell(sheet.cell(row_idx, mapping[name]).value)
            new_value = clean_cell(machine.get(name))
            if old_value != new_value:
                sheet.cell(row_idx, mapping[name], machine.get(name, ""))
                changed = True
        if machine_changed:
            for name, value in (
                ("apply", "no"),
                ("review_status", "pending"),
                ("approved_machine_digest", ""),
            ):
                if clean_cell(sheet.cell(row_idx, mapping[name]).value) != value:
                    sheet.cell(row_idx, mapping[name], value)
                    changed = True
            report.warn(f"pad_inventory {machine['pad_id']}: machine identity changed; approval reset")

    if changed:
        atomic_save_workbook(workbook, path)
        report.sync_changed = True
        report.info("synchronized canonical pad machine fields; review is required")
    else:
        workbook.close()


def _active_io_rows_by_pad(
    rows: Sequence[FormRow], stage: str, corner: str, tool: str
) -> Dict[str, List[FormRow]]:
    result: Dict[str, List[FormRow]] = defaultdict(list)
    for row in rows:
        if not row_selected_for_assembled(row, "common", stage, corner):
            continue
        if not is_apply_approved(row):
            continue
        if not commands_for_row(row, tool, True):
            continue
        values = row.values
        pad_name = clean_cell(values.get("pad_name"))
        if pad_name:
            # The emitted command targets the exact top port/bit, so one
            # reviewed row covers every canonical direct edge in its fanout.
            result[pad_name].append(row)
    return result


def _approved_na_rows_by_pad(
    rows: Sequence[FormRow], stage: str, corner: str
) -> Dict[Tuple[str, str, str], List[FormRow]]:
    result: Dict[Tuple[str, str, str], List[FormRow]] = defaultdict(list)
    for row in rows:
        if not row_selected_for_assembled(row, "common", stage, corner):
            continue
        if not is_approved_na(row):
            continue
        values = row.values
        result[
            (
                clean_cell(values.get("pad_name")),
                clean_cell(values.get("subsys_instance")),
                clean_cell(values.get("subsys_port")),
            )
        ].append(row)
    return result


def _resolved_flat_pad_rows(
    form_path: Path,
    io_rows: Sequence[FormRow],
    stage: str,
    corner: str,
    tool: str,
    accounting_after: str,
    report: Report,
) -> List[Dict[str, object]]:
    workbook = load_workbook(form_path, data_only=False)
    sheet = workbook["pad_inventory"]
    active_by_pad = _active_io_rows_by_pad(io_rows, stage, corner, tool)
    na_by_pad = _approved_na_rows_by_pad(io_rows, stage, corner)
    resolved: List[Dict[str, object]] = []
    seen: Set[Tuple[str, str]] = set()
    for row_idx in range(2, sheet.max_row + 1):
        values = row_values(sheet, row_idx, PAD_HEADERS)
        if not clean_cell(values.get("pad_name")):
            continue
        if clean_cell(values.get("stage")) != stage or clean_cell(values.get("corner")) != corner:
            continue
        pad_id = clean_cell(values.get("pad_id"))
        view_id = clean_cell(values.get("view_id"))
        identity = (pad_id, view_id)
        if not pad_id or not re.fullmatch(r"[0-9a-f]{64}", pad_id):
            report.error(f"pad_inventory row {row_idx}: pad_id must be a full SHA-256")
        if identity in seen:
            report.error(f"pad_inventory row {row_idx}: duplicate pad_id/view_id {identity}")
        seen.add(identity)
        key = (
            clean_cell(values.get("pad_name")),
            clean_cell(values.get("subsys_instance")),
            clean_cell(values.get("subsys_port")),
        )
        active_rows = active_by_pad.get(key[0], [])
        normal_timing_rows = [
            row for row in active_rows
            if normalize_key(row.values.get("constraint_type")) in DELAY_TYPES | {"false_path"}
        ]
        na_rows = na_by_pad.get(key, [])
        timing_active = bool(active_rows)
        disposition = normalize_key(values.get("pad_disposition"))
        apply_value = normalize_key(values.get("apply"))
        review_status = normalize_key(values.get("review_status"))
        integration_direction = normalize_key(
            values.get("direction_from_integration") or values.get("direction")
        )
        effective_direction = normalize_key(values.get("effective_direction"))
        if integration_direction in {"input", "output"}:
            values["effective_direction"] = integration_direction
        elif integration_direction == "inout" and timing_active and disposition != "route_to_30":
            active_directions = {
                normalize_key(row.values.get("direction"))
                for row in active_rows
                if normalize_key(row.values.get("direction"))
            }
            if len(active_directions) != 1 or not active_directions <= {"input", "output"}:
                report.error(
                    f"pad_inventory row {row_idx}: active inout IO commands must resolve to "
                    "one effective input/output direction"
                )
            else:
                values["effective_direction"] = next(iter(active_directions))
        if disposition and disposition not in {"constrained", "not_applicable", "route_to_30", "pending"}:
            report.error(f"pad_inventory row {row_idx}: invalid pad_disposition {disposition}")

        if disposition == "route_to_30":
            if apply_value != "yes" or review_status != "approved":
                report.error(f"pad_inventory row {row_idx}: route_to_30 requires apply=yes and review_status=approved")
            if not clean_cell(values.get("owner")) or not clean_cell(values.get("basis")):
                report.error(f"pad_inventory row {row_idx}: route_to_30 requires owner and basis")
            if not clean_cell(values.get("reviewer")) or not clean_cell(values.get("review_date")):
                report.error(f"pad_inventory row {row_idx}: route_to_30 requires reviewer and review_date")
            if not clean_cell(values.get("related_exception_intent")):
                report.error(f"pad_inventory row {row_idx}: route_to_30 requires related_exception_intent")
            if clean_cell(values.get("approved_machine_digest")) != clean_cell(values.get("machine_digest")):
                report.error(f"pad_inventory row {row_idx}: route_to_30 approval is stale")
            if integration_direction == "inout":
                if effective_direction not in {"input", "output"}:
                    report.error(
                        f"pad_inventory row {row_idx}: inout route_to_30 requires "
                        "effective_direction=input or output"
                    )
                elif effective_direction == "output":
                    report.error(
                        f"pad_inventory row {row_idx}: inout route_to_30 effective_direction=output "
                        "is not supported by the current 30 oriented edge contract"
                    )
                else:
                    values["effective_direction"] = "input"
                route_active_directions = {
                    normalize_key(row.values.get("direction"))
                    for row in active_rows
                    if normalize_key(row.values.get("direction"))
                }
                if route_active_directions and route_active_directions != {effective_direction}:
                    report.error(
                        f"pad_inventory row {row_idx}: inout route_to_30 effective_direction "
                        "conflicts with active 04 electrical direction"
                    )
            if normal_timing_rows:
                report.error(
                    f"pad_inventory row {row_idx}: route_to_30 conflicts with active 04 normal timing"
                )
            values["timing_active"] = "no"
            values["coverage_status"] = "route_to_30"
        elif timing_active:
            first = active_rows[0]
            values["pad_disposition"] = "constrained"
            values["timing_active"] = "yes"
            values["coverage_status"] = "constrained"
            values["apply"] = "yes"
            values["review_status"] = "approved"
            values["owner"] = clean_cell(first.values.get("owner"))
            values["basis"] = clean_cell(first.values.get("basis")) or clean_cell(first.values.get("note"))
            values["reviewer"] = clean_cell(first.values.get("reviewer"))
            values["review_date"] = clean_cell(first.values.get("review_date"))
            values["related_exception_intent"] = ""
        elif na_rows:
            first = na_rows[0]
            values["pad_disposition"] = "not_applicable"
            values["timing_active"] = "no"
            values["coverage_status"] = "not_applicable"
            values["apply"] = "yes"
            values["review_status"] = "approved"
            values["owner"] = clean_cell(first.values.get("owner"))
            values["basis"] = clean_cell(first.values.get("basis")) or clean_cell(first.values.get("note"))
            values["reviewer"] = clean_cell(first.values.get("reviewer"))
            values["review_date"] = clean_cell(first.values.get("review_date"))
            values["related_exception_intent"] = ""
        elif disposition == "not_applicable":
            missing_review = [
                name for name in ("owner", "basis", "reviewer", "review_date")
                if not clean_cell(values.get(name))
            ]
            if apply_value != "yes" or review_status != "approved" or missing_review:
                report.error(
                    f"pad_inventory row {row_idx}: not_applicable requires approved review, "
                    "owner, basis, reviewer, and review_date"
                )
            values["timing_active"] = "no"
            values["coverage_status"] = "not_applicable"
        else:
            if disposition == "constrained":
                report.error(
                    f"pad_inventory row {row_idx}: constrained requires an active approved 04 IO command"
                )
            values["pad_disposition"] = disposition or "pending"
            values["timing_active"] = "no"
            values["coverage_status"] = "pending"
        values["accounting_digest_after"] = accounting_after
        resolved.append(values)
    workbook.close()
    return resolved


def _merge_prior_pad_rows(
    path: Path,
    current_rows: Sequence[Dict[str, object]],
    current_view_id: str,
    allowed_view_ids: Set[str],
    run_context: Dict[str, str],
    structure: str,
    report: Report,
) -> Tuple[List[Dict[str, object]], Set[str]]:
    merged: Dict[Tuple[str, str], Dict[str, object]] = {}
    authenticated_prior_views: Set[str] = set()
    if path.is_file():
        meta_path = path.with_name("pad_inventory.meta")
        try:
            prior_meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except FileNotFoundError:
            report.warn(f"ignoring prior pad inventory without metadata: {meta_path}")
            prior_meta = None
        except (OSError, ValueError, TypeError) as exc:
            report.warn(f"ignoring invalid prior pad inventory metadata {meta_path}: {exc}")
            prior_meta = None

        if prior_meta is not None:
            stale_fields = []
            if str(prior_meta.get("schema_version") or "").strip() != FLAT_PAD_SCHEMA_VERSION:
                stale_fields.append("schema_version")
            for name, expected in (
                ("stage_name", "04_soc_io_pads"),
                ("run_id", clean_cell(run_context.get("run_id"))),
                ("mode_label", clean_cell(run_context.get("mode_label"))),
                ("design_revision", clean_cell(run_context.get("design_revision"))),
                ("structure_digest", structure),
                ("completion_status", "complete"),
                ("sync_changed", "no"),
            ):
                actual = (
                    normalize_key(prior_meta.get(name))
                    if name in {"completion_status", "sync_changed"}
                    else clean_cell(prior_meta.get(name))
                )
                if actual != expected:
                    stale_fields.append(name)
            if clean_cell(prior_meta.get("error_count")) not in {"", "0"}:
                stale_fields.append("error_count")
            if stale_fields:
                report.warn(
                    f"ignoring stale prior pad inventory {path}: {', '.join(stale_fields)}"
                )
            else:
                try:
                    actual_digest = hashlib.sha256(path.read_bytes()).hexdigest()
                except OSError as exc:
                    report.error(f"failed to read prior pad inventory {path}: {exc}")
                    actual_digest = ""
                declared_digests = [
                    clean_cell(prior_meta.get(name))
                    for name in ("inventory_digest", "pad_inventory_digest")
                    if clean_cell(prior_meta.get(name))
                ]
                if not declared_digests or any(
                    declared != actual_digest for declared in declared_digests
                ):
                    report.error(
                        f"prior pad inventory digest does not match {meta_path}"
                    )
                else:
                    try:
                        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
                            reader = csv.DictReader(file_obj)
                            missing_headers = sorted(set(PAD_HEADERS) - set(reader.fieldnames or []))
                            prior_view_values = prior_meta.get("view_ids")
                            if missing_headers or not isinstance(prior_view_values, list):
                                details = (
                                    "missing headers " + ", ".join(missing_headers)
                                    if missing_headers
                                    else "metadata has no canonical view_ids list"
                                )
                                report.warn(
                                    f"ignoring non-canonical prior pad inventory {path}: {details}"
                                )
                                prior_view_ids: Set[str] = set()
                            else:
                                prior_view_ids = {
                                    clean_cell(value) for value in prior_view_values
                                    if clean_cell(value)
                                }
                                authenticated_prior_views = (
                                    prior_view_ids & allowed_view_ids
                                ) - {current_view_id}
                            for row in reader:
                                pad_id = clean_cell(row.get("pad_id"))
                                view_id = clean_cell(row.get("view_id"))
                                if not prior_view_ids:
                                    continue
                                if view_id not in prior_view_ids:
                                    report.error(
                                        f"prior pad inventory row has undeclared view_id {view_id or '<blank>'}"
                                    )
                                    continue
                                if str(row.get("schema_version") or "").strip() != FLAT_PAD_SCHEMA_VERSION:
                                    report.error(
                                        f"prior pad inventory row {pad_id or '<blank>'}/{view_id}: schema_version is stale"
                                    )
                                    continue
                                if clean_cell(row.get("structure_digest")) != structure:
                                    report.error(
                                        f"prior pad inventory row {pad_id or '<blank>'}/{view_id}: structure_digest is stale"
                                    )
                                    continue
                                if not re.fullmatch(r"[0-9a-f]{64}", pad_id):
                                    report.error(
                                        f"prior pad inventory row {pad_id or '<blank>'}/{view_id}: invalid pad_id"
                                    )
                                    continue
                                if view_id == current_view_id:
                                    continue
                                if view_id not in allowed_view_ids:
                                    continue
                                identity = (pad_id, view_id)
                                if identity in merged:
                                    report.error(
                                        f"prior pad inventory duplicates pad_id/view_id {identity}"
                                    )
                                    continue
                                merged[identity] = dict(row)
                    except (OSError, csv.Error) as exc:
                        report.error(f"failed to preserve prior pad inventory views: {exc}")
    for row in current_rows:
        merged[(clean_cell(row.get("pad_id")), clean_cell(row.get("view_id")))] = dict(row)
    return [merged[key] for key in sorted(merged)], authenticated_prior_views


def _load_accounting_runtime():
    """Load the shared 00 workbook/accounting implementation for flat runs."""
    script_root = Path(__file__).resolve().parents[1]
    candidates = [
        script_root / "00_harden_port_inventory" / "00_harden_port_inventory.py",
        script_root / "00_harden_port_inventory.py",
        Path(__file__).resolve().with_name("00_harden_port_inventory.py"),
    ]
    runtime_path = next((path for path in candidates if path.is_file()), None)
    if runtime_path is None:
        raise RuntimeError("shared accounting runtime 00_harden_port_inventory.py was not found")
    spec = importlib.util.spec_from_file_location("soc_sdc_accounting_runtime_for_04", str(runtime_path))
    if spec is None or spec.loader is None:
        raise RuntimeError(f"failed to load accounting runtime: {runtime_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    module.STAGE_NAME = "04_soc_io_pads"
    return module


def _read_flat_manifest(path: Path, run_root: Path, instances: Dict[str, InstInfo], report: Report) -> RunCompleteness:
    available: List[str] = []
    missing: List[str] = []
    not_required: List[str] = []
    if not path.is_file():
        report.error(f"{path}: HARDEN_SDC_MANIFEST_MISSING")
        return RunCompleteness(status="invalid", manifest_path=str(path.resolve()))
    with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        required = {"inst_name", "module_name", "availability_status"}
        missing_fields = sorted(required - set(reader.fieldnames or []))
        if missing_fields:
            report.error(f"{path}: HARDEN_SDC_MANIFEST_SCHEMA_ERROR: missing field(s): {', '.join(missing_fields)}")
        entries: Dict[str, Dict[str, str]] = {}
        for row_idx, row in enumerate(reader, start=2):
            name = clean_cell(row.get("inst_name"))
            if not name:
                report.error(f"{path.name} row {row_idx}: empty inst_name")
                continue
            if name in entries:
                report.error(f"{path.name} row {row_idx}: duplicate inst_name {name}")
                continue
            entries[name] = {key: clean_cell(value) for key, value in row.items()}
    for name, inst in sorted(instances.items()):
        entry = entries.get(name)
        if entry is None:
            inst.sdc_status = "missing"
            missing.append(name)
            report.error(f"{path.name}: manifest missing instance {name}")
            continue
        module_name = entry.get("module_name", "")
        if module_name and module_name != inst.module_name:
            report.error(f"{path.name}: module mismatch for {name}: info_all={inst.module_name} manifest={module_name}")
        status = normalize_key(entry.get("availability_status"))
        inst.sdc_status = status
        inst.sdc_note = entry.get("note", "")
        if status == "available":
            raw_path = entry.get("sdc_path", "") or entry.get("resolved_sdc_path", "")
            if not raw_path:
                report.error(f"{path.name}: available instance {name} has no sdc_path")
                continue
            sdc_path = Path(raw_path).expanduser()
            if not sdc_path.is_absolute():
                sdc_path = run_root / sdc_path
            if not sdc_path.is_file():
                report.error(f"{path.name}: available SDC missing for {name}: {sdc_path}")
            else:
                inst.sdc_path = sdc_path.resolve()
                available.append(name)
                report.info(f"manifest selected {name}: status=available path={inst.sdc_path}")
        elif status == "missing":
            missing.append(name)
            report.warn(f"{path.name}: harden SDC missing for {name}: {inst.sdc_note or '<no note>'}")
        elif status == "not_required":
            not_required.append(name)
        else:
            report.error(f"{path.name}: invalid availability_status for {name}: {status or '<empty>'}")
    for name in sorted(set(entries) - set(instances)):
        report.error(f"{path.name}: orphan manifest instance {name}")
    status = "partial" if missing and report.error_count == 0 else ("invalid" if report.error_count else "complete")
    return RunCompleteness(status=status, available_instances=available, missing_instances=missing, not_required_instances=not_required, manifest_path=str(path.resolve()))


def _flat_owner_updates(
    runtime,
    records: Sequence[Any],
    pad_rows: Sequence[Dict[str, object]],
    report: Report,
) -> List[Dict[str, Any]]:
    updates: Dict[Tuple[str, str, int, str, int], Dict[str, Any]] = {}
    by_key: Dict[Tuple[str, str, str], List[Any]] = defaultdict(list)
    for record in records:
        by_key[(record.inst_name, record.direction, record.shape.base)].append(record)

    for values in pad_rows:
        disposition = normalize_key(values.get("pad_disposition"))
        timing_active = normalize_key(values.get("timing_active"))
        if disposition in {"route_to_30", "pending"}:
            continue
        if disposition == "constrained" and timing_active != "yes":
            continue
        if disposition not in {"constrained", "not_applicable"}:
            continue

        instance = clean_cell(values.get("harden_instance") or values.get("subsys_instance"))
        port = clean_cell(values.get("harden_port") or values.get("subsys_port"))
        direction = normalize_key(values.get("direction")) or normalize_key(values.get("direction_from_integration"))
        match = re.fullmatch(r"(?P<base>[^\[\]]+)(?:\[(?P<bit>-?\d+)\])?", port)
        if not match or not instance:
            report.error(
                f"pad_inventory {clean_cell(values.get('pad_id')) or '<unknown>'}: "
                f"cannot derive accounting port from {instance}/{port}"
            )
            continue
        base = match.group("base")
        bit_text = clean_cell(values.get("harden_bit_index"))
        bit = int(bit_text) if bit_text else (
            int(match.group("bit")) if match.group("bit") is not None else 0
        )
        candidates = by_key.get((instance, direction, base), [])
        matches = [record for record in candidates if record.shape.contains(bit)]
        if len(matches) != 1:
            report.error(
                f"pad_inventory {clean_cell(values.get('pad_id')) or '<unknown>'}: "
                f"accounting port {instance}/{direction}/{port} matches {len(matches)} workbook row(s)"
            )
            continue
        record = matches[0]
        owner = clean_cell(values.get("pad_id"))
        if not owner:
            report.error(
                f"pad_inventory {instance}/{direction}/{port}: canonical pad_id is missing for accounting owner"
            )
            continue
        added = {bit} - set(record.used_bits)
        record.used_bits.add(bit)
        record.added_bits |= added
        if added:
            cell = record.model.workbook[record.sheet].cell(record.row, record.used_col)
            cell.value = runtime.format_bits(record.used_bits)
            cell.number_format = "@"
            record.modified = True
            record.model.modified = True
        update_key = (record.workbook, record.sheet, record.row, record.direction, bit)
        if update_key not in updates:
            updates[update_key] = {
                "record": record,
                "added_bits": added,
                "owner_object_id": owner,
                "reason": (
                    "approved SoC IO/pad constraint"
                    if disposition == "constrained"
                    else "approved NA disposition"
                ),
            }
        else:
            updates[update_key]["added_bits"] |= added
    return [updates[key] for key in sorted(updates)]


def _validate_route_accounting_history(
    runtime: Any,
    pad_rows: Sequence[Dict[str, object]],
    prior_delta_rows: Sequence[Dict[str, object]],
    report: Report,
) -> None:
    normal_owner_ids = {
        clean_cell(row.get("pad_id"))
        for row in pad_rows
        if normalize_key(row.get("pad_disposition")) in {"constrained", "not_applicable"}
        and clean_cell(row.get("pad_id"))
    }
    normal_sources: Dict[Tuple[str, str, str, str], Set[int]] = defaultdict(set)
    for row in pad_rows:
        if normalize_key(row.get("pad_disposition")) not in {"constrained", "not_applicable"}:
            continue
        bit_text = clean_cell(row.get("harden_bit_index"))
        normal_sources[
            (
                clean_cell(row.get("source_workbook")),
                clean_cell(row.get("source_sheet")),
                clean_cell(row.get("source_row")),
                normalize_key(row.get("direction")),
            )
        ].add(int(bit_text) if bit_text else 0)

    route_ids: Set[str] = set()
    route_sources: Dict[Tuple[str, str, str, str], Set[int]] = defaultdict(set)
    for row in pad_rows:
        if normalize_key(row.get("pad_disposition")) != "route_to_30":
            continue
        pad_id = clean_cell(row.get("pad_id"))
        bit_text = clean_cell(row.get("harden_bit_index"))
        bit = int(bit_text) if bit_text else 0
        source_key = (
            clean_cell(row.get("source_workbook")),
            clean_cell(row.get("source_sheet")),
            clean_cell(row.get("source_row")),
            normalize_key(row.get("direction")),
        )
        if pad_id in normal_owner_ids or bit in normal_sources.get(source_key, set()):
            continue
        if pad_id:
            route_ids.add(pad_id)
        route_sources[source_key].add(bit)

    reported: Set[str] = set()
    for delta in prior_delta_rows:
        added_text = clean_cell(delta.get("added_bits"))
        if not added_text:
            continue
        try:
            added_bits = runtime.parse_bits_field(added_text)
        except ValueError as exc:
            report.error(f"invalid prior 04 accounting delta bits: {exc}")
            continue
        owner_id = clean_cell(delta.get("owner_object_id"))
        source_key = (
            clean_cell(delta.get("workbook")),
            clean_cell(delta.get("sheet")),
            clean_cell(delta.get("row")),
            normalize_key(delta.get("direction")),
        )
        matched = owner_id in route_ids or bool(
            added_bits & route_sources.get(source_key, set())
        )
        if not matched:
            continue
        label = owner_id or "/".join(source_key)
        if label in reported:
            continue
        reported.add(label)
        report.error(
            f"route_to_30 pad {label} was already claimed by a prior 04 accounting transaction; "
            "start from a fresh run root before transferring ownership to 30"
        )


def _validate_flat_upstream_completions(
    runtime: Any,
    run_root: Path,
    run_context: Dict[str, str],
    structure: str,
    clock_path: Path,
    report: Report,
) -> Dict[str, str]:
    specs = (
        (
            "00_stage_completion",
            run_root / "00_middle" / "stage_completion.meta",
            "00_harden_port_inventory",
        ),
        (
            "01_stage_completion",
            run_root / "01_middle" / "stage_completion.meta",
            "01_soc_clocks",
        ),
    )
    digests: Dict[str, str] = {}
    payloads: Dict[str, Dict[str, object]] = {}
    for label, path, expected_stage_name in specs:
        if not path.is_file():
            report.error(f"required upstream completion is missing: {path}")
            digests[label] = ""
            continue
        digests[label] = runtime.sha256_file(path)
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, ValueError, TypeError) as exc:
            report.error(f"required upstream completion is invalid: {path}: {exc}")
            continue
        if not isinstance(payload, dict):
            report.error(f"required upstream completion is invalid: {path}: expected JSON object")
            continue
        payloads[label] = payload
        checks = {
            "stage_name": clean_cell(payload.get("stage_name")) == expected_stage_name,
            "run_id": clean_cell(payload.get("run_id")) == clean_cell(run_context.get("run_id")),
            "mode_label": clean_cell(payload.get("mode_label")) == clean_cell(run_context.get("mode_label")),
            "structure_digest": clean_cell(payload.get("structure_digest")) == structure,
            "completion_status": normalize_key(payload.get("completion_status")) == "complete",
            "sync_changed": normalize_key(payload.get("sync_changed")) == "no",
            "error_count": clean_cell(payload.get("error_count")) in {"", "0"},
        }
        if label == "01_stage_completion":
            checks["design_revision"] = (
                clean_cell(payload.get("design_revision"))
                == clean_cell(run_context.get("design_revision"))
            )
        failed = [name for name, passed in checks.items() if not passed]
        if failed:
            report.error(
                f"upstream completion is stale or incomplete: {path}: {', '.join(failed)}"
            )

    digests["01_clock_inventory"] = (
        runtime.sha256_file(clock_path) if clock_path.is_file() else ""
    )
    completion_01 = payloads.get("01_stage_completion", {})
    if not clock_path.is_file():
        report.error(f"required 01 clock inventory is missing: {clock_path}")
    elif clean_cell(completion_01.get("clock_inventory_digest")) != digests["01_clock_inventory"]:
        report.error(
            f"{run_root / '01_middle' / 'stage_completion.meta'}: "
            "clock_inventory_digest does not match 01_middle/clock_inventory.csv"
        )
    return digests


def run_target_flat(args: argparse.Namespace) -> int:
    run_root = Path(args.run_root).expanduser().resolve()
    input_root = run_root / "inputs"
    middle = run_root / "04_middle"
    result = run_root / "04_result"
    form_path = middle / "04_soc_io_pads.xlsx"
    clock_path = run_root / "01_middle" / "clock_inventory.csv"
    manifest_path = run_root / "00_middle" / "harden_sdc_manifest.csv"
    stage, corner = args.stage, args.corner
    report = Report()
    runtime = _load_accounting_runtime()
    runtime.recover_transactions(run_root, report)
    run_context = runtime.read_run_context(input_root / "run_context.csv", report)
    required_views = runtime.read_required_views(input_root / "required_views.csv", report)
    required_04_views = [
        view for view in required_views if normalize_key(view.get("require_04")) == "yes"
    ]
    if stage == "all" and corner == "all" and required_04_views:
        explicit_all_view = next(
            (
                view for view in required_04_views
                if clean_cell(view.get("stage")) == "all"
                and clean_cell(view.get("corner")) == "all"
            ),
            None,
        )
        if explicit_all_view is None:
            stage, corner = required_04_views[0]["stage"], required_04_views[0]["corner"]
    current_view = next(
        (
            view for view in required_04_views
            if view["stage"] == stage and view["corner"] == corner
        ),
        None,
    )
    if current_view is None:
        report.error(
            f"04 formal generation requires a require_04=yes view, got stage={stage} corner={corner}"
        )
    view_id = current_view["view_id"] if current_view else f"{stage}_{corner}"
    output_path = result / ("04_soc_io_pads.sdc" if stage == "all" and corner == "all" else f"04_soc_io_pads_{stage}_{safe_filename_token(corner)}.sdc")
    report_path = result / "reports" / f"io_pad_check_report_{stage}_{safe_filename_token(corner)}.txt"

    info_entries, info_semantic = runtime.read_info_all(input_root / "info_all.xlsx", report)
    instances: Dict[str, InstInfo] = {}
    for name, entry in info_entries.items():
        instances[name] = InstInfo(
            module_name=entry.get("module_name", name),
            inst_name=name,
            owner=entry.get("owner", ""),
            sdc_hint=entry.get("sdc_path", ""),
        )
    completeness = _read_flat_manifest(manifest_path, run_root, instances, report)
    if args.require_complete_harden_sdc and completeness.missing_instances:
        report.error(
            "HARDEN_SDC_COMPLETENESS_REQUIRED: missing harden SDC instance(s): "
            + ", ".join(completeness.missing_instances)
        )
    port_paths = runtime.discover_port_workbooks(input_root, report)
    models, records = runtime.read_port_workbooks(port_paths, run_root, info_entries, True, report)
    runtime.validate_connections(records, info_entries, report)
    structure = runtime.structure_digest(run_context, required_views, info_semantic, records)
    accounting_before = runtime.accounting_digest(records)
    upstream_artifact_digests = _validate_flat_upstream_completions(
        runtime, run_root, run_context, structure, clock_path, report
    )
    port_sheets = read_port_workbooks(port_paths, report)
    attach_port_data(instances, port_sheets, report)
    pads = _build_flat_exact_pads(runtime, records, instances, report)
    extracted: List[ExtractedConstraint] = []
    for inst in instances.values():
        extracted.extend(extract_constraints_from_instance(inst, "common", report))
    extracted = _expand_flat_extracted_constraints(extracted, pads, report)
    if report.error_count == 0:
        sync_workbook(
            form_path,
            pads,
            extracted,
            report,
            {"view_id": view_id, "stage": stage, "corner": corner},
        )
        _sync_flat_pad_machine_fields(
            form_path,
            pads,
            run_context,
            completeness,
            structure,
            accounting_before,
            view_id,
            stage,
            corner,
            report,
        )
    common_inventory = read_clock_inventory(clock_path, report, "common", required=True)
    rows = read_form_rows(form_path) if form_path.is_file() else []
    if rows:
        validate_rows(
            rows,
            pads,
            "common",
            stage,
            corner,
            common_inventory,
            {},
            collect_current_sdc_digests(instances),
            "",
            "",
            "sta",
            report,
            allow_false_path=False,
        )
    coverage = build_coverage_lines(rows, pads, "common", stage, corner)
    if report.sync_changed or report.error_count:
        write_report(report_path, report, "common", stage, corner, "sta", form_path, output_path, coverage, completeness, [clock_path], manifest_path, None, False)
        return 1

    generated = generate_sdc(
        rows, "common", stage, corner, "sta", completeness, assembled=True
    )
    current_pad_rows = _resolved_flat_pad_rows(
        form_path, rows, stage, corner, "sta", accounting_before, report
    )
    updates = _flat_owner_updates(
        runtime, records, current_pad_rows, report
    )
    accounting_after = runtime.accounting_digest(records)
    if report.error_count:
        write_report(report_path, report, "common", stage, corner, "sta", form_path, output_path, coverage, completeness, [clock_path], manifest_path, None, False)
        return 1

    output_payload = ("\n".join(generated).rstrip() + "\n").encode("utf-8")
    for item in current_pad_rows:
        item["accounting_digest_after"] = accounting_after
    pad_rows, authenticated_prior_views = _merge_prior_pad_rows(
        middle / "pad_inventory.csv",
        current_pad_rows,
        view_id,
        {clean_cell(view.get("view_id")) for view in required_04_views},
        run_context,
        structure,
        report,
    )
    pad_payload = runtime.csv_text(PAD_HEADERS, pad_rows).encode("utf-8")
    pad_meta = {
        "schema_version": "1.0", "stage_name": "04_soc_io_pads",
        "run_id": run_context.get("run_id", ""), "mode_label": run_context.get("mode_label", ""),
        "design_revision": run_context.get("design_revision", ""), "completion_status": "complete",
        "error_count": 0, "sync_changed": "no",
        "structure_digest": structure, "accounting_digest_before": accounting_before,
        "accounting_digest_after": accounting_after, "inventory_digest": hashlib.sha256(pad_payload).hexdigest(),
        "pad_inventory_digest": hashlib.sha256(pad_payload).hexdigest(),
        "view_ids": sorted(
            {view_id}
            | authenticated_prior_views
            | {clean_cell(row.get("view_id")) for row in pad_rows if clean_cell(row.get("view_id"))}
        ),
    }
    delta_path = middle / "port_accounting_delta.csv"
    delta_meta_path = middle / "port_accounting_delta.meta"
    old_delta_rows = runtime.load_delta_rows(delta_path, report)
    _validate_route_accounting_history(
        runtime, pad_rows, old_delta_rows, report
    )
    transaction_id = "04_" + hashlib.sha256((structure + accounting_after + runtime.utc_timestamp()).encode("utf-8")).hexdigest()[:16]
    new_delta_rows = []
    for item in updates:
        record = item["record"]
        new_delta_rows.append({
            "schema_version": "1.0", "run_id": run_context.get("run_id", ""), "mode_label": run_context.get("mode_label", ""),
            "stage_name": "04_soc_io_pads", "transaction_id": transaction_id,
            "view_id": current_view["view_id"] if current_view else "", "stage": stage, "corner": corner,
            "structure_digest": structure, "accounting_digest_before": accounting_before, "accounting_digest_after": accounting_after,
            "workbook": record.workbook, "sheet": record.sheet, "row": record.row, "direction": record.direction,
            "port": record.shape.raw, "legal_bits": runtime.format_bits(set(record.shape.bits())),
            "added_bits": runtime.format_bits(item["added_bits"]), "final_used_bits": runtime.format_bits(record.used_bits),
            "owner_object_id": item["owner_object_id"], "reason": item["reason"], "evidence_status": "approved",
        })
    delta_payload = runtime.csv_text(runtime.DELTA_HEADERS, old_delta_rows + new_delta_rows).encode("utf-8")
    old_meta = {}
    if delta_meta_path.is_file():
        try:
            old_meta = json.loads(delta_meta_path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            report.warn(f"ignoring malformed prior delta meta: {delta_meta_path}")
    transaction = {
        "transaction_id": transaction_id, "committed_at": runtime.utc_timestamp(),
        "structure_digest": structure, "accounting_digest_before": accounting_before,
        "accounting_digest_after": accounting_after, "delta_rows_digest": runtime.delta_rows_digest(new_delta_rows),
    }
    delta_meta = {
        "schema_version": "1.0", "run_id": run_context.get("run_id", ""), "mode_label": run_context.get("mode_label", ""),
        "design_revision": run_context.get("design_revision", ""), "stage_name": "04_soc_io_pads",
        "completion_status": "complete", "structure_digest": structure,
        "accounting_digest_before": accounting_before, "accounting_digest_after": accounting_after,
        "delta_csv_digest": hashlib.sha256(delta_payload).hexdigest(),
        "transactions": list(old_meta.get("transactions", [])) + [transaction],
    }
    completion = {
        "schema_version": "1.0", "author": author_name(), "stage_name": "04_soc_io_pads", "stage": stage, "corner": corner,
        "run_id": run_context.get("run_id", ""), "mode_label": run_context.get("mode_label", ""), "design_revision": run_context.get("design_revision", ""),
        "completion_status": "complete", "error_count": 0, "sync_changed": "no", "structure_digest": structure,
        "accounting_digest_before": accounting_before, "accounting_digest_after": accounting_after, "port_accounting": "enabled",
        "upstream_artifact_digests": upstream_artifact_digests,
        "output_sdc_digest": hashlib.sha256(output_payload).hexdigest(), "pad_inventory_digest": hashlib.sha256(pad_payload).hexdigest(),
        "accounting_delta_digest": hashlib.sha256(delta_payload).hexdigest(), "review_workbook_digest": runtime.sha256_file(form_path), "transaction_id": transaction_id,
    }
    view_completion = dict(completion)
    view_completion.update({"view_id": view_id, "output_sdc_digest": hashlib.sha256(output_payload).hexdigest(), "completion_status": "complete"})
    pad_digest = hashlib.sha256(pad_payload).hexdigest()
    view_completion["pad_inventory_digest"] = pad_digest
    view_completion_payloads: Dict[Path, bytes] = {}
    required_view_digests: Dict[str, str] = {}
    completion_path = middle / "stage_completion.meta"
    for required_view in required_04_views:
        required_id = clean_cell(required_view.get("view_id"))
        required_stage = clean_cell(required_view.get("stage"))
        required_corner = clean_cell(required_view.get("corner"))
        required_path = (
            middle / "completion"
            / f"{safe_filename_token(required_stage)}_{safe_filename_token(required_corner)}.meta"
        )
        if required_id == view_id:
            payload = dict(view_completion)
        elif required_id not in authenticated_prior_views:
            continue
        elif required_path.is_file():
            try:
                payload = json.loads(required_path.read_text(encoding="utf-8"))
            except (OSError, ValueError, TypeError) as exc:
                report.warn(
                    f"excluding unreadable prior 04 view completion {required_path}: {exc}"
                )
                continue
            expected_output = result / (
                "04_soc_io_pads.sdc"
                if required_stage == "all" and required_corner == "all"
                else f"04_soc_io_pads_{required_stage}_{safe_filename_token(required_corner)}.sdc"
            )
            checks = {
                "stage_name": clean_cell(payload.get("stage_name")) == "04_soc_io_pads",
                "view_id": clean_cell(payload.get("view_id")) == required_id,
                "stage": clean_cell(payload.get("stage")) == required_stage,
                "corner": clean_cell(payload.get("corner")) == required_corner,
                "run_id": clean_cell(payload.get("run_id")) == clean_cell(run_context.get("run_id")),
                "mode_label": clean_cell(payload.get("mode_label")) == clean_cell(run_context.get("mode_label")),
                "design_revision": clean_cell(payload.get("design_revision")) == clean_cell(run_context.get("design_revision")),
                "structure_digest": clean_cell(payload.get("structure_digest")) == structure,
                "completion_status": normalize_key(payload.get("completion_status")) == "complete",
                "sync_changed": normalize_key(payload.get("sync_changed")) == "no",
                "error_count": clean_cell(payload.get("error_count")) in {"", "0"},
                "output_sdc_digest": (
                    expected_output.is_file()
                    and clean_cell(payload.get("output_sdc_digest"))
                    == runtime.sha256_file(expected_output)
                ),
                "upstream_artifact_digests": (
                    isinstance(payload.get("upstream_artifact_digests"), dict)
                    and all(
                        clean_cell(payload["upstream_artifact_digests"].get(name))
                        == digest
                        for name, digest in upstream_artifact_digests.items()
                    )
                ),
            }
            failed = [name for name, passed in checks.items() if not passed]
            if failed:
                report.warn(
                    f"excluding stale prior 04 view completion {required_path}: "
                    f"{', '.join(failed)}"
                )
                continue
            payload["pad_inventory_digest"] = pad_digest
        else:
            continue
        payload_bytes = (json.dumps(payload, indent=2) + "\n").encode("utf-8")
        view_completion_payloads[required_path] = payload_bytes
        required_view_digests[required_id] = hashlib.sha256(payload_bytes).hexdigest()
    completion["required_view_completions"] = required_view_digests
    required_04_ids = {clean_cell(view.get("view_id")) for view in required_04_views}
    if required_04_ids and set(required_view_digests) != required_04_ids:
        completion["completion_status"] = "review_required"
    write_report(report_path, report, "common", stage, corner, "sta", form_path, output_path, coverage, completeness, [clock_path], manifest_path, None, False)
    if report.error_count:
        return 1
    artifact_payloads = [
        (output_path, output_payload), (report_path, report_path.read_bytes()),
        (middle / "pad_inventory.csv", pad_payload), (middle / "pad_inventory.meta", (json.dumps(pad_meta, indent=2) + "\n").encode("utf-8")),
        (delta_path, delta_payload), (delta_meta_path, (json.dumps(delta_meta, indent=2) + "\n").encode("utf-8")),
        (completion_path, (json.dumps(completion, indent=2) + "\n").encode("utf-8")),
    ]
    artifact_payloads.extend(sorted(view_completion_payloads.items(), key=lambda item: str(item[0])))
    candidate_root = run_root / ".04_candidates"
    if candidate_root.exists():
        shutil.rmtree(str(candidate_root))
    candidate_root.mkdir(parents=True)
    prepared: Dict[str, Path] = {}
    for model in models:
        if model.modified:
            candidate = candidate_root / model.relative_name
            candidate.parent.mkdir(parents=True, exist_ok=True)
            model.workbook.save(str(candidate))
            prepared[model.relative_name] = candidate
    try:
        runtime.execute_transaction(run_root, models, prepared, artifact_payloads, transaction_id, run_context, structure, accounting_before, accounting_after, report)
    finally:
        if candidate_root.exists():
            shutil.rmtree(str(candidate_root))
    return 0


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate 04 SoC IO/pad SDC from extracted/reviewed constraints.")
    parser.add_argument(
        "--run-root",
        help="target runtime root; reads inputs/00_middle/01_middle and writes 04_middle/04_result",
    )
    parser.add_argument("-scenario", "--scenario", default="common", choices=sorted(SCENARIOS), help="legacy scenario; flat target runs are single-run")
    parser.add_argument("-stage", "--stage", default="all", choices=sorted(STAGES), help="target stage/view")
    parser.add_argument("-corner", "--corner", default="all", help="target corner/view")
    parser.add_argument("-input", "--input", help="common 01 assembled clock inventory CSV")
    parser.add_argument("--scenario-input", help="scenario assembled clock inventory CSV")
    parser.add_argument("--info-all", help="legacy-only integration summary xlsx")
    parser.add_argument("--port-files", nargs="*", help="legacy-only explicit integration port workbook path(s)")
    parser.add_argument("--form", help="IO pad constraint workbook")
    parser.add_argument("--output", help="output SDC path")
    parser.add_argument("--harden-sdc-manifest", help="00 harden SDC manifest CSV")
    parser.add_argument(
        "--require-complete-harden-sdc",
        action="store_true",
        help="treat any manifest/inferred missing harden SDC as an error",
    )
    parser.add_argument("--connection-inventory", help="00 bit-level connection inventory CSV")
    parser.add_argument("--tool", default="sta", choices=sorted(TOOLS), help="target tool surface; sta skips dont_touch_network")
    parser.add_argument("--time-unit", default="", help="optional expected time unit for unit_time checks, e.g. ns")
    parser.add_argument("--cap-unit", default="", help="optional expected capacitance unit for unit_cap checks, e.g. pF")
    parser.add_argument("--force-generate-after-sync", action="store_true", help="generate SDC even if workbook was synchronized")
    parser.add_argument(
        "--pending-root",
        help=(
            "optional 00 pending root override; target default is 00_middle/scenario/<scenario>, "
            "legacy default is 00_harden_port_inventory"
        ),
    )
    parser.add_argument(
        "--no-update-pending",
        action="store_true",
        help="do not consume 00_harden_port_inventory/pending even if it exists",
    )
    parser.add_argument("--report", help="output report path")
    return parser.parse_args(argv)


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)
    if args.run_root:
        flat_manifest = Path(args.run_root).expanduser().resolve() / "00_middle" / "harden_sdc_manifest.csv"
        if flat_manifest.is_file():
            return run_target_flat(args)
    cwd = Path.cwd()
    report = Report()
    print(f"Author: {author_name()}")

    target_layout = bool(args.run_root)
    run_root = Path(args.run_root).expanduser().resolve() if target_layout else cwd
    input_root = run_root / "inputs" if target_layout else cwd

    def resolve_path(base: Path, value: Optional[str], default: str) -> Path:
        path = Path(value).expanduser() if value else Path(default)
        return path if path.is_absolute() else base / path

    info_all = resolve_path(input_root, args.info_all, "info_all.xlsx")
    form_path = resolve_path(
        run_root if target_layout else cwd,
        args.form,
        "04_middle/04_soc_io_pads.xlsx" if target_layout else "04_soc_io_pads.xlsx",
    )
    output_path = resolve_path(
        run_root if target_layout else cwd,
        args.output,
        str(output_sdc_path(Path("04_result") if target_layout else Path("."), args.scenario, args.stage, args.corner)),
    )
    rpt_path = resolve_path(
        run_root if target_layout else cwd,
        args.report,
        (
            f"04_result/reports/io_pad_check_report_{args.scenario}_{args.stage}_"
            f"{safe_filename_token(args.corner)}.txt"
            if target_layout
            else report_path(Path("."), args.scenario, args.stage, args.corner).as_posix()
        ),
    )

    if target_layout:
        common_clock_path = resolve_path(
            run_root,
            args.input,
            "01_middle/assembled/common/clock_inventory.csv",
        ).resolve()
        scenario_clock_path = resolve_path(
            run_root,
            args.scenario_input,
            f"01_middle/assembled/{args.scenario}/clock_inventory.csv",
        ).resolve()
        manifest_path: Optional[Path] = resolve_path(
            run_root,
            args.harden_sdc_manifest,
            f"00_middle/scenario/{args.scenario}/harden_sdc_manifest.csv",
        )
        connection_path: Optional[Path] = resolve_path(
            run_root,
            args.connection_inventory,
            "00_middle/connection_inventory.csv",
        )
        if args.pending_root:
            pending_base = resolve_path(run_root, args.pending_root, args.pending_root)
            pending_dir = pending_base / "pending"
        else:
            pending_dir = run_root / f"00_middle/scenario/{args.scenario}/pending"
        removed_log_dir = run_root / f"04_middle/scenario/{args.scenario}/removed_log"
    else:
        common_clock_path = resolve_path(cwd, args.input, "../01_soc_clocks/clock_inventory.csv").resolve()
        scenario_clock_path = resolve_path(cwd, args.scenario_input, args.scenario_input or "").resolve()
        manifest_path = (
            resolve_path(cwd, args.harden_sdc_manifest, args.harden_sdc_manifest)
            if args.harden_sdc_manifest
            else None
        )
        connection_path = (
            resolve_path(cwd, args.connection_inventory, args.connection_inventory)
            if args.connection_inventory
            else None
        )
        pending_base = resolve_path(cwd, args.pending_root, "00_harden_port_inventory")
        pending_dir = pending_base / "pending"
        removed_log_dir = pending_base / "removed_log"

    report.info(f"resolved run root: {run_root.resolve()}")
    if not target_layout:
        report.info(f"resolved legacy input root: {input_root.resolve()}")
        report.info(f"resolved legacy info workbook: {info_all.resolve()}")
    report.info(f"resolved form workbook: {form_path.resolve()}")
    report.info(f"resolved common clock inventory: {common_clock_path.resolve()}")
    if args.scenario != "common" and (target_layout or args.scenario_input):
        report.info(f"resolved scenario clock inventory: {scenario_clock_path.resolve()}")
    if manifest_path is not None:
        report.info(f"resolved harden SDC manifest: {manifest_path.resolve()}")
    if connection_path is not None:
        report.info(f"resolved connection inventory: {connection_path.resolve()}")
    report.info(f"resolved output SDC: {output_path.resolve()}")
    report.info(f"resolved output report: {rpt_path.resolve()}")

    if target_layout:
        instances: Dict[str, InstInfo] = {}
        if args.info_all or args.port_files:
            report.error(
                "TARGET_LEGACY_INTEGRATION_INPUT_FORBIDDEN: target mode derives pad mappings from "
                "00_middle/connection_inventory.csv; --info-all/--port-files are legacy-only"
            )
        assert manifest_path is not None
        completeness = apply_harden_sdc_manifest(
            instances,
            manifest_path,
            run_root,
            args.scenario,
            args.require_complete_harden_sdc,
            report,
            create_instances=True,
        )
        if connection_path is None or not connection_path.is_file():
            report.error(f"TARGET_UPSTREAM_CONNECTION_INVENTORY_MISSING: {connection_path}")
            pad_records: List[PadRecord] = []
        else:
            target_edges = read_connection_inventory(
                connection_path,
                report,
                scenario=args.scenario,
                strict_schema=True,
            )
            pad_records = attach_target_pad_edges(instances, target_edges, report)
    else:
        if not info_all.is_file():
            raise RuntimeError(f"integration file not found: {info_all}")
        instances = read_info_all(info_all, report)
        if args.port_files:
            port_paths = [resolve_path(input_root, value, value) for value in args.port_files]
        else:
            port_paths = default_port_workbooks(input_root, info_all.name, form_path.name, report)
        if not port_paths:
            report.error("no owner port workbook found; expected port_*.xlsx/ports_*.xlsx or --port-files")
        for path in port_paths:
            if not path.is_file():
                report.error(f"port workbook not found: {path}")
        port_sheets = read_port_workbooks(port_paths, report)
        attach_port_data(instances, port_sheets, report)
        validate_integration_port_keys(instances, report)
        if manifest_path is not None:
            completeness = apply_harden_sdc_manifest(
                instances,
                manifest_path,
                run_root,
                args.scenario,
                args.require_complete_harden_sdc,
                report,
            )
        else:
            completeness = resolve_sdc_paths(instances, input_root, report)
            if args.require_complete_harden_sdc and completeness.missing_instances:
                report.error(
                    "HARDEN_SDC_COMPLETENESS_REQUIRED: missing harden SDC instance(s): "
                    + ", ".join(completeness.missing_instances)
                )
        pad_records = build_pad_records(instances)
        if connection_path is not None:
            if not connection_path.is_file():
                report.error(f"LEGACY_CONNECTION_INVENTORY_MISSING: {connection_path}")
            else:
                legacy_edges = read_connection_inventory(
                    connection_path,
                    report,
                    scenario=args.scenario,
                    strict_schema=False,
                )
                validate_pad_connections(pad_records, legacy_edges, report)

    current_digests = collect_current_sdc_digests(instances)
    extracted: List[ExtractedConstraint] = []
    for inst in instances.values():
        extracted.extend(extract_constraints_from_instance(inst, args.scenario, report))

    sync_workbook(form_path, pad_records, extracted, report)

    common_inventory = read_clock_inventory(common_clock_path, report, "common", required=target_layout)
    scenario_inventory = (
        read_clock_inventory(scenario_clock_path, report, args.scenario, required=target_layout)
        if args.scenario != "common" and (target_layout or args.scenario_input)
        else {}
    )

    rows = read_form_rows(form_path)
    validate_rows(
        rows,
        pad_records,
        args.scenario,
        args.stage,
        args.corner,
        common_inventory,
        scenario_inventory,
        current_digests,
        args.time_unit,
        args.cap_unit,
        args.tool,
        report,
    )

    if target_layout and not args.no_update_pending and not pending_dir.is_dir():
        report.error(f"TARGET_UPSTREAM_PENDING_MISSING: {pending_dir}")

    generated = False
    if report.sync_changed and not args.force_generate_after_sync:
        report.warn("workbook changed during sync; review 04_soc_io_pads.xlsx before SDC generation")
    elif report.error_count == 0:
        atomic_write_text(
            output_path,
            "\n".join(
                generate_sdc(
                    rows,
                    args.scenario,
                    args.stage,
                    args.corner,
                    args.tool,
                    completeness,
                )
            ).rstrip()
            + "\n",
        )
        report.info(f"wrote {output_path}")
        generated = True
    else:
        report.warn("SDC generation skipped because errors were reported")

    if generated and not args.no_update_pending:
        update_pending_for_04(
            run_root,
            pending_dir,
            removed_log_dir,
            rows,
            pad_records,
            args.scenario,
            args.stage,
            args.corner,
            report,
        )

    coverage_lines = build_coverage_lines(rows, pad_records, args.scenario, args.stage, args.corner)
    clock_paths = [common_clock_path]
    if args.scenario != "common" and (target_layout or args.scenario_input):
        clock_paths.append(scenario_clock_path)
    write_report(
        rpt_path,
        report,
        args.scenario,
        args.stage,
        args.corner,
        args.tool,
        form_path,
        output_path,
        coverage_lines,
        completeness,
        clock_paths,
        manifest_path,
        connection_path,
        not args.no_update_pending,
    )
    print(f"Report: {rpt_path}")
    print(f"Warnings: {report.warning_count}  Errors: {report.error_count}  Sync changed: {report.sync_changed}")
    if report.error_count:
        return 1
    if report.sync_changed and not args.force_generate_after_sync:
        return 1
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(2)
