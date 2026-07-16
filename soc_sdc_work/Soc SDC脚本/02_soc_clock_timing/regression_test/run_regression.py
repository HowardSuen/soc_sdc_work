#!/usr/bin/env python3
"""Regression for the single-run stage 02 clock-timing runtime."""

from __future__ import print_function

import csv
import hashlib
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parent
SOC = BASE.parent.parent
EX02 = SOC / "02_soc_clock_timing" / "02_extract_soc_clock_timing.py"
WORK = BASE / "work_complex"

CLOCK_BUDGET_HEADERS = [
    "stage", "corner", "clock_name", "clock_kind", "period",
    "source_latency_early", "source_latency_late",
    "network_latency_early", "network_latency_late",
    "setup_uncertainty", "hold_uncertainty",
    "transition_min", "transition_max", "propagated", "apply",
    "sync_status", "source_inventory_digest", "note",
]
INVENTORY_HEADERS = [
    "run_id", "mode_label", "design_revision", "structure_digest",
    "clock_name", "clock_kind", "period", "direction", "direct_source",
    "final_action", "final_sdc_digest", "run_completeness",
    "missing_instances", "note",
]


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def clean_dir(path):
    if path.exists():
        shutil.rmtree(str(path))
    path.mkdir(parents=True)


def sha256_file(path):
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        while True:
            chunk = file_obj.read(1024 * 1024)
            if not chunk:
                return digest.hexdigest()
            digest.update(chunk)


def write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(dict((header, row.get(header, "")) for header in headers))


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, sort_keys=True, indent=2) + "\n", encoding="utf-8")


def clocks():
    return [
        {
            "clock_name": "top_sys_clk",
            "clock_kind": "create_clock",
            "period": "10",
            "direction": "input",
            "direct_source": "top/sys_clk",
            "final_action": "emit_top_clock",
        },
        {
            "clock_name": "u_pll_core_clk",
            "clock_kind": "create_generated_clock",
            "period": "5",
            "direction": "output",
            "direct_source": "u_pll/core_clk_o",
            "final_action": "emit_output_clock",
        },
        {
            "clock_name": "v_ddr_ref",
            "clock_kind": "virtual_clock",
            "period": "8",
            "direction": "virtual",
            "direct_source": "virtual/v_ddr_ref",
            "final_action": "emit_virtual_clock",
        },
    ]


def write_upstream(root, clock_rows=None, run_id="RUN_02_TARGET", mode="func", revision="revA", structure="structure-02", completeness="complete", missing=None):
    clock_rows = list(clock_rows if clock_rows is not None else clocks())
    missing = list(missing or [])
    write_csv(
        root / "inputs/run_context.csv",
        ["run_id", "mode_label", "design_revision", "note"],
        [{"run_id": run_id, "mode_label": mode, "design_revision": revision, "note": "02 regression"}],
    )
    write_csv(
        root / "inputs/required_views.csv",
        ["view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note"],
        [
            {"view_id": "prects_ss", "stage": "prects", "corner": "ss_125", "require_02": "yes", "require_04": "yes", "require_20": "yes", "require_30": "yes", "note": "required"},
            {"view_id": "prects_ff_diag", "stage": "prects", "corner": "ff_m40", "require_02": "no", "require_04": "no", "require_20": "no", "require_30": "no", "note": "diagnostic"},
        ],
    )
    sdc_path = root / "01_result/01_soc_clocks.sdc"
    sdc_path.parent.mkdir(parents=True, exist_ok=True)
    sdc_lines = ["# 01 regression output"]
    for row in clock_rows:
        if row["clock_kind"] == "create_generated_clock":
            sdc_lines.append("create_generated_clock -name {%s} -source [get_clocks {top_sys_clk}] -divide_by 2 [get_pins {u_pll/core_clk_o}]" % row["clock_name"])
        else:
            sdc_lines.append("create_clock -name {%s} -period %s" % (row["clock_name"], row["period"]))
    sdc_path.write_text("\n".join(sdc_lines) + "\n", encoding="utf-8")
    sdc_digest = sha256_file(sdc_path)
    inventory_rows = []
    for clock in clock_rows:
        row = dict(clock)
        row.update({
            "run_id": run_id,
            "mode_label": mode,
            "design_revision": revision,
            "structure_digest": structure,
            "final_sdc_digest": sdc_digest,
            "run_completeness": completeness,
            "missing_instances": ";".join(missing),
        })
        inventory_rows.append(row)
    inventory_path = root / "01_middle/clock_inventory.csv"
    write_csv(inventory_path, INVENTORY_HEADERS, inventory_rows)
    inventory_digest = sha256_file(inventory_path)
    meta = {
        "schema_version": "1.0",
        "stage_name": "01_soc_clocks",
        "run_id": run_id,
        "mode_label": mode,
        "design_revision": revision,
        "completion_status": "complete",
        "error_count": 0,
        "structure_digest": structure,
        "inventory_path": str(inventory_path.resolve()),
        "inventory_digest": inventory_digest,
        "final_sdc_path": str(sdc_path.resolve()),
        "final_sdc_digest": sdc_digest,
        "run_completeness": completeness,
        "missing_instances": missing,
    }
    write_json(root / "01_middle/clock_inventory.meta", meta)
    completion = {
        "schema_version": "1.0",
        "stage_name": "01_soc_clocks",
        "run_id": run_id,
        "mode_label": mode,
        "design_revision": revision,
        "completion_status": "complete",
        "error_count": 0,
        "sync_changed": "no",
        "structure_digest": structure,
        "clock_inventory_digest": inventory_digest,
        "output_sdc_digest": sdc_digest,
    }
    write_json(root / "01_middle/stage_completion.meta", completion)


def run02(root, stage="prects", corner="ss_125", diagnose=False, extra=None):
    args = [sys.executable, str(EX02), "--run-root", str(root), "--stage", stage, "--corner", corner]
    if diagnose:
        args.append("--diagnose-only")
    if extra:
        args.extend(extra)
    return subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)


def workbook_path(root):
    return root / "02_middle/02_soc_clock_timing_budget_prects.xlsx"


def header_map(ws):
    return dict((str(ws.cell(1, col).value), col) for col in range(1, ws.max_column + 1))


def rows_for(ws, stage, corner):
    mapping = header_map(ws)
    result = []
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row_idx, mapping["stage"]).value or "") == stage and str(ws.cell(row_idx, mapping["corner"]).value or "") == corner:
            result.append(row_idx)
    return mapping, result


def review_view(root, stage="prects", corner="ss_125"):
    path = workbook_path(root)
    wb = load_workbook(str(path))
    ws = wb["clock_budget"]
    mapping, row_indices = rows_for(ws, stage, corner)
    require(row_indices, "no workbook rows found for review")
    for row_idx in row_indices:
        ws.cell(row_idx, mapping["setup_uncertainty"], 0.10)
        ws.cell(row_idx, mapping["hold_uncertainty"], 0.02)
        ws.cell(row_idx, mapping["transition_max"], 0.20)
        ws.cell(row_idx, mapping["propagated"], "no")
        ws.cell(row_idx, mapping["apply"], "yes")
        ws.cell(row_idx, mapping["note"], "reviewed regression budget")
    wb.save(str(path))


def settle_view(root, stage="prects", corner="ss_125", diagnose=False):
    first = run02(root, stage, corner, diagnose)
    require(first.returncode == 1, "first run must create/sync workbook")
    review_view(root, stage, corner)
    second = run02(root, stage, corner, diagnose)
    require(second.returncode == 1, "reviewed NEW_FROM_01 rows must pass through sync gate")
    third = run02(root, stage, corner, diagnose)
    require(third.returncode == 0, "clean reviewed view did not pass: %s\n%s" % (third.stdout, third.stderr))
    return third


def copy_case(source, name):
    target = WORK / name
    if target.exists():
        shutil.rmtree(str(target))
    shutil.copytree(str(source), str(target))
    meta_path = target / "01_middle/clock_inventory.meta"
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    meta["inventory_path"] = str((target / "01_middle/clock_inventory.csv").resolve())
    meta["final_sdc_path"] = str((target / "01_result/01_soc_clocks.sdc").resolve())
    write_json(meta_path, meta)
    return target


def update_upstream_digests(root, structure=None, run_id=None, mode=None, revision=None):
    inventory_path = root / "01_middle/clock_inventory.csv"
    inventory_digest = sha256_file(inventory_path)
    meta_path = root / "01_middle/clock_inventory.meta"
    completion_path = root / "01_middle/stage_completion.meta"
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    completion = json.loads(completion_path.read_text(encoding="utf-8"))
    meta["inventory_digest"] = inventory_digest
    completion["clock_inventory_digest"] = inventory_digest
    if structure is not None:
        meta["structure_digest"] = structure
        completion["structure_digest"] = structure
    for payload in (meta, completion):
        if run_id is not None:
            payload["run_id"] = run_id
        if mode is not None:
            payload["mode_label"] = mode
        if revision is not None:
            payload["design_revision"] = revision
    write_json(meta_path, meta)
    write_json(completion_path, completion)


def test_success_and_completion():
    root = WORK / "success"
    clean_dir(root)
    write_upstream(root)
    port_file = root / "inputs/port_dummy.xlsx"
    port_file.write_bytes(b"stage02-must-not-touch-port-workbook")
    port_before = sha256_file(port_file)
    result = settle_view(root)
    require("Port accounting: not_applicable; added_bits=0" in result.stdout, "stdout port-accounting contract missing")
    output = root / "02_result/02_soc_clock_timing_prects_ss_125.sdc"
    manifest = root / "02_middle/resolved/prects_ss_125.manifest"
    completion_path = root / "02_middle/completion/prects_ss_125.meta"
    report = root / "02_result/reports/clock_timing_check_report_prects_ss_125.txt"
    for path in (output, manifest, completion_path, report):
        require(path.is_file(), "missing formal artifact %s" % path)
    text = output.read_text(encoding="utf-8")
    require("set_clock_uncertainty -setup 0.1 [get_clocks {top_sys_clk}]" in text, "setup uncertainty command missing")
    require("set_clock_transition -max 0.2 [get_clocks {u_pll_core_clk}]" in text, "transition command missing")
    require("# Port accounting: not_applicable; added_bits=0" in text, "SDC port-accounting contract missing")
    completion = json.loads(completion_path.read_text(encoding="utf-8"))
    require(completion["completion_status"] == "complete", "02 completion status is not complete")
    require(completion["view_id"] == "prects_ss", "completion view_id mismatch")
    require(completion["error_count"] == 0 and completion["sync_changed"] == "no", "unsafe completion flags")
    require(completion["port_accounting"] == "not_applicable" and completion["added_bits"] == 0, "completion port accounting mismatch")
    require(completion["port_accounting_summary"] == "Port accounting: not_applicable; added_bits=0", "completion literal port-accounting summary missing")
    require(completion["output_sdc_digest"] == sha256_file(output), "completion output digest mismatch")
    require(completion["resolved_manifest_digest"] == sha256_file(manifest), "completion manifest digest mismatch")
    require(completion["upstream_01_inventory_digest"] == sha256_file(root / "01_middle/clock_inventory.csv"), "completion inventory digest mismatch")
    require(completion["upstream_01_completion_digest"] == sha256_file(root / "01_middle/stage_completion.meta"), "completion upstream completion digest mismatch")
    require(len(completion["workbook_semantic_digest"]) == 64, "workbook semantic digest missing")
    require(port_before == sha256_file(port_file), "stage 02 modified a port workbook")
    wb = load_workbook(str(workbook_path(root)), data_only=True)
    metadata = dict((str(wb["runtime_metadata"].cell(row, 1).value), str(wb["runtime_metadata"].cell(row, 2).value or "")) for row in range(2, wb["runtime_metadata"].max_row + 1))
    require(metadata["run_id"] == "RUN_02_TARGET" and metadata["view_id"] == "prects_ss", "runtime metadata provenance missing")
    require(metadata["Port accounting"] == "not_applicable" and metadata["added_bits"] == "0", "runtime metadata port-accounting mismatch")
    return root


def test_non_required_diagnostic(base_root):
    root = copy_case(base_root, "non_required")
    formal = run02(root, corner="ff_m40")
    require(formal.returncode == 1, "non-required formal view must be blocked")
    report = root / "02_result/reports/clock_timing_check_report_prects_ff_m40.txt"
    require("require_02=no" in report.read_text(encoding="utf-8"), "non-required report did not explain gate")
    first = run02(root, corner="ff_m40", diagnose=True)
    require(first.returncode == 1, "first diagnostic run should sync rows")
    review_view(root, corner="ff_m40")
    second = run02(root, corner="ff_m40", diagnose=True)
    require(second.returncode == 1, "diagnostic sync review gate missing")
    third = run02(root, corner="ff_m40", diagnose=True)
    require(third.returncode == 0, "clean diagnostic view did not pass")
    require(not (root / "02_middle/completion/prects_ff_m40.meta").exists(), "diagnostic run published completion")
    require(not (root / "02_result/02_soc_clock_timing_prects_ff_m40.sdc").exists(), "diagnostic run published formal SDC")


def test_other_corner_ignored(base_root):
    root = copy_case(base_root, "other_corner_ignored")
    path = workbook_path(root)
    wb = load_workbook(str(path))
    ws = wb["clock_budget"]
    mapping = header_map(ws)
    row = ws.max_row + 1
    values = {"stage": "prects", "corner": "ff_m40", "clock_name": "obsolete_other_corner", "apply": "yes", "setup_uncertainty": 0.2, "sync_status": "STALE_NOT_IN_01", "note": "must be ignored"}
    for key, value in values.items():
        ws.cell(row, mapping[key], value)
    wb.save(str(path))
    result = run02(root)
    require(result.returncode == 0, "other-corner stale row incorrectly blocked current view")


def test_sync_and_duplicates(base_root):
    sync_root = copy_case(base_root, "machine_sync")
    path = workbook_path(sync_root)
    wb = load_workbook(str(path))
    ws = wb["clock_budget"]
    mapping, indices = rows_for(ws, "prects", "ss_125")
    ws.cell(indices[0], mapping["clock_kind"], "wrong_kind")
    ws.cell(indices[0], mapping["period"], "999")
    ws.cell(indices[0], mapping["source_inventory_digest"], "stale")
    wb.save(str(path))
    result = run02(sync_root)
    require(result.returncode == 1, "machine-column synchronization must stop generation")
    require(not (sync_root / "02_middle/completion/prects_ss_125.meta").exists(), "sync-changing run retained completion")
    wb = load_workbook(str(path), data_only=True)
    ws = wb["clock_budget"]
    mapping, indices = rows_for(ws, "prects", "ss_125")
    require(str(ws.cell(indices[0], mapping["period"]).value) == "10", "period was not synchronized from 01")
    require(ws.cell(indices[0], mapping["clock_kind"]).value == "create_clock", "clock_kind was not synchronized")
    require(ws.cell(indices[0], mapping["source_inventory_digest"]).value == sha256_file(sync_root / "01_middle/clock_inventory.csv"), "source inventory digest was not synchronized")

    duplicate_root = copy_case(base_root, "duplicate")
    path = workbook_path(duplicate_root)
    wb = load_workbook(str(path))
    ws = wb["clock_budget"]
    mapping, indices = rows_for(ws, "prects", "ss_125")
    source = indices[0]
    target = ws.max_row + 1
    for col in range(1, ws.max_column + 1):
        ws.cell(target, col, ws.cell(source, col).value)
    wb.save(str(path))
    result = run02(duplicate_root)
    require(result.returncode == 1, "duplicate current view key must fail")
    report = duplicate_root / "02_result/reports/clock_timing_check_report_prects_ss_125.txt"
    require("duplicate stage/corner/clock_name" in report.read_text(encoding="utf-8"), "duplicate failure missing from report")


def test_auxiliary_validation(base_root):
    root = copy_case(base_root, "auxiliary")
    path = workbook_path(root)
    wb = load_workbook(str(path))
    pair = wb["clock_pair_uncertainty"]
    pair.append(["prects", "ss_125", "unknown_clock", "top_sys_clk", 0.03, 0.01, "yes", "pair review"])
    wb.save(str(path))
    result = run02(root)
    require(result.returncode == 1, "invalid auxiliary clock must fail")
    wb = load_workbook(str(path))
    pair = wb["clock_pair_uncertainty"]
    pair.cell(2, 3, "u_pll_core_clk")
    derate = wb["derate_ocv"]
    derate.append(["prects", "ss_125", "all_clocks", "clock", 0.95, 1.05, "yes", "yes", "MMMC owns derate"])
    wb.save(str(path))
    result = run02(root)
    require(result.returncode == 0, "valid auxiliary placeholder rows should pass validation")
    output = (root / "02_result/02_soc_clock_timing_prects_ss_125.sdc").read_text(encoding="utf-8")
    require("set_clock_derate" not in output and "-from" not in output, "placeholder auxiliary sheets emitted unsupported commands")
    report = (root / "02_result/reports/clock_timing_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("managed by external flow" in report, "external derate warning missing")


def test_upstream_failures(base_root):
    root = copy_case(base_root, "missing_completion")
    (root / "01_middle/stage_completion.meta").unlink()
    require(run02(root).returncode == 1, "missing 01 completion must fail")

    root = copy_case(base_root, "provenance_mismatch")
    context = root / "inputs/run_context.csv"
    write_csv(context, ["run_id", "mode_label", "design_revision", "note"], [{"run_id": "RUN_CHANGED", "mode_label": "scan", "design_revision": "revB", "note": "changed"}])
    require(run02(root).returncode == 1, "run provenance mismatch must fail")

    root = copy_case(base_root, "structure_mismatch")
    completion_path = root / "01_middle/stage_completion.meta"
    completion = json.loads(completion_path.read_text(encoding="utf-8"))
    completion["structure_digest"] = "stale-structure"
    write_json(completion_path, completion)
    require(run02(root).returncode == 1, "01 structure digest mismatch must fail")

    root = copy_case(base_root, "inventory_digest_mismatch")
    with (root / "01_middle/clock_inventory.csv").open("a", encoding="utf-8") as file_obj:
        file_obj.write("#stale\n")
    require(run02(root).returncode == 1, "01 inventory digest mismatch must fail")

    root = copy_case(base_root, "completion_flags")
    completion_path = root / "01_middle/stage_completion.meta"
    completion = json.loads(completion_path.read_text(encoding="utf-8"))
    completion["sync_changed"] = "yes"
    write_json(completion_path, completion)
    require(run02(root).returncode == 1, "01 sync_changed=yes must fail")


def test_partial_inventory(base_root):
    root = copy_case(base_root, "partial_inventory")
    remaining = clocks()[:2]
    write_upstream(root, clock_rows=remaining, completeness="partial", missing=["u_missing"])
    result = run02(root)
    require(result.returncode == 1, "partial inventory removal should synchronize workbook")
    wb = load_workbook(str(workbook_path(root)), data_only=True)
    ws = wb["clock_budget"]
    mapping, indices = rows_for(ws, "prects", "ss_125")
    statuses = dict((ws.cell(row, mapping["clock_name"]).value, ws.cell(row, mapping["sync_status"]).value) for row in indices)
    require(statuses.get("v_ddr_ref") == "BLOCKED_BY_MISSING_SDC", "partial inventory clock was marked as real stale")
    require(not (root / "02_middle/completion/prects_ss_125.meta").exists(), "partial sync-changing run retained completion")
    report = (root / "02_result/reports/clock_timing_check_report_prects_ss_125.txt").read_text(encoding="utf-8")
    require("run completeness is partial" in report, "partial inventory warning missing")


def test_cli_rejects_scenario(base_root):
    result = run02(base_root, extra=["--scenario", "func"])
    require(result.returncode == 2, "stage 02 must reject scenario CLI selection")


def main():
    clean_dir(WORK)
    base_root = test_success_and_completion()
    test_non_required_diagnostic(base_root)
    test_other_corner_ignored(base_root)
    test_sync_and_duplicates(base_root)
    test_auxiliary_validation(base_root)
    test_upstream_failures(base_root)
    test_partial_inventory(base_root)
    test_cli_rejects_scenario(base_root)
    print("02 single-run clock timing regression: PASS")


if __name__ == "__main__":
    main()
