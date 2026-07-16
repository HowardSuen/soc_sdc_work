#!/usr/bin/env python3
"""Generate one SoC clock-timing SDC for a single-run stage/corner view.

Stage 02 consumes the completed flat artifacts from stage 01.  It never
selects a scenario and never performs port accounting.  A formal completion
record is published only for a required view after a clean, review-stable run.
"""

import argparse
import csv
import hashlib
import json
import math
import os
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from dataclasses import dataclass
except ImportError:  # pragma: no cover - project still supports Python 3.6
    def dataclass(cls):
        annotations = getattr(cls, "__annotations__", {})
        names = list(annotations.keys())
        defaults = dict((name, getattr(cls, name)) for name in names if hasattr(cls, name))

        def __init__(self, *args, **kwargs):
            if len(args) > len(names):
                raise TypeError("too many positional arguments")
            values = dict(zip(names, args))
            for name in names[len(args):]:
                if name in kwargs:
                    values[name] = kwargs.pop(name)
                elif name in defaults:
                    values[name] = defaults[name]
                else:
                    raise TypeError("missing required argument: " + name)
            if kwargs:
                raise TypeError("unexpected argument: " + sorted(kwargs)[0])
            for name in names:
                setattr(self, name, values[name])

        cls.__init__ = __init__
        return cls

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover
    print("ERROR: openpyxl is required for stage 02 workbooks.", file=sys.stderr)
    raise SystemExit(2) from exc


ACTIVE_01_ACTIONS = {"emit_top_clock", "emit_output_clock", "emit_virtual_clock"}
YES_NO = {"yes", "no"}
SYNC_OK = {"", "OK"}
SYNC_VALUES = {"", "OK", "NEW_FROM_01", "STALE_NOT_IN_01", "BLOCKED_BY_MISSING_SDC"}
MACHINE_COLUMNS = ("clock_kind", "period", "source_inventory_digest")
NUMERIC_COLUMNS = (
    "source_latency_early",
    "source_latency_late",
    "network_latency_early",
    "network_latency_late",
    "setup_uncertainty",
    "hold_uncertainty",
    "transition_min",
    "transition_max",
)

CLOCK_BUDGET_HEADERS = [
    "stage", "corner", "clock_name", "clock_kind", "period",
    "source_latency_early", "source_latency_late",
    "network_latency_early", "network_latency_late",
    "setup_uncertainty", "hold_uncertainty",
    "transition_min", "transition_max", "propagated", "apply",
    "sync_status", "source_inventory_digest", "note",
]
PAIR_HEADERS = [
    "stage", "corner", "from_clock", "to_clock",
    "setup_uncertainty", "hold_uncertainty", "apply", "note",
]
DERATE_HEADERS = [
    "stage", "corner", "derate_scope", "object_type", "early", "late",
    "apply", "managed_by_flow", "note",
]
REQUIRED_VIEW_HEADERS = [
    "view_id", "stage", "corner", "require_02", "require_04",
    "require_20", "require_30", "note",
]

HEADER_FILL = PatternFill("solid", fgColor="215967")
TITLE_FILL = PatternFill("solid", fgColor="335C81")
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")
STALE_FILL = PatternFill("solid", fgColor="F4CCCC")
BLOCKED_FILL = PatternFill("solid", fgColor="FCE5CD")
CLEAR_FILL = PatternFill(fill_type=None)
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C6CC"),
    right=Side(style="thin", color="B8C6CC"),
    top=Side(style="thin", color="B8C6CC"),
    bottom=Side(style="thin", color="B8C6CC"),
)


@dataclass
class RunContext:
    run_id: str
    mode_label: str
    design_revision: str = ""
    note: str = ""


@dataclass
class RequiredView:
    view_id: str
    stage: str
    corner: str
    require_02: bool


@dataclass
class ClockInfo:
    clock_name: str
    direction: str = ""
    clock_kind: str = ""
    period: str = ""
    source: str = ""


@dataclass
class InventoryContext:
    path: Path
    meta_path: Path
    completion_path: Path
    final_sdc_path: Path
    clocks: Dict[str, ClockInfo]
    completeness: str
    missing_instances: Tuple[str, ...]
    inventory_digest: str
    final_sdc_digest: str
    structure_digest: str
    run_id: str
    mode_label: str
    design_revision: str
    completion_digest: str


@dataclass
class SheetRow:
    row_idx: int
    values: Dict[str, object]


class Report:
    def __init__(self) -> None:
        self.lines: List[str] = []
        self.warning_count = 0
        self.error_count = 0
        self.sync_changed = False

    def info(self, message: str) -> None:
        self.lines.append("INFO: " + message)

    def warn(self, message: str) -> None:
        self.warning_count += 1
        self.lines.append("WARNING: " + message)

    def error(self, message: str) -> None:
        self.error_count += 1
        self.lines.append("ERROR: " + message)


def author_name() -> str:
    return "".join(chr(value) for value in (72, 111, 119, 97, 114, 100))


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(".%s.%s.tmp" % (path.name, os.getpid()))
    temporary.write_text(content, encoding="utf-8")
    os.replace(str(temporary), str(path))


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        while True:
            chunk = file_obj.read(1024 * 1024)
            if not chunk:
                return digest.hexdigest()
            digest.update(chunk)


def read_json(path: Path, label: str, report: Report) -> Dict[str, object]:
    if not path.is_file():
        report.error("required %s is missing: %s" % (label, path))
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        report.error("cannot read %s %s: %s" % (label, path, exc))
        return {}
    if not isinstance(payload, dict):
        report.error("%s must contain a JSON object: %s" % (label, path))
        return {}
    return payload


def read_csv_rows(path: Path, label: str, report: Report) -> Tuple[List[str], List[Dict[str, str]]]:
    if not path.is_file():
        report.error("required %s is missing: %s" % (label, path))
        return [], []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            reader = csv.DictReader(file_obj)
            headers = [clean(item) for item in (reader.fieldnames or [])]
            if len(headers) != len(set(headers)):
                report.error("%s has duplicate headers: %s" % (label, path))
            return headers, [dict(row) for row in reader]
    except (OSError, csv.Error) as exc:
        report.error("cannot read %s %s: %s" % (label, path, exc))
        return [], []


def read_run_context(path: Path, report: Report) -> RunContext:
    headers, rows = read_csv_rows(path, "run context", report)
    expected = ["run_id", "mode_label", "design_revision", "note"]
    if headers != expected:
        report.error("run_context.csv headers must be exactly: " + ",".join(expected))
    if len(rows) != 1:
        report.error("run_context.csv must contain exactly one data row")
    row = rows[0] if rows else {}
    run_id = clean(row.get("run_id"))
    mode_label = clean(row.get("mode_label"))
    if not run_id:
        report.error("run_context.csv run_id is required")
    if not mode_label:
        report.error("run_context.csv mode_label is required")
    return RunContext(run_id, mode_label, clean(row.get("design_revision")), clean(row.get("note")))


def read_required_view(path: Path, stage: str, corner: str, report: Report) -> Optional[RequiredView]:
    headers, rows = read_csv_rows(path, "required views", report)
    if headers != REQUIRED_VIEW_HEADERS:
        report.error("required_views.csv headers must be exactly: " + ",".join(REQUIRED_VIEW_HEADERS))
    seen_ids: Dict[str, int] = {}
    seen_keys: Dict[Tuple[str, str], int] = {}
    selected: Optional[RequiredView] = None
    for row_idx, row in enumerate(rows, start=2):
        view_id = clean(row.get("view_id"))
        row_stage = clean(row.get("stage"))
        row_corner = clean(row.get("corner"))
        if not view_id or not row_stage or not row_corner:
            report.error("required_views.csv row %d has blank view_id/stage/corner" % row_idx)
        if view_id in seen_ids:
            report.error("required_views.csv duplicate view_id %s at rows %d and %d" % (view_id, seen_ids[view_id], row_idx))
        seen_ids[view_id] = row_idx
        key = (row_stage, row_corner)
        if key in seen_keys:
            report.error("required_views.csv duplicate stage/corner %s/%s at rows %d and %d" % (row_stage, row_corner, seen_keys[key], row_idx))
        seen_keys[key] = row_idx
        for flag in ("require_02", "require_04", "require_20", "require_30"):
            value = clean(row.get(flag)).lower()
            if value not in YES_NO:
                report.error("required_views.csv row %d %s must be yes/no" % (row_idx, flag))
        if key == (stage, corner):
            selected = RequiredView(view_id, stage, corner, clean(row.get("require_02")).lower() == "yes")
    if selected is None:
        report.error("stage/corner %s/%s is not declared in required_views.csv" % (stage, corner))
    return selected


def validate_provenance(payload: Dict[str, object], label: str, run: RunContext, report: Report) -> None:
    for field in ("run_id", "mode_label", "design_revision"):
        expected = clean(getattr(run, field))
        actual = clean(payload.get(field))
        if actual != expected:
            report.error("%s %s mismatch: expected %s, got %s" % (label, field, expected or "<blank>", actual or "<blank>"))


def load_inventory_context(run_root: Path, run: RunContext, report: Report) -> Optional[InventoryContext]:
    inventory_path = run_root / "01_middle/clock_inventory.csv"
    meta_path = run_root / "01_middle/clock_inventory.meta"
    completion_path = run_root / "01_middle/stage_completion.meta"
    default_sdc_path = run_root / "01_result/01_soc_clocks.sdc"
    meta = read_json(meta_path, "01 clock inventory meta", report)
    completion = read_json(completion_path, "01 completion meta", report)
    headers, rows = read_csv_rows(inventory_path, "01 clock inventory", report)
    required_headers = {"clock_name", "clock_kind", "period", "final_action"}
    missing_headers = sorted(required_headers - set(headers))
    if missing_headers:
        report.error("01 clock inventory missing headers: " + ",".join(missing_headers))
    if not meta or not completion or not inventory_path.is_file():
        return None

    inventory_digest = sha256_file(inventory_path)
    if clean(meta.get("inventory_digest")) != inventory_digest:
        report.error("01 clock inventory digest does not match clock_inventory.meta")
    if clean(completion.get("clock_inventory_digest")) != inventory_digest:
        report.error("01 clock inventory digest does not match stage_completion.meta")
    validate_provenance(meta, "01 clock inventory meta", run, report)
    validate_provenance(completion, "01 completion meta", run, report)
    if clean(meta.get("stage_name")) != "01_soc_clocks":
        report.error("01 clock inventory meta stage_name must be 01_soc_clocks")
    if clean(completion.get("stage_name")) != "01_soc_clocks":
        report.error("01 completion stage_name must be 01_soc_clocks")
    if clean(meta.get("completion_status")) != "complete":
        report.error("01 clock inventory meta is not complete")
    if clean(completion.get("completion_status")) != "complete":
        report.error("01 stage completion is not complete")
    try:
        if int(completion.get("error_count", 1)) != 0:
            report.error("01 stage completion error_count is not zero")
    except (TypeError, ValueError):
        report.error("01 stage completion error_count is invalid")
    if clean(completion.get("sync_changed")).lower() != "no":
        report.error("01 stage completion sync_changed must be no")
    structure_digest = clean(meta.get("structure_digest"))
    if not structure_digest or clean(completion.get("structure_digest")) != structure_digest:
        report.error("01 structure_digest is blank or inconsistent between meta/completion")

    declared_sdc = clean(meta.get("final_sdc_path"))
    final_sdc_path = Path(declared_sdc).expanduser() if declared_sdc else default_sdc_path
    if not final_sdc_path.is_absolute():
        final_sdc_path = (run_root / final_sdc_path).resolve()
    if final_sdc_path.resolve() != default_sdc_path.resolve():
        report.error("01 clock inventory meta final_sdc_path must resolve to %s" % default_sdc_path.resolve())
    final_sdc_digest = ""
    if not default_sdc_path.is_file():
        report.error("required final 01 SDC is missing: %s" % default_sdc_path)
    else:
        final_sdc_digest = sha256_file(default_sdc_path)
        if clean(meta.get("final_sdc_digest")) != final_sdc_digest:
            report.error("01 final SDC digest does not match clock_inventory.meta")
        if clean(completion.get("output_sdc_digest")) != final_sdc_digest:
            report.error("01 final SDC digest does not match stage_completion.meta")

    clocks: Dict[str, ClockInfo] = {}
    completeness_values = set()
    missing_instances = set()
    for row_idx, row in enumerate(rows, start=2):
        action = clean(row.get("final_action"))
        if action not in ACTIVE_01_ACTIONS:
            continue
        name = clean(row.get("clock_name"))
        if not name:
            report.error("01 clock inventory row %d has active action but blank clock_name" % row_idx)
            continue
        if name in clocks:
            report.error("01 clock inventory contains duplicate active clock_name %s" % name)
            continue
        row_run = clean(row.get("run_id"))
        row_mode = clean(row.get("mode_label"))
        row_revision = clean(row.get("design_revision"))
        if row_run and row_run != run.run_id:
            report.error("01 clock inventory row %d run_id is stale" % row_idx)
        if row_mode and row_mode != run.mode_label:
            report.error("01 clock inventory row %d mode_label is stale" % row_idx)
        if row_revision and row_revision != run.design_revision:
            report.error("01 clock inventory row %d design_revision is stale" % row_idx)
        row_structure = clean(row.get("structure_digest"))
        if row_structure and row_structure != structure_digest:
            report.error("01 clock inventory row %d structure_digest is stale" % row_idx)
        row_sdc_digest = clean(row.get("final_sdc_digest"))
        if row_sdc_digest and row_sdc_digest != final_sdc_digest:
            report.error("01 clock inventory row %d final_sdc_digest is stale" % row_idx)
        completeness = clean(row.get("run_completeness"))
        if completeness:
            completeness_values.add(completeness)
        for item in clean(row.get("missing_instances")).split(";"):
            if item.strip():
                missing_instances.add(item.strip())
        clocks[name] = ClockInfo(
            name,
            clean(row.get("direction")),
            clean(row.get("clock_kind")),
            clean(row.get("period")),
            clean(row.get("direct_source")),
        )
    if not clocks:
        report.error("01 clock inventory contains no active emitted clocks")
    completeness = clean(meta.get("run_completeness")) or (sorted(completeness_values)[0] if completeness_values else "complete")
    meta_missing = meta.get("missing_instances", [])
    if isinstance(meta_missing, list):
        missing_instances.update(clean(item) for item in meta_missing if clean(item))
    if completeness == "partial":
        report.warn("01 run completeness is partial; missing harden SDC: %s" % (",".join(sorted(missing_instances)) or "<unspecified>"))
    elif completeness != "complete":
        report.error("01 run completeness must be complete or partial, got %s" % (completeness or "<blank>"))
    return InventoryContext(
        inventory_path, meta_path, completion_path, default_sdc_path, clocks,
        completeness, tuple(sorted(missing_instances)), inventory_digest,
        final_sdc_digest, structure_digest, run.run_id, run.mode_label,
        run.design_revision, sha256_file(completion_path),
    )


def style_header(ws, row_idx: int, headers: Sequence[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row_idx, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(13, min(30, len(header) + 3))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s1" % get_column_letter(len(headers))


def setup_sheet(wb: Workbook, name: str, headers: Sequence[str]) -> object:
    ws = wb.create_sheet(name)
    style_header(ws, 1, headers)
    if "apply" in headers:
        col = get_column_letter(headers.index("apply") + 1)
        validation = DataValidation(type="list", formula1='"yes,no"', allow_blank=False)
        ws.add_data_validation(validation)
        validation.add("%s2:%s1048576" % (col, col))
    for flag in ("propagated", "managed_by_flow"):
        if flag in headers:
            col = get_column_letter(headers.index(flag) + 1)
            validation = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
            ws.add_data_validation(validation)
            validation.add("%s2:%s1048576" % (col, col))
    return ws


def append_budget_row(ws, stage: str, corner: str, clock: ClockInfo, inventory_digest: str, status: str) -> None:
    mapping = dict((header, index + 1) for index, header in enumerate(CLOCK_BUDGET_HEADERS))
    row_idx = ws.max_row + 1
    values = {
        "stage": stage, "corner": corner, "clock_name": clock.clock_name,
        "clock_kind": clock.clock_kind, "period": clock.period,
        "propagated": "no", "apply": "", "sync_status": status,
        "source_inventory_digest": inventory_digest, "note": "",
    }
    for header, value in values.items():
        ws.cell(row_idx, mapping[header], value)
    fill = NEW_FILL if status == "NEW_FROM_01" else CLEAR_FILL
    for col_idx in range(1, len(CLOCK_BUDGET_HEADERS) + 1):
        ws.cell(row_idx, col_idx).fill = fill
        ws.cell(row_idx, col_idx).border = THIN_BORDER


def metadata_rows(run: RunContext, view: RequiredView, inventory: InventoryContext) -> List[Tuple[str, str]]:
    return [
        ("Author", author_name()),
        ("run_id", run.run_id),
        ("mode_label", run.mode_label),
        ("design_revision", run.design_revision),
        ("view_id", view.view_id),
        ("stage", view.stage),
        ("corner", view.corner),
        ("run completeness", inventory.completeness),
        ("structure digest", inventory.structure_digest),
        ("01 inventory path", str(inventory.path.resolve())),
        ("01 inventory digest", inventory.inventory_digest),
        ("01 completion path", str(inventory.completion_path.resolve())),
        ("01 completion digest", inventory.completion_digest),
        ("Port accounting", "not_applicable"),
        ("added_bits", "0"),
    ]


def update_metadata(wb: Workbook, run: RunContext, view: RequiredView, inventory: InventoryContext) -> bool:
    expected = metadata_rows(run, view, inventory)
    if "runtime_metadata" in wb.sheetnames:
        ws = wb["runtime_metadata"]
        current = [(clean(ws.cell(row, 1).value), clean(ws.cell(row, 2).value)) for row in range(2, ws.max_row + 1)]
        if current == expected:
            return False
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("runtime_metadata", 0)
    ws.cell(1, 1, "Field")
    ws.cell(1, 2, "Value")
    style_header(ws, 1, ["Field", "Value"])
    for row_idx, (field, value) in enumerate(expected, start=2):
        ws.cell(row_idx, 1, field).font = Font(bold=True)
        ws.cell(row_idx, 2, value)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    return True


def save_workbook_atomic(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(".%s.%s.tmp.xlsx" % (path.stem, os.getpid()))
    wb.save(str(temporary))
    os.replace(str(temporary), str(path))


def create_workbook(path: Path, run: RunContext, view: RequiredView, inventory: InventoryContext) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = setup_sheet(wb, "clock_budget", CLOCK_BUDGET_HEADERS)
    setup_sheet(wb, "clock_pair_uncertainty", PAIR_HEADERS)
    setup_sheet(wb, "derate_ocv", DERATE_HEADERS)
    update_metadata(wb, run, view, inventory)
    for name in sorted(inventory.clocks):
        append_budget_row(ws, view.stage, view.corner, inventory.clocks[name], inventory.inventory_digest, "NEW_FROM_01")
    if ws.max_row >= 2:
        table = Table(displayName="clock_budget_table", ref="A1:%s%d" % (get_column_letter(len(CLOCK_BUDGET_HEADERS)), ws.max_row))
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    save_workbook_atomic(wb, path)


def sheet_mapping(ws, expected: Sequence[str], label: str, report: Report) -> Tuple[Dict[str, int], bool]:
    headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    changed = False
    duplicate = sorted(item for item in set(headers) if item and headers.count(item) > 1)
    if duplicate:
        report.error("%s has duplicate headers: %s" % (label, ",".join(duplicate)))
    for header in expected:
        if header not in headers:
            ws.cell(1, ws.max_column + 1, header)
            headers.append(header)
            changed = True
            report.warn("%s added missing header %s" % (label, header))
    mapping = dict((header, headers.index(header) + 1) for header in expected if header in headers)
    style_header(ws, 1, headers)
    return mapping, changed


def read_sheet_rows(ws, mapping: Dict[str, int], key_headers: Iterable[str]) -> List[SheetRow]:
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        values = dict((header, ws.cell(row_idx, col_idx).value) for header, col_idx in mapping.items())
        if any(clean(values.get(header)) for header in key_headers):
            rows.append(SheetRow(row_idx, values))
    return rows


def row_is_view(row: SheetRow, stage: str, corner: str) -> bool:
    return clean(row.values.get("stage")) == stage and clean(row.values.get("corner")) == corner


def apply_row_fill(ws, row_idx: int, status: str) -> None:
    fill = CLEAR_FILL
    if status == "NEW_FROM_01":
        fill = NEW_FILL
    elif status == "STALE_NOT_IN_01":
        fill = STALE_FILL
    elif status == "BLOCKED_BY_MISSING_SDC":
        fill = BLOCKED_FILL
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row_idx, col_idx).fill = fill


def row_has_intent(values: Dict[str, object]) -> bool:
    if clean(values.get("apply")).lower() == "no" and clean(values.get("note")):
        return True
    if clean(values.get("propagated")).lower() == "yes":
        return True
    return any(clean(values.get(column)) for column in NUMERIC_COLUMNS)


def sync_workbook(wb: Workbook, path: Path, view: RequiredView, inventory: InventoryContext, report: Report) -> Tuple[object, Dict[str, int], List[SheetRow]]:
    if "clock_budget" not in wb.sheetnames:
        report.error("workbook is missing required sheet clock_budget")
        return None, {}, []
    ws = wb["clock_budget"]
    mapping, header_changed = sheet_mapping(ws, CLOCK_BUDGET_HEADERS, "clock_budget", report)
    if header_changed:
        report.sync_changed = True
    rows = read_sheet_rows(ws, mapping, ("stage", "corner", "clock_name"))
    current = [row for row in rows if row_is_view(row, view.stage, view.corner)]
    by_name: Dict[str, List[SheetRow]] = {}
    for row in current:
        by_name.setdefault(clean(row.values.get("clock_name")), []).append(row)
    for name in sorted(inventory.clocks):
        if name not in by_name:
            append_budget_row(ws, view.stage, view.corner, inventory.clocks[name], inventory.inventory_digest, "NEW_FROM_01")
            report.warn("added clock %s from 01 inventory to %s/%s" % (name, view.stage, view.corner))
            report.sync_changed = True
    rows = read_sheet_rows(ws, mapping, ("stage", "corner", "clock_name"))
    for row in rows:
        if not row_is_view(row, view.stage, view.corner):
            continue
        name = clean(row.values.get("clock_name"))
        status = clean(row.values.get("sync_status"))
        if name not in inventory.clocks:
            desired = "BLOCKED_BY_MISSING_SDC" if inventory.completeness == "partial" else "STALE_NOT_IN_01"
            if status != desired:
                ws.cell(row.row_idx, mapping["sync_status"], desired)
                report.sync_changed = True
            apply_row_fill(ws, row.row_idx, desired)
            continue
        clock = inventory.clocks[name]
        expected_machine = {
            "clock_kind": clock.clock_kind,
            "period": clock.period,
            "source_inventory_digest": inventory.inventory_digest,
        }
        machine_changed = False
        for column, expected in expected_machine.items():
            if clean(row.values.get(column)) != expected:
                ws.cell(row.row_idx, mapping[column], expected)
                machine_changed = True
        if machine_changed:
            ws.cell(row.row_idx, mapping["sync_status"], "NEW_FROM_01")
            apply_row_fill(ws, row.row_idx, "NEW_FROM_01")
            report.warn("synchronized machine context for clock %s" % name)
            report.sync_changed = True
        elif status in {"NEW_FROM_01", "STALE_NOT_IN_01", "BLOCKED_BY_MISSING_SDC"} and row_has_intent(row.values):
            ws.cell(row.row_idx, mapping["sync_status"], "OK")
            apply_row_fill(ws, row.row_idx, "OK")
            report.info("reset reviewed sync_status to OK for clock %s" % name)
            report.sync_changed = True
    if report.sync_changed:
        report.info("workbook synchronization changed %s" % path)
    return ws, mapping, read_sheet_rows(ws, mapping, ("stage", "corner", "clock_name"))


def parse_number(value: object, label: str, report: Report) -> Optional[str]:
    text = clean(value)
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        report.error("%s is not numeric: %s" % (label, text))
        return None
    if not math.isfinite(number):
        report.error("%s must be finite" % label)
        return None
    if number < 0:
        report.error("%s must be non-negative" % label)
        return None
    return format(number, ".15g")


def validate_budget_rows(rows: Sequence[SheetRow], inventory: InventoryContext, view: RequiredView, report: Report) -> List[SheetRow]:
    current = [row for row in rows if row_is_view(row, view.stage, view.corner)]
    seen: Dict[Tuple[str, str, str], int] = {}
    selected_names = set()
    for row in current:
        values = row.values
        name = clean(values.get("clock_name"))
        key = (clean(values.get("stage")), clean(values.get("corner")), name)
        if key in seen:
            report.error("clock_budget duplicate stage/corner/clock_name %s/%s/%s at rows %d and %d" % (key[0], key[1], key[2], seen[key], row.row_idx))
        seen[key] = row.row_idx
        if not name:
            report.error("clock_budget row %d has blank clock_name" % row.row_idx)
            continue
        selected_names.add(name)
        if name not in inventory.clocks:
            report.error("clock_budget row %d clock %s is not in current 01 inventory" % (row.row_idx, name))
        status = clean(values.get("sync_status"))
        if status not in SYNC_VALUES:
            report.error("clock_budget row %d invalid sync_status %s" % (row.row_idx, status))
        if status not in SYNC_OK:
            report.error("clock_budget row %d clock %s has blocking sync_status %s" % (row.row_idx, name, status))
        apply_value = clean(values.get("apply")).lower()
        if apply_value not in YES_NO:
            report.error("clock_budget row %d apply must be yes/no" % row.row_idx)
        propagated = clean(values.get("propagated")).lower()
        if propagated not in {"", "yes", "no"}:
            report.error("clock_budget row %d propagated must be yes/no/blank" % row.row_idx)
        parsed = {}
        for column in NUMERIC_COLUMNS:
            parsed[column] = parse_number(values.get(column), "clock_budget row %d %s" % (row.row_idx, column), report)
        if apply_value == "yes" and not (any(parsed.values()) or propagated == "yes"):
            report.error("clock_budget row %d apply=yes but no supported command is defined" % row.row_idx)
        if apply_value == "no" and not clean(values.get("note")):
            report.error("clock_budget row %d apply=no requires a note" % row.row_idx)
        if view.stage.lower().startswith("pre") and propagated == "yes":
            report.warn("clock_budget row %d uses propagated=yes in a pre-CTS-like stage" % row.row_idx)
        if view.stage.lower().startswith("post") and propagated == "no" and apply_value == "yes":
            report.warn("clock_budget row %d keeps propagated=no in a post-CTS-like stage" % row.row_idx)
    missing = sorted(set(inventory.clocks) - selected_names)
    extra = sorted(selected_names - set(inventory.clocks))
    if missing:
        report.error("current view is missing 01 clock(s): " + ",".join(missing))
    if extra:
        report.error("current view contains stale clock(s): " + ",".join(extra))
    if not current:
        report.error("clock_budget contains no rows for %s/%s" % (view.stage, view.corner))
    return current


def validate_aux_sheets(wb: Workbook, inventory: InventoryContext, view: RequiredView, report: Report) -> None:
    for name, headers in (("clock_pair_uncertainty", PAIR_HEADERS), ("derate_ocv", DERATE_HEADERS)):
        if name not in wb.sheetnames:
            report.error("workbook is missing required sheet %s" % name)
            continue
        ws = wb[name]
        mapping, changed = sheet_mapping(ws, headers, name, report)
        if changed:
            report.sync_changed = True
        key_headers = ("stage", "corner", "from_clock", "to_clock") if name == "clock_pair_uncertainty" else ("stage", "corner", "derate_scope", "object_type")
        rows = read_sheet_rows(ws, mapping, key_headers)
        for row in rows:
            if not row_is_view(row, view.stage, view.corner):
                continue
            values = row.values
            apply_value = clean(values.get("apply")).lower()
            if apply_value not in YES_NO:
                report.error("%s row %d apply must be yes/no" % (name, row.row_idx))
            if name == "clock_pair_uncertainty":
                for field in ("from_clock", "to_clock"):
                    clock_name = clean(values.get(field))
                    if not clock_name or clock_name not in inventory.clocks:
                        report.error("%s row %d %s is not a current 01 clock: %s" % (name, row.row_idx, field, clock_name or "<blank>"))
                setup = parse_number(values.get("setup_uncertainty"), "%s row %d setup_uncertainty" % (name, row.row_idx), report)
                hold = parse_number(values.get("hold_uncertainty"), "%s row %d hold_uncertainty" % (name, row.row_idx), report)
                if apply_value == "yes" and not (setup or hold):
                    report.error("%s row %d apply=yes but setup/hold are both blank" % (name, row.row_idx))
                if apply_value == "yes":
                    report.warn("%s row %d is validated but pair command emission is not enabled" % (name, row.row_idx))
            else:
                managed = clean(values.get("managed_by_flow")).lower()
                if managed not in YES_NO:
                    report.error("%s row %d managed_by_flow must be yes/no" % (name, row.row_idx))
                early = parse_number(values.get("early"), "%s row %d early" % (name, row.row_idx), report)
                late = parse_number(values.get("late"), "%s row %d late" % (name, row.row_idx), report)
                if apply_value == "yes" and not (early or late):
                    report.error("%s row %d apply=yes but early/late are both blank" % (name, row.row_idx))
                if managed == "yes":
                    report.warn("%s row %d derate is managed by external flow; stage 02 will not emit it" % (name, row.row_idx))
                elif apply_value == "yes":
                    report.warn("%s row %d is validated but derate command emission is not enabled" % (name, row.row_idx))


def tcl_object(name: str) -> str:
    return "[get_clocks {%s}]" % name.replace("}", "\\}")


def generate_sdc(rows: Sequence[SheetRow], run: RunContext, view: RequiredView, inventory: InventoryContext) -> str:
    lines = [
        "# Auto-generated by 02_extract_soc_clock_timing.py",
        "# Author: %s" % author_name(),
        "# Run ID: %s" % run.run_id,
        "# Mode label: %s" % run.mode_label,
        "# Design revision: %s" % run.design_revision,
        "# View ID: %s" % view.view_id,
        "# Stage: %s" % view.stage,
        "# Corner: %s" % view.corner,
        "# Structure digest: %s" % inventory.structure_digest,
        "# 01 inventory digest: %s" % inventory.inventory_digest,
        "# Port accounting: not_applicable; added_bits=0",
        "",
    ]
    command_map = (
        ("setup_uncertainty", "set_clock_uncertainty -setup {value} {object}"),
        ("hold_uncertainty", "set_clock_uncertainty -hold {value} {object}"),
        ("source_latency_early", "set_clock_latency -source -early {value} {object}"),
        ("source_latency_late", "set_clock_latency -source -late {value} {object}"),
        ("network_latency_early", "set_clock_latency -early {value} {object}"),
        ("network_latency_late", "set_clock_latency -late {value} {object}"),
        ("transition_min", "set_clock_transition -min {value} {object}"),
        ("transition_max", "set_clock_transition -max {value} {object}"),
    )
    emitted = 0
    for row in sorted(rows, key=lambda item: clean(item.values.get("clock_name"))):
        values = row.values
        if clean(values.get("apply")).lower() != "yes" or clean(values.get("sync_status")) not in SYNC_OK:
            continue
        name = clean(values.get("clock_name"))
        lines.append("# clock_budget row %d: %s" % (row.row_idx, name))
        object_expr = tcl_object(name)
        for column, template in command_map:
            value = parse_number(values.get(column), "internal generation %s" % column, Report())
            if value is not None:
                lines.append(template.format(value=value, object=object_expr))
                emitted += 1
        if clean(values.get("propagated")).lower() == "yes":
            lines.append("set_propagated_clock %s" % object_expr)
            emitted += 1
        lines.append("")
    if not emitted:
        lines.append("# No clock timing commands emitted for this view.")
    return "\n".join(lines).rstrip() + "\n"


def semantic_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return str(value)
        return format(value, ".15g")
    return clean(value)


def workbook_semantic_digest(wb: Workbook) -> str:
    payload = {}
    for sheet_name, expected in (
        ("clock_budget", CLOCK_BUDGET_HEADERS),
        ("clock_pair_uncertainty", PAIR_HEADERS),
        ("derate_ocv", DERATE_HEADERS),
    ):
        ws = wb[sheet_name]
        headers = [clean(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        ordered = [header for header in expected if header in headers]
        mapping = dict((header, headers.index(header) + 1) for header in ordered)
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row = dict((header, semantic_value(ws.cell(row_idx, mapping[header]).value)) for header in ordered)
            if any(value != "" for value in row.values()):
                rows.append(row)
        payload[sheet_name] = {"headers": ordered, "rows": rows}
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return sha256_bytes(canonical)


def resolved_manifest_payload(rows: Sequence[SheetRow], run: RunContext, view: RequiredView, inventory: InventoryContext, workbook_digest: str, output_digest: str) -> Dict[str, object]:
    resolved = []
    for row in sorted(rows, key=lambda item: clean(item.values.get("clock_name"))):
        values = row.values
        resolved.append({
            "row": row.row_idx,
            "clock_name": clean(values.get("clock_name")),
            "apply": clean(values.get("apply")).lower(),
            "sync_status": clean(values.get("sync_status")),
            "emitted": clean(values.get("apply")).lower() == "yes" and clean(values.get("sync_status")) in SYNC_OK,
        })
    return {
        "schema_version": "1.0",
        "author": author_name(),
        "stage_name": "02_soc_clock_timing",
        "run_id": run.run_id,
        "mode_label": run.mode_label,
        "design_revision": run.design_revision,
        "view_id": view.view_id,
        "stage": view.stage,
        "corner": view.corner,
        "structure_digest": inventory.structure_digest,
        "upstream_01_inventory_digest": inventory.inventory_digest,
        "upstream_01_completion_digest": inventory.completion_digest,
        "workbook_semantic_digest": workbook_digest,
        "output_sdc_digest": output_digest,
        "port_accounting": "not_applicable",
        "added_bits": 0,
        "rows": resolved,
    }


def render_report(report: Report, run: RunContext, view: Optional[RequiredView], inventory: Optional[InventoryContext], diagnose_only: bool, form_path: Path, output_path: Path, completion_path: Path) -> str:
    lines = [
        "02 SoC Clock Timing Check Report",
        "Author: %s" % author_name(),
        "Run ID: %s" % run.run_id,
        "Mode label: %s" % run.mode_label,
        "Design revision: %s" % run.design_revision,
        "View ID: %s" % (view.view_id if view else ""),
        "Stage: %s" % (view.stage if view else ""),
        "Corner: %s" % (view.corner if view else ""),
        "Required by 02: %s" % ("yes" if view and view.require_02 else "no"),
        "Diagnose only: %s" % ("yes" if diagnose_only else "no"),
        "Run completeness: %s" % (inventory.completeness if inventory else "unknown"),
        "Structure digest: %s" % (inventory.structure_digest if inventory else ""),
        "01 inventory digest: %s" % (inventory.inventory_digest if inventory else ""),
        "01 completion digest: %s" % (inventory.completion_digest if inventory else ""),
        "Workbook: %s" % form_path,
        "Output SDC: %s" % output_path,
        "Completion meta: %s" % completion_path,
        "Port accounting: not_applicable; added_bits=0",
        "Sync changed: %s" % ("yes" if report.sync_changed else "no"),
        "Warning count: %d" % report.warning_count,
        "Error count: %d" % report.error_count,
        "",
    ]
    lines.extend(report.lines)
    return "\n".join(lines).rstrip() + "\n"


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate one single-run stage 02 clock timing view")
    parser.add_argument("--run-root", required=True, help="single-run SoC SDC root")
    parser.add_argument("--stage", required=True, help="timing stage, for example prects")
    parser.add_argument("--corner", required=True, help="case-sensitive timing corner/view")
    parser.add_argument("--diagnose-only", action="store_true", help="validate/synchronize without formal SDC or completion publication")
    return parser.parse_args(argv)


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)
    run_root = Path(args.run_root).expanduser().resolve()
    stage = clean(args.stage)
    corner = clean(args.corner)
    if not stage or not corner or "/" in stage or "/" in corner or "\\" in stage or "\\" in corner:
        print("ERROR: stage and corner must be non-blank filename-safe tokens", file=sys.stderr)
        return 2

    form_path = run_root / ("02_middle/02_soc_clock_timing_budget_%s.xlsx" % stage)
    output_path = run_root / ("02_result/02_soc_clock_timing_%s_%s.sdc" % (stage, corner))
    manifest_path = run_root / ("02_middle/resolved/%s_%s.manifest" % (stage, corner))
    completion_path = run_root / ("02_middle/completion/%s_%s.meta" % (stage, corner))
    report_path = run_root / ("02_result/reports/clock_timing_check_report_%s_%s.txt" % (stage, corner))
    report = Report()
    run = read_run_context(run_root / "inputs/run_context.csv", report)
    view = read_required_view(run_root / "inputs/required_views.csv", stage, corner, report)
    inventory = load_inventory_context(run_root, run, report)

    print("Run root: %s" % run_root)
    print("Run ID: %s" % run.run_id)
    print("Mode label: %s" % run.mode_label)
    print("Stage: %s" % stage)
    print("Corner: %s" % corner)
    print("Port accounting: not_applicable; added_bits=0")

    if not args.diagnose_only and completion_path.exists():
        completion_path.unlink()
        report.info("invalidated prior completion meta before formal rerun")

    if view is not None and not view.require_02 and not args.diagnose_only:
        report.error("view %s (%s/%s) has require_02=no; use --diagnose-only" % (view.view_id, stage, corner))

    if report.error_count or view is None or inventory is None:
        atomic_write_text(report_path, render_report(report, run, view, inventory, args.diagnose_only, form_path, output_path, completion_path))
        print("Report: %s" % report_path)
        return 1

    if not form_path.is_file():
        create_workbook(form_path, run, view, inventory)
        report.sync_changed = True
        report.warn("created new stage workbook; fill timing values and review NEW_FROM_01 rows")
        atomic_write_text(report_path, render_report(report, run, view, inventory, args.diagnose_only, form_path, output_path, completion_path))
        print("Workbook: %s" % form_path)
        print("Report: %s" % report_path)
        return 1

    try:
        wb = load_workbook(str(form_path))
    except Exception as exc:
        report.error("cannot load workbook %s: %s" % (form_path, exc))
        atomic_write_text(report_path, render_report(report, run, view, inventory, args.diagnose_only, form_path, output_path, completion_path))
        return 1

    ws, mapping, rows = sync_workbook(wb, form_path, view, inventory, report)
    metadata_changed = update_metadata(wb, run, view, inventory)
    validate_aux_sheets(wb, inventory, view, report)
    if report.sync_changed or metadata_changed:
        save_workbook_atomic(wb, form_path)
    if report.sync_changed:
        report.warn("workbook changed during synchronization; review and rerun before generation")
        atomic_write_text(report_path, render_report(report, run, view, inventory, args.diagnose_only, form_path, output_path, completion_path))
        print("Workbook: %s" % form_path)
        print("Report: %s" % report_path)
        return 1

    current_rows = validate_budget_rows(rows, inventory, view, report)
    if report.error_count:
        atomic_write_text(report_path, render_report(report, run, view, inventory, args.diagnose_only, form_path, output_path, completion_path))
        print("Report: %s" % report_path)
        return 1

    workbook_digest = workbook_semantic_digest(wb)
    if args.diagnose_only:
        report.info("diagnose-only completed; formal SDC, manifest, and completion were not published")
        atomic_write_text(report_path, render_report(report, run, view, inventory, True, form_path, output_path, completion_path))
        print("Diagnostic validation passed.")
        print("Report: %s" % report_path)
        return 0

    sdc_text = generate_sdc(current_rows, run, view, inventory)
    output_digest = sha256_bytes(sdc_text.encode("utf-8"))
    manifest = resolved_manifest_payload(current_rows, run, view, inventory, workbook_digest, output_digest)
    atomic_write_text(output_path, sdc_text)
    atomic_write_text(manifest_path, json.dumps(manifest, ensure_ascii=False, sort_keys=True, indent=2) + "\n")
    completion = {
        "schema_version": "1.0",
        "author": author_name(),
        "stage_name": "02_soc_clock_timing",
        "run_id": run.run_id,
        "mode_label": run.mode_label,
        "design_revision": run.design_revision,
        "view_id": view.view_id,
        "stage": stage,
        "corner": corner,
        "completion_status": "complete",
        "error_count": 0,
        "sync_changed": "no",
        "structure_digest": inventory.structure_digest,
        "upstream_01_inventory_path": str(inventory.path.resolve()),
        "upstream_01_inventory_digest": inventory.inventory_digest,
        "upstream_01_completion_path": str(inventory.completion_path.resolve()),
        "upstream_01_completion_digest": inventory.completion_digest,
        "run_completeness": inventory.completeness,
        "workbook_path": str(form_path.resolve()),
        "workbook_semantic_digest": workbook_digest,
        "resolved_manifest_path": str(manifest_path.resolve()),
        "resolved_manifest_digest": sha256_file(manifest_path),
        "output_sdc_path": str(output_path.resolve()),
        "output_sdc_digest": output_digest,
        "port_accounting": "not_applicable",
        "added_bits": 0,
        "port_accounting_summary": "Port accounting: not_applicable; added_bits=0",
    }
    atomic_write_text(completion_path, json.dumps(completion, ensure_ascii=False, sort_keys=True, indent=2) + "\n")
    report.info("published formal SDC, resolved manifest, and completion meta")
    atomic_write_text(report_path, render_report(report, run, view, inventory, False, form_path, output_path, completion_path))
    print("SDC: %s" % output_path)
    print("Manifest: %s" % manifest_path)
    print("Completion: %s" % completion_path)
    print("Report: %s" % report_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
