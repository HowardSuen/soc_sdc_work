#!/usr/bin/env python3
"""Regression for the single-run stage-30 accounting/finalization contract."""

import csv
import hashlib
import importlib.util
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parent
SOC = BASE.parent.parent
EX00 = SOC / "00_harden_port_inventory" / "00_harden_port_inventory.py"
EX30 = SOC / "30_harden_to_harden_exception" / "30_extract_harden_to_harden_exception.py"
WORK = BASE / "work_target"
DELTA_HEADERS = [
    "schema_version", "run_id", "mode_label", "stage_name", "transaction_id",
    "view_id", "stage", "corner", "structure_digest",
    "accounting_digest_before", "accounting_digest_after", "workbook", "sheet",
    "row", "direction", "port", "legal_bits", "added_bits",
    "final_used_bits", "owner_object_id", "reason", "evidence_status",
]


def require(condition, message):
    if not condition:
        raise AssertionError(message)


def load_module(path, name):
    spec = importlib.util.spec_from_file_location(name, str(path))
    require(spec is not None and spec.loader is not None, "failed to load {0}".format(path))
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def run(script, root, args):
    return subprocess.run(
        [sys.executable, str(script)] + list(args), cwd=str(root),
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True,
    )


def write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def digest(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def canonical_digest(value):
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def write_json(path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n", encoding="utf-8")


def create_inputs(root, views=None, require_upstream_sdc=False):
    if root.exists():
        shutil.rmtree(str(root))
    inputs = root / "inputs"
    inputs.mkdir(parents=True)
    write_csv(inputs / "run_context.csv", ["run_id", "mode_label", "design_revision", "note"], [{
        "run_id": "reg30", "mode_label": "func", "design_revision": "revA", "note": "regression",
    }])
    if views is None:
        views = [("all_all", "all", "all")]
    write_csv(inputs / "required_views.csv", [
        "view_id", "stage", "corner", "require_02", "require_04", "require_20", "require_30", "note",
    ], [{
        "view_id": view_id, "stage": stage, "corner": corner,
        "require_02": "yes" if require_upstream_sdc else "no",
        "require_04": "yes" if require_upstream_sdc else "no",
        "require_20": "yes" if require_upstream_sdc else "no",
        "require_30": "yes", "note": "",
    } for view_id, stage, corner in views])

    wb = Workbook()
    ws = wb.active
    ws.append(["module_name", "inst_name", "owner", "sdc_path"])
    ws.append(["ha", "u_a", "alice", "u_a.sdc"])
    ws.append(["hb", "u_b", "bob", "u_b.sdc"])
    wb.save(str(inputs / "info_all.xlsx"))
    (inputs / "u_a.sdc").write_text(
        "set_false_path -from [get_ports sig_o]\n", encoding="utf-8"
    )
    (inputs / "u_b.sdc").write_text("# no boundary exception\n", encoding="utf-8")

    headers = [
        "Input", "Input Width", "Input Used Width", "From Whom",
        "Output", "Output Width", "Output Used Width", "To Top",
        "Inout", "Inout Width", "Inout Connectivity", "Inout Name",
    ]
    port_wb = Workbook()
    port_wb.remove(port_wb.active)
    ws_a = port_wb.create_sheet("u_a")
    ws_a.append(headers)
    ws_a.append(["", "", "", "", "sig_o[1:0]", 2, "", "", "", "", "", ""])
    ws_b = port_wb.create_sheet("u_b")
    ws_b.append(headers)
    ws_b.append(["sig_i[1:0]", 2, "", "u_a.sig_o[1:0]", "", "", "", "", "", "", "", ""])
    port_wb.save(str(inputs / "port_hardens.xlsx"))


def empty_delta_stage(root, directory, stage_name, structure, accounting):
    delta = root / directory / "port_accounting_delta.csv"
    write_csv(delta, DELTA_HEADERS, [])
    txn_id = directory[:2] + "_reg"
    empty_rows_digest = canonical_digest(["1.0", []])
    meta = {
        "schema_version": "1.0", "run_id": "reg30", "mode_label": "func",
        "design_revision": "revA", "stage_name": stage_name,
        "completion_status": "complete", "structure_digest": structure,
        "accounting_digest_before": accounting, "accounting_digest_after": accounting,
        "delta_csv_digest": digest(delta),
        "transactions": [{
            "transaction_id": txn_id, "committed_at": "2026-07-17T00:00:00Z",
            "structure_digest": structure, "accounting_digest_before": accounting,
            "accounting_digest_after": accounting, "delta_rows_digest": empty_rows_digest,
        }],
    }
    write_json(root / directory / "port_accounting_delta.meta", meta)
    completion = dict(meta)
    completion.update({"error_count": 0, "sync_changed": "no", "accounting_delta_digest": digest(delta)})
    write_json(root / directory / "stage_completion.meta", completion)


def connection_id(bit):
    value = ["1.0", "u_a", "output", "sig_o", bit, "u_b", "input", "sig_i", bit]
    return "CONN_" + canonical_digest(value)


def fixture_upstream_output(root, label, stage, corner):
    if label == "02":
        return root / "02_result" / (
            "02_soc_clock_timing_{0}_{1}.sdc".format(stage, corner)
        )
    if label == "04":
        name = (
            "04_soc_io_pads.sdc"
            if stage == "all" and corner == "all"
            else "04_soc_io_pads_{0}_{1}.sdc".format(stage, corner)
        )
        return root / "04_result" / name
    if label == "20":
        name = (
            "20_harden_x_if.sdc"
            if stage == "all" and corner == "all"
            else "20_harden_x_if_{0}_{1}.sdc".format(stage, corner)
        )
        return root / "20_result" / name
    raise AssertionError("unsupported upstream fixture stage: {0}".format(label))


def create_required_upstream_view(root, label, stage_name, view,
                                  structure, accounting):
    stage = view["stage"]
    corner = view["corner"]
    view_id = view["view_id"]
    output = fixture_upstream_output(root, label, stage, corner)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        "# regression {0} formal SDC for {1}\n".format(label, view_id),
        encoding="utf-8",
    )
    directory = "{0}_middle".format(label)
    if label in {"04", "20"}:
        stage_completion_path = root / directory / "stage_completion.meta"
        completion = json.loads(stage_completion_path.read_text(encoding="utf-8"))
    else:
        completion = {
            "schema_version": "1.0", "run_id": "reg30", "mode_label": "func",
            "design_revision": "revA", "completion_status": "complete",
            "error_count": 0, "sync_changed": "no",
            "structure_digest": structure,
            "accounting_digest_before": accounting,
            "accounting_digest_after": accounting,
        }
    completion.update({
        "stage_name": stage_name, "stage": stage, "corner": corner,
        "view_id": view_id, "output_sdc_digest": digest(output),
    })
    if label == "02":
        completion["output_sdc_path"] = str(output.resolve())
    view_path = root / directory / "completion" / (
        "{0}_{1}.meta".format(stage, corner)
    )
    write_json(view_path, completion)
    if label in {"04", "20"}:
        stage_completion_path = root / directory / "stage_completion.meta"
        stage_completion = json.loads(
            stage_completion_path.read_text(encoding="utf-8")
        )
        authenticated = dict(stage_completion.get("required_view_completions", {}))
        authenticated[view_id] = digest(view_path)
        stage_completion["required_view_completions"] = authenticated
        stage_completion["output_sdc_digest"] = digest(output)
        write_json(stage_completion_path, stage_completion)


def create_upstreams(root):
    meta00 = json.loads((root / "00_middle" / "port_accounting_delta.meta").read_text(encoding="utf-8"))
    structure = meta00["structure_digest"]
    accounting = meta00["accounting_digest_after"]
    for directory, stage_name in (
        ("01_middle", "01_soc_clocks"),
        ("04_middle", "04_soc_io_pads"),
        ("10_middle", "10_feedthrough"),
        ("20_middle", "20_harden_x_if"),
    ):
        empty_delta_stage(root, directory, stage_name, structure, accounting)

    output_01 = root / "01_result" / "01_soc_clocks.sdc"
    output_01.parent.mkdir(parents=True, exist_ok=True)
    output_01.write_text("# regression 01 formal SDC\n", encoding="utf-8")
    completion_01_path = root / "01_middle" / "stage_completion.meta"
    completion_01 = json.loads(completion_01_path.read_text(encoding="utf-8"))
    completion_01["output_sdc_digest"] = digest(output_01)
    write_json(completion_01_path, completion_01)

    clock = root / "01_middle" / "clock_inventory.csv"
    write_csv(clock, [
        "run_id", "mode_label", "structure_digest", "inst_name", "port_name",
        "direction", "clock_name", "original_clock_name", "final_action",
        "target_object", "from_whom", "direct_source", "root_source",
    ], [])
    write_json(root / "01_middle" / "clock_inventory.meta", {
        "schema_version": "1.0", "run_id": "reg30", "mode_label": "func",
        "design_revision": "revA", "stage_name": "01_soc_clocks",
        "completion_status": "complete", "structure_digest": structure,
        "inventory_digest": digest(clock),
    })

    completion03 = {
        "schema_version": "1.0", "run_id": "reg30", "mode_label": "func",
        "design_revision": "revA", "stage_name": "03_soc_clock_groups",
        "completion_status": "complete", "error_count": 0, "sync_changed": "no",
        "structure_digest": structure, "accounting_digest_before": accounting,
        "accounting_digest_after": accounting,
    }
    output_03 = root / "03_result" / "03_soc_clock_groups.sdc"
    output_03.parent.mkdir(parents=True, exist_ok=True)
    output_03.write_text("# regression 03 formal SDC\n", encoding="utf-8")
    completion03["output_sdc_path"] = str(output_03.resolve())
    completion03["output_sdc_digest"] = digest(output_03)
    write_json(root / "03_middle" / "stage_completion.meta", completion03)

    relation = root / "03_middle" / "relation_map.csv"
    write_csv(relation, ["run_id", "mode_label", "structure_digest", "clock_a", "clock_b", "relation_type"], [])
    write_json(root / "03_middle" / "relation_map.meta", {
        **completion03, "relation_map_digest": digest(relation),
    })

    pad = root / "04_middle" / "pad_inventory.csv"
    write_csv(pad, ["run_id", "mode_label", "structure_digest", "pad_id", "connection_id", "disposition"], [])
    write_json(root / "04_middle" / "pad_inventory.meta", {
        "schema_version": "1.0", "run_id": "reg30", "mode_label": "func",
        "design_revision": "revA", "stage_name": "04_soc_io_pads",
        "completion_status": "complete", "structure_digest": structure,
        "pad_inventory_digest": digest(pad),
    })

    feedthrough = root / "10_middle" / "feedthrough_edge_inventory.csv"
    write_csv(feedthrough, ["run_id", "mode_label", "structure_digest", "feedthrough_edge_id", "connection_id", "channel_disposition", "apply", "review_status", "emit_max", "emit_min"], [])
    write_json(root / "10_middle" / "feedthrough_edge_inventory.meta", {
        "schema_version": "1.0", "run_id": "reg30", "mode_label": "func",
        "design_revision": "revA", "stage_name": "10_feedthrough",
        "completion_status": "complete", "structure_digest": structure,
        "feedthrough_edge_inventory_digest": digest(feedthrough),
    })

    channel = root / "20_middle" / "channel_inventory.csv"
    rows = []
    for bit in (0, 1):
        rows.append({
            "run_id": "reg30", "mode_label": "func", "structure_digest": structure,
            "channel_id": "CH20_{0}".format(bit), "connection_id": connection_id(bit),
            "stage": "all", "corner": "all", "channel_disposition": "route_to_30",
            "apply": "yes", "review_status": "approved", "emit_max": "no", "emit_min": "no",
        })
    write_csv(channel, list(rows[0]), rows)
    write_json(root / "20_middle" / "channel_inventory.meta", {
        "schema_version": "1.0", "run_id": "reg30", "mode_label": "func",
        "design_revision": "revA", "stage_name": "20_harden_x_if",
        "completion_status": "complete", "structure_digest": structure,
        "channel_inventory_digest": digest(channel),
    })

    output_10 = root / "10_result" / "10_feedthrough.sdc"
    output_10.parent.mkdir(parents=True, exist_ok=True)
    output_10.write_text("# regression 10 formal SDC\n", encoding="utf-8")
    completion_10_path = root / "10_middle" / "stage_completion.meta"
    completion_10 = json.loads(completion_10_path.read_text(encoding="utf-8"))
    completion_10["output_sdc_digest"] = digest(output_10)
    write_json(completion_10_path, completion_10)

    with (root / "inputs" / "required_views.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as file_obj:
        required_views = list(csv.DictReader(file_obj))
    for view in required_views:
        for flag, label, stage_name in (
            ("require_02", "02", "02_soc_clock_timing"),
            ("require_04", "04", "04_soc_io_pads"),
            ("require_20", "20", "20_harden_x_if"),
        ):
            if str(view.get(flag, "")).strip().lower() != "yes":
                continue
            create_required_upstream_view(
                root, label, stage_name, view, structure, accounting
            )


def approve(root):
    path = root / "30_middle" / "30_harden_to_harden_exception.xlsx"
    wb = load_workbook(str(path))
    ws = wb["exception_rule"]
    columns = {cell.value: cell.column for cell in ws[1]}
    count = 0
    for row in range(2, ws.max_row + 1):
        if not ws.cell(row, columns["exception_id"]).value:
            continue
        ws.cell(row, columns["apply"], "yes")
        ws.cell(row, columns["review_status"], "approved")
        ws.cell(row, columns["owner"], "architecture")
        ws.cell(row, columns["exception_type"], "false_path")
        ws.cell(row, columns["path_category"], "config")
        ws.cell(row, columns["basis"], "approved static configuration path with no functional timing requirement")
        count += 1
    wb.save(str(path))
    require(count >= 2 and count % 2 == 0, "expected exact bit rules for one or more views")


def force_stale_extraction_time(root):
    path = root / "30_middle" / "30_harden_to_harden_exception.xlsx"
    wb = load_workbook(str(path))
    ws = wb["exception_candidate"]
    columns = {cell.value: cell.column for cell in ws[1]}
    changed = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, columns["source_type"]).value != "extracted_harden_exception":
            continue
        ws.cell(row, columns["extraction_time"], "2000-01-01T00:00:00")
        changed += 1
    wb.save(str(path))
    require(changed == 2, "expected two extracted-exception candidates")


def setup_root(root, views=None, require_upstream_sdc=False):
    create_inputs(
        root, views=views, require_upstream_sdc=require_upstream_sdc
    )
    initialized = run(EX00, root, ["--run-root", str(root)])
    require(initialized.returncode == 0, "00 initialization failed:\n{0}\n{1}".format(initialized.stdout, initialized.stderr))
    create_upstreams(root)


def test_formal_flow():
    root = WORK / "formal"
    setup_root(root)
    first = run(EX30, root, ["--run-root", str(root), "--stage", "all", "--corner", "all"])
    require(first.returncode == 1, "first 30 run must stop for workbook review:\n{0}\n{1}".format(first.stdout, first.stderr))
    require("Completion status: review_required" in first.stdout, "review-required status missing")
    approve(root)
    second = run(EX30, root, ["--run-root", str(root), "--stage", "all", "--corner", "all"])
    require(second.returncode == 0, "approved 30 run failed:\n{0}\n{1}".format(second.stdout, second.stderr))
    sdc = (root / "30_result" / "30_harden_to_harden_exception.sdc").read_text(encoding="utf-8")
    require(sdc.count("set_false_path") == 2, "expected two exact-bit false paths")
    require("get_pins {u_a/sig_o[0]}" in sdc and "get_pins {u_b/sig_i[1]}" in sdc, "exact endpoints missing")
    require((root / "30_middle" / "port_accounting_delta.csv").is_file(), "30 delta missing")
    require((root / "30_middle" / "completion" / "all_all.meta").is_file(), "per-view completion missing")
    require((root / "30_middle" / "stage_completion.meta").is_file(), "run-wide completion missing")
    final_report = (root / "30_result" / "reports" / "port_accounting_final_report.txt").read_text(encoding="utf-8")
    require("Accounting closure: complete" in final_report, "final closure not complete")

    workbook = load_workbook(str(root / "inputs" / "port_hardens.xlsx"))
    for sheet_name, header in (("u_a", "Output Used Width"), ("u_b", "Input Used Width")):
        ws = workbook[sheet_name]
        columns = {cell.value: cell.column for cell in ws[1]}
        cell = ws.cell(2, columns[header])
        require(cell.value == "ALL USED", "final ALL USED token missing")
        require(cell.fill.fgColor.rgb[-6:] == "C6EFCE", "final green fill missing")
        require(cell.font.color.rgb[-6:] == "006100", "final green font missing")
    workbook.close()

    force_stale_extraction_time(root)
    rerun = run(EX30, root, ["--run-root", str(root), "--stage", "all", "--corner", "all"])
    require(rerun.returncode == 0, "idempotent rerun failed:\n{0}\n{1}".format(rerun.stdout, rerun.stderr))
    require("Completion status: complete" in rerun.stdout, "extraction_time refresh reopened the review gate")


def test_cli_and_diagnostic_gate():
    rejected = run(EX30, BASE, ["--run-root", str(BASE), "--scenario", "common"])
    require(rejected.returncode == 2 and "unrecognized arguments" in rejected.stderr, "--scenario must be rejected")
    root = WORK / "diagnostic"
    setup_root(root)
    before = digest(root / "inputs" / "port_hardens.xlsx")
    result = run(EX30, root, ["--run-root", str(root), "--stage", "prects", "--corner", "ss", "--diagnose-only"])
    require(result.returncode == 0, "non-required diagnostic failed:\n{0}\n{1}".format(result.stdout, result.stderr))
    require(digest(root / "inputs" / "port_hardens.xlsx") == before, "diagnostic modified port workbook")
    require(not (root / "30_middle" / "completion" / "prects_ss.meta").exists(), "diagnostic published completion")


def test_deferred_multiview_finalization():
    root = WORK / "multiview"
    setup_root(root, views=[("prects_ss", "prects", "ss"), ("postroute_ff", "postroute", "ff")])
    first_sync = run(EX30, root, ["--run-root", str(root), "--stage", "prects", "--corner", "ss"])
    require(first_sync.returncode == 1, "first multiview sync must require review")
    approve(root)
    first_done = run(EX30, root, [
        "--run-root", str(root), "--stage", "prects", "--corner", "ss", "--defer-final-accounting",
    ])
    require(first_done.returncode == 0 and "Final accounting: deferred" in first_done.stdout, "first view did not defer final accounting")
    require(not (root / "30_middle" / "stage_completion.meta").exists(), "deferred view published run-wide completion")
    workbook = load_workbook(str(root / "inputs" / "port_hardens.xlsx"), data_only=False)
    ws = workbook["u_a"]
    columns = {cell.value: cell.column for cell in ws[1]}
    require(ws.cell(2, columns["Output Used Width"]).value == "0,1", "deferred view wrote final token")
    workbook.close()

    second_sync = run(EX30, root, ["--run-root", str(root), "--stage", "postroute", "--corner", "ff"])
    require(second_sync.returncode == 1, "second multiview sync must require independent review")
    approve(root)
    second_done = run(EX30, root, ["--run-root", str(root), "--stage", "postroute", "--corner", "ff"])
    require(second_done.returncode == 0 and "Final accounting: complete" in second_done.stdout, "second view did not finalize accounting")
    require((root / "30_middle" / "stage_completion.meta").is_file(), "multiview run-wide completion missing")
    report = (root / "30_result" / "reports" / "port_accounting_final_report.txt").read_text(encoding="utf-8")
    require("prects_ss" in report and "postroute_ff" in report, "required-view completion matrix is incomplete")


def test_required_upstream_sdc_digest_gate():
    root = WORK / "upstream_sdc_digest"
    setup_root(
        root,
        views=[("prects_ss_125", "prects", "ss_125")],
        require_upstream_sdc=True,
    )
    args = ["--run-root", str(root), "--stage", "prects", "--corner", "ss_125"]
    first = run(EX30, root, args)
    require(
        first.returncode == 1,
        "upstream digest fixture did not stop for review:\n{0}\n{1}".format(
            first.stdout, first.stderr
        ),
    )
    approve(root)
    completed = run(EX30, root, args)
    require(
        completed.returncode == 0,
        "upstream digest baseline failed:\n{0}\n{1}".format(
            completed.stdout, completed.stderr
        ),
    )

    completion_30 = root / "30_middle" / "completion" / "prects_ss_125.meta"
    delta_meta_30 = root / "30_middle" / "port_accounting_delta.meta"
    output_30 = (
        root / "30_result" /
        "30_harden_to_harden_exception_prects_ss_125.sdc"
    )
    report_path = (
        root / "30_result" / "reports" /
        "harden_to_harden_exception_check_report_prects_ss_125.txt"
    )
    upstream_sdcs = [
        ("01", root / "01_result" / "01_soc_clocks.sdc"),
        ("02", fixture_upstream_output(root, "02", "prects", "ss_125")),
        ("04", fixture_upstream_output(root, "04", "prects", "ss_125")),
        ("10", root / "10_result" / "10_feedthrough.sdc"),
        ("20", fixture_upstream_output(root, "20", "prects", "ss_125")),
        ("03", root / "03_result" / "03_soc_clock_groups.sdc"),
    ]
    for label, sdc_path in upstream_sdcs:
        original = sdc_path.read_bytes()
        completion_before = digest(completion_30)
        delta_meta_before = digest(delta_meta_30)
        output_before = digest(output_30)
        sdc_path.write_bytes(
            original + "# stale {0} SDC\n".format(label).encode("utf-8")
        )
        rejected = run(EX30, root, args)
        require(
            rejected.returncode != 0,
            "30 accepted stale {0} output SDC".format(label),
        )
        report = report_path.read_text(encoding="utf-8")
        require(
            "{0} output_sdc_digest".format(label) in report
            and "declared=" in report
            and "actual=" in report,
            "stale {0} digest diagnostic is incomplete".format(label),
        )
        require(
            digest(completion_30) == completion_before,
            "stale {0} run republished 30 completion".format(label),
        )
        require(
            digest(delta_meta_30) == delta_meta_before,
            "stale {0} run appended a 30 accounting transaction".format(label),
        )
        require(
            digest(output_30) == output_before,
            "stale {0} run rewrote the formal 30 SDC".format(label),
        )
        sdc_path.write_bytes(original)
        recovered = run(EX30, root, args)
        require(
            recovered.returncode == 0,
            "30 did not recover after restoring {0} SDC:\n{1}\n{2}".format(
                label, recovered.stdout, recovered.stderr
            ),
        )


def test_clock_context_resolution():
    legacy = load_module(EX30, "reg30_legacy")
    target = load_module(EX30.with_name("30_target_runtime.py"), "reg30_target")
    stage10 = target.load_stage10(legacy)
    report = legacy.Report()
    rows = [
        {"inst_name": "u_clkgen", "port_name": "bus_clk_o", "clock_name": "u_clkgen_bus_clk_o", "original_clock_name": "bus_clk", "final_action": "emit_output_clock", "target_object": "u_clkgen/bus_clk_o"},
        {"inst_name": "u_clkgen", "port_name": "peri_clk_o", "clock_name": "u_clkgen_peri_clk_o", "original_clock_name": "peri_clk", "final_action": "emit_output_clock", "target_object": "u_clkgen/peri_clk_o"},
        {"inst_name": "u_ctrl", "port_name": "clk_i", "clock_name": "u_ctrl_clk_i", "original_clock_name": "ctrl_clk", "final_action": "check_only", "from_whom": "u_clkgen.bus_clk_o", "root_source": "u_clkgen/bus_clk_o"},
        {"inst_name": "u_dpg", "port_name": "clk_i", "clock_name": "u_dpg_clk_i", "original_clock_name": "dpg_clk", "final_action": "check_only", "from_whom": "u_clkgen.bus_clk_o", "root_source": "u_clkgen/bus_clk_o"},
        {"inst_name": "u_periph", "port_name": "clk_i", "clock_name": "u_periph_clk_i", "original_clock_name": "periph_clk", "final_action": "check_only", "from_whom": "u_clkgen.peri_clk_o", "root_source": "u_clkgen/peri_clk_o"},
    ]
    context = target.build_clock_context(stage10, rows, report)
    empty_index = target.delay_evidence_index(legacy, [])
    ctrl_clocks, _ = target.endpoint_clock_context(
        legacy, empty_index, context, "u_ctrl", "ctrl_cfg_o[0]", "output_delay"
    )
    dpg_clocks, _ = target.endpoint_clock_context(
        legacy, empty_index, context, "u_dpg", "cfg_shadow_i[0]", "input_delay"
    )
    periph_clocks, _ = target.endpoint_clock_context(
        legacy, empty_index, context, "u_periph", "async_req_i[1]", "input_delay"
    )
    relations = {
        tuple(sorted(("u_clkgen_bus_clk_o", "u_clkgen_peri_clk_o"))): "asynchronous"
    }
    require(
        target.resolved_clock_pair(ctrl_clocks, dpg_clocks, relations)
        == ("u_clkgen_bus_clk_o", "u_clkgen_bus_clk_o", "synchronous"),
        "same-clock MCP context did not resolve",
    )
    require(
        target.resolved_clock_pair(dpg_clocks, periph_clocks, relations)
        == ("u_clkgen_bus_clk_o", "u_clkgen_peri_clk_o", "asynchronous"),
        "asynchronous max/min context did not resolve",
    )
    require(report.error_count == 0, "clock context unit regression reported errors")


def test_prior_view_upstream_provenance_gate():
    root = WORK / "prior_view_upstream"
    if root.exists():
        shutil.rmtree(str(root))
    legacy = load_module(EX30, "reg30_prior_legacy")
    target = load_module(EX30.with_name("30_target_runtime.py"), "reg30_prior_target")
    prior_view = {"view_id": "prects_ss", "stage": "prects", "corner": "ss", "require_30": "yes"}
    current_view = {"view_id": "postroute_ff", "stage": "postroute", "corner": "ff", "require_30": "yes"}
    output = target.output_path(root, "prects", "ss")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("# prior 30 formal SDC\n", encoding="utf-8")
    prior_upstream = {"01:output_sdc": "old-upstream-digest"}
    write_json(target.completion_path(root, "prects", "ss"), {
        "schema_version": "1.0", "stage_name": target.STAGE_NAME,
        "view_id": "prects_ss", "stage": "prects", "corner": "ss",
        "run_id": "reg30", "mode_label": "func",
        "completion_status": "complete", "structure_digest": "structure",
        "output_sdc_digest": digest(output),
        "upstream_artifact_digests": prior_upstream,
    })
    report = legacy.Report()
    ready, _ = target.all_required_30_ready(
        root, [prior_view, current_view], current_view, True,
        {"run_id": "reg30", "mode_label": "func"}, "structure",
        {"01:output_sdc": "new-upstream-digest"}, report,
    )
    require(not ready, "prior 30 view accepted stale upstream provenance")
    require(
        any("upstream_artifact_digests are stale" in line for line in report.lines),
        "prior-view upstream mismatch diagnostic missing",
    )

    clean_report = legacy.Report()
    ready, digests = target.all_required_30_ready(
        root, [prior_view, current_view], current_view, True,
        {"run_id": "reg30", "mode_label": "func"}, "structure",
        prior_upstream, clean_report,
    )
    require(ready and "prects_ss" in digests, "matching prior-view provenance was rejected")


def main():
    test_formal_flow()
    test_cli_and_diagnostic_gate()
    test_deferred_multiview_finalization()
    test_required_upstream_sdc_digest_gate()
    test_clock_context_resolution()
    test_prior_view_upstream_provenance_gate()
    print("30 target regression: PASS")


if __name__ == "__main__":
    main()
