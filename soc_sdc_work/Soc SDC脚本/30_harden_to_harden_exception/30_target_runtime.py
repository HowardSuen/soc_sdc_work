#!/usr/bin/env python3
"""Single-run workbook/accounting runtime for stage 30.

This module is loaded by 30_extract_harden_to_harden_exception.py after that
file has initialized its review-workbook and SDC parsing helpers.
"""

import csv
import hashlib
import importlib.util
import io
import json
import os
import re
import shutil
import socket
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


SCHEMA_VERSION = "1.0"
STAGE_NAME = "30_harden_to_harden_exception"
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")

RULE_MACHINE_FIELDS = {
    "exception_id", "stage", "corner", "channel_id", "related_04_pad_id",
    "related_20_channel_id", "related_10_feedthrough_edge_id", "src_bit_index",
    "src_endpoint", "dst_bit_index", "dst_endpoint", "from_collection",
    "to_collection", "through_collection", "src_clock", "dst_clock",
    "clock_relation", "timing_contract_status", "harden_clock_context_status",
    "source_type", "source_sdc_file", "source_line", "source_command",
    "source_digest", "machine_digest",
}

# These fields may be refreshed on every extraction without invalidating a
# human review of the candidate's structural/evidence identity.
CANDIDATE_NON_INVALIDATING_FIELDS = {
    "accounting_digest_before",
    "extraction_time",
    "note",
    "port_accounting",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value):
    return clean(value).lower()


def sha_file(path):
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_bytes(value):
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")


def json_digest(value):
    return hashlib.sha256(json_bytes(value)).hexdigest()


def json_payload(value):
    return (json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")


def safe_token(value):
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", clean(value)).strip("_") or "all"


def pad_safe_token(value):
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-")
    token = "".join(
        char if char in allowed else "_" for char in clean(value)
    )
    return token or "unknown"


def load_stage10(legacy):
    path = Path(legacy.__file__).resolve().parent.parent / "10_feedthrough" / "10_extract_feedthrough.py"
    if not path.is_file():
        raise RuntimeError("migrated stage-10 runtime was not found: {0}".format(path))
    spec = importlib.util.spec_from_file_location("soc_sdc_stage10_runtime_for_30", str(path))
    if spec is None or spec.loader is None:
        raise RuntimeError("failed to load stage-10 runtime: {0}".format(path))
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def output_path(run_root, stage, corner):
    if stage == "all" and corner == "all":
        return run_root / "30_result" / "30_harden_to_harden_exception.sdc"
    return run_root / "30_result" / (
        "30_harden_to_harden_exception_{0}_{1}.sdc".format(safe_token(stage), safe_token(corner))
    )


def report_path(run_root, stage, corner):
    return run_root / "30_result" / "reports" / (
        "harden_to_harden_exception_check_report_{0}_{1}.txt".format(
            safe_token(stage), safe_token(corner)
        )
    )


def completion_path(run_root, stage, corner):
    return run_root / "30_middle" / "completion" / (
        "{0}_{1}.meta".format(safe_token(stage), safe_token(corner))
    )


def upstream_view_output_path(run_root, label, stage, corner):
    stage_token = safe_token(stage)
    corner_token = safe_token(corner)
    if label == "02":
        name = "02_soc_clock_timing_{0}_{1}.sdc".format(
            clean(stage), clean(corner)
        )
        return run_root / "02_result" / name
    if label == "04":
        name = (
            "04_soc_io_pads.sdc"
            if clean(stage) == "all" and clean(corner) == "all"
            else "04_soc_io_pads_{0}_{1}.sdc".format(
                clean(stage), pad_safe_token(corner)
            )
        )
        return run_root / "04_result" / name
    if label == "20":
        name = (
            "20_harden_x_if.sdc"
            if clean(stage) == "all" and clean(corner) == "all"
            else "20_harden_x_if_{0}_{1}.sdc".format(
                stage_token, corner_token
            )
        )
        return run_root / "20_result" / name
    raise ValueError("unsupported required upstream stage: {0}".format(label))


def upstream_view_completion_path(run_root, label, stage, corner):
    if label == "02":
        name = "{0}_{1}.meta".format(clean(stage), clean(corner))
    elif label == "04":
        name = "{0}_{1}.meta".format(
            pad_safe_token(stage), pad_safe_token(corner)
        )
    elif label == "20":
        name = "{0}_{1}.meta".format(safe_token(stage), safe_token(corner))
    else:
        raise ValueError("unsupported required upstream stage: {0}".format(label))
    return run_root / "{0}_middle".format(label) / "completion" / name


def validate_upstream_output_digest(payload, meta_path, output, run_root,
                                    label, report):
    if not payload:
        return ""
    valid = True
    declared_path_value = clean(payload.get("output_sdc_path"))
    if declared_path_value:
        declared_path = Path(declared_path_value).expanduser()
        if not declared_path.is_absolute():
            declared_path = run_root / declared_path
        if declared_path.resolve() != output.resolve():
            report.error(
                "{0}: output_sdc_path must identify canonical {1} output {2}; "
                "declared={3}".format(
                    meta_path, label, output, declared_path
                )
            )
            valid = False
    declared = clean(payload.get("output_sdc_digest"))
    if not declared:
        report.error(
            "{0}: {1} output_sdc_digest is required for {2}".format(
                meta_path, label, output
            )
        )
        return ""
    if not output.is_file():
        report.error(
            "{0}: canonical {1} output SDC is missing: {2}; declared={3}".format(
                meta_path, label, output, declared
            )
        )
        return ""
    actual = sha_file(output)
    if actual != declared:
        report.error(
            "{0}: {1} output_sdc_digest does not match {2}; "
            "declared={3}, actual={4}".format(
                meta_path, label, output, declared, actual
            )
        )
        return ""
    return actual if valid else ""


def read_json(path, label, report, required=True):
    if not path.is_file():
        if required:
            report.error("required {0} is missing: {1}".format(label, path))
        return {}
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        report.error("invalid {0} {1}: {2}".format(label, path, exc))
        return {}
    if not isinstance(value, dict):
        report.error("{0} must be a JSON object: {1}".format(label, path))
        return {}
    return value


def read_csv_rows(path, label, report, required=True):
    if not path.is_file():
        if required:
            report.error("required {0} is missing: {1}".format(label, path))
        return []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            return [dict(row) for row in csv.DictReader(file_obj)]
    except OSError as exc:
        report.error("failed to read {0} {1}: {2}".format(label, path, exc))
        return []


def current_required_view(required_views, stage, corner):
    matches = [
        row for row in required_views
        if clean(row.get("stage")) == stage
        and clean(row.get("corner")) == corner
        and norm(row.get("require_30")) == "yes"
    ]
    return matches[0] if len(matches) == 1 else None


def parse_final_used(value, shape, resume, location, report):
    text = clean(value)
    legal = set(shape.bits())
    if not text:
        return set()
    if norm(text) == "all used":
        return legal
    match = re.fullmatch(r"USED:\s*(.*?);\s*UNUSED:\s*(.*)", text, re.IGNORECASE)
    if match:
        def bits(part):
            item = clean(part)
            if item in {"", "-"}:
                return set()
            if not re.fullmatch(r"-?\d+(?:\s*,\s*-?\d+)*", item):
                raise ValueError("invalid final bit list {0!r}".format(item))
            values = [int(token.strip()) for token in item.split(",")]
            if len(set(values)) != len(values):
                raise ValueError("duplicate bit in final token")
            return set(values)
        try:
            used = bits(match.group(1))
            unused = bits(match.group(2))
        except ValueError as exc:
            report.error("{0}: {1}".format(location, exc))
            return set()
        if used & unused:
            report.error("{0}: USED and UNUSED overlap".format(location))
        if used | unused != legal:
            report.error("{0}: final token does not exactly cover legal bits".format(location))
        return used & legal
    if not resume:
        report.error("{0} must be blank in a fresh run, got {1!r}".format(location, text))
        return set()
    if not re.fullmatch(r"-?\d+(?:\s*,\s*-?\d+)*", text):
        report.error("{0} has invalid Used state {1!r}".format(location, text))
        return set()
    values = [int(item.strip()) for item in text.split(",")]
    result = set(values)
    for bit in sorted(result):
        if not shape.contains(bit):
            report.error("{0} contains out-of-range Used bit {1}".format(location, bit))
    return result & legal


def candidate_validation(model, candidate_path, report):
    try:
        candidate = load_workbook(str(candidate_path), data_only=False)
    except Exception as exc:
        report.error("failed to reopen candidate workbook {0}: {1}".format(candidate_path, exc))
        return
    if candidate.sheetnames != model.workbook.sheetnames:
        report.error("candidate workbook changed sheet order: {0}".format(model.relative_name))
    for record in model.records:
        sheet = candidate[record.sheet]
        if sheet.cell(record.row, record.port_col).value != record.port_value:
            report.error("candidate changed non-accounting port cell for {0}".format(record.location()))
        if sheet.cell(record.row, record.width_col).value != record.width_value:
            report.error("candidate changed non-accounting width cell for {0}".format(record.location()))
        if sheet.cell(record.row, record.connection_col).value != record.connection_value:
            report.error("candidate changed connectivity cell for {0}".format(record.location()))
        used_cell = sheet.cell(record.row, record.used_col)
        if record.modified:
            expected = clean(getattr(record, "target_used_text", ""))
            if clean(used_cell.value) != expected:
                report.error("candidate Used state mismatch for {0}".format(record.location()))
            if used_cell.number_format != "@":
                report.error("candidate Used state is not stored as text for {0}".format(record.location()))
        elif used_cell.value != record.used_value:
            report.error("candidate changed an unrelated Used state for {0}".format(record.location()))
    candidate.close()


def patch_accounting_runtime(runtime):
    runtime.parse_used_state = parse_final_used
    runtime.candidate_cell_validation = candidate_validation
    runtime.STAGE_NAME = STAGE_NAME


def build_clock_context(stage10, rows, report):
    """Resolve 01 local/check-only clock aliases to emitted SoC clocks."""
    aliases = {}
    alias_candidates = defaultdict(set)
    active_clocks = set()
    object_to_clock = {}
    conflicted_objects = set()
    active_rows = []
    check_rows = []

    for row_idx, row in enumerate(rows, start=2):
        action = norm(row.get("final_action"))
        clock_name = clean(row.get("clock_name"))
        if not clock_name:
            continue
        if action in stage10.ACTIVE_CLOCK_ACTIONS:
            active_rows.append((row_idx, row))
        elif action in {"check_only", "alias_only"}:
            check_rows.append((row_idx, row))

    def add_alias(inst_name, alias, mapped):
        if inst_name and alias and mapped:
            alias_candidates[(inst_name, alias)].add(mapped)

    def add_object(value, mapped, row_idx):
        key = stage10.normalize_clock_object(value)
        if not key or key in conflicted_objects:
            return
        previous = object_to_clock.get(key)
        if previous and previous != mapped:
            report.warn(
                "01 clock inventory row {0}: object {1} maps to both {2} and {3}; ignored".format(
                    row_idx, key, previous, mapped
                )
            )
            object_to_clock.pop(key, None)
            conflicted_objects.add(key)
            return
        object_to_clock[key] = mapped

    for row_idx, row in active_rows:
        clock_name = clean(row.get("clock_name"))
        inst_name = clean(row.get("inst_name"))
        active_clocks.add(clock_name)
        add_alias(inst_name, clock_name, clock_name)
        add_alias(inst_name, clean(row.get("original_clock_name")), clock_name)
        add_object(row.get("target_object"), clock_name, row_idx)
        if inst_name and clean(row.get("port_name")):
            add_object(
                "{0}/{1}".format(inst_name, clean(row.get("port_name"))),
                clock_name,
                row_idx,
            )
        if norm(row.get("final_action")) in {"emit_top_clock", "emit_virtual_clock"}:
            add_object(row.get("direct_source"), clock_name, row_idx)
            if clean(row.get("root_source")) == clean(row.get("direct_source")):
                add_object(row.get("root_source"), clock_name, row_idx)

    unresolved = list(check_rows)
    for _ in range(len(unresolved) + 1):
        if not unresolved:
            break
        next_unresolved = []
        progress = False
        for row_idx, row in unresolved:
            resolved = set()
            for value in (row.get("from_whom"), row.get("root_source"), row.get("direct_source")):
                key = stage10.normalize_clock_object(value)
                if key.startswith("clock:") and key[6:] in active_clocks:
                    resolved.add(key[6:])
                elif key in active_clocks:
                    resolved.add(key)
                elif key in object_to_clock:
                    resolved.add(object_to_clock[key])
            if len(resolved) != 1:
                if len(resolved) > 1:
                    report.warn(
                        "01 clock inventory row {0}: check-only clock maps to multiple SoC clocks; ignored".format(
                            row_idx
                        )
                    )
                    progress = True
                else:
                    next_unresolved.append((row_idx, row))
                continue
            mapped = next(iter(resolved))
            inst_name = clean(row.get("inst_name"))
            add_alias(inst_name, clean(row.get("clock_name")), mapped)
            add_alias(inst_name, clean(row.get("original_clock_name")), mapped)
            target = clean(row.get("target_object"))
            if not target and inst_name and clean(row.get("port_name")):
                target = "{0}/{1}".format(inst_name, clean(row.get("port_name")))
            add_object(target, mapped, row_idx)
            progress = True
        unresolved = next_unresolved
        if not progress:
            break

    for key, values in sorted(alias_candidates.items()):
        if len(values) == 1:
            aliases[key] = next(iter(values))
        else:
            report.warn(
                "01 clock inventory has conflicting alias mapping for {0}/{1}; ignored".format(
                    key[0], key[1]
                )
            )

    instance_clocks = defaultdict(set)
    for (inst_name, _alias), mapped in aliases.items():
        instance_clocks[inst_name].add(mapped)
    return {
        "clock_aliases": aliases,
        "active_clocks": active_clocks,
        "instance_clocks": instance_clocks,
    }


def validate_completion(stage10, path, expected_stage, run_context, structure, report):
    payload = read_json(path, expected_stage + " completion", report)
    if payload:
        stage10.target_validate_provenance(
            payload, path, run_context, structure, report,
            require_gate_fields=True, expected_stage_name=expected_stage,
        )
        if norm(payload.get("completion_status")) != "complete":
            report.error("{0}: completion_status must be complete".format(path))
    return payload


def validate_upstreams(runtime, stage10, run_root, run_context, required_views,
                       structure, accounting_before, port_records, report):
    digests = {}
    metas = {}
    previous_after = ""
    specs = (
        ("00", "00_middle", "00_harden_port_inventory"),
        ("01", "01_middle", "01_soc_clocks"),
        ("04", "04_middle", "04_soc_io_pads"),
        ("10", "10_middle", "10_feedthrough"),
        ("20", "20_middle", "20_harden_x_if"),
    )
    completions = {}
    for label, directory, expected_stage in specs:
        middle = run_root / directory
        delta_path = middle / "port_accounting_delta.csv"
        meta_path = middle / "port_accounting_delta.meta"
        completion = middle / "stage_completion.meta"
        meta = read_json(meta_path, label + " accounting delta meta", report)
        if meta:
            stage10.target_validate_provenance(
                meta, meta_path, run_context, structure, report,
                expected_stage_name=expected_stage,
            )
        payload = validate_completion(
            stage10, completion, expected_stage, run_context, structure, report
        )
        completions[label] = payload
        if not delta_path.is_file():
            report.error("required {0} accounting delta is missing: {1}".format(label, delta_path))
        elif clean(meta.get("delta_csv_digest")) != sha_file(delta_path):
            report.error("{0}: delta_csv_digest is stale".format(meta_path))
        before = clean(meta.get("accounting_digest_before"))
        after = clean(meta.get("accounting_digest_after"))
        if previous_after and before != previous_after:
            report.error("accounting digest chain is discontinuous before {0}".format(label))
        if after:
            previous_after = after
        metas[label] = meta
        for path in (delta_path, meta_path, completion):
            digests["{0}:{1}".format(label, path.name)] = sha_file(path) if path.is_file() else ""

    completion_01_path = run_root / "01_middle" / "stage_completion.meta"
    output_01 = run_root / "01_result" / "01_soc_clocks.sdc"
    digests["01:output_sdc"] = validate_upstream_output_digest(
        completions.get("01", {}), completion_01_path, output_01,
        run_root, "01", report,
    )

    completion_10_path = run_root / "10_middle" / "stage_completion.meta"
    output_10 = run_root / "10_result" / "10_feedthrough.sdc"
    digests["10:output_sdc"] = validate_upstream_output_digest(
        completions.get("10", {}), completion_10_path, output_10,
        run_root, "10", report,
    )

    completion_03 = run_root / "03_middle" / "stage_completion.meta"
    completion_03_payload = validate_completion(
        stage10, completion_03, "03_soc_clock_groups", run_context, structure, report
    )
    digests["03:stage_completion.meta"] = sha_file(completion_03) if completion_03.is_file() else ""
    output_03 = run_root / "03_result" / "03_soc_clock_groups.sdc"
    digests["03:output_sdc"] = validate_upstream_output_digest(
        completion_03_payload, completion_03, output_03,
        run_root, "03", report,
    )

    for flag, label, directory, expected_stage in (
        ("require_02", "02", "02_middle", "02_soc_clock_timing"),
        ("require_04", "04", "04_middle", "04_soc_io_pads"),
        ("require_20", "20", "20_middle", "20_harden_x_if"),
    ):
        required = [row for row in required_views if norm(row.get(flag)) == "yes"]
        authenticated = {}
        if label in {"04", "20"}:
            authenticated = completions.get(label, {}).get(
                "required_view_completions", {}
            )
            if required and not isinstance(authenticated, dict):
                report.error(
                    "{0} run-wide completion requires required_view_completions".format(
                        label
                    )
                )
                authenticated = {}
            if isinstance(authenticated, dict):
                expected_ids = {
                    clean(row.get("view_id")) for row in required
                }
                actual_ids = set(clean(key) for key in authenticated)
                if actual_ids != expected_ids:
                    report.error(
                        "{0} run-wide completion required_view_completions "
                        "keys mismatch: expected={1}, actual={2}".format(
                            label, sorted(expected_ids), sorted(actual_ids)
                        )
                    )
        for view in required:
            path = upstream_view_completion_path(
                run_root, label, view.get("stage"), view.get("corner")
            )
            payload = validate_completion(
                stage10, path, expected_stage, run_context, structure, report
            )
            if clean(payload.get("view_id")) != clean(view.get("view_id")):
                report.error("{0}: view_id does not match required_views.csv".format(path))
            if clean(payload.get("stage")) != clean(view.get("stage")) or clean(payload.get("corner")) != clean(view.get("corner")):
                report.error("{0}: stage/corner does not match required_views.csv".format(path))
            view_id = clean(view.get("view_id"))
            completion_digest = sha_file(path) if path.is_file() else ""
            digests["{0}:view:{1}".format(directory[:2], view_id)] = completion_digest
            if label in {"04", "20"}:
                expected_completion_digest = clean(authenticated.get(view_id))
                if (
                    not expected_completion_digest
                    or expected_completion_digest != completion_digest
                ):
                    report.error(
                        "{0}: {1} run-wide completion does not authenticate "
                        "required view {2}; declared={3}, actual={4}".format(
                            path, label, view_id,
                            expected_completion_digest, completion_digest,
                        )
                    )
            output = upstream_view_output_path(
                run_root, label, view.get("stage"), view.get("corner")
            )
            digests["{0}:view:{1}:output_sdc".format(
                label, view_id
            )] = validate_upstream_output_digest(
                payload, path, output, run_root, label, report
            )

    empty_digest = runtime.accounting_digest(port_records, empty=True)
    runtime.validate_resume_evidence(
        run_root, port_records, run_context, structure,
        accounting_before, empty_digest, report,
    )
    existing_30_meta_path = run_root / "30_middle" / "port_accounting_delta.meta"
    existing_30_meta = read_json(
        existing_30_meta_path, "30 accounting delta meta", report, required=False
    )
    if existing_30_meta:
        stage10.target_validate_provenance(
            existing_30_meta, existing_30_meta_path, run_context, structure, report,
            expected_stage_name=STAGE_NAME,
        )
    expected_current = clean(
        existing_30_meta.get("accounting_digest_after")
        if existing_30_meta else metas.get("20", {}).get("accounting_digest_after")
    )
    if expected_current and expected_current != accounting_before:
        report.error(
            "accounting digest chain mismatch before 30: expected={0}, current={1}".format(
                expected_current, accounting_before
            )
        )

    inventories = {}
    clock_path = run_root / "01_middle" / "clock_inventory.csv"
    clock_meta_path = run_root / "01_middle" / "clock_inventory.meta"
    clock_rows, clock_meta = stage10.target_validate_inventory_pair(
        clock_path, clock_meta_path, "01 clock inventory", run_context,
        structure, report, ("inventory_digest", "clock_inventory_digest"),
    )
    inventories.update(build_clock_context(stage10, clock_rows, report))
    inventories["clock_rows"] = clock_rows
    inventories["clock_meta"] = clock_meta
    for path in (clock_path, clock_meta_path):
        digests["inventory:{0}".format(path.relative_to(run_root))] = (
            sha_file(path) if path.is_file() else ""
        )

    for key, relpath, meta_relpath, label, digest_fields in (
        ("pad", "04_middle/pad_inventory.csv", "04_middle/pad_inventory.meta", "04 pad inventory", ("inventory_digest", "pad_inventory_digest")),
        ("feedthrough", "10_middle/feedthrough_edge_inventory.csv", "10_middle/feedthrough_edge_inventory.meta", "10 feedthrough inventory", ("inventory_digest", "feedthrough_edge_inventory_digest", "feedthrough_inventory_digest")),
        ("channel", "20_middle/channel_inventory.csv", "20_middle/channel_inventory.meta", "20 channel inventory", ("inventory_digest", "channel_inventory_digest")),
    ):
        data_path = run_root / relpath
        meta_path = run_root / meta_relpath
        rows, meta = stage10.target_validate_inventory_pair(
            data_path, meta_path, label, run_context, structure, report, digest_fields
        )
        inventories[key] = rows
        for path in (data_path, meta_path):
            digests["inventory:{0}".format(relpath if path == data_path else meta_relpath)] = (
                sha_file(path) if path.is_file() else ""
            )

    relation_path = run_root / "03_middle" / "relation_map.csv"
    relation_meta = run_root / "03_middle" / "relation_map.meta"
    relation_rows, _ = stage10.target_validate_inventory_pair(
        relation_path, relation_meta, "03 relation map", run_context, structure,
        report, ("relation_map_digest", "inventory_digest"),
    )
    relations = {}
    for row in relation_rows:
        a = clean(row.get("clock_a"))
        b = clean(row.get("clock_b"))
        relation = norm(row.get("relation_type")) or "unknown"
        if a and b:
            key = tuple(sorted((a, b)))
            previous = relations.get(key)
            if previous and previous != relation:
                report.error(
                    "03 relation map has conflicting relation for {0}/{1}".format(
                        key[0], key[1]
                    )
                )
            else:
                relations[key] = relation
    inventories["relations"] = relations

    owner_map = defaultdict(list)
    index = runtime.record_index(port_records)
    for directory in ("00_middle", "01_middle", "04_middle", "10_middle", "20_middle", "30_middle"):
        path = run_root / directory / "port_accounting_delta.csv"
        for row in runtime.load_delta_rows(path, report) if path.is_file() else []:
            key = (
                clean(row.get("workbook")), clean(row.get("sheet")),
                int(clean(row.get("row")) or "0"), norm(row.get("direction")),
                clean(row.get("port")),
            )
            record = index.get(key)
            if record is None:
                continue
            try:
                bits = runtime.parse_bits_field(row.get("added_bits"))
            except ValueError:
                continue
            for bit in bits:
                owner_map[(record.inst_name, record.direction, record.shape.base, bit)].append(
                    (clean(row.get("stage_name")), clean(row.get("owner_object_id")))
                )
    inventories["owner_map"] = owner_map
    inventories["upstream_digests"] = digests
    inventories["existing_30_meta"] = existing_30_meta
    return inventories


def exact_endpoint(stage10, edge, side):
    inst = clean(getattr(edge, side + "_instance"))
    port = clean(getattr(edge, side + "_port"))
    return stage10.target_endpoint_collection(inst, port)


def endpoint_tuple(stage10, edge, side):
    return stage10.target_edge_endpoint_tuple(edge, side)


def row_is_active(row):
    return norm(row.get("apply")) == "yes" and norm(row.get("review_status")) == "approved"


def row_routes_30(row):
    disposition = norm(row.get("channel_disposition") or row.get("pad_disposition") or row.get("disposition"))
    return row_is_active(row) and disposition == "route_to_30"


def row_has_active_timing(row):
    if not row_is_active(row):
        return False
    if norm(row.get("pad_disposition")) == "constrained":
        return True
    return any(norm(row.get(name)) == "yes" for name in (
        "emit_max", "emit_min", "emit_input_delay", "emit_output_delay",
        "emit_load", "emit_drive", "timing_active",
    ))


def matching_inventory_rows(rows, edge):
    connection_id = clean(edge.connection_id)
    direct = [row for row in rows if clean(row.get("connection_id")) == connection_id]
    if direct:
        return direct
    src = clean(edge.src_soc_object)
    dst = clean(edge.dst_soc_object)
    result = []
    for row in rows:
        row_src = clean(row.get("src_soc_object") or row.get("src_endpoint"))
        row_dst = clean(row.get("dst_soc_object") or row.get("dst_endpoint"))
        if row_src in {src, exact_collection_text(src)} and row_dst in {dst, exact_collection_text(dst)}:
            result.append(row)
    return result


def rows_for_view(rows, stage, corner):
    exact = [
        row for row in rows
        if clean(row.get("stage")) == stage and clean(row.get("corner")) == corner
    ]
    if exact:
        return exact
    common = [
        row for row in rows
        if clean(row.get("stage")) in {"", "all"}
        and clean(row.get("corner")) in {"", "all"}
    ]
    return common or list(rows)


def exact_collection_text(value):
    text = clean(value)
    if text.startswith("["):
        return text
    if "/" in text:
        return "[get_pins {{{0}}}]".format(text)
    return "[get_ports {{{0}}}]".format(text)


def manifest_instances(legacy, manifest, instances, records):
    converted = {}
    for inst_name, info in instances.items():
        entry = manifest.get(inst_name)
        item = legacy.InstInfo(
            module_name=clean(info.get("module_name")),
            inst_name=inst_name,
            owner=clean(info.get("owner")),
            sdc_hint=clean(entry.sdc_path) if entry is not None else "",
            sdc_status=norm(entry.availability_status) if entry is not None else "missing",
            sdc_note=clean(entry.note) if entry is not None else "",
        )
        if entry is not None and entry.resolved_path is not None:
            item.sdc_path = entry.resolved_path
        converted[inst_name] = item
    for record in records:
        item = converted.get(record.inst_name)
        if item is None:
            continue
        target = item.inputs if record.direction == "input" else (
            item.outputs if record.direction == "output" else item.inouts
        )
        for bit in record.shape.bits():
            name = record.shape.base if record.shape.scalar else "{0}[{1}]".format(record.shape.base, bit)
            target.setdefault(name, legacy.PortInfo(name=name))
        target.setdefault(record.shape.base, legacy.PortInfo(name=record.shape.base))
    return converted


def delay_evidence_index(legacy, delays):
    result = defaultdict(list)
    for item in delays:
        for port_name in {clean(item.port_name), legacy.port_base(item.port_name)}:
            result[(clean(item.inst_name), port_name, norm(item.constraint_type))].append(item)
    return result


def endpoint_clock_context(legacy, index, inventories, inst_name, port_name, constraint_type):
    evidence = []
    seen = set()
    for candidate_port in (clean(port_name), legacy.port_base(port_name)):
        for item in index.get((clean(inst_name), candidate_port, constraint_type), []):
            if id(item) not in seen:
                seen.add(id(item))
                evidence.append(item)

    resolved = []
    aliases = inventories.get("clock_aliases", {})
    active = inventories.get("active_clocks", set())
    for item in evidence:
        local = clean(item.clock_name)
        if not local:
            continue
        mapped = aliases.get((clean(inst_name), local))
        if mapped:
            resolved.append(mapped)
        elif local in active:
            resolved.append(local)
        else:
            resolved.append("unresolved:" + local)

    # With no endpoint delay command, a single authenticated instance clock is
    # still an unambiguous harden clock context. Multi-clock hardens stay unknown.
    if not evidence:
        instance_values = sorted(inventories.get("instance_clocks", {}).get(clean(inst_name), set()))
        if len(instance_values) == 1:
            resolved = instance_values
    return sorted(set(resolved)), evidence


def resolved_clock_pair(src_clocks, dst_clocks, relations):
    if any(item.startswith("unresolved:") for item in list(src_clocks) + list(dst_clocks)):
        return "", "", "unknown"
    src = sorted(set(item for item in src_clocks if item))
    dst = sorted(set(item for item in dst_clocks if item))
    if len(src) != 1 or len(dst) != 1:
        return "", "", "unknown"
    if src[0] == dst[0]:
        return src[0], dst[0], "synchronous"
    return src[0], dst[0], relations.get(tuple(sorted((src[0], dst[0]))), "unknown")


def build_seeds(legacy, stage10, contexts, inventories, manifest, completeness,
                instances, records, run_context, structure, accounting_before,
                stage, corner, report):
    converted = manifest_instances(legacy, manifest, instances, records)
    delays, evidence = legacy.extract_sdc_evidence(converted, report)
    delay_index = delay_evidence_index(legacy, delays)
    evidence_index = defaultdict(list)
    for item in evidence:
        # ExceptionEvidence stores collection endpoints as from_ports/to_ports;
        # index each endpoint so target candidate mapping can resolve either side.
        for port_name in list(item.from_ports) + list(item.to_ports):
            evidence_index[(item.inst_name, legacy.port_base(port_name))].append(item)
    candidates = []
    rules = []
    channels = {}
    for context in contexts:
        edge = context.edge
        canonical = [
            SCHEMA_VERSION, clean(edge.src_instance), norm(edge.src_direction),
            legacy.port_base(edge.src_port), int(clean(edge.src_bit_index) or "0"),
            clean(edge.dst_instance), norm(edge.dst_direction),
            legacy.port_base(edge.dst_port), int(clean(edge.dst_bit_index) or "0"),
        ]
        digest = json_digest(canonical)
        channel_id = "CH_" + digest
        candidate_id = "CAND_" + json_digest([SCHEMA_VERSION, channel_id, stage, corner, "base"])
        exception_id = "EXC_" + json_digest([SCHEMA_VERSION, channel_id, stage, corner, "rule"])
        ft_rows = rows_for_view(
            matching_inventory_rows(inventories.get("feedthrough", []), edge), stage, corner
        )
        pad_rows = rows_for_view(
            matching_inventory_rows(inventories.get("pad", []), edge), stage, corner
        )
        selected_20 = rows_for_view(
            matching_inventory_rows(inventories.get("channel", []), edge), stage, corner
        )
        is_feedthrough = bool(ft_rows)
        is_pad = bool(pad_rows)
        related_ft = clean(ft_rows[0].get("feedthrough_edge_id")) if ft_rows else ""
        related_pad = clean(
            (pad_rows[0].get("pad_id") or pad_rows[0].get("related_04_pad_id") or pad_rows[0].get("connection_id"))
        ) if pad_rows else ""
        related_20 = clean(selected_20[0].get("channel_id")) if selected_20 else ""
        owner_rows = ft_rows if is_feedthrough else (pad_rows if is_pad else selected_20)
        owner_ok = bool(owner_rows) and any(row_routes_30(row) for row in owner_rows)
        conflict = any(row_has_active_timing(row) for row in ft_rows + pad_rows + selected_20)
        if is_feedthrough:
            category = "feedthrough_normal"
        elif is_pad:
            category = "pad_related"
        else:
            category = "needs_review"
        src_status = norm(manifest.get(clean(edge.src_instance)).availability_status) if manifest.get(clean(edge.src_instance)) else "not_required"
        dst_status = norm(manifest.get(clean(edge.dst_instance)).availability_status) if manifest.get(clean(edge.dst_instance)) else "not_required"
        sdc_status = "incomplete_missing_sdc" if "missing" in {src_status, dst_status} else "complete"
        src_clocks, src_delay_evidence = endpoint_clock_context(
            legacy, delay_index, inventories, edge.src_instance, edge.src_port,
            "output_delay",
        )
        dst_clocks, dst_delay_evidence = endpoint_clock_context(
            legacy, delay_index, inventories, edge.dst_instance, edge.dst_port,
            "input_delay",
        )
        src_clock, dst_clock, clock_relation = resolved_clock_pair(
            src_clocks, dst_clocks, inventories.get("relations", {})
        )
        if src_clock and dst_clock:
            clock_context_status = "matched"
        elif src_clocks or dst_clocks:
            clock_context_status = "mismatch"
        else:
            clock_context_status = (
                "incomplete_missing_sdc"
                if sdc_status == "incomplete_missing_sdc" else "unknown"
            )
        if src_delay_evidence and dst_delay_evidence:
            timing_status = "both_sides_timed"
        elif src_delay_evidence:
            timing_status = "src_timed_only"
        elif dst_delay_evidence:
            timing_status = "dst_timed_only"
        else:
            timing_status = (
                "incomplete_missing_sdc"
                if sdc_status == "incomplete_missing_sdc" else "no_port_timing"
            )
        matched_evidence = evidence_index.get((clean(edge.src_instance), legacy.port_base(edge.src_port)), []) + evidence_index.get((clean(edge.dst_instance), legacy.port_base(edge.dst_port)), [])
        first = matched_evidence[0] if matched_evidence else None
        source_type = "extracted_harden_exception" if first else (
            "missing_timing_candidate" if sdc_status == "incomplete_missing_sdc" else "manual_entry"
        )
        values = {header: "" for header in legacy.EXCEPTION_CANDIDATE_HEADERS}
        values.update({
            "schema_version": SCHEMA_VERSION,
            "author": legacy.author_name(),
            "run_id": run_context.get("run_id", ""),
            "mode_label": run_context.get("mode_label", ""),
            "design_revision": run_context.get("design_revision", ""),
            "run_completeness": completeness.status,
            "port_accounting": "enabled",
            "candidate_id": candidate_id,
            "scenario": "",
            "stage": stage,
            "corner": corner,
            "channel_id": channel_id,
            "connection_id": edge.connection_id,
            "related_04_pad_id": related_pad,
            "source_type": source_type,
            "path_category": "other_reviewed" if owner_ok else "unknown",
            "timing_contract_status": timing_status,
            "src_instance": edge.src_instance,
            "src_port": edge.src_port,
            "src_bit_index": edge.src_bit_index,
            "src_endpoint": exact_endpoint(stage10, edge, "src"),
            "dst_instance": edge.dst_instance,
            "dst_port": edge.dst_port,
            "dst_bit_index": edge.dst_bit_index,
            "dst_endpoint": exact_endpoint(stage10, edge, "dst"),
            "src_clock": src_clock,
            "dst_clock": dst_clock,
            "clock_relation": clock_relation,
            "has_src_output_delay": "yes" if src_delay_evidence else "no",
            "has_dst_input_delay": "yes" if dst_delay_evidence else "no",
            "related_20_channel_id": related_20,
            "related_20_status": norm(selected_20[0].get("channel_disposition")) if selected_20 else "missing",
            "related_10_feedthrough_edge_id": related_ft,
            "source_workbook": context.source_record.workbook,
            "source_sheet": context.source_record.sheet,
            "source_row": context.source_record.row,
            "structure_digest": structure,
            "accounting_digest_before": accounting_before,
            "harden_clock_context_status": clock_context_status,
            "sdc_evidence_status": sdc_status,
            "source_sdc_file": first.source_sdc_file if first else "",
            "source_line": first.source_line if first else "",
            "source_command": first.source_command if first else "",
            "source_digest": first.source_digest if first else "",
            "extraction_time": first.extraction_time if first else "",
            "candidate_status": "active",
            "candidate_reason": "owner route_to_30" if owner_ok else "owner review is incomplete",
            "recommended_action": "review_exception_basis" if owner_ok and not conflict else "resolve_owner_or_timing_conflict",
            "note": "classification={0}; owner_ok={1}; active_conflict={2}".format(category, "yes" if owner_ok else "no", "yes" if conflict else "no"),
        })
        values["machine_digest"] = json_digest([
            [name, clean(values.get(name))]
            for name in sorted(set(legacy.EXCEPTION_CANDIDATE_HEADERS) - {
                "extraction_time", "note", "accounting_digest_before", "machine_digest",
            })
        ])
        rule = {header: "" for header in legacy.EXCEPTION_RULE_HEADERS}
        rule.update({
            "exception_id": exception_id,
            "scenario": "",
            "stage": stage,
            "corner": corner,
            "apply": "no",
            "review_status": "pending",
            "exception_type": first.exception_type if first else "needs_review",
            "path_category": "unknown",
            "channel_id": channel_id,
            "related_04_pad_id": related_pad,
            "related_20_channel_id": related_20,
            "related_10_feedthrough_edge_id": related_ft,
            "src_bit_index": edge.src_bit_index,
            "src_endpoint": values["src_endpoint"],
            "dst_bit_index": edge.dst_bit_index,
            "dst_endpoint": values["dst_endpoint"],
            "from_collection": values["src_endpoint"],
            "to_collection": values["dst_endpoint"],
            "through_collection": "",
            "src_clock": src_clock,
            "dst_clock": dst_clock,
            "clock_relation": clock_relation,
            "timing_contract_status": values["timing_contract_status"],
            "harden_clock_context_status": values["harden_clock_context_status"],
            "check_type": first.check_type if first else "both",
            "max_value": first.max_value if first else "",
            "min_value": first.min_value if first else "",
            "setup_cycles": first.setup_cycles if first else "",
            "hold_cycles": first.hold_cycles if first else "",
            "mcp_reference": first.mcp_reference if first else "",
            "datapath_only": "no",
            "tool_surface": "sta",
            "source_type": source_type,
            "source_sdc_file": values["source_sdc_file"],
            "source_line": values["source_line"],
            "source_command": values["source_command"],
            "source_digest": values["source_digest"],
        })
        rule["machine_digest"] = json_digest([
            [name, clean(rule.get(name))]
            for name in sorted(RULE_MACHINE_FIELDS - {"machine_digest"})
        ])
        candidates.append(values)
        rules.append(rule)
        channels[channel_id] = {
            "context": context, "edge": edge, "feedthrough_rows": ft_rows,
            "pad_rows": pad_rows, "channel_rows": selected_20,
            "owner_ok": owner_ok, "active_conflict": conflict,
            "is_feedthrough": is_feedthrough, "is_pad": is_pad,
            "related_ft": related_ft, "related_pad": related_pad,
            "related_20": related_20, "sdc_status": sdc_status,
        }
    candidates.sort(key=lambda row: clean(row.get("candidate_id")))
    rules.sort(key=lambda row: clean(row.get("exception_id")))
    return candidates, rules, channels, evidence


def sync_workbook(legacy, path, candidates, rules, metadata, report):
    workbook, created = legacy.create_or_load_workbook(path)
    legacy.ensure_sheet(workbook, "run_metadata", ["key", "value"])
    legacy.ensure_sheet(workbook, "exception_candidate", legacy.EXCEPTION_CANDIDATE_HEADERS)
    legacy.ensure_sheet(workbook, "exception_rule", legacy.EXCEPTION_RULE_HEADERS)
    meta_sheet = workbook["run_metadata"]
    if meta_sheet.max_row > 1:
        meta_sheet.delete_rows(2, meta_sheet.max_row - 1)
    for key, value in metadata.items():
        legacy.append_dict(meta_sheet, ["key", "value"], {"key": key, "value": clean(value)})

    candidate_sheet = workbook["exception_candidate"]
    candidate_map = legacy.header_map(candidate_sheet)
    existing = {
        clean(candidate_sheet.cell(row=row, column=candidate_map["candidate_id"]).value): row
        for row in range(2, candidate_sheet.max_row + 1)
        if clean(candidate_sheet.cell(row=row, column=candidate_map["candidate_id"]).value)
    }
    active = set()
    for values in candidates:
        key = clean(values.get("candidate_id"))
        active.add(key)
        if key not in existing:
            legacy.append_dict(candidate_sheet, legacy.EXCEPTION_CANDIDATE_HEADERS, values, legacy.NEW_FILL)
            existing[key] = candidate_sheet.max_row
            report.sync_changed = True
        else:
            row = existing[key]
            for header in legacy.EXCEPTION_CANDIDATE_HEADERS:
                old = clean(candidate_sheet.cell(row=row, column=candidate_map[header]).value)
                new = clean(values.get(header))
                if old != new:
                    candidate_sheet.cell(row=row, column=candidate_map[header], value=new)
                    if header not in CANDIDATE_NON_INVALIDATING_FIELDS:
                        report.sync_changed = True
    active_stage = clean(candidates[0].get("stage")) if candidates else ""
    active_corner = clean(candidates[0].get("corner")) if candidates else ""
    for key, row in existing.items():
        same_view = (
            clean(candidate_sheet.cell(row=row, column=candidate_map["stage"]).value) == active_stage
            and clean(candidate_sheet.cell(row=row, column=candidate_map["corner"]).value) == active_corner
        )
        if same_view and key not in active and norm(candidate_sheet.cell(row=row, column=candidate_map["candidate_status"]).value) != "stale":
            candidate_sheet.cell(row=row, column=candidate_map["candidate_status"], value="stale")
            report.sync_changed = True

    rule_sheet = workbook["exception_rule"]
    rule_map = legacy.header_map(rule_sheet)
    existing_rules = {
        clean(rule_sheet.cell(row=row, column=rule_map["exception_id"]).value): row
        for row in range(2, rule_sheet.max_row + 1)
        if clean(rule_sheet.cell(row=row, column=rule_map["exception_id"]).value)
    }
    for values in rules:
        key = clean(values.get("exception_id"))
        if key not in existing_rules:
            legacy.append_dict(rule_sheet, legacy.EXCEPTION_RULE_HEADERS, values, legacy.NEW_FILL)
            existing_rules[key] = rule_sheet.max_row
            report.sync_changed = True
            continue
        row = existing_rules[key]
        machine_changed = False
        for header in RULE_MACHINE_FIELDS:
            if header not in rule_map:
                continue
            old = clean(rule_sheet.cell(row=row, column=rule_map[header]).value)
            new = clean(values.get(header))
            if old != new:
                rule_sheet.cell(row=row, column=rule_map[header], value=new)
                machine_changed = True
        if machine_changed:
            rule_sheet.cell(row=row, column=rule_map["apply"], value="no")
            rule_sheet.cell(row=row, column=rule_map["review_status"], value="pending")
            if "approved_machine_digest" in rule_map:
                rule_sheet.cell(row=row, column=rule_map["approved_machine_digest"], value="")
            report.sync_changed = True

    legacy.add_validations(workbook)
    for sheet in workbook.worksheets:
        legacy.style_sheet(sheet)
    legacy.atomic_save_workbook(workbook, path)
    if created:
        report.sync_changed = True
    if report.sync_changed:
        report.info("synchronized 30 review workbook; review is required before formal SDC generation")


def candidates_csv(legacy, candidates):
    stream = io.StringIO()
    writer = csv.DictWriter(stream, fieldnames=legacy.EXCEPTION_CANDIDATE_HEADERS,
                            extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for row in candidates:
        writer.writerow({name: clean(row.get(name)) for name in legacy.EXCEPTION_CANDIDATE_HEADERS})
    return stream.getvalue()


def read_rules(legacy, path):
    return legacy.read_rule_rows(path)


def selected_rules(legacy, rows, stage, corner):
    return [
        row for row in rows
        if clean(row.values.get("stage") or "all") == stage
        and clean(row.values.get("corner") or "all") == corner
    ]


def validate_rules(legacy, rows, channels, relations, stage, corner, report):
    approved = []
    supported = legacy.EMITTED_EXCEPTION_TYPES
    allowed_categories = {"static", "config", "handshake", "cdc", "reset", "test_control", "other_reviewed"}
    for row in selected_rules(legacy, rows, stage, corner):
        values = row.values
        if norm(values.get("apply")) != "yes" or norm(values.get("review_status")) != "approved":
            continue
        exception_id = clean(values.get("exception_id"))
        channel_id = clean(values.get("channel_id"))
        channel = channels.get(channel_id)
        if channel is None:
            report.error("exception_rule row {0} {1}: channel_id is stale".format(row.row_idx, exception_id))
            continue
        edge = channel["edge"]
        expected = (
            exact_collection_text(edge.src_soc_object), exact_collection_text(edge.dst_soc_object),
            clean(edge.src_bit_index), clean(edge.dst_bit_index),
        )
        actual = (
            clean(values.get("from_collection") or values.get("src_endpoint")),
            clean(values.get("to_collection") or values.get("dst_endpoint")),
            clean(values.get("src_bit_index")), clean(values.get("dst_bit_index")),
        )
        if actual != expected:
            report.error("exception_rule row {0} {1}: endpoints do not match current exact raw edge".format(row.row_idx, exception_id))
        if not channel["owner_ok"]:
            report.error("exception_rule row {0} {1}: owner inventory is not approved route_to_30".format(row.row_idx, exception_id))
        if channel["active_conflict"]:
            report.error("exception_rule row {0} {1}: active 04/10/20 normal timing conflicts with 30".format(row.row_idx, exception_id))
        if channel["is_feedthrough"] and clean(values.get("related_10_feedthrough_edge_id")) != channel["related_ft"]:
            report.error("exception_rule row {0} {1}: related_10_feedthrough_edge_id is missing/stale".format(row.row_idx, exception_id))
        if channel["is_pad"]:
            if not channel["related_pad"] or clean(values.get("related_04_pad_id")) != channel["related_pad"]:
                report.error("exception_rule row {0} {1}: pad rule requires current related_04_pad_id".format(row.row_idx, exception_id))
        elif not channel["is_feedthrough"] and clean(values.get("related_20_channel_id")) != channel["related_20"]:
            report.error("exception_rule row {0} {1}: non-feedthrough rule requires current related_20_channel_id".format(row.row_idx, exception_id))
        etype = norm(values.get("exception_type"))
        if etype not in supported:
            report.error("exception_rule row {0} {1}: unsupported exception_type {2}".format(row.row_idx, exception_id, etype or "<blank>"))
        category = norm(values.get("path_category"))
        if category not in allowed_categories:
            report.error("exception_rule row {0} {1}: path_category must be reviewed and non-unknown".format(row.row_idx, exception_id))
        if not clean(values.get("owner")) or not clean(values.get("basis")):
            report.error("exception_rule row {0} {1}: owner and basis are required".format(row.row_idx, exception_id))
        if clean(values.get("through_collection")):
            report.error("exception_rule row {0} {1}: through_collection is not allowed for direct 30 edges".format(row.row_idx, exception_id))
        for field in ("from_collection", "to_collection"):
            text = clean(values.get(field))
            if not re.fullmatch(r"\[get_(?:pins|ports) \{[^{}*?]+\}\]", text):
                report.error("exception_rule row {0} {1}: {2} must be one exact SoC-visible object".format(row.row_idx, exception_id, field))
        if channel["sdc_status"] == "incomplete_missing_sdc" and not clean(values.get("sdc_independent_basis")):
            report.error("exception_rule row {0} {1}: missing SDC requires sdc_independent_basis".format(row.row_idx, exception_id))
        if norm(values.get("source_type")) == "missing_timing_candidate":
            report.error(
                "exception_rule row {0} {1}: missing_timing_candidate must be replaced by an independent reviewed source_type".format(
                    row.row_idx, exception_id
                )
            )
        if clean(values.get("source_sdc_file")):
            source = Path(clean(values.get("source_sdc_file")))
            if source.is_file() and clean(values.get("source_digest")) != sha_file(source):
                report.error("exception_rule row {0} {1}: source SDC digest is stale".format(row.row_idx, exception_id))
        src_clock = clean(values.get("src_clock"))
        dst_clock = clean(values.get("dst_clock"))
        declared_relation = norm(values.get("clock_relation")) or "unknown"
        if src_clock and dst_clock:
            expected_relation = "synchronous" if src_clock == dst_clock else relations.get(
                tuple(sorted((src_clock, dst_clock))), "unknown"
            )
            if declared_relation != expected_relation:
                report.error(
                    "exception_rule row {0} {1}: clock_relation={2} does not match 03 relation {3}".format(
                        row.row_idx, exception_id, declared_relation, expected_relation
                    )
                )
        if etype == "false_path":
            if category == "reset" and not legacy.reset_false_path_basis_ok(values):
                report.error("exception_rule row {0} {1}: reset false_path requires RDC/recovery/removal basis".format(row.row_idx, exception_id))
        elif etype == "multicycle_path":
            legacy.validate_multicycle_row(row, exception_id, norm(values.get("check_type")), report)
            if not clean(values.get("protocol_ref")):
                report.error("exception_rule row {0} {1}: multicycle_path requires protocol_ref".format(row.row_idx, exception_id))
        else:
            if etype in {"max_delay_override", "max_min_delay_override"} and legacy.parse_finite_number(values.get("max_value")) is None:
                report.error("exception_rule row {0} {1}: finite max_value is required".format(row.row_idx, exception_id))
            if etype in {"min_delay_override", "max_min_delay_override"} and legacy.parse_finite_number(values.get("min_value")) is None:
                report.error("exception_rule row {0} {1}: finite min_value is required".format(row.row_idx, exception_id))
            legacy.validate_datapath_strategy(
                row, exception_id, norm(values.get("clock_relation")), category,
                norm(values.get("datapath_only")), report,
            )
            if declared_relation == "unknown":
                report.error("exception_rule row {0} {1}: max/min override requires a resolved clock_relation".format(row.row_idx, exception_id))
        commands = legacy.commands_for_row(row)
        if not commands:
            report.error("exception_rule row {0} {1}: approved row emits no SDC command".format(row.row_idx, exception_id))
        approved.append(row)
    return approved


def render_sdc(legacy, rows, run_context, stage, corner, completeness, structure, before, after):
    lines = [
        "# Auto-generated by {0}".format(STAGE_NAME),
        "# Author: {0}".format(legacy.author_name()),
        "# Run ID: {0}".format(run_context.get("run_id", "")),
        "# Mode label: {0}".format(run_context.get("mode_label", "")),
        "# Design revision: {0}".format(run_context.get("design_revision", "")),
        "# View: stage={0}, corner={1}".format(stage, corner),
        "# Run completeness: {0}".format(completeness.status),
        "# Structure digest: {0}".format(structure),
        "# Accounting digest before: {0}".format(before),
        "# Accounting digest after: {0}".format(after),
        "",
    ]
    for row in rows:
        lines.append("# exception_id={0} channel_id={1}".format(
            clean(row.values.get("exception_id")), clean(row.values.get("channel_id"))
        ))
        lines.extend(legacy.commands_for_row(row))
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def plan_accounting(stage10, approved, channels, records, inventories, report):
    lookup = stage10.target_port_record_lookup(records)
    planned = []
    for row in approved:
        channel = channels.get(clean(row.values.get("channel_id")))
        if channel is None:
            continue
        for side in ("src", "dst"):
            resolved = stage10.target_record_for_endpoint(channel["edge"], side, lookup, report)
            if resolved is None:
                continue
            record, bit = resolved
            owner_key = (record.inst_name, record.direction, record.shape.base, bit)
            if inventories.get("owner_map", {}).get(owner_key):
                continue
            if bit not in record.used_bits:
                record.used_bits.add(bit)
                record.added_bits.add(bit)
                record.modified = True
                record.model.modified = True
                planned.append((row, record, bit, side))
    return planned


def apply_intermediate_cells(runtime, records):
    for record in records:
        if not record.modified:
            continue
        text = runtime.format_bits(record.used_bits)
        record.target_used_text = text
        cell = record.model.workbook[record.sheet].cell(record.row, record.used_col)
        cell.value = text
        cell.number_format = "@"


def apply_final_cells(records):
    all_rows = 0
    incomplete_rows = 0
    for record in records:
        legal = set(record.shape.bits())
        used = set(record.used_bits)
        unused = legal - used
        if not unused:
            text = "ALL USED"
            fill = GREEN_FILL
            font = GREEN_FONT
            all_rows += 1
        else:
            text = "USED:{0}; UNUSED:{1}".format(
                ",".join(str(bit) for bit in sorted(used)) if used else "-",
                ",".join(str(bit) for bit in sorted(unused)),
            )
            fill = RED_FILL
            font = RED_FONT
            incomplete_rows += 1
        cell = record.model.workbook[record.sheet].cell(record.row, record.used_col)
        cell.value = text
        cell.number_format = "@"
        cell.fill = fill
        cell.font = font
        record.target_used_text = text
        record.modified = True
        record.model.modified = True
    return all_rows, incomplete_rows


def transaction_id(run_id):
    stamp = __import__("datetime").datetime.utcnow().strftime("%Y%m%dT%H%M%S%fZ")
    seed = "{0}|{1}|{2}|{3}".format(run_id, stamp, os.getpid(), socket.gethostname())
    return "30_{0}_{1}".format(stamp, hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12])


def delta_rows(runtime, planned, run_context, txn_id, view_id, stage, corner,
               structure, before, after):
    rows = []
    for rule, record, bit, side in planned:
        rows.append({
            "schema_version": SCHEMA_VERSION,
            "run_id": run_context.get("run_id", ""),
            "mode_label": run_context.get("mode_label", ""),
            "stage_name": STAGE_NAME,
            "transaction_id": txn_id,
            "view_id": view_id,
            "stage": stage,
            "corner": corner,
            "structure_digest": structure,
            "accounting_digest_before": before,
            "accounting_digest_after": after,
            "workbook": record.workbook,
            "sheet": record.sheet,
            "row": record.row,
            "direction": record.direction,
            "port": record.shape.raw,
            "legal_bits": runtime.format_bits(set(record.shape.bits())),
            "added_bits": str(bit),
            "final_used_bits": runtime.format_bits(record.used_bits),
            "owner_object_id": clean(rule.values.get("exception_id")),
            "reason": "approved exception-only {0} endpoint".format(side),
            "evidence_status": "approved",
        })
    return rows


def all_required_30_ready(run_root, required_views, current_view, current_ready,
                          run_context, structure, upstream_digests, report):
    digests = {}
    ready = True
    current_id = clean(current_view.get("view_id"))
    for view in [row for row in required_views if norm(row.get("require_30")) == "yes"]:
        view_id = clean(view.get("view_id"))
        if view_id == current_id:
            if not current_ready:
                ready = False
            continue
        path = completion_path(run_root, clean(view.get("stage")), clean(view.get("corner")))
        payload = read_json(path, "30 required view completion", report, required=False)
        if not payload or norm(payload.get("completion_status")) != "complete":
            ready = False
            continue
        if clean(payload.get("stage_name")) != STAGE_NAME or clean(payload.get("view_id")) != view_id:
            report.error("{0}: stale 30 required view completion".format(path))
            ready = False
            continue
        if clean(payload.get("run_id")) != clean(run_context.get("run_id")) or clean(payload.get("structure_digest")) != structure:
            report.error("{0}: stale run/structure provenance".format(path))
            ready = False
            continue
        prior_upstream = payload.get("upstream_artifact_digests")
        if not isinstance(prior_upstream, dict) or prior_upstream != upstream_digests:
            report.error(
                "{0}: upstream_artifact_digests are stale for required view {1}".format(
                    path, view_id
                )
            )
            ready = False
            continue
        output = output_path(run_root, clean(view.get("stage")), clean(view.get("corner")))
        if not output.is_file() or clean(payload.get("output_sdc_digest")) != sha_file(output):
            report.error("{0}: output_sdc_digest is stale".format(path))
            ready = False
            continue
        digests[view_id] = sha_file(path)
    return ready and not report.error_count, digests


def completion_payload(legacy, run_context, view, stage, corner, structure,
                       before, after, output_digest, candidate_digest,
                       delta_digest, workbook_before, workbook_after,
                       upstream_digests, txn_id, completeness, final_status):
    return {
        "schema_version": SCHEMA_VERSION,
        "author": legacy.author_name(),
        "stage_name": STAGE_NAME,
        "run_id": run_context.get("run_id", ""),
        "mode_label": run_context.get("mode_label", ""),
        "design_revision": run_context.get("design_revision", ""),
        "view_id": clean(view.get("view_id")),
        "stage": stage,
        "corner": corner,
        "completion_status": "complete",
        "error_count": 0,
        "sync_changed": "no",
        "structure_digest": structure,
        "accounting_digest_before": before,
        "accounting_digest_after": after,
        "workbook_file_digest_before": workbook_before,
        "workbook_file_digest_after": workbook_after,
        "upstream_artifact_digests": upstream_digests,
        "output_sdc_digest": output_digest,
        "exception_candidates_digest": candidate_digest,
        "accounting_delta_digest": delta_digest,
        "transaction_id": txn_id,
        "run_completeness": completeness.status,
        "missing_instances": completeness.missing_instances,
        "final_accounting": final_status,
    }


def final_report_text(legacy, run_root, run_context, required_views, records,
                      owner_map, planned, structure, before, after,
                      workbook_after, closure, recovered, txn_id, completeness):
    planned_owners = defaultdict(list)
    for rule, record, bit, _side in planned:
        planned_owners[(record.inst_name, record.direction, record.shape.base, bit)].append(
            (STAGE_NAME, clean(rule.values.get("exception_id")))
        )
    lines = [
        "SoC SDC 30 Port Accounting Final Report",
        "Author: {0}".format(legacy.author_name()),
        "Run ID: {0}".format(run_context.get("run_id", "")),
        "Mode label: {0}".format(run_context.get("mode_label", "")),
        "Design revision: {0}".format(run_context.get("design_revision", "")),
        "Structure digest: {0}".format(structure),
        "Accounting digest before: {0}".format(before),
        "Accounting digest after: {0}".format(after),
        "Transaction ID: {0}".format(txn_id),
        "Recovery status: recovered_transactions={0}".format(recovered),
        "Accounting closure: {0}".format(closure),
        "",
        "Required 30 Views",
    ]
    for view in required_views:
        lines.append(
            "{0} stage={1} corner={2} require_02={3} require_04={4} require_20={5} require_30={6} complete=yes".format(
                clean(view.get("view_id")), clean(view.get("stage")), clean(view.get("corner")),
                clean(view.get("require_02")), clean(view.get("require_04")),
                clean(view.get("require_20")), clean(view.get("require_30")),
            )
        )
    lines.extend(["", "Port Rows"])
    totals = defaultdict(lambda: [0, 0, 0])
    all_rows = incomplete_rows = 0
    for record in records:
        legal = set(record.shape.bits())
        used = set(record.used_bits)
        unused = legal - used
        totals[record.inst_name][0] += len(legal)
        totals[record.inst_name][1] += len(used)
        totals[record.inst_name][2] += len(unused)
        if unused:
            incomplete_rows += 1
        else:
            all_rows += 1
        owners = []
        for bit in sorted(used):
            entries = list(owner_map.get((record.inst_name, record.direction, record.shape.base, bit), []))
            entries.extend(planned_owners.get((record.inst_name, record.direction, record.shape.base, bit), []))
            owners.append("{0}={1}".format(bit, "+".join("{0}:{1}".format(*item) for item in entries) or "UNEXPLAINED"))
        lines.append(
            "{0}:{1}:{2} instance={3} direction={4} port={5} legal={6} used={7} unused={8} owners={9}".format(
                record.workbook, record.sheet, record.row, record.inst_name,
                record.direction, record.shape.raw,
                ",".join(str(bit) for bit in sorted(legal)),
                ",".join(str(bit) for bit in sorted(used)) or "-",
                ",".join(str(bit) for bit in sorted(unused)) or "-",
                ";".join(owners) or "-",
            )
        )
    lines.extend(["", "Per Instance"])
    global_total = [0, 0, 0]
    for inst_name in sorted(totals):
        values = totals[inst_name]
        global_total = [a + b for a, b in zip(global_total, values)]
        lines.append("{0}: total={1} used={2} unused={3}".format(inst_name, *values))
    lines.extend([
        "Global: total={0} used={1} unused={2}".format(*global_total),
        "All-used rows: {0}".format(all_rows),
        "Incomplete rows: {0}".format(incomplete_rows),
        "Missing SDC related unused bits: {0}".format(sum(
            len(set(record.shape.bits()) - set(record.used_bits))
            for record in records if record.inst_name in set(completeness.missing_instances)
        )),
        "",
        "Final Workbook Digests",
    ])
    for name in sorted(workbook_after):
        lines.append("{0}={1}".format(name, workbook_after[name]))
    lines.append("")
    return "\n".join(lines)


def check_report_text(legacy, report, run_context, stage, corner, completeness,
                      structure, before, after, status, diagnostic, approved,
                      planned, final_status):
    lines = [
        "SoC SDC 30 Harden Interface Exception Check Report",
        "Author: {0}".format(legacy.author_name()),
        "Run ID: {0}".format(run_context.get("run_id", "")),
        "Mode label: {0}".format(run_context.get("mode_label", "")),
        "Design revision: {0}".format(run_context.get("design_revision", "")),
        "View: stage={0}, corner={1}".format(stage, corner),
        "Run completeness: {0}".format(completeness.status),
        "Completion status: {0}".format(status),
        "Diagnostic only: {0}".format("yes" if diagnostic else "no"),
        "Final accounting: {0}".format(final_status),
        "Structure digest: {0}".format(structure),
        "Accounting digest before: {0}".format(before),
        "Accounting digest after: {0}".format(after),
        "Approved rules: {0}".format(len(approved)),
        "New accounting bit claims: {0}".format(len(planned)),
        "Warnings: {0}".format(report.warning_count),
        "Errors: {0}".format(report.error_count),
        "",
        "Messages",
    ]
    lines.extend(report.lines or ["INFO: no messages"])
    lines.append("")
    return "\n".join(lines)


def run_target30(args, legacy):
    report = legacy.Report()
    stage10 = load_stage10(legacy)
    runtime = stage10.load_accounting_runtime()
    patch_accounting_runtime(runtime)
    run_root = Path(args.run_root).expanduser().resolve()
    input_root = run_root / "inputs"
    middle = run_root / "30_middle"
    form_path = middle / "30_harden_to_harden_exception.xlsx"
    candidate_path = middle / "exception_candidates.csv"
    delta_path = middle / "port_accounting_delta.csv"
    delta_meta_path = middle / "port_accounting_delta.meta"
    view_path = completion_path(run_root, args.stage, args.corner)
    stage_completion_path = middle / "stage_completion.meta"
    output = output_path(run_root, args.stage, args.corner)
    check_report = report_path(run_root, args.stage, args.corner)
    final_report = run_root / "30_result" / "reports" / "port_accounting_final_report.txt"
    lock_path = input_root / ".port_accounting.lock"
    run_context = {}
    required_views = []
    current_view = {}
    instances = {}
    info_semantic = []
    models = []
    records = []
    completeness = legacy.RunCompleteness(status="invalid")
    structure = before = after = ""
    approved = []
    planned = []
    upstream = {}
    preview_dir = None
    status = "failed"
    final_status = "not_attempted"
    recovered_count = 0
    diagnostic = bool(args.diagnose_only)
    try:
        if not input_root.is_dir():
            raise RuntimeError("required inputs directory is missing: {0}".format(input_root))
        with runtime.AccountingLock(lock_path, report):
            before_recovery = report.warning_count
            runtime.recover_transactions(run_root, report)
            recovered_count = report.warning_count - before_recovery
            run_context = runtime.read_run_context(input_root / "run_context.csv", report)
            required_views = runtime.read_required_views(input_root / "required_views.csv", report)
            current_view = current_required_view(required_views, args.stage, args.corner) or {}
            if not diagnostic and not current_view:
                report.error(
                    "formal 30 view stage={0}, corner={1} is not required with require_30=yes; use --diagnose-only".format(
                        args.stage, args.corner
                    )
                )
            instances, info_semantic = runtime.read_info_all(input_root / "info_all.xlsx", report)
            paths = runtime.discover_port_workbooks(input_root, report)
            models, records = runtime.read_port_workbooks(paths, run_root, instances, True, report)
            runtime.validate_connections(records, instances, report)
            structure = runtime.structure_digest(run_context, required_views, info_semantic, records)
            before = runtime.accounting_digest(records)
            after = before
            upstream = validate_upstreams(
                runtime, stage10, run_root, run_context, required_views,
                structure, before, records, report,
            )
            manifest, completeness = stage10.read_target_harden_manifest(
                run_root / "00_middle" / "harden_sdc_manifest.csv",
                run_root, instances, False, report,
            )
            contexts = stage10.build_target_direct_edges(runtime, records, report)
            candidates, rule_seeds, channels, evidence = build_seeds(
                legacy, stage10, contexts, upstream, manifest, completeness,
                instances, records, run_context, structure, before,
                args.stage, args.corner, report,
            )
            candidate_text = candidates_csv(legacy, candidates)
            candidate_digest = hashlib.sha256(candidate_text.encode("utf-8")).hexdigest()
            if report.error_count:
                raise RuntimeError("target 30 input validation failed")

            metadata = {
                "Author": legacy.author_name(),
                "Run ID": run_context.get("run_id", ""),
                "Mode Label": run_context.get("mode_label", ""),
                "Design Revision": run_context.get("design_revision", ""),
                "Stage": args.stage,
                "Corner": args.corner,
                "Structure Digest": structure,
                "Accounting Digest Before": before,
                "Run Completeness": completeness.status,
            }
            if diagnostic:
                diagnostic_path = middle / "diagnostic" / "exception_candidates.csv"
                legacy.atomic_write_text(diagnostic_path, candidate_text)
                final_status = "diagnostic"
                status = "diagnostic"
                text = check_report_text(
                    legacy, report, run_context, args.stage, args.corner,
                    completeness, structure, before, before, status, True,
                    [], [], final_status,
                )
                legacy.atomic_write_text(check_report, text)
            else:
                sync_workbook(legacy, form_path, candidates, rule_seeds, metadata, report)
                legacy.atomic_write_text(candidate_path, candidate_text)
                if report.sync_changed:
                    status = "review_required"
                    final_status = "deferred"
                    review_payload = {
                        "schema_version": SCHEMA_VERSION,
                        "stage_name": STAGE_NAME,
                        "run_id": run_context.get("run_id", ""),
                        "mode_label": run_context.get("mode_label", ""),
                        "design_revision": run_context.get("design_revision", ""),
                        "view_id": clean(current_view.get("view_id")),
                        "stage": args.stage, "corner": args.corner,
                        "completion_status": "review_required",
                        "error_count": 0, "sync_changed": "yes",
                        "structure_digest": structure,
                        "accounting_digest_before": before,
                        "accounting_digest_after": before,
                        "exception_candidates_digest": candidate_digest,
                        "final_accounting": "deferred",
                    }
                    legacy.atomic_write_text(view_path, json.dumps(review_payload, ensure_ascii=False, sort_keys=True, indent=2) + "\n")
                    legacy.atomic_write_text(check_report, check_report_text(
                        legacy, report, run_context, args.stage, args.corner,
                        completeness, structure, before, before, status, False,
                        [], [], final_status,
                    ))
                else:
                    rows = read_rules(legacy, form_path)
                    approved = validate_rules(
                        legacy, rows, channels, upstream.get("relations", {}),
                        args.stage, args.corner, report,
                    )
                    if report.error_count:
                        raise RuntimeError("target 30 review validation failed")
                    planned = plan_accounting(stage10, approved, channels, records, upstream, report)
                    apply_intermediate_cells(runtime, records)
                    after = runtime.accounting_digest(records)
                    if report.error_count:
                        raise RuntimeError("target 30 accounting plan failed")

                    ready, prior_view_digests = all_required_30_ready(
                        run_root, required_views, current_view, True,
                        run_context, structure,
                        upstream.get("upstream_digests", {}), report,
                    )
                    finalize = ready and not args.defer_final_accounting
                    if finalize:
                        _all_rows, incomplete_rows = apply_final_cells(records)
                        closure = "complete" if incomplete_rows == 0 else "incomplete"
                        if closure == "incomplete":
                            report.warn("final port accounting has unused bits")
                        if args.strict_port_closure and closure != "complete":
                            report.error("--strict-port-closure: final accounting is incomplete")
                            raise RuntimeError("strict port closure failed")
                        final_status = closure
                    else:
                        final_status = "deferred"

                    sdc_text = render_sdc(
                        legacy, approved, run_context, args.stage, args.corner,
                        completeness, structure, before, after,
                    )
                    sdc_payload = sdc_text.encode("utf-8")
                    sdc_digest = hashlib.sha256(sdc_payload).hexdigest()
                    txn_id = transaction_id(run_context.get("run_id", ""))
                    new_delta = delta_rows(
                        runtime, planned, run_context, txn_id,
                        clean(current_view.get("view_id")), args.stage, args.corner,
                        structure, before, after,
                    )
                    old_delta = runtime.load_delta_rows(delta_path, report) if delta_path.is_file() else []
                    all_delta = old_delta + new_delta
                    delta_text = runtime.csv_text(runtime.DELTA_HEADERS, all_delta)
                    delta_payload = delta_text.encode("utf-8")
                    delta_digest = hashlib.sha256(delta_payload).hexdigest()
                    old_meta = upstream.get("existing_30_meta", {})
                    txn_entry = {
                        "transaction_id": txn_id,
                        "committed_at": runtime.utc_timestamp(),
                        "structure_digest": structure,
                        "accounting_digest_before": before,
                        "accounting_digest_after": after,
                        "delta_rows_digest": runtime.delta_rows_digest(new_delta),
                    }
                    preview_dir = middle / ".30_preview_{0}".format(os.getpid())
                    if preview_dir.exists():
                        shutil.rmtree(str(preview_dir))
                    preview_dir.mkdir(parents=True, exist_ok=False)
                    prepared = {}
                    workbook_before = {model.relative_name: model.digest_before for model in models}
                    for model in models:
                        if model.modified:
                            path = preview_dir / model.path.name
                            model.workbook.save(str(path))
                            model.digest_after = sha_file(path)
                            prepared[model.relative_name] = path
                        else:
                            model.digest_after = model.digest_before
                    workbook_after = {model.relative_name: model.digest_after for model in models}
                    delta_meta = {
                        "schema_version": SCHEMA_VERSION,
                        "run_id": run_context.get("run_id", ""),
                        "mode_label": run_context.get("mode_label", ""),
                        "design_revision": run_context.get("design_revision", ""),
                        "stage_name": STAGE_NAME,
                        "completion_status": "complete",
                        "structure_digest": structure,
                        "accounting_digest_before": before,
                        "accounting_digest_after": after,
                        "workbook_file_digest_before": workbook_before,
                        "workbook_file_digest_after": workbook_after,
                        "delta_csv_digest": delta_digest,
                        "transactions": list(old_meta.get("transactions", [])) + [txn_entry],
                    }
                    view_payload = completion_payload(
                        legacy, run_context, current_view, args.stage, args.corner,
                        structure, before, after, sdc_digest, candidate_digest,
                        delta_digest, workbook_before, workbook_after,
                        upstream.get("upstream_digests", {}), txn_id,
                        completeness, final_status,
                    )
                    view_bytes = json_payload(view_payload)
                    artifacts = [
                        (candidate_path, candidate_text.encode("utf-8")),
                        (delta_path, delta_payload),
                        (delta_meta_path, json_payload(delta_meta)),
                        (output, sdc_payload),
                        (view_path, view_bytes),
                    ]
                    if finalize:
                        current_id = clean(current_view.get("view_id"))
                        required_digests = dict(prior_view_digests)
                        required_digests[current_id] = hashlib.sha256(view_bytes).hexdigest()
                        closure = final_status
                        final_text = final_report_text(
                            legacy, run_root, run_context, required_views, records,
                            upstream.get("owner_map", {}), planned, structure,
                            before, after, workbook_after, closure,
                            recovered_count, txn_id, completeness,
                        )
                        stage_payload = {
                            "schema_version": SCHEMA_VERSION,
                            "author": legacy.author_name(),
                            "stage_name": STAGE_NAME,
                            "run_id": run_context.get("run_id", ""),
                            "mode_label": run_context.get("mode_label", ""),
                            "design_revision": run_context.get("design_revision", ""),
                            "completion_status": "complete",
                            "error_count": 0, "sync_changed": "no",
                            "structure_digest": structure,
                            "accounting_digest_before": before,
                            "accounting_digest_after": after,
                            "accounting_closure": closure,
                            "required_view_completions": required_digests,
                            "port_accounting_final_report_digest": hashlib.sha256(final_text.encode("utf-8")).hexdigest(),
                            "transaction_id": txn_id,
                        }
                        artifacts.extend([
                            (final_report, final_text.encode("utf-8")),
                            (stage_completion_path, json_payload(stage_payload)),
                        ])
                    check_text = check_report_text(
                        legacy, report, run_context, args.stage, args.corner,
                        completeness, structure, before, after, "complete", False,
                        approved, planned, final_status,
                    )
                    artifacts.append((check_report, check_text.encode("utf-8")))
                    for model in models:
                        if not model.path.is_file() or sha_file(model.path) != model.digest_before:
                            report.error("concurrent workbook modification detected before 30 commit: {0}".format(model.path))
                    if report.error_count:
                        raise RuntimeError("target workbook snapshot changed before 30 commit")
                    runtime.execute_transaction(
                        run_root, models, prepared, artifacts, txn_id,
                        run_context, structure, before, after, report,
                    )
                    status = "complete"
    except Exception as exc:
        if not report.error_count:
            report.error(str(exc))
        status = "failed"
        try:
            legacy.atomic_write_text(check_report, check_report_text(
                legacy, report, run_context, args.stage, args.corner,
                completeness, structure, before, after or before, status,
                diagnostic, approved, planned, final_status,
            ))
        except Exception:
            pass
    finally:
        for model in models:
            try:
                model.workbook.close()
            except Exception:
                pass
        if preview_dir is not None and preview_dir.exists():
            shutil.rmtree(str(preview_dir))

    print("Author: {0}".format(legacy.author_name()))
    print("Run ID: {0}".format(run_context.get("run_id", "")))
    print("Mode label: {0}".format(run_context.get("mode_label", "")))
    print("Design revision: {0}".format(run_context.get("design_revision", "")))
    print("View: stage={0}, corner={1}".format(args.stage, args.corner))
    print("Run completeness: {0}".format(completeness.status))
    print("Port accounting: {0}".format("diagnostic/read-only" if diagnostic else ("committed" if status == "complete" else "not committed")))
    print("Final accounting: {0}".format(final_status))
    print("Completion status: {0}".format(status))
    print("Report: {0}".format(check_report))
    print("Warnings: {0}  Errors: {1}  Sync changed: {2}".format(
        report.warning_count, report.error_count, report.sync_changed
    ))
    if report.error_count or status in {"failed", "review_required"}:
        return 1
    return 0
